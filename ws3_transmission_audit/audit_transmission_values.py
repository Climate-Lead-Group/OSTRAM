# -*- coding: utf-8 -*-
"""
WS-3 Phase 1 — Transmission value & sourcing audit (READ-ONLY).

Builds the parameter-to-source matrix by reconciling, for every transmission
technology and parameter:

  * LIVE value      - what the SOLVED model actually used (Executables/<scen>_0/*_Input.csv)
  * v18 SOURCED     - the value in the live template's Interconnector_Params sheet
                      (interconnectors) or Config_country_codes.yaml (internal families)
  * COST-DB value   - SoAsia_OSTRAM_Cost_Database.xlsx (interconnectors only)
  * REFS citation   - SOASIA_v18_REFS.xlsx IEEE Reference column (interconnectors)

Emits a timestamped CSV matrix + a console summary. Changes NO model value.

Run:
    conda activate OSTRAM-env
    python ws3_transmission_audit/audit_transmission_values.py
"""
from __future__ import annotations
import os, re, csv, sys
from datetime import datetime

import openpyxl
import pandas as pd
import yaml

# ----------------------------------------------------------------------------- paths
HERE   = os.path.dirname(os.path.abspath(__file__))
REPO   = os.path.abspath(os.path.join(HERE, ".."))
INPUTS = os.path.join(HERE, "inputs")
OUTPUTS = os.path.join(HERE, "outputs")
os.makedirs(OUTPUTS, exist_ok=True)

V18    = os.path.join(REPO, "t1_confection", "A3_process", "SOASIA_OSeMOSYS_Template_v18.xlsx")
REFS   = os.path.join(INPUTS, "SOASIA_v18_REFS.xlsx")
COSTDB = os.path.join(INPUTS, "SoAsia_OSTRAM_Cost_Database.xlsx")
YAMLF  = os.path.join(REPO, "t1_confection", "Config_country_codes.yaml")
SOLVED = {
    "A_Calibrated_BAU": os.path.join(REPO, "t1_confection", "Executables", "A_Calibrated_BAU_0", "A_Calibrated_BAU_0_Input.csv"),
    "B_Optimised_VRE":  os.path.join(REPO, "t1_confection", "Executables", "B_Optimised_VRE_0",  "B_Optimised_VRE_0_Input.csv"),
    "C_Target_VRE":     os.path.join(REPO, "t1_confection", "Executables", "C_Target_VRE_0",     "C_Target_VRE_0_Input.csv"),
}

IC = re.compile(r"^TRN[A-Z]{5}[A-Z]{5}$")            # 13-char interconnector code
INTERNAL = ("RNWTRN", "RNWRPO", "RNWNLI", "PWRTRN", "TRNRPO", "TRNNLI")

# COST-DB corridor code  ->  live/v18 interconnector tech code
COSTDB_TO_V18 = {
    "TRNCBD_BD_IN_E":  "TRNBGDXXINDEA", "TRNCBD_BD_IN_NE": "TRNBGDXXINDNE",
    "TRNCBD_BT_IN_E":  "TRNBTNXXINDEA", "TRNCBD_BT_IN_NE": "TRNBTNXXINDNE",
    "TRNCBD_NP_IN_N":  "TRNINDNONPLXX", "TRNCBD_NP_IN_E":  "TRNINDEANPLXX",
    "TRNCBD_LK_IN_S":  "TRNINDSOLKAXX", "TRNCBD_IN_N_IN_E":"TRNINDEAINDNO",
    "TRNCBD_IN_N_IN_W":"TRNINDNOINDWE", "TRNCBD_IN_E_IN_NE":"TRNINDEAINDNE",
    "TRNCBD_IN_E_IN_S":"TRNINDEAINDSO", "TRNCBD_IN_W_IN_S":"TRNINDSOINDWE",
    "TRNCBD_MV_IN_S":  "TRNMDVXXINDSO",
}

SUBMARINE = {"TRNINDSOLKAXX", "TRNMDVXXINDSO", "TRNLKAXXMDVXX"}
INDIA_INTERNAL = {"TRNINDEAINDNO", "TRNINDNOINDWE", "TRNINDEAINDNE", "TRNINDEAINDSO",
                  "TRNINDSOINDWE", "TRNINDEAINDWE", "TRNINDNEINDNO"}

def segment(tech: str) -> str:
    if tech in SUBMARINE:
        return "submarine-HVDC"
    if tech in INDIA_INTERNAL:
        return "india-internal"
    if tech in ("TRNNPLXXBGDXX", "TRNBTNXXBGDXX"):
        return "cross-border-via-India"
    return "cross-border-overhead"

def is_year(h):
    try:
        v = int(float(h))
        return 1990 <= v <= 2100
    except (TypeError, ValueError):
        return False

def approx_eq(a, b, tol=0.5):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a) == str(b)

# ---------------------------------------------------------------- Interconnector_Params
def read_ic_params(path, want_citation=False):
    """Return {(tech,param): {'value':..,'unit':..,'citation':..}} from Interconnector_Params (BAU)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Interconnector_Params"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    hdr = list(rows[0])
    col = {name: i for i, name in enumerate(hdr) if name is not None}
    def idx(*names):
        for n in names:
            if n in col:
                return col[n]
        return None
    i_scen = idx("scenario"); i_tech = idx("Tech", "Tech.ID2", "Technology")
    i_par = idx("Parameter"); i_unit = idx("Unit")
    i_cit = idx("IEEE Reference", "Reference", "Source")
    year_cols = [i for i, h in enumerate(hdr) if is_year(h)]
    out = {}
    for r in rows[1:]:
        if not r or (i_scen is not None and str(r[i_scen]) not in ("BAU", "None")):
            if i_scen is not None and r[i_scen] not in (None, "BAU"):
                continue
        tech = r[i_tech] if i_tech is not None else None
        par = r[i_par] if i_par is not None else None
        if not tech or not par:
            continue
        yr = [r[i] for i in year_cols if i < len(r) and r[i] not in (None, "")]
        distinct = sorted(set(yr), key=lambda x: str(x))
        if not distinct:
            value = ""
        elif len(distinct) == 1:
            value = distinct[0]
        else:
            value = f"{min(yr)}–{max(yr)} (varies)"
        rec = {"value": value, "unit": (r[i_unit] if i_unit is not None else "")}
        if want_citation and i_cit is not None and i_cit < len(r):
            cit = r[i_cit]
            rec["citation"] = "" if cit is None else " ".join(str(cit).split())
        out[(str(tech), str(par))] = rec
    return out

# ---------------------------------------------------------------- COST-DB Interconnectors
def read_costdb():
    wb = openpyxl.load_workbook(COSTDB, read_only=True, data_only=True)
    ws = wb["Interconnectors"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    hdr = [str(h) if h is not None else "" for h in rows[0]]
    ci = {h: i for i, h in enumerate(hdr)}
    out = {}
    for r in rows[1:]:
        if not r or not r[ci["Code"]]:
            continue
        v18 = COSTDB_TO_V18.get(str(r[ci["Code"]]))
        if not v18:
            continue
        out[v18] = {
            "CapitalCost": r[ci.get("CAPEX_USD/kW", -1)],
            "FixedCost":   r[ci.get("FOM_USD/kW/yr", -1)],
            "loss":        r[ci.get("Loss_%", -1)],
            "life":        r[ci.get("Life_yr", -1)],
            "confidence":  r[ci.get("Confidence", -1)],
            "source":      r[ci.get("Primary_Source", -1)],
        }
    return out

# ---------------------------------------------------------------- solved _Input.csv
PARAMS_IC  = ["CapitalCost", "FixedCost", "OperationalLife", "ResidualCapacity", "CapacityFactor", "CapacityToActivityUnit"]
def read_solved(path):
    """Return {'ic':{tech:{param:val}}, 'int':{family:{param:set}}} from a solved _Input.csv."""
    usecols = ["TECHNOLOGY"] + [p for p in PARAMS_IC]
    df = pd.read_csv(path, usecols=lambda c: c in usecols, low_memory=False)
    df["TECHNOLOGY"] = df["TECHNOLOGY"].astype(str)
    ic, intl = {}, {}
    for _, row in df.iterrows():
        t = row["TECHNOLOGY"]
        if IC.match(t):
            d = ic.setdefault(t, {})
            for p in PARAMS_IC:
                if p in df.columns and pd.notna(row[p]):
                    d.setdefault(p, set()).add(round(float(row[p]), 3))
        else:
            fam = next((p for p in INTERNAL if t.startswith(p)), None)
            if fam:
                d = intl.setdefault(fam, {})
                for p in PARAMS_IC:
                    if p in df.columns and pd.notna(row[p]):
                        d.setdefault(p, set()).add(round(float(row[p]), 3))
    return {"ic": ic, "int": intl}

def fmt_set(s):
    if not s:
        return ""
    vals = sorted(s)
    if len(vals) == 1:
        return vals[0]
    return f"{min(vals)}–{max(vals)}"

# ---------------------------------------------------------------- YAML internal families
def read_yaml_internal():
    with open(YAMLF, "r", encoding="utf-8") as fh:
        y = yaml.safe_load(fh)
    out = {}
    for fam in INTERNAL:
        if fam in y:
            out[fam] = y[fam]
    return out

# =============================================================================== main
def main():
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    v18 = read_ic_params(V18)
    refs = read_ic_params(REFS, want_citation=True)
    costdb = read_costdb()
    yaml_int = read_yaml_internal()
    solved = {scen: read_solved(p) for scen, p in SOLVED.items() if os.path.isfile(p)}
    base = solved.get("A_Calibrated_BAU", {"ic": {}, "int": {}})

    # cross-scenario identity check (transmission only)
    ident = True
    for scen, s in solved.items():
        if s["ic"] != base["ic"] or s["int"] != base["int"]:
            ident = False
    print(f"[cross-scenario] transmission inputs identical across {list(solved)}: {ident}")

    rows = []
    all_ic_techs = sorted(set(base["ic"]) | {t for (t, _) in v18})
    for tech in all_ic_techs:
        seg = segment(tech)
        cdb = costdb.get(tech, {})
        for param in ["CapitalCost", "FixedCost", "OperationalLife", "ResidualCapacity", "CapacityFactor"]:
            live = fmt_set(base["ic"].get(tech, {}).get(param))
            src = v18.get((tech, param), {})
            v18v = src.get("value", "")
            unit = src.get("unit", "")
            cit = refs.get((tech, param), {}).get("citation", "")
            cdbv = {"CapitalCost": cdb.get("CapitalCost"), "FixedCost": cdb.get("FixedCost"),
                    "OperationalLife": cdb.get("life")}.get(param, "")
            in_v18 = (tech, param) in v18
            match = approx_eq(live, v18v) if (live != "" and v18v != "") else None
            # gap classification
            if not in_v18:
                gap = "EXTRA-not-in-v18"
            elif cit.startswith("[Pending") or "not in" in cit:
                gap = "GAP-no-citation"
            elif param == "CapitalCost" and str(cdb.get("confidence", "")).lower() == "low":
                gap = "LOW-CONFIDENCE"
            elif cit:
                gap = "SOURCED"
            else:
                gap = "unclassified"
            if match is False:
                gap += " | LIVE≠v18(stale?)"
            rows.append({
                "domain": "INTERCONNECTOR", "segment": seg, "tech_or_family": tech,
                "parameter": param, "unit": unit,
                "live_solved": live, "v18_sourced": v18v, "costdb": cdbv,
                "live==v18": "" if match is None else match,
                "confidence": cdb.get("confidence", ""),
                "gap_flag": gap,
                "refs_citation": (cit[:180] + "…") if len(cit) > 180 else cit,
                "costdb_source": cdb.get("source", ""),
            })

    # internal families
    for fam in INTERNAL:
        y = yaml_int.get(fam, {})
        for param in ["CapitalCost", "FixedCost", "OperationalLife", "ResidualCapacity", "CapacityToActivityUnit"]:
            live = fmt_set(base["int"].get(fam, {}).get(param))
            yv = y.get(param, "")
            match = approx_eq(live, yv) if (live != "" and yv != "") else None
            gap = "PLACEHOLDER-unsourced"
            if param == "OperationalLife" and match is False:
                gap = "DRIFT: live≠YAML"
            rows.append({
                "domain": "INTERNAL", "segment": ("internal-RE" if fam.startswith("RNW") else "internal-nonRE"),
                "tech_or_family": fam, "parameter": param, "unit": "",
                "live_solved": live, "v18_sourced": f"YAML:{yv}", "costdb": "",
                "live==v18": "" if match is None else match, "confidence": "",
                "gap_flag": gap, "refs_citation": "", "costdb_source": "",
            })

    out_csv = os.path.join(OUTPUTS, f"parameter_source_matrix_{ts}.csv")
    cols = ["domain", "segment", "tech_or_family", "parameter", "unit", "live_solved",
            "v18_sourced", "costdb", "live==v18", "confidence", "gap_flag",
            "refs_citation", "costdb_source"]
    with open(out_csv, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.DictWriter(fh, fieldnames=cols)
        w.writeheader()
        w.writerows(rows)

    # summary
    from collections import Counter
    print(f"\n[matrix] {len(rows)} rows -> {out_csv}")
    print("[gap flags]")
    for g, n in Counter(r["gap_flag"] for r in rows).most_common():
        print(f"    {n:>3}  {g}")
    stale = [r for r in rows if r["live==v18"] is False and r["domain"] == "INTERCONNECTOR"]
    print(f"\n[interconnector params where LIVE != v18-sourced]: {len(stale)}")
    for r in sorted(set((r["tech_or_family"], r["parameter"], str(r['live_solved']), str(r['v18_sourced'])) for r in stale)):
        print(f"    {r[0]:15} {r[1]:14} live={r[2]:<14} v18={r[3]}")

if __name__ == "__main__":
    main()
