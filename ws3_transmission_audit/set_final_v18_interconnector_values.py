# -*- coding: utf-8 -*-
"""
set_final_v18_interconnector_values.py  (WS-3 Phase 3)

One-shot editor of the v18 template's Interconnector_Params sheet, applying the
gate-approved final values:

  * Submarine raise (research, 2023 USD):
        TRNINDSOLKAXX (LK<->IN_S) CapitalCost 1031 -> 1250, FixedCost -> 18.75
        TRNMDVXXINDSO (MV<->IN_S) CapitalCost 1600 -> 2800, FixedCost -> 42.0
  * OperationalLife = 40 for every interconnector corridor (was 40 XB / 50 internal / 30 MV).
  * Add the 3 corridors present in the model but missing from Interconnector_Params:
        TRNINDEAINDWE (IN_E<->IN_W)  CapEx 691.399  (legacy CEA-consistent, now documented)
        TRNINDNEINDNO (IN_NE<->IN_N) CapEx 645.703  (legacy CEA-consistent, now documented)
        TRNLKAXXMDVXX (LK<->MV)      CapEx 1250     (repriced subsea; was 508 mispriced as overhead)
    FixedCost = 1.5% x CapEx; OperationalLife = 40 for all three.

FOM convention: FixedCost = 1.5% x CapitalCost.
Edits the workbook IN PLACE (backup written first). Read-only elsewhere.
"""
import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

V18 = Path(r"C:\Users\luisfernando\Desktop\OSeMOSYS\OSTRAM_ws3_workcopy\t1_confection\A3_process\SOASIA_OSeMOSYS_Template_v18.xlsx")
SHEET = "Interconnector_Params"
APPLIED = ("CapitalCost", "FixedCost", "OperationalLife")
FOM_RATE = 0.015

# CapitalCost updates for EXISTING techs (FixedCost recomputed at 1.5%)
CAPEX_UPDATES = {
    "TRNINDSOLKAXX": 1250.0,   # LK<->IN_S submarine (research ~1250)
    "TRNMDVXXINDSO": 2800.0,   # MV<->IN_S submarine (research ~2800 @400MW)
}
LIFE_ALL = 40  # OperationalLife for every interconnector corridor

# New corridors to ADD (present in model / OG_csvs, absent from Interconnector_Params)
ADDITIONS = {
    "TRNINDEAINDWE": {"name": "Cross-border interconnector: India East ↔ India West (internal)",
                      "CapitalCost": 691.399},
    "TRNINDNEINDNO": {"name": "Cross-border interconnector: India Northeast ↔ India North (internal)",
                      "CapitalCost": 645.703},
    "TRNLKAXXMDVXX": {"name": "Cross-border interconnector: Sri Lanka ↔ Maldives (submarine cable)",
                      "CapitalCost": 1250.0},
}


def main() -> int:
    if not V18.is_file():
        print(f"ERROR: v18 not found: {V18}", file=sys.stderr)
        return 1
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = V18.with_name(V18.stem + f"_PRE_WS3_VALUES_{stamp}.xlsx")
    shutil.copy(V18, backup)
    print(f"Backup: {backup.name}")

    wb = openpyxl.load_workbook(V18)
    ws = wb[SHEET]

    hdr = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    c_scen = hdr.get("scenario"); c_tid = hdr.get("Tech.ID"); c_tech = hdr.get("Tech")
    c_tname = hdr.get("Tech.Name"); c_pid = hdr.get("Parameter.ID"); c_par = hdr.get("Parameter")
    c_unit = hdr.get("Unit"); c_pmode = hdr.get("Projection.Mode"); c_pparam = hdr.get("Projection.Parameter")
    year_cols = [c for c in range(1, ws.max_column + 1)
                 if isinstance(ws.cell(row=1, column=c).value, int)
                 and 1900 <= ws.cell(row=1, column=c).value <= 2200]
    if not all([c_tech, c_par]) or not year_cols:
        print("ERROR: could not locate Tech/Parameter/year columns", file=sys.stderr)
        return 2

    def set_row_years(row_idx, value):
        for yc in year_cols:
            ws.cell(row=row_idx, column=yc, value=value)

    # index existing (tech, param) -> row; capture a template row per param for cloning
    rowmap = {}
    param_template = {}
    max_tid = 0
    for r in range(2, ws.max_row + 1):
        t = ws.cell(row=r, column=c_tech).value
        p = ws.cell(row=r, column=c_par).value
        if t is None:
            continue
        rowmap[(str(t), str(p))] = r
        if c_tid and isinstance(ws.cell(row=r, column=c_tid).value, (int, float)):
            max_tid = max(max_tid, int(ws.cell(row=r, column=c_tid).value))
        if p in APPLIED and p not in param_template:
            param_template[p] = r

    changes = []

    # 1. CapitalCost updates + FixedCost recompute for existing submarine techs
    for tech, capex in CAPEX_UPDATES.items():
        r = rowmap.get((tech, "CapitalCost"))
        if r:
            old = ws.cell(row=r, column=year_cols[0]).value
            set_row_years(r, capex); changes.append(f"{tech} CapitalCost {old} -> {capex}")
        rf = rowmap.get((tech, "FixedCost"))
        if rf:
            fom = round(capex * FOM_RATE, 4)
            old = ws.cell(row=rf, column=year_cols[0]).value
            set_row_years(rf, fom); changes.append(f"{tech} FixedCost {old} -> {fom}")

    # 2. OperationalLife = 40 for every existing interconnector corridor
    for (tech, par), r in list(rowmap.items()):
        if par == "OperationalLife" and tech.startswith("TRN") and len(tech) == 13:
            old = ws.cell(row=r, column=year_cols[0]).value
            if old != LIFE_ALL:
                set_row_years(r, LIFE_ALL); changes.append(f"{tech} OperationalLife {old} -> {LIFE_ALL}")

    # 3. Add the 3 missing corridors (CapitalCost, FixedCost, OperationalLife)
    next_row = ws.max_row + 1
    next_tid = max_tid + 1
    for tech, spec in ADDITIONS.items():
        if (tech, "CapitalCost") in rowmap:
            changes.append(f"{tech} already present — skipped add")
            continue
        capex = spec["CapitalCost"]
        vals = {"CapitalCost": capex, "FixedCost": round(capex * FOM_RATE, 4), "OperationalLife": LIFE_ALL}
        for par in APPLIED:
            tmpl = param_template[par]
            # clone the template row cell-by-cell, then override identity + value
            for c in range(1, ws.max_column + 1):
                ws.cell(row=next_row, column=c, value=ws.cell(row=tmpl, column=c).value)
            if c_scen: ws.cell(row=next_row, column=c_scen, value="BAU")
            if c_tid: ws.cell(row=next_row, column=c_tid, value=next_tid)
            ws.cell(row=next_row, column=c_tech, value=tech)
            if c_tname: ws.cell(row=next_row, column=c_tname, value=spec["name"])
            ws.cell(row=next_row, column=c_par, value=par)
            if c_pmode: ws.cell(row=next_row, column=c_pmode, value="User defined")
            set_row_years(next_row, vals[par])
            changes.append(f"ADD {tech} {par} = {vals[par]}")
            next_row += 1
        next_tid += 1

    wb.save(V18); wb.close()
    print(f"\nApplied {len(changes)} changes to {SHEET}:")
    for ch in changes:
        print("   ", ch)
    return 0


if __name__ == "__main__":
    sys.exit(main())
