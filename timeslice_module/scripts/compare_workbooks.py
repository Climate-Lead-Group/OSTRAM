# -*- coding: utf-8 -*-
"""Task 3: cell-by-cell workbook comparison (NOT hash-based).

Numeric cells   : equal within rtol=1e-9 (reports max abs and max rel diff per sheet)
Text cells      : exact string equality
Structure       : sheet names, sheet order, used range (max_row x max_col)
Type mismatches : counted separately (number vs text in the same cell)
"""
import sys, math, json
import openpyxl

RTOL = 1e-9
MAX_EXAMPLES = 12

def colname(i):
    s = ""
    while i > 0:
        i, r = divmod(i - 1, 26)
        s = chr(65 + r) + s
    return s

def load(path):
    return openpyxl.load_workbook(path, data_only=True, read_only=False)

def compare(ref_path, new_path, label_ref, label_new):
    wb_r, wb_n = load(ref_path), load(new_path)
    rs, ns = wb_r.sheetnames, wb_n.sheetnames

    report = {
        "ref": ref_path, "new": new_path,
        "sheets_ref": rs, "sheets_new": ns,
        "sheet_names_identical": rs == ns,
        "only_in_ref": [s for s in rs if s not in ns],
        "only_in_new": [s for s in ns if s not in rs],
        "sheets": [],
    }

    for name in rs:
        if name not in ns:
            continue
        a, b = wb_r[name], wb_n[name]
        ent = {
            "sheet": name,
            "dims_ref": [a.max_row, a.max_column],
            "dims_new": [b.max_row, b.max_column],
            "dims_match": (a.max_row == b.max_row and a.max_column == b.max_column),
            "cells_compared": 0,
            "numeric_cells": 0, "text_cells": 0, "blank_cells": 0,
            "numeric_diffs": 0, "text_diffs": 0, "type_diffs": 0,
            "max_abs_diff": 0.0, "max_rel_diff": 0.0,
            "max_abs_at": None, "max_rel_at": None,
            "examples": [],
        }
        R = max(a.max_row, b.max_row)
        C = max(a.max_column, b.max_column)
        for r in range(1, R + 1):
            for c in range(1, C + 1):
                va = a.cell(row=r, column=c).value
                vb = b.cell(row=r, column=c).value
                ent["cells_compared"] += 1
                addr = f"{colname(c)}{r}"

                if va is None and vb is None:
                    ent["blank_cells"] += 1
                    continue

                na = isinstance(va, (int, float)) and not isinstance(va, bool)
                nb = isinstance(vb, (int, float)) and not isinstance(vb, bool)

                if na and nb:
                    ent["numeric_cells"] += 1
                    if math.isnan(va) and math.isnan(vb):
                        continue
                    ad = abs(va - vb)
                    den = max(abs(va), abs(vb))
                    rd = ad / den if den > 0 else 0.0
                    if ad > ent["max_abs_diff"]:
                        ent["max_abs_diff"], ent["max_abs_at"] = ad, addr
                    if rd > ent["max_rel_diff"]:
                        ent["max_rel_diff"], ent["max_rel_at"] = rd, addr
                    if not (ad == 0.0 or rd <= RTOL):
                        ent["numeric_diffs"] += 1
                        if len(ent["examples"]) < MAX_EXAMPLES:
                            ent["examples"].append(
                                f"{addr}: NUM ref={va!r} new={vb!r} abs={ad:.3e} rel={rd:.3e}")
                elif (not na) and (not nb):
                    ent["text_cells"] += 1
                    if str(va) != str(vb):
                        ent["text_diffs"] += 1
                        if len(ent["examples"]) < MAX_EXAMPLES:
                            ent["examples"].append(f"{addr}: TXT ref={va!r} new={vb!r}")
                else:
                    ent["type_diffs"] += 1
                    if len(ent["examples"]) < MAX_EXAMPLES:
                        ent["examples"].append(f"{addr}: TYPE ref={va!r} new={vb!r}")

        ent["verdict"] = ("PARITY" if (ent["dims_match"] and ent["numeric_diffs"] == 0
                                       and ent["text_diffs"] == 0 and ent["type_diffs"] == 0)
                          else "NO-PARITY")
        report["sheets"].append(ent)

    report["overall"] = (
        "PARITY" if (report["sheet_names_identical"]
                     and all(s["verdict"] == "PARITY" for s in report["sheets"]))
        else "NO-PARITY")
    wb_r.close(); wb_n.close()
    return report


if __name__ == "__main__":
    ref, new, out = sys.argv[1], sys.argv[2], sys.argv[3]
    rep = compare(ref, new, "REFERENCE", "FRESH")
    with open(out, "w", encoding="utf-8") as f:
        json.dump(rep, f, indent=2)

    print(f"sheet names identical : {rep['sheet_names_identical']}")
    print(f"sheets: {len(rep['sheets'])}")
    print(f"{'sheet':<16}{'dims':<14}{'cells':>8}{'num':>7}{'txt':>7}"
          f"{'ndiff':>7}{'tdiff':>7}{'typed':>7}{'maxabs':>12}{'maxrel':>12}  verdict")
    for s in rep["sheets"]:
        print(f"{s['sheet']:<16}"
              f"{str(s['dims_ref'][0])+'x'+str(s['dims_ref'][1]):<14}"
              f"{s['cells_compared']:>8}{s['numeric_cells']:>7}{s['text_cells']:>7}"
              f"{s['numeric_diffs']:>7}{s['text_diffs']:>7}{s['type_diffs']:>7}"
              f"{s['max_abs_diff']:>12.3e}{s['max_rel_diff']:>12.3e}  {s['verdict']}")
        for ex in s["examples"]:
            print(f"    {ex}")
    print(f"\nOVERALL: {rep['overall']}")
