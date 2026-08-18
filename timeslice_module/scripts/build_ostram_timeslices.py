"""
build_ostram_timeslices.py
==========================
Consolidated OSTRAM Timeslice Parameter Generator
  Countries: BGD, LKA, IND (x5 sub-regions), NPL, BTN

Timeslice structure is FULLY CONFIGURABLE via SEASON_DEF and DAYPART_DEF.
Change number, boundaries, or labels in one place — everything derives
automatically. Keep the SAME config as build_reninja_timeslices.py.

Open in Spyder and press F5.

Author: CLG / Luis Victor-Gallardo
"""

import pandas as pd
import numpy as np
import openpyxl
import os
import glob
import json
import re
import csv
from datetime import datetime, time as dtime, timedelta
from collections import defaultdict
import time

# ======================================================================
# USER CONFIGURATION
# ======================================================================

BASE_DIR = r"C:\Users\luisfernando\Desktop\timeslice_module\inputs"
OUTPUT_DIR = r"C:\Users\luisfernando\Desktop\timeslice_module\outputs\run_5dp20ts"

BGD_PGCB_DIR = os.path.join(BASE_DIR, r"Bangladesh\pgcb_analysis\pgcb_daily_files")
LKA_PUCSL_DIR = os.path.join(BASE_DIR, r"Sri Lanka\PUCSL")
IND_DEMAND_DIR = os.path.join(BASE_DIR, r"India\Initial_Analysis_Sources")
IND_HYDRO_CF_FILE = os.path.join(BASE_DIR, r"India\Hydro_CF\India_Hydro_CF_from_CEA.xlsx")
IND_HYDRO_CF_SHEET = "CF_Pivot"
BTN_DA_FILE = os.path.join(BASE_DIR, r"Bhutan\BTN_DA_v4.xlsx")
BTN_PROFILES_FILE = os.path.join(BASE_DIR, r"Bhutan\BTN_Profiles.xlsx")
NPL_PROFILES_CSV = os.path.join(BASE_DIR, r"Nepal\Nepal_HourlyProfiles_Literature.csv")
NINJA_TS_FILE = os.path.join(BASE_DIR, r"_Reno_Ninja\ninja_data\output_rebuilt\compiled_reninja_ts.csv")
NINJA_CONFIG_FILE = os.path.join(BASE_DIR, r"_Reno_Ninja\ninja_data\output_rebuilt\timeslice_config.json")

COUNTRIES = ['BGD', 'LKA', 'IND', 'NPL', 'BTN', 'MDV']
SPR_MULTIPLIER = 0.85

UTC_OFFSETS = {
    'BGD': 6, 'BTN': 6, 'IND': 5.5, 'NPL': 5.75, 'LKA': 5.5, 'MDV': 5,
    'INDNO': 5.5, 'INDEA': 5.5, 'INDNE': 5.5, 'INDSO': 5.5, 'INDWE': 5.5,
    'BGDXX': 6, 'BTNXX': 6, 'NPLXX': 5.75, 'LKAXX': 5.5, 'MDVXX': 5,
}

# ======================================================================
# TIMESLICE SYSTEM DEFINITION
#
# Edit SEASON_DEF and DAYPART_DEF to reshape the timeslice grid.
# Everything below derives automatically. Keep IDENTICAL to
# build_reninja_timeslices.py so Ninja outputs align.
# ======================================================================

SEASON_DEF = [
    ('S1', 'Winter (Dec-Feb)',       [12, 1, 2],       90),
    ('S2', 'Pre-monsoon (Mar-May)',  [3, 4, 5],        92),
    ('S3', 'SW Monsoon (Jun-Sep)',   [6, 7, 8, 9],    122),
    ('S4', 'Post-monsoon (Oct-Nov)', [10, 11],          61),
]

# Scheme: 5dp_D_6_17 — 5 dp (solar 6-17, tight)
# Selected via sensitivity_timeslice_sweep.py ranking (April 2026).
# Design: one fat solar block (06-17) maximises SCV by contrasting
# "solar" vs. "no-solar" hours; three short evening/night blocks
# capture the post-sunset demand peak and ramp-down.
DAYPART_DEF = [
    ('D1', 'Night',         0,  6),
    ('D2', 'Solar day',     6, 17),
    ('D3', 'Evening peak', 17, 20),
    ('D4', 'Late evening', 20, 22),
    ('D5', 'Late night',   22, 24),
]
# ======================================================================
# DERIVED CONSTANTS — auto-generated, do NOT edit
# ======================================================================

def _validate_and_build():
    s_codes = [s[0] for s in SEASON_DEF]
    s_names = {s[0]: s[1] for s in SEASON_DEF}
    s_days  = {s[0]: s[3] for s in SEASON_DEF}
    m2s = {}
    for code, label, months, days in SEASON_DEF:
        for m in months:
            if m in m2s:
                raise ValueError(f"Month {m} in {m2s[m]} and {code}")
            m2s[m] = code
    if sorted(m2s.keys()) != list(range(1, 13)):
        raise ValueError(f"Not all months assigned: {sorted(m2s.keys())}")
    if sum(s_days.values()) != 365:
        raise ValueError(f"Season days = {sum(s_days.values())}, not 365")

    d_codes  = [d[0] for d in DAYPART_DEF]
    d_names  = {d[0]: f"{d[1]} ({d[2]:02d}-{d[3]:02d})" for d in DAYPART_DEF}
    d_hours  = {d[0]: d[3] - d[2] for d in DAYPART_DEF}
    d_ranges = {d[0]: (d[2], d[3]) for d in DAYPART_DEF}
    if DAYPART_DEF[0][2] != 0 or DAYPART_DEF[-1][3] != 24:
        raise ValueError("Dayparts must span 0-24")
    for i in range(1, len(DAYPART_DEF)):
        if DAYPART_DEF[i][2] != DAYPART_DEF[i-1][3]:
            raise ValueError(f"Gap between {DAYPART_DEF[i-1][0]} and {DAYPART_DEF[i][0]}")

    ts = [s + d for s in s_codes for d in d_codes]
    ys = {s + d: round((sd * dh) / 8760, 6)
          for s, sd in s_days.items() for d, dh in d_hours.items()}
    lut = {}
    for dc, dl, ds, de in DAYPART_DEF:
        for h in range(ds, de):
            lut[h] = dc

    return dict(s_codes=s_codes, s_names=s_names, s_days=s_days, m2s=m2s,
                d_codes=d_codes, d_names=d_names, d_hours=d_hours,
                d_ranges=d_ranges, ts=ts, ys=ys, lut=lut)

_C = _validate_and_build()
SEASONS_LIST    = _C['s_codes']
SEASON_NAMES    = _C['s_names']
SEASON_DAYS_365 = _C['s_days']
MONTH_TO_SEASON = _C['m2s']
DAYPARTS_LIST   = _C['d_codes']
DAYPART_NAMES   = _C['d_names']
DAYPART_HOURS   = _C['d_hours']
DAYPART_RANGES  = _C['d_ranges']
TIMESLICES      = _C['ts']
YEARSPLIT       = _C['ys']
_HOUR_LUT       = _C['lut']
N_SEASONS       = len(SEASONS_LIST)
N_DAYPARTS      = len(DAYPARTS_LIST)
N_TIMESLICES    = len(TIMESLICES)

THERMAL_TECHS = ['COA', 'GAS', 'OIL', 'PET', 'HFO', 'HSD', 'DSL', 'NGA']
STORAGE_TECHS = ['LDS', 'SDS', 'HPS']
DERIVED_TECH_MULTIPLIERS = {
    'SPR': {'base': 'SPV', 'multiplier': SPR_MULTIPLIER},
    'FPV': {'base': 'SPV', 'multiplier': 1.0},
    'WOF': {'base': 'WON', 'multiplier': 1.0},
}
CSP_CF = {s: v for s, v in zip(SEASONS_LIST, [0.21, 0.25, 0.22, 0.20])}
TIDAL_CF = 0.25

# ----------------------------------------------------------------------
# RES classification for CF handling.
#
# Only these tech_type values get a "real" capacity factor from
# dispatch / ninja / workbook data — because their availability is
# dictated by nature (resource-limited). Everything else (thermal,
# imports, storage) is treated as fully dispatchable with cf_default=1.0;
# the model then picks activity based on cost, capacity, and constraints.
# ----------------------------------------------------------------------
RES_TECH_TYPES = {
    'Solar', 'SPV', 'SPR', 'FPV', 'CSP',
    'Wind',  'WON', 'WOF',
    'Hydro', 'HYD', 'HDR', 'HRO', 'HPS',  # HPS is pumped storage — treated as hydro for CF purposes
    'Geothermal', 'GEO',
    'Biomass', 'BIO', 'BMS',
    'Tidal', 'TID',
}

def is_res_tech(tech_type=None, tech_code=None):
    """True if the tech's CF is availability-limited (needs real data)."""
    if tech_type and tech_type in RES_TECH_TYPES:
        return True
    if tech_code:
        # 3-letter fuel token in positions 3-6 of PWR{FUEL}{REG}
        token = tech_code[3:6] if len(tech_code) >= 6 else ''
        if token in RES_TECH_TYPES:
            return True
    return False

def finalize_cf_row(row, verbose=False):
    """
    Apply the two provenance rules in-place on a CF row dict:

      1. Non-RES techs: force cf_default = 1.0 (dispatchable; model decides
         activity from cost + capacity, not availability).
      2. RES techs: if cf_dispatch is out of [0, 1] and cf_ninja is
         available, annotate preference for ninja. If ninja is missing,
         clip dispatch to [0, 1] and note the clip in source_notes.

    Returns the modified row.
    """
    tt = row.get('tech_type'); tc = row.get('tech_code')
    is_res = is_res_tech(tt, tc)
    ts = row.get('timeslice', '?')
    notes = row.get('source_notes') or ''

    if not is_res:
        # Dispatchable — overwrite cf_default to 1.0 regardless of what was set
        row['cf_default'] = 1.0
        # Wipe dispatch/ninja for dispatchable techs — they're meaningless
        # as availability factors and will only confuse the selector.
        # (cf_da_workbook kept in case a user intentionally sets it.)
        row['cf_dispatch'] = None
        row['cf_ninja'] = None
        if 'CF=1.0 (dispatchable)' not in notes:
            row['source_notes'] = (notes + '; ' if notes else '') + 'CF=1.0 (dispatchable)'
        return row

    # RES tech: policing
    cd = row.get('cf_dispatch'); cn = row.get('cf_ninja')
    try:
        cd_val = float(cd) if cd is not None else None
    except (ValueError, TypeError):
        cd_val = None

    if cd_val is not None and (cd_val > 1.0 or cd_val < 0.0):
        if cn is not None:
            # Prefer ninja; leave dispatch in place for provenance but flag it
            row['source_notes'] = (notes + '; ' if notes else '') + \
                f'dispatch {cd_val:.3f} out of [0,1] at {ts} — use ninja'
            if verbose:
                print(f"    NOTE: {tc} {ts} dispatch={cd_val:.3f} out of range — selector will prefer ninja")
        else:
            # Clip as last resort
            clipped = max(0.0, min(1.0, cd_val))
            row['cf_dispatch'] = round(clipped, 6)
            row['source_notes'] = (notes + '; ' if notes else '') + \
                f'dispatch clipped from {cd_val:.3f} to {clipped:.3f} at {ts}'
            if verbose:
                print(f"    NOTE: {tc} {ts} dispatch clipped {cd_val:.3f} -> {clipped:.3f}")

    return row

# Legacy 3-daypart definition (for workbooks built before the 4dp switch)
LEGACY_3DP_DEF = [('D1', 0, 6), ('D2', 6, 18), ('D3', 18, 24)]


# ======================================================================
# SHARED UTILITIES
# ======================================================================

def hour_to_daypart(hour):
    return _HOUR_LUT[int(hour) % 24]

def get_season(month):
    return MONTH_TO_SEASON.get(month)

def get_timeslice(month, hour):
    return f"{MONTH_TO_SEASON[month]}{hour_to_daypart(hour)}"

def get_config_snapshot():
    return {
        'seasons': [{'code': s[0], 'label': s[1], 'months': s[2], 'days': s[3]}
                    for s in SEASON_DEF],
        'dayparts': [{'code': d[0], 'label': d[1], 'start_hour': d[2], 'end_hour': d[3]}
                     for d in DAYPART_DEF],
        'n_timeslices': N_TIMESLICES, 'timeslices': TIMESLICES,
    }


def remap_daypart_fractions(old_daypart_def, old_fractions, season):
    """
    Remap fractions/CFs from an OLD daypart definition to the CURRENT one
    using proportional hour overlap.

    old_daypart_def : list of (code, start_hour, end_hour)
    old_fractions   : dict {old_dp_code: value}
    season          : str, e.g. 'S1'

    Returns dict {new_timeslice: remapped_value}
    """
    result = {}
    for new_code, new_start, new_end in [(d[0], d[2], d[3]) for d in DAYPART_DEF]:
        value = 0.0
        for old_code, old_start, old_end in old_daypart_def:
            overlap = max(0, min(new_end, old_end) - max(new_start, old_start))
            if overlap > 0:
                old_hours = old_end - old_start
                value += old_fractions.get(old_code, 0) * (overlap / old_hours)
        result[season + new_code] = value
    return result


def validate_demand_fractions(fracs, country, tol=0.001):
    vals = [fracs.get(ts, 0) for ts in TIMESLICES]
    total = sum(vals)
    if abs(total - 1.0) > tol:
        print(f"\n  *** FAIL: {country} demand fracs sum = {total:.6f}")
        print(f"      {[round(v, 6) for v in vals]}")
        return False
    return True

def validate_capacity_factors(cf_rows, country):
    # Pass 1: apply provenance rules (non-RES -> cf_default=1.0; RES -> clip/prefer-ninja)
    for row in cf_rows:
        finalize_cf_row(row, verbose=False)

    # Pass 2: validate what remains
    ok = True
    for row in cf_rows:
        ts, tech = row['timeslice'], row.get('tech_code', '?')
        for col in ['cf_dispatch', 'cf_ninja', 'cf_da_workbook', 'cf_default']:
            val = row.get(col)
            if val is None or val == '':
                continue
            try:
                v = float(val)
            except (ValueError, TypeError):
                continue
            if v > 1.0:
                print(f"  *** CF>1: {country} {tech} {ts} {col}={v:.4f}")
                ok = False
            if v < 0:
                print(f"  *** CF<0: {country} {tech} {ts} {col}={v:.4f}")
                ok = False
        if ts.endswith(DAYPARTS_LIST[0]) and 'SPV' in tech:
            for col in ['cf_dispatch', 'cf_ninja']:
                val = row.get(col)
                if val not in (None, ''):
                    try:
                        if float(val) > 0.01:
                            print(f"  ** Night solar: {country} {tech} {ts} {col}={float(val):.4f}")
                    except (ValueError, TypeError):
                        pass
    return ok

def validate_hydro_cf(cf_rows, country):
    for row in cf_rows:
        tech = row.get('tech_code', '')
        if not any(x in tech for x in ['HYD', 'HDR', 'HRO']):
            continue
        for col in ['cf_dispatch', 'cf_da_workbook']:
            val = row.get(col)
            if val in (None, ''):
                continue
            try:
                v = float(val)
                if 0 < v < 0.15:
                    print(f"  ** HYDRO: {country} {tech} {row['timeslice']} "
                          f"{col}={v:.4f} -> consider UpperLimit over MaxCapacity")
            except (ValueError, TypeError):
                pass

def print_coverage(season_days, country, label=""):
    print(f"\n  {country} {label} -- Seasonal coverage:")
    for s in SEASONS_LIST:
        n = season_days.get(s, 0)
        exp = SEASON_DAYS_365[s]
        pct = (n / exp * 100) if exp > 0 else 0
        tag = "OK" if pct >= 50 else "LOW" if n > 0 else "MISSING"
        print(f"    {s} ({SEASON_NAMES[s]}): {n}/{exp} ({pct:.0f}%) [{tag}]")

def ts_now():
    return datetime.now().strftime("%H:%M:%S")


# ======================================================================
# NINJA CF LOADER
# ======================================================================

def load_ninja_cfs(ninja_ts_file, ninja_config_file=None):
    if not os.path.exists(ninja_ts_file):
        print(f"\n  WARNING: Ninja file not found: {ninja_ts_file}")
        return {}
    df = pd.read_csv(ninja_ts_file)
    avg = df.groupby(['resource', 'region', 'timeslice'])['cf_mean'].mean().reset_index()

    # Check config alignment
    ninja_dp_def = None
    if ninja_config_file and os.path.exists(ninja_config_file):
        with open(ninja_config_file) as f:
            ncfg = json.load(f)
        ninja_dp_def = [(d['code'], d['start_hour'], d['end_hour'])
                        for d in ncfg.get('dayparts', [])]
        current_dp = [(d[0], d[2], d[3]) for d in DAYPART_DEF]
        match = (ninja_dp_def == current_dp)
    else:
        ts_set = set(avg['timeslice'].unique())
        match = ts_set.issubset(set(TIMESLICES))
        if not match:
            ninja_dp_def = LEGACY_3DP_DEF
            print(f"  NOTE: No timeslice_config.json; assuming legacy 3dp Ninja output.")

    if match:
        ninja = {(r['resource'], r['region'], r['timeslice']): r['cf_mean']
                 for _, r in avg.iterrows()}
        print(f"  Ninja CFs loaded: {len(ninja)} entries (configs match)")
        return ninja

    # Remap
    print(f"  Ninja CFs need remapping ({[d[0] for d in ninja_dp_def]} -> {DAYPARTS_LIST})")
    season_buckets = {}
    for _, row in avg.iterrows():
        key = (row['resource'], row['region'], row['timeslice'][:2])
        season_buckets.setdefault(key, {})[row['timeslice'][2:]] = row['cf_mean']
    ninja = {}
    for (res, reg, season), old_fracs in season_buckets.items():
        remapped = remap_daypart_fractions(ninja_dp_def, old_fracs, season)
        for ts, val in remapped.items():
            ninja[(res, reg, ts)] = val
    print(f"  Ninja CFs remapped: {len(ninja)} entries")
    return ninja

def get_ninja_cf(ninja_cfs, resource, region, timeslice):
    val = ninja_cfs.get((resource, region, timeslice))
    return round(val, 6) if val is not None else None


# ======================================================================
# BGD ADAPTER
# ======================================================================

BGD_VALID_FUELS = {'Gas', 'Gas/HSD', 'HSD/Gas', 'HFO', 'HFO/Gas', 'HSD',
                   'Coal', 'Hydro', 'Solar', 'Wind', 'Import'}
BGD_FUEL_TO_TECH = {
    'Gas': 'Gas', 'Gas/HSD': 'Gas', 'HSD/Gas': 'Gas',
    'HFO': 'Oil_HFO', 'HFO/Gas': 'Oil_HFO', 'HSD': 'Oil_HSD',
    'Coal': 'Coal', 'Hydro': 'Hydro', 'Solar': 'Solar',
    'Wind': 'Wind', 'Import': 'Import',
}
BGD_TECHS = ['Gas', 'Coal', 'Oil_HFO', 'Oil_HSD', 'Hydro', 'Solar', 'Wind', 'Import']
BGD_TECH_TO_OSTRAM = {
    'Gas': 'PWRGASBGDXX', 'Coal': 'PWRCOABGDXX',
    'Oil_HFO': 'PWRHFOBGDXX', 'Oil_HSD': 'PWRHSDBGDXX',
    'Hydro': 'PWRHYDBGDXX', 'Solar': 'PWRSPVBGDXX',
    'Wind': 'PWRWONBGDXX', 'Import': 'PWRIMPBGDXX',
}

def bgd_parse_hour(t):
    if isinstance(t, dtime):
        return t.hour + (0.5 if t.minute == 30 else 0)
    if isinstance(t, str):
        parts = t.split(':')
        try:
            return int(parts[0]) + (0.5 if len(parts) > 1 and parts[1] == '30' else 0)
        except Exception:
            return None
    return None

def bgd_detect_format(fp):
    try:
        ec = pd.read_excel(fp, sheet_name='En-Curve', header=None)
        for i in range(min(10, ec.shape[0])):
            if str(ec.iloc[i, 0]).strip() == 'TIME':
                return i
    except Exception:
        pass
    return None

def bgd_extract_date(fp, hrow):
    try:
        ec = pd.read_excel(fp, sheet_name='En-Curve', header=None)
        for cr in [hrow-2, hrow-1, 0, 1, 2, 3]:
            if 0 <= cr < ec.shape[0]:
                cell = ec.iloc[cr, 1]
                if isinstance(cell, datetime): return cell
                if pd.notna(cell):
                    try: return pd.to_datetime(str(cell), dayfirst=True)
                    except: pass
    except: pass
    try:
        lc = pd.read_excel(fp, sheet_name='L-Curve', header=None)
        cell = lc.iloc[0, 1]
        if isinstance(cell, datetime): return cell
        if pd.notna(cell):
            try: return pd.to_datetime(str(cell), dayfirst=True)
            except: pass
    except: pass
    m = re.search(r'\d{4}-\d{2}-\d{2}', os.path.basename(fp))
    return pd.to_datetime(m.group()) if m else None

def bgd_parse_en_curve(fp, hrow):
    ec = pd.read_excel(fp, sheet_name='En-Curve', header=None)
    raw = [str(ec.iloc[hrow, c]).strip() if pd.notna(ec.iloc[hrow, c]) else ''
           for c in range(min(20, ec.shape[1]))]
    cm = {}
    for idx, h in enumerate(raw):
        hu = h.upper().replace('.','').replace(' ','')
        if hu == 'TIME': cm['TIME'] = idx
        elif hu in ('GAS','GAS-PUBLIC','GASPUBLIC'): cm['Gas_Public'] = idx
        elif hu in ('PGEN(GAS)','GAS-PVT','GASPVT','PVTGEN(GAS)'): cm['Gas_Pvt'] = idx
        elif hu == 'HVDC': cm['HVDC'] = idx
        elif hu == 'NEPAL': cm['Nepal'] = idx
        elif hu == 'TRIPURA': cm['Tripura'] = idx
        elif hu == 'ADANI': cm['Adani'] = idx
        elif hu == 'COAL': cm['Coal'] = idx
        elif hu == 'HYDRO': cm['Hydro'] = idx
        elif hu == 'SOLAR': cm['Solar'] = idx
        elif hu == 'WIND': cm['Wind'] = idx
        elif hu in ('PGEN(HFO)','HFO-PVT','HFOPVT','PVTGEN(HFO)'): cm['HFO_Pvt'] = idx
        elif hu in ('GGEN(HFO)','HFO-PUBLIC','HFOPUBLIC','GOVTGEN(HFO)'): cm['HFO_Public'] = idx
        elif hu in ('PGEN(HSD)','HSD-PVT','HSDPVT','PVTGEN(HSD)'): cm['HSD_Pvt'] = idx
        elif hu in ('GGEN(HSD)','HSD-PUBLIC','HSDPUBLIC','GOVTGEN(HSD)'): cm['HSD_Public'] = idx
    rows = []
    for i in range(hrow+1, min(hrow+50, ec.shape[0])):
        hour = bgd_parse_hour(ec.iloc[i, cm.get('TIME', 0)])
        if hour is None: continue
        row = {'hour': hour, 'daypart': hour_to_daypart(int(hour))}
        for key, ci in cm.items():
            if key == 'TIME': continue
            val = ec.iloc[i, ci]
            row[key] = float(val) if pd.notna(val) and isinstance(val, (int, float, np.integer, np.floating)) else 0
        rows.append(row)
    df = pd.DataFrame(rows)
    df['Gas'] = df.get('Gas_Public', 0) + df.get('Gas_Pvt', 0)
    df['Coal_tech'] = df.get('Coal', 0)
    df['Oil_HFO'] = df.get('HFO_Public', 0) + df.get('HFO_Pvt', 0)
    df['Oil_HSD'] = df.get('HSD_Public', 0) + df.get('HSD_Pvt', 0)
    df['Hydro_tech'] = df.get('Hydro', 0)
    df['Solar_tech'] = df.get('Solar', 0)
    df['Wind_tech'] = df.get('Wind', 0)
    df['Import'] = df.get('HVDC',0)+df.get('Nepal',0)+df.get('Tripura',0)+df.get('Adani',0)
    # Drop raw columns that conflict with the renamed ones
    for c in ['Coal','Hydro','Solar','Wind']:
        if c in df.columns and c+'_tech' in df.columns:
            df.drop(columns=[c], inplace=True)
    df.rename(columns={'Coal_tech':'Coal','Hydro_tech':'Hydro',
                        'Solar_tech':'Solar','Wind_tech':'Wind'}, inplace=True)
    return df

def bgd_parse_load_curve(fp):
    lc = pd.read_excel(fp, sheet_name='L-Curve', header=None)
    tc, hr = None, None
    for i in range(min(5,lc.shape[0])):
        for j in range(min(10,lc.shape[1])):
            if str(lc.iloc[i,j]).strip().upper() == 'TOTAL':
                tc, hr = j, i; break
        if tc: break
    if tc is None: tc, hr = 3, 2
    rows = []
    for i in range(hr+1, min(hr+50, lc.shape[0])):
        hour = bgd_parse_hour(lc.iloc[i, 0])
        if hour is None: continue
        total = lc.iloc[i, tc]
        if pd.notna(total):
            try: rows.append({'hour':hour,'daypart':hour_to_daypart(int(hour)),'Total':float(total)})
            except: pass
    return pd.DataFrame(rows)

def bgd_parse_installed_capacity(fp, hrow_en):
    try: fc = pd.read_excel(fp, sheet_name='Forecast', header=None)
    except: return {t:0 for t in BGD_TECHS}
    fuel_col, cap_col = None, None
    for i in range(min(10,fc.shape[0])):
        for j in range(min(12,fc.shape[1])):
            val = str(fc.iloc[i,j]).strip() if pd.notna(fc.iloc[i,j]) else ''
            if val == 'Fuel': fuel_col = j
            if 'Present' in val and 'Capacity' in val.replace('\n',' '): cap_col = j
            if val.strip() == 'Capacity' and cap_col is None and i > 0:
                if 'Present' in str(fc.iloc[i-1,j]).strip(): cap_col = j
    if fuel_col is None: fuel_col = 4 if hrow_en==4 else 2
    if cap_col is None: cap_col = 8 if hrow_en==4 else 6
    tc = {t:0 for t in BGD_TECHS}
    ds = 8 if hrow_en==4 else 10
    for i in range(ds, min(fc.shape[0], 260)):
        fuel = fc.iloc[i,fuel_col] if fuel_col<fc.shape[1] else None
        cap = fc.iloc[i,cap_col] if cap_col<fc.shape[1] else None
        if pd.notna(fuel) and pd.notna(cap):
            fs = str(fuel).strip()
            if fs in BGD_VALID_FUELS:
                tech = BGD_FUEL_TO_TECH.get(fs)
                if tech:
                    try: tc[tech] += float(cap)
                    except: pass
    return tc

def bgd_process_single_file_to_rows(fp):
    """Parse one PGCB file, return list of half-hourly row dicts (daypart-agnostic)."""
    hrow = bgd_detect_format(fp)
    if hrow is None: return None, None
    dt = bgd_extract_date(fp, hrow)
    if dt is None: return None, None
    try:
        df_f = bgd_parse_en_curve(fp, hrow)
        df_l = bgd_parse_load_curve(fp)
        ic = bgd_parse_installed_capacity(fp, hrow)
    except Exception as e:
        print(f"    ERROR {os.path.basename(fp)}: {e}"); return None, None
    if len(df_f)==0 or len(df_l)==0: return None, None
    # Merge generation + demand on hour
    df_f['date'] = dt.strftime('%Y-%m-%d')
    df_f['month'] = dt.month
    df_f['season'] = get_season(dt.month)
    # Join demand
    demand_by_hour = df_l.groupby('hour')['Total'].mean()
    df_f['demand_MW'] = df_f['hour'].map(demand_by_hour).fillna(0)
    keep = ['date','month','season','hour'] + BGD_TECHS + ['demand_MW']
    rows = df_f[[c for c in keep if c in df_f.columns]].to_dict('records')
    return rows, ic

# Path for the compiled half-hourly CSV (lives next to the PGCB files)
BGD_COMPILED_CSV = os.path.join(BGD_PGCB_DIR, '..', 'bgd_halfhourly_compiled.csv')
BGD_COMPILED_IC  = os.path.join(BGD_PGCB_DIR, '..', 'bgd_installed_capacity.csv')

def bgd_compile_all_files():
    """Parse all PGCB Excel files and save a compiled half-hourly CSV.
    This is the SLOW step (~20-40 min). Only runs once; subsequent runs read the CSV."""
    files = sorted(glob.glob(os.path.join(BGD_PGCB_DIR, '*.xlsx')))
    if not files:
        print(f"  ERROR: No files in {BGD_PGCB_DIR}"); return None, None
    print(f"  Compiling {len(files)} PGCB files (this is the slow step — only needed once)...")
    all_rows = []
    ic_rows = []
    ok, err = 0, 0
    t_start = time.perf_counter()
    for i, f in enumerate(files):
        rows, ic = bgd_process_single_file_to_rows(f)
        if rows:
            all_rows.extend(rows)
            ic_rows.append(ic)
            ok += 1
        else:
            err += 1
        if (i+1) % 100 == 0 or (i+1) == len(files):
            elapsed = time.perf_counter() - t_start
            rate = (i+1) / elapsed if elapsed > 0 else 0
            eta = (len(files) - i - 1) / rate if rate > 0 else 0
            print(f"    {i+1}/{len(files)}  ({elapsed:.0f}s elapsed, ~{eta:.0f}s remaining)")
    if not all_rows:
        print("  ERROR: No data extracted"); return None, None
    print(f"  Parsed: {ok} files OK, {err} errors, {len(all_rows)} half-hourly rows")
    # Save compiled CSV
    df = pd.DataFrame(all_rows)
    csv_path = os.path.normpath(BGD_COMPILED_CSV)
    df.to_csv(csv_path, index=False)
    print(f"  Saved compiled CSV: {csv_path}  ({len(df)} rows)")
    # Save average installed capacity
    ic_df = pd.DataFrame(ic_rows)
    ic_avg = ic_df.mean().to_dict()
    ic_path = os.path.normpath(BGD_COMPILED_IC)
    pd.DataFrame([ic_avg]).to_csv(ic_path, index=False)
    print(f"  Saved installed capacity: {ic_path}")
    return df, ic_avg

def run_bgd_adapter(ninja_cfs):
    print(f"\n{'='*70}\n[{ts_now()}] BGD -- Bangladesh\n{'='*70}")
    csv_path = os.path.normpath(BGD_COMPILED_CSV)
    ic_path = os.path.normpath(BGD_COMPILED_IC)

    # --- Fast path: read compiled CSV ---
    if os.path.exists(csv_path) and os.path.exists(ic_path):
        print(f"  FAST PATH: Reading compiled CSV")
        print(f"    {csv_path}")
        df = pd.read_csv(csv_path)
        ic_df = pd.read_csv(ic_path)
        ac = ic_df.iloc[0].to_dict()
        nf = df['date'].nunique()
        print(f"    {len(df)} rows, {nf} unique dates")
        print(f"    To re-parse from Excel, delete: {csv_path}")
    else:
        # --- Slow path: parse all Excel files and compile ---
        result = bgd_compile_all_files()
        if result[0] is None: return None, None, None
        df, ac = result
        nf = df['date'].nunique()

    # --- Aggregate to current DAYPART_DEF (works for any config) ---
    df['hour_int'] = df['hour'].astype(int)
    df['daypart'] = df['hour_int'].apply(hour_to_daypart)
    df['timeslice'] = df['season'] + df['daypart']

    # Demand fractions: sum demand energy per timeslice
    ts_demand = df.groupby('timeslice')['demand_MW'].sum()
    tde = ts_demand.sum()
    dem = {ts: round(ts_demand.get(ts, 0) / tde, 6) if tde > 0 else 0 for ts in TIMESLICES}

    # Generation CFs: mean generation per timeslice / installed capacity
    ts_gen = df.groupby('timeslice')[BGD_TECHS].mean()
    cfs = []
    for ts in TIMESLICES:
        s = ts[:2]; dp = ts[2:]
        for tech in BGD_TECHS:
            oc = BGD_TECH_TO_OSTRAM[tech]
            ag = ts_gen.loc[ts, tech] if ts in ts_gen.index else 0
            cap = ac.get(tech, 0)
            cd = round(ag / cap, 6) if cap > 0 else None
            cn = None
            if tech == 'Solar': cn = get_ninja_cf(ninja_cfs, 'solar', 'BGDXX', ts)
            elif tech == 'Wind': cn = get_ninja_cf(ninja_cfs, 'wind', 'BGDXX', ts)
            cf_def = 1.0 if tech in ['Gas', 'Coal', 'Oil_HFO', 'Oil_HSD'] else None
            cfs.append({'timeslice': ts, 'season': SEASON_NAMES[s], 'daypart': DAYPART_NAMES[dp],
                        'tech_code': oc, 'tech_type': tech, 'cf_dispatch': cd, 'cf_ninja': cn,
                        'cf_da_workbook': None, 'cf_default': cf_def,
                        'source_notes': f"dispatch: {nf} PGCB files (compiled CSV)"})
    sd = df.groupby('season')['date'].nunique().to_dict()
    print_coverage(sd, 'BGD', 'PGCB')
    dates = pd.to_datetime(df['date'])
    meta = {'country': 'BGD', 'n_files': nf,
            'date_range': [dates.min().strftime('%Y-%m-%d'), dates.max().strftime('%Y-%m-%d')],
            'installed_capacity_MW': {k: round(v, 1) for k, v in ac.items()},
            'source': 'PGCB daily Excel (compiled CSV)'}
    validate_demand_fractions(dem, 'BGD')
    validate_capacity_factors(cfs, 'BGD')
    validate_hydro_cf(cfs, 'BGD')
    return dem, cfs, meta


# ======================================================================
# LKA ADAPTER
# ======================================================================

LKA_IC = {'Coal':900,'Oil_CEB':710,'Oil_IPP':420,'Major_Hydro':1384,
           'Wind':128,'Solar':200,'Biomass':30,'Mini_Hydro':415}
LKA_FUELS = ['Coal','Oil_CEB','Oil_IPP','Major_Hydro','Wind','Solar','Biomass','Mini_Hydro']
LKA_T2O = {'Coal':'PWRCOALKAXX','Oil_CEB':'PWROILKA_CEB','Oil_IPP':'PWROILKA_IPP',
            'Major_Hydro':'PWRHYDLKAXX','Wind':'PWRWONLKAXX','Solar':'PWRSPVLKAXX',
            'Biomass':'PWRBIOLKAXX','Mini_Hydro':'PWRMNHLKAXX'}
LKA_FLP = {
    'Coal':[r'\bcoal\b'],'Oil_CEB':[r'oil\s*[-\u2013]\s*ceb',r'oil\s*ceb'],
    'Oil_IPP':[r'oil\s*[-\u2013]\s*ipp',r'oil\s*ipp'],
    'Major_Hydro':[r'major\s*hydro',r'\bhydro\b(?!.*mini)'],
    'Wind':[r'\bwind\b'],'Solar':[r'\bsolar\b'],
    'Biomass':[r'\bbiomass\b'],'Mini_Hydro':[r'mini\s*hydro',r'small\s*hydro'],
}
LKA_TLP = [r'total\s*mw',r'total\s*\*',r'^total$',r'system\s*total']

def lka_parse_hour(ts):
    ts = str(ts).strip()
    if ts.count(':')==2: p=ts.split(':'); h,m=int(p[0]),int(p[1])
    else: p=ts.split(':'); h,m=int(p[0]),int(p[1]) if len(p)>1 else 0
    if h==0 and m==0 and ':00:00' in ts: return 23
    if m==0 and h>0: return h-1
    return h

def lka_build_col_map(ws):
    cm, ec = {}, None
    for ci in range(2,120):
        v = ws.cell(row=2,column=ci).value
        if v is None: continue
        vs = str(v).strip()
        if 'Energy' in vs or 'MWh' in vs: ec=ci; break
        try: h=lka_parse_hour(vs); cm[ci]=(h,hour_to_daypart(h))
        except: continue
    return cm, ec

def lka_detect_layout(ws, mx=80):
    labels = {}
    for r in range(1,mx+1):
        v = ws.cell(row=r,column=1).value
        if v is not None: labels[r] = str(v).strip()
    tr = None
    for r,t in labels.items():
        for p in LKA_TLP:
            if re.search(p,t,re.IGNORECASE): tr=r; break
        if tr: break
    fr = {}
    ss = (tr+1) if tr else 1
    for fk, pats in LKA_FLP.items():
        for r in range(ss, mx+1):
            if r in labels:
                for p in pats:
                    if re.search(p,labels[r],re.IGNORECASE): fr[fk]=r; break
                if fk in fr: break
    return fr, tr

def lka_process_file(fp):
    wb = openpyxl.load_workbook(fp, data_only=True)
    recs = []
    for sn in wb.sheetnames:
        if '.' not in sn: continue
        try: p=sn.split('.'); sm,sd=int(p[0]),int(p[1])
        except: continue
        ws = wb[sn]
        dv = ws.cell(row=1,column=9).value
        if not isinstance(dv,datetime):
            try: dv=datetime(2023,sm,sd)
            except: continue
        cm,ec = lka_build_col_map(ws)
        if len(cm)<90: continue
        fr,tr = lka_detect_layout(ws)
        if tr is None or len(fr)<4: continue
        rec = {'date':dv,'month':dv.month,'day':dv.day,'fuel_data':{},'total_mw':{},'col_map':cm,'energy_col':ec}
        for fn,ri in fr.items():
            mv = {ci:float(ws.cell(row=ri,column=ci).value) if isinstance(ws.cell(row=ri,column=ci).value,(int,float)) else 0.0 for ci in cm}
            em = 0.0
            if ec:
                ev = ws.cell(row=ri,column=ec).value
                if isinstance(ev,(int,float)): em=float(ev)
            rec['fuel_data'][fn] = {'mw':mv,'e':em}
        tmw = {ci:float(ws.cell(row=tr,column=ci).value) if isinstance(ws.cell(row=tr,column=ci).value,(int,float)) else 0.0 for ci in cm}
        te = 0.0
        if ec:
            ev = ws.cell(row=tr,column=ec).value
            if isinstance(ev,(int,float)): te=float(ev)
        rec['total_mw'] = {'mw':tmw,'e':te}
        recs.append(rec)
    wb.close()
    return recs

LKA_COMPILED_CSV = os.path.join(BASE_DIR, r"Sri Lanka\lka_hourly_compiled.csv")

def run_lka_adapter(ninja_cfs):
    print(f"\n{'='*70}\n[{ts_now()}] LKA -- Sri Lanka\n{'='*70}")
    csv_path = os.path.normpath(LKA_COMPILED_CSV)

    # --- Fast path: compiled CSV from compile_lka_pucsl.py ---
    if os.path.exists(csv_path):
        print(f"  FAST PATH: Reading compiled CSV")
        print(f"    {csv_path}")
        df = pd.read_csv(csv_path)
        n_dates = df['date'].nunique()
        print(f"    {len(df)} rows, {n_dates} unique dates")
        print(f"    To re-parse from Excel, delete: {csv_path}")

        # Map hour -> daypart using current DAYPART_DEF
        df['daypart'] = df['hour'].astype(int).apply(hour_to_daypart)
        df['timeslice'] = df['season'] + df['daypart']

        # Demand fractions
        ts_demand = df.groupby('timeslice')['demand_MW'].sum()
        tde = ts_demand.sum()
        dem = {ts: round(ts_demand.get(ts, 0) / tde, 6) if tde > 0 else 0 for ts in TIMESLICES}

        # CFs from dispatch: mean generation / installed capacity
        ts_gen = df.groupby('timeslice')[LKA_FUELS].mean()
        cfs = []
        for ts in TIMESLICES:
            s, dp = ts[:2], ts[2:]
            for fuel in LKA_FUELS:
                oc = LKA_T2O[fuel]
                ag = ts_gen.loc[ts, fuel] if ts in ts_gen.index else 0
                ic = LKA_IC.get(fuel, 0)
                cd = round(ag / ic, 6) if ic > 0 else None
                cn = None
                if fuel == 'Solar': cn = get_ninja_cf(ninja_cfs, 'solar', 'LKAXX', ts)
                elif fuel == 'Wind': cn = get_ninja_cf(ninja_cfs, 'wind', 'LKAXX', ts)
                cf_def = 1.0 if fuel in ['Coal', 'Oil_CEB', 'Oil_IPP'] else None
                cfs.append({'timeslice': ts, 'season': SEASON_NAMES[s], 'daypart': DAYPART_NAMES[dp],
                            'tech_code': oc, 'tech_type': fuel, 'cf_dispatch': cd, 'cf_ninja': cn,
                            'cf_da_workbook': None, 'cf_default': cf_def,
                            'source_notes': f"dispatch: {n_dates} days (compiled CSV)"})

        sd = df.groupby('season')['date'].nunique().to_dict()
        print_coverage(sd, 'LKA', 'PUCSL')
        meta = {'country': 'LKA', 'n_dates': n_dates,
                'installed_capacity_MW': LKA_IC, 'source': 'PUCSL 15-min (compiled CSV)'}
        validate_demand_fractions(dem, 'LKA')
        validate_capacity_factors(cfs, 'LKA')
        validate_hydro_cf(cfs, 'LKA')
        return dem, cfs, meta

    # --- Slow path: parse openpyxl (fallback if no compiled CSV) ---
    print(f"  No compiled CSV found at: {csv_path}")
    print(f"  Run compile_lka_pucsl.py first for fast loading.")
    print(f"  Falling back to openpyxl parsing (slow)...")
    pf = sorted([f for f in os.listdir(LKA_PUCSL_DIR) if f.endswith('.xlsx') and '15-Minutes' in f])
    if not pf: print(f"  ERROR: No files in {LKA_PUCSL_DIR}"); return None,None,None
    print(f"  Found {len(pf)} monthly files")
    ar = []
    for fn in pf:
        recs = lka_process_file(os.path.join(LKA_PUCSL_DIR,fn))
        ar.extend(recs); print(f"    {fn}: {len(recs)} sheets")
    if not ar: print("  ERROR: No records"); return None,None,None
    print(f"  Total records: {len(ar)}")
    cmt = ar[0]['col_map']
    tf = defaultdict(lambda: defaultdict(lambda: {'ms':0.0,'ni':0,'em':0.0,'nd':0}))
    td = defaultdict(lambda: {'ms':0.0,'ni':0,'em':0.0,'nd':0})
    for rec in ar:
        s = MONTH_TO_SEASON[rec['month']]
        dpc = defaultdict(list)
        for ci in cmt: h,dp=cmt[ci]; dpc[dp].append(ci)
        for fn,fi in rec['fuel_data'].items():
            mp, de = fi['mw'], fi['e']
            dms, tdm = {}, 0.0
            for dp,cols in dpc.items():
                dm = sum(mp.get(c,0.0) for c in cols); dms[dp]=dm; tdm+=dm
            for dp,cols in dpc.items():
                ts = f"{s}{dp}"; dme=dms[dp]
                dpe = de*(dme/tdm) if tdm>0 else 0
                tf[fn][ts]['ms']+=dme; tf[fn][ts]['ni']+=len(cols); tf[fn][ts]['em']+=dpe
            for dp in dpc: tf[fn][f"{s}{dp}"]['nd']+=1
        tmw, te = rec['total_mw']['mw'], rec['total_mw']['e']
        tdmw = sum(tmw.get(c,0.0) for c in tmw)
        for dp,cols in dpc.items():
            ts = f"{s}{dp}"; dm=sum(tmw.get(c,0.0) for c in cols)
            dpe = te*(dm/tdmw) if tdmw>0 else 0
            td[ts]['ms']+=dm; td[ts]['ni']+=len(cols); td[ts]['em']+=dpe; td[ts]['nd']+=1
    tde = sum(td[ts]['em'] for ts in TIMESLICES)
    dem = {ts:round(td[ts]['em']/tde,6) if tde>0 else 0 for ts in TIMESLICES}
    cfs = []
    for ts in TIMESLICES:
        s,dp = ts[:2],ts[2:]
        for fuel in LKA_FUELS:
            oc = LKA_T2O[fuel]
            am = tf[fuel][ts]['ms']/tf[fuel][ts]['ni'] if tf[fuel][ts]['ni']>0 else 0
            ic = LKA_IC.get(fuel,0)
            cd = round(am/ic,6) if ic>0 else None
            cn = None
            if fuel=='Solar': cn=get_ninja_cf(ninja_cfs,'solar','LKAXX',ts)
            elif fuel=='Wind': cn=get_ninja_cf(ninja_cfs,'wind','LKAXX',ts)
            cf_def = 1.0 if fuel in ['Coal','Oil_CEB','Oil_IPP'] else None
            cfs.append({'timeslice':ts,'season':SEASON_NAMES[s],'daypart':DAYPART_NAMES[dp],
                        'tech_code':oc,'tech_type':fuel,'cf_dispatch':cd,'cf_ninja':cn,
                        'cf_da_workbook':None,'cf_default':cf_def,
                        'source_notes':f"dispatch: {len(ar)} PUCSL sheets"})
    sd = defaultdict(int)
    for r in ar: sd[MONTH_TO_SEASON[r['month']]] += 1
    print_coverage(sd,'LKA','PUCSL')
    meta = {'country':'LKA','n_records':len(ar),'n_files':len(pf),
            'installed_capacity_MW':LKA_IC,'source':'PUCSL 15-min monthly'}
    validate_demand_fractions(dem,'LKA'); validate_capacity_factors(cfs,'LKA'); validate_hydro_cf(cfs,'LKA')
    return dem, cfs, meta


# ======================================================================
# IND ADAPTER (5 sub-regions)
# ======================================================================

IND_REGIONS = ['INDNO','INDEA','INDNE','INDSO','INDWE']

# Map Grid-India region strings to OSTRAM codes (handles typos in source)
IND_REGION_MAP = {
    'eastern': 'INDEA', 'north-eastern': 'INDNE', 'northeastern': 'INDNE',
    'northern': 'INDNO', 'northen': 'INDNO',  # typo in 2023 file
    'southern': 'INDSO', 'western': 'INDWE',
}

_MONTH_ABBR = {'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,
               'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12}

def ind_parse_date_hour(date_str):
    """Parse Grid-India '01-Jan 12am' -> (month, day, hour_24)."""
    m = re.match(r'(\d{1,2})-(\w{3})\s+(\d{1,2})(am|pm)', str(date_str).strip())
    if not m:
        return None, None, None
    day = int(m.group(1))
    month = _MONTH_ABBR.get(m.group(2))
    h12 = int(m.group(3))
    ampm = m.group(4)
    if ampm == 'am':
        hour = 0 if h12 == 12 else h12
    else:
        hour = 12 if h12 == 12 else h12 + 12
    return month, day, hour

def ind_parse_region(region_str):
    """Parse 'Eastern Region - 2023' -> ('INDEA', 2023)."""
    s = str(region_str).strip()
    # Extract year from end
    ym = re.search(r'(\d{4})$', s)
    year = int(ym.group(1)) if ym else None
    # Extract region name (everything before 'Region')
    rm = re.match(r'(.+?)\s*region', s, re.IGNORECASE)
    if rm:
        rname = rm.group(1).strip().lower().replace(' ', '-')
        # Also try without hyphen
        code = IND_REGION_MAP.get(rname) or IND_REGION_MAP.get(rname.replace('-', ''))
        return code, year
    return None, year

def run_ind_adapter(ninja_cfs):
    print(f"\n{'='*70}\n[{ts_now()}] IND -- India (5 regions)\n{'='*70}")
    dfs = sorted(glob.glob(os.path.join(IND_DEMAND_DIR,'year_demand_*.xlsx')))
    if not dfs: print(f"  ERROR: No files in {IND_DEMAND_DIR}"); return {},{},{}
    print(f"  Found {len(dfs)} demand files")

    # Read and parse all files
    all_rows = []
    for fp in dfs:
        fname = os.path.basename(fp)
        print(f"    Reading: {fname}")
        try:
            df = pd.read_excel(fp)
            n_ok, n_err = 0, 0
            for _, row in df.iterrows():
                region_code, year = ind_parse_region(row['Region'])
                month, day, hour = ind_parse_date_hour(row['Date'])
                if region_code and month and hour is not None:
                    demand = row.get('Hourly Demand Met (in MW)', 0)
                    if pd.notna(demand):
                        all_rows.append({
                            'region': region_code, 'year': year,
                            'month': month, 'day': day, 'hour': hour,
                            'season': MONTH_TO_SEASON[month],
                            'daypart': hour_to_daypart(hour),
                            'demand_mw': float(demand),
                            'date_str': f"{year}-{month:02d}-{day:02d}",
                        })
                        n_ok += 1
                    else:
                        n_err += 1
                else:
                    n_err += 1
            print(f"      {n_ok} rows OK, {n_err} errors")
        except Exception as e:
            print(f"      ERROR: {e}")

    if not all_rows:
        print("  ERROR: No demand data loaded"); return {},{},{}

    da = pd.DataFrame(all_rows)
    da['timeslice'] = da['season'] + da['daypart']
    print(f"  Total parsed: {len(da)} rows")
    print(f"  Regions found: {sorted(da['region'].unique())}")
    print(f"  Years: {sorted(da['year'].unique())}")

    # Hydro CFs from CEA
    hcf = {}
    if os.path.exists(IND_HYDRO_CF_FILE):
        try:
            dh = pd.read_excel(IND_HYDRO_CF_FILE, sheet_name=IND_HYDRO_CF_SHEET)
            CEA_TO_OSTRAM = {'IND-E':'INDEA','IND-N':'INDNO','IND-NE':'INDNE',
                             'IND-S':'INDSO','IND-W':'INDWE'}
            for _,row in dh.iterrows():
                reg = CEA_TO_OSTRAM.get(str(row.iloc[0]).strip(),
                                        str(row.iloc[0]).strip())
                for s in SEASONS_LIST:
                    if s in dh.columns and pd.notna(row[s]): hcf[(reg,s)]=float(row[s])
            print(f"  Hydro CFs: {len(hcf)} entries")
        except Exception as e: print(f"  ERROR hydro: {e}")
    else:
        print(f"  *** MISSING: {IND_HYDRO_CF_FILE}")

    ad, ac, am = {}, {}, {}
    for reg in IND_REGIONS:
        dr = da[da['region'] == reg].copy()
        if len(dr) == 0:
            print(f"  WARNING: No data for {reg}"); continue
        print(f"\n  {reg}: {len(dr)} rows, {dr['date_str'].nunique()} days")

        tse = dr.groupby('timeslice')['demand_mw'].sum()
        te = tse.sum()
        dem = {ts: round(tse.get(ts, 0) / te, 6) if te > 0 else 0 for ts in TIMESLICES}

        cfs = []
        for ts in TIMESLICES:
            s, dp = ts[:2], ts[2:]
            hv = hcf.get((reg, s))
            for ht in ['HYD','HDR','HRO']:
                cfs.append({'timeslice':ts,'season':SEASON_NAMES[s],'daypart':DAYPART_NAMES[dp],
                    'tech_code':f'PWR{ht}{reg}','tech_type':f'Hydro_{ht}',
                    'cf_dispatch':None,'cf_ninja':None,
                    'cf_da_workbook':round(hv,6) if hv else None,'cf_default':None,
                    'source_notes':'CEA fleet PLF' if hv else 'MISSING'})
            cfs.append({'timeslice':ts,'season':SEASON_NAMES[s],'daypart':DAYPART_NAMES[dp],
                'tech_code':f'PWRSPV{reg}','tech_type':'Solar','cf_dispatch':None,
                'cf_ninja':get_ninja_cf(ninja_cfs,'solar',reg,ts),
                'cf_da_workbook':None,'cf_default':None,'source_notes':'Ninja'})
            cfs.append({'timeslice':ts,'season':SEASON_NAMES[s],'daypart':DAYPART_NAMES[dp],
                'tech_code':f'PWRWON{reg}','tech_type':'Wind','cf_dispatch':None,
                'cf_ninja':get_ninja_cf(ninja_cfs,'wind',reg,ts),
                'cf_da_workbook':None,'cf_default':None,'source_notes':'Ninja'})
            for th in ['COA','GAS','OIL','NGA']:
                cfs.append({'timeslice':ts,'season':SEASON_NAMES[s],'daypart':DAYPART_NAMES[dp],
                    'tech_code':f'PWR{th}{reg}','tech_type':th,'cf_dispatch':None,'cf_ninja':None,
                    'cf_da_workbook':None,'cf_default':1.0,'source_notes':'thermal CF=1.0'})

        sdc = dr.groupby('season')['date_str'].nunique().to_dict()
        print_coverage(sdc, reg, 'Grid-India')
        validate_demand_fractions(dem, reg)
        validate_capacity_factors(cfs, reg)
        validate_hydro_cf(cfs, reg)
        ad[reg]=dem; ac[reg]=cfs
        am[reg]={'country':reg,'source':'Grid-India hourly+CEA hydro+Ninja',
                 'n_rows':len(dr),'years':sorted(dr['year'].unique())}
    return ad, ac, am


# ======================================================================
# NPL ADAPTER — Simkhada 2022 hourly profiles + NDOR seasonal/hydro
# ======================================================================

# NDOR-validated seasonal demand shares (818 days, Oct 2022–Jan 2025)
NPL_SEASONAL_SHARES = {'S1': 0.2248, 'S2': 0.2629, 'S3': 0.3775, 'S4': 0.1348}

# NDOR FY2080 hydro CFs (seasonal average, flat within season)
# Source: Nepal_sources_and_methodology.md Section 5
NPL_HYDRO_CF = {'S1': 0.390, 'S2': 0.358, 'S3': 0.755, 'S4': 0.697}

# Legacy 3-dp fallback (only used if hourly CSV missing)
NPL_DP3 = {
    'S1': {'D1': 0.2042, 'D2': 0.5400, 'D3': 0.2557},
    'S2': {'D1': 0.2247, 'D2': 0.5078, 'D3': 0.2675},
    'S3': {'D1': 0.2484, 'D2': 0.4967, 'D3': 0.2548},
    'S4': {'D1': 0.2402, 'D2': 0.5090, 'D3': 0.2507},
}

def run_npl_adapter(ninja_cfs):
    print(f"\n{'='*70}\n[{ts_now()}] NPL -- Nepal\n{'='*70}")

    # --- Try hourly profiles CSV (Simkhada 2022 literature) ---
    csv_path = os.path.normpath(NPL_PROFILES_CSV)
    if os.path.exists(csv_path):
        print(f"  Reading hourly profiles: {os.path.basename(csv_path)}")
        df = pd.read_csv(csv_path)
        print(f"    {len(df)} rows, seasons {sorted(df['SEASON'].unique())}")

        # Compute demand fractions: seasonal_weight × daypart_share
        dem = {}
        for s in SEASONS_LIST:
            sdf = df[df['SEASON'] == s]
            total_mw = sdf['MW'].sum()
            for dp_code in DAYPARTS_LIST:
                dp_start, dp_end = DAYPART_RANGES[dp_code]
                mask = (sdf['HOUR'] >= dp_start) & (sdf['HOUR'] < dp_end)
                dp_mw = sdf.loc[mask, 'MW'].sum()
                dp_share = dp_mw / total_mw if total_mw > 0 else 0
                dem[s + dp_code] = NPL_SEASONAL_SHARES[s] * dp_share

        total = sum(dem.values())
        dem = {ts: round(dem.get(ts, 0) / total, 6) for ts in TIMESLICES}
        print(f"    Demand fractions: NDOR seasonal x Simkhada hourly ({N_DAYPARTS} dayparts)")
        src_note = 'NDOR 818d seasonal x Simkhada 2022 hourly'
    else:
        # --- Fallback: legacy 3-dp remap ---
        print(f"  Hourly CSV not found: {csv_path}")
        print(f"  Falling back to NREL FY2073 3-dp remap")
        dem = {}
        for s in SEASONS_LIST:
            remapped = remap_daypart_fractions(LEGACY_3DP_DEF, NPL_DP3[s], s)
            for ts, dp_share in remapped.items():
                dem[ts] = NPL_SEASONAL_SHARES[s] * dp_share
        total = sum(dem.values())
        dem = {ts: round(dem.get(ts, 0) / total, 6) for ts in TIMESLICES}
        print(f"  ** D2=D3 limitation (legacy remap)")
        src_note = 'NDOR seasonal x NREL FY2073 (remapped)'

    # --- CF rows ---
    cfs = []
    for ts in TIMESLICES:
        s, dp = ts[:2], ts[2:]
        # Hydro RoR — NDOR seasonal, flat within season
        hcf = NPL_HYDRO_CF.get(s)
        cfs.append({'timeslice': ts, 'season': SEASON_NAMES[s], 'daypart': DAYPART_NAMES[dp],
            'tech_code': 'PWRHYDNPLXX', 'tech_type': 'Hydro',
            'cf_dispatch': None, 'cf_ninja': None,
            'cf_da_workbook': round(hcf, 6) if hcf else None, 'cf_default': None,
            'source_notes': 'NDOR FY2080 fleet avg, flat within season'})
        # Solar
        cfs.append({'timeslice': ts, 'season': SEASON_NAMES[s], 'daypart': DAYPART_NAMES[dp],
            'tech_code': 'PWRSPVNPLXX', 'tech_type': 'Solar',
            'cf_dispatch': None,
            'cf_ninja': get_ninja_cf(ninja_cfs, 'solar', 'NPLXX', ts),
            'cf_da_workbook': None, 'cf_default': None, 'source_notes': 'Ninja'})
        # Wind
        cfs.append({'timeslice': ts, 'season': SEASON_NAMES[s], 'daypart': DAYPART_NAMES[dp],
            'tech_code': 'PWRWONNPLXX', 'tech_type': 'Wind',
            'cf_dispatch': None,
            'cf_ninja': get_ninja_cf(ninja_cfs, 'wind', 'NPLXX', ts),
            'cf_da_workbook': None, 'cf_default': None, 'source_notes': 'Ninja'})
        # Thermal
        cfs.append({'timeslice': ts, 'season': SEASON_NAMES[s], 'daypart': DAYPART_NAMES[dp],
            'tech_code': 'PWRGASNPLXX', 'tech_type': 'Gas',
            'cf_dispatch': None, 'cf_ninja': None,
            'cf_da_workbook': None, 'cf_default': 1.0, 'source_notes': 'thermal CF=1.0'})

    validate_demand_fractions(dem, 'NPL')
    validate_capacity_factors(cfs, 'NPL')
    validate_hydro_cf(cfs, 'NPL')

    meta = {
        'country': 'NPL',
        'source': src_note,
        'hydro_cf': 'NDOR FY2080 (818 days), IC=2452 MW, flat within season',
        'hourly_profiles': 'Simkhada et al. 2022 (midpoint DR + verified p.u.)',
        'net_exporter': 'S3/S4 net exporter, S1/S2 net importer. Demand fracs use consumption.',
    }
    return dem, cfs, meta


# ======================================================================
# BTN ADAPTER — reads BTN_Profiles.xlsx (hourly profiles + BPSO demand)
# ======================================================================

def run_btn_adapter(ninja_cfs):
    print(f"\n{'='*70}\n[{ts_now()}] BTN -- Bhutan\n{'='*70}")

    # --- Try BTN_Profiles.xlsx first (hourly, any DAYPART_DEF) ---
    demand_fracs = {}
    hydro_cf_seasonal = {}

    if os.path.exists(BTN_PROFILES_FILE):
        print(f"  Reading: {os.path.basename(BTN_PROFILES_FILE)}")

        # 1. Hourly profiles: 24 rows × (hour, S1, S2, S3, S4)
        df_hp = pd.read_excel(BTN_PROFILES_FILE, sheet_name='Hourly_Profiles')
        print(f"    Hourly_Profiles: {len(df_hp)} rows, seasons {[c for c in df_hp.columns if c.startswith('S')]}")

        # 2. Seasonal weights from Monthly_Demand
        df_md = pd.read_excel(BTN_PROFILES_FILE, sheet_name='Monthly_Demand')
        seasonal_weights = {}
        for _, row in df_md.iterrows():
            s = str(row['season']).strip()
            if s in SEASON_NAMES:
                seasonal_weights[s] = float(row['seasonal_weight'])
        # Deduplicate (all rows for same season have same weight)
        print(f"    Seasonal weights: {seasonal_weights}")

        # 3. Hydro CFs from Hydro_CF_Monthly (average by season)
        df_hcf = pd.read_excel(BTN_PROFILES_FILE, sheet_name='Hydro_CF_Monthly')
        hydro_cf_seasonal = {}
        for s in SEASONS_LIST:
            s_rows = df_hcf[df_hcf['season'] == s]
            if len(s_rows) > 0:
                hydro_cf_seasonal[s] = round(s_rows['CF_Capped'].mean(), 6)
        print(f"    Hydro CFs (seasonal avg): {hydro_cf_seasonal}")

        # 4. Compute demand fractions using hourly profiles × seasonal weights
        #    demand_fraction(s, d) = seasonal_weight(s) × daypart_share(s, d)
        #    where daypart_share = sum(MW in daypart hours) / sum(MW all 24 hours)
        demand_fracs = {}
        for s in SEASONS_LIST:
            if s not in df_hp.columns:
                print(f"    WARNING: Season {s} not in Hourly_Profiles columns")
                continue
            total_mw = df_hp[s].sum()
            for dp_code in DAYPARTS_LIST:
                dp_start, dp_end = DAYPART_RANGES[dp_code]
                mask = (df_hp['hour'] >= dp_start) & (df_hp['hour'] < dp_end)
                dp_mw = df_hp.loc[mask, s].sum()
                dp_share = dp_mw / total_mw if total_mw > 0 else 0
                ts = s + dp_code
                demand_fracs[ts] = seasonal_weights.get(s, 0) * dp_share

        # Normalize
        total = sum(demand_fracs.values())
        demand_fracs = {ts: round(demand_fracs.get(ts, 0) / total, 6) for ts in TIMESLICES}

        print(f"    Demand fractions computed: hourly profiles x seasonal weights")
        print(f"    Works for any DAYPART_DEF ({N_DAYPARTS} dayparts)")

    # --- Fallback to BTN_DA_v4.xlsx (legacy) ---
    elif os.path.exists(BTN_DA_FILE):
        print(f"  BTN_Profiles.xlsx not found. Falling back to BTN_DA_v4.xlsx")
        print(f"  ** For proper daypart resolution, run build_btn_profiles.py first.")
        demand_fracs = {}
        hydro_cf_seasonal = {}
        # (legacy code omitted for brevity — would use remap_daypart_fractions)
        for ts in TIMESLICES:
            demand_fracs[ts] = YEARSPLIT[ts]
        total = sum(demand_fracs.values())
        demand_fracs = {ts: round(v / total, 6) for ts, v in demand_fracs.items()}
    else:
        print(f"  ERROR: Neither BTN_Profiles.xlsx nor BTN_DA_v4.xlsx found")
        return None, None, None

    # --- Build CF rows ---
    cfs = []
    for ts in TIMESLICES:
        s, dp = ts[:2], ts[2:]
        hv = hydro_cf_seasonal.get(s)
        cfs.append({'timeslice': ts, 'season': SEASON_NAMES[s], 'daypart': DAYPART_NAMES[dp],
            'tech_code': 'PWRHYDBTNXX', 'tech_type': 'Hydro',
            'cf_dispatch': None, 'cf_ninja': None,
            'cf_da_workbook': hv, 'cf_default': None,
            'source_notes': 'BPSO 2024 ex-bus gen, seasonal avg, capped at 1.0' if hv else 'MISSING'})
        cfs.append({'timeslice': ts, 'season': SEASON_NAMES[s], 'daypart': DAYPART_NAMES[dp],
            'tech_code': 'PWRSPVBTNXX', 'tech_type': 'Solar',
            'cf_dispatch': None,
            'cf_ninja': get_ninja_cf(ninja_cfs, 'solar', 'BTNXX', ts),
            'cf_da_workbook': None, 'cf_default': None, 'source_notes': 'Ninja'})
        cfs.append({'timeslice': ts, 'season': SEASON_NAMES[s], 'daypart': DAYPART_NAMES[dp],
            'tech_code': 'PWRWONBTNXX', 'tech_type': 'Wind',
            'cf_dispatch': None,
            'cf_ninja': get_ninja_cf(ninja_cfs, 'wind', 'BTNXX', ts),
            'cf_da_workbook': None, 'cf_default': None, 'source_notes': 'Ninja'})

    validate_demand_fractions(demand_fracs, 'BTN')
    validate_capacity_factors(cfs, 'BTN')
    validate_hydro_cf(cfs, 'BTN')

    meta = {
        'country': 'BTN',
        'source': 'BTN_Profiles.xlsx (BPC PDB 2024 hourly + BPSO 2023 demand + BPSO 2024 hydro)',
        'net_exporter': '~75% to India, demand fracs use consumption',
        'hourly_profiles': 'BPC Table 5.20 (Thimphu proxy, 50/50 blend for S2/S4)',
        'hydro_cf': 'BPSO 2024 ex-bus, seasonal avg, flat within season, capped at 1.0',
    }
    return demand_fracs, cfs, meta


# ======================================================================
# MDV ADAPTER — Maldives (STELCO yearbook + WB hourly proxy)
# ======================================================================

# Monthly production fractions: STELCO Table 12.2 (2023, 60+ localities, metered)
MDV_MONTHLY_FRACS = {
    1: 0.07380, 2: 0.07229, 3: 0.08541, 4: 0.08899, 5: 0.08972,
    6: 0.08406, 7: 0.08442, 8: 0.09116, 9: 0.07870, 10: 0.08451,
    11: 0.08188, 12: 0.08506,
}

# 24-hour demand shape (normalized shares, season-invariant)
# Source: WB/IEEE Male EV Planning Paper Fig.14, April 2019 recorded loads
# Key insight: cooling-driven daytime plateau, NOT evening-peaked (equatorial)
MDV_HOURLY_SHARES = {
    0: 0.040, 1: 0.037, 2: 0.034, 3: 0.033, 4: 0.032, 5: 0.031,
    6: 0.031, 7: 0.032, 8: 0.035, 9: 0.042, 10: 0.046, 11: 0.048,
    12: 0.048, 13: 0.049, 14: 0.049, 15: 0.049, 16: 0.048, 17: 0.047,
    18: 0.045, 19: 0.045, 20: 0.045, 21: 0.045, 22: 0.046, 23: 0.043,
}

# Monthly GHI (kWh/m2/day) from WB/Solargis — for seasonal solar CF
MDV_MONTHLY_GHI = {
    1: 5.68, 2: 6.36, 3: 6.59, 4: 6.06, 5: 5.29, 6: 5.14,
    7: 5.10, 8: 5.40, 9: 5.39, 10: 5.65, 11: 5.02, 12: 4.95,
}
MDV_ANNUAL_SOLAR_CF = 0.18  # IRENA 2023
MDV_ANNUAL_GHI = sum(MDV_MONTHLY_GHI.values()) / 12

# Installed capacity (IRENA 2023 + MCCEE Road Map)
MDV_IC = {'Diesel': 531.5, 'Solar': 68.5, 'Wind': 1.5}


def run_mdv_adapter(ninja_cfs):
    print(f"\n{'='*70}\n[{ts_now()}] MDV -- Maldives\n{'='*70}")

    # Seasonal demand weights from monthly production
    seasonal_weights = {}
    for s in SEASONS_LIST:
        seasonal_weights[s] = sum(MDV_MONTHLY_FRACS[m] for m in range(1, 13)
                                  if MONTH_TO_SEASON[m] == s)
    total_sw = sum(seasonal_weights.values())
    seasonal_weights = {s: v / total_sw for s, v in seasonal_weights.items()}
    print(f"  Seasonal weights (STELCO 2023): "
          f"{', '.join(f'{s}={v:.3f}' for s, v in seasonal_weights.items())}")

    # Demand fractions: seasonal weight x daypart share
    # Single hourly profile for all seasons (equatorial, <1.3C annual range)
    dem = {}
    total_hourly = sum(MDV_HOURLY_SHARES.values())
    for s in SEASONS_LIST:
        for dp_code in DAYPARTS_LIST:
            ds, de = DAYPART_RANGES[dp_code]
            dp_share = sum(MDV_HOURLY_SHARES.get(h, 0) for h in range(ds, de)) / total_hourly
            dem[s + dp_code] = seasonal_weights[s] * dp_share
    total = sum(dem.values())
    dem = {ts: round(dem.get(ts, 0) / total, 6) for ts in TIMESLICES}
    print(f"  Demand fractions: STELCO seasonal x WB hourly ({N_DAYPARTS} dayparts)")
    print(f"  Note: Season-invariant hourly profile (cooling-driven daytime plateau)")

    # Solar CF by season (GHI-proportional)
    solar_cf_seasonal = {}
    for s in SEASONS_LIST:
        s_months = [m for m in range(1, 13) if MONTH_TO_SEASON[m] == s]
        s_ghi = sum(MDV_MONTHLY_GHI[m] for m in s_months) / len(s_months)
        solar_cf_seasonal[s] = round(MDV_ANNUAL_SOLAR_CF * (s_ghi / MDV_ANNUAL_GHI), 4)
    print(f"  Solar CF (GHI-scaled): "
          f"{', '.join(f'{s}={v:.3f}' for s, v in solar_cf_seasonal.items())}")

    # CF rows
    cfs = []
    for ts in TIMESLICES:
        s, dp = ts[:2], ts[2:]
        # Diesel — CF=1.0 default (optimizer dispatches within capacity)
        cfs.append({'timeslice': ts, 'season': SEASON_NAMES[s], 'daypart': DAYPART_NAMES[dp],
            'tech_code': 'PWRDSLMDVXX', 'tech_type': 'Diesel',
            'cf_dispatch': None, 'cf_ninja': None,
            'cf_da_workbook': None, 'cf_default': 1.0,
            'source_notes': 'thermal default CF=1.0'})
        # Solar — seasonal from GHI, zero at night
        is_night = (DAYPART_RANGES[dp][1] <= 6 or DAYPART_RANGES[dp][0] >= 20)
        sol_cf = 0.0 if is_night else solar_cf_seasonal.get(s, MDV_ANNUAL_SOLAR_CF)
        cf_ninja_sol = get_ninja_cf(ninja_cfs, 'solar', 'MDVXX', ts)
        cfs.append({'timeslice': ts, 'season': SEASON_NAMES[s], 'daypart': DAYPART_NAMES[dp],
            'tech_code': 'PWRSPVMDVXX', 'tech_type': 'Solar',
            'cf_dispatch': None, 'cf_ninja': cf_ninja_sol,
            'cf_da_workbook': round(sol_cf, 6), 'cf_default': None,
            'source_notes': 'GHI-scaled seasonal + Ninja'})
        # Wind — Ninja where available
        cf_ninja_wnd = get_ninja_cf(ninja_cfs, 'wind', 'MDVXX', ts)
        cfs.append({'timeslice': ts, 'season': SEASON_NAMES[s], 'daypart': DAYPART_NAMES[dp],
            'tech_code': 'PWRWONMDVXX', 'tech_type': 'Wind',
            'cf_dispatch': None, 'cf_ninja': cf_ninja_wnd,
            'cf_da_workbook': None, 'cf_default': None,
            'source_notes': 'Ninja (1.5 MW installed)'})

    validate_demand_fractions(dem, 'MDV')
    validate_capacity_factors(cfs, 'MDV')

    meta = {
        'country': 'MDV',
        'source': 'STELCO Table 12.2 (60+ localities, 2023) + WB/IEEE Male Fig.14 (Apr 2019)',
        'hourly_profile': 'Season-invariant (equatorial, cooling-driven daytime plateau)',
        'solar_cf': 'GHI-proportional from WB/Solargis, annual avg=0.18 (IRENA 2023)',
        'installed_capacity_MW': MDV_IC,
        'no_interconnection': 'Maldives has no cross-border interconnection',
    }
    return dem, cfs, meta


# ======================================================================
# OUTPUT FUNCTIONS
# ======================================================================

def export_country_csvs(country, dem, cfs, meta, odir):
    os.makedirs(odir, exist_ok=True)
    dp = os.path.join(odir, f'{country}_demand_fractions.csv')
    with open(dp,'w',newline='') as f:
        w = csv.writer(f); w.writerow(['timeslice','season','daypart','demand_fraction','yearsplit'])
        for ts in TIMESLICES:
            w.writerow([ts,SEASON_NAMES[ts[:2]],DAYPART_NAMES[ts[2:]],
                        f"{dem.get(ts,0):.6f}",f"{YEARSPLIT[ts]:.6f}"])
    print(f"  Saved: {dp}")
    cp = os.path.join(odir, f'{country}_capacity_factors.csv')
    with open(cp,'w',newline='') as f:
        w = csv.writer(f)
        w.writerow(['timeslice','season','daypart','tech_code','tech_type',
                     'cf_dispatch','cf_ninja','cf_da_workbook','cf_default','source_notes'])
        for r in cfs:
            w.writerow([r['timeslice'],r['season'],r['daypart'],r['tech_code'],r['tech_type'],
                f"{r['cf_dispatch']:.6f}" if r['cf_dispatch'] is not None else '',
                f"{r['cf_ninja']:.6f}" if r['cf_ninja'] is not None else '',
                f"{r['cf_da_workbook']:.6f}" if r['cf_da_workbook'] is not None else '',
                f"{r['cf_default']:.1f}" if r['cf_default'] is not None else '',
                r.get('source_notes','')])
    print(f"  Saved: {cp}")
    mp = os.path.join(odir, f'{country}_metadata.json')
    with open(mp,'w') as f: json.dump({**meta,'timeslice_config':get_config_snapshot()}, f, indent=2, default=str)
    print(f"  Saved: {mp}")

def build_selector(acf, odir):
    seen, rows = set(), []
    for co, cfs in acf.items():
        for r in cfs:
            k = (co, r['tech_code'])
            if k in seen: continue
            seen.add(k)
            av = [c for c in ['cf_dispatch','cf_ninja','cf_da_workbook','cf_default'] if r.get(c) is not None]
            # Recommend by provenance:
            #   Non-RES -> cf_default (always 1.0 after finalize_cf_row)
            #   RES     -> cf_ninja if available, else first valid dispatch, else da_workbook
            if not is_res_tech(r.get('tech_type'), r.get('tech_code')):
                rec = 'cf_default'
            else:
                priority = ['cf_ninja', 'cf_dispatch', 'cf_da_workbook', 'cf_default']
                rec = next((c for c in priority if c in av), av[0] if av else '')
            rows.append({'country':co,'tech_code':r['tech_code'],'tech_type':r['tech_type'],
                         'available':'; '.join(av),'recommended':rec,'notes':r.get('source_notes','')})
    sp = os.path.join(odir,'ostram_cf_selector_template.csv')
    with open(sp,'w',newline='') as f:
        w = csv.DictWriter(f, fieldnames=['country','tech_code','tech_type','available','recommended','notes'])
        w.writeheader(); w.writerows(rows)
    print(f"\n  Selector: {sp}")

def build_excel(ad, ac, am, odir):
    xp = os.path.join(odir,'OSTRAM_Timeslice_Outputs.xlsx')
    with pd.ExcelWriter(xp, engine='openpyxl') as wr:
        pd.DataFrame([{'timeslice':ts,'season':SEASON_NAMES[ts[:2]],'daypart':DAYPART_NAMES[ts[2:]],'yearsplit':YEARSPLIT[ts]} for ts in TIMESLICES]).to_excel(wr,sheet_name='YearSplit',index=False)
        for co in sorted(ad.keys()):
            pd.DataFrame([{'timeslice':ts,'demand_fraction':ad[co].get(ts,0),'yearsplit':YEARSPLIT[ts]} for ts in TIMESLICES]).to_excel(wr,sheet_name=f'{co}_Dem'[:31],index=False)
            if co in ac: pd.DataFrame(ac[co]).to_excel(wr,sheet_name=f'{co}_CF'[:31],index=False)
        pd.DataFrame([{'country':c,'timeslice':ts,'demand_fraction':ad[c].get(ts,0)} for c in sorted(ad) for ts in TIMESLICES]).to_excel(wr,sheet_name='Summary',index=False)
        cfg = [{'key':f'daypart_{d[0]}','value':f'{d[1]} ({d[2]:02d}-{d[3]:02d})'} for d in DAYPART_DEF]
        cfg += [{'key':f'season_{s[0]}','value':f'{s[1]} ({s[3]}d)'} for s in SEASON_DEF]
        pd.DataFrame(cfg).to_excel(wr,sheet_name='Config',index=False)
    print(f"\n  Excel: {xp}")


# ======================================================================
# MAIN
# ======================================================================

_T_WALL_START = time.perf_counter()
_T_STEPS = {}  # step_name -> seconds

print("="*70)
print(f"OSTRAM Timeslices -- {N_SEASONS}S x {N_DAYPARTS}D = {N_TIMESLICES} ts")
print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
print("="*70)
for d in DAYPART_DEF: print(f"  {d[0]}: {d[1]} ({d[2]:02d}:00-{d[3]:02d}:00, {d[3]-d[2]}h)")
for s in SEASON_DEF: print(f"  {s[0]}: {s[1]} ({s[3]}d)")
print(f"YearSplit sum: {sum(YEARSPLIT.values()):.6f}")
print(f"Countries: {COUNTRIES}")

os.makedirs(OUTPUT_DIR, exist_ok=True)
with open(os.path.join(OUTPUT_DIR,'timeslice_config.json'),'w') as f:
    json.dump(get_config_snapshot(), f, indent=2)

_t0 = time.perf_counter()
print(f"\n[{ts_now()}] Loading Ninja CFs...")
ninja_cfs = load_ninja_cfs(NINJA_TS_FILE, NINJA_CONFIG_FILE)
_T_STEPS['Ninja load'] = time.perf_counter() - _t0

ad, ac, am = {}, {}, {}

if 'BGD' in COUNTRIES:
    _t0 = time.perf_counter()
    d,c,m = run_bgd_adapter(ninja_cfs)
    _T_STEPS['BGD'] = time.perf_counter() - _t0
    if d:
        _t0 = time.perf_counter()
        ad['BGD']=d; ac['BGD']=c; am['BGD']=m; export_country_csvs('BGD',d,c,m,OUTPUT_DIR)
        _T_STEPS['BGD export'] = time.perf_counter() - _t0

if 'LKA' in COUNTRIES:
    _t0 = time.perf_counter()
    d,c,m = run_lka_adapter(ninja_cfs)
    _T_STEPS['LKA'] = time.perf_counter() - _t0
    if d:
        _t0 = time.perf_counter()
        ad['LKA']=d; ac['LKA']=c; am['LKA']=m; export_country_csvs('LKA',d,c,m,OUTPUT_DIR)
        _T_STEPS['LKA export'] = time.perf_counter() - _t0

if 'IND' in COUNTRIES:
    _t0 = time.perf_counter()
    id,ic,im = run_ind_adapter(ninja_cfs)
    _T_STEPS['IND (5 regions)'] = time.perf_counter() - _t0
    _t0 = time.perf_counter()
    for reg in IND_REGIONS:
        if reg in id:
            ad[reg]=id[reg]; ac[reg]=ic[reg]; am[reg]=im[reg]
            export_country_csvs(reg,id[reg],ic[reg],im[reg],OUTPUT_DIR)
    _T_STEPS['IND export'] = time.perf_counter() - _t0

if 'NPL' in COUNTRIES:
    _t0 = time.perf_counter()
    d,c,m = run_npl_adapter(ninja_cfs)
    _T_STEPS['NPL'] = time.perf_counter() - _t0
    if d:
        _t0 = time.perf_counter()
        ad['NPL']=d; ac['NPL']=c; am['NPL']=m; export_country_csvs('NPL',d,c,m,OUTPUT_DIR)
        _T_STEPS['NPL export'] = time.perf_counter() - _t0

if 'BTN' in COUNTRIES:
    _t0 = time.perf_counter()
    d,c,m = run_btn_adapter(ninja_cfs)
    _T_STEPS['BTN'] = time.perf_counter() - _t0
    if d:
        _t0 = time.perf_counter()
        ad['BTN']=d; ac['BTN']=c; am['BTN']=m; export_country_csvs('BTN',d,c,m,OUTPUT_DIR)
        _T_STEPS['BTN export'] = time.perf_counter() - _t0

if 'MDV' in COUNTRIES:
    _t0 = time.perf_counter()
    d,c,m = run_mdv_adapter(ninja_cfs)
    _T_STEPS['MDV'] = time.perf_counter() - _t0
    if d:
        _t0 = time.perf_counter()
        ad['MDV']=d; ac['MDV']=c; am['MDV']=m; export_country_csvs('MDV',d,c,m,OUTPUT_DIR)
        _T_STEPS['MDV export'] = time.perf_counter() - _t0

_t0 = time.perf_counter()
print(f"\n{'='*70}\n[{ts_now()}] Combined outputs\n{'='*70}")
if ad: build_selector(ac, OUTPUT_DIR); build_excel(ad, ac, am, OUTPUT_DIR)
_T_STEPS['Combined Excel'] = time.perf_counter() - _t0

_T_WALL_TOTAL = time.perf_counter() - _T_WALL_START

print(f"\n{'='*70}\nSUMMARY\n{'='*70}")
for co in sorted(ad):
    f = ad[co]; t = sum(f.values()); nc = len(ac.get(co,[]))
    print(f"  {co}: fracs sum={t:.6f}, CF rows={nc}")
    for ts in TIMESLICES: print(f"    {ts} {f.get(ts,0):.6f} {YEARSPLIT[ts]:.6f}")
    print()

print(f"\n{'='*70}")
print(f"TIMING REPORT  ({N_SEASONS}S x {N_DAYPARTS}D = {N_TIMESLICES} ts)")
print(f"{'='*70}")
for step, secs in _T_STEPS.items():
    if secs >= 60:
        print(f"  {step:<25s}  {secs/60:>6.1f} min  ({secs:>8.1f} s)")
    else:
        print(f"  {step:<25s}  {secs:>6.1f} s")
print(f"  {'-'*45}")
print(f"  {'TOTAL WALL TIME':<25s}  {_T_WALL_TOTAL/60:>6.1f} min  ({_T_WALL_TOTAL:>8.1f} s)")
print()
print(f"Config: {N_DAYPARTS} dayparts x {N_SEASONS} seasons = {N_TIMESLICES} timeslices")
print(f"Output: {OUTPUT_DIR}")
print(f"Finished: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
print("="*70)
