#!/usr/bin/env python
"""Build the OSTRAM timeslice fabric companion HTML from the computed workbooks.

Reads the four timeslice workbooks in outputs/ plus the per-fabric metadata in
_session_logs/fabric_menu_summary.json, and emits a single self-contained HTML
(outputs/Timeslice_Fabric_Companion.html) in which the fabric tabs switch every
section, including the zone-level demand chart, the CF chart and the heatmap,
to that fabric's real per-zone data.

The hand-built companion at the module root is the design template. It is read
at build time and patched with strictly anchored replacements; if any anchor is
not found exactly once, the build aborts rather than emit a wrong page.

SCHEMA MAPPING (validated against the ground-truth DATA object embedded in
Timeslice_Explainer_OSTRAM_SOASIA.html: 200/200 demand rows and 1200/1200 CF
values matched for the 5dp20ts reference workbook, e.g. BGD S1D2 demand
fraction 0.107446, YearSplit 0.113014, BGD S1D2 Solar CF 0.3843):

  Sheet "YearSplit"   columns: timeslice | season | daypart | yearsplit
                      one row per timeslice, defines the canonical timeslice
                      order (S1D1 .. S4Dn).
  Sheet "<ZONE>_Dem"  columns: timeslice | demand_fraction | yearsplit
                      demand_fraction -> df, yearsplit -> ys in the contract.
  Sheet "<ZONE>_CF"   columns: timeslice | season | daypart | tech_code |
                      tech_type | cf_dispatch | cf_ninja | cf_da_workbook |
                      cf_default | source_notes
                      The CF value per (timeslice, tech) is the first non-null
                      of: cf_ninja, cf_dispatch, cf_da_workbook, cf_default,
                      rounded to 4 decimal places (this precedence reproduces
                      the explainer's values exactly).
  Sheet "Config"      key/value rows: daypart_Dn -> "Label (HH-HH)",
                      season_Sn -> "Label (Mon-Mon) (NNd)".

  Technology normalisation (matches the explainer's tech set):
    - The India zones carry three hydro rows per timeslice (Hydro_HYD,
      Hydro_HDR, Hydro_HRO) with identical values; they collapse to a single
      "Hydro" series (the build asserts they are identical).
    - LKA "Major_Hydro" is renamed "Hydro"; "Mini_Hydro" (a constant 1.0
      baseload series) is dropped, as in the explainer, and logged.

  Per-fabric metadata comes from _session_logs/fabric_menu_summary.json only:
  peak solar block CF <- dom_mean_cf. The file carries no phantom solar
  percentages, so that stat is omitted from the generated grid-explorer stat
  cards (phantom:null) and logged; the hand-built summary table and teaching
  prose are template content and stay verbatim.

Usage:
  python scripts/build_companion_html.py [--out PATH]
"""

import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl

MODULE_DIR = Path(__file__).resolve().parent.parent
TEMPLATE = MODULE_DIR / "Timeslice_Fabric_Companion.html"
SUMMARY_JSON = MODULE_DIR / "_session_logs" / "fabric_menu_summary.json"
DEFAULT_OUT = MODULE_DIR / "outputs" / "Timeslice_Fabric_Companion.html"

# Fabric order matches the template's FABRICS array and the summary JSON.
FABRICS = [
    ("3dp12ts", "3dp/12ts", 12, "OSTRAM_Timeslice_Outputs_3dp12ts.xlsx"),
    ("4dp16ts", "4dp/16ts", 16, "OSTRAM_Timeslice_Outputs_4dp16ts.xlsx"),
    ("5dp20ts", "5dp/20ts", 20, "OSTRAM_Timeslice_Outputs_REFERENCE_5dp20ts.xlsx"),
    ("6dp24ts", "6dp/24ts", 24, "OSTRAM_Timeslice_Outputs_6dp24ts.xlsx"),
]

# Daypart boundaries per fabric as specified in the session brief. The build
# asserts each workbook's Config sheet agrees before emitting anything.
EXPECTED_DAYPARTS = {
    "3dp12ts": [(0, 8), (8, 16), (16, 24)],
    "4dp16ts": [(0, 6), (6, 12), (12, 18), (18, 24)],
    "5dp20ts": [(0, 6), (6, 17), (17, 20), (20, 22), (22, 24)],
    "6dp24ts": [(0, 5), (5, 8), (8, 17), (17, 20), (20, 22), (22, 24)],
}

EXPECTED_ZONES = ["BGD", "BTN", "INDEA", "INDNE", "INDNO", "INDSO", "INDWE", "LKA", "MDV", "NPL"]

# Month-range abbreviations used in the workbook season labels.
MONTH_NUM = {"Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
             "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12}

CF_PRECEDENCE = ("cf_ninja", "cf_dispatch", "cf_da_workbook", "cf_default")

log_lines = []


def log(msg):
    print(msg)
    log_lines.append(msg)


def fail(msg):
    print("BUILD ABORTED: " + msg, file=sys.stderr)
    sys.exit(1)


def months_from_range(rng):
    """'Dec-Feb' -> [12, 1, 2]; 'Jun-Sep' -> [6, 7, 8, 9]."""
    a, b = rng.split("-")
    start, end = MONTH_NUM[a], MONTH_NUM[b]
    months = [start]
    m = start
    while m != end:
        m = m % 12 + 1
        months.append(m)
    return months


def parse_config(wb):
    """Read the Config sheet into dayparts and seasons in contract shape."""
    cfg = {}
    for key, value in wb["Config"].iter_rows(values_only=True):
        if key != "key":
            cfg[key] = value
    dayparts, seasons = [], []
    for key in sorted(k for k in cfg if k.startswith("daypart_")):
        m = re.match(r"^(.*) \((\d{2})-(\d{2})\)$", cfg[key])
        if not m:
            fail(f"unparseable daypart label {cfg[key]!r}")
        dayparts.append({"code": key.split("_")[1], "label": m.group(1),
                         "start_hour": int(m.group(2)), "end_hour": int(m.group(3))})
    for key in sorted(k for k in cfg if k.startswith("season_")):
        m = re.match(r"^(.*) \(([A-Za-z]{3}-[A-Za-z]{3})\) \((\d+)d\)$", cfg[key])
        if not m:
            fail(f"unparseable season label {cfg[key]!r}")
        seasons.append({"code": key.split("_")[1],
                        "label": f"{m.group(1)} ({m.group(2)})",
                        "months": months_from_range(m.group(2)),
                        "days": int(m.group(3))})
    return dayparts, seasons


def cf_value(ninja, dispatch, da, default):
    for v in (ninja, dispatch, da, default):
        if v is not None:
            return round(float(v), 4)
    return None


def extract_fabric(path, fabric_id, n_ts):
    """Extract one workbook into the explainer DATA contract shape."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    dayparts, seasons = parse_config(wb)
    bounds = [(d["start_hour"], d["end_hour"]) for d in dayparts]
    if bounds != EXPECTED_DAYPARTS[fabric_id]:
        fail(f"{fabric_id}: Config dayparts {bounds} do not match the brief "
             f"{EXPECTED_DAYPARTS[fabric_id]}")

    ys_rows = [r for r in wb["YearSplit"].iter_rows(values_only=True)][1:]
    timeslices = [r[0] for r in ys_rows]
    if len(timeslices) != n_ts:
        fail(f"{fabric_id}: expected {n_ts} timeslices, found {len(timeslices)}")

    zones = [s[:-4] for s in wb.sheetnames if s.endswith("_Dem")]
    if zones != EXPECTED_ZONES:
        fail(f"{fabric_id}: zone set {zones} does not match expected {EXPECTED_ZONES}")

    demand = {}
    for z in zones:
        rows = [r for r in wb[f"{z}_Dem"].iter_rows(values_only=True)][1:]
        by_ts = {r[0]: r for r in rows}
        demand[z] = [{"ts": ts, "df": round(float(by_ts[ts][1]), 6),
                      "ys": round(float(by_ts[ts][2]), 6)} for ts in timeslices]
        sdf = sum(e["df"] for e in demand[z])
        sys_ = sum(e["ys"] for e in demand[z])
        if abs(sdf - 1) > 1e-3 or abs(sys_ - 1) > 1e-3:
            fail(f"{fabric_id}/{z}: demand fractions sum to {sdf}, yearsplit to {sys_}")

    cf = {}
    for z in zones:
        raw = {}       # (ts, tech_type) -> value
        tech_order = []
        for row in [r for r in wb[f"{z}_CF"].iter_rows(values_only=True)][1:]:
            ts, _season, _daypart, _code, tech, disp, ninja, da, default, _notes = row
            v = cf_value(ninja, disp, da, default)
            if v is None:
                fail(f"{fabric_id}/{z}: no CF value for {ts}/{tech}")
            if tech.startswith("Hydro_H"):          # India hydro triplet
                prev = raw.get((ts, "Hydro"))
                if prev is not None and prev != v:
                    fail(f"{fabric_id}/{z}: hydro variants differ at {ts}: {prev} vs {v}")
                tech = "Hydro"
            elif tech == "Major_Hydro":
                tech = "Hydro"
            elif tech == "Mini_Hydro":
                continue                             # dropped, logged below
            raw[(ts, tech)] = v
            if tech not in tech_order:
                tech_order.append(tech)
        cf[z] = {t: [{"ts": ts, "v": raw[(ts, t)]} for ts in timeslices]
                 for t in tech_order}
        for t, arr in cf[z].items():
            if len(arr) != n_ts:
                fail(f"{fabric_id}/{z}/{t}: {len(arr)} CF values, expected {n_ts}")
            for e in arr:
                if not (0 <= e["v"] <= 1.0001):
                    fail(f"{fabric_id}/{z}/{t}: CF out of range at {e['ts']}: {e['v']}")

    return {
        "config": {"seasons": seasons, "dayparts": dayparts,
                   "n_timeslices": n_ts, "timeslices": timeslices},
        "demand": demand,
        "cf": cf,
    }


def load_summary_meta():
    """Per-fabric stats from fabric_menu_summary.json; never recomputed."""
    if not SUMMARY_JSON.exists():
        log(f"NOTE: {SUMMARY_JSON.name} missing; peak CF and phantom stats omitted")
        return {}
    entries = json.loads(SUMMARY_JSON.read_text(encoding="utf-8")).get("fabrics", [])
    return {e["n_ts"]: e for e in entries}


def patch(html, old, new, what, count=1):
    """Anchored replacement; abort unless the anchor occurs exactly `count` times.

    `old` may be a plain string (literal match) or a compiled regex pattern
    (in which case `new` should be a callable to avoid escape processing).
    """
    if isinstance(old, re.Pattern):
        n = len(old.findall(html))
    else:
        n = html.count(old)
    if n != count:
        fail(f"template anchor for {what!r} found {n} times, expected {count}")
    return old.sub(new, html) if isinstance(old, re.Pattern) else html.replace(old, new)


def json_for_html(obj):
    """Compact JSON, </ escaped so the script block cannot be terminated early."""
    return json.dumps(obj, ensure_ascii=True, separators=(",", ":")).replace("</", "<\\/")


def build(out_path):
    if not TEMPLATE.exists():
        fail(f"design template not found: {TEMPLATE}")
    html = TEMPLATE.read_text(encoding="utf-8")
    meta = load_summary_meta()

    fabric_data = {}
    for fabric_id, fabric_name, n_ts, filename in FABRICS:
        path = MODULE_DIR / "outputs" / filename
        if not path.exists():
            fail(f"workbook not found: {path}")
        fabric_data[fabric_id] = extract_fabric(path, fabric_id, n_ts)
        log(f"extracted {fabric_id}: {n_ts} timeslices x "
            f"{len(fabric_data[fabric_id]['demand'])} zones")
    log("note: LKA Mini_Hydro (constant CF 1.0) dropped in all fabrics, "
        "matching the explainer's technology set")

    # 1. Embedded data: one DATA object per fabric, same contract as the explainer.
    data_js = ("const FABRIC_DATA = " + json_for_html(fabric_data) + ";\n"
               "let DATA = FABRIC_DATA['5dp20ts'];")
    html = patch(html, re.compile(r"^const DATA = \{.*\};$", re.M),
                 lambda m: data_js, "embedded DATA object")

    # 2. FABRICS stats: peak CF from the summary JSON; phantom solar is not in
    #    that file, so it is emitted as null and its stat card is omitted.
    stats_by_index = []
    for _fabric_id, fabric_name, n_ts, _f in FABRICS:
        e = meta.get(n_ts, {})
        peak = e.get("dom_mean_cf")
        if peak is None:
            log(f"NOTE: no dom_mean_cf for {fabric_name} in summary JSON; "
                "peak CF stat omitted for this fabric")
        stats_by_index.append((peak, fabric_name))
    if meta and all(m.get("dom_mean_cf") is None for m in meta.values()):
        log("NOTE: summary JSON has no dom_mean_cf values at all")
    log("NOTE: phantom solar percentages are not present in "
        "fabric_menu_summary.json; the phantom stat card is omitted from the "
        "grid explorer (the hand-built summary table keeps its template values)")

    counter = {"i": 0}

    def stat_sub(m):
        peak, fabric_name = stats_by_index[counter["i"]]
        counter["i"] += 1
        peak_js = "null" if peak is None else repr(peak)
        return f"phantom:null,peakCF:{peak_js},fabricName:'{fabric_name}'"

    html = patch(html, re.compile(r"phantom:[\d.]+,peakCF:[\d.]+"), stat_sub,
                 "FABRICS stat fields", count=4)

    # 3. Stat cards: built conditionally so a missing stat is omitted, not faked.
    old_stats = re.compile(
        r"  const vs20 = \(activeFab === 2\).*?document\.getElementById\('statsArea'\)\.innerHTML = `.*?`;",
        re.S)
    new_stats = """  const adopted = FABRICS[2];
  let cards = '<div class="stat"><div class="lab">Timeslices</div><div class="val">'+f.ts+'</div><div class="sub">'+SEASONS.length+' seasons x '+f.dayparts.length+' dayparts</div></div>';
  if (f.phantom != null) {
    cards += '<div class="stat"><div class="lab">Phantom solar</div><div class="val">'+f.phantom+'%</div><div class="sub">Solar credited to dark hours</div></div>';
  }
  if (f.peakCF != null) {
    const vs20 = (activeFab === 2 || adopted.peakCF == null) ? 'baseline' : ((f.peakCF/adopted.peakCF-1)*100).toFixed(1)+'%';
    cards += '<div class="stat"><div class="lab">Peak solar block CF</div><div class="val">'+f.peakCF.toFixed(3)+'</div><div class="sub">vs adopted: '+vs20+'</div></div>';
  }
  cards += '<div class="stat"><div class="lab">Build time</div><div class="val">~45s</div><div class="sub">Fabric-independent</div></div>';
  document.getElementById('statsArea').innerHTML = cards;"""
    html = patch(html, old_stats, lambda m: new_stats, "stat cards block")

    # 4. The fabric tabs drive the whole page, not just the grid.
    html = patch(html,
                 "b.onclick = () => { activeFab=i; renderGrid(); };",
                 "b.onclick = () => { setFabric(i); };",
                 "tab onclick")

    # 5. Zone-level section heading and prose follow the selected fabric.
    html = patch(html,
                 "<!-- SECTION 3: Zone-level explorer (20ts only, real data) -->",
                 "<!-- SECTION 3: Zone-level explorer (per fabric, real data) -->",
                 "section 3 comment")
    html = patch(html,
                 "<h2>Zone-level explorer (adopted 5dp/20ts)</h2>",
                 '<h2>Zone-level explorer <span id="zoneFabricLabel" style="color:#0F8B7D"></span></h2>',
                 "section 3 heading")
    html = patch(html,
                 "<p>Real pipeline data for all ten model zones. This section uses the "
                 "adopted fabric's computed values. The variant workbooks contain "
                 "equivalent per-zone data for each fabric.</p>",
                 "<p>Real pipeline data for all ten model zones, extracted from the "
                 "computed workbooks. The charts below follow the fabric selected in "
                 "the tabs above: the demand profile, the capacity factors and the "
                 "heatmap all switch to that fabric's per-zone values.</p>",
                 "section 3 intro")
    html = patch(html,
                 "how each responds to the 20-timeslice grid.",
                 "how each responds to the selected timeslice grid.",
                 "CF prose")

    # 6. A zone without the selected technology clears the chart rather than
    #    leaving the previous zone's data on screen.
    html = patch(html,
                 "  if(!cfData) return;",
                 "  if(!cfData) { if(cfChart) { cfChart.destroy(); cfChart = null; } return; }",
                 "missing-tech chart guard")

    # 7. Init goes through setFabric so every section starts consistent.
    html = patch(html,
                 """// Init
renderGrid();
renderDemandZones();
renderDemandChart();
renderCFZones();
renderCFChart();
renderHeatmap();""",
                 """// Fabric switching: one entry point keeps every section on the same fabric
function setFabric(i) {
  activeFab = i;
  DATA = FABRIC_DATA[FABRICS[i].id];
  const lbl = document.getElementById('zoneFabricLabel');
  if (lbl) lbl.textContent = '(' + (FABRICS[i].canonical ? 'adopted ' : '') + FABRICS[i].fabricName + ')';
  renderGrid();
  renderDemandChart();
  renderCFChart();
  renderHeatmap();
}

// Init
renderDemandZones();
renderCFZones();
setFabric(activeFab);""",
                 "init block")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(html, encoding="utf-8")

    # Final verdict
    verdict = "PASS: all workbook extractions validated (dayparts, zone sets, "
    verdict += "row counts, sums, CF ranges); all template anchors applied"
    print()
    print(f"output:  {out_path}")
    print(f"fabrics: {len(fabric_data)}")
    for fabric_id in fabric_data:
        print(f"  {fabric_id}: {len(fabric_data[fabric_id]['demand'])} zones, "
              f"{fabric_data[fabric_id]['config']['n_timeslices']} timeslices")
    print(f"verdict: {verdict}")


def main():
    ap = argparse.ArgumentParser(description="Build the timeslice fabric companion HTML")
    ap.add_argument("--out", type=Path, default=DEFAULT_OUT,
                    help=f"output HTML path (default: {DEFAULT_OUT})")
    args = ap.parse_args()
    build(args.out)


if __name__ == "__main__":
    main()
