# -*- coding: utf-8 -*-
"""
test_scenarios_lite.py — fast pre-pipeline checks for SOASIA v18 multi-scenario.

Covers plan tests #9 (bad config errors), #10 (cycle detection) and #11 (HTML
build). The expensive tests #1-#8 require running the full A3 pipeline and
should be exercised separately via `python run.py`.

Run:
    python tests/validation/test_scenarios_lite.py
"""

from __future__ import annotations

import shutil
import sys
import traceback
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[2]
A3_PROCESS_DIR = REPO_ROOT / "t1_confection" / "A3_process"
sys.path.insert(0, str(A3_PROCESS_DIR))

import _scenarios as S


V18 = A3_PROCESS_DIR / "SOASIA_OSeMOSYS_Template_v18.xlsx"
USER_GUIDE_HTML = A3_PROCESS_DIR / "docs" / "USER_GUIDE.html"


PASS = "\033[92mPASS\033[0m"
FAIL = "\033[91mFAIL\033[0m"


def _add_scenario_row(wb_path: Path, scenario: str, *, active=True,
                      rules_script="add_max_cap_investment_lid_rule.py",
                      inherit_from="") -> None:
    """Append a scenario row to the Control sheet."""
    wb = load_workbook(wb_path)
    ws = wb["Control"]
    # Find first empty row in column A
    next_row = 2
    while ws.cell(row=next_row, column=1).value is not None:
        next_row += 1
    ws.cell(row=next_row, column=1, value=scenario)
    ws.cell(row=next_row, column=2, value=active)
    ws.cell(row=next_row, column=3, value=rules_script)
    ws.cell(row=next_row, column=4, value=inherit_from)
    ws.cell(row=next_row, column=5, value="test fixture")
    wb.save(wb_path)


def _make_v18_copy(td: Path, label: str) -> Path:
    dst = td / f"v18_{label}.xlsx"
    shutil.copy(V18, dst)
    return dst


# -----------------------------------------------------------------------------
# Test 9 — bad config errors
# -----------------------------------------------------------------------------

def test_bad_rules_script() -> tuple[bool, str]:
    """A scenario referencing a nonexistent rules_script must fail with a
    message naming the scenario and available scripts."""
    with TemporaryDirectory(ignore_cleanup_errors=True) as td:
        td = Path(td)
        v18 = _make_v18_copy(td, "bad_rules")
        _add_scenario_row(v18, "NDC", rules_script="nonexistent.py")
        try:
            S.read_control_sheet(v18)
            return False, "read_control_sheet did NOT raise on bad rules_script"
        except ValueError as e:
            msg = str(e)
            if "NDC" in msg and "nonexistent.py" in msg:
                return True, f"OK: {msg[:120]}..."
            return False, f"raised but message lacks scenario/script: {msg}"


def test_bad_inherit_reference() -> tuple[bool, str]:
    """A scenario inheriting from an unknown name must fail."""
    with TemporaryDirectory(ignore_cleanup_errors=True) as td:
        td = Path(td)
        v18 = _make_v18_copy(td, "bad_inherit")
        _add_scenario_row(v18, "NDC", inherit_from="UNKNOWN_SCEN")
        try:
            S.read_control_sheet(v18)
            return False, "did NOT raise on unknown inherit reference"
        except ValueError as e:
            msg = str(e)
            if "NDC" in msg and "UNKNOWN_SCEN" in msg:
                return True, f"OK: {msg[:120]}..."
            return False, f"raised but message lacks names: {msg}"


def test_duplicate_scenario_name() -> tuple[bool, str]:
    """Two rows with the same scenario name must fail."""
    with TemporaryDirectory(ignore_cleanup_errors=True) as td:
        td = Path(td)
        v18 = _make_v18_copy(td, "dup")
        _add_scenario_row(v18, "BAU")  # duplicate of the seeded BAU
        try:
            S.read_control_sheet(v18)
            return False, "did NOT raise on duplicate scenario"
        except ValueError as e:
            msg = str(e)
            if "BAU" in msg or "Duplicate" in msg:
                return True, f"OK: {msg[:120]}..."
            return False, f"unexpected message: {msg}"


# -----------------------------------------------------------------------------
# Test 10 — cycle detection
# -----------------------------------------------------------------------------

def test_inheritance_cycle() -> tuple[bool, str]:
    """A <- B and B <- A must be detected as a cycle by topological_order."""
    with TemporaryDirectory(ignore_cleanup_errors=True) as td:
        td = Path(td)
        v18 = _make_v18_copy(td, "cycle")
        _add_scenario_row(v18, "A", inherit_from="B")
        _add_scenario_row(v18, "B", inherit_from="A")
        configs = S.read_control_sheet(v18)
        try:
            S.topological_order(configs)
            return False, "topological_order did NOT raise on cycle"
        except ValueError as e:
            msg = str(e)
            if "cycle" in msg.lower():
                return True, f"OK: {msg[:120]}..."
            return False, f"raised but message lacks 'cycle': {msg}"


def test_topo_order_chain() -> tuple[bool, str]:
    """When NDC inherits from BAU and both are active, BAU must come first."""
    with TemporaryDirectory(ignore_cleanup_errors=True) as td:
        td = Path(td)
        v18 = _make_v18_copy(td, "topo_chain")
        _add_scenario_row(v18, "NDC", inherit_from="BAU")
        configs = S.read_control_sheet(v18)
        ordered = S.topological_order(configs)
        names = [c.scenario for c in ordered]
        if "BAU" in names and "NDC" in names and names.index("BAU") < names.index("NDC"):
            return True, f"OK: {names}"
        return False, f"unexpected order: {names}"


# -----------------------------------------------------------------------------
# Test 11 — HTML build
# -----------------------------------------------------------------------------

def test_html_build() -> tuple[bool, str]:
    """USER_GUIDE.html must exist, be non-trivial, and reference key sections."""
    if not USER_GUIDE_HTML.is_file():
        return False, f"missing: {USER_GUIDE_HTML}"
    size = USER_GUIDE_HTML.stat().st_size
    if size < 5000:
        return False, f"suspiciously small ({size} B)"
    text = USER_GUIDE_HTML.read_text(encoding="utf-8")
    needles = [
        "<title>",
        "<style>",
        "Control",
        "Restrictions",
        "rules_script",
        "inherit_restrictions_from",
        "<table",
    ]
    missing = [n for n in needles if n not in text]
    if missing:
        return False, f"HTML missing markers: {missing}"
    return True, f"OK ({size // 1024} KB, all markers present)"


# -----------------------------------------------------------------------------
# Bonus — confirm read_restrictions guard works (empty Restrictions case)
# -----------------------------------------------------------------------------

def test_read_restrictions_empty_source() -> tuple[bool, str]:
    """Inheriting from a scenario with no rows in Restrictions must error."""
    with TemporaryDirectory(ignore_cleanup_errors=True) as td:
        v18 = _make_v18_copy(Path(td), "empty_restrictions")
        wb = load_workbook(v18)
        ws = wb["Restrictions"]
        if ws.max_row >= 2:
            ws.delete_rows(2, ws.max_row - 1)
        wb.save(v18)
        try:
            S.read_restrictions(v18, ["BAU"])
            return False, "read_restrictions did NOT raise on empty source"
        except ValueError as e:
            msg = str(e)
            if "No Restrictions rows" in msg and "BAU" in msg:
                return True, f"OK: {msg[:120]}..."
            return False, f"raised but message wrong: {msg}"


# -----------------------------------------------------------------------------
# Bonus — sanity-check materialize on the seeded v18 (BAU only) matches v17
# -----------------------------------------------------------------------------

def test_materialize_bau_contract() -> tuple[bool, str]:
    """materialize(BAU) drops multi-scenario sheets and scenario columns."""
    v17 = A3_PROCESS_DIR / "SOASIA_OSeMOSYS_Template_v17.xlsx"
    if not v17.is_file():
        return False, f"v17 missing: {v17}"
    with TemporaryDirectory(ignore_cleanup_errors=True) as td:
        out = Path(td) / "materialized.xlsx"
        S.materialize_scenario_template(V18, "BAU", out)
        wb17 = load_workbook(v17, data_only=False)
        wb18 = load_workbook(V18, data_only=False)
        wbm = load_workbook(out, data_only=False)
        if wb17.sheetnames != wbm.sheetnames:
            return False, (
                f"sheet name mismatch: v17={wb17.sheetnames} "
                f"mat={wbm.sheetnames}"
            )
        for s in S.PARAMETRIC_SHEETS:
            h18 = [c.value for c in next(wb18[s].iter_rows(min_row=1, max_row=1))]
            hm = [c.value for c in next(wbm[s].iter_rows(min_row=1, max_row=1))]
            expected = [value for value in h18 if value != "scenario"]
            if hm != expected:
                return False, f"{s} materialized header differs from v18 minus scenario"
        return True, "OK: v17 sheet set + v18 headers without scenario"


# -----------------------------------------------------------------------------
# Driver
# -----------------------------------------------------------------------------

TESTS = [
    ("9a  bad rules_script reference",     test_bad_rules_script),
    ("9b  bad inherit_from reference",     test_bad_inherit_reference),
    ("9c  duplicate scenario name",        test_duplicate_scenario_name),
    ("10a cycle detection",                test_inheritance_cycle),
    ("10b topo order chain (BAU<-NDC)",    test_topo_order_chain),
    ("11  USER_GUIDE.html sanity",         test_html_build),
    ("+   read_restrictions empty source", test_read_restrictions_empty_source),
    ("+   materialize(BAU) contract",       test_materialize_bau_contract),
]


def main() -> int:
    n_pass = n_fail = 0
    for label, fn in TESTS:
        try:
            ok, msg = fn()
        except Exception as exc:
            ok = False
            msg = f"EXCEPTION: {exc.__class__.__name__}: {exc}"
            traceback.print_exc()
        tag = PASS if ok else FAIL
        print(f"  [{tag}] {label}: {msg}")
        if ok:
            n_pass += 1
        else:
            n_fail += 1
    print()
    print(f"  {n_pass} passed, {n_fail} failed")
    return 0 if n_fail == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
