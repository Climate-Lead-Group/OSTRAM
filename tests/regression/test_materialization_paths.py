from __future__ import annotations

import argparse
from contextlib import redirect_stdout
import hashlib
import io
import os
from pathlib import Path
import shutil
import tempfile
from types import SimpleNamespace
import unittest
from unittest import mock

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from ostram import paths as path_module
from ostram.pipeline.scenarios import orchestrator as a3_orchestrator
from ostram.pipeline.scenarios import transform
from ostram.pipeline.scenarios.transformations import cap_trn_to_residual
from ostram.pipeline.scenarios.transformations import clear_stale_unbinding_caps
from ostram.pipeline.scenarios.transformations import fix_trn_residuals
from ostram.pipeline.scenarios.transformations import scenario_workbooks


REPO_ROOT = Path(__file__).resolve().parents[2]
SCENARIO_AUTHORITY = (
    REPO_ROOT / "inputs" / "scenarios" / "OSTRAM_Scenario_Inputs.xlsx"
)
FAILED_STAGE3_NAME = (
    "A-O_Parametrization_c2a_patched_FIXED_POST_CAP_RESET_20260801_140148_"
    "POST_TRN_CAP_20260801_140151.xlsx"
)
INPUT_FILES = (
    "A-O_AR_Model_Base_Year.xlsx",
    "A-O_AR_Projections.xlsx",
    "A-O_Demand.xlsx",
    "A-O_Parametrization.xlsx",
)


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _comparable_stage3_parent(temp_root: Path) -> tuple[Path, Path]:
    """Return a creatable parent whose legacy desired path is at least 265."""

    tail = (
        Path("workspaces")
        / "governed 15 café"
        / "scenarios"
        / "_run_20260801_135946"
        / "stage3"
    )
    parent = temp_root / tail
    legacy = parent / FAILED_STAGE3_NAME
    missing = max(0, 265 - path_module.windows_path_units(legacy))
    if missing:
        padding = "root-padding-" + ("x" * max(1, missing - 14))
        parent = temp_root / padding / tail
        legacy = parent / FAILED_STAGE3_NAME
    parent.mkdir(parents=True)
    if path_module.windows_path_units(legacy) < 265:
        raise AssertionError(f"long-path fixture is too short: {legacy}")
    return parent, legacy


def _build_cap_fixture(path: Path) -> None:
    """Build a representative Stage-3 workbook with exactly 777 cap changes."""

    authority = fix_trn_residuals.load_rc_authority(SCENARIO_AUTHORITY)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Secondary Techs"
    years = list(cap_trn_to_residual.YEARS)
    worksheet.append(
        [
            "Tech.ID",
            "Tech",
            "Tech.Name",
            "Parameter.ID",
            "Parameter",
            "Unit",
            "Projection.Mode",
            "Projection.Parameter",
            *years,
        ]
    )

    remaining_max_investment_changes = 273
    parameter_ids = {
        fix_trn_residuals.RESIDUAL_PARAM: 3,
        fix_trn_residuals.MIN_INV_PARAM: 7,
        cap_trn_to_residual.PARAM_MAXCAP: 8,
        cap_trn_to_residual.PARAM_MAXCAPINV: 9,
    }
    for index, tech in enumerate(sorted(cap_trn_to_residual.TRN_TECHS), start=1):
        for parameter in (
            fix_trn_residuals.RESIDUAL_PARAM,
            fix_trn_residuals.MIN_INV_PARAM,
            cap_trn_to_residual.PARAM_MAXCAP,
            cap_trn_to_residual.PARAM_MAXCAPINV,
        ):
            if parameter == fix_trn_residuals.RESIDUAL_PARAM:
                values = [authority[tech][year] for year in years]
            elif parameter == cap_trn_to_residual.PARAM_MAXCAP:
                values = [None] * len(years)
            elif parameter == cap_trn_to_residual.PARAM_MAXCAPINV:
                changed = min(remaining_max_investment_changes, len(years))
                values = [0] * changed + [9999] * (len(years) - changed)
                remaining_max_investment_changes -= changed
            else:
                values = [0.0] * len(years)
            worksheet.append(
                [
                    index,
                    tech,
                    f"Representative {tech}",
                    parameter_ids[parameter],
                    parameter,
                    "GW",
                    fix_trn_residuals.DEFAULT_PROJECTION_MODE,
                    0,
                    *values,
                ]
            )
    if remaining_max_investment_changes != 0:
        raise AssertionError("fixture did not allocate exactly 273 changes")

    audit = workbook.create_sheet("Formula Audit")
    audit["A1"] = "=SUM(1,2)"
    audit["A1"].font = Font(bold=True, color="FFFFFF")
    audit["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    audit["A1"].number_format = "0.00"
    audit["A2"] = "Unicode café Ω and spaces"
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()


def _workbook_snapshot(path: Path) -> tuple:
    workbook = load_workbook(path, data_only=False)
    try:
        sheets = []
        for worksheet in workbook.worksheets:
            cells = []
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is None and not cell.has_style:
                        continue
                    cells.append(
                        (
                            cell.coordinate,
                            cell.value,
                            cell.data_type,
                            cell.number_format,
                            cell.style_id,
                        )
                    )
            sheets.append((worksheet.title, tuple(cells)))
        return tuple(sheets)
    finally:
        workbook.close()


def _argument_map(arguments: list[object]) -> dict[str, Path | str]:
    values: dict[str, Path | str] = {}
    index = 0
    while index < len(arguments):
        key = str(arguments[index])
        if key.startswith("--") and index + 1 < len(arguments):
            values[key] = arguments[index + 1]
            index += 2
        else:
            index += 1
    return values


def _stage3_dispatch(module: str, arguments=None, cwd=None, label=None) -> str:
    del label
    root = Path(cwd).resolve()
    values = _argument_map(list(arguments or []))

    def resolve(value: Path | str) -> Path:
        path = Path(value)
        return path.resolve() if path.is_absolute() else (root / path).resolve()

    if module.endswith("fix_trn_residuals"):
        diffs, plans, skipped, warnings = fix_trn_residuals.run_fix(
            resolve(values["--input"]),
            resolve(values["--output"]),
            mode=str(values["--mode"]),
            cutoff_year=int(str(values["--cutoff-year"])),
            authority_path=resolve(values["--authority"]),
        )
        fix_trn_residuals.write_diff_csv(diffs, resolve(values["--diff-csv"]))
        fix_trn_residuals.write_diff_md(
            diffs,
            plans,
            skipped,
            warnings,
            mode=str(values["--mode"]),
            cutoff_year=int(str(values["--cutoff-year"])),
            path=resolve(values["--diff-md"]),
        )
        return ""
    if module.endswith("clear_stale_unbinding_caps"):
        clear_stale_unbinding_caps.patch_workbook(
            resolve(values["--input"]),
            False,
            resolve(values["--output"]),
        )
        return ""
    if module.endswith("cap_trn_to_residual"):
        cap_trn_to_residual.patch_workbook(
            resolve(values["--input"]),
            False,
            resolve(values["--output"]),
        )
        return ""
    raise AssertionError(f"unexpected Stage-3 module: {module}")


def _run_stage3_chain(root: Path) -> tuple[Path, Path]:
    stage1 = root / "stage1"
    stage2 = root / "stage2"
    stage3 = root / "stage3"
    stage5 = root / "stage5"
    for directory in (stage1, stage2, stage3, stage5):
        directory.mkdir(parents=True, exist_ok=True)
    source = stage2 / "A-O_Parametrization_c2a_patched.xlsx"
    _build_cap_fixture(source)
    aligned = stage1 / "wvaligned_outputs_v2"
    aligned.mkdir()
    for name in (
        "A-O_AR_Model_Base_Year_wvaligned_v2.xlsx",
        "A-O_AR_Projections_wvaligned_v2.xlsx",
        "A-O_Demand_wvaligned_v2.xlsx",
    ):
        shutil.copy(source, aligned / name)
    authority = root / "_materialized_A_Calibrated_BAU.xlsx"
    scenario_workbooks.materialize_scenario_template(
        SCENARIO_AUTHORITY,
        "A_Calibrated_BAU",
        authority,
    )
    previous = os.environ.get("OSTRAM_TEMPLATE_PATH")
    os.environ["OSTRAM_TEMPLATE_PATH"] = str(authority)
    try:
        with mock.patch.object(transform, "run_subproc", side_effect=_stage3_dispatch):
            final_stage3 = transform.stage_3_fix_2(stage2, stage3)
        transform.stage_4_consolidate(stage1, stage3, stage5, final_stage3)
    finally:
        if previous is None:
            os.environ.pop("OSTRAM_TEMPLATE_PATH", None)
        else:
            os.environ["OSTRAM_TEMPLATE_PATH"] = previous
    return final_stage3, stage5 / "A-O_Parametrization.xlsx"


class BoundedWorkbookPathTests(unittest.TestCase):
    def test_bounded_name_unit_behavior_preserves_safe_and_compacts_long_paths(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir=REPO_ROOT.parent) as temp:
            root = Path(temp).resolve()
            safe = root / "safe café workspace" / "POST_TRN_CAP.xlsx"
            self.assertEqual(
                path_module.bounded_workspace_workbook_path(
                    safe,
                    stage_identity="POST_TRN_CAP",
                ),
                safe.resolve(),
            )

            parent, desired = _comparable_stage3_parent(root)
            bounded = path_module.bounded_workspace_workbook_path(
                desired,
                stage_identity="POST_TRN_CAP",
            )
            repeated = path_module.bounded_workspace_workbook_path(
                desired,
                stage_identity="POST_TRN_CAP",
            )
            distinct = path_module.bounded_workspace_workbook_path(
                desired.with_name("distinct_" + desired.name),
                stage_identity="POST_TRN_CAP",
            )
            self.assertEqual(bounded, repeated)
            self.assertNotEqual(bounded, distinct)
            self.assertEqual(bounded.parent, parent)
            self.assertEqual(bounded.suffix, ".xlsx")
            self.assertIn("POST_TRN_CAP", bounded.stem)
            self.assertLess(
                path_module.windows_path_units(bounded),
                path_module.WINDOWS_SAFE_ABSOLUTE_PATH_BUDGET,
            )
            self.assertGreaterEqual(path_module.windows_path_units(desired), 265)

    def test_impossibly_long_parent_fails_before_transformation(self) -> None:
        with tempfile.TemporaryDirectory(dir=REPO_ROOT.parent) as temp:
            parent = Path(temp).resolve()
            while path_module.windows_path_units(parent) < 222:
                parent /= "parent-segment-with-no-safe-filename-budget"
            desired = parent / FAILED_STAGE3_NAME
            with self.assertRaisesRegex(
                path_module.WorkspacePathBudgetError,
                "parent leaves no Windows-safe workbook filename budget",
            ):
                path_module.bounded_workspace_workbook_path(
                    desired,
                    stage_identity="POST_TRN_CAP",
                )


class WorkbookRoundTripAndMaterializationTests(unittest.TestCase):
    def test_real_openpyxl_long_unicode_space_round_trip(self) -> None:
        with tempfile.TemporaryDirectory(dir=REPO_ROOT.parent) as temp:
            root = Path(temp).resolve()
            parent, desired = _comparable_stage3_parent(root)
            source = parent / "A-O_Parametrization_FIXED.xlsx"
            _build_cap_fixture(source)
            source_hash = _sha256(source)
            output = io.StringIO()
            with redirect_stdout(output):
                actual = cap_trn_to_residual.patch_workbook(
                    source,
                    False,
                    desired,
                )
            self.assertIsNotNone(actual)
            assert actual is not None
            self.assertTrue(actual.is_file())
            self.assertLess(
                path_module.windows_path_units(actual),
                path_module.WINDOWS_SAFE_ABSOLUTE_PATH_BUDGET,
            )
            self.assertEqual(_sha256(source), source_hash)
            self.assertIn("TRN techs processed: 18 / 18 allowlisted", output.getvalue())
            self.assertIn("777 cells changed", output.getvalue())
            workbook = load_workbook(actual, data_only=False)
            try:
                self.assertEqual(workbook["Formula Audit"]["A1"].value, "=SUM(1,2)")
                self.assertTrue(workbook["Formula Audit"]["A1"].font.bold)
                self.assertEqual(
                    workbook["Formula Audit"]["A2"].value,
                    "Unicode café Ω and spaces",
                )
            finally:
                workbook.close()
            self.assertFalse(any(path.name == ".dvc" for path in root.rglob("*")))

    def test_actual_a_calibrated_bau_materialization_route_reaches_consumer(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir=REPO_ROOT.parent) as temp:
            root = Path(temp).resolve()
            stage3_parent, legacy = _comparable_stage3_parent(root)
            process_dir = stage3_parent.parent.parent
            expected_stage3 = (
                process_dir / "_run_20260801_135946" / "stage3"
            )
            if expected_stage3 != stage3_parent:
                raise AssertionError((expected_stage3, stage3_parent))
            preparation = root / "preparation"
            snapshot = (
                preparation
                / "A1_Outputs"
                / "_post_a2_snapshot_A_Calibrated_BAU"
            )
            snapshot.mkdir(parents=True)
            _build_cap_fixture(snapshot / "A-O_Parametrization.xlsx")
            for filename in INPUT_FILES[:3]:
                shutil.copy(snapshot / "A-O_Parametrization.xlsx", snapshot / filename)
            output_dir = (
                preparation
                / "A1_Outputs"
                / "A1_Outputs_A_Calibrated_BAU"
            )
            authority_hash = _sha256(SCENARIO_AUTHORITY)

            def stage1(stage1_dir: Path) -> None:
                aligned = stage1_dir / "wvaligned_outputs_v2"
                aligned.mkdir()
                shutil.copy(
                    stage1_dir / "A-O_Parametrization.xlsx",
                    aligned / "A-O_Parametrization_wvaligned_v2_ts20.xlsx",
                )
                for source_name, target_name in (
                    (
                        "A-O_AR_Model_Base_Year.xlsx",
                        "A-O_AR_Model_Base_Year_wvaligned_v2.xlsx",
                    ),
                    (
                        "A-O_AR_Projections.xlsx",
                        "A-O_AR_Projections_wvaligned_v2.xlsx",
                    ),
                    ("A-O_Demand.xlsx", "A-O_Demand_wvaligned_v2.xlsx"),
                ):
                    shutil.copy(stage1_dir / source_name, aligned / target_name)

            def stage1b(_wd: Path, stage1_dir: Path, stage1b_dir: Path) -> None:
                aligned = stage1_dir / "wvaligned_outputs_v2"
                shutil.copy(
                    aligned / "A-O_Parametrization_wvaligned_v2_ts20.xlsx",
                    stage1b_dir / "A-O_Parametrization.xlsx",
                )

            def stage2(_wd: Path, stage1b_dir: Path, stage2_dir: Path) -> None:
                shutil.copy(
                    stage1b_dir / "A-O_Parametrization.xlsx",
                    stage2_dir / "A-O_Parametrization_c2a_patched.xlsx",
                )

            def deliver(stage5: Path, destination: Path) -> None:
                destination.mkdir(parents=True, exist_ok=True)
                for filename in INPUT_FILES:
                    shutil.copy(stage5 / filename, destination / filename)

            fixed_time = SimpleNamespace(
                strftime=lambda _format: "20260801_135946"
            )
            clock_values = iter((0.0, 1.0))
            dependencies = a3_orchestrator.A3Dependencies(
                resolve_scenario_config=lambda _args, _soasia: (
                    "A_Calibrated_BAU",
                    [],
                    [],
                ),
                resolve_path=lambda path: Path(path).resolve(),
                build_workdir=transform.build_workdir,
                materialize_scenario_template=(
                    scenario_workbooks.materialize_scenario_template
                ),
                stage_1_scripts_1_to_5=stage1,
                stage_1b=stage1b,
                stage_2_and_2_5=stage2,
                stage_3_fix_2=transform.stage_3_fix_2,
                stage_4_consolidate=transform.stage_4_consolidate,
                stage_4_5_apply_inherited_restrictions=lambda *_args: None,
                stage_5_rules_scripts=lambda *_args: None,
                stage_ws3_interconnector_costs=lambda *_args: None,
                stage_ws3_internal_transmission=lambda *_args: None,
                stage_ws3_internal_tx_losses=lambda *_args: None,
                stage_ws4_pwr_min_pin=lambda *_args: None,
                stage_6_sync_og_to_ts20=lambda *_args: None,
                stage_6_persist_restrictions=lambda *_args: None,
                deliver_outputs=deliver,
                remove_tree=shutil.rmtree,
                copy_tree=shutil.copytree,
                copy_file=shutil.copy,
                environment=os.environ,
                clock=lambda: next(clock_values),
                timestamp_now=lambda: fixed_time,
                banner=lambda _message: None,
                emit=lambda _message: None,
            )
            plan = a3_orchestrator.A3Plan(
                scenario="A_Calibrated_BAU",
                rules_scripts=(),
                inherit_from=(),
                soasia=SCENARIO_AUTHORITY,
                input_dir=output_dir,
                output_dir=output_dir,
                snapshot_dir=snapshot,
                workdir_base=process_dir,
                keep_workdir=True,
            )
            stdout = io.StringIO()
            previous = os.environ.get("OSTRAM_TEMPLATE_PATH")
            try:
                with (
                    redirect_stdout(stdout),
                    mock.patch.object(
                        transform,
                        "run_subproc",
                        side_effect=_stage3_dispatch,
                    ),
                ):
                    result = a3_orchestrator.execute_plan(
                        plan,
                        dependencies,
                        INPUT_FILES,
                    )
            finally:
                if previous is None:
                    os.environ.pop("OSTRAM_TEMPLATE_PATH", None)
                else:
                    os.environ["OSTRAM_TEMPLATE_PATH"] = previous

            completed = output_dir / "A-O_Parametrization.xlsx"
            self.assertEqual(result, 0)
            self.assertTrue(completed.is_file())
            self.assertGreaterEqual(path_module.windows_path_units(legacy), 265)
            self.assertIn("TRN techs processed: 18 / 18 allowlisted", stdout.getvalue())
            self.assertIn("777 cells changed", stdout.getvalue())
            workbook = load_workbook(completed, data_only=False)
            try:
                self.assertEqual(workbook["Formula Audit"]["A1"].value, "=SUM(1,2)")
            finally:
                workbook.close()
            self.assertEqual(_sha256(SCENARIO_AUTHORITY), authority_hash)
            self.assertFalse(any(path.name == ".dvc" for path in root.rglob("*")))

    def test_short_and_bounded_long_materialization_content_is_identical(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir=REPO_ROOT.parent) as temp:
            root = Path(temp).resolve()
            short_root = root / "short"
            long_parent, _legacy = _comparable_stage3_parent(root / "long")
            long_root = long_parent.parent
            authority_hash = _sha256(SCENARIO_AUTHORITY)
            short_stage3, short_output = _run_stage3_chain(short_root)
            long_stage3, long_output = _run_stage3_chain(long_root)
            self.assertTrue(short_stage3.is_file())
            self.assertTrue(long_stage3.is_file())
            self.assertEqual(
                _workbook_snapshot(short_output),
                _workbook_snapshot(long_output),
            )
            for filename in INPUT_FILES[:3]:
                self.assertEqual(
                    _workbook_snapshot(short_output.parent / filename),
                    _workbook_snapshot(long_output.parent / filename),
                )
            self.assertEqual(_sha256(SCENARIO_AUTHORITY), authority_hash)


if __name__ == "__main__":
    unittest.main()
