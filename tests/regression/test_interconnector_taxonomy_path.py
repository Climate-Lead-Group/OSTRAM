from __future__ import annotations

import csv
from contextlib import contextmanager
import hashlib
import os
from pathlib import Path
import tempfile
import unittest
from unittest import mock

from openpyxl import Workbook, load_workbook

from ostram import paths as path_module
from ostram.pipeline.scenarios.rules import set_interconnector_direction as rule


REPO_ROOT = Path(__file__).resolve().parents[2]
AUTHORITATIVE_TAXONOMY = (
    REPO_ROOT / "config" / "scenarios" / "technology_types.csv"
)
OBSOLETE_TAXONOMY = REPO_ROOT / "ostram" / "pipeline" / "scenarios" / "TECH_TYPES.csv"
CONTRACTUAL_OVERLAY = (
    REPO_ROOT
    / "config"
    / "scenarios"
    / "B_Opt_DirContractual"
    / rule.YAML_FILE_NAME
)
EXPECTED_INTERCONNECTORS = {
    "TRNBGDXXINDEA",
    "TRNBGDXXINDNE",
    "TRNBTNXXBGDXX",
    "TRNBTNXXINDEA",
    "TRNBTNXXINDNE",
    "TRNINDEAINDNE",
    "TRNINDEAINDNO",
    "TRNINDEAINDWE",
    "TRNINDEAINDSO",
    "TRNINDNEINDNO",
    "TRNINDNOINDWE",
    "TRNINDNONPLXX",
    "TRNINDSOINDWE",
    "TRNINDSOLKAXX",
    "TRNLKAXXMDVXX",
    "TRNMDVXXINDSO",
    "TRNNPLXXBGDXX",
    "TRNINDEANPLXX",
}


@contextmanager
def _working_directory(path: Path):
    previous = Path.cwd()
    os.chdir(path)
    try:
        yield
    finally:
        os.chdir(previous)


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _cell_snapshot(path: Path) -> tuple:
    workbook = load_workbook(path, data_only=False)
    try:
        sheets = []
        for worksheet in workbook.worksheets:
            cells = []
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is None and not cell.has_style:
                        continue
                    cells.append(
                        (
                            cell.coordinate,
                            cell.value,
                            cell.data_type,
                            cell.style_id,
                            cell.number_format,
                        )
                    )
            sheets.append(
                (
                    worksheet.title,
                    worksheet.max_row,
                    worksheet.max_column,
                    worksheet.freeze_panes,
                    tuple(cells),
                )
            )
        return tuple(sheets)
    finally:
        workbook.close()


def _build_projection_fixture(
    path: Path,
    directions: dict[str, str],
) -> dict[tuple[str, str], object]:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = rule.AR_SHEET
    worksheet.freeze_panes = "F2"
    worksheet.append(
        [
            rule.COL_MODE_OP,
            rule.COL_TECH,
            rule.COL_FUEL,
            rule.COL_DIRECTION,
            rule.COL_PMODE,
            2026,
            2027,
            2030,
            "Unrelated metadata",
        ]
    )

    intended: dict[tuple[str, str], object] = {}
    value = 100.0
    for tech, desired in directions.items():
        source, destination = rule.parse_tech_regions(tech)
        disabled_mode = 2 if desired == "forward" else 1
        for mode, input_region in ((1, source), (2, destination)):
            for activity_direction in ("Input", "Output"):
                value += 10.0
                worksheet.append(
                    [
                        mode,
                        tech,
                        f"ELC{input_region}",
                        activity_direction,
                        "Default",
                        value + 1,
                        value + 2,
                        value + 3,
                        f"preserve {tech} mode {mode} {activity_direction}",
                    ]
                )
                row = worksheet.max_row
                if mode == disabled_mode:
                    intended[(rule.AR_SHEET, f"E{row}")] = rule.PMODE_USER
                    intended[(rule.AR_SHEET, f"G{row}")] = 0.0
                    intended[(rule.AR_SHEET, f"H{row}")] = 0.0

    worksheet.append(
        [
            1,
            "PWRWONNPLXX",
            "ELCNPLXX",
            "Input",
            "Generated",
            901.0,
            902.0,
            903.0,
            "unrelated row must remain unchanged",
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()
    return intended


def _build_base_year_fixture(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = rule.AR_SHEET
    worksheet.append(
        [
            rule.COL_MODE_OP,
            rule.COL_TECH,
            rule.COL_BASE_VALUE_I,
            rule.COL_BASE_VALUE_O,
            "Preserved",
        ]
    )
    worksheet.append([1, "TRNBGDXXINDEA", 17.5, 18.5, "base window"])
    workbook.save(path)
    workbook.close()


def _expected_snapshot(before: tuple, intended: dict[tuple[str, str], object]) -> tuple:
    expected_sheets = []
    for title, max_row, max_column, freeze_panes, cells in before:
        expected_cells = []
        for coordinate, value, data_type, style_id, number_format in cells:
            replacement = intended.get((title, coordinate), value)
            expected_cells.append(
                (coordinate, replacement, data_type, style_id, number_format)
            )
        expected_sheets.append(
            (title, max_row, max_column, freeze_panes, tuple(expected_cells))
        )
    return tuple(expected_sheets)


class InterconnectorTaxonomyPathTests(unittest.TestCase):
    def test_authoritative_path_columns_and_classification_are_unchanged(self) -> None:
        paths = path_module.resolve_paths(project_root=REPO_ROOT, environ={})
        taxonomy = paths.scenario_config_root / rule.TECH_TYPES_FILE

        self.assertEqual(taxonomy, AUTHORITATIVE_TAXONOMY)
        self.assertTrue(taxonomy.is_file())
        self.assertFalse(OBSOLETE_TAXONOMY.exists())
        with taxonomy.open("r", encoding="utf-8-sig", newline="") as stream:
            reader = csv.DictReader(stream)
            self.assertEqual(
                reader.fieldnames,
                [rule.TECH_TYPES_CATEGORY_COL, rule.TECH_TYPES_TECH_COL],
            )
        self.assertEqual(rule.load_interconnector_techs(taxonomy), EXPECTED_INTERCONNECTORS)

        configured = rule.load_config(CONTRACTUAL_OVERLAY)["directions"]
        self.assertEqual(len(configured), 9)
        self.assertLessEqual(set(configured), EXPECTED_INTERCONNECTORS)

    def test_configured_rule_from_external_cwd_changes_only_direction_cells(self) -> None:
        configured = rule.load_config(CONTRACTUAL_OVERLAY)["directions"]
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp).resolve()
            input_dir = root / "configured scenario input"
            caller = root / "unrelated external cwd"
            caller.mkdir(parents=True)
            projection = input_dir / rule.AR_PROJ_FILE
            base_year = input_dir / rule.AR_BASE_FILE
            intended = _build_projection_fixture(projection, configured)
            _build_base_year_fixture(base_year)
            before = _cell_snapshot(projection)
            base_hash = _sha256(base_year)

            with (
                mock.patch.dict(
                    os.environ,
                    {
                        path_module.PROJECT_ROOT_ENV: str(REPO_ROOT),
                        path_module.WORKSPACE_ENV: "",
                    },
                ),
                mock.patch.object(
                    rule,
                    "load_interconnector_techs",
                    wraps=rule.load_interconnector_techs,
                ) as load_taxonomy,
                _working_directory(caller),
            ):
                self.assertEqual(
                    path_module.resolve_paths().scenario_config_root,
                    REPO_ROOT / "config" / "scenarios",
                )
                log = rule.run(
                    input_dir,
                    skip_backup=True,
                    yaml_path=CONTRACTUAL_OVERLAY,
                    study_start_year=2027,
                )

            load_taxonomy.assert_called_once_with(AUTHORITATIVE_TAXONOMY)
            self.assertEqual(len(log["projections"]["changes"]), 9)
            self.assertEqual(log["projections"]["warnings"], [])
            self.assertEqual(
                {change["tech"] for change in log["projections"]["changes"]},
                set(configured),
            )
            self.assertEqual(log["base_year"]["skipped"], True)
            self.assertEqual(_sha256(base_year), base_hash)
            self.assertEqual(
                _cell_snapshot(projection),
                _expected_snapshot(before, intended),
            )
            self.assertFalse(OBSOLETE_TAXONOMY.exists())

    def test_missing_authoritative_taxonomy_fails_clearly(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            project = Path(temp).resolve() / "valid project without taxonomy"
            for directory in (
                project / "ostram",
                project / "inputs",
                project / "config" / "scenarios",
                project / "model",
            ):
                directory.mkdir(parents=True)
            (project / "ostram" / "__init__.py").write_text("", encoding="utf-8")
            (project / "environment.yaml").write_text("name: fixture\n", encoding="utf-8")
            missing = project / "config" / "scenarios" / rule.TECH_TYPES_FILE
            obsolete = project / "ostram" / "pipeline" / "scenarios" / "TECH_TYPES.csv"

            with mock.patch.dict(
                os.environ,
                {
                    path_module.PROJECT_ROOT_ENV: str(project),
                    path_module.WORKSPACE_ENV: "",
                },
            ):
                with self.assertRaisesRegex(
                    FileNotFoundError,
                    "Authoritative technology taxonomy not found",
                ) as raised:
                    rule.run(
                        project / "unused input",
                        skip_backup=True,
                        yaml_path=CONTRACTUAL_OVERLAY,
                        study_start_year=2027,
                    )

            self.assertIn(str(missing), str(raised.exception))
            self.assertFalse(missing.exists())
            self.assertFalse(obsolete.exists())


if __name__ == "__main__":
    unittest.main()
