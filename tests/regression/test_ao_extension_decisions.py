from __future__ import annotations

import importlib.util
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[2]
HELPER = (
    ROOT
    / "ostram"
    / "pipeline"
    / "scenarios"
    / "transformations"
    / "ao_extension_decisions.py"
)


def _load_helper():
    spec = importlib.util.spec_from_file_location(
        "apply_ao_extension_decisions_tested",
        HELPER,
    )
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _write_authority(path: Path, rows: list[tuple[object, ...]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "AO_Extension_Decisions"
    worksheet.append(
        [
            "AO_Code_To_Add",
            "Include",
            "Override_Template_AO",
            "Override_Tech.Name_AO",
            "Notes",
        ]
    )
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()


def _write_generated(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "1_Extensions_To_Add"
    worksheet.append(
        [
            "AO_Code_To_Add",
            "Suggested_Template_AO",
            "Add_To_Param",
            "Add_To_AR_Base",
            "Add_To_AR_Proj",
            "Add_To_Demand",
            "Include",
            "Override_Template_AO",
            "Override_Tech.Name_AO",
            "Notes",
        ]
    )
    worksheet.append(
        ["AUTO_SHP", "PWRHYD", "Y", "Y", "Y", "N", "Y", "", "", "mechanical"]
    )
    worksheet.append(
        ["MANUAL_A", "PWRNGS", "Y", "Y", "Y", "N", "", "", "", "generated"]
    )
    worksheet.append(
        ["MANUAL_B", "TRNBASE", "N", "N", "N", "N", "", "", "", "generated"]
    )
    audit = workbook.create_sheet("2_Parameter_Rows_To_Replicate")
    audit.append(["Source_Sheet", "Tech", "Parameter", 2023])
    audit.append(["Secondary_Techs", "MANUAL_A", "CapitalCost", 7])
    audit.append(
        [
            "Interconnector_Params",
            "MANUAL_B",
            "TotalAnnualMinCapacityInvestment",
            1,
        ]
    )
    audit.append(
        [
            "Interconnector_Params",
            "MANUAL_B",
            "MinimumInvestmentClampBoundary",
            1.04,
        ]
    )
    workbook.save(path)
    workbook.close()


def _write_ao_inputs(root: Path) -> None:
    for filename, contains_override in (
        ("A-O_Parametrization.xlsx", True),
        ("A-O_AR_Model_Base_Year.xlsx", True),
        ("A-O_AR_Projections.xlsx", True),
        ("A-O_Demand.xlsx", False),
    ):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["Tech", "Parameter"])
        worksheet.append(
            ["TRN_OVERRIDE" if contains_override else "OTHER", "fixture"]
        )
        workbook.save(root / filename)
        workbook.close()


class AOExtensionDecisionTests(unittest.TestCase):
    def test_only_maintained_decision_fields_are_overlaid(self) -> None:
        helper = _load_helper()
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as temp:
            root = Path(temp)
            authority = root / "authority.xlsx"
            generated = root / "generated.xlsx"
            _write_authority(
                authority,
                [
                    ("MANUAL_A", "Y", "", "", ""),
                    (
                        "MANUAL_B",
                        "Y",
                        "TRN_OVERRIDE",
                        "Named interconnector",
                        "reviewed",
                    ),
                ],
            )
            _write_generated(generated)
            _write_ao_inputs(root)

            count = helper.apply_decisions(generated, authority)

            self.assertEqual(count, 2)
            workbook = load_workbook(generated, read_only=True, data_only=True)
            try:
                worksheet = workbook["1_Extensions_To_Add"]
                rows = list(worksheet.iter_rows(min_row=2, values_only=True))
                self.assertEqual(
                    [row[0] for row in rows],
                    ["AUTO_SHP", "MANUAL_A", "MANUAL_B"],
                )
                self.assertEqual(rows[0], (
                    "AUTO_SHP",
                    "PWRHYD",
                    "Y",
                    "Y",
                    "Y",
                    "N",
                    "Y",
                    None,
                    None,
                    "mechanical",
                ))
                self.assertEqual(rows[1], (
                    "MANUAL_A",
                    "PWRNGS",
                    "Y",
                    "Y",
                    "Y",
                    "N",
                    "Y",
                    None,
                    None,
                    "generated",
                ))
                self.assertEqual(rows[2], (
                    "MANUAL_B",
                    "TRNBASE",
                    "Y",
                    "Y",
                    "Y",
                    "N",
                    "Y",
                    "TRN_OVERRIDE",
                    "Named interconnector",
                    "reviewed",
                ))
                self.assertEqual(
                    list(
                        workbook[
                            "2_Parameter_Rows_To_Replicate"
                        ].iter_rows(values_only=True)
                    ),
                    [
                        ("Source_Sheet", "Tech", "Parameter", 2023),
                        (
                            "Secondary_Techs",
                            "MANUAL_A",
                            "CapitalCost",
                            7,
                        ),
                    ],
                )
            finally:
                workbook.close()

    def test_unknown_or_duplicate_decisions_fail_closed(self) -> None:
        helper = _load_helper()
        cases = (
            [("UNKNOWN", "Y", "", "", "")],
            [
                ("MANUAL_A", "Y", "", "", ""),
                ("MANUAL_A", "Y", "", "", ""),
            ],
        )
        for index, rows in enumerate(cases):
            with self.subTest(index=index):
                with tempfile.TemporaryDirectory(
                    ignore_cleanup_errors=True
                ) as temp:
                    root = Path(temp)
                    authority = root / "authority.xlsx"
                    generated = root / "generated.xlsx"
                    _write_authority(authority, rows)
                    _write_generated(generated)
                    _write_ao_inputs(root)
                    before = generated.read_bytes()

                    with self.assertRaises(ValueError):
                        helper.apply_decisions(generated, authority)

                    self.assertEqual(generated.read_bytes(), before)


if __name__ == "__main__":
    unittest.main()
