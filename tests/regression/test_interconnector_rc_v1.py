from __future__ import annotations

import ast
import hashlib
import importlib.util
import inspect
import json
import shutil
import sys
import tempfile
import unittest
from decimal import Decimal
from pathlib import Path
from unittest import mock

from openpyxl import Workbook, load_workbook


TEST_ROOT = Path(__file__).resolve().parent
REPO_ROOT = TEST_ROOT.parents[1]
T1_ROOT = REPO_ROOT / "t1_confection"
TEMPLATE = T1_ROOT / "A3_process" / "SOASIA_OSeMOSYS_Template_v18.xlsx"
FIX_SCRIPT = T1_ROOT / "A3_process" / "fix_trn_residuals.py"
A3_SCRIPT = T1_ROOT / "A3_process.py"
PATCH_SCRIPT = T1_ROOT / "sensitivity_expansion" / "apply_patches.py"
TXCAP_CONFIG = T1_ROOT / "A3_process" / "rules_scripts" / "configs" / "B_Opt_TxCap150" / "patches.json"
LINKFREEZE_CONFIG = T1_ROOT / "A3_process" / "rules_scripts" / "configs" / "B_Opt_LinkFreeze" / "patches.json"
SCENARIO_REGISTRY = TEST_ROOT / "scenarios.yaml"

YEARS = tuple(range(2023, 2051))
STUDY_YEARS = tuple(range(2027, 2051))
AUTHORITY = {
    "TRNBGDXXINDEA": Decimal("2.496"),
    "TRNBGDXXINDNE": Decimal("0.160"),
    "TRNBTNXXBGDXX": Decimal("0"),
    "TRNBTNXXINDEA": Decimal("5.992"),
    "TRNBTNXXINDNE": Decimal("0.170"),
    "TRNINDEAINDNE": Decimal("2.86"),
    "TRNINDEAINDNO": Decimal("19.53"),
    "TRNINDEAINDSO": Decimal("5.53"),
    "TRNINDEAINDWE": Decimal("21.19"),
    "TRNINDEANPLXX": Decimal("1.200"),
    "TRNINDNEINDNO": Decimal("3.00"),
    "TRNINDNOINDWE": Decimal("29.22"),
    "TRNINDNONPLXX": Decimal("0.500"),
    "TRNINDSOINDWE": Decimal("15.12"),
    "TRNINDSOLKAXX": Decimal("0"),
    "TRNLKAXXMDVXX": Decimal("0"),
    "TRNMDVXXINDSO": Decimal("0"),
    "TRNNPLXXBGDXX": Decimal("0"),
}
ZERO_TECHS = frozenset(tech for tech, value in AUTHORITY.items() if value == 0)
AUTHORITY_SEMANTIC_SHA256 = "6c4420017b4a2df0b0e4ab6cc11f0bb45c79aa139bed858bdea5fec1aa54584b"
TXCAP_FLOORS = {
    "TRNBGDXXINDEA": Decimal("8.236183089312995"),
    "TRNBGDXXINDNE": Decimal("0.32"),
    "TRNBTNXXBGDXX": Decimal("0"),
    "TRNBTNXXINDEA": Decimal("5.526"),
    "TRNBTNXXINDNE": Decimal("0.11"),
    "TRNINDEANPLXX": Decimal("1.3"),
    "TRNINDNONPLXX": Decimal("1.2"),
    "TRNINDSOLKAXX": Decimal("0"),
    "TRNLKAXXMDVXX": Decimal("0"),
    "TRNMDVXXINDSO": Decimal("0"),
    "TRNNPLXXBGDXX": Decimal("0.04"),
}
CAP_TARGETS = {"TotalAnnualMaxCapacity", "TotalAnnualMaxCapacityInvestment"}


def _load_module(path: Path, label: str):
    name = f"_ostram_rc_v1_{label}"
    spec = importlib.util.spec_from_file_location(name, path)
    if spec is None or spec.loader is None:
        raise AssertionError(f"could not load module spec for {path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[name] = module
    try:
        spec.loader.exec_module(module)
    finally:
        sys.modules.pop(name, None)
    return module


def _decimal(value) -> Decimal:
    if value is None or isinstance(value, bool):
        raise AssertionError(f"expected numeric value, got {value!r}")
    try:
        result = Decimal(str(value))
    except Exception as exc:
        raise AssertionError(f"expected numeric value, got {value!r}") from exc
    if not result.is_finite():
        raise AssertionError(f"expected finite value, got {value!r}")
    return result


def _normal_decimal(value) -> str:
    number = _decimal(value)
    return "0" if number == 0 else format(number.normalize(), "f")


def _authority_digest(values: dict[str, dict[int, object]]) -> str:
    payload = "".join(
        f"{tech}|{year}|{_normal_decimal(values[tech][year])}\n"
        for tech in sorted(values)
        for year in YEARS
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _read_authority_rows(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=True)
    try:
        if "Interconnector_Params" not in workbook.sheetnames:
            raise AssertionError("v18 workbook is missing Interconnector_Params")
        worksheet = workbook["Interconnector_Params"]
        headers = {cell.value: cell.column for cell in worksheet[1] if cell.value is not None}
        required = {
            "scenario", "Tech.ID", "Tech", "Tech.Name", "Parameter.ID",
            "Parameter", "Unit", "Projection.Mode", "Projection.Parameter",
            "Source", *YEARS,
        }
        missing_headers = required - set(headers)
        if missing_headers:
            raise AssertionError(f"missing authority headers: {sorted(missing_headers)!r}")
        values: dict[str, dict[int, object]] = {}
        metadata: dict[str, dict[str, object]] = {}
        duplicates: list[str] = []
        for row in range(2, worksheet.max_row + 1):
            tech = worksheet.cell(row, headers["Tech"]).value
            parameter = worksheet.cell(row, headers["Parameter"]).value
            if tech not in AUTHORITY or parameter != "ResidualCapacity":
                continue
            if tech in values:
                duplicates.append(tech)
                continue
            values[tech] = {year: worksheet.cell(row, headers[year]).value for year in YEARS}
            metadata[tech] = {
                key: worksheet.cell(row, headers[key]).value
                for key in (
                    "scenario", "Tech.ID", "Tech.Name", "Parameter.ID", "Unit",
                    "Projection.Mode", "Projection.Parameter", "Source",
                )
            }
        return values, metadata, duplicates
    finally:
        workbook.close()


def _snapshot_parameter(worksheet, parameter: str):
    headers = {cell.value: cell.column for cell in worksheet[1] if cell.value is not None}
    result = {}
    for row in range(2, worksheet.max_row + 1):
        if worksheet.cell(row, headers["Parameter"]).value != parameter:
            continue
        tech = worksheet.cell(row, headers["Tech"]).value
        result[tech] = tuple(worksheet.cell(row, headers[year]).value for year in YEARS)
    return result


def _secondary_fixture():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Secondary Techs"
    worksheet.append([
        "Tech.ID", "Tech", "Tech.Name", "Parameter.ID", "Parameter", "Unit",
        "Projection.Mode", "Projection.Parameter", *YEARS,
    ])
    for index, tech in enumerate(sorted(AUTHORITY), start=1):
        residual = [0.110] * len(YEARS)
        if tech == "TRNBTNXXINDNE":
            residual = [0.110 if year < 2033 else 0.150 for year in YEARS]
        worksheet.append([
            index, tech, tech, 3, "ResidualCapacity", "GW", "User defined", 0,
            *residual,
        ])
        minimum = [None] * len(YEARS)
        if tech == "TRNBTNXXINDNE":
            minimum[YEARS.index(2024)] = 0.0
            minimum[YEARS.index(2033)] = 0.02
        worksheet.append([
            index, tech, tech, 4, "TotalAnnualMinCapacityInvestment", "GW",
            "User defined", 0, *minimum,
        ])
    return workbook, worksheet


def _patch_fixture(tech: str, residual: object, target_value: object = -1.0):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Secondary Techs"
    worksheet.append([
        "Tech.ID", "Tech", "Tech.Name", "Parameter.ID", "Parameter", "Unit",
        "Projection.Mode", "Projection.Parameter", *YEARS,
    ])
    worksheet.append([
        1, tech, tech, 3, "ResidualCapacity", "GW", "User defined", 0,
        *([residual] * len(YEARS)),
    ])
    for parameter in sorted(CAP_TARGETS | {"TotalAnnualMinCapacityInvestment"}):
        worksheet.append([
            1, tech, tech, 4, parameter, "GW", "User defined", 0,
            *([target_value] * len(YEARS)),
        ])
    return workbook, worksheet


def _patch_log():
    return {"cells": [], "rows_created": [], "skipped": []}


def _canonicalize_sparse_rc(serialized):
    expected_keys = {(tech, year) for tech in AUTHORITY for year in YEARS}
    extra = set(serialized) - expected_keys
    if extra:
        raise AssertionError(f"unexpected ResidualCapacity keys: {sorted(extra)!r}")
    canonical = {}
    for tech, year in sorted(expected_keys):
        key = (tech, year)
        if key in serialized:
            canonical[key] = _decimal(serialized[key])
        elif tech in ZERO_TECHS:
            canonical[key] = Decimal("0")
        else:
            raise AssertionError(f"missing nonzero ResidualCapacity key: {key!r}")
    return canonical


def _assert_finite_caps_not_below_residual(residuals, caps):
    for key, residual in residuals.items():
        if key not in caps:
            raise AssertionError(f"missing capacity key: {key!r}")
        cap = _decimal(caps[key])
        if cap == Decimal("-1"):
            continue
        if cap < 0:
            raise AssertionError(f"invalid finite capacity for {key!r}: {cap}")
        if cap < _decimal(residual):
            raise AssertionError(
                f"finite capacity below residual for {key!r}: {cap} < {residual}"
            )


class InterconnectorAuthorityWorkbookTests(unittest.TestCase):
    def test_v18_contains_exact_dense_18_by_28_authority_and_digest(self) -> None:
        values, metadata, duplicates = _read_authority_rows(TEMPLATE)
        self.assertEqual(duplicates, [])
        self.assertEqual(set(values), set(AUTHORITY))
        self.assertEqual(sum(len(profile) for profile in values.values()), 18 * 28)
        for tech, expected in AUTHORITY.items():
            with self.subTest(tech=tech):
                self.assertEqual(set(values[tech]), set(YEARS))
                self.assertEqual({_decimal(v) for v in values[tech].values()}, {expected})
                self.assertEqual(metadata[tech]["scenario"], "BAU")
                self.assertIsInstance(metadata[tech]["Tech.ID"], int)
                self.assertIsInstance(metadata[tech]["Tech.Name"], str)
                self.assertEqual(metadata[tech]["Parameter.ID"], 3)
                self.assertEqual(metadata[tech]["Unit"], "GW")
                self.assertEqual(metadata[tech]["Projection.Mode"], "User defined")
                self.assertEqual(metadata[tech]["Projection.Parameter"], 0)
                self.assertIn(metadata[tech]["Source"], (None, ""))
        self.assertEqual(_authority_digest(values), AUTHORITY_SEMANTIC_SHA256)

    def test_public_authority_loader_accepts_only_exact_rc_v1(self) -> None:
        module = _load_module(FIX_SCRIPT, "authority_loader")
        loaded = module.load_rc_authority(TEMPLATE)
        self.assertEqual(set(module.AUTHORITY_TECHS), set(AUTHORITY))
        self.assertEqual(tuple(module.AUTHORITY_YEARS), YEARS)
        self.assertEqual(module.AUTHORITY_SEMANTIC_SHA256, AUTHORITY_SEMANTIC_SHA256)
        self.assertEqual(set(loaded), set(AUTHORITY))
        for tech, expected in AUTHORITY.items():
            self.assertEqual(set(loaded[tech]), set(YEARS))
            self.assertEqual({_decimal(v) for v in loaded[tech].values()}, {expected})

        with tempfile.TemporaryDirectory() as tmp:
            corrupt = Path(tmp) / "corrupt_authority.xlsx"
            shutil.copy2(TEMPLATE, corrupt)
            workbook = load_workbook(corrupt)
            worksheet = workbook["Interconnector_Params"]
            headers = {cell.value: cell.column for cell in worksheet[1] if cell.value is not None}
            for row in range(2, worksheet.max_row + 1):
                if (
                    worksheet.cell(row, headers["Tech"]).value == "TRNBTNXXINDNE"
                    and worksheet.cell(row, headers["Parameter"]).value == "ResidualCapacity"
                ):
                    worksheet.cell(row, headers[2023]).value = 0.171
                    break
            workbook.save(corrupt)
            workbook.close()
            with self.assertRaises(ValueError):
                module.load_rc_authority(corrupt)


class InterconnectorRuntimeRouteTests(unittest.TestCase):
    def test_a3_uses_materialized_v18_authority_before_unchanged_writers(self) -> None:
        source = A3_SCRIPT.read_text(encoding="utf-8-sig")
        tree = ast.parse(source, filename=str(A3_SCRIPT))
        stage = next(
            node for node in tree.body
            if isinstance(node, ast.FunctionDef) and node.name == "stage_3_fix_2"
        )
        stage_source = ast.get_source_segment(source, stage)
        self.assertIsNotNone(stage_source)
        assert stage_source is not None
        fix_at = stage_source.index('label="fix_trn_residuals.py"')
        clear_at = stage_source.index('label="clear_stale_unbinding_caps.py"')
        cap_at = stage_source.index('label="cap_trn_to_residual.py"')
        self.assertLess(fix_at, clear_at)
        self.assertLess(clear_at, cap_at)
        self.assertIn('"--reference", "A-O_Parametrization_NATY.xlsx"', stage_source)
        self.assertIn('os.environ.get("OSTRAM_TEMPLATE_PATH")', stage_source)
        self.assertIn('"--authority", authority_path', stage_source)

        module = _load_module(FIX_SCRIPT, "runtime_route")
        run_source = inspect.getsource(module.run_fix)
        self.assertIn("write_residual=False", run_source)
        self.assertLess(run_source.index("load_rc_authority("), run_source.index("load_workbook("))
        self.assertLess(run_source.rindex("apply_fix("), run_source.index("apply_rc_authority("))
        self.assertEqual(run_source.count("apply_rc_authority("), 1)

    def test_rc_suppression_preserves_complete_minimum_family(self) -> None:
        module = _load_module(FIX_SCRIPT, "minimum_isolation")
        baseline_book, baseline_sheet = _secondary_fixture()
        candidate_book, candidate_sheet = _secondary_fixture()
        profile = {year: 0.110 if year < 2033 else 0.150 for year in YEARS}
        plan = module.TechFix(
            tech="TRNBTNXXINDNE",
            base_value=0.110,
            original_profile=profile,
            commissionings=[
                module.Commissioning(
                    tech="TRNBTNXXINDNE", year=2033, capacity_added=0.040
                )
            ],
            flatten_only=True,
            base_source="reference",
            profile_source="reference",
        )
        module.apply_fix(
            baseline_sheet, plan, mode="min", cutoff_year=2023,
            write_residual=True,
        )
        module.apply_fix(
            candidate_sheet, plan, mode="min", cutoff_year=2023,
            write_residual=False,
        )
        authority = {
            tech: {year: float(value) for year in YEARS}
            for tech, value in AUTHORITY.items()
        }
        module.apply_rc_authority(candidate_sheet, authority)
        self.assertEqual(
            _snapshot_parameter(candidate_sheet, module.MIN_INV_PARAM),
            _snapshot_parameter(baseline_sheet, module.MIN_INV_PARAM),
        )
        candidate_rc = _snapshot_parameter(candidate_sheet, module.RESIDUAL_PARAM)
        for tech, expected in AUTHORITY.items():
            self.assertEqual({_decimal(v) for v in candidate_rc[tech]}, {expected})
        baseline_book.close()
        candidate_book.close()

    def test_protected_cap_and_relax_implementations_are_byte_identical(self) -> None:
        protected = {
            T1_ROOT / "A3_process" / "cap_trn_to_residual.py":
                "f9f876d1e58cc8dd1339aea703477fe7a85bef776ad227ab99d972d25f7c6a36",
            T1_ROOT / "A3_process" / "rules_scripts" / "relax_interconnectors.py":
                "e496d54157459e7da2eb460d0cc76264eeee26a386e6a8c811cad1285424fbb7",
        }
        for path, expected in protected.items():
            with self.subTest(path=path.name):
                self.assertEqual(hashlib.sha256(path.read_bytes()).hexdigest(), expected)

    def test_writer_order_and_scenario_predicates_remain_unchanged(self) -> None:
        workbook = load_workbook(TEMPLATE, read_only=True, data_only=True)
        try:
            control = {
                row[0]: row[1:]
                for row in workbook["Control"].iter_rows(min_row=2, values_only=True)
                if row[0]
            }
        finally:
            workbook.close()
        self.assertEqual(
            control["A_Calibrated_BAU"][:3],
            (
                True,
                "set_retirement_schedule.py, add_max_cap_investment_lid_rule.py, "
                "set_min_capacity_floors.py, relax_interconnectors.py",
                "BAU",
            ),
        )
        self.assertEqual(
            control["B_Optimised_VRE"][:3],
            (
                True,
                "set_retirement_schedule.py, add_max_cap_investment_lid_rule.py, "
                "relax_interconnectors.py, add_storage_min_investment.py",
                None,
            ),
        )
        registry = json.loads(SCENARIO_REGISTRY.read_text(encoding="utf-8"))
        scenarios = {entry["name"]: entry for entry in registry["scenarios"]}
        self.assertEqual(
            scenarios["B_Opt_LinkFreeze"],
            {
                "name": "B_Opt_LinkFreeze",
                "tier": "superseded-protected",
                "source_scenario": "B_Optimised_VRE",
                "recipe": "link-freeze",
                "cleanup_acceptance": False,
                "cleanup_exclusion_reason": (
                    "superseded provenance scenario; no A2, otoole, compiled, "
                    "or direct output evidence"
                ),
            },
        )
        self.assertEqual(
            scenarios["B_Opt_TxCap150"],
            {
                "name": "B_Opt_TxCap150",
                "tier": "final",
                "source_scenario": "B_Optimised_VRE",
                "recipe": "transmission-capex-150",
                "cleanup_acceptance": True,
            },
        )
        linkfreeze = json.loads(LINKFREEZE_CONFIG.read_text(encoding="utf-8"))
        self.assertEqual(linkfreeze["base_scenario"], "B_Optimised_VRE")
        self.assertEqual(
            linkfreeze["frozen_corridors"],
            ["TRNBGDXXINDEA", "TRNNPLXXBGDXX"],
        )
        edits = linkfreeze["edits"]
        self.assertEqual(len(edits), 6)
        self.assertEqual(
            [(e["tech"], e["param"], e["op"]) for e in edits],
            [
                ("TRNBGDXXINDEA", "TotalAnnualMaxCapacity", "set_to_residual"),
                ("TRNBGDXXINDEA", "TotalAnnualMaxCapacityInvestment", "set_to_residual"),
                ("TRNNPLXXBGDXX", "TotalAnnualMaxCapacity", "set_to_residual"),
                ("TRNNPLXXBGDXX", "TotalAnnualMaxCapacityInvestment", "set_to_residual"),
                ("TRNBGDXXINDEA", "TotalAnnualMinCapacityInvestment", "clamp_to_residual"),
                ("TRNNPLXXBGDXX", "TotalAnnualMinCapacityInvestment", "clamp_to_residual"),
            ],
        )
        for edit in edits[:4]:
            self.assertEqual(edit.get("residual_source", "effective"), "effective")
        for edit in edits[4:]:
            self.assertEqual(edit.get("residual_source"), "legacy_min_reference")


class TxCap150FormulaTests(unittest.TestCase):
    def test_config_preserves_formula_domain_targets_and_non_cap_edits(self) -> None:
        config = json.loads(TXCAP_CONFIG.read_text(encoding="utf-8"))
        self.assertEqual(config["scenario"], "B_Opt_TxCap150")
        self.assertEqual(config["base_scenario"], "B_Optimised_VRE")
        dynamic = [
            edit for edit in config["edits"]
            if edit.get("op") == "set_to_residual_factor_floor"
        ]
        self.assertEqual(len(dynamic), 22)
        self.assertEqual({edit["tech"] for edit in dynamic}, set(TXCAP_FLOORS))
        self.assertEqual({edit["param"] for edit in dynamic}, CAP_TARGETS)
        self.assertEqual(
            {(edit["tech"], edit["param"]) for edit in dynamic},
            {(tech, param) for tech in TXCAP_FLOORS for param in CAP_TARGETS},
        )
        for edit in dynamic:
            with self.subTest(tech=edit["tech"], parameter=edit["param"]):
                self.assertEqual(_decimal(edit["factor"]), Decimal("1.5"))
                self.assertEqual(edit["residual_year"], 2023)
                self.assertEqual(tuple(edit["years"]), STUDY_YEARS)
                self.assertEqual(
                    _decimal(edit["base_window_floor"]),
                    TXCAP_FLOORS[edit["tech"]],
                )

        minimum_edits = [
            edit for edit in config["edits"]
            if edit["param"] == "TotalAnnualMinCapacityInvestment"
        ]
        activity_edits = [
            edit for edit in config["edits"]
            if edit["param"] == "TotalTechnologyAnnualActivityUpperLimit"
        ]
        def canonical(rows):
            return json.dumps(
                rows, sort_keys=True, separators=(",", ":"), ensure_ascii=False
            ).encode("utf-8")
        self.assertEqual(len(minimum_edits), 8)
        self.assertEqual(
            hashlib.sha256(canonical(minimum_edits)).hexdigest(),
            "51c2c7613eff840f9a20193a06057026070a81f08eb58777386d7ba7c184f7b8",
        )
        self.assertEqual(len(activity_edits), 10)
        self.assertEqual(
            hashlib.sha256(canonical(activity_edits)).hexdigest(),
            "19c1d41469c19144a40764c8033533eaafb01f8b2bd85381d17e368f233caef5",
        )

    def test_dynamic_formula_reads_effective_rc_for_both_targets(self) -> None:
        module = _load_module(PATCH_SCRIPT, "txcap_formula")
        config = json.loads(TXCAP_CONFIG.read_text(encoding="utf-8"))
        edits = [
            edit for edit in config["edits"]
            if edit.get("op") == "set_to_residual_factor_floor"
            and edit["tech"] == "TRNBTNXXINDNE"
        ]
        workbook, worksheet = _patch_fixture("TRNBTNXXINDNE", 0.170)
        columns, year_columns = module.scan_columns(worksheet)
        log = _patch_log()
        for edit in edits:
            module.apply_edit(worksheet, columns, year_columns, edit, log)
        for parameter in CAP_TARGETS:
            row = module.find_row(worksheet, columns, "TRNBTNXXINDNE", parameter)
            self.assertIsNotNone(row)
            assert row is not None
            self.assertEqual(worksheet.cell(row, year_columns[2026]).value, -1.0)
            for year in STUDY_YEARS:
                self.assertEqual(
                    _decimal(worksheet.cell(row, year_columns[year]).value),
                    Decimal("0.255"),
                )
        self.assertEqual(len(log["cells"]), 2 * len(STUDY_YEARS))
        workbook.close()

    def test_dynamic_formula_fails_closed_on_missing_or_invalid_rc(self) -> None:
        module = _load_module(PATCH_SCRIPT, "txcap_fail_closed")
        edit = {
            "sheet": "Secondary Techs",
            "tech": "TRNBTNXXINDNE",
            "param": "TotalAnnualMaxCapacity",
            "op": "set_to_residual_factor_floor",
            "factor": 1.5,
            "residual_year": 2023,
            "base_window_floor": 0.11,
            "years": list(STUDY_YEARS),
        }
        for invalid in (None, -0.1, float("nan"), float("inf")):
            with self.subTest(invalid=invalid):
                workbook, worksheet = _patch_fixture("TRNBTNXXINDNE", invalid)
                columns, year_columns = module.scan_columns(worksheet)
                with self.assertRaises(ValueError):
                    module.apply_edit(
                        worksheet, columns, year_columns, edit, _patch_log()
                    )
                workbook.close()

    def test_linkfreeze_minimum_clamp_uses_legacy_boundary_only(self) -> None:
        module = _load_module(PATCH_SCRIPT, "linkfreeze_minimum")
        config = json.loads(LINKFREEZE_CONFIG.read_text(encoding="utf-8"))
        edits = {
            edit["tech"]: edit for edit in config["edits"]
            if edit["param"] == "TotalAnnualMinCapacityInvestment"
        }
        cases = (
            ("TRNBGDXXINDEA", 2.496, 2.5),
            ("TRNNPLXXBGDXX", 0.0, 0.02),
        )
        legacy = {
            "TRNBGDXXINDEA": {year: 2.5 for year in YEARS},
            "TRNNPLXXBGDXX": {year: 0.04 for year in YEARS},
        }
        with mock.patch.object(
            module, "load_legacy_min_residuals", return_value=legacy
        ):
            for tech, effective_rc, minimum in cases:
                with self.subTest(tech=tech):
                    workbook, worksheet = _patch_fixture(tech, effective_rc, minimum)
                    columns, year_columns = module.scan_columns(worksheet)
                    module.apply_edit(
                        worksheet, columns, year_columns, edits[tech], _patch_log()
                    )
                    row = module.find_row(
                        worksheet, columns, tech,
                        "TotalAnnualMinCapacityInvestment",
                    )
                    self.assertIsNotNone(row)
                    assert row is not None
                    self.assertEqual(
                        _decimal(worksheet.cell(row, year_columns[2033]).value),
                        Decimal(str(minimum)),
                    )
                    workbook.close()


class SemanticComparisonBehaviorTests(unittest.TestCase):
    def test_sparse_zero_serialization_canonicalizes_only_approved_zeros(self) -> None:
        serialized = {
            (tech, year): value
            for tech, value in AUTHORITY.items()
            for year in YEARS
            if tech not in ZERO_TECHS
        }
        canonical = _canonicalize_sparse_rc(serialized)
        self.assertEqual(len(serialized), 13 * 28)
        self.assertEqual(len(canonical), 18 * 28)
        for tech, expected in AUTHORITY.items():
            for year in YEARS:
                self.assertEqual(canonical[(tech, year)], expected)
        missing_nonzero = dict(serialized)
        missing_nonzero.pop(("TRNBTNXXINDNE", 2023))
        with self.assertRaisesRegex(AssertionError, "missing nonzero"):
            _canonicalize_sparse_rc(missing_nonzero)

    def test_minus_one_is_the_only_unbounded_cap_sentinel(self) -> None:
        residuals = {("TRNBTNXXINDNE", 2027): Decimal("0.170")}
        _assert_finite_caps_not_below_residual(
            residuals, {("TRNBTNXXINDNE", 2027): Decimal("-1")}
        )
        _assert_finite_caps_not_below_residual(
            residuals, {("TRNBTNXXINDNE", 2027): Decimal("0.170")}
        )
        with self.assertRaisesRegex(AssertionError, "below residual"):
            _assert_finite_caps_not_below_residual(
                residuals, {("TRNBTNXXINDNE", 2027): Decimal("0.169")}
            )
        with self.assertRaisesRegex(AssertionError, "invalid finite capacity"):
            _assert_finite_caps_not_below_residual(
                residuals, {("TRNBTNXXINDNE", 2027): Decimal("-1.1")}
            )


if __name__ == "__main__":
    unittest.main()
