from __future__ import annotations

import importlib.util
import sys
import unittest
from pathlib import Path

from openpyxl import Workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
SCRIPT = (
    REPO_ROOT
    / "ostram"
    / "pipeline"
    / "scenarios"
    / "rules"
    / "set_vre_targets.py"
)
PRODUCTION_CONFIG = (
    REPO_ROOT
    / "config"
    / "scenarios"
    / "C_Target_VRE"
    / "set_vre_targets.yaml"
)


def _load_module():
    name = "_ostram_set_vre_targets_boundary_cap_test"
    spec = importlib.util.spec_from_file_location(name, SCRIPT)
    if spec is None or spec.loader is None:
        raise AssertionError(f"cannot import {SCRIPT}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[name] = module
    try:
        spec.loader.exec_module(module)
    finally:
        sys.modules.pop(name, None)
    return module


VRE = _load_module()


class SetVreTargetsBoundaryCapTests(unittest.TestCase):
    def test_production_config_has_exact_transition_cap_series(self):
        config = VRE.load_config(PRODUCTION_CONFIG)

        self.assertEqual(
            config["activity_lower_caps"],
            [
                {
                    "region": "GLOBAL",
                    "technology": "PWRWONINDWE",
                    "year": 2027,
                    "maximum": 227.414,
                },
                {
                    "region": "GLOBAL",
                    "technology": "PWRWONINDWE",
                    "year": 2028,
                    "maximum": 255.078,
                },
            ],
        )

    def test_caps_reduce_raw_values_but_never_increase_lower_values(self):
        caps = [
            {
                "region": "GLOBAL",
                "technology": "PWRWONINDWE",
                "year": 2027,
                "maximum": 227.414,
            },
            {
                "region": "GLOBAL",
                "technology": "PWRWONINDWE",
                "year": 2028,
                "maximum": 255.078,
            },
        ]

        capped_2027, evidence_2027 = VRE.cap_activity_lower_value(
            236.974,
            region="GLOBAL",
            technology="PWRWONINDWE",
            year=2027,
            caps=caps,
        )
        capped_2028, evidence_2028 = VRE.cap_activity_lower_value(
            260.596,
            region="GLOBAL",
            technology="PWRWONINDWE",
            year=2028,
            caps=caps,
        )

        self.assertEqual(capped_2027, 227.414)
        self.assertEqual(evidence_2027["raw_derived_value"], 236.974)
        self.assertTrue(evidence_2027["applied"])
        self.assertEqual(capped_2028, 255.078)
        self.assertEqual(evidence_2028["raw_derived_value"], 260.596)
        self.assertTrue(evidence_2028["applied"])

        for year, value in ((2027, 220.0), (2028, 250.0)):
            with self.subTest(year=year):
                already_lower, evidence = VRE.cap_activity_lower_value(
                    value,
                    region="GLOBAL",
                    technology="PWRWONINDWE",
                    year=year,
                    caps=caps,
                )
                self.assertEqual(already_lower, value)
                self.assertFalse(evidence["applied"])

    def test_only_exact_transition_keys_are_eligible(self):
        caps = VRE.load_config(PRODUCTION_CONFIG)["activity_lower_caps"]
        ineligible = [
            ("GLOBAL", "PWRWONINDWE", 2029),
            ("GLOBAL", "PWRWONINDWE", 2050),
            ("GLOBAL", "PWRSPVINDSO", 2027),
            ("GLOBAL", "PWRWONINDNO", 2028),
            ("INDWE", "PWRWONINDWE", 2028),
        ]

        for region, technology, year in ineligible:
            with self.subTest(
                region=region,
                technology=technology,
                year=year,
            ):
                unchanged, evidence = VRE.cap_activity_lower_value(
                    999.0,
                    region=region,
                    technology=technology,
                    year=year,
                    caps=caps,
                )
                self.assertEqual(unchanged, 999.0)
                self.assertIsNone(evidence)

    def test_apply_changes_only_the_authorized_parameter_key(self):
        def make_sheet():
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Primary Techs"
            sheet.append([
                "Tech", "Parameter", "Projection.Mode", 2027, 2028, 2029,
            ])
            sheet.append([
                "PWRWONINDWE",
                VRE.ACTIVITY_LOWER_PARAM,
                VRE.PROJ_MODE_EMPTY,
                0.0,
                0.0,
                0.0,
            ])
            sheet.append([
                "PWRWONINDWE",
                VRE.ACTIVITY_UPPER_PARAM,
                VRE.PROJ_MODE_USER,
                999.0,
                999.0,
                999.0,
            ])
            sheet.append([
                "PWRSPVINDWE",
                VRE.ACTIVITY_LOWER_PARAM,
                VRE.PROJ_MODE_USER,
                12.0,
                13.0,
                14.0,
            ])
            return sheet

        def snapshot(sheet):
            return {
                (sheet.cell(row=row, column=1).value,
                 sheet.cell(row=row, column=2).value, year):
                sheet.cell(row=row, column=col).value
                for row in range(2, sheet.max_row + 1)
                for year, col in {2027: 4, 2028: 5, 2029: 6}.items()
            }

        config = {
            "constraint_type": "activity",
            "max_floor_share": 1.0,
            "targets": [{
                "cr": "INDWE",
                "tech": "PWRWON*",
                "schedule": {2027: 1.0, 2028: 1.0, 2029: 1.0},
                "cap_envelope": False,
            }],
        }
        total_prod = {
            ("INDWE", 2027): 236.974,
            ("INDWE", 2028): 260.596,
            ("INDWE", 2029): 284.453,
        }
        gen_techs = {"PWRWONINDWE", "PWRSPVINDWE"}
        year_cols = {2027: 4, 2028: 5, 2029: 6}

        uncapped_sheet = make_sheet()
        VRE.apply_vre_targets(
            uncapped_sheet,
            {**config, "activity_lower_caps": []},
            total_prod,
            gen_techs,
            year_cols,
        )
        capped_sheet = make_sheet()
        log = VRE.apply_vre_targets(
            capped_sheet,
            {
                **config,
                "activity_lower_caps": VRE.load_config(
                    PRODUCTION_CONFIG
                )["activity_lower_caps"],
            },
            total_prod,
            gen_techs,
            year_cols,
        )
        uncapped = snapshot(uncapped_sheet)
        capped = snapshot(capped_sheet)

        changed_keys = {
            key for key in uncapped if uncapped[key] != capped[key]
        }
        authorized_keys = {
            ("PWRWONINDWE", VRE.ACTIVITY_LOWER_PARAM, 2027),
            ("PWRWONINDWE", VRE.ACTIVITY_LOWER_PARAM, 2028),
        }
        self.assertEqual(changed_keys, authorized_keys)
        self.assertEqual(
            capped[("PWRWONINDWE", VRE.ACTIVITY_LOWER_PARAM, 2027)],
            227.414,
        )
        self.assertEqual(
            capped[("PWRWONINDWE", VRE.ACTIVITY_LOWER_PARAM, 2028)],
            255.078,
        )
        self.assertEqual(
            capped[("PWRWONINDWE", VRE.ACTIVITY_LOWER_PARAM, 2029)],
            284.453,
        )
        self.assertEqual(
            log["activity_lower_caps"],
            [
                {
                    "region": "GLOBAL",
                    "technology": "PWRWONINDWE",
                    "year": 2027,
                    "raw_derived_value": 236.974,
                    "maximum": 227.414,
                    "final_value": 227.414,
                    "applied": True,
                },
                {
                    "region": "GLOBAL",
                    "technology": "PWRWONINDWE",
                    "year": 2028,
                    "raw_derived_value": 260.596,
                    "maximum": 255.078,
                    "final_value": 255.078,
                    "applied": True,
                },
            ],
        )


if __name__ == "__main__":
    unittest.main()
