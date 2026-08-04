from __future__ import annotations

import hashlib
import io
import os
from pathlib import Path
import sys
import tempfile
import unittest
from contextlib import redirect_stderr, redirect_stdout
from unittest import mock

from openpyxl import Workbook

from ostram import paths as path_module
from ostram.pipeline.execution.patches import reserve_margin_repair_xlsx


FAILED_WARNING_NAME = (
    "Pre_processed_A_Calibrated_BAU_0_StorageDelayN5_OpenBCK_"
    "RMCarefulXLSX.warnings.txt"
)
WARNING_PARENT_UNITS = 177
FAILED_WARNING_UNITS = 260


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _warning_parent(root: Path) -> Path:
    prefix = root / "governed acceptance é Ω"
    suffix = Path("execution") / "Executables" / "A_Calibrated_BAU_0"
    for filler_units in range(1, 200):
        parent = prefix / ("p" * filler_units) / suffix
        if path_module.windows_path_units(parent) == WARNING_PARENT_UNITS:
            parent.mkdir(parents=True)
            return parent
    raise AssertionError(f"could not build exact warning parent below {root}")


def _exact_failed_warning_path(root: Path) -> Path:
    desired = _warning_parent(root) / FAILED_WARNING_NAME
    if path_module.windows_path_units(desired) != FAILED_WARNING_UNITS:
        raise AssertionError(desired)
    return desired


def _write_fallback_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "fallbacks"
    worksheet.append(
        [
            "CR",
            "TECH_PREFIX",
            "YEAR",
            "TotalAnnualMaxCapacity",
            "TotalAnnualMaxCapacityInvestment",
        ]
    )
    workbook.save(path)
    workbook.close()


def _write_datafile(path: Path) -> None:
    path.write_text(
        """set YEAR := 2023 ;
set TECHNOLOGY := PWRNGSTSTXX ;
param default 0 : ResidualCapacity :=
GLOBAL PWRNGSTSTXX 2023 0
;
param default 0 : TotalAnnualMinCapacity :=
GLOBAL PWRNGSTSTXX 2023 0
;
param default 0 : TotalAnnualMinCapacityInvestment :=
GLOBAL PWRNGSTSTXX 2023 0
;
param default 0 : TotalAnnualMaxCapacity :=
GLOBAL PWRNGSTSTXX 2023 1
;
param default 0 : TotalAnnualMaxCapacityInvestment :=
GLOBAL PWRNGSTSTXX 2023 0
;
""",
        encoding="utf-8",
    )


def _run_patcher(
    input_path: Path,
    output_path: Path,
    workbook: Path,
    warnings: Path,
) -> None:
    opened_workbooks = []
    real_load_workbook = reserve_margin_repair_xlsx.load_workbook

    def tracked_load_workbook(*args, **kwargs):
        loaded = real_load_workbook(*args, **kwargs)
        opened_workbooks.append(loaded)
        return loaded

    arguments = [
        "reserve_margin_repair_xlsx",
        str(input_path),
        "-o",
        str(output_path),
        "--fallback-xlsx",
        str(workbook),
        "--xlsx-sheet",
        "fallbacks",
        "--warnings-file",
        str(warnings),
    ]
    try:
        with (
            mock.patch.object(sys, "argv", arguments),
            mock.patch.object(
                reserve_margin_repair_xlsx,
                "load_workbook",
                side_effect=tracked_load_workbook,
            ),
            redirect_stdout(io.StringIO()),
            redirect_stderr(io.StringIO()),
        ):
            result = reserve_margin_repair_xlsx.main()
    finally:
        for opened_workbook in opened_workbooks:
            opened_workbook.close()
    if result != 0:
        raise AssertionError(result)


class ReserveMarginWarningPathTests(unittest.TestCase):
    @unittest.skipUnless(
        os.name == "nt",
        "the established 260-unit failure is Windows-specific",
    )
    def test_exact_260_unit_raw_failure_is_bounded_written_and_reopened(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            desired = _exact_failed_warning_path(Path(temp).resolve())
            self.assertTrue(desired.parent.is_dir())
            with self.assertRaises(FileNotFoundError):
                desired.write_text("established failure é Ω\n", encoding="utf-8")

            bounded = reserve_margin_repair_xlsx.bounded_warnings_path(desired)
            self.assertNotEqual(bounded, desired)
            self.assertLess(
                path_module.windows_path_units(bounded),
                path_module.WINDOWS_SAFE_ABSOLUTE_PATH_BUDGET,
            )
            bounded.write_text("established failure é Ω\n", encoding="utf-8")
            self.assertEqual(
                bounded.read_text(encoding="utf-8"),
                "established failure é Ω\n",
            )

    def test_safe_short_warning_paths_are_identical_without_cwd_resolution(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            desired = Path(temp).resolve() / "short warnings é Ω.txt"
            self.assertEqual(
                reserve_margin_repair_xlsx.bounded_warnings_path(desired),
                desired,
            )

        relative = Path("relative warning output") / "warnings.txt"
        with mock.patch.object(Path, "resolve", side_effect=AssertionError("resolved")):
            actual = reserve_margin_repair_xlsx.bounded_warnings_path(relative)
        self.assertEqual(actual, relative)

    def test_independently_generated_long_warning_paths_do_not_collide(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            first = _exact_failed_warning_path(Path(temp).resolve())
            second = first.with_name("independent_" + first.name)
            first_bounded = reserve_margin_repair_xlsx.bounded_warnings_path(first)
            repeated = reserve_margin_repair_xlsx.bounded_warnings_path(first)
            second_bounded = reserve_margin_repair_xlsx.bounded_warnings_path(second)
            self.assertEqual(first_bounded, repeated)
            self.assertNotEqual(first_bounded, second_bounded)
            self.assertEqual(first_bounded.parent, second_bounded.parent)
            for bounded in (first_bounded, second_bounded):
                self.assertLess(
                    path_module.windows_path_units(bounded),
                    path_module.WINDOWS_SAFE_ABSOLUTE_PATH_BUDGET,
                )

    def test_short_and_long_runs_preserve_warning_output_and_workbook_bytes(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp).resolve()
            input_path = root / "input.txt"
            workbook = root / "fallbacks.xlsx"
            short_output = root / "short-output.txt"
            short_warnings = root / "short.warnings.txt"
            _write_datafile(input_path)
            _write_fallback_workbook(workbook)
            workbook_hash = _sha256(workbook)

            long_warnings = _exact_failed_warning_path(root)
            long_output = long_warnings.parent / "long-output.txt"
            bounded_long_warnings = reserve_margin_repair_xlsx.bounded_warnings_path(
                long_warnings
            )

            _run_patcher(input_path, short_output, workbook, short_warnings)
            self.assertEqual(_sha256(workbook), workbook_hash)
            _run_patcher(input_path, long_output, workbook, long_warnings)
            self.assertEqual(_sha256(workbook), workbook_hash)

            expected_warning_text = (
                "TotalAnnualMaxCapacityInvestment: skipped PWRNGSTSTXX 2023; "
                "current=0 is sentinel but no fallback was provided for CR=TSTXX, "
                "prefix=PWRNGS.\n"
                "Investment blocked: PWRNGSTSTXX 2023 has "
                "TotalAnnualMaxCapacity=1 but TotalAnnualMaxCapacityInvestment=0.\n"
            )
            expected_warning = expected_warning_text.replace(
                "\n", os.linesep
            ).encode("utf-8")
            self.assertEqual(short_warnings.read_bytes(), expected_warning)
            self.assertEqual(bounded_long_warnings.read_bytes(), expected_warning)
            self.assertEqual(short_output.read_bytes(), long_output.read_bytes())
            self.assertEqual(short_output.read_bytes(), input_path.read_bytes())
            self.assertFalse(long_warnings.exists())

    def test_overlong_parent_fails_explicitly(self) -> None:
        relative = Path("relative") / ("p" * 230) / FAILED_WARNING_NAME
        with self.assertRaisesRegex(
            path_module.WorkspacePathBudgetError,
            "must be absolute",
        ):
            reserve_margin_repair_xlsx.bounded_warnings_path(relative)

        desired = Path("C:/") / ("p" * 230) / FAILED_WARNING_NAME
        with self.assertRaisesRegex(
            path_module.WorkspacePathBudgetError,
            "parent leaves no Windows-safe filename budget",
        ):
            reserve_margin_repair_xlsx.bounded_warnings_path(desired)


if __name__ == "__main__":
    unittest.main()
