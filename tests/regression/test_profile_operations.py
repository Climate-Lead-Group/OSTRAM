from __future__ import annotations

import json
from pathlib import Path
import tempfile
import unittest
from unittest import mock

import pandas as pd
from openpyxl import Workbook, load_workbook

from ostram.pipeline.preparation.scenario_country_sync import (
    SHEET_MAP,
    synchronize_country,
    technology_belongs_to_country,
)
from ostram.pipeline.scenarios.registry import load_registry
from ostram.pipeline.scenarios.transformations.ao_extension_decisions import (
    _decision_rows,
)
from ostram.pipeline.scenarios.transformations.fix_elc_pmode_revert import (
    country_region_map,
    configured_elc_dispatch_techs,
    elc_dispatch_techs,
)
from ostram.reporting.training_dashboard import (
    build_dashboard_data,
    render_html,
)
from ostram.terminal import RunReporter
from ostram.pipeline.scenarios.rules import add_max_cap_investment_lid_rule as lid


REPO_ROOT = Path(__file__).resolve().parents[2]


def source_workbook(path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for source, _target in SHEET_MAP:
        sheet = workbook.create_sheet(source)
        sheet.append(["Tech", "Value"])
        sheet.append(["PWRSPVMMRXX01", 2.496])
        sheet.append(["PWRSPVBGDXX01", 9.0])
    workbook.save(path)


def target_workbook(path: Path) -> None:
    workbook = Workbook()
    control = workbook.active
    control.title = "Control"
    control.append(["scenario"])
    control.append(["A"])
    control.append(["B"])
    for _source, target in SHEET_MAP:
        sheet = workbook.create_sheet(target)
        sheet.append(["scenario", "Tech", "Value"])
    workbook.save(path)


class ScenarioCountrySyncTests(unittest.TestCase):
    def test_structural_country_matching(self) -> None:
        self.assertTrue(technology_belongs_to_country("PWRSPVMMRXX01", "MMR"))
        self.assertTrue(technology_belongs_to_country("TRNBGDXXMMRXX", "MMR"))
        self.assertFalse(technology_belongs_to_country("PWRSPVBGDXX01", "MMR"))
        self.assertFalse(technology_belongs_to_country("NOT_MMR_TEXT", "MMR"))

    def test_sync_is_schema_driven_atomic_and_idempotent(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            ao = root / "ao.xlsx"
            target = root / "scenario.xlsx"
            source_workbook(ao)
            target_workbook(target)
            dry = synchronize_country(
                country="MMR", ao_path=ao, scenario_path=target, dry_run=True
            )
            self.assertEqual(sum(change.rows_added for change in dry), 6)
            workbook = load_workbook(target, read_only=True)
            self.assertEqual(workbook[SHEET_MAP[0][1]].max_row, 1)
            workbook.close()

            first = synchronize_country(country="MMR", ao_path=ao, scenario_path=target)
            second = synchronize_country(country="MMR", ao_path=ao, scenario_path=target)
            self.assertEqual(sum(change.rows_added for change in first), 6)
            self.assertEqual(sum(change.rows_added for change in second), 0)
            workbook = load_workbook(target, read_only=True, data_only=True)
            rows = list(workbook[SHEET_MAP[0][1]].iter_rows(min_row=2, values_only=True))
            workbook.close()
            self.assertEqual(rows, [
                ("A", "PWRSPVMMRXX01", 2.496),
            ])


class SelectedRegistryTests(unittest.TestCase):
    def test_registry_content_drives_non_full_scenarios(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            patch_dir = root / "Derived"
            patch_dir.mkdir()
            (patch_dir / "patches.json").write_text(json.dumps({
                "scenario": "Derived", "base_scenario": "Decision", "edits": []
            }), encoding="utf-8")
            registry_path = root / "registry.json"
            registry_path.write_text(json.dumps({
                "schema": "ostram-scenario-registry-v1",
                "support_scenarios": ["Seed"],
                "root_scenarios": [
                    {"name": "Seed", "role": "support", "dependencies": []},
                    {"name": "Decision", "role": "decision", "dependencies": []},
                ],
                "decision_scenarios": ["Decision", "Derived"],
                "derived_scenarios": [{
                    "name": "Derived", "base_scenario": "Decision",
                    "patches": "Derived/patches.json",
                }],
            }), encoding="utf-8")
            registry = load_registry(registry_path)
            self.assertEqual(registry.scenario_names, ("Seed", "Decision", "Derived"))
            self.assertEqual(registry.required_roots(["Derived"]), ("Decision",))


class AoDecisionAuthorityTests(unittest.TestCase):
    def test_csv_sidecar_is_exact_schema_and_duplicate_safe(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "ao_decisions.csv"
            path.write_text(
                "AO_Code_To_Add,Include,Override_Template_AO,"
                "Override_Tech.Name_AO,Notes\n"
                "PWRSPVMMRXX01,Y,PWRSPVBGDXX01,,training\n",
                encoding="utf-8",
            )
            rows = _decision_rows(path)
            self.assertEqual(rows[0]["AO_Code_To_Add"], "PWRSPVMMRXX01")
            self.assertEqual(rows[0]["Include"], "Y")
            path.write_text(
                path.read_text(encoding="utf-8")
                + "PWRSPVMMRXX01,N,,,duplicate\n",
                encoding="utf-8",
            )
            with self.assertRaisesRegex(ValueError, "duplicate"):
                _decision_rows(path)


class ProfileLidPolicyTests(unittest.TestCase):
    def _worksheet(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Secondary Techs"
        worksheet.append(["Tech", "Parameter", "Projection.Mode", 2025])
        worksheet.append([
            "PWRSPVMMRXX", "TotalAnnualMaxCapacityInvestment", "User defined", 0.166
        ])
        return workbook, worksheet

    def _apply(self, policies: str | None):
        workbook, worksheet = self._worksheet()
        environment = {} if policies is None else {"OSTRAM_PROFILE_POLICIES": policies}
        patches = (
            mock.patch.object(lid, "LID_RULE_MODE", "uniform"),
            mock.patch.object(lid, "LID_PERCENTAGE_DEFAULT", 0.5),
            mock.patch.object(lid, "LID_PERCENTAGE_BY_YEAR", {}),
            mock.patch.object(lid, "LID_RELAXATION_SCHEDULE", {2025: 2.0}),
            mock.patch.object(lid, "LID_FAMILY_RELAXATION_CEILINGS", {}),
            mock.patch.object(lid, "LID_EXEMPT_PREFIXES", []),
            mock.patch.object(lid, "LID_FLOOR_GW", 0.0),
        )
        with mock.patch.dict("os.environ", environment, clear=True):
            with patches[0], patches[1], patches[2], patches[3], patches[4], patches[5], patches[6]:
                result = lid.apply_lid_to_sheet(
                    worksheet,
                    {"PWRSPVMMRXX"},
                    {("MMRXX", 2025): 10.0},
                    {},
                    demand_mult_map={},
                )
        value = worksheet.cell(row=2, column=4).value
        workbook.close()
        return value, result

    def test_full_default_preserves_manual_value_and_profile_policy_can_override(self) -> None:
        full_value, full_log = self._apply(None)
        enabled_value, enabled_log = self._apply('{"lid_rule_new_semantics": true}')
        self.assertEqual(full_value, 0.166)
        self.assertEqual(full_log["changes"], [])
        self.assertEqual(enabled_value, 10.0)
        self.assertEqual(enabled_log["changes"][0]["reason"], "lid_relaxed")


class ProfileElcDispatchTests(unittest.TestCase):
    def test_dispatch_nodes_are_derived_from_profile_country_regions(self) -> None:
        self.assertEqual(
            elc_dispatch_techs(["BGD", "INDEA", "MMR"]),
            frozenset({"ELCBGDXX01", "ELCINDEA01", "ELCMMRXX01"}),
        )
        self.assertEqual(
            country_region_map(["BGD", "INDEA", "MMR"]),
            {"BGDXX": "BGD", "INDEA": "INDEA", "MMRXX": "MMR"},
        )
        self.assertEqual(
            configured_elc_dispatch_techs(
                REPO_ROOT
                / "examples"
                / "unescap"
                / "config"
                / "preparation"
                / "Config_country_codes.yaml"
            ),
            frozenset({"ELCBGDXX01", "ELCINDEA01"}),
        )

    def test_full_country_regions_preserve_the_historical_node_set(self) -> None:
        self.assertEqual(
            configured_elc_dispatch_techs(
                REPO_ROOT / "config" / "preparation" / "Config_country_codes.yaml"
            ),
            frozenset({
                "ELCBGDXX01", "ELCBTNXX01", "ELCINDEA01", "ELCINDNE01",
                "ELCINDNO01", "ELCINDSO01", "ELCINDWE01", "ELCLKAXX01",
                "ELCMDVXX01", "ELCNPLXX01",
            }),
        )


class ProfileReportingTests(unittest.TestCase):
    def test_synthetic_results_inject_profile_metadata_and_exact_source_value(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            csv_path = root / "results.csv"
            pd.DataFrame([
                {
                    "Scenario": "Training", "REGION": "GLOBAL", "YEAR": 2050,
                    "TECHNOLOGY": "PWRSPVMMRXX01",
                    "ProductionByTechnologyAnnual": 10.0,
                    "TotalCapacityAnnual": 4.0, "AnnualEmissions": 1.0,
                    "TotalDiscountedCost": 20.0, "CapitalInvestment": 5.0,
                },
                {
                    "Scenario": "Training", "REGION": "GLOBAL", "YEAR": 2050,
                    "TECHNOLOGY": "TRNBGDXXMMRXX",
                    "ProductionByTechnologyAnnual": 2.0,
                    "TotalCapacityAnnual": 2.496,
                },
            ]).to_csv(csv_path, index=False)
            data = build_dashboard_data(
                [("before", csv_path)], profile_id="tiny",
                manifest=root / "profile.yaml", workspace=root / "workspace",
                metadata={
                    "country_regions": [{"region": "MMRXX", "label": "Myanmar"}],
                    "interconnectors": [{"technology": "TRNBGDXXMMRXX"}],
                    "effective_values": {"interconnector_capacity_gw": 2.496},
                },
            )
            self.assertEqual(data["profile_id"], "tiny")
            self.assertEqual(data["effective_values"]["interconnector_capacity_gw"], 2.496)
            metrics = data["snapshots"]["before"]["Training"]["System"]
            self.assertEqual(
                metrics["interconnectors"]["TotalCapacityAnnual"][2050], 2.496
            )
            html = render_html(data)
            self.assertIn("ostram-profile-data", html)
            self.assertIn("2.496", html)

    def test_terminal_log_includes_profile_manifest_workspace_and_compile_state(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            reporter = RunReporter(
                project_root=REPO_ROOT, workspace=root / "workspace",
                scenarios=("Training",), verbose=False, profile_id="tiny",
                manifest=root / "profile.yaml", compile_only=True,
            )
            reporter.finish(outcome="COMPILE_ONLY_SUCCESS", exit_code=0,
                            final_message="complete")
            log = reporter.log_path.read_text(encoding="utf-8")
            self.assertIn("profile_id=tiny", log)
            self.assertIn("manifest=", log)
            self.assertIn("workspace=", log)
            self.assertIn("scenarios=[\"Training\"]", log)
            self.assertIn("compile_only=True", log)
            self.assertIn("outcome=COMPILE_ONLY_SUCCESS", log)


if __name__ == "__main__":
    unittest.main()
