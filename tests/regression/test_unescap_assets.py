from __future__ import annotations

import csv
from collections import Counter
import hashlib
from pathlib import Path
import tempfile
import unittest

from openpyxl import load_workbook

from ostram.profiles import load_manifest
from ostram.validation.profile import (
    ProfileDomainError,
    validate_compiled_domain,
    validate_seed_domain,
)


REPO_ROOT = Path(__file__).resolve().parents[2]
UNESCAP = REPO_ROOT / "examples" / "unescap"


def _read_values(path: Path) -> set[str]:
    with path.open("r", encoding="utf-8-sig", newline="") as stream:
        return {row["VALUE"] for row in csv.DictReader(stream)}


def _write_values(path: Path, values: set[str]) -> None:
    with path.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.writer(stream, lineterminator="\n")
        writer.writerow(["VALUE"])
        writer.writerows([value] for value in sorted(values))


class UnescapWorkbookTests(unittest.TestCase):
    def test_restrictions_is_a_header_only_seed_and_control_is_unchanged(self) -> None:
        workbook = load_workbook(
            UNESCAP / "inputs" / "scenarios" / "OSTRAM_Scenario_Inputs.xlsx",
            read_only=True,
            data_only=False,
        )
        self.assertEqual(len(workbook.sheetnames), 20)
        self.assertEqual(
            list(workbook["Restrictions"].values),
            [(
                "scenario", "source_sheet", "tech", "parameter", "year",
                "value", "rule_applied", "source_run_timestamp",
            )],
        )
        self.assertEqual(
            [row[0] for row in workbook["Control"].iter_rows(min_row=2, values_only=True)],
            ["BAU", "A_Calibrated_BAU", "B_Optimised_VRE", "C_Target_VRE"],
        )
        workbook.close()

    def test_workbook_retains_exact_interconnector_authority(self) -> None:
        workbook = load_workbook(
            UNESCAP / "inputs" / "scenarios" / "OSTRAM_Scenario_Inputs.xlsx",
            read_only=True,
            data_only=True,
        )
        sheet = workbook["Interconnector_Params"]
        row = next(
            values
            for values in sheet.iter_rows(min_row=2, values_only=True)
            if values[2] == "TRNBGDXXINDEA" and values[5] == "ResidualCapacity"
        )
        self.assertEqual(row[9:15], (2.496,) * 6)
        workbook.close()


class UnescapMappingTests(unittest.TestCase):
    def test_stage_specific_domain_contract_accepts_only_the_exact_delta(self) -> None:
        manifest = load_manifest(UNESCAP / "profile.yaml")
        seed_root = UNESCAP / "inputs" / "osemosys_global"
        seed = validate_seed_domain(manifest, seed_root)
        assert seed is not None
        self.assertEqual(seed["seed"]["TECHNOLOGY"]["count"], 89)
        self.assertEqual(seed["seed"]["FUEL"]["count"], 43)
        self.assertEqual(seed["projected_seed"]["TECHNOLOGY"]["count"], 89)

        with (
            UNESCAP / "config" / "scenarios" / "technology_types.csv"
        ).open("r", encoding="utf-8-sig", newline="") as stream:
            compiled_technology = {
                row["Technology"] for row in csv.DictReader(stream)
                if row["Technology"]
            }
        compiled_fuel = _read_values(seed_root / "FUEL.csv") | {
            "ELCBGDXX00", "ELCBGDXX03", "ELCBGDXX04",
            "ELCINDEA00", "ELCINDEA03", "ELCINDEA04",
        }
        with tempfile.TemporaryDirectory() as temp:
            compiled_root = Path(temp) / "compiled"
            compiled_root.mkdir()
            _write_values(compiled_root / "TECHNOLOGY.csv", compiled_technology)
            _write_values(compiled_root / "FUEL.csv", compiled_fuel)
            compiled = validate_compiled_domain(
                manifest,
                osemosys_inputs=seed_root,
                compiled_root=compiled_root,
            )
            assert compiled is not None
            self.assertEqual(
                compiled["generated_delta"],
                {
                    "TECHNOLOGY": ["PWRSHPINDEA"],
                    "FUEL": [
                        "ELCBGDXX00", "ELCBGDXX03", "ELCBGDXX04",
                        "ELCINDEA00", "ELCINDEA03", "ELCINDEA04",
                    ],
                },
            )
            self.assertEqual(compiled["compiled"]["TECHNOLOGY"]["count"], 90)
            self.assertEqual(compiled["compiled"]["FUEL"]["count"], 49)

            compiled_technology.remove("PWRSHPINDEA")
            compiled_technology.add("PWRUNDECLARED")
            _write_values(compiled_root / "TECHNOLOGY.csv", compiled_technology)
            with self.assertRaisesRegex(
                ProfileDomainError, "unexpected_additions.*PWRUNDECLARED"
            ):
                validate_compiled_domain(
                    manifest,
                    osemosys_inputs=seed_root,
                    compiled_root=compiled_root,
                )

    def test_seed_domain_rejects_same_count_membership_drift(self) -> None:
        manifest = load_manifest(UNESCAP / "profile.yaml")
        source = UNESCAP / "inputs" / "osemosys_global"
        technology = _read_values(source / "TECHNOLOGY.csv")
        technology.remove(next(iter(technology)))
        technology.add("UNDECLARED_SEED_TECHNOLOGY")
        with tempfile.TemporaryDirectory() as temp:
            seed_root = Path(temp)
            _write_values(seed_root / "TECHNOLOGY.csv", technology)
            _write_values(seed_root / "FUEL.csv", _read_values(source / "FUEL.csv"))
            with self.assertRaisesRegex(ProfileDomainError, "seed TECHNOLOGY"):
                validate_seed_domain(manifest, seed_root)

    def test_profile_presents_seed_and_compiled_counts_as_distinct_stages(self) -> None:
        metadata = load_manifest(UNESCAP / "profile.yaml").metadata
        self.assertEqual(metadata["seed_set_sizes"]["TECHNOLOGY"], 89)
        self.assertEqual(metadata["seed_set_sizes"]["FUEL"], 43)
        self.assertEqual(metadata["compiled_set_sizes"]["TECHNOLOGY"], 90)
        self.assertEqual(metadata["compiled_set_sizes"]["FUEL"], 49)
        delta = metadata["domain_contract"]["generated_delta"]
        self.assertEqual(delta["TECHNOLOGY"], ["PWRSHPINDEA"])
        self.assertEqual(
            delta["FUEL"],
            [
                "ELCBGDXX00", "ELCBGDXX03", "ELCBGDXX04",
                "ELCINDEA00", "ELCINDEA03", "ELCINDEA04",
            ],
        )

    def test_reduced_pwr_min_pin_is_complete_for_its_domain(self) -> None:
        authority = (
            UNESCAP / "config" / "scenarios" / "rules"
            / "pwr_min_2023_2026_pin.csv"
        )
        self.assertEqual(
            hashlib.sha256(authority.read_bytes()).hexdigest(),
            "984c3885f7bcee992d634c602402e3c6183a2f3bc8a1d8a0620ae214c1a1d872",
        )
        with authority.open(encoding="utf-8-sig", newline="") as stream:
            rows = list(csv.DictReader(stream))
        self.assertEqual(len(rows), 510)
        self.assertEqual(len({row["technology"] for row in rows}), 50)
        self.assertEqual(
            Counter(row["canonical_country"] for row in rows),
            {"BGD": 252, "IND": 258},
        )
        self.assertEqual(
            {
                scenario: sum(
                    scenario in row["root_scenarios_with_actual_change"].split(";")
                    for row in rows
                )
                for scenario in (
                    "A_Calibrated_BAU", "B_Optimised_VRE", "C_Target_VRE",
                )
            },
            {
                "A_Calibrated_BAU": 501,
                "B_Optimised_VRE": 510,
                "C_Target_VRE": 510,
            },
        )

    def test_ao_decisions_match_the_reduced_live_extension_set(self) -> None:
        authority = UNESCAP / "config" / "scenarios" / "ao_extension_decisions.csv"
        with authority.open(encoding="utf-8", newline="") as stream:
            codes = {row["AO_Code_To_Add"] for row in csv.DictReader(stream)}
        self.assertEqual(codes, {"PWRNGSBGDXX", "PWRSHPINDEA"})

    def test_reduced_taxonomy_is_complete_without_full_model_padding(self) -> None:
        taxonomy_path = UNESCAP / "config" / "scenarios" / "technology_types.csv"
        with taxonomy_path.open("r", encoding="utf-8-sig", newline="") as stream:
            rows = list(csv.DictReader(stream))
        mappings = [row["Technology"] for row in rows if row["Technology"]]
        self.assertEqual(len(taxonomy_path.read_text(encoding="utf-8").splitlines()), 92)
        self.assertEqual(len(mappings), 90)
        self.assertEqual(len(set(mappings)), len(mappings))
        self.assertIn("TRNBGDXXINDEA", mappings)

        # Every raw reduced-model technology that survives the declared matrix
        # filter/PWR cleanup is mapped. NGS and SHP are derived by preparation.
        with (
            UNESCAP / "inputs" / "osemosys_global" / "TECHNOLOGY.csv"
        ).open("r", encoding="utf-8-sig", newline="") as stream:
            raw = {row["VALUE"] for row in csv.DictReader(stream)}
        normalized = {
            tech[:-2]
            if tech.startswith("PWR") and len(tech) == 13 and tech[-2:] in {"00", "01"}
            else tech
            for tech in raw
        }
        deliberately_filtered = {
            "PWRCCGBGDXX", "PWRCCGINDEA", "PWROCGBGDXX", "PWROCGINDEA",
            "PWRCSPBGDXX", "PWRGEOBGDXX", "PWRGEOINDEA", "PWRHETINDEA",
            "PWRWAVBGDXX", "PWRWAVINDEA",
        }
        self.assertTrue((normalized - deliberately_filtered).issubset(set(mappings)))

        exercise = (
            UNESCAP / "exercises" / "add-interconnector.html"
        ).read_text(encoding="utf-8")
        self.assertIn("INTERCONNECTORS,TRNBGDXXMMRXX", exercise)

    def test_timeslice_authority_contains_only_reduced_model_sources(self) -> None:
        workbook = load_workbook(
            UNESCAP / "inputs" / "scenarios" / "OSTRAM_Timeslice_Inputs.xlsx",
            read_only=True,
            data_only=False,
        )
        self.assertEqual(
            workbook.sheetnames,
            ["YearSplit", "BGD_Dem", "BGD_CF", "INDEA_Dem", "INDEA_CF", "Config"],
        )
        workbook.close()


if __name__ == "__main__":
    unittest.main()
