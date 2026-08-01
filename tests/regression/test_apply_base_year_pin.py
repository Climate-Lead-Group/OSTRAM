from __future__ import annotations

import csv
import hashlib
import importlib.util
import json
import sys
import tempfile
import unittest
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
SCRIPT = (
    REPO_ROOT
    / "ostram"
    / "pipeline"
    / "scenarios"
    / "rules"
    / "apply_base_year_pin.py"
)
PRODUCTION_RULES = REPO_ROOT / "config" / "scenarios" / "rules" / "pwr_min_2023_2026_pin.csv"


def _load_module():
    name = "_ostram_apply_base_year_pin_test"
    spec = importlib.util.spec_from_file_location(name, SCRIPT)
    if spec is None or spec.loader is None:
        raise AssertionError(f"cannot import {SCRIPT}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[name] = module
    try:
        spec.loader.exec_module(module)
    finally:
        sys.modules.pop(name, None)
    return module


PIN = _load_module()


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _rule(
    *,
    technology: str = "PWRSPVINDEA",
    parameter: str = "TotalAnnualMaxCapacityInvestment",
    year: int = 2023,
    value: str = "1.25",
    state: str = "POSITIVE",
    root_scenarios: str = (
        "A_Calibrated_BAU;B_Optimised_VRE;C_Target_VRE"
    ),
) -> dict[str, str]:
    return {
        "source_rule_id": (
            f"PWR_MIN_PIN::{parameter}::GLOBAL::{technology}::{year}"
        ),
        "parameter": parameter,
        "ordered_parameter_indices": json.dumps(
            ["REGION", "TECHNOLOGY", "YEAR"],
            separators=(",", ":"),
        ),
        "region": "GLOBAL",
        "technology": technology,
        "semantic_technology_group": "PWR:SPV",
        "canonical_country": "IND",
        "year": str(year),
        "required_root_present": "true",
        "required_root_value": value,
        "required_root_state": state,
        "verified_physical_unit": (
            "PJ_per_year_activity"
            if "Activity" in parameter
            else "GW"
        ),
        "verified_compiled_unit": (
            "PJ_per_year" if "Activity" in parameter else "GW"
        ),
        "root_scenarios_with_actual_change": root_scenarios,
        "authority_classification": "BENCHMARK_SUPPORTED",
        "authority_lineage_class": (
            "ACCEPTED_WS4_BASE_YEAR_PIN_2023_2026"
        ),
    }


def _write_rules(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(
            stream, fieldnames=PIN.EXPECTED_FIELDS, lineterminator="\n"
        )
        writer.writeheader()
        writer.writerows(rows)


def _add_headers(worksheet) -> None:
    headers = (
        "Tech",
        "Parameter",
        "Projection.Mode",
        "Projection.Parameter",
        2022,
        2023,
        2024,
        2025,
        2026,
        2027,
    )
    for column, value in enumerate(headers, start=1):
        worksheet.cell(row=1, column=column, value=value)


def _write_workbook(
    directory: Path,
    rows: list[tuple[object, ...]],
    *,
    duplicate_target: bool = False,
) -> Path:
    directory.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    primary = workbook.active
    primary.title = "Primary Techs"
    secondary = workbook.create_sheet("Secondary Techs")
    _add_headers(primary)
    _add_headers(secondary)
    for row_number, values in enumerate(rows, start=2):
        for column, value in enumerate(values, start=1):
            secondary.cell(row=row_number, column=column, value=value)
    if duplicate_target:
        for column, value in enumerate(rows[0], start=1):
            primary.cell(row=2, column=column, value=value)
    path = directory / PIN.PARAM_FILE
    workbook.save(path)
    workbook.close()
    return path


def _row_values(path: Path, technology: str, parameter: str):
    workbook = load_workbook(path, data_only=False)
    try:
        for sheet_name in PIN.SHEETS:
            worksheet = workbook[sheet_name]
            headers = {
                cell.value: cell.column for cell in worksheet[1]
            }
            for row in range(2, worksheet.max_row + 1):
                if (
                    worksheet.cell(row, headers["Tech"]).value == technology
                    and worksheet.cell(row, headers["Parameter"]).value
                    == parameter
                ):
                    return {
                        header: worksheet.cell(row, column).value
                        for header, column in headers.items()
                    }
    finally:
        workbook.close()
    raise AssertionError(f"row not found: {technology}/{parameter}")


def _workbook_structure(path: Path) -> dict[str, tuple[object, ...]]:
    workbook = load_workbook(path, data_only=False)
    try:
        result: dict[str, tuple[object, ...]] = {}
        for sheet_name in PIN.SHEETS:
            worksheet = workbook[sheet_name]
            headers = {
                cell.value: cell.column for cell in worksheet[1]
            }
            identities = tuple(
                (
                    worksheet.cell(row, headers["Tech"]).value,
                    worksheet.cell(row, headers["Parameter"]).value,
                )
                for row in range(2, worksheet.max_row + 1)
            )
            result[sheet_name] = (
                worksheet.max_row,
                worksheet.max_column,
                identities,
            )
        return result
    finally:
        workbook.close()


class ProductionRuleContractTests(unittest.TestCase):
    def test_production_allowlist_is_exact_and_non_maldives(self) -> None:
        self.assertEqual(_sha256(PRODUCTION_RULES), PIN.RULES_SHA256)
        rules = PIN.load_pin_rules()
        self.assertEqual(len(rules), 1956)
        self.assertEqual(len({rule.complete_key for rule in rules}), 1956)
        self.assertEqual(
            Counter(rule.parameter for rule in rules),
            PIN.EXPECTED_PARAMETER_COUNTS,
        )
        self.assertEqual(
            Counter(rule.state for rule in rules),
            PIN.EXPECTED_STATE_COUNTS,
        )
        self.assertEqual(len({rule.technology for rule in rules}), 182)
        self.assertFalse(
            any(
                rule.canonical_country == "MDV"
                or "MDV" in rule.technology
                for rule in rules
            )
        )
        self.assertTrue(all(rule.present for rule in rules))
        self.assertTrue(
            all(rule.year in {2023, 2024, 2025, 2026} for rule in rules)
        )
        scenario_counts = {
            scenario: sum(
                scenario in rule.root_scenarios for rule in rules
            )
            for scenario in PIN.PIN_ROOT_SCENARIOS
        }
        self.assertEqual(scenario_counts, PIN.EXPECTED_SCENARIO_COUNTS)

    def test_source_rules_collapse_to_complete_workbook_rows_without_modes(self):
        rules = PIN.load_pin_rules()
        years_by_row: dict[tuple[str, str], set[int]] = defaultdict(set)
        for rule in rules:
            years_by_row[(rule.technology, rule.parameter)].add(rule.year)
        self.assertEqual(len(years_by_row), 517)
        self.assertEqual(
            Counter(len(years) for years in years_by_row.values()),
            {4: 432, 3: 66, 2: 11, 1: 8},
        )
        with PRODUCTION_RULES.open(
            "r", encoding="utf-8", newline=""
        ) as stream:
            self.assertTrue(
                all(
                    tuple(json.loads(raw["ordered_parameter_indices"]))
                    == PIN.ORDERED_INDICES
                    for raw in csv.DictReader(stream)
                )
            )


class StaticPinApplicationTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name)
        self.rules = self.root / "rules.csv"

    def tearDown(self) -> None:
        self.temporary.cleanup()

    def test_exact_keys_zero_absence_window_idempotence_and_no_other_change(
        self,
    ) -> None:
        _write_rules(
            self.rules,
            [
                _rule(year=2023, value="2.5"),
                _rule(year=2024, value="0", state="ZERO"),
            ],
        )
        input_dir = self.root / "input"
        path = _write_workbook(
            input_dir,
            [
                (
                    "PWRSPVINDEA",
                    "TotalAnnualMaxCapacityInvestment",
                    "User defined",
                    0,
                    22,
                    9,
                    None,
                    None,
                    None,
                    77,
                ),
                (
                    "PWRSPVINDEA",
                    "TotalAnnualMinCapacityInvestment",
                    "User defined",
                    0,
                    122,
                    None,
                    None,
                    None,
                    None,
                    177,
                ),
                (
                    "PWRSPVINDWE",
                    "TotalAnnualMaxCapacityInvestment",
                    "User defined",
                    0,
                    222,
                    8,
                    8,
                    8,
                    8,
                    277,
                ),
            ],
        )
        before_structure = _workbook_structure(path)
        first = PIN.apply_pin_rules(
            input_dir,
            "B_Optimised_VRE",
            self.rules,
            skip_backup=True,
        )
        target = _row_values(
            path,
            "PWRSPVINDEA",
            "TotalAnnualMaxCapacityInvestment",
        )
        self.assertEqual(target[2022], 22)
        self.assertEqual(target[2023], 2.5)
        self.assertEqual(target[2024], 0)
        self.assertIsNone(target[2025])
        self.assertIsNone(target[2026])
        self.assertEqual(target[2027], 77)
        absent = _row_values(
            path,
            "PWRSPVINDEA",
            "TotalAnnualMinCapacityInvestment",
        )
        self.assertIsNone(absent[2023])
        self.assertIsNone(absent[2024])
        other_tech = _row_values(
            path,
            "PWRSPVINDWE",
            "TotalAnnualMaxCapacityInvestment",
        )
        self.assertEqual(
            [other_tech[year] for year in (2022, 2023, 2024, 2025, 2026, 2027)],
            [222, 8, 8, 8, 8, 277],
        )
        self.assertEqual(first["rules_applied"], 2)
        self.assertEqual(first["workbook_rows_matched"], 1)
        self.assertEqual(first["changed_value_cells"], 2)
        self.assertEqual(first["zero_rules"], 1)
        self.assertEqual(_workbook_structure(path), before_structure)
        after_first_hash = _sha256(path)
        second = PIN.apply_pin_rules(
            input_dir,
            "B_Optimised_VRE",
            self.rules,
            skip_backup=True,
        )
        self.assertFalse(second["saved"])
        self.assertEqual(second["changed_value_cells"], 0)
        self.assertEqual(second["changed_projection_modes"], 0)
        self.assertEqual(_sha256(path), after_first_hash)

    def test_scenario_membership_filters_rules_without_touching_other_rows(self):
        _write_rules(
            self.rules,
            [
                _rule(
                    technology="PWRSPVINDEA",
                    value="2",
                    root_scenarios="A_Calibrated_BAU",
                ),
                _rule(
                    technology="PWRSPVINDWE",
                    value="3",
                    root_scenarios="B_Optimised_VRE;C_Target_VRE",
                ),
            ],
        )
        input_dir = self.root / "input"
        path = _write_workbook(
            input_dir,
            [
                (
                    "PWRSPVINDEA",
                    "TotalAnnualMaxCapacityInvestment",
                    "User defined",
                    0,
                    None,
                    9,
                    None,
                    None,
                    None,
                    None,
                ),
                (
                    "PWRSPVINDWE",
                    "TotalAnnualMaxCapacityInvestment",
                    "User defined",
                    0,
                    None,
                    8,
                    None,
                    None,
                    None,
                    None,
                ),
            ],
        )
        before_structure = _workbook_structure(path)
        result = PIN.apply_pin_rules(
            input_dir,
            "A_Calibrated_BAU",
            self.rules,
            skip_backup=True,
        )
        included = _row_values(
            path,
            "PWRSPVINDEA",
            "TotalAnnualMaxCapacityInvestment",
        )
        excluded = _row_values(
            path,
            "PWRSPVINDWE",
            "TotalAnnualMaxCapacityInvestment",
        )
        self.assertEqual(included[2023], 2)
        self.assertEqual(excluded[2023], 8)
        self.assertEqual(result["rules_loaded"], 2)
        self.assertEqual(result["rules_applied"], 1)
        self.assertEqual(result["workbook_rows_matched"], 1)
        self.assertEqual(_workbook_structure(path), before_structure)

    def test_empty_projection_mode_flips_only_when_non_targets_are_blank(self):
        _write_rules(self.rules, [_rule(year=2023, value="1")])
        input_dir = self.root / "input"
        path = _write_workbook(
            input_dir,
            [
                (
                    "PWRSPVINDEA",
                    "TotalAnnualMaxCapacityInvestment",
                    "EMPTY",
                    0,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                )
            ],
        )
        result = PIN.apply_pin_rules(
            input_dir,
            "C_Target_VRE",
            self.rules,
            skip_backup=True,
        )
        row = _row_values(
            path,
            "PWRSPVINDEA",
            "TotalAnnualMaxCapacityInvestment",
        )
        self.assertEqual(row["Projection.Mode"], "User defined")
        self.assertEqual(row[2023], 1)
        self.assertTrue(
            all(row[year] is None for year in (2022, 2024, 2025, 2026, 2027))
        )
        self.assertEqual(result["changed_projection_modes"], 1)

    def test_explicit_zero_in_non_target_year_blocks_row_activation(self):
        _write_rules(self.rules, [_rule(year=2023, value="1")])
        input_dir = self.root / "input"
        path = _write_workbook(
            input_dir,
            [
                (
                    "PWRSPVINDEA",
                    "TotalAnnualMaxCapacityInvestment",
                    "EMPTY",
                    0,
                    None,
                    None,
                    None,
                    None,
                    None,
                    0,
                )
            ],
        )
        before = _sha256(path)
        with self.assertRaisesRegex(ValueError, "populated non-target"):
            PIN.apply_pin_rules(
                input_dir,
                "A_Calibrated_BAU",
                self.rules,
                skip_backup=True,
            )
        self.assertEqual(_sha256(path), before)

    def test_missing_duplicate_bad_mode_and_bad_header_fail_before_write(self):
        _write_rules(self.rules, [_rule(year=2023, value="1")])
        cases = ("missing", "duplicate", "mode", "header")
        for case in cases:
            with self.subTest(case=case):
                input_dir = self.root / case
                tech = (
                    "PWRSPVINDWE"
                    if case == "missing"
                    else "PWRSPVINDEA"
                )
                mode = "Flat" if case == "mode" else "User defined"
                path = _write_workbook(
                    input_dir,
                    [
                        (
                            tech,
                            "TotalAnnualMaxCapacityInvestment",
                            mode,
                            0,
                            None,
                            9,
                            None,
                            None,
                            None,
                            None,
                        )
                    ],
                    duplicate_target=(case == "duplicate"),
                )
                if case == "header":
                    workbook = load_workbook(path)
                    for sheet in PIN.SHEETS:
                        workbook[sheet].cell(row=1, column=3, value="Bad.Mode")
                    workbook.save(path)
                    workbook.close()
                before = _sha256(path)
                with self.assertRaises(ValueError):
                    PIN.apply_pin_rules(
                        input_dir,
                        "B_Optimised_VRE",
                        self.rules,
                        skip_backup=True,
                    )
                self.assertEqual(_sha256(path), before)

    def test_rule_validation_rejects_duplicates_absence_mdv_and_bad_keys(self):
        duplicate = _rule()
        invalid_cases: dict[str, list[dict[str, str]]] = {
            "duplicate": [duplicate, dict(duplicate)],
            "absence": [
                {
                    **_rule(),
                    "required_root_present": "false",
                    "required_root_value": "",
                    "required_root_state": "ABSENT",
                }
            ],
            "maldives": [
                _rule(technology="PWRSPVMDVXX")
                | {
                    "canonical_country": "MDV",
                    "source_rule_id": (
                        "PWR_MIN_PIN::TotalAnnualMaxCapacityInvestment::"
                        "GLOBAL::PWRSPVMDVXX::2023"
                    ),
                }
            ],
            "post-window": [
                _rule(year=2027)
            ],
            "sentinel-placeholder": [
                _rule(value="9999")
            ],
            "bad-indices": [
                {
                    **_rule(),
                    "ordered_parameter_indices": json.dumps(
                        ["TECHNOLOGY", "YEAR"]
                    ),
                }
            ],
            "nonfinite": [
                _rule(value="NaN")
            ],
        }
        for case, rows in invalid_cases.items():
            with self.subTest(case=case):
                path = self.root / f"{case}.csv"
                _write_rules(path, rows)
                with self.assertRaises(ValueError):
                    PIN.load_pin_rules(path)

    def test_unsupported_scenario_is_rejected_without_opening_workbook(self):
        _write_rules(self.rules, [_rule()])
        with self.assertRaisesRegex(ValueError, "unsupported pin scenario"):
            PIN.apply_pin_rules(
                self.root / "does-not-exist",
                "BAU",
                self.rules,
                skip_backup=True,
            )


if __name__ == "__main__":
    unittest.main()
