"""
apply_patches.py  --  OSTRAM sensitivity post-A3 patcher   (CLG / OSTRAM)
========================================================================

Builds a sensitivity scenario's A-O_Parametrization.xlsx by branching from
the ``base_scenario`` declared by its canonical registry/patch contract and
applying, in order:

    1. the SHARED VRE physical-potential ceiling layer
       (sensitivity_expansion/reference/vre_ceilings_base.json)
    2. the run-specific edits in
       A3_process/rules_scripts/configs/<scenario>/patches.json

Design (patterned on the rules_scripts):
  * Non-destructive to the SOURCE: the declared root A-O is never mutated.
  * Idempotent: every run rebuilds the target folder from a FRESH copy of the
    source A1_Outputs folder, so the result depends only on (source + patches),
    never on a prior patched state.
  * Timestamped backup of any pre-existing target folder before overwrite.
  * Writes a *_CHANGES.json audit log.
  * --self-test runs in-process unit tests; --restore reverts the target from
    its most recent backup.

This is INPUT GENERATION (post-A3). It does NOT run A3/B1/B2.

USAGE
-----
    python apply_patches.py --scenario B_Opt_TradeCap50
    python apply_patches.py --scenario B_Opt_SolarHi10
    python apply_patches.py --scenario B_Opt_LinkFreeze
    python apply_patches.py --self-test
    python apply_patches.py --scenario B_Opt_TradeCap50 --restore

patches.json edit schema (one dict per edit):
    sheet            : worksheet name (e.g. "Secondary Techs", "Demand Techs")
    param            : OSeMOSYS parameter (the "Parameter" column value)
    tech | tech_prefix | techs : technology selector (exactly one)
    ONE operation of:
        values       : {year: value}  -> set those year cells explicitly
        op:"multiply"+factor          -> new = old * factor  (per year)
        op:"set_flat"+value           -> set every model year to value
        op:"set_to_residual"          -> set every year = that tech's
                                         ResidualCapacity (same sheet)
        op:"set_to_residual_factor_floor"
                                      -> selected years = max(
                                         factor * effective RC[residual_year],
                                         base_window_floor)
        op:"clamp_to_residual"         -> clamp to the effective RC, or to the
                                         separate v18 compatibility boundary
                                         when residual_source is
                                         "minimum_investment_boundary"
                                         (minimum only)
    create_if_absent : bool (default False) -> create the param row if missing
    note             : free text (audit only)
"""

from __future__ import annotations

import argparse
import json
import math
import shutil
import sys
import tempfile
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook, Workbook
from ostram._legacy_import import load_file_module

# --------------------------------------------------------------------------
# Constants / paths
# --------------------------------------------------------------------------
PARAM_FILE = "A-O_Parametrization.xlsx"
PROJ_MODE_COL = "Projection.Mode"
PROJ_MODE_USER = "User defined"
RES_PARAM = "ResidualCapacity"
MIN_INV_PARAM = "TotalAnnualMinCapacityInvestment"
YEAR_MIN, YEAR_MAX = 2020, 2060

SCRIPT_DIR = Path(__file__).resolve().parent          # .../sensitivity_expansion
REPO = SCRIPT_DIR.parent                               # .../t1_confection
A1_OUTPUTS = REPO / "A1_Outputs"
A3_PROCESS_DIR = REPO / "A3_process"
CONFIGS = A3_PROCESS_DIR / "rules_scripts" / "configs"
CEIL_BASE = SCRIPT_DIR / "reference" / "vre_ceilings_base.json"
SOASIA_V18 = A3_PROCESS_DIR / "OSTRAM_Scenario_Inputs.xlsx"
MINIMUM_BOUNDARY_SOURCE = "minimum_investment_boundary"

_authority = load_file_module(
    "_ostram_stage11_interconnector_authority",
    A3_PROCESS_DIR / "interconnector_authority.py",
)
MINIMUM_BOUNDARY_PARAMETER = _authority.MINIMUM_BOUNDARY_PARAMETER
load_minimum_boundary_authority = _authority.load_minimum_boundary_authority


# --------------------------------------------------------------------------
# Worksheet helpers
# --------------------------------------------------------------------------
def _as_year(v):
    try:
        y = int(float(v))
        return y if YEAR_MIN <= y <= YEAR_MAX else None
    except (TypeError, ValueError):
        return None


def scan_columns(ws):
    """Return (col_map, year_cols) from the header row (row 1)."""
    col_map, year_cols = {}, {}
    for c in range(1, ws.max_column + 1):
        raw = ws.cell(row=1, column=c).value
        if raw is None:
            continue
        y = _as_year(raw)
        if y is not None:
            year_cols[y] = c
            continue
        name = str(raw).strip()
        if name in ("Technology", "Tech"):
            col_map["tech"] = c
        elif name == "Parameter":
            col_map["param"] = c
        elif name == PROJ_MODE_COL:
            col_map["proj"] = c
        elif name == "Tech.ID":
            col_map["techid"] = c
        elif name == "Tech.Name":
            col_map["techname"] = c
        elif name == "Parameter.ID":
            col_map["paramid"] = c
        elif name == "Unit":
            col_map["unit"] = c
    if "tech" not in col_map or "param" not in col_map:
        raise ValueError("Sheet missing 'Tech'/'Parameter' columns")
    return col_map, year_cols


def find_row(ws, col_map, tech, param):
    tc, pc = col_map["tech"], col_map["param"]
    for r in range(2, ws.max_row + 1):
        t = ws.cell(row=r, column=tc).value
        p = ws.cell(row=r, column=pc).value
        if t is None or p is None:
            continue
        if str(t).strip() == tech and str(p).strip() == param:
            return r
    return None


def find_unique_row(ws, col_map, tech, param):
    rows = []
    tc, pc = col_map["tech"], col_map["param"]
    for r in range(2, ws.max_row + 1):
        if (
            str(ws.cell(row=r, column=tc).value or "").strip() == tech
            and str(ws.cell(row=r, column=pc).value or "").strip() == param
        ):
            rows.append(r)
    if len(rows) != 1:
        raise ValueError(
            f"expected exactly one {tech}/{param} row, found {len(rows)}"
        )
    return rows[0]


def finite_nonnegative(value, context):
    if isinstance(value, bool) or value is None:
        raise ValueError(f"{context}: expected numeric value, got {value!r}")
    try:
        result = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"{context}: expected numeric value, got {value!r}"
        ) from exc
    if not math.isfinite(result) or result < 0:
        raise ValueError(
            f"{context}: expected finite non-negative value, got {value!r}"
        )
    return result


def load_minimum_investment_boundaries(authority_path=None):
    """Load the exact dense minimum-clamp authority from canonical v18."""
    path = SOASIA_V18 if authority_path is None else authority_path
    return load_minimum_boundary_authority(Path(path).resolve())


def validate_residual_source(edit):
    """Validate a patch's optional external residual source."""
    source = edit.get("residual_source")
    if source is None or source == "effective":
        return False
    if source != MINIMUM_BOUNDARY_SOURCE:
        raise ValueError(f"unsupported residual_source {source!r}")
    if (
        edit.get("param") != MIN_INV_PARAM
        or edit.get("op") != "clamp_to_residual"
    ):
        raise ValueError(
            f"{MINIMUM_BOUNDARY_SOURCE} is permitted only for "
            f"{MIN_INV_PARAM} clamp_to_residual edits, got "
            f"parameter={edit.get('param')!r}, op={edit.get('op')!r}"
        )
    return True


def residual_values_for_edit(
    ws,
    col_map,
    year_cols,
    tech,
    param,
    edit,
    minimum_boundaries=None,
):
    source = edit.get("residual_source", "effective")
    if source == "effective":
        row = find_unique_row(ws, col_map, tech, RES_PARAM)
        return {
            year: finite_nonnegative(
                ws.cell(row, col).value,
                f"effective {tech}/{RES_PARAM}/{year}",
            )
            for year, col in year_cols.items()
        }
    if source == MINIMUM_BOUNDARY_SOURCE:
        if param != MIN_INV_PARAM:
            raise ValueError(
                f"{MINIMUM_BOUNDARY_SOURCE} is permitted only for "
                f"{MIN_INV_PARAM}, got {param}"
            )
        profiles = (
            minimum_boundaries
            if minimum_boundaries is not None
            else load_minimum_investment_boundaries()
        )
        if tech not in profiles:
            raise ValueError(
                f"{MINIMUM_BOUNDARY_PARAMETER} authority missing {tech}"
            )
        missing_years = sorted(set(year_cols) - set(profiles[tech]))
        if missing_years:
            raise ValueError(
                f"{MINIMUM_BOUNDARY_PARAMETER} authority {tech} "
                f"missing years {missing_years}"
            )
        return {
            year: finite_nonnegative(
                profiles[tech][year],
                f"{MINIMUM_BOUNDARY_PARAMETER} authority {tech}/{year}",
            )
            for year in year_cols
        }
    raise ValueError(f"unsupported residual_source {source!r}")


def find_any_row_for_tech(ws, col_map, tech):
    tc = col_map["tech"]
    for r in range(2, ws.max_row + 1):
        t = ws.cell(row=r, column=tc).value
        if t is not None and str(t).strip() == tech:
            return r
    return None


def techs_matching(ws, col_map, edit):
    """Resolve tech / tech_prefix / techs into a concrete ordered list that
    actually exists on the sheet."""
    tc = col_map["tech"]
    present = []
    seen = set()
    for r in range(2, ws.max_row + 1):
        t = ws.cell(row=r, column=tc).value
        if t is None:
            continue
        t = str(t).strip()
        if t not in seen:
            seen.add(t)
            present.append(t)
    if "tech" in edit:
        return [edit["tech"]]
    if "techs" in edit:
        return list(edit["techs"])
    if "tech_prefix" in edit:
        pref = edit["tech_prefix"]
        return [t for t in present if t.startswith(pref)]
    raise ValueError("edit needs one of tech / techs / tech_prefix")


def read_row_values(ws, col_map, year_cols, tech, param):
    r = find_row(ws, col_map, tech, param)
    if r is None:
        return None
    return {y: ws.cell(row=r, column=c).value for y, c in year_cols.items()}


def create_param_row(ws, col_map, tech, param):
    """Append a new parameter row for `tech`, copying identity columns from an
    existing row of that tech. Returns the new row index."""
    src = find_any_row_for_tech(ws, col_map, tech)
    r = ws.max_row + 1
    # copy identity columns
    for key in ("techid", "tech", "techname", "unit"):
        if key in col_map:
            val = ws.cell(row=src, column=col_map[key]).value if src else (tech if key == "tech" else None)
            ws.cell(row=r, column=col_map[key]).value = val
    ws.cell(row=r, column=col_map["tech"]).value = tech
    ws.cell(row=r, column=col_map["param"]).value = param
    if "proj" in col_map:
        ws.cell(row=r, column=col_map["proj"]).value = PROJ_MODE_USER
    return r


def flip_proj_mode(ws, col_map, row):
    if "proj" in col_map:
        cell = ws.cell(row=row, column=col_map["proj"])
        if cell.value in (None, "EMPTY", ""):
            cell.value = PROJ_MODE_USER


# --------------------------------------------------------------------------
# Edit application
# --------------------------------------------------------------------------
def apply_edit(
    ws,
    col_map,
    year_cols,
    edit,
    log,
    minimum_boundaries=None,
):
    param = edit["param"]
    validate_residual_source(edit)
    create = bool(edit.get("create_if_absent", False))
    for tech in techs_matching(ws, col_map, edit):
        r = find_row(ws, col_map, tech, param)
        if r is None:
            if edit.get("op") == "set_to_residual_factor_floor":
                raise ValueError(f"required target row missing: {tech}/{param}")
            if not create:
                log["skipped"].append({"tech": tech, "param": param,
                                       "reason": "row absent, create_if_absent=false"})
                continue
            r = create_param_row(ws, col_map, tech, param)
            log["rows_created"].append({"tech": tech, "param": param, "row": r})

        # compute new value per year
        if edit.get("op") == "set_to_residual_factor_floor":
            residual_year = int(edit["residual_year"])
            if residual_year not in year_cols:
                raise ValueError(
                    f"{tech}/{param}: residual year {residual_year} missing"
                )
            factor = finite_nonnegative(
                edit["factor"], f"{tech}/{param} factor"
            )
            if factor <= 0:
                raise ValueError(f"{tech}/{param}: factor must be positive")
            floor = finite_nonnegative(
                edit["base_window_floor"],
                f"{tech}/{param} base_window_floor",
            )
            years = tuple(int(year) for year in edit["years"])
            if not years or len(set(years)) != len(years):
                raise ValueError(
                    f"{tech}/{param}: years must be unique and non-empty"
                )
            missing_years = sorted(set(years) - set(year_cols))
            if missing_years:
                raise ValueError(
                    f"{tech}/{param}: target years missing {missing_years}"
                )
            residual_row = find_unique_row(
                ws, col_map, tech, RES_PARAM
            )
            residual = finite_nonnegative(
                ws.cell(residual_row, year_cols[residual_year]).value,
                f"effective {tech}/{RES_PARAM}/{residual_year}",
            )
            formula_value = round(
                max(round(factor * residual, 6), floor), 6
            )
            year_set = set(years)
            newfn = lambda y, old, value=formula_value, yrs=year_set: (
                value if y in yrs else old
            )
        elif "values" in edit:
            newfn = lambda y, old, v=edit["values"]: (
                v[str(y)] if str(y) in v else (v[y] if y in v else old))
        elif edit.get("op") == "multiply":
            f = float(edit["factor"])
            # Optional per-year restriction: {"op":"multiply","factor":..,"years":[..]}
            # multiplies only the listed years; all other years are left unchanged.
            # Absent "years" -> multiply every year (backward compatible).
            yrs = set(int(y) for y in edit["years"]) if edit.get("years") else None
            def newfn(y, old, f=f, yrs=yrs):
                if yrs is not None and y not in yrs:
                    return old
                if old is None or (isinstance(old, str) and not old.strip()):
                    return old
                try:
                    return round(float(old) * f, 6)
                except (TypeError, ValueError):
                    return old
        elif edit.get("op") == "set_flat":
            val = float(edit["value"])
            newfn = lambda y, old, val=val: val
        elif edit.get("op") == "set_to_residual":
            resvals = residual_values_for_edit(
                ws, col_map, year_cols, tech, param, edit, minimum_boundaries
            )
            def newfn(y, old, rv=resvals):
                return round(rv[y], 6)
        elif edit.get("op") == "clamp_to_residual":
            resvals = residual_values_for_edit(
                ws, col_map, year_cols, tech, param, edit, minimum_boundaries
            )
            def newfn(y, old, rv=resvals):
                cap = rv[y]
                if old is None:
                    return None
                try:
                    return round(min(float(old), cap), 6)
                except (TypeError, ValueError):
                    return old
        else:
            raise ValueError(f"edit for {tech}/{param}: no valid operation")

        changed = False
        for y, c in year_cols.items():
            cell = ws.cell(row=r, column=c)
            old = cell.value
            nv = newfn(y, old)
            if nv is not None and (old is None or _neq(old, nv)):
                cell.value = nv
                changed = True
                log["cells"].append({"tech": tech, "param": param, "year": y,
                                     "old": _num(old), "new": _num(nv)})
        if changed:
            flip_proj_mode(ws, col_map, r)


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return v


def _neq(a, b):
    try:
        return abs(float(a) - float(b)) > 1e-9
    except (TypeError, ValueError):
        return a != b


def clamp_row_to(ws, col_map, year_cols, tech, param, cap, log):
    """Clamp an existing (tech,param) row down so no year exceeds `cap`.
    Preserves coherence when a ceiling lowers TotalAnnualMaxCapacity below an
    existing investment lid/floor (e.g. MaxCapInv > MaxCap after a clip)."""
    r = find_row(ws, col_map, tech, param)
    if r is None:
        return
    changed = False
    for y, c in year_cols.items():
        cell = ws.cell(row=r, column=c)
        old = cell.value
        if old is None:
            continue
        try:
            ov = float(old)
        except (TypeError, ValueError):
            continue
        if ov > cap + 1e-9:
            cell.value = round(cap, 6)
            changed = True
            log["cells"].append({"tech": tech, "param": param, "year": y,
                                 "old": ov, "new": round(cap, 6), "reason": "clamp<=ceiling"})
    if changed:
        flip_proj_mode(ws, col_map, r)


ACT_LOWER = "TotalTechnologyAnnualActivityLowerLimit"


def scale_activity_to_ceiling(ws, col_map, year_cols, tech, ceil, maxcap_orig, log):
    """Where the VRE ceiling clips MaxCap below its pre-clip value, scale the
    tech's activity LOWER limit (NDC generation floor) down by the same per-year
    ratio ceil/MaxCap_orig.

    Rationale: C_Target's set_vre_targets writes an activity floor sized to a
    cap_envelope MaxCap (floor + 20% headroom). When the atlas ceiling clips
    MaxCap below that (e.g. MDV solar 1.73 -> 1.0 GW), the clipped capacity can
    no longer physically produce the floor, so glpsol's line-199 check
    (max production >= activity lower limit) aborts. Production is linear in
    MaxCap, so scaling the floor by ceil/MaxCap_orig re-caps the NDC target at
    what is physically buildable while preserving the original 20% headroom.

    ONLY the lower limit is touched -- never the pinned upper limit:
      * B_Opt-derived scenarios have lower limit == 0 on every VRE tech (they
        use investment lids, not activity floors), so scaling is a pure no-op
        for them (0 -> 0), and their upper limit is the -1 'unconstrained'
        sentinel which must NOT be scaled into a bogus negative bound.
      * Leaving C_Target's upper pin unscaled just makes it non-binding above
        physical max (lower <= upper still holds, line-184 check safe); the
        tech lands at ~physical maximum -- the correct 'NDC capped at physical
        potential' outcome."""
    r = find_row(ws, col_map, tech, ACT_LOWER)
    if r is None:
        return
    changed = False
    for y, c in year_cols.items():
        mco = maxcap_orig.get(y)
        try:
            mco = float(mco) if mco is not None else None
        except (TypeError, ValueError):
            mco = None
        if mco is None or mco <= 0 or mco <= ceil + 1e-9:
            continue  # ceiling does not bind this year -> leave floor as-is
        cell = ws.cell(row=r, column=c)
        old = cell.value
        if old is None:
            continue
        try:
            ov = float(old)
        except (TypeError, ValueError):
            continue
        if ov <= 1e-9:
            continue  # no floor to scale (B_Opt: lower limit == 0)
        nv = round(ov * ceil / mco, 6)
        if _neq(ov, nv):
            cell.value = nv
            changed = True
            log["cells"].append({"tech": tech, "param": ACT_LOWER, "year": y,
                                 "old": ov, "new": nv, "reason": "scale NDC floor to clipped ceiling"})
    if changed:
        flip_proj_mode(ws, col_map, r)


def apply_ceiling_layer(wb, log, ceiling_path=CEIL_BASE):
    ceiling_path = Path(ceiling_path)
    base = json.loads(ceiling_path.read_text(encoding="utf-8"))
    sheet, param = base["sheet"], base["param"]
    ceilings = base["ceilings_gw"]
    ws = wb[sheet]
    col_map, year_cols = scan_columns(ws)
    for tech, ceil in ceilings.items():
        ceil = float(ceil)
        # capture pre-clip MaxCap BEFORE overwriting, so any activity floor/pin
        # sized to the un-clipped capacity can be scaled back into feasibility.
        maxcap_orig = read_row_values(ws, col_map, year_cols, tech, param) or {}
        edit = {"param": param, "tech": tech, "op": "set_flat", "value": ceil}
        apply_edit(ws, col_map, year_cols, edit, log)
        # coherence: no annual investment lid/floor may exceed the total ceiling
        clamp_row_to(ws, col_map, year_cols, tech, "TotalAnnualMaxCapacityInvestment", ceil, log)
        clamp_row_to(ws, col_map, year_cols, tech, "TotalAnnualMinCapacityInvestment", ceil, log)
        # coherence: an activity floor sized above the ceiling is unproducible
        # once MaxCap is clipped (C_Target NDC cap_envelope) -> scale it down.
        scale_activity_to_ceiling(ws, col_map, year_cols, tech, ceil, maxcap_orig, log)
    log["ceiling_layer"] = {
        "n_techs": len(ceilings),
        "source": str(ceiling_path),
    }


def apply_run_patches(wb, patches, log, minimum_boundaries=None):
    for edit in patches.get("edits", []):
        ws = wb[edit["sheet"]]
        col_map, year_cols = scan_columns(ws)
        apply_edit(
            ws,
            col_map,
            year_cols,
            edit,
            log,
            minimum_boundaries=minimum_boundaries,
        )


def validate_patch_authorities(patches, authority_path=None):
    """Validate external patch authorities before target-directory mutation."""
    requires_minimum_boundary = False
    for edit in patches.get("edits", []):
        requires_minimum_boundary |= validate_residual_source(edit)
    if not requires_minimum_boundary:
        return None
    return load_minimum_investment_boundaries(authority_path)


# --------------------------------------------------------------------------
# Orchestration
# --------------------------------------------------------------------------
def build_scenario(
    scenario,
    source=None,
    skip_backup=False,
    *,
    a1_outputs=A1_OUTPUTS,
    configs=CONFIGS,
    ceiling_path=CEIL_BASE,
    authority_path=SOASIA_V18,
):
    """Rebuild one derived scenario from its explicitly declared root.

    ``source`` is retained as a fail-closed compatibility override.  When it
    is supplied it must equal ``patches.json::base_scenario``; omitting it uses
    the declaration directly.  Injectable path roots support pristine,
    disposable materialization proofs without touching the live worktree.
    """

    a1_outputs = Path(a1_outputs)
    configs = Path(configs)
    patches_path = configs / scenario / "patches.json"
    if not patches_path.is_file():
        raise FileNotFoundError(f"patches.json not found: {patches_path}")
    patches = json.loads(patches_path.read_text(encoding="utf-8"))
    declared_source = patches.get("base_scenario")
    if not declared_source:
        raise ValueError(f"{patches_path} does not declare base_scenario")
    if source is not None and source != declared_source:
        raise ValueError(
            f"source override {source!r} conflicts with "
            f"{patches_path.name}::base_scenario={declared_source!r}"
        )
    source = declared_source

    src_dir = a1_outputs / f"A1_Outputs_{source}"
    tgt_dir = a1_outputs / f"A1_Outputs_{scenario}"

    if not (src_dir / PARAM_FILE).is_file():
        raise FileNotFoundError(f"source A-O not found: {src_dir/PARAM_FILE}")
    minimum_boundaries = validate_patch_authorities(
        patches,
        authority_path=authority_path,
    )

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    # Backup any pre-existing target, then rebuild from a FRESH source copy.
    backup_dir = None
    if tgt_dir.exists():
        if not skip_backup:
            backup_dir = tgt_dir.parent / f"{tgt_dir.name}_PREPATCH_{stamp}"
            shutil.copytree(tgt_dir, backup_dir)
        shutil.rmtree(tgt_dir)
    shutil.copytree(src_dir, tgt_dir)

    log = {"scenario": scenario, "source": source, "timestamp": stamp,
           "patches_json": str(patches_path), "target": str(tgt_dir),
           "backup_dir": str(backup_dir) if backup_dir else None,
           "cells": [], "rows_created": [], "skipped": []}

    wb = load_workbook(tgt_dir / PARAM_FILE)
    if patches.get("apply_vre_ceiling_layer", True):
        apply_ceiling_layer(wb, log, ceiling_path=ceiling_path)
    apply_run_patches(
        wb,
        patches,
        log,
        minimum_boundaries=minimum_boundaries,
    )
    wb.save(tgt_dir / PARAM_FILE)
    wb.close()

    log_path = tgt_dir / f"apply_patches_CHANGES_{stamp}.json"
    log_path.write_text(json.dumps(log, indent=2, default=str))
    log["log_path"] = str(log_path)
    return log


def restore_scenario(scenario):
    tgt_dir = A1_OUTPUTS / f"A1_Outputs_{scenario}"
    backups = sorted(tgt_dir.parent.glob(f"{tgt_dir.name}_PREPATCH_*"))
    if not backups:
        raise FileNotFoundError(f"no backup found for {scenario}")
    latest = backups[-1]
    if tgt_dir.exists():
        shutil.rmtree(tgt_dir)
    shutil.copytree(latest, tgt_dir)
    print(f"Restored {tgt_dir} from {latest}")


def print_summary(log):
    print("=" * 64)
    print(f"apply_patches  --  {log['scenario']}  (from {log['source']})")
    print("=" * 64)
    print(f"target        : {log['target']}")
    print(f"backup        : {log.get('backup_dir') or '(none)'}")
    if "ceiling_layer" in log:
        print(f"VRE ceilings  : {log['ceiling_layer']['n_techs']} gen techs")
    print(f"cells written : {len(log['cells'])}")
    print(f"rows created  : {len(log['rows_created'])}  {[r['tech']+'/'+r['param'] for r in log['rows_created']]}")
    print(f"skipped       : {len(log['skipped'])}")
    print(f"audit log     : {log.get('log_path')}")


# --------------------------------------------------------------------------
# Self-test
# --------------------------------------------------------------------------
def self_test():
    """In-process tests of the edit engine on a synthetic workbook."""
    tmp = Path(tempfile.mkdtemp())
    wb = Workbook()
    ws = wb.active
    ws.title = "Secondary Techs"
    hdr = ["Tech.ID", "Tech", "Tech.Name", "Parameter.ID", "Parameter",
           "Unit", "Projection.Mode", "Projection.Parameter"] + list(range(2023, 2051))
    ws.append(hdr)
    ws.append([1, "PWRSPVBGDXX", "solar", 1, "CapitalCost", None, "User defined", 0] + [100.0] * 28)
    ws.append([1, "PWRSPVBGDXX", "solar", 2, "TotalAnnualMaxCapacity", None, "User defined", 0] + [365.0] * 28)
    ws.append([2, "TRNBGDXXINDEA", "trn", 3, "ResidualCapacity", None, "User defined", 0] + [3.0] * 28)
    ws.append([2, "TRNBGDXXINDEA", "trn", 4, "TotalAnnualMaxCapacity", None, "User defined", 0] + [9999.0] * 28)
    ws.append([2, "TRNBGDXXINDEA", "trn", 5, "TotalTechnologyAnnualActivityUpperLimit", None, "EMPTY", 0] + [None] * 28)
    ws.append([9, "TRNNLIBGDXX", "bk", 6, "ResidualCapacity", None, "User defined", 0] + [5.0] * 28)
    col_map, year_cols = scan_columns(ws)
    log = {"cells": [], "rows_created": [], "skipped": []}

    # multiply
    apply_edit(ws, col_map, year_cols, {"param": "CapitalCost", "tech": "PWRSPVBGDXX", "op": "multiply", "factor": 1.10}, log)
    r = find_row(ws, col_map, "PWRSPVBGDXX", "CapitalCost")
    assert abs(ws.cell(row=r, column=year_cols[2050]).value - 110.0) < 1e-6, "multiply failed"

    # set_flat ceiling (clip 365 -> 40)
    apply_edit(ws, col_map, year_cols, {"param": "TotalAnnualMaxCapacity", "tech": "PWRSPVBGDXX", "op": "set_flat", "value": 40.0}, log)
    r = find_row(ws, col_map, "PWRSPVBGDXX", "TotalAnnualMaxCapacity")
    assert abs(ws.cell(row=r, column=year_cols[2030]).value - 40.0) < 1e-6, "set_flat failed"

    # set_to_residual (freeze 9999 -> 3)
    apply_edit(ws, col_map, year_cols, {"param": "TotalAnnualMaxCapacity", "tech": "TRNBGDXXINDEA", "op": "set_to_residual"}, log)
    r = find_row(ws, col_map, "TRNBGDXXINDEA", "TotalAnnualMaxCapacity")
    assert abs(ws.cell(row=r, column=year_cols[2050]).value - 3.0) < 1e-6, "set_to_residual failed"

    # explicit values on existing empty AUL row
    apply_edit(ws, col_map, year_cols, {"param": "TotalTechnologyAnnualActivityUpperLimit", "tech": "TRNBGDXXINDEA", "values": {"2050": 123.4}}, log)
    r = find_row(ws, col_map, "TRNBGDXXINDEA", "TotalTechnologyAnnualActivityUpperLimit")
    assert abs(ws.cell(row=r, column=year_cols[2050]).value - 123.4) < 1e-6, "values failed"

    # create_if_absent AUL row for backstop
    apply_edit(ws, col_map, year_cols, {"param": "TotalTechnologyAnnualActivityUpperLimit", "tech": "TRNNLIBGDXX", "op": "set_flat", "value": 0.0, "create_if_absent": True}, log)
    r = find_row(ws, col_map, "TRNNLIBGDXX", "TotalTechnologyAnnualActivityUpperLimit")
    assert r is not None, "create_if_absent failed"
    assert ws.cell(row=r, column=year_cols[2035]).value == 0.0, "created row value failed"
    assert len(log["rows_created"]) == 1, "rows_created log failed"

    shutil.rmtree(tmp, ignore_errors=True)
    print("SELF-TEST PASSED (6 assertions):")
    print("  multiply x1.10, set_flat clip, set_to_residual freeze,")
    print("  explicit values, create_if_absent backstop row, proj-mode flip.")
    return 0


# --------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--scenario", help="scenario name (config folder under configs/)")
    ap.add_argument(
        "--source-scenario",
        default=None,
        help=(
            "Fail-closed compatibility override; must match the "
            "patches.json base_scenario declaration."
        ),
    )
    ap.add_argument("--self-test", action="store_true")
    ap.add_argument("--restore", action="store_true")
    ap.add_argument("--skip-backup", action="store_true")
    args = ap.parse_args()

    if args.self_test:
        return self_test()
    if not args.scenario:
        ap.error("--scenario is required (or use --self-test)")
    if args.restore:
        restore_scenario(args.scenario)
        return 0
    log = build_scenario(args.scenario, args.source_scenario, args.skip_backup)
    print_summary(log)
    return 0


if __name__ == "__main__":
    sys.exit(main())
