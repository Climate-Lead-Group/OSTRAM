# -*- coding: utf-8 -*-
"""
Created on 2025

@author: ClimateLeadGroup, Andrey Salazar-Vargas
"""

import os
import re
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
from typing import List, Dict, Any
from pathlib import Path
import yaml
from Z_AUX_config_loader import (
    get_olade_country_mapping, get_iso_country_map, get_code_to_energy,
    get_first_year, get_add_missing_countries_from_olade, get_pwr_cleanup_mode
)

def list_scenario_suffixes(base_dir: Path) -> List[str]:
    """Return list like ['BAU_NoRPO','NDC','NDC+ELC'] from folders 'A1_Outputs_*'."""
    suffixes: List[str] = []
    for item in sorted(base_dir.iterdir()):
        if item.is_dir() and item.name.startswith("A1_Outputs_"):
            suffix = item.name.split("A1_Outputs_", 1)[1]
            if suffix:  # Ensure non-empty
                suffixes.append(suffix)
    return suffixes


# Define folder paths relative to the script location (not cwd)
SCRIPT_DIR = Path(__file__).resolve().parent  # t1_confection/
INPUT_FOLDER = SCRIPT_DIR / "OG_csvs_inputs"
OUTPUT_FOLDER = SCRIPT_DIR / "A1_Outputs"
MISCELLANEOUS_FOLDER = SCRIPT_DIR / "Miscellaneous"
A2_EXTRA_INPUTS_FOLDER = SCRIPT_DIR / "A2_Extra_Inputs"
REGION_CONSOLIDATION_CONFIG = SCRIPT_DIR / "Config_region_consolidation.yaml"
TECH_COUNTRY_MATRIX_FILE = SCRIPT_DIR / "Tech_Country_Matrix.xlsx"
OLADE_GENERATION_FILE = SCRIPT_DIR / "OLADE - Capacidad instalada por fuente - Anual.xlsx"

# Model horizon years - data outside this range will be filtered/adjusted
LAST_YEAR = 2050

# Get FIRST_YEAR from centralized config
FIRST_YEAR = get_first_year()

# Default year range for sheets when no data is available
MODEL_YEARS = list(range(FIRST_YEAR, LAST_YEAR + 1))

# Country and technology mappings from centralized config
OLADE_COUNTRY_MAPPING = get_olade_country_mapping()
iso_country_map = get_iso_country_map()
code_to_energy = get_code_to_energy()

#-------------------------------------Formated functions--------------------------------------------#
def read_csv_files(input_dir):
    """Reads all CSV files in the given directory and returns a dictionary of DataFrames."""
    data_dict = {}
    for filename in os.listdir(input_dir):
        if filename.endswith(".csv"):
            file_path = os.path.join(input_dir, filename)
            df = pd.read_csv(file_path)
            key = os.path.splitext(filename)[0]
            data_dict[key] = df
    return data_dict


def normalize_temporal_profiles(og_data: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
    """
    Normalize temporal profiles to ensure they sum to 1.0 for proper OSeMOSYS input.

    This function normalizes:
    1. SpecifiedDemandProfile - by (REGION, FUEL, YEAR)
    2. YearSplit - by YEAR
    3. DaySplit - by grouping columns (YEAR, DAYTYPE, REGION if present)

    Args:
        og_data: Dictionary of DataFrames from OG_csvs_inputs

    Returns:
        Modified dictionary with normalized profiles
    """
    print("\n" + "🔧" * 40)
    print("🔧" + " " * 78 + "🔧")
    print("🔧" + " " * 20 + "⚡ NORMALIZING TEMPORAL PROFILES ⚡" + " " * 20 + "🔧")
    print("🔧" + " " * 78 + "🔧")
    print("🔧" * 40)
    print("\n[INFO] Ensuring all temporal profiles sum to 1.0 for OSeMOSYS compliance")
    print("=" * 80)

    normalization_results = {}

    # 1. Normalize SpecifiedDemandProfile
    if "SpecifiedDemandProfile" in og_data:
        print("\n🔍 [1/3] Checking SpecifiedDemandProfile...")
        df = og_data["SpecifiedDemandProfile"].copy()

        if not df.empty and "VALUE" in df.columns:
            # Check current state
            sums = df.groupby(["REGION", "FUEL", "YEAR"])["VALUE"].sum()
            problems_before = len(sums[abs(sums - 1.0) > 0.0001])

            if problems_before > 0:
                print(f"   ⚠️  Found {problems_before} combinations with incorrect sums")
                print(f"   🔧 Normalizing by (REGION, FUEL, YEAR)...")

                # Normalize
                df["sum_check"] = df.groupby(["REGION", "FUEL", "YEAR"])["VALUE"].transform("sum")
                df["VALUE"] = df["VALUE"] / df["sum_check"]
                df = df.drop(columns=["sum_check"])

                # Verify
                sums_after = df.groupby(["REGION", "FUEL", "YEAR"])["VALUE"].sum()
                all_correct = all(abs(sums_after - 1.0) < 0.0001)

                if all_correct:
                    og_data["SpecifiedDemandProfile"] = df
                    print(f"   ✅ SpecifiedDemandProfile normalized successfully!")
                    print(f"   📊 All {len(sums_after)} combinations now sum to 1.0")
                    normalization_results["SpecifiedDemandProfile"] = "NORMALIZED"
                else:
                    print(f"   ❌ Normalization failed for SpecifiedDemandProfile")
                    normalization_results["SpecifiedDemandProfile"] = "FAILED"
            else:
                print(f"   ✅ SpecifiedDemandProfile already normalized (all sums = 1.0)")
                normalization_results["SpecifiedDemandProfile"] = "OK"
        else:
            print(f"   ⚠️  SpecifiedDemandProfile is empty or missing VALUE column")
            normalization_results["SpecifiedDemandProfile"] = "SKIPPED"
    else:
        print("\n⚠️  [1/3] SpecifiedDemandProfile not found in data")
        normalization_results["SpecifiedDemandProfile"] = "NOT_FOUND"

    # 2. Normalize YearSplit
    if "YearSplit" in og_data:
        print("\n🔍 [2/3] Checking YearSplit...")
        df = og_data["YearSplit"].copy()

        if not df.empty and "VALUE" in df.columns:
            # Check if YEAR column exists
            if "YEAR" in df.columns:
                sums_before = df.groupby("YEAR")["VALUE"].sum()
                problems_before = len(sums_before[abs(sums_before - 1.0) > 0.0001])

                if problems_before > 0:
                    print(f"   ⚠️  Found {problems_before} years with incorrect sums")
                    print(f"   🔧 Normalizing by YEAR...")

                    # Normalize
                    df["sum_check"] = df.groupby("YEAR")["VALUE"].transform("sum")
                    df["VALUE"] = df["VALUE"] / df["sum_check"]
                    df = df.drop(columns=["sum_check"])

                    # Verify
                    sums_after = df.groupby("YEAR")["VALUE"].sum()
                    all_correct = all(abs(sums_after - 1.0) < 0.0001)

                    if all_correct:
                        og_data["YearSplit"] = df
                        print(f"   ✅ YearSplit normalized successfully!")
                        print(f"   📊 All {len(sums_after)} years now sum to 1.0")
                        normalization_results["YearSplit"] = "NORMALIZED"
                    else:
                        print(f"   ❌ Normalization failed for YearSplit")
                        normalization_results["YearSplit"] = "FAILED"
                else:
                    print(f"   ✅ YearSplit already normalized (all sums = 1.0)")
                    normalization_results["YearSplit"] = "OK"
            else:
                # No YEAR column - normalize entire dataset
                total = df["VALUE"].sum()
                if abs(total - 1.0) > 0.0001:
                    print(f"   ⚠️  Total sum = {total:.6f} (expected 1.0)")
                    print(f"   🔧 Normalizing entire dataset...")
                    df["VALUE"] = df["VALUE"] / total
                    og_data["YearSplit"] = df
                    print(f"   ✅ YearSplit normalized successfully!")
                    normalization_results["YearSplit"] = "NORMALIZED"
                else:
                    print(f"   ✅ YearSplit already normalized (sum = 1.0)")
                    normalization_results["YearSplit"] = "OK"
        else:
            print(f"   ⚠️  YearSplit is empty or missing VALUE column")
            normalization_results["YearSplit"] = "SKIPPED"
    else:
        print("\n⚠️  [2/3] YearSplit not found in data")
        normalization_results["YearSplit"] = "NOT_FOUND"

    # 3. Normalize DaySplit
    if "DaySplit" in og_data:
        print("\n🔍 [3/3] Checking DaySplit...")
        df = og_data["DaySplit"].copy()

        if not df.empty and "VALUE" in df.columns:
            # Identify grouping columns
            group_cols = [col for col in ["YEAR", "DAYTYPE", "REGION"] if col in df.columns]

            if group_cols:
                sums_before = df.groupby(group_cols)["VALUE"].sum()
                problems_before = len(sums_before[abs(sums_before - 1.0) > 0.0001])

                if problems_before > 0:
                    print(f"   ⚠️  Found {problems_before} combinations with incorrect sums")
                    print(f"   🔧 Normalizing by {group_cols}...")

                    # Normalize
                    df["sum_check"] = df.groupby(group_cols)["VALUE"].transform("sum")
                    df["VALUE"] = df["VALUE"] / df["sum_check"]
                    df = df.drop(columns=["sum_check"])

                    # Verify
                    sums_after = df.groupby(group_cols)["VALUE"].sum()
                    all_correct = all(abs(sums_after - 1.0) < 0.0001)

                    if all_correct:
                        og_data["DaySplit"] = df
                        print(f"   ✅ DaySplit normalized successfully!")
                        print(f"   📊 All {len(sums_after)} combinations now sum to 1.0")
                        normalization_results["DaySplit"] = "NORMALIZED"
                    else:
                        print(f"   ❌ Normalization failed for DaySplit")
                        normalization_results["DaySplit"] = "FAILED"
                else:
                    print(f"   ✅ DaySplit already normalized (all sums = 1.0)")
                    normalization_results["DaySplit"] = "OK"
            else:
                # No grouping columns - normalize entire dataset
                total = df["VALUE"].sum()
                if abs(total - 1.0) > 0.0001:
                    print(f"   ⚠️  Total sum = {total:.6f} (expected 1.0)")
                    print(f"   🔧 Normalizing entire dataset...")
                    df["VALUE"] = df["VALUE"] / total
                    og_data["DaySplit"] = df
                    print(f"   ✅ DaySplit normalized successfully!")
                    normalization_results["DaySplit"] = "NORMALIZED"
                else:
                    print(f"   ✅ DaySplit already normalized (sum = 1.0)")
                    normalization_results["DaySplit"] = "OK"
        else:
            print(f"   ⚠️  DaySplit is empty or missing VALUE column")
            normalization_results["DaySplit"] = "SKIPPED"
    else:
        print("\n⚠️  [3/3] DaySplit not found in data")
        normalization_results["DaySplit"] = "NOT_FOUND"

    # Summary
    print("\n" + "=" * 80)
    print("📋 NORMALIZATION SUMMARY:")
    print("=" * 80)
    for param, status in normalization_results.items():
        status_icon = {
            "NORMALIZED": "✅ ✨",
            "OK": "✅",
            "FAILED": "❌",
            "SKIPPED": "⚠️ ",
            "NOT_FOUND": "❓"
        }.get(status, "?")
        print(f"   {status_icon} {param}: {status}")

    print("\n" + "🔧" * 40)
    print("🔧" + " " * 78 + "🔧")
    print("🔧" + " " * 15 + "⚡ TEMPORAL PROFILES NORMALIZATION COMPLETE ⚡" + " " * 14 + "🔧")
    print("🔧" + " " * 78 + "🔧")
    print("🔧" * 40 + "\n")

    return og_data


def replace_country_codes(og_data: Dict[str, pd.DataFrame], old_code: str, new_code: str) -> Dict[str, pd.DataFrame]:
    """
    Replace country codes in all DataFrames.

    This function replaces occurrences of old_code with new_code in:
    - String columns that contain country codes (e.g., TECHNOLOGY, FUEL, REGION)
    - The replacement is done for codes embedded in longer strings (e.g., PWRHYDJAMXX -> PWRHYDBRBXX)
    - Also adds the new emission code (e.g., CO2BRB) to the EMISSION set if CO2+old_code exists

    Args:
        og_data: Dictionary of DataFrames from OG_csvs_inputs
        old_code: Country code to replace (e.g., 'JAM')
        new_code: New country code (e.g., 'BRB')

    Returns:
        Modified dictionary with replaced country codes
    """
    print(f"\n[Info] Replacing country code {old_code} -> {new_code} in all data")
    total_replacements = 0

    # First, add new emission code to EMISSION set if old code exists
    # Note: We keep the old emission code (e.g., CO2JAM) as a backup for future use
    if 'EMISSION' in og_data:
        emission_df = og_data['EMISSION']
        old_emission = f"CO2{old_code}"
        new_emission = f"CO2{new_code}"

        # Check if old emission exists and new one doesn't
        if 'VALUE' in emission_df.columns:
            has_old = (emission_df['VALUE'] == old_emission).any()
            has_new = (emission_df['VALUE'] == new_emission).any()

            if has_old and not has_new:
                # Add new emission code (keeping the old one as backup)
                new_row = pd.DataFrame({'VALUE': [new_emission]})
                og_data['EMISSION'] = pd.concat([emission_df, new_row], ignore_index=True)
                print(f"[Info] Added {new_emission} to EMISSION set (keeping {old_emission} as backup)")

    # Now replace all occurrences EXCEPT in the EMISSION set (to preserve CO2JAM as backup)
    for param_name, df in og_data.items():
        # Skip the EMISSION set - we want to keep both CO2JAM and CO2BRB there
        if param_name == 'EMISSION':
            continue

        for col in df.columns:
            if df[col].dtype == 'object':  # String columns
                # Count replacements before applying
                mask = df[col].astype(str).str.contains(old_code, na=False)
                replacements_in_col = mask.sum()

                if replacements_in_col > 0:
                    df[col] = df[col].astype(str).str.replace(old_code, new_code, regex=False)
                    total_replacements += replacements_in_col

    print(f"[Info] Replaced {total_replacements} occurrences of {old_code} -> {new_code}")
    return og_data


def write_csv_files(og_data: Dict[str, pd.DataFrame], output_folder: Path) -> None:
    """
    Write all DataFrames back to CSV files in the output folder.

    This function writes the processed OG_Input_Data back to OG_csvs_inputs
    so that B1_Compiler picks up the correct EMISSION set (with CO2BRB).

    Args:
        og_data: Dictionary of DataFrames to write
        output_folder: Path to the output folder (typically OG_csvs_inputs)
    """
    print(f"\n[Info] Writing updated CSV files to {output_folder}")
    written_count = 0

    for param_name, df in og_data.items():
        output_path = output_folder / f"{param_name}.csv"
        df.to_csv(output_path, index=False)
        written_count += 1

    print(f"[Info] Written {written_count} CSV files to {output_folder}")


def filter_data_by_first_year(og_data: Dict[str, pd.DataFrame], first_year: int) -> Dict[str, pd.DataFrame]:
    """
    Filter all DataFrames to only include data from first_year onwards.

    Args:
        og_data: Dictionary of DataFrames from OG_csvs_inputs
        first_year: First year to include in the data (e.g., 2021)

    Returns:
        Modified dictionary with filtered DataFrames
    """
    print(f"\n[Info] Filtering data to start from year {first_year}")
    total_rows_filtered = 0

    for param_name, df in og_data.items():
        if "YEAR" not in df.columns:
            continue

        rows_before = len(df)
        og_data[param_name] = df[df["YEAR"] >= first_year].copy()
        rows_filtered = rows_before - len(og_data[param_name])

        if rows_filtered > 0:
            total_rows_filtered += rows_filtered

    # Also filter the YEAR set itself (which has years in VALUE column, not YEAR column)
    if "YEAR" in og_data and "VALUE" in og_data["YEAR"].columns:
        rows_before = len(og_data["YEAR"])
        og_data["YEAR"] = og_data["YEAR"][og_data["YEAR"]["VALUE"] >= first_year].copy()
        rows_filtered = rows_before - len(og_data["YEAR"])
        if rows_filtered > 0:
            total_rows_filtered += rows_filtered
            print(f"[Info] Filtered YEAR set: removed {rows_filtered} years before {first_year}")

    if total_rows_filtered > 0:
        print(f"[Info] Filtered out {total_rows_filtered} rows with YEAR < {first_year}")

    return og_data


def read_olade_generation_data():
    """
    Read OLADE electricity generation data from Excel file.
    Used to add missing countries (like HTI) that may not be in the OSeMOSYS input CSVs.

    Note: OLADE data is in GWh, converted to PJ for the model (1 GWh = 0.0036 PJ)

    Returns:
        dict: {
            'reference_year': int,
            'data': {
                country_iso3: generation_pj
            }
        }
        Returns None if file not found.
    """
    if not OLADE_GENERATION_FILE.exists():
        print(f"[Warning] OLADE generation file not found: {OLADE_GENERATION_FILE}")
        return None

    try:
        from openpyxl import load_workbook as load_wb_olade
        wb = load_wb_olade(OLADE_GENERATION_FILE, data_only=True)

        # Find the sheet with generation data (format: "1.YYYY")
        ref_year = FIRST_YEAR  # Use the dynamically detected year
        sheet_name = f"1.{ref_year}"
        if sheet_name not in wb.sheetnames:
            # Fallback: find any sheet starting with "1."
            for sn in wb.sheetnames:
                if sn.startswith("1."):
                    sheet_name = sn
                    ref_year = int(sn.split(".")[1])
                    break
        ws = wb[sheet_name]

        # Get country columns from row 5 (starting at column 3)
        country_columns = {}
        for col_idx in range(3, ws.max_column + 1):
            country_name = ws.cell(5, col_idx).value
            if country_name and str(country_name) in OLADE_COUNTRY_MAPPING:
                iso3_code = OLADE_COUNTRY_MAPPING[str(country_name)]
                country_columns[col_idx] = iso3_code

        # Read Total generation from row 21
        data = {}
        for col_idx, country_iso3 in country_columns.items():
            total_gwh = ws.cell(21, col_idx).value  # Row 21 = "Total"

            if total_gwh is not None and total_gwh != '':
                try:
                    generation_gwh = float(total_gwh)
                    # Convert from GWh to PJ (1 GWh = 0.0036 PJ)
                    generation_pj = generation_gwh * 0.0036
                    data[country_iso3] = generation_pj
                except ValueError:
                    pass

        wb.close()

        return {
            'reference_year': ref_year,
            'data': data
        }
    except Exception as e:
        print(f"[Warning] Could not read OLADE generation data: {e}")
        return None


#-------------------------------------Region Consolidation Functions---------------------------------#
def load_region_consolidation_config() -> Dict[str, Any]:
    """Load region consolidation configuration from YAML file."""
    if not REGION_CONSOLIDATION_CONFIG.exists():
        print(f"[Warning] Region consolidation config not found: {REGION_CONSOLIDATION_CONFIG}")
        return {"enabled": False}

    with open(REGION_CONSOLIDATION_CONFIG, "r", encoding="utf-8") as f:
        config = yaml.safe_load(f)
    return config


def get_column_for_consolidation(df: pd.DataFrame) -> str:
    """Determine which column contains technology/fuel/storage codes for consolidation."""
    if "TECHNOLOGY" in df.columns:
        return "TECHNOLOGY"
    elif "FUEL" in df.columns:
        return "FUEL"
    elif "STORAGE" in df.columns:
        return "STORAGE"
    return None


def find_country_region_in_code(code: str, country: str, regions: List[str]) -> str:
    """
    Find if a code contains a country+region pattern and return the region found.
    Returns None if no match found.

    Patterns checked:
    - Position 3-5 for country (e.g., ELCBRACN01 -> BRA at pos 3-6, CN at pos 6-8)
    - Position 6-8 for country (e.g., PWRGASBRACN00 -> BRA at pos 6-9, CN at pos 9-11)
    - For TRN interconnections: check both endpoints
    """
    code_upper = code.upper()

    # Pattern 1: Country at position 3-6 (e.g., ELCBRACN01, SDSBRAXX01)
    if len(code) >= 8 and code_upper[3:6] == country:
        region = code_upper[6:8]
        if region in regions:
            return region

    # Pattern 2: Country at position 6-9 (e.g., PWRGASBRACN00)
    if len(code) >= 11 and code_upper[6:9] == country:
        region = code_upper[9:11]
        if region in regions:
            return region

    # Pattern 3: TRN interconnections (TRN + 5chars + 5chars, e.g., TRNBRACNARGXX)
    if code_upper.startswith("TRN") and len(code) >= 13:
        # First endpoint: positions 3-8 (country 3-6, region 6-8)
        if code_upper[3:6] == country:
            region = code_upper[6:8]
            if region in regions:
                return region
        # Second endpoint: positions 8-13 (country 8-11, region 11-13)
        if code_upper[8:11] == country:
            region = code_upper[11:13]
            if region in regions:
                return region

    return None


def replace_region_in_code(code: str, country: str, regions: List[str], unified: str) -> str:
    """
    Replace regional codes with unified region code.
    Handles multiple patterns and TRN interconnections with both endpoints.
    """
    code_upper = code.upper()

    # Pattern 3: TRN interconnections - replace both endpoints if needed (check first!)
    if code_upper.startswith("TRN") and len(code) >= 13:
        modified = list(code)
        # First endpoint (positions 3-8: country at 3-6, region at 6-8)
        if code_upper[3:6] == country and code_upper[6:8] in regions:
            modified[6:8] = list(unified)
        # Second endpoint (positions 8-13: country at 8-11, region at 11-13)
        if code_upper[8:11] == country and code_upper[11:13] in regions:
            modified[11:13] = list(unified)
        return "".join(modified)

    # Pattern 1: Country at position 3-6 (e.g., ELCBRACN01 -> ELCBRAXX01)
    if len(code) >= 8 and code_upper[3:6] == country:
        region = code_upper[6:8]
        if region in regions:
            return code[:6] + unified + code[8:]

    # Pattern 2: Country at position 6-9 (e.g., PWRGASBRACN00 -> PWRGASBRAXX00)
    if len(code) >= 11 and code_upper[6:9] == country:
        region = code_upper[9:11]
        if region in regions:
            return code[:9] + unified + code[11:]

    return code


def create_grouping_key(code: str, country: str, regions: List[str], unified: str) -> str:
    """Create a normalized key for grouping rows that should be consolidated."""
    return replace_region_in_code(code, country, regions, unified)


def consolidate_dataframe(
    df: pd.DataFrame,
    param_name: str,
    column: str,
    country: str,
    regions: List[str],
    unified: str,
    agg_method: str
) -> pd.DataFrame:
    """
    Consolidate a DataFrame by replacing regional codes with unified codes.

    Args:
        df: Input DataFrame
        param_name: Parameter name (for logging)
        column: Column containing codes to consolidate (TECHNOLOGY/FUEL/STORAGE)
        country: Country code (e.g., "BRA")
        regions: List of region codes to consolidate (e.g., ["CN", "NW", ...])
        unified: Unified region code (e.g., "XX")
        agg_method: Aggregation method ("avg" or "sum")

    Returns:
        Consolidated DataFrame
    """
    if df.empty:
        return df

    # First, transform FUEL column if it exists and has regional codes
    # This ensures FUEL codes like BIOBRACN become BIOBRAXX before grouping
    has_fuel_with_regions = False
    if "FUEL" in df.columns and column != "FUEL":
        fuel_mask = df["FUEL"].apply(
            lambda x: find_country_region_in_code(str(x), country, regions) is not None if pd.notna(x) else False
        )
        if fuel_mask.any():
            has_fuel_with_regions = True
            df = df.copy()
            df.loc[fuel_mask, "FUEL"] = df.loc[fuel_mask, "FUEL"].apply(
                lambda x: replace_region_in_code(str(x), country, regions, unified)
            )

    # Also transform STORAGE column if it exists and has regional codes
    # This ensures STORAGE codes like LDSBRACN01 become LDSBRAXX01 before grouping
    has_storage_with_regions = False
    if "STORAGE" in df.columns and column != "STORAGE":
        storage_mask = df["STORAGE"].apply(
            lambda x: find_country_region_in_code(str(x), country, regions) is not None if pd.notna(x) else False
        )
        if storage_mask.any():
            has_storage_with_regions = True
            if not has_fuel_with_regions:  # df not yet copied
                df = df.copy()
            df.loc[storage_mask, "STORAGE"] = df.loc[storage_mask, "STORAGE"].apply(
                lambda x: replace_region_in_code(str(x), country, regions, unified)
            )

    # Identify rows that need consolidation (based on main column)
    mask = df[column].apply(lambda x: find_country_region_in_code(str(x), country, regions) is not None)

    # Also check if FUEL column has regional codes that need consolidation
    if "FUEL" in df.columns and column != "FUEL":
        fuel_mask = df["FUEL"].apply(
            lambda x: find_country_region_in_code(str(x), country, regions) is not None if pd.notna(x) else False
        )
        mask = mask | fuel_mask

    # Also check if STORAGE column has regional codes that need consolidation
    if "STORAGE" in df.columns and column != "STORAGE":
        storage_mask = df["STORAGE"].apply(
            lambda x: find_country_region_in_code(str(x), country, regions) is not None if pd.notna(x) else False
        )
        mask = mask | storage_mask

    if not mask.any():
        return df  # No rows to consolidate

    # Separate rows: those to consolidate and those to keep as-is
    df_to_consolidate = df[mask].copy()
    df_unchanged = df[~mask].copy()

    # Create grouping key (normalized code for main column)
    df_to_consolidate["_group_key"] = df_to_consolidate[column].apply(
        lambda x: create_grouping_key(str(x), country, regions, unified)
    )

    # Also normalize FUEL column for grouping if it exists
    if "FUEL" in df_to_consolidate.columns and column != "FUEL":
        df_to_consolidate["_fuel_key"] = df_to_consolidate["FUEL"].apply(
            lambda x: replace_region_in_code(str(x), country, regions, unified) if pd.notna(x) else x
        )

    # Also normalize STORAGE column for grouping if it exists
    if "STORAGE" in df_to_consolidate.columns and column != "STORAGE":
        df_to_consolidate["_storage_key"] = df_to_consolidate["STORAGE"].apply(
            lambda x: replace_region_in_code(str(x), country, regions, unified) if pd.notna(x) else x
        )

    # Determine grouping columns (all columns except VALUE and the target column)
    # For most CSVs, we group by all non-numeric columns except the one being consolidated
    group_cols = ["_group_key"]

    # Add other categorical columns to grouping
    for col in df_to_consolidate.columns:
        if col in ["_group_key", "_fuel_key", "_storage_key", column, "VALUE", "FUEL", "STORAGE"]:
            continue
        if col in ["YEAR", "TIMESLICE", "MODE_OF_OPERATION", "EMISSION", "DAILYTIMEBRACKET", "REGION"]:
            group_cols.append(col)

    # Add normalized FUEL key to grouping if it exists
    if "_fuel_key" in df_to_consolidate.columns:
        group_cols.append("_fuel_key")

    # Add normalized STORAGE key to grouping if it exists
    if "_storage_key" in df_to_consolidate.columns:
        group_cols.append("_storage_key")

    # Perform aggregation
    if "VALUE" in df_to_consolidate.columns:
        if agg_method == "avg":
            agg_func = "mean"
        else:  # sum
            agg_func = "sum"

        # Group and aggregate
        df_consolidated = df_to_consolidate.groupby(group_cols, as_index=False).agg({"VALUE": agg_func})

        # Restore the consolidated code
        df_consolidated[column] = df_consolidated["_group_key"]
        df_consolidated.drop(columns=["_group_key"], inplace=True)

        # Restore FUEL from _fuel_key if it exists
        if "_fuel_key" in df_consolidated.columns:
            df_consolidated["FUEL"] = df_consolidated["_fuel_key"]
            df_consolidated.drop(columns=["_fuel_key"], inplace=True)

        # Restore STORAGE from _storage_key if it exists
        if "_storage_key" in df_consolidated.columns:
            df_consolidated["STORAGE"] = df_consolidated["_storage_key"]
            df_consolidated.drop(columns=["_storage_key"], inplace=True)
    else:
        # For DataFrames without VALUE column, just deduplicate
        df_consolidated = df_to_consolidate.drop_duplicates(subset=["_group_key"]).copy()
        df_consolidated[column] = df_consolidated["_group_key"]
        df_consolidated.drop(columns=["_group_key"], inplace=True)
        if "_fuel_key" in df_consolidated.columns:
            df_consolidated["FUEL"] = df_consolidated["_fuel_key"]
            df_consolidated.drop(columns=["_fuel_key"], inplace=True)
        if "_storage_key" in df_consolidated.columns:
            df_consolidated["STORAGE"] = df_consolidated["_storage_key"]
            df_consolidated.drop(columns=["_storage_key"], inplace=True)

    # Combine unchanged and consolidated rows
    result = pd.concat([df_unchanged, df_consolidated], ignore_index=True)

    # Log consolidation
    original_count = len(df_to_consolidate)
    consolidated_count = len(df_consolidated)
    if original_count > consolidated_count:
        print(f"    {param_name}: {original_count} rows -> {consolidated_count} rows ({agg_method})")

    return result


def is_internal_interconnection(code: str, country: str, unified: str) -> bool:
    """
    Check if a TRN code represents an internal interconnection within the same country.

    Internal interconnections are those where both endpoints are the same country
    with the unified region code (e.g., TRNBRAXXBRAXX).

    Args:
        code: Technology code to check
        country: Country code (e.g., "BRA")
        unified: Unified region code (e.g., "XX")

    Returns:
        True if this is an internal interconnection that should be removed
    """
    code_upper = str(code).upper()

    # Check if it's a TRN interconnection code
    if not code_upper.startswith("TRN") or len(code) < 13:
        return False

    # Extract both endpoints
    # Pattern: TRN + COUNTRY1(3) + REGION1(2) + COUNTRY2(3) + REGION2(2)
    endpoint1_country = code_upper[3:6]
    endpoint1_region = code_upper[6:8]
    endpoint2_country = code_upper[8:11]
    endpoint2_region = code_upper[11:13]

    # Check if both endpoints are the same country with unified region
    if (endpoint1_country == country and endpoint1_region == unified and
        endpoint2_country == country and endpoint2_region == unified):
        return True

    return False


def remove_internal_interconnections(
    og_data: Dict[str, pd.DataFrame],
    country: str,
    unified: str
) -> Dict[str, pd.DataFrame]:
    """
    Remove internal interconnections for a consolidated country.

    Args:
        og_data: Dictionary of DataFrames
        country: Country code (e.g., "BRA")
        unified: Unified region code (e.g., "XX")

    Returns:
        Modified dictionary with internal interconnections removed
    """
    total_removed = 0

    for param_name, df in og_data.items():
        column = get_column_for_consolidation(df)
        if column is None:
            continue

        # Find internal interconnections
        mask = df[column].apply(lambda x: is_internal_interconnection(str(x), country, unified))

        if mask.any():
            rows_before = len(df)
            og_data[param_name] = df[~mask].copy()
            rows_removed = rows_before - len(og_data[param_name])
            if rows_removed > 0:
                total_removed += rows_removed
                print(f"    {param_name}: removed {rows_removed} internal interconnection rows")

    return og_data, total_removed


def consolidate_regions(og_data: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
    """
    Consolidate multiple regions into unified regions based on configuration.

    Args:
        og_data: Dictionary of DataFrames from OG_csvs_inputs

    Returns:
        Modified dictionary with consolidated DataFrames
    """
    config = load_region_consolidation_config()

    if not config.get("enabled", False):
        print("[Info] Region consolidation is disabled.")
        return og_data

    print("\n" + "=" * 70)
    print("REGION CONSOLIDATION")
    print("=" * 70)

    countries = config.get("countries", {})
    agg_rules = config.get("aggregation_rules", {})

    avg_params = set(agg_rules.get("avg", []))
    sum_params = set(agg_rules.get("sum", []))
    disabled_params = set(agg_rules.get("disabled", []))

    # Print disabled parameters
    if disabled_params:
        print("\n[Info] Disabled parameters (no consolidation):")
        for param in sorted(disabled_params):
            if param in og_data:
                print(f"    - {param}")

    # Process each country
    for country_code, country_config in countries.items():
        regions = country_config.get("regions", [])
        unified = country_config.get("unified_region", "XX")

        print(f"\n[Processing] Country: {country_code}")
        print(f"    Regions to consolidate: {regions} -> {unified}")

        # Process each DataFrame
        for param_name, df in og_data.items():
            # Skip disabled parameters
            if param_name in disabled_params:
                continue

            # Determine aggregation method
            if param_name in avg_params:
                agg_method = "avg"
            elif param_name in sum_params:
                agg_method = "sum"
            else:
                # Parameter not in any list - skip
                continue

            # Find the column to consolidate
            column = get_column_for_consolidation(df)
            if column is None:
                continue

            # Consolidate the DataFrame
            og_data[param_name] = consolidate_dataframe(
                df=df,
                param_name=param_name,
                column=column,
                country=country_code,
                regions=regions,
                unified=unified,
                agg_method=agg_method
            )

        # Remove internal interconnections for this country
        print(f"\n    Removing internal interconnections ({country_code}{unified}{country_code}{unified}):")
        og_data, removed_count = remove_internal_interconnections(og_data, country_code, unified)
        if removed_count == 0:
            print(f"    (no internal interconnections found)")

    print("\n" + "=" * 70)
    print("Region consolidation completed.")
    print("=" * 70 + "\n")

    return og_data


def clean_pwr_technologies(og_data: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
    """
    Clean PWR technology codes by:
    1. Removing PWR technologies ending in '00' when a '01' version exists
    2. Renaming PWR technologies ending in '01' to remove the suffix

    Args:
        og_data: Dictionary of DataFrames from OG_csvs_inputs

    Returns:
        Modified dictionary with cleaned PWR technology codes
    """
    print("\n" + "=" * 70)
    print("PWR TECHNOLOGY CLEANUP")
    print("=" * 70)

    # First, collect all PWR technologies across all DataFrames
    all_pwr_techs = set()
    for param_name, df in og_data.items():
        if "TECHNOLOGY" in df.columns:
            pwr_techs = df[df["TECHNOLOGY"].str.startswith("PWR", na=False)]["TECHNOLOGY"].unique()
            all_pwr_techs.update(pwr_techs)

    # Find PWR00 technologies that have a PWR01 counterpart
    pwr00_to_remove = set()
    for tech in all_pwr_techs:
        if tech.endswith("00"):
            tech01 = tech[:-2] + "01"
            if tech01 in all_pwr_techs:
                pwr00_to_remove.add(tech)

    if pwr00_to_remove:
        print(f"\n[Step 1] Removing PWR00 technologies (PWR01 version exists):")
        print(f"    Found {len(pwr00_to_remove)} PWR00 technologies to remove")
        for tech in sorted(list(pwr00_to_remove))[:10]:
            print(f"    - {tech}")
        if len(pwr00_to_remove) > 10:
            print(f"    ... and {len(pwr00_to_remove) - 10} more")
    else:
        print(f"\n[Step 1] No PWR00 technologies to remove (no PWR01 counterparts found)")

    # Remove PWR00 technologies from all DataFrames
    total_rows_removed = 0
    for param_name, df in og_data.items():
        if "TECHNOLOGY" in df.columns:
            mask = df["TECHNOLOGY"].isin(pwr00_to_remove)
            if mask.any():
                rows_before = len(df)
                og_data[param_name] = df[~mask].copy()
                rows_removed = rows_before - len(og_data[param_name])
                total_rows_removed += rows_removed

    if total_rows_removed > 0:
        print(f"    Total rows removed: {total_rows_removed}")

    # Now rename PWR01 to PWR (remove '01' suffix)
    print(f"\n[Step 2] Renaming PWR01 technologies (removing '01' suffix):")

    # Find all PWR01 technologies (after removal of PWR00)
    pwr01_techs = set()
    for param_name, df in og_data.items():
        if "TECHNOLOGY" in df.columns:
            pwr_techs = df[df["TECHNOLOGY"].str.startswith("PWR", na=False) &
                          df["TECHNOLOGY"].str.endswith("01", na=False)]["TECHNOLOGY"].unique()
            pwr01_techs.update(pwr_techs)

    if pwr01_techs:
        print(f"    Found {len(pwr01_techs)} PWR01 technologies to rename")
        for tech in sorted(list(pwr01_techs))[:10]:
            print(f"    - {tech} -> {tech[:-2]}")
        if len(pwr01_techs) > 10:
            print(f"    ... and {len(pwr01_techs) - 10} more")
    else:
        print(f"    No PWR01 technologies found to rename")

    # Create mapping for renaming
    rename_map = {tech: tech[:-2] for tech in pwr01_techs}

    # Apply renaming to all DataFrames
    total_rows_renamed = 0
    for param_name, df in og_data.items():
        if "TECHNOLOGY" in df.columns:
            mask = df["TECHNOLOGY"].isin(pwr01_techs)
            if mask.any():
                og_data[param_name] = df.copy()
                og_data[param_name].loc[mask, "TECHNOLOGY"] = og_data[param_name].loc[mask, "TECHNOLOGY"].map(
                    lambda x: rename_map.get(x, x)
                )
                total_rows_renamed += mask.sum()

    if total_rows_renamed > 0:
        print(f"    Total rows renamed: {total_rows_renamed}")

    print("\n" + "=" * 70)
    print("PWR technology cleanup completed.")
    print("=" * 70 + "\n")

    return og_data


def merge_pwr_technologies(og_data: Dict[str, pd.DataFrame], matrix_config: Dict[str, Any]) -> Dict[str, pd.DataFrame]:
    """
    Merge PWR technology codes by aggregating PWR00 values into PWR01 before removal.

    Unlike clean_pwr_technologies (which simply drops PWR00), this function:
    1. Finds PWR00 technologies that have a PWR01 counterpart
    2. Aggregates PWR00 VALUE into the matching PWR01 rows using the aggregation
       rules from Tech_Country_Matrix.xlsx (sum or avg per parameter)
    3. Removes PWR00 rows (their contribution is now in PWR01)
    4. Renames PWR01 technologies by removing the '01' suffix

    Args:
        og_data: Dictionary of DataFrames from OG_csvs_inputs
        matrix_config: Configuration from load_tech_country_matrix() with aggregation_rules

    Returns:
        Modified dictionary with merged PWR technology codes
    """
    print("\n" + "=" * 70)
    print("PWR TECHNOLOGY MERGE (00 -> 01)")
    print("=" * 70)

    # Load aggregation rules
    agg_rules = matrix_config.get("aggregation_rules", {})
    avg_params = set(agg_rules.get("avg", []))
    sum_params = set(agg_rules.get("sum", []))
    disabled_params = set(agg_rules.get("disabled", []))

    # Collect all PWR technologies across all DataFrames
    all_pwr_techs = set()
    for param_name, df in og_data.items():
        if "TECHNOLOGY" in df.columns:
            pwr_techs = df[df["TECHNOLOGY"].str.startswith("PWR", na=False)]["TECHNOLOGY"].unique()
            all_pwr_techs.update(pwr_techs)

    # Find PWR00 technologies that have a PWR01 counterpart
    pwr00_with_01 = set()
    for tech in all_pwr_techs:
        if tech.endswith("00"):
            tech01 = tech[:-2] + "01"
            if tech01 in all_pwr_techs:
                pwr00_with_01.add(tech)

    if pwr00_with_01:
        print(f"\n[Step 1] Merging PWR00 values into PWR01:")
        print(f"    Found {len(pwr00_with_01)} PWR00 technologies to merge")
        for tech in sorted(list(pwr00_with_01))[:10]:
            print(f"    - {tech} -> {tech[:-2]}01")
        if len(pwr00_with_01) > 10:
            print(f"    ... and {len(pwr00_with_01) - 10} more")
    else:
        print(f"\n[Step 1] No PWR00 technologies to merge (no PWR01 counterparts found)")

    if disabled_params:
        print(f"\n    Disabled parameters (no merge): {', '.join(sorted(disabled_params))}")

    # Merge PWR00 values into PWR01 for each DataFrame
    total_rows_merged = 0
    for param_name, df in og_data.items():
        if "TECHNOLOGY" not in df.columns:
            continue

        # Skip disabled parameters
        if param_name in disabled_params:
            continue

        # Determine aggregation method for this parameter
        if param_name in avg_params:
            agg_func = "mean"
        elif param_name in sum_params:
            agg_func = "sum"
        else:
            # Parameter not in any aggregation list - skip merge, just drop PWR00
            continue

        # Check if this DataFrame has any PWR00 rows to merge
        has_pwr00 = df["TECHNOLOGY"].isin(pwr00_with_01).any()
        if not has_pwr00:
            continue

        if "VALUE" in df.columns:
            # Identify grouping columns (everything except TECHNOLOGY and VALUE)
            group_cols = [c for c in df.columns if c not in ("TECHNOLOGY", "VALUE")]

            # Separate: PWR00, PWR01 (with counterpart), and everything else
            mask00 = df["TECHNOLOGY"].isin(pwr00_with_01)
            pwr01_set = {tech[:-2] + "01" for tech in pwr00_with_01}
            mask01 = df["TECHNOLOGY"].isin(pwr01_set)

            df_00 = df[mask00].copy()
            df_01 = df[mask01].copy()
            df_rest = df[~mask00 & ~mask01].copy()

            # Map PWR00 tech names to their PWR01 equivalents for joining
            df_00["TECHNOLOGY"] = df_00["TECHNOLOGY"].apply(lambda x: x[:-2] + "01")

            # Concatenate PWR00 (now renamed to 01) with PWR01 and aggregate
            df_combined = pd.concat([df_01, df_00], ignore_index=True)
            if group_cols:
                df_merged = df_combined.groupby(
                    ["TECHNOLOGY"] + group_cols, as_index=False
                ).agg({"VALUE": agg_func})
            else:
                df_merged = df_combined.groupby(
                    ["TECHNOLOGY"], as_index=False
                ).agg({"VALUE": agg_func})

            rows_before = len(df_01) + len(df_00)
            rows_after = len(df_merged)
            if rows_before > rows_after:
                total_rows_merged += (rows_before - rows_after)
                print(f"    {param_name}: {len(df_00)} PWR00 + {len(df_01)} PWR01 -> {rows_after} rows ({agg_func})")

            og_data[param_name] = pd.concat([df_rest, df_merged], ignore_index=True)
        else:
            # For set DataFrames (no VALUE column), just remove PWR00 entries
            og_data[param_name] = df[~df["TECHNOLOGY"].isin(pwr00_with_01)].copy()

    if total_rows_merged > 0:
        print(f"\n    Total rows merged: {total_rows_merged}")

    # Now rename PWR01 to remove '01' suffix (same as clean_pwr_technologies Step 2)
    print(f"\n[Step 2] Renaming PWR01 technologies (removing '01' suffix):")

    pwr01_techs = set()
    for param_name, df in og_data.items():
        if "TECHNOLOGY" in df.columns:
            pwr_techs = df[df["TECHNOLOGY"].str.startswith("PWR", na=False) &
                          df["TECHNOLOGY"].str.endswith("01", na=False)]["TECHNOLOGY"].unique()
            pwr01_techs.update(pwr_techs)

    if pwr01_techs:
        print(f"    Found {len(pwr01_techs)} PWR01 technologies to rename")
        for tech in sorted(list(pwr01_techs))[:10]:
            print(f"    - {tech} -> {tech[:-2]}")
        if len(pwr01_techs) > 10:
            print(f"    ... and {len(pwr01_techs) - 10} more")
    else:
        print(f"    No PWR01 technologies found to rename")

    rename_map = {tech: tech[:-2] for tech in pwr01_techs}

    total_rows_renamed = 0
    for param_name, df in og_data.items():
        if "TECHNOLOGY" in df.columns:
            mask = df["TECHNOLOGY"].isin(pwr01_techs)
            if mask.any():
                og_data[param_name] = df.copy()
                og_data[param_name].loc[mask, "TECHNOLOGY"] = og_data[param_name].loc[mask, "TECHNOLOGY"].map(
                    lambda x: rename_map.get(x, x)
                )
                total_rows_renamed += mask.sum()

    if total_rows_renamed > 0:
        print(f"    Total rows renamed: {total_rows_renamed}")

    print("\n" + "=" * 70)
    print("PWR technology merge completed.")
    print("=" * 70 + "\n")

    return og_data


#-------------------------------------Tech-Country Matrix Functions---------------------------------#
def load_tech_country_matrix() -> Dict[str, Any]:
    """
    Load the technology-country matrix configuration from Excel file.

    Returns:
        Dictionary with:
        - 'enabled': bool - whether matrix filtering is enabled
        - 'matrix': dict - {tech: {country: bool}} mapping
        - 'ngs_enabled': bool - whether NGS unification is enabled
        - 'aggregation_rules': dict - aggregation rules for NGS unification
    """
    if not TECH_COUNTRY_MATRIX_FILE.exists():
        print(f"[Info] Tech-Country Matrix file not found: {TECH_COUNTRY_MATRIX_FILE}")
        return {"enabled": False}

    try:
        # Read the "Enable Matrix Filtering" flag from row 1
        df_flag = pd.read_excel(TECH_COUNTRY_MATRIX_FILE, sheet_name="Matrix", header=None, nrows=1)
        matrix_filtering_enabled = str(df_flag.iloc[0, 1]).upper() == "YES"

        # Read the Matrix sheet (skip flag row + empty spacer = 2 rows)
        df_matrix = pd.read_excel(TECH_COUNTRY_MATRIX_FILE, sheet_name="Matrix", index_col=0, skiprows=2)

        # Convert to dict: {tech: {country: True/False}}
        matrix = {}
        for tech in df_matrix.index:
            matrix[tech] = {}
            for country in df_matrix.columns:
                value = df_matrix.loc[tech, country]
                matrix[tech][country] = str(value).upper() == "YES"

        # Read NGS_Unification sheet
        df_ngs = pd.read_excel(TECH_COUNTRY_MATRIX_FILE, sheet_name="NGS_Unification", header=None)
        ngs_enabled = False
        for idx, row in df_ngs.iterrows():
            if pd.notna(row[0]) and "Enable NGS Unification" in str(row[0]):
                ngs_enabled = str(row[1]).upper() == "YES"
                break

        # Read aggregation rules
        df_agg = pd.read_excel(TECH_COUNTRY_MATRIX_FILE, sheet_name="Aggregation_Rules", skiprows=4)

        agg_rules = {"avg": [], "sum": [], "disabled": []}
        for _, row in df_agg.iterrows():
            param = row.get("Parameter", "")
            agg_type = str(row.get("Aggregation Type", "")).upper()
            if pd.notna(param) and param:
                if agg_type == "AVG":
                    agg_rules["avg"].append(param)
                elif agg_type == "SUM":
                    agg_rules["sum"].append(param)
                elif agg_type == "DISABLED":
                    agg_rules["disabled"].append(param)

        return {
            "enabled": True,
            "matrix_filtering_enabled": matrix_filtering_enabled,
            "matrix": matrix,
            "ngs_enabled": ngs_enabled,
            "aggregation_rules": agg_rules
        }

    except Exception as e:
        print(f"[Error] Failed to load Tech-Country Matrix: {e}")
        return {"enabled": False}


def extract_tech_country_from_code(code: str) -> tuple:
    """
    Extract technology sub-code and country from a technology code.

    Args:
        code: Technology code (e.g., PWRBIOARGXX, PWRCCGCOLXX01)

    Returns:
        Tuple of (tech_subcode, country) or (None, None) if not parseable
    """
    code_upper = str(code).upper()

    # PWR technologies: PWRXXXCCCRRSS or PWRXXXCCCRR (with or without suffix)
    # PWR(3) + subcode(3) + country(3) + region(2) = 11 chars minimum
    if code_upper.startswith("PWR") and len(code) >= 11:
        sub_code = code_upper[3:6]  # e.g., BIO, CCG, OCG, BCK
        country = code_upper[6:9]   # e.g., ARG, COL
        return (sub_code, country)

    return (None, None)


def filter_by_tech_country_matrix(og_data: Dict[str, pd.DataFrame], matrix_config: Dict[str, Any]) -> Dict[str, pd.DataFrame]:
    """
    Filter technologies based on the tech-country matrix configuration.

    Args:
        og_data: Dictionary of DataFrames
        matrix_config: Configuration loaded from Tech_Country_Matrix.xlsx

    Returns:
        Filtered dictionary of DataFrames
    """
    if not matrix_config.get("enabled", False):
        return og_data

    # Check the "Enable Matrix Filtering" flag from the Excel
    if not matrix_config.get("matrix_filtering_enabled", True):
        print("\n" + "=" * 70)
        print("TECH-COUNTRY MATRIX FILTERING: DISABLED")
        print("  (All technologies from data will be passed through)")
        print("=" * 70 + "\n")
        return og_data

    matrix = matrix_config.get("matrix", {})
    if not matrix:
        return og_data

    print("\n" + "=" * 70)
    print("TECH-COUNTRY MATRIX FILTERING")
    print("=" * 70)

    total_filtered = 0

    for param_name, df in og_data.items():
        if "TECHNOLOGY" not in df.columns:
            continue

        # Skip empty DataFrames to preserve column structure
        if df.empty:
            continue

        rows_before = len(df)

        def is_allowed(tech):
            sub_code, country = extract_tech_country_from_code(tech)
            if sub_code is None or country is None:
                return True  # Allow non-PWR technologies

            # Map CCG/OCG to NGS for lookup
            if sub_code in ["CCG", "OCG"]:
                sub_code = "NGS"

            # Check matrix
            if sub_code in matrix and country in matrix[sub_code]:
                return matrix[sub_code][country]

            return True  # Default: allow if not in matrix

        mask = df["TECHNOLOGY"].apply(is_allowed)
        og_data[param_name] = df[mask].copy()

        rows_filtered = rows_before - len(og_data[param_name])
        if rows_filtered > 0:
            total_filtered += rows_filtered
            print(f"    {param_name}: filtered {rows_filtered} rows")

    print(f"\n    Total rows filtered: {total_filtered}")
    print("\n" + "=" * 70)
    print("Tech-country matrix filtering completed.")
    print("=" * 70 + "\n")

    return og_data


def unify_ngs_technologies(og_data: Dict[str, pd.DataFrame], matrix_config: Dict[str, Any]) -> Dict[str, pd.DataFrame]:
    """
    Unify CCG (Combined Cycle) and OCG (Open Cycle) technologies into NGS (Natural Gas).

    This follows the same aggregation pattern as region consolidation:
    - avg parameters: values are averaged
    - sum parameters: values are summed
    - disabled parameters: skipped

    Args:
        og_data: Dictionary of DataFrames
        matrix_config: Configuration loaded from Tech_Country_Matrix.xlsx

    Returns:
        Modified dictionary with unified NGS technologies
    """
    if not matrix_config.get("ngs_enabled", False):
        print("[Info] NGS unification is disabled.")
        return og_data

    print("\n" + "=" * 70)
    print("NGS UNIFICATION (CCG + OCG -> NGS)")
    print("=" * 70)

    agg_rules = matrix_config.get("aggregation_rules", {})
    avg_params = set(agg_rules.get("avg", []))
    sum_params = set(agg_rules.get("sum", []))
    disabled_params = set(agg_rules.get("disabled", []))

    if disabled_params:
        print("\n[Info] Disabled parameters (no unification):")
        for param in sorted(disabled_params):
            if param in og_data:
                print(f"    - {param}")

    for param_name, df in og_data.items():
        if "TECHNOLOGY" not in df.columns:
            continue

        if param_name in disabled_params:
            continue

        # Determine aggregation method
        if param_name in avg_params:
            agg_method = "avg"
        elif param_name in sum_params:
            agg_method = "sum"
        else:
            continue  # Parameter not in any list

        # Find CCG and OCG technologies
        ccg_mask = df["TECHNOLOGY"].str.contains("PWRCCG", na=False)
        ocg_mask = df["TECHNOLOGY"].str.contains("PWROCG", na=False)

        if not (ccg_mask.any() or ocg_mask.any()):
            continue

        # Separate gas and non-gas rows
        gas_mask = ccg_mask | ocg_mask
        df_gas = df[gas_mask].copy()
        df_other = df[~gas_mask].copy()

        if df_gas.empty:
            continue

        # Create NGS technology codes (replace CCG/OCG with NGS)
        df_gas["_ngs_tech"] = df_gas["TECHNOLOGY"].str.replace("PWRCCG", "PWRNGS", regex=False)
        df_gas["_ngs_tech"] = df_gas["_ngs_tech"].str.replace("PWROCG", "PWRNGS", regex=False)

        # Also transform FUEL column if it contains CCG/OCG
        if "FUEL" in df_gas.columns:
            df_gas["_ngs_fuel"] = df_gas["FUEL"].str.replace("CCG", "NGS", regex=False)
            df_gas["_ngs_fuel"] = df_gas["_ngs_fuel"].str.replace("OCG", "NGS", regex=False)

        # Determine grouping columns
        group_cols = ["_ngs_tech"]
        for col in df_gas.columns:
            if col in ["_ngs_tech", "_ngs_fuel", "TECHNOLOGY", "VALUE", "FUEL"]:
                continue
            if col in ["YEAR", "TIMESLICE", "MODE_OF_OPERATION", "EMISSION", "REGION"]:
                group_cols.append(col)

        if "_ngs_fuel" in df_gas.columns:
            group_cols.append("_ngs_fuel")

        # Aggregate
        if "VALUE" in df_gas.columns:
            agg_func = "mean" if agg_method == "avg" else "sum"
            df_unified = df_gas.groupby(group_cols, as_index=False).agg({"VALUE": agg_func})
            df_unified["TECHNOLOGY"] = df_unified["_ngs_tech"]
            df_unified.drop(columns=["_ngs_tech"], inplace=True)

            if "_ngs_fuel" in df_unified.columns:
                df_unified["FUEL"] = df_unified["_ngs_fuel"]
                df_unified.drop(columns=["_ngs_fuel"], inplace=True)
        else:
            df_unified = df_gas.drop_duplicates(subset=["_ngs_tech"]).copy()
            df_unified["TECHNOLOGY"] = df_unified["_ngs_tech"]
            df_unified.drop(columns=["_ngs_tech"], inplace=True)
            if "_ngs_fuel" in df_unified.columns:
                df_unified["FUEL"] = df_unified["_ngs_fuel"]
                df_unified.drop(columns=["_ngs_fuel"], inplace=True)

        # Combine with non-gas rows
        result = pd.concat([df_other, df_unified], ignore_index=True)

        original_gas_count = len(df_gas)
        unified_count = len(df_unified)
        if original_gas_count > unified_count:
            print(f"    {param_name}: {original_gas_count} CCG/OCG rows -> {unified_count} NGS rows ({agg_method})")

        og_data[param_name] = result

    print("\n" + "=" * 70)
    print("NGS unification completed.")
    print("=" * 70 + "\n")

    return og_data


#--------------------------------------------------------------------------------------------------#

def update_sheet_year_headers(output_excel_path, sheet_name, model_years, fixed_cols_count=8):
    """
    Update only the year headers in a sheet without modifying data rows.
    Used for sheets that have no new data but need their year columns aligned to MODEL_YEARS.

    Args:
        output_excel_path: Path to the Excel file
        sheet_name: Name of the sheet to update
        model_years: List of years to use as headers (e.g., MODEL_YEARS)
        fixed_cols_count: Number of fixed (non-year) columns at the start
    """
    wb = load_workbook(output_excel_path)
    if sheet_name not in wb.sheetnames:
        print(f"[Warning] Sheet '{sheet_name}' not found in file. Skipping header update.")
        wb.close()
        return

    ws = wb[sheet_name]

    # Get current headers
    current_headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]

    # Find where year columns start (first numeric header after fixed columns)
    year_start_col = None
    for col_idx, header in enumerate(current_headers, 1):
        if header is not None:
            try:
                year = int(header)
                if 2000 <= year <= 2100:
                    year_start_col = col_idx
                    break
            except (ValueError, TypeError):
                continue

    if year_start_col is None:
        # No year columns found, use fixed_cols_count + 1
        year_start_col = fixed_cols_count + 1

    # Get old year columns for data migration
    old_year_cols = {}
    for col_idx, header in enumerate(current_headers, 1):
        if header is not None:
            try:
                year = int(header)
                if 2000 <= year <= 2100:
                    old_year_cols[year] = col_idx
            except (ValueError, TypeError):
                continue

    # Create mapping from old column to new column for years we're keeping
    new_year_cols = {year: year_start_col + idx for idx, year in enumerate(model_years)}

    # Read all data rows first (to avoid issues with column shifting)
    data_rows = []
    for row_idx in range(2, ws.max_row + 1):
        row_data = {}
        # Fixed columns
        for col_idx in range(1, year_start_col):
            row_data[col_idx] = ws.cell(row_idx, col_idx).value
        # Year data (by year, not column)
        for year, col_idx in old_year_cols.items():
            if year in model_years:  # Only keep years in MODEL_YEARS
                row_data[('year', year)] = ws.cell(row_idx, col_idx).value
        data_rows.append(row_data)

    # Calculate new total columns needed
    new_max_col = year_start_col + len(model_years) - 1

    # Clear existing year headers and data beyond fixed columns
    for col_idx in range(year_start_col, ws.max_column + 1):
        ws.cell(1, col_idx).value = None
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row_idx, col_idx).value = None

    # Write new year headers
    for idx, year in enumerate(model_years):
        ws.cell(1, year_start_col + idx, year)

    # Write back data for years that exist in MODEL_YEARS
    for row_idx, row_data in enumerate(data_rows, 2):
        for year in model_years:
            if ('year', year) in row_data:
                new_col = new_year_cols[year]
                ws.cell(row_idx, new_col, row_data[('year', year)])

    wb.save(output_excel_path)
    print(f"[Success] Sheet '{sheet_name}' year headers updated to {model_years[0]}-{model_years[-1]}.")

def write_sheet(sheet_name, records, all_years, output_excel_path):
    if not records:
        # Even without data, update the year headers
        print(f"[Info] No data to write to sheet '{sheet_name}'. Updating year headers only.")
        update_sheet_year_headers(output_excel_path, sheet_name, all_years, fixed_cols_count=8)
        return

    df_out = pd.DataFrame(records)

    for y in all_years:
        if y not in df_out.columns:
            df_out[y] = np.nan

    df_out = df_out.sort_values(by=["Tech.ID", "Parameter.ID"])
    df_out = df_out[[
        "Tech.ID", "Tech", "Tech.Name", "Parameter.ID", "Parameter", "Unit",
        "Projection.Mode", "Projection.Parameter"
    ] + all_years]

    wb = load_workbook(output_excel_path)
    ws = wb[sheet_name]
    ws.delete_rows(1, ws.max_row)

    for r in dataframe_to_rows(df_out, index=False, header=True):
        ws.append(r)

    wb.save(output_excel_path)
    print(f"[Success] Sheet '{sheet_name}' in Parametrization file updated.")

def parse_tech_name(tech):
    """
    Returns a descriptive name for a technology code with structure interpretation.
    Handles special formats for TRN, SDS, LDS technologies and adds investability notes.
    """
    main_code = tech[0:3]

    # Handle transmission interconnection codes
    if main_code == "TRN" and len(tech) >= 13:
        iso1 = tech[3:6]
        region1 = tech[6:8]
        iso2 = tech[8:11]
        region2 = tech[11:13]
        country1 = iso_country_map.get(iso1, f"Unknown ({iso1})")
        country2 = iso_country_map.get(iso2, f"Unknown ({iso2})")
        return f"Transmission interconnection from {country1}, region {region1} to {country2}, region {region2}"
    
    # Handle storage codes (SDS, LDS)
    if (main_code == "SDS" or main_code == "LDS") and len(tech) <= 10:
        iso1 = tech[3:6]
        region1 = tech[6:8]
        storage_code = code_to_energy.get(main_code, "specific technology")
        country1 = iso_country_map.get(iso1, f"Unknown ({iso1})")
        return f"{storage_code} {country1}, region {region1}"

    iso = tech[6:9]
    region = tech[9:11]
    country = iso_country_map.get(iso, f"Unknown ({iso})")
    sub_code = tech[3:6]
    main_desc = code_to_energy.get(main_code, "General technology")
    sub_desc = code_to_energy.get(sub_code, "specific technology")

    # Use consistent naming base
    base = f"{sub_desc} ({main_desc})" if main_desc != sub_desc else sub_desc
    name = f"{base} {country}"

    # Add region information (omit for MIN)
    if not tech.startswith("MIN") and region != "XX":
        name += f", region {region}"
    elif region == "XX":
        name += f", region XX"

    return name


def parse_fuel_name(fuel):
    """
    Generates a readable name for a fuel code based on its structure.
    Structure format:
    - First 3 characters: fuel type (e.g., OIL, HYD, PET)
    - Characters 3-5: ISO-3 country code
    - Characters 6-7 (if present): region
    - Ending in '01' or '02' is interpreted as output type
    """
    prefix = fuel[0:3]
    iso = fuel[3:6]
    region = fuel[6:8] if len(fuel) >= 8 else None
    suffix = None

    if fuel.endswith("01"):
        suffix = "power plant output"
    elif fuel.endswith("02"):
        suffix = "transmission line output"

    fuel_type = code_to_energy.get(prefix, "Unknown")
    country = iso_country_map.get(iso, f"Unknown ({iso})")

    name_parts = [fuel_type, country]
    if region and region != "XX":
        name_parts.append(f"region {region}")
    elif region == "XX":
        name_parts.append("region XX")
    if suffix:
        name_parts.append(suffix)

    return ", ".join(name_parts)

def assign_tech_type(tech):
    if tech.startswith("MIN") or tech.startswith("RNW"):
        return "Primary"
    elif tech.startswith("PWRTRN"):
        return "Demand"
    else:
        return "Secondary"

#--------------------------------------------------------------------------------------------------#

#-------------------------------------Updated intermediate functions-------------------------------#
def update_demand_profiles(df, output_excel_path, input_excel_path):
    """Updates the Profiles sheet in the given Excel file using the specified DataFrame."""
    # Identify unique years - use MODEL_YEARS as fallback if no data
    unique_years = sorted(df["YEAR"].unique()) if not df.empty else MODEL_YEARS
    year_cols = [str(y) for y in unique_years]

    records = []
    for (timeslice, fuel), group in df.groupby(["TIMESLICE", "FUEL"]):
        record = {
            "Timeslices": timeslice,
            "Demand/Share": "Demand",
            "Fuel/Tech": fuel,
            "Ref.Cap.BY": "not needed",
            "Ref.OAR.BY": "not needed",
            "Ref.km.BY": "not needed",
            "Projection.Mode": "User defined",
            "Projection.Parameter": 0
        }

        # Generate descriptive Name field
        iso = fuel[3:6]
        region = fuel[6:8]
        demand = fuel[8:10]
        country = iso_country_map.get(iso, f"Unknown country ({iso})")

        if demand == "01":
            name = f"Output demand of power plants in {country}"
        elif demand == "02":
            name = f"Output demand of transmission lines in {country}"
        else:
            name = f"Unknown demand type for {fuel} in {country}"

        if region != "XX":
            name += f", in region {region}."

        record["Name"] = name

        for _, row in group.iterrows():
            record[str(row["YEAR"])] = row["VALUE"]

        records.append(record)

    # Create DataFrame
    df_timeslices = pd.DataFrame(records)
    fixed_cols = [
        "Timeslices", "Demand/Share", "Fuel/Tech", "Name",
        "Ref.Cap.BY", "Ref.OAR.BY", "Ref.km.BY", "Projection.Mode", "Projection.Parameter"
    ]
    df_timeslices = df_timeslices[fixed_cols + year_cols]

    # Update the Excel sheet
    wb = load_workbook(input_excel_path)
    ws = wb["Profiles"]
    ws.delete_rows(1, ws.max_row)

    for r in dataframe_to_rows(df_timeslices, index=False, header=True):
        ws.append(r)

    os.makedirs(os.path.dirname(output_excel_path), exist_ok=True)
    wb.save(output_excel_path)
    print("[Success] Sheet 'Profiles' in Demand file updated.")

def update_demand_demand_projection(df, output_excel_path, input_excel_path):
    """Updates the Demand_Projection sheet in the given Excel file using the specified DataFrame.

    Also adds missing countries from OLADE generation data (e.g., HTI) that may not be
    present in the OSeMOSYS input CSVs.
    """
    # Identify unique years - use MODEL_YEARS as fallback if no data
    unique_years = sorted(df["YEAR"].unique()) if not df.empty else MODEL_YEARS
    year_cols = [str(y) for y in unique_years]

    records = []
    existing_countries = set()  # Track which countries we have in the CSV data

    for fuel, group in df.groupby("FUEL"):
        record = {
            "Demand/Share": "Demand",
            "Fuel/Tech": fuel,
            "Ref.Cap.BY": "not needed",
            "Ref.OAR.BY": "not needed",
            "Ref.km.BY": "not needed",
            "Projection.Mode": "User defined",
            "Projection.Parameter": 0
        }

        # Generate descriptive Name field
        iso = fuel[3:6]
        region = fuel[6:8]
        demand = fuel[8:10]
        country = iso_country_map.get(iso, f"Unknown country ({iso})")

        # Track country codes for ELC*XX02 entries (transmission demand)
        if fuel.startswith("ELC") and region == "XX" and demand == "02":
            existing_countries.add(iso)

        if demand == "01":
            name = f"Output demand of power plants in {country}"
        elif demand == "02":
            name = f"Output demand of transmission lines in {country}"
        else:
            name = f"Unknown demand type for {fuel} in {country}"

        if region != "XX":
            name += f", in region {region}."

        record["Name"] = name

        for _, row in group.iterrows():
            record[str(row["YEAR"])] = row["VALUE"]

        records.append(record)

    # Add missing countries from OLADE data (controlled by config flag)
    if not get_add_missing_countries_from_olade():
        olade_data = None
    else:
        olade_data = read_olade_generation_data()
    if olade_data:
        olade_countries = set(olade_data['data'].keys())
        # Only add countries that are in iso_country_map (model countries)
        model_countries = set(iso_country_map.keys()) - {'INT'}  # Exclude international markets
        missing_countries = (olade_countries & model_countries) - existing_countries

        if missing_countries:
            print(f"[Info] Adding missing countries from OLADE data: {', '.join(sorted(missing_countries))}")

            ref_year = olade_data['reference_year']
            default_growth_rate = 0.02  # 2% annual growth rate for demand projection

            for country_code in sorted(missing_countries):
                base_demand_pj = olade_data['data'][country_code]
                fuel_code = f"ELC{country_code}XX02"
                country_name = iso_country_map.get(country_code, f"Unknown country ({country_code})")

                record = {
                    "Demand/Share": "Demand",
                    "Fuel/Tech": fuel_code,
                    "Name": f"Output demand of transmission lines in {country_name}",
                    "Ref.Cap.BY": "not needed",
                    "Ref.OAR.BY": "not needed",
                    "Ref.km.BY": "not needed",
                    "Projection.Mode": "User defined",
                    "Projection.Parameter": 0
                }

                # Calculate demand for each year using linear growth from reference year
                for year_str in year_cols:
                    year = int(year_str)
                    years_diff = year - ref_year
                    demand_value = base_demand_pj * (1 + default_growth_rate * years_diff)
                    record[year_str] = max(0, demand_value)  # Ensure non-negative

                records.append(record)
                print(f"  Added {fuel_code}: base demand {base_demand_pj:.2f} PJ from OLADE ({ref_year})")

    # Create DataFrame
    df_demand_projection  = pd.DataFrame(records)
    fixed_cols = [
        "Demand/Share", "Fuel/Tech", "Name",
        "Ref.Cap.BY", "Ref.OAR.BY", "Ref.km.BY", "Projection.Mode", "Projection.Parameter"
    ]
    df_demand_projection = df_demand_projection[fixed_cols + year_cols]

    # Update the Excel sheet
    wb = load_workbook(output_excel_path)
    ws = wb["Demand_Projection"]
    ws.delete_rows(1, ws.max_row)

    for r in dataframe_to_rows(df_demand_projection, index=False, header=True):
        ws.append(r)

    wb.save(output_excel_path)
    print("[Success] Sheet 'Demand_Projection' in Demand file updated.")

def update_parametrization_capacities(df, output_excel_path):
    """Updates Capacities sheet in A-O_Parametrization.xlsx using CapacityFactor data."""
    # Use MODEL_YEARS as fallback if no data
    unique_years = sorted(df["YEAR"].unique()) if not df.empty else MODEL_YEARS
    year_cols = [str(y) for y in unique_years]

    tech_id_map = {}
    tech_id_counter = 1
    records = []

    for (timeslice, tech), group in df.groupby(["TIMESLICE", "TECHNOLOGY"]):
        if tech not in tech_id_map:
            tech_id_map[tech] = tech_id_counter
            tech_id_counter += 1

        record = {
            "Timeslices": timeslice,
            "Tech.ID": tech_id_map[tech],
            "Tech": tech,
            "Parameter.ID": 13,
            "Parameter": "CapacityFactor",
            "Unit": None,
            "Projection.Mode": "User defined",
            "Projection.Parameter": 0
        }

        record["Tech.Name"] = parse_tech_name(tech)

        for _, row in group.iterrows():
            record[str(row["YEAR"])] = row["VALUE"]

        records.append(record)

    df_cap = pd.DataFrame(records)
    fixed_cols = [
        "Timeslices", "Tech.ID", "Tech", "Tech.Name", "Parameter.ID",
        "Parameter", "Unit", "Projection.Mode", "Projection.Parameter"
    ]
    df_cap = df_cap[fixed_cols + year_cols]
    # Apply sorting: first by Tech alphabetically, then by Parameter.ID
    df_cap = df_cap.sort_values(by=["Tech.ID", "Timeslices"], ascending=[True, True])


    wb = load_workbook(output_excel_path)
    ws = wb["Capacities"]
    ws.delete_rows(1, ws.max_row)

    for row in dataframe_to_rows(df_cap, index=False, header=True):
        ws.append(row)

    wb.save(output_excel_path)
    print("[Success] Sheet 'Capacities' in Parametrization file updated.")

def update_parametrization_yearsplit(df, output_excel_path):
    """Updates Yearsplit sheet in A-O_Parametrization.xlsx using YearSplit data."""
    # Use MODEL_YEARS as fallback if no data
    unique_years = sorted(df["YEAR"].unique()) if not df.empty else MODEL_YEARS
    year_cols = [str(y) for y in unique_years]

    records = []

    for timeslice, group in df.groupby("TIMESLICE"):

        record = {
            "Timeslices": timeslice,
            "Parameter.ID": 14,
            "Parameter": "YearSplit",
            "Unit": None,
            "Projection.Mode": "User defined",
            "Projection.Parameter": 0
        }

        for _, row in group.iterrows():
            record[str(row["YEAR"])] = row["VALUE"]

        records.append(record)

    df_cap = pd.DataFrame(records)
    fixed_cols = [
        "Timeslices", "Parameter.ID",
        "Parameter", "Unit", "Projection.Mode", "Projection.Parameter"
    ]
    df_cap = df_cap[fixed_cols + year_cols]

    wb = load_workbook(output_excel_path)
    ws = wb["Yearsplit"]
    ws.delete_rows(1, ws.max_row)

    for row in dataframe_to_rows(df_cap, index=False, header=True):
        ws.append(row)

    wb.save(output_excel_path)
    print("[Success] Sheet 'Yearsplit' in Parametrization file updated.")

def update_parametrization_daysplit(df, output_excel_path):
    """Updates DaySplit sheet in A-O_Parametrization.xlsx using DaySplit data."""
    # Use MODEL_YEARS as fallback if no data
    unique_years = sorted(df["YEAR"].unique()) if not df.empty else MODEL_YEARS
    year_cols = [str(y) for y in unique_years]
    
    records = []
    
    for dailytimebracket, group in df.groupby("DAILYTIMEBRACKET"):

        record = {
            "DAILYTIMEBRACKET": dailytimebracket,
            "Parameter.ID": 12,
            "Parameter": "DaySplit",
            "Unit": None,
            "Projection.Mode": "User defined",
            "Projection.Parameter": 0
        }

        for _, row in group.iterrows():
            record[str(int(row["YEAR"]))] = row["VALUE"]

        records.append(record)
        
    df_cap = pd.DataFrame(records)
    fixed_cols = [
        "DAILYTIMEBRACKET", "Parameter.ID",
        "Parameter", "Unit", "Projection.Mode", "Projection.Parameter"
    ]

    df_cap = df_cap[fixed_cols + year_cols]

    wb = load_workbook(output_excel_path)
    ws = wb["DaySplit"]
    ws.delete_rows(1, ws.max_row)

    for row in dataframe_to_rows(df_cap, index=False, header=True):
        ws.append(row)

    wb.save(output_excel_path)
    print("[Success] Sheet 'DaySplit' in Parametrization file updated.")

def update_parametrization_fixed_horizon_parameters(df_ctau, df_oplife, output_excel_path, input_excel_path):
    """
    Updates the 'Fixed Horizon Parameters' sheet using CapacityToActivityUnit and OperationalLife data.
    Applies parameter values, fills missing with default = 1, assigns Tech.Type based on naming rules.
    Sorts results by Tech and Parameter.ID before writing to Excel.
    """

    PARAMETERS = [
        ("CapacityToActivityUnit", 1, df_ctau),
        ("OperationalLife", 2, df_oplife)
    ]

    all_techs = set()
    param_data = {}

    for param_name, param_id, df in PARAMETERS:
        param_data[param_name] = {}
        for _, row in df.iterrows():
            tech = row["TECHNOLOGY"]
            value = row["VALUE"]
            param_data[param_name][tech] = value
            all_techs.add(tech)

    tech_ids = {tech: idx + 1 for idx, tech in enumerate(sorted(all_techs))}

    output_rows = []
    for tech in all_techs:
        for param_name, param_id, _ in PARAMETERS:
            value = param_data[param_name].get(tech, 1)
            output_rows.append({
                "Tech.Type": assign_tech_type(tech),
                "Tech.ID": tech_ids[tech],
                "Tech": tech,
                "Tech.Name": parse_tech_name(tech),
                "Parameter.ID": param_id,
                "Parameter": param_name,
                "Unit": None,
                "Value": value
            })

    df_fixed = pd.DataFrame(output_rows)
    df_fixed = df_fixed.sort_values(by=["Tech", "Parameter.ID"])

    wb = load_workbook(input_excel_path)
    ws = wb["Fixed Horizon Parameters"]
    ws.delete_rows(1, ws.max_row)

    for r in dataframe_to_rows(df_fixed, index=False, header=True):
        ws.append(r)

    os.makedirs(os.path.dirname(output_excel_path), exist_ok=True)
    wb.save(output_excel_path)
    print("[Success] Sheet 'Fixed Horizon Parameters' in Parametrization updated.")

def update_parametrization_primary_secondary_demand_techs(og_data, output_excel_path):
    """
    Updates Primary, Secondary, and Demand Tech sheets using parameter data.
    Tech type is determined by prefix. Naming logic is conditional:
    - parse_tech_name for MIN, RNW, PWR, TRN
    - parse_fuel_name otherwise
    """
    PARAMETERS = [
        "CapitalCost", "FixedCost", "ResidualCapacity",
        "TotalAnnualMaxCapacity", "TotalTechnologyAnnualActivityUpperLimit",
        "TotalTechnologyAnnualActivityLowerLimit", "TotalAnnualMinCapacityInvestment",
        "AvailabilityFactor", "ReserveMarginTagFuel",
        "ReserveMarginTagTechnology", "TotalAnnualMaxCapacityInvestment"
    ]

    PARAMETER_IDS = {name: idx + 1 for idx, name in enumerate(PARAMETERS)}
    primary_records, secondary_records, demand_records = [], [], []
    tech_ids = {}
    tech_counter = 1
    all_years = set()
    techs_by_param = {}

    for param in PARAMETERS:
        if param not in og_data:
            techs_by_param[param] = set()  # Empty set for missing parameters
            continue
        df = og_data[param]
        key_col = "FUEL" if param == "ReserveMarginTagFuel" else "TECHNOLOGY"
        # Handle empty DataFrames - still track the parameter but with empty tech set
        if df.empty or key_col not in df.columns:
            techs_by_param[param] = set()
            continue
        techs_by_param[param] = set(df[key_col].unique())
        if param != "ReserveMarginTagFuel" and "YEAR" in df.columns:
            all_years.update(df["YEAR"].unique())

    all_techs = set().union(*techs_by_param.values())

    for tech in all_techs:
        is_demand_tech = tech.startswith("PWRTRN")
        main_prefix = tech[0:3]

        if tech not in tech_ids:
            tech_ids[tech] = tech_counter
            tech_counter += 1

        # Select naming function based on tech prefix
        if main_prefix in ["MIN", "RNW", "PWR", "TRN"]:
            tech_name = parse_tech_name(tech)
        else:
            tech_name = parse_fuel_name(tech)

        for param in PARAMETERS:
            # Determine target list based on tech type and parameter
            if is_demand_tech and param in ["CapitalCost", "FixedCost", "ResidualCapacity"]:
                target = demand_records
            elif is_demand_tech:
                continue
            elif main_prefix in ["MIN", "RNW"]:
                target = primary_records
            else:
                target = secondary_records

            record = {
                "Tech.ID": tech_ids[tech],
                "Tech": tech,
                "Tech.Name": tech_name,
                "Parameter.ID": PARAMETER_IDS[param],
                "Parameter": param,
                "Unit": None,
                "Projection.Parameter": 0
            }

            # Check if parameter has data for this tech
            group = pd.DataFrame()  # Empty by default
            if param in og_data:
                df = og_data[param]
                key_col = "FUEL" if param == "ReserveMarginTagFuel" else "TECHNOLOGY"
                if not df.empty and key_col in df.columns:
                    group = df[df[key_col] == tech]

            if group.empty:
                record["Projection.Mode"] = "EMPTY"
                for y in all_years:
                    record[int(y)] = float("nan")
            else:
                available_years = sorted(group["YEAR"].unique())
                year_values = {int(row["YEAR"]): row["VALUE"] for _, row in group.iterrows()}
                values = [year_values.get(y, float("nan")) for y in available_years]

                non_nan_count = sum(pd.notna(values))
                if non_nan_count == 0:
                    mode = "EMPTY"
                elif non_nan_count == 1 and not pd.isna(values[0]):
                    mode = "Flat"
                elif non_nan_count == len(values):
                    mode = "User defined"
                else:
                    mode = "interpolation"

                record["Projection.Mode"] = mode
                for y in available_years:
                    record[int(y)] = year_values.get(y, float("nan"))

            target.append(record)

    # Use MODEL_YEARS as fallback if no data provided years
    all_years = sorted(all_years) if all_years else MODEL_YEARS
    write_sheet("Primary Techs", primary_records, all_years, output_excel_path)
    write_sheet("Secondary Techs", secondary_records, all_years, output_excel_path)
    write_sheet("Demand Techs", demand_records, all_years, output_excel_path)

def update_parametrization_variable_cost(og_data, output_excel_path):
    """
    Updates the 'VariableCost' sheet in the Parametrization file.
    Includes an additional 'Mode.Operation' column from the MODE_OF_OPERATION column in the data.
    Applies the same projection logic and naming rules as other tech sheets.
    """

    param = "VariableCost"
    if param not in og_data:
        print(f"[Warning] {param} not found in OG_Input_Data.")
        return

    df = og_data[param]
    tech_ids = {}
    tech_counter = 1
    # Use MODEL_YEARS as fallback if no data
    all_years = sorted(df["YEAR"].unique()) if not df.empty else MODEL_YEARS
    records = []

    for (tech, mode), group in df.groupby(["TECHNOLOGY", "MODE_OF_OPERATION"]):
        if tech not in tech_ids:
            tech_ids[tech] = tech_counter
            tech_counter += 1

        prefix = tech[0:3]
        if prefix in ["MIN", "RNW", "PWR", "TRN"]:
            tech_name = parse_tech_name(tech)
        else:
            tech_name = parse_fuel_name(tech)

        record = {
            "Mode.Operation": int(mode),
            "Tech.ID": tech_ids[tech],
            "Tech": tech,
            "Tech.Name": tech_name,
            "Parameter.ID": 12,
            "Parameter": "VariableCost",
            "Unit": None,
            "Projection.Parameter": 0
        }

        year_values = {int(row["YEAR"]): row["VALUE"] for _, row in group.iterrows()}
        values = [year_values.get(y, float("nan")) for y in all_years]
        non_nan_count = sum(pd.notna(values))

        if non_nan_count == 0:
            mode_str = "EMPTY"
        elif non_nan_count == 1 and not pd.isna(values[0]):
            mode_str = "Flat"
        elif non_nan_count == len(values):
            mode_str = "User defined"
        else:
            mode_str = "interpolation"

        record["Projection.Mode"] = mode_str

        for y in all_years:
            record[int(y)] = year_values.get(y, float("nan"))

        records.append(record)

    # Write to Excel
    df_out = pd.DataFrame(records)
    df_out = df_out.sort_values(by=["Tech", "Mode.Operation"])

    wb = load_workbook(output_excel_path)
    ws = wb["VariableCost"]
    ws.delete_rows(1, ws.max_row)

    for r in dataframe_to_rows(df_out, index=False, header=True):
        ws.append(r)

    wb.save(output_excel_path)
    print("[Success] Sheet 'VariableCost' in Parametrization file updated.")


def update_xtra_emissions_ghg(og_data, workbook):
    """
    Updates the 'GHGs' sheet in A-Xtra_Emissions.xlsx using EmissionActivityRatio data.
    Keeps only one row per (TECHNOLOGY, EMISSION, MODE_OF_OPERATION), taking the first available value.
    Includes the Mode_Of_Operation column in the output.
    """
    if "EmissionActivityRatio" not in og_data:
        print("[Warning] 'EmissionActivityRatio' not found in OG_Input_Data.")
        return

    df = og_data["EmissionActivityRatio"]

    # Group by (TECHNOLOGY, EMISSION, MODE_OF_OPERATION) and take the first row
    grouped = (
        df.groupby(["TECHNOLOGY", "EMISSION", "MODE_OF_OPERATION"], as_index=False)
        .first()[["TECHNOLOGY", "EMISSION", "MODE_OF_OPERATION", "VALUE"]]
        .rename(columns={
            "TECHNOLOGY": "Tech",
            "EMISSION": "Emission",
            "MODE_OF_OPERATION": "Mode_Of_Operation",
            "VALUE": "EmissionActivityRatio"
        })
    )

    grouped["Unit"] = "MT"
    grouped = grouped[["Mode_Of_Operation", "Tech", "Emission", "EmissionActivityRatio", "Unit"]]

    ws = workbook["GHGs"]
    ws.delete_rows(1, ws.max_row)

    for row in dataframe_to_rows(grouped, index=False, header=True):
        ws.append(row)

    print("[Success] Sheet 'GHGs' in Extra Emissions file updated.")

def update_xtra_emissions_externalities(og_data, workbook):
    """
    Updates the 'Externalities' sheet in A-Xtra_Emissions.xlsx using EmissionsPenalty data.
    Fills in Emission, EmissionsPenalty, and Final Unit columns. Other columns remain blank or NaN.
    """

    if "EmissionsPenalty" not in og_data:
        print("[Warning] 'EmissionsPenalty' not found in OG_Input_Data.")
        return

    df = og_data["EmissionsPenalty"]

    # Drop duplicate EMISSION values (value is constant per emission)
    grouped = (
        df.groupby("EMISSION", as_index=False)
        .first()[["EMISSION", "VALUE"]]
        .rename(columns={
            "EMISSION": "Emission",
            "VALUE": "EmissionsPenalty"
        })
    )

    grouped["Tech"] = None
    grouped["External Cost"] = None
    grouped["Mode_Of_Operation"] = None
    grouped["EmissionActivityRatio"] = None
    grouped["Final Unit"] = None

    # Reorder to match Excel columns
    grouped = grouped[[
        "Tech", "Emission", "External Cost", "Mode_Of_Operation",
        "EmissionActivityRatio", "EmissionsPenalty", "Final Unit"
    ]]

    # Write to workbook
    ws = workbook["Externalities"]
    ws.delete_rows(1, ws.max_row)

    for row in dataframe_to_rows(grouped, index=False, header=True):
        ws.append(row)

    print("[Success] Sheet 'Externalities' in Extra Emissions file updated.")


def update_model_base_year_primary(og_data, workbook):
    """
    Updates the 'Primary' sheet in the base year model Excel workbook
    using the 'OutputActivityRatio' parameter data from OG_Input_Data.
    Only technologies starting with 'MIN' or 'RNW' are included.
    """
    if "OutputActivityRatio" not in og_data:
        print("[Warning] 'OutputActivityRatio' not found in OG_Input_Data.")
        return

    df = og_data["OutputActivityRatio"]
    df_filtered = df[df["TECHNOLOGY"].str.startswith(("MIN", "RNW"))]

    # Group by unique combinations to extract representative row
    grouped = df_filtered.groupby(["TECHNOLOGY", "FUEL", "MODE_OF_OPERATION"], as_index=False).first()

    records = []
    for _, row in grouped.iterrows():
        tech = row["TECHNOLOGY"]
        fuel = row["FUEL"]
        mode = int(row["MODE_OF_OPERATION"])

        record = {
            "Mode.Operation": mode,
            "Tech": tech,
            "Tech.Name": parse_tech_name(tech),
            "Fuel.O": fuel,
            "Fuel.O.Name": parse_fuel_name(fuel),
            "Value.Fuel.O": 1,  # Always fixed to int(1)
            "Unit.Fuel.O": None
        }
        records.append(record)

    df_final = pd.DataFrame(records)

    # Clear and write to 'Primary' sheet
    ws = workbook["Primary"]
    ws.delete_rows(1, ws.max_row)

    for r in dataframe_to_rows(df_final, index=False, header=True):
        ws.append(r)

    print("[Success] Sheet 'Primary' in Base Year Model file updated.")



def update_model_base_year_secondary(og_data, workbook):
    """
    Updates the 'Secondary' sheet in the base year model Excel workbook
    using both 'InputActivityRatio' and 'OutputActivityRatio' data.
    Includes 'PWRBCK' technologies with missing inputs by filling input fields with None.
    Excludes technologies starting with 'MIN' or 'RNW', and fuels ending in '02'.
    Each row combines matching input/output records for the same technology and mode of operation.
    """

    if "InputActivityRatio" not in og_data or "OutputActivityRatio" not in og_data:
        print("[Warning] Missing one or both parameters: 'InputActivityRatio', 'OutputActivityRatio'.")
        return

    df_input = og_data["InputActivityRatio"]
    df_output = og_data["OutputActivityRatio"]

    # Filter inputs and outputs by prefix and suffix rules
    df_input = df_input[
        (~df_input["TECHNOLOGY"].str.startswith(("MIN", "RNW"), na=False))
    ]
    df_output = df_output[
        (~df_output["TECHNOLOGY"].str.startswith(("MIN", "RNW"), na=False)) &
        (~df_output["FUEL"].str.endswith("02", na=False))
    ]

    # Group and merge
    input_grouped = df_input.groupby(["TECHNOLOGY", "MODE_OF_OPERATION"], as_index=False).first()
    output_grouped = df_output.groupby(["TECHNOLOGY", "MODE_OF_OPERATION"], as_index=False).first()

    merged = pd.merge(
        input_grouped, output_grouped,
        on=["TECHNOLOGY", "MODE_OF_OPERATION"],
        suffixes=("_I", "_O")
    )

    # Handle 'PWRBCK' techs missing input data
    techs_output_only = df_output[
        df_output["TECHNOLOGY"].str.startswith("PWRBCK") &
        ~df_output["TECHNOLOGY"].isin(input_grouped["TECHNOLOGY"])
    ].copy()

    if not techs_output_only.empty:
        techs_output_only = techs_output_only.groupby(["TECHNOLOGY", "MODE_OF_OPERATION"], as_index=False).first()
        techs_output_only["VALUE_O"] = techs_output_only["VALUE"]
        techs_output_only["VALUE_I"] = None
        techs_output_only["FUEL_I"] = None

        merged_extra = techs_output_only.rename(columns={
            "TECHNOLOGY": "TECHNOLOGY",
            "MODE_OF_OPERATION": "MODE_OF_OPERATION",
            "FUEL": "FUEL_O"
        })[["TECHNOLOGY", "MODE_OF_OPERATION", "FUEL_I", "VALUE_I", "FUEL_O", "VALUE_O"]]

        # Add missing columns explicitly to avoid future warning
        required_columns = ["TECHNOLOGY", "MODE_OF_OPERATION", "FUEL_I", "VALUE_I", "FUEL_O", "VALUE_O"]
        for col in required_columns:
            if col not in merged_extra.columns:
                merged_extra[col] = pd.NA

        # Set proper data types to avoid future warnings
        merged_extra = merged_extra.astype({
            "TECHNOLOGY": "string",
            "MODE_OF_OPERATION": "Int64",
            "FUEL_I": "string",
            "VALUE_I": "float",
            "FUEL_O": "string",
            "VALUE_O": "float"
        }, errors="ignore")

        # Reorder columns
        merged_extra = merged_extra[required_columns]

        # Concatenate merged_extra with merged
        merged = pd.concat([merged, merged_extra], ignore_index=True)

    # Handle 'PWRLDS' techs missing input data
    techs_output_only = df_output[
        df_output["TECHNOLOGY"].str.startswith("PWRLDS") 
    ].copy()

    if not techs_output_only.empty:
        techs_output_only = techs_output_only.groupby(["TECHNOLOGY", "MODE_OF_OPERATION"], as_index=False).first()
        techs_output_only["VALUE_O"] = techs_output_only["VALUE"]
        techs_output_only["VALUE_I"] = None
        techs_output_only["FUEL_I"] = None

        merged_extra = techs_output_only.rename(columns={
            "TECHNOLOGY": "TECHNOLOGY",
            "MODE_OF_OPERATION": "MODE_OF_OPERATION",
            "FUEL": "FUEL_O"
        })[["TECHNOLOGY", "MODE_OF_OPERATION", "FUEL_I", "VALUE_I", "FUEL_O", "VALUE_O"]]

        # Add missing columns explicitly to avoid future warning
        required_columns = ["TECHNOLOGY", "MODE_OF_OPERATION", "FUEL_I", "VALUE_I", "FUEL_O", "VALUE_O"]
        for col in required_columns:
            if col not in merged_extra.columns:
                merged_extra[col] = pd.NA

        # Set proper data types to avoid future warnings
        merged_extra = merged_extra.astype({
            "TECHNOLOGY": "string",
            "MODE_OF_OPERATION": "Int64",
            "FUEL_I": "string",
            "VALUE_I": "float",
            "FUEL_O": "string",
            "VALUE_O": "float"
        }, errors="ignore")

        # Reorder columns
        merged_extra = merged_extra[required_columns]

        # Concatenate merged_extra with merged
        merged = pd.concat([merged, merged_extra], ignore_index=True)

    # Handle 'PWRSDS' techs missing input data
    techs_output_only = df_output[
        df_output["TECHNOLOGY"].str.startswith("PWRSDS")
    ].copy()

    if not techs_output_only.empty:
        techs_output_only = techs_output_only.groupby(["TECHNOLOGY", "MODE_OF_OPERATION"], as_index=False).first()
        techs_output_only["VALUE_O"] = techs_output_only["VALUE"]
        techs_output_only["VALUE_I"] = None
        techs_output_only["FUEL_I"] = None

        merged_extra = techs_output_only.rename(columns={
            "TECHNOLOGY": "TECHNOLOGY",
            "MODE_OF_OPERATION": "MODE_OF_OPERATION",
            "FUEL": "FUEL_O"
        })[["TECHNOLOGY", "MODE_OF_OPERATION", "FUEL_I", "VALUE_I", "FUEL_O", "VALUE_O"]]

        # Add missing columns explicitly to avoid future warning
        required_columns = ["TECHNOLOGY", "MODE_OF_OPERATION", "FUEL_I", "VALUE_I", "FUEL_O", "VALUE_O"]
        for col in required_columns:
            if col not in merged_extra.columns:
                merged_extra[col] = pd.NA

        # Set proper data types to avoid future warnings
        merged_extra = merged_extra.astype({
            "TECHNOLOGY": "string",
            "MODE_OF_OPERATION": "Int64",
            "FUEL_I": "string",
            "VALUE_I": "float",
            "FUEL_O": "string",
            "VALUE_O": "float"
        }, errors="ignore")

        # Reorder columns
        merged_extra = merged_extra[required_columns]

        # Concatenate merged_extra with merged
        merged = pd.concat([merged, merged_extra], ignore_index=True)

    # Handle 'PWRLDS' techs missing input data
    techs_output_only = df_input[
        df_input["TECHNOLOGY"].str.startswith("PWRLDS")
    ].copy()

    if not techs_output_only.empty:
        techs_output_only = techs_output_only.groupby(["TECHNOLOGY", "MODE_OF_OPERATION"], as_index=False).first()
        techs_output_only["VALUE_O"] = None
        techs_output_only["VALUE_I"] = techs_output_only["VALUE"]
        techs_output_only["FUEL_O"] = None

        merged_extra = techs_output_only.rename(columns={
            "TECHNOLOGY": "TECHNOLOGY",
            "MODE_OF_OPERATION": "MODE_OF_OPERATION",
            "FUEL": "FUEL_I"
        })[["TECHNOLOGY", "MODE_OF_OPERATION", "FUEL_I", "VALUE_I", "FUEL_O", "VALUE_O"]]

        # Add missing columns explicitly to avoid future warning
        required_columns = ["TECHNOLOGY", "MODE_OF_OPERATION", "FUEL_I", "VALUE_I", "FUEL_O", "VALUE_O"]
        for col in required_columns:
            if col not in merged_extra.columns:
                merged_extra[col] = pd.NA

        # Set proper data types to avoid future warnings
        merged_extra = merged_extra.astype({
            "TECHNOLOGY": "string",
            "MODE_OF_OPERATION": "Int64",
            "FUEL_I": "string",
            "VALUE_I": "float",
            "FUEL_O": "string",
            "VALUE_O": "float"
        }, errors="ignore")

        # Reorder columns
        merged_extra = merged_extra[required_columns]

        # Concatenate merged_extra with merged
        merged = pd.concat([merged, merged_extra], ignore_index=True)

    # Handle 'PWRSDS' techs missing input data
    techs_output_only = df_input[
        df_input["TECHNOLOGY"].str.startswith("PWRSDS")
    ].copy()

    if not techs_output_only.empty:
        techs_output_only = techs_output_only.groupby(["TECHNOLOGY", "MODE_OF_OPERATION"], as_index=False).first()
        techs_output_only["VALUE_O"] = None
        techs_output_only["VALUE_I"] = techs_output_only["VALUE"]
        techs_output_only["FUEL_O"] = None

        merged_extra = techs_output_only.rename(columns={
            "TECHNOLOGY": "TECHNOLOGY",
            "MODE_OF_OPERATION": "MODE_OF_OPERATION",
            "FUEL": "FUEL_I"
        })[["TECHNOLOGY", "MODE_OF_OPERATION", "FUEL_I", "VALUE_I", "FUEL_O", "VALUE_O"]]

        # Add missing columns explicitly to avoid future warning
        required_columns = ["TECHNOLOGY", "MODE_OF_OPERATION", "FUEL_I", "VALUE_I", "FUEL_O", "VALUE_O"]
        for col in required_columns:
            if col not in merged_extra.columns:
                merged_extra[col] = pd.NA

        # Set proper data types to avoid future warnings
        merged_extra = merged_extra.astype({
            "TECHNOLOGY": "string",
            "MODE_OF_OPERATION": "Int64",
            "FUEL_I": "string",
            "VALUE_I": "float",
            "FUEL_O": "string",
            "VALUE_O": "float"
        }, errors="ignore")

        # Reorder columns
        merged_extra = merged_extra[required_columns]

        # Concatenate merged_extra with merged
        merged = pd.concat([merged, merged_extra], ignore_index=True)

    # Build output records
    records = []
    for _, row in merged.iterrows():
        tech = row["TECHNOLOGY"]
        mode = int(row["MODE_OF_OPERATION"])
        fuel_i = row["FUEL_I"]
        fuel_o = row["FUEL_O"]

        record = {
            "Mode.Operation": mode,
            "Fuel.I": fuel_i,
            "Fuel.I.Name": parse_fuel_name(fuel_i) if pd.notna(fuel_i) else None,
            "Value.Fuel.I": 1 if pd.notna(fuel_i) else None,
            "Unit.Fuel.I": None,
            "Tech": tech,
            "Tech.Name": parse_tech_name(tech),
            "Fuel.O": fuel_o,
            "Fuel.O.Name": parse_fuel_name(fuel_o) if pd.notna(fuel_o) else None,
            "Value.Fuel.O": 1 if pd.notna(fuel_o) else None,
            "Unit.Fuel.O": None
        }
        records.append(record)

    df_final = pd.DataFrame(records)

    # Clear and write to the Excel sheet
    ws = workbook["Secondary"]
    ws.delete_rows(1, ws.max_row)

    for r in dataframe_to_rows(df_final, index=False, header=True):
        ws.append(r)

    print("[Success] Sheet 'Secondary' in Base Year Model file updated.")
    return df_input,df_output,merged

def update_model_base_year_demand_techs(og_data, workbook):
    """
    Updates the 'Demand Techs' sheet in the base year model Excel workbook
    using filtered data from 'InputActivityRatio' and 'OutputActivityRatio'.
    Includes only technologies starting with 'PWRTRN' and excludes any starting with 'MIN' or 'RNW'.
    Requires input fuels ending in '01' and output fuels ending in '02'.
    """
    if "InputActivityRatio" not in og_data or "OutputActivityRatio" not in og_data:
        print("[Warning] Missing one or both parameters: 'InputActivityRatio', 'OutputActivityRatio'.")
        return

    df_input = og_data["InputActivityRatio"]
    df_output = og_data["OutputActivityRatio"]

    df_input = df_input[
        df_input["TECHNOLOGY"].str.startswith("PWRTRN") &
        (~df_input["TECHNOLOGY"].str.startswith(("MIN", "RNW"))) &
        df_input["FUEL"].str.endswith("01")
    ]
    df_output = df_output[
        df_output["TECHNOLOGY"].str.startswith("PWRTRN") &
        (~df_output["TECHNOLOGY"].str.startswith(("MIN", "RNW"))) &
        df_output["FUEL"].str.endswith("02")
    ]

    input_grouped = df_input.groupby(["TECHNOLOGY", "MODE_OF_OPERATION"], as_index=False).first()
    output_grouped = df_output.groupby(["TECHNOLOGY", "MODE_OF_OPERATION"], as_index=False).first()

    merged = pd.merge(
        input_grouped, output_grouped,
        on=["TECHNOLOGY", "MODE_OF_OPERATION"],
        suffixes=("_I", "_O")
    )

    records = []
    for _, row in merged.iterrows():
        tech = row["TECHNOLOGY"]
        mode = int(row["MODE_OF_OPERATION"])
        fuel_i = row["FUEL_I"]
        fuel_o = row["FUEL_O"]

        record = {
            "Mode.Operation": mode,
            "Fuel.I": fuel_i,
            "Fuel.I.Name": parse_fuel_name(fuel_i),
            "Value.Fuel.I": 1,
            "Unit.Fuel.I": None,
            "Tech": tech,
            "Tech.Name": parse_tech_name(tech),
            "Fuel.O": fuel_o,
            "Fuel.O.Name": parse_fuel_name(fuel_o),
            "Value.Fuel.O": 1,
            "Unit.Fuel.O": None
        }
        records.append(record)

    df_final = pd.DataFrame(records)

    ws = workbook["Demand Techs"]
    ws.delete_rows(1, ws.max_row)

    for r in dataframe_to_rows(df_final, index=False, header=True):
        ws.append(r)

    print("[Success] Sheet 'Demand Techs' in Base Year Model file updated.")
    
def update_projection_primary(og_data, workbook):
    """
    Updates the 'Primary' sheet in the projection Excel workbook using InputActivityRatio and OutputActivityRatio.
    Filters for technologies starting with 'MIN' or 'RNW'. Sets direction and adapts year columns based on parameter range.
    """
    if "InputActivityRatio" not in og_data or "OutputActivityRatio" not in og_data:
        print("[Warning] Missing one or both required parameters for Primary projection.")
        return

    df_input = og_data["InputActivityRatio"]
    df_output = og_data["OutputActivityRatio"]

    # Filter for technologies that start with MIN or RNW
    df_input = df_input[df_input["TECHNOLOGY"].str.startswith(("MIN", "RNW"))]
    df_output = df_output[df_output["TECHNOLOGY"].str.startswith(("MIN", "RNW"))]

    # Determine the union of all years used - use MODEL_YEARS as fallback if no data
    years_set = set(df_input["YEAR"]).union(df_output["YEAR"]) if not (df_input.empty and df_output.empty) else set()
    all_years = sorted(years_set) if years_set else MODEL_YEARS

    def build_records(df, direction):
        records = []
        for (tech, mode, fuel), group in df.groupby(["TECHNOLOGY", "MODE_OF_OPERATION", "FUEL"]):
            row = {
                "Mode.Operation": int(mode),
                "Tech": tech,
                "Tech.Name": parse_tech_name(tech),
                "Fuel": fuel,
                "Fuel.Name": parse_fuel_name(fuel),
                "Direction": direction,
                "Projection.Parameter": 0
            }

            year_values = {int(y): v for y, v in zip(group["YEAR"], group["VALUE"])}
            values = [year_values.get(y, np.nan) for y in all_years]

            # Determine Projection.Mode
            non_nan = [v for v in values if pd.notna(v)]
            if not non_nan:
                row["Projection.Mode"] = "EMPTY"
            elif len(non_nan) == 1 and not pd.isna(non_nan[0]):
                row["Projection.Mode"] = "Flat"
            elif len(non_nan) == len(all_years):
                row["Projection.Mode"] = "User defined"
            else:
                row["Projection.Mode"] = "interpolation"

            for y in all_years:
                row[str(y)] = year_values.get(y, np.nan)

            records.append(row)
        return records

    input_records = build_records(df_input, "Input")
    output_records = build_records(df_output, "Output")
    df_final = pd.DataFrame(input_records + output_records)

    # Reorder columns: fixed ones first, then years
    fixed_cols = [
        "Mode.Operation", "Tech", "Tech.Name", "Fuel", "Fuel.Name",
        "Direction", "Projection.Mode", "Projection.Parameter"
    ]
    df_final = df_final[fixed_cols + [str(y) for y in all_years]]
    df_final = df_final.sort_values(by=["Tech", "Direction"])

    # Write to sheet
    ws = workbook["Primary"]
    ws.delete_rows(1, ws.max_row)
    for r in dataframe_to_rows(df_final, index=False, header=True):
        ws.append(r)

    print("[Success] Sheet 'Primary' in Projections file updated.")
    
def update_projection_secondary(og_data, workbook):
    """
    Updates the 'Secondary' sheet in the projection Excel workbook using InputActivityRatio and OutputActivityRatio.
    Includes only technologies that do not start with 'MIN', 'RNW', or 'PWRTRN'.
    Sets direction as 'Input' or 'Output' and adapts year columns to the parameter data.
    """
    if "InputActivityRatio" not in og_data or "OutputActivityRatio" not in og_data:
        print("[Warning] Missing one or both required parameters for Secondary projection.")
        return

    df_input = og_data["InputActivityRatio"]
    df_output = og_data["OutputActivityRatio"]

    df_input = df_input[
        ~df_input["TECHNOLOGY"].str.startswith(("MIN", "RNW", "PWRTRN"))
    ]
    df_output = df_output[
        ~df_output["TECHNOLOGY"].str.startswith(("MIN", "RNW", "PWRTRN"))
    ]

    # Use MODEL_YEARS as fallback if no data
    years_set = set(df_input["YEAR"]).union(df_output["YEAR"]) if not (df_input.empty and df_output.empty) else set()
    all_years = sorted(years_set) if years_set else MODEL_YEARS

    def build_records(df, direction):
        records = []
        for (tech, mode, fuel), group in df.groupby(["TECHNOLOGY", "MODE_OF_OPERATION", "FUEL"]):
            row = {
                "Mode.Operation": int(mode),
                "Tech": tech,
                "Tech.Name": parse_tech_name(tech),
                "Fuel": fuel,
                "Fuel.Name": parse_fuel_name(fuel),
                "Direction": direction,
                "Projection.Parameter": 0
            }

            year_values = {int(y): v for y, v in zip(group["YEAR"], group["VALUE"])}
            values = [year_values.get(y, np.nan) for y in all_years]

            non_nan = [v for v in values if pd.notna(v)]
            if not non_nan:
                row["Projection.Mode"] = "EMPTY"
            elif len(non_nan) == 1 and not pd.isna(non_nan[0]):
                row["Projection.Mode"] = "Flat"
            elif len(non_nan) == len(all_years):
                row["Projection.Mode"] = "User defined"
            else:
                row["Projection.Mode"] = "interpolation"

            for y in all_years:
                row[str(y)] = year_values.get(y, np.nan)

            records.append(row)
        return records

    input_records = build_records(df_input, "Input")
    output_records = build_records(df_output, "Output")
    df_final = pd.DataFrame(input_records + output_records)

    fixed_cols = [
        "Mode.Operation", "Tech", "Tech.Name", "Fuel", "Fuel.Name",
        "Direction", "Projection.Mode", "Projection.Parameter"
    ]
    df_final = df_final[fixed_cols + [str(y) for y in all_years]]
    df_final = df_final.sort_values(by=["Tech", "Direction"])

    ws = workbook["Secondary"]
    ws.delete_rows(1, ws.max_row)
    for r in dataframe_to_rows(df_final, index=False, header=True):
        ws.append(r)

    print("[Success] Sheet 'Secondary' in Projections file updated.")

def update_projection_demand_techs(og_data, workbook):
    """
    Updates the 'Demand Techs' sheet in the projection Excel workbook using InputActivityRatio and OutputActivityRatio.
    Includes only technologies that start with 'PWRTRN', with input fuels ending in '01' and output fuels ending in '02'.
    Sorts the final result by 'Tech' and 'Direction', and adapts year columns to the parameter data.
    """
    if "InputActivityRatio" not in og_data or "OutputActivityRatio" not in og_data:
        print("[Warning] Missing one or both required parameters for Demand Techs projection.")
        return

    df_input = og_data["InputActivityRatio"]
    df_output = og_data["OutputActivityRatio"]

    df_input = df_input[
        df_input["TECHNOLOGY"].str.startswith("PWRTRN") &
        df_input["FUEL"].str.endswith("01")
    ]
    df_output = df_output[
        df_output["TECHNOLOGY"].str.startswith("PWRTRN") &
        df_output["FUEL"].str.endswith("02")
    ]

    # Use MODEL_YEARS as fallback if no data
    years_set = set(df_input["YEAR"]).union(df_output["YEAR"]) if not (df_input.empty and df_output.empty) else set()
    all_years = sorted(years_set) if years_set else MODEL_YEARS

    def build_records(df, direction):
        records = []
        for (tech, mode, fuel), group in df.groupby(["TECHNOLOGY", "MODE_OF_OPERATION", "FUEL"]):
            row = {
                "Mode.Operation": int(mode),
                "Tech": tech,
                "Tech.Name": parse_tech_name(tech),
                "Fuel": fuel,
                "Fuel.Name": parse_fuel_name(fuel),
                "Direction": direction,
                "Projection.Parameter": 0
            }

            year_values = {int(y): v for y, v in zip(group["YEAR"], group["VALUE"])}
            values = [year_values.get(y, np.nan) for y in all_years]

            non_nan = [v for v in values if pd.notna(v)]
            if not non_nan:
                row["Projection.Mode"] = "EMPTY"
            elif len(non_nan) == 1 and not pd.isna(non_nan[0]):
                row["Projection.Mode"] = "Flat"
            elif len(non_nan) == len(all_years):
                row["Projection.Mode"] = "User defined"
            else:
                row["Projection.Mode"] = "interpolation"

            for y in all_years:
                row[str(y)] = year_values.get(y, np.nan)

            records.append(row)
        return records

    input_records = build_records(df_input, "Input")
    output_records = build_records(df_output, "Output")
    df_final = pd.DataFrame(input_records + output_records)

    fixed_cols = [
        "Mode.Operation", "Tech", "Tech.Name", "Fuel", "Fuel.Name",
        "Direction", "Projection.Mode", "Projection.Parameter"
    ]
    df_final = df_final[fixed_cols + [str(y) for y in all_years]]
    df_final = df_final.sort_values(by=["Tech", "Direction"])

    ws = workbook["Demand Techs"]
    ws.delete_rows(1, ws.max_row)
    for r in dataframe_to_rows(df_final, index=False, header=True):
        ws.append(r)

    print("[Success] Sheet 'Demand Techs' in Projections file updated.")

def update_xtra_storage_fixed_horizon_parameters(og_data, workbook):
    """
    Updates the 'Fixed Horizon Parameters' sheet in A-Xtra_Storage.xlsx
    using StorageLevelStart and OperationalLifeStorage data.
    Missing values are filled with default = 1. Includes storage name column.
    """

    parameters = [
        ("StorageLevelStart", 1),
        ("OperationalLifeStorage", 2),
    ]

    all_techs = set()
    param_data = {}

    for param_name, param_id in parameters:
        if param_name not in og_data:
            continue
        df = og_data[param_name]
        param_data[param_name] = {}
        for _, row in df.iterrows():
            tech = row["STORAGE"]
            param_data[param_name][tech] = row["VALUE"]
            all_techs.add(tech)

    tech_ids = {tech: idx + 1 for idx, tech in enumerate(sorted(all_techs))}

    rows = []
    for tech in sorted(all_techs):
        for param_name, param_id in parameters:
            value = param_data.get(param_name, {}).get(tech, 1)
            rows.append({
                "STORAGE.ID": tech_ids[tech],
                "STORAGE": tech,
                "STORAGE.Name": parse_tech_name(tech),
                "Parameter.ID": param_id,
                "Parameter": param_name,
                "Unit": None,
                "Value": value
            })

    df_out = pd.DataFrame(rows).sort_values(by=["STORAGE", "Parameter.ID"])

    ws = workbook["Fixed Horizon Parameters"]
    ws.delete_rows(1, ws.max_row)

    for r in dataframe_to_rows(df_out, index=False, header=True):
        ws.append(r)

    print("[Success] Sheet 'Fixed Horizon Parameters' in Extra Storage updated.")


def update_xtra_storage_capital_cost_storage(og_data, workbook):
    """
    Updates the 'CapitalCostStorage' sheet in the A-Xtra_Storage Excel workbook
    using the 'CapitalCostStorage' and 'ResidualStorageCapacity' parameters.
    Missing storage technologies in one parameter are filled with EMPTY mode in the other.
    Rows are sorted by 'STORAGE.ID'.
    """
    param_list = [("CapitalCostStorage", 1), ("ResidualStorageCapacity", 2)]
    all_storages = set()
    param_data = {}

    for param_name, param_id in param_list:
        if param_name not in og_data:
            print(f"[Warning] '{param_name}' not found in OG_Input_Data.")
            continue
        df = og_data[param_name]
        df_grouped = df.groupby("STORAGE")
        all_storages.update(df["STORAGE"].unique())

        records = {}
        for storage, group in df_grouped:
            year_values = {int(row["YEAR"]): row["VALUE"] for _, row in group.iterrows()}
            available_years = sorted(group["YEAR"].unique())
            values = [year_values.get(y, np.nan) for y in available_years]
            non_nan_count = sum(pd.notna(values))

            if non_nan_count == 0:
                mode = "EMPTY"
            elif non_nan_count == 1:
                mode = "Flat"
            elif non_nan_count == len(values):
                mode = "User defined"
            else:
                mode = "interpolation"

            records[storage] = {
                "Projection.Mode": mode,
                "Year.Values": year_values,
                "Years": available_years
            }
        param_data[param_name] = records

    # Use MODEL_YEARS as fallback if no data
    years_set = set(
        y for pdata in param_data.values()
        for pinfo in pdata.values()
        for y in pinfo["Years"]
    )
    all_years = sorted(years_set) if years_set else MODEL_YEARS

    storage_ids = {name: idx + 1 for idx, name in enumerate(sorted(all_storages))}
    final_records = []

    for storage in sorted(all_storages):
        for param_name, param_id in param_list:
            pdata = param_data.get(param_name, {})
            pinfo = pdata.get(storage)

            record = {
                "STORAGE.ID": storage_ids[storage],
                "STORAGE": storage,
                "STORAGE.Name": parse_tech_name(storage),
                "Parameter.ID": param_id,
                "Parameter": param_name,
                "Unit": None,
                "Projection.Parameter": None
            }

            if pinfo is None:
                record["Projection.Mode"] = "EMPTY"
                for y in all_years:
                    record[int(y)] = np.nan
            else:
                record["Projection.Mode"] = pinfo["Projection.Mode"]
                for y in all_years:
                    record[int(y)] = pinfo["Year.Values"].get(y, np.nan)

            final_records.append(record)

    df_out = pd.DataFrame(final_records)
    df_out = df_out.sort_values(by=["STORAGE.ID"])
    df_out = df_out[
        ["STORAGE.ID", "STORAGE", "STORAGE.Name", "Parameter.ID", "Parameter",
         "Unit", "Projection.Mode", "Projection.Parameter"] + all_years
    ]

    ws = workbook["CapitalCostStorage"]
    ws.delete_rows(1, ws.max_row)
    for r in dataframe_to_rows(df_out, index=False, header=True):
        ws.append(r)

    print("[Success] Sheet 'CapitalCostStorage' in Extra Storage updated.")
    return workbook, df_out.head()

def update_xtra_storage_technology_storage(og_data, workbook):
    """
    Updates the 'TechnologyStorage' sheet in A-Xtra_Storage.xlsx
    using TechnologyToStorage and TechnologyFromStorage parameters.
    Combines both into a unified format with clear naming and direction.
    """

    if "TechnologyToStorage" not in og_data or "TechnologyFromStorage" not in og_data:
        print("[Warning] Missing TechnologyToStorage or TechnologyFromStorage in OG_Input_Data.")
        return workbook

    df_to = og_data["TechnologyToStorage"].copy()
    df_from = og_data["TechnologyFromStorage"].copy()

    # Common fields
    def prepare_df(df, param_name, param_id):
        return pd.DataFrame({
            "MODE_OF_OPERATION": df["MODE_OF_OPERATION"],
            "TECHNOLOGY": df["TECHNOLOGY"],
            "TECHNOLOGY.Name": df["TECHNOLOGY"].apply(parse_tech_name),
            "STORAGE": df["STORAGE"],
            "STORAGE.Name": df["STORAGE"].apply(parse_tech_name),
            "Parameter.ID": param_id,
            "Parameter": param_name,
            "Value.STORAGE": df["VALUE"],
            "Unit.STORAGE": None
        })

    df_to_prepared = prepare_df(df_to, "TechnologyToStorage", 1)
    df_from_prepared = prepare_df(df_from, "TechnologyFromStorage", 2)

    df_out = pd.concat([df_to_prepared, df_from_prepared], ignore_index=True)
    df_out = df_out.sort_values(by=["TECHNOLOGY", "STORAGE", "MODE_OF_OPERATION"])

    # Write to the Excel workbook
    ws = workbook["TechnologyStorage"]
    ws.delete_rows(1, ws.max_row)

    for r in dataframe_to_rows(df_out, index=False, header=True):
        ws.append(r)

    print("[Success] Sheet 'TechnologyStorage' in Extra Storage updated.")
    return workbook    

def update_yaml_conversions(og_data, yaml_path):
    """
    Updates Conversionls, Conversionld, Conversionlh values in a YAML file using OG_Input_Data.
    Replaces the lists while preserving inline comments and formatting.
    """
    params = ["Conversionls", "Conversionld", "Conversionlh"]

    # Extract values from OG_Input_Data
    replacements = {}
    for param in params:
        if param in og_data:
            values = og_data[param]["VALUE"].astype(int).tolist()
            replacements[param] = values
        else:
            print(f"[Warning] {param} not found in OG_Input_Data.")

    # Read the YAML file as plain text
    with open(yaml_path, "r", encoding="utf-8") as f:
        lines = f.readlines()

    # Update each conversion line by pattern
    updated_lines = []
    for line in lines:
        matched = False
        for param in params:
            pattern = rf"^({param}:\s*)\[[^\]]*\](\s*#.*)$"
            match = re.match(pattern, line)
            if match:
                prefix, suffix = match.groups()
                new_list = ", ".join(map(str, replacements.get(param, [])))
                line = f"{prefix}[{new_list}]{suffix}\n"
                matched = True
                break
        updated_lines.append(line if not matched else line)

    # Write the updated YAML content back
    with open(yaml_path, "w", encoding="utf-8") as f:
        f.writelines(updated_lines)

    print("[Success] Lists: 'Conversionls', 'Conversionlh' and 'Conversionld'\n in MOMF_T1_A file updated.")

def update_yaml_xtra_scen(og_data, yaml_path):
    """
    Updates the xtra_scen block of a YAML file using mapped keys from OG_Input_Data.
    Replaces only the values (inside [] or '') in the corresponding xtra_scen lines.
    Preserves formatting and comments.
    """

    # Mapping of OG_Input_Data keys to xtra_scen keys
    key_map = {
        "REGION": "Region",
        "MODE_OF_OPERATION": "Mode_of_Operation",
        "SEASON": "Season",
        "DAYTYPE": "DayType",
        "DAILYTIMEBRACKET": "DailyTimeBracket",
        "TIMESLICE": "Timeslices",
        "STORAGE": "Storage"
    }

    # Parameters that must be strings in the YAML list
    force_str_keys = {"Season", "DayType", "DailyTimeBracket", "Timeslices", "Storage"}

    replacements = {}
    for og_key, yaml_key in key_map.items():
        if og_key in og_data:
            values = og_data[og_key]["VALUE"].tolist()
            if yaml_key == "Region":
                replacements[yaml_key] = str(values[0]) if values else ""
            else:
                if yaml_key in force_str_keys:
                    replacements[yaml_key] = [f"'{str(v)}'" for v in values]
                else:
                    replacements[yaml_key] = [int(v) if isinstance(v, (int, float)) else str(v) for v in values]

    # Read YAML as plain text lines
    with open(yaml_path, "r", encoding="utf-8") as f:
        lines = f.readlines()

    updated_lines = []
    i = 0
    while i < len(lines):
        line = lines[i]
        matched = False

        for yaml_key, new_values in replacements.items():
            if yaml_key == "Region":
                # Match Region with or without quotes
                pattern = rf"^(\s*{yaml_key}:\s*).*$"
                match = re.match(pattern, line)
                if match:
                    prefix = match.group(1)
                    updated_lines.append(f"{prefix}{new_values}\n")
                    i += 1
                    matched = True
                    break
            else:
                # Check for inline list: Key: [item1, item2]
                inline_pattern = rf"^(\s*{yaml_key}:\s*)\[.*\](.*)$"
                inline_match = re.match(inline_pattern, line)
                if inline_match:
                    prefix, suffix = inline_match.groups()
                    formatted = ", ".join(map(str, new_values))
                    updated_lines.append(f"{prefix}[{formatted}]{suffix}\n")
                    i += 1
                    matched = True
                    break

                # Check for multi-line list: Key:\n  - item1\n  - item2
                multiline_pattern = rf"^(\s*){yaml_key}:\s*$"
                multiline_match = re.match(multiline_pattern, line)
                if multiline_match:
                    indent = multiline_match.group(1)
                    updated_lines.append(line)
                    i += 1
                    # Skip old list items
                    while i < len(lines) and re.match(rf"^{indent}- ", lines[i]):
                        i += 1
                    # Insert new list items
                    for val in new_values:
                        updated_lines.append(f"{indent}- {val}\n")
                    matched = True
                    break

        if not matched:
            updated_lines.append(line)
            i += 1

    with open(yaml_path, "w", encoding="utf-8") as f:
        f.writelines(updated_lines)

    print("[Success] Dict 'xtra_scen' in MOMF_T1_A file updated.")

def update_yaml_years(og_data, yaml_path):
    """
    Updates the base_year, initial_year, and final_year fields in the YAML file
    based on the first and last YEAR value found in the OG_Input_Data dictionary.
    Properly clears and replaces quoted year values.
    """

    if "YEAR" not in og_data or "VALUE" not in og_data["YEAR"].columns:
        print("[Warning] 'YEAR' parameter with column 'VALUE' not found in OG_Input_Data.")
        return

    years = sorted(og_data["YEAR"]["VALUE"].unique())
    if not years:
        print("[Warning] YEAR data is empty.")
        return

    base_year = str(int(years[0]))
    final_year = str(int(years[-1]))

    # Read YAML as plain text
    with open(yaml_path, "r", encoding="utf-8") as f:
        lines = f.readlines()

    updated_lines = []
    for line in lines:
        if "base_year:" in line:
            line = re.sub(r"(base_year:\s*)['\"].*?['\"]", rf"\1'{base_year}'", line)
        elif "initial_year:" in line:
            line = re.sub(r"(initial_year:\s*)['\"].*?['\"]", rf"\1'{base_year}'", line)
        elif "final_year:" in line:
            line = re.sub(r"(final_year:\s*)['\"].*?['\"]", rf"\1'{final_year}'", line)
        updated_lines.append(line)

    with open(yaml_path, "w", encoding="utf-8") as f:
        f.writelines(updated_lines)

    print("[Success] YAML years variables in MOMF_T1_A file updated.")
#--------------------------------------------------------------------------------------------------#

#-------------------------------------Updated main functions---------------------------------------#

def update_model_base_year(og_data, input_excel_path, output_excel_path):
    """
    Orchestrates the update process for the base year model Excel file.
    Updates the 'Primary', 'Secondary', and 'Demand Techs' sheets using OG_Input_Data.
    Loads the base workbook from input_excel_path and saves to output_excel_path.
    """
    os.makedirs(os.path.dirname(output_excel_path), exist_ok=True)
    wb = load_workbook(input_excel_path)

    # Update Primary sheet
    update_model_base_year_primary(og_data, wb)

    # Update Secondary sheet
    df_input,df_output,merged=update_model_base_year_secondary(og_data, wb)

    # Update Demand Techs sheet
    update_model_base_year_demand_techs(og_data, wb)

    # Save final workbook
    wb.save(output_excel_path)
    print("[Success] Excel file 'Model Base Year' updated.")
    print("-------------------------------------------------------------------------\n")
    return df_input,df_output,merged

def update_demand(og_data, input_excel_path, output_excel_path):
    """
    Orchestrates updates to the demand Excel file.
    Applies profile and demand projection updates using OG_Input_Data.
    Assumes all necessary keys exist in og_data.
    """
    os.makedirs(os.path.dirname(output_excel_path), exist_ok=True)

    update_demand_profiles(
        df=og_data["SpecifiedDemandProfile"],
        output_excel_path=output_excel_path,
        input_excel_path=input_excel_path
    )

    update_demand_demand_projection(
        df=og_data["SpecifiedAnnualDemand"],
        output_excel_path=output_excel_path,
        input_excel_path=input_excel_path
    )
    
    print("[Success] Excel file 'Demand' updated.")
    print("-------------------------------------------------------------------------\n")
    
def update_parametrization(og_data, output_excel_path, input_excel_path):
    """
    Executes all update routines for the A-O_Parametrization Excel file.
    Applies fixed horizon parameters, timeslices, tech parameter sheets, and yearsplit.
    Assumes all required keys exist in og_data.
    """
    os.makedirs(os.path.dirname(output_excel_path), exist_ok=True)

    update_parametrization_fixed_horizon_parameters(
        df_ctau=og_data["CapacityToActivityUnit"],
        df_oplife=og_data["OperationalLife"],
        output_excel_path=output_excel_path,
        input_excel_path=input_excel_path
    )

    update_parametrization_capacities(
        df=og_data["CapacityFactor"],
        output_excel_path=output_excel_path
    )

    update_parametrization_primary_secondary_demand_techs(
        og_data=og_data,
        output_excel_path=output_excel_path
    )
    
    update_parametrization_variable_cost(
        og_data=og_data,
        output_excel_path=output_excel_path
    )

    update_parametrization_yearsplit(
        df=og_data["YearSplit"],
        output_excel_path=output_excel_path
    )
    
    update_parametrization_daysplit(
        df=og_data["DaySplit"],
        output_excel_path=output_excel_path
    )

    # Update year headers for sheets that may not have new data but need aligned years
    # Other_Techs has 7 fixed columns: Application, Tech, Tech.Name, Fuel, Parameter, Unit, Projection.Mode
    update_sheet_year_headers(output_excel_path, "Other_Techs", MODEL_YEARS, fixed_cols_count=7)

    print("[Success] Excel file 'Parametrization' updated.")
    print("-------------------------------------------------------------------------\n")
    
def update_projections(og_data, input_excel_path, output_excel_path):
    """
    Coordinates the update of the A-O_AR_Projections Excel file.
    Updates the 'Primary', 'Secondary', and 'Demand Techs' sheets using OG_Input_Data.
    Each update adapts year columns based on actual parameter data ranges and sorts by 'Tech' and 'Direction'.
    """
    os.makedirs(os.path.dirname(output_excel_path), exist_ok=True)
    wb = load_workbook(input_excel_path)

    # Update each sheet with sorting logic inside each function
    update_projection_primary(og_data, wb)
    update_projection_secondary(og_data, wb)
    update_projection_demand_techs(og_data, wb)

    wb.save(output_excel_path)
    print("[Success] Excel file 'Projections' updated.")
    print("-------------------------------------------------------------------------\n")

def update_xtra_emissions(og_data, input_excel_path, output_excel_path):
    """
    Coordinates update of A-Xtra_Emissions.xlsx.
    Applies updates to GHGs and Externalities sheets.
    """
    os.makedirs(os.path.dirname(output_excel_path), exist_ok=True)
    wb = load_workbook(input_excel_path)

    update_xtra_emissions_ghg(og_data, wb)
    update_xtra_emissions_externalities(og_data, wb)

    wb.save(output_excel_path)
    print("[Success] Excel file 'Xtra Emissions' updated.")
    print("-------------------------------------------------------------------------\n")

def update_xtra_storage(og_data, input_excel_path, output_excel_path):
    """
    Main function to update A-Xtra_Storage.xlsx using multiple storage-related parameters.
    Calls individual update functions for each sheet.
    """
    os.makedirs(os.path.dirname(output_excel_path), exist_ok=True)
    wb = load_workbook(input_excel_path)

    update_xtra_storage_fixed_horizon_parameters(og_data, wb)
    update_xtra_storage_capital_cost_storage(og_data, wb)
    update_xtra_storage_technology_storage(og_data, wb)

    wb.save(output_excel_path)
    print("[Success] Excel file 'Xtra Storage' updated.")

def update_yaml_structure(og_data, yaml_path):
    """
    Executes YAML updates using OG_Input_Data:
    - Updates Conversionls, Conversionld, Conversionlh
    - Updates xtra_scen block
    - Updates base_year, initial_year, final_year
    """
    update_yaml_conversions(og_data, yaml_path)
    update_yaml_xtra_scen(og_data, yaml_path)
    update_yaml_years(og_data, yaml_path)
    
    print("[Success] Yaml file 'MOMF_T1_A' updated.")
    print("-------------------------------------------------------------------------\n")

def sort_csv_files_in_folder(folder_path):
    if not os.path.isdir(folder_path):
        print(f"The path is invalid: {folder_path}")
        return
    print('################################################################')
    print('Sort csv files.')
    for filename in os.listdir(folder_path):
        if filename.endswith(".csv"):
            file_path = os.path.join(folder_path, filename)
            print(f"Processing: {filename}")
            try:
                # Read the CSV preserving the header
                df = pd.read_csv(file_path)

                # Sort using all columns
                df_sorted = df.sort_values(by=list(df.columns))

                # Overwrite the original file
                df_sorted.to_csv(file_path, index=False)
            except Exception as e:
                print(f"Error processing {filename}: {e}")

    print("✅ All files were sort.")
    print('################################################################\n')

#--------------------------------------------------------------------------------------------------#
def main():
    """Main execution function."""
    os.makedirs(INPUT_FOLDER, exist_ok=True)
    sort_csv_files_in_folder(INPUT_FOLDER)
    global OG_Input_Data
    OG_Input_Data = read_csv_files(INPUT_FOLDER)

    # ⚡ NORMALIZE TEMPORAL PROFILES (SpecifiedDemandProfile, YearSplit, DaySplit)
    # This ensures all profiles sum to 1.0 for proper OSeMOSYS input compliance
    OG_Input_Data = normalize_temporal_profiles(OG_Input_Data)

    # Replace JAM -> BRB for Barbados (original model used JAM incorrectly)
    OG_Input_Data = replace_country_codes(OG_Input_Data, 'JAM', 'BRB')

    # Filter data to start from FIRST_YEAR (removes earlier years to avoid overlap with old model)
    OG_Input_Data = filter_data_by_first_year(OG_Input_Data, FIRST_YEAR)

    # Load tech-country matrix configuration (runs BEFORE other processing)
    matrix_config = load_tech_country_matrix()

    # Apply tech-country matrix filtering (if enabled)
    OG_Input_Data = filter_by_tech_country_matrix(OG_Input_Data, matrix_config)

    # Unify CCG + OCG -> NGS (if enabled)
    OG_Input_Data = unify_ngs_technologies(OG_Input_Data, matrix_config)

    # Apply region consolidation if enabled
    OG_Input_Data = consolidate_regions(OG_Input_Data)

    # Clean/merge PWR technologies based on config flag
    pwr_mode = get_pwr_cleanup_mode()
    if pwr_mode == "drop":
        OG_Input_Data = clean_pwr_technologies(OG_Input_Data)
    elif pwr_mode == "merge":
        OG_Input_Data = merge_pwr_technologies(OG_Input_Data, matrix_config)
    else:
        print("[Info] PWR technology cleanup is disabled (pwr_cleanup_mode = false).")

    scenario_suffixes = list_scenario_suffixes(OUTPUT_FOLDER)
    for scen in scenario_suffixes:
        print('\nScenario process: ',scen)
        # File A-O_Demand.xlsx
        try:
            update_demand(
                og_data=OG_Input_Data,
                input_excel_path=MISCELLANEOUS_FOLDER / "A-O_Demand.xlsx",
                output_excel_path=OUTPUT_FOLDER / f"A1_Outputs_{scen}" / "A-O_Demand.xlsx"
            )
        except KeyError as e:
            print(f"[KeyError] Missing key in OG_Input_Data: {e}")
        except Exception as e:
            print(f"[Error] Failed to update demand file: {e}")
    
        # File A-O_Parametrization.xlsx
        try:
            update_parametrization(
                og_data=OG_Input_Data,
                input_excel_path=MISCELLANEOUS_FOLDER / "A-O_Parametrization.xlsx",
                output_excel_path=OUTPUT_FOLDER / f"A1_Outputs_{scen}" / "A-O_Parametrization.xlsx"
            )
        except KeyError as e:
            print(f"[KeyError] Missing key in OG_Input_Data: {e}")
        except Exception as e:
            print(f"[Error] Failed to update parametrization file: {e}")
        
        # File A-Xtra_Emissions.xlsx
        try:
            update_xtra_emissions(
                og_data=OG_Input_Data,
                input_excel_path=MISCELLANEOUS_FOLDER / "A-Xtra_Emissions.xlsx",
                output_excel_path=A2_EXTRA_INPUTS_FOLDER / "A-Xtra_Emissions.xlsx"
            )
        except KeyError as e:
            print(f"[KeyError] Missing key in OG_Input_Data: {e}")
        except Exception as e:
            print(f"Failed to update extra emissions file: {e}")
    
        # File A-O_AR_Model_Base_Year.xlsx
        # try:
        df_input,df_output,merged=update_model_base_year(
            og_data=OG_Input_Data,
            input_excel_path=MISCELLANEOUS_FOLDER / "A-O_AR_Model_Base_Year.xlsx",
            output_excel_path=OUTPUT_FOLDER / f"A1_Outputs_{scen}" / "A-O_AR_Model_Base_Year.xlsx"
        )
        # except KeyError as e:
        #     print(f"[KeyError] Missing key in OG_Input_Data: {e}")
        # except Exception as e:
        #     print(f"Failed to update model base year file: {e}")
    
        try:
            update_projections(
                og_data=OG_Input_Data,
                input_excel_path=MISCELLANEOUS_FOLDER / "A-O_AR_Projections.xlsx",
                output_excel_path=OUTPUT_FOLDER / f"A1_Outputs_{scen}" / "A-O_AR_Projections.xlsx"
            )
        except Exception as e:
            print(f"[Error] Failed to update projections file: {e}")
            
        try:
            update_yaml_structure(
                og_data=OG_Input_Data,
                yaml_path=SCRIPT_DIR / "Config_MOMF_T1_A.yaml"
            )
        except Exception as e:
            print(f"[Error] Failed to update YAML structure: {e}")
    
    
        try:
            update_xtra_storage(
                og_data=OG_Input_Data,
                input_excel_path=MISCELLANEOUS_FOLDER / "A-Xtra_Storage.xlsx",
                output_excel_path=A2_EXTRA_INPUTS_FOLDER / "A-Xtra_Storage.xlsx"
            )
        except Exception as e:
            print(f"[Error] Failed to update storage file: {e}")

    # Write updated OG_csvs_inputs files (including EMISSION.csv with CO2BRB)
    # This ensures B1_Compiler picks up the correct EMISSION set
    try:
        write_csv_files(OG_Input_Data, INPUT_FOLDER)
    except Exception as e:
        print(f"[Error] Failed to write CSV files: {e}")

    return df_input,df_output,merged


if __name__ == "__main__":
    df_input,df_output,merged=main()
