"""
Update Secondary Techs from Editor Template

This script reads the Secondary_Techs_Editor.xlsx file and applies
the changes to the corresponding A-O_Parametrization.xlsx files.

Usage:
    python t1_confection/D2_update_secondary_techs.py
"""
import openpyxl
import sys
from pathlib import Path
from datetime import datetime
import shutil
import yaml
from Z_AUX_config_loader import (
    get_olade_country_mapping, get_olade_country_mapping_normalized,
    get_olade_tech_mapping, get_shares_country_mapping, get_shares_tech_mapping,
    strip_accents
)

# Country and technology mappings from centralized config
OLADE_COUNTRY_MAPPING = get_olade_country_mapping()
OLADE_COUNTRY_MAPPING_NORMALIZED = get_olade_country_mapping_normalized()
OLADE_TECH_MAPPING = get_olade_tech_mapping()


def read_base_scenario():
    """
    Read base_scenario from Config_MOMF_T1_AB.yaml

    Returns:
        str: base scenario name (default: 'BAU')
    """
    yaml_path = Path(__file__).parent / "Config_MOMF_T1_AB.yaml"
    try:
        with open(yaml_path, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f)
            base_scenario = config.get('base_scenario', 'BAU')
            return base_scenario
    except Exception as e:
        print(f"Warning: Could not read base_scenario from YAML: {e}")
        print("Using default base_scenario: BAU")
        return 'BAU'


def read_olade_config(editor_path):
    """
    Read OLADE configuration from the editor Excel file

    Returns:
        dict with config: {enabled, petroleum_split_mode, demand_enabled, activity_lower_limit_enabled,
                          activity_upper_limit_enabled, demand_growth_rates, scenarios_demand_adjustments,
                          renewability_targets, technology_weights}
    """
    wb = openpyxl.load_workbook(editor_path, data_only=True)

    if 'OLADE_Config' not in wb.sheetnames:
        wb.close()
        return {
            'enabled': False,
            'petroleum_split_mode': 'Split_PET_OIL',
            'demand_enabled': False,
            'activity_lower_limit_enabled': False,
            'activity_upper_limit_enabled': False,
            'trade_balance_enabled': False,
            'interconnections_enabled': False,
            'demand_growth_rates': {},
            'scenarios_demand_adjustments': {},
            'renewability_targets': {},
            'technology_weights': {}
        }

    ws = wb['OLADE_Config']

    # Read configuration values
    # Row 5 = ResidualCapacitiesFromOLADE, Row 6 = PetroleumSplitMode, Row 7 = DemandFromOLADE
    # Row 8 = ActivityLowerLimitFromOLADE, Row 9 = ActivityUpperLimitFromOLADE
    # Row 10 = TradeBalanceDemandAdjustment, Row 11 = InterconnectionsControl
    enabled = str(ws['B5'].value).upper() == 'YES' if ws['B5'].value else False
    petroleum_split_mode = str(ws['B6'].value) if ws['B6'].value else 'Split_PET_OIL'
    demand_enabled = str(ws['B7'].value).upper() == 'YES' if ws['B7'].value else False
    activity_lower_limit_enabled = str(ws['B8'].value).upper() == 'YES' if ws['B8'].value else False
    activity_upper_limit_enabled = str(ws['B9'].value).upper() == 'YES' if ws['B9'].value else False
    trade_balance_enabled = str(ws['B10'].value).upper() == 'YES' if ws['B10'].value else False
    interconnections_enabled = str(ws['B11'].value).upper() == 'YES' if ws['B11'].value else False

    # Read demand growth rates from Demand_Growth sheet
    demand_growth_rates = {}
    if 'Demand_Growth' in wb.sheetnames:
        ws_demand = wb['Demand_Growth']
        # Data starts at row 5, Column A = country code, Column C = growth rate
        for row_idx in range(5, ws_demand.max_row + 1):
            country_code = ws_demand.cell(row_idx, 1).value
            growth_rate = ws_demand.cell(row_idx, 3).value
            if country_code and growth_rate is not None:
                try:
                    # Convert percentage to decimal (e.g., 2.0 -> 0.02)
                    demand_growth_rates[str(country_code).strip()] = float(growth_rate) / 100.0
                except (ValueError, TypeError):
                    pass

    # Read renewability targets from Renewability_Targets sheet
    # Structure: {(country, scenario): {'interpolation': str, 'targets': {year: percentage}}}
    renewability_targets = {}
    if 'Renewability_Targets' in wb.sheetnames:
        ws_renew = wb['Renewability_Targets']
        # Get year columns from header (row 5)
        year_cols = {}
        for col_idx in range(4, ws_renew.max_column + 1):
            header = ws_renew.cell(5, col_idx).value
            if header and str(header).isdigit():
                year_cols[int(header)] = col_idx

        # Data starts at row 6
        for row_idx in range(6, ws_renew.max_row + 1):
            country = ws_renew.cell(row_idx, 1).value
            scenario = ws_renew.cell(row_idx, 2).value
            interpolation = ws_renew.cell(row_idx, 3).value

            if not country or not scenario:
                continue

            country_str = str(country).strip().upper()
            scenario_str = str(scenario).strip()
            interpolation_str = str(interpolation).strip().lower() if interpolation else 'flat_step'

            targets = {}
            for year, col_idx in year_cols.items():
                value = ws_renew.cell(row_idx, col_idx).value
                if value is not None:
                    try:
                        # Value might be percentage (0.5) or decimal (50)
                        pct = float(value)
                        if pct > 1:  # Assume it's a percentage like 50 instead of 0.5
                            pct = pct / 100.0
                        targets[year] = pct
                    except (ValueError, TypeError):
                        pass

            if targets:  # Only add if there are any targets defined
                renewability_targets[(country_str, scenario_str)] = {
                    'interpolation': interpolation_str,
                    'targets': targets
                }

    # Read scenario-specific demand growth adjustments from Scenarios_Demand_Growth sheet
    # Structure: {(country, scenario): {year: adjustment_percentage}}
    scenarios_demand_adjustments = {}
    if 'Scenarios_Demand_Growth' in wb.sheetnames:
        ws_scen_demand = wb['Scenarios_Demand_Growth']
        # Get year columns from header (row 5)
        year_cols = {}
        for col_idx in range(3, ws_scen_demand.max_column + 1):
            header = ws_scen_demand.cell(5, col_idx).value
            if header and str(header).isdigit():
                year_cols[int(header)] = col_idx

        # Data starts at row 6
        for row_idx in range(6, ws_scen_demand.max_row + 1):
            country = ws_scen_demand.cell(row_idx, 1).value
            scenario = ws_scen_demand.cell(row_idx, 2).value

            if not country or not scenario:
                continue

            country_str = str(country).strip().upper()
            scenario_str = str(scenario).strip()

            adjustments = {}
            for year, col_idx in year_cols.items():
                value = ws_scen_demand.cell(row_idx, col_idx).value
                if value is not None:
                    try:
                        # Value is stored as percentage (e.g., 0.05 for 5%)
                        # Convert to decimal if needed
                        adj_pct = float(value)
                        # If value > 1, assume it's like 5 instead of 0.05
                        if abs(adj_pct) > 1:
                            adj_pct = adj_pct / 100.0
                        adjustments[year] = adj_pct
                    except (ValueError, TypeError):
                        pass

            if adjustments:  # Only add if there are any adjustments defined
                scenarios_demand_adjustments[(country_str, scenario_str)] = adjustments

    # Read technology weights from Technology_Weights sheet (formerly Renewable_Weights)
    # Structure: {(country, scenario): {'renewable': {'HYD': w, ...}, 'non_renewable': {'COA': w, ...}}}
    technology_weights = {}
    weights_sheet_name = 'Technology_Weights' if 'Technology_Weights' in wb.sheetnames else 'Renewable_Weights'
    if weights_sheet_name in wb.sheetnames:
        ws_weights = wb[weights_sheet_name]

        # Find header row (contains 'Country')
        header_row = None
        for row_idx in range(1, min(15, ws_weights.max_row + 1)):
            if ws_weights.cell(row_idx, 1).value == 'Country':
                header_row = row_idx
                break

        if header_row:
            # Get column indices for each technology from header
            tech_cols = {}
            for col_idx in range(3, ws_weights.max_column + 1):
                header = ws_weights.cell(header_row, col_idx).value
                if header:
                    tech_cols[str(header).strip().upper()] = col_idx

            renewable_techs = ['HYD', 'SPV', 'WON', 'GEO', 'BIO']
            non_renewable_techs = ['COA', 'PET', 'NGS', 'URN', 'OIL']

            # Data starts after header row
            for row_idx in range(header_row + 1, ws_weights.max_row + 1):
                country = ws_weights.cell(row_idx, 1).value
                scenario = ws_weights.cell(row_idx, 2).value

                if not country or not scenario:
                    continue

                country_str = str(country).strip().upper()
                scenario_str = str(scenario).strip()

                renewable_weights = {}
                non_renewable_weights = {}
                has_weights = False

                for tech, col_idx in tech_cols.items():
                    value = ws_weights.cell(row_idx, col_idx).value
                    if value is not None:
                        try:
                            weight = float(value)
                            if tech in renewable_techs:
                                renewable_weights[tech] = weight
                            elif tech in non_renewable_techs:
                                non_renewable_weights[tech] = weight
                            has_weights = True
                        except (ValueError, TypeError):
                            pass

                if has_weights:
                    technology_weights[(country_str, scenario_str)] = {
                        'renewable': renewable_weights,
                        'non_renewable': non_renewable_weights
                    }

    wb.close()

    return {
        'enabled': enabled,
        'petroleum_split_mode': petroleum_split_mode,
        'demand_enabled': demand_enabled,
        'activity_lower_limit_enabled': activity_lower_limit_enabled,
        'activity_upper_limit_enabled': activity_upper_limit_enabled,
        'trade_balance_enabled': trade_balance_enabled,
        'interconnections_enabled': interconnections_enabled,
        'demand_growth_rates': demand_growth_rates,
        'scenarios_demand_adjustments': scenarios_demand_adjustments,
        'renewability_targets': renewability_targets,
        'technology_weights': technology_weights
    }


def read_shares_data(shares_file_path):
    """
    Read Shares_PET_OIL_Split.xlsx file to get Diésel, Fuel oil, and Búnker shares by scenario, country, and year

    The file should be normalized so that Diésel + Fuel oil + Búnker = 1.0

    Returns:
        dict: {
            scenario: {
                country_iso3: {
                    year: {
                        'Diésel': share_value,
                        'Fuel oil': share_value,
                        'Búnker': share_value
                    }
                }
            }
        }
    """
    if not shares_file_path.exists():
        raise FileNotFoundError(f"Shares file not found: {shares_file_path}")

    wb = openpyxl.load_workbook(shares_file_path, data_only=True)

    shares_data = {}

    # Map sheet names to scenario codes
    sheet_scenario_map = {
        'SharesBAU': 'BAU',
        'SharesNDC': 'NDC',
        'SharesNDC_NoRPO': 'NDC_NoRPO',
        'SharesNDC+ELC': 'NDC+ELC'
    }

    # Map country names from Shares to ISO3 codes (from centralized config)
    shares_country_to_iso3 = get_shares_country_mapping()

    for sheet_name, scenario_code in sheet_scenario_map.items():
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        shares_data[scenario_code] = {}

        # Read years from row 1 (starting from column 2)
        years = []
        for col_idx in range(2, ws.max_column + 1):
            year_val = ws.cell(1, col_idx).value
            if year_val:
                try:
                    year = int(float(year_val))
                    years.append((col_idx, year))
                except:
                    pass

        # Process rows: each country has a block of rows
        current_country_iso3 = None
        row_idx = 2

        while row_idx <= ws.max_row:
            # First column: could be country name or fuel name
            cell_value = ws.cell(row_idx, 1).value

            if not cell_value:
                row_idx += 1
                continue

            cell_str = str(cell_value).strip()

            # Check if this is a country row
            if cell_str in shares_country_to_iso3:
                current_country_iso3 = shares_country_to_iso3[cell_str]
                shares_data[scenario_code][current_country_iso3] = {}
                row_idx += 1
                continue

            # If we have a current country, check for fuel types
            if current_country_iso3 and cell_str in ['Diésel', 'Fuel oil', 'Búnker']:
                fuel_name = cell_str

                # Read shares for all years
                for col_idx, year in years:
                    share_value = ws.cell(row_idx, col_idx).value

                    if year not in shares_data[scenario_code][current_country_iso3]:
                        shares_data[scenario_code][current_country_iso3][year] = {}

                    if share_value is not None:
                        try:
                            shares_data[scenario_code][current_country_iso3][year][fuel_name] = float(share_value)
                        except (ValueError, TypeError):
                            shares_data[scenario_code][current_country_iso3][year][fuel_name] = 0.0
                    else:
                        shares_data[scenario_code][current_country_iso3][year][fuel_name] = 0.0

            row_idx += 1

    wb.close()
    return shares_data


def read_olade_data(olade_file_path):
    """
    Read OLADE capacity data from Excel file

    Note: OLADE data is in MW, but is converted to GW for the model (1 GW = 1000 MW)

    Returns:
        dict: {
            'reference_year': int,
            'data': {
                country_iso3: {
                    tech_code: capacity_gw
                }
            }
        }
    """
    if not olade_file_path.exists():
        raise FileNotFoundError(f"OLADE file not found: {olade_file_path}")

    wb = openpyxl.load_workbook(olade_file_path, data_only=True)
    ws = wb['1.2023']

    # Extract reference year from A4
    ref_year_cell = ws['A4'].value
    ref_year = 2023  # Default
    if ref_year_cell and str(ref_year_cell).startswith('2023'):
        ref_year = 2023

    # Get country columns from row 5 (starting at column 3)
    country_columns = {}
    for col_idx in range(3, ws.max_column + 1):
        country_name = ws.cell(5, col_idx).value
        if country_name and str(country_name) in OLADE_COUNTRY_MAPPING:
            iso3_code = OLADE_COUNTRY_MAPPING[str(country_name)]
            country_columns[col_idx] = iso3_code

    # Read technology data (rows 6-20)
    data = {}

    # Process each technology
    for row_idx in range(6, 21):
        tech_name = ws.cell(row_idx, 1).value
        if not tech_name:
            continue

        tech_name_str = str(tech_name).strip()

        # Skip non-applicable technologies
        if tech_name_str in ['Térmica no renovable (combustión)', 'Otras fuentes',
                             'Térmica renovable (combustión)', 'Fuentes renovable (no combustión)',
                             'Biocombustibles líquidos', 'Total']:
            continue

        # Map OLADE tech name to model tech code
        tech_code = None
        if tech_name_str in OLADE_TECH_MAPPING:
            tech_code = OLADE_TECH_MAPPING[tech_name_str]

        # Special handling for BIO (sum of Biogás and Biomasa sólida)
        if tech_name_str == 'Biogás':
            tech_code = 'BIO'
        elif tech_name_str == 'Biomasa sólida':
            tech_code = 'BIO'

        # Special handling for Petróleo y derivados (will be split into PET and OIL later)
        elif tech_name_str == 'Petróleo y derivados':
            tech_code = 'PETROLEUM'  # Temporary code, will be split later

        if not tech_code:
            continue

        # Read capacity values for each country
        for col_idx, country_iso3 in country_columns.items():
            capacity = ws.cell(row_idx, col_idx).value

            if capacity is not None and capacity != '':
                try:
                    capacity_mw = float(capacity)

                    # Convert from MW to GW (1 GW = 1000 MW)
                    capacity_gw = capacity_mw / 1000.0

                    # Initialize country if not exists
                    if country_iso3 not in data:
                        data[country_iso3] = {}

                    # For BIO, sum Biogás + Biomasa sólida
                    if tech_code == 'BIO':
                        if tech_code in data[country_iso3]:
                            data[country_iso3][tech_code] += capacity_gw
                        else:
                            data[country_iso3][tech_code] = capacity_gw
                    else:
                        data[country_iso3][tech_code] = capacity_gw

                except ValueError:
                    pass

    wb.close()

    return {
        'reference_year': ref_year,
        'data': data
    }


def read_olade_generation_data(generation_file_path):
    """
    Read OLADE electricity generation data from Excel file

    Note: OLADE data is in GWh, converted to PJ for the model (1 GWh = 0.0036 PJ)

    Returns:
        dict: {
            'reference_year': int,
            'data': {
                country_iso3: generation_pj
            },
            'tech_shares': {
                country_iso3: {
                    tech_code: share_value
                }
            }
        }
    """
    if not generation_file_path.exists():
        raise FileNotFoundError(f"OLADE generation file not found: {generation_file_path}")

    wb = openpyxl.load_workbook(generation_file_path, data_only=True)
    ws = wb['1.2023']

    ref_year = 2023  # Reference year

    # Technology row mapping (OLADE row -> model tech code)
    # Row 6: Nuclear
    # Row 8: Petróleo y derivados (will be mapped to OIL+PET combined as 'PETROLEUM')
    # Row 9: Gas natural
    # Row 10: Carbón mineral
    # Row 13: Biogás
    # Row 14: Biomasa sólida
    # Row 15: Biocombustibles líquidos
    # Row 17: Hidro
    # Row 18: Geotermia
    # Row 19: Eólica
    # Row 20: Solar
    # Row 21: Total

    tech_row_mapping = {
        6: 'URN',      # Nuclear
        8: 'PETROLEUM', # Petróleo y derivados (combined PET+OIL)
        9: 'NGS',      # Gas natural
        10: 'COA',     # Carbón mineral
        13: 'BIO',     # Biogás (will be summed with biomass)
        14: 'BIO',     # Biomasa sólida
        15: 'BIO',     # Biocombustibles líquidos
        17: 'HYD',     # Hidro
        18: 'GEO',     # Geotermia
        19: 'WON',     # Eólica
        20: 'SPV',     # Solar
    }

    # Get country columns from row 5 (starting at column 3)
    country_columns = {}
    for col_idx in range(3, ws.max_column + 1):
        country_name = ws.cell(5, col_idx).value
        if country_name and str(country_name) in OLADE_COUNTRY_MAPPING:
            iso3_code = OLADE_COUNTRY_MAPPING[str(country_name)]
            country_columns[col_idx] = iso3_code

    # Read Total generation from row 21 and individual technologies
    data = {}
    tech_shares = {}

    for col_idx, country_iso3 in country_columns.items():
        total_gwh = ws.cell(21, col_idx).value  # Row 21 = "Total"

        if total_gwh is not None and total_gwh != '':
            try:
                generation_gwh = float(total_gwh)
                # Convert from GWh to PJ (1 GWh = 0.0036 PJ)
                generation_pj = generation_gwh * 0.0036
                data[country_iso3] = generation_pj

                # Calculate technology shares for this country
                tech_generation = {}
                for row_idx, tech_code in tech_row_mapping.items():
                    tech_gwh = ws.cell(row_idx, col_idx).value
                    if tech_gwh is not None and tech_gwh != '':
                        try:
                            tech_gwh_float = float(tech_gwh)
                            # Accumulate for technologies that are summed (e.g., BIO)
                            if tech_code in tech_generation:
                                tech_generation[tech_code] += tech_gwh_float
                            else:
                                tech_generation[tech_code] = tech_gwh_float
                        except (ValueError, TypeError):
                            pass

                # Calculate shares (percentage of total generation)
                tech_shares[country_iso3] = {}
                if generation_gwh > 0:
                    for tech_code, tech_gwh_value in tech_generation.items():
                        share = tech_gwh_value / generation_gwh
                        tech_shares[country_iso3][tech_code] = share

            except ValueError:
                pass

    wb.close()

    return {
        'reference_year': ref_year,
        'data': data,
        'tech_shares': tech_shares
    }


def read_trade_balance_data(trade_balance_file_path):
    """
    Read per-country electricity import/export data from the 'Imp-Exp por País' sheet.

    Uses accent-normalized country names to match against OLADE_COUNTRY_MAPPING.

    Args:
        trade_balance_file_path: Path to flujos_energia_estimados_optimizacion.xlsx

    Returns:
        dict: {
            country_iso3: {
                year: {
                    'importaciones_gwh': float,
                    'exportaciones_gwh': float
                }
            }
        }
    """
    if not trade_balance_file_path.exists():
        raise FileNotFoundError(f"Trade balance file not found: {trade_balance_file_path}")

    wb = openpyxl.load_workbook(trade_balance_file_path, data_only=True)

    # Find the 'Imp-Exp por País' sheet (accent-insensitive match)
    target_sheet = None
    for name in wb.sheetnames:
        if strip_accents(name) == 'Imp-Exp por Pais':
            target_sheet = name
            break

    if not target_sheet:
        wb.close()
        raise ValueError(
            f"Sheet 'Imp-Exp por País' not found in {trade_balance_file_path}. "
            f"Available sheets: {wb.sheetnames}"
        )

    ws = wb[target_sheet]

    # Parse header row to find column indices
    col_map = {}
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(1, col_idx).value
        if header:
            col_map[strip_accents(str(header).strip())] = col_idx

    year_col = col_map.get('Ano')
    country_col = col_map.get('Pais')
    exp_col = col_map.get('Exportaciones (GWh)')
    imp_col = col_map.get('Importaciones (GWh)')

    if not all([year_col, country_col, exp_col, imp_col]):
        wb.close()
        raise ValueError(
            f"Missing required columns in '{target_sheet}'. "
            f"Found headers: {list(col_map.keys())}. "
            f"Required: Año, País, Exportaciones (GWh), Importaciones (GWh)"
        )

    # Read data rows
    trade_data = {}

    for row_idx in range(2, ws.max_row + 1):
        year_val = ws.cell(row_idx, year_col).value
        country_val = ws.cell(row_idx, country_col).value
        exp_val = ws.cell(row_idx, exp_col).value
        imp_val = ws.cell(row_idx, imp_col).value

        if year_val is None or country_val is None:
            continue

        year = int(year_val)
        country_name_clean = strip_accents(str(country_val).strip())
        iso3 = OLADE_COUNTRY_MAPPING_NORMALIZED.get(country_name_clean)

        if not iso3:
            continue  # Country not in our mapping

        if iso3 not in trade_data:
            trade_data[iso3] = {}

        trade_data[iso3][year] = {
            'importaciones_gwh': float(imp_val) if imp_val is not None else 0.0,
            'exportaciones_gwh': float(exp_val) if exp_val is not None else 0.0
        }

    wb.close()
    return trade_data


def read_demand_data(demand_file_path):
    """
    Read projected electricity demand from A-O_Demand.xlsx

    The demand data comes from the Demand_Projection sheet with technology codes
    in format ELCxxxXX02 where xxx is the country code (e.g., ELCARGXX02 for Argentina).

    Args:
        demand_file_path: Path to A-O_Demand.xlsx file

    Returns:
        dict: {
            country_code: {
                year: demand_pj
            }
        }
    """
    if not demand_file_path.exists():
        raise FileNotFoundError(f"Demand file not found: {demand_file_path}")

    wb = openpyxl.load_workbook(demand_file_path, data_only=True)

    if 'Demand_Projection' not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Demand_Projection sheet not found in {demand_file_path}")

    ws = wb['Demand_Projection']

    # Build year column map from header row (row 1)
    year_col_map = {}
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(1, col_idx).value
        if header and str(header).isdigit():
            try:
                year = int(header)
                if 2000 <= year <= 2100:
                    year_col_map[year] = col_idx
            except (ValueError, TypeError):
                pass

    demand_data = {}

    # Read demand rows (starting at row 2)
    for row_idx in range(2, ws.max_row + 1):
        demand_type = ws.cell(row_idx, 1).value  # Column A: Demand/Share
        tech_code = ws.cell(row_idx, 2).value    # Column B: Fuel/Tech (e.g., ELCARGXX02)

        if not demand_type or str(demand_type).strip() != 'Demand':
            continue

        if not tech_code:
            continue

        tech_str = str(tech_code).strip().upper()

        # Extract country code from ELCxxxXX02 format (positions 3-6, 0-indexed)
        # Example: ELCARGXX02 -> ARG
        if not tech_str.startswith('ELC') or len(tech_str) < 6:
            continue

        country_code = tech_str[3:6]

        if country_code not in demand_data:
            demand_data[country_code] = {}

        # Read demand values for each year
        for year, col_idx in year_col_map.items():
            value = ws.cell(row_idx, col_idx).value
            if value is not None:
                try:
                    demand_pj = float(value)
                    demand_data[country_code][year] = demand_pj
                except (ValueError, TypeError):
                    pass

    wb.close()

    return demand_data


def read_shares_total_data(shares_total_path):
    """
    Read Shares_Power_Generation_Technologies.xlsx file to get technology shares by scenario, country, technology, and year

    The file has a structure where countries are followed by their technology rows.
    Technologies in Shares_Power_Generation_Technologies are mapped to model tech codes:
    - Biomasa → BIO
    - Búnker + Fuel oil → OIL
    - Carbón → COA
    - Diésel → PET
    - Eólica → WON
    - Gas natural → NGS (not CCG, since NGS is the generic natural gas code)
    - Geotérmica → GEO
    - Hidroeléctrica → HYD
    - Nuclear → URN
    - Solar (GD) + Solar (gran escala) → SPV

    Returns:
        dict: {
            scenario: {
                country_iso3: {
                    tech_code: {
                        year: share_value
                    }
                }
            }
        }
    """
    if not shares_total_path.exists():
        raise FileNotFoundError(f"Shares_Total file not found: {shares_total_path}")

    wb = openpyxl.load_workbook(shares_total_path, data_only=True)

    shares_data = {}

    # Map sheet names to scenario codes
    sheet_scenario_map = {
        'SharesBAU': 'BAU',
        'SharesNDC': 'NDC',
        'SharesNDC_NoRPO': 'NDC_NoRPO',
        'SharesNDC+ELC': 'NDC+ELC'
    }

    # Map country and technology names from Shares_Total (from centralized config)
    shares_country_to_iso3 = get_shares_country_mapping()
    shares_tech_to_code = get_shares_tech_mapping()

    for sheet_name, scenario_code in sheet_scenario_map.items():
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        shares_data[scenario_code] = {}

        # Read years from row 1 (starting from column 2)
        years = []
        for col_idx in range(2, ws.max_column + 1):
            year_val = ws.cell(1, col_idx).value
            if year_val:
                try:
                    year = int(float(year_val))
                    years.append((col_idx, year))
                except:
                    pass

        # Process rows: each country has a block of rows
        current_country_iso3 = None
        row_idx = 2

        while row_idx <= ws.max_row:
            cell_value = ws.cell(row_idx, 1).value

            if not cell_value:
                row_idx += 1
                continue

            cell_str = str(cell_value).strip()

            # Check if this is a country row
            if cell_str in shares_country_to_iso3:
                current_country_iso3 = shares_country_to_iso3[cell_str]
                shares_data[scenario_code][current_country_iso3] = {}
                row_idx += 1
                continue

            # If we have a current country, check for technology rows
            if current_country_iso3 and cell_str in shares_tech_to_code:
                tech_code = shares_tech_to_code[cell_str]

                # Initialize tech_code dict if not exists
                if tech_code not in shares_data[scenario_code][current_country_iso3]:
                    shares_data[scenario_code][current_country_iso3][tech_code] = {}

                # Read shares for all years
                for col_idx, year in years:
                    share_value = ws.cell(row_idx, col_idx).value

                    if share_value is not None:
                        try:
                            share_float = float(share_value)
                            # For combined techs (OIL, SPV), sum the values
                            if year in shares_data[scenario_code][current_country_iso3][tech_code]:
                                shares_data[scenario_code][current_country_iso3][tech_code][year] += share_float
                            else:
                                shares_data[scenario_code][current_country_iso3][tech_code][year] = share_float
                        except (ValueError, TypeError):
                            if year not in shares_data[scenario_code][current_country_iso3][tech_code]:
                                shares_data[scenario_code][current_country_iso3][tech_code][year] = 0.0
                    else:
                        if year not in shares_data[scenario_code][current_country_iso3][tech_code]:
                            shares_data[scenario_code][current_country_iso3][tech_code][year] = 0.0

            row_idx += 1

    wb.close()
    return shares_data


class SecondaryTechsUpdater:
    def __init__(self, editor_path, base_path, olade_file_path=None, shares_file_path=None, generation_file_path=None, shares_total_file_path=None, trade_balance_file_path=None):
        self.editor_path = editor_path
        self.base_path = base_path
        self.olade_file_path = olade_file_path
        self.shares_file_path = shares_file_path
        self.generation_file_path = generation_file_path
        self.shares_total_file_path = shares_total_file_path
        self.trade_balance_file_path = trade_balance_file_path
        self.scenarios = ["BAU", "NDC", "NDC+ELC", "NDC_NoRPO"]
        self.log_lines = []
        self.changes_applied = 0
        self.rows_failed = 0
        self.olade_config = None
        self.olade_data = None
        self.shares_data = None
        self.generation_data = None
        self.shares_total_data = None
        self.trade_balance_data = None
        # DEBUG: Cache for DemandBased normalized shares
        self.demand_based_shares_cache = {}  # {(scenario, country): {fuel: {year: share}}}

    def log(self, message, level="INFO"):
        """Add message to log"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_line = f"[{timestamp}] {level}: {message}"
        self.log_lines.append(log_line)
        print(log_line)

    def read_editor_file(self):
        """
        Read and parse the editor Excel file

        Returns:
            list of dicts with editing instructions
        """
        self.log("Reading editor file...")

        if not self.editor_path.exists():
            raise FileNotFoundError(f"Editor file not found: {self.editor_path}")

        wb = openpyxl.load_workbook(self.editor_path, data_only=True)

        if 'Editor' not in wb.sheetnames:
            raise ValueError("'Editor' sheet not found in template file")

        ws = wb['Editor']

        # Read header to get year columns
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value))
            else:
                break

        # Identify year columns (after Scenario, Country, Tech.Name, Tech, Parameter)
        # Years start from column 6 (F)
        year_columns = []
        for idx, header in enumerate(headers[5:], 6):  # Start from column 6 (F)
            if header.isdigit():
                year_columns.append((idx, int(header)))

        self.log(f"Found {len(year_columns)} year columns: {year_columns[0][1]} to {year_columns[-1][1]}")

        # Read data rows
        # Columns: 1=Scenario, 2=Country, 3=Tech.Name, 4=Tech, 5=Parameter
        edit_instructions = []
        for row_idx in range(2, ws.max_row + 1):
            scenario = ws.cell(row_idx, 1).value
            country = ws.cell(row_idx, 2).value
            tech_name = ws.cell(row_idx, 3).value
            tech = ws.cell(row_idx, 4).value  # This is the Tech code (auto-filled from Tech.Name)
            parameter = ws.cell(row_idx, 5).value

            # Skip empty rows
            if not scenario and not country and not tech and not parameter:
                continue

            # Read year values
            year_values = {}
            for col_idx, year in year_columns:
                value = ws.cell(row_idx, col_idx).value
                if value is not None and value != "":
                    year_values[year] = value

            # Only add if we have at least one year value
            if year_values:
                edit_instructions.append({
                    'row': row_idx,
                    'scenario': str(scenario).strip() if scenario else None,
                    'country': str(country).strip() if country else None,
                    'tech_name': str(tech_name).strip() if tech_name else None,  # Keep for logging
                    'tech': str(tech).strip() if tech else None,  # This is what we'll use for matching
                    'parameter': str(parameter).strip() if parameter else None,
                    'year_values': year_values
                })

        wb.close()

        self.log(f"Found {len(edit_instructions)} rows with data to process")
        return edit_instructions

    def validate_instruction(self, instruction):
        """
        Validate an edit instruction

        Returns:
            (is_valid, error_message)
        """
        if not instruction['scenario']:
            return False, "Scenario is empty"

        if not instruction['country']:
            return False, "Country is empty"

        if not instruction['tech']:
            return False, "Tech is empty"

        if not instruction['parameter']:
            return False, "Parameter is empty"

        if not instruction['year_values']:
            return False, "No year values provided"

        # Validate scenario
        valid_scenarios = self.scenarios + ['ALL']
        if instruction['scenario'] not in valid_scenarios:
            return False, f"Invalid scenario '{instruction['scenario']}'. Must be one of: {valid_scenarios}"

        # Skip tech validation for OLADE instructions (they use 3-char codes)
        if instruction.get('is_olade'):
            return True, None

        # Validate that tech contains country code (for PWR technologies: PWRTRNARGXX -> ARG is at position 6-8)
        tech = instruction['tech'].upper()
        country = instruction['country'].upper()

        if tech.startswith('PWR'):
            # For PWR technologies, country code is at positions 6-8
            if len(tech) >= 9:
                tech_country = tech[6:9]
                if tech_country != country:
                    return False, f"Tech '{instruction['tech']}' contains country code '{tech_country}', but '{country}' was specified"
            else:
                return False, f"Tech '{instruction['tech']}' has invalid format (too short for PWR technology)"
        else:
            # For non-PWR technologies, country code might be at the beginning or elsewhere
            # We'll just issue a warning but allow it
            pass

        return True, None

    def create_backup(self, file_path):
        """Create backup of file before modifying"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = file_path.with_name(f"{file_path.stem}_backup_{timestamp}{file_path.suffix}")
        shutil.copy2(file_path, backup_path)
        return backup_path

    def find_and_update_row(self, ws, tech, parameter, year_values, year_col_map, projection_mode_col, is_olade=False, country=None):
        """
        Find the row matching tech and parameter, and update year values

        Args:
            is_olade: If True, tech is a 3-char code and we need to match PWR technologies by first 9 chars
            country: Country code (needed for OLADE matching)

        Returns:
            (success, message, values_updated)
        """
        # Search for matching row
        # Columns in Secondary Techs: 1=Tech.Id, 2=Tech, 3=Tech.Name, 5=Parameter
        target_rows = []
        max_capacity_rows = {}  # {full_tech_code: row_idx} for TotalAnnualMaxCapacity

        for row_idx in range(2, ws.max_row + 1):
            row_tech = ws.cell(row_idx, 2).value  # Column 2: Tech
            row_param = ws.cell(row_idx, 5).value  # Column 5: Parameter

            if not row_tech or not row_param:
                continue

            row_tech_str = str(row_tech).strip()
            row_param_str = str(row_param).strip()

            # Track TotalAnnualMaxCapacity rows for constraint checking
            if row_param_str == 'TotalAnnualMaxCapacity':
                max_capacity_rows[row_tech_str.upper()] = row_idx

            # Check parameter match
            if row_param_str != parameter:
                continue

            # Check tech match
            if is_olade:
                # For OLADE: match PWR technologies by first 9 chars (PWR + 3 chars + country code)
                # Example: Looking for tech_code='URN' country='ARG' should match 'PWRURNARGXX'
                if row_tech_str.upper().startswith('PWR') and len(row_tech_str) >= 9:
                    # Extract positions 4-6 (tech type) and 7-9 (country)
                    row_tech_type = row_tech_str[3:6].upper()
                    row_country = row_tech_str[6:9].upper()

                    if row_tech_type == tech.upper() and row_country == country.upper():
                        target_rows.append((row_idx, row_tech_str.upper()))
            else:
                # Exact match for manual instructions
                if row_tech_str == tech:
                    target_rows.append((row_idx, row_tech_str.upper()))

        if not target_rows:
            tech_desc = f"Tech type='{tech}' Country='{country}'" if is_olade else f"Tech='{tech}'"
            return False, f"No matching row found for {tech_desc} and Parameter='{parameter}'", 0

        # Update year values for all matching rows
        total_values_updated = 0
        rows_updated = []
        capped_count = 0

        for target_row, full_tech_code in target_rows:
            values_updated = 0

            # For ResidualCapacity, check against TotalAnnualMaxCapacity constraint
            max_cap_row = max_capacity_rows.get(full_tech_code) if parameter == 'ResidualCapacity' else None

            for year, value in year_values.items():
                if year in year_col_map:
                    col_idx = year_col_map[year]
                    final_value = value

                    # Check ResidualCapacity <= TotalAnnualMaxCapacity constraint
                    if max_cap_row and parameter == 'ResidualCapacity':
                        max_cap_value = ws.cell(max_cap_row, col_idx).value
                        if max_cap_value is not None:
                            try:
                                max_cap = float(max_cap_value)
                                if value > max_cap:
                                    # Cap ResidualCapacity to TotalAnnualMaxCapacity
                                    final_value = max_cap
                                    capped_count += 1
                            except (ValueError, TypeError):
                                pass

                    ws.cell(target_row, col_idx, final_value)
                    values_updated += 1

            # Set Projection.Mode to "User defined" if column exists
            if projection_mode_col:
                current_value = ws.cell(target_row, projection_mode_col).value
                if current_value != "User defined":
                    ws.cell(target_row, projection_mode_col, "User defined")

            total_values_updated += values_updated
            rows_updated.append(target_row)

        capped_msg = f" ({capped_count} capped to MaxCapacity)" if capped_count > 0 else ""
        if len(rows_updated) > 1:
            return True, f"Rows {rows_updated} updated with {total_values_updated} total year values{capped_msg}", total_values_updated
        else:
            return True, f"Row {rows_updated[0]} updated with {total_values_updated} year values{capped_msg}", total_values_updated

    def apply_instruction_to_scenario(self, instruction, scenario, ws, year_col_map, projection_mode_col):
        """
        Apply a single edit instruction to an already-open worksheet

        Args:
            instruction: instruction dict
            scenario: scenario name
            ws: openpyxl worksheet (already open)
            year_col_map: dict mapping years to column indices
            projection_mode_col: column index for Projection.Mode

        Returns:
            (success, message)
        """
        try:
            # Apply update
            is_olade = instruction.get('is_olade', False)
            country = instruction.get('country')

            success, message, values_updated = self.find_and_update_row(
                ws,
                instruction['tech'],
                instruction['parameter'],
                instruction['year_values'],
                year_col_map,
                projection_mode_col,
                is_olade=is_olade,
                country=country
            )

            if success:
                self.changes_applied += 1
                return True, f"{scenario}: {message}"
            else:
                return False, f"{scenario}: {message}"

        except Exception as e:
            return False, f"{scenario}: Error - {str(e)}"

    def apply_instructions_batch(self, instructions):
        """
        Apply all instructions grouped by scenario (batch processing)
        This opens each workbook only once, creates one backup, and saves only once.
        """
        # Group instructions by scenario
        from collections import defaultdict
        scenario_instructions = defaultdict(list)

        for instruction in instructions:
            # Validate first
            is_valid, error_msg = self.validate_instruction(instruction)
            if not is_valid:
                row_num = instruction['row']
                self.log(f"\nRow {row_num} FAILED: {error_msg}", "ERROR")
                self.rows_failed += 1
                continue

            # Determine target scenarios
            if instruction['scenario'] == 'ALL':
                target_scenarios = self.scenarios
            else:
                target_scenarios = [instruction['scenario']]

            # Add to each target scenario's list
            for scenario in target_scenarios:
                scenario_instructions[scenario].append(instruction)

        # Process each scenario
        for scenario in self.scenarios:
            if scenario not in scenario_instructions:
                self.log(f"\nScenario {scenario}: No instructions to apply")
                continue

            instructions_for_scenario = scenario_instructions[scenario]
            self.log(f"\nProcessing scenario {scenario}: {len(instructions_for_scenario)} instructions")

            scenario_path = self.base_path / f"A1_Outputs_{scenario}" / "A-O_Parametrization.xlsx"

            if not scenario_path.exists():
                self.log(f"  ✗ File not found: {scenario_path}", "ERROR")
                self.rows_failed += len(instructions_for_scenario)
                continue

            # Create backup ONCE for this scenario
            backup_path = self.create_backup(scenario_path)
            self.log(f"  Backup created: {backup_path.name}")

            try:
                # Open workbook ONCE
                wb = openpyxl.load_workbook(scenario_path)

                if 'Secondary Techs' not in wb.sheetnames:
                    wb.close()
                    self.log(f"  ✗ 'Secondary Techs' sheet not found", "ERROR")
                    self.rows_failed += len(instructions_for_scenario)
                    continue

                ws = wb['Secondary Techs']

                # Store initial row count for validation
                initial_row_count = ws.max_row

                # Build year column map and find Projection.Mode column ONCE
                headers = [cell.value for cell in ws[1]]
                year_col_map = {}
                projection_mode_col = None

                for col_idx, header in enumerate(headers, 1):
                    if header:
                        if str(header).isdigit():
                            try:
                                year = int(header)
                                if 2000 <= year <= 2100:
                                    year_col_map[year] = col_idx
                            except:
                                pass
                        elif str(header).strip() == "Projection.Mode":
                            projection_mode_col = col_idx

                # Apply all instructions for this scenario
                for instruction in instructions_for_scenario:
                    row_num = instruction['row']
                    country = instruction.get('country', 'N/A')
                    self.log(f"  Row {row_num} [{country}]: {instruction['tech']} - {instruction['parameter']}")

                    success, message = self.apply_instruction_to_scenario(
                        instruction, scenario, ws, year_col_map, projection_mode_col
                    )

                    if success:
                        self.log(f"    ✓ {message}", "SUCCESS")
                    else:
                        self.log(f"    ✗ {message}", "ERROR")
                        self.rows_failed += 1

                # Validate row count before saving
                final_row_count = ws.max_row
                if final_row_count != initial_row_count:
                    self.log(f"  ⚠ WARNING: Row count changed from {initial_row_count} to {final_row_count}!", "WARNING")

                # Save ONCE after all instructions
                self.log(f"  Saving {scenario}...")
                wb.save(scenario_path)
                wb.close()
                self.log(f"  ✓ {scenario} saved successfully (rows: {final_row_count})")

            except Exception as e:
                self.log(f"  ✗ Error processing {scenario}: {e}", "ERROR")
                self.rows_failed += len(instructions_for_scenario)
                try:
                    wb.close()
                except:
                    pass

    def generate_olade_instructions(self, all_years):
        """
        Generate instructions from OLADE data

        Uses flat capacity values (same for all years, no growth applied).

        For PETROLEUM:
        - PET = Petroleum × Diésel share
        - OIL = Petroleum × (Fuel oil + Búnker) share

        Args:
            all_years: list of years from Secondary Techs

        Returns:
            list of instruction dicts
        """
        if not self.olade_config['enabled']:
            return []

        self.log("")
        self.log("=" * 80)
        self.log("PROCESSING OLADE DATA")
        self.log("=" * 80)
        self.log(f"Reference year: {self.olade_data['reference_year']}")
        self.log(f"Petroleum split mode: {self.olade_config.get('petroleum_split_mode', 'Split_PET_OIL')}")
        self.log("Using FLAT capacity values (same for all years)")
        self.log("")

        instructions = []

        # Generate instructions for each country and technology
        for country_iso3, techs in self.olade_data['data'].items():
            self.log(f"Processing country: {country_iso3}")
            for tech_code, base_capacity in techs.items():

                # Special handling for PETROLEUM
                if tech_code == 'PETROLEUM':
                    # Check split mode
                    split_mode = self.olade_config.get('petroleum_split_mode', 'Split_PET_OIL')

                    if split_mode == 'OIL_only':
                        # Option 1: Assign all petroleum to OIL only
                        for scenario in self.scenarios:
                            # Flat value for all years (6 decimals to preserve small values)
                            oil_year_values = {year: round(base_capacity, 6) for year in all_years}

                            instruction_oil = {
                                'row': 'OLADE',
                                'scenario': scenario,
                                'country': country_iso3,
                                'tech_name': 'PWR-OIL',
                                'tech': 'OIL',
                                'parameter': 'ResidualCapacity',
                                'year_values': oil_year_values,
                                'is_olade': True
                            }
                            instructions.append(instruction_oil)

                    else:
                        # Option 2: Split petroleum into PET and OIL using shares
                        # PET = Petroleum × Diésel
                        # OIL = Petroleum × (Fuel oil + Búnker)
                        for scenario in self.scenarios:
                            pet_year_values = {}
                            oil_year_values = {}

                            for year in all_years:
                                # Get shares from shares_data
                                diesel_share = 0.0
                                fuel_oil_share = 0.0
                                bunker_share = 0.0

                                if (self.shares_data and
                                    scenario in self.shares_data and
                                    country_iso3 in self.shares_data[scenario] and
                                    year in self.shares_data[scenario][country_iso3]):

                                    year_shares = self.shares_data[scenario][country_iso3][year]
                                    diesel_share = year_shares.get('Diésel', 0.0)
                                    fuel_oil_share = year_shares.get('Fuel oil', 0.0)
                                    bunker_share = year_shares.get('Búnker', 0.0)

                                # Shares should already be normalized (sum to 1.0)
                                # PET = Petroleum × Diésel
                                # OIL = Petroleum × (Fuel oil + Búnker)
                                pet_capacity = base_capacity * diesel_share
                                oil_capacity = base_capacity * (fuel_oil_share + bunker_share)

                                # Use 6 decimals to preserve small capacity values
                                pet_year_values[year] = round(pet_capacity, 6)
                                oil_year_values[year] = round(oil_capacity, 6)

                            # Create instruction for PET (Diésel)
                            instruction_pet = {
                                'row': 'OLADE',
                                'scenario': scenario,
                                'country': country_iso3,
                                'tech_name': 'PWR-PET',
                                'tech': 'PET',
                                'parameter': 'ResidualCapacity',
                                'year_values': pet_year_values,
                                'is_olade': True
                            }
                            instructions.append(instruction_pet)

                            # Create instruction for OIL (Fuel oil + Búnker)
                            instruction_oil = {
                                'row': 'OLADE',
                                'scenario': scenario,
                                'country': country_iso3,
                                'tech_name': 'PWR-OIL',
                                'tech': 'OIL',
                                'parameter': 'ResidualCapacity',
                                'year_values': oil_year_values,
                                'is_olade': True
                            }
                            instructions.append(instruction_oil)

                else:
                    # Normal handling for other technologies
                    # Flat capacity value for all years (6 decimals to preserve small values)
                    year_values = {year: round(base_capacity, 6) for year in all_years}

                    # Create instruction for each scenario
                    for scenario in self.scenarios:
                        instruction = {
                            'row': 'OLADE',
                            'scenario': scenario,
                            'country': country_iso3,
                            'tech_name': f'PWR-{tech_code}',
                            'tech': tech_code,  # This will be used to match PWR technologies
                            'parameter': 'ResidualCapacity',
                            'year_values': year_values,
                            'is_olade': True
                        }
                        instructions.append(instruction)

        self.log(f"Generated {len(instructions)} OLADE instructions")
        self.log("")

        return instructions

    def save_log(self, log_path):
        """Save log to file"""
        with open(log_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(self.log_lines))

    def export_shares_to_csv(self, all_shares_data, all_years, output_path):
        """
        Export technology shares to CSV for debugging.

        Format:
        - Column 1: Country
        - Column 2: Scenario
        - Column 3: Technology
        - Columns 4+: Years (time series)

        Args:
            all_shares_data: dict {(scenario, country): {tech_type: {year: share}}}
            all_years: list of years
            output_path: Path to output CSV file
        """
        import csv

        self.log(f"  Starting CSV export...")
        self.log(f"  Output path: {output_path}")
        self.log(f"  Scenario/country combinations: {len(all_shares_data)}")
        self.log(f"  Years to export: {len(all_years)}")

        try:
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)

                # Write header
                header = ['Country', 'Scenario', 'Technology'] + [str(year) for year in sorted(all_years)]
                writer.writerow(header)
                self.log(f"  ✓ Header written with {len(header)} columns")

                # Sort by scenario and country for consistent output
                sorted_items = sorted(all_shares_data.items(), key=lambda x: (x[0][0], x[0][1]))

                # Write data rows
                rows_written = 0
                for (scenario, country), tech_shares in sorted_items:
                    # Sort technologies for consistent output
                    sorted_techs = sorted(tech_shares.keys())

                    for tech_type in sorted_techs:
                        year_shares = tech_shares[tech_type]

                        # Build row: country, scenario, tech, then year values
                        row = [country, scenario, tech_type]

                        for year in sorted(all_years):
                            share_value = year_shares.get(year, 0.0)
                            # Format with 6 decimals
                            row.append(f"{share_value:.6f}")

                        writer.writerow(row)
                        rows_written += 1

                self.log(f"  ✓ Data written: {rows_written} rows")

            self.log(f"✓ CSV export completed successfully!")
            self.log(f"  File: {output_path}")

        except Exception as e:
            self.log(f"✗ ERROR exporting CSV: {e}", "ERROR")
            import traceback
            traceback.print_exc()

    def update_demand_files(self, all_years):
        """
        Update A-O_Demand.xlsx files with electricity demand from OLADE generation data

        Applies linear growth: Demand(year) = Demand(2023) × (1 + rate × (year - 2023))

        Args:
            all_years: list of years to populate
        """
        if not self.olade_config.get('demand_enabled') or not self.generation_data:
            return

        self.log("")
        self.log("=" * 80)
        self.log("UPDATING ELECTRICITY DEMAND")
        self.log("=" * 80)

        ref_year = self.generation_data['reference_year']
        growth_rates = self.olade_config.get('demand_growth_rates', {})

        self.log(f"Reference year: {ref_year}")
        self.log(f"Growth type: Linear")
        self.log("")

        demand_changes = 0

        for scenario in self.scenarios:
            demand_path = self.base_path / f"A1_Outputs_{scenario}" / "A-O_Demand.xlsx"

            if not demand_path.exists():
                self.log(f"  ✗ Demand file not found: {demand_path}", "WARNING")
                continue

            self.log(f"Processing {scenario}...")

            # Create backup
            backup_path = self.create_backup(demand_path)
            self.log(f"  Backup created: {backup_path.name}")

            try:
                wb = openpyxl.load_workbook(demand_path)

                if 'Demand_Projection' not in wb.sheetnames:
                    wb.close()
                    self.log(f"  ✗ 'Demand_Projection' sheet not found", "ERROR")
                    continue

                ws = wb['Demand_Projection']

                # Build year column map from headers
                year_col_map = {}
                for col_idx in range(1, ws.max_column + 1):
                    header = ws.cell(1, col_idx).value
                    if header and str(header).isdigit():
                        try:
                            year = int(header)
                            if 2000 <= year <= 2100:
                                year_col_map[year] = col_idx
                        except:
                            pass

                # Find existing ELC*XX02 country codes and their row indices
                # Format: ELC + country(3) + XX02 = 10 characters (e.g., ELCARGXX02)
                existing_countries = {}  # {country_code: row_idx}
                for row_idx in range(2, ws.max_row + 1):
                    fuel_code = ws.cell(row_idx, 2).value
                    if fuel_code:
                        fuel_str = str(fuel_code).strip().upper()
                        if fuel_str.startswith('ELC') and fuel_str.endswith('XX02') and len(fuel_str) == 10:
                            country_code = fuel_str[3:6]
                            existing_countries[country_code] = row_idx

                # Update existing ELC*XX02 rows with generation data
                for country_code, row_idx in existing_countries.items():
                    # Check if we have generation data for this country
                    if country_code not in self.generation_data['data']:
                        continue

                    base_demand_pj = self.generation_data['data'][country_code]
                    growth_rate = growth_rates.get(country_code, 0.02)  # Default 2%

                    self.log(f"  {country_code}: Base={base_demand_pj:.2f} PJ, Growth={growth_rate*100:.1f}%")

                    # Update each year with linear growth
                    for year in all_years:
                        if year in year_col_map:
                            # Linear growth: Demand(year) = Demand(ref_year) × (1 + rate × (year - ref_year))
                            years_diff = year - ref_year
                            demand_year = base_demand_pj * (1 + growth_rate * years_diff)

                            # Apply scenario-specific demand adjustment if defined
                            # Formula: Demand(scenario, year) = Base_Demand(year) × (1 + adjustment_percentage)
                            adjustments_dict = self.olade_config.get('scenarios_demand_adjustments', {})
                            adjustment_key = (country_code, scenario)
                            if adjustment_key in adjustments_dict:
                                year_adjustments = adjustments_dict[adjustment_key]
                                if year in year_adjustments:
                                    adjustment_pct = year_adjustments[year]
                                    demand_year = demand_year * (1 + adjustment_pct)

                            demand_year = round(demand_year, 2)

                            ws.cell(row_idx, year_col_map[year], demand_year)
                            demand_changes += 1

                # Add missing countries that have generation data but aren't in the sheet
                missing_countries = set(self.generation_data['data'].keys()) - set(existing_countries.keys())
                if missing_countries:
                    self.log(f"  Adding missing countries: {', '.join(sorted(missing_countries))}")

                    # Get column positions for fixed columns
                    # Headers: Demand/Share, Fuel/Tech, Name, Ref.Cap.BY, Ref.OAR.BY, Ref.km.BY, Projection.Mode, Projection.Parameter, [years...]
                    fixed_col_map = {}
                    for col_idx in range(1, ws.max_column + 1):
                        header = ws.cell(1, col_idx).value
                        if header:
                            fixed_col_map[str(header)] = col_idx

                    # Add a row for each missing country
                    for country_code in sorted(missing_countries):
                        base_demand_pj = self.generation_data['data'][country_code]
                        growth_rate = growth_rates.get(country_code, 0.02)

                        # Find country name from OLADE mapping
                        country_name = None
                        for name, code in OLADE_COUNTRY_MAPPING.items():
                            if code == country_code:
                                country_name = name
                                break

                        # Create new row
                        new_row = ws.max_row + 1
                        fuel_code = f"ELC{country_code}XX02"

                        ws.cell(new_row, fixed_col_map.get('Demand/Share', 1), 'Demand')
                        ws.cell(new_row, fixed_col_map.get('Fuel/Tech', 2), fuel_code)
                        ws.cell(new_row, fixed_col_map.get('Name', 3), f"Output demand of transmission lines in {country_name or country_code}")
                        ws.cell(new_row, fixed_col_map.get('Ref.Cap.BY', 4), 'not needed')
                        ws.cell(new_row, fixed_col_map.get('Ref.OAR.BY', 5), 'not needed')
                        ws.cell(new_row, fixed_col_map.get('Ref.km.BY', 6), 'not needed')
                        ws.cell(new_row, fixed_col_map.get('Projection.Mode', 7), 'User defined')
                        ws.cell(new_row, fixed_col_map.get('Projection.Parameter', 8), 0)

                        self.log(f"  + {country_code}: Base={base_demand_pj:.2f} PJ, Growth={growth_rate*100:.1f}%")

                        # Add year values
                        for year in all_years:
                            if year in year_col_map:
                                years_diff = year - ref_year
                                demand_year = base_demand_pj * (1 + growth_rate * years_diff)

                                # Apply scenario-specific demand adjustment if defined
                                adjustments_dict = self.olade_config.get('scenarios_demand_adjustments', {})
                                adjustment_key = (country_code, scenario)
                                if adjustment_key in adjustments_dict:
                                    year_adjustments = adjustments_dict[adjustment_key]
                                    if year in year_adjustments:
                                        adjustment_pct = year_adjustments[year]
                                        demand_year = demand_year * (1 + adjustment_pct)

                                demand_year = round(demand_year, 2)

                                ws.cell(new_row, year_col_map[year], demand_year)
                                demand_changes += 1

                # Save
                wb.save(demand_path)
                wb.close()
                self.log(f"  ✓ {scenario} demand updated")

            except Exception as e:
                self.log(f"  ✗ Error updating {scenario}: {e}", "ERROR")
                try:
                    wb.close()
                except:
                    pass

        self.log("")
        self.log(f"Demand updates completed: {demand_changes} values written")

    def _get_trade_for_year(self, country_iso3, year):
        """
        Get trade balance data for a given country and year.

        For years beyond the available data, uses the last available year's values.
        For years before the earliest data, uses the earliest year's values.

        Returns:
            tuple: (importaciones_gwh, exportaciones_gwh)
        """
        country_data = self.trade_balance_data.get(country_iso3, {})
        if not country_data:
            return 0.0, 0.0

        if year in country_data:
            return (country_data[year]['importaciones_gwh'],
                    country_data[year]['exportaciones_gwh'])

        # Use last available year for future years
        max_available_year = max(country_data.keys())
        if year > max_available_year:
            return (country_data[max_available_year]['importaciones_gwh'],
                    country_data[max_available_year]['exportaciones_gwh'])

        # Use earliest available year for past years
        min_available_year = min(country_data.keys())
        return (country_data[min_available_year]['importaciones_gwh'],
                country_data[min_available_year]['exportaciones_gwh'])

    def update_demand_with_trade_balance(self, all_years):
        """
        Adjust electricity demand in A-O_Demand.xlsx by trade balance (imports/exports).

        Formula: Adjusted_Demand = Current_Demand - Exports_PJ + Imports_PJ
        Where Imports/Exports are converted from GWh to PJ (1 GWh = 0.0036 PJ).

        For years beyond available trade data, uses last available year as constant.
        This method runs LAST to avoid propagating changes to activity limits.

        Args:
            all_years: list of years to process
        """
        if not self.olade_config.get('trade_balance_enabled') or not self.trade_balance_data:
            return

        GWH_TO_PJ = 0.0036

        self.log("")
        self.log("=" * 80)
        self.log("ADJUSTING DEMAND BY TRADE BALANCE (Imports/Exports)")
        self.log("=" * 80)

        # Log available years in trade data
        all_trade_years = set()
        for country_data in self.trade_balance_data.values():
            all_trade_years.update(country_data.keys())
        max_trade_year = max(all_trade_years) if all_trade_years else None
        self.log(f"Trade data years: {sorted(all_trade_years)}")
        if max_trade_year:
            self.log(f"Years beyond {max_trade_year} will use {max_trade_year} values as constant")
        self.log(f"Conversion: 1 GWh = {GWH_TO_PJ} PJ")
        self.log(f"Formula: New_Demand = Current_Demand - Exports_PJ + Imports_PJ")
        self.log("")

        trade_balance_changes = 0

        for scenario in self.scenarios:
            demand_path = self.base_path / f"A1_Outputs_{scenario}" / "A-O_Demand.xlsx"

            if not demand_path.exists():
                self.log(f"  ✗ Demand file not found: {demand_path}", "WARNING")
                continue

            self.log(f"Processing {scenario}...")

            # Create backup
            backup_path = self.create_backup(demand_path)
            self.log(f"  Backup: {backup_path.name}")

            try:
                wb = openpyxl.load_workbook(demand_path)

                if 'Demand_Projection' not in wb.sheetnames:
                    wb.close()
                    self.log(f"  ✗ 'Demand_Projection' sheet not found", "ERROR")
                    continue

                ws = wb['Demand_Projection']

                # Build year column map from headers
                year_col_map = {}
                for col_idx in range(1, ws.max_column + 1):
                    header = ws.cell(1, col_idx).value
                    if header and str(header).isdigit():
                        try:
                            year = int(header)
                            if 2000 <= year <= 2100:
                                year_col_map[year] = col_idx
                        except (ValueError, TypeError):
                            pass

                # Find ELC*XX02 rows and apply adjustments
                for row_idx in range(2, ws.max_row + 1):
                    fuel_code = ws.cell(row_idx, 2).value
                    if not fuel_code:
                        continue

                    fuel_str = str(fuel_code).strip().upper()
                    if not (fuel_str.startswith('ELC') and fuel_str.endswith('XX02') and len(fuel_str) == 10):
                        continue

                    country_code = fuel_str[3:6]  # e.g., 'ARG' from 'ELCARGXX02'

                    if country_code not in self.trade_balance_data:
                        continue  # No trade data for this country

                    first_year_logged = False

                    for year in all_years:
                        if year not in year_col_map:
                            continue

                        current_value = ws.cell(row_idx, year_col_map[year]).value
                        if current_value is None:
                            continue

                        try:
                            current_demand_pj = float(current_value)
                        except (ValueError, TypeError):
                            continue

                        imports_gwh, exports_gwh = self._get_trade_for_year(country_code, year)

                        # Adjust: New_Demand = Current - Exports + Imports (all in PJ)
                        imports_pj = imports_gwh * GWH_TO_PJ
                        exports_pj = exports_gwh * GWH_TO_PJ
                        new_demand_pj = round(current_demand_pj - exports_pj + imports_pj, 2)

                        ws.cell(row_idx, year_col_map[year], new_demand_pj)
                        trade_balance_changes += 1

                        if not first_year_logged:
                            self.log(f"  {country_code}: Imp={imports_gwh:.1f} GWh ({imports_pj:.4f} PJ), "
                                     f"Exp={exports_gwh:.1f} GWh ({exports_pj:.4f} PJ), "
                                     f"Demand {current_demand_pj:.2f} -> {new_demand_pj:.2f} PJ (year {year})")
                            first_year_logged = True

                wb.save(demand_path)
                wb.close()
                self.log(f"  ✓ {scenario} trade balance adjustment applied")

            except Exception as e:
                self.log(f"  ✗ Error adjusting {scenario}: {e}", "ERROR")
                try:
                    wb.close()
                except:
                    pass

        self.log("")
        self.log(f"Trade balance adjustments completed: {trade_balance_changes} values updated")

    def calculate_technology_shares(self, country_code, scenario, all_years):
        """
        Calculate technology shares for each year based on renewability targets.

        Uses OLADE base year data and interpolates to reach renewability targets.
        Renewable techs: HYD, SPV, WND, GEO, BIO
        Non-renewable techs: COA, NGS, OIL, PET, URN

        Args:
            country_code: ISO3 country code
            scenario: Scenario name (BAU, NDC, etc.)
            all_years: list of years

        Returns:
            dict: {tech_type: {year: share}} for all technologies
        """
        renewable_techs = ['HYD', 'SPV', 'WON', 'GEO', 'BIO']
        non_renewable_techs = ['COA', 'NGS', 'OIL', 'PET', 'URN']

        # Get renewability target configuration for this country/scenario
        target_key = (country_code, scenario)
        target_config = self.olade_config.get('renewability_targets', {}).get(target_key)

        # Get custom weights if defined (new structure: {'renewable': {...}, 'non_renewable': {...}})
        custom_weights = self.olade_config.get('technology_weights', {}).get(target_key)

        # Get base year shares from OLADE (via shares_total_data)
        if scenario not in self.shares_total_data or country_code not in self.shares_total_data[scenario]:
            return {}

        base_shares = self.shares_total_data[scenario][country_code]

        # If no renewability targets defined, use original shares_total_data
        if not target_config:
            return base_shares

        # Get target years and values
        targets = target_config['targets']  # {year: renewable_percentage}
        interpolation = target_config['interpolation']  # 'linear' or 'flat_step'

        # Find base year (first year in all_years that has OLADE data)
        ref_year = self.generation_data['reference_year']
        base_year = min(all_years) if all_years else ref_year

        # Get base year shares from OLADE direct calculations (tech_shares)
        # These are the actual percentages calculated from OLADE generation data for 2023
        olade_tech_shares = {}
        if 'tech_shares' in self.generation_data and country_code in self.generation_data['tech_shares']:
            olade_tech_shares = self.generation_data['tech_shares'][country_code]

        # Calculate base renewable share from OLADE direct data
        # Note: OLADE uses 'PETROLEUM' (combined PET+OIL), so we need to handle this
        if olade_tech_shares:
            # Use OLADE direct shares for base year
            base_renewable_share = 0.0
            for tech in renewable_techs:
                base_renewable_share += olade_tech_shares.get(tech, 0.0)

            base_non_renewable_share = 1.0 - base_renewable_share
        else:
            # Fallback to Shares_Total if OLADE data not available
            base_renewable_share = sum(base_shares.get(tech, {}).get(base_year, 0.0) for tech in renewable_techs)
            base_non_renewable_share = 1.0 - base_renewable_share

        # Calculate renewable tech proportions (for distributing renewable target)
        if custom_weights and custom_weights.get('renewable'):
            # Use custom renewable weights
            renewable_proportions = custom_weights['renewable']
            # Normalize if needed
            total_weight = sum(renewable_proportions.values())
            if total_weight > 0 and abs(total_weight - 1.0) > 0.001:
                renewable_proportions = {k: v / total_weight for k, v in renewable_proportions.items()}
        else:
            # Use proportional distribution based on OLADE direct data
            if olade_tech_shares:
                total_renewable = sum(olade_tech_shares.get(tech, 0.0) for tech in renewable_techs)
                if total_renewable > 0:
                    renewable_proportions = {
                        tech: olade_tech_shares.get(tech, 0.0) / total_renewable
                        for tech in renewable_techs
                    }
                else:
                    # Default equal distribution if no renewable data
                    renewable_proportions = {tech: 1.0 / len(renewable_techs) for tech in renewable_techs}
            else:
                # Fallback to Shares_Total
                total_renewable = sum(base_shares.get(tech, {}).get(base_year, 0.0) for tech in renewable_techs)
                if total_renewable > 0:
                    renewable_proportions = {
                        tech: base_shares.get(tech, {}).get(base_year, 0.0) / total_renewable
                        for tech in renewable_techs
                    }
                else:
                    # Default equal distribution if no renewable data
                    renewable_proportions = {tech: 1.0 / len(renewable_techs) for tech in renewable_techs}

        # Calculate non-renewable tech proportions (for distributing non-renewable portion)
        if custom_weights and custom_weights.get('non_renewable'):
            # Use custom non-renewable weights
            non_renewable_proportions = custom_weights['non_renewable']
            # Normalize if needed
            total_weight = sum(non_renewable_proportions.values())
            if total_weight > 0 and abs(total_weight - 1.0) > 0.001:
                non_renewable_proportions = {k: v / total_weight for k, v in non_renewable_proportions.items()}
        else:
            # Use proportional distribution based on OLADE direct data
            # Note: OLADE has 'PETROLEUM' instead of separate 'PET' and 'OIL'
            # We need to handle this specially
            if olade_tech_shares:
                # Build non-renewable shares, handling PETROLEUM specially
                non_renewable_shares = {}
                for tech in non_renewable_techs:
                    if tech in ['PET', 'OIL']:
                        # Split PETROLEUM between PET and OIL
                        # Use Shares_Power_Generation_Technologies.xlsx to get the split ratio for base year
                        petroleum_share = olade_tech_shares.get('PETROLEUM', 0.0)
                        if petroleum_share > 0:
                            # Get PET and OIL shares from Shares_Power_Generation_Technologies for the split ratio
                            pet_share_total = base_shares.get('PET', {}).get(base_year, 0.0)
                            oil_share_total = base_shares.get('OIL', {}).get(base_year, 0.0)
                            total_petroleum_shares = pet_share_total + oil_share_total

                            if total_petroleum_shares > 0:
                                # Split proportionally based on Shares_Total
                                if tech == 'PET':
                                    non_renewable_shares[tech] = petroleum_share * (pet_share_total / total_petroleum_shares)
                                else:  # OIL
                                    non_renewable_shares[tech] = petroleum_share * (oil_share_total / total_petroleum_shares)
                            else:
                                # Default 50/50 split
                                non_renewable_shares[tech] = petroleum_share * 0.5
                        else:
                            non_renewable_shares[tech] = 0.0
                    else:
                        non_renewable_shares[tech] = olade_tech_shares.get(tech, 0.0)

                total_non_renewable = sum(non_renewable_shares.values())
                if total_non_renewable > 0:
                    non_renewable_proportions = {
                        tech: non_renewable_shares[tech] / total_non_renewable
                        for tech in non_renewable_techs
                    }
                else:
                    # Default equal distribution if no non-renewable data
                    non_renewable_proportions = {tech: 1.0 / len(non_renewable_techs) for tech in non_renewable_techs}
            else:
                # Fallback to Shares_Total
                total_non_renewable = sum(base_shares.get(tech, {}).get(base_year, 0.0) for tech in non_renewable_techs)
                if total_non_renewable > 0:
                    non_renewable_proportions = {
                        tech: base_shares.get(tech, {}).get(base_year, 0.0) / total_non_renewable
                        for tech in non_renewable_techs
                    }
                else:
                    # Default equal distribution if no non-renewable data
                    non_renewable_proportions = {tech: 1.0 / len(non_renewable_techs) for tech in non_renewable_techs}

        # Sort target years
        sorted_target_years = sorted(targets.keys())

        # Build the interpolated shares for each year
        result = {tech: {} for tech in renewable_techs + non_renewable_techs}

        for year in all_years:
            # Determine renewable percentage for this year
            if year <= base_year:
                # Use base year values
                renewable_pct = base_renewable_share
            elif not sorted_target_years:
                # No targets, use base
                renewable_pct = base_renewable_share
            else:
                # Find where this year falls relative to targets
                if year >= sorted_target_years[-1]:
                    # Beyond last target year
                    last_target_year = sorted_target_years[-1]
                    last_target_pct = targets[last_target_year]

                    if interpolation == 'linear':
                        # Calculate average annual growth rate from interpolated years
                        # Use the growth rate between the last two target points
                        if len(sorted_target_years) >= 2:
                            # Use last two targets to calculate rate
                            second_last_year = sorted_target_years[-2]
                            second_last_pct = targets[second_last_year]
                            years_diff = last_target_year - second_last_year
                            if years_diff > 0:
                                annual_growth_rate = (last_target_pct - second_last_pct) / years_diff
                            else:
                                annual_growth_rate = 0.0
                        else:
                            # Only one target, use growth from base year
                            years_diff = last_target_year - base_year
                            if years_diff > 0:
                                annual_growth_rate = (last_target_pct - base_renewable_share) / years_diff
                            else:
                                annual_growth_rate = 0.0

                        # Apply linear growth beyond last target
                        years_beyond = year - last_target_year
                        renewable_pct = last_target_pct + (annual_growth_rate * years_beyond)

                        # Cap at 100% (1.0)
                        renewable_pct = min(renewable_pct, 1.0)

                    else:  # flat_step
                        # After target year, grow proportionally with demand
                        # Get demand growth rate for this country
                        growth_rates = self.olade_config.get('demand_growth_rates', {})
                        demand_growth_rate = growth_rates.get(country_code, 0.02)  # Default 2%

                        # Apply compound growth from last target year
                        years_beyond = year - last_target_year
                        renewable_pct = last_target_pct * ((1 + demand_growth_rate) ** years_beyond)

                        # Cap at 100% (1.0)
                        renewable_pct = min(renewable_pct, 1.0)
                else:
                    # Find the surrounding target years
                    prev_year = base_year
                    prev_pct = base_renewable_share
                    next_year = None
                    next_pct = None

                    for ty in sorted_target_years:
                        if ty <= year:
                            prev_year = ty
                            prev_pct = targets[ty]
                        else:
                            next_year = ty
                            next_pct = targets[ty]
                            break

                    if next_year is None:
                        renewable_pct = prev_pct
                    elif interpolation == 'linear':
                        # Linear interpolation
                        if next_year != prev_year:
                            ratio = (year - prev_year) / (next_year - prev_year)
                            renewable_pct = prev_pct + ratio * (next_pct - prev_pct)
                        else:
                            renewable_pct = prev_pct
                    else:  # flat_step
                        # Keep flat until target year, then step
                        renewable_pct = prev_pct

            # Distribute renewable percentage among renewable techs
            for tech in renewable_techs:
                tech_share = renewable_pct * renewable_proportions.get(tech, 0.0)
                result[tech][year] = tech_share

            # Distribute non-renewable percentage among non-renewable techs
            non_renewable_pct = 1.0 - renewable_pct
            for tech in non_renewable_techs:
                tech_share = non_renewable_pct * non_renewable_proportions.get(tech, 0.0)
                result[tech][year] = tech_share

        return result

    def calculate_fuel_shares_from_olade(self, country_code, renewable_pct, year, all_years):
        """
        Distribute renewable percentage among renewable fuels and non-renewable percentage
        among non-renewable fuels using OLADE generation weights.

        Fuel groups:
        - Renewable: HYD, SPV, WON, GEO, BIO, CSP, WOF, WAV
        - Non-renewable: COA, NGS, OIL, PET, URN
        - Other (weight=0 if no OLADE data): BCK, CCS, COG, LDS, OTH, SDS, WAS

        Args:
            country_code: ISO3 country code
            renewable_pct: Renewable percentage for this year (0.0-1.0)
            year: Target year
            all_years: List of all years (for reference)

        Returns:
            dict: {fuel_code: share} where sum of all shares = 1.0
        """
        renewable_fuels = ['HYD', 'SPV', 'WON', 'GEO', 'BIO', 'CSP', 'WOF', 'WAV']
        non_renewable_fuels = ['COA', 'NGS', 'OIL', 'PET', 'URN']
        other_fuels = ['BCK', 'CCS', 'COG', 'LDS', 'OTH', 'SDS', 'WAS']
        all_fuels = renewable_fuels + non_renewable_fuels + other_fuels

        result = {fuel: 0.0 for fuel in all_fuels}

        # Get OLADE tech shares for this country
        olade_tech_shares = {}
        if 'tech_shares' in self.generation_data and country_code in self.generation_data['tech_shares']:
            olade_tech_shares = self.generation_data['tech_shares'][country_code]

        # Calculate renewable fuel proportions from OLADE
        total_renewable_olade = 0.0
        renewable_olade_shares = {}
        for fuel in renewable_fuels:
            if fuel == 'WON':
                # OLADE uses 'Eólica' -> mapped to 'WON' in OLADE_TECH_MAPPING
                share = olade_tech_shares.get(fuel, 0.0)
            else:
                share = olade_tech_shares.get(fuel, 0.0)
            renewable_olade_shares[fuel] = share
            total_renewable_olade += share

        # Calculate non-renewable fuel proportions from OLADE
        total_non_renewable_olade = 0.0
        non_renewable_olade_shares = {}
        for fuel in non_renewable_fuels:
            if fuel in ['PET', 'OIL']:
                # OLADE has 'PETROLEUM' for combined PET+OIL
                # Split using proportions from Shares_Total if available
                petroleum_share = olade_tech_shares.get('PETROLEUM', 0.0)
                if petroleum_share > 0:
                    # Try to get split ratio from shares_total_data
                    base_year = min(all_years) if all_years else 2023
                    for scenario in self.shares_total_data:
                        if country_code in self.shares_total_data[scenario]:
                            pet_share_total = self.shares_total_data[scenario][country_code].get('PET', {}).get(base_year, 0.0)
                            oil_share_total = self.shares_total_data[scenario][country_code].get('OIL', {}).get(base_year, 0.0)
                            total_petroleum = pet_share_total + oil_share_total
                            if total_petroleum > 0:
                                if fuel == 'PET':
                                    non_renewable_olade_shares[fuel] = petroleum_share * (pet_share_total / total_petroleum)
                                else:
                                    non_renewable_olade_shares[fuel] = petroleum_share * (oil_share_total / total_petroleum)
                            else:
                                non_renewable_olade_shares[fuel] = petroleum_share * 0.5
                            break
                    else:
                        # No Shares_Total data, use 50/50 split
                        non_renewable_olade_shares[fuel] = petroleum_share * 0.5
                else:
                    non_renewable_olade_shares[fuel] = 0.0
            else:
                non_renewable_olade_shares[fuel] = olade_tech_shares.get(fuel, 0.0)
            total_non_renewable_olade += non_renewable_olade_shares.get(fuel, 0.0)

        # Distribute renewable percentage among renewable fuels
        if total_renewable_olade > 0:
            for fuel in renewable_fuels:
                proportion = renewable_olade_shares.get(fuel, 0.0) / total_renewable_olade
                result[fuel] = renewable_pct * proportion
        else:
            # No OLADE data for renewables, distribute equally among HYD, SPV, WON
            default_renewables = ['HYD', 'SPV', 'WON']
            for fuel in default_renewables:
                result[fuel] = renewable_pct / len(default_renewables)

        # Distribute non-renewable percentage among non-renewable fuels
        non_renewable_pct = 1.0 - renewable_pct
        if total_non_renewable_olade > 0:
            for fuel in non_renewable_fuels:
                proportion = non_renewable_olade_shares.get(fuel, 0.0) / total_non_renewable_olade
                result[fuel] = non_renewable_pct * proportion
        else:
            # No OLADE data for non-renewables, distribute equally among COA, NGS
            default_non_renewables = ['COA', 'NGS']
            for fuel in default_non_renewables:
                result[fuel] = non_renewable_pct / len(default_non_renewables)

        # Other fuels stay at 0.0 (no OLADE data)
        # print(country_code,year,result)
        return result

    def normalize_shares_by_year(self, tech_shares, all_years):
        """
        Normalize shares to ensure they sum to 1.0 for each year.

        Args:
            tech_shares: {tech_type: {year: share}}
            all_years: list of years

        Returns:
            dict: {tech_type: {year: normalized_share}}
        """
        result = {tech: {} for tech in tech_shares}

        for year in all_years:
            total_share = sum(tech_shares.get(tech, {}).get(year, 0.0) for tech in tech_shares)

            if total_share > 0:
                for tech in tech_shares:
                    original_share = tech_shares.get(tech, {}).get(year, 0.0)
                    result[tech][year] = original_share / total_share
            else:
                # If total is 0, keep shares as 0
                for tech in tech_shares:
                    result[tech][year] = 0.0

        return result

    def read_osemosys_defaults(self):
        """
        Read OSeMOSYS parameter default values from conversion_format.yaml.

        Returns:
            dict: Parameter defaults (e.g., {'CapacityFactor': 1, 'AvailabilityFactor': 1, ...})
        """
        import yaml
        from pathlib import Path

        yaml_path = Path(__file__).parent / 'Miscellaneous' / 'conversion_format.yaml'
        if not yaml_path.exists():
            self.log("WARNING: conversion_format.yaml not found, using hardcoded defaults")
            return {'CapacityFactor': 1.0}

        try:
            with open(yaml_path, 'r', encoding='utf-8') as f:
                config = yaml.safe_load(f)

            defaults = {}
            for param_name, param_config in config.items():
                if isinstance(param_config, dict) and 'default' in param_config:
                    defaults[param_name] = param_config['default']

            return defaults
        except Exception as e:
            self.log(f"WARNING: Error reading conversion_format.yaml: {e}")
            return {'CapacityFactor': 1.0}

    def apply_ndc_bau_override(self, tech_shares_ndc, tech_shares_bau, override_years=None):
        """
        For NDC scenarios, replace shares for years 2023-2025 with BAU values.
        This ensures policy effects don't apply until 2026.

        Args:
            tech_shares_ndc: {tech_type: {year: share}} for NDC scenario
            tech_shares_bau: {tech_type: {year: share}} for BAU scenario
            override_years: list of years to override (default: [2023, 2024, 2025])

        Returns:
            dict: Modified NDC shares with BAU values for override years
        """
        if override_years is None:
            override_years = [2023, 2024, 2025]

        result = {tech: dict(years) for tech, years in tech_shares_ndc.items()}

        for tech_type in result:
            for year in override_years:
                if year in tech_shares_bau.get(tech_type, {}):
                    result[tech_type][year] = tech_shares_bau[tech_type][year]

        return result

    def read_yearsplit_values(self, wb):
        """
        Read YearSplit values from Yearsplit sheet.

        Returns:
            dict: {timeslice: proportion} where sum of all proportions = 1.0
        """
        yearsplit_values = {}

        if 'Yearsplit' not in wb.sheetnames:
            return yearsplit_values

        ws_yearsplit = wb['Yearsplit']

        # Find first year column (YearSplit values are in year columns, e.g., column 7 for 2023)
        # Look for a column with a numeric year header (2023, 2024, etc.)
        value_col = None
        for col_idx in range(1, min(15, ws_yearsplit.max_column + 1)):
            header = ws_yearsplit.cell(1, col_idx).value
            if header and str(header).strip().isdigit():
                value_col = col_idx
                break

        if not value_col:
            # Fallback: try column 5 (old format)
            value_col = 5

        # Read timeslice and value from rows
        for row_idx in range(2, ws_yearsplit.max_row + 1):
            timeslice = ws_yearsplit.cell(row_idx, 1).value
            value = ws_yearsplit.cell(row_idx, value_col).value

            if timeslice and value is not None:
                try:
                    yearsplit_values[str(timeslice).strip()] = float(value)
                except (ValueError, TypeError):
                    pass

        return yearsplit_values

    def read_capacity_factors_sum(self, wb, tech_str, year):
        """
        Read CapacityFactor values from Capacities sheet and sum with YearSplit weights.

        Returns: sum(CapacityFactor[l,y] × YearSplit[l,y]) for all timeslices
        """
        # Get default from conversion_format.yaml
        defaults = self.read_osemosys_defaults()
        default_cf = defaults.get('CapacityFactor', 1.0)

        if 'Capacities' not in wb.sheetnames:
            return default_cf  # Fallback to default from yaml

        ws_cap = wb['Capacities']

        # Read YearSplit from Yearsplit sheet
        yearsplit_values = self.read_yearsplit_values(wb)

        # Find year column
        year_col = None
        for col_idx in range(1, ws_cap.max_column + 1):
            header = ws_cap.cell(1, col_idx).value
            if header and str(header) == str(year):
                year_col = col_idx
                break

        if not year_col:
            return default_cf  # Fallback to default from yaml

        # Sum CF × YearSplit for all timeslices of this tech
        cf_sum = 0.0
        for row_idx in range(2, ws_cap.max_row + 1):
            row_tech = ws_cap.cell(row_idx, 3).value  # Column 3: Tech (Col 2 is Tech.ID)
            param = ws_cap.cell(row_idx, 6).value     # Column 6: Parameter
            timeslice = ws_cap.cell(row_idx, 1).value # Column 1: Timeslice

            if (row_tech and str(row_tech).strip().upper() == tech_str and
                param and str(param).strip() == 'CapacityFactor'):

                cf_value = ws_cap.cell(row_idx, year_col).value
                if cf_value is not None and timeslice:
                    try:
                        cf = float(cf_value)
                        yearsplit = yearsplit_values.get(str(timeslice).strip(), 0.0)
                        cf_sum += cf * yearsplit
                    except (ValueError, TypeError):
                        pass

        # If no CapacityFactor found, use default from conversion_format.yaml
        return cf_sum if cf_sum > 0 else default_cf

    def calculate_max_possible_activity(self, wb, ws, tech_str, year, year_col_map,
                                         capacity_to_activity, max_capacity_rows,
                                         residual_capacity_rows, availability_rows):
        """
        Calculate maximum possible activity for a technology in a given year.

        IMPORTANT: Uses ONLY TotalAnnualMaxCapacity (not ResidualCapacity).
        Formula: MaxPossibleActivity = MaxCapacity × C2A × AvailabilityFactor × sum(CF × YearSplit)
        Note: C2A already includes the 8760 hours/year conversion, so we do NOT multiply by 8760.

        All parameters use defaults from conversion_format.yaml if not defined:
        - TotalAnnualMaxCapacity: default = -1 (no limit in OSeMOSYS)
        - AvailabilityFactor: default = 1.0
        - CapacityFactor: default = 1.0
        - CapacityToActivityUnit: default = 1.0 (but 31.536 for PWR* power plants)

        Returns: (max_possible_activity, capacity_source, components_dict)
        where components_dict = {
            'max_cap': float,
            'residual_cap': float,
            'avail': float,
            'c2a': float,
            'cf_yearsplit_sum': float
        }
        """
        if year not in year_col_map:
            return (None, None, {})

        col_idx = year_col_map[year]

        # Get defaults from conversion_format.yaml
        defaults = self.read_osemosys_defaults()

        # Get C2A with fallback to appropriate default
        # For power generation technologies (PWR*), use 31.536 PJ/GW/year
        # For other technologies, use YAML default (1.0)
        if tech_str not in capacity_to_activity:
            if tech_str.startswith('PWR'):
                c2a = 31.536  # Standard conversion for power plants (GW to PJ/year)
            else:
                c2a = defaults.get('CapacityToActivityUnit', 1.0)
        else:
            c2a = capacity_to_activity[tech_str]

        avail_row = availability_rows.get(tech_str)
        max_cap_row = max_capacity_rows.get(tech_str)
        residual_cap_row = residual_capacity_rows.get(tech_str)

        # Read AvailabilityFactor with YAML default
        avail = None
        if avail_row:
            avail_value = ws.cell(avail_row, col_idx).value
            if avail_value is not None:
                try:
                    avail = float(avail_value)
                except (ValueError, TypeError):
                    pass

        # Use default from YAML if not found
        if avail is None:
            avail = defaults.get('AvailabilityFactor', 1.0)

        # Read CapacityFactor from Capacities sheet with YAML fallback
        cf_yearsplit_sum = self.read_capacity_factors_sum(wb, tech_str, year)

        # Read TotalAnnualMaxCapacity with YAML default
        max_cap = None
        max_cap_activity = None

        if max_cap_row:
            max_cap_value = ws.cell(max_cap_row, col_idx).value
            # Handle None or empty cells - use YAML default
            if max_cap_value is None or max_cap_value == '':
                max_cap = defaults.get('TotalAnnualMaxCapacity', -1)
            else:
                try:
                    max_cap = float(max_cap_value)
                except (ValueError, TypeError):
                    max_cap = defaults.get('TotalAnnualMaxCapacity', -1)
        else:
            # Row doesn't exist, use YAML default
            max_cap = defaults.get('TotalAnnualMaxCapacity', -1)

        # Calculate activity only if MaxCapacity is valid (not -1 which means no limit)
        if max_cap is not None and max_cap >= 0:
            # Formula: MaxPossibleActivity = MaxCap × C2A × Avail × sum(CF × YearSplit)
            max_cap_activity = max_cap * c2a * avail * cf_yearsplit_sum
        elif max_cap == -1:
            # -1 means no limit in OSeMOSYS, so max_possible is effectively infinite
            max_cap_activity = float('inf')

        # Read ResidualCapacity for logging purposes only (NOT used in calculation per requirement)
        residual_cap = None
        if residual_cap_row:
            residual_cap_value = ws.cell(residual_cap_row, col_idx).value
            if residual_cap_value is not None:
                try:
                    residual_cap = float(residual_cap_value)
                except (ValueError, TypeError):
                    pass

        # FIXED: Use ONLY TotalAnnualMaxCapacity (per user requirement)
        if max_cap_activity is not None:
            max_possible = max_cap_activity
            capacity_source = 'MaxCapacity'
        else:
            max_possible = 0
            capacity_source = 'None'

        components = {
            'max_cap': max_cap,
            'residual_cap': residual_cap,
            'avail': avail,
            'c2a': c2a,
            'cf_yearsplit_sum': cf_yearsplit_sum
        }

        return (max_possible, capacity_source, components)

    def increase_max_capacity_for_limit(self, ws, tech_str, year, year_col_map,
                                        target_lower_limit, capacity_components,
                                        max_capacity_rows, residual_capacity_rows):
        """
        Increase TotalAnnualMaxCapacity to make target_lower_limit achievable.

        Formula: Required_Capacity = (LowerLimit / (C2A × Avail × sum(CF × YearSplit))) × 1.05
        Note: C2A already includes the 8760 hours/year conversion, so we do NOT divide by 8760.
        A 5% safety margin is applied to ensure MaxCapacity > LowerLimit (strictly greater, not equal).

        Also adjusts ResidualCapacity if needed to maintain constraint:
        ResidualCapacity <= TotalAnnualMaxCapacity

        Returns: (new_max_capacity, max_adjusted, residual_adjusted)
        """
        if year not in year_col_map:
            return (None, False, False)

        c2a = capacity_components.get('c2a', 31.536)
        avail = capacity_components.get('avail')
        cf_sum = capacity_components.get('cf_yearsplit_sum', 0.28)

        if not avail or cf_sum <= 0 or c2a <= 0:
            return (None, False, False)

        # FIXED: Removed / 8760 because C2A already includes hours/year conversion
        # Formula: Required_Capacity = LowerLimit / (C2A × Avail × sum(CF × YearSplit))
        # Add 5% safety margin to ensure MaxCapacity is STRICTLY GREATER than LowerLimit (not equal)
        required_cap = (target_lower_limit / (c2a * avail * cf_sum)) * 1.05

        # Update TotalAnnualMaxCapacity
        max_cap_row = max_capacity_rows.get(tech_str)
        if not max_cap_row:
            return (None, False, False)

        col_idx = year_col_map[year]
        ws.cell(max_cap_row, col_idx, round(required_cap, 6))
        max_adjusted = True

        # Ensure ResidualCapacity <= TotalAnnualMaxCapacity
        # Note: required_cap already includes the 5% safety margin
        residual_adjusted = False
        residual_row = residual_capacity_rows.get(tech_str)
        if residual_row:
            current_residual = ws.cell(residual_row, col_idx).value
            if current_residual:
                try:
                    current_residual_float = float(current_residual)
                    if current_residual_float > required_cap:
                        # Residual exceeds new max - cap it to maintain OSeMOSYS constraint
                        # ResidualCapacity is capped to the new MaxCapacity (which has the 5% margin)
                        ws.cell(residual_row, col_idx, round(required_cap, 6))
                        residual_adjusted = True
                except (ValueError, TypeError):
                    pass

        return (required_cap, max_adjusted, residual_adjusted)

    def calculate_demand_based_limits(self, scenario, country_code, all_years, demand_data,
                                        base_scenario='BAU', base_shares=None):
        """
        Calculate LowerLimits using DemandBased method.

        Formula: LowerLimit[tech,year] = Demand[country,year] × Share[tech,year]

        Steps:
        1. Read % renewable from Renewability_Targets (with interpolation)
        2. Calculate shares from renewability using OLADE weights
        3. Normalize shares year-by-year (Σ = 1.0)
        4. For non-base scenarios, apply base scenario override for 2023-2025
        5. LowerLimit = Demand × Share for each tech/year

        Args:
            scenario: scenario name (BAU, NDC, etc.)
            country_code: ISO3 country code
            all_years: list of years
            demand_data: {country_code: {year: demand_pj}} from read_demand_data()
            base_scenario: base scenario name from YAML (default: 'BAU')
            base_shares: Optional pre-calculated base scenario shares for override

        Returns:
            dict: {tech_str: {year: lower_limit_pj}}
        """
        # Get demand for this country
        country_demand = demand_data.get(country_code, {})
        if not country_demand:
            self.log(f"  No demand data for {country_code}", "WARNING")
            return {}

        # Get renewability target configuration for this country/scenario
        target_key = (country_code, scenario)
        target_config = self.olade_config.get('renewability_targets', {}).get(target_key)

        # Get base renewable share from OLADE
        olade_tech_shares = {}
        if 'tech_shares' in self.generation_data and country_code in self.generation_data['tech_shares']:
            olade_tech_shares = self.generation_data['tech_shares'][country_code]

        renewable_fuels = ['HYD', 'SPV', 'WON', 'GEO', 'BIO', 'CSP', 'WOF', 'WAV']
        base_renewable_share = sum(olade_tech_shares.get(tech, 0.0) for tech in renewable_fuels if tech in olade_tech_shares)

        ref_year = self.generation_data['reference_year']
        base_year = min(all_years) if all_years else ref_year

        # Get targets and interpolation method
        if target_config:
            targets = target_config['targets']  # {year: renewable_percentage}
            interpolation = target_config['interpolation']  # 'linear' or 'flat_step'
        else:
            targets = {}
            interpolation = 'flat_step'

        sorted_target_years = sorted(targets.keys()) if targets else []

        # Calculate shares for each year
        all_fuels = ['HYD', 'SPV', 'WON', 'GEO', 'BIO', 'CSP', 'WOF', 'WAV',
                     'COA', 'NGS', 'OIL', 'PET', 'URN',
                     'BCK', 'CCS', 'COG', 'LDS', 'OTH', 'SDS', 'WAS']
        tech_shares = {fuel: {} for fuel in all_fuels}

        for year in all_years:
            # Interpolate renewable percentage for this year
            if year <= base_year:
                renewable_pct = base_renewable_share
            elif not sorted_target_years:
                renewable_pct = base_renewable_share
            elif year >= sorted_target_years[-1]:
                # Beyond last target year
                last_target_year = sorted_target_years[-1]
                last_target_pct = targets[last_target_year]

                if interpolation == 'linear':
                    # Calculate average annual growth rate from interpolated years
                    # Use the growth rate between the last two target points
                    if len(sorted_target_years) >= 2:
                        # Use last two targets to calculate rate
                        second_last_year = sorted_target_years[-2]
                        second_last_pct = targets[second_last_year]
                        years_diff = last_target_year - second_last_year
                        if years_diff > 0:
                            annual_growth_rate = (last_target_pct - second_last_pct) / years_diff
                        else:
                            annual_growth_rate = 0.0
                    else:
                        # Only one target, use growth from base year
                        years_diff = last_target_year - base_year
                        if years_diff > 0:
                            annual_growth_rate = (last_target_pct - base_renewable_share) / years_diff
                        else:
                            annual_growth_rate = 0.0

                    # Apply linear growth beyond last target
                    years_beyond = year - last_target_year
                    renewable_pct = last_target_pct + (annual_growth_rate * years_beyond)

                    # Cap at 100% (1.0)
                    renewable_pct = min(renewable_pct, 1.0)

                else:  # flat_step
                    # After target year, grow proportionally with demand
                    # Get demand growth rate for this country
                    growth_rates = self.olade_config.get('demand_growth_rates', {})
                    demand_growth_rate = growth_rates.get(country_code, 0.02)  # Default 2%

                    # Apply compound growth from last target year
                    years_beyond = year - last_target_year
                    renewable_pct = last_target_pct * ((1 + demand_growth_rate) ** years_beyond)

                    # Cap at 100% (1.0)
                    renewable_pct = min(renewable_pct, 1.0)
            else:
                # Find surrounding target years
                prev_year = base_year
                prev_pct = base_renewable_share
                next_year = None
                next_pct = None

                for ty in sorted_target_years:
                    if ty <= year:
                        prev_year = ty
                        prev_pct = targets[ty]
                    else:
                        next_year = ty
                        next_pct = targets[ty]
                        break

                if next_year is None:
                    renewable_pct = prev_pct
                elif interpolation == 'linear':
                    if next_year != prev_year:
                        ratio = (year - prev_year) / (next_year - prev_year)
                        renewable_pct = prev_pct + ratio * (next_pct - prev_pct)
                    else:
                        renewable_pct = prev_pct
                else:  # flat_step
                    renewable_pct = prev_pct

            # Calculate shares for this year using OLADE weights
            year_shares = self.calculate_fuel_shares_from_olade(country_code, renewable_pct, year, all_years)
            for fuel in year_shares:
                tech_shares[fuel][year] = year_shares[fuel]

        # Normalize shares
        tech_shares = self.normalize_shares_by_year(tech_shares, all_years)

        # Apply base scenario override for years 2023-2025 if this is NOT the base scenario
        if scenario != base_scenario and base_shares:
            tech_shares = self.apply_ndc_bau_override(tech_shares, base_shares, [2023, 2024, 2025])

        # DEBUG: Store normalized shares in cache for CSV export
        self.demand_based_shares_cache[(scenario, country_code)] = tech_shares
        # Count non-zero shares for logging
        non_zero_techs = sum(1 for fuel in tech_shares if any(tech_shares[fuel].values()))
        self.log(f"  DEBUG: Stored {non_zero_techs} technologies with shares for {country_code}/{scenario}")

        # Calculate LowerLimits: Demand × Share
        result = {}
        for fuel in all_fuels:
            tech_str = f"PWR{fuel}{country_code}XX"
            result[tech_str] = {}

            for year in all_years:
                demand_pj = country_demand.get(year, 0.0)
                share = tech_shares.get(fuel, {}).get(year, 0.0)
                limit_value = demand_pj * share
                #limit_value = share
                result[tech_str][year] = round(limit_value, 4)
                #print(country_code, year, result)
        return result

    def adjust_capacity_factors_for_share_based_limits(self, wb, scenario, country_code, tech_shares,
                                                         all_years, ref_year, base_generation_pj, growth_rate):
        """
        Adjust CapacityFactor values in the Capacities sheet to meet share-based LowerLimits.

        This method:
        1. Calculates target LowerLimit based on shares
        2. Checks if current CapacityFactor allows meeting that target with TotalAnnualMinCapacityInvestment
        3. Adjusts CapacityFactor proportionally if needed
        4. Normalizes all adjustments to ensure total share = 100%
        5. Caps individual CapacityFactors to maximum 1.0

        Args:
            wb: openpyxl workbook
            scenario: scenario name
            country_code: ISO3 country code
            tech_shares: dict {tech_type: {year: share}}
            all_years: list of years
            ref_year: OLADE reference year
            base_generation_pj: base generation in PJ
            growth_rate: annual growth rate

        Returns:
            dict: {tech_str: {year: adjusted_lower_limit}}
        """
        if 'Secondary Techs' not in wb.sheetnames or 'Capacities' not in wb.sheetnames:
            self.log("  ✗ Required sheets not found for ShareBased method", "WARNING")
            return {}

        ws_secondary = wb['Secondary Techs']
        ws_capacities = wb['Capacities']

        # Build indices for Secondary Techs sheet
        year_col_map = {}
        headers = [cell.value for cell in ws_secondary[1]]
        for col_idx, header in enumerate(headers, 1):
            if header and str(header).isdigit():
                try:
                    year = int(header)
                    if 2000 <= year <= 2100:
                        year_col_map[year] = col_idx
                except:
                    pass

        # Find TotalAnnualMinCapacityInvestment rows (Parameter.ID = 7)
        min_capacity_rows = {}  # {tech_str: row_idx}
        availability_rows = {}  # {tech_str: row_idx}

        for row_idx in range(2, ws_secondary.max_row + 1):
            tech_code = ws_secondary.cell(row_idx, 2).value
            param_id = ws_secondary.cell(row_idx, 4).value
            parameter = ws_secondary.cell(row_idx, 5).value

            if tech_code:
                tech_str = str(tech_code).strip().upper()
                if param_id == 7 or (parameter and str(parameter).strip() == 'TotalAnnualMinCapacityInvestment'):
                    min_capacity_rows[tech_str] = row_idx
                elif parameter and str(parameter).strip() == 'AvailabilityFactor':
                    availability_rows[tech_str] = row_idx

        # Read CapacityToActivityUnit from Fixed Horizon Parameters
        capacity_to_activity = {}
        if 'Fixed Horizon Parameters' in wb.sheetnames:
            ws_fixed = wb['Fixed Horizon Parameters']
            for row_idx in range(2, ws_fixed.max_row + 1):
                tech_code = ws_fixed.cell(row_idx, 2).value
                parameter = ws_fixed.cell(row_idx, 6).value
                value = ws_fixed.cell(row_idx, 8).value
                if tech_code and parameter and value:
                    tech_str = str(tech_code).strip().upper()
                    if str(parameter).strip() == 'CapacityToActivityUnit':
                        try:
                            capacity_to_activity[tech_str] = float(value)
                        except (ValueError, TypeError):
                            pass

        # Build index for Capacities sheet
        # Structure: Timeslices | Tech.ID | Tech | Tech.Name | Parameter.ID | Parameter | Unit | Projection.Mode | Projection.Parameter | Year columns...
        cap_year_col_map = {}
        cap_headers = [cell.value for cell in ws_capacities[1]]
        for col_idx, header in enumerate(cap_headers, 1):
            if header and str(header).isdigit():
                try:
                    year = int(header)
                    if 2000 <= year <= 2100:
                        cap_year_col_map[year] = col_idx
                except:
                    pass

        # Find CapacityFactor rows for each technology (Parameter.ID = 6)
        # Group by technology and timeslice
        capacity_factor_rows = {}  # {tech_str: {timeslice: row_idx}}
        for row_idx in range(2, ws_capacities.max_row + 1):
            timeslice = ws_capacities.cell(row_idx, 1).value
            tech_code = ws_capacities.cell(row_idx, 2).value
            param_id = ws_capacities.cell(row_idx, 5).value
            parameter = ws_capacities.cell(row_idx, 6).value

            if tech_code and timeslice:
                tech_str = str(tech_code).strip().upper()
                timeslice_str = str(timeslice).strip()
                if param_id == 6 or (parameter and str(parameter).strip() == 'CapacityFactor'):
                    if tech_str not in capacity_factor_rows:
                        capacity_factor_rows[tech_str] = {}
                    capacity_factor_rows[tech_str][timeslice_str] = row_idx

        # Read YearSplit values (proportion of year for each timeslice)
        # These should be in Yearsplit sheet
        yearsplit_values = {}  # {timeslice: proportion}
        if 'Yearsplit' in wb.sheetnames:
            ws_yearsplit = wb['Yearsplit']
            for row_idx in range(2, ws_yearsplit.max_row + 1):
                timeslice = ws_yearsplit.cell(row_idx, 1).value
                value = ws_yearsplit.cell(row_idx, 5).value  # Assuming YearSplit is in column 5
                if timeslice and value is not None:
                    try:
                        yearsplit_values[str(timeslice).strip()] = float(value)
                    except (ValueError, TypeError):
                        pass

        # If no YearSplit data, use equal distribution across timeslices
        if not yearsplit_values and capacity_factor_rows:
            # Get all unique timeslices
            all_timeslices = set()
            for tech_timeslices in capacity_factor_rows.values():
                all_timeslices.update(tech_timeslices.keys())
            if all_timeslices:
                equal_split = 1.0 / len(all_timeslices)
                yearsplit_values = {ts: equal_split for ts in all_timeslices}

        self.log(f"  ShareBased: Processing {country_code} with {len(min_capacity_rows)} technologies")

        # Process each technology for each year
        adjusted_limits = {}  # {tech_str: {year: lower_limit}}
        adjustment_factors = {}  # {(tech_str, year): adjustment_factor}

        for tech_type, year_shares in tech_shares.items():
            tech_str = f"PWR{tech_type}{country_code}XX"

            # Skip if no TotalAnnualMinCapacityInvestment defined
            if tech_str not in min_capacity_rows:
                continue

            # Skip if no CapacityFactor data
            if tech_str not in capacity_factor_rows:
                continue

            min_cap_row = min_capacity_rows[tech_str]
            avail_row = availability_rows.get(tech_str)
            c2a = capacity_to_activity.get(tech_str, 31.536)

            if tech_str not in adjusted_limits:
                adjusted_limits[tech_str] = {}

            for year in all_years:
                if year not in year_col_map or year not in cap_year_col_map:
                    continue

                share = year_shares.get(year, 0.0)
                if share <= 0:
                    continue

                # Get TotalAnnualMinCapacityInvestment for this year
                min_cap_value = ws_secondary.cell(min_cap_row, year_col_map[year]).value
                if min_cap_value is None or float(min_cap_value) <= 0:
                    continue

                try:
                    min_cap = float(min_cap_value)
                except (ValueError, TypeError):
                    continue

                # Get AvailabilityFactor
                avail = 1.0  # Default
                if avail_row:
                    avail_value = ws_secondary.cell(avail_row, year_col_map[year]).value
                    if avail_value is not None:
                        try:
                            avail = float(avail_value)
                        except (ValueError, TypeError):
                            pass

                # Calculate target LowerLimit based on shares
                years_diff = year - ref_year
                generation_year = base_generation_pj * (1 + growth_rate * years_diff)
                target_lower_limit = generation_year * share

                # Calculate current maximum activity from minimum capacity
                # Sum all CapacityFactor * YearSplit for this technology
                sum_cf_yearsplit = 0.0
                cf_values = {}  # Store current CF values for later adjustment

                for timeslice, cf_row_idx in capacity_factor_rows[tech_str].items():
                    cf_value = ws_capacities.cell(cf_row_idx, cap_year_col_map[year]).value
                    if cf_value is not None:
                        try:
                            cf = float(cf_value)
                            yearsplit = yearsplit_values.get(timeslice, 0.0)
                            sum_cf_yearsplit += cf * yearsplit
                            cf_values[timeslice] = cf
                        except (ValueError, TypeError):
                            pass

                # MaxActivity = MinCapacity × C2A × Avail × Σ(CF × YearSplit) × 8760
                current_max_activity = min_cap * c2a * avail * sum_cf_yearsplit * 8760

                # Calculate adjustment factor if needed
                if current_max_activity > 0 and target_lower_limit > current_max_activity:
                    adjustment_factor = target_lower_limit / current_max_activity
                    adjustment_factors[(tech_str, year)] = adjustment_factor

                    # Store the target (will be normalized later)
                    adjusted_limits[tech_str][year] = target_lower_limit
                else:
                    # No adjustment needed
                    adjusted_limits[tech_str][year] = min(target_lower_limit, current_max_activity)

        # NORMALIZATION STEP
        # For each year, normalize all adjustments to ensure total share = 100%
        for year in all_years:
            if year not in year_col_map or year not in cap_year_col_map:
                continue

            # Collect all technologies that need adjustment for this year
            techs_to_adjust = []
            total_requested_share = 0.0

            for tech_type, year_shares in tech_shares.items():
                tech_str = f"PWR{tech_type}{country_code}XX"
                if (tech_str, year) in adjustment_factors:
                    share = year_shares.get(year, 0.0)
                    techs_to_adjust.append((tech_str, tech_type, share))
                    total_requested_share += share

            # Apply adjustments to CapacityFactor
            if techs_to_adjust:
                self.log(f"    Year {year}: Adjusting {len(techs_to_adjust)} technologies (total share: {total_requested_share:.2%})")

                for tech_str, tech_type, share in techs_to_adjust:
                    adjustment_factor = adjustment_factors.get((tech_str, year), 1.0)

                    # Normalize the adjustment factor
                    normalized_factor = adjustment_factor / total_requested_share if total_requested_share > 0 else 1.0

                    # Apply to all timeslices for this technology
                    if tech_str in capacity_factor_rows:
                        for timeslice, cf_row_idx in capacity_factor_rows[tech_str].items():
                            current_cf = ws_capacities.cell(cf_row_idx, cap_year_col_map[year]).value
                            if current_cf is not None:
                                try:
                                    cf = float(current_cf)
                                    new_cf = cf * normalized_factor
                                    # Cap at 1.0
                                    new_cf = min(new_cf, 1.0)
                                    ws_capacities.cell(cf_row_idx, cap_year_col_map[year], round(new_cf, 4))
                                except (ValueError, TypeError):
                                    pass

                        self.log(f"      {tech_type}: Adjusted CF by factor {normalized_factor:.4f}")

                # Recalculate adjusted limits after CF changes
                for tech_str, tech_type, share in techs_to_adjust:
                    min_cap_row = min_capacity_rows.get(tech_str)
                    avail_row = availability_rows.get(tech_str)
                    c2a = capacity_to_activity.get(tech_str, 31.536)

                    if not min_cap_row:
                        continue

                    min_cap_value = ws_secondary.cell(min_cap_row, year_col_map[year]).value
                    if min_cap_value is None or float(min_cap_value) <= 0:
                        continue

                    min_cap = float(min_cap_value)

                    avail = 1.0
                    if avail_row:
                        avail_value = ws_secondary.cell(avail_row, year_col_map[year]).value
                        if avail_value is not None:
                            try:
                                avail = float(avail_value)
                            except (ValueError, TypeError):
                                pass

                    # Recalculate sum_cf_yearsplit with adjusted values
                    sum_cf_yearsplit = 0.0
                    if tech_str in capacity_factor_rows:
                        for timeslice, cf_row_idx in capacity_factor_rows[tech_str].items():
                            cf_value = ws_capacities.cell(cf_row_idx, cap_year_col_map[year]).value
                            if cf_value is not None:
                                try:
                                    cf = float(cf_value)
                                    yearsplit = yearsplit_values.get(timeslice, 0.0)
                                    sum_cf_yearsplit += cf * yearsplit
                                except (ValueError, TypeError):
                                    pass

                    # Recalculate final LowerLimit
                    adjusted_max_activity = min_cap * c2a * avail * sum_cf_yearsplit * 8760

                    # Calculate target again
                    years_diff = year - ref_year
                    generation_year = base_generation_pj * (1 + growth_rate * years_diff)
                    target_lower_limit = generation_year * share

                    # Set final limit (can't exceed adjusted max activity)
                    final_limit = min(target_lower_limit, adjusted_max_activity)
                    adjusted_limits[tech_str][year] = final_limit

        return adjusted_limits

    def update_activity_limits(self, all_years):
        """
        Update TotalTechnologyAnnualActivityLowerLimit and/or TotalTechnologyAnnualActivityUpperLimit
        in A-O_Parametrization.xlsx files.

        Uses renewability targets from Renewability_Targets sheet to calculate shares.
        Formula: Generation_OLADE × (1 + growth_rate × (year - ref_year)) × Share_technology
        UpperLimit = LowerLimit + 0.1

        IMPORTANT: LowerLimit is capped to ensure it doesn't exceed what the MaxCapacity can produce.
        MaxPossibleActivity = MaxCapacity × CapacityToActivityUnit × AvailabilityFactor × sum(CapacityFactor × YearSplit)

        Args:
            all_years: list of years to populate
        """
        # Only DemandBased method is used after removing CapacityBased and ShareBased
        limit_method = 'DemandBased'

        update_lower = self.olade_config.get('activity_lower_limit_enabled', False)
        update_upper = self.olade_config.get('activity_upper_limit_enabled', False)

        if not update_lower and not update_upper:
            return

        if not self.generation_data:
            self.log("Cannot update Activity Limits: Missing generation data", "WARNING")
            return

        self.log("")
        self.log("=" * 80)
        self.log("UPDATING ACTIVITY LIMITS")
        self.log("=" * 80)

        ref_year = self.generation_data['reference_year']
        growth_rates = self.olade_config.get('demand_growth_rates', {})
        renewability_targets = self.olade_config.get('renewability_targets', {})

        # Read base_scenario from YAML for DemandBased override logic
        base_scenario = read_base_scenario()

        self.log(f"Reference year: {ref_year}")
        self.log(f"Update LowerLimit: {'YES' if update_lower else 'NO'}")
        self.log(f"Update UpperLimit: {'YES' if update_upper else 'NO'}")
        self.log(f"Base scenario: {base_scenario}")
        self.log(f"Renewability targets defined for: {len(renewability_targets)} country/scenario combinations")
        self.log("Calculation Method: LowerLimit = Demand × Normalized_Share")
        self.log("  - Reads demand from A-O_Demand.xlsx")
        self.log("  - Distributes shares using OLADE weights")
        self.log(f"  - Non-base scenarios: years 2023-2025 use {base_scenario} shares")
        self.log("")

        activity_changes = 0

        # DEBUG: Collect all shares for CSV export
        all_shares_data = {}  # {(scenario, country): {tech_type: {year: share}}}
        # Clear the DemandBased shares cache
        self.demand_based_shares_cache = {}

        for scenario in self.scenarios:
            param_path = self.base_path / f"A1_Outputs_{scenario}" / "A-O_Parametrization.xlsx"

            if not param_path.exists():
                self.log(f"  ✗ Parametrization file not found: {param_path}", "WARNING")
                continue

            self.log(f"Processing {scenario}...")

            # Create backup before modifying
            backup_path = self.create_backup(param_path)
            self.log(f"  Backup created: {backup_path.name}")

            try:
                wb = openpyxl.load_workbook(param_path)

                if 'Secondary Techs' not in wb.sheetnames:
                    wb.close()
                    self.log(f"  ✗ 'Secondary Techs' sheet not found", "ERROR")
                    continue

                ws = wb['Secondary Techs']

                # Store initial row count for validation
                initial_row_count = ws.max_row
                self.log(f"  Initial row count: {initial_row_count}")

                # Build year column map from headers
                year_col_map = {}
                projection_mode_col = None
                headers = [cell.value for cell in ws[1]]

                for col_idx, header in enumerate(headers, 1):
                    if header:
                        if str(header).isdigit():
                            try:
                                year = int(header)
                                if 2000 <= year <= 2100:
                                    year_col_map[year] = col_idx
                            except:
                                pass
                        elif str(header).strip() == "Projection.Mode":
                            projection_mode_col = col_idx

                # Build index of rows by tech and parameter
                lower_limit_rows = {}  # {tech_code: row_idx}
                upper_limit_rows = {}  # {tech_code: row_idx}
                max_capacity_rows = {}  # {tech_code: row_idx}
                residual_capacity_rows = {}  # {tech_code: row_idx}
                availability_rows = {}  # {tech_code: row_idx}

                for row_idx in range(2, ws.max_row + 1):
                    tech_code = ws.cell(row_idx, 2).value
                    parameter = ws.cell(row_idx, 5).value
                    if tech_code and parameter:
                        tech_str = str(tech_code).strip().upper()
                        param_str = str(parameter).strip()
                        if param_str == 'TotalTechnologyAnnualActivityLowerLimit':
                            lower_limit_rows[tech_str] = row_idx
                        elif param_str == 'TotalTechnologyAnnualActivityUpperLimit':
                            upper_limit_rows[tech_str] = row_idx
                        elif param_str == 'TotalAnnualMaxCapacity':
                            max_capacity_rows[tech_str] = row_idx
                        elif param_str == 'ResidualCapacity':
                            residual_capacity_rows[tech_str] = row_idx
                        elif param_str == 'AvailabilityFactor':
                            availability_rows[tech_str] = row_idx

                # Read CapacityToActivityUnit from Fixed Horizon Parameters
                capacity_to_activity = {}  # {tech_code: value}
                if 'Fixed Horizon Parameters' in wb.sheetnames:
                    ws_fixed = wb['Fixed Horizon Parameters']
                    for row_idx in range(2, ws_fixed.max_row + 1):
                        tech_code = ws_fixed.cell(row_idx, 2).value  # Column B: Tech
                        parameter = ws_fixed.cell(row_idx, 6).value  # Column F: Parameter
                        value = ws_fixed.cell(row_idx, 8).value      # Column H: Value
                        if tech_code and parameter and value:
                            tech_str = str(tech_code).strip().upper()
                            param_str = str(parameter).strip()
                            if param_str == 'CapacityToActivityUnit':
                                try:
                                    capacity_to_activity[tech_str] = float(value)
                                except (ValueError, TypeError):
                                    pass

                # Get unique countries from the technology codes
                countries_in_sheet = set()
                for tech_str in set(lower_limit_rows.keys()) | set(upper_limit_rows.keys()):
                    if tech_str.startswith('PWR') and len(tech_str) >= 9:
                        country_code = tech_str[6:9]
                        countries_in_sheet.add(country_code)

                # For DemandBased method: read demand data and prepare BAU shares for NDC override
                demand_data = {}
                bau_demand_data = {}
                bau_shares_cache = {}  # {country_code: {fuel: {year: share}}}

                if limit_method == 'DemandBased' and update_lower:
                    # Read demand data for this scenario
                    demand_path = self.base_path / f"A1_Outputs_{scenario}" / "A-O_Demand.xlsx"
                    if demand_path.exists():
                        try:
                            demand_data = read_demand_data(demand_path)
                            self.log(f"  Loaded demand data for {len(demand_data)} countries")
                        except Exception as e:
                            self.log(f"  ✗ Error reading demand data: {e}", "WARNING")
                    else:
                        self.log(f"  ✗ Demand file not found: {demand_path}", "WARNING")

                    # For non-base scenarios, also read base scenario demand for override
                    if scenario != base_scenario:
                        base_demand_path = self.base_path / f"A1_Outputs_{base_scenario}" / "A-O_Demand.xlsx"
                        if base_demand_path.exists():
                            try:
                                bau_demand_data = read_demand_data(base_demand_path)
                                self.log(f"  Loaded {base_scenario} demand data for override")
                            except Exception as e:
                                self.log(f"  ✗ Error reading {base_scenario} demand data: {e}", "WARNING")

                # Process each country
                for country_code in countries_in_sheet:
                    # Check if we have generation data for this country
                    if country_code not in self.generation_data['data']:
                        continue

                    base_generation_pj = self.generation_data['data'][country_code]
                    growth_rate = growth_rates.get(country_code, 0.02)

                    # Calculate technology shares using renewability targets
                    tech_shares = self.calculate_technology_shares(country_code, scenario, all_years)

                    if not tech_shares:
                        # Fall back to shares_total_data if available
                        if self.shares_total_data and scenario in self.shares_total_data:
                            tech_shares = self.shares_total_data[scenario].get(country_code, {})

                    if not tech_shares:
                        continue

                    # DEBUG: Store shares for CSV export
                    all_shares_data[(scenario, country_code)] = tech_shares

                    # Use DemandBased method: LowerLimit = Demand × Normalized_Share
                    demand_based_limits = {}

                    if update_lower:
                        # Calculate base scenario shares for override if needed (non-base scenarios)
                        base_shares = None
                        if scenario != base_scenario and country_code not in bau_shares_cache:
                            # Calculate base scenario shares for this country
                            base_limits = self.calculate_demand_based_limits(
                                base_scenario, country_code, all_years,
                                bau_demand_data if bau_demand_data else demand_data,
                                base_scenario=base_scenario  # No override for base scenario itself
                            )
                            # Convert from {tech_str: {year: limit}} to {fuel: {year: share}}
                            # We need the shares, not the limits, for the override
                            if base_limits and bau_demand_data:
                                base_country_demand = bau_demand_data.get(country_code, {})
                                base_shares_fuel = {}
                                for tech_str, year_limits in base_limits.items():
                                    # Extract fuel code from tech_str (PWRHYDARGXX -> HYD)
                                    if tech_str.startswith('PWR') and len(tech_str) >= 6:
                                        fuel = tech_str[3:6]
                                        base_shares_fuel[fuel] = {}
                                        for year, limit in year_limits.items():
                                            demand = base_country_demand.get(year, 0.0)
                                            if demand > 0:
                                                base_shares_fuel[fuel][year] = limit / demand
                                            else:
                                                base_shares_fuel[fuel][year] = 0.0
                                bau_shares_cache[country_code] = base_shares_fuel

                        base_shares_for_country = bau_shares_cache.get(country_code)
                        demand_based_limits = self.calculate_demand_based_limits(
                            scenario, country_code, all_years, demand_data,
                            base_scenario=base_scenario, base_shares=base_shares_for_country
                        )

                    # Update each technology
                    for tech_type, year_shares in tech_shares.items():
                        tech_str = f"PWR{tech_type}{country_code}XX"

                        lower_row = lower_limit_rows.get(tech_str)
                        upper_row = upper_limit_rows.get(tech_str)
                        max_cap_row = max_capacity_rows.get(tech_str)
                        residual_cap_row = residual_capacity_rows.get(tech_str)
                        avail_row = availability_rows.get(tech_str)

                        if not lower_row and not upper_row:
                            continue

                        # Get CapacityToActivityUnit for this tech (default 31.536 for power plants)
                        c2a = capacity_to_activity.get(tech_str, 31.536)

                        values_updated = 0
                        for year in all_years:
                            if year not in year_col_map:
                                continue

                            share = year_shares.get(year, 0.0)

                            # Calculate limit_value using DemandBased method
                            # IMPORTANT: Always calculate and write, even if share=0, to clear old values
                            if tech_str in demand_based_limits:
                                limit_value = demand_based_limits[tech_str].get(year)
                                if limit_value is None:
                                    # If no value in pre-calculated limits, use 0.0
                                    limit_value = 0.0
                            else:
                                # Fallback: use 0.0 if tech not in demand_based_limits
                                limit_value = 0.0

                            # Capacity validation now handled by Universal Validation
                            # No pre-capping needed - Universal Validation will increase MaxCapacity if needed
                            limit_value = round(limit_value, 4)

                            # Update LowerLimit - ALWAYS write, even if 0, to clear old values
                            if update_lower and lower_row:
                                ws.cell(lower_row, year_col_map[year], limit_value)
                                activity_changes += 1
                                values_updated += 1

                            # Update UpperLimit = LowerLimit × 1.05 + 0.1 (proportional + fixed margin)
                            if update_upper and upper_row:
                                upper_value = round(limit_value * 1.05 + 0.1, 4)
                                ws.cell(upper_row, year_col_map[year], upper_value)
                                activity_changes += 1

                        # Set Projection.Mode to "User defined"
                        if values_updated > 0 and projection_mode_col:
                            if update_lower and lower_row:
                                ws.cell(lower_row, projection_mode_col, "User defined")
                            if update_upper and upper_row:
                                ws.cell(upper_row, projection_mode_col, "User defined")
                            self.log(f"  {country_code}-{tech_type}: Updated {values_updated} years")

                # ============================================================
                # UNIVERSAL VALIDATION (UNIFIED FOR ALL METHODS)
                # Check power technologies with LowerLimit
                # When LowerLimit > MaxPossibleActivity: INCREASE MaxCapacity instead of capping limit
                # Applies to the 10 main PWR generation technologies from OLADE:
                # URN, NGS, COA, HYD, GEO, WON, SPV, BIO, PET, OIL
                # Excludes storage (PWRSDS, PWRLDS), backup (PWRBCK), and transmission (TRN)
                # ============================================================
                capacity_increases = 0

                # Only validate these specific technology types
                validated_tech_types = ['URN', 'NGS', 'COA', 'HYD', 'GEO', 'WON', 'SPV', 'BIO', 'PET', 'OIL']

                # Log validation start
                if lower_limit_rows:
                    self.log(f"  Running Universal Validation: checking capacity constraints for {len(lower_limit_rows)} technologies...")
                else:
                    self.log("  No technologies to validate")

                for tech_str, lower_row in lower_limit_rows.items():
                    # Only process PWR technologies (power generation)
                    if not tech_str.startswith('PWR'):
                        continue

                    # Only validate the 10 main generation technologies
                    # Format: PWR{TYPE}{COUNTRY}XX (e.g., PWRHYDARGXX)
                    if len(tech_str) >= 6:
                        tech_type = tech_str[3:6]  # e.g., 'HYD', 'SPV', 'BIO'
                        if tech_type not in validated_tech_types:
                            # Skip storage (SDS, LDS), backup (BCK), and any other tech
                            continue
                    else:
                        continue

                    for year in all_years:
                        if year not in year_col_map:
                            continue

                        col_idx = year_col_map[year]
                        current_limit = ws.cell(lower_row, col_idx).value
                        if current_limit is None:
                            continue
                        try:
                            current_limit = float(current_limit)
                        except (ValueError, TypeError):
                            continue

                        if current_limit <= 0:
                            continue

                        # Calculate max possible activity using new helper function
                        max_possible, source, components = self.calculate_max_possible_activity(
                            wb, ws, tech_str, year, year_col_map, capacity_to_activity,
                            max_capacity_rows, residual_capacity_rows, availability_rows
                        )

                        # Skip validation if max_possible is infinite (TotalAnnualMaxCapacity = -1, no limit)
                        if max_possible is not None and not (isinstance(max_possible, float) and max_possible == float('inf')):
                            # Use >= with 5% margin to force re-correction when MaxCapacity is "almost" sufficient
                            # This prevents issues where rounding during CSV export causes constraint violations
                            # Example: MaxCap=0.123714 gives max_possible=3.5113002, but rounds to 0.1237 -> 3.5109 < 3.5113
                            if current_limit >= max_possible * 0.95:
                                # UNIFIED BEHAVIOR: Always increase MaxCapacity to ensure 5% safety margin
                                new_cap, max_adjusted, res_adjusted = self.increase_max_capacity_for_limit(
                                    ws, tech_str, year, year_col_map, current_limit,
                                    components, max_capacity_rows, residual_capacity_rows
                                )

                                if max_adjusted:
                                    capacity_increases += 1

                                    # Recalculate UpperLimit with new formula: LowerLimit × 1.05 + 0.1
                                    upper_row = upper_limit_rows.get(tech_str)
                                    if upper_row:
                                        new_upper = current_limit * 1.05 + 0.1
                                        ws.cell(upper_row, col_idx, round(new_upper, 4))

                        # DEBUG: Special validation for PWRGEOHNDXX, year 2038, scenario NDC
                        if scenario == "NDC" and year == 2038 and tech_str == "PWRGEOHNDXX":
                            self.log(f"")
                            self.log(f"=" * 80)
                            self.log(f"DEBUG: Special validation for PWRGEOHNDXX in NDC 2038")
                            self.log(f"=" * 80)
                            self.log(f"Technology: {tech_str}")
                            self.log(f"Year: {year}")
                            self.log(f"Scenario: {scenario}")
                            self.log(f"")
                            self.log(f"C2A Diagnostic:")
                            c2a_value = components.get('c2a', 0)
                            if tech_str in capacity_to_activity:
                                self.log(f"  - C2A source: Found in Fixed Horizon Parameters")
                                self.log(f"  - C2A value from dict: {capacity_to_activity[tech_str]:.6f}")
                            else:
                                self.log(f"  - C2A source: NOT FOUND in Fixed Horizon Parameters (using default)")
                                self.log(f"  - Available technologies in C2A dict: {len(capacity_to_activity)}")
                                self.log(f"  - Sample C2A keys: {list(capacity_to_activity.keys())[:5]}")
                            self.log(f"  - C2A used in calculation: {c2a_value:.6f}")
                            self.log(f"")
                            self.log(f"Formula Components:")
                            self.log(f"  - CapacityFactor × YearSplit sum: {components.get('cf_yearsplit_sum', 0):.6f}")
                            self.log(f"  - TotalAnnualMaxCapacity: {components.get('max_cap', 0):.6f} GW")
                            self.log(f"  - AvailabilityFactor: {components.get('avail', 0):.6f}")
                            self.log(f"  - CapacityToActivityUnit: {c2a_value:.6f}")
                            self.log(f"")
                            self.log(f"Calculation:")
                            manual_calc = components.get('max_cap', 0) * c2a_value * components.get('avail', 0) * components.get('cf_yearsplit_sum', 0)
                            self.log(f"  - Formula: MaxCap × C2A × Avail × CF_sum")
                            self.log(f"  - Manual: {components.get('max_cap', 0):.6f} × {c2a_value:.6f} × {components.get('avail', 0):.6f} × {components.get('cf_yearsplit_sum', 0):.6f}")
                            self.log(f"  - Manual result: {manual_calc:.6f} PJ")
                            self.log(f"")
                            self.log(f"Constraint Check:")
                            self.log(f"  - MaxPossibleActivity: {max_possible:.6f} PJ")
                            self.log(f"  - TotalTechnologyAnnualActivityLowerLimit: {current_limit:.6f} PJ")
                            self.log(f"")

                            # # Check constraint: max_possible >= current_limit
                            # if max_possible >= current_limit:
                            #     self.log(f"✓ Constraint SATISFIED: {max_possible:.6f} >= {current_limit:.6f}")
                            #     self.log(f"Exiting with code 2 (constraint satisfied)")
                            #     wb.close()
                            #     sys.exit(2)
                            # else:
                            #     self.log(f"✗ Constraint VIOLATED: {max_possible:.6f} < {current_limit:.6f}")
                            #     deficit = current_limit - max_possible
                            #     required_cap = current_limit / (c2a_value * components.get('avail', 0) * components.get('cf_yearsplit_sum', 0))
                            #     self.log(f"  - Deficit: {deficit:.6f} PJ")
                            #     self.log(f"  - Required MaxCapacity to meet constraint: {required_cap:.6f} GW")
                            #     self.log(f"Exiting with code 3 (constraint violated)")
                            #     wb.close()
                            #     sys.exit(3)

                # Final summary
                if capacity_increases > 0:
                    self.log(f"  ✓ Universal Validation: Increased MaxCapacity for {capacity_increases} technology-year combinations")
                else:
                    self.log(f"  ✓ Universal Validation: No capacity adjustments needed")

                # Validate row count before saving
                final_row_count = ws.max_row
                if final_row_count != initial_row_count:
                    self.log(f"  ⚠ WARNING: Row count changed from {initial_row_count} to {final_row_count}!", "WARNING")

                # Save
                wb.save(param_path)
                wb.close()
                self.log(f"  ✓ {scenario} activity limits updated (rows: {final_row_count})")

            except Exception as e:
                self.log(f"  ✗ Error updating {scenario}: {e}", "ERROR")
                import traceback
                traceback.print_exc()
                try:
                    wb.close()
                except:
                    pass

        self.log("")
        self.log(f"Activity limits updates completed: {activity_changes} values written")

        # DEBUG: Export shares to CSV
        self.log("")
        self.log("=" * 80)
        self.log("DEBUG: SHARES EXPORT")
        self.log("=" * 80)
        self.log(f"Limit method: {limit_method}")
        self.log(f"DemandBased shares cache size: {len(self.demand_based_shares_cache)}")
        self.log(f"All shares data size: {len(all_shares_data)}")

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        # Decide which shares to export based on method
        if limit_method == 'DemandBased' and self.demand_based_shares_cache:
            # Export DemandBased normalized shares
            csv_path = self.base_path / f"DEBUG_DemandBased_Normalized_Shares_{timestamp}.csv"
            self.log(f"Exporting DemandBased shares to: {csv_path}")
            self.export_shares_to_csv(self.demand_based_shares_cache, all_years, csv_path)
            self.log(f"✓ DemandBased normalized shares exported to {csv_path.name}")
        elif all_shares_data:
            # Export CapacityBased/ShareBased shares
            csv_path = self.base_path / f"DEBUG_Technology_Shares_{timestamp}.csv"
            self.log(f"Exporting Technology shares to: {csv_path}")
            self.export_shares_to_csv(all_shares_data, all_years, csv_path)
            self.log(f"✓ Technology shares exported to {csv_path.name}")
        else:
            self.log("⚠ No shares data to export!", "WARNING")
            if limit_method != 'DemandBased':
                self.log(f"  Reason: limit_method is '{limit_method}' (expected 'DemandBased')", "WARNING")
            if not self.demand_based_shares_cache and not all_shares_data:
                self.log(f"  Reason: Both shares caches are empty", "WARNING")

    def update_activity_lower_limit(self, all_years):
        """Legacy wrapper - calls update_activity_limits"""
        self.update_activity_limits(all_years)

    def read_interconnections_config(self):
        """
        Read the Interconnections sheet from Secondary_Techs_Editor.xlsx.

        Returns a list of dicts: [{tech, mode, status}, ...]
        where status is 'ON' or 'OFF' and mode is 1 or 2.
        """
        try:
            wb = openpyxl.load_workbook(self.editor_path, data_only=True)
        except Exception as e:
            self.log(f"ERROR: Could not open editor file for interconnections: {e}", "ERROR")
            return []

        if 'Interconnections' not in wb.sheetnames:
            self.log("WARNING: 'Interconnections' sheet not found in editor file", "WARNING")
            wb.close()
            return []

        ws = wb['Interconnections']
        config = []
        # Data starts at row 5: Col A=Tech, Col C=Mode, Col D=Status
        for row_idx in range(5, ws.max_row + 1):
            tech = ws.cell(row_idx, 1).value
            mode = ws.cell(row_idx, 3).value
            status = ws.cell(row_idx, 4).value

            if tech is None or mode is None or status is None:
                continue

            tech = str(tech).strip()
            try:
                mode = int(mode)
            except (ValueError, TypeError):
                continue
            status = str(status).strip().upper()

            if status in ('ON', 'OFF'):
                config.append({'tech': tech, 'mode': mode, 'status': status})

        wb.close()
        self.log(f"Read {len(config)} interconnection entries from editor")
        on_count = sum(1 for c in config if c['status'] == 'ON')
        off_count = sum(1 for c in config if c['status'] == 'OFF')
        self.log(f"  ON: {on_count}, OFF: {off_count}")
        return config

    def apply_interconnections(self, interconnections_config):
        """
        Apply interconnection ON/OFF controls to AR files for all scenarios.

        For OFF: Base Year Value.Fuel.I=0, Value.Fuel.O=0; Projections Projection.Mode='EMPTY'
        For ON:  Base Year Value.Fuel.I=1, Value.Fuel.O=1; Projections Projection.Mode='User defined'

        Args:
            interconnections_config: list of {tech, mode, status} dicts
        """
        self.log("")
        self.log("=" * 80)
        self.log("APPLYING INTERCONNECTION CONTROLS")
        self.log("=" * 80)

        # Build lookup: (tech, mode) -> status
        status_lookup = {}
        for entry in interconnections_config:
            key = (entry['tech'], entry['mode'])
            status_lookup[key] = entry['status']

        off_entries = [(t, m) for (t, m), s in status_lookup.items() if s == 'OFF']
        if off_entries:
            self.log(f"Interconnections to disable: {len(off_entries)}")
            for tech, mode in off_entries:
                self.log(f"  OFF: {tech} Mode {mode}")
        else:
            self.log("All interconnections are ON - no changes needed")
            return

        total_base_year_changes = 0
        total_projections_changes = 0

        for scenario in self.scenarios:
            scenario_path = self.base_path / f"A1_Outputs_{scenario}"
            if not scenario_path.exists():
                continue

            self.log(f"\n--- Scenario: {scenario} ---")

            # ===== Part A: A-O_AR_Model_Base_Year.xlsx =====
            base_year_path = scenario_path / "A-O_AR_Model_Base_Year.xlsx"
            if base_year_path.exists():
                try:
                    wb_by = openpyxl.load_workbook(base_year_path)
                    if 'Secondary' in wb_by.sheetnames:
                        ws_by = wb_by['Secondary']

                        # Find column indices by header name (row 1)
                        headers_by = {}
                        for col_idx in range(1, ws_by.max_column + 1):
                            val = ws_by.cell(1, col_idx).value
                            if val:
                                headers_by[str(val).strip()] = col_idx

                        col_mode = headers_by.get('Mode.Operation')
                        col_tech = headers_by.get('Tech')
                        col_val_fi = headers_by.get('Value.Fuel.I')
                        col_val_fo = headers_by.get('Value.Fuel.O')

                        if not all([col_mode, col_tech, col_val_fi, col_val_fo]):
                            self.log(f"  WARNING: Missing required columns in Base Year Secondary sheet", "WARNING")
                            self.log(f"  Found headers: {list(headers_by.keys())}", "WARNING")
                        else:
                            base_year_changes = 0
                            for row_idx in range(2, ws_by.max_row + 1):
                                tech = ws_by.cell(row_idx, col_tech).value
                                mode = ws_by.cell(row_idx, col_mode).value
                                if tech is None or mode is None:
                                    continue
                                tech = str(tech).strip()
                                try:
                                    mode = int(mode)
                                except (ValueError, TypeError):
                                    continue

                                key = (tech, mode)
                                if key in status_lookup:
                                    status = status_lookup[key]
                                    if status == 'OFF':
                                        ws_by.cell(row_idx, col_val_fi).value = 0
                                        ws_by.cell(row_idx, col_val_fo).value = 0
                                        base_year_changes += 1
                                    else:  # ON
                                        ws_by.cell(row_idx, col_val_fi).value = 1
                                        ws_by.cell(row_idx, col_val_fo).value = 1
                                        base_year_changes += 1

                            if base_year_changes > 0:
                                wb_by.save(base_year_path)
                                self.log(f"  Base Year: {base_year_changes} rows updated")
                                total_base_year_changes += base_year_changes
                            else:
                                self.log(f"  Base Year: no matching rows found")
                    else:
                        self.log(f"  WARNING: 'Secondary' sheet not found in Base Year file", "WARNING")
                    wb_by.close()
                except Exception as e:
                    self.log(f"  ERROR processing Base Year for {scenario}: {e}", "ERROR")
            else:
                self.log(f"  Base Year file not found: {base_year_path}")

            # ===== Part B: A-O_AR_Projections.xlsx =====
            proj_path = scenario_path / "A-O_AR_Projections.xlsx"
            if proj_path.exists():
                try:
                    wb_proj = openpyxl.load_workbook(proj_path)
                    if 'Secondary' in wb_proj.sheetnames:
                        ws_proj = wb_proj['Secondary']

                        # Find column indices by header name (row 1)
                        headers_proj = {}
                        for col_idx in range(1, ws_proj.max_column + 1):
                            val = ws_proj.cell(1, col_idx).value
                            if val:
                                headers_proj[str(val).strip()] = col_idx

                        col_mode_p = headers_proj.get('Mode.Operation')
                        col_tech_p = headers_proj.get('Tech')
                        col_proj_mode = headers_proj.get('Projection.Mode')

                        if not all([col_mode_p, col_tech_p, col_proj_mode]):
                            self.log(f"  WARNING: Missing required columns in Projections Secondary sheet", "WARNING")
                            self.log(f"  Found headers: {list(headers_proj.keys())}", "WARNING")
                        else:
                            proj_changes = 0
                            for row_idx in range(2, ws_proj.max_row + 1):
                                tech = ws_proj.cell(row_idx, col_tech_p).value
                                mode = ws_proj.cell(row_idx, col_mode_p).value
                                if tech is None or mode is None:
                                    continue
                                tech = str(tech).strip()
                                try:
                                    mode = int(mode)
                                except (ValueError, TypeError):
                                    continue

                                key = (tech, mode)
                                if key in status_lookup:
                                    status = status_lookup[key]
                                    if status == 'OFF':
                                        ws_proj.cell(row_idx, col_proj_mode).value = 'EMPTY'
                                        proj_changes += 1
                                    else:  # ON
                                        ws_proj.cell(row_idx, col_proj_mode).value = 'User defined'
                                        proj_changes += 1

                            if proj_changes > 0:
                                wb_proj.save(proj_path)
                                self.log(f"  Projections: {proj_changes} rows updated")
                                total_projections_changes += proj_changes
                            else:
                                self.log(f"  Projections: no matching rows found")
                    else:
                        self.log(f"  WARNING: 'Secondary' sheet not found in Projections file", "WARNING")
                    wb_proj.close()
                except Exception as e:
                    self.log(f"  ERROR processing Projections for {scenario}: {e}", "ERROR")
            else:
                self.log(f"  Projections file not found: {proj_path}")

        self.log(f"\nInterconnection controls completed:")
        self.log(f"  Base Year rows updated: {total_base_year_changes}")
        self.log(f"  Projections rows updated: {total_projections_changes}")

    def run(self):
        """Main execution"""
        self.log("=" * 80)
        self.log("SECONDARY TECHS UPDATER")
        self.log("=" * 80)
        self.log(f"Start time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        self.log(f"Editor file: {self.editor_path}")
        self.log("")

        try:
            # Read OLADE configuration
            self.olade_config = read_olade_config(self.editor_path)

            if self.olade_config['enabled']:
                self.log("OLADE integration: ENABLED")
                if self.olade_file_path and self.olade_file_path.exists():
                    self.log(f"OLADE file: {self.olade_file_path}")
                    try:
                        self.olade_data = read_olade_data(self.olade_file_path)
                        self.log(f"OLADE data loaded: {len(self.olade_data['data'])} countries")

                        # Load Shares data for petroleum split
                        if self.shares_file_path and self.shares_file_path.exists():
                            self.log(f"Shares file: {self.shares_file_path}")
                            try:
                                self.shares_data = read_shares_data(self.shares_file_path)
                                self.log(f"Shares data loaded: {len(self.shares_data)} scenarios")
                            except Exception as e:
                                self.log(f"ERROR loading Shares data: {e}", "ERROR")
                                self.log("Petroleum will be split 50/50 between PET and OIL", "WARNING")
                                self.shares_data = None
                        else:
                            self.log(f"WARNING: Shares file not found: {self.shares_file_path}", "WARNING")
                            self.log("Petroleum will be split 50/50 between PET and OIL", "WARNING")
                            self.shares_data = None

                    except Exception as e:
                        self.log(f"ERROR loading OLADE data: {e}", "ERROR")
                        self.log("Continuing without OLADE data...", "WARNING")
                        self.olade_config['enabled'] = False
                else:
                    self.log(f"WARNING: OLADE file not found: {self.olade_file_path}", "WARNING")
                    self.log("Continuing without OLADE data...", "WARNING")
                    self.olade_config['enabled'] = False
            else:
                self.log("OLADE integration: DISABLED")

            # Check Demand integration (also needed for ActivityLowerLimit)
            demand_or_activity_enabled = (self.olade_config.get('demand_enabled') or
                                          self.olade_config.get('activity_lower_limit_enabled'))
            if demand_or_activity_enabled:
                self.log("")
                self.log("Demand/ActivityLowerLimit integration: ENABLED")
                if self.generation_file_path and self.generation_file_path.exists():
                    self.log(f"Generation file: {self.generation_file_path}")
                    try:
                        self.generation_data = read_olade_generation_data(self.generation_file_path)
                        self.log(f"Generation data loaded: {len(self.generation_data['data'])} countries")
                        self.log(f"Growth rates configured: {len(self.olade_config.get('demand_growth_rates', {}))} countries")
                    except Exception as e:
                        self.log(f"ERROR loading generation data: {e}", "ERROR")
                        self.log("Continuing without demand/activity update...", "WARNING")
                        self.olade_config['demand_enabled'] = False
                        self.olade_config['activity_lower_limit_enabled'] = False
                else:
                    self.log(f"WARNING: Generation file not found: {self.generation_file_path}", "WARNING")
                    self.log("Continuing without demand/activity update...", "WARNING")
                    self.olade_config['demand_enabled'] = False
                    self.olade_config['activity_lower_limit_enabled'] = False
            else:
                self.log("")
                self.log("Demand integration: DISABLED")

            # Check Activity Limits integration - needs Shares_Power_Generation_Technologies.xlsx
            lower_enabled = self.olade_config.get('activity_lower_limit_enabled')
            upper_enabled = self.olade_config.get('activity_upper_limit_enabled')

            if lower_enabled or upper_enabled:
                self.log("")
                self.log(f"ActivityLowerLimit integration: {'ENABLED' if lower_enabled else 'DISABLED'}")
                self.log(f"ActivityUpperLimit integration: {'ENABLED' if upper_enabled else 'DISABLED'}")

                # Log renewability targets info
                renewability_targets = self.olade_config.get('renewability_targets', {})
                technology_weights = self.olade_config.get('technology_weights', {})
                if renewability_targets:
                    self.log(f"Renewability targets: {len(renewability_targets)} country/scenario combinations")
                if technology_weights:
                    self.log(f"Custom technology weights: {len(technology_weights)} country/scenario combinations")

                if self.shares_total_file_path and self.shares_total_file_path.exists():
                    self.log(f"Shares_Total file: {self.shares_total_file_path}")
                    try:
                        self.shares_total_data = read_shares_total_data(self.shares_total_file_path)
                        self.log(f"Shares_Total data loaded: {len(self.shares_total_data)} scenarios")
                    except Exception as e:
                        self.log(f"ERROR loading Shares_Total data: {e}", "ERROR")
                        self.log("Continuing without activity limits update...", "WARNING")
                        self.olade_config['activity_lower_limit_enabled'] = False
                        self.olade_config['activity_upper_limit_enabled'] = False
                else:
                    self.log(f"WARNING: Shares_Total file not found: {self.shares_total_file_path}", "WARNING")
                    self.log("Continuing without activity limits update...", "WARNING")
                    self.olade_config['activity_lower_limit_enabled'] = False
                    self.olade_config['activity_upper_limit_enabled'] = False
            else:
                self.log("")
                self.log("Activity Limits integration: DISABLED")

            # Check Trade Balance Demand Adjustment
            if self.olade_config.get('trade_balance_enabled'):
                self.log("")
                self.log("Trade Balance Demand Adjustment: ENABLED")
                if self.trade_balance_file_path and self.trade_balance_file_path.exists():
                    self.log(f"Trade balance file: {self.trade_balance_file_path}")
                    try:
                        self.trade_balance_data = read_trade_balance_data(self.trade_balance_file_path)
                        countries_with_data = len(self.trade_balance_data)
                        self.log(f"Trade balance data loaded: {countries_with_data} countries")
                    except Exception as e:
                        self.log(f"ERROR loading trade balance data: {e}", "ERROR")
                        self.log("Continuing without trade balance adjustment...", "WARNING")
                        self.olade_config['trade_balance_enabled'] = False
                else:
                    self.log(f"WARNING: Trade balance file not found: {self.trade_balance_file_path}", "WARNING")
                    self.log("Continuing without trade balance adjustment...", "WARNING")
                    self.olade_config['trade_balance_enabled'] = False
            else:
                self.log("")
                self.log("Trade Balance Demand Adjustment: DISABLED")

            # Check Interconnections Control
            if self.olade_config.get('interconnections_enabled'):
                self.log("")
                self.log("Interconnections Control: ENABLED")
            else:
                self.log("")
                self.log("Interconnections Control: DISABLED")

            self.log("")

            # Read editor file
            instructions = self.read_editor_file()

            # Get all years from first available scenario to determine year range
            all_years = set()
            for scenario in self.scenarios:
                scenario_path = self.base_path / f"A1_Outputs_{scenario}" / "A-O_Parametrization.xlsx"
                if scenario_path.exists():
                    self.log(f"Reading year columns from {scenario} scenario")
                    wb = openpyxl.load_workbook(scenario_path, data_only=True)
                    if 'Secondary Techs' in wb.sheetnames:
                        ws = wb['Secondary Techs']
                        headers = [cell.value for cell in ws[1]]
                        for header in headers:
                            if header and str(header).isdigit():
                                try:
                                    year = int(header)
                                    if 2000 <= year <= 2100:
                                        all_years.add(year)
                                except:
                                    pass
                    wb.close()
                    break  # Found a valid scenario, stop searching

            if not all_years:
                self.log("WARNING: Could not find any scenario with year columns!", "WARNING")

            # Generate OLADE instructions if enabled
            olade_instructions = []
            if self.olade_config['enabled'] and self.olade_data:
                olade_instructions = self.generate_olade_instructions(sorted(all_years))

            # Combine instructions: OLADE takes priority for ResidualCapacity
            # Filter out manual ResidualCapacity instructions for PWR techs if OLADE is enabled
            if olade_instructions:
                filtered_instructions = []
                for instr in instructions:
                    # Check if this is a ResidualCapacity instruction for a PWR tech
                    if (instr.get('parameter') == 'ResidualCapacity' and
                        instr.get('tech') and str(instr.get('tech')).upper().startswith('PWR')):
                        # Skip it - OLADE will handle it
                        self.log(f"Skipping manual ResidualCapacity for {instr['tech']} - using OLADE data", "DEBUG")
                        continue
                    filtered_instructions.append(instr)
                instructions = filtered_instructions + olade_instructions
            else:
                # No OLADE, use manual instructions as-is
                pass

            # Apply instructions in batch mode (one file open per scenario) if any
            if instructions:
                self.apply_instructions_batch(instructions)
            else:
                self.log("No manual instructions to process.")

            # Update demand files if enabled
            if self.olade_config.get('demand_enabled') and self.generation_data:
                self.update_demand_files(sorted(all_years))

            # Update activity limits if enabled (LowerLimit and/or UpperLimit)
            if (self.olade_config.get('activity_lower_limit_enabled') or
                self.olade_config.get('activity_upper_limit_enabled')) and self.generation_data:
                self.update_activity_limits(sorted(all_years))

            # Adjust demand by trade balance if enabled (runs LAST - must not affect activity limits)
            if self.olade_config.get('trade_balance_enabled') and self.trade_balance_data:
                self.update_demand_with_trade_balance(sorted(all_years))

            # Apply interconnection controls if enabled
            if self.olade_config.get('interconnections_enabled'):
                interconnections_config = self.read_interconnections_config()
                if interconnections_config:
                    self.apply_interconnections(interconnections_config)

            # Summary
            self.log("")
            self.log("=" * 80)
            self.log("SUMMARY")
            self.log("=" * 80)
            self.log(f"Total instructions processed: {len(instructions)}")
            self.log(f"Changes applied: {self.changes_applied}")
            self.log(f"Rows failed: {self.rows_failed}")

            # Save log
            log_path = self.base_path / f"secondary_techs_update_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            self.save_log(log_path)
            self.log(f"\nLog saved: {log_path}")

            if self.rows_failed > 0:
                self.log("\n⚠ Some rows failed. Please review the log.", "WARNING")
                return 1
            else:
                self.log("\n✓ All changes applied successfully!", "SUCCESS")
                return 0

        except Exception as e:
            self.log(f"\nFATAL ERROR: {e}", "ERROR")
            import traceback
            traceback.print_exc()
            return 1


def main():
    try:
        # Paths
        script_dir = Path(__file__).parent
        editor_path = script_dir / "Secondary_Techs_Editor.xlsx"
        base_path = script_dir / "A1_Outputs"
        olade_file_path = script_dir / "OLADE - Capacidad instalada por fuente - Anual.xlsx"
        shares_file_path = script_dir / "Shares_PET_OIL_Split.xlsx"
        generation_file_path = script_dir / "OLADE - Generación eléctrica por fuente - Anual.xlsx"
        shares_total_file_path = script_dir / "Shares_Power_Generation_Technologies.xlsx"
        trade_balance_file_path = script_dir / "Matriz Balance energético" / "flujos_energia_estimados_optimizacion.xlsx"

        # Create updater and run
        updater = SecondaryTechsUpdater(editor_path, base_path, olade_file_path, shares_file_path, generation_file_path, shares_total_file_path, trade_balance_file_path)
        return updater.run()

    except Exception as e:
        print(f"\nERROR: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
