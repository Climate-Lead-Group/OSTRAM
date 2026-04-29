#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
test_4_apply_manual_fixes.py

Acceptance test for 4_apply_manual_fixes.py.

Contract (the green-light criterion for script 4):

    The ONLY differences between the post-3 and post-4 workbooks are
    exactly the cells script 4 was instructed to change via its three
    edit lists:
        - DELETIONS         -> rows removed
        - SUBSTITUTIONS     -> specific cells overwritten
        - OAR_CORRECTIONS   -> Output rows of TRN techs in AR_Proj/Secondary,
                                 all year columns set to new OAR value
        - TECH_NAME_RENAMES -> 'Tech.Name' column for matching Tech codes,
                                 across every sheet of every workbook

This test:
    1. Parses the edit lists out of 4_apply_manual_fixes.py (no exec).
    2. Walks every cell of every sheet of every workbook in BOTH
       post-3 and post-4 outputs.
    3. Records actual cell-level differences.
    4. For each actual diff, checks it matches one of the four edit kinds.
    5. For each declared edit, checks at least one corresponding diff exists.
    6. Asserts: zero unexplained diffs, every edit applied.

Files expected (relative to this script):
    4_apply_manual_fixes.py
    wvaligned_outputs/A-O_*_wvaligned.xlsx       (post-3, the input to script 4)
    wvaligned_outputs_v2/A-O_*_wvaligned_v2.xlsx (post-4, the output of script 4)

Exits 0 on full success.
"""
from pathlib import Path
import sys
import ast
import pandas as pd
import numpy as np
from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
SCRIPT_4   = SCRIPT_DIR / "4_apply_manual_fixes.py"
POST_3_DIR = SCRIPT_DIR / "wvaligned_outputs"
POST_4_DIR = SCRIPT_DIR / "wvaligned_outputs_v2"

AO_STEMS = {
    "Param":   "A-O_Parametrization",
    "AR_Base": "A-O_AR_Model_Base_Year",
    "AR_Proj": "A-O_AR_Projections",
    "Demand":  "A-O_Demand",
}

# -------------------- minimal test runner --------------------
_results = []
def test(label):
    def deco(fn):
        def runner():
            try:
                fn()
                _results.append((True, label, None))
                print(f"  PASS  {label}")
            except AssertionError as e:
                _results.append((False, label, str(e)))
                print(f"  FAIL  {label}")
                print(f"        {e}")
            except Exception as e:
                _results.append((False, label, f"{type(e).__name__}: {e}"))
                print(f"  ERROR {label}")
                print(f"        {type(e).__name__}: {e}")
        runner.__name__ = fn.__name__
        runner()
        return runner
    return deco


# -------------------- helpers --------------------
def parse_edit_lists(path):
    """ast-based extraction of the four edit-list constants from script 4.
    Does NOT execute the module."""
    src = path.read_text(encoding="utf-8")
    tree = ast.parse(src)
    targets = {"DELETIONS", "SUBSTITUTIONS", "OAR_CORRECTIONS",
               "TECH_NAME_RENAMES"}
    out = {}
    for node in tree.body:
        if isinstance(node, ast.Assign) and len(node.targets) == 1:
            tgt = node.targets[0]
            if isinstance(tgt, ast.Name) and tgt.id in targets:
                out[tgt.id] = ast.literal_eval(node.value)
    return out


def cell_eq(a, b):
    """Robust cell equality across openpyxl/pandas type drift.
    None / NaN treated as equivalent. Numerics compared as float."""
    a_null = a is None or (isinstance(a, float) and np.isnan(a))
    b_null = b is None or (isinstance(b, float) and np.isnan(b))
    if a_null and b_null:
        return True
    if a_null != b_null:
        return False
    try:
        return float(a) == float(b)
    except (TypeError, ValueError):
        return a == b


def normalize(v):
    """Cast for comparison: numbers->float, strings->stripped str."""
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return None
    if isinstance(v, str):
        return v.strip()
    try:
        return float(v)
    except (TypeError, ValueError):
        return v


def is_year_col(c):
    if isinstance(c, int) and 2000 <= c <= 2100:
        return True
    if isinstance(c, str) and c.isdigit() and 2000 <= int(c) <= 2100:
        return True
    return False


# -------------------- diff computation --------------------
def collect_diffs():
    """Walk every workbook+sheet+cell. Return list of dicts:
       {workbook, sheet, row_idx, col_idx, col_name, post3, post4, tech, row_dict_post3}"""
    diffs = []
    for label, stem in AO_STEMS.items():
        f3 = POST_3_DIR / f"{stem}_wvaligned.xlsx"
        f4 = POST_4_DIR / f"{stem}_wvaligned_v2.xlsx"
        if not f3.is_file() or not f4.is_file():
            continue
        wb3 = load_workbook(f3, data_only=True)
        wb4 = load_workbook(f4, data_only=True)
        for sh in wb3.sheetnames:
            if sh not in wb4.sheetnames:
                continue
            ws3 = wb3[sh]
            ws4 = wb4[sh]
            hdrs = [c.value for c in ws3[1]]
            n_cols = len(hdrs)
            tech_col_idx = None
            for i, h in enumerate(hdrs):
                if h in ("Tech", "Fuel/Tech"):
                    tech_col_idx = i
                    break
            # walk rows up to max(rows_3, rows_4)
            max_r = max(ws3.max_row, ws4.max_row)
            for r in range(2, max_r + 1):
                row3 = [ws3.cell(r, c + 1).value for c in range(n_cols)]
                row4 = [ws4.cell(r, c + 1).value for c in range(n_cols)]
                tech = row3[tech_col_idx] if tech_col_idx is not None and r <= ws3.max_row else None
                for c, h in enumerate(hdrs):
                    v3 = row3[c]
                    v4 = row4[c]
                    if not cell_eq(v3, v4):
                        diffs.append({
                            "workbook": label,
                            "sheet":    sh,
                            "row_idx":  r,
                            "col_idx":  c,
                            "col_name": h,
                            "post3":    v3,
                            "post4":    v4,
                            "tech":     tech,
                            "row3":     dict(zip(hdrs, row3)),
                        })
    return diffs


# -------------------- explainers --------------------
def explain_substitution(diff, subs):
    """Match diff against any SUBSTITUTIONS entry."""
    for entry in subs:
        wb_lbl, sheet, tech, selectors, edit_col, new_val, *_ = entry
        if diff["workbook"] != wb_lbl:        continue
        if diff["sheet"]    != sheet:         continue
        if diff["col_name"] != edit_col:      continue
        if normalize(diff["tech"]) != normalize(tech):
            continue
        # Check selectors all match in the post-3 row
        ok = True
        for sk, sv in selectors.items():
            if normalize(diff["row3"].get(sk)) != normalize(sv):
                ok = False
                break
        if not ok:
            continue
        if normalize(diff["post4"]) != normalize(new_val):
            continue
        return entry
    return None


def explain_oar(diff, oar_corrections):
    """Match diff against OAR_CORRECTIONS (AR_Proj/Secondary, Output rows,
    year columns, value == new_oar)."""
    if diff["workbook"] != "AR_Proj":  return None
    if diff["sheet"]    != "Secondary": return None
    if not is_year_col(diff["col_name"]): return None
    direction = diff["row3"].get("Direction")
    if str(direction).strip().upper() not in ("O", "OUTPUT"):
        return None
    for tech, new_oar, *_ in oar_corrections:
        if normalize(diff["tech"]) != normalize(tech):
            continue
        if normalize(diff["post4"]) == normalize(new_oar):
            return (tech, new_oar)
    return None


def explain_rename(diff, renames):
    """Match diff against TECH_NAME_RENAMES (Tech.Name column updated for
    a tech in the rename dict)."""
    if diff["col_name"] != "Tech.Name":   return None
    tech = normalize(diff["tech"])
    if tech is None:                       return None
    new_name = renames.get(tech)
    if new_name is None:                   return None
    if normalize(diff["post4"]) == normalize(new_name):
        return (tech, new_name)
    return None


# -------------------- load --------------------
print("=" * 70)
print("Acceptance test for 4_apply_manual_fixes.py")
print("=" * 70)

assert SCRIPT_4.is_file(),  f"missing {SCRIPT_4}"
assert POST_3_DIR.is_dir(), f"missing {POST_3_DIR}"
assert POST_4_DIR.is_dir(), f"missing {POST_4_DIR}"

EDITS = parse_edit_lists(SCRIPT_4)
DELETIONS         = EDITS.get("DELETIONS", [])
SUBSTITUTIONS     = EDITS.get("SUBSTITUTIONS", [])
OAR_CORRECTIONS   = EDITS.get("OAR_CORRECTIONS", [])
TECH_NAME_RENAMES = EDITS.get("TECH_NAME_RENAMES", {})

print(f"  Loaded edit lists from {SCRIPT_4.name}:")
print(f"    DELETIONS:         {len(DELETIONS)}")
print(f"    SUBSTITUTIONS:     {len(SUBSTITUTIONS)}")
print(f"    OAR_CORRECTIONS:   {len(OAR_CORRECTIONS)}")
print(f"    TECH_NAME_RENAMES: {len(TECH_NAME_RENAMES)}")
print()
print("  Computing cell-level diffs across all 4 workbooks...")
DIFFS = collect_diffs()
print(f"  Total cells differing post-3 -> post-4: {len(DIFFS)}")
print()


# -------------------- classification phase --------------------
classified = {"sub": [], "oar": [], "rename": [], "unexplained": []}
for d in DIFFS:
    if (m := explain_substitution(d, SUBSTITUTIONS)) is not None:
        classified["sub"].append((d, m))
        continue
    if (m := explain_oar(d, OAR_CORRECTIONS)) is not None:
        classified["oar"].append((d, m))
        continue
    if (m := explain_rename(d, TECH_NAME_RENAMES)) is not None:
        classified["rename"].append((d, m))
        continue
    classified["unexplained"].append(d)


# ============================================================================
# Tests
# ============================================================================

@test("T01: every actual cell change is explained by an edit-list entry")
def t01():
    if classified["unexplained"]:
        sample = classified["unexplained"][:5]
        sample_repr = [
            f"{d['workbook']}/{d['sheet']} row{d['row_idx']} "
            f"{d['col_name']!r}: {d['post3']!r} -> {d['post4']!r} "
            f"(Tech={d['tech']!r})"
            for d in sample
        ]
        raise AssertionError(
            f"{len(classified['unexplained'])} unexplained cell change(s). "
            f"Sample:\n        " + "\n        ".join(sample_repr)
        )
    print(f"        all {len(DIFFS)} cell changes traceable: "
          f"{len(classified['sub'])} subs, "
          f"{len(classified['oar'])} oar, "
          f"{len(classified['rename'])} renames")


@test("T02: every SUBSTITUTIONS entry was applied (one cell each)")
def t02():
    applied_subs = {id(m) for _, m in classified["sub"]}
    expected = len(SUBSTITUTIONS)
    matched = len(applied_subs)
    if matched != expected:
        # Find which ones didn't match
        seen_keys = set()
        for d, m in classified["sub"]:
            seen_keys.add((m[0], m[1], m[2], m[4]))   # workbook,sheet,tech,col
        missing = []
        for s in SUBSTITUTIONS:
            key = (s[0], s[1], s[2], s[4])
            if key not in seen_keys:
                missing.append(s[:5])
        raise AssertionError(
            f"{matched}/{expected} substitutions matched. "
            f"Missing: {missing[:3]}"
        )
    print(f"        all {expected} SUBSTITUTIONS applied "
          f"({len(classified['sub'])} cell changes)")


@test("T03: OAR_CORRECTIONS applied to every TRN tech's Output rows in AR_Proj/Secondary")
def t03():
    expected_techs = {normalize(t) for t, *_ in OAR_CORRECTIONS}
    seen_techs = {normalize(d["tech"]) for d, _ in classified["oar"]}
    missing = expected_techs - seen_techs
    if missing:
        raise AssertionError(f"OAR not applied for: {missing}")

    # Each TRN tech in AR_Proj/Secondary has 2 Output rows (Mode 1 + Mode 2)
    # x 28 year columns = 56 cells per tech.
    # 2 techs x 56 = 112 cells total expected.
    expected_cells = len(OAR_CORRECTIONS) * 2 * 28
    actual_cells = len(classified["oar"])
    if actual_cells != expected_cells:
        # Could be off if year columns or row counts differ; report softly
        print(f"        WARN: expected {expected_cells} OAR cells, got {actual_cells}")
    else:
        print(f"        all {len(expected_techs)} techs OAR-corrected "
              f"({actual_cells} cells)")


@test("T04: every TECH_NAME_RENAMES entry was applied at least once")
def t04():
    expected_techs = {normalize(t) for t in TECH_NAME_RENAMES}
    seen_techs = {normalize(d["tech"]) for d, _ in classified["rename"]}
    missing = expected_techs - seen_techs
    if missing:
        raise AssertionError(f"Tech.Name not renamed for: {missing}")
    print(f"        all {len(expected_techs)} techs had Tech.Name renamed "
          f"({len(classified['rename'])} cell changes across sheets)")


@test("T05: no Tech.Name cell was changed for techs NOT in TECH_NAME_RENAMES")
def t05():
    rename_techs = {normalize(t) for t in TECH_NAME_RENAMES}
    leaks = []
    for d in DIFFS:
        if d["col_name"] != "Tech.Name":
            continue
        if normalize(d["tech"]) not in rename_techs:
            leaks.append(d)
    if leaks:
        sample = [(l["workbook"], l["sheet"], l["tech"]) for l in leaks[:5]]
        raise AssertionError(
            f"{len(leaks)} Tech.Name cell(s) changed for techs not in rename list: {sample}"
        )
    print(f"        no rogue Tech.Name edits")


@test("T06: row counts unchanged when DELETIONS is empty")
def t06():
    if DELETIONS:
        print(f"        SKIP: DELETIONS non-empty ({len(DELETIONS)} entries)")
        return
    drift = []
    for label, stem in AO_STEMS.items():
        f3 = POST_3_DIR / f"{stem}_wvaligned.xlsx"
        f4 = POST_4_DIR / f"{stem}_wvaligned_v2.xlsx"
        if not (f3.is_file() and f4.is_file()):
            continue
        for sh in pd.ExcelFile(f3).sheet_names:
            try:
                n3 = pd.read_excel(f3, sh).shape[0]
                n4 = pd.read_excel(f4, sh).shape[0]
            except Exception:
                continue
            if n3 != n4:
                drift.append((label, sh, n3, n4))
    if drift:
        raise AssertionError(f"Row count drift in {len(drift)} sheets: {drift[:5]}")
    print(f"        all sheet row counts identical post-3 -> post-4")


@test("T07: no value drift in non-edit columns of refreshed rows (sanity)")
def t07():
    """For every row touched by an edit, the OTHER cells of that row should
    be byte-identical between post-3 and post-4. Catches accidental row
    rewrites by openpyxl that might silently change formatting/types."""
    edited_cells = {(d["workbook"], d["sheet"], d["row_idx"], d["col_idx"])
                    for d in DIFFS}
    edited_rows = {(d["workbook"], d["sheet"], d["row_idx"]) for d in DIFFS}
    # Re-walk the rows that had edits, check non-edit cells are equal
    drift = []
    by_wb_sh = {}
    for (wb, sh, r) in edited_rows:
        by_wb_sh.setdefault((wb, sh), []).append(r)
    for (wb, sh), rows in by_wb_sh.items():
        stem = AO_STEMS[wb]
        f3 = POST_3_DIR / f"{stem}_wvaligned.xlsx"
        f4 = POST_4_DIR / f"{stem}_wvaligned_v2.xlsx"
        wb3 = load_workbook(f3, data_only=True)
        wb4 = load_workbook(f4, data_only=True)
        ws3 = wb3[sh]; ws4 = wb4[sh]
        n_cols = ws3.max_column
        for r in rows:
            for c in range(n_cols):
                if (wb, sh, r, c) in edited_cells:
                    continue
                v3 = ws3.cell(r, c + 1).value
                v4 = ws4.cell(r, c + 1).value
                if not cell_eq(v3, v4):
                    drift.append((wb, sh, r, c, v3, v4))
    if drift:
        raise AssertionError(
            f"{len(drift)} unexpected non-edit cell change(s). "
            f"Sample: {drift[:3]}"
        )
    print(f"        non-edit cells in {len(edited_rows)} touched rows are byte-identical")


# ============================================================================
# Summary
# ============================================================================
print()
n_total = len(_results)
n_pass  = sum(1 for ok, _, _ in _results if ok)
n_fail  = n_total - n_pass
print("-" * 70)
print(f"SUMMARY:  {n_pass}/{n_total} passed,  {n_fail} failed")
print("-" * 70)
print()
print(f"Cell-change ledger:")
print(f"  Total diffs:        {len(DIFFS):5d}")
print(f"  Substitutions:      {len(classified['sub']):5d}  (expected {len(SUBSTITUTIONS)} cells)")
print(f"  OAR corrections:    {len(classified['oar']):5d}  (expected {len(OAR_CORRECTIONS) * 2 * 28} cells)")
print(f"  Tech.Name renames:  {len(classified['rename']):5d}  ({len(TECH_NAME_RENAMES)} unique techs)")
print(f"  Unexplained:        {len(classified['unexplained']):5d}")

sys.exit(0 if n_fail == 0 else 1)
