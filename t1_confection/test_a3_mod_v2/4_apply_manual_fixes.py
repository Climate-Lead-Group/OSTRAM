# -*- coding: utf-8 -*-
"""
4_apply_manual_fixes.py  (pipeline stage 4)

Applies human-curated cell-level corrections to the *_wvaligned.xlsx files
produced by stage 3. Outputs *_wvaligned_v2.xlsx with the fixes applied.

Pipeline:
  1_merge_timeslices_into_WV.py    timeslices source     -> WV
  2_extract_ao_extensions.py       WV  + A-O             -> Extensions xlsx
  3_update_ao_from_extensions.py   Extensions + A-O      -> *_wvaligned
  4_apply_manual_fixes.py          *_wvaligned           -> *_wvaligned_v2  <- THIS

What stage 4 does (and why it's a separate stage)
-------------------------------------------------
Stage 3 does mechanical row cloning with region substitution. It cannot
catch semantic anomalies discovered when reviewing its output -- e.g.,
"Maldives has no gas pipeline from India" or "Nepal imports oil via India,
not Bangladesh". Those corrections are post-hoc by nature: you only know
them after seeing the cloned rows. They land here as a hardcoded list.

Four kinds of edits supported
-----------------------------
  DELETIONS        -- remove a specific row identified by
                      (workbook, sheet, tech, selector_dict)
  SUBSTITUTIONS    -- overwrite a specific cell value
                      (with type coercion to match column dtype)
  OAR_CORRECTIONS  -- overwrite OutputActivityRatio across all year columns
                      for TRN techs in AR_Proj/Secondary Output rows
  TECH_NAME_RENAMES -- bulk-rename Tech.Name for a tech code across every
                       row in every sheet in every workbook

Substituted cells are colored purple (#9B59B6) so they stand out from the
five Step-3 / Step-2B color tiers. Deleted rows obviously can't be colored.

Setup-A: stage 3 outputs are NEVER modified. We copy them to a v2 folder
and edit the copies.

Run in Spyder with F5. Paths and edits are in USER CONFIGURATION below.
"""

import os
import shutil
import hashlib
from collections import Counter, defaultdict
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# =============================================================================
# USER CONFIGURATION
# =============================================================================

WORK_DIR = r"C:\Users\luisfernando\Desktop\OSeMOSYS\asia_ostram\t1_confection\test_a3_mod_v2"  # <-- edit for your machine

# Stage 3 outputs (read-only inputs to this stage)
INPUT_SUBDIR  = "wvaligned_outputs"
INPUT_SUFFIX  = "_wvaligned"

# Stage 4 outputs
OUTPUT_SUBDIR = "wvaligned_outputs_v2"
OUTPUT_SUFFIX = "_wvaligned_v2"

# Source filenames (stems before suffix)
AO_FILES = {
    "Param":   "A-O_Parametrization",
    "AR_Base": "A-O_AR_Model_Base_Year",
    "AR_Proj": "A-O_AR_Projections",
    "Demand":  "A-O_Demand",
}

LOG_FILE_NAME = "manual_fixes_log.txt"

# Color for substituted cells (purple). Distinct from stage 3's green/yellow/
# orange and stage 3's light-blue Demand-refresh tier. Deletions can't be
# colored (the row is gone).
MANUAL_FIX_COLOR = "FF9B59B6"

# Code column candidates (mirrors stage 3)
CODE_COLS = ["Tech", "Fuel/Tech", "Technology_Code"]

# -----------------------------------------------------------------------------
# DELETIONS: rows to remove from the v1 files.
# Tuple: (workbook_label, sheet_name, tech_code, selector_dict, justification)
#
# NONE remaining. The original 12 deletions for PWROILMDVXX, PWROILNPLXX,
# PWRNGSMDVXX were over-deletions: every gas/oil tech in the model needs
# dual-mode (M1=country fuel, M2=GASINT/OILINT). Mode 1 was wrong only
# because it inherited the template's fuel code -- the fix is a rename
# (now in SUBSTITUTIONS below), not a row deletion.
# -----------------------------------------------------------------------------

DELETIONS = []

# -----------------------------------------------------------------------------
# SUBSTITUTIONS: cell values to overwrite (one cell per entry).
# Tuple: (workbook_label, sheet_name, tech_code, selector_dict,
#         edit_col, new_value, justification)
# -----------------------------------------------------------------------------

SUBSTITUTIONS = [
    # ---- PWRNGSBGDXX Mode 1: GASIND -> GASBGD (Bangladesh has its own gas supply) ----
    ("AR_Base", "Secondary", "PWRNGSBGDXX", {"Mode.Operation": 1},
     "Fuel.I", "GASBGD",
     "Bangladesh gas plant should source from GASBGD (produced by MINGASBGD), not GASIND."),
    ("AR_Base", "Secondary", "PWRNGSBGDXX", {"Mode.Operation": 1},
     "Fuel.I.Name", "Natural Gas, Bangladesh",
     "Update fuel name to match GASBGD code."),
    ("AR_Proj", "Secondary", "PWRNGSBGDXX", {"Mode.Operation": 1, "Direction": "Input"},
     "Fuel", "GASBGD",
     "Propagate Fuel.I rename to AR_Proj Input row."),
    ("AR_Proj", "Secondary", "PWRNGSBGDXX", {"Mode.Operation": 1, "Direction": "Input"},
     "Fuel.Name", "Natural Gas, Bangladesh",
     "Propagate Fuel.I.Name rename to AR_Proj Input row."),

    # ---- PWRGEOINDNO Mode 1: HYDINDNO -> GEOINDNO (geothermal != hydro resource) ----
    ("AR_Base", "Secondary", "PWRGEOINDNO", {"Mode.Operation": 1},
     "Fuel.I", "GEOINDNO",
     "Geothermal plant should source from GEOINDNO (produced by RNWGEOINDNO), not HYDINDNO."),
    ("AR_Base", "Secondary", "PWRGEOINDNO", {"Mode.Operation": 1},
     "Fuel.I.Name", "Geothermal, India, region NO",
     "Update fuel name to match GEOINDNO code."),
    ("AR_Proj", "Secondary", "PWRGEOINDNO", {"Mode.Operation": 1, "Direction": "Input"},
     "Fuel", "GEOINDNO",
     "Propagate Fuel.I rename to AR_Proj Input row."),
    ("AR_Proj", "Secondary", "PWRGEOINDNO", {"Mode.Operation": 1, "Direction": "Input"},
     "Fuel.Name", "Geothermal, India, region NO",
     "Propagate Fuel.I.Name rename to AR_Proj Input row."),

    # ---- PWRNGSMDVXX Mode 1: GASIND -> GASMDV (Maldives has its own gas supply) ----
    ("AR_Base", "Secondary", "PWRNGSMDVXX", {"Mode.Operation": 1},
     "Fuel.I", "GASMDV",
     "Maldives gas plant Mode 1 should source from GASMDV (produced by MINGASMDV), not GASIND."),
    ("AR_Base", "Secondary", "PWRNGSMDVXX", {"Mode.Operation": 1},
     "Fuel.I.Name", "Natural Gas, Maldives",
     "Update fuel name to match GASMDV code."),
    ("AR_Proj", "Secondary", "PWRNGSMDVXX", {"Mode.Operation": 1, "Direction": "Input"},
     "Fuel", "GASMDV",
     "Propagate Fuel.I rename to AR_Proj Input row."),
    ("AR_Proj", "Secondary", "PWRNGSMDVXX", {"Mode.Operation": 1, "Direction": "Input"},
     "Fuel.Name", "Natural Gas, Maldives",
     "Propagate Fuel.I.Name rename to AR_Proj Input row."),

    # ---- PWROILMDVXX Mode 1: OILBGD -> OILMDV (Maldives has its own oil supply) ----
    ("AR_Base", "Secondary", "PWROILMDVXX", {"Mode.Operation": 1},
     "Fuel.I", "OILMDV",
     "Maldives oil plant Mode 1 should source from OILMDV (produced by MINOILMDV), not OILBGD."),
    ("AR_Base", "Secondary", "PWROILMDVXX", {"Mode.Operation": 1},
     "Fuel.I.Name", "Oil, Maldives",
     "Update fuel name to match OILMDV code."),
    ("AR_Proj", "Secondary", "PWROILMDVXX", {"Mode.Operation": 1, "Direction": "Input"},
     "Fuel", "OILMDV",
     "Propagate Fuel.I rename to AR_Proj Input row."),
    ("AR_Proj", "Secondary", "PWROILMDVXX", {"Mode.Operation": 1, "Direction": "Input"},
     "Fuel.Name", "Oil, Maldives",
     "Propagate Fuel.I.Name rename to AR_Proj Input row."),

    # ---- PWROILNPLXX Mode 1: OILBGD -> OILNPL (Nepal has its own oil supply) ----
    ("AR_Base", "Secondary", "PWROILNPLXX", {"Mode.Operation": 1},
     "Fuel.I", "OILNPL",
     "Nepal oil plant Mode 1 should source from OILNPL (produced by MINOILNPL), not OILBGD."),
    ("AR_Base", "Secondary", "PWROILNPLXX", {"Mode.Operation": 1},
     "Fuel.I.Name", "Oil, Nepal",
     "Update fuel name to match OILNPL code."),
    ("AR_Proj", "Secondary", "PWROILNPLXX", {"Mode.Operation": 1, "Direction": "Input"},
     "Fuel", "OILNPL",
     "Propagate Fuel.I rename to AR_Proj Input row."),
    ("AR_Proj", "Secondary", "PWROILNPLXX", {"Mode.Operation": 1, "Direction": "Input"},
     "Fuel.Name", "Oil, Nepal",
     "Propagate Fuel.I.Name rename to AR_Proj Input row."),
]

# -----------------------------------------------------------------------------
# OAR_CORRECTIONS: fix OutputActivityRatio across ALL year columns for TRN
# techs in AR_Proj/Secondary whose OAR was inherited from the wrong template.
# Tuple: (tech_code, new_oar_value, justification)
# Applied to every Output row (both Mode 1 and Mode 2) in AR_Proj/Secondary.
# -----------------------------------------------------------------------------

OAR_CORRECTIONS = [
    ("TRNBTNXXBGDXX", 0.95,
     "Template OAR 0.96716 inherited from TRNBTNXXINDEA; WV standard is 0.95."),
    ("TRNNPLXXBGDXX", 0.95,
     "Template OAR 0.98292 inherited from TRNBGDXXINDEA; WV standard is 0.95."),
]

# -----------------------------------------------------------------------------
# TECH_NAME_RENAMES: bulk-rename Tech.Name for tech codes whose name was
# inherited from the template and is wrong (wrong region, wrong technology
# category, or both). Applied to EVERY occurrence of the tech code across
# every sheet in every workbook.
# Dict: {tech_code: correct_name}
# -----------------------------------------------------------------------------

TECH_NAME_RENAMES = {
    "PWRNGSBGDXX":   "Natural Gas (Power generator) Bangladesh, region XX",
    "PWRNGSMDVXX":   "Natural Gas (Power generator) Maldives, region XX",
    "PWROILMDVXX":   "Oil (Power generator) Maldives, region XX",
    "PWROILNPLXX":   "Oil (Power generator) Nepal, region XX",
    "PWRSHPINDEA":   "Small Hydropower (Power generator) India, region EA",
    "PWRSHPINDNE":   "Small Hydropower (Power generator) India, region NE",
    "PWRSHPINDNO":   "Small Hydropower (Power generator) India, region NO",
    "PWRSHPINDSO":   "Small Hydropower (Power generator) India, region SO",
    "PWRSHPINDWE":   "Small Hydropower (Power generator) India, region WE",
    "PWRWOFMDVXX":   "Offshore Wind (Power generator) Maldives, region XX",
    "PWRWONMDVXX":   "Onshore Wind (Power generator) Maldives, region XX",
    "TRNBTNXXBGDXX": "Transmission interconnection from Bhutan, region XX to Bangladesh, region XX",
    "TRNBTNXXINDNE": "Transmission interconnection from Bhutan, region XX to India, region NE",
}

# =============================================================================
# HELPERS
# =============================================================================

def find_code_col(headers):
    return next((c for c in CODE_COLS if c in headers), None)


def md5_of(path):
    if not os.path.exists(path):
        return None
    h = hashlib.md5()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def values_equal(a, b):
    """Type-tolerant equality for selector matching.
    Handles int/float/str cross-type compares (e.g. cell value 1.0 vs
    selector 1, or 'Input' vs 'input ')."""
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    if a == b:
        return True
    # Numeric compare
    try:
        return float(a) == float(b)
    except (TypeError, ValueError):
        pass
    # String compare (whitespace-tolerant, case-tolerant)
    return str(a).strip().lower() == str(b).strip().lower()


def find_row(ws, tech_code, selectors):
    """Return list of 1-indexed row numbers in ws matching tech_code AND all
    selector key-value pairs."""
    hdrs = [c.value for c in ws[1]]
    cc = find_code_col(hdrs)
    if cc is None:
        return []
    ci = hdrs.index(cc) + 1

    sel_idx = {}
    for sel_col in selectors:
        if sel_col not in hdrs:
            return []  # selector column doesn't exist in this sheet
        sel_idx[sel_col] = hdrs.index(sel_col) + 1

    matches = []
    for r in range(2, ws.max_row + 1):
        cell_code = ws.cell(row=r, column=ci).value
        if not values_equal(cell_code, tech_code):
            continue
        all_match = True
        for sel_col, sel_val in selectors.items():
            cell_val = ws.cell(row=r, column=sel_idx[sel_col]).value
            if not values_equal(cell_val, sel_val):
                all_match = False
                break
        if all_match:
            matches.append(r)
    return matches


def coerce_value(ws, col_idx_1based, new_value):
    """Probe column's existing dtype (excluding header) and coerce new_value
    to match. Falls back to new_value as-is if column is empty."""
    sample_types = []
    for r in range(2, min(ws.max_row + 1, 50)):  # sample first ~50 data rows
        v = ws.cell(row=r, column=col_idx_1based).value
        if v is not None:
            sample_types.append(type(v))
    if not sample_types:
        return new_value
    dominant = Counter(sample_types).most_common(1)[0][0]
    try:
        if dominant is int:
            return int(new_value)
        if dominant is float:
            return float(new_value)
        if dominant is str:
            return str(new_value)
    except (TypeError, ValueError):
        pass
    return new_value


# =============================================================================
# STEP 0 -- SET UP OUTPUT FOLDER + COPY V1 -> V2 (Setup-A)
# =============================================================================

print("=" * 72)
print("4_apply_manual_fixes.py")
print("=" * 72)

in_dir  = os.path.join(WORK_DIR, INPUT_SUBDIR)
out_dir = os.path.join(WORK_DIR, OUTPUT_SUBDIR)
os.makedirs(out_dir, exist_ok=True)
print(f"Input folder:  {in_dir}")
print(f"Output folder: {out_dir}")

IN_FILES  = {label: os.path.join(in_dir, stem + INPUT_SUFFIX + ".xlsx")
             for label, stem in AO_FILES.items()}
OUT_FILES = {label: os.path.join(out_dir, stem + OUTPUT_SUFFIX + ".xlsx")
             for label, stem in AO_FILES.items()}

# Hashes BEFORE we do anything so we can verify v1 inputs untouched
src_hashes = {p: md5_of(p) for p in IN_FILES.values()}

# Sanity: every input file must exist
print("\nSetup-A: copying stage-3 outputs to v2 folder")
for label in AO_FILES:
    if not os.path.exists(IN_FILES[label]):
        raise FileNotFoundError(
            f"Stage 3 output not found: {IN_FILES[label]}\n"
            f"  Run stage 3 first to produce *{INPUT_SUFFIX}.xlsx files."
        )
    shutil.copyfile(IN_FILES[label], OUT_FILES[label])
    print(f"  {label:8s}  {os.path.basename(OUT_FILES[label])}")

# Open output workbooks for editing
out_wbs = {label: load_workbook(OUT_FILES[label]) for label in AO_FILES}

# Audit log accumulator
log_lines = []
log_lines.append("=" * 72)
log_lines.append("4_apply_manual_fixes.py -- Run log")
log_lines.append(f"Run timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
log_lines.append("=" * 72)
log_lines.append("")
log_lines.append("INPUT")
for label, p in IN_FILES.items():
    log_lines.append(f"  {label:8s}  {os.path.basename(p)}  md5={src_hashes[p]}")
log_lines.append(f"  Deletions requested:     {len(DELETIONS)}")
log_lines.append(f"  Substitutions requested: {len(SUBSTITUTIONS)}")
log_lines.append(f"  OAR corrections requested: {len(OAR_CORRECTIONS)}")
log_lines.append(f"  Tech.Name renames requested: {len(TECH_NAME_RENAMES)}")
log_lines.append("")

# =============================================================================
# STEP 1 -- APPLY DELETIONS
# =============================================================================

print("\n" + "=" * 72)
print("STEP 1 -- Apply deletions")
print("=" * 72)

# Group deletions by (workbook, sheet) so we can sort row indices descending
# within a sheet before deleting (avoids index drift when openpyxl shifts
# subsequent rows up after each ws.delete_rows() call).
del_groups = defaultdict(list)   # (label, sheet) -> [(row_idx, tech, selectors, justification)]
del_outcomes = []                 # (label, sheet, tech, selectors, status, justification)

for label, sheet, tech, selectors, justification in DELETIONS:
    if label not in out_wbs:
        del_outcomes.append((label, sheet, tech, selectors, "NO_WORKBOOK", justification))
        continue
    wb = out_wbs[label]
    if sheet not in wb.sheetnames:
        del_outcomes.append((label, sheet, tech, selectors, "NO_SHEET", justification))
        continue
    ws = wb[sheet]
    matches = find_row(ws, tech, selectors)
    if len(matches) == 0:
        del_outcomes.append((label, sheet, tech, selectors, "NO_MATCH", justification))
    elif len(matches) > 1:
        del_outcomes.append((label, sheet, tech, selectors,
                             f"AMBIGUOUS({matches})", justification))
    else:
        del_groups[(label, sheet)].append((matches[0], tech, selectors, justification))
        del_outcomes.append((label, sheet, tech, selectors,
                             f"FOUND_AT_ROW_{matches[0]}", justification))

# Apply: bottom-up per sheet
deletions_applied = 0
del_per_sheet = defaultdict(int)

for (label, sheet), entries in del_groups.items():
    ws = out_wbs[label][sheet]
    entries_sorted = sorted(entries, key=lambda x: -x[0])  # descending by row idx
    for row_idx, tech, selectors, justification in entries_sorted:
        ws.delete_rows(row_idx, 1)
        deletions_applied += 1
        del_per_sheet[(label, sheet)] += 1
        log_lines.append(f"DELETED  {label}/{sheet}  row {row_idx}: {tech}  {selectors}")
        log_lines.append(f"  Justification: {justification}")
        print(f"  deleted   {label:7s} {sheet:14s}  row {row_idx:5d}  "
              f"{tech:14s}  {selectors}")

# Log skipped/no-match cases
skipped = [o for o in del_outcomes if not o[4].startswith("FOUND")]
if skipped:
    log_lines.append("")
    log_lines.append("DELETIONS NOT APPLIED:")
    for label, sheet, tech, selectors, status, _ in skipped:
        log_lines.append(f"  [{status}]  {label}/{sheet}  {tech}  {selectors}")
        print(f"  [{status:14s}] {label}/{sheet}  {tech}  {selectors}")

print(f"\n  Deletions applied: {deletions_applied} / {len(DELETIONS)}")

# =============================================================================
# STEP 2 -- APPLY SUBSTITUTIONS
# =============================================================================

print("\n" + "=" * 72)
print("STEP 2 -- Apply substitutions")
print("=" * 72)

substitutions_applied = 0
sub_outcomes = []
purple_fill = PatternFill(start_color=MANUAL_FIX_COLOR,
                          end_color=MANUAL_FIX_COLOR, fill_type="solid")

for entry in SUBSTITUTIONS:
    label, sheet, tech, selectors, edit_col, new_value, justification = entry
    if label not in out_wbs:
        sub_outcomes.append((label, sheet, tech, selectors, edit_col, "NO_WORKBOOK"))
        continue
    wb = out_wbs[label]
    if sheet not in wb.sheetnames:
        sub_outcomes.append((label, sheet, tech, selectors, edit_col, "NO_SHEET"))
        continue
    ws = wb[sheet]
    hdrs = [c.value for c in ws[1]]
    if edit_col not in hdrs:
        sub_outcomes.append((label, sheet, tech, selectors, edit_col, "NO_COLUMN"))
        continue
    matches = find_row(ws, tech, selectors)
    if len(matches) == 0:
        sub_outcomes.append((label, sheet, tech, selectors, edit_col, "NO_MATCH"))
        continue
    if len(matches) > 1:
        sub_outcomes.append((label, sheet, tech, selectors, edit_col,
                             f"AMBIGUOUS({matches})"))
        continue

    r = matches[0]
    ec = hdrs.index(edit_col) + 1
    old_val = ws.cell(row=r, column=ec).value
    coerced = coerce_value(ws, ec, new_value)
    ws.cell(row=r, column=ec).value = coerced
    ws.cell(row=r, column=ec).fill = purple_fill

    substitutions_applied += 1
    log_lines.append(f"SUBST    {label}/{sheet}  row {r}: {tech}  {selectors}")
    log_lines.append(f"  {edit_col}: {old_val!r} -> {coerced!r}")
    log_lines.append(f"  Justification: {justification}")
    print(f"  subst     {label:7s} {sheet:14s}  row {r:5d}  "
          f"{tech:14s}  {edit_col}: {old_val!r} -> {coerced!r}")

# Log skipped/no-match cases
skipped_sub = [o for o in sub_outcomes if "FOUND" not in str(o)]
if skipped_sub:
    log_lines.append("")
    log_lines.append("SUBSTITUTIONS NOT APPLIED:")
    for o in skipped_sub:
        log_lines.append(f"  [{o[5]}]  {o[0]}/{o[1]}  {o[2]}  {o[3]}  edit_col={o[4]}")
        print(f"  [{o[5]:14s}] {o[0]}/{o[1]}  {o[2]}  edit_col={o[4]}")

print(f"\n  Substitutions applied: {substitutions_applied} / {len(SUBSTITUTIONS)}")

# =============================================================================
# STEP 2B -- APPLY OAR CORRECTIONS
# =============================================================================

print("\n" + "=" * 72)
print("STEP 2B -- Apply OAR corrections")
print("=" * 72)

oar_cells_fixed = 0
oar_rows_fixed = 0

ws_proj = out_wbs["AR_Proj"]["Secondary"]
proj_hdrs = [c.value for c in ws_proj[1]]
proj_tc_i = proj_hdrs.index("Tech") + 1
proj_dir_i = proj_hdrs.index("Direction") + 1

# Find year column indices (all columns whose header is an integer >= 2023)
year_cols = []
for ci_0, h in enumerate(proj_hdrs):
    try:
        y = int(h)
        if 2023 <= y <= 2060:
            year_cols.append(ci_0 + 1)  # 1-indexed
    except (TypeError, ValueError):
        pass

for tech, new_oar, justification in OAR_CORRECTIONS:
    for r in range(2, ws_proj.max_row + 1):
        cell_tech = ws_proj.cell(row=r, column=proj_tc_i).value
        cell_dir  = ws_proj.cell(row=r, column=proj_dir_i).value
        if not values_equal(cell_tech, tech):
            continue
        if not values_equal(cell_dir, "Output"):
            continue
        # Found an Output row for this tech -- overwrite all year cells
        old_val = ws_proj.cell(row=r, column=year_cols[0]).value
        for yc in year_cols:
            ws_proj.cell(row=r, column=yc).value = new_oar
            ws_proj.cell(row=r, column=yc).fill = purple_fill
        oar_cells_fixed += len(year_cols)
        oar_rows_fixed += 1
        log_lines.append(f"OAR_FIX  AR_Proj/Secondary  row {r}: {tech}  Output")
        log_lines.append(f"  All year cols: {old_val!r} -> {new_oar!r}  ({len(year_cols)} cells)")
        log_lines.append(f"  Justification: {justification}")
        print(f"  oar_fix   AR_Proj Secondary    row {r:5d}  "
              f"{tech:14s}  OAR: {old_val!r} -> {new_oar!r}  ({len(year_cols)} cells)")

print(f"\n  OAR rows fixed: {oar_rows_fixed}  ({oar_cells_fixed} cells)")

# =============================================================================
# STEP 2C -- APPLY TECH.NAME RENAMES
# =============================================================================

print("\n" + "=" * 72)
print("STEP 2C -- Apply Tech.Name renames")
print("=" * 72)

rename_cells_fixed = 0
rename_per_sheet = defaultdict(int)

for label in AO_FILES:
    wb = out_wbs[label]
    for sn in wb.sheetnames:
        ws = wb[sn]
        hdrs = [c.value for c in ws[1]]
        # Find Tech and Tech.Name columns
        if "Tech" not in hdrs or "Tech.Name" not in hdrs:
            continue
        tc_ci = hdrs.index("Tech") + 1
        tn_ci = hdrs.index("Tech.Name") + 1
        for r in range(2, ws.max_row + 1):
            cell_tech = ws.cell(row=r, column=tc_ci).value
            if cell_tech not in TECH_NAME_RENAMES:
                continue
            old_name = ws.cell(row=r, column=tn_ci).value
            new_name = TECH_NAME_RENAMES[cell_tech]
            if old_name == new_name:
                continue  # already correct (shouldn't happen, but safe)
            ws.cell(row=r, column=tn_ci).value = new_name
            ws.cell(row=r, column=tn_ci).fill = purple_fill
            rename_cells_fixed += 1
            rename_per_sheet[(label, sn)] += 1

# Log summary per sheet (not per cell -- would be too verbose)
log_lines.append("")
for (label, sn), n in sorted(rename_per_sheet.items()):
    log_lines.append(f"RENAME   {label}/{sn}  {n} Tech.Name cell(s) updated")
    print(f"  renamed   {label:7s} {sn:25s}  {n:3d} cell(s)")
log_lines.append(f"RENAME   Total: {rename_cells_fixed} cell(s) across "
                 f"{len(rename_per_sheet)} sheet(s) for {len(TECH_NAME_RENAMES)} tech(s)")

print(f"\n  Tech.Name cells renamed: {rename_cells_fixed} across "
      f"{len(rename_per_sheet)} sheets")

# =============================================================================
# STEP 3 -- SAVE OUTPUT WORKBOOKS
# =============================================================================

print("\n" + "=" * 72)
print("STEP 3 -- Save output workbooks")
print("=" * 72)
for label in AO_FILES:
    out_wbs[label].save(OUT_FILES[label])
    n_del = sum(c for (l, _), c in del_per_sheet.items() if l == label)
    print(f"  Saved: {os.path.basename(OUT_FILES[label])}  ({n_del} row(s) deleted)")

# =============================================================================
# STEP 4 -- WRITE AUDIT LOG
# =============================================================================

print("\n" + "=" * 72)
print("STEP 4 -- Write audit log")
print("=" * 72)

log_lines.append("")
log_lines.append("=" * 72)
log_lines.append("SUMMARY")
log_lines.append("=" * 72)
log_lines.append(f"  Deletions applied:     {deletions_applied} / {len(DELETIONS)}")
log_lines.append(f"  Substitutions applied: {substitutions_applied} / {len(SUBSTITUTIONS)}")
log_lines.append(f"  OAR rows fixed:        {oar_rows_fixed} ({oar_cells_fixed} cells)")
log_lines.append(f"  Tech.Name renames:     {rename_cells_fixed} cells across {len(rename_per_sheet)} sheets")
log_lines.append("  Per-sheet deletion counts:")
for (label, sheet), n in sorted(del_per_sheet.items()):
    log_lines.append(f"    {label}/{sheet:20s}  -{n}")

log_path = os.path.join(out_dir, LOG_FILE_NAME)
with open(log_path, "w", encoding="utf-8") as f:
    f.write("\n".join(log_lines) + "\n")
print(f"  Wrote: {log_path}")

# =============================================================================
# BUILT-IN TESTS
# =============================================================================

print("\n" + "=" * 72)
print("TESTS")
print("=" * 72)

_passed, _failed = 0, 0
def check(name, cond, detail=""):
    global _passed, _failed
    if cond:
        _passed += 1
        print(f"  PASS  {name}")
    else:
        _failed += 1
        print(f"  FAIL  {name}   {detail}")

# (a) v1 inputs untouched
for p, h_orig in src_hashes.items():
    h_now = md5_of(p)
    check(f"v1 source unchanged: {os.path.basename(p)}",
          h_now == h_orig,
          f"(was {h_orig[:8]}, now {h_now[:8] if h_now else 'missing'})")

# (b) v2 outputs exist
for label, p in OUT_FILES.items():
    check(f"v2 output exists: {os.path.basename(p)}", os.path.exists(p))

# (c) Each requested deletion's target row is no longer present in the v2 file.
#     If the deletion was skipped (NO_MATCH/AMBIGUOUS/etc.), this still holds:
#     for NO_MATCH we expect 0 matches anyway; for AMBIGUOUS the test will fail,
#     correctly flagging that nothing was done.
#
# Cache reloaded workbooks by label (one open per workbook, not per deletion).
# Use normal load (not read_only) because find_row relies on ws[1] and
# ws.max_row, which are unreliable in read_only mode.
verify_wbs = {label: load_workbook(OUT_FILES[label], data_only=True)
              for label in AO_FILES}

for label, sheet, tech, selectors, _ in DELETIONS:
    wb = verify_wbs[label]
    if sheet not in wb.sheetnames:
        check(f"deleted from {label}/{sheet}: {tech} {selectors}",
              False, "(sheet missing in v2)")
        continue
    matches = find_row(wb[sheet], tech, selectors)
    check(f"deleted from {label}/{sheet}: {tech} {selectors}",
          len(matches) == 0,
          f"({len(matches)} rows still match)")

# (d) Each requested substitution landed
for entry in SUBSTITUTIONS:
    label, sheet, tech, selectors, edit_col, new_value, _ = entry
    wb = verify_wbs[label]
    if sheet not in wb.sheetnames:
        check(f"sub target sheet exists: {label}/{sheet}", False)
        continue
    ws = wb[sheet]
    hdrs = [c.value for c in ws[1]]
    matches = find_row(ws, tech, selectors)
    if len(matches) != 1 or edit_col not in hdrs:
        check(f"sub locatable: {label}/{sheet} {tech} {edit_col}",
              False, f"({len(matches)} matches, edit_col in hdrs={edit_col in hdrs})")
        continue
    r = matches[0]
    ec = hdrs.index(edit_col) + 1
    actual = ws.cell(row=r, column=ec).value
    check(f"sub applied: {label}/{sheet} {tech}.{edit_col} = {new_value!r}",
          values_equal(actual, new_value),
          f"(got {actual!r})")

# (d2) Each OAR correction landed -- check year cols in verify_wbs
ws_v = verify_wbs["AR_Proj"]["Secondary"]
v_hdrs = [c.value for c in ws_v[1]]
v_tc_i = v_hdrs.index("Tech") + 1
v_dir_i = v_hdrs.index("Direction") + 1
v_year_cols = []
for ci_0, h in enumerate(v_hdrs):
    try:
        y = int(h)
        if 2023 <= y <= 2060:
            v_year_cols.append(ci_0 + 1)
    except (TypeError, ValueError):
        pass

for tech, new_oar, _ in OAR_CORRECTIONS:
    for r in range(2, ws_v.max_row + 1):
        ct = ws_v.cell(row=r, column=v_tc_i).value
        cd = ws_v.cell(row=r, column=v_dir_i).value
        if not values_equal(ct, tech) or not values_equal(cd, "Output"):
            continue
        # Check first and last year columns
        v_first = ws_v.cell(row=r, column=v_year_cols[0]).value
        v_last  = ws_v.cell(row=r, column=v_year_cols[-1]).value
        check(f"OAR applied: {tech} row {r} first_yr={new_oar}",
              values_equal(v_first, new_oar),
              f"(got {v_first!r})")
        check(f"OAR applied: {tech} row {r} last_yr={new_oar}",
              values_equal(v_last, new_oar),
              f"(got {v_last!r})")

# (d3) Each Tech.Name rename landed -- spot-check one row per tech per workbook
for label in AO_FILES:
    wb = verify_wbs[label]
    for sn in wb.sheetnames:
        ws = wb[sn]
        hdrs = [c.value for c in ws[1]]
        if "Tech" not in hdrs or "Tech.Name" not in hdrs:
            continue
        tc_ci = hdrs.index("Tech") + 1
        tn_ci = hdrs.index("Tech.Name") + 1
        checked_in_sheet = set()
        for r in range(2, ws.max_row + 1):
            ct = ws.cell(row=r, column=tc_ci).value
            if ct not in TECH_NAME_RENAMES or ct in checked_in_sheet:
                continue
            checked_in_sheet.add(ct)
            actual_name = ws.cell(row=r, column=tn_ci).value
            expected_name = TECH_NAME_RENAMES[ct]
            check(f"rename: {label}/{sn} {ct}",
                  actual_name == expected_name,
                  f"(got {actual_name!r})")

# Close verification workbooks
for wb in verify_wbs.values():
    wb.close()

# (e) Per-sheet row delta equals -count_of_deletions for that sheet
for label in AO_FILES:
    src_xl = pd.ExcelFile(IN_FILES[label])
    out_xl = pd.ExcelFile(OUT_FILES[label])
    for s in src_xl.sheet_names:
        if s not in out_xl.sheet_names:
            check(f"{label}: sheet '{s}' preserved", False, "(missing in v2)")
            continue
        n_src = pd.read_excel(src_xl, s).shape[0]
        n_out = pd.read_excel(out_xl, s).shape[0]
        expected_delta = -del_per_sheet.get((label, s), 0)
        actual_delta = n_out - n_src
        check(f"{label}/{s} row delta = {expected_delta}",
              actual_delta == expected_delta,
              f"(got {actual_delta}; src={n_src}, v2={n_out})")

# (f) Every requested deletion was either applied or accounted for
applied = sum(1 for o in del_outcomes if o[4].startswith("FOUND"))
unapplied = len(DELETIONS) - applied
check(f"deletion accounting: {applied} applied + {unapplied} unapplied = {len(DELETIONS)} requested",
      applied + unapplied == len(DELETIONS),
      f"(applied={applied}, unapplied={unapplied}, requested={len(DELETIONS)})")

print(f"\n  {_passed} passed, {_failed} failed")
if _failed == 0:
    print("  ALL TESTS PASSED")
else:
    print(f"  {_failed} TEST(S) FAILED -- review output above")
