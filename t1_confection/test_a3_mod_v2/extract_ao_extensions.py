# -*- coding: utf-8 -*-
"""
OSTRAM v13 — Extract A-O Extension List
========================================

Answers the question: which v13 tech codes need to be added to A-O?

Two ways to detect them — this script uses ABSENCE (the reliable one) and
OPTIONALLY cross-checks against COLOR (the visual flag the team painted).

Why absence is the right primary signal
---------------------------------------
A code is "missing from A-O" if it's used in v13 but isn't in any A-O
parametrization file. That's a binary, file-content fact — no risk of
false positives from misapplied highlights or false negatives from rows
the team forgot to color.

Color (orange = FCD5B4) is the human-applied flag. It can be useful as
a second check: if a code is orange but exists in A-O, it means the
team flagged something we shouldn't worry about (or there's a stale flag).
If a code is absent from A-O but NOT orange, it slipped through.

Output: OSTRAM_AO_Extensions.xlsx
  - Tab 1: Extensions_To_Add  — codes to add to A-O, with their Tech.Name,
    where they appear in v13, and the parameter rows from v13 that should
    be replicated in A-O
  - Tab 2: Disagreements      — orange-but-in-AO and absent-but-not-orange
    cases (sanity check)

Requirements: pandas, openpyxl
"""

import pandas as pd
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook


# ════════════════════════════════════════════════════════════════════════
# CONFIG — edit these paths
# ════════════════════════════════════════════════════════════════════════
WORK_DIR = Path(r"C:\Users\luisfernando\Desktop\OSeMOSYS\asia_ostram\t1_confection\test_a3_mod_v2")

V13_PATH = WORK_DIR / "SOASIA_OSeMOSYS_Template_v15.xlsx"

AO_FILES = [
    WORK_DIR / "A-O_Parametrization.xlsx",
    WORK_DIR / "A-O_AR_Model_Base_Year.xlsx",
    WORK_DIR / "A-O_AR_Projections.xlsx",
    WORK_DIR / "A-O_Demand.xlsx",
]

OUTPUT_PATH = WORK_DIR / "OSTRAM_AO_Extensions.xlsx"


# Sheets that are parsed by the AO-based model that runs.
# Only codes in these sheets matter for A-O extension decisions.
BLUE_SHEETS = ['Fixed_Horizon_Parameters', 'Primary_Techs', 'Secondary_Techs',
               'VariableCost', 'Demand_Projection', 'Demand_Techs', 'Emissions']

# Orange highlight color used by the team to flag rows for decision.
ORANGE_COLORS = {'FFFCD5B4', 'FFFCD5B5'}


# ════════════════════════════════════════════════════════════════════════
# PRIMARY SIGNAL: ABSENCE
# ════════════════════════════════════════════════════════════════════════
def collect_v13_codes(path):
    """Return {sheet: {code: tech_name}} for blue sheets only."""
    xl = pd.ExcelFile(path)
    out = {}
    for s in xl.sheet_names:
        if s not in BLUE_SHEETS:
            continue
        df = pd.read_excel(xl, s)
        code_col = next((c for c in ['Tech', 'Fuel/Tech', 'Technology_Code']
                         if c in df.columns), None)
        if not code_col:
            continue
        name_col = next((c for c in ['Tech.Name', 'Name', 'Technology_Name']
                         if c in df.columns), None)
        codes = {}
        for _, r in df.iterrows():
            code = r[code_col]
            if pd.isna(code) or str(code).strip() == '':
                continue
            code = str(code)
            name = ''
            if name_col and pd.notna(r[name_col]):
                name = str(r[name_col]).strip()
            if code not in codes or len(name) > len(codes[code]):
                codes[code] = name
        out[s] = codes
    return out


def collect_ao_codes(paths):
    """Return set of all tech codes present in A-O parametrization files."""
    codes = set()
    for fp in paths:
        xl = pd.ExcelFile(fp)
        for s in xl.sheet_names:
            df = pd.read_excel(xl, s)
            for c in ['Tech', 'Fuel/Tech']:
                if c in df.columns:
                    codes.update(df[c].dropna().astype(str))
    return codes


# ════════════════════════════════════════════════════════════════════════
# SECONDARY SIGNAL: COLOR (cross-check only)
# ════════════════════════════════════════════════════════════════════════
def collect_orange_codes(path):
    """Find every code that has at least one orange-fill cell in a blue sheet."""
    wb = load_workbook(path, data_only=True)
    orange_codes = set()
    for sheet_name in wb.sheetnames:
        if sheet_name not in BLUE_SHEETS:
            continue
        ws = wb[sheet_name]
        # Headers (row 1) → find code column
        headers = [c.value for c in ws[1]]
        code_col_idx = None
        for i, h in enumerate(headers):
            if h in ('Tech', 'Fuel/Tech', 'Technology_Code'):
                code_col_idx = i + 1
                break
        if code_col_idx is None:
            continue
        for row_idx in range(2, ws.max_row + 1):
            # If any cell in this row is orange, capture its code
            row_is_orange = False
            for cell in ws[row_idx]:
                if cell.fill and cell.fill.fgColor and \
                        cell.fill.fgColor.rgb in ORANGE_COLORS:
                    row_is_orange = True
                    break
            if row_is_orange:
                code = ws.cell(row_idx, code_col_idx).value
                if isinstance(code, str) and code.strip():
                    orange_codes.add(code)
    return orange_codes


# ════════════════════════════════════════════════════════════════════════
# EXTRACT PARAMETER ROWS FOR EACH EXTENSION CODE
# ════════════════════════════════════════════════════════════════════════
def extract_parameter_rows(path, codes_to_extract):
    """Return list of dicts — one per parameter row referencing an extension code,
    keyed by source v13 sheet so the team can replicate them in A-O."""
    rows = []
    xl = pd.ExcelFile(path)
    for s in xl.sheet_names:
        if s not in BLUE_SHEETS:
            continue
        df = pd.read_excel(xl, s)
        code_col = next((c for c in ['Tech', 'Fuel/Tech', 'Technology_Code']
                         if c in df.columns), None)
        if not code_col:
            continue
        sub = df[df[code_col].astype(str).isin(codes_to_extract)].copy()
        if len(sub) == 0:
            continue
        sub.insert(0, 'Source_Sheet', s)
        rows.append(sub)
    if not rows:
        return pd.DataFrame()
    return pd.concat(rows, ignore_index=True, sort=False)


# ════════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════════
def main():
    print("=" * 70)
    print("OSTRAM v15 — A-O Extension Extractor")
    print("=" * 70)

    print("\nLoading v15 (blue sheets only)...")
    v13_by_sheet = collect_v13_codes(V13_PATH)
    v13_all = {}
    for sheet, codes in v13_by_sheet.items():
        for code, name in codes.items():
            if code not in v13_all or len(name) > len(v13_all[code]):
                v13_all[code] = name
    print(f"  {len(v13_all)} unique codes across {len(v13_by_sheet)} sheets")

    print("Loading A-O reference...")
    ao_codes = collect_ao_codes(AO_FILES)
    print(f"  {len(ao_codes)} unique codes")

    print("Detecting orange highlights in v13...")
    orange_codes = collect_orange_codes(V13_PATH)
    print(f"  {len(orange_codes)} codes have at least one orange cell")

    # The two signals
    absent = {c for c in v13_all if c not in ao_codes}

    # Cross-check
    orange_and_absent     = orange_codes & absent           # both signals agree
    orange_but_in_ao      = orange_codes - absent           # flagged but already in A-O
    absent_but_not_orange = absent - orange_codes           # missing in A-O but unflagged

    print(f"\n  Absent from A-O:                {len(absent)}")
    print(f"    of which orange-flagged:       {len(orange_and_absent)}")
    print(f"    of which NOT orange-flagged:   {len(absent_but_not_orange)}")
    print(f"  Orange but already in A-O:       {len(orange_but_in_ao)}")

    # Build the canonical extension list = absent (the binary fact)
    extensions = sorted(absent)
    print(f"\nExtension list ({len(extensions)} codes):")
    for c in extensions:
        flag = '  ✓ orange' if c in orange_codes else '  ⚠ not flagged'
        print(f"  {c:<22} {v13_all[c]}{flag}")

    # Build sheet-presence summary per extension code
    presence = []
    for c in extensions:
        sheets_with = [s for s, codes in v13_by_sheet.items() if c in codes]
        presence.append({
            'AO_Code_To_Add': c,
            'Tech_Name': v13_all[c],
            'Orange_Flagged': 'Y' if c in orange_codes else 'N',
            'Sheets_in_v13': ', '.join(sheets_with),
            'Sheet_Count': len(sheets_with),
        })
    presence_df = pd.DataFrame(presence)

    # Extract the actual parameter rows so the team can replicate them in A-O
    print("\nExtracting parameter rows for A-O insertion...")
    param_rows = extract_parameter_rows(V13_PATH, set(extensions))
    print(f"  {len(param_rows)} parameter rows referencing the {len(extensions)} extension codes")

    # Disagreements between signals
    disagree_rows = []
    for c in sorted(orange_but_in_ao):
        disagree_rows.append({
            'Code': c,
            'Issue': 'Orange-flagged but already in A-O',
            'Action': 'Probably stale flag — clear the orange highlight',
        })
    for c in sorted(absent_but_not_orange):
        disagree_rows.append({
            'Code': c,
            'Issue': 'Absent from A-O but NOT orange-flagged',
            'Action': 'Verify — likely needs A-O extension; team may have missed flagging',
        })
    disagree_df = pd.DataFrame(disagree_rows) if disagree_rows else \
                  pd.DataFrame(columns=['Code', 'Issue', 'Action'])

    # Write workbook
    print(f"\nWriting {OUTPUT_PATH.name}...")
    with pd.ExcelWriter(OUTPUT_PATH, engine='openpyxl') as w:
        presence_df.to_excel(w, sheet_name='1_Extensions_To_Add', index=False)
        if len(param_rows):
            param_rows.to_excel(w, sheet_name='2_Parameter_Rows_To_Replicate', index=False)
        else:
            pd.DataFrame({'Note': ['No parameter rows found']}).to_excel(
                w, sheet_name='2_Parameter_Rows_To_Replicate', index=False)
        disagree_df.to_excel(w, sheet_name='3_Signal_Disagreements', index=False)

    print(f"  ✓ {OUTPUT_PATH.name}")
    print("\nDone.\n")
    print("To answer your question: ABSENCE is the reliable signal.")
    print("Color is a useful cross-check — if a code is absent and orange, both signals agree.")
    print("If they disagree, see Tab 3 of the output workbook.")


if __name__ == '__main__':
    main()
