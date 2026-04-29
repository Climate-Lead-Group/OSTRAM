# -*- coding: utf-8 -*-
"""
OSTRAM v13 — Tech.ID Renumberer
================================

After deleting rows from v13, the Tech.ID column has gaps (e.g.,
1, 2, 3, ..., 10, [skip 11–15], 16, 17, ...). This script renumbers
Tech.ID sequentially from 1 within each sheet, preserving the order
in which Techs currently appear and keeping all parameter rows for
the same Tech grouped together with the same new ID.

Tech.ID is for human readability — it is not used by the AO-based
parser — so this is a cosmetic cleanup. All parameter values, codes,
and other content are preserved untouched.

Output: SOASIA_OSeMOSYS_Template_v13_renumbered.xlsx (formatting preserved)

Requirements: openpyxl
"""

import shutil
from pathlib import Path
from openpyxl import load_workbook


# ════════════════════════════════════════════════════════════════════════
# CONFIG
# ════════════════════════════════════════════════════════════════════════
WORK_DIR = Path(r"C:\Users\luisfernando\Desktop\OSeMOSYS\asia_ostram\t1_confection\test_a3_mod_v2")

V13_INPUT  = WORK_DIR / "SOASIA_OSeMOSYS_Template_v14.xlsx"
V13_OUTPUT = WORK_DIR / "SOASIA_OSeMOSYS_Template_v15.xlsx"


# Sheets to renumber. Limited to the in-scope (blue) sheets that have Tech.ID.
SHEETS_TO_RENUMBER = [
    'Fixed_Horizon_Parameters',
    'Primary_Techs',
    'Secondary_Techs',
    'Demand_Techs',
    'VariableCost',
    'Emissions',
]


def find_col(ws, header_name):
    """Return 1-based column index for a header name in row 1, or None."""
    for col in range(1, ws.max_column + 1):
        if ws.cell(1, col).value == header_name:
            return col
    return None


def renumber_sheet(ws, sheet_name):
    """Renumber the Tech.ID column sequentially based on the order Techs appear.

    Strategy: many parameter rows share the same Tech (with the same Tech.ID
    appearing on multiple consecutive rows). We assign each unique Tech a new
    sequential ID by first appearance, then propagate that new ID to all rows
    sharing the same Tech code.
    """
    techid_col = find_col(ws, 'Tech.ID')
    tech_col = find_col(ws, 'Tech')
    if not tech_col:
        tech_col = find_col(ws, 'Fuel/Tech')
    if not techid_col or not tech_col:
        return 0  # nothing to do

    # First pass: collect (row_idx, tech_code) for every data row
    tech_first_seen = {}  # tech_code → row index of first appearance
    rows = []
    for r in range(2, ws.max_row + 1):
        code = ws.cell(r, tech_col).value
        if code is None or (isinstance(code, str) and code.strip() == ''):
            rows.append((r, None))
            continue
        code = str(code)
        rows.append((r, code))
        if code not in tech_first_seen:
            tech_first_seen[code] = r

    # Assign new IDs in order of first appearance
    ordered_techs = sorted(tech_first_seen.keys(),
                           key=lambda t: tech_first_seen[t])
    new_id = {tech: i + 1 for i, tech in enumerate(ordered_techs)}

    # Second pass: write new IDs to every data row
    rewritten = 0
    for r, code in rows:
        if code is None:
            continue
        ws.cell(r, techid_col).value = new_id[code]
        rewritten += 1

    return rewritten


def main():
    print(f"Copying {V13_INPUT.name} → {V13_OUTPUT.name}...")
    shutil.copy2(V13_INPUT, V13_OUTPUT)

    print(f"Opening {V13_OUTPUT.name}...")
    wb = load_workbook(V13_OUTPUT)

    print("\nRenumbering Tech.ID per sheet:")
    print(f"  {'Sheet':<32}{'rows updated':>16}")
    print("  " + "=" * 50)
    total = 0
    for s in SHEETS_TO_RENUMBER:
        if s not in wb.sheetnames:
            continue
        ws = wb[s]
        n = renumber_sheet(ws, s)
        total += n
        print(f"  {s:<32}{n:>16}")
    print("  " + "=" * 50)
    print(f"  {'TOTAL':<32}{total:>16}")

    print(f"\nSaving {V13_OUTPUT.name}...")
    wb.save(V13_OUTPUT)
    print("Done.")


if __name__ == '__main__':
    main()
