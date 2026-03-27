# -*- coding: utf-8 -*-
"""
A3_update_csvs_from_datapackage.py
Updates OG_csvs_inputs/ CSVs from DATA_PACKAGE_V2 sources.

Reads cost data, power infrastructure, energy statistics, demand analyses,
and VRE capacity factors from DATA_PACKAGE_V2/ Excel/CSV files, then
merges them into the existing OG_csvs_inputs/ CSVs — preserving any
technologies not present in the data package.

Run from t1_confection/:
    python A3_update_csvs_from_datapackage.py

Author: Climate Lead Group
"""

import os
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import yaml

# Reuse shared config loader
from openpyxl.utils.dataframe import dataframe_to_rows

from Z_AUX_config_loader import (
    get_code_to_energy,
    get_countries,
    get_enable_dsptrn,
    get_first_year,
    get_iso_country_map,
    get_model_countries_list,
    get_multi_region_map,
)

SCRIPT_DIR = Path(__file__).resolve().parent

# ============================================================
# 1. CONFIGURATION
# ============================================================

def load_datapackage_config():
    """Load Config_datapackage.yaml."""
    config_path = SCRIPT_DIR / "Config_datapackage.yaml"
    with open(config_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


CFG = load_datapackage_config()
DATA_PKG = SCRIPT_DIR / CFG["data_package_path"]
OG_DIR = SCRIPT_DIR / "OG_csvs_inputs"
FIRST_YEAR = CFG["first_year"]
FINAL_YEAR = CFG["final_year"]
TARGET_YEARS = list(range(FIRST_YEAR, FINAL_YEAR + 1))
COST_SCENARIO = CFG["cost_scenario"]
TECH_NAME_MAP = CFG["tech_name_map"]
AGGREGATED_TECHS = CFG["aggregated_techs"]
AGG_RULES = CFG["aggregation_rules"]
RNW_TECH_CODES = set(CFG.get("rnw_tech_codes", []))
FUEL_TO_TECH = CFG.get("fuel_to_tech", CFG.get("fuel_to_min_tech", {}))

# Milestone years in Cost DB
MILESTONE_YEARS = [2023, 2025, 2030, 2035, 2040, 2045, 2050]
MILESTONE_YEAR_STRS = [str(y) for y in MILESTONE_YEARS]

# Timeslices
TIMESLICES = [f"S{s}D{d}" for s in range(1, 5) for d in range(1, 4)]

# Lazy-loaded lookup dicts (populated once on first use)
_CODE_TO_ENERGY = None
_ISO_COUNTRY_MAP = None


def _get_code_to_energy():
    global _CODE_TO_ENERGY
    if _CODE_TO_ENERGY is None:
        _CODE_TO_ENERGY = get_code_to_energy()
    return _CODE_TO_ENERGY


def _get_iso_country_map():
    global _ISO_COUNTRY_MAP
    if _ISO_COUNTRY_MAP is None:
        _ISO_COUNTRY_MAP = get_iso_country_map()
    return _ISO_COUNTRY_MAP


# ------------------------------------------------------------------
# Utility functions (mirrors of A1_Pre_processing logic)
# ------------------------------------------------------------------

def parse_tech_name(tech):
    """Return a descriptive name for a technology code."""
    code_to_energy = _get_code_to_energy()
    iso_country_map = _get_iso_country_map()
    main_code = tech[0:3]

    # Transmission interconnection codes
    if main_code == "TRN" and len(tech) >= 13:
        iso1, region1 = tech[3:6], tech[6:8]
        iso2, region2 = tech[8:11], tech[11:13]
        c1 = iso_country_map.get(iso1, f"Unknown ({iso1})")
        c2 = iso_country_map.get(iso2, f"Unknown ({iso2})")
        return f"Transmission interconnection from {c1}, region {region1} to {c2}, region {region2}"

    # Storage codes (SDS, LDS)
    if main_code in ("SDS", "LDS") and len(tech) <= 10:
        iso1, region1 = tech[3:6], tech[6:8]
        storage_desc = code_to_energy.get(main_code, "specific technology")
        c1 = iso_country_map.get(iso1, f"Unknown ({iso1})")
        return f"{storage_desc} {c1}, region {region1}"

    iso = tech[6:9]
    region = tech[9:11] if len(tech) > 9 else "XX"
    country = iso_country_map.get(iso, f"Unknown ({iso})")
    sub_code = tech[3:6]
    main_desc = code_to_energy.get(main_code, "General technology")
    sub_desc = code_to_energy.get(sub_code, "specific technology")

    base = f"{sub_desc} ({main_desc})" if main_desc != sub_desc else sub_desc
    name = f"{base} {country}"

    if not tech.startswith("MIN") and region != "XX":
        name += f", region {region}"
    elif region == "XX":
        name += f", region XX"

    return name


def parse_fuel_name(fuel):
    """Return a descriptive name for a fuel code."""
    code_to_energy = _get_code_to_energy()
    iso_country_map = _get_iso_country_map()
    prefix = fuel[0:3]
    iso = fuel[3:6]
    region = fuel[6:8] if len(fuel) >= 8 else None
    suffix = None

    if fuel.endswith("01"):
        suffix = "power plant output"
    elif fuel.endswith("02"):
        suffix = "transmission line output"
    elif fuel.endswith("03"):
        suffix = "dispatch output"

    fuel_type = code_to_energy.get(prefix, "Unknown")
    country = iso_country_map.get(iso, f"Unknown ({iso})")

    parts = [fuel_type, country]
    if region and region != "XX":
        parts.append(f"region {region}")
    elif region == "XX":
        parts.append("region XX")
    if suffix:
        parts.append(suffix)
    return ", ".join(parts)


def assign_tech_type(tech):
    """Classify a technology code into Primary / Secondary / Demand."""
    if tech.startswith("MIN") or tech.startswith("RNW"):
        return "Primary"
    elif tech.startswith("PWRTRN"):
        return "Demand"
    return "Secondary"


# ============================================================
# 2. INTERPOLATION / EXTRAPOLATION
# ============================================================

def interpolate_linear(known_points, target_years):
    """Linear interpolation between known points.
    For years before the first point: hold flat.
    Does NOT extrapolate beyond the last point.

    known_points: {year: value, ...}
    target_years: [2023, 2024, ..., 2050]
    Returns: {year: value, ...}
    """
    if not known_points:
        return {}
    sorted_years = sorted(known_points.keys())
    first_year = sorted_years[0]
    last_year = sorted_years[-1]
    result = {}
    for y in target_years:
        if y in known_points:
            result[y] = known_points[y]
        elif y < first_year:
            result[y] = known_points[first_year]
        elif y > last_year:
            continue  # Don't extrapolate
        else:
            # Find bracketing years
            lo = max(yr for yr in sorted_years if yr <= y)
            hi = min(yr for yr in sorted_years if yr >= y)
            if lo == hi:
                result[y] = known_points[lo]
            else:
                frac = (y - lo) / (hi - lo)
                result[y] = known_points[lo] + frac * (known_points[hi] - known_points[lo])
    return result


def extrapolate_flat_rate(last_known_year, last_known_value, annual_rate, target_years):
    """Compound growth extrapolation from a known endpoint.
    value(y) = value(y-1) * (1 + annual_rate)
    Only for years > last_known_year.
    """
    result = {}
    for y in sorted(target_years):
        if y <= last_known_year:
            continue
        dt = y - last_known_year
        result[y] = last_known_value * (1 + annual_rate) ** dt
    return result


def fill_annual_values(known_points, target_years, extrapolation_rate=None):
    """Unified function:
    1. Linear interpolation between known points
    2. Hold flat for backward extrapolation (before first point)
    3. Flat rate compound growth for forward extrapolation (after last point)
    """
    result = interpolate_linear(known_points, target_years)
    if extrapolation_rate is not None and known_points:
        last_year = max(known_points.keys())
        last_value = known_points[last_year]
        missing_years = [y for y in target_years if y > last_year and y not in result]
        if missing_years:
            ext = extrapolate_flat_rate(last_year, last_value, extrapolation_rate, missing_years)
            result.update(ext)
    return result


# ============================================================
# 3. COST DATABASE READER
# ============================================================

def read_cost_database():
    """Read SoAsia_OSTRAM_Cost_Database.xlsx.
    Returns dict with keys: technologies, fuel_costs, node_anchors
    Each is a list of dicts (row-wise).
    """
    path = DATA_PKG / "costs" / "SoAsia_OSTRAM_Cost_Database.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # Technologies sheet
    ws = wb["Technologies"]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    header = rows[0]
    techs = []
    for row in rows[1:]:
        if row[0] is None:
            continue
        d = dict(zip(header, row))
        techs.append(d)

    # Fuel_Costs sheet
    ws = wb["Fuel_Costs"]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    header = rows[0]
    fuels = []
    for row in rows[1:]:
        if row[0] is None:
            continue
        d = dict(zip(header, row))
        fuels.append(d)

    # Node_Anchors sheet
    ws = wb["Node_Anchors"]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    header = rows[0]
    anchors = []
    for row in rows[1:]:
        if row[0] is None:
            continue
        d = dict(zip(header, row))
        anchors.append(d)

    wb.close()
    return {"technologies": techs, "fuel_costs": fuels, "node_anchors": anchors}


def _parse_multiplier(val):
    """Parse a multiplier value. Handles: float, 'N/A', None, ranges like '1.5-2.0'."""
    if val is None or val == "N/A" or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if s.upper() == "N/A":
        return None
    # Check for range like "1.5-2.0"
    if "-" in s and not s.startswith("-"):
        parts = s.split("-")
        try:
            return (float(parts[0]) + float(parts[1])) / 2.0
        except ValueError:
            return None
    try:
        return float(s)
    except ValueError:
        return None


def get_country_multipliers(node_anchors):
    """Build {costdb_code: {iso3: multiplier}} from Node_Anchors.
    India (IN) is base = 1.0.
    """
    node_to_iso3 = CFG["node_to_iso3"]
    # Column name patterns: BD_mult, LK_mult, NP_mult, BT_mult, MV_mult
    mult_cols = {
        "BD": "BD_mult",
        "LK": "LK_mult",
        "NP": "NP_mult",
        "BT": "BT_mult",
        "MV": "MV_mult",
    }
    result = {}
    for anchor in node_anchors:
        code = anchor["Code"]
        mults = {"IND": 1.0}  # India is always base
        for node_2char, col_name in mult_cols.items():
            iso3 = node_to_iso3[node_2char]
            val = _parse_multiplier(anchor.get(col_name))
            if val is not None:
                mults[iso3] = val
        result[code] = mults
    return result


def _get_milestone_values(row_dict):
    """Extract {year: value} for milestone years from a Cost DB row."""
    result = {}
    for ys in MILESTONE_YEAR_STRS:
        val = row_dict.get(ys)
        if val is not None and val != "" and not (isinstance(val, str) and val.strip() == ""):
            try:
                result[int(ys)] = float(val)
            except (ValueError, TypeError):
                pass
    return result


def _get_regions_for_country(iso3):
    """Get list of OSTRAM region codes for a country.
    IND → [INDEA, INDNE, INDNO, INDSO, INDWE]
    BGD → [BGDXX]
    """
    multi = get_multi_region_map()
    if iso3 in multi:
        return [f"{iso3}{reg}" for reg in multi[iso3]]
    return [f"{iso3}XX"]


def _is_implausible(tech_code, iso3):
    """Check if tech-country combo is implausible (from Config_country_codes.yaml)."""
    from Z_AUX_config_loader import get_raw_config
    cfg = get_raw_config()
    impl = cfg.get("implausible_combinations", {})
    blocked_countries = impl.get(tech_code, [])
    return iso3 in blocked_countries


def build_fuel_tech_name(tech_code, region):
    """Build the full technology name for a fuel supply technology.

    Non-renewable codes use MIN prefix + country 3-char:
        e.g. COA + INDEA → MINCOAIND, COA + NPLXX → MINCOANPL
    Renewable codes use RNW prefix + full region 5-char:
        e.g. BIO + INDEA → RNWBIOINDEA, BIO + LKAXX → RNWBIOLKAXX
    """
    if tech_code in RNW_TECH_CODES:
        return f"RNW{tech_code}{region}"
    return f"MIN{tech_code}{region[:3]}"


def generate_capital_cost(cost_db, existing_df):
    """Generate CapitalCost.csv from Cost DB Technologies + Node_Anchors."""
    techs = cost_db["technologies"]
    multipliers = get_country_multipliers(cost_db["node_anchors"])
    countries = get_countries()  # ['BGD', 'BTN', 'IND', 'LKA', 'MDV', 'NPL']

    rows = []
    # Filter for CAPEX rows with matching scenario
    capex_rows = [t for t in techs
                  if t["Parameter"] == "CAPEX"
                  and t["Scenario"] in (COST_SCENARIO, "All")]

    for crow in capex_rows:
        costdb_code = crow["Code"]
        ostram_code = TECH_NAME_MAP.get(costdb_code)
        if ostram_code is None:
            continue

        milestone_vals = _get_milestone_values(crow)
        if not milestone_vals:
            continue

        # Interpolate to annual
        annual_vals = fill_annual_values(milestone_vals, TARGET_YEARS)

        # Get multipliers for this tech
        tech_mults = multipliers.get(costdb_code, {"IND": 1.0})

        for iso3 in countries:
            if _is_implausible(ostram_code, iso3):
                continue
            mult = tech_mults.get(iso3)
            if mult is None:
                # Use India base if no specific multiplier
                mult = tech_mults.get("IND", 1.0)

            regions = _get_regions_for_country(iso3)
            for region in regions:
                tech_full = f"PWR{ostram_code}{region}"
                for y, v in annual_vals.items():
                    # Convert USD/kW to M$/GW: USD/kW = k$/MW = M$/GW
                    # OSeMOSYS uses M$/GW which equals USD/kW numerically
                    rows.append({
                        "REGION": region,
                        "TECHNOLOGY": tech_full,
                        "YEAR": y,
                        "VALUE": round(v * mult, 2),
                    })

    new_df = pd.DataFrame(rows)

    # Aggregate N:1 technologies (e.g., PWRSPV from SPV+SPR+FPV)
    new_df = _aggregate_technologies(new_df, "CapitalCost")

    return new_df


def generate_fixed_cost(cost_db, existing_df):
    """Generate FixedCost.csv from Cost DB Technologies + Node_Anchors."""
    techs = cost_db["technologies"]
    multipliers = get_country_multipliers(cost_db["node_anchors"])
    countries = get_countries()

    rows = []
    fom_rows = [t for t in techs
                if t["Parameter"] == "Fixed O&M"
                and t["Scenario"] in (COST_SCENARIO, "All")]

    for frow in fom_rows:
        costdb_code = frow["Code"]
        ostram_code = TECH_NAME_MAP.get(costdb_code)
        if ostram_code is None:
            continue

        milestone_vals = _get_milestone_values(frow)
        if not milestone_vals:
            continue

        annual_vals = fill_annual_values(milestone_vals, TARGET_YEARS)
        tech_mults = multipliers.get(costdb_code, {"IND": 1.0})

        for iso3 in countries:
            if _is_implausible(ostram_code, iso3):
                continue
            mult = tech_mults.get(iso3, tech_mults.get("IND", 1.0))
            regions = _get_regions_for_country(iso3)
            for region in regions:
                tech_full = f"PWR{ostram_code}{region}"
                for y, v in annual_vals.items():
                    # USD/kW/yr = M$/GW/yr in OSeMOSYS
                    rows.append({
                        "REGION": region,
                        "TECHNOLOGY": tech_full,
                        "YEAR": y,
                        "VALUE": round(v * mult, 2),
                    })

    new_df = pd.DataFrame(rows)
    new_df = _aggregate_technologies(new_df, "FixedCost")
    return new_df


def generate_variable_cost(cost_db, existing_df):
    """Generate VariableCost.csv from Cost DB (Variable O&M for PWR + Fuel_Costs for MIN)."""
    techs = cost_db["technologies"]
    multipliers = get_country_multipliers(cost_db["node_anchors"])
    countries = get_countries()
    fuel_to_tech = FUEL_TO_TECH
    fuel_node_to_regions = CFG["fuel_node_to_regions"]

    rows = []

    # --- Part 1: Variable O&M for PWR technologies (MODE_OF_OPERATION=1) ---
    vom_rows = [t for t in techs
                if t["Parameter"] == "Variable O&M"
                and t["Scenario"] in (COST_SCENARIO, "All")]

    for vrow in vom_rows:
        costdb_code = vrow["Code"]
        ostram_code = TECH_NAME_MAP.get(costdb_code)
        if ostram_code is None:
            continue

        milestone_vals = _get_milestone_values(vrow)
        if not milestone_vals:
            continue
        # Check if all zeros
        if all(v == 0 for v in milestone_vals.values()):
            continue

        annual_vals = fill_annual_values(milestone_vals, TARGET_YEARS)
        tech_mults = multipliers.get(costdb_code, {"IND": 1.0})

        for iso3 in countries:
            if _is_implausible(ostram_code, iso3):
                continue
            mult = tech_mults.get(iso3, tech_mults.get("IND", 1.0))
            regions = _get_regions_for_country(iso3)
            for region in regions:
                tech_full = f"PWR{ostram_code}{region}"
                for y, v in annual_vals.items():
                    # Cost DB gives USD/MWh; OSeMOSYS VariableCost is M$/PJ
                    # 1 PJ = 277,778 MWh → USD/MWh × 277778 / 1e6 = USD/MWh × 0.277778
                    v_mpj = v * 0.277778
                    rows.append({
                        "REGION": region,
                        "TECHNOLOGY": tech_full,
                        "MODE_OF_OPERATION": 1,
                        "YEAR": y,
                        "VALUE": round(v_mpj * mult, 4),
                    })

    # --- Part 2: Fuel costs for MIN/RNW technologies (MODE_OF_OPERATION=1) ---
    fuel_costs = cost_db["fuel_costs"]
    fc_rows = [f for f in fuel_costs
               if f["Scenario"] in (COST_SCENARIO, "All")]

    for frow in fc_rows:
        fuel_name = frow["Fuel"]
        ostram_tech = fuel_to_tech.get(fuel_name)
        if ostram_tech is None:
            continue

        node_override = str(frow.get("Node_Override", "All") or "All").strip()
        target_regions = fuel_node_to_regions.get(node_override, [])
        if not target_regions:
            continue

        milestone_vals = _get_milestone_values(frow)
        if not milestone_vals:
            continue

        annual_vals = fill_annual_values(milestone_vals, TARGET_YEARS)

        for region in target_regions:
            iso3 = region[:3]
            if _is_implausible(ostram_tech, iso3):
                continue
            tech_full = build_fuel_tech_name(ostram_tech, region)
            for y, v in annual_vals.items():
                rows.append({
                    "REGION": region,
                    "TECHNOLOGY": tech_full,
                    "MODE_OF_OPERATION": 1,
                    "YEAR": y,
                    "VALUE": round(v, 4),
                })

    new_df = pd.DataFrame(rows)
    # Aggregate N:1 for PWR techs
    new_df = _aggregate_technologies(new_df, "VariableCost")
    return new_df


def generate_operational_life(cost_db, existing_df):
    """Generate OperationalLife.csv from Cost DB Technologies."""
    techs = cost_db["technologies"]
    countries = get_countries()

    rows = []
    ol_rows = [t for t in techs
               if t["Parameter"] == "Operational Life"
               and t["Scenario"] in (COST_SCENARIO, "All")]

    for orow in ol_rows:
        costdb_code = orow["Code"]
        ostram_code = TECH_NAME_MAP.get(costdb_code)
        if ostram_code is None:
            continue

        # Operational life is constant — take value from any milestone year
        milestone_vals = _get_milestone_values(orow)
        if not milestone_vals:
            continue
        op_life = list(milestone_vals.values())[0]  # Should be same for all years

        for iso3 in countries:
            if _is_implausible(ostram_code, iso3):
                continue
            regions = _get_regions_for_country(iso3)
            for region in regions:
                tech_full = f"PWR{ostram_code}{region}"
                rows.append({
                    "REGION": region,
                    "TECHNOLOGY": tech_full,
                    "VALUE": op_life,
                })

    new_df = pd.DataFrame(rows)
    new_df = _aggregate_technologies(new_df, "OperationalLife")
    return new_df


def _aggregate_technologies(df, param_name):
    """Aggregate N:1 tech mappings using the appropriate rule (AVG or SUM).
    Multiple rows for the same (REGION, TECHNOLOGY, YEAR, ...) are combined.
    """
    if df.empty:
        return df

    # Determine aggregation method
    if param_name in AGG_RULES.get("AVG", []):
        agg_func = "mean"
    elif param_name in AGG_RULES.get("SUM", []):
        agg_func = "sum"
    else:
        agg_func = "mean"  # Default to average

    # Group by all columns except VALUE
    group_cols = [c for c in df.columns if c != "VALUE"]
    if not group_cols:
        return df

    df = df.groupby(group_cols, as_index=False)["VALUE"].agg(agg_func)
    df["VALUE"] = df["VALUE"].round(4)
    return df


# ============================================================
# 4. VRE CAPACITY FACTORS & YEARSPLIT
# ============================================================

def generate_capacity_factor(cost_db, existing_df):
    """Generate CapacityFactor.csv from VRE CSV + Demand Analysis thermal/hydro CFs."""
    vre_resource_to_tech = CFG["vre_resource_to_tech"]

    # --- Part 1: VRE from Renewable Ninja CSV ---
    vre_path = DATA_PKG / "reno_ninja" / "SoAsia_OSTRAM_VRE_Capacity_Factors.csv"
    vre_df = pd.read_csv(vre_path)

    # Average cf_mean across all model_years for each (resource, region, timeslice)
    vre_avg = vre_df.groupby(["resource", "region", "timeslice"])["cf_mean"].mean().reset_index()

    vre_rows = []
    for _, r in vre_avg.iterrows():
        tech_code = vre_resource_to_tech.get(r["resource"])
        if tech_code is None:
            continue
        region = r["region"]
        iso3 = region[:3]
        if _is_implausible(tech_code, iso3):
            continue
        tech_full = f"PWR{tech_code}{region}"
        ts = r["timeslice"]
        cf = r["cf_mean"]
        for y in TARGET_YEARS:
            vre_rows.append({
                "REGION": region,
                "TECHNOLOGY": tech_full,
                "TIMESLICE": ts,
                "YEAR": y,
                "VALUE": round(cf, 6),
            })

    # --- Part 2: Thermal/Hydro CFs from Demand Analysis ---
    thermal_rows = []
    countries = get_countries()
    # Tech column mapping in OSTRAM_Profiles sheets
    demand_analysis_tech_map = {
        "Gas": "NGS",
        "Coal": "COA",
        "Oil_HFO": "OIL",
        "Oil_HSD": "PET",
        "Hydro": "HYD",
        "Import": None,  # Skip
    }

    for iso3 in countries:
        if iso3 == "MDV":
            # MDV has no Demand_Analysis.xlsx, only assessment markdown
            continue
        if iso3 == "IND":
            # India has different format — skip thermal CFs from demand analysis
            # India thermal CFs come from the ICED PLF data
            continue

        da_path = DATA_PKG / "more_timeslices" / f"{iso3}_Demand_Analysis.xlsx"
        if not da_path.exists():
            continue

        wb = openpyxl.load_workbook(da_path, read_only=True, data_only=True)
        ws = wb["1_OSTRAM_Profiles"]
        all_rows = list(ws.iter_rows(min_row=1, values_only=True))
        wb.close()

        # Find the header row with timeslice data
        header_row_idx = None
        for i, row in enumerate(all_rows):
            if row[0] and str(row[0]).strip() == "Timeslice":
                header_row_idx = i
                break
        if header_row_idx is None:
            continue

        header = all_rows[header_row_idx]
        # Parse column names — strip newlines and MW values
        col_names = []
        for h in header:
            if h is None:
                col_names.append(None)
                continue
            name = str(h).split("\n")[0].strip()
            col_names.append(name)

        # Read data rows (12 timeslices: S1D1..S4D3)
        data_rows = all_rows[header_row_idx + 1: header_row_idx + 13]
        regions = _get_regions_for_country(iso3)

        for drow in data_rows:
            ts = drow[0]
            if ts is None or not str(ts).startswith("S"):
                continue
            ts = str(ts).strip()

            for col_idx, col_name in enumerate(col_names):
                if col_name is None or col_idx >= len(drow):
                    continue
                tech_code = demand_analysis_tech_map.get(col_name)
                if tech_code is None:
                    continue
                if _is_implausible(tech_code, iso3):
                    continue

                val = drow[col_idx]
                if val is None:
                    continue
                try:
                    cf = float(val)
                except (ValueError, TypeError):
                    continue

                for region in regions:
                    tech_full = f"PWR{tech_code}{region}"
                    for y in TARGET_YEARS:
                        thermal_rows.append({
                            "REGION": region,
                            "TECHNOLOGY": tech_full,
                            "TIMESLICE": ts,
                            "YEAR": y,
                            "VALUE": round(cf, 6),
                        })

    all_cf_rows = vre_rows + thermal_rows
    new_df = pd.DataFrame(all_cf_rows)
    if not new_df.empty:
        new_df = _aggregate_technologies(new_df, "CapacityFactor")
    return new_df


def generate_year_split(existing_df):
    """Generate YearSplit.csv from VRE CSV yearsplit column."""
    vre_path = DATA_PKG / "reno_ninja" / "SoAsia_OSTRAM_VRE_Capacity_Factors.csv"
    vre_df = pd.read_csv(vre_path)

    # Average yearsplit across all resources/regions/years for each timeslice
    ys_avg = vre_df.groupby("timeslice")["yearsplit"].mean().reset_index()

    # Normalize to sum = 1.0
    total = ys_avg["yearsplit"].sum()
    if total > 0:
        ys_avg["yearsplit"] = ys_avg["yearsplit"] / total

    rows = []
    for _, r in ys_avg.iterrows():
        ts = r["timeslice"]
        val = r["yearsplit"]
        for y in TARGET_YEARS:
            rows.append({
                "TIMESLICE": ts,
                "YEAR": y,
                "VALUE": round(val, 10),
            })

    return pd.DataFrame(rows)


# ============================================================
# 5. RESIDUAL CAPACITY
# ============================================================

def _read_power_infrastructure(iso3):
    """Read Existing Generation sheet from a country's Power Infrastructure file."""
    path = DATA_PKG / "power" / f"{iso3}_Power_Infrastructure.xlsx"
    if not path.exists():
        print(f"  WARNING: {path} not found")
        return pd.DataFrame()

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Existing Generation"]
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    # Find header row (contains "Plant Name" or "Installed Capacity")
    header_idx = None
    for i, row in enumerate(all_rows):
        if row and any(str(c).strip() == "Plant Name" for c in row if c):
            header_idx = i
            break
    if header_idx is None:
        return pd.DataFrame()

    header = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(all_rows[header_idx])]
    data = []
    for row in all_rows[header_idx + 1:]:
        if row[0] is None:
            continue
        d = {}
        for j, h in enumerate(header):
            d[h] = row[j] if j < len(row) else None
        data.append(d)

    return pd.DataFrame(data)


def _get_operational_life_map(cost_db):
    """Build {ostram_code: op_life_years} from Cost DB."""
    techs = cost_db["technologies"]
    result = {}
    for t in techs:
        if t["Parameter"] == "Operational Life" and t["Scenario"] in ("All", COST_SCENARIO):
            code = TECH_NAME_MAP.get(t["Code"])
            if code and code not in result:
                vals = _get_milestone_values(t)
                if vals:
                    result[code] = list(vals.values())[0]
    return result


def generate_residual_capacity(cost_db, existing_df):
    """Generate ResidualCapacity.csv from Power Infrastructure files."""
    countries = get_countries()
    power_fuel_map = CFG["power_fuel_type_to_tech"]
    op_life_map = _get_operational_life_map(cost_db)
    india_zone_map = CFG["india_zone_to_region"]

    rows = []

    for iso3 in countries:
        pi_df = _read_power_infrastructure(iso3)
        if pi_df.empty:
            continue

        # Get the capacity column name
        cap_col = None
        for c in pi_df.columns:
            if "Installed Capacity" in c:
                cap_col = c
                break
        if cap_col is None:
            print(f"  WARNING: No 'Installed Capacity' column in {iso3} Power Infrastructure")
            continue

        # Get fuel type column
        fuel_col = "Fuel Type"
        if fuel_col not in pi_df.columns:
            print(f"  WARNING: No 'Fuel Type' column in {iso3} Power Infrastructure")
            continue

        # Determine region assignment
        if iso3 == "IND":
            zone_col = "Zone"
        else:
            zone_col = None

        # Commission date column (IND has it)
        comm_col = None
        for c in pi_df.columns:
            if "Commissioning" in c or "Commission" in c:
                comm_col = c
                break

        # Aggregate by (fuel_type, region) → total MW
        for _, plant in pi_df.iterrows():
            fuel_type = str(plant.get(fuel_col, "")).strip()
            tech_code = power_fuel_map.get(fuel_type)
            if tech_code is None:
                continue

            cap_mw = plant.get(cap_col)
            if cap_mw is None:
                continue
            try:
                cap_mw = float(cap_mw)
            except (ValueError, TypeError):
                continue
            if cap_mw <= 0:
                continue

            if _is_implausible(tech_code, iso3):
                continue

            # Determine region
            if iso3 == "IND" and zone_col and zone_col in plant.index:
                zone = str(plant.get(zone_col, "")).strip()
                region = india_zone_map.get(zone)
                if region is None:
                    continue
            else:
                regions = _get_regions_for_country(iso3)
                region = regions[0]  # Single-region country

            # Get commissioning year if available
            comm_year = None
            if comm_col and comm_col in plant.index:
                cv = plant.get(comm_col)
                if cv is not None:
                    cv_str = str(cv).strip()
                    # Try to extract year
                    match = re.search(r"(\d{4})", cv_str)
                    if match:
                        comm_year = int(match.group(1))
                    elif cv_str.startswith("Pre-"):
                        # "Pre-2000" → assume 2000
                        match2 = re.search(r"Pre-(\d{4})", cv_str)
                        if match2:
                            comm_year = int(match2.group(1))

            # Get operational life
            ol = op_life_map.get(tech_code, 30)  # Default 30 years

            # Calculate residual capacity curve
            tech_full = f"PWR{tech_code}{region}"
            cap_gw = cap_mw / 1000.0

            if comm_year and comm_year > 1900:
                # Known commissioning year: retire after ol years
                retire_year = comm_year + int(ol)
                for y in TARGET_YEARS:
                    if y < retire_year:
                        rows.append({
                            "REGION": region,
                            "TECHNOLOGY": tech_full,
                            "YEAR": y,
                            "VALUE": cap_gw,
                        })
            else:
                # Unknown commissioning: linear retirement from base year
                for y in TARGET_YEARS:
                    fraction = max(0.0, 1.0 - (y - FIRST_YEAR) / ol)
                    if fraction > 0:
                        rows.append({
                            "REGION": region,
                            "TECHNOLOGY": tech_full,
                            "YEAR": y,
                            "VALUE": cap_gw * fraction,
                        })

    new_df = pd.DataFrame(rows)
    if new_df.empty:
        return new_df

    # Sum by (REGION, TECHNOLOGY, YEAR) — multiple plants aggregate
    new_df = new_df.groupby(["REGION", "TECHNOLOGY", "YEAR"], as_index=False)["VALUE"].sum()
    new_df["VALUE"] = new_df["VALUE"].round(4)
    return new_df


# ============================================================
# 6. SPECIFIED ANNUAL DEMAND
# ============================================================

def _read_bgd_demand():
    """Read Bangladesh demand: IEPMP 2023 milestones from 2_Demand_Forecasts."""
    path = DATA_PKG / "energy" / "BGD_Energy_Statistics.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["2_Demand_Forecasts"]
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    # Find section B header row, then data
    known_points = {}
    for row in all_rows:
        if row[0] is None:
            continue
        try:
            year = int(row[0])
        except (ValueError, TypeError):
            continue
        if 2020 <= year <= 2060:
            # Peak Demand GW Low (col B = index 1)
            val = row[1]
            if val is not None:
                try:
                    peak_gw = float(val)
                    # Convert peak GW to annual GWh using load factor ~0.6
                    # Then GWh to PJ: GWh * 0.0036
                    # Peak GW * 8760 * LF = GWh
                    # IEPMP uses ~60% load factor for Bangladesh
                    gwh = peak_gw * 8760 * 0.60
                    known_points[year] = gwh * 0.0036  # PJ
                except (ValueError, TypeError):
                    pass
    return known_points


def _read_ind_demand():
    """Read India demand by region from 5_Demand_Energy_State."""
    path = DATA_PKG / "energy" / "IND_Energy_Statistics.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["5_Demand_Energy_State"]
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    # Row 2 is header: Region, State/UT, FY 2016-17, ..., FY 2036-37
    header = all_rows[1]
    # Extract fiscal years from header
    fy_cols = {}
    for i, h in enumerate(header):
        if h and str(h).startswith("FY "):
            # "FY 2016-17" → calendar year 2016 (start of FY)
            match = re.search(r"FY (\d{4})", str(h))
            if match:
                fy_cols[i] = int(match.group(1))

    india_region_map = CFG["india_state_region_to_ostram"]

    # Accumulate demand by region
    # Look for summary rows (► Northern Region, etc.)
    region_demand = {}  # {ostram_region: {year: GWh}}
    for row in all_rows[2:]:
        if row[0] is None and row[1] and str(row[1]).startswith("►"):
            # Summary row
            region_name = str(row[1]).replace("►", "").strip()
            ostram_region = india_region_map.get(region_name)
            if ostram_region is None:
                continue
            points = {}
            for col_idx, cal_year in fy_cols.items():
                if col_idx < len(row) and row[col_idx] is not None:
                    try:
                        gwh = float(row[col_idx])  # MU = GWh
                        points[cal_year] = gwh
                    except (ValueError, TypeError):
                        pass
            if points:
                region_demand[ostram_region] = points

    # Convert GWh to PJ
    for region in region_demand:
        region_demand[region] = {y: v * 0.0036 for y, v in region_demand[region].items()}

    return region_demand


def _read_lka_demand():
    """Read Sri Lanka demand from Demand Forecast sheet."""
    path = DATA_PKG / "energy" / "LKA_Energy_Statistics.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Demand Forecast"]
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    known_points = {}
    for row in all_rows:
        if row[0] is None:
            continue
        try:
            year = int(row[0])
        except (ValueError, TypeError):
            continue
        if 2020 <= year <= 2060 and row[1] is not None:
            try:
                gwh = float(row[1])
                known_points[year] = gwh * 0.0036  # PJ
            except (ValueError, TypeError):
                pass
    return known_points


def _read_npl_demand():
    """Read Nepal demand from CES Reference scenario."""
    path = DATA_PKG / "energy" / "NPL_Energy_Statistics.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Demand Forecasts"]
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    # Find "CES / LTS Forecast (Reference 7% GDP)" row
    known_points = {}
    for row in all_rows:
        if row[0] is None:
            continue
        if "CES" in str(row[0]) and "Reference" in str(row[0]):
            # Row 13: columns correspond to years 2020, 2025, 2030, ..., 2050
            milestone_years_npl = [2020, 2025, 2030, 2035, 2040, 2045, 2050]
            for i, my in enumerate(milestone_years_npl):
                col_idx = i + 2  # Starts at column C (index 2)
                if col_idx < len(row) and row[col_idx] is not None:
                    try:
                        twh = float(row[col_idx])
                        known_points[my] = twh * 3.6  # TWh to PJ
                    except (ValueError, TypeError):
                        pass
            break
    return known_points


def _read_btn_demand():
    """Read Bhutan demand: PSMP 2040 milestones."""
    # PSMP gives: 2018 actual = 2328 GWh, 2030 forecast = 5317 GWh
    # Use BPC PDB 2024 for more recent: 2024 = 7125 GWh (total sales)
    # However 7125 includes HV industrial (much of which is ferroalloy/heavy industry)
    # Use PSMP figures for consistency with planning
    known_points = {
        2018: 2328 * 0.0036,   # GWh to PJ
        2030: 5317 * 0.0036,
    }
    return known_points


def _read_mdv_demand():
    """Read Maldives demand: MCCEE Road Map 2024-2033."""
    # 2023: ~1284 GWh (IRENA); 2028: ~2400 GWh (Road Map target)
    known_points = {
        2023: 1284 * 0.0036,
        2028: 2400 * 0.0036,
    }
    return known_points


def generate_specified_annual_demand(existing_df):
    """Generate SpecifiedAnnualDemand.csv from energy statistics."""
    extrapolation_rates = CFG["extrapolation_rates"]
    rows = []

    # --- Bangladesh ---
    bgd_points = _read_bgd_demand()
    if bgd_points:
        bgd_rate = extrapolation_rates["BGD"]["rate"]
        bgd_annual = fill_annual_values(bgd_points, TARGET_YEARS, bgd_rate)
        for y, v in bgd_annual.items():
            rows.append({"REGION": "BGDXX", "FUEL": "ELCBGDXX02", "YEAR": y, "VALUE": round(v, 4)})

    # --- India (5 regions) ---
    ind_demand = _read_ind_demand()
    ind_rate = extrapolation_rates["IND"]["rate"]
    for region, points in ind_demand.items():
        annual = fill_annual_values(points, TARGET_YEARS, ind_rate)
        # Fuel code: ELC{region}02
        fuel_code = f"ELC{region}02"
        for y, v in annual.items():
            rows.append({"REGION": region, "FUEL": fuel_code, "YEAR": y, "VALUE": round(v, 4)})

    # --- Sri Lanka ---
    lka_points = _read_lka_demand()
    if lka_points:
        lka_rate = extrapolation_rates["LKA"]["rate"]
        lka_annual = fill_annual_values(lka_points, TARGET_YEARS, lka_rate)
        for y, v in lka_annual.items():
            rows.append({"REGION": "LKAXX", "FUEL": "ELCLKAXX02", "YEAR": y, "VALUE": round(v, 4)})

    # --- Nepal ---
    npl_points = _read_npl_demand()
    if npl_points:
        npl_rate = extrapolation_rates["NPL"]["rate"]
        npl_annual = fill_annual_values(npl_points, TARGET_YEARS, npl_rate)
        for y, v in npl_annual.items():
            rows.append({"REGION": "NPLXX", "FUEL": "ELCNPLXX02", "YEAR": y, "VALUE": round(v, 4)})

    # --- Bhutan ---
    btn_points = _read_btn_demand()
    if btn_points:
        btn_rate = extrapolation_rates["BTN"]["rate"]
        btn_annual = fill_annual_values(btn_points, TARGET_YEARS, btn_rate)
        for y, v in btn_annual.items():
            rows.append({"REGION": "BTNXX", "FUEL": "ELCBTNXX02", "YEAR": y, "VALUE": round(v, 4)})

    # --- Maldives ---
    mdv_points = _read_mdv_demand()
    if mdv_points:
        mdv_rate = extrapolation_rates["MDV"]["rate"]
        mdv_annual = fill_annual_values(mdv_points, TARGET_YEARS, mdv_rate)
        for y, v in mdv_annual.items():
            rows.append({"REGION": "MDVXX", "FUEL": "ELCMDVXX02", "YEAR": y, "VALUE": round(v, 4)})

    return pd.DataFrame(rows)


# ============================================================
# 7. SPECIFIED DEMAND PROFILE
# ============================================================

def _read_demand_fractions_standard(iso3):
    """Read demand fractions from 1_OSTRAM_Profiles for non-IND countries.
    Returns {timeslice: fraction} (12 entries summing to ~1.0).

    All country sheets share a consistent pattern:
    - A header row containing "Demand Fraction" (always column index 4)
    - 12 data rows below (S1D1 through S4D3) with the fraction in that column
    """
    da_path = DATA_PKG / "more_timeslices" / f"{iso3}_Demand_Analysis.xlsx"
    if not da_path.exists():
        return None

    wb = openpyxl.load_workbook(da_path, read_only=True, data_only=True)
    ws = wb["1_OSTRAM_Profiles"]
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    # Find the header row containing "Demand Fraction"
    demand_frac_col = None
    header_row_idx = None
    for i, row in enumerate(all_rows):
        for j, cell in enumerate(row):
            if cell and "Demand Fraction" in str(cell):
                demand_frac_col = j
                header_row_idx = i
                break
        if header_row_idx is not None:
            break

    if header_row_idx is None or demand_frac_col is None:
        return None

    # Read the 12 timeslice rows below the header
    fractions = {}
    for row in all_rows[header_row_idx + 1:]:
        if row[0] is None:
            continue
        ts = str(row[0]).strip()
        if ts in TIMESLICES:
            val = row[demand_frac_col]
            if val is not None:
                try:
                    fractions[ts] = float(val)
                except (ValueError, TypeError):
                    pass
        if ts == "TOTAL":
            break

    return fractions if fractions else None


def _read_demand_fractions_india():
    """Read India demand fractions from IND_Demand_Analysis.xlsx → Demand Fractions sheet.
    Returns {ostram_region: {timeslice: fraction}}.
    """
    da_path = DATA_PKG / "more_timeslices" / "IND_Demand_Analysis.xlsx"
    wb = openpyxl.load_workbook(da_path, read_only=True, data_only=True)
    ws = wb["Demand Fractions"]
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    # Structure: blocks per region (India_East, India_North, etc.)
    # Each block has: Season, D1_Night, D2_Day, D3_Evening, Season Total
    region_name_map = {
        "India_East": "INDEA",
        "India_North": "INDNO",
        "India_West": "INDWE",
        "India_South": "INDSO",
        "India_NorthEast": "INDNE",
    }
    # Also try alternative names
    alt_names = {
        "India East": "INDEA",
        "India North": "INDNO",
        "India West": "INDWE",
        "India South": "INDSO",
        "India NorthEast": "INDNE",
        "India North East": "INDNE",
    }
    region_name_map.update(alt_names)

    result = {}
    current_region = None

    season_to_code = {
        "S1_Winter": "S1",
        "S2_PreMonsoon": "S2",
        "S3_SWMonsoon": "S3",
        "S4_PostMonsoon": "S4",
    }

    for row in all_rows:
        if row[0] is None and row[1] is None:
            continue

        # Check for region header
        cell0 = str(row[0]).strip() if row[0] else ""
        for rname, rcode in region_name_map.items():
            if cell0 == rname:
                current_region = rcode
                result[current_region] = {}
                break

        # Check for season data row
        if current_region and cell0 in season_to_code:
            s_code = season_to_code[cell0]
            # D1_Night = col 1, D2_Day = col 2, D3_Evening = col 3
            for d_idx, d_code in [(1, "D1"), (2, "D2"), (3, "D3")]:
                if d_idx < len(row) and row[d_idx] is not None:
                    try:
                        frac = float(row[d_idx])
                        ts = f"{s_code}{d_code}"
                        result[current_region][ts] = frac
                    except (ValueError, TypeError):
                        pass

    return result


def generate_specified_demand_profile(existing_df):
    """Generate SpecifiedDemandProfile.csv from Demand Analysis files."""
    countries = get_countries()
    rows = []

    # India — special handling with 5 regions
    ind_fractions = _read_demand_fractions_india()
    for region, fracs in ind_fractions.items():
        fuel_code = f"ELC{region}02"
        # Normalize
        total = sum(fracs.values())
        if total > 0:
            fracs = {k: v / total for k, v in fracs.items()}
        for ts in TIMESLICES:
            val = fracs.get(ts, 0.0)
            for y in TARGET_YEARS:
                rows.append({
                    "REGION": region,
                    "FUEL": fuel_code,
                    "TIMESLICE": ts,
                    "YEAR": y,
                    "VALUE": round(val, 10),
                })

    # Other countries
    for iso3 in countries:
        if iso3 == "IND":
            continue

        fracs = _read_demand_fractions_standard(iso3)
        regions = _get_regions_for_country(iso3)
        region = regions[0]
        fuel_code = f"ELC{region}02"

        if fracs is None:
            # No data available — use flat profile
            flat_val = 1.0 / len(TIMESLICES)
            fracs = {ts: flat_val for ts in TIMESLICES}
            print(f"  INFO: Using flat demand profile for {iso3}")

        # Normalize
        total = sum(fracs.values())
        if total > 0:
            fracs = {k: v / total for k, v in fracs.items()}

        for ts in TIMESLICES:
            val = fracs.get(ts, 0.0)
            for y in TARGET_YEARS:
                rows.append({
                    "REGION": region,
                    "FUEL": fuel_code,
                    "TIMESLICE": ts,
                    "YEAR": y,
                    "VALUE": round(val, 10),
                })

    return pd.DataFrame(rows)


# ============================================================
# 8. EXCEL-MODE HELPERS (write_sheet_overlay, merge, formatters)
# ============================================================

def write_sheet_overlay(excel_path, sheet_name, df):
    """Overwrite a single sheet in an existing workbook, preserving all others."""
    from openpyxl import load_workbook
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]
    ws.delete_rows(1, ws.max_row)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    wb.save(excel_path)


def read_excel_sheet(excel_path, sheet_name):
    """Read a sheet from an Excel workbook into a DataFrame."""
    from openpyxl import load_workbook
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()
    if not rows:
        return pd.DataFrame()
    header = rows[0]
    data = rows[1:]
    return pd.DataFrame(data, columns=header)


def _build_tech_id_map(existing_df, tech_col="Tech"):
    """Build {tech_code: id} from an existing wide-format DataFrame."""
    if existing_df.empty or tech_col not in existing_df.columns or "Tech.ID" not in existing_df.columns:
        return {}
    mapping = {}
    for _, row in existing_df.iterrows():
        t = row[tech_col]
        tid = row["Tech.ID"]
        if t is not None and tid is not None:
            try:
                mapping[str(t)] = int(tid)
            except (ValueError, TypeError):
                pass
    return mapping


def _assign_tech_ids(df, existing_id_map, tech_col="Tech"):
    """Assign Tech.ID preserving existing IDs, new ones from max+1."""
    if df.empty:
        return df
    max_id = max(existing_id_map.values()) if existing_id_map else 0
    ids = []
    for t in df[tech_col]:
        if t in existing_id_map:
            ids.append(existing_id_map[t])
        else:
            max_id += 1
            existing_id_map[t] = max_id
            ids.append(max_id)
    df = df.copy()
    df["Tech.ID"] = ids
    return df


# ------------------------------------------------------------------
# Merge helpers for Excel wide-format
# ------------------------------------------------------------------

def merge_wide_by_tech_param(existing_df, new_df, tech_col="Tech", param_col="Parameter"):
    """Replace rows (tech, param) covered by new_df, preserve the rest."""
    if new_df.empty:
        return existing_df
    if existing_df.empty:
        return new_df
    new_keys = set(zip(new_df[tech_col], new_df[param_col]))
    preserved = existing_df[~existing_df.apply(
        lambda r: (r[tech_col], r[param_col]) in new_keys, axis=1)]
    return pd.concat([preserved, new_df], ignore_index=True)


def merge_wide_by_tech(existing_df, new_df, tech_col="Tech"):
    """Replace rows by tech only (for Capacities, VariableCost)."""
    if new_df.empty:
        return existing_df
    if existing_df.empty:
        return new_df
    new_techs = set(new_df[tech_col])
    preserved = existing_df[~existing_df[tech_col].isin(new_techs)]
    return pd.concat([preserved, new_df], ignore_index=True)


def merge_wide_by_fuel(existing_df, new_df, fuel_col="Fuel/Tech"):
    """Replace rows by fuel code (for Demand sheets)."""
    if new_df.empty:
        return existing_df
    if existing_df.empty:
        return new_df
    new_fuels = set(new_df[fuel_col])
    preserved = existing_df[~existing_df[fuel_col].isin(new_fuels)]
    return pd.concat([preserved, new_df], ignore_index=True)


# ------------------------------------------------------------------
# 7 Formatters: long-format → wide-format for Excel sheets
# ------------------------------------------------------------------

def format_param_yearly(long_df, param_name, param_id, all_years):
    """Format CapitalCost/FixedCost/ResidualCapacity → Primary/Secondary Techs wide.

    Input columns: REGION, TECHNOLOGY, YEAR, VALUE
    Output columns: Tech.ID, Tech, Tech.Name, Parameter.ID, Parameter, Unit,
                    Projection.Mode, Projection.Parameter, [years...]
    """
    if long_df.empty:
        return pd.DataFrame()

    techs = long_df["TECHNOLOGY"].unique()
    records = []
    for tech in techs:
        sub = long_df[long_df["TECHNOLOGY"] == tech]
        row = {
            "Tech.ID": 0,  # placeholder, assigned later
            "Tech": tech,
            "Tech.Name": parse_tech_name(tech),
            "Parameter.ID": param_id,
            "Parameter": param_name,
            "Unit": "",
            "Projection.Mode": "User defined",
            "Projection.Parameter": 0,
        }
        year_vals = dict(zip(sub["YEAR"], sub["VALUE"]))
        for y in all_years:
            row[y] = year_vals.get(y, np.nan)
        records.append(row)

    return pd.DataFrame(records)


def format_fixed_horizon(long_df):
    """Format OperationalLife → Fixed Horizon Parameters wide.

    Input columns: REGION, TECHNOLOGY, VALUE
    Output columns: Tech.Type, Tech.ID, Tech, Tech.Name, Parameter.ID, Parameter, Unit, Value
    """
    if long_df.empty:
        return pd.DataFrame()

    techs = long_df["TECHNOLOGY"].unique()
    records = []
    for tech in techs:
        sub = long_df[long_df["TECHNOLOGY"] == tech]
        val = sub["VALUE"].iloc[0]
        records.append({
            "Tech.Type": assign_tech_type(tech),
            "Tech.ID": 0,
            "Tech": tech,
            "Tech.Name": parse_tech_name(tech),
            "Parameter.ID": 2,
            "Parameter": "OperationalLife",
            "Unit": "",
            "Value": val,
        })

    return pd.DataFrame(records)


def format_capacities(long_df, all_years):
    """Format CapacityFactor → Capacities sheet wide.

    Input columns: REGION, TECHNOLOGY, TIMESLICE, YEAR, VALUE
    Output columns: Timeslices, Tech.ID, Tech, Tech.Name, Parameter.ID, Parameter,
                    Unit, Projection.Mode, Projection.Parameter, [years...]
    """
    if long_df.empty:
        return pd.DataFrame()

    groups = long_df.groupby(["TIMESLICE", "TECHNOLOGY"])
    records = []
    for (ts, tech), sub in groups:
        row = {
            "Timeslices": ts,
            "Tech.ID": 0,
            "Tech": tech,
            "Tech.Name": parse_tech_name(tech),
            "Parameter.ID": 13,
            "Parameter": "CapacityFactor",
            "Unit": "",
            "Projection.Mode": "User defined",
            "Projection.Parameter": 0,
        }
        year_vals = dict(zip(sub["YEAR"], sub["VALUE"]))
        for y in all_years:
            row[y] = year_vals.get(y, np.nan)
        records.append(row)

    return pd.DataFrame(records)


def format_variable_cost(long_df, all_years):
    """Format VariableCost → VariableCost sheet wide.

    Input columns: REGION, TECHNOLOGY, MODE_OF_OPERATION, YEAR, VALUE
    Output columns: Mode.Operation, Tech.ID, Tech, Tech.Name, Parameter.ID, Parameter,
                    Unit, Projection.Mode, Projection.Parameter, [years...]
    """
    if long_df.empty:
        return pd.DataFrame()

    groups = long_df.groupby(["TECHNOLOGY", "MODE_OF_OPERATION"])
    records = []
    for (tech, mode), sub in groups:
        row = {
            "Mode.Operation": int(mode),
            "Tech.ID": 0,
            "Tech": tech,
            "Tech.Name": parse_tech_name(tech),
            "Parameter.ID": 12,
            "Parameter": "VariableCost",
            "Unit": "",
            "Projection.Mode": "User defined",
            "Projection.Parameter": 0,
        }
        year_vals = dict(zip(sub["YEAR"], sub["VALUE"]))
        for y in all_years:
            row[y] = year_vals.get(y, np.nan)
        records.append(row)

    return pd.DataFrame(records)


def format_yearsplit(long_df, all_years):
    """Format YearSplit → Yearsplit sheet wide.

    Input columns: TIMESLICE, YEAR, VALUE
    Output columns: Timeslices, Parameter.ID, Parameter, Unit,
                    Projection.Mode, Projection.Parameter, [years...]
    """
    if long_df.empty:
        return pd.DataFrame()

    records = []
    for ts in long_df["TIMESLICE"].unique():
        sub = long_df[long_df["TIMESLICE"] == ts]
        row = {
            "Timeslices": ts,
            "Parameter.ID": 14,
            "Parameter": "YearSplit",
            "Unit": "",
            "Projection.Mode": "User defined",
            "Projection.Parameter": 0,
        }
        year_vals = dict(zip(sub["YEAR"], sub["VALUE"]))
        for y in all_years:
            row[y] = year_vals.get(y, np.nan)
        records.append(row)

    return pd.DataFrame(records)


def _rewrite_fuel_dsptrn(fuel_code):
    """If DSPTRN enabled, rewrite ELC*02 → ELC*03."""
    if fuel_code.startswith("ELC") and fuel_code.endswith("02"):
        return fuel_code[:-2] + "03"
    return fuel_code


def format_demand_projection(long_df, enable_dsptrn, all_years):
    """Format SpecifiedAnnualDemand → Demand_Projection sheet wide.

    Input columns: REGION, FUEL, YEAR, VALUE
    Output columns: Demand/Share, Fuel/Tech, Name, Ref.Cap.BY, Ref.OAR.BY,
                    Ref.km.BY, Projection.Mode, Projection.Parameter, [years...]
    """
    if long_df.empty:
        return pd.DataFrame()

    records = []
    for fuel in long_df["FUEL"].unique():
        sub = long_df[long_df["FUEL"] == fuel]
        fuel_out = _rewrite_fuel_dsptrn(fuel) if enable_dsptrn else fuel
        row = {
            "Demand/Share": "Demand",
            "Fuel/Tech": fuel_out,
            "Name": parse_fuel_name(fuel_out),
            "Ref.Cap.BY": "not needed",
            "Ref.OAR.BY": "not needed",
            "Ref.km.BY": "not needed",
            "Projection.Mode": "User defined",
            "Projection.Parameter": 0,
        }
        year_vals = dict(zip(sub["YEAR"], sub["VALUE"]))
        for y in all_years:
            row[y] = year_vals.get(y, np.nan)
        records.append(row)

    return pd.DataFrame(records)


def format_demand_profiles(long_df, enable_dsptrn, all_years):
    """Format SpecifiedDemandProfile → Profiles sheet wide.

    Input columns: REGION, FUEL, TIMESLICE, YEAR, VALUE
    Output columns: Timeslices, Demand/Share, Fuel/Tech, Name, Ref.Cap.BY,
                    Ref.OAR.BY, Ref.km.BY, Projection.Mode, Projection.Parameter, [years...]
    """
    if long_df.empty:
        return pd.DataFrame()

    groups = long_df.groupby(["TIMESLICE", "FUEL"])
    records = []
    for (ts, fuel), sub in groups:
        fuel_out = _rewrite_fuel_dsptrn(fuel) if enable_dsptrn else fuel
        row = {
            "Timeslices": ts,
            "Demand/Share": "Demand",
            "Fuel/Tech": fuel_out,
            "Name": parse_fuel_name(fuel_out),
            "Ref.Cap.BY": "not needed",
            "Ref.OAR.BY": "not needed",
            "Ref.km.BY": "not needed",
            "Projection.Mode": "User defined",
            "Projection.Parameter": 0,
        }
        year_vals = dict(zip(sub["YEAR"], sub["VALUE"]))
        for y in all_years:
            row[y] = year_vals.get(y, np.nan)
        records.append(row)

    return pd.DataFrame(records)


# ============================================================
# 8b. ORCHESTRATORS (Excel mode)
# ============================================================

def backup_a1_outputs(a1_path):
    """Create timestamped backup of A1_Outputs Excel files."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_dir = a1_path / f"_backup_{timestamp}"
    backup_dir.mkdir(exist_ok=True)
    excel_files = list(a1_path.glob("*.xlsx"))
    for f in excel_files:
        shutil.copy2(f, backup_dir / f.name)
    print(f"  Backup created: {backup_dir} ({len(excel_files)} files)")


def update_parametrization(excel_path, cost_db):
    """Orchestrator: update A-O_Parametrization.xlsx in-place from data package.

    Updates sheets: Primary Techs, Secondary Techs, Fixed Horizon Parameters,
                    Capacities, VariableCost, Yearsplit.
    Does NOT touch: Demand Techs (preserves PWRTRN/RNWTRN/DSPTRN from A2_AddTx).
    """
    all_years = TARGET_YEARS
    excel_path = str(excel_path)

    # --- Generate long-format data ---
    print("  Generating long-format data...")
    cap_cost_df = generate_capital_cost(cost_db, pd.DataFrame())
    fix_cost_df = generate_fixed_cost(cost_db, pd.DataFrame())
    res_cap_df = generate_residual_capacity(cost_db, pd.DataFrame())
    op_life_df = generate_operational_life(cost_db, pd.DataFrame())
    cap_factor_df = generate_capacity_factor(cost_db, pd.DataFrame())
    var_cost_df = generate_variable_cost(cost_db, pd.DataFrame())
    year_split_df = generate_year_split(pd.DataFrame())

    # --- Format to wide ---
    print("  Formatting to wide format...")
    wide_capex = format_param_yearly(cap_cost_df, "CapitalCost", 3, all_years)
    wide_fixom = format_param_yearly(fix_cost_df, "FixedCost", 4, all_years)
    wide_rescap = format_param_yearly(res_cap_df, "ResidualCapacity", 5, all_years)
    wide_oplife = format_fixed_horizon(op_life_df)
    wide_capfac = format_capacities(cap_factor_df, all_years)
    wide_varcost = format_variable_cost(var_cost_df, all_years)
    wide_ysplit = format_yearsplit(year_split_df, all_years)

    # --- Classify techs: Primary (MIN/RNW) vs Secondary (PWR, not PWRTRN) ---
    def split_primary_secondary(wide_df):
        if wide_df.empty:
            return pd.DataFrame(), pd.DataFrame()
        primary = wide_df[wide_df["Tech"].apply(
            lambda t: t.startswith("MIN") or t.startswith("RNW"))]
        secondary = wide_df[wide_df["Tech"].apply(
            lambda t: not t.startswith("MIN") and not t.startswith("RNW")
            and not t.startswith("PWRTRN"))]
        return primary, secondary

    pri_capex, sec_capex = split_primary_secondary(wide_capex)
    pri_fixom, sec_fixom = split_primary_secondary(wide_fixom)
    pri_rescap, sec_rescap = split_primary_secondary(wide_rescap)

    # Combine per sheet
    new_primary = pd.concat([pri_capex, pri_fixom, pri_rescap], ignore_index=True)
    new_secondary = pd.concat([sec_capex, sec_fixom, sec_rescap], ignore_index=True)

    # --- Read existing sheets and merge ---
    print("  Reading existing sheets...")
    exist_primary = read_excel_sheet(excel_path, "Primary Techs")
    exist_secondary = read_excel_sheet(excel_path, "Secondary Techs")
    exist_fixed = read_excel_sheet(excel_path, "Fixed Horizon Parameters")
    exist_capacities = read_excel_sheet(excel_path, "Capacities")
    exist_varcost = read_excel_sheet(excel_path, "VariableCost")
    exist_ysplit = read_excel_sheet(excel_path, "Yearsplit")

    print("  Merging...")
    # Primary/Secondary: merge by (Tech, Parameter)
    merged_primary = merge_wide_by_tech_param(exist_primary, new_primary)
    merged_secondary = merge_wide_by_tech_param(exist_secondary, new_secondary)

    # Fixed Horizon: merge by Tech only (one param per tech)
    merged_fixed = merge_wide_by_tech(exist_fixed, wide_oplife)

    # Capacities: merge by Tech (all timeslices replaced together)
    merged_capacities = merge_wide_by_tech(exist_capacities, wide_capfac)

    # VariableCost: merge by Tech
    merged_varcost = merge_wide_by_tech(exist_varcost, wide_varcost)

    # YearSplit: full replacement
    merged_ysplit = wide_ysplit if not wide_ysplit.empty else exist_ysplit

    # --- Assign Tech.IDs ---
    # Build unified ID map from all existing sheets
    id_map = {}
    for edf in [exist_primary, exist_secondary, exist_fixed, exist_capacities, exist_varcost]:
        id_map.update(_build_tech_id_map(edf))

    merged_primary = _assign_tech_ids(merged_primary, id_map)
    merged_secondary = _assign_tech_ids(merged_secondary, id_map)
    merged_fixed = _assign_tech_ids(merged_fixed, id_map)
    merged_capacities = _assign_tech_ids(merged_capacities, id_map)
    merged_varcost = _assign_tech_ids(merged_varcost, id_map)

    # --- Sort ---
    for df in [merged_primary, merged_secondary]:
        if not df.empty and "Tech.ID" in df.columns and "Parameter.ID" in df.columns:
            df.sort_values(["Tech.ID", "Parameter.ID"], inplace=True)

    # --- Ensure correct column order ---
    yearly_cols = [c for c in merged_primary.columns if isinstance(c, int)]
    yearly_cols.sort()
    param_fixed_cols = ["Tech.ID", "Tech", "Tech.Name", "Parameter.ID", "Parameter",
                        "Unit", "Projection.Mode", "Projection.Parameter"]
    for df_name, df_ref in [("merged_primary", merged_primary), ("merged_secondary", merged_secondary)]:
        if not df_ref.empty:
            ordered = [c for c in param_fixed_cols if c in df_ref.columns] + yearly_cols
            # Keep only columns that exist
            ordered = [c for c in ordered if c in df_ref.columns]
            # Re-assign in place via index
            idx = df_ref.index
            reordered = df_ref[ordered]
            if df_name == "merged_primary":
                merged_primary = reordered
            else:
                merged_secondary = reordered

    # Capacities column order
    if not merged_capacities.empty:
        cap_fixed = ["Timeslices", "Tech.ID", "Tech", "Tech.Name", "Parameter.ID",
                      "Parameter", "Unit", "Projection.Mode", "Projection.Parameter"]
        cap_cols = [c for c in cap_fixed if c in merged_capacities.columns] + yearly_cols
        cap_cols = [c for c in cap_cols if c in merged_capacities.columns]
        merged_capacities = merged_capacities[cap_cols]

    # VariableCost column order
    if not merged_varcost.empty:
        vc_fixed = ["Mode.Operation", "Tech.ID", "Tech", "Tech.Name", "Parameter.ID",
                     "Parameter", "Unit", "Projection.Mode", "Projection.Parameter"]
        vc_cols = [c for c in vc_fixed if c in merged_varcost.columns] + yearly_cols
        vc_cols = [c for c in vc_cols if c in merged_varcost.columns]
        merged_varcost = merged_varcost[vc_cols]

    # Yearsplit column order
    if not merged_ysplit.empty:
        ys_fixed = ["Timeslices", "Parameter.ID", "Parameter", "Unit",
                     "Projection.Mode", "Projection.Parameter"]
        ys_yearly = [c for c in merged_ysplit.columns if isinstance(c, int)]
        ys_yearly.sort()
        ys_cols = [c for c in ys_fixed if c in merged_ysplit.columns] + ys_yearly
        ys_cols = [c for c in ys_cols if c in merged_ysplit.columns]
        merged_ysplit = merged_ysplit[ys_cols]

    # Fixed Horizon column order
    if not merged_fixed.empty:
        fh_cols = ["Tech.Type", "Tech.ID", "Tech", "Tech.Name", "Parameter.ID",
                    "Parameter", "Unit", "Value"]
        fh_cols = [c for c in fh_cols if c in merged_fixed.columns]
        merged_fixed = merged_fixed[fh_cols]

    # --- Write sheets ---
    print("  Writing sheets...")
    write_sheet_overlay(excel_path, "Primary Techs", merged_primary)
    print("    Primary Techs updated")
    write_sheet_overlay(excel_path, "Secondary Techs", merged_secondary)
    print("    Secondary Techs updated")
    write_sheet_overlay(excel_path, "Fixed Horizon Parameters", merged_fixed)
    print("    Fixed Horizon Parameters updated")
    write_sheet_overlay(excel_path, "Capacities", merged_capacities)
    print("    Capacities updated")
    write_sheet_overlay(excel_path, "VariableCost", merged_varcost)
    print("    VariableCost updated")
    write_sheet_overlay(excel_path, "Yearsplit", merged_ysplit)
    print("    Yearsplit updated")

    print("  Parametrization update complete.")


def update_demand(excel_path, cost_db):
    """Orchestrator: update A-O_Demand.xlsx in-place from data package.

    Updates sheets: Demand_Projection, Profiles.
    """
    all_years = TARGET_YEARS
    excel_path = str(excel_path)
    enable_dsptrn = get_enable_dsptrn()

    # --- Generate long-format data ---
    print("  Generating demand data...")
    sad_df = generate_specified_annual_demand(pd.DataFrame())
    sdp_df = generate_specified_demand_profile(pd.DataFrame())

    # --- Format to wide ---
    print("  Formatting to wide format...")
    wide_proj = format_demand_projection(sad_df, enable_dsptrn, all_years)
    wide_prof = format_demand_profiles(sdp_df, enable_dsptrn, all_years)

    # --- Read existing and merge ---
    print("  Reading existing sheets...")
    exist_proj = read_excel_sheet(excel_path, "Demand_Projection")
    exist_prof = read_excel_sheet(excel_path, "Profiles")

    print("  Merging...")
    merged_proj = merge_wide_by_fuel(exist_proj, wide_proj)
    merged_prof = merge_wide_by_fuel(exist_prof, wide_prof)

    # --- Column order ---
    proj_fixed = ["Demand/Share", "Fuel/Tech", "Name", "Ref.Cap.BY", "Ref.OAR.BY",
                  "Ref.km.BY", "Projection.Mode", "Projection.Parameter"]
    prof_fixed = ["Timeslices", "Demand/Share", "Fuel/Tech", "Name", "Ref.Cap.BY",
                  "Ref.OAR.BY", "Ref.km.BY", "Projection.Mode", "Projection.Parameter"]

    if not merged_proj.empty:
        yearly = [c for c in merged_proj.columns if isinstance(c, int)]
        yearly.sort()
        cols = [c for c in proj_fixed if c in merged_proj.columns] + yearly
        cols = [c for c in cols if c in merged_proj.columns]
        merged_proj = merged_proj[cols]

    if not merged_prof.empty:
        yearly = [c for c in merged_prof.columns if isinstance(c, int)]
        yearly.sort()
        cols = [c for c in prof_fixed if c in merged_prof.columns] + yearly
        cols = [c for c in cols if c in merged_prof.columns]
        merged_prof = merged_prof[cols]

    # --- Write sheets ---
    print("  Writing sheets...")
    write_sheet_overlay(excel_path, "Demand_Projection", merged_proj)
    print("    Demand_Projection updated")
    write_sheet_overlay(excel_path, "Profiles", merged_prof)
    print("    Profiles updated")

    print("  Demand update complete.")


# ============================================================
# 9. MERGE & MAIN (CSV legacy)
# ============================================================

def read_existing_csv(filename):
    """Read an existing CSV from OG_csvs_inputs/."""
    path = OG_DIR / filename
    if path.exists():
        return pd.read_csv(path)
    return pd.DataFrame()


def merge_dataframes(new_df, existing_df, key_cols):
    """Merge new data into existing, preserving rows not in new_df.
    Rows in new_df override matching rows in existing_df.
    Rows only in existing_df are preserved (techs without data package coverage).
    """
    if new_df.empty:
        return existing_df
    if existing_df.empty:
        return new_df

    # Identify which technologies are covered by new data
    if "TECHNOLOGY" in new_df.columns:
        new_techs = set(new_df["TECHNOLOGY"].unique())
        # Keep existing rows whose TECHNOLOGY is NOT in new_techs
        preserved = existing_df[~existing_df["TECHNOLOGY"].isin(new_techs)]
    elif "FUEL" in new_df.columns:
        new_fuels = set(new_df["FUEL"].unique())
        preserved = existing_df[~existing_df["FUEL"].isin(new_fuels)]
    elif "TIMESLICE" in key_cols and "TECHNOLOGY" not in new_df.columns:
        # YearSplit — replace entirely
        preserved = pd.DataFrame()
    else:
        preserved = pd.DataFrame()

    result = pd.concat([new_df, preserved], ignore_index=True)

    # Sort
    sort_cols = [c for c in key_cols if c in result.columns]
    if sort_cols:
        result = result.sort_values(sort_cols).reset_index(drop=True)

    return result


def backup_csvs():
    """Create timestamped backup of current OG_csvs_inputs/."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_dir = OG_DIR / f"_backup_{timestamp}"
    backup_dir.mkdir(exist_ok=True)

    csv_files = list(OG_DIR.glob("*.csv"))
    for f in csv_files:
        shutil.copy2(f, backup_dir / f.name)

    print(f"  Backup created: {backup_dir} ({len(csv_files)} files)")
    return backup_dir


def write_csv(df, filename):
    """Write DataFrame to CSV in OG_csvs_inputs/."""
    path = OG_DIR / filename
    df.to_csv(path, index=False)
    print(f"  Written: {filename} ({len(df)} rows)")


def _main_csv_mode(cost_db):
    """Legacy CSV mode: write to OG_csvs_inputs/."""
    # Step 1: Backup
    if CFG.get("backup_existing_csvs", True):
        print("\n--- Step 1: Backup ---")
        backup_csvs()

    # Step 2: Generate each CSV
    csv_generators = {
        "CapitalCost.csv": {
            "func": lambda: generate_capital_cost(cost_db, read_existing_csv("CapitalCost.csv")),
            "key_cols": ["REGION", "TECHNOLOGY", "YEAR"],
        },
        "FixedCost.csv": {
            "func": lambda: generate_fixed_cost(cost_db, read_existing_csv("FixedCost.csv")),
            "key_cols": ["REGION", "TECHNOLOGY", "YEAR"],
        },
        "VariableCost.csv": {
            "func": lambda: generate_variable_cost(cost_db, read_existing_csv("VariableCost.csv")),
            "key_cols": ["REGION", "TECHNOLOGY", "MODE_OF_OPERATION", "YEAR"],
        },
        "OperationalLife.csv": {
            "func": lambda: generate_operational_life(cost_db, read_existing_csv("OperationalLife.csv")),
            "key_cols": ["REGION", "TECHNOLOGY"],
        },
        "CapacityFactor.csv": {
            "func": lambda: generate_capacity_factor(cost_db, read_existing_csv("CapacityFactor.csv")),
            "key_cols": ["REGION", "TECHNOLOGY", "TIMESLICE", "YEAR"],
        },
        "ResidualCapacity.csv": {
            "func": lambda: generate_residual_capacity(cost_db, read_existing_csv("ResidualCapacity.csv")),
            "key_cols": ["REGION", "TECHNOLOGY", "YEAR"],
        },
        "SpecifiedAnnualDemand.csv": {
            "func": lambda: generate_specified_annual_demand(read_existing_csv("SpecifiedAnnualDemand.csv")),
            "key_cols": ["REGION", "FUEL", "YEAR"],
        },
        "SpecifiedDemandProfile.csv": {
            "func": lambda: generate_specified_demand_profile(read_existing_csv("SpecifiedDemandProfile.csv")),
            "key_cols": ["REGION", "FUEL", "TIMESLICE", "YEAR"],
        },
        "YearSplit.csv": {
            "func": lambda: generate_year_split(read_existing_csv("YearSplit.csv")),
            "key_cols": ["TIMESLICE", "YEAR"],
        },
    }

    print("\n--- Generating CSVs ---")
    for csv_name, spec in csv_generators.items():
        print(f"\n  Processing {csv_name}...")
        try:
            new_df = spec["func"]()
            existing_df = read_existing_csv(csv_name)
            merged_df = merge_dataframes(new_df, existing_df, spec["key_cols"])
            write_csv(merged_df, csv_name)

            n_existing = len(existing_df)
            n_new = len(new_df)
            n_merged = len(merged_df)
            print(f"    Existing: {n_existing} | New: {n_new} | Merged: {n_merged}")
        except Exception as e:
            print(f"    ERROR generating {csv_name}: {e}")
            import traceback
            traceback.print_exc()

    print("\n" + "=" * 70)
    print("A3 (csv mode): Complete. Next step: run A1_Pre_processing_OG_csvs.py")
    print("=" * 70)


def _main_excel_mode(cost_db):
    """Excel mode: write directly to A1_Outputs Excel files (preserves Tx layers)."""
    a1_subdir = CFG.get("a1_outputs_subdir", "A1_Outputs/A1_Outputs_BAU")
    a1_path = SCRIPT_DIR / a1_subdir

    param_excel = a1_path / "A-O_Parametrization.xlsx"
    demand_excel = a1_path / "A-O_Demand.xlsx"

    # Verify files exist
    if not param_excel.exists():
        print(f"  ERROR: {param_excel} not found. Run A1_Pre_processing + A2_AddTx first.")
        sys.exit(1)
    if not demand_excel.exists():
        print(f"  ERROR: {demand_excel} not found. Run A1_Pre_processing + A2_AddTx first.")
        sys.exit(1)

    # Backup
    if CFG.get("backup_existing_csvs", True):
        print("\n--- Backup ---")
        backup_a1_outputs(a1_path)

    # Update Parametrization
    print("\n--- Updating A-O_Parametrization.xlsx ---")
    try:
        update_parametrization(param_excel, cost_db)
    except Exception as e:
        print(f"  ERROR updating Parametrization: {e}")
        import traceback
        traceback.print_exc()

    # Update Demand
    print("\n--- Updating A-O_Demand.xlsx ---")
    try:
        update_demand(demand_excel, cost_db)
    except Exception as e:
        print(f"  ERROR updating Demand: {e}")
        import traceback
        traceback.print_exc()

    print("\n" + "=" * 70)
    print("A3 (excel mode): Complete. Next step: run B1_Compiler.py")
    print("=" * 70)


def main():
    """Main entry point."""
    output_mode = CFG.get("output_mode", "csv")

    print("=" * 70)
    print(f"A3: Updating from DATA_PACKAGE_V2 (mode: {output_mode})")
    print(f"  Scenario: {CFG['scenario']}")
    print(f"  Cost scenario: {COST_SCENARIO}")
    print(f"  Data package: {DATA_PKG}")
    print(f"  Model horizon: {FIRST_YEAR}-{FINAL_YEAR}")
    print("=" * 70)

    # Read Cost Database
    print("\n--- Reading Cost Database ---")
    cost_db = read_cost_database()
    print(f"  Technologies: {len(cost_db['technologies'])} rows")
    print(f"  Fuel costs: {len(cost_db['fuel_costs'])} rows")
    print(f"  Node anchors: {len(cost_db['node_anchors'])} rows")

    if output_mode == "excel":
        _main_excel_mode(cost_db)
    else:
        _main_csv_mode(cost_db)


if __name__ == "__main__":
    main()
