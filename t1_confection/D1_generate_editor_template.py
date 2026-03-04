"""
Generate Secondary Techs Editor Template

This script reads all A-O_Parametrization.xlsx files and generates
a user-friendly Excel template for editing Secondary Techs data.

Usage:
    python t1_confection/generate_editor_template.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Protection, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
import sys
from pathlib import Path
from datetime import datetime
import yaml
from Z_AUX_config_loader import get_ostram_country_mapping, get_ostram_tech_mapping, get_country_names

# Country and technology mappings from centralized config
OSTRAM_COUNTRY_MAPPING = get_ostram_country_mapping()
OSTRAM_TECH_MAPPING = get_ostram_tech_mapping()


def read_base_scenario():
    """
    Read base_scenario from Config_MOMF_T1_AB.yaml

    Returns:
        str: The base scenario name (default: 'BAU')
    """
    yaml_path = Path(__file__).parent / "Config_MOMF_T1_AB.yaml"
    try:
        with open(yaml_path, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f)
            base_scenario = config.get('base_scenario', 'BAU')
            print(f"Base scenario from YAML: {base_scenario}")
            return base_scenario
    except Exception as e:
        print(f"Warning: Could not read base_scenario from YAML: {e}")
        print("Using default base_scenario: BAU")
        return 'BAU'


def collect_data_from_all_scenarios():
    """
    Collect all unique values from all scenario files

    Returns:
        dict with scenarios, countries, tech_mapping (Tech.Name -> Tech), parameters, years,
        tech_by_country (country_code -> list of Tech.Name)
    """
    base_path = Path(__file__).parent / "A1_Outputs"

    # Auto-discover scenarios from A1_Outputs_* folders
    scenarios = []
    for item in sorted(base_path.iterdir()):
        if item.is_dir() and item.name.startswith("A1_Outputs_"):
            suffix = item.name.split("A1_Outputs_", 1)[1]
            if suffix:
                scenarios.append(suffix)

    if not scenarios:
        print("WARNING: No 'A1_Outputs_*' folders found. Nothing to do.")
        return None

    all_countries = set()
    tech_mapping = {}  # Tech.Name -> Tech code
    tech_by_country = {}  # country_code -> set of Tech.Name
    all_parameters = set()
    all_years = set()

    print("Collecting data from all scenarios...")
    print()

    for scenario in scenarios:
        scenario_path = base_path / f"A1_Outputs_{scenario}" / "A-O_Parametrization.xlsx"

        if not scenario_path.exists():
            print(f"WARNING: File not found: {scenario_path}")
            continue

        print(f"Reading {scenario}...")

        wb = openpyxl.load_workbook(scenario_path, data_only=True)

        if 'Secondary Techs' not in wb.sheetnames:
            print(f"  WARNING: 'Secondary Techs' sheet not found in {scenario}")
            wb.close()
            continue

        ws = wb['Secondary Techs']

        # Get headers to find year columns
        headers = [cell.value for cell in ws[1]]

        # Find year columns (columns with numeric values representing years)
        year_col_indices = []
        for idx, header in enumerate(headers, 1):
            if header and str(header).isdigit():
                try:
                    year = int(header)
                    if 2000 <= year <= 2100:
                        all_years.add(year)
                        year_col_indices.append(idx)
                except:
                    pass

        # Collect tech mapping and parameters (skip header row)
        # Columns: 1=Tech.Id, 2=Tech, 3=Tech.Name, 5=Parameter
        for row_idx in range(2, ws.max_row + 1):
            tech_code = ws.cell(row_idx, 2).value      # Column 2: Tech
            tech_name = ws.cell(row_idx, 3).value      # Column 3: Tech.Name
            parameter = ws.cell(row_idx, 5).value      # Column 5: Parameter

            if tech_code and tech_name:
                tech_code_str = str(tech_code).strip()
                tech_name_str = str(tech_name).strip()

                # Build mapping: Tech.Name -> Tech
                tech_mapping[tech_name_str] = tech_code_str

                # Extract country code from PWR and TRN technologies
                # PWR Format: PWRURNARGXX -> country code is ARG (characters 6-8)
                # TRN Format: TRNBRAXXPRYXX -> origin country is BRA (characters 3-5)
                country_code = None

                if tech_code_str.upper().startswith('PWR') and len(tech_code_str) >= 9:
                    # PWR technologies: country at positions 6-8
                    country_code = tech_code_str[6:9].upper()
                elif tech_code_str.upper().startswith('TRN') and len(tech_code_str) >= 8:
                    # TRN technologies: origin country at positions 3-5
                    # Example: TRNBRAXXPRYXX -> BRA is origin
                    country_code = tech_code_str[3:6].upper()

                if country_code:
                    all_countries.add(country_code)

                    # Group Tech.Name by country
                    if country_code not in tech_by_country:
                        tech_by_country[country_code] = set()
                    tech_by_country[country_code].add(tech_name_str)

            if parameter:
                all_parameters.add(str(parameter).strip())

        wb.close()
        print(f"  Found: {len(tech_mapping)} technologies, {len(all_parameters)} parameters")

    # Convert sets to sorted lists
    for country in tech_by_country:
        tech_by_country[country] = sorted(tech_by_country[country])

    print()
    print(f"Summary:")
    print(f"  Scenarios: {len(scenarios)}")
    print(f"  Countries: {len(all_countries)} - {sorted(all_countries)}")
    print(f"  Technologies (Tech.Name): {len(tech_mapping)}")
    print(f"  Parameters: {len(all_parameters)}")
    print(f"  Years: {len(all_years)} - {min(all_years) if all_years else 'N/A'} to {max(all_years) if all_years else 'N/A'}")
    print()

    return {
        'scenarios': sorted(scenarios),
        'countries': sorted(all_countries),
        'tech_mapping': tech_mapping,  # Tech.Name -> Tech code
        'tech_names': sorted(tech_mapping.keys()),  # List of Tech.Name for dropdown
        'tech_by_country': tech_by_country,  # country_code -> list of Tech.Name
        'parameters': sorted(all_parameters),
        'years': sorted(all_years)
    }


def collect_trn_interconnections():
    """
    Collect TRN interconnection technologies from A-O_AR_Model_Base_Year.xlsx.

    Reads the base scenario's Base Year file to discover all cross-border and
    intra-country interconnection technologies (13-character TRN codes).

    Returns:
        list of dicts sorted by tech code:
        [
            {
                'tech': 'TRNBGDXXINDEA',
                'mode1_label': 'BGD-XX -> IND-EA',
                'mode2_label': 'IND-EA -> BGD-XX',
                'mode1_input_fuel': 'ELCBGDXX02',
                'mode1_output_fuel': 'ELCINDEA01',
                'mode2_input_fuel': 'ELCINDEA02',
                'mode2_output_fuel': 'ELCBGDXX01',
            },
            ...
        ]
    """
    base_scenario = read_base_scenario()
    base_year_path = Path(__file__).parent / "A1_Outputs" / f"A1_Outputs_{base_scenario}" / "A-O_AR_Model_Base_Year.xlsx"

    if not base_year_path.exists():
        print(f"WARNING: Base Year file not found: {base_year_path}")
        return []

    print(f"Reading TRN interconnections from {base_year_path.name}...")

    wb = openpyxl.load_workbook(base_year_path, data_only=True)
    if 'Secondary' not in wb.sheetnames:
        print("WARNING: 'Secondary' sheet not found in Base Year file")
        wb.close()
        return []

    ws = wb['Secondary']

    # Headers: Mode.Operation(1), Fuel.I(2), Fuel.I.Name(3), Value.Fuel.I(4),
    #          Unit.Fuel.I(5), Tech(6), Tech.Name(7), Fuel.O(8), Fuel.O.Name(9),
    #          Value.Fuel.O(10), Unit.Fuel.O(11)
    # Find column indices by header name for robustness
    headers = {str(cell.value).strip(): idx for idx, cell in enumerate(ws[1], 1) if cell.value}
    mode_col = headers.get('Mode.Operation', 1)
    fuel_i_col = headers.get('Fuel.I', 2)
    tech_col = headers.get('Tech', 6)
    fuel_o_col = headers.get('Fuel.O', 8)

    # Collect TRN interconnection rows
    # Pattern: TRN + origin_country(3) + origin_region(2) + dest_country(3) + dest_region(2) = 13 chars
    trn_data = {}  # tech_code -> {mode -> {input_fuel, output_fuel}}

    for row_idx in range(2, ws.max_row + 1):
        tech = ws.cell(row_idx, tech_col).value
        if not tech:
            continue
        tech_str = str(tech).strip()

        # Only 13-char TRN codes are cross-border/intra-country interconnections
        if not tech_str.startswith('TRN') or len(tech_str) != 13:
            continue

        mode = ws.cell(row_idx, mode_col).value
        fuel_i = ws.cell(row_idx, fuel_i_col).value
        fuel_o = ws.cell(row_idx, fuel_o_col).value

        if tech_str not in trn_data:
            trn_data[tech_str] = {}

        mode_int = int(mode) if mode else 0
        trn_data[tech_str][mode_int] = {
            'input_fuel': str(fuel_i).strip() if fuel_i else '',
            'output_fuel': str(fuel_o).strip() if fuel_o else '',
        }

    wb.close()

    # Build result list
    result = []
    for tech_code in sorted(trn_data.keys()):
        modes = trn_data[tech_code]

        # Derive labels from tech code
        origin_country = tech_code[3:6]
        origin_region = tech_code[6:8]
        dest_country = tech_code[8:11]
        dest_region = tech_code[11:13]

        mode1_label = f"{origin_country}-{origin_region} -> {dest_country}-{dest_region}"
        mode2_label = f"{dest_country}-{dest_region} -> {origin_country}-{origin_region}"

        entry = {
            'tech': tech_code,
            'mode1_label': mode1_label,
            'mode2_label': mode2_label,
            'mode1_input_fuel': modes.get(1, {}).get('input_fuel', ''),
            'mode1_output_fuel': modes.get(1, {}).get('output_fuel', ''),
            'mode2_input_fuel': modes.get(2, {}).get('input_fuel', ''),
            'mode2_output_fuel': modes.get(2, {}).get('output_fuel', ''),
        }
        result.append(entry)

    print(f"  Found {len(result)} TRN interconnections ({sum(len(m) for m in trn_data.values())} total directions)")
    return result


def create_editor_template(data, output_path):
    """
    Create the Excel editor template with dropdowns and validation

    Args:
        data: dict with scenarios, countries, technologies, parameters, years
        output_path: Path where to save the template
    """
    from openpyxl.workbook.defined_name import DefinedName

    print("Creating Excel template...")

    wb = openpyxl.Workbook()

    # Force automatic calculation in Excel
    wb.calculation.calcMode = 'auto'
    wb.calculation.fullCalcOnLoad = True

    # Main sheet for data entry
    ws_main = wb.active
    ws_main.title = "Editor"

    # Create hidden sheets for dropdown lists and mappings
    ws_scenarios = wb.create_sheet("_Scenarios")
    ws_countries = wb.create_sheet("_Countries")
    ws_tech_names = wb.create_sheet("_TechNames")
    ws_tech_mapping = wb.create_sheet("_TechMapping")  # Tech.Name -> Tech code
    ws_parameters = wb.create_sheet("_Parameters")
    ws_tech_by_country = wb.create_sheet("_TechByCountry")  # Tech.Name organized by country

    # Hide validation sheets
    ws_scenarios.sheet_state = 'hidden'
    ws_countries.sheet_state = 'hidden'
    ws_tech_names.sheet_state = 'hidden'
    ws_tech_mapping.sheet_state = 'hidden'
    ws_parameters.sheet_state = 'hidden'
    ws_tech_by_country.sheet_state = 'hidden'

    # Populate validation sheets
    for idx, scenario in enumerate(['ALL'] + data['scenarios'], 1):
        ws_scenarios.cell(idx, 1, scenario)

    for idx, country in enumerate(data['countries'], 1):
        ws_countries.cell(idx, 1, country)

    # Populate Tech.Name list and Tech.Name -> Tech mapping
    for idx, (tech_name, tech_code) in enumerate(sorted(data['tech_mapping'].items()), 1):
        ws_tech_names.cell(idx, 1, tech_name)  # Column A: Tech.Name for dropdown
        ws_tech_mapping.cell(idx, 1, tech_name)  # Column A: Tech.Name
        ws_tech_mapping.cell(idx, 2, tech_code)  # Column B: Tech code

    for idx, param in enumerate(data['parameters'], 1):
        ws_parameters.cell(idx, 1, param)

    # Populate _TechByCountry sheet and create named ranges for each country
    # Each column will have: Row 1 = Country code, Rows 2+ = Tech.Name list for that country
    sorted_countries = sorted(data['tech_by_country'].keys())
    for col_idx, country_code in enumerate(sorted_countries, 1):
        tech_list = data['tech_by_country'][country_code]
        col_letter = openpyxl.utils.get_column_letter(col_idx)

        # Row 1: Country code as header
        ws_tech_by_country.cell(1, col_idx, country_code)

        # Rows 2+: Tech.Name list
        for row_idx, tech_name in enumerate(tech_list, 2):
            ws_tech_by_country.cell(row_idx, col_idx, tech_name)

        # Create named range for this country (e.g., "Tech_ARG")
        # Range is from row 2 to row (1 + number of techs)
        num_techs = len(tech_list)
        range_ref = f"'_TechByCountry'!${col_letter}$2:${col_letter}${1 + num_techs}"
        defined_name = DefinedName(f"Tech_{country_code}", attr_text=range_ref)
        wb.defined_names.add(defined_name)

    # Create header row in main sheet
    # Columns: Scenario, Country, Tech.Name, Tech (auto-filled), Parameter, Years...
    headers = ['Scenario', 'Country', 'Tech.Name', 'Tech', 'Parameter'] + [str(year) for year in data['years']]

    # Style for header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(headers, 1):
        cell = ws_main.cell(1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Set column widths
    ws_main.column_dimensions['A'].width = 15  # Scenario
    ws_main.column_dimensions['B'].width = 12  # Country
    ws_main.column_dimensions['C'].width = 40  # Tech.Name
    ws_main.column_dimensions['D'].width = 20  # Tech (auto-filled, read-only)
    ws_main.column_dimensions['E'].width = 30  # Parameter

    # Year columns (start from column F = 6)
    for col_idx in range(6, len(headers) + 1):
        ws_main.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 10

    # Add data validations (dropdowns) for first 100 rows
    max_data_rows = 100

    # Scenario dropdown (column A)
    dv_scenario = DataValidation(
        type="list",
        formula1=f"=_Scenarios!$A$1:$A${len(data['scenarios']) + 1}",
        allow_blank=False
    )
    dv_scenario.error = 'Please select a valid scenario'
    dv_scenario.errorTitle = 'Invalid Scenario'
    ws_main.add_data_validation(dv_scenario)
    dv_scenario.add(f'A2:A{max_data_rows + 1}')

    # Country dropdown (column B)
    dv_country = DataValidation(
        type="list",
        formula1=f"=_Countries!$A$1:$A${len(data['countries'])}",
        allow_blank=False
    )
    dv_country.error = 'Please select a valid country'
    dv_country.errorTitle = 'Invalid Country'
    ws_main.add_data_validation(dv_country)
    dv_country.add(f'B2:B{max_data_rows + 1}')

    # Tech.Name dropdown (column C) - depends on Country selected in column B
    # Uses INDIRECT to reference named range "Tech_ARG", "Tech_BOL", etc.
    # The formula uses relative reference B2, but Excel will adjust it for each row
    dv_tech_name = DataValidation(
        type="list",
        formula1='=INDIRECT("Tech_"&$B2)',
        allow_blank=False
    )
    dv_tech_name.error = 'Please select a valid technology for this country'
    dv_tech_name.errorTitle = 'Invalid Tech.Name'
    ws_main.add_data_validation(dv_tech_name)
    dv_tech_name.add(f'C2:C{max_data_rows + 1}')

    # Column D (Tech) will be auto-filled with VLOOKUP formula
    # Add formulas to map Tech.Name to Tech code
    for row_idx in range(2, max_data_rows + 2):
        formula = f'=IFERROR(VLOOKUP(C{row_idx},_TechMapping!$A:$B,2,FALSE),"")'
        ws_main.cell(row_idx, 4, formula)  # Column D

    # Protect column D (Tech) from manual editing
    for row_idx in range(2, max_data_rows + 2):
        ws_main.cell(row_idx, 4).protection = Protection(locked=True)

    # Parameter dropdown (column E)
    dv_parameter = DataValidation(
        type="list",
        formula1=f"=_Parameters!$A$1:$A${len(data['parameters'])}",
        allow_blank=False
    )
    dv_parameter.error = 'Please select a valid parameter'
    dv_parameter.errorTitle = 'Invalid Parameter'
    ws_main.add_data_validation(dv_parameter)
    dv_parameter.add(f'E2:E{max_data_rows + 1}')

    # Freeze panes (freeze first row and first 5 columns)
    ws_main.freeze_panes = 'F2'

    # Add instructions in a separate sheet (first sheet - index 0)
    ws_instructions = wb.create_sheet("Instructions", 0)
    ws_instructions.column_dimensions['A'].width = 80

    # Add OSTRAM Configuration sheet (second sheet - index 1)
    ws_ostram = wb.create_sheet("OSTRAM_Config", 1)
    ws_ostram.column_dimensions['A'].width = 80
    ws_ostram.column_dimensions['B'].width = 20

    # Style for configuration sheet
    header_fill_config = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    header_font_config = Font(bold=True, color="FFFFFF", size=12)
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    cell = ws_ostram.cell(1, 1, "OSTRAM CONFIGURATION")
    cell.font = Font(size=14, bold=True, color="366092")
    ws_ostram.merge_cells('A1:B1')

    # Instructions
    ws_ostram.cell(2, 1, "Fill in the configuration below to automatically populate ResidualCapacity from OSTRAM data")
    ws_ostram.merge_cells('A2:B2')
    ws_ostram.cell(2, 1).font = Font(italic=True)

    # Configuration headers
    ws_ostram.cell(4, 1, "Configuration Parameter").fill = header_fill_config
    ws_ostram.cell(4, 1).font = header_font_config
    ws_ostram.cell(4, 1).border = border_style
    ws_ostram.cell(4, 2, "Value").fill = header_fill_config
    ws_ostram.cell(4, 2).font = header_font_config
    ws_ostram.cell(4, 2).border = border_style

    # ResidualCapacitiesFromOSTRAM
    ws_ostram.cell(5, 1, "ResidualCapacitiesFromOSTRAM").border = border_style
    ws_ostram.cell(5, 2, "NO").border = border_style
    dv_yes_no = DataValidation(type="list", formula1='"YES,NO"', allow_blank=False)
    ws_ostram.add_data_validation(dv_yes_no)
    dv_yes_no.add('B5')

    # PetroleumSplitMode
    ws_ostram.cell(6, 1, "PetroleumSplitMode").border = border_style
    ws_ostram.cell(6, 2, "Split_PET_OIL").border = border_style
    dv_petroleum = DataValidation(type="list", formula1='"OIL_only,Split_PET_OIL"', allow_blank=False)
    ws_ostram.add_data_validation(dv_petroleum)
    dv_petroleum.add('B6')

    # DemandFromOSTRAM
    ws_ostram.cell(7, 1, "DemandFromOSTRAM").border = border_style
    ws_ostram.cell(7, 2, "NO").border = border_style
    dv_yes_no.add('B7')

    # ActivityLowerLimitFromOSTRAM
    ws_ostram.cell(8, 1, "ActivityLowerLimitFromOSTRAM").border = border_style
    ws_ostram.cell(8, 2, "NO").border = border_style
    dv_yes_no.add('B8')

    # ActivityUpperLimitFromOSTRAM
    ws_ostram.cell(9, 1, "ActivityUpperLimitFromOSTRAM").border = border_style
    ws_ostram.cell(9, 2, "NO").border = border_style
    dv_yes_no.add('B9')

    # TradeBalanceDemandAdjustment
    ws_ostram.cell(10, 1, "TradeBalanceDemandAdjustment").border = border_style
    ws_ostram.cell(10, 2, "NO").border = border_style
    dv_yes_no.add('B10')

    # InterconnectionsControl
    ws_ostram.cell(11, 1, "InterconnectionsControl").border = border_style
    ws_ostram.cell(11, 2, "NO").border = border_style
    dv_yes_no.add('B11')

    # Add detailed descriptions with formulas and data sources
    ws_ostram.cell(13, 1, "DETAILED DESCRIPTIONS AND FORMULAS")
    ws_ostram.cell(13, 1).font = Font(bold=True, size=12, color="366092")
    ws_ostram.merge_cells('A13:B13')

    current_row = 15

    # ResidualCapacitiesFromOSTRAM
    ws_ostram.cell(current_row, 1, "1. ResidualCapacitiesFromOSTRAM")
    ws_ostram.cell(current_row, 1).font = Font(bold=True, size=11)
    ws_ostram.merge_cells(f'A{current_row}:B{current_row}')
    current_row += 1

    residual_desc = [
        "Populates the ResidualCapacity parameter for PWR (power generation) technologies.",
        "",
        "DATA SOURCE:",
        "  File: 'OSTRAM - Installed Capacity by Source - Annual.xlsx'",
        "  Sheet: '1.2023' (or corresponding year)",
        "  Row 5: Country names",
        "  Rows 6-20: Technology capacities in MW",
        "",
        "FORMULA:",
        "  ResidualCapacity (GW) = OSTRAM_Capacity (MW) / 1000",
        "",
        "NOTES:",
        "  - Same flat capacity value is used for all model years",
        "  - Only applies to PWR technologies (power plants)",
    ]
    for line in residual_desc:
        ws_ostram.cell(current_row, 1, line)
        ws_ostram.cell(current_row, 1).font = Font(size=9)
        ws_ostram.merge_cells(f'A{current_row}:B{current_row}')
        current_row += 1
    current_row += 1

    # PetroleumSplitMode
    ws_ostram.cell(current_row, 1, "2. PetroleumSplitMode")
    ws_ostram.cell(current_row, 1).font = Font(bold=True, size=11)
    ws_ostram.merge_cells(f'A{current_row}:B{current_row}')
    current_row += 1

    petroleum_desc = [
        "Controls how petroleum-based generation is split between PET (Diesel) and OIL (Fuel oil).",
        "",
        "OPTIONS:",
        "  OIL_only: Assign all petroleum to OIL technology",
        "  Split_PET_OIL: Split using shares from Shares_PET_OIL_Split.xlsx",
        "",
        "DATA SOURCE (for Split_PET_OIL):",
        "  File: 'Shares_PET_OIL_Split.xlsx'",
        "  Sheets: SharesBAU, SharesNDC, SharesNDC_NoRPO, SharesNDC+ELC",
        "  Contains: Diesel, Fuel oil, Bunker shares by country and year",
        "",
        "FORMULA (Split_PET_OIL):",
        "  PET_share = Diesel_share",
        "  OIL_share = Fuel_oil_share + Bunker_share",
    ]
    for line in petroleum_desc:
        ws_ostram.cell(current_row, 1, line)
        ws_ostram.cell(current_row, 1).font = Font(size=9)
        ws_ostram.merge_cells(f'A{current_row}:B{current_row}')
        current_row += 1
    current_row += 1

    # DemandFromOSTRAM
    ws_ostram.cell(current_row, 1, "3. DemandFromOSTRAM")
    ws_ostram.cell(current_row, 1).font = Font(bold=True, size=11)
    ws_ostram.merge_cells(f'A{current_row}:B{current_row}')
    current_row += 1

    demand_desc = [
        "Populates SpecifiedAnnualDemand in A-O_Demand.xlsx for electricity demand.",
        "",
        "DATA SOURCE:",
        "  File: 'OSTRAM - Electric Generation by Source - Annual.xlsx'",
        "  Sheet: '1.2023'",
        "  Row 21: Total generation by country (GWh)",
        "  Reference year: 2023 (from cell A4)",
        "",
        "FORMULA:",
        "  Demand(year) = Generation_OSTRAM(PJ) x (1 + growth_rate x (year - 2023))",
        "",
        "UNIT CONVERSION:",
        "  1 GWh = 0.0036 PJ",
        "  Generation_PJ = Generation_GWh x 0.0036",
        "",
        "GROWTH RATES:",
        "  Configured per country in 'Demand_Growth' sheet (default: 2% annual)",
    ]
    for line in demand_desc:
        ws_ostram.cell(current_row, 1, line)
        ws_ostram.cell(current_row, 1).font = Font(size=9)
        ws_ostram.merge_cells(f'A{current_row}:B{current_row}')
        current_row += 1
    current_row += 1

    # ActivityLowerLimitFromOSTRAM
    ws_ostram.cell(current_row, 1, "4. ActivityLowerLimitFromOSTRAM")
    ws_ostram.cell(current_row, 1).font = Font(bold=True, size=11)
    ws_ostram.merge_cells(f'A{current_row}:B{current_row}')
    current_row += 1

    lower_desc = [
        "Populates TotalTechnologyAnnualActivityLowerLimit in A-O_Parametrization.xlsx.",
        "",
        "DATA SOURCES:",
        "  1. 'OSTRAM - Electric Generation by Source - Annual.xlsx' - Total generation",
        "  2. 'Shares_Power_Generation_Technologies.xlsx' - Technology shares by country/scenario/year",
        "  3. 'Renewability_Targets' sheet - Target renewable % by year (optional)",
        "  4. 'Technology_Weights' sheet - Custom tech distribution (optional)",
        "",
        "FORMULA:",
        "  LowerLimit(tech,year) = Generation_Total(PJ) x (1 + growth_rate x (year - 2023)) x Share(tech,year)",
        "",
        "SHARE CALCULATION:",
        "  - If Renewability_Targets defined: Interpolate shares to reach renewable % targets",
        "  - If not defined: Use shares directly from Shares_Power_Generation_Technologies.xlsx",
        "",
        "INTERPOLATION MODES (in Renewability_Targets):",
        "  - 'linear': Linear interpolation between base year and target years",
        "  - 'flat_step': Keep flat until target year, then step up (staircase)",
    ]
    for line in lower_desc:
        ws_ostram.cell(current_row, 1, line)
        ws_ostram.cell(current_row, 1).font = Font(size=9)
        ws_ostram.merge_cells(f'A{current_row}:B{current_row}')
        current_row += 1
    current_row += 1

    # ActivityUpperLimitFromOSTRAM
    ws_ostram.cell(current_row, 1, "5. ActivityUpperLimitFromOSTRAM")
    ws_ostram.cell(current_row, 1).font = Font(bold=True, size=11)
    ws_ostram.merge_cells(f'A{current_row}:B{current_row}')
    current_row += 1

    upper_desc = [
        "Populates TotalTechnologyAnnualActivityUpperLimit in A-O_Parametrization.xlsx.",
        "",
        "FORMULA:",
        "  UpperLimit(tech,year) = LowerLimit(tech,year) + 0.1",
        "",
        "NOTES:",
        "  - Can be enabled independently of LowerLimit",
        "  - If only UpperLimit is enabled, base value is calculated using same formula as LowerLimit",
        "  - The +0.1 margin allows slight flexibility in the optimization",
    ]
    for line in upper_desc:
        ws_ostram.cell(current_row, 1, line)
        ws_ostram.cell(current_row, 1).font = Font(size=9)
        ws_ostram.merge_cells(f'A{current_row}:B{current_row}')
        current_row += 1
    current_row += 1

    # ActivityLowerLimitMethod (DemandBased)
    ws_ostram.cell(current_row, 1, "6. ActivityLowerLimit Calculation Method")
    ws_ostram.cell(current_row, 1).font = Font(bold=True, size=11)
    ws_ostram.merge_cells(f'A{current_row}:B{current_row}')
    current_row += 1

    method_desc = [
        "Calculates TotalTechnologyAnnualActivityLowerLimit directly from projected demand and normalized shares.",
        "Only applies when ActivityLowerLimitFromOSTRAM = YES.",
        "",
        "CALCULATION METHOD:",
        "  Formula: LowerLimit[tech,year] = Demand[country,year] × Share_normalized[tech,year]",
        "  Step 1: Read projected demand from A-O_Demand.xlsx (SpecifiedAnnualDemand)",
        "  Step 2: Read % renewable from Renewability_Targets (with interpolation)",
        "  Step 3: Distribute % renewable among renewable fuels using OSTRAM generation weights",
        "  Step 4: Distribute % non-renewable among non-renewable fuels using OSTRAM generation weights",
        "  Step 5: Normalize all shares year-by-year to ensure Σ(shares) = 1.0",
        "  Step 6: For non-base scenarios, override years 2023-2025 with base scenario shares",
        "  Step 7: Calculate LowerLimit = Demand × Share for each technology/year",
        "",
        "REQUIREMENTS:",
        "  - ActivityLowerLimitFromOSTRAM must be YES",
        "  - A-O_Demand.xlsx must exist in the scenario output folder",
        "  - Renewability_Targets sheet must define % renewable targets",
        "  - OSTRAM generation data must be available for weight calculation",
        "  - base_scenario defined in Config_MOMF_T1_AB.yaml",
        "",
        "FUEL GROUPS:",
        "  Renewable: HYD, SPV, WON, GEO, BIO, CSP, WOF, WAV",
        "  Non-renewable: COA, NGS, OIL, PET, URN",
        "  Other (weight=0 if no OSTRAM data): BCK, CCS, COG, LDS, OTH, SDS, WAS",
        "",
        "BASE SCENARIO OVERRIDE:",
        "  - For all scenarios different from base_scenario (defined in Config_MOMF_T1_AB.yaml)",
        "  - Years 2023, 2024, 2025 use base scenario shares (policy effects start from 2026)",
        "  - This ensures realistic transition timing",
    ]
    for line in method_desc:
        ws_ostram.cell(current_row, 1, line)
        ws_ostram.cell(current_row, 1).font = Font(size=9)
        ws_ostram.merge_cells(f'A{current_row}:B{current_row}')
        current_row += 1
    current_row += 1

    # Technology mapping section
    ws_ostram.cell(current_row, 1, "TECHNOLOGY MAPPING (Source files -> Model)")
    ws_ostram.cell(current_row, 1).font = Font(bold=True, size=11, color="366092")
    ws_ostram.merge_cells(f'A{current_row}:B{current_row}')
    current_row += 1

    tech_mapping_desc = [
        "",
        "RENEWABLE TECHNOLOGIES:",
        "  Biomass         -> BIO (Biomass)",
        "  Wind            -> WON (Wind)",
        "  Geothermal      -> GEO (Geothermal)",
        "  Hydro           -> HYD (Hydro)",
        "  Solar           -> SPV (Solar PV)",
        "",
        "NON-RENEWABLE TECHNOLOGIES:",
        "  Mineral coal    -> COA (Coal)",
        "  Petroleum       -> PET (Petroleum/Diesel)",
        "  Natural gas     -> NGS (Natural Gas)",
        "  Nuclear         -> URN (Uranium/Nuclear)",
        "  Fuel oil + Bunker -> OIL (Oil/Fuel oil)",
    ]
    for line in tech_mapping_desc:
        ws_ostram.cell(current_row, 1, line)
        ws_ostram.cell(current_row, 1).font = Font(size=9)
        ws_ostram.merge_cells(f'A{current_row}:B{current_row}')
        current_row += 1
    current_row += 1

    # InterconnectionsControl
    ws_ostram.cell(current_row, 1, "7. InterconnectionsControl")
    ws_ostram.cell(current_row, 1).font = Font(bold=True, size=11)
    ws_ostram.merge_cells(f'A{current_row}:B{current_row}')
    current_row += 1

    intercon_desc = [
        "Controls ON/OFF state of TRN (transmission interconnection) technologies.",
        "Each interconnection is bidirectional and can be toggled per direction.",
        "",
        "WHEN ENABLED (InterconnectionsControl = YES):",
        "  The 'Interconnections' sheet controls each direction independently.",
        "  Changes are applied to both A-O_AR_Model_Base_Year.xlsx and A-O_AR_Projections.xlsx.",
        "",
        "OFF DIRECTION:",
        "  - Base Year: Value.Fuel.I = 0, Value.Fuel.O = 0",
        "  - Projections: Projection.Mode = 'EMPTY'",
        "",
        "ON DIRECTION:",
        "  - Base Year: Value.Fuel.I = 1, Value.Fuel.O = 1",
        "  - Projections: Projection.Mode = 'User defined'",
        "",
        "NOTE: Year values in Projections (efficiency/loss factors) are NOT modified.",
    ]
    for line in intercon_desc:
        ws_ostram.cell(current_row, 1, line)
        ws_ostram.cell(current_row, 1).font = Font(size=9)
        ws_ostram.merge_cells(f'A{current_row}:B{current_row}')
        current_row += 1

    # Populate instructions sheet
    instructions = [
        ["Secondary Techs Editor - Instructions", ""],
        ["", ""],
        ["This Excel file allows you to edit Secondary Techs data easily.", ""],
        ["", ""],
        ["HOW TO USE:", ""],
        ["1. (Optional) Configure OSTRAM settings in 'OSTRAM_Config' sheet", ""],
        ["2. Go to the 'Editor' sheet", ""],
        ["3. Fill in each row with:", ""],
        ["   - Scenario: Select BAU, NDC, NDC+ELC, NDC_NoRPO, or ALL (applies to all scenarios)", ""],
        ["   - Country: Select the country code (ARG, BOL, CHL, COL, ECU, GTM, etc.)", ""],
        ["   - Tech.Name: Select from dropdown (options depend on country selected)", ""],
        ["   - Tech: This column is auto-filled based on Tech.Name (READ-ONLY)", ""],
        ["   - Parameter: Select the parameter to modify", ""],
        ["   - Year values: Enter numeric values for each year (leave empty to keep current value)", ""],
        ["", ""],
        ["4. Save and close this file", ""],
        ["5. Run: python t1_confection/D2_update_secondary_techs.py", ""],
        ["", ""],
        ["OSTRAM INTEGRATION:", ""],
        ["- If ResidualCapacitiesFromOSTRAM = YES in OSTRAM_Config sheet:", ""],
        ["  * The script will automatically populate ResidualCapacity for PWR technologies", ""],
        ["  * Data comes from 'OSTRAM - Installed Capacity by Source - Annual.xlsx'", ""],
        ["  * OSTRAM data takes priority over manual Editor entries for ResidualCapacity", ""],
        ["  * Flat capacity values are used (same value for all years)", ""],
        ["", ""],
        ["- If DemandFromOSTRAM = YES in OSTRAM_Config sheet:", ""],
        ["  * The script will populate electricity demand in A-O_Demand.xlsx", ""],
        ["  * Data comes from 'OSTRAM - Electric Generation by Source - Annual.xlsx'", ""],
        ["  * Configure growth rates per country in the 'Demand_Growth' sheet", ""],
        ["  * Data is converted from GWh to PJ (1 GWh = 0.0036 PJ)", ""],
        ["  * Linear growth is applied from the OSTRAM reference year (2023)", ""],
        ["", ""],
        ["- If ActivityLowerLimitFromOSTRAM = YES in OSTRAM_Config sheet:", ""],
        ["  * The script will populate TotalTechnologyAnnualActivityLowerLimit in A-O_Parametrization.xlsx", ""],
        ["  * Formula: LowerLimit[tech,year] = Demand[country,year] × Share_normalized[tech,year]", ""],
        ["  * Shares are calculated from Renewability_Targets and OSTRAM generation weights", ""],
        ["  * Demand values come from A-O_Demand.xlsx (with projected growth)", ""],
        ["", ""],
        ["INTERCONNECTION CONTROL:", ""],
        ["- If InterconnectionsControl = YES in OSTRAM_Config sheet:", ""],
        ["  * The 'Interconnections' sheet controls TRN cross-border/intra-country links", ""],
        ["  * Set each direction to ON or OFF independently", ""],
        ["  * OFF directions: InputActivityRatio=0, OutputActivityRatio=0, Projection.Mode='EMPTY'", ""],
        ["  * ON directions: InputActivityRatio=1, OutputActivityRatio=1, Projection.Mode='User defined'", ""],
        ["  * Changes apply to A-O_AR_Model_Base_Year.xlsx and A-O_AR_Projections.xlsx", ""],
        ["", ""],
        ["IMPORTANT NOTES:", ""],
        ["- You can add as many rows as needed", ""],
        ["- The 'Tech' column is automatically filled when you select a 'Tech.Name'", ""],
        ["- Empty year cells will NOT modify the current value in the destination file", ""],
        ["- Scenario 'ALL' will apply changes to all 4 scenario files", ""],
        ["- A backup will be created automatically before making changes", ""],
        ["- Check the log file after running the update script", ""],
        ["", ""],
        [f"Template generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ""],
    ]

    for row_idx, row_data in enumerate(instructions, 1):
        cell = ws_instructions.cell(row_idx, 1, row_data[0])
        if row_idx == 1:
            cell.font = Font(size=16, bold=True, color="366092")
        elif (row_data[0].startswith("HOW TO USE:") or row_data[0].startswith("IMPORTANT NOTES:") or
              row_data[0].startswith("OSTRAM INTEGRATION:") or row_data[0].startswith("INTERCONNECTION CONTROL:")):
            cell.font = Font(size=12, bold=True)

        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Create Demand_Growth sheet for country-specific growth rates
    ws_demand = wb.create_sheet("Demand_Growth", 2)
    ws_demand.column_dimensions['A'].width = 15
    ws_demand.column_dimensions['B'].width = 30
    ws_demand.column_dimensions['C'].width = 20

    # Title
    cell = ws_demand.cell(1, 1, "ELECTRICITY DEMAND GROWTH RATES")
    cell.font = Font(size=14, bold=True, color="366092")
    ws_demand.merge_cells('A1:C1')

    # Instructions
    ws_demand.cell(2, 1, "Configure the annual linear growth rate (%) for each country's electricity demand")
    ws_demand.merge_cells('A2:C2')
    ws_demand.cell(2, 1).font = Font(italic=True)

    # Headers
    demand_header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    demand_header_font = Font(bold=True, color="FFFFFF")

    ws_demand.cell(4, 1, "Country Code").fill = demand_header_fill
    ws_demand.cell(4, 1).font = demand_header_font
    ws_demand.cell(4, 1).border = border_style
    ws_demand.cell(4, 1).alignment = center_align = Alignment(horizontal="center", vertical="center")

    ws_demand.cell(4, 2, "Country Name").fill = demand_header_fill
    ws_demand.cell(4, 2).font = demand_header_font
    ws_demand.cell(4, 2).border = border_style

    ws_demand.cell(4, 3, "Growth Rate (%)").fill = demand_header_fill
    ws_demand.cell(4, 3).font = demand_header_font
    ws_demand.cell(4, 3).border = border_style
    ws_demand.cell(4, 3).alignment = Alignment(horizontal="center", vertical="center")

    # Country list with growth rates (default 2.0%) from centralized config
    country_names = get_country_names()
    demand_countries = [
        (iso3, name, 2.0) for iso3, name in sorted(country_names.items())
    ]

    for row_idx, (code, name, rate) in enumerate(demand_countries, 5):
        ws_demand.cell(row_idx, 1, code).border = border_style
        ws_demand.cell(row_idx, 1).alignment = Alignment(horizontal="center")
        ws_demand.cell(row_idx, 2, name).border = border_style
        ws_demand.cell(row_idx, 3, rate).border = border_style
        ws_demand.cell(row_idx, 3).number_format = '0.00'
        ws_demand.cell(row_idx, 3).alignment = Alignment(horizontal="center")

    # Add note at the bottom
    note_row = 5 + len(demand_countries) + 1
    ws_demand.cell(note_row, 1, "Note: Growth rate is applied linearly from the OSTRAM reference year (2023).")
    ws_demand.merge_cells(f'A{note_row}:C{note_row}')
    ws_demand.cell(note_row, 1).font = Font(italic=True, size=9)

    ws_demand.cell(note_row + 1, 1, "Formula: Demand(year) = Demand(2023) × (1 + rate × (year - 2023))")
    ws_demand.merge_cells(f'A{note_row + 1}:C{note_row + 1}')
    ws_demand.cell(note_row + 1, 1).font = Font(italic=True, size=9)

    # =========================================================================
    # Create Renewability_Targets sheet
    # =========================================================================
    ws_renew = wb.create_sheet("Renewability_Targets", 3)
    ws_renew.column_dimensions['A'].width = 12
    ws_renew.column_dimensions['B'].width = 15
    ws_renew.column_dimensions['C'].width = 15

    # OSTRAM base year is 2023 - filter years to start from 2023
    ostram_base_year = 2023
    renew_years = [y for y in data['years'] if y >= ostram_base_year]

    # Title
    cell = ws_renew.cell(1, 1, "RENEWABILITY TARGETS")
    cell.font = Font(size=14, bold=True, color="366092")
    ws_renew.merge_cells('A1:C1')

    # Instructions
    ws_renew.cell(2, 1, "Define the target percentage of RENEWABLE generation for each country/scenario.")
    ws_renew.merge_cells(f'A2:{openpyxl.utils.get_column_letter(3 + len(renew_years))}2')
    ws_renew.cell(2, 1).font = Font(italic=True)

    ws_renew.cell(3, 1, f"Base year is {ostram_base_year} (from OSTRAM data). Only define targets for years where you want to set a specific renewable %.")
    ws_renew.merge_cells(f'A3:{openpyxl.utils.get_column_letter(3 + len(renew_years))}3')
    ws_renew.cell(3, 1).font = Font(italic=True, size=9)

    # Headers
    renew_header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    renew_header_font = Font(bold=True, color="FFFFFF")

    renew_headers = ['Country', 'Scenario', 'Interpolation'] + [str(year) for year in renew_years]
    for col_idx, header in enumerate(renew_headers, 1):
        cell = ws_renew.cell(5, col_idx, header)
        cell.fill = renew_header_fill
        cell.font = renew_header_font
        cell.border = border_style
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col_idx > 3:
            ws_renew.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 8

    # Pre-populate rows for each country+scenario combination
    scenarios = data['scenarios']
    countries_list = sorted(data['countries'])

    row_idx = 6
    for country in countries_list:
        for scenario in scenarios:
            ws_renew.cell(row_idx, 1, country).border = border_style
            ws_renew.cell(row_idx, 1).alignment = Alignment(horizontal="center")
            ws_renew.cell(row_idx, 2, scenario).border = border_style
            ws_renew.cell(row_idx, 2).alignment = Alignment(horizontal="center")
            ws_renew.cell(row_idx, 3, "flat_step").border = border_style
            ws_renew.cell(row_idx, 3).alignment = Alignment(horizontal="center")
            # Year columns - leave empty for user input
            for col_idx in range(4, 4 + len(renew_years)):
                ws_renew.cell(row_idx, col_idx).border = border_style
                ws_renew.cell(row_idx, col_idx).number_format = '0.00%'
            row_idx += 1

    # Add Interpolation dropdown
    dv_interpolation = DataValidation(type="list", formula1='"linear,flat_step"', allow_blank=False)
    ws_renew.add_data_validation(dv_interpolation)
    dv_interpolation.add(f'C6:C{row_idx - 1}')

    # Add detailed notes at the bottom
    note_row = row_idx + 1
    ws_renew.cell(note_row, 1, "HOW IT WORKS:")
    ws_renew.merge_cells(f'A{note_row}:C{note_row}')
    ws_renew.cell(note_row, 1).font = Font(bold=True, size=10)
    note_row += 1

    notes = [
        f"1. Base year ({ostram_base_year}): Uses actual renewable % from OSTRAM/Shares_Power_Generation_Technologies.xlsx data",
        "2. Target years: Define your desired renewable % in specific years (e.g., 50% in 2030, 70% in 2040)",
        "3. Empty cells: Values are interpolated based on the interpolation mode selected",
        "",
        "INTERPOLATION MODES:",
        "  - linear: Smooth linear progression from base year to each target year",
        "  - flat_step: Keep flat at previous value until target year, then jump to new value (staircase)",
        "",
        "RENEWABLE SOURCES (summed for renewable %):",
        "  HYD (Hydro), SPV (Solar), WND (Wind), GEO (Geothermal), BIO (Biomass)",
        "",
        "EXAMPLE: If base year has 40% renewable and you set 60% for 2030:",
        "  - linear: 40%(2023) -> 44%(2025) -> 48%(2027) -> 52%(2029) -> 60%(2030)",
        "  - flat_step: 40%(2023-2029) -> 60%(2030+)",
    ]
    for note in notes:
        ws_renew.cell(note_row, 1, note)
        ws_renew.cell(note_row, 1).font = Font(size=9)
        ws_renew.merge_cells(f'A{note_row}:{openpyxl.utils.get_column_letter(3 + len(renew_years))}{note_row}')
        note_row += 1

    # Freeze panes
    ws_renew.freeze_panes = 'D6'

    # =========================================================================
    # Create Technology_Weights sheet (renamed from Renewable_Weights)
    # =========================================================================
    ws_weights = wb.create_sheet("Technology_Weights", 4)
    ws_weights.column_dimensions['A'].width = 12
    ws_weights.column_dimensions['B'].width = 15

    # All technology columns (same as OSTRAM sources)
    # Renewables: HYD, SPV, WND, GEO, BIO
    # Non-renewables: COA, PET, NGS, URN, OIL
    all_techs = ['HYD', 'SPV', 'WND', 'GEO', 'BIO', 'COA', 'PET', 'NGS', 'URN', 'OIL']
    renewable_techs = ['HYD', 'SPV', 'WND', 'GEO', 'BIO']
    non_renewable_techs = ['COA', 'PET', 'NGS', 'URN', 'OIL']

    # Title
    cell = ws_weights.cell(1, 1, "TECHNOLOGY WEIGHTS (OPTIONAL)")
    cell.font = Font(size=14, bold=True, color="366092")
    ws_weights.merge_cells(f'A1:{openpyxl.utils.get_column_letter(2 + len(all_techs))}1')

    # Warning that this is optional
    ws_weights.cell(2, 1, "*** THIS SHEET IS OPTIONAL - LEAVE EMPTY TO USE DEFAULT BEHAVIOR ***")
    ws_weights.merge_cells(f'A2:{openpyxl.utils.get_column_letter(2 + len(all_techs))}2')
    ws_weights.cell(2, 1).font = Font(bold=True, size=11, color="FF0000")

    # Default behavior explanation
    ws_weights.cell(3, 1, "DEFAULT BEHAVIOR (when this sheet is empty):")
    ws_weights.merge_cells(f'A3:{openpyxl.utils.get_column_letter(2 + len(all_techs))}3')
    ws_weights.cell(3, 1).font = Font(bold=True, size=10)

    default_explanation = [
        "  The renewable % target is distributed PROPORTIONALLY based on each technology's share in OSTRAM base year data.",
        "  Example: If OSTRAM shows HYD=60%, SPV=25%, WND=10%, GEO=3%, BIO=2% and you set 70% renewable target:",
        "    -> HYD gets 70% x (60/100) = 42%, SPV gets 70% x (25/100) = 17.5%, etc.",
        "  Non-renewable technologies are also distributed proportionally within the remaining (1 - renewable%) share.",
    ]
    row_start = 4
    for i, line in enumerate(default_explanation):
        ws_weights.cell(row_start + i, 1, line)
        ws_weights.merge_cells(f'A{row_start + i}:{openpyxl.utils.get_column_letter(2 + len(all_techs))}{row_start + i}')
        ws_weights.cell(row_start + i, 1).font = Font(size=9)

    # Headers row
    header_row = 9
    weights_header_fill_renew = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")  # Green for renewables
    weights_header_fill_nonrenew = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")  # Orange for non-renewables
    weights_header_font = Font(bold=True, color="FFFFFF")

    # Fixed columns
    ws_weights.cell(header_row, 1, "Country").fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws_weights.cell(header_row, 1).font = weights_header_font
    ws_weights.cell(header_row, 1).border = border_style
    ws_weights.cell(header_row, 1).alignment = Alignment(horizontal="center", vertical="center")

    ws_weights.cell(header_row, 2, "Scenario").fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws_weights.cell(header_row, 2).font = weights_header_font
    ws_weights.cell(header_row, 2).border = border_style
    ws_weights.cell(header_row, 2).alignment = Alignment(horizontal="center", vertical="center")

    # Technology columns
    for col_idx, tech in enumerate(all_techs, 3):
        cell = ws_weights.cell(header_row, col_idx, tech)
        if tech in renewable_techs:
            cell.fill = weights_header_fill_renew
        else:
            cell.fill = weights_header_fill_nonrenew
        cell.font = weights_header_font
        cell.border = border_style
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_weights.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 8

    # Add subheader for renewable/non-renewable grouping
    ws_weights.cell(header_row - 1, 3, "RENEWABLE")
    ws_weights.merge_cells(f'C{header_row - 1}:G{header_row - 1}')
    ws_weights.cell(header_row - 1, 3).fill = weights_header_fill_renew
    ws_weights.cell(header_row - 1, 3).font = weights_header_font
    ws_weights.cell(header_row - 1, 3).alignment = Alignment(horizontal="center")

    ws_weights.cell(header_row - 1, 8, "NON-RENEWABLE")
    ws_weights.merge_cells(f'H{header_row - 1}:L{header_row - 1}')
    ws_weights.cell(header_row - 1, 8).fill = weights_header_fill_nonrenew
    ws_weights.cell(header_row - 1, 8).font = weights_header_font
    ws_weights.cell(header_row - 1, 8).alignment = Alignment(horizontal="center")

    # Pre-populate rows for each country+scenario combination
    row_idx = header_row + 1
    for country in countries_list:
        for scenario in scenarios:
            ws_weights.cell(row_idx, 1, country).border = border_style
            ws_weights.cell(row_idx, 1).alignment = Alignment(horizontal="center")
            ws_weights.cell(row_idx, 2, scenario).border = border_style
            ws_weights.cell(row_idx, 2).alignment = Alignment(horizontal="center")
            # Weight columns - leave empty for default (proportional to OSTRAM)
            for col_idx in range(3, 3 + len(all_techs)):
                ws_weights.cell(row_idx, col_idx).border = border_style
                ws_weights.cell(row_idx, col_idx).number_format = '0.00'
            row_idx += 1

    # Add detailed notes at the bottom
    note_row = row_idx + 2
    ws_weights.cell(note_row, 1, "WHEN TO USE THIS SHEET:")
    ws_weights.cell(note_row, 1).font = Font(bold=True, size=10)
    note_row += 1

    notes = [
        "Only fill this sheet if you want to OVERRIDE the default proportional distribution.",
        "For example, if you want Solar to grow faster than its historical proportion suggests.",
        "",
        "RULES FOR WEIGHTS:",
        "  - Renewable weights (HYD+SPV+WND+GEO+BIO) must sum to 1.0",
        "  - Non-renewable weights (COA+PET+NGS+URN+OIL) must sum to 1.0",
        "  - The renewable % from Renewability_Targets determines how much goes to each group",
        "",
        "TECHNOLOGY CODES (same as OSTRAM sources):",
        "  RENEWABLE: HYD=Hydro, SPV=Solar, WND=Wind, GEO=Geothermal, BIO=Biomass",
        "  NON-RENEWABLE: COA=Coal, PET=Diesel, NGS=Natural gas, URN=Nuclear, OIL=Fuel oil+Bunker",
        "",
        "EXAMPLE:",
        "  If Renewability_Targets says 60% renewable for 2030, and you set weights:",
        "    HYD=0.40, SPV=0.35, WND=0.20, GEO=0.03, BIO=0.02 (sum=1.0)",
        "    COA=0.10, PET=0.15, NGS=0.60, URN=0.10, OIL=0.05 (sum=1.0)",
        "  Then for 2030: HYD=60%x0.40=24%, SPV=60%x0.35=21%, NGS=40%x0.60=24%, etc.",
    ]
    for note in notes:
        ws_weights.cell(note_row, 1, note)
        ws_weights.cell(note_row, 1).font = Font(size=9)
        ws_weights.merge_cells(f'A{note_row}:{openpyxl.utils.get_column_letter(2 + len(all_techs))}{note_row}')
        note_row += 1

    # Freeze panes
    ws_weights.freeze_panes = f'C{header_row + 1}'

    # =========================================================================
    # Create Interconnections sheet (toggle ON/OFF for TRN technologies)
    # =========================================================================
    trn_data = collect_trn_interconnections()

    if trn_data:
        ws_intercon = wb.create_sheet("Interconnections", 6)
        ws_intercon.column_dimensions['A'].width = 18   # Technology
        ws_intercon.column_dimensions['B'].width = 22   # Direction
        ws_intercon.column_dimensions['C'].width = 8    # Mode
        ws_intercon.column_dimensions['D'].width = 10   # Status
        ws_intercon.column_dimensions['E'].width = 45   # Description

        # Colors
        intercon_header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        intercon_header_font = Font(bold=True, color="FFFFFF")
        readonly_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        alt_fill_a = PatternFill(start_color="E8EDF3", end_color="E8EDF3", fill_type="solid")
        alt_fill_b = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # Title
        cell = ws_intercon.cell(1, 1, "INTERCONNECTION CONTROLS")
        cell.font = Font(size=14, bold=True, color="366092")
        ws_intercon.merge_cells('A1:E1')

        # Instructions
        ws_intercon.cell(2, 1, "Toggle ON/OFF each direction of TRN interconnection technologies. Requires InterconnectionsControl = YES in OSTRAM_Config.")
        ws_intercon.merge_cells('A2:E2')
        ws_intercon.cell(2, 1).font = Font(italic=True)

        # Header row (row 4)
        intercon_headers = ['Technology', 'Direction', 'Mode', 'Status', 'Description']
        for col_idx, header in enumerate(intercon_headers, 1):
            cell = ws_intercon.cell(4, col_idx, header)
            cell.fill = intercon_header_fill
            cell.font = intercon_header_font
            cell.border = border_style
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data rows starting at row 5
        row_idx = 5
        for trn_idx, trn in enumerate(trn_data):
            # Determine alternating color for this pair
            pair_fill = alt_fill_a if trn_idx % 2 == 0 else alt_fill_b

            # Mode 1 row
            ws_intercon.cell(row_idx, 1, trn['tech']).border = border_style
            ws_intercon.cell(row_idx, 1).fill = pair_fill
            ws_intercon.cell(row_idx, 1).alignment = Alignment(horizontal="center")
            ws_intercon.cell(row_idx, 1).protection = Protection(locked=True)

            ws_intercon.cell(row_idx, 2, trn['mode1_label']).border = border_style
            ws_intercon.cell(row_idx, 2).fill = pair_fill
            ws_intercon.cell(row_idx, 2).alignment = Alignment(horizontal="center")
            ws_intercon.cell(row_idx, 2).protection = Protection(locked=True)

            ws_intercon.cell(row_idx, 3, 1).border = border_style
            ws_intercon.cell(row_idx, 3).fill = pair_fill
            ws_intercon.cell(row_idx, 3).alignment = Alignment(horizontal="center")
            ws_intercon.cell(row_idx, 3).protection = Protection(locked=True)

            ws_intercon.cell(row_idx, 4, "ON").border = border_style
            ws_intercon.cell(row_idx, 4).alignment = Alignment(horizontal="center")
            ws_intercon.cell(row_idx, 4).protection = Protection(locked=False)

            desc_1 = f"{trn['mode1_input_fuel']} -> {trn['mode1_output_fuel']}"
            ws_intercon.cell(row_idx, 5, desc_1).border = border_style
            ws_intercon.cell(row_idx, 5).fill = pair_fill
            ws_intercon.cell(row_idx, 5).protection = Protection(locked=True)
            ws_intercon.cell(row_idx, 5).font = Font(size=9)
            row_idx += 1

            # Mode 2 row
            ws_intercon.cell(row_idx, 1, trn['tech']).border = border_style
            ws_intercon.cell(row_idx, 1).fill = pair_fill
            ws_intercon.cell(row_idx, 1).alignment = Alignment(horizontal="center")
            ws_intercon.cell(row_idx, 1).protection = Protection(locked=True)

            ws_intercon.cell(row_idx, 2, trn['mode2_label']).border = border_style
            ws_intercon.cell(row_idx, 2).fill = pair_fill
            ws_intercon.cell(row_idx, 2).alignment = Alignment(horizontal="center")
            ws_intercon.cell(row_idx, 2).protection = Protection(locked=True)

            ws_intercon.cell(row_idx, 3, 2).border = border_style
            ws_intercon.cell(row_idx, 3).fill = pair_fill
            ws_intercon.cell(row_idx, 3).alignment = Alignment(horizontal="center")
            ws_intercon.cell(row_idx, 3).protection = Protection(locked=True)

            ws_intercon.cell(row_idx, 4, "ON").border = border_style
            ws_intercon.cell(row_idx, 4).alignment = Alignment(horizontal="center")
            ws_intercon.cell(row_idx, 4).protection = Protection(locked=False)

            desc_2 = f"{trn['mode2_input_fuel']} -> {trn['mode2_output_fuel']}"
            ws_intercon.cell(row_idx, 5, desc_2).border = border_style
            ws_intercon.cell(row_idx, 5).fill = pair_fill
            ws_intercon.cell(row_idx, 5).protection = Protection(locked=True)
            ws_intercon.cell(row_idx, 5).font = Font(size=9)
            row_idx += 1

        # Add ON/OFF dropdown validation for column D
        dv_on_off = DataValidation(type="list", formula1='"ON,OFF"', allow_blank=False)
        dv_on_off.error = 'Please select ON or OFF'
        dv_on_off.errorTitle = 'Invalid Status'
        ws_intercon.add_data_validation(dv_on_off)
        dv_on_off.add(f'D5:D{row_idx - 1}')

        # Notes at bottom
        note_row = row_idx + 1
        notes = [
            "HOW IT WORKS:",
            "  - ON: InputActivityRatio=1, OutputActivityRatio=1, Projection.Mode='User defined'",
            "  - OFF: InputActivityRatio=0, OutputActivityRatio=0, Projection.Mode='EMPTY'",
            "  - Changes are applied to A-O_AR_Model_Base_Year.xlsx and A-O_AR_Projections.xlsx",
            "  - All scenario folders are updated",
            "",
            "NOTE: Year values in Projections (efficiency/loss factors) are NOT modified by this control.",
        ]
        for note in notes:
            ws_intercon.cell(note_row, 1, note)
            ws_intercon.cell(note_row, 1).font = Font(size=9)
            ws_intercon.merge_cells(f'A{note_row}:E{note_row}')
            note_row += 1

        # Freeze panes
        ws_intercon.freeze_panes = 'D5'

        print(f"  Interconnections sheet created with {len(trn_data)} technologies ({(row_idx - 5)} rows)")
    else:
        print("  No TRN interconnections found - skipping Interconnections sheet")

    # =========================================================================
    # Create Scenarios_Demand_Growth sheet (after Demand_Growth, before Renewability_Targets)
    # =========================================================================
    # Read base_scenario from YAML to exclude it from this sheet
    base_scenario = read_base_scenario()
    scenarios_for_demand = [s for s in scenarios if s != base_scenario]

    ws_scenarios_demand = wb.create_sheet("Scenarios_Demand_Growth", 3)
    ws_scenarios_demand.column_dimensions['A'].width = 15
    ws_scenarios_demand.column_dimensions['B'].width = 15

    # OSTRAM base year is 2023 - filter years to start from 2023
    ostram_base_year = 2023
    demand_adj_years = [y for y in data['years'] if y >= ostram_base_year]

    # Title
    cell = ws_scenarios_demand.cell(1, 1, "SCENARIO-SPECIFIC DEMAND GROWTH ADJUSTMENTS")
    cell.font = Font(size=14, bold=True, color="366092")
    ws_scenarios_demand.merge_cells(f'A1:{openpyxl.utils.get_column_letter(2 + len(demand_adj_years))}1')

    # Instructions
    ws_scenarios_demand.cell(2, 1, f"Define percentage adjustments to electricity demand for each scenario (excluding base scenario '{base_scenario}').")
    ws_scenarios_demand.merge_cells(f'A2:{openpyxl.utils.get_column_letter(2 + len(demand_adj_years))}2')
    ws_scenarios_demand.cell(2, 1).font = Font(italic=True)

    ws_scenarios_demand.cell(3, 1, "Each percentage is applied INDEPENDENTLY to the base demand of that year (not cumulative).")
    ws_scenarios_demand.merge_cells(f'A3:{openpyxl.utils.get_column_letter(2 + len(demand_adj_years))}3')
    ws_scenarios_demand.cell(3, 1).font = Font(italic=True, bold=True, size=9)

    # Headers
    scenarios_demand_header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    scenarios_demand_header_font = Font(bold=True, color="FFFFFF")

    scenarios_demand_headers = ['Country', 'Scenario'] + [str(year) for year in demand_adj_years]
    for col_idx, header in enumerate(scenarios_demand_headers, 1):
        cell = ws_scenarios_demand.cell(5, col_idx, header)
        cell.fill = scenarios_demand_header_fill
        cell.font = scenarios_demand_header_font
        cell.border = border_style
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col_idx > 2:
            ws_scenarios_demand.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 8

    # Pre-populate rows for each country+scenario combination (excluding base scenario)
    row_idx = 6
    for country in countries_list:
        for scenario in scenarios_for_demand:
            ws_scenarios_demand.cell(row_idx, 1, country).border = border_style
            ws_scenarios_demand.cell(row_idx, 1).alignment = Alignment(horizontal="center")
            ws_scenarios_demand.cell(row_idx, 2, scenario).border = border_style
            ws_scenarios_demand.cell(row_idx, 2).alignment = Alignment(horizontal="center")
            # Year columns - default 0%
            for col_idx in range(3, 3 + len(demand_adj_years)):
                ws_scenarios_demand.cell(row_idx, col_idx, 0).border = border_style
                ws_scenarios_demand.cell(row_idx, col_idx).number_format = '0.00%'
                ws_scenarios_demand.cell(row_idx, col_idx).alignment = Alignment(horizontal="center")
            row_idx += 1

    # Add detailed notes at the bottom
    note_row = row_idx + 1
    ws_scenarios_demand.cell(note_row, 1, "HOW IT WORKS:")
    ws_scenarios_demand.merge_cells(f'A{note_row}:C{note_row}')
    ws_scenarios_demand.cell(note_row, 1).font = Font(bold=True, size=10)
    note_row += 1

    notes = [
        f"Base Scenario: '{base_scenario}' uses the demand growth rates from the 'Demand_Growth' sheet",
        "Other Scenarios: Apply additional percentage adjustments defined in this sheet",
        "",
        "CALCULATION METHOD (Independent, Non-Cumulative):",
        f"  Demand(scenario, year) = Base_Demand({base_scenario}, year) × (1 + adjustment_percentage)",
        "",
        "EXAMPLE:",
        f"  If {base_scenario} demand in 2030 is 100 PJ and you set +5% for NDC in 2030:",
        "    → NDC demand in 2030 = 100 × (1 + 0.05) = 105 PJ",
        "  If you set -3% for 2035:",
        f"    → NDC demand in 2035 = Base_Demand({base_scenario}, 2035) × (1 - 0.03)",
        "  Each year's percentage is independent and applies to that year's base demand only",
        "",
        "USE CASES:",
        "  - Electric vehicle adoption scenarios (increases demand)",
        "  - Energy efficiency improvements (decreases demand)",
        "  - Different economic growth assumptions",
        "",
        "NOTES:",
        "  - Positive percentages increase demand, negative percentages decrease demand",
        "  - Default is 0% (same as base scenario)",
        f"  - Base scenario '{base_scenario}' is excluded from this sheet",
    ]
    for note in notes:
        ws_scenarios_demand.cell(note_row, 1, note)
        ws_scenarios_demand.cell(note_row, 1).font = Font(size=9)
        ws_scenarios_demand.merge_cells(f'A{note_row}:{openpyxl.utils.get_column_letter(2 + len(demand_adj_years))}{note_row}')
        note_row += 1

    # Freeze panes
    ws_scenarios_demand.freeze_panes = 'C6'

    # =========================================================================
    # Create Documentation sheet for Activity Limits validation (moved after Instructions)
    # =========================================================================
    # This will be created after all other sheets, then moved to index 2 (after Instructions)
    ws_doc = wb.create_sheet("Documentation")
    ws_doc.column_dimensions['A'].width = 100
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws_doc.column_dimensions[col].width = 12

    # Title
    ws_doc['A1'] = 'ACTIVITY LIMITS CALCULATION & VALIDATION'
    ws_doc['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws_doc['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    ws_doc['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_doc.merge_cells('A1:H1')
    ws_doc.row_dimensions[1].height = 25

    # Section 1: Formula
    doc_row = 3
    ws_doc[f'A{doc_row}'] = '1. BASE FORMULA'
    ws_doc[f'A{doc_row}'].font = Font(size=12, bold=True, color='FFFFFF')
    ws_doc[f'A{doc_row}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws_doc.merge_cells(f'A{doc_row}:H{doc_row}')

    doc_row += 1
    ws_doc[f'A{doc_row}'] = 'ActivityLimit = Generation_OSTRAM × (1 + growth_rate × (year - ref_year)) × Share_technology'
    ws_doc[f'A{doc_row}'].font = Font(italic=True)
    ws_doc.merge_cells(f'A{doc_row}:H{doc_row}')

    doc_row += 2
    ws_doc[f'A{doc_row}'] = 'Where:'
    ws_doc[f'A{doc_row}'].font = Font(bold=True)
    doc_row += 1
    ws_doc[f'A{doc_row}'] = '  • Generation_OSTRAM: Total electricity generation for country in base year (PJ)'
    doc_row += 1
    ws_doc[f'A{doc_row}'] = '  • growth_rate: Annual growth rate (e.g., 0.02 = 2%)'
    doc_row += 1
    ws_doc[f'A{doc_row}'] = '  • ref_year: OSTRAM reference year (2023)'
    doc_row += 1
    ws_doc[f'A{doc_row}'] = '  • Share_technology: Technology share (from Shares_Total or Renewability_Targets)'

    doc_row += 2
    ws_doc[f'A{doc_row}'] = 'UpperLimit = LowerLimit + 0.1'
    ws_doc[f'A{doc_row}'].font = Font(italic=True)

    # Section 2: Validation 1
    doc_row += 3
    ws_doc[f'A{doc_row}'] = '2. INDIVIDUAL VALIDATION (during calculation)'
    ws_doc[f'A{doc_row}'].font = Font(size=12, bold=True, color='FFFFFF')
    ws_doc[f'A{doc_row}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws_doc.merge_cells(f'A{doc_row}:H{doc_row}')

    doc_row += 1
    ws_doc[f'A{doc_row}'] = 'Constraint: LowerLimit ≤ MaxPossibleActivity'
    ws_doc[f'A{doc_row}'].font = Font(bold=True)
    ws_doc.merge_cells(f'A{doc_row}:H{doc_row}')

    doc_row += 2
    ws_doc[f'A{doc_row}'] = 'MaxPossibleActivity = Capacity × CapacityToActivityUnit × AvailabilityFactor × CapacityFactor'
    ws_doc[f'A{doc_row}'].font = Font(italic=True)
    ws_doc.merge_cells(f'A{doc_row}:H{doc_row}')

    doc_row += 2
    ws_doc[f'A{doc_row}'] = 'Where Capacity = min(TotalAnnualMaxCapacity, ResidualCapacity)'
    ws_doc[f'A{doc_row}'].font = Font(italic=True)
    ws_doc.merge_cells(f'A{doc_row}:H{doc_row}')

    doc_row += 2
    ws_doc[f'A{doc_row}'] = 'Validation Logic:'
    ws_doc[f'A{doc_row}'].font = Font(bold=True)
    doc_row += 1
    ws_doc[f'A{doc_row}'] = '  IF MaxPossibleActivity ≤ 0:'
    doc_row += 1
    ws_doc[f'A{doc_row}'] = '    → No capacity available'
    doc_row += 1
    ws_doc[f'A{doc_row}'] = '    → LowerLimit = 0'
    doc_row += 1
    ws_doc[f'A{doc_row}'] = '    → UpperLimit = 0.1'
    doc_row += 1
    ws_doc[f'A{doc_row}'] = '    → Value is CAPPED'
    doc_row += 2
    ws_doc[f'A{doc_row}'] = '  ELSE:'
    doc_row += 1
    ws_doc[f'A{doc_row}'] = '    → MaxAllowed = MaxPossibleActivity - 0.05  (safety margin)'
    doc_row += 1
    ws_doc[f'A{doc_row}'] = '    → IF LowerLimit > MaxAllowed:'
    doc_row += 1
    ws_doc[f'A{doc_row}'] = '         → LowerLimit = max(0, MaxAllowed)'
    doc_row += 1
    ws_doc[f'A{doc_row}'] = '         → UpperLimit = LowerLimit + 0.1'
    doc_row += 1
    ws_doc[f'A{doc_row}'] = '         → Value is CAPPED'

    # Section 3: Validation 2
    doc_row += 3
    ws_doc[f'A{doc_row}'] = '3. UNIVERSAL VALIDATION (after all calculations)'
    ws_doc[f'A{doc_row}'].font = Font(size=12, bold=True, color='FFFFFF')
    ws_doc[f'A{doc_row}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws_doc.merge_cells(f'A{doc_row}:H{doc_row}')

    doc_row += 1
    ws_doc[f'A{doc_row}'] = 'Applied ONLY to these 10 PWR generation technologies:'
    ws_doc[f'A{doc_row}'].font = Font(bold=True)
    ws_doc.merge_cells(f'A{doc_row}:H{doc_row}')

    doc_row += 1
    ws_doc[f'A{doc_row}'] = '  URN (Nuclear), NGS (Natural Gas), COA (Coal), HYD (Hydro), GEO (Geothermal)'
    doc_row += 1
    ws_doc[f'A{doc_row}'] = '  WON (Wind), SPV (Solar), BIO (Biomass), PET (Petroleum/Diesel), OIL (Fuel Oil)'

    doc_row += 2
    ws_doc[f'A{doc_row}'] = 'Excludes: Storage (SDS, LDS), Backup (BCK), Transmission (TRN)'
    ws_doc[f'A{doc_row}'].font = Font(italic=True)

    doc_row += 2
    ws_doc[f'A{doc_row}'] = 'Purpose: Re-verify that NO technology has LowerLimit > MaxPossibleActivity'
    ws_doc[f'A{doc_row}'].font = Font(bold=True)
    ws_doc.merge_cells(f'A{doc_row}:H{doc_row}')

    doc_row += 1
    ws_doc[f'A{doc_row}'] = 'This catches manual errors from the Editor sheet and ensures model feasibility.'
    ws_doc.merge_cells(f'A{doc_row}:H{doc_row}')

    # Section 4: Example
    doc_row += 3
    ws_doc[f'A{doc_row}'] = '4. NUMERICAL EXAMPLE'
    ws_doc[f'A{doc_row}'].font = Font(size=12, bold=True, color='FFFFFF')
    ws_doc[f'A{doc_row}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws_doc.merge_cells(f'A{doc_row}:H{doc_row}')

    doc_row += 1
    ws_doc[f'A{doc_row}'] = 'Country: Argentina | Technology: PWRHYDARGXX (Hydro) | Year: 2030'
    ws_doc[f'A{doc_row}'].font = Font(bold=True)
    ws_doc.merge_cells(f'A{doc_row}:H{doc_row}')

    doc_row += 2
    ws_doc[f'A{doc_row}'] = 'Input Data:'
    ws_doc[f'A{doc_row}'].font = Font(bold=True)
    doc_row += 1
    ws_doc[f'A{doc_row}'] = '  • Generation_OSTRAM (2023) = 100 PJ'
    doc_row += 1
    ws_doc[f'A{doc_row}'] = '  • Growth rate = 2% (0.02)'
    doc_row += 1
    ws_doc[f'A{doc_row}'] = '  • Share_HYD = 40%'
    doc_row += 1
    ws_doc[f'A{doc_row}'] = '  • ResidualCapacity = 2.5 GW'
    doc_row += 1
    ws_doc[f'A{doc_row}'] = '  • TotalAnnualMaxCapacity = 3.0 GW'
    doc_row += 1
    ws_doc[f'A{doc_row}'] = '  • CapacityToActivityUnit = 31.536'
    doc_row += 1
    ws_doc[f'A{doc_row}'] = '  • AvailabilityFactor = 0.95'
    doc_row += 1
    ws_doc[f'A{doc_row}'] = '  • CapacityFactor = 0.28'

    doc_row += 2
    ws_doc[f'A{doc_row}'] = 'Step 1: Calculate base LowerLimit'
    ws_doc[f'A{doc_row}'].font = Font(bold=True, underline='single')
    doc_row += 1
    ws_doc[f'A{doc_row}'] = '  years_diff = 2030 - 2023 = 7'
    doc_row += 1
    ws_doc[f'A{doc_row}'] = '  Generation_2030 = 100 × (1 + 0.02 × 7) = 114 PJ'
    doc_row += 1
    ws_doc[f'A{doc_row}'] = '  LowerLimit = 114 × 0.40 = 45.6 PJ'

    doc_row += 2
    ws_doc[f'A{doc_row}'] = 'Step 2: Validate against capacity'
    ws_doc[f'A{doc_row}'].font = Font(bold=True, underline='single')
    doc_row += 1
    ws_doc[f'A{doc_row}'] = '  Capacity = min(2.5, 3.0) = 2.5 GW  (ResidualCapacity is more restrictive)'
    doc_row += 1
    ws_doc[f'A{doc_row}'] = '  MaxPossibleActivity = 2.5 × 31.536 × 0.95 × 0.28 = 20.95 PJ'
    doc_row += 1
    ws_doc[f'A{doc_row}'] = '  MaxAllowed = 20.95 - 0.05 = 20.90 PJ'

    doc_row += 2
    ws_doc[f'A{doc_row}'] = 'Step 3: Apply validation'
    ws_doc[f'A{doc_row}'].font = Font(bold=True, underline='single')
    doc_row += 1
    ws_doc[f'A{doc_row}'] = '  LowerLimit (45.6) > MaxAllowed (20.90)  →  CAPPED!'
    ws_doc[f'A{doc_row}'].font = Font(color='FF0000', bold=True)
    doc_row += 1
    ws_doc[f'A{doc_row}'] = '  Final LowerLimit = 20.90 PJ'
    ws_doc[f'A{doc_row}'].font = Font(color='00B050', bold=True)
    doc_row += 1
    ws_doc[f'A{doc_row}'] = '  Final UpperLimit = 20.90 + 0.1 = 21.00 PJ'
    ws_doc[f'A{doc_row}'].font = Font(color='00B050', bold=True)

    doc_row += 2
    ws_doc[f'A{doc_row}'] = 'Note: This indicates additional capacity investment is needed to meet the demand target.'
    ws_doc[f'A{doc_row}'].font = Font(italic=True)
    ws_doc[f'A{doc_row}'].fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    ws_doc.merge_cells(f'A{doc_row}:H{doc_row}')

    # Reorder sheets: Documentation should come after Instructions (index 1)
    # Current order after creation: Instructions(0), OSTRAM_Config(1), Demand_Growth(2),
    # Scenarios_Demand_Growth(3), Renewability_Targets(4), Technology_Weights(5),
    # Editor, _hidden sheets, Documentation(last)
    # Target order: Instructions(0), Documentation(1), OSTRAM_Config(2), ...
    doc_sheet_index = wb.sheetnames.index('Documentation')
    wb.move_sheet('Documentation', offset=-(doc_sheet_index - 1))

    # Save the workbook
    wb.save(output_path)
    print(f"Template saved: {output_path}")
    print()


def main():
    try:
        print("=" * 80)
        print("SECONDARY TECHS EDITOR - TEMPLATE GENERATOR")
        print("=" * 80)
        print()

        # Collect data from all scenarios
        data = collect_data_from_all_scenarios()

        if not data['tech_mapping'] or not data['parameters'] or not data['years']:
            print("ERROR: Could not collect enough data from scenario files")
            print("Please check that A-O_Parametrization.xlsx files exist and contain Secondary Techs sheet")
            return 1

        # Create template
        output_path = Path(__file__).parent / "Secondary_Techs_Editor.xlsx"
        create_editor_template(data, output_path)

        print("=" * 80)
        print("TEMPLATE GENERATION COMPLETE")
        print("=" * 80)
        print()
        # print("Next steps:")
        # print("1. Open t1_confection/Secondary_Techs_Editor.xlsx")
        # print("2. Fill in the 'Editor' sheet with your changes")
        # print("3. Save and close the file")
        # print("4. Run: python t1_confection/update_secondary_techs.py")
        print()

        return 0

    except Exception as e:
        print(f"\nERROR: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
