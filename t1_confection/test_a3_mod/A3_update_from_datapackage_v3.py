# -*- coding: utf-8 -*-
"""
A3_update_from_datapackage_v3.py
Reads a single DataPackage V3 Excel workbook and produces A-O files for B1.

Two modes (set in Config_datapackage_v3.yaml → compute.*):
  - compute=true  : derive values from green inventory sheets
  - compute=false : passthrough blue A-O sheets as-is

Run from t1_confection/:
    python A3_update_from_datapackage_v3.py

Author: Climate Lead Group
"""

import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import yaml
from openpyxl.utils.dataframe import dataframe_to_rows

from Z_AUX_config_loader import (
    get_code_to_energy,
    get_countries,
    get_enable_dsptrn,
    get_iso_country_map,
)

SCRIPT_DIR = Path(__file__).resolve().parent

# ============================================================
# 1. CONFIGURATION
# ============================================================

def load_config():
    """Load Config_datapackage_v3.yaml."""
    path = SCRIPT_DIR / "Config_datapackage_v3.yaml"
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


CFG = load_config()
XLSX_PATH = SCRIPT_DIR / CFG["datapackage_v3_path"]
FIRST_YEAR = CFG["first_year"]
FINAL_YEAR = CFG["final_year"]
TARGET_YEARS = list(range(FIRST_YEAR, FINAL_YEAR + 1))
COST_SCENARIO = CFG["cost_scenario"]
MILESTONE_YEARS = CFG["milestone_years"]
MILESTONE_YEAR_STRS = [str(y) for y in MILESTONE_YEARS]
TIMESLICES = CFG["timeslices"]
RNW_TECH_CODES = set(CFG.get("rnw_tech_codes", []))
FUEL_TO_TECH = CFG.get("fuel_to_tech", {})
POWER_FUEL_MAP = CFG.get("power_fuel_type_to_tech", {})
AGGREGATED_TECHS = CFG.get("aggregated_techs", {})
AGG_RULES = CFG.get("aggregation_rules", {})

# Lazy-loaded lookups
_CODE_TO_ENERGY = None
_ISO_COUNTRY_MAP = None


def _get_code_to_energy():
    global _CODE_TO_ENERGY
    if _CODE_TO_ENERGY is None:
        _CODE_TO_ENERGY = get_code_to_energy()
    return _CODE_TO_ENERGY


def _get_iso_country_map():
    global _ISO_COUNTRY_MAP
    if _ISO_COUNTRY_MAP is None:
        _ISO_COUNTRY_MAP = get_iso_country_map()
    return _ISO_COUNTRY_MAP


# ============================================================
# 2. TEMPLATE READER
# ============================================================

def read_template(xlsx_path):
    """Read all relevant sheets from the DataPackage V3 workbook.

    Returns dict: {sheet_name: pd.DataFrame}
    """
    sheets = {}
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    for sname in wb.sheetnames:
        ws = wb[sname]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        if not rows:
            sheets[sname] = pd.DataFrame()
            continue
        header = list(rows[0])
        data = [list(r) for r in rows[1:] if any(c is not None for c in r)]
        sheets[sname] = pd.DataFrame(data, columns=header)

    wb.close()
    return sheets


def read_node_mapping(sheets):
    """Build node-mapping lookup dicts from the Node_Mapping sheet.

    Returns:
        short_to_ostram: {"BD": "BGDXX", "IN_N": "INDNO", ...}
        short_to_iso3:   {"BD": "BGD", "IN_N": "IND", ...}
        iso3_to_regions: {"BGD": ["BGDXX"], "IND": ["INDNO", "INDSO", ...], ...}
    """
    df = sheets["Node_Mapping"]
    short_to_ostram = dict(zip(df["Short_Node"], df["OSTRAM_Region"]))
    short_to_iso3 = dict(zip(df["Short_Node"], df["ISO3"]))

    iso3_to_regions = {}
    for _, row in df.iterrows():
        iso3 = row["ISO3"]
        region = row["OSTRAM_Region"]
        iso3_to_regions.setdefault(iso3, []).append(region)

    return short_to_ostram, short_to_iso3, iso3_to_regions


def read_fuel_node_mapping(sheets):
    """Build fuel Node_Override → list of OSTRAM regions from Node_Mapping.

    E.g. "BD,LK" → ["BGDXX", "LKAXX"]
         "IN_all" → ["INDNO", "INDSO", "INDEA", "INDWE", "INDNE"]
         "All"    → all regions
    """
    short_to_ostram, _, iso3_to_regions = read_node_mapping(sheets)
    all_regions = sorted(short_to_ostram.values())

    def resolve(node_override):
        node_override = str(node_override).strip()
        if node_override == "All":
            return all_regions
        if node_override == "IN_all":
            return iso3_to_regions.get("IND", [])
        if node_override == "IN":
            return iso3_to_regions.get("IND", [])
        # Comma-separated short nodes
        parts = [p.strip() for p in node_override.split(",")]
        result = []
        for p in parts:
            if p in short_to_ostram:
                result.append(short_to_ostram[p])
            # Also handle "IN,BD,LK" where IN expands to all India regions
            elif p == "IN":
                result.extend(iso3_to_regions.get("IND", []))
        return result

    return resolve


# ============================================================
# 3. NAMING UTILITIES (from current A3)
# ============================================================

def parse_tech_name(tech):
    """Return a descriptive name for a technology code."""
    code_to_energy = _get_code_to_energy()
    iso_map = _get_iso_country_map()
    main_code = tech[0:3]

    if main_code == "TRN" and len(tech) >= 13:
        iso1, r1 = tech[3:6], tech[6:8]
        iso2, r2 = tech[8:11], tech[11:13]
        c1 = iso_map.get(iso1, f"Unknown ({iso1})")
        c2 = iso_map.get(iso2, f"Unknown ({iso2})")
        return f"Transmission interconnection from {c1}, region {r1} to {c2}, region {r2}"

    if main_code in ("SDS", "LDS") and len(tech) <= 10:
        iso1, r1 = tech[3:6], tech[6:8]
        desc = code_to_energy.get(main_code, "specific technology")
        c1 = iso_map.get(iso1, f"Unknown ({iso1})")
        return f"{desc} {c1}, region {r1}"

    iso = tech[6:9]
    region = tech[9:11] if len(tech) > 9 else "XX"
    country = iso_map.get(iso, f"Unknown ({iso})")
    sub_code = tech[3:6]
    main_desc = code_to_energy.get(main_code, "General technology")
    sub_desc = code_to_energy.get(sub_code, "specific technology")

    base = f"{sub_desc} ({main_desc})" if main_desc != sub_desc else sub_desc
    name = f"{base} {country}"
    if not tech.startswith("MIN") and region != "XX":
        name += f", region {region}"
    elif region == "XX":
        name += f", region XX"
    return name


def parse_fuel_name(fuel):
    """Return a descriptive name for a fuel code."""
    code_to_energy = _get_code_to_energy()
    iso_map = _get_iso_country_map()
    prefix = fuel[0:3]
    iso = fuel[3:6]
    region = fuel[6:8] if len(fuel) >= 8 else None
    suffix = None
    if fuel.endswith("01"):
        suffix = "power plant output"
    elif fuel.endswith("02"):
        suffix = "transmission line output"
    elif fuel.endswith("03"):
        suffix = "dispatch output"

    fuel_type = code_to_energy.get(prefix, "Unknown")
    country = iso_map.get(iso, f"Unknown ({iso})")
    parts = [fuel_type, country]
    if region and region != "XX":
        parts.append(f"region {region}")
    elif region == "XX":
        parts.append("region XX")
    if suffix:
        parts.append(suffix)
    return ", ".join(parts)


def assign_tech_type(tech):
    """Classify technology: Primary / Secondary / Demand."""
    if tech.startswith("MIN") or tech.startswith("RNW"):
        return "Primary"
    elif tech.startswith("PWRTRN"):
        return "Demand"
    return "Secondary"


# ============================================================
# 4. INTERPOLATION / EXTRAPOLATION (unchanged from current A3)
# ============================================================

def interpolate_linear(known_points, target_years):
    """Linear interpolation. Holds flat before first point. No extrapolation."""
    if not known_points:
        return {}
    sorted_years = sorted(known_points.keys())
    first_year, last_year = sorted_years[0], sorted_years[-1]
    result = {}
    for y in target_years:
        if y in known_points:
            result[y] = known_points[y]
        elif y < first_year:
            result[y] = known_points[first_year]
        elif y > last_year:
            continue
        else:
            lo = max(yr for yr in sorted_years if yr <= y)
            hi = min(yr for yr in sorted_years if yr >= y)
            if lo == hi:
                result[y] = known_points[lo]
            else:
                frac = (y - lo) / (hi - lo)
                result[y] = known_points[lo] + frac * (known_points[hi] - known_points[lo])
    return result


def fill_annual_values(known_points, target_years, extrapolation_rate=None):
    """Interpolate + optional compound-growth extrapolation."""
    result = interpolate_linear(known_points, target_years)
    if extrapolation_rate is not None and known_points:
        last_year = max(known_points.keys())
        last_value = known_points[last_year]
        for y in sorted(target_years):
            if y > last_year and y not in result:
                dt = y - last_year
                result[y] = last_value * (1 + extrapolation_rate) ** dt
    return result


# ============================================================
# 5. COMPUTE FUNCTIONS (green sheets → long-format DataFrames)
# ============================================================

def _get_milestone_values_from_row(row, milestone_years):
    """Extract {year: float} from a row dict keyed by milestone years."""
    result = {}
    for y in milestone_years:
        val = row.get(y) if y in row else row.get(str(y))
        if val is not None and val != "" and not (isinstance(val, str) and val.strip() == ""):
            try:
                result[int(y)] = float(val)
            except (ValueError, TypeError):
                pass
    return result


def _aggregate_technologies(df, param_name):
    """Aggregate N:1 tech mappings using AVG or SUM rule."""
    if df.empty:
        return df
    if param_name in AGG_RULES.get("AVG", []):
        agg_func = "mean"
    elif param_name in AGG_RULES.get("SUM", []):
        agg_func = "sum"
    else:
        agg_func = "mean"

    group_cols = [c for c in df.columns if c != "VALUE"]
    if not group_cols:
        return df
    df = df.groupby(group_cols, as_index=False)["VALUE"].agg(agg_func)
    df["VALUE"] = df["VALUE"].round(4)
    return df


def build_fuel_tech_name(tech_code, region):
    """Build MIN/RNW technology name from fuel code + region."""
    if tech_code in RNW_TECH_CODES:
        return f"RNW{tech_code}{region}"
    return f"MIN{tech_code}{region[:3]}"


def compute_costs(sheets, param_filter, param_name):
    """Compute CapitalCost or FixedCost from Technology_Costs + Country_Multipliers.

    param_filter: "CAPEX" or "Fixed O&M"
    param_name: "CapitalCost" or "FixedCost"
    """
    tech_costs = sheets["Technology_Costs"]
    multipliers_df = sheets["Country_Multipliers"]
    _, short_to_iso3, iso3_to_regions = read_node_mapping(sheets)

    # Build multipliers: {tech_code: {short_node: mult}}
    mult_map = {}
    for _, mrow in multipliers_df.iterrows():
        tc = mrow["Technology_Code"]
        mults = {}
        for node in ["BD", "BT", "IN", "LK", "MV", "NP"]:
            v = mrow.get(node)
            if v is not None and v != "":
                try:
                    mults[node] = float(v)
                except (ValueError, TypeError):
                    pass
        mult_map[tc] = mults

    # Build reverse map: OSTRAM tech code → list of Cost DB codes
    # (for aggregated techs like SPV = [PWRSPV, PWRSPR, PWRFPV])
    # The Technology_Costs sheet uses OSTRAM codes directly (e.g., PWRSPV)

    rows = []
    cost_rows = tech_costs[tech_costs["Parameter"] == param_filter]
    if cost_rows.empty:
        return pd.DataFrame()

    # All short nodes that exist
    short_nodes = list(short_to_iso3.keys())

    for _, crow in cost_rows.iterrows():
        tech_code = crow["Technology_Code"]  # e.g., "PWRSPV"
        milestone_vals = _get_milestone_values_from_row(crow, MILESTONE_YEARS)
        if not milestone_vals:
            continue

        annual_vals = fill_annual_values(milestone_vals, TARGET_YEARS)
        tech_mults = mult_map.get(tech_code, {})

        # For each node, look up multiplier and expand to regions
        for short_node in short_nodes:
            # Determine multiplier
            if short_node.startswith("IN"):
                mult = tech_mults.get("IN", 1.0)
            else:
                mult = tech_mults.get(short_node)
                if mult is None:
                    mult = tech_mults.get("IN", 1.0)

            iso3 = short_to_iso3[short_node]
            # Get OSTRAM region for this node
            node_to_ostram, _, _ = read_node_mapping(sheets)
            region = node_to_ostram[short_node]

            tech_full = f"{tech_code}{region}"
            # Fix: tech_code already has PWR prefix from template
            # Template uses codes like PWRSPV, PWRCOASCPC etc.
            # But region in template is like BDXX01, not BGDXX
            # Actually the template Secondary_Techs uses codes like PWRSPVBDXX01
            # Let's match that convention
            # For now, generate the standard OSTRAM code: PWR{3char}{region}
            # Extract the sub-code from template tech code
            # PWRSPV → SPV, PWRCOASCPC → COASCPC, etc.
            # Actually, keep the template convention: {tech_code}{node}{XX}{01}
            # But this depends on how B1 reads it...
            # Looking at the existing template data: PWRSPVBDXX01, PWRCOASCPCIN_NXX01
            # This is: {TechCode}{ShortNode}XX01
            # Let's use that convention
            tech_full = f"{tech_code}{short_node}XX01" if short_node.startswith("IN") else f"{tech_code}{short_node}XX01"
            # Simplify: for non-India single-region, it's {tech_code}{short_node}XX01
            # For India regions: PWRSPVIN_NXX01

            for y, v in annual_vals.items():
                rows.append({
                    "REGION": region,
                    "TECHNOLOGY": tech_full,
                    "YEAR": y,
                    "VALUE": round(v * mult, 2),
                })

    new_df = pd.DataFrame(rows)
    if not new_df.empty:
        new_df = _aggregate_technologies(new_df, param_name)
    return new_df


def compute_variable_cost(sheets):
    """Compute VariableCost from Technology_Costs (VOM) + Fuel_Costs."""
    tech_costs = sheets["Technology_Costs"]
    fuel_costs_df = sheets["Fuel_Costs"]
    resolve_nodes = read_fuel_node_mapping(sheets)
    _, short_to_iso3, _ = read_node_mapping(sheets)

    rows = []

    # --- Part 1: Variable O&M for PWR technologies ---
    vom_rows = tech_costs[tech_costs["Parameter"] == "Variable O&M"]
    for _, vrow in vom_rows.iterrows():
        tech_code = vrow["Technology_Code"]
        milestone_vals = _get_milestone_values_from_row(vrow, MILESTONE_YEARS)
        if not milestone_vals or all(v == 0 for v in milestone_vals.values()):
            continue

        annual_vals = fill_annual_values(milestone_vals, TARGET_YEARS)
        # VOM applies to all nodes where this tech exists
        node_mapping, _, _ = read_node_mapping(sheets)
        for short_node, region in node_mapping.items():
            tech_full = f"{tech_code}{short_node}XX01"
            for y, v in annual_vals.items():
                # USD/MWh → M$/PJ: × 0.277778
                v_mpj = v * 0.277778
                rows.append({
                    "REGION": region,
                    "TECHNOLOGY": tech_full,
                    "MODE_OF_OPERATION": 1,
                    "YEAR": y,
                    "VALUE": round(v_mpj, 4),
                })

    # --- Part 2: Fuel costs for MIN/RNW technologies ---
    fc_rows = fuel_costs_df[
        fuel_costs_df["Scenario"].isin([COST_SCENARIO, "All"])
    ]
    for _, frow in fc_rows.iterrows():
        fuel_name = frow["Fuel"]
        tech_code = FUEL_TO_TECH.get(fuel_name)
        if tech_code is None:
            continue

        node_override = str(frow.get("Node_Override", "All") or "All").strip()
        target_regions = resolve_nodes(node_override)
        if not target_regions:
            continue

        milestone_vals = _get_milestone_values_from_row(frow, MILESTONE_YEARS)
        if not milestone_vals:
            continue

        annual_vals = fill_annual_values(milestone_vals, TARGET_YEARS)

        for region in target_regions:
            tech_full = build_fuel_tech_name(tech_code, region)
            for y, v in annual_vals.items():
                rows.append({
                    "REGION": region,
                    "TECHNOLOGY": tech_full,
                    "MODE_OF_OPERATION": 1,
                    "YEAR": y,
                    "VALUE": round(v, 4),
                })

    new_df = pd.DataFrame(rows)
    if not new_df.empty:
        new_df = _aggregate_technologies(new_df, "VariableCost")
    return new_df


def compute_operational_life(sheets):
    """Compute OperationalLife from Technology_Costs."""
    tech_costs = sheets["Technology_Costs"]
    node_mapping, _, _ = read_node_mapping(sheets)

    rows = []
    ol_rows = tech_costs[tech_costs["Parameter"] == "Operational Life"]

    for _, orow in ol_rows.iterrows():
        tech_code = orow["Technology_Code"]
        milestone_vals = _get_milestone_values_from_row(orow, MILESTONE_YEARS)
        if not milestone_vals:
            continue
        op_life = list(milestone_vals.values())[0]

        for short_node, region in node_mapping.items():
            tech_full = f"{tech_code}{short_node}XX01"
            rows.append({
                "REGION": region,
                "TECHNOLOGY": tech_full,
                "VALUE": op_life,
            })

    new_df = pd.DataFrame(rows)
    if not new_df.empty:
        new_df = _aggregate_technologies(new_df, "OperationalLife")
    return new_df


def _get_op_life_map(sheets):
    """Build {tech_code: op_life_years} from Technology_Costs."""
    tech_costs = sheets["Technology_Costs"]
    ol_rows = tech_costs[tech_costs["Parameter"] == "Operational Life"]
    result = {}
    for _, orow in ol_rows.iterrows():
        tc = orow["Technology_Code"]
        vals = _get_milestone_values_from_row(orow, MILESTONE_YEARS)
        if vals and tc not in result:
            result[tc] = list(vals.values())[0]
    return result


def compute_residual_capacity(sheets):
    """Compute ResidualCapacity from Existing_Generation + Node_Mapping."""
    eg_df = sheets["Existing_Generation"]
    node_mapping, _, _ = read_node_mapping(sheets)
    op_life_map = _get_op_life_map(sheets)

    rows = []
    for _, plant in eg_df.iterrows():
        tech_code = plant.get("Technology_Code")
        if tech_code is None or pd.isna(tech_code):
            continue
        tech_code = str(tech_code).strip()

        cap_mw = plant.get("Capacity_MW")
        if cap_mw is None or pd.isna(cap_mw):
            continue
        try:
            cap_mw = float(cap_mw)
        except (ValueError, TypeError):
            continue
        if cap_mw <= 0:
            continue

        # Determine region from Node column
        node = str(plant.get("Node", "")).strip()
        region = node_mapping.get(node)
        if region is None:
            continue

        # Build full tech name matching template convention
        tech_full = f"{tech_code}"
        # The Existing_Generation already has the full tech code (e.g., PWRNGABD)
        # We just need it as-is since it matches the template convention

        # Commissioning and retirement
        comm_year = None
        cv = plant.get("Commissioning_Year")
        if cv is not None and not pd.isna(cv):
            match = re.search(r"(\d{4})", str(cv))
            if match:
                comm_year = int(match.group(1))

        retire_year_explicit = None
        rv = plant.get("Expected_Retirement_Year")
        if rv is not None and not pd.isna(rv):
            match = re.search(r"(\d{4})", str(rv))
            if match:
                retire_year_explicit = int(match.group(1))

        ol = op_life_map.get(tech_code, 30)
        cap_gw = cap_mw / 1000.0

        if retire_year_explicit and retire_year_explicit > 1900:
            retire_year = retire_year_explicit
        elif comm_year and comm_year > 1900:
            retire_year = comm_year + int(ol)
        else:
            retire_year = None

        if retire_year:
            for y in TARGET_YEARS:
                if y < retire_year:
                    rows.append({
                        "REGION": region,
                        "TECHNOLOGY": tech_full,
                        "YEAR": y,
                        "VALUE": cap_gw,
                    })
        else:
            for y in TARGET_YEARS:
                fraction = max(0.0, 1.0 - (y - FIRST_YEAR) / ol)
                if fraction > 0:
                    rows.append({
                        "REGION": region,
                        "TECHNOLOGY": tech_full,
                        "YEAR": y,
                        "VALUE": cap_gw * fraction,
                    })

    new_df = pd.DataFrame(rows)
    if new_df.empty:
        return new_df

    # Sum by (REGION, TECHNOLOGY, YEAR) — multiple plants aggregate
    new_df = new_df.groupby(["REGION", "TECHNOLOGY", "YEAR"], as_index=False)["VALUE"].sum()
    new_df["VALUE"] = new_df["VALUE"].round(4)
    return new_df


def compute_planned_capacity(sheets):
    """Compute TotalAnnualMinCapacity from Planned_Generation."""
    pg_df = sheets.get("Planned_Generation", pd.DataFrame())
    if pg_df.empty:
        return pd.DataFrame()

    node_mapping, _, _ = read_node_mapping(sheets)
    rows = []

    for _, proj in pg_df.iterrows():
        tech_code = proj.get("Technology_Code")
        if tech_code is None or pd.isna(tech_code):
            continue
        tech_code = str(tech_code).strip()

        cap_mw = proj.get("Capacity_MW")
        if cap_mw is None or pd.isna(cap_mw):
            continue
        try:
            cap_mw = float(cap_mw)
        except (ValueError, TypeError):
            continue
        if cap_mw <= 0:
            continue

        node = str(proj.get("Node", "")).strip()
        region = node_mapping.get(node)
        if region is None:
            continue

        cod = proj.get("Expected_COD")
        if cod is None or pd.isna(cod):
            continue
        try:
            cod_year = int(float(str(cod)))
        except (ValueError, TypeError):
            continue

        cap_gw = cap_mw / 1000.0
        tech_full = str(tech_code)

        # From COD year onward, this capacity is committed
        for y in TARGET_YEARS:
            if y >= cod_year:
                rows.append({
                    "REGION": region,
                    "TECHNOLOGY": tech_full,
                    "YEAR": y,
                    "VALUE": cap_gw,
                })

    new_df = pd.DataFrame(rows)
    if new_df.empty:
        return new_df

    new_df = new_df.groupby(["REGION", "TECHNOLOGY", "YEAR"], as_index=False)["VALUE"].sum()
    new_df["VALUE"] = new_df["VALUE"].round(4)
    return new_df


def compute_demand(sheets):
    """Compute SpecifiedAnnualDemand from Demand_Assumptions + Node_Mapping."""
    da_df = sheets.get("Demand_Assumptions", pd.DataFrame())
    if da_df.empty:
        return pd.DataFrame()

    node_mapping, _, _ = read_node_mapping(sheets)
    rows = []

    for _, drow in da_df.iterrows():
        fuel_code = drow.get("Fuel_Code")
        if fuel_code is None or pd.isna(fuel_code):
            continue
        fuel_code = str(fuel_code).strip()

        node = str(drow.get("Country_Node", "")).strip()
        region = node_mapping.get(node)
        if region is None:
            continue

        extrap_rate = drow.get("Extrap_Rate")
        if extrap_rate is not None and not pd.isna(extrap_rate):
            try:
                extrap_rate = float(extrap_rate)
            except (ValueError, TypeError):
                extrap_rate = None

        milestone_vals = _get_milestone_values_from_row(drow, MILESTONE_YEARS)
        if not milestone_vals:
            continue

        annual_vals = fill_annual_values(milestone_vals, TARGET_YEARS, extrap_rate)
        for y, v in annual_vals.items():
            rows.append({
                "REGION": region,
                "FUEL": fuel_code,
                "YEAR": y,
                "VALUE": round(v, 4),
            })

    return pd.DataFrame(rows)


def _infer_emission_from_unit(unit):
    """Infer emission species from the Unit string (e.g. 'tCO2/PJ' -> 'CO2').

    Recognizes CO2, CH4, N2O, NOx, SOx, SO2. Falls back to 'CO2' if no match.
    TODO: replace with an explicit 'Emission' column on the SoT Emissions sheet.
    """
    if unit is None or (isinstance(unit, float) and pd.isna(unit)):
        return "CO2"
    s = str(unit).upper()
    for candidate in ("CO2", "CH4", "N2O", "NOX", "SOX", "SO2"):
        if candidate in s:
            return candidate
    return "CO2"


def build_ghgs_from_emissions(sheets):
    """Transform the wide Emissions sheet into the GHGs sheet B1 expects.

    Input (SoT Emissions, wide): Tech, Parameter, Unit, <years...>, plus metadata.
    Output columns required by B1 (see A1_Pre_processing_OG_csvs.update_xtra_emissions_ghg):
        Mode_Of_Operation, Tech, Emission, EmissionActivityRatio, Unit

    Behaviour:
    - Only rows with Parameter == 'EmissionActivityRatio' are kept.
    - A single scalar per row is chosen (first non-null across year columns),
      matching A1's `.first()` aggregation.
    - Emission species is inferred from Unit (e.g. 'tCO2/PJ' -> 'CO2').
    - Mode_Of_Operation defaults to 1 (add an explicit column to the SoT to override).
    """
    em = sheets.get("Emissions", pd.DataFrame())
    if em.empty or "Parameter" not in em.columns:
        return pd.DataFrame(columns=["Mode_Of_Operation", "Tech", "Emission",
                                     "EmissionActivityRatio", "Unit"])

    ear = em[em["Parameter"] == "EmissionActivityRatio"].copy()
    if ear.empty:
        return pd.DataFrame(columns=["Mode_Of_Operation", "Tech", "Emission",
                                     "EmissionActivityRatio", "Unit"])

    year_cols = [c for c in ear.columns if isinstance(c, int)]

    out_rows = []
    for _, r in ear.iterrows():
        tech = r.get("Tech")
        if tech is None or (isinstance(tech, float) and pd.isna(tech)):
            continue
        val = None
        for y in year_cols:
            v = r.get(y)
            if v is not None and not (isinstance(v, float) and pd.isna(v)):
                val = v
                break
        if val is None:
            continue
        unit = r.get("Unit")
        out_rows.append({
            "Mode_Of_Operation": r.get("Mode_Of_Operation", 1) or 1,
            "Tech": tech,
            "Emission": _infer_emission_from_unit(unit),
            "EmissionActivityRatio": val,
            "Unit": "MT" if unit is None else unit,
        })

    return pd.DataFrame(out_rows, columns=["Mode_Of_Operation", "Tech", "Emission",
                                           "EmissionActivityRatio", "Unit"])


# ============================================================
# 6. FORMATTERS (long-format → wide-format for A-O Excel sheets)
# ============================================================

def format_param_yearly(long_df, param_name, param_id, all_years):
    """Format CapitalCost/FixedCost/ResidualCapacity → wide rows.

    Output: Tech.ID, Tech, Tech.Name, Parameter.ID, Parameter, Unit,
            Projection.Mode, Projection.Parameter, [years...]
    """
    if long_df.empty:
        return pd.DataFrame()

    records = []
    for tech in long_df["TECHNOLOGY"].unique():
        sub = long_df[long_df["TECHNOLOGY"] == tech]
        row = {
            "Tech.ID": 0,
            "Tech": tech,
            "Tech.Name": parse_tech_name(tech),
            "Parameter.ID": param_id,
            "Parameter": param_name,
            "Unit": "",
            "Projection.Mode": "User defined",
            "Projection.Parameter": 0,
        }
        year_vals = dict(zip(sub["YEAR"], sub["VALUE"]))
        for y in all_years:
            row[y] = year_vals.get(y, np.nan)
        records.append(row)

    return pd.DataFrame(records)


def format_fixed_horizon(long_df):
    """Format OperationalLife → Fixed Horizon Parameters wide."""
    if long_df.empty:
        return pd.DataFrame()

    records = []
    for tech in long_df["TECHNOLOGY"].unique():
        sub = long_df[long_df["TECHNOLOGY"] == tech]
        val = sub["VALUE"].iloc[0]
        records.append({
            "Tech.Type": assign_tech_type(tech),
            "Tech.ID": 0,
            "Tech": tech,
            "Tech.Name": parse_tech_name(tech),
            "Parameter.ID": 2,
            "Parameter": "OperationalLife",
            "Unit": "",
            "Value": val,
        })
    return pd.DataFrame(records)


def format_variable_cost(long_df, all_years):
    """Format VariableCost → wide."""
    if long_df.empty:
        return pd.DataFrame()

    groups = long_df.groupby(["TECHNOLOGY", "MODE_OF_OPERATION"])
    records = []
    for (tech, mode), sub in groups:
        row = {
            "Mode.Operation": int(mode),
            "Tech.ID": 0,
            "Tech": tech,
            "Tech.Name": parse_tech_name(tech),
            "Parameter.ID": 12,
            "Parameter": "VariableCost",
            "Unit": "",
            "Projection.Mode": "User defined",
            "Projection.Parameter": 0,
        }
        year_vals = dict(zip(sub["YEAR"], sub["VALUE"]))
        for y in all_years:
            row[y] = year_vals.get(y, np.nan)
        records.append(row)

    return pd.DataFrame(records)


def _rewrite_fuel_dsptrn(fuel_code):
    """If DSPTRN enabled, rewrite ELC*02 → ELC*03."""
    if fuel_code.startswith("ELC") and fuel_code.endswith("02"):
        return fuel_code[:-2] + "03"
    return fuel_code


def format_demand_projection(long_df, enable_dsptrn, all_years):
    """Format SpecifiedAnnualDemand → Demand_Projection wide."""
    if long_df.empty:
        return pd.DataFrame()

    records = []
    for fuel in long_df["FUEL"].unique():
        sub = long_df[long_df["FUEL"] == fuel]
        fuel_out = _rewrite_fuel_dsptrn(fuel) if enable_dsptrn else fuel
        row = {
            "Demand/Share": "Demand",
            "Fuel/Tech": fuel_out,
            "Name": parse_fuel_name(fuel_out),
            "Ref.Cap.BY": "not needed",
            "Ref.OAR.BY": "not needed",
            "Ref.km.BY": "not needed",
            "Projection.Mode": "User defined",
            "Projection.Parameter": 0,
        }
        year_vals = dict(zip(sub["YEAR"], sub["VALUE"]))
        for y in all_years:
            row[y] = year_vals.get(y, np.nan)
        records.append(row)

    return pd.DataFrame(records)


# ============================================================
# 7. MERGE HELPERS
# ============================================================

def merge_wide_by_tech_param(existing_df, new_df, tech_col="Tech", param_col="Parameter",
                             clear_on_replace=False):
    """Merge by (tech, param) key.

    clear_on_replace=False (default): replace matching (tech, param) rows,
        preserve all other existing rows.
    clear_on_replace=True: new_df is authoritative — discard existing_df
        entirely and return new_df. SoT defines what exists.
    """
    if new_df.empty:
        return existing_df
    if clear_on_replace:
        return new_df.reset_index(drop=True)
    if existing_df.empty:
        return new_df
    new_keys = set(zip(new_df[tech_col], new_df[param_col]))
    preserved = existing_df[~existing_df.apply(
        lambda r: (r[tech_col], r[param_col]) in new_keys, axis=1)]
    return pd.concat([preserved, new_df], ignore_index=True)


def merge_wide_by_tech(existing_df, new_df, tech_col="Tech", clear_on_replace=False):
    """Merge by tech key.

    clear_on_replace=False (default): replace matching-tech rows, preserve rest.
    clear_on_replace=True: discard existing_df, return new_df as authoritative.
    """
    if new_df.empty:
        return existing_df
    if clear_on_replace:
        return new_df.reset_index(drop=True)
    if existing_df.empty:
        return new_df
    new_techs = set(new_df[tech_col])
    preserved = existing_df[~existing_df[tech_col].isin(new_techs)]
    return pd.concat([preserved, new_df], ignore_index=True)


def merge_wide_by_fuel(existing_df, new_df, fuel_col="Fuel/Tech", clear_on_replace=False):
    """Merge by fuel code.

    clear_on_replace=False (default): replace matching-fuel rows, preserve rest.
    clear_on_replace=True: discard existing_df, return new_df as authoritative.
    """
    if new_df.empty:
        return existing_df
    if clear_on_replace:
        return new_df.reset_index(drop=True)
    if existing_df.empty:
        return new_df
    new_fuels = set(new_df[fuel_col])
    preserved = existing_df[~existing_df[fuel_col].isin(new_fuels)]
    return pd.concat([preserved, new_df], ignore_index=True)


# ============================================================
# 7b. SCENARIO RULES ENGINE
# ============================================================

def _tech_filter_matches(filter_str, tech):
    """Match a Tech_Filter against a tech code.

    Supports exact match, '*' wildcard, and trailing-'*' prefix match.
    """
    if filter_str is None or tech is None:
        return False
    f = str(filter_str).strip()
    t = str(tech)
    if f == "*":
        return True
    if f.endswith("*"):
        return t.startswith(f[:-1])
    return t == f


def _apply_rule_to_frame(df, rule, key_col, year_cols, value_col=None):
    """Apply a single scenario rule to one wide-format frame.

    Parameters
    ----------
    df : pd.DataFrame
        Frame to mutate. Returned unchanged if it doesn't carry the target parameter.
    rule : dict-like with keys Rule_Type, Parameter, Tech_Filter, Value
    key_col : str
        Column holding the tech code used to match Tech_Filter (e.g. 'Tech', 'Fuel/Tech').
    year_cols : list
        Year columns holding per-year values (may be empty for single-Value frames).
    value_col : str or None
        If the frame stores a single scalar per row (e.g. 'Value' in Fixed Horizon),
        name it here; year_cols should then be empty.
    """
    if df is None or df.empty:
        return df

    # Parameter filter (skip frames that don't carry Parameter column or don't match)
    if "Parameter" in df.columns:
        mask = df["Parameter"] == rule["Parameter"]
    else:
        # Frames without a Parameter column (e.g. passthrough DaySplit) — skip
        return df

    if not mask.any():
        return df

    # Tech_Filter
    tech_mask = df[key_col].apply(lambda t: _tech_filter_matches(rule["Tech_Filter"], t))
    target = mask & tech_mask

    if not target.any():
        return df

    rtype = str(rule["Rule_Type"]).strip().lower()

    if rtype == "clear":
        return df[~target].reset_index(drop=True)

    if rtype in ("cap", "set_value"):
        val = rule["Value"]
        if val is None:
            print(f"    WARN scenario rule '{rtype}' missing Value — skipping "
                  f"(param={rule['Parameter']} filter={rule['Tech_Filter']})")
            return df
        try:
            val = float(val)
        except (TypeError, ValueError):
            print(f"    WARN non-numeric Value {val!r} for rule '{rtype}' — skipping")
            return df
        df = df.copy()
        if value_col is not None and value_col in df.columns:
            if rtype == "cap":
                df.loc[target, value_col] = df.loc[target, value_col].where(
                    ~(df.loc[target, value_col] > val), val)
            else:  # set_value
                df.loc[target, value_col] = val
        else:
            cols_in_df = [c for c in year_cols if c in df.columns]
            for c in cols_in_df:
                if rtype == "cap":
                    df.loc[target, c] = df.loc[target, c].where(
                        ~(df.loc[target, c] > val), val)
                else:  # set_value
                    df.loc[target, c] = val
        return df

    print(f"    WARN unknown Rule_Type {rule['Rule_Type']!r} — skipping")
    return df


def apply_scenario_rules(frames, scenario_name, rules_df, year_cols):
    """Apply scenario rules to a dict of wide-format frames.

    Parameters
    ----------
    frames : dict
        {frame_name: (df, key_col, value_col_or_None)}. Each entry describes one
        merged frame and how to address rows in it.
    scenario_name : str
        Only rules whose Scenario column equals this (case-insensitive) are applied.
    rules_df : pd.DataFrame
        Scenario_Rules sheet as read from the template.
    year_cols : list
        Year columns used by year-based frames.

    Returns
    -------
    dict of the same shape, with (possibly) mutated DataFrames.
    """
    if rules_df is None or rules_df.empty or not scenario_name:
        return {k: v[0] for k, v in frames.items()}

    required = {"Scenario", "Rule_Type", "Parameter", "Tech_Filter", "Value"}
    missing = required - set(rules_df.columns)
    if missing:
        print(f"  Scenario_Rules sheet missing columns {missing} — skipping")
        return {k: v[0] for k, v in frames.items()}

    target_rules = rules_df[
        rules_df["Scenario"].astype(str).str.strip().str.lower()
        == scenario_name.strip().lower()
    ]
    if target_rules.empty:
        print(f"  No scenario rules matched '{scenario_name}'")
        return {k: v[0] for k, v in frames.items()}

    print(f"  Applying {len(target_rules)} scenario rule(s) for '{scenario_name}'...")

    out = {k: v[0].copy() if v[0] is not None and not v[0].empty else v[0]
           for k, v in frames.items()}

    for _, rule in target_rules.iterrows():
        for fname, (_, key_col, value_col) in frames.items():
            before = out[fname]
            if before is None or before.empty:
                continue
            after = _apply_rule_to_frame(before, rule, key_col, year_cols, value_col)
            if after is not before:
                out[fname] = after
                delta = len(before) - len(after)
                if delta != 0:
                    print(f"    [{fname}] rule {rule['Rule_Type']}/"
                          f"{rule['Parameter']}/{rule['Tech_Filter']}: "
                          f"{delta:+d} rows")

    return out


# ============================================================
# 8. TECH ID MANAGEMENT
# ============================================================

def _build_tech_id_map(df, tech_col="Tech"):
    """Build {tech_code: id} from existing wide-format DataFrame."""
    if df.empty or tech_col not in df.columns or "Tech.ID" not in df.columns:
        return {}
    mapping = {}
    for _, row in df.iterrows():
        t = row[tech_col]
        tid = row["Tech.ID"]
        if t is not None and tid is not None:
            try:
                mapping[str(t)] = int(tid)
            except (ValueError, TypeError):
                pass
    return mapping


def _assign_tech_ids(df, id_map, tech_col="Tech"):
    """Assign Tech.ID preserving existing, new from max+1."""
    if df.empty:
        return df
    max_id = max(id_map.values()) if id_map else 0
    ids = []
    for t in df[tech_col]:
        if t in id_map:
            ids.append(id_map[t])
        else:
            max_id += 1
            id_map[t] = max_id
            ids.append(max_id)
    df = df.copy()
    df["Tech.ID"] = ids
    return df


# ============================================================
# 9. EXCEL I/O
# ============================================================

def write_sheet_overlay(excel_path, sheet_name, df):
    """Overwrite a single sheet in an existing workbook, preserving others."""
    wb = openpyxl.load_workbook(excel_path)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(sheet_name)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    wb.save(excel_path)


def read_excel_sheet(excel_path, sheet_name):
    """Read a sheet from an Excel workbook into a DataFrame."""
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return pd.DataFrame()
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()
    if not rows:
        return pd.DataFrame()
    return pd.DataFrame(rows[1:], columns=list(rows[0]))


# ============================================================
# 10. ORCHESTRATORS
# ============================================================

def backup_outputs(a1_path):
    """Create timestamped backup of A-O files."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_dir = a1_path / f"_backup_{timestamp}"
    backup_dir.mkdir(exist_ok=True)
    for f in a1_path.glob("*.xlsx"):
        shutil.copy2(f, backup_dir / f.name)
    print(f"  Backup created: {backup_dir}")


def run_compute_mode(sheets, a1_path):
    """Compute mode: green sheets → A-O Excel files."""
    all_years = TARGET_YEARS
    compute_flags = CFG.get("compute", {})
    enable_dsptrn = get_enable_dsptrn()

    param_excel = a1_path / "A-O_Parametrization.xlsx"
    demand_excel = a1_path / "A-O_Demand.xlsx"

    # --- Generate long-format data from green sheets ---
    print("\n--- Computing from green sheets ---")

    cap_cost_df = pd.DataFrame()
    fix_cost_df = pd.DataFrame()
    res_cap_df = pd.DataFrame()
    op_life_df = pd.DataFrame()
    var_cost_df = pd.DataFrame()
    demand_df = pd.DataFrame()
    planned_df = pd.DataFrame()

    if compute_flags.get("costs", False):
        print("  Computing CapitalCost...")
        cap_cost_df = compute_costs(sheets, "CAPEX", "CapitalCost")
        print(f"    {len(cap_cost_df)} rows")
        print("  Computing FixedCost...")
        fix_cost_df = compute_costs(sheets, "Fixed O&M", "FixedCost")
        print(f"    {len(fix_cost_df)} rows")

    if compute_flags.get("residual_capacity", False):
        print("  Computing ResidualCapacity...")
        res_cap_df = compute_residual_capacity(sheets)
        print(f"    {len(res_cap_df)} rows")

    if compute_flags.get("planned_capacity", False):
        print("  Computing PlannedCapacity (TotalAnnualMinCapacity)...")
        planned_df = compute_planned_capacity(sheets)
        print(f"    {len(planned_df)} rows")

    if compute_flags.get("operational_life", False):
        print("  Computing OperationalLife...")
        op_life_df = compute_operational_life(sheets)
        print(f"    {len(op_life_df)} rows")

    if compute_flags.get("variable_cost", False):
        print("  Computing VariableCost...")
        var_cost_df = compute_variable_cost(sheets)
        print(f"    {len(var_cost_df)} rows")

    if compute_flags.get("demand", False):
        print("  Computing SpecifiedAnnualDemand...")
        demand_df = compute_demand(sheets)
        print(f"    {len(demand_df)} rows")

    # --- Format to wide ---
    print("\n--- Formatting to wide ---")
    wide_capex = format_param_yearly(cap_cost_df, "CapitalCost", 3, all_years)
    wide_fixom = format_param_yearly(fix_cost_df, "FixedCost", 4, all_years)
    wide_rescap = format_param_yearly(res_cap_df, "ResidualCapacity", 5, all_years)
    wide_planned = format_param_yearly(planned_df, "TotalAnnualMinCapacity", 6, all_years)
    wide_oplife = format_fixed_horizon(op_life_df)
    wide_varcost = format_variable_cost(var_cost_df, all_years)
    wide_demand = format_demand_projection(demand_df, enable_dsptrn, all_years)

    # --- Split Primary / Secondary ---
    def split_pri_sec(df):
        if df.empty:
            return pd.DataFrame(), pd.DataFrame()
        pri = df[df["Tech"].apply(lambda t: str(t).startswith("MIN") or str(t).startswith("RNW"))]
        sec = df[df["Tech"].apply(
            lambda t: not str(t).startswith("MIN") and not str(t).startswith("RNW")
            and not str(t).startswith("PWRTRN"))]
        return pri, sec

    pri_capex, sec_capex = split_pri_sec(wide_capex)
    pri_fixom, sec_fixom = split_pri_sec(wide_fixom)
    pri_rescap, sec_rescap = split_pri_sec(wide_rescap)
    _, sec_planned = split_pri_sec(wide_planned)

    new_primary = pd.concat([pri_capex, pri_fixom, pri_rescap], ignore_index=True)
    new_secondary = pd.concat([sec_capex, sec_fixom, sec_rescap, sec_planned], ignore_index=True)

    # --- Passthrough sheets (read blue sheets from template as-is) ---
    print("\n--- Reading passthrough blue sheets ---")

    # Read existing A-O files if they exist, otherwise use template blue sheets
    if param_excel.exists():
        exist_primary = read_excel_sheet(str(param_excel), "Primary Techs")
        exist_secondary = read_excel_sheet(str(param_excel), "Secondary Techs")
        exist_fixed = read_excel_sheet(str(param_excel), "Fixed Horizon Parameters")
        exist_capacities = read_excel_sheet(str(param_excel), "Capacities")
        exist_varcost = read_excel_sheet(str(param_excel), "VariableCost")
        exist_ysplit = read_excel_sheet(str(param_excel), "Yearsplit")
    else:
        # Bootstrap from template blue sheets
        exist_primary = sheets.get("Primary_Techs", pd.DataFrame())
        exist_secondary = sheets.get("Secondary_Techs", pd.DataFrame())
        exist_fixed = sheets.get("Fixed_Horizon_Parameters", pd.DataFrame())
        exist_capacities = sheets.get("Capacities_CF", pd.DataFrame())
        exist_varcost = sheets.get("VariableCost", pd.DataFrame())
        exist_ysplit = sheets.get("Yearsplit_Template", pd.DataFrame())

    if demand_excel.exists():
        exist_demand_proj = read_excel_sheet(str(demand_excel), "Demand_Projection")
        exist_demand_prof = read_excel_sheet(str(demand_excel), "Profiles")
    else:
        exist_demand_proj = sheets.get("Demand_Projection", pd.DataFrame())
        exist_demand_prof = sheets.get("Demand_Profiles", pd.DataFrame())

    # --- Merge computed data into existing ---
    print("  Merging...")

    cor_flags = CFG.get("clear_on_replace", {})

    if not new_primary.empty:
        merged_primary = merge_wide_by_tech_param(
            exist_primary, new_primary,
            clear_on_replace=cor_flags.get("Primary_Techs", False))
    else:
        merged_primary = exist_primary

    if not new_secondary.empty:
        merged_secondary = merge_wide_by_tech_param(
            exist_secondary, new_secondary,
            clear_on_replace=cor_flags.get("Secondary_Techs", False))
    else:
        merged_secondary = exist_secondary

    if not wide_oplife.empty:
        merged_fixed = merge_wide_by_tech(
            exist_fixed, wide_oplife,
            clear_on_replace=cor_flags.get("OperationalLife", False))
    else:
        merged_fixed = exist_fixed

    # Passthrough sheets (not computed)
    merged_capacities = exist_capacities
    if not compute_flags.get("capacity_factors", False):
        pass  # Use existing

    merged_varcost = exist_varcost
    if not wide_varcost.empty:
        merged_varcost = merge_wide_by_tech(
            exist_varcost, wide_varcost,
            clear_on_replace=cor_flags.get("VariableCost", False))

    merged_ysplit = exist_ysplit

    merged_demand_proj = exist_demand_proj
    if not wide_demand.empty:
        merged_demand_proj = merge_wide_by_fuel(
            exist_demand_proj, wide_demand,
            clear_on_replace=cor_flags.get("Demand_Projection", False))

    merged_demand_prof = exist_demand_prof

    # --- Apply scenario rules (after merge, before write) ---
    scenario_name = CFG.get("scenario")
    rules_df = sheets.get("Scenario_Rules", pd.DataFrame())
    if scenario_name and not rules_df.empty:
        print(f"\n--- Applying scenario rules ({scenario_name}) ---")
        scenario_frames = {
            "Primary Techs":            (merged_primary,     "Tech",       None),
            "Secondary Techs":          (merged_secondary,   "Tech",       None),
            "Fixed Horizon Parameters": (merged_fixed,       "Tech",       "Value"),
            "Capacities":               (merged_capacities,  "Tech",       None),
            "VariableCost":             (merged_varcost,     "Tech",       None),
            "Demand_Projection":        (merged_demand_proj, "Fuel/Tech",  None),
        }
        updated = apply_scenario_rules(scenario_frames, scenario_name, rules_df, all_years)
        merged_primary     = updated["Primary Techs"]
        merged_secondary   = updated["Secondary Techs"]
        merged_fixed       = updated["Fixed Horizon Parameters"]
        merged_capacities  = updated["Capacities"]
        merged_varcost     = updated["VariableCost"]
        merged_demand_proj = updated["Demand_Projection"]

    # --- Assign Tech IDs ---
    id_map = {}
    for edf in [exist_primary, exist_secondary, exist_fixed, exist_capacities, exist_varcost]:
        id_map.update(_build_tech_id_map(edf))

    if "Tech" in merged_primary.columns:
        merged_primary = _assign_tech_ids(merged_primary, id_map)
    if "Tech" in merged_secondary.columns:
        merged_secondary = _assign_tech_ids(merged_secondary, id_map)
    if "Tech" in merged_fixed.columns:
        merged_fixed = _assign_tech_ids(merged_fixed, id_map)
    if "Tech" in merged_capacities.columns:
        merged_capacities = _assign_tech_ids(merged_capacities, id_map)
    if "Tech" in merged_varcost.columns:
        merged_varcost = _assign_tech_ids(merged_varcost, id_map)

    # --- Sort ---
    for df in [merged_primary, merged_secondary]:
        if not df.empty and "Tech.ID" in df.columns and "Parameter.ID" in df.columns:
            df.sort_values(["Tech.ID", "Parameter.ID"], inplace=True)

    # --- Write A-O_Parametrization.xlsx ---
    print("\n--- Writing A-O_Parametrization.xlsx ---")
    if not param_excel.exists():
        # Create from scratch — need a template workbook
        wb = openpyxl.Workbook()
        for sname in ["Primary Techs", "Secondary Techs", "Fixed Horizon Parameters",
                       "Capacities", "VariableCost", "Yearsplit", "DaySplit", "Demand Techs"]:
            wb.create_sheet(sname)
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        wb.save(str(param_excel))

    write_sheet_overlay(str(param_excel), "Primary Techs", merged_primary)
    print("    Primary Techs updated")
    write_sheet_overlay(str(param_excel), "Secondary Techs", merged_secondary)
    print("    Secondary Techs updated")
    write_sheet_overlay(str(param_excel), "Fixed Horizon Parameters", merged_fixed)
    print("    Fixed Horizon Parameters updated")
    write_sheet_overlay(str(param_excel), "Capacities", merged_capacities)
    print("    Capacities updated")
    write_sheet_overlay(str(param_excel), "VariableCost", merged_varcost)
    print("    VariableCost updated")
    write_sheet_overlay(str(param_excel), "Yearsplit", merged_ysplit)
    print("    Yearsplit updated")

    # DaySplit — passthrough from template
    if "DaySplit" in sheets and not sheets["DaySplit"].empty:
        write_sheet_overlay(str(param_excel), "DaySplit", sheets["DaySplit"])
        print("    DaySplit updated")

    # Demand Techs — passthrough from template (PWRTRN/RNWTRN/DSPTRN)
    if "Demand_Techs" in sheets and not sheets["Demand_Techs"].empty:
        write_sheet_overlay(str(param_excel), "Demand Techs", sheets["Demand_Techs"])
        print("    Demand Techs updated")

    # --- Write A-O_Demand.xlsx ---
    print("\n--- Writing A-O_Demand.xlsx ---")
    if not demand_excel.exists():
        wb = openpyxl.Workbook()
        for sname in ["Demand_Projection", "Profiles"]:
            wb.create_sheet(sname)
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        wb.save(str(demand_excel))

    write_sheet_overlay(str(demand_excel), "Demand_Projection", merged_demand_proj)
    print("    Demand_Projection updated")
    write_sheet_overlay(str(demand_excel), "Profiles", merged_demand_prof)
    print("    Profiles updated")

    # --- Write A-Xtra_Emissions.xlsx (GHGs) ---
    if compute_flags.get("emissions", False):
        print("\n--- Writing A-Xtra_Emissions.xlsx ---")
        emissions_excel = a1_path / "A-Xtra_Emissions.xlsx"
        ghgs_df = build_ghgs_from_emissions(sheets)
        if ghgs_df.empty:
            print("    WARN no EmissionActivityRatio rows found in SoT Emissions sheet — skipping GHGs")
        else:
            if not emissions_excel.exists():
                wb = openpyxl.Workbook()
                wb.create_sheet("GHGs")
                if "Sheet" in wb.sheetnames:
                    del wb["Sheet"]
                wb.save(str(emissions_excel))
            write_sheet_overlay(str(emissions_excel), "GHGs", ghgs_df)
            print(f"    GHGs updated ({len(ghgs_df)} rows)")

    print("\n  A-O files written successfully.")


def run_passthrough_mode(sheets, a1_path):
    """Passthrough mode: read blue sheets from template → write A-O files directly."""
    param_excel = a1_path / "A-O_Parametrization.xlsx"
    demand_excel = a1_path / "A-O_Demand.xlsx"

    # Sheet name mapping: template name → A-O sheet name
    param_sheet_map = {
        "Primary_Techs": "Primary Techs",
        "Secondary_Techs": "Secondary Techs",
        "Demand_Techs": "Demand Techs",
        "Fixed_Horizon_Parameters": "Fixed Horizon Parameters",
        "Capacities_CF": "Capacities",
        "VariableCost": "VariableCost",
        "Yearsplit_Template": "Yearsplit",
        "DaySplit": "DaySplit",
        "Emissions": "Emissions",
    }
    demand_sheet_map = {
        "Demand_Projection": "Demand_Projection",
        "Demand_Profiles": "Profiles",
    }

    # --- Write A-O_Parametrization.xlsx ---
    print("\n--- Writing A-O_Parametrization.xlsx (passthrough) ---")
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    for template_name, ao_name in param_sheet_map.items():
        df = sheets.get(template_name, pd.DataFrame())
        if df.empty:
            print(f"  WARNING: {template_name} is empty, skipping")
            continue
        ws = wb.create_sheet(ao_name)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        print(f"    {ao_name} ← {template_name} ({len(df)} rows)")

    # Also include Interconnector_Params merged into Secondary Techs if present
    if "Interconnector_Params" in sheets and not sheets["Interconnector_Params"].empty:
        ic_df = sheets["Interconnector_Params"]
        # Append to Secondary Techs sheet
        if "Secondary Techs" in wb.sheetnames:
            ws = wb["Secondary Techs"]
            for r in dataframe_to_rows(ic_df, index=False, header=False):
                ws.append(r)
            print(f"    Interconnector_Params appended to Secondary Techs ({len(ic_df)} rows)")

    wb.save(str(param_excel))

    # --- Write A-O_Demand.xlsx ---
    print("\n--- Writing A-O_Demand.xlsx (passthrough) ---")
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    for template_name, ao_name in demand_sheet_map.items():
        df = sheets.get(template_name, pd.DataFrame())
        if df.empty:
            print(f"  WARNING: {template_name} is empty, skipping")
            continue
        ws = wb.create_sheet(ao_name)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        print(f"    {ao_name} ← {template_name} ({len(df)} rows)")

    wb.save(str(demand_excel))

    print("\n  A-O files written successfully (passthrough).")


# ============================================================
# 11. MAIN
# ============================================================

def main():
    print("=" * 70)
    print(f"A3 DataPackage V3: {XLSX_PATH.name}")
    print(f"  Scenario: {CFG['scenario']}")
    print(f"  Cost scenario: {COST_SCENARIO}")
    print(f"  Horizon: {FIRST_YEAR}-{FINAL_YEAR}")
    print("=" * 70)

    # Verify template exists
    if not XLSX_PATH.exists():
        print(f"ERROR: {XLSX_PATH} not found.")
        sys.exit(1)

    # Read all sheets
    print("\n--- Reading template ---")
    sheets = read_template(XLSX_PATH)
    print(f"  Loaded {len(sheets)} sheets: {', '.join(sheets.keys())}")

    # Verify required sheets
    required_blue = ["Primary_Techs", "Secondary_Techs", "Demand_Projection", "Demand_Profiles"]
    required_green = ["Node_Mapping"]
    for s in required_blue + required_green:
        if s not in sheets or sheets[s].empty:
            print(f"ERROR: Required sheet '{s}' is missing or empty.")
            sys.exit(1)

    # Determine output path
    a1_subdir = CFG.get("a1_outputs_subdir", "A1_Outputs/A1_Outputs_BAU")
    a1_path = SCRIPT_DIR / a1_subdir
    a1_path.mkdir(parents=True, exist_ok=True)

    # Backup
    if CFG.get("backup", True) and any(a1_path.glob("*.xlsx")):
        print("\n--- Backup ---")
        backup_outputs(a1_path)

    # Check if any compute flag is true
    compute_flags = CFG.get("compute", {})
    any_compute = any(compute_flags.values())

    if any_compute:
        # Verify green sheets needed for compute
        if compute_flags.get("costs", False):
            for s in ["Technology_Costs", "Country_Multipliers"]:
                if s not in sheets or sheets[s].empty:
                    print(f"ERROR: compute.costs requires '{s}' sheet.")
                    sys.exit(1)
        if compute_flags.get("residual_capacity", False):
            if "Existing_Generation" not in sheets or sheets["Existing_Generation"].empty:
                print("ERROR: compute.residual_capacity requires 'Existing_Generation' sheet.")
                sys.exit(1)
        if compute_flags.get("variable_cost", False):
            if "Fuel_Costs" not in sheets or sheets["Fuel_Costs"].empty:
                print("ERROR: compute.variable_cost requires 'Fuel_Costs' sheet.")
                sys.exit(1)
        if compute_flags.get("demand", False):
            if "Demand_Assumptions" not in sheets or sheets["Demand_Assumptions"].empty:
                print("ERROR: compute.demand requires 'Demand_Assumptions' sheet.")
                sys.exit(1)

        run_compute_mode(sheets, a1_path)
    else:
        run_passthrough_mode(sheets, a1_path)

    # Optionally update blue sheets in the template itself
    if CFG.get("update_blue_sheets_in_template", False) and any_compute:
        print("\n--- Updating blue sheets in template ---")
        # Read back what we just wrote and update the template
        param_excel = a1_path / "A-O_Parametrization.xlsx"
        demand_excel = a1_path / "A-O_Demand.xlsx"

        ao_to_template = {
            "Primary Techs": "Primary_Techs",
            "Secondary Techs": "Secondary_Techs",
            "Fixed Horizon Parameters": "Fixed_Horizon_Parameters",
            "Capacities": "Capacities_CF",
            "VariableCost": "VariableCost",
            "Yearsplit": "Yearsplit_Template",
        }
        for ao_name, tpl_name in ao_to_template.items():
            df = read_excel_sheet(str(param_excel), ao_name)
            if not df.empty:
                write_sheet_overlay(str(XLSX_PATH), tpl_name, df)
                print(f"    {tpl_name} updated")

        demand_to_template = {
            "Demand_Projection": "Demand_Projection",
            "Profiles": "Demand_Profiles",
        }
        for ao_name, tpl_name in demand_to_template.items():
            df = read_excel_sheet(str(demand_excel), ao_name)
            if not df.empty:
                write_sheet_overlay(str(XLSX_PATH), tpl_name, df)
                print(f"    {tpl_name} updated")

    print("\n" + "=" * 70)
    print("A3 DataPackage V3: Complete. Next step: run B1_Compiler.py")
    print("=" * 70)


if __name__ == "__main__":
    main()
