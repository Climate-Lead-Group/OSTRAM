# -*- coding: utf-8 -*-
"""
merge_timeslices_into_SoT.py

Reads OSTRAM_Timeslice_Outputs.xlsx and writes the timeslice data into
the Source-of-Truth workbook (SOASIA_DataPackage_V3.xlsx) at three slots:

    YearSplit              -> Yearsplit_Template
    {region}_Dem (x10)     -> Demand_Profiles    (unioned across regions)
    {region}_CF  (x10)     -> Capacities_CF      (unioned across regions)

Idempotent: re-running replaces the target sheets' contents entirely.

Usage from t1_confection/test_a3_mod/:
    python merge_timeslices_into_SoT.py
Optional positional args:
    python merge_timeslices_into_SoT.py <timeslice_xlsx> <sot_xlsx>

Tech.IDs in the generated Capacities_CF are back-filled from the SoT's
Primary_Techs + Secondary_Techs sheets where possible; missing ones default
to 0 and are resolved by A3 downstream.
"""
import sys
from pathlib import Path

import openpyxl
import pandas as pd
import yaml
from openpyxl.utils.dataframe import dataframe_to_rows


SCRIPT_DIR = Path(__file__).resolve().parent


# Timeslice-file region prefix → SoT OSTRAM region code
TIMESLICE_REGION_MAP = {
    "BGD":   "BGDXX",
    "BTN":   "BTNXX",
    "LKA":   "LKAXX",
    "MDV":   "MDVXX",
    "NPL":   "NPLXX",
    "INDEA": "INDEA",
    "INDNE": "INDNE",
    "INDNO": "INDNO",
    "INDSO": "INDSO",
    "INDWE": "INDWE",
}


def load_config():
    path = SCRIPT_DIR / "Config_datapackage_v3.yaml"
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def read_sheet(xlsx_path, sheet_name):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return pd.DataFrame()
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return pd.DataFrame()
    header = list(rows[0])
    data = [list(r) for r in rows[1:] if any(c is not None for c in r)]
    return pd.DataFrame(data, columns=header)


def write_sheet_overlay(xlsx_path, sheet_name, df):
    """Overwrite one sheet in an existing workbook, preserving others.

    Same pattern as A3_update_from_datapackage_v3.write_sheet_overlay().
    """
    wb = openpyxl.load_workbook(xlsx_path)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(sheet_name)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    wb.save(xlsx_path)


def build_tech_id_map(sot_path):
    """Build {tech_code: tech_id} from Primary_Techs + Secondary_Techs in SoT."""
    id_map = {}
    for sname in ("Primary_Techs", "Secondary_Techs"):
        df = read_sheet(sot_path, sname)
        if df.empty or "Tech" not in df.columns or "Tech.ID" not in df.columns:
            continue
        for _, r in df.iterrows():
            tech = r["Tech"]
            tid = r["Tech.ID"]
            if tech is None or tid is None:
                continue
            try:
                id_map[str(tech)] = int(tid)
            except (ValueError, TypeError):
                continue
    return id_map


# ------------------------------------------------------------------
# Builders — each produces a wide-format DataFrame matching the SoT sheet
# ------------------------------------------------------------------

def build_yearsplit(src_df, years):
    """YearSplit source (timeslice, season, daypart, yearsplit)
    → Timeslices, Parameter.ID=14, Parameter='YearSplit', Unit, Projection.Mode,
      Projection.Parameter=0, <years...>
    """
    rows = []
    for _, r in src_df.iterrows():
        row = {
            "Timeslices": r["timeslice"],
            "Parameter.ID": 14,
            "Parameter": "YearSplit",
            "Unit": None,
            "Projection.Mode": "User defined",
            "Projection.Parameter": 0,
        }
        for y in years:
            row[y] = r["yearsplit"]
        rows.append(row)
    return pd.DataFrame(rows)


def _best_cf(r):
    """Pick the best capacity factor from the CF sources, in priority order."""
    for key in ("cf_dispatch", "cf_ninja", "cf_da_workbook", "cf_default"):
        v = r.get(key)
        if v is not None and not (isinstance(v, float) and pd.isna(v)):
            return v
    return None


def build_capacities_cf(timeslice_xlsx, years, tech_id_map):
    """Concatenate {region}_CF sheets into Capacities_CF rows."""
    out = []
    for prefix in TIMESLICE_REGION_MAP:
        sheet = f"{prefix}_CF"
        df = read_sheet(timeslice_xlsx, sheet)
        if df.empty:
            print(f"    {sheet}: empty/missing — skipped")
            continue
        for _, r in df.iterrows():
            cf = _best_cf(r)
            tech = r["tech_code"]
            row = {
                "Timeslices": r["timeslice"],
                "Tech.ID": tech_id_map.get(str(tech), 0),
                "Tech": tech,
                "Tech.Name": None,   # A3 / B1 tolerate blank name; can be enriched later
                "Parameter.ID": 13,
                "Parameter": "CapacityFactor",
                "Unit": None,
                "Projection.Mode": "User defined",
                "Projection.Parameter": 0,
            }
            for y in years:
                row[y] = cf
            out.append(row)
        print(f"    {sheet}: {len(df)} rows")
    return pd.DataFrame(out)


def build_demand_profiles(timeslice_xlsx, years):
    """Concatenate {region}_Dem sheets into Demand_Profiles rows.

    Fuel/Tech code = 'ELC{OSTRAM_Region}03' to match the dispatch-ready suffix
    used in _CONTROL_TIMESLICES/A-O_Demand.xlsx Profiles.
    """
    out = []
    for prefix, region in TIMESLICE_REGION_MAP.items():
        sheet = f"{prefix}_Dem"
        df = read_sheet(timeslice_xlsx, sheet)
        if df.empty:
            print(f"    {sheet}: empty/missing — skipped")
            continue
        fuel_code = f"ELC{region}03"
        for _, r in df.iterrows():
            row = {
                "Timeslices": r["timeslice"],
                "Demand/Share": "Demand",
                "Fuel/Tech": fuel_code,
                "Name": None,
                "Ref.Cap.BY": "not needed",
                "Ref.OAR.BY": "not needed",
                "Ref.km.BY": "not needed",
                "Projection.Mode": "User defined",
                "Projection.Parameter": 0,
            }
            for y in years:
                row[y] = r["demand_fraction"]
            out.append(row)
        print(f"    {sheet}: {len(df)} rows -> {fuel_code}")
    return pd.DataFrame(out)


# ------------------------------------------------------------------
# Main
# ------------------------------------------------------------------

def main():
    cfg = load_config()
    years = list(range(cfg["first_year"], cfg["final_year"] + 1))

    ts_path = Path(sys.argv[1]) if len(sys.argv) > 1 else \
        SCRIPT_DIR / "OSTRAM_Timeslice_Outputs.xlsx"
    sot_path = Path(sys.argv[2]) if len(sys.argv) > 2 else \
        SCRIPT_DIR / cfg["datapackage_v3_path"]

    if not ts_path.exists():
        sys.exit(f"ERROR: timeslice file not found: {ts_path}")
    if not sot_path.exists():
        sys.exit(f"ERROR: SoT file not found: {sot_path}")

    print(f"Source : {ts_path}")
    print(f"Target : {sot_path}")
    print(f"Years  : {years[0]}..{years[-1]} ({len(years)})")

    print("\n[1/3] YearSplit -> Yearsplit_Template")
    yr = read_sheet(str(ts_path), "YearSplit")
    if yr.empty:
        print("  WARN YearSplit sheet empty/missing — skipping")
    else:
        # The timeslice pipeline is authoritative for the timeslice set.
        src_ts = sorted(set(yr["timeslice"].dropna().astype(str).tolist()))
        print(f"  source timeslices ({len(src_ts)}): {src_ts}")
        ys_df = build_yearsplit(yr, years)
        write_sheet_overlay(str(sot_path), "Yearsplit_Template", ys_df)
        print(f"  -> {len(ys_df)} rows written")

    print("\n[2/3] {region}_CF -> Capacities_CF")
    tech_id_map = build_tech_id_map(str(sot_path))
    print(f"  (resolved {len(tech_id_map)} Tech.IDs from SoT Primary/Secondary Techs)")
    cf_df = build_capacities_cf(str(ts_path), years, tech_id_map)
    if cf_df.empty:
        print("  WARN no CF rows produced")
    else:
        write_sheet_overlay(str(sot_path), "Capacities_CF", cf_df)
        print(f"  -> {len(cf_df)} rows written")

    print("\n[3/3] {region}_Dem -> Demand_Profiles")
    dp_df = build_demand_profiles(str(ts_path), years)
    if dp_df.empty:
        print("  WARN no demand profile rows produced")
    else:
        write_sheet_overlay(str(sot_path), "Demand_Profiles", dp_df)
        print(f"  -> {len(dp_df)} rows written")

    print("\nDone.")


if __name__ == "__main__":
    main()
