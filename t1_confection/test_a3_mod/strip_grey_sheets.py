# -*- coding: utf-8 -*-
"""
strip_grey_sheets.py

One-off housekeeping script for the Source-of-Truth workbook.

Reads SOASIA_DataPackage_V3.xlsx, copies the grey (reference / audit) sheets
into SOASIA_DataPackage_V3_REFERENCE.xlsx, then removes them from the
working template so only model-driving sheets remain.

Idempotent: running it again on an already-clean template is a no-op for
sheets that don't exist and simply refreshes the REFERENCE file.

Usage from t1_confection/test_a3_mod/:
    python strip_grey_sheets.py
Optional positional args:
    python strip_grey_sheets.py <working_xlsx> <reference_xlsx>
"""
import shutil
import sys
from pathlib import Path

import openpyxl


SCRIPT_DIR = Path(__file__).resolve().parent

# Sheets to move out of the working template.
GREY_SHEETS = [
    "Provenance_Matrix",
    "Cross_Validation_Log",
    "ATB_Comparison",
    "Capacities_CF_Options",
    "Validation_ThermalDispatch",
    "Integration_Log",
    "Yearsplit_ReNinja",
    "RE_Potential_Ceilings",
    "RE_Targets_Policies",
    "README",
]


def copy_sheet(src_ws, dst_wb, sheet_name):
    """Copy values from src_ws into a new sheet in dst_wb (values only, no styles)."""
    if sheet_name in dst_wb.sheetnames:
        del dst_wb[sheet_name]
    dst_ws = dst_wb.create_sheet(sheet_name)
    for row in src_ws.iter_rows(values_only=True):
        dst_ws.append(row)


def main():
    working = Path(sys.argv[1]) if len(sys.argv) > 1 else \
        SCRIPT_DIR / "SOASIA_DataPackage_V3.xlsx"
    reference = Path(sys.argv[2]) if len(sys.argv) > 2 else \
        SCRIPT_DIR / "SOASIA_DataPackage_V3_REFERENCE.xlsx"

    if not working.exists():
        sys.exit(f"ERROR: working template not found: {working}")

    print(f"Working   : {working}")
    print(f"Reference : {reference}")

    # Load working template
    src_wb = openpyxl.load_workbook(working)

    # Open (or create fresh) reference workbook
    if reference.exists():
        dst_wb = openpyxl.load_workbook(reference)
    else:
        dst_wb = openpyxl.Workbook()
        if "Sheet" in dst_wb.sheetnames:
            del dst_wb["Sheet"]

    # Copy each grey sheet into reference, then remove from working
    copied = []
    missing = []
    for name in GREY_SHEETS:
        if name not in src_wb.sheetnames:
            missing.append(name)
            continue
        copy_sheet(src_wb[name], dst_wb, name)
        del src_wb[name]
        copied.append(name)

    # Save. If the reference workbook ended up with zero sheets (nothing to copy
    # and it didn't previously exist), add a placeholder so openpyxl can save.
    if len(dst_wb.sheetnames) == 0:
        dst_wb.create_sheet("_empty_")

    dst_wb.save(reference)
    src_wb.save(working)

    print()
    print(f"Moved to REFERENCE ({len(copied)}): {copied}")
    if missing:
        print(f"Not present in template ({len(missing)}): {missing}")
    print()
    print(f"Working template sheets ({len(src_wb.sheetnames)}):")
    for s in src_wb.sheetnames:
        print(f"  - {s}")
    print("\nDone.")


if __name__ == "__main__":
    main()
