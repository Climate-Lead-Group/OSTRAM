# -*- coding: utf-8 -*-
"""
A3_setup_template_v3.py
One-time script to add the 4 new inventory sheets to the DataPackage V3 template.

Adds:
  - Node_Mapping
  - Country_Multipliers
  - Fuel_Costs
  - Demand_Assumptions

Run once from t1_confection/:
    python A3_setup_template_v3.py <path_to_template.xlsx> <output_path.xlsx>
"""
import shutil
import sys
from pathlib import Path

import openpyxl


def add_node_mapping(wb):
    """Add Node_Mapping sheet."""
    ws = wb.create_sheet("Node_Mapping")
    header = ["Short_Node", "OSTRAM_Region", "ISO3", "Country_Name", "Multi_Region"]
    ws.append(header)
    rows = [
        ["BD",    "BGDXX", "BGD", "Bangladesh",       "No"],
        ["BT",    "BTNXX", "BTN", "Bhutan",            "No"],
        ["IN_N",  "INDNO", "IND", "India - North",     "Yes"],
        ["IN_S",  "INDSO", "IND", "India - South",     "Yes"],
        ["IN_E",  "INDEA", "IND", "India - East",      "Yes"],
        ["IN_W",  "INDWE", "IND", "India - West",      "Yes"],
        ["IN_NE", "INDNE", "IND", "India - NorthEast", "Yes"],
        ["LK",    "LKAXX", "LKA", "Sri Lanka",         "No"],
        ["MV",    "MDVXX", "MDV", "Maldives",          "No"],
        ["NP",    "NPLXX", "NPL", "Nepal",             "No"],
    ]
    for r in rows:
        ws.append(r)
    print(f"  Added Node_Mapping ({len(rows)} rows)")


def add_country_multipliers(wb):
    """Add Country_Multipliers sheet with cost scaling factors."""
    ws = wb.create_sheet("Country_Multipliers")
    header = ["Technology_Code", "BD", "BT", "IN", "LK", "MV", "NP", "Source"]
    ws.append(header)
    # India is always 1.0 (base). N/A = use India base.
    rows = [
        ["PWRSPV",     1.5,  1.8, 1.0, 1.5, 3.0, 1.6, "Cost DB Node_Anchors"],
        ["PWRSPR",     1.5,  1.8, 1.0, 1.2, 3.0, 1.6, "Cost DB Node_Anchors"],
        ["PWRWON",     1.6,  None, 1.0, 1.8, None, None, "Cost DB Node_Anchors"],
        ["PWRWOF",     None, None, 1.0, 1.6, None, None, "Cost DB Node_Anchors"],
        ["PWRHDR",     None, 0.8, 1.0, 1.7, None, 1.8, "Cost DB Node_Anchors"],
        ["PWRHRO",     None, 0.85, 1.0, 1.4, None, 1.2, "Cost DB Node_Anchors"],
        ["PWRHPS",     None, None, 1.0, None, None, None, "Cost DB Node_Anchors"],
        ["PWRCOASCPC", 1.6,  None, 1.0, 2.0, None, None, "Cost DB Node_Anchors"],
        ["PWRCCG",     1.1,  None, 1.0, 1.9, None, None, "Cost DB Node_Anchors"],
        ["PWROCG",     1.1,  None, 1.0, 0.8, None, None, "Cost DB Node_Anchors"],
        ["PWROIL",     1.1,  None, 1.0, 1.4, None, 1.0, "Cost DB Node_Anchors"],
        ["PWRBIO",     None, None, 1.0, 1.6, None, None, "Cost DB Node_Anchors"],
        ["PWRURN",     2.6,  None, 1.0, 2.6, None, None, "Cost DB Node_Anchors"],
        ["PWRSDS",     1.2,  1.4, 1.0, 1.5, 1.5, 1.3, "Cost DB Node_Anchors"],
    ]
    for r in rows:
        ws.append(r)
    print(f"  Added Country_Multipliers ({len(rows)} rows)")


def add_fuel_costs(wb):
    """Add Fuel_Costs sheet with milestone-year fuel prices."""
    ws = wb.create_sheet("Fuel_Costs")
    header = ["Fuel", "Node_Override", "Scenario", "Unit",
              2023, 2025, 2030, 2035, 2040, 2045, 2050, "Source"]
    ws.append(header)
    rows = [
        ["Imported coal (FOB Aus)", "BD,LK", "Conservative", "USD/GJ", 6.9, 4.3, 4.2, 4.2, 4.2, 4.2, 4.2, "IEA WEO"],
        ["Imported coal (FOB Aus)", "BD,LK", "Moderate",     "USD/GJ", 6.9, 4.3, 3.8, 3.5, 3.2, 3.0, 2.8, "IEA WEO"],
        ["Imported coal (FOB Aus)", "BD,LK", "Ambitious",    "USD/GJ", 6.9, 3.8, 2.5, 1.8, 1.5, 1.2, 0.9, "IEA WEO"],
        ["India domestic coal", "IN_all", "Conservative", "USD/GJ", 2.2, 2.2, 2.4, 2.5, 2.6, 2.7, 2.8, "Coal India"],
        ["India domestic coal", "IN_all", "Moderate",     "USD/GJ", 2.2, 2.2, 2.3, 2.3, 2.4, 2.4, 2.5, "Coal India"],
        ["India domestic coal", "IN_all", "Ambitious",    "USD/GJ", 2.2, 2.2, 2.2, 2.1, 2.0, 1.9, 1.8, "Coal India"],
        ["Bangladesh piped gas", "BD", "Conservative", "USD/GJ", 5.5, 5.5, 6.5, 8.0, 9.0, 9.5, 10.0, "Petrobangla"],
        ["Bangladesh piped gas", "BD", "Moderate",     "USD/GJ", 5.5, 5.5, 5.7, 5.8, 5.9, 6.0, 6.0, "Petrobangla"],
        ["Bangladesh piped gas", "BD", "Ambitious",    "USD/GJ", 5.5, 5.5, 5.0, 4.5, 4.0, 3.5, 3.0, "Petrobangla"],
        ["LNG (Asia/JKM)", "IN_W,IN_S,LK", "Conservative", "USD/GJ", 13.6, 11.9, 10.0, 10.0, 10.5, 10.5, 10.5, "Platts JKM"],
        ["LNG (Asia/JKM)", "IN_W,IN_S,LK", "Moderate",     "USD/GJ", 13.6, 11.9, 10.0, 10.0, 10.0, 10.0, 10.0, "Platts JKM"],
        ["LNG (Asia/JKM)", "IN_W,IN_S,LK", "Ambitious",    "USD/GJ", 13.6, 10.0, 7.0, 5.0, 4.5, 4.0, 4.0, "Platts JKM"],
        ["Diesel / HFO", "LK,MV", "Conservative", "USD/GJ", 18.0, 14.8, 14.5, 14.5, 14.5, 14.5, 14.5, "IEA"],
        ["Diesel / HFO", "LK,MV", "Moderate",     "USD/GJ", 18.0, 14.8, 14.0, 14.0, 14.0, 14.0, 14.0, "IEA"],
        ["Diesel / HFO", "LK,MV", "Ambitious",    "USD/GJ", 18.0, 12.0, 10.0, 6.0, 5.5, 5.0, 4.5, "IEA"],
        ["Diesel (Maldives)", "MV", "Conservative", "USD/GJ", 24.5, 21.0, 20.0, 20.0, 20.0, 20.0, 20.0, "STO Maldives"],
        ["Diesel (Maldives)", "MV", "Moderate",     "USD/GJ", 24.5, 21.0, 19.5, 18.5, 17.5, 16.5, 15.5, "STO Maldives"],
        ["Diesel (Maldives)", "MV", "Ambitious",    "USD/GJ", 24.5, 18.0, 14.0, 10.0, 9.0, 8.0, 7.5, "STO Maldives"],
        ["Uranium", "IN,BD,LK", "All", "USD/GJ", 0.56, 0.56, 0.56, 0.56, 0.56, 0.56, 0.56, "WNA"],
        ["Biomass", "IN,LK", "Conservative", "USD/GJ", 4.0, 4.0, 4.2, 4.4, 4.6, 4.8, 5.0, "IEA"],
        ["Biomass", "IN,LK", "Moderate",     "USD/GJ", 4.0, 4.0, 4.1, 4.2, 4.3, 4.4, 4.5, "IEA"],
        ["Biomass", "IN,LK", "Ambitious",    "USD/GJ", 4.0, 4.0, 3.8, 3.6, 3.4, 3.2, 3.0, "IEA"],
        ["Waste / MSW", "IN", "All", "USD/GJ", 0, 0, 0, 0, 0, 0, 0, "Assumed zero"],
    ]
    for r in rows:
        ws.append(r)
    print(f"  Added Fuel_Costs ({len(rows)} rows)")


def add_scenario_rules(wb):
    """Add Scenario_Rules sheet.

    Columns:
      Scenario      — scenario name this rule applies to (BAU / Green / NoCaps / ...)
      Rule_Type     — one of: clear | cap | set_value
      Parameter     — target parameter name (e.g. TotalAnnualMaxCapacity, CapitalCost)
      Tech_Filter   — tech code or glob-like prefix match (e.g. 'PWRSPVBDXX01', 'PWRSPV*')
      Value         — numeric value used by 'cap' and 'set_value' (ignored by 'clear')
      Notes         — free-text provenance/explanation

    Rules apply AFTER compute/merge, BEFORE write. See apply_scenario_rules()
    in A3_update_from_datapackage_v3.py.
    """
    ws = wb.create_sheet("Scenario_Rules")
    header = ["Scenario", "Rule_Type", "Parameter", "Tech_Filter", "Value", "Notes"]
    ws.append(header)
    # Example rows — team edits these to drive per-scenario behaviour.
    rows = [
        ["Green",   "clear",     "TotalAnnualMaxCapacity", "PWRCOA*",     None,    "Green: no new coal caps"],
        ["Green",   "cap",       "TotalAnnualMaxCapacity", "PWRSPV*",     10000.0, "Green: 10 GW/yr solar build"],
        ["NoCaps",  "clear",     "TotalAnnualMaxCapacity", "*",           None,    "NoCaps: remove all upper activity limits"],
        ["NoCaps",  "clear",     "TotalAnnualMinCapacity", "*",           None,    "NoCaps: remove all lower activity limits"],
        ["BAU",     "set_value", "CapitalCost",            "PWRSPVBDXX01", 900.0,  "BAU: override solar CAPEX (BD)"],
    ]
    for r in rows:
        ws.append(r)
    print(f"  Added Scenario_Rules ({len(rows)} example rows)")


def add_demand_assumptions(wb):
    """Add Demand_Assumptions sheet with milestone demand values."""
    ws = wb.create_sheet("Demand_Assumptions")
    header = ["Country_Node", "Fuel_Code", "Unit",
              2023, 2025, 2030, 2035, 2040, 2045, 2050,
              "Extrap_Rate", "Source"]
    ws.append(header)
    # Values from current A3's hardcoded readers + Energy Statistics files
    # These are PJ values (already converted)
    rows = [
        ["BD", "ELCBDXX02", "PJ",
         310.5, None, 498.0, None, 875.0, None, 1270.0,
         0.038, "IEPMP 2023"],
        ["BT", "ELCBTXX02", "PJ",
         20.5, None, 29.0, None, 47.0, None, 77.0,
         0.050, "PSMP 2040"],
        ["IN_N", "ELCIN_NXX02", "PJ",
         1555.2, None, 2167.1, None, 3358.3, None, 4590.0,
         0.049, "CEA NEP / Long Term Demand"],
        ["IN_S", "ELCIN_SXX02", "PJ",
         1267.2, None, 1777.6, None, 2770.5, None, 3756.0,
         0.049, "CEA NEP / Long Term Demand"],
        ["IN_E", "ELCIN_EXX02", "PJ",
         666.0, None, 924.0, None, 1422.0, None, 1926.0,
         0.049, "CEA NEP / Long Term Demand"],
        ["IN_W", "ELCIN_WXX02", "PJ",
         1744.8, None, 2434.8, None, 3789.0, None, 5148.0,
         0.049, "CEA NEP / Long Term Demand"],
        ["IN_NE", "ELCIN_NEXX02", "PJ",
         180.0, None, 253.8, None, 396.0, None, 540.0,
         0.049, "CEA NEP / Long Term Demand"],
        ["LK", "ELCLKXX02", "PJ",
         63.0, None, 93.6, None, 140.4, None, 194.4,
         0.038, "LTGEP 2023"],
        ["MV", "ELCMVXX02", "PJ",
         4.6, None, 8.6, None, None, None, None,
         0.050, "MCCEE Road Map"],
        ["NP", "ELCNPXX02", "PJ",
         36.0, None, 64.8, None, 115.2, None, 180.0,
         0.050, "CES Reference 7% GDP"],
    ]
    for r in rows:
        ws.append(r)
    print(f"  Added Demand_Assumptions ({len(rows)} rows)")


def main():
    if len(sys.argv) < 3:
        print("Usage: python A3_setup_template_v3.py <input_template.xlsx> <output.xlsx>")
        sys.exit(1)

    src = Path(sys.argv[1])
    dst = Path(sys.argv[2])

    print(f"Copying {src} -> {dst}")
    shutil.copy2(src, dst)

    print("Adding new sheets...")
    wb = openpyxl.load_workbook(dst)

    # Remove sheets if they already exist (idempotent)
    for name in ["Node_Mapping", "Country_Multipliers", "Fuel_Costs",
                 "Demand_Assumptions", "Scenario_Rules"]:
        if name in wb.sheetnames:
            del wb[name]

    add_node_mapping(wb)
    add_country_multipliers(wb)
    add_fuel_costs(wb)
    add_demand_assumptions(wb)
    add_scenario_rules(wb)

    wb.save(dst)
    print(f"\nDone. Template saved to: {dst}")


if __name__ == "__main__":
    main()
