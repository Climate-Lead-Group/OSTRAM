#!/usr/bin/env python3
"""
diagnose_reserve_margin.py

Pre-measurement diagnostic for the OSTRAM ReserveMargin parametrization.
Tests the hypothesis: is ReserveMargin (combined with the Tag-Fuel and
Tag-Technology coefficients) forcing structurally infeasible build volumes
INDEPENDENT of demand magnitude?

Reads:
  - A-O_Parametrization.xlsx     (sheet 'System Parameters' -> ReserveMargin)
  - Pre_processed_BAU_0.txt      (otoole-style .txt with all RM params +
                                  SpecifiedAnnualDemand, SpecifiedDemandProfile,
                                  YearSplit, CapacityToActivityUnit)

Optional overrides (so this script can be re-run against the parallel
demand-chat's findings or alternative fleet baselines):
  --demand-override demand.json     -> {fuel: {year: PJ}}
  --actual-fleet    fleet.json      -> {country: GW_2023}

Produces a sectioned console report:

  A. ReserveMargin trajectory (xlsx vs .txt cross-check + flag rule)
  B. ReserveMarginTagTechnology distribution by tech class
  C. ReserveMarginTagFuel distribution by fuel/country
  D. Implied minimum firm-capacity requirement per country x year
     (peak demand x RMTagFuel x ReserveMargin / CTA)
     Decomposed by binding timeslice and by country.
  E. Comparison vs actual 2023 fleet (built-in defaults or --actual-fleet)
  F. Verdict: is reserve-margin parametrization the structural cause?

Self-test (-T / --self-test): builds synthetic inputs in a temp dir, runs
the full pipeline, and asserts the numerical results against a hand-
computed expected. Run this BEFORE trusting real-data output.

Read-only on inputs. Only writes --report-json (if specified) and the temp
files created by --self-test.

Usage (Windows / Anaconda Prompt / cmd):

  python diagnose_reserve_margin.py ^
      --xlsx A-O_Parametrization.xlsx ^
      --txt  Pre_processed_BAU_0.txt ^
      --report-json rm_diag.json

Self-test:
  python diagnose_reserve_margin.py --self-test
"""

from __future__ import annotations
import argparse
import json
import re
import sys
import tempfile
from collections import defaultdict
from pathlib import Path


# ----------------------------------------------------------------------------
# Country / tech-class extraction (matches diagnose_feasopt_run.py conventions)
# ----------------------------------------------------------------------------

# OSTRAM tech codes: PWRXXXYYYZZ (11 chars) -- e.g. PWRWONNPLXX -> WON, NPLXX
# OSTRAM fuel codes: ELCXXXYYNN  (10 chars) -- e.g. ELCNPLXX01  -> NPLXX

def extract_country_from_tech(tech: str | None) -> str | None:
    if not tech: return None
    if len(tech) >= 11 and tech[:3] in {"PWR", "DSP", "MIN", "REF", "ENG"}:
        return tech[6:11]
    return None

def extract_country_from_fuel(fuel: str | None) -> str | None:
    if not fuel: return None
    if len(fuel) >= 8 and fuel[:3] == "ELC":
        return fuel[3:8]
    return None

def tech_class_of(tech: str) -> str:
    """PWRWONNPLXX -> WON, PWRSDSLKAXX -> SDS, PWRHYDINDNO -> HYD."""
    if len(tech) >= 6 and tech[:3] == "PWR":
        return tech[3:6]
    return tech


# ----------------------------------------------------------------------------
# .txt parser (otoole AMPL-style param blocks)
# ----------------------------------------------------------------------------

PARAM_HEADER_RE = re.compile(r"^\s*param\s+default\s+(\S+)\s*:\s*(\w+)\s*:=\s*$")

def parse_param_blocks(txt_path: Path, wanted: set[str]) -> dict[str, list[tuple]]:
    """Stream the .txt; return {param_name: [(field1, field2, ..., value_float), ...]}.

    Only blocks whose name appears in `wanted` are kept. Each row is split on
    whitespace; the last token is parsed as float, the rest are kept as strings.
    Defaults from the param header are stored under key '__defaults__'.
    """
    out: dict[str, list[tuple]] = {}
    defaults: dict[str, float] = {}
    current = None
    with open(txt_path, encoding="utf-8") as f:
        for raw in f:
            line = raw.rstrip("\r\n")
            m = PARAM_HEADER_RE.match(line)
            if m:
                default_val_str, name = m.group(1), m.group(2)
                if name in wanted:
                    current = name
                    out.setdefault(name, [])
                    try:
                        defaults[name] = float(default_val_str)
                    except ValueError:
                        defaults[name] = 0.0
                else:
                    current = None
                continue
            if current is None:
                continue
            stripped = line.strip()
            if stripped == ";":
                current = None
                continue
            if not stripped:
                continue
            parts = stripped.split()
            try:
                val = float(parts[-1])
            except ValueError:
                continue
            out[current].append(tuple(parts[:-1]) + (val,))
    out["__defaults__"] = defaults  # type: ignore[assignment]
    return out


# ----------------------------------------------------------------------------
# A-O_Parametrization.xlsx reader (System Parameters sheet, ReserveMargin row)
# ----------------------------------------------------------------------------

def read_reservemargin_from_xlsx(xlsx_path: Path) -> dict[int, float]:
    """Read 'System Parameters' sheet, return {year: ReserveMargin_value}.

    The sheet layout is:
        row 1: ['Parameter', 'Unit', 2023, 2024, ..., 2050]
        row 2+: [param_name, unit, val_2023, val_2024, ...]
    We pick up the row whose Parameter == 'ReserveMargin'.
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if "System Parameters" not in wb.sheetnames:
        raise RuntimeError(
            f"Sheet 'System Parameters' not found in {xlsx_path}. "
            f"Available: {wb.sheetnames}"
        )
    ws = wb["System Parameters"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise RuntimeError("System Parameters sheet is empty")
    header = rows[0]
    year_cols: list[tuple[int, int]] = []
    for i, h in enumerate(header):
        if isinstance(h, int) and 2000 <= h <= 2100:
            year_cols.append((i, h))
        elif isinstance(h, str) and h.strip().isdigit():
            y = int(h.strip())
            if 2000 <= y <= 2100:
                year_cols.append((i, y))
    out: dict[int, float] = {}
    for r in rows[1:]:
        if not r: continue
        pname = (r[0] or "").strip() if isinstance(r[0], str) else r[0]
        if pname != "ReserveMargin":
            continue
        for col_i, year in year_cols:
            v = r[col_i] if col_i < len(r) else None
            if v is None: continue
            try:
                out[year] = float(v)
            except (TypeError, ValueError):
                continue
        break
    if not out:
        raise RuntimeError(
            "No ReserveMargin row found in 'System Parameters'. "
            "Expected a row whose first cell == 'ReserveMargin'."
        )
    return out


# ----------------------------------------------------------------------------
# Indexing helpers from parsed .txt
# ----------------------------------------------------------------------------

def index_reserve_margin(rows: list[tuple]) -> dict[tuple[str, int], float]:
    """rows: [(region, year_str, value), ...] -> {(region, year_int): val}"""
    out = {}
    for r in rows:
        if len(r) != 3: continue
        region, year, val = r
        try:
            out[(region, int(year))] = val
        except (TypeError, ValueError):
            continue
    return out

def index_rmtag_fuel(rows: list[tuple]) -> dict[tuple[str, str, int], float]:
    """rows: [(region, fuel, year, val), ...] -> {(region, fuel, year): val}"""
    out = {}
    for r in rows:
        if len(r) != 4: continue
        region, fuel, year, val = r
        try:
            out[(region, fuel, int(year))] = val
        except (TypeError, ValueError):
            continue
    return out

def index_rmtag_tech(rows: list[tuple]) -> dict[tuple[str, str, int], float]:
    """Same shape as fuel."""
    return index_rmtag_fuel(rows)

def index_specified_demand(rows: list[tuple]) -> dict[tuple[str, str, int], float]:
    """rows: [(region, fuel, year, val), ...]"""
    return index_rmtag_fuel(rows)

def index_demand_profile(rows: list[tuple]) -> dict[tuple[str, str, str, int], float]:
    """rows: [(region, fuel, ts, year, val), ...]"""
    out = {}
    for r in rows:
        if len(r) != 5: continue
        region, fuel, ts, year, val = r
        try:
            out[(region, fuel, ts, int(year))] = val
        except (TypeError, ValueError):
            continue
    return out

def index_yearsplit(rows: list[tuple]) -> dict[tuple[str, int], float]:
    """rows: [(ts, year, val), ...]"""
    out = {}
    for r in rows:
        if len(r) != 3: continue
        ts, year, val = r
        try:
            out[(ts, int(year))] = val
        except (TypeError, ValueError):
            continue
    return out

def index_cta(rows: list[tuple]) -> dict[tuple[str, str], float]:
    """rows: [(region, tech, val), ...]"""
    out = {}
    for r in rows:
        if len(r) != 3: continue
        region, tech, val = r
        out[(region, tech)] = val
    return out


# ----------------------------------------------------------------------------
# Implied firm-capacity computation
# ----------------------------------------------------------------------------

def compute_implied_firm_capacity(
    reservemargin: dict[tuple[str, int], float],
    rmtag_fuel:    dict[tuple[str, str, int], float],
    sad:           dict[tuple[str, str, int], float],
    sdp:           dict[tuple[str, str, str, int], float],
    yearsplit:     dict[tuple[str, int], float],
    cta_default:   float,
    rm_default:    float,
    rmtag_fuel_default: float,
    region: str = "GLOBAL",
) -> dict:
    """For each year, compute the implied minimum firm-capacity requirement
    (in GW) given the OSeMOSYS reserve-margin formulation:

        DemandNeedingReserveMargin[r,l,y] =
            sum_f( RateOfDemand[r,l,f,y] * ReserveMarginTagFuel[r,f,y] )
        RateOfDemand[r,l,f,y] = SAD[r,f,y] * SDP[r,f,l,y] / YearSplit[l,y]
        Constraint: TotalCapacityInReserveMargin[r,y] >= DNRM[r,l,y] * ReserveMargin[r,y]
        TotalCapacityInReserveMargin[r,y] = sum_t( Cap[r,t,y] * RMTagTech[r,t,y] * CTA[r,t] )

    The constraint binds at the timeslice maximizing DNRM. The implied
    firm-capacity floor (in GW, dividing through by a uniform CTA) is the
    binding-timeslice DNRM * ReserveMargin / CTA.

    Returns dict keyed by year:
        {
          year: {
            "binding_ts":         str,
            "dnrm_max":           float,    # PJ/yr-equivalent rate
            "reserve_margin":     float,
            "rhs_pjyr":           float,    # DNRM * RM (PJ/yr-equivalent)
            "firm_cap_required_gw": float,  # rhs / CTA
            "by_country":         {country: {"dnrm_share": float,
                                             "firm_share_gw": float}},
          }
        }
    """
    # Discover the years and timeslices present
    years = sorted({y for (_, y) in reservemargin.keys()} | {y for (_, _, y) in sad.keys()})
    timeslices = sorted({ts for (ts, _) in yearsplit.keys()} | {ts for (_, _, ts, _) in sdp.keys()})
    fuels = sorted({f for (_, f, _) in sad.keys()})

    out: dict[int, dict] = {}
    for y in years:
        rm = reservemargin.get((region, y), rm_default)
        # Compute DNRM for each timeslice, then take max
        dnrm_per_ts: dict[str, float] = {}
        contrib_per_ts_country: dict[str, dict[str, float]] = {}
        for ts in timeslices:
            ys = yearsplit.get((ts, y))
            if ys is None or ys <= 0:
                continue
            total = 0.0
            by_c: dict[str, float] = defaultdict(float)
            for f in fuels:
                sad_v = sad.get((region, f, y), 0.0)
                if sad_v == 0.0: continue
                sdp_v = sdp.get((region, f, ts, y), 0.0)
                if sdp_v == 0.0: continue
                tag = rmtag_fuel.get((region, f, y), rmtag_fuel_default)
                rate = sad_v * sdp_v / ys                  # PJ/yr-equivalent
                contrib = rate * tag
                total += contrib
                country = extract_country_from_fuel(f) or "(unknown)"
                by_c[country] += contrib
            dnrm_per_ts[ts] = total
            contrib_per_ts_country[ts] = dict(by_c)
        if not dnrm_per_ts:
            continue
        binding_ts = max(dnrm_per_ts, key=dnrm_per_ts.get)
        dnrm_max = dnrm_per_ts[binding_ts]
        rhs = dnrm_max * rm
        firm_gw = rhs / cta_default if cta_default > 0 else float("nan")
        # By-country at binding timeslice
        by_country = {}
        binding_contribs = contrib_per_ts_country.get(binding_ts, {})
        for c, share in binding_contribs.items():
            by_country[c] = {
                "dnrm_share":    share,
                "firm_share_gw": (share * rm) / cta_default if cta_default > 0 else float("nan"),
            }
        out[y] = {
            "binding_ts":          binding_ts,
            "dnrm_max":            dnrm_max,
            "reserve_margin":      rm,
            "rhs_pjyr":            rhs,
            "firm_cap_required_gw": firm_gw,
            "by_country":          by_country,
        }
    return out


# ----------------------------------------------------------------------------
# Reports
# ----------------------------------------------------------------------------

def fmt_num(x):
    if x is None: return "—"
    if isinstance(x, str): return x
    try:
        if abs(x) >= 1000: return f"{x:,.0f}"
        if abs(x) >= 1:    return f"{x:,.3f}"
        return f"{x:.5f}"
    except (TypeError, ValueError):
        return str(x)


def report_a_reservemargin(
    rm_xlsx: dict[int, float],
    rm_txt:  dict[tuple[str, int], float],
    flag_threshold: float,
):
    print("=" * 78)
    print("A. RESERVEMARGIN TRAJECTORY (xlsx vs .txt cross-check)")
    print("=" * 78)
    # OSeMOSYS convention note
    print("  Convention: in OSeMOSYS the value V in ReserveMargin is the multiplier:")
    print("              firm capacity >= V * peak fuel-demand-needing-reserve.")
    print(f"              V=1.15 -> 15% reserve. V=1.30 -> 30% (your flag).")
    print()
    txt_regions = sorted({r for (r, _) in rm_txt.keys()})
    txt_years   = sorted({y for (_, y) in rm_txt.keys()})
    xlsx_years  = sorted(rm_xlsx.keys())
    print(f"  xlsx years: {xlsx_years[0]}-{xlsx_years[-1]} ({len(xlsx_years)} entries)" if xlsx_years else "  xlsx: (no years parsed)")
    print(f"  .txt regions: {txt_regions}")
    print(f"  .txt years:   {txt_years[0]}-{txt_years[-1]} ({len(txt_years)} entries)" if txt_years else "  .txt: (no entries)")

    # Cross-check
    print()
    print("  Consistency between xlsx and .txt (per year, GLOBAL region):")
    diffs = []
    common = sorted(set(xlsx_years) & set(txt_years))
    for y in common:
        v_x = rm_xlsx.get(y)
        v_t = rm_txt.get(("GLOBAL", y))
        if v_x is None or v_t is None: continue
        if abs(v_x - v_t) > 1e-9:
            diffs.append((y, v_x, v_t))
    if not diffs:
        print(f"    OK: all {len(common)} common years agree (xlsx == .txt).")
    else:
        print(f"    !! {len(diffs)} year(s) differ between xlsx and .txt:")
        for y, vx, vt in diffs:
            print(f"      {y}: xlsx={fmt_num(vx)} | .txt={fmt_num(vt)}")

    # Trajectory and flags
    print()
    print(f"  Trajectory (.txt, by region) -- flagged values: > {flag_threshold:.2f}")
    print(f"    {'REGION':<10} {'YEAR':>6} {'VALUE':>10} {'FLAG':<6}")
    flagged = 0
    for (r, y), v in sorted(rm_txt.items()):
        flag = "FLAG" if v > flag_threshold else ""
        if flag: flagged += 1
        # Compress: only print every 5th year unless flagged
        if flag or y % 5 == 0 or y in (min(txt_years), max(txt_years)):
            print(f"    {r:<10} {y:>6} {fmt_num(v):>10} {flag:<6}")
    print(f"    -> {flagged} (region, year) entries above threshold.")
    if flagged == 0 and rm_txt:
        max_v = max(rm_txt.values())
        if max_v <= 1.20:
            print(f"    -> Max value {fmt_num(max_v)} = {(max_v - 1.0) * 100:.0f}% reserve. NORMAL — RM itself is not a likely structural cause.")
        elif max_v <= 1.30:
            print(f"    -> Max value {fmt_num(max_v)} = {(max_v - 1.0) * 100:.0f}% reserve. Aggressive but plausible.")
        else:
            print(f"    -> Max value {fmt_num(max_v)} = {(max_v - 1.0) * 100:.0f}% reserve. Suspect.")


# Heuristic firmness expectations for tech classes (real-world physics-based)
EXPECTED_FIRMNESS = {
    # tech_class: (lo, hi, "rationale")
    "BIO": (0.7, 0.95, "biomass = dispatchable thermal"),
    "COA": (0.85, 0.95, "coal = baseload"),
    "NGS": (0.85, 0.98, "gas = peaking/baseload, very firm"),
    "OIL": (0.85, 0.98, "oil = peaking, very firm"),
    "PET": (0.85, 0.98, "petroleum = peaking, very firm"),
    "URN": (0.85, 0.98, "nuclear = baseload"),
    "OTH": (0.7,  0.95, "other thermal"),
    "WAS": (0.6,  0.9,  "waste-to-energy"),
    "COG": (0.5,  0.8,  "cogeneration depends on host process"),
    "CCS": (0.85, 0.95, "CCS = thermal baseload with capture"),
    "HYD": (0.5,  0.9,  "large hydro = dispatchable when reservoir filled"),
    "SHP": (0.2,  0.5,  "small/run-of-river hydro = limited firm credit"),
    "GEO": (0.85, 0.95, "geothermal = baseload"),
    "CSP": (0.2,  0.5,  "CSP firm credit depends on thermal storage"),
    "SPV": (0.0,  0.2,  "solar PV = no firm credit at peak (unless storage)"),
    "UPV": (0.0,  0.2,  "utility PV = no firm credit at peak (unless storage)"),
    "WON": (0.05, 0.2,  "onshore wind = minimal firm credit"),
    "WOF": (0.1,  0.3,  "offshore wind = minimal firm credit"),
    "SDS": (0.3,  0.9,  "short-duration storage = firm only if energy avail at peak"),
    "LDS": (0.5,  0.95, "long-duration storage = firm if dispatch policy aligned"),
}

def report_b_tech_tags(rmtag_tech: dict[tuple[str, str, int], float]):
    print()
    print("=" * 78)
    print("B. RESERVEMARGINTAGTECHNOLOGY DISTRIBUTION")
    print("=" * 78)
    if not rmtag_tech:
        print("  (No RMTagTechnology entries.)")
        return

    # Per-tech: distinct values across years (catch time-varying tags)
    tech_values: dict[str, set[float]] = defaultdict(set)
    for (_, t, _), v in rmtag_tech.items():
        tech_values[t].add(v)
    techs_with_var = {t: vals for t, vals in tech_values.items() if len(vals) > 1}

    # By tech class
    class_to_value: dict[str, dict[float, list[str]]] = defaultdict(lambda: defaultdict(list))
    for tech, vals in tech_values.items():
        cls = tech_class_of(tech)
        if len(vals) == 1:
            class_to_value[cls][next(iter(vals))].append(tech)

    # Distinct value histogram
    val_count = defaultdict(int)
    for vals in tech_values.values():
        for v in vals:
            val_count[v] += 1  # per-tech presence (not per cell)
    print(f"  Distinct tag values among {len(tech_values)} techs:")
    print(f"    {'VALUE':>6}  {'TECH COUNT':>12}")
    for v in sorted(val_count.keys()):
        print(f"    {v:>6}  {val_count[v]:>12}")

    print()
    print("  By tech class (showing distinct tagged value(s) and example techs):")
    print(f"    {'CLASS':<6} {'VALUE(S)':<20} {'#TECHS':>7}  {'EXPECTED':<14}  {'FLAG':<6}")
    flags_class = []
    for cls in sorted(class_to_value.keys()):
        for val in sorted(class_to_value[cls].keys()):
            techs = class_to_value[cls][val]
            exp = EXPECTED_FIRMNESS.get(cls)
            if exp is None:
                expected_str = "(no rule)"
                flag = ""
            else:
                lo, hi, _why = exp
                expected_str = f"[{lo:.2f}, {hi:.2f}]"
                if val < lo:
                    flag = "LOW"
                    flags_class.append((cls, val, lo, hi, _why))
                elif val > hi:
                    flag = "HIGH"
                    flags_class.append((cls, val, lo, hi, _why))
                else:
                    flag = ""
            print(f"    {cls:<6} {fmt_num(val):<20} {len(techs):>7}  {expected_str:<14}  {flag:<6}")

    if flags_class:
        print()
        print("  !! Tech-class tag flags (value outside expected firmness range):")
        for cls, val, lo, hi, why in flags_class:
            direction = "below" if val < lo else "above"
            print(f"    {cls}: tagged at {fmt_num(val)}, expected [{lo:.2f}, {hi:.2f}] ({why}).")
            print(f"          -> {direction} expectation: see remediation table at end.")

    # Techs missing from RM-tag table altogether (defaulting to 0)
    # We can't enumerate them here without the SET TECHNOLOGY list; flag in report D once we have demand fuels.
    if techs_with_var:
        print()
        print(f"  !! {len(techs_with_var)} tech(s) have time-varying RMTagTechnology values:")
        for t, vals in list(techs_with_var.items())[:10]:
            print(f"    {t}: values = {sorted(vals)}")


def report_c_fuel_tags(rmtag_fuel: dict[tuple[str, str, int], float], sad: dict[tuple[str, str, int], float]):
    print()
    print("=" * 78)
    print("C. RESERVEMARGINTAGFUEL DISTRIBUTION")
    print("=" * 78)
    if not rmtag_fuel:
        print("  (No RMTagFuel entries -- all fuels default to 0, RM constraint trivially satisfied.)")
        return
    fuel_values: dict[str, set[float]] = defaultdict(set)
    for (_, fl, _), v in rmtag_fuel.items():
        fuel_values[fl].add(v)
    print(f"  Tagged fuels: {len(fuel_values)}")
    print(f"  Distinct values: {sorted({v for vs in fuel_values.values() for v in vs})}")

    # By country
    print()
    print("  By country (electricity fuels only):")
    country_to_fuels: dict[str, list[tuple[str, list[float]]]] = defaultdict(list)
    for fl, vals in fuel_values.items():
        c = extract_country_from_fuel(fl) or "(unknown)"
        country_to_fuels[c].append((fl, sorted(vals)))
    print(f"    {'COUNTRY':<10} {'FUEL':<14} {'VALUE(S)':<20} {'2023 SAD (PJ)':>14}")
    for c in sorted(country_to_fuels.keys()):
        for fl, vals in sorted(country_to_fuels[c]):
            sad_2023 = sad.get(("GLOBAL", fl, 2023), 0.0)
            print(f"    {c:<10} {fl:<14} {str(vals):<20} {fmt_num(sad_2023):>14}")

    # Fuels with demand but no tag -> contribute zero to RM RHS (silent under-coverage)
    demand_fuels = {fl for (_, fl, _), v in sad.items() if v != 0.0}
    tagged_fuels = set(fuel_values.keys())
    untagged = demand_fuels - tagged_fuels
    intersection = demand_fuels & tagged_fuels
    if untagged and not intersection:
        print()
        print("  *** PATHOLOGY: tagged-fuel set and demand-fuel set are DISJOINT ***")
        print(f"      Tagged fuels (RM constraint scope): {sorted(tagged_fuels)}")
        print(f"      Fuels with non-zero demand:          {sorted(demand_fuels)}")
        print(f"      No overlap -> the RM constraint RHS is zero for every (region, year).")
        print(f"      RM is structurally inactive in this model run. (See Section F.)")
    elif untagged:
        print()
        print(f"  !! {len(untagged)} demand fuel(s) have NO RMTagFuel entry (default 0):")
        for fl in sorted(untagged):
            print(f"    {fl}")
        print(f"     ({len(intersection)} demand fuel(s) ARE tagged.)")


def report_d_implied_capacity(implied: dict, year_focus: list[int]):
    print()
    print("=" * 78)
    print("D. IMPLIED MINIMUM FIRM-CAPACITY REQUIREMENT (from RM constraint)")
    print("=" * 78)
    if not implied:
        print("  (Could not compute — missing demand or yearsplit data.)")
        return
    print("  Method: for each year y, find the timeslice maximizing")
    print("          DNRM = sum_f( SAD[f,y] * SDP[f,ts,y] / YearSplit[ts,y] * RMTagFuel[f,y] )")
    print("          then RHS_y = DNRM * ReserveMargin[y], firm_GW = RHS_y / CTA (=31.536).")
    print("          This is a HARD floor on weighted-firm capacity even before any")
    print("          energy-balance constraint.")
    print()
    # Detect identically-zero RHS and shout
    max_dnrm_all = max((d.get("dnrm_max", 0.0) for d in implied.values()), default=0.0)
    if max_dnrm_all < 1e-6:
        print("  *** DNRM is IDENTICALLY ZERO across all years ***")
        print("      The reserve-margin RHS is zero -> the RM constraint is vacuous.")
        print("      (This is NOT because demand is zero -- it's because RMTagFuel and")
        print("       SpecifiedAnnualDemand point at disjoint fuel sets. See Section C/F.)")
        print()
    print(f"  Per-year summary (showing focus years {year_focus}):")
    print(f"    {'YEAR':>6} {'BIND.TS':<6} {'RM':>6} {'DNRM_MAX (PJ/yr)':>20} {'FIRM_GW':>12}")
    for y in sorted(implied.keys()):
        if year_focus and y not in year_focus and y % 5 != 0: continue
        d = implied[y]
        print(f"    {y:>6} {d['binding_ts']:<6} {fmt_num(d['reserve_margin']):>6} "
              f"{fmt_num(d['dnrm_max']):>20} {fmt_num(d['firm_cap_required_gw']):>12}")

    # Detailed by-country breakdown for focus years
    if year_focus:
        for y in year_focus:
            d = implied.get(y)
            if d is None: continue
            print()
            print(f"  Year {y} -- per-country breakdown at binding timeslice ({d['binding_ts']}):")
            print(f"    {'COUNTRY':<10} {'DNRM_share (PJ/yr)':>20} {'FIRM_GW_share':>14} "
                  f"{'%':>6}")
            total = sum(c["firm_share_gw"] for c in d["by_country"].values())
            for c, ent in sorted(d["by_country"].items(), key=lambda kv: -kv[1]["firm_share_gw"]):
                pct = (ent["firm_share_gw"] / total * 100) if total > 0 else 0.0
                print(f"    {c:<10} {fmt_num(ent['dnrm_share']):>20} "
                      f"{fmt_num(ent['firm_share_gw']):>14} {pct:>5.1f}%")
            print(f"    {'TOTAL':<10} {'':<20} {fmt_num(total):>14}")


# Built-in 2023 fleet snapshot (rough, public sources). Override via --actual-fleet.
DEFAULT_ACTUAL_FLEET_GW = {
    "BGDXX": 26.0,    # Bangladesh
    "BTNXX": 2.4,     # Bhutan
    "INDEA": 70.0,    # India East
    "INDNE": 11.0,    # India NE
    "INDNO": 110.0,   # India North
    "INDSO": 120.0,   # India South
    "INDWE": 110.0,   # India West
    "LKAXX": 5.0,     # Sri Lanka
    "MDVXX": 0.4,     # Maldives
    "NPLXX": 2.5,     # Nepal
}

def report_e_actual_fleet_compare(implied: dict, fleet: dict[str, float], compare_year: int = 2023):
    print()
    print("=" * 78)
    print(f"E. COMPARISON vs ACTUAL {compare_year} FLEET")
    print("=" * 78)
    d = implied.get(compare_year)
    if d is None:
        print(f"  (No implied result for {compare_year}.)")
        return
    print(f"  {'COUNTRY':<10} {'IMPLIED FIRM (GW)':>20} {'ACTUAL TOTAL (GW)':>20} {'RATIO':>8} {'INTERP':<28}")
    flags = []
    for c in sorted(set(d["by_country"].keys()) | set(fleet.keys())):
        impl = d["by_country"].get(c, {}).get("firm_share_gw", 0.0)
        actual = fleet.get(c)
        if actual is None:
            interp = "(no actual reference)"
            ratio_str = "—"
        elif actual <= 0:
            interp = "(actual=0)"
            ratio_str = "—"
        else:
            ratio = impl / actual
            ratio_str = f"{ratio:.2f}"
            if ratio > 1.5:
                interp = "IMPLIED >> ACTUAL: RM forces overbuild"
                flags.append((c, ratio))
            elif ratio > 1.0:
                interp = "implied > actual: tight"
            elif ratio > 0.5:
                interp = "comfortable headroom"
            else:
                interp = "loose"
        print(f"  {c:<10} {fmt_num(impl):>20} {fmt_num(actual) if actual is not None else '—':>20} "
              f"{ratio_str:>8} {interp:<28}")
    if flags:
        print()
        print(f"  !! {len(flags)} country(-ies) where implied firm > 1.5x actual fleet:")
        for c, r in flags:
            print(f"    {c}: ratio={r:.2f} (RM is forcing overbuild here)")


def report_f_verdict(rm_txt, rmtag_tech, rmtag_fuel, sad, implied, fleet, flag_threshold):
    print()
    print("=" * 78)
    print("F. VERDICT")
    print("=" * 78)
    rm_max = max(rm_txt.values()) if rm_txt else None
    rm_clean = (rm_max is not None and rm_max <= flag_threshold)

    # Detect the most damning structural pathology: RMTagFuel and SAD on
    # disjoint fuel sets. If true, RM constraint is identically vacuous.
    tagged_fuels  = {f for (_, f, _) in rmtag_fuel.keys()}
    demand_fuels  = {f for (_, f, _), v in sad.items() if v != 0.0}
    intersection  = tagged_fuels & demand_fuels
    rm_inactive   = (len(tagged_fuels) > 0 and len(demand_fuels) > 0 and len(intersection) == 0)
    # Also detect: max DNRM across all years ~ 0
    max_dnrm = max((d.get("dnrm_max", 0.0) for d in implied.values()), default=0.0)
    rm_dnrm_zero = max_dnrm < 1e-6

    if rm_inactive or rm_dnrm_zero:
        print()
        print("  ************************************************************")
        print("  *  HEADLINE FINDING: ReserveMargin constraint is INACTIVE  *")
        print("  ************************************************************")
        print()
        print(f"  RMTagFuel covers {len(tagged_fuels)} fuel(s); SpecifiedAnnualDemand has")
        print(f"  non-zero values on {len(demand_fuels)} fuel(s); intersection = {len(intersection)} fuel(s).")
        print(f"  Max DNRM across all 28 years: {fmt_num(max_dnrm)} PJ/yr.")
        print()
        print("  Implication: the RHS of the RM constraint is identically zero. RM is")
        print("  not binding any build decision in OSTRAM right now -- it's a no-op.")
        print()
        print("  This RULES OUT the user's hypothesis that RM is forcing the 12,000 GW")
        print("  build. Whatever is driving the LP toward that volume is NOT RM.")
        print()
        if tagged_fuels and demand_fuels and not intersection:
            print("  Surgical fix to make RM actually bite (when needed):")
            print(f"    RMTagFuel is on:           {sorted(tagged_fuels)[:3]}{'...' if len(tagged_fuels) > 3 else ''}")
            print(f"    SpecifiedAnnualDemand on:  {sorted(demand_fuels)[:3]}{'...' if len(demand_fuels) > 3 else ''}")
            print("    -> Either move RMTagFuel onto the demand-side fuels (ELC...03),")
            print("       or add RMTagFuel entries on those fuels in addition to the")
            print("       upstream ones. Without this, RM constraint cannot constrain.")
            print()
            print("  Where the tag lives: RMTagFuel is generated upstream of the .txt")
            print("  by B2/preprocessing -- check SOASIA_OSeMOSYS_WV.xlsx and the A-O")
            print("  fuel-tagging logic. (Not visible in A-O_Parametrization.xlsx itself.)")
            print()
        print("  Caution: leaving RM inactive is fine for the current diagnostic, but")
        print("  any fix to other parameters that is supposed to interact with RM (e.g.")
        print("  storage credits) won't behave as intended until the tag is corrected.")
        print()
        print("  -- Supplementary checks (RM-value, tag-class, fleet) below for completeness --")
        print()
    # Tag-class flags: re-derive from the same logic as section B
    tech_values: dict[str, set[float]] = defaultdict(set)
    for (_, t, _), v in rmtag_tech.items():
        tech_values[t].add(v)
    class_to_value: dict[str, set[float]] = defaultdict(set)
    for tech, vals in tech_values.items():
        if len(vals) == 1:
            class_to_value[tech_class_of(tech)].add(next(iter(vals)))
    tag_flags = []
    for cls, vals in class_to_value.items():
        exp = EXPECTED_FIRMNESS.get(cls)
        if exp is None: continue
        lo, hi, _ = exp
        for v in vals:
            if v < lo or v > hi:
                tag_flags.append((cls, v, lo, hi))
    # Implied vs actual
    impl_2023 = implied.get(2023, {})
    overbuild_countries = []
    for c, ent in (impl_2023.get("by_country") or {}).items():
        impl = ent.get("firm_share_gw", 0.0)
        actual = fleet.get(c)
        if actual and actual > 0 and impl / actual > 1.5:
            overbuild_countries.append((c, impl / actual))

    print(f"  ReserveMargin trajectory:    {'CLEAN' if rm_clean else 'SUSPECT'}"
          f" (max value {fmt_num(rm_max)}, threshold {flag_threshold:.2f})")
    print(f"  Tech-tag class anomalies:    {len(tag_flags)} class/value pair(s) outside expected band")
    print(f"  Implied >> actual countries: {len(overbuild_countries)} country(-ies) at 2023 baseline")
    print()
    if rm_clean and not overbuild_countries:
        print("  ==> RM parametrization itself is NOT the structural cause of the 12,000 GW")
        print("      build. Look elsewhere (demand magnitude, lid distribution, residual")
        print("      capacity, technology costs).")
        if tag_flags:
            print()
            print("      However, the following tag inconsistencies should be fixed for hygiene:")
            for cls, v, lo, hi in tag_flags:
                print(f"        - {cls} tagged at {v}, expected [{lo:.2f}, {hi:.2f}]")
            print("      These either UNDER-credit firm sources (forcing modest overbuild)")
            print("      or OVER-credit storage/wind (under-building capacity).")
    elif rm_clean and overbuild_countries:
        print("  ==> RM value is normal but per-country implied firm requirement exceeds")
        print("      actual fleet by >1.5x in some countries. This is driven by demand")
        print("      magnitude in those countries, not by RM. Check the demand-chat results.")
    else:
        print("  ==> RM value itself is suspect. Surgical fixes:")
        print("        - Edit System Parameters sheet, ReserveMargin row, set values to ~1.15.")
        print("        - Ensure xlsx and .txt agree (re-run preprocessing if not).")
    print()
    print("  Surgical-fix suggestions for tag-class anomalies (sheet: not in A-O xlsx;")
    print("  RMTagTechnology lives in Pre_processed_BAU_0.txt only, generated upstream).")
    print("  Locate the upstream data source (likely SOASIA_OSeMOSYS_WV.xlsx or B2's")
    print("  intermediate). Apply class-by-class:")
    for cls, v, lo, hi in tag_flags:
        target = (lo + hi) / 2
        why = EXPECTED_FIRMNESS[cls][2]
        print(f"    {cls}: {fmt_num(v)} -> ~{fmt_num(target)}  ({why})")


# ----------------------------------------------------------------------------
# Self-test (synthetic inputs)
# ----------------------------------------------------------------------------

SELF_TEST_TXT = """\
# Synthetic self-test fixture (mimics otoole .txt format)
param default 1 : ReserveMargin :=
GLOBAL 2023 1.20
GLOBAL 2024 1.25
;
param default 0 : ReserveMarginTagFuel :=
GLOBAL ELCAAAXX01 2023 1.0
GLOBAL ELCAAAXX01 2024 1.0
GLOBAL ELCBBBXX01 2023 0.5
GLOBAL ELCBBBXX01 2024 0.5
;
param default 0 : ReserveMarginTagTechnology :=
GLOBAL PWRCOAAAAXX 2023 0.9
GLOBAL PWRWONAAAXX 2023 0.1
GLOBAL PWRHYDAAAXX 2023 0.3
;
param default 0 : SpecifiedAnnualDemand :=
GLOBAL ELCAAAXX01 2023 100.0
GLOBAL ELCAAAXX01 2024 110.0
GLOBAL ELCBBBXX01 2023 50.0
GLOBAL ELCBBBXX01 2024 55.0
;
param default 0 : SpecifiedDemandProfile :=
GLOBAL ELCAAAXX01 S1D1 2023 0.4
GLOBAL ELCAAAXX01 S2D1 2023 0.6
GLOBAL ELCAAAXX01 S1D1 2024 0.4
GLOBAL ELCAAAXX01 S2D1 2024 0.6
GLOBAL ELCBBBXX01 S1D1 2023 0.5
GLOBAL ELCBBBXX01 S2D1 2023 0.5
GLOBAL ELCBBBXX01 S1D1 2024 0.5
GLOBAL ELCBBBXX01 S2D1 2024 0.5
;
param default 0 : YearSplit :=
S1D1 2023 0.5
S2D1 2023 0.5
S1D1 2024 0.5
S2D1 2024 0.5
;
param default 1 : CapacityToActivityUnit :=
GLOBAL PWRCOAAAAXX 31.536
GLOBAL PWRWONAAAXX 31.536
GLOBAL PWRHYDAAAXX 31.536
;
"""

def write_synthetic_xlsx(path: Path):
    """Build a minimal A-O_Parametrization.xlsx with just the System Parameters sheet."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "System Parameters"
    ws.append(["Parameter", "Unit", 2023, 2024])
    ws.append(["ReserveMargin", "ratio", 1.20, 1.25])
    wb.save(path)

def run_self_test() -> int:
    print("Self-test: building synthetic inputs in temp dir ...")
    tmp = Path(tempfile.mkdtemp(prefix="rm_diag_test_"))
    txt = tmp / "Pre_processed_synthetic.txt"
    xlsx = tmp / "A-O_synthetic.xlsx"
    txt.write_text(SELF_TEST_TXT, encoding="utf-8")
    write_synthetic_xlsx(xlsx)
    print(f"  txt:  {txt}")
    print(f"  xlsx: {xlsx}")

    # Parse
    rm_xlsx = read_reservemargin_from_xlsx(xlsx)
    blocks = parse_param_blocks(txt, {
        "ReserveMargin", "ReserveMarginTagFuel", "ReserveMarginTagTechnology",
        "SpecifiedAnnualDemand", "SpecifiedDemandProfile", "YearSplit",
        "CapacityToActivityUnit",
    })
    rm_txt = index_reserve_margin(blocks["ReserveMargin"])
    rmtag_fuel = index_rmtag_fuel(blocks["ReserveMarginTagFuel"])
    rmtag_tech = index_rmtag_tech(blocks["ReserveMarginTagTechnology"])
    sad = index_specified_demand(blocks["SpecifiedAnnualDemand"])
    sdp = index_demand_profile(blocks["SpecifiedDemandProfile"])
    ys  = index_yearsplit(blocks["YearSplit"])
    cta = index_cta(blocks["CapacityToActivityUnit"])

    # Assertions on parsed values
    assert rm_xlsx == {2023: 1.20, 2024: 1.25}, f"xlsx RM read wrong: {rm_xlsx}"
    assert rm_txt[("GLOBAL", 2023)] == 1.20
    assert rm_txt[("GLOBAL", 2024)] == 1.25
    assert rmtag_fuel[("GLOBAL", "ELCAAAXX01", 2023)] == 1.0
    assert rmtag_fuel[("GLOBAL", "ELCBBBXX01", 2023)] == 0.5
    assert rmtag_tech[("GLOBAL", "PWRCOAAAAXX", 2023)] == 0.9
    assert sad[("GLOBAL", "ELCAAAXX01", 2023)] == 100.0
    assert sdp[("GLOBAL", "ELCAAAXX01", "S2D1", 2023)] == 0.6
    assert ys[("S1D1", 2023)] == 0.5
    assert cta[("GLOBAL", "PWRCOAAAAXX")] == 31.536
    print("  PASS: parsing.")

    # Country extraction
    assert extract_country_from_fuel("ELCAAAXX01") == "AAAXX"
    assert extract_country_from_tech("PWRCOAAAAXX") == "AAAXX"
    assert tech_class_of("PWRWONAAAXX") == "WON"
    print("  PASS: country/class extraction.")

    # Implied firm-capacity computation
    # Hand-computed expected for 2023:
    #   For S1D1: rate(A) = 100*0.4/0.5 = 80; tag=1.0 -> 80
    #             rate(B) = 50*0.5/0.5 = 50; tag=0.5 -> 25
    #             DNRM(S1D1) = 105
    #   For S2D1: rate(A) = 100*0.6/0.5 = 120; tag=1.0 -> 120
    #             rate(B) = 50*0.5/0.5 = 50;  tag=0.5 -> 25
    #             DNRM(S2D1) = 145    <- binding
    #   RHS = 145 * 1.20 = 174
    #   firm_GW = 174 / 31.536 = 5.518...
    implied = compute_implied_firm_capacity(
        rm_txt, rmtag_fuel, sad, sdp, ys,
        cta_default=31.536, rm_default=1.0, rmtag_fuel_default=0.0,
    )
    d23 = implied[2023]
    assert d23["binding_ts"] == "S2D1", f"expected S2D1 binding, got {d23['binding_ts']}"
    assert abs(d23["dnrm_max"] - 145.0) < 1e-6, f"DNRM_max wrong: {d23['dnrm_max']}"
    assert abs(d23["rhs_pjyr"] - 174.0) < 1e-6, f"rhs wrong: {d23['rhs_pjyr']}"
    expected_gw = 174.0 / 31.536
    assert abs(d23["firm_cap_required_gw"] - expected_gw) < 1e-6
    # Per-country shares at S2D1: A=120*1.0=120; B=50*0.5=25; total=145 -> consistent
    assert "AAAXX" in d23["by_country"]
    assert abs(d23["by_country"]["AAAXX"]["dnrm_share"] - 120.0) < 1e-6
    assert abs(d23["by_country"]["BBBXX"]["dnrm_share"] - 25.0) < 1e-6
    print(f"  PASS: implied firm-cap (2023 = {expected_gw:.4f} GW as expected).")

    # 2024 spot-check: SAD = 110/55, RM=1.25
    # S2D1 binding: rate(A)=110*0.6/0.5=132 -> 132; rate(B)=55*0.5/0.5=55 -> 27.5; DNRM=159.5
    # rhs = 159.5 * 1.25 = 199.375; firm_gw = 199.375/31.536
    d24 = implied[2024]
    assert d24["binding_ts"] == "S2D1"
    assert abs(d24["dnrm_max"] - 159.5) < 1e-6
    assert abs(d24["rhs_pjyr"] - 199.375) < 1e-6
    print(f"  PASS: implied firm-cap (2024 = {199.375/31.536:.4f} GW as expected).")

    # xlsx/.txt consistency
    assert all(rm_xlsx[y] == rm_txt[("GLOBAL", y)] for y in rm_xlsx)
    print("  PASS: xlsx/.txt RM consistency.")

    print("\nSelf-test: ALL PASSED.")
    return 0


# ----------------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--xlsx", type=Path,
                    help="Path to A-O_Parametrization.xlsx")
    ap.add_argument("--txt", type=Path,
                    help="Path to Pre_processed_BAU_0.txt (or NoStorage variant)")
    ap.add_argument("--demand-override", type=Path, default=None,
                    help="Optional JSON {fuel: {year: PJ}} to override SAD from .txt "
                         "(for re-running with corrected demand from the parallel chat)")
    ap.add_argument("--actual-fleet", type=Path, default=None,
                    help="Optional JSON {country: GW_in_2023} to override built-in fleet defaults")
    ap.add_argument("--flag-threshold", type=float, default=1.30,
                    help="Flag ReserveMargin values strictly above this multiplier (default 1.30 = 30%% reserve)")
    ap.add_argument("--cta", type=float, default=31.536,
                    help="CapacityToActivityUnit assumed uniform (default 31.536 PJ/(GW*yr))")
    ap.add_argument("--focus-years", type=str, default="2023,2030,2040,2050",
                    help="Comma-separated years for the per-country detail in section D (default 2023,2030,2040,2050)")
    ap.add_argument("--report-json", type=Path, default=None,
                    help="Optional: write a structured JSON summary to this path")
    ap.add_argument("-T", "--self-test", action="store_true",
                    help="Run synthetic self-test and exit (no real inputs needed)")
    args = ap.parse_args()

    if args.self_test:
        return run_self_test()

    if args.xlsx is None or args.txt is None:
        ap.error("--xlsx and --txt are required (or use --self-test)")

    for p in [args.xlsx, args.txt]:
        if not p.exists():
            sys.exit(f"Not found: {p}")
    if args.actual_fleet and not args.actual_fleet.exists():
        sys.exit(f"Not found: {args.actual_fleet}")
    if args.demand_override and not args.demand_override.exists():
        sys.exit(f"Not found: {args.demand_override}")

    print(f"Inputs:")
    print(f"  xlsx:            {args.xlsx}")
    print(f"  txt:             {args.txt}")
    print(f"  demand-override: {args.demand_override or '(none)'}")
    print(f"  actual-fleet:    {args.actual_fleet or '(built-in defaults)'}")
    print(f"  flag-threshold:  {args.flag_threshold}")
    print(f"  cta:             {args.cta}")
    print()

    print("Reading ReserveMargin from xlsx ...")
    rm_xlsx = read_reservemargin_from_xlsx(args.xlsx)
    print(f"  -> {len(rm_xlsx)} year entries")

    print("Parsing param blocks from .txt ...")
    blocks = parse_param_blocks(args.txt, {
        "ReserveMargin", "ReserveMarginTagFuel", "ReserveMarginTagTechnology",
        "SpecifiedAnnualDemand", "SpecifiedDemandProfile", "YearSplit",
        "CapacityToActivityUnit",
    })
    rm_txt    = index_reserve_margin(blocks["ReserveMargin"])
    rmtag_fuel = index_rmtag_fuel(blocks["ReserveMarginTagFuel"])
    rmtag_tech = index_rmtag_tech(blocks["ReserveMarginTagTechnology"])
    sad        = index_specified_demand(blocks["SpecifiedAnnualDemand"])
    sdp        = index_demand_profile(blocks["SpecifiedDemandProfile"])
    ys         = index_yearsplit(blocks["YearSplit"])
    cta_idx    = index_cta(blocks["CapacityToActivityUnit"])
    print(f"  -> ReserveMargin:           {len(rm_txt)} cells")
    print(f"  -> ReserveMarginTagFuel:    {len(rmtag_fuel)} cells")
    print(f"  -> ReserveMarginTagTech:    {len(rmtag_tech)} cells")
    print(f"  -> SpecifiedAnnualDemand:   {len(sad)} cells")
    print(f"  -> SpecifiedDemandProfile:  {len(sdp)} cells")
    print(f"  -> YearSplit:               {len(ys)} cells")
    print(f"  -> CapacityToActivityUnit:  {len(cta_idx)} cells")

    # Apply demand override (if any) BEFORE computing implied capacity
    if args.demand_override:
        with open(args.demand_override) as f:
            override = json.load(f)
        n_replaced = 0; n_added = 0
        for fuel, year_map in override.items():
            for y_str, val in year_map.items():
                key = ("GLOBAL", fuel, int(y_str))
                if key in sad: n_replaced += 1
                else: n_added += 1
                sad[key] = float(val)
        print(f"  -> demand-override applied: {n_replaced} replaced, {n_added} added.")

    # CTA: use first encountered if uniform, else fall back to --cta
    cta_vals = set(cta_idx.values())
    if len(cta_vals) > 1:
        print(f"  ! CapacityToActivityUnit varies across techs: {sorted(cta_vals)[:5]}{'...' if len(cta_vals) > 5 else ''}")
        print(f"    Using --cta={args.cta} for the implied-capacity calculation.")
        cta = args.cta
    elif cta_vals:
        cta = next(iter(cta_vals))
        if abs(cta - args.cta) > 1e-9:
            print(f"  CTA from .txt = {cta} (overriding --cta={args.cta} since uniform).")
    else:
        cta = args.cta

    # Compute implied firm capacity
    print("Computing implied firm-capacity requirement per year ...")
    rm_default          = blocks["__defaults__"].get("ReserveMargin", 1.0)
    rmtag_fuel_default  = blocks["__defaults__"].get("ReserveMarginTagFuel", 0.0)
    implied = compute_implied_firm_capacity(
        rm_txt, rmtag_fuel, sad, sdp, ys,
        cta_default=cta, rm_default=rm_default, rmtag_fuel_default=rmtag_fuel_default,
    )
    print(f"  -> {len(implied)} year(s) computed.")
    print()

    # Reports
    report_a_reservemargin(rm_xlsx, rm_txt, args.flag_threshold)
    report_b_tech_tags(rmtag_tech)
    report_c_fuel_tags(rmtag_fuel, sad)
    focus_years = [int(y.strip()) for y in args.focus_years.split(",") if y.strip()]
    report_d_implied_capacity(implied, focus_years)

    # Actual fleet
    if args.actual_fleet:
        with open(args.actual_fleet) as f:
            fleet = {k: float(v) for k, v in json.load(f).items()}
    else:
        fleet = dict(DEFAULT_ACTUAL_FLEET_GW)
    report_e_actual_fleet_compare(implied, fleet, compare_year=2023)

    report_f_verdict(rm_txt, rmtag_tech, rmtag_fuel, sad, implied, fleet, args.flag_threshold)

    # JSON dump
    if args.report_json:
        out = {
            "rm_xlsx":              {str(k): v for k, v in rm_xlsx.items()},
            "rm_txt":                [{"region": r, "year": y, "value": v}
                                       for (r, y), v in rm_txt.items()],
            "rmtag_fuel":            [{"region": r, "fuel": f, "year": y, "value": v}
                                       for (r, f, y), v in rmtag_fuel.items()],
            "rmtag_tech":            [{"region": r, "tech": t, "year": y, "value": v}
                                       for (r, t, y), v in rmtag_tech.items()],
            "implied_per_year":      {str(y): {**d,
                                                "by_country": {c: ent for c, ent in d["by_country"].items()}}
                                       for y, d in implied.items()},
            "actual_fleet_2023":     fleet,
            "flag_threshold":        args.flag_threshold,
            "cta":                   cta,
        }
        with open(args.report_json, "w") as f:
            json.dump(out, f, indent=1)
        print(f"\nStructured JSON written: {args.report_json}")

    print()
    print("Done.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
