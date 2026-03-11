"""
Set TRN Interconnection Activity Limits from Bilateral Flow Data

Reads estimated bilateral energy flows from flujos_energia_estimados_optimizacion.xlsx
and writes Lower/Upper activity limits into Secondary_Techs_Editor.xlsx.

- LowerLimit = Flujo Total (PJ) * 0.95  (-5%)
- UpperLimit = Flujo Total (PJ) * 1.05  (+5%)
- Years 2023-2025: actual flow data
- Years 2026-2050: flat projection from 2025 value

Usage:
    python t1_confection/D1b_set_trn_limits_from_flows.py

Author: Climate Lead Group, Andrey Salazar-Vargas
"""
import openpyxl
import unicodedata
from pathlib import Path

# ============================================================================
# CONSTANTS
# ============================================================================

GWH_TO_PJ = 0.0036  # 1 GWh = 0.0036 PJ

LOWER_LIMIT_FACTOR = 0.95  # -5%
UPPER_LIMIT_FACTOR = 1.05  # +5%

FLAT_PROJECTION_BASE_YEAR = 2025
EDITOR_FIRST_YEAR = 2023
EDITOR_LAST_YEAR = 2050

# Country name mapping (Spanish names with accents → 3-letter model codes)
# Same as D2_update_secondary_techs.py lines 19-40
COUNTRY_NAME_TO_CODE = {
    'Argentina': 'ARG',
    'Bolivia': 'BOL',
    'Brasil': 'BRA',
    'Chile': 'CHL',
    'Colombia': 'COL',
    'Costa Rica': 'CRI',
    'Ecuador': 'ECU',
    'El Salvador': 'SLV',
    'Guatemala': 'GTM',
    'Haiti': 'HTI',
    'Honduras': 'HND',
    'Mexico': 'MEX',
    'Nicaragua': 'NIC',
    'Panama': 'PAN',
    'Paraguay': 'PRY',
    'Peru': 'PER',
    'Republica Dominicana': 'DOM',
    'Uruguay': 'URY',
}


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def strip_accents(text):
    """Remove accents from text (e.g., 'México' → 'Mexico')"""
    nfkd = unicodedata.normalize('NFKD', str(text))
    return ''.join(c for c in nfkd if not unicodedata.combining(c))


def read_flow_data(flow_file_path):
    """
    Read bilateral energy flow data from the source Excel.

    Returns:
        dict: {frozenset({code_a, code_b}): {year: flujo_total_gwh}}
    """
    print(f"Reading flow data from: {flow_file_path.name}")

    wb = openpyxl.load_workbook(flow_file_path, data_only=True)

    # Find the sheet (handle accent in name)
    target_sheet = None
    for name in wb.sheetnames:
        if strip_accents(name) == 'Flujos por Interconexion':
            target_sheet = name
            break

    if not target_sheet:
        wb.close()
        raise ValueError(f"Sheet 'Flujos por Interconexión' not found. Available: {wb.sheetnames}")

    ws = wb[target_sheet]

    # Parse header row to find column indices
    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(1, col_idx).value
        if header:
            headers[strip_accents(str(header).strip())] = col_idx

    year_col = headers.get('Ano')
    pais_a_col = headers.get('Pais A')
    pais_b_col = headers.get('Pais B')
    flujo_total_col = headers.get('Flujo Total (GWh)')

    if not all([year_col, pais_a_col, pais_b_col, flujo_total_col]):
        wb.close()
        raise ValueError(f"Missing required columns. Found headers: {list(headers.keys())}")

    # Read flow data
    flow_data = {}  # {frozenset({code_a, code_b}): {year: flujo_gwh}}

    for row_idx in range(2, ws.max_row + 1):
        year_val = ws.cell(row_idx, year_col).value
        pais_a = ws.cell(row_idx, pais_a_col).value
        pais_b = ws.cell(row_idx, pais_b_col).value
        flujo = ws.cell(row_idx, flujo_total_col).value

        if year_val is None or flujo is None or pais_a is None or pais_b is None:
            continue

        pais_a_clean = strip_accents(str(pais_a).strip())
        pais_b_clean = strip_accents(str(pais_b).strip())

        code_a = COUNTRY_NAME_TO_CODE.get(pais_a_clean)
        code_b = COUNTRY_NAME_TO_CODE.get(pais_b_clean)

        if not code_a:
            print(f"  WARNING: Unknown country '{pais_a}' (cleaned: '{pais_a_clean}')")
            continue
        if not code_b:
            print(f"  WARNING: Unknown country '{pais_b}' (cleaned: '{pais_b_clean}')")
            continue

        pair = frozenset({code_a, code_b})
        if pair not in flow_data:
            flow_data[pair] = {}
        flow_data[pair][int(year_val)] = float(flujo)

    wb.close()

    print(f"  Found {len(flow_data)} country pairs")
    for pair, years in sorted(flow_data.items(), key=lambda x: sorted(x[0])):
        codes = sorted(pair)
        available_years = sorted(years.keys())
        print(f"  {codes[0]}-{codes[1]}: years {available_years[0]}-{available_years[-1]}, "
              f"2025 flow = {years.get(2025, 'N/A')} GWh")

    return flow_data


def fill_editor(editor_path, flow_data):
    """
    Write Lower/Upper activity limits into the Editor sheet.

    Two-pass approach:
    1. First read with data_only=True to resolve VLOOKUP formulas in column D
    2. Then open without data_only to write values (preserving formulas)
    """
    print(f"\nProcessing Editor: {editor_path.name}")

    # --- Pass 1: Read tech codes (resolving formulas) ---
    wb_read = openpyxl.load_workbook(editor_path, data_only=True)
    ws_read = wb_read['Editor']

    # Build year-column map from header
    year_col_map = {}
    for col_idx in range(1, ws_read.max_column + 1):
        header = ws_read.cell(1, col_idx).value
        if header and str(header).strip().isdigit():
            year = int(str(header).strip())
            if EDITOR_FIRST_YEAR <= year <= EDITOR_LAST_YEAR:
                year_col_map[year] = col_idx

    print(f"  Year columns: {min(year_col_map)} to {max(year_col_map)} ({len(year_col_map)} years)")

    # Read row info: tech codes and parameters
    row_info = {}  # {row_idx: {'tech_code': str, 'parameter': str}}
    for row_idx in range(2, ws_read.max_row + 1):
        tech_code = ws_read.cell(row_idx, 4).value  # Column D (resolved VLOOKUP)
        parameter = ws_read.cell(row_idx, 5).value   # Column E

        if tech_code and parameter:
            tech_str = str(tech_code).strip().upper()
            param_str = str(parameter).strip()
            if tech_str.startswith('TRN'):
                row_info[row_idx] = {
                    'tech_code': tech_str,
                    'parameter': param_str
                }

    wb_read.close()

    print(f"  Found {len(row_info)} TRN rows in Editor")

    # --- Pass 2: Write values (preserving formulas) ---
    wb_write = openpyxl.load_workbook(editor_path)
    ws_write = wb_write['Editor']

    updated_count = 0
    unmatched_techs = []

    for row_idx, info in sorted(row_info.items()):
        tech_code = info['tech_code']
        parameter = info['parameter']

        # Determine limit factor
        if parameter == 'TotalTechnologyAnnualActivityLowerLimit':
            factor = LOWER_LIMIT_FACTOR
            limit_type = 'Lower'
        elif parameter == 'TotalTechnologyAnnualActivityUpperLimit':
            factor = UPPER_LIMIT_FACTOR
            limit_type = 'Upper'
        else:
            continue

        # Extract country pair from tech code
        # Format: TRN[3-char]XX[3-char]XX (e.g., TRNARGXXBOLXX)
        # Positions: [3:6] = origin, then skip XX, [8:11] = destination
        origin = tech_code[3:6]
        dest = tech_code[8:11]

        # Bidirectional lookup in flow data
        pair = frozenset({origin, dest})

        if pair not in flow_data:
            unmatched_techs.append(tech_code)
            continue

        pair_flows = flow_data[pair]
        base_value_gwh = pair_flows.get(FLAT_PROJECTION_BASE_YEAR, 0.0)

        # Write values for each year
        for year in range(EDITOR_FIRST_YEAR, EDITOR_LAST_YEAR + 1):
            if year not in year_col_map:
                continue

            # Use actual data if available, otherwise flat from base year
            if year in pair_flows:
                gwh = pair_flows[year]
            else:
                gwh = base_value_gwh

            # Convert GWh → PJ and apply factor
            pj = gwh * GWH_TO_PJ
            limit_value = round(pj * factor, 6)

            ws_write.cell(row_idx, year_col_map[year], limit_value)

        updated_count += 1
        # Log first few for verification
        sample_2023 = pair_flows.get(2023, base_value_gwh)
        sample_2025 = pair_flows.get(2025, 0.0)
        print(f"  Row {row_idx:3d}: {tech_code} {limit_type:5s} | "
              f"2023: {sample_2023:.1f} GWh → {round(sample_2023 * GWH_TO_PJ * factor, 6):.6f} PJ | "
              f"2025 (flat): {sample_2025:.1f} GWh → {round(sample_2025 * GWH_TO_PJ * factor, 6):.6f} PJ")

    # Save
    wb_write.save(editor_path)
    wb_write.close()

    print(f"\n  Updated {updated_count} rows")
    if unmatched_techs:
        print(f"  WARNING: {len(unmatched_techs)} TRN techs without matching flow data:")
        for tech in unmatched_techs:
            print(f"    - {tech}")


# ============================================================================
# MAIN
# ============================================================================

def main():
    script_dir = Path(__file__).parent

    flow_file = script_dir / "Matriz Balance energético" / "flujos_energia_estimados_optimizacion.xlsx"
    editor_file = script_dir / "Secondary_Techs_Editor.xlsx"

    # Validate files exist
    if not flow_file.exists():
        print(f"ERROR: Flow data file not found: {flow_file}")
        return
    if not editor_file.exists():
        print(f"ERROR: Editor file not found: {editor_file}")
        return

    print("=" * 80)
    print("D1b: Set TRN Interconnection Activity Limits from Flow Data")
    print("=" * 80)
    print(f"  Source: {flow_file.name}")
    print(f"  Target: {editor_file.name}")
    print(f"  Conversion: GWh × {GWH_TO_PJ} = PJ")
    print(f"  Lower factor: {LOWER_LIMIT_FACTOR} (-5%)")
    print(f"  Upper factor: {UPPER_LIMIT_FACTOR} (+5%)")
    print(f"  Flat projection from: {FLAT_PROJECTION_BASE_YEAR}")
    print()

    # Step 1: Read flow data
    flow_data = read_flow_data(flow_file)

    # Step 2 & 3: Fill editor with limits
    fill_editor(editor_file, flow_data)

    print("\nDone!")


if __name__ == '__main__':
    main()
