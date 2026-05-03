#!/usr/bin/env python3
"""
demand_magnitude_check.py

Pre-measurement diagnostic for OSTRAM SpecifiedAnnualDemand magnitudes.
Tests the hypothesis: "Demand magnitude error somewhere in A-O_Demand.xlsx
or introduced during the multi-script A-O pipeline that produces
Pre_processed_BAU_0.txt, leading to ~12,000 GW of new capacity 2023-2030."

Reads
-----
  --xlsx  : A-O_Demand.xlsx          (source workbook, optional)
  --txt   : Pre_processed_BAU_0.txt  (post-pipeline GMPL data file, required)

Produces (siblings of inputs, never mutates inputs)
---------------------------------------------------
  <prefix>_by_country_year.csv   Total demand per country, year
                                 (raw model units AND TWh-equivalent assuming PJ)
  <prefix>_growth_flags.csv      YoY > 15% flag, 2030/2023 > 2x flag
  <prefix>_xlsx_vs_txt_diff.csv  Per-cell xlsx vs txt comparison (only if --xlsx)
  <prefix>_summary.txt           Human-readable summary, top offenders,
                                 unit-sanity verdict against published refs

Self-test
---------
  --self_test  Build synthetic xlsx + txt on disk, run the full pipeline,
               assert known answers, print PASS/FAIL.

Usage (Windows cmd, line continuation = ^)
------------------------------------------
  python demand_magnitude_check.py --self_test

  python demand_magnitude_check.py ^
      --xlsx  ..\\inputs\\A-O_Demand.xlsx ^
      --txt   ..\\inputs\\Pre_processed_BAU_0.txt ^
      --out_prefix demand_check
"""

from __future__ import annotations

import argparse
import csv
import math
import re
import sys
import tempfile
from collections import defaultdict
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple


# Reference 2023 electricity consumption (TWh) for the unit-sanity verdict.
REFERENCE_2023_TWH: Dict[str, float] = {
    "IND": 1500.0,   # India total final electricity ~1500 TWh
    "BGD": 80.0,
    "LKA": 16.0,
    "NPL": 9.0,
    "BTN": 3.0,
    "MDV": 0.6,
}

# 1 PJ = 0.27778 TWh
PJ_TO_TWH = 1.0 / 3.6

# OSTRAM demand fuel codes look like ELC<RRRRR><NN> where chars 3:6 are the
# ISO3 country code (or the 'IND' prefix for India sub-regions).
_FUEL_RE = re.compile(r"^ELC([A-Z]{3})([A-Z0-9]{2})(\d{2})$")


def country_of(fuel: str) -> Optional[str]:
    """ELCBGDXX03 -> 'BGD', ELCINDNO03 -> 'IND'. None if unrecognised."""
    m = _FUEL_RE.match(fuel)
    return m.group(1) if m else None


def subregion_of(fuel: str) -> Optional[str]:
    """ELCBGDXX03 -> 'BGDXX', ELCINDNO03 -> 'INDNO'."""
    m = _FUEL_RE.match(fuel)
    return (m.group(1) + m.group(2)) if m else None


# ---------------------------------------------------------------------------
# 1. GMPL .txt parser
# ---------------------------------------------------------------------------

# OSTRAM Pre_processed header is:
#   param default 0 : SpecifiedAnnualDemand :=
_HEADER_RE = re.compile(
    r"param\s+default\s+\S+\s*:\s*SpecifiedAnnualDemand\s*:=", re.IGNORECASE
)


def parse_txt_specified_annual_demand(
    txt_path: Path,
) -> List[Tuple[str, str, int, float]]:
    """
    Parse the SpecifiedAnnualDemand block. Handles the 'Shape B' enumeration
    (one tuple per line: REGION FUEL YEAR VALUE) we observed in OSTRAM, and
    falls back to 'Shape A' (sliced table) if needed.
    """
    rows: List[Tuple[str, str, int, float]] = []
    pending: List[str] = []
    in_block = False

    with Path(txt_path).open("r", encoding="utf-8", errors="replace") as f:
        for raw in f:
            line = raw.split("#", 1)[0].strip()
            if not in_block:
                if _HEADER_RE.match(line):
                    in_block = True
                continue
            # Terminator: line is exactly ';' or ends with ';' as a lone token.
            if line == ";":
                break
            if line.endswith(";"):
                line = line[:-1].strip()
                if not line:
                    break
            if not line:
                continue
            toks = line.split()
            if len(toks) == 4:
                region, fuel, ystr, vstr = toks
                try:
                    rows.append((region, fuel, int(ystr), float(vstr)))
                    continue
                except ValueError:
                    pass
            pending.append(line)

    if rows:
        return rows

    # Shape A fallback: alternating slice header [REG, FUEL, *]: then years
    # then values. Best-effort.
    slice_re = re.compile(r"\[\s*([^,\]]+)\s*,\s*([^,\]]+)\s*,\s*\*\s*\]\s*:")
    i = 0
    while i < len(pending):
        m = slice_re.match(pending[i])
        if not m:
            i += 1
            continue
        region, fuel = m.group(1).strip(), m.group(2).strip()
        i += 1
        if i >= len(pending):
            break
        head = pending[i]
        year_part = head.split(":=", 1)[0] if ":=" in head else head
        i += 1
        years = [int(x) for x in year_part.split() if x.isdigit()]
        if i >= len(pending):
            break
        vals = []
        for x in pending[i].split():
            try:
                vals.append(float(x))
            except ValueError:
                pass
        i += 1
        for y, v in zip(years, vals):
            rows.append((region, fuel, y, v))
    return rows


# ---------------------------------------------------------------------------
# 2. xlsx parser
# ---------------------------------------------------------------------------

def parse_xlsx_specified_annual_demand(
    xlsx_path: Path,
) -> List[Tuple[str, str, int, float]]:
    """
    Read SpecifiedAnnualDemand from A-O_Demand.xlsx. Prefers sheet
    'Demand_Projection'; locates 'Fuel/Tech' column and year columns by
    header content, so the parser tolerates column-position changes.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise SystemExit(
            "openpyxl is required for --xlsx. Install: pip install openpyxl"
        ) from e

    wb = load_workbook(filename=str(xlsx_path), data_only=True, read_only=True)
    candidate = None
    if "Demand_Projection" in wb.sheetnames:
        candidate = "Demand_Projection"
    else:
        for sn in wb.sheetnames:
            ws = wb[sn]
            first = next(ws.iter_rows(max_row=1, values_only=True), None)
            if not first:
                continue
            n_years = sum(
                1 for c in first if isinstance(c, int) and 2000 <= c <= 2100
            )
            if n_years >= 5:
                candidate = sn
                break
    if candidate is None:
        raise ValueError(
            f"No demand sheet found in {xlsx_path}. Sheets: {wb.sheetnames}"
        )

    ws = wb[candidate]
    it = ws.iter_rows(values_only=True)
    header = next(it, None)
    if header is None:
        raise ValueError(f"Empty sheet '{candidate}' in {xlsx_path}")

    fuel_col: Optional[int] = None
    type_col: Optional[int] = None
    year_cols: List[Tuple[int, int]] = []
    for idx, h in enumerate(header):
        hs = ("" if h is None else str(h)).strip()
        norm = hs.lower().replace(" ", "").replace("/", "").replace("_", "")
        if norm in {"fueltech", "fuel"}:
            fuel_col = idx
        if norm == "demandshare":
            type_col = idx
        if isinstance(h, int) and 2000 <= h <= 2100:
            year_cols.append((idx, int(h)))
        elif isinstance(h, str) and h.isdigit() and 2000 <= int(h) <= 2100:
            year_cols.append((idx, int(h)))

    if fuel_col is None:
        raise ValueError(
            f"No 'Fuel/Tech' header in '{candidate}'. Headers: {header[:10]}"
        )
    if not year_cols:
        raise ValueError(f"No year columns in '{candidate}'.")

    out: List[Tuple[str, str, int, float]] = []
    for row in it:
        if row is None:
            continue
        if type_col is not None:
            t = row[type_col] if type_col < len(row) else None
            if t is None or str(t).strip().lower() != "demand":
                continue
        fuel_raw = row[fuel_col] if fuel_col < len(row) else None
        if fuel_raw is None:
            continue
        fuel = str(fuel_raw).strip()
        if not fuel:
            continue
        for col_idx, year in year_cols:
            v = row[col_idx] if col_idx < len(row) else None
            if v is None:
                continue
            try:
                fv = float(v)
            except (TypeError, ValueError):
                continue
            out.append(("GLOBAL", fuel, year, fv))
    return out


# ---------------------------------------------------------------------------
# 3. Aggregation & flagging
# ---------------------------------------------------------------------------

def aggregate_by_country_year(
    rows: Iterable[Tuple[str, str, int, float]],
) -> Tuple[Dict[Tuple[str, int], float],
           Dict[Tuple[str, int], float],
           List[int]]:
    by_c: Dict[Tuple[str, int], float] = defaultdict(float)
    by_sr: Dict[Tuple[str, int], float] = defaultdict(float)
    years: set = set()
    for _r, fuel, year, val in rows:
        c = country_of(fuel)
        sr = subregion_of(fuel)
        if c is None:
            continue
        by_c[(c, year)] += val
        if sr is not None:
            by_sr[(sr, year)] += val
        years.add(year)
    return dict(by_c), dict(by_sr), sorted(years)


def compute_growth_flags(
    by_cy: Dict[Tuple[str, int], float],
    years: List[int],
    yoy_threshold: float = 0.15,
    ratio_2030_threshold: float = 2.0,
) -> List[Dict]:
    countries = sorted({c for c, _ in by_cy.keys()})
    out: List[Dict] = []
    y0 = years[0] if years else None
    for c in countries:
        for i, y in enumerate(years):
            if i == 0:
                continue
            yp = years[i - 1]
            v = by_cy.get((c, y))
            vp = by_cy.get((c, yp))
            if v is None or vp is None or vp == 0:
                continue
            yoy = (v / vp) - 1.0
            flags: List[str] = []
            if yoy > yoy_threshold:
                flags.append(f"YOY>{int(yoy_threshold*100)}%")
            v0 = by_cy.get((c, y0)) if y0 is not None else None
            ratio = (v / v0) if v0 else None
            if y == 2030 and ratio is not None and ratio > ratio_2030_threshold:
                flags.append(f"2030/{y0}>{ratio_2030_threshold:g}x")
            out.append({
                "country": c, "year": y, "prev_year": yp,
                "value": v, "prev_value": vp,
                "yoy_growth": yoy,
                "ratio_to_first_year": ratio,
                "flags": flags,
            })
    return out


def crosscheck_xlsx_vs_txt(
    xlsx_rows: List[Tuple[str, str, int, float]],
    txt_rows: List[Tuple[str, str, int, float]],
    rel_tol: float = 1e-3,
) -> List[Dict]:
    x_map = {(f, y): v for _r, f, y, v in xlsx_rows}
    t_map = {(f, y): v for _r, f, y, v in txt_rows}
    keys = sorted(set(x_map) | set(t_map))
    diffs: List[Dict] = []
    for k in keys:
        f, y = k
        xv = x_map.get(k)
        tv = t_map.get(k)
        if xv is None:
            diffs.append({"fuel": f, "year": y, "xlsx": None, "txt": tv,
                          "abs_diff": None, "rel_diff": None,
                          "status": "txt_only"})
            continue
        if tv is None:
            diffs.append({"fuel": f, "year": y, "xlsx": xv, "txt": None,
                          "abs_diff": None, "rel_diff": None,
                          "status": "xlsx_only"})
            continue
        ad = abs(xv - tv)
        rd = ad / max(abs(xv), abs(tv), 1e-12)
        if rd > rel_tol:
            diffs.append({"fuel": f, "year": y, "xlsx": xv, "txt": tv,
                          "abs_diff": ad, "rel_diff": rd,
                          "status": "DIFFER"})
    return diffs


# ---------------------------------------------------------------------------
# 4. Reports
# ---------------------------------------------------------------------------

def fmt_num(x):
    if x is None:
        return ""
    if isinstance(x, float) and (math.isnan(x) or math.isinf(x)):
        return str(x)
    if abs(x) >= 1000:
        return f"{x:,.1f}"
    if abs(x) >= 1:
        return f"{x:,.3f}"
    return f"{x:.5f}"


def write_by_country_year_csv(path, by_cy, by_sr, years):
    countries = sorted({c for c, _ in by_cy.keys()})
    subregions = sorted({s for s, _ in by_sr.keys()})
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["scope", "code", "year", "value_model_units",
                    "value_TWh_if_PJ"])
        for c in countries:
            for y in years:
                v = by_cy.get((c, y))
                if v is None:
                    continue
                w.writerow(["country", c, y, f"{v:.6f}",
                            f"{v * PJ_TO_TWH:.6f}"])
        for s in subregions:
            for y in years:
                v = by_sr.get((s, y))
                if v is None:
                    continue
                w.writerow(["subregion", s, y, f"{v:.6f}",
                            f"{v * PJ_TO_TWH:.6f}"])


def write_growth_flags_csv(path, records):
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["country", "year", "prev_year",
                    "value", "prev_value", "yoy_growth",
                    "ratio_to_first_year", "flags"])
        for r in records:
            w.writerow([
                r["country"], r["year"], r["prev_year"],
                f"{r['value']:.6f}", f"{r['prev_value']:.6f}",
                f"{r['yoy_growth']:.6f}",
                "" if r["ratio_to_first_year"] is None
                    else f"{r['ratio_to_first_year']:.6f}",
                ";".join(r["flags"]),
            ])


def write_xlsx_vs_txt_csv(path, diffs):
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["fuel", "year", "xlsx", "txt",
                    "abs_diff", "rel_diff", "status"])
        for d in diffs:
            w.writerow([
                d["fuel"], d["year"],
                "" if d["xlsx"] is None else f"{d['xlsx']:.6f}",
                "" if d["txt"] is None else f"{d['txt']:.6f}",
                "" if d["abs_diff"] is None else f"{d['abs_diff']:.6g}",
                "" if d["rel_diff"] is None else f"{d['rel_diff']:.6g}",
                d["status"],
            ])


def write_summary(path, txt_path, xlsx_path,
                  by_cy, by_sr, years, growth, diffs):
    countries = sorted({c for c, _ in by_cy.keys()})
    subregions = sorted({s for s, _ in by_sr.keys()})
    y_first = years[0]
    y_2030 = 2030 if 2030 in years else None
    y_last = years[-1]

    L: List[str] = []

    def p(s=""):
        L.append(s)

    p("=" * 78)
    p("OSTRAM SpecifiedAnnualDemand magnitude diagnostic")
    p("=" * 78)
    p(f"  txt:  {txt_path}")
    p(f"  xlsx: {xlsx_path or '(not supplied — cross-check skipped)'}")
    p(f"  Year span: {y_first}..{y_last}  ({len(years)} years)")
    p(f"  Countries: {', '.join(countries)}")
    p(f"  Subregions: {', '.join(subregions)}")

    p()
    p("-" * 78)
    p(f"  Country totals at {y_first} / 2030 / {y_last} "
      "(model units & TWh-equivalent assuming PJ)")
    p("-" * 78)
    p(f"    {'COUNTRY':<8} {'YEAR':>6} {'MODEL_UNITS':>14} "
      f"{'TWh_if_PJ':>12} {'REF_2023_TWh':>14} {'RATIO':>8}")
    for c in countries:
        for y in (y_first, y_2030, y_last):
            if y is None:
                continue
            v = by_cy.get((c, y))
            if v is None:
                continue
            twh = v * PJ_TO_TWH
            ref = REFERENCE_2023_TWH.get(c) if y == y_first else None
            ratio = (twh / ref) if (ref and ref > 0) else None
            p(f"    {c:<8} {y:>6} {fmt_num(v):>14} {fmt_num(twh):>12} "
              f"{(fmt_num(ref) if ref else ''):>14} "
              f"{(f'{ratio:.2f}x' if ratio else ''):>8}")

    p()
    p("-" * 78)
    p(f"  Sub-region totals — {y_first} / 2030 / {y_last} "
      "(TWh-equivalent assuming PJ)")
    p("-" * 78)
    p(f"    {'SUBREGION':<10} {'2023':>10} {'2030':>10} "
      f"{f'{y_last}':>10} {'2030/2023':>10} {f'{y_last}/2023':>10}")
    for s in subregions:
        v23 = by_sr.get((s, y_first))
        v30 = by_sr.get((s, 2030))
        vlast = by_sr.get((s, y_last))
        r30 = (v30 / v23) if (v23 and v30) else None
        rlast = (vlast / v23) if (v23 and vlast) else None
        p(f"    {s:<10} "
          f"{(fmt_num(v23 * PJ_TO_TWH) if v23 else ''):>10} "
          f"{(fmt_num(v30 * PJ_TO_TWH) if v30 else ''):>10} "
          f"{(fmt_num(vlast * PJ_TO_TWH) if vlast else ''):>10} "
          f"{(f'{r30:.2f}x' if r30 else ''):>10} "
          f"{(f'{rlast:.2f}x' if rlast else ''):>10}")

    flagged = [r for r in growth if r["flags"]]
    p()
    p("-" * 78)
    p(f"  Growth-rate flags (YoY > 15% or 2030/{y_first} > 2x)")
    p("-" * 78)
    if not flagged:
        p("    (none — per-country YoY and 2030 ratio within sane range)")
    else:
        p(f"    {len(flagged)} flagged rows. Top 30 by YoY growth:")
        top = sorted(flagged, key=lambda r: -r["yoy_growth"])[:30]
        p(f"    {'COUNTRY':<8} {'YEAR':>6} {'PREV':>12} {'CURR':>12} "
          f"{'YOY':>8}  FLAGS")
        for r in top:
            p(f"    {r['country']:<8} {r['year']:>6} "
              f"{fmt_num(r['prev_value']):>12} {fmt_num(r['value']):>12} "
              f"{r['yoy_growth']*100:>7.1f}%  {','.join(r['flags'])}")

    if diffs is not None:
        p()
        p("-" * 78)
        p("  xlsx vs txt cross-check (rel_tol = 1e-3)")
        p("-" * 78)
        differ = [d for d in diffs if d["status"] == "DIFFER"]
        only_x = [d for d in diffs if d["status"] == "xlsx_only"]
        only_t = [d for d in diffs if d["status"] == "txt_only"]
        p(f"    Cells differing: {len(differ)}")
        p(f"    Cells in xlsx but not txt: {len(only_x)}")
        p(f"    Cells in txt but not xlsx: {len(only_t)}")
        if differ:
            p("    Top 10 by relative diff:")
            top = sorted(differ, key=lambda d: -d["rel_diff"])[:10]
            p(f"      {'FUEL':<14} {'YEAR':>6} {'XLSX':>14} {'TXT':>14} {'REL':>10}")
            for d in top:
                p(f"      {d['fuel']:<14} {d['year']:>6} "
                  f"{fmt_num(d['xlsx']):>14} {fmt_num(d['txt']):>14} "
                  f"{d['rel_diff']*100:>9.4f}%")

    p()
    p("-" * 78)
    p("  Unit-sanity verdict")
    p("-" * 78)
    rows_v = []
    for c, ref in REFERENCE_2023_TWH.items():
        v23 = by_cy.get((c, y_first))
        if v23 is None:
            continue
        twh = v23 * PJ_TO_TWH
        ratio = twh / ref if ref else None
        rows_v.append((c, v23, twh, ref, ratio))
    if rows_v:
        rs = sorted([r[4] for r in rows_v if r[4] is not None])
        if rs:
            mid = rs[len(rs) // 2]
            p(f"    Median (modeled-2023 TWh / published-2023 TWh) "
              f"across {len(rs)} countries assuming PJ units: {mid:.2f}x")
            if 0.7 <= mid <= 1.5:
                p("    -> Units consistent with PJ; magnitudes plausibly correct.")
            elif 1.5 < mid <= 5:
                p("    -> Magnitudes HIGH (1.5-5x). Likely a parametrization "
                  "issue, not a unit error.")
            elif mid > 5:
                p("    -> Magnitudes WAY HIGH (>5x). Suspect unit error or "
                  "double-counting (e.g. transmission losses + final "
                  "consumption added together).")
            else:
                p("    -> Magnitudes LOW (<0.7x). Suspect unit error in the "
                  "other direction (EJ being read as PJ?).")
        for c, v23, twh, ref, ratio in rows_v:
            if ratio is not None:
                p(f"    {c}: 2023 modeled = {twh:>8.2f} TWh, "
                  f"published = {ref:>8.2f} TWh, ratio = {ratio:.2f}x")
            else:
                p(f"    {c}: 2023 modeled = {twh:>8.2f} TWh, no reference")

    p()
    p("=" * 78)
    p("End of summary.")
    p("=" * 78)

    text = "\n".join(L)
    path.write_text(text, encoding="utf-8")
    return text


# ---------------------------------------------------------------------------
# 5. Self-test
# ---------------------------------------------------------------------------

def _self_test() -> int:
    print("Running self-test on synthetic inputs...")
    try:
        from openpyxl import Workbook
    except ImportError:
        print("  SKIP — openpyxl not available")
        return 1

    fails: List[str] = []

    def expect(cond, msg):
        if not cond:
            fails.append(msg)
            print(f"  FAIL: {msg}")
        else:
            print(f"  OK:   {msg}")

    tmp = Path(tempfile.mkdtemp(prefix="ostram_demand_test_"))

    # Synthetic xlsx
    xlsx_path = tmp / "synthetic_demand.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Demand_Projection"
    years = [2023, 2024, 2025]
    ws.append(["Demand/Share", "Fuel/Tech", "Name",
               "Ref.Cap.BY", "Ref.OAR.BY", "Ref.km.BY",
               "Projection.Mode", "Projection.Parameter"] + years)
    ws.append(["Demand", "ELCBGDXX03", "BGD", "n", "n", "n",
               "User defined", 0, 100.0, 110.0, 121.0])
    ws.append(["Demand", "ELCINDNO03", "IND-NO", "n", "n", "n",
               "User defined", 0, 50.0, 200.0, 220.0])
    ws.append(["Demand", "ELCINDWE03", "IND-WE", "n", "n", "n",
               "User defined", 0, 30.0, 33.0, 36.3])
    wb.save(str(xlsx_path))

    # Synthetic txt
    txt_path = tmp / "synthetic_pre.txt"
    txt_path.write_text("\n".join([
        "# preamble",
        "param default 0 : SpecifiedAnnualDemand :=",
        "GLOBAL ELCBGDXX03 2023 100.0",
        "GLOBAL ELCBGDXX03 2024 110.0",
        "GLOBAL ELCBGDXX03 2025 121.0",
        "GLOBAL ELCINDNO03 2023 50.0",
        "GLOBAL ELCINDNO03 2024 200.0",
        "GLOBAL ELCINDNO03 2025 220.0",
        "GLOBAL ELCINDWE03 2023 30.0",
        "GLOBAL ELCINDWE03 2024 33.0",
        "GLOBAL ELCINDWE03 2025 36.3",
        ";",
        "param other_param := ;",
    ]), encoding="utf-8")

    txt_rows = parse_txt_specified_annual_demand(txt_path)
    expect(len(txt_rows) == 9,
           f"txt parse: 9 rows expected, got {len(txt_rows)}")
    expect(("GLOBAL", "ELCBGDXX03", 2024, 110.0) in txt_rows,
           "txt has BGD 2024 = 110.0")

    xlsx_rows = parse_xlsx_specified_annual_demand(xlsx_path)
    expect(len(xlsx_rows) == 9,
           f"xlsx parse: 9 rows expected, got {len(xlsx_rows)}")
    expect(("GLOBAL", "ELCBGDXX03", 2024, 110.0) in xlsx_rows,
           "xlsx has BGD 2024 = 110.0")

    by_cy, by_sr, ys = aggregate_by_country_year(txt_rows)
    expect(by_cy[("BGD", 2023)] == 100.0, "BGD 2023 country total = 100")
    expect(by_cy[("IND", 2023)] == 80.0,
           f"IND 2023 = 80, got {by_cy.get(('IND', 2023))}")
    expect(by_cy[("IND", 2024)] == 233.0,
           f"IND 2024 = 233, got {by_cy.get(('IND', 2024))}")
    expect(by_sr[("INDNO", 2024)] == 200.0,
           f"INDNO 2024 = 200, got {by_sr.get(('INDNO', 2024))}")
    expect(set(ys) == {2023, 2024, 2025}, f"years = 2023..2025, got {ys}")

    growth = compute_growth_flags(by_cy, ys)
    bgd_24 = next((r for r in growth
                   if r["country"] == "BGD" and r["year"] == 2024), None)
    expect(bgd_24 is not None, "growth row for BGD 2024 exists")
    expect(bgd_24 and abs(bgd_24["yoy_growth"] - 0.10) < 1e-9,
           f"BGD 2024 YoY = 0.10, got {bgd_24 and bgd_24['yoy_growth']}")
    expect(bgd_24 and not bgd_24["flags"],
           "BGD 2024 has no flags (10% YoY)")
    ind_24 = next((r for r in growth
                   if r["country"] == "IND" and r["year"] == 2024), None)
    expect(ind_24 and ind_24["yoy_growth"] > 1.0,
           f"IND 2024 YoY huge, got {ind_24 and ind_24['yoy_growth']}")
    expect(ind_24 and any("YOY>" in fl for fl in ind_24["flags"]),
           f"IND 2024 flagged for YoY, got {ind_24 and ind_24['flags']}")

    diffs_clean = crosscheck_xlsx_vs_txt(xlsx_rows, txt_rows)
    expect(all(d["status"] != "DIFFER" for d in diffs_clean),
           f"identical inputs => zero DIFFER, got {len(diffs_clean)}")

    # Inject a synthetic mismatch
    bad = [(r, f, y, (200.0 if (f == "ELCBGDXX03" and y == 2025) else v))
           for (r, f, y, v) in txt_rows]
    diffs_bad = crosscheck_xlsx_vs_txt(xlsx_rows, bad)
    rec = next((d for d in diffs_bad
                if d["fuel"] == "ELCBGDXX03" and d["year"] == 2025), None)
    expect(rec is not None and rec["status"] == "DIFFER",
           "injected BGD 2025 mismatch detected as DIFFER")

    write_by_country_year_csv(tmp / "x_by.csv", by_cy, by_sr, ys)
    write_growth_flags_csv(tmp / "x_grw.csv", growth)
    write_xlsx_vs_txt_csv(tmp / "x_dif.csv", diffs_clean)
    write_summary(tmp / "x_sum.txt", txt_path, xlsx_path,
                  by_cy, by_sr, ys, growth, diffs_clean)
    expect((tmp / "x_sum.txt").stat().st_size > 0,
           "summary written non-empty")

    expect(country_of("ELCBGDXX03") == "BGD",
           "country_of(ELCBGDXX03) == BGD")
    expect(country_of("ELCINDNO03") == "IND",
           "country_of(ELCINDNO03) == IND")
    expect(country_of("PWRSDSLKAXX") is None,
           "country_of on a non-fuel returns None")

    if fails:
        print(f"\nSelf-test: {len(fails)} FAILURES")
        return 1
    print(f"\nSelf-test: PASS  (artefacts in {tmp})")
    return 0


# ---------------------------------------------------------------------------
# 6. Main CLI
# ---------------------------------------------------------------------------

def main(argv: Optional[List[str]] = None) -> int:
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("--xlsx", type=Path, default=None,
                    help="Path to A-O_Demand.xlsx (optional)")
    ap.add_argument("--txt", type=Path, default=None,
                    help="Path to Pre_processed_BAU_0.txt "
                         "(required unless --self_test)")
    ap.add_argument("--out_prefix", type=str, default="demand_check",
                    help="Output filename prefix")
    ap.add_argument("--self_test", action="store_true",
                    help="Run synthetic self-test and exit")
    ap.add_argument("--yoy_threshold", type=float, default=0.15,
                    help="YoY threshold for flagging (default 0.15)")
    args = ap.parse_args(argv)

    if args.self_test:
        return _self_test()

    if args.txt is None:
        ap.error("--txt is required (or pass --self_test)")
    if not args.txt.exists():
        sys.exit(f"Not found: {args.txt}")
    if args.xlsx is not None and not args.xlsx.exists():
        sys.exit(f"Not found: {args.xlsx}")

    print(f"Inputs:")
    print(f"  txt:  {args.txt}")
    print(f"  xlsx: {args.xlsx or '(skipped)'}")
    print()

    print("Parsing .txt...")
    txt_rows = parse_txt_specified_annual_demand(args.txt)
    print(f"  -> {len(txt_rows)} (region, fuel, year, value) tuples")

    xlsx_rows: List[Tuple[str, str, int, float]] = []
    if args.xlsx is not None:
        print("Parsing .xlsx...")
        xlsx_rows = parse_xlsx_specified_annual_demand(args.xlsx)
        print(f"  -> {len(xlsx_rows)} (region, fuel, year, value) tuples")

    print("Aggregating by country and subregion...")
    by_cy, by_sr, years = aggregate_by_country_year(txt_rows)
    print(f"  -> {len(by_cy)} (country, year) cells over {len(years)} years")

    print("Computing growth flags...")
    growth = compute_growth_flags(by_cy, years, args.yoy_threshold)
    flagged = [r for r in growth if r["flags"]]
    print(f"  -> {len(growth)} growth records, {len(flagged)} flagged")

    diffs: Optional[List[Dict]] = None
    if xlsx_rows:
        print("Cross-checking xlsx vs txt...")
        diffs = crosscheck_xlsx_vs_txt(xlsx_rows, txt_rows)
        n_d = sum(1 for d in diffs if d["status"] == "DIFFER")
        n_x = sum(1 for d in diffs if d["status"] == "xlsx_only")
        n_t = sum(1 for d in diffs if d["status"] == "txt_only")
        print(f"  -> {n_d} differing, {n_x} xlsx-only, {n_t} txt-only")

    out_dir = args.txt.parent
    p_byc = out_dir / f"{args.out_prefix}_by_country_year.csv"
    p_grw = out_dir / f"{args.out_prefix}_growth_flags.csv"
    p_dif = out_dir / f"{args.out_prefix}_xlsx_vs_txt_diff.csv"
    p_sum = out_dir / f"{args.out_prefix}_summary.txt"

    print("Writing reports...")
    write_by_country_year_csv(p_byc, by_cy, by_sr, years)
    print(f"  -> {p_byc}")
    write_growth_flags_csv(p_grw, growth)
    print(f"  -> {p_grw}")
    if diffs is not None:
        write_xlsx_vs_txt_csv(p_dif, diffs)
        print(f"  -> {p_dif}")
    text = write_summary(p_sum, args.txt, args.xlsx,
                         by_cy, by_sr, years, growth, diffs)
    print(f"  -> {p_sum}")
    print()
    print(text)
    return 0


if __name__ == "__main__":
    sys.exit(main())
