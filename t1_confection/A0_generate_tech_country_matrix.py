# -*- coding: utf-8 -*-
"""
Technology-Country Matrix Generator
====================================
Generates an Excel configuration file that allows users to specify which
technology-country combinations should be processed by A1_Pre_processing_OG_csvs.py.

This script creates:
1. A matrix sheet where each cell indicates if a tech-country combo is enabled
2. An NGS unification configuration sheet (CCG + OCG → NGS)
3. Aggregation rules following the same pattern as region consolidation

Author: Climate Lead Group, Andrey Salazar-Vargas
"""

import yaml
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path
from Z_AUX_config_loader import get_countries, get_country_names

# Script directory
SCRIPT_DIR = Path(__file__).resolve().parent

# Country codes and names from centralized config
COUNTRIES = get_countries()
COUNTRY_NAMES = get_country_names()

# Energy/technology sub-codes from code_to_energy dictionary
# These are the specific energy types that can be filtered by country
# Note: CCG and OCG are unified into NGS
# Excluded prefixes (ELC, MIN, PWR, RNW, TRN) - these are structural prefixes
# that combine with the codes below to form complete technology names
TECHNOLOGIES = [
    "BCK",  # Backstop
    "BIO",  # Biomass
    "CCS",  # Carbon Capture Storage with Coal
    "COA",  # Coal
    "COG",  # Cogeneration
    "CSP",  # Concentrated Solar Power
    "GAS",  # Natural Gas
    "GEO",  # Geothermal
    "HYD",  # Hydroelectric
    "LDS",  # Long duration storage
    "NGS",  # Natural Gas (unified CCG + OCG)
    "OIL",  # Oil
    "OTH",  # Other
    "PET",  # Petroleum
    "SDS",  # Short duration storage
    "SPV",  # Solar Photovoltaic
    "URN",  # Nuclear
    "WAS",  # Waste
    "WAV",  # Wave
    "WOF",  # Offshore Wind
    "WON",  # Onshore Wind
]

TECH_DESCRIPTIONS = {
    "BCK": "Backstop",
    "BIO": "Biomass",
    "CCS": "Carbon Capture Storage with Coal",
    "COA": "Coal",
    "COG": "Cogeneration",
    "CSP": "Concentrated Solar Power",
    "GAS": "Natural Gas",
    "GEO": "Geothermal",
    "HYD": "Hydroelectric",
    "LDS": "Long duration storage",
    "NGS": "Natural Gas (CCG + OCG unified)",
    "OIL": "Oil",
    "OTH": "Other",
    "PET": "Petroleum",
    "SDS": "Short duration storage",
    "SPV": "Solar Photovoltaic",
    "URN": "Nuclear",
    "WAS": "Waste",
    "WAV": "Wave",
    "WOF": "Offshore Wind",
    "WON": "Onshore Wind",
}

# Load implausible combinations from region consolidation config
# Format in YAML: {tech: [country1, country2, ...]}
# Converted here to set of (tech, country) tuples for fast lookup
def _load_implausible_combinations():
    config_path = SCRIPT_DIR / "Config_country_codes.yaml"
    with open(config_path, "r", encoding="utf-8") as f:
        config = yaml.safe_load(f)
    combos = set()
    for tech, countries in config.get("implausible_combinations", {}).items():
        for country in countries:
            combos.add((tech, country))
    return combos

IMPLAUSIBLE_COMBINATIONS = _load_implausible_combinations()

# Aggregation rules (same as region_consolidation.yaml)
AGGREGATION_RULES = {
    "avg": [
        "AvailabilityFactor",
        "CapacityFactor",
        "CapacityToActivityUnit",
        "CapitalCost",
        "CapitalCostStorage",
        "EmissionActivityRatio",
        "FixedCost",
        "InputActivityRatio",
        "OperationalLife",
        "OperationalLifeStorage",
        "OutputActivityRatio",
        "ReserveMarginTagFuel",
        "ReserveMarginTagTechnology",
        "SpecifiedDemandProfile",
        "TechnologyFromStorage",
        "TechnologyToStorage",
        "VariableCost",
    ],
    "sum": [
        "ResidualCapacity",
        "ResidualStorageCapacity",
        "StorageLevelStart",
        "SpecifiedAnnualDemand",
        "TotalAnnualMaxCapacity",
        "TotalAnnualMaxCapacityInvestment",
        "TotalAnnualMinCapacityInvestment",
        "TotalTechnologyAnnualActivityLowerLimit",
        "TotalTechnologyAnnualActivityUpperLimit",
    ],
    "disabled": [
        "CapacityOfOneTechnologyUnit",
        "RETagTechnology",
        "TotalAnnualMinCapacity",
        "TotalTechnologyModelPeriodActivityLowerLimit",
        "TotalTechnologyModelPeriodActivityUpperLimit",
    ]
}


def create_tech_country_matrix():
    """Creates the technology-country matrix Excel file."""

    output_file = SCRIPT_DIR / "Tech_Country_Matrix.xlsx"

    wb = Workbook()

    # =========================================================================
    # Sheet 1: Matrix
    # =========================================================================
    ws_matrix = wb.active
    ws_matrix.title = "Matrix"

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    tech_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    implausible_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Light red
    implausible_font = Font(color="CC0000")  # Dark red text
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Row 1: Enable Matrix Filtering flag
    flag_label_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # Gold
    flag_font = Font(bold=True, size=11)

    cell_label = ws_matrix.cell(row=1, column=1, value="Enable Matrix Filtering:")
    cell_label.fill = flag_label_fill
    cell_label.font = flag_font
    cell_label.alignment = center_align
    cell_label.border = thin_border

    cell_flag = ws_matrix.cell(row=1, column=2, value="YES")
    cell_flag.fill = flag_label_fill
    cell_flag.font = flag_font
    cell_flag.alignment = center_align
    cell_flag.border = thin_border

    # Data validation for the flag cell
    dv_flag = DataValidation(type="list", formula1='"YES,NO"', allow_blank=False)
    dv_flag.error = "Please select YES or NO"
    dv_flag.errorTitle = "Invalid Input"
    ws_matrix.add_data_validation(dv_flag)
    dv_flag.add(cell_flag)

    # Note explaining the flag
    note_cell = ws_matrix.cell(row=1, column=3, value="YES = only pass technologies marked YES below  |  NO = pass ALL technologies from data")
    note_cell.font = Font(italic=True, color="666666")

    # Row 2: empty spacer

    # Row 3: Header row - Countries (shifted down by 2)
    MATRIX_START_ROW = 3
    ws_matrix.cell(row=MATRIX_START_ROW, column=1, value="Tech \\ Country")
    ws_matrix.cell(row=MATRIX_START_ROW, column=1).fill = header_fill
    ws_matrix.cell(row=MATRIX_START_ROW, column=1).font = header_font
    ws_matrix.cell(row=MATRIX_START_ROW, column=1).alignment = center_align
    ws_matrix.cell(row=MATRIX_START_ROW, column=1).border = thin_border

    for col_idx, country in enumerate(COUNTRIES, start=2):
        cell = ws_matrix.cell(row=MATRIX_START_ROW, column=col_idx, value=country)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # Technology rows (shifted down by 2)
    for row_idx, tech in enumerate(TECHNOLOGIES, start=MATRIX_START_ROW + 1):
        # Tech code in first column
        cell = ws_matrix.cell(row=row_idx, column=1, value=tech)
        cell.fill = tech_fill
        cell.font = Font(bold=True)
        cell.alignment = center_align
        cell.border = thin_border

        # YES/NO cells for each country
        for col_idx, country in enumerate(COUNTRIES, start=2):
            # Check if this is an implausible combination
            is_implausible = (tech, country) in IMPLAUSIBLE_COMBINATIONS
            value = "NO" if is_implausible else "YES"

            cell = ws_matrix.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = center_align
            cell.border = thin_border

            # Apply red highlighting for implausible combinations
            if is_implausible:
                cell.fill = implausible_fill
                cell.font = implausible_font

    # Data validation for YES/NO in matrix cells
    dv = DataValidation(type="list", formula1='"YES,NO"', allow_blank=False)
    dv.error = "Please select YES or NO"
    dv.errorTitle = "Invalid Input"
    ws_matrix.add_data_validation(dv)

    # Apply validation to all data cells (shifted down by 2)
    for row_idx in range(MATRIX_START_ROW + 1, MATRIX_START_ROW + 1 + len(TECHNOLOGIES)):
        for col_idx in range(2, len(COUNTRIES) + 2):
            dv.add(ws_matrix.cell(row=row_idx, column=col_idx))

    # Adjust column widths
    ws_matrix.column_dimensions['A'].width = 18
    for col_idx in range(2, len(COUNTRIES) + 2):
        ws_matrix.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A' + chr(64 + col_idx - 26)].width = 6

    # =========================================================================
    # Sheet 2: NGS_Unification
    # =========================================================================
    ws_ngs = wb.create_sheet("NGS_Unification")

    # Title
    ws_ngs.cell(row=1, column=1, value="NGS Unification Configuration")
    ws_ngs.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws_ngs.merge_cells('A1:D1')

    # Description
    ws_ngs.cell(row=3, column=1, value="This configuration unifies CCG (Combined Cycle) and OCG (Open Cycle) into NGS (Natural Gas).")
    ws_ngs.merge_cells('A3:F3')

    # Enable/Disable
    ws_ngs.cell(row=5, column=1, value="Enable NGS Unification:")
    ws_ngs.cell(row=5, column=1).font = Font(bold=True)
    ws_ngs.cell(row=5, column=2, value="YES")

    dv_enable = DataValidation(type="list", formula1='"YES,NO"', allow_blank=False)
    ws_ngs.add_data_validation(dv_enable)
    dv_enable.add(ws_ngs.cell(row=5, column=2))

    # Source technologies
    ws_ngs.cell(row=7, column=1, value="Source Technologies:")
    ws_ngs.cell(row=7, column=1).font = Font(bold=True)
    ws_ngs.cell(row=7, column=2, value="CCG, OCG")

    ws_ngs.cell(row=8, column=1, value="Target Technology:")
    ws_ngs.cell(row=8, column=1).font = Font(bold=True)
    ws_ngs.cell(row=8, column=2, value="NGS")

    # =========================================================================
    # Sheet 3: Aggregation_Rules
    # =========================================================================
    ws_agg = wb.create_sheet("Aggregation_Rules")

    # Title
    ws_agg.cell(row=1, column=1, value="Aggregation Rules for NGS Unification")
    ws_agg.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws_agg.merge_cells('A1:C1')

    ws_agg.cell(row=3, column=1, value="These rules define how parameters are aggregated when unifying CCG + OCG → NGS")
    ws_agg.merge_cells('A3:E3')

    # Headers
    row = 5
    for col, header in enumerate(["Parameter", "Aggregation Type", "Description"], start=1):
        cell = ws_agg.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    row = 6

    # Average parameters
    for param in AGGREGATION_RULES["avg"]:
        ws_agg.cell(row=row, column=1, value=param).border = thin_border
        ws_agg.cell(row=row, column=2, value="AVG").border = thin_border
        ws_agg.cell(row=row, column=2).alignment = center_align
        ws_agg.cell(row=row, column=3, value="Values are averaged").border = thin_border
        row += 1

    # Sum parameters
    for param in AGGREGATION_RULES["sum"]:
        ws_agg.cell(row=row, column=1, value=param).border = thin_border
        ws_agg.cell(row=row, column=2, value="SUM").border = thin_border
        ws_agg.cell(row=row, column=2).alignment = center_align
        ws_agg.cell(row=row, column=3, value="Values are summed").border = thin_border
        row += 1

    # Disabled parameters
    for param in AGGREGATION_RULES["disabled"]:
        ws_agg.cell(row=row, column=1, value=param).border = thin_border
        ws_agg.cell(row=row, column=2, value="DISABLED").border = thin_border
        ws_agg.cell(row=row, column=2).alignment = center_align
        ws_agg.cell(row=row, column=3, value="Parameter is skipped").border = thin_border
        row += 1

    # Adjust column widths
    ws_agg.column_dimensions['A'].width = 40
    ws_agg.column_dimensions['B'].width = 18
    ws_agg.column_dimensions['C'].width = 25

    # =========================================================================
    # Sheet 4: Tech_Reference
    # =========================================================================
    ws_ref = wb.create_sheet("Tech_Reference")

    # Title
    ws_ref.cell(row=1, column=1, value="Technology Reference")
    ws_ref.cell(row=1, column=1).font = Font(bold=True, size=14)

    # Headers
    row = 3
    for col, header in enumerate(["Code", "Description", "Notes"], start=1):
        cell = ws_ref.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    row = 4
    for tech, desc in TECH_DESCRIPTIONS.items():
        ws_ref.cell(row=row, column=1, value=tech).border = thin_border
        ws_ref.cell(row=row, column=2, value=desc).border = thin_border
        notes = ""
        if tech == "NGS":
            notes = "Unified from CCG + OCG"
        ws_ref.cell(row=row, column=3, value=notes).border = thin_border
        row += 1

    # Adjust column widths
    ws_ref.column_dimensions['A'].width = 10
    ws_ref.column_dimensions['B'].width = 35
    ws_ref.column_dimensions['C'].width = 25

    # =========================================================================
    # Sheet 5: Country_Reference
    # =========================================================================
    ws_country = wb.create_sheet("Country_Reference")

    # Title
    ws_country.cell(row=1, column=1, value="Country Reference")
    ws_country.cell(row=1, column=1).font = Font(bold=True, size=14)

    # Headers
    row = 3
    for col, header in enumerate(["Code", "Country Name"], start=1):
        cell = ws_country.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    row = 4
    for code in COUNTRIES:
        ws_country.cell(row=row, column=1, value=code).border = thin_border
        ws_country.cell(row=row, column=2, value=COUNTRY_NAMES.get(code, "")).border = thin_border
        row += 1

    # Adjust column widths
    ws_country.column_dimensions['A'].width = 10
    ws_country.column_dimensions['B'].width = 25

    # Save workbook
    wb.save(output_file)
    print(f"✓ Created: {output_file}")
    print(f"  - Matrix sheet: {len(TECHNOLOGIES)} technologies × {len(COUNTRIES)} countries")
    print(f"    └─ {len(IMPLAUSIBLE_COMBINATIONS)} implausible combinations marked as NO (red highlighted)")
    print(f"  - NGS_Unification sheet: CCG + OCG → NGS configuration")
    print(f"  - Aggregation_Rules sheet: {len(AGGREGATION_RULES['avg'])} avg, {len(AGGREGATION_RULES['sum'])} sum, {len(AGGREGATION_RULES['disabled'])} disabled")
    print(f"  - Tech_Reference sheet: Technology descriptions")
    print(f"  - Country_Reference sheet: Country names")

    return output_file


if __name__ == "__main__":
    create_tech_country_matrix()
