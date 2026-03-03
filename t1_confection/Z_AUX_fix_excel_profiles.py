"""
Script to normalize demand profiles in Excel files A-O_Demand.xlsx

This script fixes the SpecifiedDemandProfile sums in the "Profiles" sheet
of A-O_Demand.xlsx files so they sum to exactly 1.0 per country/year.

Problem: Some countries have profiles that sum to 1.01 or 0.99 instead of 1.0,
which causes errors in the OSeMOSYS model.

Author: Claude Code
Date: 2026-02-03
"""

import openpyxl
from pathlib import Path
from datetime import datetime
import shutil

def make_backup(file_path):
    """Create backup of an Excel file"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_file = str(file_path).replace(".xlsx", f"_backup_{timestamp}.xlsx")
    shutil.copy(file_path, backup_file)
    return backup_file

def normalize_excel_profiles(excel_path):
    """
    Normalizes demand profiles in the Profiles sheet of A-O_Demand.xlsx

    Expected structure:
    - Column 0: Timeslices (S1D1, S1D2, etc.)
    - Column 1: Demand/Share
    - Column 2: Fuel/Tech (e.g.: ELCPERXX02)
    - Columns 9+: Years (2023, 2024, ..., 2050)

    For each Fuel/Tech, the sum of all timeslices must be 1.0 per year.
    """
    print(f"\n{'='*70}")
    print(f"📊 Normalizing: {excel_path.name}")
    print(f"{'='*70}")

    if not excel_path.exists():
        print(f"❌ File not found: {excel_path}")
        return False

    # Create backup
    backup = make_backup(excel_path)
    print(f"✓ Backup created: {Path(backup).name}")

    # Load workbook
    wb = openpyxl.load_workbook(excel_path)

    if 'Profiles' not in wb.sheetnames:
        print(f"⚠️  Sheet 'Profiles' not found in {excel_path.name}")
        wb.close()
        return False

    ws = wb['Profiles']

    # Identify year columns (from column 9 onwards)
    # Column 9 (J) should be 2023
    first_year_col = 10  # Column J in Excel (1-indexed)
    last_col = ws.max_column

    # Get list of unique fuels
    fuels_data = {}  # {fuel: [(row_num, timeslice), ...]}

    for row_num in range(2, ws.max_row + 1):
        fuel = ws.cell(row_num, 3).value  # Column C (Fuel/Tech)
        timeslice = ws.cell(row_num, 1).value  # Column A (Timeslice)

        if fuel and timeslice:
            if fuel not in fuels_data:
                fuels_data[fuel] = []
            fuels_data[fuel].append(row_num)

    print(f"\n📍 Fuels found: {len(fuels_data)}")

    # Normalize each fuel for each year
    problemas_antes = 0
    problemas_despues = 0
    correcciones = 0

    for fuel, row_numbers in fuels_data.items():
        # For each year column
        for col_num in range(first_year_col, last_col + 1):
            # Read current values
            values = []
            for row_num in row_numbers:
                cell = ws.cell(row_num, col_num)
                val = cell.value
                if val is not None:
                    try:
                        values.append((row_num, float(val)))
                    except:
                        values.append((row_num, 0.0))
                else:
                    values.append((row_num, 0.0))

            # Calculate current sum
            suma_actual = sum(v[1] for v in values)

            # If sum is not 1.0, normalize
            if abs(suma_actual - 1.0) > 0.0001:
                problemas_antes += 1

                if suma_actual > 0:  # Only normalize if sum is positive
                    # Normalize by dividing by the sum
                    for row_num, val in values:
                        new_val = val / suma_actual
                        ws.cell(row_num, col_num).value = new_val

                    correcciones += 1

                    # Verify sum after correction
                    suma_nueva = sum(ws.cell(r, col_num).value for r, _ in values)
                    if abs(suma_nueva - 1.0) > 0.0001:
                        problemas_despues += 1

    print(f"\n📊 Results:")
    print(f"  - Problems found: {problemas_antes}")
    print(f"  - Corrections applied: {correcciones}")
    print(f"  - Remaining problems: {problemas_despues}")

    # Save changes
    wb.save(excel_path)
    wb.close()

    if problemas_despues == 0:
        print(f"\n✅ File normalized successfully")
        return True
    else:
        print(f"\n⚠️  Some problems persist after normalization")
        return False

def main():
    """Main function"""
    print("="*70)
    print("🔧 PROFILE NORMALIZATION IN EXCEL FILES")
    print("="*70)

    base_dir = Path(__file__).parent

    # List of scenarios to process
    scenarios = ["BAU", "NDC", "NDC+ELC", "NDC_NoRPO"]

    results = {}

    for scenario in scenarios:
        scenario_dir = base_dir / "A1_Outputs" / f"A1_Outputs_{scenario}"
        excel_file = scenario_dir / "A-O_Demand.xlsx"

        if excel_file.exists():
            results[scenario] = normalize_excel_profiles(excel_file)
        else:
            print(f"\n⚠️  Scenario {scenario}: file not found")
            print(f"    {excel_file}")
            results[scenario] = None

    # Final summary
    print("\n" + "="*70)
    print("📋 FINAL SUMMARY")
    print("="*70)

    for scenario, result in results.items():
        if result is True:
            print(f"✅ {scenario}: Normalized successfully")
        elif result is False:
            print(f"❌ {scenario}: Problems during normalization")
        else:
            print(f"⚠️  {scenario}: File not found")

    all_success = all(r == True for r in results.values() if r is not None)

    if all_success:
        print("\n🎉 All files normalized successfully!")
    else:
        print("\n⚠️  Some files could not be normalized correctly")

    print("\n💡 IMPORTANT:")
    print("   Profiles in Excel files now sum to exactly 1.0.")
    print("   If you run A2_AddTx.py, these values will be exported to CSVs.")
    print("="*70)

    return 0 if all_success else 1

if __name__ == "__main__":
    exit(main())
