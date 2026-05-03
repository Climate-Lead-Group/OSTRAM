"""
test_add_max_cap_investment_lid_rule.py
========================================

Tests for the standalone lid + untie patch.

Run with:  pytest test_add_max_cap_investment_lid_rule.py -v
"""

from __future__ import annotations

import hashlib
import shutil
import sys
from pathlib import Path

import pandas as pd
import pytest

THIS_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(THIS_DIR))

from add_max_cap_investment_lid_rule import (  # noqa: E402
    LID_PERCENTAGE_DEFAULT,
    MAX_INV_PARAM,
    MIN_INV_PARAM,
    PLACEHOLDER_VALUE,
    PROJ_MODE_COL,
    RES_PARAM,
    UNTIE_MULTIPLIER,
    identify_allowed_techs,
    run,
)

SOURCE_DIR = THIS_DIR / "A1_Outputs" / "A1_Outputs_BAU"
PARAM = "A-O_Parametrization.xlsx"
OTHER_FILES = (
    "A-O_AR_Model_Base_Year.xlsx",
    "A-O_AR_Projections.xlsx",
    "A-O_Demand.xlsx",
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------
@pytest.fixture
def working_dir(tmp_path):
    """Fresh copy of A1_Outputs_BAU per test."""
    target = tmp_path / "A1_Outputs_BAU"
    shutil.copytree(SOURCE_DIR, target)
    return target


def _md5(path: Path) -> str:
    return hashlib.md5(path.read_bytes()).hexdigest()


def _read(path: Path, sheet: str) -> pd.DataFrame:
    return pd.read_excel(path, sheet_name=sheet)


def _allowed_techs(working_dir):
    df = _read(working_dir / PARAM, "Secondary Techs")
    year_cols = [c for c in df.columns if isinstance(c, int)]
    from add_max_cap_investment_lid_rule import (
        load_generation_techs, RESTRICT_TO_GENERATION, TECH_TYPES_FILE,
    )
    gen = None
    if RESTRICT_TO_GENERATION:
        gen = load_generation_techs(THIS_DIR / TECH_TYPES_FILE)
    return identify_allowed_techs(df, year_cols, gen)


def _gen_techs():
    """Convenience: load the GENERATION set from TECH_TYPES.csv (next to script)."""
    from add_max_cap_investment_lid_rule import load_generation_techs, TECH_TYPES_FILE
    return load_generation_techs(THIS_DIR / TECH_TYPES_FILE)


def _demand_mult(working_dir):
    """Convenience: load the demand multiplier map for the working_dir's data."""
    from add_max_cap_investment_lid_rule import (
        build_demand_multiplier_map, DEMAND_FILE_NAME,
    )
    return build_demand_multiplier_map(working_dir / DEMAND_FILE_NAME)


# ---------------------------------------------------------------------------
# Backup + log
# ---------------------------------------------------------------------------
class TestBackup:
    def test_backup_folder_created(self, working_dir):
        log = run(working_dir)
        assert Path(log["backup_dir"]).is_dir()

    def test_backup_param_file_byte_identical_to_pre_edit(self, working_dir):
        pre_hash = _md5(working_dir / PARAM)
        log = run(working_dir)
        bak_hash = _md5(Path(log["backup_dir"]) / PARAM)
        assert bak_hash == pre_hash

    def test_change_log_written(self, working_dir):
        log = run(working_dir)
        assert Path(log["log_path"]).is_file()

    def test_change_log_records_lid_config(self, working_dir):
        log = run(working_dir)
        assert log["lid_percentage_default"] == LID_PERCENTAGE_DEFAULT


# ---------------------------------------------------------------------------
# Lid math
# ---------------------------------------------------------------------------
class TestLidMath:
    def test_empty_cells_filled_with_pool_based_lid(self, working_dir):
        """For ALLOWED PWR* techs, empty MaxInv cells become pct(cr, y) * pool,
        unless the untie rule kicks in. The pool is the per-(country+region,
        year) sum of ResidualCapacity across ALLOWED techs in that
        country+region."""
        from add_max_cap_investment_lid_rule import (
            country_region_for, build_pool_map, identify_allowed_techs,
            lid_pct_for_cr_year,
        )
        pre = _read(working_dir / PARAM, "Secondary Techs")
        year_cols = [c for c in pre.columns if isinstance(c, int)]
        pre_max_inv = pre[pre["Parameter"] == MAX_INV_PARAM].set_index("Tech")[year_cols]
        mininvs = pre[pre["Parameter"] == MIN_INV_PARAM].set_index("Tech")[year_cols]
        gen = _gen_techs()
        allowed = identify_allowed_techs(pre, year_cols, gen)
        pool_map = build_pool_map(pre, allowed, year_cols)
        demand_mult = _demand_mult(working_dir)

        run(working_dir)
        post = _read(working_dir / PARAM, "Secondary Techs")
        post_max_inv = post[post["Parameter"] == MAX_INV_PARAM].set_index("Tech")[year_cols]

        for tech in allowed:
            cr = country_region_for(tech)
            if cr is None:
                continue  # TRN* etc. are skipped
            for yr in year_cols:
                old = pre_max_inv.loc[tech, yr]
                if not pd.isna(old):
                    continue
                new = post_max_inv.loc[tech, yr]
                pool = pool_map.get((cr, yr), 0.0)
                mci = mininvs.loc[tech, yr]
                mci = 0.0 if pd.isna(mci) else float(mci)
                expected_lid = lid_pct_for_cr_year(cr, yr, demand_mult) * pool
                if mci > 0 and mci >= expected_lid:
                    expected = mci * UNTIE_MULTIPLIER
                else:
                    expected = expected_lid
                assert abs(float(new) - expected) < 1e-9, (
                    f"{tech} {yr}: expected {expected}, got {new}"
                )

    def test_all_techs_in_same_country_region_share_lid(self, working_dir):
        """Two ALLOWED techs in the same country+region must end up with the
        same MaxInv value for years where neither has a manual cal and
        neither hits the untie rule."""
        run(working_dir)
        post = _read(working_dir / PARAM, "Secondary Techs")
        # PWRCOABGDXX and PWROILBGDXX are both BGDXX, both empty in 2030, both
        # MinInv=0 in 2030 — should share lid.
        a = post[
            (post["Tech"] == "PWRCOABGDXX") & (post["Parameter"] == MAX_INV_PARAM)
        ].iloc[0][2030]
        b = post[
            (post["Tech"] == "PWROILBGDXX") & (post["Parameter"] == MAX_INV_PARAM)
        ].iloc[0][2030]
        assert abs(float(a) - float(b)) < 1e-9, (
            f"BGDXX techs 2030 should share lid: PWRCOABGDXX={a}, PWROILBGDXX={b}"
        )

    def test_lid_for_pwrhydindno_2050(self, working_dir):
        """PWRHYDINDNO is in INDNO; lid = pct(INDNO, 2050) * pool(INDNO, 2050).
        Under the non-linear demand ramp (pct = base * mult ** LID_DEMAND_EXP)
        this is many multiples of 0.005 * pool — the test recomputes the
        expected value via lid_pct_for_cr_year so it stays correct if the
        exponent changes."""
        from add_max_cap_investment_lid_rule import (
            build_pool_map, identify_allowed_techs, lid_pct_for_cr_year,
        )
        pre = _read(working_dir / PARAM, "Secondary Techs")
        year_cols = [c for c in pre.columns if isinstance(c, int)]
        gen = _gen_techs()
        allowed = identify_allowed_techs(pre, year_cols, gen)
        pool_map = build_pool_map(pre, allowed, year_cols)
        demand_mult = _demand_mult(working_dir)
        expected = lid_pct_for_cr_year("INDNO", 2050, demand_mult) * pool_map[("INDNO", 2050)]

        run(working_dir)
        post = _read(working_dir / PARAM, "Secondary Techs")
        new_val = post[
            (post["Tech"] == "PWRHYDINDNO") & (post["Parameter"] == MAX_INV_PARAM)
        ].iloc[0][2050]
        assert abs(float(new_val) - expected) < 1e-9

    def test_uniform_mode_formula(self, working_dir):
        """Pin the uniform-mode formula:
            pct(cr, y) = base_pct(y) * mult(cr, y)
            lid(t, y)  = pct(cr, y) * pool(cr, y)
        where base_pct(y) = LID_PERCENTAGE_BY_YEAR.get(y, LID_PERCENTAGE_DEFAULT).

        At the reference year, mult=1 so pct=base_pct and lid=base_pct*pool.
        At 2050, pct=base_pct(2050)*mult(2050) — verified by direct comparison
        against the production helper. This is a regression net: anyone who
        changes the formula will fail here.

        Proportional mode is exercised separately in TestProportionalMode.
        """
        from add_max_cap_investment_lid_rule import (
            lid_pct_for_cr_year,
            LID_PERCENTAGE_DEFAULT,
            LID_PERCENTAGE_BY_YEAR,
            LID_RULE_MODE,
            DEMAND_REFERENCE_YEAR,
        )
        # This test only makes sense in uniform mode.
        assert LID_RULE_MODE == "uniform", (
            f"This test pins uniform-mode behaviour; "
            f"current mode is {LID_RULE_MODE!r}"
        )

        demand_mult = _demand_mult(working_dir)

        # Reference year: pct should equal base_pct(ref_year) (mult=1).
        ref_base = LID_PERCENTAGE_BY_YEAR.get(
            DEMAND_REFERENCE_YEAR, LID_PERCENTAGE_DEFAULT
        )
        for cr in ("LKAXX", "INDNO", "BGDXX"):
            if (cr, DEMAND_REFERENCE_YEAR) not in demand_mult:
                continue
            pct_ref = lid_pct_for_cr_year(cr, DEMAND_REFERENCE_YEAR, demand_mult)
            assert abs(pct_ref - ref_base) < 1e-12, (
                f"{cr} {DEMAND_REFERENCE_YEAR}: pct={pct_ref}, "
                f"expected base_pct={ref_base} (mult=1)"
            )

        # 2050: pct should equal base_pct(2050) * mult(cr, 2050).
        base_2050 = LID_PERCENTAGE_BY_YEAR.get(2050, LID_PERCENTAGE_DEFAULT)
        for cr in ("LKAXX", "INDNO", "BGDXX"):
            if (cr, 2050) not in demand_mult:
                continue
            mult_2050 = demand_mult[(cr, 2050)]
            expected = base_2050 * mult_2050
            actual = lid_pct_for_cr_year(cr, 2050, demand_mult)
            assert abs(actual - expected) < 1e-12, (
                f"{cr} 2050: pct={actual}, expected {expected} "
                f"(base={base_2050}, mult={mult_2050})"
            )
            assert actual > LID_PERCENTAGE_DEFAULT, (
                f"{cr} 2050: pct {actual} should exceed base default "
                f"{LID_PERCENTAGE_DEFAULT}"
            )


class TestPoolComputation:
    def test_pool_includes_only_allowed_pwr_techs(self, working_dir):
        """Pool sums ResidualCapacity across ALLOWED PWR* techs in the same
        country+region. Manual recompute should match build_pool_map output."""
        from add_max_cap_investment_lid_rule import (
            build_pool_map, identify_allowed_techs,
        )
        df = _read(working_dir / PARAM, "Secondary Techs")
        year_cols = [c for c in df.columns if isinstance(c, int)]
        allowed = identify_allowed_techs(df, year_cols)
        pool_map = build_pool_map(df, allowed, year_cols)

        # Manually compute BGDXX 2030
        bgdxx_techs = [t for t in allowed
                       if isinstance(t, str) and len(t) == 11 and t[6:11] == "BGDXX"]
        res = df[df["Parameter"] == "ResidualCapacity"].set_index("Tech")
        manual_pool = sum(
            float(res.loc[t, 2030]) if not pd.isna(res.loc[t, 2030]) else 0.0
            for t in bgdxx_techs
        )
        assert abs(pool_map[("BGDXX", 2030)] - manual_pool) < 1e-9

    def test_pool_keys_only_from_pwr_techs(self, working_dir):
        """Country+region keys in the pool map must come from PWR* allowed
        techs only — no spurious keys from TRN* misparse."""
        from add_max_cap_investment_lid_rule import (
            build_pool_map, identify_allowed_techs, country_region_for,
        )
        df = _read(working_dir / PARAM, "Secondary Techs")
        year_cols = [c for c in df.columns if isinstance(c, int)]
        allowed = identify_allowed_techs(df, year_cols)
        pool_map = build_pool_map(df, allowed, year_cols)
        crs = {k[0] for k in pool_map}
        valid = {country_region_for(t) for t in allowed if country_region_for(t)}
        assert crs <= valid

    def test_trn_techs_marked_as_skipped(self, working_dir):
        """TRN* are INTERCONNECTORS in TECH_TYPES.csv → not GENERATION → not in
        the ALLOWED set. They never reach the apply loop, so the
        skipped_non_pwr_techs trace will only show non-PWR techs that pass the
        category filter (rare; usually empty). The important contract is that
        TRN* are absent from allowed_techs."""
        log = run(working_dir)
        allowed = set(log["sheets"][0]["allowed_techs"])
        for trn in ["TRNINDEAINDNE", "TRNINDSOLKAXX", "TRNNPLXXBGDXX"]:
            assert trn not in allowed, (
                f"{trn} should not be in allowed (not GENERATION)"
            )

    def test_trn_max_inv_unchanged(self, working_dir):
        """TRN* MaxInv rows are untouched even though they're in ALLOWED."""
        pre = _read(working_dir / PARAM, "Secondary Techs")
        year_cols = [c for c in pre.columns if isinstance(c, int)]
        run(working_dir)
        post = _read(working_dir / PARAM, "Secondary Techs")
        for trn in ["TRNINDEAINDNE", "TRNINDSOLKAXX"]:
            pre_row = pre[
                (pre["Tech"] == trn) & (pre["Parameter"] == MAX_INV_PARAM)
            ].iloc[0]
            post_row = post[
                (post["Tech"] == trn) & (post["Parameter"] == MAX_INV_PARAM)
            ].iloc[0]
            for yr in year_cols:
                pre_v = pre_row[yr]
                post_v = post_row[yr]
                if pd.isna(pre_v) and pd.isna(post_v):
                    continue
                assert pre_v == post_v, (
                    f"{trn} {yr}: TRN* MaxInv should be untouched, "
                    f"pre={pre_v} post={post_v}"
                )


# ---------------------------------------------------------------------------
# Untie rule
# ---------------------------------------------------------------------------
class TestUntieRule:
    def test_untie_pushes_max_above_min(self, working_dir):
        log = run(working_dir)
        post = _read(working_dir / PARAM, "Secondary Techs")
        year_cols = [c for c in post.columns if isinstance(c, int)]
        max_invs = post[post["Parameter"] == MAX_INV_PARAM].set_index("Tech")[year_cols]
        untie_changes = [
            c for c in log["sheets"][0]["changes"]
            if c.get("reason") == "untie_min_inv"
        ]
        assert len(untie_changes) > 0  # sanity: there should be cases
        for c in untie_changes:
            actual = max_invs.loc[c["tech"], c["year"]]
            expected = c["min_inv"] * UNTIE_MULTIPLIER
            assert abs(float(actual) - expected) < 1e-9

    def test_post_run_invariant_max_strictly_greater_than_min(self, working_dir):
        """After the run, every ALLOWED MaxInv cell must be > the corresponding
        MinInv cell wherever MinInv > 0 (this is the whole point of untie)."""
        log = run(working_dir)
        post = _read(working_dir / PARAM, "Secondary Techs")
        year_cols = [c for c in post.columns if isinstance(c, int)]
        allowed = log["sheets"][0]["allowed_techs"]
        max_invs = post[post["Parameter"] == MAX_INV_PARAM].set_index("Tech")[year_cols]
        mininvs = post[post["Parameter"] == MIN_INV_PARAM].set_index("Tech")[year_cols]
        for tech in allowed:
            for yr in year_cols:
                mci = mininvs.loc[tech, yr]
                if pd.isna(mci) or mci <= 0:
                    continue
                maxinv = max_invs.loc[tech, yr]
                assert not pd.isna(maxinv), (
                    f"{tech} {yr}: MaxInv NaN despite MinInv={mci}"
                )
                assert float(maxinv) > float(mci), (
                    f"{tech} {yr}: MaxInv={maxinv} not > MinInv={mci}"
                )

    def test_explicit_zero_with_min_inv_zero_stays_zero(self, working_dir):
        """If both MaxInv and MinInv were 0, MaxInv stays 0 (no untie needed)."""
        pre = _read(working_dir / PARAM, "Secondary Techs")
        year_cols = [c for c in pre.columns if isinstance(c, int)]
        pre_max = pre[pre["Parameter"] == MAX_INV_PARAM].set_index("Tech")[year_cols]
        pre_min = pre[pre["Parameter"] == MIN_INV_PARAM].set_index("Tech")[year_cols]

        log = run(working_dir)
        allowed = log["sheets"][0]["allowed_techs"]
        post = _read(working_dir / PARAM, "Secondary Techs")
        post_max = post[post["Parameter"] == MAX_INV_PARAM].set_index("Tech")[year_cols]

        for tech in allowed:
            for yr in year_cols:
                old = pre_max.loc[tech, yr]
                if pd.isna(old) or old != 0:
                    continue
                mci = pre_min.loc[tech, yr]
                if not pd.isna(mci) and mci > 0:
                    continue
                new = post_max.loc[tech, yr]
                assert new == 0, f"{tech} {yr}: zero with MinInv=0 should stay 0, got {new}"


# ---------------------------------------------------------------------------
# Manual calibration preservation
# ---------------------------------------------------------------------------
class TestManualPreservation:
    def test_manual_positive_value_preserved(self, working_dir):
        """PWRHYDBGDXX 2024 has 0.166 (manual cal), MinInv there is 0/NaN
        -> must be preserved exactly."""
        run(working_dir)
        post = _read(working_dir / PARAM, "Secondary Techs")
        row = post[
            (post["Tech"] == "PWRHYDBGDXX") & (post["Parameter"] == MAX_INV_PARAM)
        ].iloc[0]
        assert abs(float(row[2024]) - 0.166) < 1e-9

    def test_manual_value_preserved_unless_untie_fires(self, working_dir):
        pre = _read(working_dir / PARAM, "Secondary Techs")
        year_cols = [c for c in pre.columns if isinstance(c, int)]
        pre_max = pre[pre["Parameter"] == MAX_INV_PARAM].set_index("Tech")[year_cols]
        pre_min = pre[pre["Parameter"] == MIN_INV_PARAM].set_index("Tech")[year_cols]

        log = run(working_dir)
        allowed = log["sheets"][0]["allowed_techs"]
        post = _read(working_dir / PARAM, "Secondary Techs")
        post_max = post[post["Parameter"] == MAX_INV_PARAM].set_index("Tech")[year_cols]

        for tech in allowed:
            for yr in year_cols:
                old = pre_max.loc[tech, yr]
                if pd.isna(old) or old <= 0:
                    continue
                if old == PLACEHOLDER_VALUE:
                    continue  # 9999 isn't a manual cal — it's a placeholder
                mci = pre_min.loc[tech, yr]
                mci = 0.0 if pd.isna(mci) else float(mci)
                new = post_max.loc[tech, yr]
                if mci > 0 and mci >= float(old):
                    assert abs(float(new) - mci * UNTIE_MULTIPLIER) < 1e-9
                else:
                    assert abs(float(new) - float(old)) < 1e-9, (
                        f"{tech} {yr}: manual {old} clobbered to {new}"
                    )


# ---------------------------------------------------------------------------
# 9999 placeholder handling (so this script can run after the first patch)
# ---------------------------------------------------------------------------
class TestPlaceholderReplacement:
    def test_no_9999_remains_after_run(self, working_dir):
        """No 9999 placeholder remains in MaxInv for the techs Patch 2 owns.

        Patch 2's scope = GENERATION ∩ ALLOWED. Anything else (TRN*, storage,
        non-residual non-pipeline candidates) is intentionally untouched.
        Storage with leftover 9999 from a prior Patch 1 run is expected
        behavior — not a Patch 2 bug. Patch 1 is the right place to handle
        non-generation placeholders if that ever becomes a need.
        """
        log = run(working_dir)
        post = _read(working_dir / PARAM, "Secondary Techs")
        year_cols = [c for c in post.columns if isinstance(c, int)]
        allowed = set(log["sheets"][0]["allowed_techs"])
        in_scope = post[
            (post["Parameter"] == MAX_INV_PARAM) & (post["Tech"].isin(allowed))
        ]
        for yr in year_cols:
            col = in_scope[yr]
            assert (col != PLACEHOLDER_VALUE).all(), (
                f"9999 placeholder still in patched MaxInv column {yr}"
            )

    def test_injected_9999_replaced_with_lid(self, working_dir):
        """Inject 9999 into a known cell (PWRHYDINDNO 2050, originally NaN).
        The lid logic should recognize it as a placeholder and replace with
        pct(INDNO, 2050) * pool(INDNO, 2050)."""
        from openpyxl import load_workbook
        from add_max_cap_investment_lid_rule import (
            build_pool_map, identify_allowed_techs, lid_pct_for_cr_year,
        )
        wb = load_workbook(working_dir / PARAM)
        ws = wb["Secondary Techs"]
        df = _read(working_dir / PARAM, "Secondary Techs")
        target_row = df.index[
            (df["Tech"] == "PWRHYDINDNO") & (df["Parameter"] == MAX_INV_PARAM)
        ][0] + 2
        target_col = next(
            c for c in range(1, ws.max_column + 1)
            if ws.cell(1, c).value == 2050
        )
        ws.cell(target_row, target_col).value = PLACEHOLDER_VALUE
        wb.save(working_dir / PARAM)
        wb.close()

        # Compute expected from the post-injection state.
        df2 = _read(working_dir / PARAM, "Secondary Techs")
        year_cols = [c for c in df2.columns if isinstance(c, int)]
        gen = _gen_techs()
        allowed = identify_allowed_techs(df2, year_cols, gen)
        pool_map = build_pool_map(df2, allowed, year_cols)
        demand_mult = _demand_mult(working_dir)
        expected = lid_pct_for_cr_year("INDNO", 2050, demand_mult) * pool_map[("INDNO", 2050)]

        run(working_dir)
        post = _read(working_dir / PARAM, "Secondary Techs")
        new_val = post[
            (post["Tech"] == "PWRHYDINDNO") & (post["Parameter"] == MAX_INV_PARAM)
        ].iloc[0][2050]
        assert abs(float(new_val) - expected) < 1e-9


# ---------------------------------------------------------------------------
# Per-year overrides
# ---------------------------------------------------------------------------
class TestPerYearOverride:
    def test_override_applied_only_to_specific_year(self, working_dir, monkeypatch):
        """An entry in LID_PERCENTAGE_BY_YEAR sets the BASE pct for that year
        (formerly the literal pct, now the base before mult is applied).

        Adding {2040: 0.10} for testing should make pct(INDNO, 2040) =
        0.10 * mult(INDNO, 2040), which differs from years using the schedule
        default. We compute the expectation via lid_pct_for_cr_year so the
        test is robust to formula changes (uniform mode only).
        """
        from add_max_cap_investment_lid_rule import (
            build_pool_map, identify_allowed_techs, lid_pct_for_cr_year,
        )
        import add_max_cap_investment_lid_rule as mod

        pre = _read(working_dir / PARAM, "Secondary Techs")
        year_cols = [c for c in pre.columns if isinstance(c, int)]
        gen = _gen_techs()
        allowed = identify_allowed_techs(pre, year_cols, gen)
        pool_map = build_pool_map(pre, allowed, year_cols)
        demand_mult = _demand_mult(working_dir)

        # Inject an artificial override at 2040 to force pct=0.10 base for
        # this test specifically. Other year entries already exist in the
        # production schedule; we only override 2040.
        monkeypatch.setitem(mod.LID_PERCENTAGE_BY_YEAR, 2040, 0.10)
        run(working_dir)
        post = _read(working_dir / PARAM, "Secondary Techs")
        row = post[
            (post["Tech"] == "PWRHYDINDNO") & (post["Parameter"] == MAX_INV_PARAM)
        ].iloc[0]

        # 2040: override now declares base_pct=0.10; mult layers on top.
        # expected = lid_pct_for_cr_year(INDNO, 2040, dm) * pool(INDNO, 2040)
        expected_2040 = (
            lid_pct_for_cr_year("INDNO", 2040, demand_mult)
            * pool_map[("INDNO", 2040)]
        )
        assert abs(float(row[2040]) - expected_2040) < 1e-9, (
            f"PWRHYDINDNO 2040: expected {expected_2040}, got {row[2040]}"
        )
        # 2050: no monkeypatch; falls through to whatever the production
        # schedule has at 2050. Same computation pattern.
        expected_2050 = (
            lid_pct_for_cr_year("INDNO", 2050, demand_mult)
            * pool_map[("INDNO", 2050)]
        )
        assert abs(float(row[2050]) - expected_2050) < 1e-9, (
            f"PWRHYDINDNO 2050: expected {expected_2050}, got {row[2050]}"
        )


# ---------------------------------------------------------------------------
# GENERATION-only filter (TECH_TYPES.csv)
# ---------------------------------------------------------------------------
class TestGenerationFilter:
    """Patch 2 should only touch techs categorized as GENERATION in
    TECH_TYPES.csv. Storage (e.g. PWRSDSLKAXX → STORAGE_SHORT) and other
    non-generation categories must be left alone."""

    def test_storage_tech_not_in_allowed(self, working_dir):
        """PWRSDSLKAXX is STORAGE_SHORT — it must not appear in allowed_techs."""
        log = run(working_dir)
        allowed = set(log["sheets"][0]["allowed_techs"])
        assert "PWRSDSLKAXX" not in allowed, (
            "PWRSDSLKAXX is storage; should be filtered out of allowed_techs"
        )

    def test_storage_max_inv_row_unchanged(self, working_dir):
        """PWRSDSLKAXX MaxInv row is byte-identical pre vs post."""
        pre = _read(working_dir / PARAM, "Secondary Techs")
        year_cols = [c for c in pre.columns if isinstance(c, int)]
        run(working_dir)
        post = _read(working_dir / PARAM, "Secondary Techs")
        pre_row = pre[
            (pre["Tech"] == "PWRSDSLKAXX") & (pre["Parameter"] == MAX_INV_PARAM)
        ]
        post_row = post[
            (post["Tech"] == "PWRSDSLKAXX") & (post["Parameter"] == MAX_INV_PARAM)
        ]
        if len(pre_row) == 0:
            pytest.skip("PWRSDSLKAXX not in this dataset")
        for yr in year_cols:
            pv = pre_row.iloc[0][yr]
            qv = post_row.iloc[0][yr]
            if pd.isna(pv) and pd.isna(qv):
                continue
            assert pv == qv, (
                f"PWRSDSLKAXX {yr}: storage row was modified (pre={pv}, post={qv})"
            )

    def test_only_generation_techs_in_allowed(self, working_dir):
        """Every tech in allowed_techs must be in the GENERATION category."""
        log = run(working_dir)
        allowed = set(log["sheets"][0]["allowed_techs"])
        gen = _gen_techs()
        non_gen = allowed - gen
        assert not non_gen, (
            f"Non-GENERATION techs leaked into allowed: {sorted(non_gen)[:10]}"
        )

    def test_log_records_filter_metadata(self, working_dir):
        """The change log captures the filter state so it's reproducible."""
        log = run(working_dir)
        assert log["restrict_to_generation"] is True
        assert log["generation_techs_count"] is not None and log["generation_techs_count"] > 0
        assert "TECH_TYPES.csv" in str(log["tech_types_file"])

    def test_missing_tech_types_file_raises(self, tmp_path, monkeypatch):
        """If TECH_TYPES.csv is missing, load_generation_techs raises clearly."""
        from add_max_cap_investment_lid_rule import load_generation_techs
        with pytest.raises(FileNotFoundError, match="TECH_TYPES.csv"):
            load_generation_techs(tmp_path / "does_not_exist.csv")


# ---------------------------------------------------------------------------
# Demand-anchored ramp
# ---------------------------------------------------------------------------
class TestDemandRamp:
    """The lid percentage for (cr, year) scales with that region's demand
    growth: pct(cr, y) = LID_PERCENTAGE_DEFAULT × demand(cr, y) / demand(cr, ref)."""

    def test_pct_at_reference_year_equals_default(self, working_dir):
        """In the anchor year, the ramp multiplier is 1.0, so pct = base."""
        from add_max_cap_investment_lid_rule import (
            lid_pct_for_cr_year, DEMAND_REFERENCE_YEAR,
        )
        demand_mult = _demand_mult(working_dir)
        for cr in ["BGDXX", "INDNO", "INDWE", "LKAXX"]:
            pct = lid_pct_for_cr_year(cr, DEMAND_REFERENCE_YEAR, demand_mult)
            assert abs(pct - LID_PERCENTAGE_DEFAULT) < 1e-12, (
                f"{cr} {DEMAND_REFERENCE_YEAR}: pct={pct} should equal "
                f"default {LID_PERCENTAGE_DEFAULT}"
            )

    def test_lid_grows_year_over_year_with_demand(self, working_dir):
        """For a region with growing demand, the lid in 2050 is strictly
        greater than the lid in 2024 (residual is roughly flat, demand 3x)."""
        run(working_dir)
        post = _read(working_dir / PARAM, "Secondary Techs")
        row = post[
            (post["Tech"] == "PWRHYDINDNO") & (post["Parameter"] == MAX_INV_PARAM)
        ].iloc[0]
        # Skip years that fall under the untie rule (where MinInv > lid)
        # by looking at years where lid is the dominant decision (e.g. 2031+
        # for PWRHYDINDNO, after the manual cal block ends).
        v_2031 = float(row[2031])
        v_2050 = float(row[2050])
        assert v_2050 > v_2031, (
            f"PWRHYDINDNO: 2050 lid ({v_2050}) should be > 2031 lid ({v_2031})"
        )

    def test_demand_mult_loaded_in_log(self, working_dir):
        """Log records that the demand ramp was applied with a populated map."""
        log = run(working_dir)
        assert log["lid_ramp_from_demand"] is True
        assert log["demand_mult_loaded"] is True
        assert log["demand_reference_year"] == 2024


class TestPerCRRamp:
    """Different country+regions get different ramps because demand grows
    at different rates (BGDXX ~5.3% CAGR vs INDWE ~4.5% CAGR)."""

    def test_bgdxx_ramp_steeper_than_indwe(self, working_dir):
        """BGDXX demand grows faster than INDWE → BGDXX lid pct in 2050
        relative to 2024 is higher than INDWE's."""
        from add_max_cap_investment_lid_rule import lid_pct_for_cr_year
        demand_mult = _demand_mult(working_dir)
        bgdxx_24 = lid_pct_for_cr_year("BGDXX", 2024, demand_mult)
        bgdxx_50 = lid_pct_for_cr_year("BGDXX", 2050, demand_mult)
        indwe_24 = lid_pct_for_cr_year("INDWE", 2024, demand_mult)
        indwe_50 = lid_pct_for_cr_year("INDWE", 2050, demand_mult)
        assert (bgdxx_50 / bgdxx_24) > (indwe_50 / indwe_24), (
            f"BGDXX ramp ratio {bgdxx_50/bgdxx_24:.3f} should exceed "
            f"INDWE ratio {indwe_50/indwe_24:.3f}"
        )

    def test_log_records_per_cr_pct_schedule(self, working_dir):
        """Log includes the (cr, year) pct values actually used in the run."""
        log = run(working_dir)
        pct_map = log["sheets"][0].get("lid_pct_by_cr_year", {})
        assert pct_map, "lid_pct_by_cr_year missing from log"
        # Spot-check: BGDXX 2050 pct should differ from INDWE 2050 pct.
        b50 = pct_map.get("BGDXX_2050")
        i50 = pct_map.get("INDWE_2050")
        assert b50 is not None and i50 is not None
        assert b50 != i50, "Different crs should produce different pct in 2050"


class TestRampDisable:
    """When LID_RAMP_FROM_DEMAND is False, the lid drops the demand multiplier
    but still respects LID_PERCENTAGE_BY_YEAR. So pct(cr, y) = base_pct(y)
    uniformly across crs (no per-cr differentiation), and lid = base_pct(y) *
    pool(cr, y). Two crs in the same year should produce identical pct."""

    def test_flat_lid_when_ramp_disabled(self, working_dir, monkeypatch):
        from add_max_cap_investment_lid_rule import (
            lid_pct_for_cr_year, build_pool_map, identify_allowed_techs,
        )
        import add_max_cap_investment_lid_rule as mod

        monkeypatch.setattr(mod, "LID_RAMP_FROM_DEMAND", False)
        # With ramp off, pct(cr, y) = base_pct(y) — same for every cr in a
        # given year, but still varies across years per the schedule.
        run(working_dir)
        post = _read(working_dir / PARAM, "Secondary Techs")
        year_cols = [c for c in post.columns if isinstance(c, int)]
        gen = _gen_techs()
        allowed = identify_allowed_techs(post, year_cols, gen)
        pool_map = build_pool_map(post, allowed, year_cols)

        # PWRHYDINDNO 2050: expected = base_pct(2050) * pool(INDNO, 2050).
        # We compute via lid_pct_for_cr_year (with demand_mult_map=None, since
        # ramp is monkeypatched off) to be robust to formula changes.
        row = post[
            (post["Tech"] == "PWRHYDINDNO") & (post["Parameter"] == MAX_INV_PARAM)
        ].iloc[0]
        expected = lid_pct_for_cr_year("INDNO", 2050, None) * pool_map[("INDNO", 2050)]
        assert abs(float(row[2050]) - expected) < 1e-9, (
            f"With ramp disabled, PWRHYDINDNO 2050 should be base_pct(2050) * "
            f"pool = {expected}, got {row[2050]}"
        )

        # Cross-cr check: with ramp off, two different crs share the same
        # pct in a given year (the schedule is uniform across crs).
        pct_indno_2050 = lid_pct_for_cr_year("INDNO", 2050, None)
        pct_lkaxx_2050 = lid_pct_for_cr_year("LKAXX", 2050, None)
        assert abs(pct_indno_2050 - pct_lkaxx_2050) < 1e-12, (
            f"Ramp off: pct should not vary across crs. "
            f"INDNO 2050 = {pct_indno_2050}, LKAXX 2050 = {pct_lkaxx_2050}"
        )


# ---------------------------------------------------------------------------
# Non-interference with the rest of the workbook
# ---------------------------------------------------------------------------
class TestNonInterference:
    def test_only_allowed_techs_max_inv_rows_changed(self, working_dir):
        """Every parameter row OTHER than MaxInv for ALLOWED techs must be
        byte-identical between pre and post."""
        pre = _read(working_dir / PARAM, "Secondary Techs")
        run(working_dir)
        post = _read(working_dir / PARAM, "Secondary Techs")
        # Compare rows where (Parameter, Tech) is NOT (MaxInv, allowed-tech)
        allowed = _allowed_techs(working_dir)
        for idx in pre.index:
            tech = pre.loc[idx, "Tech"]
            param = pre.loc[idx, "Parameter"]
            if param == MAX_INV_PARAM and tech in allowed:
                continue  # this row is allowed to change
            year_cols = [c for c in pre.columns if isinstance(c, int)]
            pre_vals = pre.loc[idx, year_cols].fillna("NA").tolist()
            post_vals = post.loc[idx, year_cols].fillna("NA").tolist()
            assert pre_vals == post_vals, (
                f"Row {idx} ({tech}, {param}) changed but should not have"
            )

    def test_other_AO_files_byte_identical(self, working_dir):
        pre = {f: _md5(working_dir / f) for f in OTHER_FILES}
        run(working_dir)
        post = {f: _md5(working_dir / f) for f in OTHER_FILES}
        assert pre == post

    def test_non_allowed_techs_max_inv_unchanged(self, working_dir):
        """Techs with no residual and no min-cap (the "not allowed" set)
        have their MaxInv rows entirely untouched. That includes both the
        zeroed real-generator candidates AND the ELC pseudo-fuels."""
        pre = _read(working_dir / PARAM, "Secondary Techs")
        year_cols = [c for c in pre.columns if isinstance(c, int)]
        allowed = _allowed_techs(working_dir)
        all_techs = set(pre["Tech"].dropna().unique())
        not_allowed = all_techs - allowed

        run(working_dir)
        post = _read(working_dir / PARAM, "Secondary Techs")

        for tech in not_allowed:
            pre_row = pre[(pre["Tech"] == tech) & (pre["Parameter"] == MAX_INV_PARAM)]
            post_row = post[(post["Tech"] == tech) & (post["Parameter"] == MAX_INV_PARAM)]
            if len(pre_row) == 0:
                continue
            pre_vals = pre_row.iloc[0][year_cols].fillna("NA").tolist()
            post_vals = post_row.iloc[0][year_cols].fillna("NA").tolist()
            assert pre_vals == post_vals, (
                f"Non-allowed tech {tech}: MaxInv row was modified"
            )


# ---------------------------------------------------------------------------
# Projection.Mode flips
# ---------------------------------------------------------------------------
class TestProjectionModeFlip:
    def test_modified_rows_get_user_defined(self, working_dir):
        """For any row where we wrote at least one cell, Projection.Mode
        must end up != EMPTY."""
        log = run(working_dir)
        post = _read(working_dir / PARAM, "Secondary Techs")
        techs_with_changes = {c["tech"] for c in log["sheets"][0]["changes"]}
        for tech in techs_with_changes:
            mode = post[
                (post["Tech"] == tech) & (post["Parameter"] == MAX_INV_PARAM)
            ][PROJ_MODE_COL].iloc[0]
            assert mode != "EMPTY", f"{tech}: Projection.Mode still EMPTY"

    def test_unmodified_rows_mode_unchanged(self, working_dir):
        """Tech rows we didn't touch keep their original Projection.Mode."""
        pre = _read(working_dir / PARAM, "Secondary Techs")
        log = run(working_dir)
        post = _read(working_dir / PARAM, "Secondary Techs")
        techs_with_changes = {c["tech"] for c in log["sheets"][0]["changes"]}
        for tech in pre["Tech"].dropna().unique():
            if tech in techs_with_changes:
                continue
            pre_mode = pre[
                (pre["Tech"] == tech) & (pre["Parameter"] == MAX_INV_PARAM)
            ][PROJ_MODE_COL].iloc[0]
            post_mode = post[
                (post["Tech"] == tech) & (post["Parameter"] == MAX_INV_PARAM)
            ][PROJ_MODE_COL].iloc[0]
            assert pre_mode == post_mode, (
                f"{tech}: mode changed {pre_mode!r} -> {post_mode!r} "
                f"despite no cell changes"
            )


# ---------------------------------------------------------------------------
# Idempotency
# ---------------------------------------------------------------------------
class TestIdempotency:
    def test_second_run_no_changes(self, working_dir):
        run(working_dir)
        log2 = run(working_dir)
        assert log2["sheets"][0]["changes"] == [], (
            f"Second run produced changes: {log2['sheets'][0]['changes'][:3]}"
        )

    def test_second_run_file_identical(self, working_dir):
        run(working_dir)
        df1 = _read(working_dir / PARAM, "Secondary Techs")
        run(working_dir)
        df2 = _read(working_dir / PARAM, "Secondary Techs")
        pd.testing.assert_frame_equal(df1, df2, check_dtype=False)


# ---------------------------------------------------------------------------
# Change-log self-consistency
# ---------------------------------------------------------------------------
class TestChangeLogConsistency:
    def test_each_logged_change_matches_post_state(self, working_dir):
        log = run(working_dir)
        df = _read(working_dir / PARAM, "Secondary Techs")
        for c in log["sheets"][0]["changes"]:
            row = df[(df["Tech"] == c["tech"]) & (df["Parameter"] == MAX_INV_PARAM)]
            assert len(row) == 1
            actual = row.iloc[0][c["year"]]
            assert abs(float(actual) - float(c["new"])) < 1e-9, (
                f"{c['tech']} {c['year']}: logged new={c['new']}, actual={actual}"
            )

    def test_change_reasons_are_valid(self, working_dir):
        log = run(working_dir)
        valid = {"lid_fill", "preserved_manual", "untie_min_inv"}
        for c in log["sheets"][0]["changes"]:
            assert c["reason"] in valid, f"unexpected reason: {c['reason']!r}"


# ---------------------------------------------------------------------------
# Composability with the first-patch script
# ---------------------------------------------------------------------------
class TestRestore:
    """Tests for the --restore / restore_from_backup feature."""

    def test_restore_reverts_to_backup_state(self, working_dir):
        """After running the lid script then restoring, the input dir is
        byte-equivalent to its pre-edit state."""
        from add_max_cap_investment_lid_rule import restore_from_backup
        pre_hash = _md5(working_dir / PARAM)
        run(working_dir)
        post_hash = _md5(working_dir / PARAM)
        assert post_hash != pre_hash
        restore_from_backup(working_dir)
        restored_hash = _md5(working_dir / PARAM)
        assert restored_hash == pre_hash

    def test_restore_creates_post_lid_safety_snapshot(self, working_dir):
        """--restore makes a _POST_LID_pre_restore_* snapshot of the
        modified state before overwriting, so the restore is reversible."""
        from add_max_cap_investment_lid_rule import restore_from_backup
        run(working_dir)
        modified_hash = _md5(working_dir / PARAM)
        restore_from_backup(working_dir)
        snapshots = list(working_dir.parent.glob(
            f"{working_dir.name}_POST_LID_pre_restore_*"
        ))
        assert len(snapshots) == 1, (
            f"Expected exactly one safety snapshot, got {snapshots}"
        )
        snap_hash = _md5(snapshots[0] / PARAM)
        assert snap_hash == modified_hash

    def test_restore_from_specific_backup(self, working_dir):
        """restore_from_backup with explicit backup_dir uses that one."""
        from add_max_cap_investment_lid_rule import restore_from_backup
        log = run(working_dir)
        backup1 = Path(log["backup_dir"])
        run(working_dir)  # second modification, creates a second backup
        restore_from_backup(working_dir, backup1)
        assert _md5(working_dir / PARAM) == _md5(backup1 / PARAM)

    def test_restore_with_no_backup_raises(self, tmp_path):
        """If no _PRE_LID_* backup exists, restore raises FileNotFoundError."""
        from add_max_cap_investment_lid_rule import restore_from_backup
        target = tmp_path / "fresh_dir"
        shutil.copytree(SOURCE_DIR, target)
        with pytest.raises(FileNotFoundError):
            restore_from_backup(target)


class TestComposability:
    def test_running_after_first_patch_still_correct_for_pwr_techs(self, working_dir):
        """For PWR* allowed techs (which the lid script does modify), running
        'lid script alone' vs 'first patch then lid script' must produce
        identical MaxInv values. TRN* techs are excluded from this invariant
        because the lid script skips them — they may differ depending on
        whether the first patch ran (which fills empty cells with 9999).
        """
        from add_max_capacity_investment_rule import run as first_patch_run
        from add_max_cap_investment_lid_rule import country_region_for

        path_a = working_dir
        path_b = working_dir.parent / "path_b"
        shutil.copytree(SOURCE_DIR, path_b)

        run(path_a)
        first_patch_run(path_b, sheets=["Secondary Techs"])
        run(path_b)

        a = _read(path_a / PARAM, "Secondary Techs")
        b = _read(path_b / PARAM, "Secondary Techs")
        year_cols = [c for c in a.columns if isinstance(c, int)]
        a_max = a[a["Parameter"] == MAX_INV_PARAM].set_index("Tech")[year_cols]
        b_max = b[b["Parameter"] == MAX_INV_PARAM].set_index("Tech")[year_cols]

        allowed = _allowed_techs(path_a)
        # Only compare PWR* techs (length-11). TRN* are skipped by lid script,
        # so they retain whatever the previous step (or no previous step) left.
        pwr_allowed = {t for t in allowed if country_region_for(t) is not None}
        for tech in pwr_allowed:
            for yr in year_cols:
                va = a_max.loc[tech, yr]
                vb = b_max.loc[tech, yr]
                if pd.isna(va) and pd.isna(vb):
                    continue
                assert abs(float(va) - float(vb)) < 1e-9, (
                    f"{tech} {yr}: path A={va} vs path B={vb}"
                )


# ---------------------------------------------------------------------------
# Proportional mode
# ---------------------------------------------------------------------------
class TestProportionalMode:
    """Proportional mode distributes pool growth among allowed gen techs in
    proportion to their 2024 fleet share, with a security factor for slack.

        tech_share(t)     = ResidualCapacity(t, ref_year) / pool(cr(t), ref_year)
        pool_delta(cr, y) = max(0, scaled_pool(cr, y) - scaled_pool(cr, y-1))
        lid(t, y)         = LID_SECURITY_FACTOR * tech_share(t) * pool_delta(cr, y)

    Switched on via LID_RULE_MODE = "proportional"; each test in this class
    monkeypatches the mode for its duration.
    """

    def test_tech_shares_sum_to_one_per_cr(self, working_dir):
        """Per-cr shares should sum to 1.0 (mod float) over the techs that
        contribute to the cr's 2024 pool. Allowed techs with zero residual at
        ref_year contribute share=0 and don't break the invariant."""
        from add_max_cap_investment_lid_rule import (
            build_tech_share_map, identify_allowed_techs, country_region_for,
        )
        pre = _read(working_dir / PARAM, "Secondary Techs")
        year_cols = [c for c in pre.columns if isinstance(c, int)]
        gen = _gen_techs()
        allowed = identify_allowed_techs(pre, year_cols, gen)
        shares = build_tech_share_map(pre, allowed)

        # Group shares by cr
        per_cr: dict = {}
        for t, s in shares.items():
            cr = country_region_for(t)
            if cr is None:
                continue
            per_cr.setdefault(cr, []).append(s)

        for cr, vals in per_cr.items():
            total = sum(vals)
            # Allow either ~1.0 (cr has nonzero pool) or 0.0 (cr's allowed
            # techs all have zero residual at ref_year — edge case).
            assert abs(total - 1.0) < 1e-9 or total == 0.0, (
                f"{cr}: shares sum to {total}, expected 1.0 or 0.0"
            )

    def test_pool_delta_first_year_is_zero(self, working_dir):
        """The earliest year in year_cols has no prior year, so pool_delta=0."""
        from add_max_cap_investment_lid_rule import (
            build_pool_map, build_pool_delta_map, build_scaled_pool_map,
            identify_allowed_techs,
        )
        pre = _read(working_dir / PARAM, "Secondary Techs")
        year_cols = [c for c in pre.columns if isinstance(c, int)]
        gen = _gen_techs()
        allowed = identify_allowed_techs(pre, year_cols, gen)
        pool_map = build_pool_map(pre, allowed, year_cols)
        demand_mult = _demand_mult(working_dir)
        scaled = build_scaled_pool_map(pool_map, demand_mult)
        delta = build_pool_delta_map(scaled, year_cols)

        first_year = min(year_cols)
        # All crs should have delta=0 at the first year.
        crs_with_data = {cr for (cr, _) in delta.keys()}
        assert crs_with_data, "No crs found in pool_delta_map"
        for cr in crs_with_data:
            assert delta[(cr, first_year)] == 0.0, (
                f"{cr} {first_year}: pool_delta={delta[(cr, first_year)]}, "
                f"expected 0.0 (no prior year)"
            )

    def test_pool_delta_nonneg_guard(self, working_dir):
        """Even if a year has lower scaled_pool than the prior year, the
        pool_delta should be 0 (not negative). We can't easily force a dip
        with the production demand data (monotone growth), so we exercise the
        guard with a synthetic scaled_pool map."""
        from add_max_cap_investment_lid_rule import build_pool_delta_map

        years = [2024, 2025, 2026, 2027]
        synthetic_scaled = {
            ("XX1", 2024): 100.0,
            ("XX1", 2025): 110.0,
            ("XX1", 2026): 105.0,   # decline year
            ("XX1", 2027): 120.0,
        }
        delta = build_pool_delta_map(synthetic_scaled, years)
        assert delta[("XX1", 2024)] == 0.0       # first year
        assert delta[("XX1", 2025)] == 10.0      # +10
        assert delta[("XX1", 2026)] == 0.0       # would be -5, guarded
        assert delta[("XX1", 2027)] == 15.0      # +15

    def test_proportional_lid_formula(self, working_dir):
        """For a representative tech (PWRHYDLKAXX 2050), verify the lid
        equals LID_SECURITY_FACTOR * tech_share * pool_delta."""
        from add_max_cap_investment_lid_rule import (
            build_tech_share_map, build_pool_map, build_pool_delta_map,
            build_scaled_pool_map, identify_allowed_techs,
            proportional_lid_for_tech_year, LID_SECURITY_FACTOR,
        )
        pre = _read(working_dir / PARAM, "Secondary Techs")
        year_cols = [c for c in pre.columns if isinstance(c, int)]
        gen = _gen_techs()
        allowed = identify_allowed_techs(pre, year_cols, gen)
        pool_map = build_pool_map(pre, allowed, year_cols)
        demand_mult = _demand_mult(working_dir)
        scaled = build_scaled_pool_map(pool_map, demand_mult)
        delta = build_pool_delta_map(scaled, year_cols)
        shares = build_tech_share_map(pre, allowed)

        tech, year = "PWRHYDLKAXX", 2050
        if tech not in allowed:
            pytest.skip(f"{tech} not in allowed; cannot exercise this test")

        expected = LID_SECURITY_FACTOR * shares[tech] * delta[("LKAXX", year)]
        actual = proportional_lid_for_tech_year(tech, year, shares, delta)
        assert abs(actual - expected) < 1e-12, (
            f"{tech} {year}: lid={actual}, expected {expected} "
            f"(security={LID_SECURITY_FACTOR}, share={shares[tech]}, "
            f"delta={delta[('LKAXX', year)]})"
        )
        # Sanity: a tech with positive 2024 residual in a growing-demand cr
        # should get a positive lid.
        assert actual > 0, f"{tech} {year}: lid={actual}, expected > 0"

    def test_run_in_proportional_mode_writes_per_tech_lids(
        self, working_dir, monkeypatch
    ):
        """End-to-end: switch to proportional mode, run the script, and verify
        that two techs in the same cr with different 2024 fleet shares end up
        with different MaxInv values in 2050. (In uniform mode they'd be
        identical because the same lid is applied to all techs in cr.)"""
        import add_max_cap_investment_lid_rule as mod
        monkeypatch.setattr(mod, "LID_RULE_MODE", "proportional")

        log = run(working_dir)
        assert log["lid_rule_mode"] == "proportional", log["lid_rule_mode"]
        assert log["lid_security_factor"] == mod.LID_SECURITY_FACTOR

        post = _read(working_dir / PARAM, "Secondary Techs")
        # PWRHYDLKAXX has the largest 2024 share; PWRWONLKAXX is small.
        # Their 2050 lids should differ — and the bigger-share tech should
        # have the bigger lid.
        hyd = post[
            (post["Tech"] == "PWRHYDLKAXX") & (post["Parameter"] == MAX_INV_PARAM)
        ].iloc[0][2050]
        won = post[
            (post["Tech"] == "PWRWONLKAXX") & (post["Parameter"] == MAX_INV_PARAM)
        ].iloc[0][2050]
        # If either is overridden by the untie rule (MinCapInv > lid), the
        # comparison may not be meaningful. Floor-untie is fine; we just check
        # they're not identical (which would indicate uniform-mode behaviour
        # leaked through).
        if pd.notna(hyd) and pd.notna(won):
            assert float(hyd) != float(won), (
                f"Proportional mode should give per-tech lids; both LKAXX "
                f"techs have MaxInv 2050 = {hyd} (uniform-mode leak?)"
            )
            # Hyd's 2024 share is ~0.36, Won's is ~0.05. Hyd's lid should
            # therefore be larger (modulo untie corrections).
            assert float(hyd) > float(won), (
                f"PWRHYDLKAXX 2050 lid ({hyd}) should exceed "
                f"PWRWONLKAXX 2050 lid ({won}) — share order is "
                f"hyd > won by ~7x in residual."
            )

    def test_security_factor_scales_lid_linearly(self, working_dir):
        """Lid scales linearly in LID_SECURITY_FACTOR — doubling the factor
        doubles every per-tech lid."""
        from add_max_cap_investment_lid_rule import (
            build_tech_share_map, build_pool_map, build_pool_delta_map,
            build_scaled_pool_map, identify_allowed_techs,
            proportional_lid_for_tech_year,
        )
        import add_max_cap_investment_lid_rule as mod

        pre = _read(working_dir / PARAM, "Secondary Techs")
        year_cols = [c for c in pre.columns if isinstance(c, int)]
        gen = _gen_techs()
        allowed = identify_allowed_techs(pre, year_cols, gen)
        pool_map = build_pool_map(pre, allowed, year_cols)
        demand_mult = _demand_mult(working_dir)
        scaled = build_scaled_pool_map(pool_map, demand_mult)
        delta = build_pool_delta_map(scaled, year_cols)
        shares = build_tech_share_map(pre, allowed)

        tech, year = "PWRHYDLKAXX", 2050
        if tech not in allowed:
            pytest.skip(f"{tech} not in allowed")

        original = mod.LID_SECURITY_FACTOR
        try:
            mod.LID_SECURITY_FACTOR = 1.0
            base = proportional_lid_for_tech_year(tech, year, shares, delta)
            mod.LID_SECURITY_FACTOR = 2.0
            doubled = proportional_lid_for_tech_year(tech, year, shares, delta)
            assert abs(doubled - 2.0 * base) < 1e-12, (
                f"Doubling security factor should double lid: "
                f"base={base}, doubled={doubled}"
            )
        finally:
            mod.LID_SECURITY_FACTOR = original
