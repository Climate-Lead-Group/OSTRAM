# -*- coding: utf-8 -*-
"""
Reference Energy System (RES) Diagram Generator for OSTRAM

Generates a standalone interactive HTML file with a Sankey-style RES diagram
showing the energy flow from primary resources through power generation,
transmission, and dispatch for each country/region.

Features:
  - Multi-region selector (checkboxes) to observe interconnections
  - Show technology/fuel codes as labels, full names on hover
  - Zoom & pan: mouse wheel to zoom, click-drag to pan
  - Color-coded flows: green for renewable, brown for fossil, blue for nuclear
  - Interactive Plotly Sankey with hover info
  - Fully standalone (no server required)

No countries, technologies, or regions are hardcoded.
Everything is discovered dynamically from the input data.

Author: Climate Lead Group
"""

import pandas as pd
import json
import re
import yaml
from pathlib import Path
from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent

with open(SCRIPT_DIR / 'Config_country_codes.yaml', 'r', encoding='utf-8') as _f:
    _CONFIG = yaml.safe_load(_f)

COUNTRY_NAME_MAP = {
    code: info['english_name']
    for code, info in _CONFIG['country_data'].items()
}
COUNTRY_NAME_MAP.update(_CONFIG.get('special_entries', {}))

RENEWABLE_FUELS = {'BIO', 'HYD', 'CSP', 'GEO', 'SPV', 'WAS', 'WON', 'WOF', 'WAV'}
FOSSIL_FUELS = {'COA', 'GAS', 'OIL', 'PET', 'COG', 'OTH', 'LDS', 'SDS', 'BCK'}
NUCLEAR_FUELS = {'URN'}

# Color palette
COLOR_RENEWABLE   = 'rgba(46, 160, 67, 0.6)'
COLOR_FOSSIL      = 'rgba(180, 120, 60, 0.6)'
COLOR_NUCLEAR     = 'rgba(80, 120, 200, 0.6)'
COLOR_ELECTRICITY = 'rgba(255, 180, 30, 0.6)'
COLOR_DISPATCH    = 'rgba(130, 90, 160, 0.6)'
COLOR_IMPORT      = 'rgba(200, 80, 80, 0.6)'
COLOR_TRANSMISS   = 'rgba(100, 160, 200, 0.6)'

NODE_COLOR_RESOURCE_REN  = 'rgba(46, 160, 67, 0.85)'
NODE_COLOR_RESOURCE_FOS  = 'rgba(180, 120, 60, 0.85)'
NODE_COLOR_RESOURCE_NUC  = 'rgba(80, 120, 200, 0.85)'
NODE_COLOR_TECH          = 'rgba(90, 90, 90, 0.75)'
NODE_COLOR_ELEC          = 'rgba(255, 180, 30, 0.85)'
NODE_COLOR_DISPATCH      = 'rgba(130, 90, 160, 0.85)'
NODE_COLOR_IMPORT        = 'rgba(200, 80, 80, 0.85)'


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------
def load_base_year_data(xlsx_path):
    """Load all sheets from the base year Excel file into structured dicts.

    Returns a list of link dicts:
        {fuel_in, fuel_in_name, tech, tech_name, fuel_out, fuel_out_name, mode}
    """
    wb = load_workbook(xlsx_path, data_only=True)
    links = []

    # --- Primary sheet: no fuel_in, only tech -> fuel_out ---
    ws = wb['Primary']
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        mode, tech, tech_name, fuel_o, fuel_o_name, val_o, unit_o = row[:7]
        if tech is None:
            continue
        links.append({
            'fuel_in': None,
            'fuel_in_name': None,
            'tech': str(tech),
            'tech_name': str(tech_name) if tech_name else str(tech),
            'fuel_out': str(fuel_o),
            'fuel_out_name': str(fuel_o_name) if fuel_o_name else str(fuel_o),
            'mode': mode,
        })

    # --- Secondary, Demand Techs, Transport Groups: fuel_in -> tech -> fuel_out ---
    for sheet_name in ['Secondary', 'Demand Techs', 'Transport Groups']:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            vals = list(row)
            if all(v is None for v in vals):
                continue
            mode = vals[0]
            fuel_i = vals[1]
            fuel_i_name = vals[2]
            tech = vals[5]
            tech_name = vals[6]
            fuel_o = vals[7]
            fuel_o_name = vals[8]
            if tech is None:
                continue
            links.append({
                'fuel_in': str(fuel_i) if fuel_i else None,
                'fuel_in_name': str(fuel_i_name) if fuel_i_name else None,
                'tech': str(tech),
                'tech_name': str(tech_name) if tech_name else str(tech),
                'fuel_out': str(fuel_o) if fuel_o else None,
                'fuel_out_name': str(fuel_o_name) if fuel_o_name else None,
                'mode': mode,
            })

    wb.close()
    return links


def extract_region(code):
    """Extract the 5-char region identifier from a technology or fuel code."""
    if not code:
        return None
    m = re.match(r'^ELC([A-Z]{5})\d{2}$', code)
    if m:
        return m.group(1)
    if re.match(r'^TRN[A-Z]{10}$', code):
        return code[3:8]
    m = re.search(r'([A-Z]{3}(?:XX|EA|NE|NO|SO|WE))$', code)
    if m:
        return m.group(1)
    m = re.match(r'^(?:MIN|RNW)?[A-Z]{3}([A-Z]{3})$', code)
    if m:
        return m.group(1) + 'XX'
    return None


def discover_regions(links):
    """Discover all unique regions from ELC fuel codes."""
    regions = set()
    elc_pattern = re.compile(r'^ELC([A-Z]{5})\d{2}$')
    for link in links:
        for code in [link.get('fuel_in'), link.get('fuel_out')]:
            if code:
                m = elc_pattern.match(code)
                if m:
                    regions.add(m.group(1))
    return sorted(regions)


def _region_label(region_code):
    """Human-readable label for a region code like BGDXX or INDEA."""
    sub_map = {
        'XX': '', 'EA': 'East', 'NE': 'Northeast',
        'NO': 'North', 'SO': 'South', 'WE': 'West',
    }
    country = COUNTRY_NAME_MAP.get(region_code[:3], region_code[:3])
    sub = sub_map.get(region_code[3:], region_code[3:])
    if sub:
        return f"{country} ({sub})"
    return country


# ---------------------------------------------------------------------------
# HTML generation
# ---------------------------------------------------------------------------
def generate_html(all_links, regions, output_path):
    """Generate a standalone interactive HTML file with the RES diagram."""

    # Serialize raw links for client-side filtering (drop mode, not needed for visual)
    serializable_links = []
    for link in all_links:
        serializable_links.append({
            'fi': link.get('fuel_in'),
            'fin': link.get('fuel_in_name'),
            'te': link['tech'],
            'ten': link['tech_name'],
            'fo': link.get('fuel_out'),
            'fon': link.get('fuel_out_name'),
        })

    # Deduplicate
    seen = set()
    deduped = []
    for l in serializable_links:
        key = (l['fi'], l['te'], l['fo'])
        if key not in seen:
            seen.add(key)
            deduped.append(l)
    serializable_links = deduped

    # Region options
    region_options = {}
    for r in sorted(regions):
        region_options[r] = _region_label(r)

    links_json = json.dumps(serializable_links)
    regions_json = json.dumps(region_options)

    # Color constants as JS-friendly values
    renewable_fuels_js = json.dumps(sorted(RENEWABLE_FUELS))
    fossil_fuels_js = json.dumps(sorted(FOSSIL_FUELS))
    nuclear_fuels_js = json.dumps(sorted(NUCLEAR_FUELS))

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>OSTRAM - Reference Energy System (RES)</title>
<script src="https://cdn.plot.ly/plotly-2.27.0.min.js"></script>
<style>
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  body {{
    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
    background: #1a1a2e;
    color: #e0e0e0;
    min-height: 100vh;
    overflow: hidden;
  }}
  .header {{
    background: linear-gradient(135deg, #16213e 0%, #0f3460 100%);
    padding: 12px 24px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    box-shadow: 0 2px 10px rgba(0,0,0,0.3);
    z-index: 100;
    position: relative;
  }}
  .header h1 {{
    font-size: 1.3em;
    font-weight: 600;
    color: #e0e0e0;
  }}
  .header h1 span {{
    color: #f5a623;
    font-weight: 700;
  }}
  .controls {{
    display: flex;
    gap: 12px;
    align-items: center;
  }}

  /* --- Multi-select dropdown --- */
  .region-dropdown {{
    position: relative;
    display: inline-block;
  }}
  .region-btn {{
    padding: 6px 16px;
    border-radius: 6px;
    border: 1px solid #3a3a5c;
    background: #16213e;
    color: #e0e0e0;
    font-size: 0.9em;
    cursor: pointer;
    user-select: none;
    display: flex;
    align-items: center;
    gap: 6px;
  }}
  .region-btn:hover {{ border-color: #f5a623; }}
  .region-btn .arrow {{ font-size: 0.7em; }}
  .region-panel {{
    display: none;
    position: absolute;
    top: 100%;
    right: 0;
    margin-top: 4px;
    background: #16213e;
    border: 1px solid #3a3a5c;
    border-radius: 8px;
    padding: 10px;
    min-width: 260px;
    max-height: 400px;
    overflow-y: auto;
    z-index: 200;
    box-shadow: 0 8px 24px rgba(0,0,0,0.5);
  }}
  .region-panel.open {{ display: block; }}
  .region-panel .panel-actions {{
    display: flex;
    gap: 8px;
    margin-bottom: 8px;
    padding-bottom: 8px;
    border-bottom: 1px solid #3a3a5c;
  }}
  .panel-action-btn {{
    padding: 3px 10px;
    border-radius: 4px;
    border: 1px solid #3a3a5c;
    background: transparent;
    color: #a0a0b0;
    font-size: 0.8em;
    cursor: pointer;
  }}
  .panel-action-btn:hover {{ background: #0f3460; color: #f5a623; }}
  .region-panel label {{
    display: flex;
    align-items: center;
    gap: 8px;
    padding: 4px 6px;
    border-radius: 4px;
    cursor: pointer;
    font-size: 0.85em;
    color: #c0c0d0;
  }}
  .region-panel label:hover {{ background: rgba(245,166,35,0.1); }}
  .region-panel input[type="checkbox"] {{
    accent-color: #f5a623;
    width: 15px;
    height: 15px;
  }}
  .region-code {{
    font-family: 'Consolas', 'Courier New', monospace;
    font-size: 0.85em;
    color: #f5a623;
    min-width: 50px;
  }}

  /* --- Zoom controls --- */
  .zoom-controls {{
    display: flex;
    gap: 4px;
    align-items: center;
  }}
  .zoom-btn {{
    width: 32px;
    height: 32px;
    border-radius: 6px;
    border: 1px solid #3a3a5c;
    background: #16213e;
    color: #e0e0e0;
    font-size: 1.1em;
    cursor: pointer;
    display: flex;
    align-items: center;
    justify-content: center;
  }}
  .zoom-btn:hover {{ border-color: #f5a623; color: #f5a623; }}
  .zoom-label {{
    font-size: 0.8em;
    color: #707090;
    min-width: 40px;
    text-align: center;
  }}

  /* --- Legend --- */
  .legend {{
    display: flex;
    gap: 16px;
    padding: 6px 24px;
    background: #16213e;
    flex-wrap: wrap;
    justify-content: center;
    z-index: 90;
    position: relative;
  }}
  .legend-item {{
    display: flex;
    align-items: center;
    gap: 5px;
    font-size: 0.75em;
    color: #c0c0d0;
  }}
  .legend-dot {{
    width: 12px;
    height: 12px;
    border-radius: 3px;
    flex-shrink: 0;
  }}

  /* --- Sankey viewport with pan/zoom --- */
  #viewport {{
    width: 100%;
    height: calc(100vh - 95px);
    overflow: hidden;
    position: relative;
    cursor: grab;
  }}
  #viewport.dragging {{ cursor: grabbing; }}
  #sankey-inner {{
    transform-origin: 0 0;
    position: absolute;
    top: 0;
    left: 0;
  }}

  .info-text {{
    position: fixed;
    bottom: 4px;
    left: 0;
    right: 0;
    text-align: center;
    font-size: 0.7em;
    color: #505070;
    pointer-events: none;
    z-index: 50;
  }}
</style>
</head>
<body>

<div class="header">
  <h1><span>OSTRAM</span> &mdash; Reference Energy System (RES)</h1>
  <div class="controls">
    <div class="zoom-controls">
      <button class="zoom-btn" id="zoom-out" title="Zoom out">&#8722;</button>
      <span class="zoom-label" id="zoom-label">100%</span>
      <button class="zoom-btn" id="zoom-in" title="Zoom in">+</button>
      <button class="zoom-btn" id="zoom-reset" title="Reset view" style="font-size:0.8em;">&#8634;</button>
    </div>
    <div class="region-dropdown">
      <div class="region-btn" id="region-toggle">
        <span id="region-summary">All regions</span>
        <span class="arrow">&#9660;</span>
      </div>
      <div class="region-panel" id="region-panel">
        <div class="panel-actions">
          <button class="panel-action-btn" id="btn-all">Select All</button>
          <button class="panel-action-btn" id="btn-none">Clear All</button>
        </div>
        <div id="region-list"></div>
      </div>
    </div>
  </div>
</div>

<div class="legend">
  <div class="legend-item"><div class="legend-dot" style="background:{NODE_COLOR_RESOURCE_REN}"></div> Renewable</div>
  <div class="legend-item"><div class="legend-dot" style="background:{NODE_COLOR_RESOURCE_FOS}"></div> Fossil</div>
  <div class="legend-item"><div class="legend-dot" style="background:{NODE_COLOR_RESOURCE_NUC}"></div> Nuclear</div>
  <div class="legend-item"><div class="legend-dot" style="background:{NODE_COLOR_TECH}"></div> Technology</div>
  <div class="legend-item"><div class="legend-dot" style="background:{NODE_COLOR_ELEC}"></div> Electricity</div>
  <div class="legend-item"><div class="legend-dot" style="background:{NODE_COLOR_DISPATCH}"></div> Dispatch</div>
  <div class="legend-item"><div class="legend-dot" style="background:{NODE_COLOR_IMPORT}"></div> Import</div>
  <div class="legend-item"><div class="legend-dot" style="background:{COLOR_TRANSMISS}"></div> Transmission</div>
</div>

<div id="viewport">
  <div id="sankey-inner"></div>
</div>

<div class="info-text">
  Scroll to zoom &bull; Drag to pan &bull; Drag nodes to rearrange &bull; Hover for full names
</div>

<script>
// =========================================================================
// Data
// =========================================================================
const ALL_LINKS = {links_json};
const REGION_OPTIONS = {regions_json};
const RENEWABLE_SET = new Set({renewable_fuels_js});
const NUCLEAR_SET = new Set({nuclear_fuels_js});

// =========================================================================
// Color functions
// =========================================================================
function fuelColor(code) {{
  const p = code.slice(0,3);
  if (RENEWABLE_SET.has(p)) return '{COLOR_RENEWABLE}';
  if (NUCLEAR_SET.has(p))   return '{COLOR_NUCLEAR}';
  if (p === 'ELC') {{
    const s = code.slice(-2);
    if (s === '00') return '{COLOR_RENEWABLE}';
    if (s === '01') return '{COLOR_FOSSIL}';
    if (s === '02') return '{COLOR_TRANSMISS}';
    if (s === '03') return '{COLOR_DISPATCH}';
    if (s === '04') return '{COLOR_IMPORT}';
    return '{COLOR_ELECTRICITY}';
  }}
  return '{COLOR_FOSSIL}';
}}

function nodeColor(code, ntype) {{
  if (ntype === 'fuel') {{
    const p = code.slice(0,3);
    if (RENEWABLE_SET.has(p)) return '{NODE_COLOR_RESOURCE_REN}';
    if (NUCLEAR_SET.has(p))   return '{NODE_COLOR_RESOURCE_NUC}';
    if (p === 'ELC') {{
      const s = code.slice(-2);
      if (s === '04') return '{NODE_COLOR_IMPORT}';
      if (s === '03') return '{NODE_COLOR_DISPATCH}';
      return '{NODE_COLOR_ELEC}';
    }}
    return '{NODE_COLOR_RESOURCE_FOS}';
  }}
  return '{NODE_COLOR_TECH}';
}}

// =========================================================================
// Region extraction (mirrors Python logic)
// =========================================================================
function extractRegion(code) {{
  if (!code) return null;
  let m = code.match(/^ELC([A-Z]{{5}})\d{{2}}$/);
  if (m) return m[1];
  if (/^TRN[A-Z]{{10}}$/.test(code)) return code.slice(3,8);
  m = code.match(/([A-Z]{{3}}(?:XX|EA|NE|NO|SO|WE))$/);
  if (m) return m[1];
  m = code.match(/^(?:MIN|RNW)?[A-Z]{{3}}([A-Z]{{3}})$/);
  if (m) return m[1] + 'XX';
  return null;
}}

function extractTrnRegions(techCode) {{
  if (/^TRN[A-Z]{{10}}$/.test(techCode)) {{
    return [techCode.slice(3,8), techCode.slice(8,13)];
  }}
  return null;
}}

// =========================================================================
// Filtering
// =========================================================================
function filterLinks(links, selectedRegions) {{
  if (selectedRegions.size === 0) return [];
  const allMode = selectedRegions.has('__ALL__');
  if (allMode) return links;

  const result = [];
  for (const link of links) {{
    const techRegion = extractRegion(link.te);

    // Direct region match
    if (techRegion && selectedRegions.has(techRegion)) {{
      result.push(link);
      continue;
    }}

    // National XX match: if region is XXXXX and tech is from same country
    if (techRegion) {{
      for (const sel of selectedRegions) {{
        if (sel.slice(0,3) === techRegion.slice(0,3) && sel.slice(3) === 'XX') {{
          result.push(link);
          break;
        }}
      }}
      if (result[result.length - 1] === link) continue;
    }}

    // Interconnection: include if either end is selected
    const trn = extractTrnRegions(link.te);
    if (trn) {{
      if (selectedRegions.has(trn[0]) || selectedRegions.has(trn[1])) {{
        result.push(link);
      }}
    }}
  }}
  return result;
}}

// =========================================================================
// Build Sankey data
// =========================================================================
function buildSankey(links) {{
  const nodeMap = {{}};
  const nodeLabels = [];
  const nodeColors = [];
  const nodeCustom = [];
  let idx = 0;

  function addNode(code, name, ntype) {{
    if (!(code in nodeMap)) {{
      nodeMap[code] = idx++;
      nodeLabels.push(code);                          // <-- CODE as label
      nodeColors.push(nodeColor(code, ntype));
      nodeCustom.push(name ? (code + '<br>' + name) : code);  // hover: code + name
    }}
  }}

  const sources = [], targets = [], values = [];
  const linkColors = [], linkLabels = [];

  for (const link of links) {{
    addNode(link.te, link.ten, 'tech');

    if (link.fi) {{
      addNode(link.fi, link.fin, 'fuel');
      sources.push(nodeMap[link.fi]);
      targets.push(nodeMap[link.te]);
      values.push(1);
      linkColors.push(fuelColor(link.fi));
      linkLabels.push((link.fin || link.fi) + ' \\u2192 ' + (link.ten || link.te));
    }}

    if (link.fo) {{
      addNode(link.fo, link.fon, 'fuel');
      sources.push(nodeMap[link.te]);
      targets.push(nodeMap[link.fo]);
      values.push(1);
      linkColors.push(fuelColor(link.fo));
      linkLabels.push((link.ten || link.te) + ' \\u2192 ' + (link.fon || link.fo));
    }}
  }}

  return {{
    node_labels: nodeLabels,
    node_colors: nodeColors,
    node_customdata: nodeCustom,
    sources, targets, values,
    link_colors: linkColors,
    link_labels: linkLabels,
  }};
}}

// =========================================================================
// Region selector UI
// =========================================================================
const regionList = document.getElementById('region-list');
const regionToggle = document.getElementById('region-toggle');
const regionPanel = document.getElementById('region-panel');
const regionSummary = document.getElementById('region-summary');

// Build checkboxes
const regionKeys = Object.keys(REGION_OPTIONS);
for (const key of regionKeys) {{
  const lbl = document.createElement('label');
  const cb = document.createElement('input');
  cb.type = 'checkbox';
  cb.value = key;
  cb.checked = true;
  cb.addEventListener('change', onRegionChange);
  const codeSpan = document.createElement('span');
  codeSpan.className = 'region-code';
  codeSpan.textContent = key;
  const nameSpan = document.createElement('span');
  nameSpan.textContent = REGION_OPTIONS[key];
  lbl.appendChild(cb);
  lbl.appendChild(codeSpan);
  lbl.appendChild(nameSpan);
  regionList.appendChild(lbl);
}}

// Toggle panel
regionToggle.addEventListener('click', (e) => {{
  e.stopPropagation();
  regionPanel.classList.toggle('open');
}});
document.addEventListener('click', (e) => {{
  if (!regionPanel.contains(e.target) && e.target !== regionToggle) {{
    regionPanel.classList.remove('open');
  }}
}});

// Select all / clear all
document.getElementById('btn-all').addEventListener('click', () => {{
  regionList.querySelectorAll('input').forEach(cb => cb.checked = true);
  onRegionChange();
}});
document.getElementById('btn-none').addEventListener('click', () => {{
  regionList.querySelectorAll('input').forEach(cb => cb.checked = false);
  onRegionChange();
}});

function getSelectedRegions() {{
  const sel = new Set();
  regionList.querySelectorAll('input:checked').forEach(cb => sel.add(cb.value));
  return sel;
}}

function updateSummary() {{
  const sel = getSelectedRegions();
  if (sel.size === 0) {{
    regionSummary.textContent = 'No regions';
  }} else if (sel.size === regionKeys.length) {{
    regionSummary.textContent = 'All regions';
  }} else if (sel.size <= 3) {{
    regionSummary.textContent = [...sel].join(', ');
  }} else {{
    regionSummary.textContent = sel.size + ' regions';
  }}
}}

function onRegionChange() {{
  updateSummary();
  renderSankey();
}}

// =========================================================================
// Pan & Zoom
// =========================================================================
const viewport = document.getElementById('viewport');
const inner = document.getElementById('sankey-inner');

let scale = 1;
let panX = 0, panY = 0;
let isDragging = false;
let dragStartX, dragStartY, panStartX, panStartY;

function applyTransform() {{
  inner.style.transform = `translate(${{panX}}px, ${{panY}}px) scale(${{scale}})`;
  document.getElementById('zoom-label').textContent = Math.round(scale * 100) + '%';
}}

viewport.addEventListener('wheel', (e) => {{
  e.preventDefault();
  const rect = viewport.getBoundingClientRect();
  const mx = e.clientX - rect.left;
  const my = e.clientY - rect.top;

  const oldScale = scale;
  const delta = e.deltaY > 0 ? 0.9 : 1.1;
  scale = Math.min(Math.max(scale * delta, 0.1), 5);

  // Zoom toward cursor
  panX = mx - (mx - panX) * (scale / oldScale);
  panY = my - (my - panY) * (scale / oldScale);
  applyTransform();
}}, {{ passive: false }});

viewport.addEventListener('mousedown', (e) => {{
  // Only pan on left click and not on a Plotly node
  if (e.button !== 0) return;
  isDragging = true;
  dragStartX = e.clientX;
  dragStartY = e.clientY;
  panStartX = panX;
  panStartY = panY;
  viewport.classList.add('dragging');
}});

window.addEventListener('mousemove', (e) => {{
  if (!isDragging) return;
  panX = panStartX + (e.clientX - dragStartX);
  panY = panStartY + (e.clientY - dragStartY);
  applyTransform();
}});

window.addEventListener('mouseup', () => {{
  isDragging = false;
  viewport.classList.remove('dragging');
}});

// Zoom buttons
document.getElementById('zoom-in').addEventListener('click', () => {{
  scale = Math.min(scale * 1.2, 5);
  applyTransform();
}});
document.getElementById('zoom-out').addEventListener('click', () => {{
  scale = Math.max(scale / 1.2, 0.1);
  applyTransform();
}});
document.getElementById('zoom-reset').addEventListener('click', () => {{
  scale = 1; panX = 0; panY = 0;
  applyTransform();
}});

// =========================================================================
// Render
// =========================================================================
function renderSankey() {{
  const selected = getSelectedRegions();
  const isAll = selected.size === regionKeys.length;
  const filtered = isAll
    ? ALL_LINKS
    : filterLinks(ALL_LINKS, selected);

  if (filtered.length === 0) {{
    Plotly.purge('sankey-inner');
    inner.innerHTML = '<div style="text-align:center;padding:80px;color:#707090;font-size:1.2em;">Select one or more regions to display.</div>';
    inner.style.width = '100%';
    inner.style.height = '100%';
    return;
  }}

  const d = buildSankey(filtered);

  // Scale canvas width with node count for readability
  const nNodes = d.node_labels.length;
  const canvasW = Math.max(window.innerWidth, nNodes * 20, 1600);
  const canvasH = Math.max(window.innerHeight - 95, nNodes * 8, 900);

  inner.style.width = canvasW + 'px';
  inner.style.height = canvasH + 'px';

  const trace = {{
    type: 'sankey',
    orientation: 'h',
    arrangement: 'snap',
    node: {{
      pad: 18,
      thickness: 20,
      line: {{ color: 'rgba(255,255,255,0.15)', width: 0.5 }},
      label: d.node_labels,
      color: d.node_colors,
      customdata: d.node_customdata,
      hovertemplate: '%{{customdata}}<extra></extra>',
    }},
    link: {{
      source: d.sources,
      target: d.targets,
      value: d.values,
      color: d.link_colors,
      customdata: d.link_labels,
      hovertemplate: '%{{customdata}}<extra></extra>',
    }},
  }};

  const layout = {{
    font: {{ size: 10, color: '#c0c0d0', family: 'Consolas, Courier New, monospace' }},
    paper_bgcolor: 'rgba(0,0,0,0)',
    plot_bgcolor: 'rgba(0,0,0,0)',
    margin: {{ l: 10, r: 10, t: 10, b: 10 }},
    width: canvasW,
    height: canvasH,
  }};

  Plotly.react('sankey-inner', [trace], layout, {{
    responsive: false,
    displayModeBar: true,
    modeBarButtonsToRemove: ['lasso2d', 'select2d', 'zoom2d', 'pan2d',
                              'zoomIn2d', 'zoomOut2d', 'autoScale2d', 'resetScale2d'],
    displaylogo: false,
    toImageButtonOptions: {{
      format: 'png',
      filename: 'OSTRAM_RES',
      height: 2400,
      width: 4800,
      scale: 2,
    }},
  }});
}}

// Initial render
updateSummary();
renderSankey();
</script>
</body>
</html>"""

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"  RES diagram saved to: {output_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    """Entry point: load data, build RES, generate HTML."""
    import time
    t0 = time.time()

    xlsx_path = SCRIPT_DIR / 'A1_Outputs' / 'A1_Outputs_BAU' / 'A-O_AR_Model_Base_Year.xlsx'
    output_path = SCRIPT_DIR / 'Figures' / 'RES_Diagram.html'

    print("OSTRAM RES Diagram Generator")
    print("=" * 50)

    if not xlsx_path.exists():
        print(f"  ERROR: File not found: {xlsx_path}")
        return

    print(f"  Loading base year data from: {xlsx_path.name}")
    links = load_base_year_data(xlsx_path)
    print(f"  Loaded {len(links)} activity ratio links")

    regions = discover_regions(links)
    print(f"  Discovered {len(regions)} regions: {', '.join(regions)}")

    print("  Generating HTML ...")
    generate_html(links, regions, output_path)

    elapsed = time.time() - t0
    print(f"  Done in {elapsed:.1f}s")


if __name__ == '__main__':
    main()
