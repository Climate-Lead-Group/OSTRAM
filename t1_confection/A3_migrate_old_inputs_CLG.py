"""
Migrate Old Inputs to New Model Structure

This script migrates data from Old_Inputs/ to the current model files,
applying technology name transformations (CCG+OCG -> NGS, suffix removal, etc.)
based on Config_tech_equivalences.yaml and Tech_Country_Matrix.xlsx.

Usage:
    python t1_confection/A0_migrate_old_inputs.py [--dry-run]

Options:
    --dry-run    Show what would be migrated without making changes

Author: Climate Lead Group, Andrey Salazar-Vargas
"""
import openpyxl
import yaml
import sys
import re
from pathlib import Path
from datetime import datetime
import shutil
from collections import defaultdict
from Z_AUX_config_loader import get_force_empty_max_capacity_investment_pwr                                                                           

# Import the CSV profile normalization script
try:
    from Z_AUX_fix_all_profiles_normalization import (
        normalize_specified_demand_profile,
        normalize_year_split,
        normalize_day_split
    )
    NORMALIZATION_AVAILABLE = True
    print("[INIT] ✓ Profile normalization module (CSV) loaded successfully")
except ImportError as e:
    NORMALIZATION_AVAILABLE = False
    print(f"[INIT] ✗ Failed to import profile normalization module: {e}")
    print("[INIT]   Normalization will be skipped")

# Import the Excel profile normalization script
try:
    from Z_AUX_fix_excel_profiles import normalize_excel_profiles
    EXCEL_NORMALIZATION_AVAILABLE = True
    print("[INIT] ✓ Excel profile normalization module loaded successfully")
except ImportError as e:
    EXCEL_NORMALIZATION_AVAILABLE = False
    print(f"[INIT] ✗ Failed to import Excel profile normalization module: {e}")
    print("[INIT]   Excel normalization will be skipped")


class TechCountryMatrix:
    """Reads and queries the Tech_Country_Matrix.xlsx file"""

    def __init__(self, matrix_path):
        self.matrix_path = matrix_path
        self.availability = {}  # {tech_code: {country_code: bool}}
        self._load_matrix()

    def _load_matrix(self):
        """Load the matrix from Excel"""
        if not self.matrix_path.exists():
            raise FileNotFoundError(f"Tech_Country_Matrix not found: {self.matrix_path}")

        wb = openpyxl.load_workbook(self.matrix_path, data_only=True)
        ws = wb.active

        # Get countries from row 1 (starting col 2)
        countries = {}
        for col in range(2, ws.max_column + 1):
            val = ws.cell(1, col).value
            if val:
                countries[col] = str(val).strip()

        # Get tech availability from rows 2+
        for row in range(2, ws.max_row + 1):
            tech = ws.cell(row, 1).value
            if not tech:
                continue
            tech_code = str(tech).strip().upper()
            self.availability[tech_code] = {}

            for col, country in countries.items():
                val = ws.cell(row, col).value
                is_available = val and str(val).upper() == 'YES'
                self.availability[tech_code][country] = is_available

        wb.close()

    def is_available(self, tech_code, country_code):
        """Check if a technology is available for a country"""
        tech_upper = tech_code.upper()
        country_upper = country_code.upper()

        if tech_upper not in self.availability:
            # If tech not in matrix, assume available (for techs like BCK, ELC, TRN)
            return True

        if country_upper not in self.availability[tech_upper]:
            # If country not in matrix, assume not available
            return False

        return self.availability[tech_upper][country_upper]

    def get_available_countries(self, tech_code):
        """Get list of countries where a technology is available"""
        tech_upper = tech_code.upper()
        if tech_upper not in self.availability:
            return []
        return [c for c, avail in self.availability[tech_upper].items() if avail]


class TechEquivalences:
    """Loads and applies technology equivalence rules from YAML"""

    # Country code transformations: old_code -> new_code
    # JAM was incorrectly used for Barbados, BRB is the correct ISO-3166 code
    COUNTRY_CODE_TRANSFORMS = {
        'JAM': 'BRB',
    }

    def __init__(self, yaml_path):
        self.yaml_path = yaml_path
        self.config = {}
        self.direct_mappings = {}
        self.gas_unification = {}
        self.aggregation_rules = {}
        self._load_yaml()

    def _load_yaml(self):
        """Load the YAML configuration"""
        if not self.yaml_path.exists():
            raise FileNotFoundError(f"Config_tech_equivalences.yaml not found: {self.yaml_path}")

        with open(self.yaml_path, 'r', encoding='utf-8') as f:
            self.config = yaml.safe_load(f)

        self.direct_mappings = self.config.get('direct_mappings', {})
        self.gas_unification = self.config.get('gas_unification', {})
        self.aggregation_rules = self.config.get('aggregation_rules', {})

    def transform_country_code(self, tech_code):
        """
        Transform old country codes to new ones in a technology code.

        For example: PWRHYDJAMXX01 -> PWRHYDBRBXX01

        Args:
            tech_code: The technology code string

        Returns:
            str: The technology code with transformed country codes
        """
        if not tech_code:
            return tech_code

        result = tech_code
        for old_code, new_code in self.COUNTRY_CODE_TRANSFORMS.items():
            # Replace the country code wherever it appears in the tech code
            result = result.replace(old_code, new_code)
        return result

    def get_new_tech_code(self, old_tech):
        """
        Get the new technology code for an old one.

        This method also applies country code transformations (e.g., JAM -> BRB).

        Returns:
            tuple: (new_tech_code, mapping_type) where mapping_type is:
                   'direct' - simple suffix removal
                   'gas_aggregation' - CCG/OCG to NGS
                   'unchanged' - no change needed
                   'removed' - technology was removed from model
                   None if no mapping found
        """
        # First, apply country code transformation to the input
        transformed_tech = self.transform_country_code(old_tech)
        old_upper = transformed_tech.upper()

        # Check direct mappings first (using transformed code)
        if transformed_tech in self.direct_mappings:
            new_code = self.direct_mappings[transformed_tech]
            if new_code is None:
                return None, 'removed'
            return self.transform_country_code(new_code), 'direct'

        # Check if it's a CCG or OCG technology (gas unification)
        if 'CCG' in old_upper or 'OCG' in old_upper:
            # Extract country code from tech (positions 6-8 for PWR technologies)
            match = re.match(r'PWR(CCG|OCG)([A-Z]{3})XX(\d{2})?', old_upper)
            if match:
                country = match.group(2)
                # Country is already transformed since we used old_upper from transformed_tech
                new_tech = f'PWRNGS{country}XX'
                return new_tech, 'gas_aggregation'

        # Check if it's a simple suffix removal (XX01 -> XX)
        match = re.match(r'^(PWR[A-Z]{3}[A-Z]{3}XX)(0[01])$', old_upper)
        if match:
            new_tech = match.group(1)
            return new_tech, 'direct'

        # Check unchanged techs (BCK, ELC, TRN)
        if old_upper.startswith(('PWRBCK', 'ELC', 'TRN')):
            if not old_upper[-2:].isdigit():  # No suffix
                return transformed_tech, 'unchanged'

        return None, None

    def get_aggregation_method(self, parameter):
        """Get the aggregation method for a parameter when merging CCG+OCG"""
        if parameter in self.aggregation_rules:
            return self.aggregation_rules[parameter].get('method', 'sum')
        return 'sum'  # Default to sum


class OldInputsMigrator:
    """Main class for migrating old inputs to new model structure"""

    # Technologies containing these strings will be excluded from migration
    EXCLUDED_TECH_STRINGS = ['BCK', 'CCS', 'COG', 'OTH', 'WAV']

    def __init__(self, base_path, dry_run=False):
        self.base_path = Path(base_path)
        self.old_inputs_path = self.base_path / "Old_Inputs"
        self.dry_run = dry_run
        self.log_lines = []
        self.stats = {
            'files_processed': 0,
            'values_migrated': 0,
            'values_skipped_no_mapping': 0,
            'values_skipped_not_available': 0,
            'values_skipped_excluded': 0,
            'values_aggregated': 0,
            'errors': 0
        }

        # Load configuration
        self.matrix = TechCountryMatrix(self.base_path / "Tech_Country_Matrix.xlsx")
        self.equivalences = TechEquivalences(self.base_path / "Config_tech_equivalences.yaml")

        # Load country codes config for DSPTRN support
        yaml_path = self.base_path / "Config_country_codes.yaml"
        if yaml_path.exists():
            with open(yaml_path, 'r', encoding='utf-8') as fh:
                self.country_codes_config = yaml.safe_load(fh) or {}
        else:
            self.country_codes_config = {}
        self.enable_dsptrn = self.country_codes_config.get('enable_dsptrn', False)
        
        # Scenarios
        self.scenarios = ["BAU", "NDC", "NDC+ELC", "NDC_NoRPO"]

    def log(self, message, level="INFO"):
        """Add message to log"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_line = f"[{timestamp}] {level}: {message}"
        self.log_lines.append(log_line)
        print(log_line)

    def should_exclude_tech(self, tech_code):
        """
        Check if a technology should be excluded from migration.

        Args:
            tech_code: The technology code string

        Returns:
            bool: True if the technology should be excluded
        """
        if not tech_code:
            return False
        tech_upper = tech_code.upper()
        for excluded in self.EXCLUDED_TECH_STRINGS:
            if excluded in tech_upper:
                return True
        return False

    def create_backup(self, file_path):
        """Create backup of file before modifying"""
        if self.dry_run:
            return None
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = file_path.with_name(f"{file_path.stem}_backup_{timestamp}{file_path.suffix}")
        shutil.copy2(file_path, backup_path)
        return backup_path

    def extract_tech_info(self, tech_code):
        """
        Extract technology type and country from a tech code.

        Returns:
            tuple: (tech_type, country_code) or (None, None) if not parseable
        """
        tech_upper = tech_code.upper()

        # PWR technologies: PWR{TYPE}{COUNTRY}XX{SUFFIX}
        match = re.match(r'PWR([A-Z]{3})([A-Z]{3})XX(\d{2})?', tech_upper)
        if match:
            return match.group(1), match.group(2)

        # ELC commodities: ELC{COUNTRY}XX{SUFFIX}
        match = re.match(r'ELC([A-Z]{3})XX(\d{2})?', tech_upper)
        if match:
            return 'ELC', match.group(1)

        # TRN technologies: TRN{COUNTRY1}XX{COUNTRY2}XX
        match = re.match(r'TRN([A-Z]{3})XX([A-Z]{3})XX', tech_upper)
        if match:
            return 'TRN', match.group(1)  # Return first country

        return None, None

    def read_sheet_data(self, ws, tech_col=2, param_col=5, has_param=True, timeslice_col=None, direction_col=None, extra_cols=None):
        """
        Read data from a worksheet into a structured format.

        Args:
            ws: Worksheet object
            tech_col: Column index for technology (1-based)
            param_col: Column index for parameter (1-based), ignored if has_param=False
            has_param: Whether the sheet has a parameter column
            timeslice_col: Column index for timeslice (1-based), if applicable
            direction_col: Column index for direction (1-based), for AR_Projections sheets
            extra_cols: List of additional column indices to read (1-based), e.g., [7] for Projection.Mode

        Returns:
            dict: {key: {year: value, 'extra': {col: value}}} where key includes tech, param, timeslice, and/or direction
            dict: {year: col_index}
        """
        data = {}

        # Get year columns from header - handle both int and string years
        year_cols = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(1, col).value
            if header is not None:
                header_str = str(header).strip()
                # Check if it's a year (could be int like 2021 or string like '2021')
                if header_str.isdigit():
                    year = int(header_str)
                    if 2000 <= year <= 2100:
                        year_cols[year] = col

        # Read data rows
        for row in range(2, ws.max_row + 1):
            tech = ws.cell(row, tech_col).value

            if not tech:
                continue

            tech_str = str(tech).strip()

            # Get direction if applicable
            direction_str = None
            if direction_col:
                direction = ws.cell(row, direction_col).value
                if direction:
                    direction_str = str(direction).strip()

            # Build the key based on available columns
            if timeslice_col:
                timeslice = ws.cell(row, timeslice_col).value
                if not timeslice:
                    continue
                timeslice_str = str(timeslice).strip()

                if has_param:
                    param = ws.cell(row, param_col).value
                    if not param:
                        continue
                    param_str = str(param).strip()
                    if direction_str:
                        key = (timeslice_str, tech_str, param_str, direction_str)
                    else:
                        key = (timeslice_str, tech_str, param_str)
                else:
                    if direction_str:
                        key = (timeslice_str, tech_str, direction_str)
                    else:
                        key = (timeslice_str, tech_str)
            elif has_param:
                param = ws.cell(row, param_col).value
                if not param:
                    continue
                param_str = str(param).strip()
                if direction_str:
                    key = (tech_str, param_str, direction_str)
                else:
                    key = (tech_str, param_str)
            else:
                if direction_str:
                    key = (tech_str, direction_str)
                else:
                    key = tech_str

            if key not in data:
                data[key] = {'row': row, 'years': {}, 'extra': {}}

            for year, col in year_cols.items():
                val = ws.cell(row, col).value
                if val is not None and val != '':
                    try:
                        data[key]['years'][year] = float(val)
                    except (ValueError, TypeError):
                        pass

            # Read extra columns if specified
            if extra_cols:
                for extra_col in extra_cols:
                    extra_val = ws.cell(row, extra_col).value
                    if extra_val is not None:
                        data[key]['extra'][extra_col] = extra_val

        return data, year_cols

    def find_target_row(self, ws, tech, param=None, tech_col=2, param_col=5, has_param=True,
                        timeslice=None, timeslice_col=None, direction=None, direction_col=None):
        """Find the row in target worksheet matching tech, param, timeslice, and/or direction"""
        for row in range(2, ws.max_row + 1):
            row_tech = ws.cell(row, tech_col).value

            if not row_tech:
                continue

            if str(row_tech).strip() != tech:
                continue

            # Check timeslice if required
            if timeslice_col and timeslice:
                row_ts = ws.cell(row, timeslice_col).value
                if not row_ts or str(row_ts).strip() != timeslice:
                    continue

            # Check direction if required
            if direction_col and direction:
                row_dir = ws.cell(row, direction_col).value
                if not row_dir or str(row_dir).strip() != direction:
                    continue

            if has_param:
                row_param = ws.cell(row, param_col).value
                if not row_param:
                    continue
                if str(row_param).strip() == param:
                    return row
            else:
                return row

        return None

    def migrate_parametrization_sheet(self, old_ws, new_ws, sheet_name, scenario,
                                        tech_col=2, param_col=5, has_param=True,
                                        timeslice_col=None, direction_col=None, extra_cols=None):
        """
        Migrate data from old to new parametrization sheet.

        Handles:
        - Direct mappings (suffix removal)
        - Gas aggregation (CCG+OCG -> NGS)
        - Tech availability check
        - Timeslice-aware matching for sheets like Capacities
        - Direction-aware matching for AR_Projections sheets

        Args:
            old_ws: Old worksheet
            new_ws: New worksheet
            sheet_name: Name of the sheet
            scenario: Scenario name
            tech_col: Column index for technology (1-based)
            param_col: Column index for parameter (1-based)
            has_param: Whether this sheet has a parameter column
            timeslice_col: Column index for timeslice (1-based), if applicable
            direction_col: Column index for direction (1-based), for AR_Projections sheets
            extra_cols: List of additional column indices to migrate (1-based), e.g., [7] for Projection.Mode
        """
        self.log(f"    Migrating {sheet_name}...")

        # Read old data with correct column configuration
        old_data, old_year_cols = self.read_sheet_data(old_ws, tech_col=tech_col,
                                                        param_col=param_col, has_param=has_param,
                                                        timeslice_col=timeslice_col,
                                                        direction_col=direction_col,
                                                        extra_cols=extra_cols)

        if not old_year_cols:
            self.log(f"      WARNING: No year columns found in {sheet_name}")
            return

        # Build year_cols dict from NEW worksheet to map years correctly
        new_year_cols = {}
        for col in range(1, new_ws.max_column + 1):
            header = new_ws.cell(1, col).value
            if header is not None:
                header_str = str(header).strip()
                if header_str.isdigit():
                    year = int(header_str)
                    if 2000 <= year <= 2100:
                        new_year_cols[year] = col

        # Group gas technologies for aggregation
        # Key format: (timeslice, country, param) if timeslice_col else (country, param)
        gas_techs = defaultdict(list)

        # Process each old tech/param combination
        migrated = 0
        skipped_no_mapping = 0
        skipped_not_available = 0
        skipped_no_target = 0
        skipped_excluded = 0
        aggregated = 0

        for key, info in old_data.items():
            # Parse the key based on structure (may include direction)
            direction = None
            if timeslice_col:
                if has_param:
                    if direction_col:
                        timeslice, old_tech, param, direction = key
                    else:
                        timeslice, old_tech, param = key
                else:
                    if direction_col:
                        timeslice, old_tech, direction = key
                    else:
                        timeslice, old_tech = key
                    param = None
            else:
                timeslice = None
                if has_param:
                    if direction_col:
                        old_tech, param, direction = key
                    else:
                        old_tech, param = key
                else:
                    if direction_col:
                        old_tech, direction = key
                    else:
                        old_tech = key
                    param = None

            # Check if technology should be excluded from migration
            if self.should_exclude_tech(old_tech):
                skipped_excluded += 1
                continue

            new_tech, mapping_type = self.equivalences.get_new_tech_code(old_tech)

            if mapping_type == 'removed':
                skipped_no_mapping += 1
                continue

            if mapping_type is None:
                # Try to determine if it's unchanged
                # ELC codes like ELCARGXX02 (demand commodities) stay the same
                # MIN codes (mining commodities) stay the same
                # TRN codes (transport interconnections) stay the same
                # PWRBCK codes (backup) stay the same
                # Storage codes stay the same:
                #   - LDS (Long Duration Storage) e.g., LDSARGXX01
                #   - SDS (Short Duration Storage) e.g., SDSARGXX01
                # Transmission technologies stay the same:
                #   - RNWTRN, RNWRPO, RNWNLI (renewable transmission)
                #   - PWRTRN, TRNRPO, TRNNLI (non-renewable transmission)
                transmission_prefixes = ('PWRBCK', 'TRN', 'MIN',
                                        'RNWTRN', 'RNWRPO', 'RNWNLI',
                                        'PWRTRN', 'TRNRPO', 'TRNNLI',
                                        'DSPTRN')
                storage_prefixes = ('LDS', 'SDS')
                if old_tech.startswith(transmission_prefixes):
                    if not old_tech[-2:].isdigit():
                        # Apply country code transformation (e.g., JAM -> BRB)
                        new_tech = self.equivalences.transform_country_code(old_tech)
                        mapping_type = 'unchanged'
                    else:
                        skipped_no_mapping += 1
                        continue
                elif old_tech.startswith('ELC'):
                    # ELC commodities keep the same code (including XX02 suffix)
                    # Apply country code transformation (e.g., JAM -> BRB)
                    new_tech = self.equivalences.transform_country_code(old_tech)
                    mapping_type = 'unchanged'
                elif old_tech.startswith(storage_prefixes):
                    # Storage codes keep the same (LDS/SDS with XX01 suffix)
                    # Apply country code transformation (e.g., JAM -> BRB)
                    new_tech = self.equivalences.transform_country_code(old_tech)
                    mapping_type = 'unchanged'
                else:
                    skipped_no_mapping += 1
                    continue

            # Extract tech type and country for availability check
            tech_type, country = self.extract_tech_info(new_tech)

            if tech_type and country:
                # Check if available in new model
                if not self.matrix.is_available(tech_type, country):
                    skipped_not_available += 1
                    continue

            if mapping_type == 'gas_aggregation':
                # Collect for later aggregation with timeslice info
                if timeslice_col:
                    gas_techs[(timeslice, country, param)].append((old_tech, info))
                else:
                    gas_techs[(country, param)].append((old_tech, info))
            else:
                # Direct migration
                target_row = self.find_target_row(new_ws, new_tech, param,
                                                   tech_col=tech_col, param_col=param_col,
                                                   has_param=has_param,
                                                   timeslice=timeslice, timeslice_col=timeslice_col,
                                                   direction=direction, direction_col=direction_col)
                if target_row:
                    if not self.dry_run:
                        # First, clear all year cells in NEW worksheet
                        for year, col in new_year_cols.items():
                            new_ws.cell(target_row, col).value = None
                        # Then write only the values that exist in source AND in target
                        for year, val in info['years'].items():
                            if year in new_year_cols:  # Only write if year exists in NEW model
                                new_ws.cell(target_row, new_year_cols[year]).value = val
                        # Write extra columns (e.g., Projection.Mode) - always migrate these
                        if extra_cols and 'extra' in info:
                            for extra_col, extra_val in info['extra'].items():
                                new_ws.cell(target_row, extra_col, extra_val)
                    # Count year values migrated, plus 1 for extra cols if any were migrated
                    migrated += len(info['years'])
                    if extra_cols and info.get('extra'):
                        migrated += len(info['extra'])
                else:
                    skipped_no_target += 1

        # Process gas aggregation
        for gas_key, tech_list in gas_techs.items():
            # Parse key based on whether we have timeslices
            if timeslice_col:
                timeslice, country, param = gas_key
            else:
                timeslice = None
                country, param = gas_key

            new_tech = f'PWRNGS{country}XX'

            # Check availability
            if not self.matrix.is_available('NGS', country):
                skipped_not_available += len(tech_list)
                continue

            # Aggregate values by year
            # Special case: For AR_Projections sheets (direction_col is set),
            # when Fuel (param) is GASINT, use average instead of sum
            # because InputActivityRatio represents efficiency, not capacity
            if direction_col and param == 'GASINT':
                method = 'average'
            else:
                method = self.equivalences.get_aggregation_method(param) if param else 'sum'
            aggregated_years = defaultdict(list)

            for old_tech, info in tech_list:
                for year, val in info['years'].items():
                    aggregated_years[year].append(val)

            # Apply aggregation method
            final_values = {}
            for year, values in aggregated_years.items():
                if method == 'sum':
                    final_values[year] = sum(values)
                elif method == 'max':
                    final_values[year] = max(values)
                elif method == 'average':
                    # Simple average for efficiency parameters (e.g., InputActivityRatio with GASINT)
                    final_values[year] = sum(values) / len(values)
                elif method == 'weighted_average_by_capacity':
                    # For now, use simple average (would need capacity data for weighted)
                    final_values[year] = sum(values) / len(values)
                else:
                    final_values[year] = sum(values)

            # Write to target
            target_row = self.find_target_row(new_ws, new_tech, param,
                                               tech_col=tech_col, param_col=param_col,
                                               has_param=has_param,
                                               timeslice=timeslice, timeslice_col=timeslice_col)
            if target_row:
                if not self.dry_run:
                    # First, clear all year cells in NEW worksheet
                    for year, col in new_year_cols.items():
                        new_ws.cell(target_row, col).value = None
                    # Then write only the aggregated values that exist in target
                    for year, val in final_values.items():
                        if year in new_year_cols:  # Only write if year exists in NEW model
                            new_ws.cell(target_row, new_year_cols[year]).value = val
                aggregated += len(final_values)

        self.log(f"      Migrated: {migrated}, Aggregated: {aggregated}, "
                 f"Skipped (no mapping): {skipped_no_mapping}, "
                 f"Skipped (not available): {skipped_not_available}, "
                 f"Skipped (excluded): {skipped_excluded}, "
                 f"Skipped (no target row): {skipped_no_target}")

        self.stats['values_migrated'] += migrated
        self.stats['values_aggregated'] += aggregated
        self.stats['values_skipped_no_mapping'] += skipped_no_mapping
        self.stats['values_skipped_not_available'] += skipped_not_available
        self.stats['values_skipped_excluded'] += skipped_excluded

    def migrate_fixed_horizon_parameters(self, old_ws, new_ws, sheet_name, scenario):
        """
        Migrate Fixed Horizon Parameters sheet which has a single Value column
        instead of year columns.

        Structure:
        - Tech.Type (col 1), Tech.ID (col 2), Tech (col 3), Tech.Name (col 4)
        - Parameter.ID (col 5), Parameter (col 6), Unit (col 7), Value (col 8)
        """
        self.log(f"    Migrating {sheet_name}...")

        # Build lookup from old data: key = (tech_code, parameter) -> value
        old_values = {}
        for row in range(2, old_ws.max_row + 1):
            tech = old_ws.cell(row, 3).value  # Tech column (3)
            param = old_ws.cell(row, 6).value  # Parameter column (6)
            value = old_ws.cell(row, 8).value  # Value column (8)

            if not tech or not param:
                continue

            tech_str = str(tech).strip()
            param_str = str(param).strip()

            # Check if technology should be excluded from migration
            if self.should_exclude_tech(tech_str):
                continue

            # Get new tech code
            new_tech, mapping_type = self.equivalences.get_new_tech_code(tech_str)

            if mapping_type == 'removed' or mapping_type is None:
                # Check if it's an unchanged tech
                transmission_prefixes = ('PWRBCK', 'TRN', 'MIN',
                                        'RNWTRN', 'RNWRPO', 'RNWNLI',
                                        'PWRTRN', 'TRNRPO', 'TRNNLI',
                                        'DSPTRN')
                storage_prefixes = ('LDS', 'SDS')
                if tech_str.startswith(transmission_prefixes):
                    if not tech_str[-2:].isdigit():
                        # Apply country code transformation (e.g., JAM -> BRB)
                        new_tech = self.equivalences.transform_country_code(tech_str)
                    else:
                        continue
                elif tech_str.startswith('ELC'):
                    # Apply country code transformation (e.g., JAM -> BRB)
                    new_tech = self.equivalences.transform_country_code(tech_str)
                elif tech_str.startswith(storage_prefixes):
                    # Apply country code transformation (e.g., JAM -> BRB)
                    new_tech = self.equivalences.transform_country_code(tech_str)
                else:
                    continue

            key = (new_tech, param_str)
            if key not in old_values:
                old_values[key] = value
            elif mapping_type == 'gas_aggregation' and param_str in ['ResidualCapacity', 'TotalTechnologyAnnualActivityLowerLimit']:
                # Sum values for gas aggregation
                if old_values[key] is not None and value is not None:
                    old_values[key] = float(old_values[key]) + float(value)

        # Update new worksheet
        migrated = 0
        for row in range(2, new_ws.max_row + 1):
            tech = new_ws.cell(row, 3).value  # Tech column (3)
            param = new_ws.cell(row, 6).value  # Parameter column (6)

            if not tech or not param:
                continue

            tech_str = str(tech).strip()
            param_str = str(param).strip()

            key = (tech_str, param_str)
            if key in old_values:
                old_val = old_values[key]
                current_val = new_ws.cell(row, 8).value

                if old_val != current_val:
                    if not self.dry_run:
                        new_ws.cell(row, 8, old_val)
                    migrated += 1

        self.log(f"      Values migrated: {migrated}")
        self.stats['values_migrated'] += migrated

    def update_pwrngs_projection_mode(self, ws, sheet_name):
        """
        Update Projection.Mode for PWRNGS technologies in Secondary Techs sheet.

        For parameters TotalTechnologyAnnualActivityLowerLimit,
        TotalAnnualMinCapacityInvestment, and ResidualCapacity:
        - Set Projection.Mode to "User defined" if any year column has a value
        - Set Projection.Mode to "EMPTY" if all year columns are empty
        """
        self.log(f"    Updating Projection.Mode for PWRNGS in {sheet_name}...")

        # Get year columns
        year_cols = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(1, col).value
            if header is not None:
                header_str = str(header).strip()
                if header_str.isdigit():
                    year = int(header_str)
                    if 2000 <= year <= 2100:
                        year_cols[year] = col

        if not year_cols:
            self.log(f"      WARNING: No year columns found")
            return

        # Parameters to check
        target_params = [
            'TotalTechnologyAnnualActivityLowerLimit',
            'TotalAnnualMinCapacityInvestment',
            'ResidualCapacity'
        ]

        # Tech col 2, Parameter col 5, Projection.Mode col 7
        tech_col = 2
        param_col = 5
        proj_mode_col = 7

        updated = 0
        for row in range(2, ws.max_row + 1):
            tech = ws.cell(row, tech_col).value
            param = ws.cell(row, param_col).value

            if not tech or not param:
                continue

            tech_str = str(tech).strip()
            param_str = str(param).strip()

            # Check if it's a PWRNGS technology and target parameter
            if not tech_str.startswith('PWRNGS'):
                continue
            if param_str not in target_params:
                continue

            # Check if any year column has a value
            has_value = False
            for year, col in year_cols.items():
                val = ws.cell(row, col).value
                if val is not None and val != '':
                    has_value = True
                    break

            # Set Projection.Mode
            new_mode = "User defined" if has_value else "EMPTY"
            current_mode = ws.cell(row, proj_mode_col).value

            if current_mode != new_mode:
                if not self.dry_run:
                    ws.cell(row, proj_mode_col, new_mode)
                updated += 1

        self.log(f"      Projection.Mode updated: {updated}")

    def update_pwr_max_capacity_investment_mode(self, ws, sheet_name):
        """
        For all PWR technologies, set Projection.Mode to "EMPTY" for
        TotalAnnualMaxCapacityInvestment.
        Controlled by force_empty_max_capacity_investment_pwr flag in Config_country_codes.yaml.
        """
        if not get_force_empty_max_capacity_investment_pwr():
            self.log(f"    Skipping PWR TotalAnnualMaxCapacityInvestment override (flag disabled)")
            return

        self.log(f"    Updating Projection.Mode for PWR TotalAnnualMaxCapacityInvestment in {sheet_name}...")

        tech_col = 2
        param_col = 5
        proj_mode_col = 7

        updated = 0
        for row in range(2, ws.max_row + 1):
            tech = ws.cell(row, tech_col).value
            param = ws.cell(row, param_col).value
            proj_mode = ws.cell(row, proj_mode_col).value

            if not tech or not param or not proj_mode:
                continue

            tech_str = str(tech).strip()
            param_str = str(param).strip()
            proj_mode_str = str(proj_mode).strip()

            if param_str != 'TotalAnnualMaxCapacityInvestment':
                continue

            if proj_mode_str == 'EMPTY':
                continue

            if tech_str.startswith('PWR'):
                if not self.dry_run:
                    ws.cell(row, proj_mode_col, "EMPTY")
                updated += 1

        self.log(f"      Projection.Mode updated: {updated}")
    def update_capacity_investment_projection_mode(self, ws, sheet_name):
        """
        Update Projection.Mode for specific parameters in Secondary Techs sheet.

        For TotalAnnualMaxCapacityInvestment and TotalAnnualMaxCapacity:
        - Change "User defined" to "EMPTY" in Projection.Mode column
          (includes interconnection technologies TRN{COUNTRY}XX{COUNTRY}XX)
        """
        self.log(f"    Updating Projection.Mode for capacity parameters in {sheet_name}...")

        # Tech col 2, Parameter col 5, Projection.Mode col 7
        tech_col = 2
        param_col = 5
        proj_mode_col = 7

        # Parameters to update
        target_params = {
            'TotalAnnualMaxCapacityInvestment': False,  # No exception - includes interconnections
            'TotalAnnualMaxCapacity': False             # False = no exception
        }

        # Regex pattern for interconnection technologies: TRN{3letters}XX{3letters}XX
        interconnection_pattern = re.compile(r'^TRN[A-Z]{3}XX[A-Z]{3}XX$')

        updated = 0
        for row in range(2, ws.max_row + 1):
            tech = ws.cell(row, tech_col).value
            param = ws.cell(row, param_col).value
            proj_mode = ws.cell(row, proj_mode_col).value

            if not tech or not param or not proj_mode:
                continue

            tech_str = str(tech).strip()
            param_str = str(param).strip()
            proj_mode_str = str(proj_mode).strip()

            # Check if this parameter needs updating
            if param_str not in target_params:
                continue

            # Only update if current value is "User defined"
            if proj_mode_str != "User defined":
                continue

            # Check interconnection exception for TotalAnnualMaxCapacityInvestment
            has_exception = target_params[param_str]
            if has_exception and interconnection_pattern.match(tech_str):
                # Skip this row - it's an interconnection and has exception
                continue

            # Update Projection.Mode to EMPTY
            if not self.dry_run:
                ws.cell(row, proj_mode_col, "EMPTY")
            updated += 1

        self.log(f"      Projection.Mode updated: {updated}")

    def migrate_parametrization(self, scenario):
        """Migrate A-O_Parametrization.xlsx for a scenario"""
        old_path = self.old_inputs_path / "A1_Outputs" / f"A1_Outputs_{scenario}" / "A-O_Parametrization.xlsx"
        new_path = self.base_path / "A1_Outputs" / f"A1_Outputs_{scenario}" / "A-O_Parametrization.xlsx"

        if not old_path.exists():
            self.log(f"  Old file not found: {old_path}", "WARNING")
            return

        if not new_path.exists():
            self.log(f"  New file not found: {new_path}", "WARNING")
            return

        self.log(f"  Processing A-O_Parametrization.xlsx...")

        # Create backup
        if not self.dry_run:
            backup = self.create_backup(new_path)
            self.log(f"    Backup: {backup.name if backup else 'N/A'}")

        # Sheets to migrate with their column configurations
        # (sheet_name, tech_col, param_col, has_param, timeslice_col, extra_cols)
        # timeslice_col is None for sheets without timeslices
        # extra_cols includes Projection.Mode column for each sheet
        sheets_config = [
            # Fixed Horizon Parameters is handled separately (no year columns)
            ('Primary Techs', 2, 5, True, None, [7]),              # Tech col 2, Parameter col 5, Projection.Mode col 7
            ('Secondary Techs', 2, 5, True, None, [7]),            # Tech col 2, Parameter col 5, Projection.Mode col 7
            ('Capacities', 3, 6, True, 1, [8]),                    # Tech col 3, Parameter col 6, Timeslice col 1, Projection.Mode col 8
            ('VariableCost', 3, 6, True, None, [9]),               # Tech col 3, Parameter col 6, Projection.Mode col 9
            ('Demand Techs', 2, 5, True, None, [7]),               # Tech col 2, Parameter col 5, Projection.Mode col 7
        ]

        try:
            old_wb = openpyxl.load_workbook(old_path, data_only=True)
            new_wb = openpyxl.load_workbook(new_path)

            # Migrate Fixed Horizon Parameters (special handling - no year columns)
            if 'Fixed Horizon Parameters' in old_wb.sheetnames and 'Fixed Horizon Parameters' in new_wb.sheetnames:
                self.migrate_fixed_horizon_parameters(
                    old_wb['Fixed Horizon Parameters'],
                    new_wb['Fixed Horizon Parameters'],
                    'Fixed Horizon Parameters',
                    scenario
                )

            # Migrate other sheets with year columns
            for sheet_name, tech_col, param_col, has_param, timeslice_col, extra_cols in sheets_config:
                if sheet_name not in old_wb.sheetnames:
                    self.log(f"    Sheet '{sheet_name}' not in old file", "WARNING")
                    continue
                if sheet_name not in new_wb.sheetnames:
                    self.log(f"    Sheet '{sheet_name}' not in new file", "WARNING")
                    continue

                old_ws = old_wb[sheet_name]
                new_ws = new_wb[sheet_name]
                self.migrate_parametrization_sheet(
                    old_ws, new_ws, sheet_name, scenario,
                    tech_col=tech_col, param_col=param_col, has_param=has_param,
                    timeslice_col=timeslice_col, extra_cols=extra_cols
                )

            # Update Projection.Mode for PWRNGS technologies in Secondary Techs
            if 'Secondary Techs' in new_wb.sheetnames:
                self.update_pwrngs_projection_mode(new_wb['Secondary Techs'], 'Secondary Techs')
                self.update_capacity_investment_projection_mode(new_wb['Secondary Techs'], 'Secondary Techs')

            if not self.dry_run:
                new_wb.save(new_path)
                self.log(f"    Saved: {new_path.name}")

            old_wb.close()
            new_wb.close()
            self.stats['files_processed'] += 1

        except Exception as e:
            self.log(f"  Error processing Parametrization: {e}", "ERROR")
            self.stats['errors'] += 1

    def migrate_demand(self, scenario):
        """Migrate A-O_Demand.xlsx for a scenario

        Structure of Demand files:
        - Demand_Projection: Fuel/Tech in col 2, no Parameter column, years start col 9
        - Profiles: Fuel/Tech in col 3, no Parameter column, years start col 10
        """
        old_path = self.old_inputs_path / "A1_Outputs" / f"A1_Outputs_{scenario}" / "A-O_Demand.xlsx"
        new_path = self.base_path / "A1_Outputs" / f"A1_Outputs_{scenario}" / "A-O_Demand.xlsx"

        if not old_path.exists():
            self.log(f"  Old Demand file not found: {old_path}", "WARNING")
            return

        if not new_path.exists():
            self.log(f"  New Demand file not found: {new_path}", "WARNING")
            return

        self.log(f"  Processing A-O_Demand.xlsx...")

        if not self.dry_run:
            backup = self.create_backup(new_path)
            self.log(f"    Backup: {backup.name if backup else 'N/A'}")

        try:
            old_wb = openpyxl.load_workbook(old_path, data_only=True)
            new_wb = openpyxl.load_workbook(new_path)

            # Sheet configurations: sheet_name -> (tech_col, has_param, timeslice_col, extra_cols)
            # extra_cols includes Projection.Mode column for each sheet
            sheet_configs = {
                'Demand_Projection': (2, False, None, [7]),  # Fuel/Tech in col 2, no param, no timeslice, Projection.Mode col 7
                'Profiles': (3, False, 1, [8]),              # Fuel/Tech in col 3, no param, Timeslice col 1, Projection.Mode col 8
            }

            for sheet_name, (tech_col, has_param, timeslice_col, extra_cols) in sheet_configs.items():
                if sheet_name not in old_wb.sheetnames:
                    continue
                if sheet_name not in new_wb.sheetnames:
                    continue

                old_ws = old_wb[sheet_name]
                new_ws = new_wb[sheet_name]

                self.migrate_parametrization_sheet(
                    old_ws, new_ws, sheet_name, scenario,
                    tech_col=tech_col, param_col=5, has_param=has_param,
                    timeslice_col=timeslice_col, extra_cols=extra_cols
                )

            if not self.dry_run:
                new_wb.save(new_path)
                self.log(f"    Saved: {new_path.name}")

            old_wb.close()
            new_wb.close()
            self.stats['files_processed'] += 1

        except Exception as e:
            self.log(f"  Error processing Demand: {e}", "ERROR")
            self.stats['errors'] += 1

    def migrate_ar_projections(self, scenario):
        """Migrate A-O_AR_Projections.xlsx for a scenario

        Structure:
        - Tech in col 2, NO Parameter column (has Direction in col 6)
        - Years start col 9
        - Use (Tech, Fuel, Direction) as unique key for matching
        """
        old_path = self.old_inputs_path / "A1_Outputs" / f"A1_Outputs_{scenario}" / "A-O_AR_Projections.xlsx"
        new_path = self.base_path / "A1_Outputs" / f"A1_Outputs_{scenario}" / "A-O_AR_Projections.xlsx"

        if not old_path.exists():
            self.log(f"  Old AR_Projections not found: {old_path}", "WARNING")
            return

        if not new_path.exists():
            self.log(f"  New AR_Projections not found: {new_path}", "WARNING")
            return

        self.log(f"  Processing A-O_AR_Projections.xlsx...")

        if not self.dry_run:
            backup = self.create_backup(new_path)
            self.log(f"    Backup: {backup.name if backup else 'N/A'}")

        # Sheets: Tech is in col 2, no traditional parameter column
        # All sheets have Projection.Mode in col 7
        sheets_to_migrate = ['Primary', 'Secondary', 'Demand Techs']

        try:
            old_wb = openpyxl.load_workbook(old_path, data_only=True)
            new_wb = openpyxl.load_workbook(new_path)

            for sheet_name in sheets_to_migrate:
                if sheet_name not in old_wb.sheetnames:
                    continue
                if sheet_name not in new_wb.sheetnames:
                    continue

                old_ws = old_wb[sheet_name]
                new_ws = new_wb[sheet_name]
                # Tech col 2, Fuel col 4, Direction col 6, Projection.Mode col 7
                # Use (Tech, Fuel, Direction) as unique key for matching
                self.migrate_parametrization_sheet(
                    old_ws, new_ws, sheet_name, scenario,
                    tech_col=2, param_col=4, has_param=True,
                    direction_col=6,  # Direction column for Input/Output distinction
                    extra_cols=[7]    # Projection.Mode column
                )

            if not self.dry_run:
                new_wb.save(new_path)
                self.log(f"    Saved: {new_path.name}")

            old_wb.close()
            new_wb.close()
            self.stats['files_processed'] += 1

        except Exception as e:
            self.log(f"  Error processing AR_Projections: {e}", "ERROR")
            self.stats['errors'] += 1

    def migrate_ar_model_base_year(self, scenario):
        """Migrate A-O_AR_Model_Base_Year.xlsx for a scenario

        This file contains model structure/topology (tech-fuel mappings) with
        InputActivityRatio and OutputActivityRatio values. Most values are 1,
        but some may differ (e.g., transmission losses with value 0).
        We need to migrate these values by matching Tech and Fuel columns.
        """
        old_path = self.old_inputs_path / "A1_Outputs" / f"A1_Outputs_{scenario}" / "A-O_AR_Model_Base_Year.xlsx"
        new_path = self.base_path / "A1_Outputs" / f"A1_Outputs_{scenario}" / "A-O_AR_Model_Base_Year.xlsx"

        if not old_path.exists():
            self.log(f"  Old AR_Model_Base_Year not found: {old_path}", "WARNING")
            return

        if not new_path.exists():
            self.log(f"  New AR_Model_Base_Year not found: {new_path}", "WARNING")
            return

        self.log(f"  Processing A-O_AR_Model_Base_Year.xlsx...")

        if not self.dry_run:
            backup = self.create_backup(new_path)
            self.log(f"    Backup: {backup.name if backup else 'N/A'}")

        try:
            old_wb = openpyxl.load_workbook(old_path, data_only=True)
            new_wb = openpyxl.load_workbook(new_path)

            # Sheet configurations: (sheet_name, tech_col, value_configs)
            # value_configs: list of (fuel_col, value_col, fuel_col_name)
            sheets_config = [
                ('Primary', 2, [(4, 6, 'Fuel.O')]),  # Tech col 2, Fuel.O col 4, Value.Fuel.O col 6
                ('Secondary', 6, [(2, 4, 'Fuel.I'), (8, 10, 'Fuel.O')]),  # Tech col 6
                ('Demand Techs', 6, [(2, 4, 'Fuel.I'), (8, 10, 'Fuel.O')]),  # Tech col 6
            ]

            for sheet_name, tech_col, value_configs in sheets_config:
                if sheet_name not in old_wb.sheetnames or sheet_name not in new_wb.sheetnames:
                    continue

                old_ws = old_wb[sheet_name]
                new_ws = new_wb[sheet_name]

                self.log(f"    Migrating {sheet_name}...")

                # Build lookup from old data: key = (tech, fuel) -> value
                old_values = {}
                for row in range(2, old_ws.max_row + 1):
                    tech = old_ws.cell(row, tech_col).value
                    if not tech:
                        continue
                    tech_str = str(tech).strip()

                    # Check if technology should be excluded from migration
                    if self.should_exclude_tech(tech_str):
                        continue

                    for fuel_col, value_col, fuel_name in value_configs:
                        fuel = old_ws.cell(row, fuel_col).value
                        value = old_ws.cell(row, value_col).value
                        if fuel and value is not None:
                            key = (tech_str, str(fuel).strip(), fuel_name)
                            old_values[key] = value

                # Update new sheet where tech/fuel matches
                migrated = 0
                for row in range(2, new_ws.max_row + 1):
                    tech = new_ws.cell(row, tech_col).value
                    if not tech:
                        continue
                    tech_str = str(tech).strip()

                    for fuel_col, value_col, fuel_name in value_configs:
                        fuel = new_ws.cell(row, fuel_col).value
                        if not fuel:
                            continue
                        fuel_str = str(fuel).strip()

                        key = (tech_str, fuel_str, fuel_name)
                        if key in old_values:
                            old_val = old_values[key]
                            current_val = new_ws.cell(row, value_col).value

                            if old_val != current_val:
                                if not self.dry_run:
                                    new_ws.cell(row, value_col, old_val)
                                migrated += 1

                self.log(f"      Values updated: {migrated}")

            if not self.dry_run:
                new_wb.save(new_path)
                self.log(f"    Saved: {new_path.name}")

            old_wb.close()
            new_wb.close()

        except Exception as e:
            self.log(f"    ERROR migrating Base Year: {e}", "ERROR")

    def migrate_storage(self):
        """Migrate A-Xtra_Storage.xlsx (shared across scenarios)"""
        old_path = self.old_inputs_path / "A2_Extra_Inputs" / "A-Xtra_Storage.xlsx"
        new_path = self.base_path / "A2_Extra_Inputs" / "A-Xtra_Storage.xlsx"

        if not old_path.exists():
            self.log(f"Old Storage file not found: {old_path}", "WARNING")
            return

        if not new_path.exists():
            self.log(f"New Storage file not found: {new_path}", "WARNING")
            return

        self.log(f"Processing A-Xtra_Storage.xlsx...")

        if not self.dry_run:
            backup = self.create_backup(new_path)
            self.log(f"  Backup: {backup.name if backup else 'N/A'}")

        try:
            old_wb = openpyxl.load_workbook(old_path, data_only=True)
            new_wb = openpyxl.load_workbook(new_path)

            for sheet_name in old_wb.sheetnames:
                if sheet_name not in new_wb.sheetnames:
                    continue

                self.log(f"  Migrating {sheet_name}...")
                old_ws = old_wb[sheet_name]
                new_ws = new_wb[sheet_name]
                # Storage sheets may have different structure - handle appropriately
                # For CapitalCostStorage, also migrate Projection.Mode (column 7)
                if sheet_name == 'CapitalCostStorage':
                    self.migrate_parametrization_sheet(old_ws, new_ws, sheet_name, "Storage",
                                                       extra_cols=[7])
                else:
                    self.migrate_parametrization_sheet(old_ws, new_ws, sheet_name, "Storage")

            if not self.dry_run:
                new_wb.save(new_path)

            old_wb.close()
            new_wb.close()
            self.stats['files_processed'] += 1

        except Exception as e:
            self.log(f"Error processing Storage: {e}", "ERROR")
            self.stats['errors'] += 1

    def normalize_profiles(self):
        """Normalize temporal profiles after demand migration"""
        # 🔔 ENTRY POINT - Always log that we reached this method
        self.log("")
        self.log("🔔" * 40)
        self.log("🔔 NORMALIZE_PROFILES() METHOD CALLED 🔔")
        self.log("🔔" * 40)

        # Check 1: Module availability
        self.log(f"[CHECK 1] NORMALIZATION_AVAILABLE = {NORMALIZATION_AVAILABLE}")
        if not NORMALIZATION_AVAILABLE:
            self.log("❌ WARNING: Profile normalization module not available", "WARNING")
            self.log("❌ NORMALIZATION SKIPPED - Module import failed")
            return

        # Check 2: Dry-run mode
        self.log(f"[CHECK 2] dry_run mode = {self.dry_run}")
        if self.dry_run:
            self.log("⚠️  Skipping profile normalization in dry-run mode")
            self.log("⚠️  NORMALIZATION SKIPPED - Dry-run mode active")
            return

        # All checks passed - proceeding with normalization
        self.log("")
        self.log("✅ All checks passed - proceeding with normalization")
        self.log("")
        self.log("=" * 70)
        self.log("⚡ NORMALIZING TEMPORAL PROFILES ⚡")
        self.log("=" * 70)

        inputs_dir = self.base_path / "OG_csvs_inputs"
        self.log(f"📁 Target directory: {inputs_dir}")
        self.log(f"📁 Directory exists: {inputs_dir.exists()}")

        try:
            self.log("")
            self.log("🔄 Starting normalization process...")
            self.log("")

            results = {
                "SpecifiedDemandProfile": normalize_specified_demand_profile(inputs_dir),
                "YearSplit": normalize_year_split(inputs_dir),
                "DaySplit": normalize_day_split(inputs_dir)
            }

            # Log results
            self.log("")
            self.log("=" * 70)
            self.log("📊 PROFILE NORMALIZATION RESULTS:")
            self.log("=" * 70)
            for profile_name, success in results.items():
                status = "✅ SUCCESS" if success else "❌ FAILED"
                self.log(f"  {status} - {profile_name}")

            if all(results.values()):
                self.log("")
                self.log("🎉 ALL PROFILES NORMALIZED SUCCESSFULLY! 🎉")
            else:
                self.log("")
                self.log("⚠️  SOME PROFILES FAILED NORMALIZATION", "WARNING")

        except Exception as e:
            self.log("")
            self.log("=" * 70)
            self.log(f"❌ ERROR DURING PROFILE NORMALIZATION: {e}", "ERROR")
            self.log("=" * 70)
            import traceback
            self.log(traceback.format_exc())
            self.stats['errors'] += 1

    def normalize_excel_profiles(self):
        """Normalize demand profiles in Excel files (A-O_Demand.xlsx)"""
        # 🔔 ENTRY POINT
        self.log("")
        self.log("📊" * 40)
        self.log("📊 NORMALIZE_EXCEL_PROFILES() METHOD CALLED 📊")
        self.log("📊" * 40)

        # Check 1: Module availability
        self.log(f"[CHECK 1] EXCEL_NORMALIZATION_AVAILABLE = {EXCEL_NORMALIZATION_AVAILABLE}")
        if not EXCEL_NORMALIZATION_AVAILABLE:
            self.log("❌ WARNING: Excel profile normalization module not available", "WARNING")
            self.log("❌ EXCEL NORMALIZATION SKIPPED - Module import failed")
            return

        # Check 2: Dry-run mode
        self.log(f"[CHECK 2] dry_run mode = {self.dry_run}")
        if self.dry_run:
            self.log("⚠️  Skipping Excel profile normalization in dry-run mode")
            self.log("⚠️  EXCEL NORMALIZATION SKIPPED - Dry-run mode active")
            return

        # All checks passed
        self.log("")
        self.log("✅ All checks passed - proceeding with Excel normalization")
        self.log("")
        self.log("=" * 70)
        self.log("📊 NORMALIZING EXCEL DEMAND PROFILES")
        self.log("=" * 70)

        results = {}

        for scenario in self.scenarios:
            scenario_dir = self.base_path / "A1_Outputs" / f"A1_Outputs_{scenario}"
            excel_file = scenario_dir / "A-O_Demand.xlsx"

            self.log("")
            self.log(f"Processing scenario: {scenario}")
            self.log(f"  File: {excel_file}")

            if excel_file.exists():
                try:
                    # Call the normalization function
                    success = normalize_excel_profiles(excel_file)
                    results[scenario] = success

                    if success:
                        self.log(f"  ✅ {scenario}: Normalized successfully")
                    else:
                        self.log(f"  ⚠️  {scenario}: Normalization completed with warnings", "WARNING")

                except Exception as e:
                    self.log(f"  ❌ {scenario}: Error during normalization - {e}", "ERROR")
                    results[scenario] = False
                    self.stats['errors'] += 1
            else:
                self.log(f"  ⚠️  {scenario}: File not found, skipping")
                results[scenario] = None

        # Summary
        self.log("")
        self.log("=" * 70)
        self.log("📋 EXCEL NORMALIZATION SUMMARY:")
        self.log("=" * 70)

        for scenario, result in results.items():
            if result is True:
                self.log(f"  ✅ {scenario}: Success")
            elif result is False:
                self.log(f"  ❌ {scenario}: Failed")
            else:
                self.log(f"  ⚠️  {scenario}: Skipped (file not found)")

        successful = sum(1 for r in results.values() if r is True)
        total_attempted = sum(1 for r in results.values() if r is not None)

        if total_attempted > 0 and successful == total_attempted:
            self.log("")
            self.log("🎉 ALL EXCEL FILES NORMALIZED SUCCESSFULLY! 🎉")
        elif successful > 0:
            self.log("")
            self.log(f"⚠️  {successful}/{total_attempted} files normalized successfully")
        else:
            self.log("")
            self.log("❌ No Excel files were normalized")

    def _has_dsptrn(self, scenario):
        """Check if the migrated data already contains DSPTRN/ELC03/ELC04 technologies."""
        base_year_path = (self.base_path / "A1_Outputs" / f"A1_Outputs_{scenario}"
                          / "A-O_AR_Model_Base_Year.xlsx")
        if not base_year_path.exists():
            return False
        try:
            import pandas as pd
            sec = pd.read_excel(base_year_path, sheet_name='Secondary', engine='openpyxl')
            # Check if any tech starts with DSPTRN or any fuel contains ELC...03/04
            has_dsptrn_tech = sec['Tech'].str.startswith('DSPTRN', na=False).any()
            has_elc03 = False
            has_elc04 = False
            for col in ['Fuel.I', 'Fuel.O']:
                if col in sec.columns:
                    has_elc03 = has_elc03 or sec[col].str.endswith('03', na=False).any()
                    has_elc04 = has_elc04 or sec[col].str.endswith('04', na=False).any()
            return has_dsptrn_tech or has_elc03 or has_elc04
        except Exception:
            return False

    def inject_dsptrn_if_missing(self, scenario):
        """If enable_dsptrn is True and data doesn't have DSPTRN/ELC03, inject them.

        This replicates the A2_AddTx logic: rewrites TRN interconnection fuel codes
        to ELC03 and adds DSPTRN dispatch technology rows.
        """
        if not self.enable_dsptrn:
            return
        if self._has_dsptrn(scenario):
            self.log(f"  DSPTRN/ELC03/ELC04 already present in {scenario}, skipping injection")
            return
        if self.dry_run:
            self.log(f"  [DRY RUN] Would inject DSPTRN into {scenario}")
            return

        self.log(f"  Injecting DSPTRN dispatch technology into {scenario}...")

        # Import A2_AddTx functions
        try:
            from A2_AddTx import (load_country_region_pairs, process_base_year,
                                  process_projections, process_parametrization)
        except ImportError:
            self.log("  WARNING: Could not import A2_AddTx, skipping DSPTRN injection", "WARNING")
            return

        yaml_path = self.base_path / "Config_country_codes.yaml"
        pairs = load_country_region_pairs(str(yaml_path))
        if not pairs:
            self.log("  WARNING: No country pairs found, skipping DSPTRN injection", "WARNING")
            return

        with open(yaml_path, 'r', encoding='utf-8') as fh:
            yaml_data = yaml.safe_load(fh) or {}

        scenario_dir = self.base_path / "A1_Outputs" / f"A1_Outputs_{scenario}"
        base_path = str(scenario_dir / "A-O_AR_Model_Base_Year.xlsx")
        proj_path = str(scenario_dir / "A-O_AR_Projections.xlsx")
        param_path = str(scenario_dir / "A-O_Parametrization.xlsx")

        try:
            if Path(base_path).exists():
                process_base_year(base_path, pairs, enable_dsptrn=True)
            if Path(proj_path).exists():
                process_projections(proj_path, pairs, enable_dsptrn=True)
            if Path(param_path).exists():
                process_parametrization(param_path, pairs, yaml_data, enable_dsptrn=True)
            self.log(f"  ✔ DSPTRN injected into {scenario}")
        except Exception as e:
            self.log(f"  ERROR injecting DSPTRN into {scenario}: {e}", "ERROR")
            self.stats['errors'] += 1
    def run(self):
        """Main execution"""
        self.log("=" * 80)
        self.log("OLD INPUTS MIGRATION")
        self.log("=" * 80)
        self.log(f"Start time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        self.log(f"Mode: {'DRY RUN' if self.dry_run else 'LIVE'}")
        self.log(f"Old Inputs: {self.old_inputs_path}")
        self.log(f"Target: {self.base_path}")
        self.log("")

        # Log country code transformations
        if self.equivalences.COUNTRY_CODE_TRANSFORMS:
            self.log("Country code transformations:")
            for old_code, new_code in self.equivalences.COUNTRY_CODE_TRANSFORMS.items():
                self.log(f"  {old_code} -> {new_code}")
            self.log("")

        # Log excluded technology strings
        if self.EXCLUDED_TECH_STRINGS:
            self.log("Technologies containing these strings will be EXCLUDED:")
            self.log(f"  {', '.join(self.EXCLUDED_TECH_STRINGS)}")
            self.log("")

        # Check Old_Inputs exists
        if not self.old_inputs_path.exists():
            self.log(f"ERROR: Old_Inputs folder not found: {self.old_inputs_path}", "ERROR")
            return 1

        # Migrate Storage (shared)
        self.log("")
        self.log("=" * 40)
        self.log("EXTRA INPUTS")
        self.log("=" * 40)
        self.migrate_storage()

        # Migrate per scenario
        for scenario in self.scenarios:
            self.log("")
            self.log("=" * 40)
            self.log(f"SCENARIO: {scenario}")
            self.log("=" * 40)

            old_scenario_path = self.old_inputs_path / "A1_Outputs" / f"A1_Outputs_{scenario}"
            if not old_scenario_path.exists():
                self.log(f"Old scenario folder not found: {old_scenario_path}", "WARNING")
                continue

            self.migrate_parametrization(scenario)
            self.migrate_demand(scenario)
            self.migrate_ar_projections(scenario)
            self.migrate_ar_model_base_year(scenario)

        # 🎯 NORMALIZE TEMPORAL PROFILES AFTER ALL DEMAND MIGRATIONS
        self.log("")
        self.log("=" * 80)
        self.log("🎯 ALL SCENARIO MIGRATIONS COMPLETE")
        self.log("=" * 80)

        # Step 1: Normalize CSV profiles (OG_csvs_inputs)
        self.log("")
        self.log("📄 Step 1: Normalizing CSV profiles (OG_csvs_inputs)")
        self.normalize_profiles()

        # Step 2: Normalize Excel profiles (A-O_Demand.xlsx files)
        self.log("")
        self.log("📊 Step 2: Normalizing Excel profiles (A-O_Demand.xlsx)")
        self.normalize_excel_profiles()

        # Summary
        self.log("")
        self.log("=" * 80)
        self.log("MIGRATION SUMMARY")
        self.log("=" * 80)
        self.log(f"Files processed: {self.stats['files_processed']}")
        self.log(f"Values migrated: {self.stats['values_migrated']}")
        self.log(f"Values aggregated (CCG+OCG->NGS): {self.stats['values_aggregated']}")
        self.log(f"Values skipped (no mapping): {self.stats['values_skipped_no_mapping']}")
        self.log(f"Values skipped (not available in matrix): {self.stats['values_skipped_not_available']}")
        self.log(f"Values skipped (excluded techs: {', '.join(self.EXCLUDED_TECH_STRINGS)}): {self.stats['values_skipped_excluded']}")
        self.log(f"Errors: {self.stats['errors']}")

        if self.dry_run:
            self.log("")
            self.log("DRY RUN - No changes were made")

        # Save log
        log_path = self.base_path / f"migration_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        with open(log_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(self.log_lines))
        self.log(f"\nLog saved: {log_path}")

        return 0 if self.stats['errors'] == 0 else 1


def main():
    import argparse

    parser = argparse.ArgumentParser(description='Migrate old inputs to new model structure')
    parser.add_argument('--dry-run', action='store_true', help='Show what would be migrated without making changes')
    args = parser.parse_args()

    try:
        script_dir = Path(__file__).parent
        migrator = OldInputsMigrator(script_dir, dry_run=args.dry_run)
        return migrator.run()

    except Exception as e:
        print(f"\nFATAL ERROR: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
