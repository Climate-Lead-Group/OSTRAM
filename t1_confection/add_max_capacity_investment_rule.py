"""
add_max_capacity_investment_rule.py
====================================

Apply the "MaxCapacityInvestment guard" rule to the A-O Parametrization workbook.

PROBLEM
-------
After running the AO workflow, awkward techs with no residual or planned capacity
were getting invested in starting in 2023, generating more than physical residual
capacity. This is a parameter-space artifact: with no lower limit, the optimizer
freely picks any technology unless an upper bound says otherwise.

RULE
----
For each technology in the target sheet(s), check if it has either of:
  - ResidualCapacity > 0 in any year, OR
  - TotalAnnualMinCapacityInvestment > 0 in any year

If yes ("ALLOWED"):
    For TotalAnnualMaxCapacityInvestment, fill EMPTY (None/blank) cells with
    9999. Existing explicit values are PRESERVED — including explicit zeros,
    which encode real modeling decisions (e.g. "no investment in 2023 base
    year because residual covers it", "transmission interconnect not online
    until 2030"). Only blank cells mean "no upper bound was set."

If no ("ZEROED"):
    For TotalAnnualMaxCapacity AND TotalAnnualMaxCapacityInvestment, set every
    year cell to 0 (overwriting whatever was there). The tech is locked out.

PROJECTION MODE
---------------
For every row we touch (above), if Projection.Mode is "EMPTY", flip it to
"User defined" so the values get picked up downstream. Rows whose Projection
Mode is already something else (e.g. linear projection, growth rate) are left
alone — those are deliberate.

Additionally, MinCapacityInvestment rows whose year cells are non-null but
mode is "EMPTY" also get flipped to "User defined". This activates planned-
build values (e.g. hydro 1.138 GW commissioned in 2026) that the upstream
pipeline left with mode=EMPTY, where downstream would silently ignore them.

SCOPE
-----
Defaults to acting on the 'Secondary Techs' sheet only:
  - Primary Techs: ALL cells are empty by design (mining/imports are unbounded
    in this model). Applying the rule would zero out every fuel source and
    make the model infeasible.
  - Demand Techs: end-use sinks. Different role; not the source of the bug.
  - Secondary Techs: where the PWR* generators that exhibit the bug live.

Use --sheets to override.

OUTPUT
------
1. A timestamped backup of the entire input directory (sibling folder).
2. In-place edit of A-O_Parametrization.xlsx (preserves all other content,
   formatting, and the 3 untouched AO files).
3. A JSON change log next to the backup, listing every cell that was modified
   and every existing nonzero value that was preserved.

USAGE
-----
    # From the t1_confection directory:
    python add_max_capacity_investment_rule.py

    # Override defaults:
    python add_max_capacity_investment_rule.py \\
        --input-dir A1_Outputs/A1_Outputs_BAU \\
        --sheets "Secondary Techs"
"""

from __future__ import annotations

import argparse
import json
import shutil
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
DEFAULT_TARGET_SHEETS = ["Secondary Techs"]
ALLOWED_FILL_VALUE = 9999
PARAM_FILE_NAME = "A-O_Parametrization.xlsx"

RES_PARAM = "ResidualCapacity"
MIN_INV_PARAM = "TotalAnnualMinCapacityInvestment"
MAX_CAP_PARAM = "TotalAnnualMaxCapacity"
MAX_INV_PARAM = "TotalAnnualMaxCapacityInvestment"

PROJ_MODE_COL = "Projection.Mode"
PROJ_MODE_EMPTY = "EMPTY"
PROJ_MODE_USER = "User defined"


# ---------------------------------------------------------------------------
# Backup
# ---------------------------------------------------------------------------
def make_backup(input_dir: Path) -> Path:
    """Copy `input_dir` to a timestamped sibling folder and return its path."""
    if not input_dir.is_dir():
        raise FileNotFoundError(f"Input directory does not exist: {input_dir}")
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = input_dir.parent / f"{input_dir.name}_PRE_MAXCAP_{stamp}"
    if backup.exists():
        raise FileExistsError(f"Backup folder already exists: {backup}")
    shutil.copytree(input_dir, backup)
    return backup


# ---------------------------------------------------------------------------
# Rule logic
# ---------------------------------------------------------------------------
def categorize_techs(df: pd.DataFrame, year_cols: list) -> tuple[set, set]:
    """
    Partition techs in `df` into (allowed, zeroed) sets.

    A tech is ALLOWED if either of these holds for any year column:
      - ResidualCapacity > 0
      - TotalAnnualMinCapacityInvestment > 0
    Otherwise it's ZEROED.
    """
    res = df[df["Parameter"] == RES_PARAM]
    mci = df[df["Parameter"] == MIN_INV_PARAM]

    res_max = res.set_index("Tech")[year_cols].fillna(0).max(axis=1)
    mci_max = mci.set_index("Tech")[year_cols].fillna(0).max(axis=1)

    techs_with_res = set(res_max[res_max > 0].index)
    techs_with_mci = set(mci_max[mci_max > 0].index)
    allowed = techs_with_res | techs_with_mci

    all_techs = set(df["Tech"].dropna().unique())
    zeroed = all_techs - allowed
    return allowed, zeroed


def find_year_columns(ws) -> dict:
    """Scan row 1 for integer year headers; return {year: column_index_1based}."""
    year_to_col = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if isinstance(val, int) and 1900 <= val <= 2200:
            year_to_col[val] = col_idx
    return year_to_col


def find_named_columns(ws, names) -> dict:
    """Return {name: column_index_1based} for headers matching `names`."""
    found = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val in names:
            found[val] = col_idx
    return found


def apply_rule_to_sheet(
    ws,
    allowed: set,
    zeroed: set,
    allowed_fill: int = ALLOWED_FILL_VALUE,
) -> dict:
    """
    Edit a worksheet in place, applying the rule. Returns a sheet-level log.
    """
    year_cols = find_year_columns(ws)
    headers = find_named_columns(ws, ["Tech", "Parameter", PROJ_MODE_COL])
    if "Tech" not in headers or "Parameter" not in headers:
        raise ValueError(
            f"Sheet '{ws.title}' missing required columns: "
            f"found {list(headers.keys())}"
        )
    tech_col = headers["Tech"]
    param_col = headers["Parameter"]
    proj_mode_col = headers.get(PROJ_MODE_COL)  # may be absent in some sheets

    log = {
        "sheet": ws.title,
        "years_found": sorted(year_cols.keys()),
        "allowed_count": len(allowed),
        "zeroed_count": len(zeroed),
        "changes_zeroed_techs": [],
        "changes_allowed_techs": [],
        "preserved_existing_values": [],
        "projection_mode_flips": [],
    }

    for row_idx in range(2, ws.max_row + 1):
        tech = ws.cell(row=row_idx, column=tech_col).value
        param = ws.cell(row=row_idx, column=param_col).value
        if tech is None:
            continue

        # Whether this row needs its Projection.Mode flipped at the end of
        # the iteration. True for any row where (a) we wrote at least one
        # cell, or (b) MinCapInvestment with any pre-existing non-null cell
        # (downstream ignores rows with mode=EMPTY, even if they have data).
        row_needs_mode_flip = False

        # ZEROED techs: blanket-zero MaxCap and MaxInv rows
        if tech in zeroed and param in (MAX_CAP_PARAM, MAX_INV_PARAM):
            row_needs_mode_flip = True
            for year, col in year_cols.items():
                cell = ws.cell(row=row_idx, column=col)
                old = cell.value
                # Treat None and any non-zero as a change to 0
                if old != 0:
                    cell.value = 0
                    log["changes_zeroed_techs"].append(
                        {
                            "tech": tech,
                            "parameter": param,
                            "year": year,
                            "old": old,
                            "new": 0,
                        }
                    )

        # ALLOWED techs: fill EMPTY MaxInv cells with 9999. Preserve every
        # explicit value, including explicit 0s — those encode real modeling
        # decisions (e.g. "no investment in 2023 base year", "transmission
        # interconnect can't come online until 2030"). Only None/blank means
        # "no upper bound was set," which is what 9999 stands in for.
        elif tech in allowed and param == MAX_INV_PARAM:
            row_needs_mode_flip = True
            for year, col in year_cols.items():
                cell = ws.cell(row=row_idx, column=col)
                old = cell.value
                if old is None:
                    cell.value = allowed_fill
                    log["changes_allowed_techs"].append(
                        {
                            "tech": tech,
                            "parameter": param,
                            "year": year,
                            "old": old,
                            "new": allowed_fill,
                        }
                    )
                else:
                    log["preserved_existing_values"].append(
                        {
                            "tech": tech,
                            "parameter": param,
                            "year": year,
                            "value": old,
                        }
                    )

        # MinCapInvestment rows: don't write any cells, but if the row has
        # any non-null data, flag for the mode flip below. This activates
        # planned-build values (e.g. hydro 1.138 in 2026) that the upstream
        # pipeline left with mode=EMPTY, where they'd be silently ignored.
        elif param == MIN_INV_PARAM:
            has_any_data = any(
                ws.cell(row=row_idx, column=col).value is not None
                for col in year_cols.values()
            )
            if has_any_data:
                row_needs_mode_flip = True

        # If this row needs it, flip Projection.Mode from "EMPTY" to
        # "User defined" so the values get picked up downstream. Existing
        # non-EMPTY modes are left alone.
        if row_needs_mode_flip and proj_mode_col is not None:
            mode_cell = ws.cell(row=row_idx, column=proj_mode_col)
            if mode_cell.value == PROJ_MODE_EMPTY:
                mode_cell.value = PROJ_MODE_USER
                log["projection_mode_flips"].append(
                    {
                        "tech": tech,
                        "parameter": param,
                        "old": PROJ_MODE_EMPTY,
                        "new": PROJ_MODE_USER,
                    }
                )

    return log


def edit_parametrization(filepath: Path, sheets: list) -> dict:
    """Apply the rule to `sheets` in the parametrization workbook (in place)."""
    df_all = pd.read_excel(filepath, sheet_name=None)
    wb = load_workbook(filepath)

    file_log = {"file": str(filepath), "sheets": []}

    for sheet in sheets:
        if sheet not in wb.sheetnames:
            file_log["sheets"].append(
                {"sheet": sheet, "skipped": "sheet not present in workbook"}
            )
            continue

        df = df_all[sheet]
        year_cols = [c for c in df.columns if isinstance(c, int)]
        if not year_cols:
            file_log["sheets"].append(
                {"sheet": sheet, "skipped": "no integer year columns found"}
            )
            continue

        allowed, zeroed = categorize_techs(df, year_cols)
        ws = wb[sheet]
        sheet_log = apply_rule_to_sheet(ws, allowed, zeroed)
        sheet_log["allowed_techs"] = sorted(allowed)
        sheet_log["zeroed_techs"] = sorted(zeroed)
        file_log["sheets"].append(sheet_log)

    wb.save(filepath)
    return file_log


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------
def run(
    input_dir,
    sheets: list = None,
    skip_backup: bool = False,
) -> dict:
    """End-to-end: backup, edit, write log. Returns the log dict."""
    input_dir = Path(input_dir)
    sheets = sheets or DEFAULT_TARGET_SHEETS

    backup_dir = None if skip_backup else make_backup(input_dir)

    paramfile = input_dir / PARAM_FILE_NAME
    if not paramfile.exists():
        raise FileNotFoundError(f"{paramfile} not found")

    log = edit_parametrization(paramfile, sheets)
    log["backup_dir"] = str(backup_dir) if backup_dir else None
    log["timestamp"] = datetime.now().isoformat()

    if backup_dir is not None:
        log_path = backup_dir.parent / f"{backup_dir.name}_CHANGES.json"
        log_path.write_text(json.dumps(log, indent=2, default=str))
        log["log_path"] = str(log_path)

    return log


def print_summary(log: dict) -> None:
    """Pretty-print the run summary."""
    bar = "=" * 72
    print(bar)
    print("MaxCapacityInvestment guard rule — applied")
    print(bar)
    print(f"Backup folder : {log.get('backup_dir', '(skipped)')}")
    print(f"Edited file   : {log['file']}")
    print()
    for s in log["sheets"]:
        if "skipped" in s:
            print(f"[SKIPPED] '{s['sheet']}': {s['skipped']}")
            continue
        years = s["years_found"]
        print(f"Sheet: '{s['sheet']}'")
        print(f"  Years        : {years[0]}..{years[-1]} ({len(years)} years)")
        print(f"  ALLOWED techs (fill EMPTY MaxInv with {ALLOWED_FILL_VALUE}): "
              f"{s['allowed_count']}")
        print(f"  ZEROED  techs (MaxCap=0 and MaxInv=0)               : "
              f"{s['zeroed_count']}")
        print(f"  Empty cells filled in ALLOWED techs : "
              f"{len(s['changes_allowed_techs'])}")
        print(f"  Cells changed in ZEROED  techs      : "
              f"{len(s['changes_zeroed_techs'])}")
        print(f"  Existing values preserved (incl. explicit 0s) : "
              f"{len(s['preserved_existing_values'])}")
        print(f"  Projection.Mode flips (EMPTY -> User defined) : "
              f"{len(s['projection_mode_flips'])}")
        if s["preserved_existing_values"]:
            preserved_techs = sorted(
                {p["tech"] for p in s["preserved_existing_values"]}
            )
            print(f"    -> across {len(preserved_techs)} tech(s): "
                  f"{', '.join(preserved_techs[:6])}"
                  f"{' ...' if len(preserved_techs) > 6 else ''}")
    if log.get("log_path"):
        print(f"\nDetailed change log written to: {log['log_path']}")


def main() -> int:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=Path("A1_Outputs/A1_Outputs_BAU"),
        help="Directory containing the AO files (default: A1_Outputs/A1_Outputs_BAU)",
    )
    parser.add_argument(
        "--sheets",
        nargs="+",
        default=DEFAULT_TARGET_SHEETS,
        help=f"Sheets to apply the rule to (default: {DEFAULT_TARGET_SHEETS})",
    )
    parser.add_argument(
        "--skip-backup",
        action="store_true",
        help="Skip backup creation (DANGEROUS — for testing only)",
    )
    args = parser.parse_args()

    try:
        log = run(args.input_dir, args.sheets, args.skip_backup)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    print_summary(log)
    return 0


if __name__ == "__main__":
    sys.exit(main())
