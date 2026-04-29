# -*- coding: utf-8 -*-
"""
Created on 2026
@author: Climate Lead Group, Luis Victor-Gallardo

Inserts a 'System Parameters' sheet into A-O_Parametrization with
ReserveMargin = 1.15 (flat, all years). Re-runnable: if the sheet
already exists it is replaced.

Run with F5 in Spyder. Edit USER CONFIGURATION below.
"""
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# ======================================================================
# USER CONFIGURATION
# ======================================================================
WORK_DIR = r'C:\Users\luisfernando\Desktop\OSeMOSYS\asia_ostram_refactored\t1_confection\A1_Outputs\A1_Outputs_BAU'
FILE_NAME = 'A-O_Parametrization.xlsx'
BASE_YEAR = 2023
END_YEAR  = 2050
RESERVE_MARGIN = 1.15   # 15% reserve margin, flat across all years
# ======================================================================

import os
file_path = os.path.join(WORK_DIR, FILE_NAME)

wb = openpyxl.load_workbook(file_path)

# Remove existing sheet if re-running
if 'System Parameters' in wb.sheetnames:
    del wb['System Parameters']

ws = wb.create_sheet('System Parameters')

years = list(range(BASE_YEAR, END_YEAR + 1))

# Header row
headers = ['Parameter', 'Unit'] + years
for c, val in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=val)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

# ReserveMargin row
ws.cell(row=2, column=1, value='ReserveMargin')
ws.cell(row=2, column=2, value='ratio')
for c, yr in enumerate(years, 3):
    ws.cell(row=2, column=c, value=RESERVE_MARGIN)

# Column widths
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 10
for c in range(3, 3 + len(years)):
    ws.column_dimensions[get_column_letter(c)].width = 8

wb.save(file_path)
print(f'Done — "System Parameters" sheet written to {FILE_NAME}')
print(f'  ReserveMargin = {RESERVE_MARGIN} for {BASE_YEAR}–{END_YEAR}')
