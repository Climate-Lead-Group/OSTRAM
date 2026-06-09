"""fix_elc_pmode_revert.py
=============================
Reproduce la edicion manual que Luis hizo entre la corrida 2 de
add_max_capacity_investment_rule.py (10:55) y B1b_Pre_solver_validation.py
(12:01) el 30-abril-2026.

QUE HACE
--------
En la hoja 'Secondary Techs', revierte la columna 'Projection.Mode' de
"User defined" a "EMPTY" para 10 techs electricos x 2 parametros = 20 cells:

  Techs (10):
    ELCBGDXX01, ELCBTNXX01, ELCINDEA01, ELCINDNE01, ELCINDNO01,
    ELCINDSO01, ELCINDWE01, ELCLKAXX01, ELCMDVXX01, ELCNPLXX01
  Parametros (2):
    TotalAnnualMaxCapacity, TotalAnnualMaxCapacityInvestment

POR QUE
-------
El segundo run de add_max_capacity_investment_rule.py (commit 2be1616)
flippea Projection.Mode EMPTY -> User defined para todos los techs ZEROED
en MaxCap/MaxCapInv. Los ELC*01 son ZEROED (lockout: el modelo no debe
invertir en una "tech de electricidad placeholder", solo usar los PWR*).
Luis decidio que el 'User defined' era engañoso para esos lockout rows
(las celdas de año son 0, sin valor real configurable) y los volvio a
"EMPTY" a mano. Los rows estan funcionalmente desactivados igual.

Idempotente: re-correr es no-op (ya estan en EMPTY).

Uso:
    python fix_elc_pmode_revert.py --input <A-O_Parametrization.xlsx>
    python fix_elc_pmode_revert.py --input <in> --output <out>  # write a copy
"""
from __future__ import annotations
import argparse
from pathlib import Path
import sys
from openpyxl import load_workbook

TARGET_SHEET = "Secondary Techs"
TECH_COL_NAME = "Tech"
PARAM_COL_NAME = "Parameter"
PMODE_COL_NAME = "Projection.Mode"

OLD_VALUE = "User defined"
NEW_VALUE = "EMPTY"

TARGET_TECHS = {
    "ELCBGDXX01", "ELCBTNXX01", "ELCINDEA01", "ELCINDNE01", "ELCINDNO01",
    "ELCINDSO01", "ELCINDWE01", "ELCLKAXX01", "ELCMDVXX01", "ELCNPLXX01",
}
TARGET_PARAMS = {
    "TotalAnnualMaxCapacity", "TotalAnnualMaxCapacityInvestment",
}


def revert_pmode(input_path: Path, output_path: Path | None = None) -> int:
    """Revert Projection.Mode for ELC*01 lockout rows. Returns cells changed."""
    if output_path is None:
        output_path = input_path
    wb = load_workbook(input_path)
    if TARGET_SHEET not in wb.sheetnames:
        sys.exit(f"ERROR: sheet '{TARGET_SHEET}' not in workbook")
    ws = wb[TARGET_SHEET]

    headers = {c.value: c.column for c in ws[1] if c.value is not None}
    for col in (TECH_COL_NAME, PARAM_COL_NAME, PMODE_COL_NAME):
        if col not in headers:
            sys.exit(f"ERROR: required column '{col}' missing in '{TARGET_SHEET}'")
    tech_col = headers[TECH_COL_NAME]
    param_col = headers[PARAM_COL_NAME]
    pmode_col = headers[PMODE_COL_NAME]

    cells_changed = 0
    rows_touched: list[tuple[int, str, str, object, object]] = []
    rows_skipped_already_empty: list[tuple[int, str, str]] = []
    for r in range(2, ws.max_row + 1):
        tech = ws.cell(r, tech_col).value
        if tech not in TARGET_TECHS:
            continue
        param = ws.cell(r, param_col).value
        if param not in TARGET_PARAMS:
            continue
        cell = ws.cell(r, pmode_col)
        old = cell.value
        if old == NEW_VALUE:
            rows_skipped_already_empty.append((r, str(tech), str(param)))
            continue
        cell.value = NEW_VALUE
        rows_touched.append((r, str(tech), str(param), old, NEW_VALUE))
        cells_changed += 1

    wb.save(output_path)

    print(f"Reverted {cells_changed} cells (Projection.Mode -> '{NEW_VALUE}')")
    if rows_touched:
        for r, t, p, old, new in rows_touched:
            print(f"  row {r:>5}  {t:<14}  {p:<35}  {old!r} -> {new!r}")
    if rows_skipped_already_empty:
        print(f"Skipped {len(rows_skipped_already_empty)} cells already at '{NEW_VALUE}' "
              f"(idempotent re-run)")
    print(f"Saved: {output_path}")
    return cells_changed


if __name__ == "__main__":
    p = argparse.ArgumentParser(description=__doc__.split("\n")[1])
    p.add_argument("--input", type=Path, required=True)
    p.add_argument("--output", type=Path, default=None)
    args = p.parse_args()
    if not args.input.is_file():
        sys.exit(f"ERROR: input not found: {args.input}")
    revert_pmode(args.input, args.output)
