"""
add_storage_min_investment.py
==============================

Independent patch — writes small TotalAnnualMinCapacityInvestment values
for storage technologies into A-O_Parametrization.xlsx, nudging the
optimizer to deploy storage gradually rather than deferring to end-of-
horizon.

This is a non-destructive patcher: reads input, writes a timestamped
sibling output, never mutates the source file.

USAGE
-----
    python add_storage_min_investment.py
    python add_storage_min_investment.py --input-dir A1_Outputs/A1_Outputs_BAU
    python add_storage_min_investment.py --yaml storage_floors.yaml

CONFIG
------
Reads storage_floors.yaml (next to this script, or --yaml override):

    storage_floors:
      SDS:          # family code (chars 3..6 of tech)
        2027: 0.1   # GW MinCapInv per tech in that family
        2030: 0.2
        2035: 0.3
      LDS:
        2030: 0.1
        2035: 0.2

Years between anchors are linearly interpolated. Years before the first
anchor get 0 (no floor). Years after the last anchor hold the last value.
Values are per-tech (e.g. 0.1 GW for PWRSDSINDNO, 0.1 GW for PWRSDSINDSO,
etc.) — NOT summed across the family.

Only cells that are currently empty (None/NaN) or zero are written.
Existing non-zero MinCapInv values (real planning data) are preserved.
"""

from __future__ import annotations

import argparse
import json
import shutil
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

try:
    import yaml as _yaml
    def _load_yaml(path: Path) -> dict:
        with open(path, "r", encoding="utf-8") as f:
            return _yaml.safe_load(f)
except ImportError:
    _yaml = None
    def _load_yaml(path: Path) -> dict:
        raise ImportError("PyYAML is required. Install with: pip install pyyaml")


# ---------------------------------------------------------------------------
# Defaults
# ---------------------------------------------------------------------------
PARAM_FILE_NAME = "A-O_Parametrization.xlsx"
YAML_FILE_NAME = "storage_floors.yaml"
TARGET_SHEET = "Secondary Techs"
MIN_INV_PARAM = "TotalAnnualMinCapacityInvestment"
PROJ_MODE_COL = "Projection.Mode"
PROJ_MODE_EMPTY = "EMPTY"
PROJ_MODE_USER = "User defined"


# ---------------------------------------------------------------------------
# Interpolation (same logic as lid script)
# ---------------------------------------------------------------------------
def interpolate_schedule(schedule: dict[int, float], year: int) -> float:
    """Linearly interpolate a year→value schedule.

    Years before the first anchor return 0.
    Years after the last anchor return the last anchor's value.
    """
    if not schedule:
        return 0.0
    anchors = sorted(schedule.items())
    if year < anchors[0][0]:
        return 0.0
    if year >= anchors[-1][0]:
        return anchors[-1][1]
    for i in range(len(anchors) - 1):
        y0, v0 = anchors[i]
        y1, v1 = anchors[i + 1]
        if y0 <= year <= y1:
            frac = (year - y0) / (y1 - y0)
            return v0 + frac * (v1 - v0)
    return 0.0


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
def load_config(yaml_path: Path) -> dict[str, dict[int, float]]:
    """Load and validate the YAML config.

    Returns dict of family_code -> {year: floor_gw}.
    """
    cfg = _load_yaml(yaml_path)
    if cfg is None:
        cfg = {}
    raw = cfg.get("storage_floors", {})
    if not raw:
        raise ValueError(
            f"No 'storage_floors' key found in {yaml_path}. "
            f"Expected a dict of family_code -> year -> GW."
        )
    out: dict[str, dict[int, float]] = {}
    for fam, schedule in raw.items():
        fam = str(fam).strip().upper()
        if not isinstance(schedule, dict):
            raise ValueError(
                f"storage_floors.{fam} must be a dict of year: GW, "
                f"got {type(schedule).__name__}"
            )
        out[fam] = {int(y): float(v) for y, v in schedule.items()}
    return out


# ---------------------------------------------------------------------------
# Main patch logic
# ---------------------------------------------------------------------------
def run(input_dir: Path, yaml_path: Path | None = None,
        skip_backup: bool = False) -> dict:
    """Apply storage MinCapInv floors to the parametrization workbook."""

    # Resolve paths
    input_dir = Path(input_dir)
    filepath = input_dir / PARAM_FILE_NAME
    if not filepath.is_file():
        raise FileNotFoundError(f"Parametrization file not found: {filepath}")

    script_dir = Path(__file__).resolve().parent
    if yaml_path is None:
        yaml_path = script_dir / YAML_FILE_NAME
    if not yaml_path.is_file():
        raise FileNotFoundError(f"YAML config not found: {yaml_path}")

    floors = load_config(yaml_path)
    print(f"Loaded storage floors for families: {list(floors.keys())}")
    for fam, sched in floors.items():
        anchors = ", ".join(f"{y}: {v} GW" for y, v in sorted(sched.items()))
        print(f"  {fam}: {anchors}")

    # Backup
    backup_dir = None
    if not skip_backup:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_dir = input_dir.parent / f"{input_dir.name}_PRE_STOFLOOR_{stamp}"
        if backup_dir.exists():
            raise FileExistsError(f"Backup already exists: {backup_dir}")
        shutil.copytree(input_dir, backup_dir)
        print(f"Backup: {backup_dir}")

    # Load workbook
    wb = load_workbook(filepath)
    if TARGET_SHEET not in wb.sheetnames:
        raise ValueError(f"Sheet '{TARGET_SHEET}' not found in {filepath}")
    ws = wb[TARGET_SHEET]

    # Read as DataFrame for column discovery
    df = pd.read_excel(filepath, sheet_name=TARGET_SHEET)

    # Find year columns
    year_cols: dict[int, int] = {}
    for col_idx, col_name in enumerate(df.columns, start=1):
        try:
            y = int(col_name)
            if 2020 <= y <= 2060:
                year_cols[y] = col_idx
        except (ValueError, TypeError):
            pass
    years = sorted(year_cols.keys())
    print(f"Year columns found: {years[0]}..{years[-1]} ({len(years)} years)")

    # Find key columns
    tech_col = None
    param_col = None
    proj_mode_col_idx = None
    for col_idx, col_name in enumerate(df.columns, start=1):
        cn = str(col_name).strip()
        if cn == "Technology":
            tech_col = col_idx
        elif cn == "Parameter":
            param_col = col_idx
        elif cn == PROJ_MODE_COL:
            proj_mode_col_idx = col_idx

    if tech_col is None or param_col is None:
        raise ValueError(
            f"Could not find 'Technology' and 'Parameter' columns in "
            f"'{TARGET_SHEET}'"
        )

    # Scan rows
    log = {
        "file": str(filepath),
        "yaml": str(yaml_path),
        "backup_dir": str(backup_dir) if backup_dir else None,
        "floors": {f: dict(s) for f, s in floors.items()},
        "changes": [],
        "preserved": [],
        "projection_mode_flips": [],
    }

    header_row = 1  # openpyxl is 1-indexed; row 1 = header
    for row_idx in range(header_row + 1, ws.max_row + 1):
        tech = ws.cell(row=row_idx, column=tech_col).value
        param = ws.cell(row=row_idx, column=param_col).value

        if tech is None or param is None:
            continue
        tech = str(tech).strip()
        param = str(param).strip()

        # Only MinCapInv rows for storage techs
        if param != MIN_INV_PARAM:
            continue
        if not tech.startswith("PWR"):
            continue
        fam = tech[3:6]
        if fam not in floors:
            continue

        schedule = floors[fam]
        row_modified = False

        for year in years:
            col = year_cols[year]
            cell = ws.cell(row=row_idx, column=col)
            old = cell.value

            floor_val = interpolate_schedule(schedule, year)
            if floor_val <= 0:
                continue

            # Only write if cell is empty or zero
            is_empty = (
                old is None
                or (isinstance(old, (int, float))
                    and (pd.isna(old) or float(old) == 0.0))
            )

            if is_empty:
                cell.value = round(floor_val, 4)
                row_modified = True
                log["changes"].append({
                    "tech": tech, "year": year,
                    "old": old, "new": floor_val,
                    "reason": "storage_floor",
                })
            else:
                log["preserved"].append({
                    "tech": tech, "year": year,
                    "value": old, "floor_would_be": floor_val,
                })

        # Flip Projection.Mode if we modified this row
        if row_modified and proj_mode_col_idx is not None:
            mode_cell = ws.cell(row=row_idx, column=proj_mode_col_idx)
            if mode_cell.value == PROJ_MODE_EMPTY:
                mode_cell.value = PROJ_MODE_USER
                log["projection_mode_flips"].append({"tech": tech})

    # Save
    wb.save(filepath)
    wb.close()

    # Write change log
    if backup_dir is not None:
        log_path = backup_dir.parent / f"{backup_dir.name}_CHANGES.json"
        log_path.write_text(json.dumps(log, indent=2, default=str))
        log["log_path"] = str(log_path)

    return log


def print_summary(log: dict) -> None:
    """Pretty-print the run summary."""
    bar = "=" * 60
    print(bar)
    print("Storage MinCapacityInvestment floors — applied")
    print(bar)
    print(f"File          : {log['file']}")
    print(f"YAML config   : {log['yaml']}")
    print(f"Backup        : {log.get('backup_dir', '(skipped)')}")
    print(f"Cells written : {len(log['changes'])}")
    print(f"Cells preserved (existing non-zero): {len(log['preserved'])}")
    print(f"Proj.Mode flips: {len(log['projection_mode_flips'])}")

    if log["changes"]:
        # Summary by tech
        from collections import Counter
        tech_counts = Counter(c["tech"] for c in log["changes"])
        print(f"\nTechs patched ({len(tech_counts)}):")
        for t, n in sorted(tech_counts.items()):
            print(f"  {t}: {n} year-cells")

    if log.get("log_path"):
        print(f"\nDetailed log: {log['log_path']}")


def main() -> int:
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--input-dir", type=Path,
        default=Path("A1_Outputs/A1_Outputs_BAU"),
        help="Directory containing the AO files",
    )
    parser.add_argument(
        "--yaml", type=Path, default=None,
        help=f"YAML config path (default: {YAML_FILE_NAME} next to script)",
    )
    parser.add_argument(
        "--skip-backup", action="store_true",
        help="Skip backup (DANGEROUS — testing only)",
    )
    args = parser.parse_args()

    try:
        log = run(args.input_dir, args.yaml, args.skip_backup)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    print_summary(log)
    return 0


if __name__ == "__main__":
    sys.exit(main())
