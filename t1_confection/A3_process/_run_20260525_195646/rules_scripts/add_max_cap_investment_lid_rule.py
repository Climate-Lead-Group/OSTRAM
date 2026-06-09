"""
add_max_cap_investment_lid_rule.py
===================================

Independent patch — applies a per-year LID and the V1 UNTIE RULE to
TotalAnnualMaxCapacityInvestment for GENERATION technologies that have
either ResidualCapacity > 0 or TotalAnnualMinCapacityInvestment > 0 in any
year.

This script is INDEPENDENT of add_max_capacity_investment_rule.py. It can run:
  - On a fresh A-O_Parametrization.xlsx (operates on empty MaxInv cells).
  - On the output of the first-patch script (recognizes 9999 placeholders and
    replaces them with the lid).
  - Repeatedly (idempotent — second run produces no changes).

PROBLEM
-------
For ALLOWED generation techs (those with residual or planned capacity), MaxInv
was being left effectively unbounded — either NaN (no constraint at all) or
9999 (a placeholder that's far above any realistic ramp). Without a real upper
bound, the optimizer can over-invest in candidate techs whose annual ramp
should be physically constrained. And without a year-by-year *ramp* tied to
demand, a constant lid either binds in late years (when demand has scaled
3x) or is too generous in early years.

RULE
----
For each tech in the ALLOWED ∩ GENERATION set, the per-year
MaxCapacityInvestment lid is computed in one of two modes, selected by
LID_RULE_MODE.

Common quantities (used by both modes):
    pool(cr, y)         = sum of ResidualCapacity(t, y) for every ALLOWED
                          GENERATION tech t in cr
    mult(cr, y)         = demand(cr, y) / demand(cr, DEMAND_REFERENCE_YEAR)
    scaled_pool(cr, y)  = pool(cr, y) * mult(cr, y)
    country_region(t)   = chars 6..10 of the tech code (e.g. BGDXX, INDNE)

Mode "uniform" (default):
    Every allowed tech in the same cr gets the SAME lid value.

        pct(cr, y) = base_pct(y) * mult(cr, y)
        where base_pct(y) = LID_PERCENTAGE_BY_YEAR.get(y, LID_PERCENTAGE_DEFAULT)
        lid(t, y)  = pct(cr, y) * pool(cr, y)
                   = base_pct(y) * scaled_pool(cr, y)

    The base_pct schedule is a per-decade plateau encoding "tight while we
    have planning data, looser as the horizon gets speculative." Demand
    growth is layered on linearly via mult — fast-growing crs get
    proportionally more headroom.

Mode "proportional":
    Each tech gets a lid sized to its share of the 2024 fleet. Total
    headroom is the year-over-year growth in scaled_pool, distributed
    proportionally and slackened by a security factor.

        tech_share(t)     = ResidualCapacity(t, ref_year) / pool(cr, ref_year)
        pool_delta(cr, y) = max(0, scaled_pool(cr, y) - scaled_pool(cr, y-1))
        lid(t, y)         = LID_SECURITY_FACTOR * tech_share(t) * pool_delta(cr, y)

    The max(0, ...) guard ensures that flat-or-declining demand years
    yield zero new headroom (no negative lids). For the reference year and
    earlier, lid = 0 (no prior year to delta against), so MinCapInv must
    cover any required ref-year build via the untie rule.

Both modes are floored by the V1 untie rule: if MinCapInv(t, y) > 0 and
the proposed lid <= MinCapInv(t, y), we push the lid to
MinCapInv(t, y) * UNTIE_MULTIPLIER. This guarantees the LP-feasibility
invariant MinCapInv < MaxCapInv whenever MinCapInv > 0.

GENERATION FILTER
-----------------
"Generation" is defined by TECH_TYPES.csv (columns: 'Technology (PWR)',
'Technology'). Only techs categorized as 'GENERATION' are eligible. This
correctly excludes storage (e.g. PWRSDSLKAXX), interconnectors, primary
fuels, etc., which need separate lid policies.

DEMAND-ANCHORED RAMP
--------------------
Demand multiplier per (country+region, year) is read from A-O_Demand.xlsx
(sheet 'Demand_Projection', rows where Demand/Share == 'Demand'). Country+
region is parsed from chars 3..7 of the Fuel/Tech code (ELCBGDXX03 -> BGDXX).
Demand rows are summed within each cr and divided by the value at the
reference year (default 2024) to obtain mult(cr, y).

Cell-by-cell, for each MaxInv cell of an ALLOWED tech:
    1. Empty (None) or 9999 placeholder  ->  proposed = lid
    2. Other explicit value (manual cal)  ->  proposed = current value (preserved)
    3. V1 UNTIE RULE (cf. B1b_Pre_solver_validation.py):
         if MinCapacityInvestment(tech, y) >= proposed:
             proposed = MinCapacityInvestment(tech, y) * UNTIE_MULTIPLIER

This guarantees Max_inv > Min_inv per year, eliminating that class of solver
infeasibility.

SCOPE
-----
- ALLOWED set: techs with residual > 0 OR min-cap-investment > 0 in any year.
- TRN* transmission interconnects (length-13 codes) are SKIPPED entirely —
  they span two country+region pairs and don't fit a single pool. Their
  manually-calibrated MaxInv values are left untouched.
- Non-allowed techs are also untouched (handled by a separate patch).

CONFIGURATION
-------------
Edit LID_PERCENTAGE_DEFAULT and LID_PERCENTAGE_BY_YEAR at the top of this
file to change the lid:

    LID_PERCENTAGE_DEFAULT = 0.005       # default 0.5% of pool per year
    LID_PERCENTAGE_BY_YEAR = {            # per-year overrides (optional)
        2030: 0.01,
        2040: 0.02,
    }

OUTPUT
------
1. A timestamped backup of the input directory.
2. In-place edit of A-O_Parametrization.xlsx.
3. A JSON change log next to the backup with per-cell reasons and the
   per-(country_region, year) pool sizes used.

USAGE
-----
    # From the t1_confection directory:
    python add_max_cap_investment_lid_rule.py

    # Override defaults:
    python add_max_cap_investment_lid_rule.py \\
        --input-dir A1_Outputs/A1_Outputs_BAU \\
        --sheets "Secondary Techs"

    # Restore from the most recent _PRE_LID_* backup:
    python add_max_cap_investment_lid_rule.py --restore

    # Restore from a specific backup folder:
    python add_max_cap_investment_lid_rule.py \\
        --restore-from A1_Outputs/A1_Outputs_BAU_PRE_LID_20260430_204529
"""

from __future__ import annotations

import argparse
import gc
import json
import shutil
import sys
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

try:
    import yaml as _yaml
    def _load_yaml(path: Path) -> dict:
        with open(path, "r", encoding="utf-8") as f:
            return _yaml.safe_load(f)
except ImportError:
    _yaml = None
    def _load_yaml(path: Path) -> dict:
        raise ImportError("PyYAML is required. Install with: pip install pyyaml")

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
DEFAULT_TARGET_SHEETS = ["Secondary Techs"]
PARAM_FILE_NAME = "A-O_Parametrization.xlsx"

# Optional YAML override. The orchestrator (A3_process.py) detects this
# constant and stages the matching YAML from rules_scripts/configs/<scenario>/
# alongside this script when the scenario provides one. When present, its
# values override the module-level constants below for the lifetime of this
# process. Absent YAML = current module defaults remain in effect (so BAU and
# scenarios without a YAML override behave exactly as before).
YAML_FILE_NAME = "lid_rule.yaml"

RES_PARAM = "ResidualCapacity"
MIN_INV_PARAM = "TotalAnnualMinCapacityInvestment"
MAX_INV_PARAM = "TotalAnnualMaxCapacityInvestment"

PROJ_MODE_COL = "Projection.Mode"
PROJ_MODE_EMPTY = "EMPTY"
PROJ_MODE_USER = "User defined"

# A leftover sentinel from the first-patch script. If we encounter this value,
# treat the cell as a placeholder and replace it with the lid (rather than as
# a manually calibrated value to preserve).
PLACEHOLDER_VALUE = 9999

# Tech naming conventions
PWR_TECH_LENGTH = 11    # e.g. PWRHYDBGDXX (3-letter prefix + 3 fuel + 3 country + 2 region)
TRN_TECH_LENGTH = 13    # transmission interconnects, e.g. TRNINDEAINDNE -- skipped
COUNTRY_REGION_SLICE = slice(6, 11)  # for length-11 PWR* codes: chars 6..10

# ---------------------------------------------------------------------------
# Tech-type filter — only patch GENERATION technologies
# ---------------------------------------------------------------------------
# TECH_TYPES.csv lives next to this script. It lists techs by category
# (GENERATION, INTERCONNECTORS, STORAGE_LONG, STORAGE_SHORT, etc.). Only
# techs in GENERATION_CATEGORY are eligible for the lid + untie rule.
# Set RESTRICT_TO_GENERATION = False to fall back to the prior behavior
# (any allowed tech with a length-11 PWR* shape).
TECH_TYPES_FILE = "TECH_TYPES.csv"
TECH_TYPES_CATEGORY_COL = "Technology (PWR)"
TECH_TYPES_TECH_COL = "Technology"
GENERATION_CATEGORY = "GENERATION"
RESTRICT_TO_GENERATION = True

# ---------------------------------------------------------------------------
# Lid rule configuration
# ---------------------------------------------------------------------------
# LID_RULE_MODE selects which lid formula to use. Both modes share the same
# pool definition and demand multiplier; they differ in *how* the lid is
# distributed among techs. See top-of-file RULE block for full formulas.
#
#   "uniform"      — Per-decade pct schedule, applied uniformly across all
#                    techs in the same cr. Same lid value for every allowed
#                    gen tech in cr per year. Use this for "let the optimizer
#                    pick winners" scenarios and for late-horizon stress tests
#                    where the planning anchor is intentionally relaxed.
#
#   "proportional" — Each tech gets a lid sized to its 2024 fleet share,
#                    times the year-over-year growth in scaled_pool, times
#                    a slack factor. Use this when the BAU narrative is
#                    "fleet evolves proportionally to current composition."
LID_RULE_MODE = "proportional"

# --- Uniform mode parameters --------------------------------------------------
# Per-year base percentage. Encodes "tight near-term, loose late-horizon."
# In uniform mode: pct(cr, y) = LID_PERCENTAGE_BY_YEAR[y] * mult(cr, y)
# Schedule rationale:
#   2023-2030 = 0.5%   - matches current near-term lid; respects national IRPs
#                        which typically have visibility through ~2030.
#   2031-2040 = 10%    - planning data thins past 2030; 20x jump frees the
#                        optimizer enough to substitute for storage if needed.
#   2041-2050 = 50%    - effectively unbounded; "we don't know what build
#                        rates will be in 2045+, let the model decide."
# Years not in this dict fall back to LID_PERCENTAGE_DEFAULT.
LID_PERCENTAGE_DEFAULT = 0.5
LID_PERCENTAGE_BY_YEAR: dict = {
    2023: 0.05, 2024: 0.05, 2025: 0.05, 2026: 0.05, 2027: 0.05,
    2028: 0.05, 2029: 0.05, 2030: 0.1,
    2031: 0.1,  2032: 0.1,  2033: 0.1,  2034: 0.1,  2035: 0.1,
    2036: 0.1,  2037: 0.1,  2038: 0.1,  2039: 0.1,  2040: 0.2,
    2041: 0.20,  2042: 0.20,  2043: 0.20,  2044: 0.20,  2045: 0.20,
    2046: 0.20,  2047: 0.20,  2048: 0.20,  2049: 0.20,  2050: 0.20,
}

# --- Proportional mode parameters --------------------------------------------
# Slack on the proportional-share lid:
#   lid(t, y) = LID_SECURITY_FACTOR * tech_share(t) * pool_delta(cr, y)
# Setting this to exactly 1.0 means each tech's lid equals its proportional
# share of the year's pool growth. Values >= 1.0 add slack to avoid binding
# the optimizer at the strict proportional split. 1.1 is a debugging knob
# for unblocking solver edge cases; values much above 1.5 dilute the
# proportional-allocation narrative.
LID_SECURITY_FACTOR = 1.1

# --- Relaxation schedule (used by both modes) --------------------------------
# Year-keyed multiplier on the computed lid.  Default {}: multiplier = 1.0 for
# every year (BAU behaviour, identical to the old script).
#
# For optimisation scenarios, set anchor points in the YAML, e.g.:
#   relaxation_schedule:
#     2023: 1.0
#     2030: 1.5
#     2040: 3.0
#     2050: 5.0
#
# Years between anchors are linearly interpolated. Years before the first
# anchor get its value; years after the last anchor get the last value.
# The final lid is  min(relaxation(year) * lid_base, 9999).
#
# This lets the optimizer gradually diverge from the BAU fleet-share trajectory
# without a spike (near-term ≈ BAU) and without single-tech dominance (the
# proportional structure is preserved — all families get the same multiplier).
LID_RELAXATION_SCHEDULE: dict = {}

# --- Exempt prefixes (used by both modes) ------------------------------------
# When non-empty, any tech whose code starts with one of these prefixes gets
# MaxCapInv = 9999 (uncapped) instead of the computed lid.
# Kept for backward compatibility; relaxation_schedule is the preferred
# mechanism.  Default: empty list (all techs get the lid).
LID_EXEMPT_PREFIXES: list = []

# --- Demand ramp (used by both modes) ----------------------------------------
# When LID_RAMP_FROM_DEMAND is True, mult(cr, y) is computed and applied:
#   uniform mode      -> pct(cr, y) = base_pct(y) * mult(cr, y)
#   proportional mode -> scaled_pool(cr, y) = pool(cr, y) * mult(cr, y)
# When False, mult collapses to 1.0 everywhere (uniform mode becomes flat
# pct schedule; proportional mode becomes share * unscaled pool delta,
# which is typically zero since residual is roughly flat).
LID_RAMP_FROM_DEMAND = True

DEMAND_FILE_NAME = "A-O_Demand.xlsx"
DEMAND_SHEET = "Demand_Projection"
DEMAND_REFERENCE_YEAR = 2024
DEMAND_TYPE_FILTER = "Demand"   # value of "Demand/Share" column to keep
DEMAND_FUEL_COL = "Fuel/Tech"
DEMAND_TYPE_COL = "Demand/Share"
DEMAND_CR_SLICE = slice(3, 8)   # ELCBGDXX03 -> BGDXX

# V1 untie rule: when MinCapInvestment(y) >= proposed MaxCapInv(y), push
# MaxCapInv up by this multiplier. Matches MAX_MULTIPLIER in B1b.
UNTIE_MULTIPLIER = 1.01


# ---------------------------------------------------------------------------
# Backup
# ---------------------------------------------------------------------------
def _rmtree_robust(path: Path, attempts: int = 5) -> None:
    """shutil.rmtree with retries.

    On Windows, .xlsx files saved via openpyxl are sometimes briefly held by
    the OS after `wb.save()` returns, causing `shutil.rmtree` to raise
    PermissionError [WinError 32]. Force a GC to release any lingering Python
    refs, back off, and retry. Linux/macOS hits the success path on attempt 0.
    """
    for i in range(attempts):
        try:
            shutil.rmtree(path)
            return
        except PermissionError:
            gc.collect()
            time.sleep(0.1 * (i + 1))
    # Final attempt — let it raise if the lock still hasn't released.
    shutil.rmtree(path)


def find_latest_backup(input_dir: Path) -> Path | None:
    """Return the most recent _PRE_LID_* sibling backup of input_dir, or None."""
    parent = input_dir.parent
    candidates = sorted(
        (p for p in parent.iterdir()
         if p.is_dir() and p.name.startswith(f"{input_dir.name}_PRE_LID_")),
        key=lambda p: p.name,
    )
    return candidates[-1] if candidates else None


def restore_from_backup(input_dir: Path, backup_dir: Path | None = None) -> Path:
    """Restore the input directory from a _PRE_LID_* backup.

    If `backup_dir` is None, use the most recent _PRE_LID_* sibling backup.
    Removes the current input_dir (after a safety copy in case the user wants
    to undo the restore) and replaces it with the backup contents.

    Returns the path of the backup that was used.
    """
    input_dir = Path(input_dir)
    if backup_dir is None:
        backup_dir = find_latest_backup(input_dir)
        if backup_dir is None:
            raise FileNotFoundError(
                f"No _PRE_LID_* backup found next to {input_dir}. "
                f"Pass --restore-from <folder> to specify one."
            )
    else:
        backup_dir = Path(backup_dir)
        if not backup_dir.is_dir():
            raise FileNotFoundError(f"Backup folder does not exist: {backup_dir}")

    # Safety: keep what's currently in input_dir as a "POST_LID" snapshot in
    # case the user runs --restore by mistake. Same parent, timestamped.
    if input_dir.is_dir():
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        snapshot = input_dir.parent / f"{input_dir.name}_POST_LID_pre_restore_{stamp}"
        if not snapshot.exists():
            shutil.copytree(input_dir, snapshot)
        _rmtree_robust(input_dir)
    shutil.copytree(backup_dir, input_dir)
    return backup_dir


def make_backup(input_dir: Path) -> Path:
    """Copy `input_dir` to a timestamped sibling folder and return its path."""
    if not input_dir.is_dir():
        raise FileNotFoundError(f"Input directory does not exist: {input_dir}")
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = input_dir.parent / f"{input_dir.name}_PRE_LID_{stamp}"
    if backup.exists():
        raise FileExistsError(f"Backup folder already exists: {backup}")
    shutil.copytree(input_dir, backup)
    return backup


# ---------------------------------------------------------------------------
# YAML config loading
# ---------------------------------------------------------------------------
def _parse_year_key(key) -> list[int]:
    """Accept '2030', 2030, or '2031-2040' and return the list of years it covers."""
    if isinstance(key, int):
        return [key]
    s = str(key).strip()
    if "-" in s:
        lo, hi = s.split("-", 1)
        return list(range(int(lo), int(hi) + 1))
    return [int(s)]


def load_config(yaml_path: Path) -> dict:
    """Load and validate the YAML configuration.

    All keys are optional; missing keys fall through to the module-level
    defaults. The schedule accepts both single-year keys (2030) and ranges
    (2031-2040), the latter expanded year-by-year.
    """
    cfg = _load_yaml(yaml_path)
    if cfg is None:
        cfg = {}

    out: dict = {}

    rule_mode = cfg.get("rule_mode")
    if rule_mode is not None:
        rule_mode = str(rule_mode).strip()
        if rule_mode not in ("uniform", "proportional"):
            raise ValueError(
                f"rule_mode={rule_mode!r} is not recognized in {yaml_path}. "
                f"Expected 'uniform' or 'proportional'."
            )
        out["rule_mode"] = rule_mode

    if "percentage_default" in cfg:
        out["percentage_default"] = float(cfg["percentage_default"])

    if "percentage_by_year" in cfg:
        expanded: dict = {}
        for raw_key, raw_val in (cfg["percentage_by_year"] or {}).items():
            value = float(raw_val)
            for y in _parse_year_key(raw_key):
                expanded[y] = value
        out["percentage_by_year"] = expanded

    if "security_factor" in cfg:
        out["security_factor"] = float(cfg["security_factor"])

    if "ramp_from_demand" in cfg:
        out["ramp_from_demand"] = bool(cfg["ramp_from_demand"])

    return out


def apply_config(cfg: dict) -> None:
    """Mutate the module-level lid constants in place.

    The script's helper functions reference these constants directly, so
    overriding them here propagates to every downstream call without touching
    any function signatures.
    """
    global LID_RULE_MODE, LID_PERCENTAGE_DEFAULT, LID_PERCENTAGE_BY_YEAR
    global LID_SECURITY_FACTOR, LID_RAMP_FROM_DEMAND
    global LID_EXEMPT_PREFIXES, LID_RELAXATION_SCHEDULE
    if "rule_mode" in cfg:
        LID_RULE_MODE = cfg["rule_mode"]
    if "percentage_default" in cfg:
        LID_PERCENTAGE_DEFAULT = cfg["percentage_default"]
    if "percentage_by_year" in cfg:
        LID_PERCENTAGE_BY_YEAR = cfg["percentage_by_year"]
    if "security_factor" in cfg:
        LID_SECURITY_FACTOR = cfg["security_factor"]
    if "ramp_from_demand" in cfg:
        LID_RAMP_FROM_DEMAND = cfg["ramp_from_demand"]
    if "exempt_prefixes" in cfg:
        LID_EXEMPT_PREFIXES = list(cfg["exempt_prefixes"])
    if "relaxation_schedule" in cfg:
        LID_RELAXATION_SCHEDULE = {
            int(k): float(v) for k, v in cfg["relaxation_schedule"].items()
        }


def relaxation_multiplier_for_year(year: int) -> float:
    """Return the relaxation multiplier for *year* by linearly interpolating
    LID_RELAXATION_SCHEDULE.

    If the schedule is empty, returns 1.0 (BAU behaviour).
    Years before the first anchor get the first anchor's value.
    Years after the last anchor get the last anchor's value.
    """
    if not LID_RELAXATION_SCHEDULE:
        return 1.0
    anchors = sorted(LID_RELAXATION_SCHEDULE.items())  # [(y, mult), ...]
    if year <= anchors[0][0]:
        return anchors[0][1]
    if year >= anchors[-1][0]:
        return anchors[-1][1]
    # Find the bounding pair and interpolate.
    for i in range(len(anchors) - 1):
        y0, m0 = anchors[i]
        y1, m1 = anchors[i + 1]
        if y0 <= year <= y1:
            frac = (year - y0) / (y1 - y0)
            return m0 + frac * (m1 - m0)
    return 1.0  # should not reach here


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def lid_pct_for_cr_year(cr: str, year: int,
                        demand_mult_map: dict | None = None) -> float:
    """Return the *uniform-mode* lid percentage for (country_region, year).

    This function is used by uniform mode to compute pct(cr, y). Proportional
    mode does not call this — it computes lid directly via tech_share and
    pool_delta (see proportional_lid_for_tech_year).

    Formula:
        base_pct(y) = LID_PERCENTAGE_BY_YEAR.get(y, LID_PERCENTAGE_DEFAULT)
        pct(cr, y)  = base_pct(y) * mult(cr, y)              if ramp on
                    = base_pct(y)                            if ramp off

    The base_pct schedule encodes the per-decade plateau ("tight near-term,
    loose late-horizon"). The demand multiplier layers on top, scaling pct
    linearly with each cr's demand growth from the reference year. So a
    region whose demand triples gets a lid that is 3x what the schedule
    alone would imply.

    At the reference year, mult=1 by construction, so pct equals base_pct
    regardless of the ramp setting.
    """
    if year in LID_PERCENTAGE_BY_YEAR:
        base_pct = float(LID_PERCENTAGE_BY_YEAR[year])
    else:
        base_pct = LID_PERCENTAGE_DEFAULT
    if LID_RAMP_FROM_DEMAND and demand_mult_map:
        mult = demand_mult_map.get((cr, year), 1.0)
        return base_pct * mult
    return base_pct


def lid_pct_for_year(year: int) -> float:
    """Backwards-compat shim: flat lid for `year`, ignoring cr and demand.
    Equivalent to lid_pct_for_cr_year with no ramp."""
    return float(LID_PERCENTAGE_BY_YEAR.get(year, LID_PERCENTAGE_DEFAULT))


def load_generation_techs(tech_types_path: Path) -> set:
    """Load TECH_TYPES.csv and return the set of techs in GENERATION_CATEGORY.

    Raises FileNotFoundError if the file is missing — TECH_TYPES.csv is the
    authoritative source for what counts as a generation tech, and silently
    falling back to a heuristic would be a footgun. To opt out, set
    RESTRICT_TO_GENERATION = False at the top of this script.
    """
    tech_types_path = Path(tech_types_path)
    if not tech_types_path.is_file():
        raise FileNotFoundError(
            f"TECH_TYPES.csv not found at {tech_types_path}. "
            f"Place it next to this script, or set "
            f"RESTRICT_TO_GENERATION = False to disable the filter."
        )
    df = pd.read_csv(tech_types_path)
    cat_col = TECH_TYPES_CATEGORY_COL
    tech_col = TECH_TYPES_TECH_COL
    missing = [c for c in (cat_col, tech_col) if c not in df.columns]
    if missing:
        raise ValueError(
            f"TECH_TYPES.csv missing columns {missing}. "
            f"Found {list(df.columns)}."
        )
    return set(df.loc[df[cat_col] == GENERATION_CATEGORY, tech_col].dropna())


def build_demand_multiplier_map(demand_path: Path,
                                ref_year: int = DEMAND_REFERENCE_YEAR) -> dict:
    """Read A-O_Demand.xlsx and return {(cr, year): demand(y) / demand(ref_year)}.

    Aggregates demand-type rows by country+region (chars 3..7 of the Fuel/Tech
    code, e.g. ELCBGDXX03 -> BGDXX). Returns an empty dict if the demand file
    is missing or doesn't have the expected structure — callers should treat
    an empty map as 'no ramp data, fall back to flat default'.
    """
    demand_path = Path(demand_path)
    if not demand_path.is_file():
        return {}
    try:
        dp = pd.read_excel(demand_path, sheet_name=DEMAND_SHEET)
    except Exception:
        return {}
    if DEMAND_TYPE_COL not in dp.columns or DEMAND_FUEL_COL not in dp.columns:
        return {}
    # Year headers may arrive as strings ("2024") or ints (2024) depending on
    # upstream writers. Accept either; map back to int for the output keys.
    year_to_col: dict = {}
    for c in dp.columns:
        if isinstance(c, int) and 1900 <= c <= 2200:
            year_to_col[c] = c
        elif isinstance(c, str) and c.isdigit() and 1900 <= int(c) <= 2200:
            year_to_col[int(c)] = c
    if ref_year not in year_to_col:
        return {}

    rows = dp[dp[DEMAND_TYPE_COL] == DEMAND_TYPE_FILTER].copy()
    if rows.empty:
        return {}
    rows["cr"] = rows[DEMAND_FUEL_COL].astype(str).str[DEMAND_CR_SLICE]
    orig_cols = [year_to_col[y] for y in sorted(year_to_col)]
    by_cr = rows.groupby("cr")[orig_cols].sum()

    ref_col = year_to_col[ref_year]
    out: dict = {}
    for cr, series in by_cr.iterrows():
        ref = float(series[ref_col])
        if ref <= 0:
            continue
        for y, col in year_to_col.items():
            out[(cr, y)] = float(series[col]) / ref
    return out


def values_differ(a, b, tol: float = 1e-12) -> bool:
    """Return True if `a` and `b` should be considered different cell values."""
    if a is None and b is None:
        return False
    if a is None or b is None:
        return True
    try:
        return abs(float(a) - float(b)) > tol
    except (TypeError, ValueError):
        return a != b


def find_year_columns(ws) -> dict:
    """Scan row 1 for integer year headers; return {year: column_index_1based}."""
    year_to_col = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if isinstance(val, int) and 1900 <= val <= 2200:
            year_to_col[val] = col_idx
    return year_to_col


def find_named_columns(ws, names) -> dict:
    """Return {name: column_index_1based} for headers matching `names`."""
    found = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val in names:
            found[val] = col_idx
    return found


def identify_allowed_techs(df: pd.DataFrame, year_cols: list,
                           generation_techs: set | None = None) -> set:
    """Return the set of techs eligible for the lid + untie rule.

    Eligibility = (ResidualCapacity > 0 OR MinCapInvestment > 0 in any year)
    AND (in `generation_techs` if RESTRICT_TO_GENERATION and the set is given).

    `generation_techs` is the set returned by load_generation_techs(). When
    None, no category filter is applied — useful for tests of the unfiltered
    behavior. Production callers should always pass it.
    """
    res = df[df["Parameter"] == RES_PARAM]
    mci = df[df["Parameter"] == MIN_INV_PARAM]
    res_max = res.set_index("Tech")[year_cols].fillna(0).max(axis=1)
    mci_max = mci.set_index("Tech")[year_cols].fillna(0).max(axis=1)
    allowed = set(res_max[res_max > 0].index) | set(mci_max[mci_max > 0].index)
    if generation_techs is not None and RESTRICT_TO_GENERATION:
        allowed = allowed & generation_techs
    return allowed


def country_region_for(tech: str) -> str | None:
    """Extract country+region code from a PWR* tech name.

    Returns e.g. 'BGDXX' for PWRHYDBGDXX, 'INDNE' for PWRHYDINDNE.
    Returns None for tech codes that don't fit the PWR* length-11 convention
    (e.g. TRN* transmission interconnects of length 13) — those should be
    excluded from pool computation and pool-based lid application.
    """
    if not isinstance(tech, str) or len(tech) != PWR_TECH_LENGTH:
        return None
    return tech[COUNTRY_REGION_SLICE]


def build_pool_map(df: pd.DataFrame, allowed: set, year_cols: list) -> dict:
    """Return {(country_region, year): pool_total} where pool_total is the
    sum of ResidualCapacity across all ALLOWED PWR* techs in that
    country+region for that year. Non-PWR* techs (TRN*, etc.) and non-allowed
    techs do not contribute to the pool.
    """
    res = df[(df["Parameter"] == RES_PARAM) & (df["Tech"].isin(allowed))].copy()
    res["cr"] = res["Tech"].apply(country_region_for)
    # Drop techs whose country_region couldn't be parsed (e.g. TRN*)
    res = res[res["cr"].notna()]
    pool_map: dict = {}
    for cr, sub in res.groupby("cr"):
        for y in year_cols:
            pool_map[(cr, y)] = float(sub[y].fillna(0).sum())
    return pool_map


def build_mininv_map(df: pd.DataFrame, year_cols: list) -> dict:
    """Return {(tech, year): min_inv} lookup, NaN normalized to 0.0."""
    mci = df[df["Parameter"] == MIN_INV_PARAM]
    mininv_map: dict = {}
    for _, row in mci.iterrows():
        tech = row["Tech"]
        for y in year_cols:
            v = row[y]
            mininv_map[(tech, y)] = 0.0 if pd.isna(v) else float(v)
    return mininv_map


def build_tech_share_map(df: pd.DataFrame, allowed: set,
                         ref_year: int = DEMAND_REFERENCE_YEAR) -> dict:
    """Return {tech: share} where share = ResidualCapacity(t, ref_year)
    / pool(cr(t), ref_year). Used by proportional mode to distribute
    pool growth among allowed gen techs in a cr.

    Shares for techs in the same cr sum to 1.0 (modulo float). Techs with
    zero residual at ref_year get share=0 — they receive no proportional
    allocation, only whatever the untie rule provides via MinCapInv.
    """
    res = df[(df["Parameter"] == RES_PARAM) & (df["Tech"].isin(allowed))].copy()
    res["cr"] = res["Tech"].apply(country_region_for)
    res = res[res["cr"].notna()]
    # Per-cr ref-year totals
    cr_totals: dict = {}
    for cr, sub in res.groupby("cr"):
        cr_totals[cr] = float(sub[ref_year].fillna(0).sum())
    # Per-tech share
    share_map: dict = {}
    for _, row in res.iterrows():
        tech = row["Tech"]
        cr = row["cr"]
        cr_total = cr_totals.get(cr, 0.0)
        ref_val = float(row[ref_year]) if pd.notna(row[ref_year]) else 0.0
        share_map[tech] = (ref_val / cr_total) if cr_total > 0 else 0.0
    return share_map


def build_scaled_pool_map(pool_map: dict,
                          demand_mult_map: dict | None) -> dict:
    """Return {(cr, year): scaled_pool} where scaled_pool = pool * mult.

    If demand_mult_map is None or LID_RAMP_FROM_DEMAND is False, mult collapses
    to 1.0 and scaled_pool == pool (consistent with how lid_pct_for_cr_year
    treats the ramp-off case).
    """
    use_mult = LID_RAMP_FROM_DEMAND and demand_mult_map is not None
    out: dict = {}
    for (cr, y), pool in pool_map.items():
        if use_mult:
            mult = demand_mult_map.get((cr, y), 1.0)
        else:
            mult = 1.0
        out[(cr, y)] = pool * mult
    return out


def build_pool_delta_map(scaled_pool_map: dict, year_cols: list) -> dict:
    """Return {(cr, year): pool_delta} where pool_delta = max(0,
    scaled_pool(y) - scaled_pool(y-1)). For the earliest year in year_cols,
    delta = 0 (no prior year to delta against).

    The max(0, ...) guard ensures negative or flat demand growth in any year
    yields zero new headroom rather than a negative lid. Per project hygiene:
    we don't model demand dips, but if a year happens to be flat or slightly
    declining due to projection methodology, this prevents propagation of
    nonsense values into the LP.
    """
    sorted_years = sorted(year_cols)
    crs = sorted({cr for (cr, _) in scaled_pool_map.keys()})
    out: dict = {}
    for cr in crs:
        for i, y in enumerate(sorted_years):
            if i == 0:
                out[(cr, y)] = 0.0
                continue
            prev_y = sorted_years[i - 1]
            cur = scaled_pool_map.get((cr, y), 0.0)
            prev = scaled_pool_map.get((cr, prev_y), 0.0)
            delta = cur - prev
            out[(cr, y)] = delta if delta > 0 else 0.0
    return out


def proportional_lid_for_tech_year(tech: str, year: int,
                                   tech_share_map: dict,
                                   pool_delta_map: dict) -> float:
    """Compute proportional-mode lid for (tech, year):
        lid = LID_SECURITY_FACTOR * tech_share(t) * pool_delta(cr, y)

    Returns 0.0 if cr cannot be parsed from tech (TRN* etc.) or if either
    share or delta is missing/zero.
    """
    cr = country_region_for(tech)
    if cr is None:
        return 0.0
    share = tech_share_map.get(tech, 0.0)
    delta = pool_delta_map.get((cr, year), 0.0)
    return LID_SECURITY_FACTOR * share * delta


# ---------------------------------------------------------------------------
# Core logic
# ---------------------------------------------------------------------------
def apply_lid_to_sheet(ws, allowed: set, pool_map: dict,
                       mininv_map: dict,
                       demand_mult_map: dict | None = None,
                       tech_share_map: dict | None = None,
                       pool_delta_map: dict | None = None) -> dict:
    """
    Edit a worksheet in place, applying the lid + untie rule to the
    TotalAnnualMaxCapacityInvestment row of every PWR* tech in `allowed`.

    The lid formula depends on LID_RULE_MODE:
      "uniform"      - lid(t, y) = pct(cr, y) * pool(cr, y)
                       where pct(cr, y) = base_pct(y) * mult(cr, y).
                       Same lid for every allowed tech in cr per year.
      "proportional" - lid(t, y) = LID_SECURITY_FACTOR * tech_share(t)
                                   * pool_delta(cr, y).
                       Per-tech, weighted by 2024 fleet share.

    For proportional mode, callers must pass tech_share_map and
    pool_delta_map; for uniform mode they're ignored.

    Techs not in `allowed`, and TRN* transmission interconnects (whose names
    don't fit the PWR* country+region format), are entirely untouched.
    """
    if LID_RULE_MODE not in ("uniform", "proportional"):
        raise ValueError(
            f"LID_RULE_MODE={LID_RULE_MODE!r} is not recognized. "
            f"Expected 'uniform' or 'proportional'."
        )
    if LID_RULE_MODE == "proportional" and (
        tech_share_map is None or pool_delta_map is None
    ):
        raise ValueError(
            "proportional mode requires tech_share_map and pool_delta_map; "
            "got None. The caller (run) should build these via "
            "build_tech_share_map and build_pool_delta_map."
        )

    year_cols = find_year_columns(ws)
    headers = find_named_columns(ws, ["Tech", "Parameter", PROJ_MODE_COL])
    if "Tech" not in headers or "Parameter" not in headers:
        raise ValueError(
            f"Sheet '{ws.title}' missing required columns: "
            f"found {list(headers.keys())}"
        )
    tech_col = headers["Tech"]
    param_col = headers["Parameter"]
    proj_mode_col = headers.get(PROJ_MODE_COL)

    # Snapshot the uniform-mode lid pct per (cr, year). Always logged for
    # traceability — under proportional mode this is informational only.
    pct_used = {
        f"{cr}_{y}": lid_pct_for_cr_year(cr, y, demand_mult_map)
        for (cr, y) in pool_map.keys()
    }

    log = {
        "sheet": ws.title,
        "years_found": sorted(year_cols.keys()),
        "allowed_count": len(allowed),
        "rule_mode": LID_RULE_MODE,
        "changes": [],
        "preserved": [],
        "projection_mode_flips": [],
        "skipped_non_pwr_techs": [],
        "pool_map": {  # serializable version: {"BGDXX_2030": 7.123, ...}
            f"{cr}_{y}": v for (cr, y), v in pool_map.items()
        },
        "lid_pct_by_cr_year": pct_used,
        "demand_mult_by_cr_year": (
            {f"{cr}_{y}": v for (cr, y), v in demand_mult_map.items()}
            if demand_mult_map else {}
        ),
        "tech_share": (
            {t: s for t, s in tech_share_map.items()}
            if tech_share_map else {}
        ),
        "pool_delta_by_cr_year": (
            {f"{cr}_{y}": v for (cr, y), v in pool_delta_map.items()}
            if pool_delta_map else {}
        ),
        "security_factor": (
            LID_SECURITY_FACTOR if LID_RULE_MODE == "proportional" else None
        ),
    }

    for row_idx in range(2, ws.max_row + 1):
        tech = ws.cell(row=row_idx, column=tech_col).value
        param = ws.cell(row=row_idx, column=param_col).value
        if tech is None or tech not in allowed or param != MAX_INV_PARAM:
            continue

        cr = country_region_for(tech)
        if cr is None:
            # TRN* and other non-PWR* techs: skip — leave manual cal alone.
            if tech not in log["skipped_non_pwr_techs"]:
                log["skipped_non_pwr_techs"].append(tech)
            continue

        row_was_modified = False

        for year, col in year_cols.items():
            cell = ws.cell(row=row_idx, column=col)
            old = cell.value
            pool = pool_map.get((cr, year), 0.0)
            min_inv = mininv_map.get((tech, year), 0.0)

            # Compute lid per the active mode.
            # --- Exempt prefixes: uncap techs matching any prefix in the list.
            is_exempt = any(tech.startswith(p) for p in LID_EXEMPT_PREFIXES)
            if is_exempt:
                lid = float(PLACEHOLDER_VALUE)
            elif LID_RULE_MODE == "uniform":
                pct = lid_pct_for_cr_year(cr, year, demand_mult_map)
                # Layer the demand multiplier onto pct * pool. Note that
                # lid_pct_for_cr_year already applies mult once, so the
                # final formula is lid = base_pct(y) * pool * mult(cr, y).
                lid = pct * pool
            else:  # "proportional"
                lid = proportional_lid_for_tech_year(
                    tech, year, tech_share_map, pool_delta_map
                )

            # --- Relaxation schedule: multiply the base lid by a year-varying
            # factor.  Has no effect when the schedule is empty (mult = 1.0)
            # or when the tech is already exempt (lid = 9999).
            relax_mult = relaxation_multiplier_for_year(year)
            if not is_exempt and relax_mult != 1.0:
                lid = min(lid * relax_mult, float(PLACEHOLDER_VALUE))

            # Decide the proposed value (before untie):
            #   placeholder cell -> lid; manual value -> preserve.
            is_placeholder = (
                old is None
                or (isinstance(old, (int, float))
                    and not pd.isna(old)
                    and float(old) == float(PLACEHOLDER_VALUE))
            )
            if is_placeholder:
                proposed = lid
                if is_exempt:
                    reason = "exempt_uncapped"
                elif relax_mult != 1.0:
                    reason = "lid_relaxed"
                else:
                    reason = "lid_fill"
            else:
                proposed = float(old) if isinstance(old, (int, float)) else old
                reason = "preserved_manual"

            # V1 untie rule: ensure proposed > min_inv.
            if min_inv > 0 and (proposed is None or proposed <= min_inv):
                proposed = min_inv * UNTIE_MULTIPLIER
                reason = "untie_min_inv"

            if values_differ(old, proposed):
                cell.value = proposed
                row_was_modified = True
                log["changes"].append(
                    {
                        "tech": tech,
                        "country_region": cr,
                        "year": year,
                        "old": old,
                        "new": proposed,
                        "reason": reason,
                        "pool": pool,
                        "min_inv": min_inv,
                        "lid": lid,
                        "relax_mult": relax_mult,
                    }
                )
            else:
                log["preserved"].append(
                    {"tech": tech, "year": year, "value": old}
                )

        # Flip Projection.Mode for any row we modified, EMPTY -> User defined.
        if row_was_modified and proj_mode_col is not None:
            mode_cell = ws.cell(row=row_idx, column=proj_mode_col)
            if mode_cell.value == PROJ_MODE_EMPTY:
                mode_cell.value = PROJ_MODE_USER
                log["projection_mode_flips"].append({"tech": tech})

    return log


def edit_parametrization(filepath: Path, sheets: list,
                         generation_techs: set | None = None,
                         demand_mult_map: dict | None = None) -> dict:
    """Apply the lid + untie rule to `sheets` in the parametrization workbook."""
    df_all = pd.read_excel(filepath, sheet_name=None)
    wb = load_workbook(filepath)

    file_log = {"file": str(filepath), "sheets": []}

    try:
        for sheet in sheets:
            if sheet not in wb.sheetnames:
                file_log["sheets"].append(
                    {"sheet": sheet, "skipped": "sheet not present in workbook"}
                )
                continue

            df = df_all[sheet]
            year_cols = [c for c in df.columns if isinstance(c, int)]
            if not year_cols:
                file_log["sheets"].append(
                    {"sheet": sheet, "skipped": "no integer year columns found"}
                )
                continue

            allowed = identify_allowed_techs(df, year_cols, generation_techs)
            pool_map = build_pool_map(df, allowed, year_cols)
            mininv_map = build_mininv_map(df, year_cols)

            # Proportional-mode auxiliary maps. Cheap to build, computed
            # regardless of mode so the log captures them either way.
            tech_share_map = build_tech_share_map(df, allowed)
            scaled_pool_map = build_scaled_pool_map(pool_map, demand_mult_map)
            pool_delta_map = build_pool_delta_map(scaled_pool_map, year_cols)

            ws = wb[sheet]
            sheet_log = apply_lid_to_sheet(
                ws, allowed, pool_map, mininv_map, demand_mult_map,
                tech_share_map=tech_share_map,
                pool_delta_map=pool_delta_map,
            )
            sheet_log["allowed_techs"] = sorted(allowed)
            file_log["sheets"].append(sheet_log)

        wb.save(filepath)
    finally:
        # Explicitly release Windows file handles so a subsequent
        # shutil.rmtree (e.g. in restore_from_backup or pytest tmp_path
        # cleanup) doesn't hit PermissionError [WinError 32].
        wb.close()
    return file_log


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------
def run(input_dir, sheets: list = None, skip_backup: bool = False,
        yaml_path: Path | None = None) -> dict:
    """End-to-end: backup, edit, write log. Returns the log dict.

    When `yaml_path` is None, the script looks for YAML_FILE_NAME next to
    itself (where the orchestrator stages the per-scenario override). If
    found, its values replace the module-level defaults for this process.
    """
    input_dir = Path(input_dir)
    sheets = sheets or DEFAULT_TARGET_SHEETS

    # Optional YAML override (per-scenario). Located next to this script —
    # the orchestrator stages it there from rules_scripts/configs/<scenario>/.
    if yaml_path is None:
        yaml_path = Path(__file__).resolve().parent / YAML_FILE_NAME
    yaml_loaded = False
    if yaml_path.is_file():
        apply_config(load_config(yaml_path))
        yaml_loaded = True

    backup_dir = None if skip_backup else make_backup(input_dir)

    paramfile = input_dir / PARAM_FILE_NAME
    if not paramfile.exists():
        raise FileNotFoundError(f"{paramfile} not found")

    # Load the GENERATION tech list. TECH_TYPES.csv is shared by other A3
    # stages (e.g. patch_ao_c2a.py), so it lives in A3_process/, one level
    # above this script (which now lives in A3_process/rules_scripts/).
    generation_techs = None
    tech_types_path = None
    if RESTRICT_TO_GENERATION:
        script_dir = Path(__file__).resolve().parent
        tech_types_path = script_dir.parent / TECH_TYPES_FILE
        generation_techs = load_generation_techs(tech_types_path)

    # Load the per-cr demand multipliers (from the input dir).
    demand_mult_map: dict = {}
    demand_path = input_dir / DEMAND_FILE_NAME
    if LID_RAMP_FROM_DEMAND:
        demand_mult_map = build_demand_multiplier_map(
            demand_path, ref_year=DEMAND_REFERENCE_YEAR
        )

    log = edit_parametrization(
        paramfile, sheets, generation_techs, demand_mult_map
    )
    log["backup_dir"] = str(backup_dir) if backup_dir else None
    log["timestamp"] = datetime.now().isoformat()
    log["lid_percentage_default"] = LID_PERCENTAGE_DEFAULT
    log["lid_percentage_by_year"] = {
        str(k): v for k, v in LID_PERCENTAGE_BY_YEAR.items()
    }
    log["lid_ramp_from_demand"] = LID_RAMP_FROM_DEMAND
    log["lid_rule_mode"] = LID_RULE_MODE
    log["lid_security_factor"] = (
        LID_SECURITY_FACTOR if LID_RULE_MODE == "proportional" else None
    )
    log["restrict_to_generation"] = RESTRICT_TO_GENERATION
    log["exempt_prefixes"] = list(LID_EXEMPT_PREFIXES)
    log["relaxation_schedule"] = dict(LID_RELAXATION_SCHEDULE)
    log["generation_techs_count"] = (
        len(generation_techs) if generation_techs is not None else None
    )
    log["tech_types_file"] = str(tech_types_path) if tech_types_path else None
    log["demand_file"] = str(demand_path) if LID_RAMP_FROM_DEMAND else None
    log["demand_reference_year"] = (
        DEMAND_REFERENCE_YEAR if LID_RAMP_FROM_DEMAND else None
    )
    log["demand_mult_loaded"] = bool(demand_mult_map)
    log["yaml_config_path"] = str(yaml_path) if yaml_loaded else None

    if backup_dir is not None:
        log_path = backup_dir.parent / f"{backup_dir.name}_CHANGES.json"
        log_path.write_text(json.dumps(log, indent=2, default=str))
        log["log_path"] = str(log_path)

    return log


def print_summary(log: dict) -> None:
    """Pretty-print the run summary."""
    bar = "=" * 72
    print(bar)
    print("MaxCapacityInvestment lid + untie rule — applied")
    print(bar)
    print(f"Backup folder : {log.get('backup_dir', '(skipped)')}")
    print(f"Edited file   : {log['file']}")
    print(f"Rule mode     : {LID_RULE_MODE}"
          + (f"  (security factor = {LID_SECURITY_FACTOR})"
             if LID_RULE_MODE == "proportional" else ""))
    print(f"Lid base pct  : {LID_PERCENTAGE_DEFAULT * 100:.3f}% "
          f"(anchored at {DEMAND_REFERENCE_YEAR})")
    print(f"Demand ramp   : {'ON' if log.get('lid_ramp_from_demand') else 'OFF'}"
          f"{' (per country+region)' if log.get('demand_mult_loaded') else ''}")
    if LID_EXEMPT_PREFIXES:
        print(f"Exempt (9999) : {', '.join(LID_EXEMPT_PREFIXES)}")
    if LID_RELAXATION_SCHEDULE:
        anchors = sorted(LID_RELAXATION_SCHEDULE.items())
        sched = ", ".join(f"{y}:{m:.2f}×" for y, m in anchors)
        print(f"Relaxation    : {sched}")
    print(f"GEN-only      : {'ON' if log.get('restrict_to_generation') else 'OFF'}"
          f" ({log.get('generation_techs_count')} GENERATION techs"
          f" loaded from {log.get('tech_types_file')})"
          if log.get('restrict_to_generation') else "")
    if LID_PERCENTAGE_BY_YEAR and LID_RULE_MODE == "uniform":
        # Collapse contiguous-equal pct runs for readable display.
        items = sorted(LID_PERCENTAGE_BY_YEAR.items())
        groups = []
        cur_start, cur_pct = items[0]
        cur_end = cur_start
        for y, p in items[1:]:
            if p == cur_pct and y == cur_end + 1:
                cur_end = y
            else:
                groups.append((cur_start, cur_end, cur_pct))
                cur_start, cur_end, cur_pct = y, y, p
        groups.append((cur_start, cur_end, cur_pct))
        sched = ", ".join(
            f"{a}-{b}:{p*100:g}%" if a != b else f"{a}:{p*100:g}%"
            for a, b, p in groups
        )
        print(f"Year schedule : {sched}")
    print()
    for s in log["sheets"]:
        if "skipped" in s:
            print(f"[SKIPPED] '{s['sheet']}': {s['skipped']}")
            continue
        years = s["years_found"]
        print(f"Sheet: '{s['sheet']}'")
        print(f"  Years          : {years[0]}..{years[-1]} ({len(years)} years)")
        print(f"  ALLOWED techs  : {s['allowed_count']}")
        print(f"  Skipped non-PWR techs (e.g. TRN*) : "
              f"{len(s.get('skipped_non_pwr_techs', []))}")
        # Show the country+region pools at first and last year
        pool_keys = sorted(set(k.rsplit("_", 1)[0] for k in s["pool_map"].keys()))
        if pool_keys and years:
            print(f"  Country+region pools: {len(pool_keys)} "
                  f"({', '.join(pool_keys[:6])}{' ...' if len(pool_keys) > 6 else ''})")
        # Show the spread of pct used at first vs last year
        pct_map = s.get("lid_pct_by_cr_year", {})
        if pct_map and pool_keys:
            y_first, y_last = years[0], years[-1]
            sample_cr = pool_keys[0]
            pct_first = pct_map.get(f"{sample_cr}_{y_first}")
            pct_last = pct_map.get(f"{sample_cr}_{y_last}")
            if pct_first is not None and pct_last is not None:
                print(f"  Lid pct ({sample_cr}): "
                      f"{pct_first*100:.3f}% in {y_first} -> "
                      f"{pct_last*100:.3f}% in {y_last}")
        from collections import Counter
        reason_counts = Counter(c.get("reason", "?") for c in s["changes"])
        n_lid = reason_counts.get("lid_fill", 0)
        n_relaxed = reason_counts.get("lid_relaxed", 0)
        n_untie = reason_counts.get("untie_min_inv", 0)
        n_other = sum(reason_counts.values()) - n_lid - n_relaxed - n_untie
        print(f"  MaxInv cells written:")
        print(f"    - filled with lid (pct * pool)    : {n_lid}")
        if n_relaxed:
            print(f"    - filled with relaxed lid          : {n_relaxed}")
        print(f"    - bumped by untie rule (>= MinInv) : {n_untie}")
        if n_other:
            print(f"    - other                            : {n_other}")
        print(f"  Manual values preserved              : {len(s['preserved'])}")
        print(f"  Projection.Mode flips (EMPTY -> User defined) : "
              f"{len(s['projection_mode_flips'])}")
    if log.get("log_path"):
        print(f"\nDetailed change log written to: {log['log_path']}")


def main() -> int:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=Path("A1_Outputs/A1_Outputs_BAU"),
        help="Directory containing the AO files (default: A1_Outputs/A1_Outputs_BAU)",
    )
    parser.add_argument(
        "--sheets",
        nargs="+",
        default=DEFAULT_TARGET_SHEETS,
        help=f"Sheets to apply the rule to (default: {DEFAULT_TARGET_SHEETS})",
    )
    parser.add_argument(
        "--skip-backup",
        action="store_true",
        help="Skip backup creation (DANGEROUS — for testing only)",
    )
    parser.add_argument(
        "--restore",
        action="store_true",
        help="Restore input dir from the most recent _PRE_LID_* backup, then exit. "
             "Saves a snapshot of the current input dir as _POST_LID_pre_restore_<ts>/ "
             "before overwriting, so the restore itself is reversible.",
    )
    parser.add_argument(
        "--restore-from",
        type=Path,
        default=None,
        help="Restore input dir from this specific backup folder, then exit.",
    )
    parser.add_argument(
        "--yaml",
        type=Path,
        default=None,
        help=f"Override YAML config path (default: {YAML_FILE_NAME} next to this script).",
    )
    args = parser.parse_args()

    # Restore-only paths: do nothing else.
    if args.restore or args.restore_from is not None:
        try:
            used = restore_from_backup(args.input_dir, args.restore_from)
        except Exception as exc:
            print(f"ERROR: {exc}", file=sys.stderr)
            return 1
        print(f"Restored {args.input_dir} from {used}")
        return 0

    try:
        log = run(args.input_dir, args.sheets, args.skip_backup, args.yaml)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    print_summary(log)
    return 0


if __name__ == "__main__":
    sys.exit(main())
