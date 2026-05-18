# -*- coding: utf-8 -*-
"""
_scenarios.py — multi-scenario reader/writer for SOASIA Template v18.

This module isolates all scenario-aware logic so the rest of the A3 pipeline
(stage-1 timeslice merge, ao extensions, lid rules, etc.) keeps consuming a
flat v17-shaped template per scenario, unchanged.

Public surface:

    read_control_sheet(soasia_path)
        -> list[ScenarioConfig]

    materialize_scenario_template(soasia_path, scenario, out_path)
        Write a v17-shaped xlsx for `scenario` (BAU base + scenario overrides),
        WITHOUT the scenario column, Control sheet or Restrictions sheet.

    read_restrictions(soasia_path, scenarios)
        -> dict[(source_sheet, tech, parameter, year)] -> value
        Resolves multi-source inheritance by list order (last wins).

    apply_restrictions(parametrization_xlsx, restrictions)
        Write restriction cells into A-O_Parametrization.xlsx.

    persist_run_restrictions(soasia_path, scenario, changes_json_path, ...)
        After a rules_script run, parse its *_CHANGES.json and rewrite the
        Restrictions rows for that scenario (clear-and-write atomic).

    available_rules_scripts(a3_process_dir=None)
        -> list[str]
        Filenames of .py scripts under rules_scripts/, for validation against
        the rules_script column in Control.
"""

from __future__ import annotations

import json
import shutil
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_SOASIA = SCRIPT_DIR / "SOASIA_OSeMOSYS_Template_v18.xlsx"
RULES_SCRIPTS_DIR = SCRIPT_DIR / "rules_scripts"

CONTROL_SHEET = "Control"
RESTRICTIONS_SHEET = "Restrictions"

BAU_SCENARIO = "BAU"

CONTROL_HEADERS = [
    "scenario",
    "active",
    "rules_script",
    "inherit_restrictions_from",
    "notes",
]

RESTRICTIONS_HEADERS = [
    "scenario",
    "source_sheet",
    "tech",
    "parameter",
    "year",
    "value",
    "rule_applied",
    "source_run_timestamp",
]

# Parametric sheets: scenario column is the first column; rows can be BAU base
# or scenario overrides keyed by the columns below.
PARAMETRIC_SHEETS = [
    "Fixed_Horizon_Parameters",
    "Primary_Techs",
    "Secondary_Techs",
    "Capacities_CF",
    "VariableCost",
    "Demand_Projection",
    "Demand_Profiles",
    "Demand_Techs",
    "Emissions",
    "Interconnectors",
    "Interconnector_Params",
    "Existing_Generation",
    "Planned_Generation",
    "Technology_Costs",
    "RE_Targets_Policies",
]

# Sheets copied verbatim into the materialized template (no scenario filtering).
PASS_THROUGH_SHEETS = ["README", "Yearsplit_Template", "DaySplit"]

# Per-sheet identity keys for the BAU + overrides merge.
# These define when a scenario row "overrides" a BAU row.
# Each key tuple uniquely identifies a row within a single scenario in v17
# (validated: zero duplicates across all 15 parametric sheets). A scenario
# override row whose key matches a BAU row replaces the BAU row; an override
# row whose key is new is appended at the end.
#
# Notes on non-obvious choices:
#   - We prefer the textual `Parameter` over `Parameter.ID` because some rows
#     leave Parameter.ID blank, collapsing keys to duplicates.
#   - Capacities_CF needs `Tech.ID` (not just `Tech`) because a single tech
#     code like `PWRHYDNPLXX` covers 3 distinct units (Reservoir / Run-of-River
#     / Pumped Storage) with different Tech.IDs and different CF profiles.
#   - Existing/Planned_Generation include Commissioning_Year / Expected_COD
#     and Capacity_MW because real plant lists in v17 contain reused names for
#     different units (e.g. "Wind (various)" x10 in Sri Lanka planned).
SCENARIO_OVERRIDE_KEYS: dict[str, list[str]] = {
    "Fixed_Horizon_Parameters": ["Tech", "Parameter"],
    "Primary_Techs": ["Tech", "Parameter"],
    "Secondary_Techs": ["Tech", "Parameter"],
    "Capacities_CF": ["Timeslices", "Tech.ID", "Parameter"],
    "VariableCost": ["Mode.Operation", "Tech", "Parameter"],
    "Demand_Projection": ["Fuel/Tech"],
    "Demand_Profiles": ["Timeslices", "Fuel/Tech"],
    "Demand_Techs": ["Tech", "Parameter"],
    "Emissions": ["Tech", "Parameter"],
    "Interconnectors": ["NO"],
    "Interconnector_Params": ["Tech", "Parameter"],
    "Existing_Generation": ["Country", "Node", "Plant_Name", "Commissioning_Year", "Capacity_MW"],
    "Planned_Generation": ["Country", "Node", "Project_Name", "Expected_COD", "Capacity_MW"],
    "Technology_Costs": ["Technology_Code", "Parameter"],
    "RE_Targets_Policies": ["Country", "Policy_Name"],
}

# Final sheet order in the materialized template (matches v17).
MATERIALIZED_SHEET_ORDER = [
    "README",
    "Fixed_Horizon_Parameters",
    "Primary_Techs",
    "Secondary_Techs",
    "Capacities_CF",
    "VariableCost",
    "Demand_Projection",
    "Demand_Profiles",
    "Demand_Techs",
    "Emissions",
    "Yearsplit_Template",
    "DaySplit",
    "Interconnectors",
    "Interconnector_Params",
    "Existing_Generation",
    "Planned_Generation",
    "Technology_Costs",
    "RE_Targets_Policies",
]

# When persisting rules_script output to Restrictions, this is the parameter
# the script edits today. Future rules_scripts targeting other parameters
# would need to declare them in their JSON output and we'd extend the parser.
DEFAULT_RESTRICTION_PARAMETER = "TotalAnnualMaxCapacityInvestment"
DEFAULT_RESTRICTION_SOURCE_SHEET = "Secondary Techs"


# =============================================================================
# Data class
# =============================================================================

@dataclass
class ScenarioConfig:
    scenario: str
    active: bool
    rules_script: str | None
    inherit_restrictions_from: list[str] = field(default_factory=list)
    notes: str | None = None


# =============================================================================
# Helpers
# =============================================================================

def _to_bool(value: Any) -> bool:
    """Coerce Excel-style cell values to bool. TRUE/FALSE strings and booleans."""
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    s = str(value).strip().upper()
    return s in ("TRUE", "T", "1", "YES", "Y")


def _parse_inherit_list(value: Any) -> list[str]:
    """Parse 'BAU, NDC' into ['BAU', 'NDC']. None/empty -> []."""
    if value is None:
        return []
    parts = [p.strip() for p in str(value).split(",")]
    return [p for p in parts if p]


def _read_header(ws: Worksheet) -> list[str]:
    """Return row-1 header values as strings (None -> '')."""
    return [
        ("" if c.value is None else str(c.value))
        for c in next(ws.iter_rows(min_row=1, max_row=1))
    ]


# =============================================================================
# Control sheet
# =============================================================================

def read_control_sheet(
    soasia_path: Path | str = DEFAULT_SOASIA,
) -> list[ScenarioConfig]:
    """Read the Control sheet and return a validated list of ScenarioConfig.

    Validations:
      - Scenario names are unique.
      - Inherited names exist as other scenarios in the same Control sheet.
      - rules_script values, if present, reference an existing .py file under
        rules_scripts/ (skipped if the directory is absent, e.g. early bring-up).
      - At least one scenario named BAU exists.
    """
    path = Path(soasia_path)
    if not path.exists():
        raise FileNotFoundError(f"SOASIA workbook not found: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    if CONTROL_SHEET not in wb.sheetnames:
        raise ValueError(
            f"Workbook {path.name} has no '{CONTROL_SHEET}' sheet. "
            f"Was it built with _build_v18_from_v17.py?"
        )
    ws = wb[CONTROL_SHEET]

    headers = _read_header(ws)
    expected = set(CONTROL_HEADERS)
    if not expected.issubset(headers):
        missing = expected - set(headers)
        raise ValueError(f"Control sheet missing columns: {missing}")
    col_idx = {h: i for i, h in enumerate(headers)}

    configs: list[ScenarioConfig] = []
    seen: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        # Skip entirely blank rows
        if all(v is None or v == "" for v in row):
            continue
        scenario = row[col_idx["scenario"]]
        if scenario is None or str(scenario).strip() == "":
            continue
        scenario = str(scenario).strip()
        if scenario in seen:
            raise ValueError(
                f"Duplicate scenario '{scenario}' in Control sheet."
            )
        seen.add(scenario)
        configs.append(
            ScenarioConfig(
                scenario=scenario,
                active=_to_bool(row[col_idx["active"]]),
                rules_script=(
                    str(row[col_idx["rules_script"]]).strip()
                    if row[col_idx["rules_script"]] not in (None, "")
                    else None
                ),
                inherit_restrictions_from=_parse_inherit_list(
                    row[col_idx["inherit_restrictions_from"]]
                ),
                notes=(
                    str(row[col_idx["notes"]])
                    if row[col_idx["notes"]] is not None
                    else None
                ),
            )
        )

    wb.close()

    if BAU_SCENARIO not in seen:
        raise ValueError(
            f"Control sheet must contain a scenario named '{BAU_SCENARIO}'."
        )

    # Validate inherit references
    for cfg in configs:
        for src in cfg.inherit_restrictions_from:
            if src not in seen:
                raise ValueError(
                    f"Scenario '{cfg.scenario}' inherits from unknown "
                    f"scenario '{src}'. Defined scenarios: {sorted(seen)}"
                )

    # Validate rules_script (only if rules_scripts dir exists)
    if RULES_SCRIPTS_DIR.is_dir():
        available = {p.name for p in RULES_SCRIPTS_DIR.glob("*.py")
                     if not p.name.startswith("_")}
        for cfg in configs:
            if cfg.rules_script and cfg.rules_script not in available:
                raise ValueError(
                    f"Scenario '{cfg.scenario}' references rules_script "
                    f"'{cfg.rules_script}' which is not present in "
                    f"{RULES_SCRIPTS_DIR}. Available: {sorted(available)}"
                )

    return configs


# =============================================================================
# Materialize a scenario template (BAU base + overrides, no scenario column)
# =============================================================================

def materialize_scenario_template(
    soasia_path: Path | str,
    scenario: str,
    out_path: Path | str,
) -> Path:
    """Write a v17-shaped xlsx with the effective data for `scenario`.

    For each parametric sheet, BAU rows form the base. Rows tagged with
    `scenario` override BAU rows that share the same identity key
    (SCENARIO_OVERRIDE_KEYS), and new keys are appended.

    The output:
      - has no 'scenario' column,
      - has no Control or Restrictions sheets,
      - keeps README / Yearsplit_Template / DaySplit verbatim,
      - keeps the v17 sheet order.

    Returns the path to the materialized file.
    """
    src = Path(soasia_path)
    dst = Path(out_path)
    if not src.exists():
        raise FileNotFoundError(f"SOASIA workbook not found: {src}")

    # Copy first so READMEs / Yearsplit / DaySplit are preserved cell-for-cell.
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(src, dst)

    wb = load_workbook(dst)

    # Drop multi-scenario-only sheets
    for s in (CONTROL_SHEET, RESTRICTIONS_SHEET):
        if s in wb.sheetnames:
            del wb[s]

    for sheet_name in PARAMETRIC_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        _materialize_parametric_sheet(wb[sheet_name], scenario, sheet_name)

    # Reorder sheets to match v17
    name_to_sheet = {s.title: s for s in wb._sheets}
    final = [name_to_sheet[n] for n in MATERIALIZED_SHEET_ORDER if n in name_to_sheet]
    # Append any other sheets at the end (audit / extras), preserving their order
    others = [s for s in wb._sheets if s.title not in MATERIALIZED_SHEET_ORDER]
    wb._sheets = final + others

    wb.save(dst)
    return dst


def _materialize_parametric_sheet(
    ws: Worksheet,
    scenario: str,
    sheet_name: str,
) -> None:
    """In-place: merge BAU + scenario rows, then drop the scenario column."""
    headers = _read_header(ws)
    if not headers or headers[0] != "scenario":
        # Sheet was not migrated to v18; treat as pass-through.
        return

    keys = SCENARIO_OVERRIDE_KEYS.get(sheet_name)
    if keys is None:
        # Unknown sheet — leave alone after dropping the scenario column.
        ws.delete_cols(1)
        return

    key_col_indices = []
    for k in keys:
        if k not in headers:
            raise ValueError(
                f"Sheet '{sheet_name}' override key '{k}' not in headers: {headers}"
            )
        key_col_indices.append(headers.index(k))

    scen_col_idx = headers.index("scenario")

    # Read every data row, partition by scenario, preserving order.
    all_rows: list[list[Any]] = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        # Skip rows with no data outside the scenario column
        if all(v is None or v == "" for i, v in enumerate(row) if i != scen_col_idx):
            continue
        all_rows.append(list(row))

    def key_of(row: list[Any]) -> tuple:
        return tuple(row[i] for i in key_col_indices)

    # Index BAU rows by key, preserving order of first occurrence.
    merged: dict[tuple, list[Any]] = {}
    for row in all_rows:
        if str(row[scen_col_idx]).strip() == BAU_SCENARIO:
            merged[key_of(row)] = row

    if scenario != BAU_SCENARIO:
        for row in all_rows:
            scen_val = row[scen_col_idx]
            if scen_val is None:
                continue
            if str(scen_val).strip() == scenario:
                merged[key_of(row)] = row

    # Clear all data rows on the sheet.
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    # Drop the scenario column (col 1).
    ws.delete_cols(1)

    # Write merged rows back without the scenario column.
    for r_idx, row in enumerate(merged.values(), start=2):
        out_row = [v for i, v in enumerate(row) if i != scen_col_idx]
        for c_idx, value in enumerate(out_row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)


# =============================================================================
# Restrictions: read / apply / persist
# =============================================================================

def read_restrictions(
    soasia_path: Path | str,
    scenarios: list[str],
) -> dict[tuple[str, str, str, int], Any]:
    """Read Restrictions rows for the given source scenarios.

    Conflict resolution: when the same (source_sheet, tech, parameter, year)
    key appears in multiple sources, the source listed LAST in `scenarios`
    wins. This matches the documented semantics of `inherit_restrictions_from`.
    """
    path = Path(soasia_path)
    wb = load_workbook(path, read_only=True, data_only=True)
    if RESTRICTIONS_SHEET not in wb.sheetnames:
        wb.close()
        raise ValueError(
            f"Workbook {path.name} has no '{RESTRICTIONS_SHEET}' sheet."
        )
    ws = wb[RESTRICTIONS_SHEET]
    headers = _read_header(ws)
    expected = set(RESTRICTIONS_HEADERS)
    if not expected.issubset(headers):
        missing = expected - set(headers)
        wb.close()
        raise ValueError(f"Restrictions sheet missing columns: {missing}")

    col = {h: headers.index(h) for h in RESTRICTIONS_HEADERS}

    # Collect rows per scenario.
    rows_per_scenario: dict[str, list[dict[str, Any]]] = {s: [] for s in scenarios}
    scenario_set = set(scenarios)

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(v is None or v == "" for v in row):
            continue
        scen = row[col["scenario"]]
        if scen is None:
            continue
        scen = str(scen).strip()
        if scen not in scenario_set:
            continue
        rows_per_scenario[scen].append(
            {
                "source_sheet": row[col["source_sheet"]],
                "tech": row[col["tech"]],
                "parameter": row[col["parameter"]],
                "year": row[col["year"]],
                "value": row[col["value"]],
            }
        )

    wb.close()

    # Validate that every listed scenario produced at least one row
    empty = [s for s in scenarios if not rows_per_scenario[s]]
    if empty:
        raise ValueError(
            f"No Restrictions rows found for scenario(s): {empty}. "
            f"Run those scenarios first, or remove them from "
            f"inherit_restrictions_from."
        )

    # Merge with last-wins semantics by following the order in `scenarios`.
    out: dict[tuple[str, str, str, int], Any] = {}
    for scen in scenarios:
        for r in rows_per_scenario[scen]:
            key = (
                str(r["source_sheet"]),
                str(r["tech"]),
                str(r["parameter"]),
                int(r["year"]) if r["year"] is not None else 0,
            )
            out[key] = r["value"]
    return out


def apply_restrictions(
    parametrization_xlsx: Path | str,
    restrictions: dict[tuple[str, str, str, int], Any],
    *,
    source_sheet_filter: str | None = DEFAULT_RESTRICTION_SOURCE_SHEET,
) -> int:
    """Write inherited restriction values into A-O_Parametrization.xlsx.

    For each (source_sheet, tech, parameter, year) -> value, find the matching
    row in the workbook (tech + parameter) and set the year column. Only
    `source_sheet_filter` is honored today (other sheets are skipped, see plan).

    Returns the number of cells written.
    """
    path = Path(parametrization_xlsx)
    wb = load_workbook(path)
    if source_sheet_filter not in wb.sheetnames:
        wb.close()
        raise ValueError(
            f"Sheet '{source_sheet_filter}' not present in {path.name}."
        )
    ws = wb[source_sheet_filter]

    headers = _read_header(ws)
    if "Tech" not in headers or "Parameter" not in headers:
        wb.close()
        raise ValueError(
            f"Sheet '{source_sheet_filter}' missing Tech or Parameter columns."
        )
    tech_idx = headers.index("Tech") + 1
    param_idx = headers.index("Parameter") + 1

    # Build a column index map for year columns
    year_to_col: dict[int, int] = {}
    for i, h in enumerate(headers, start=1):
        try:
            y = int(h)
            year_to_col[y] = i
        except (TypeError, ValueError):
            pass

    # Build an index of (tech, parameter) -> row index for fast lookup
    row_index: dict[tuple[str, str], int] = {}
    for r in range(2, ws.max_row + 1):
        t = ws.cell(row=r, column=tech_idx).value
        p = ws.cell(row=r, column=param_idx).value
        if t is None or p is None:
            continue
        row_index[(str(t), str(p))] = r

    written = 0
    for (rsheet, tech, parameter, year), value in restrictions.items():
        if source_sheet_filter is not None and rsheet != source_sheet_filter:
            continue
        row = row_index.get((tech, parameter))
        col = year_to_col.get(year)
        if row is None or col is None:
            continue
        ws.cell(row=row, column=col, value=value)
        written += 1

    wb.save(path)
    return written


def persist_run_restrictions(
    soasia_path: Path | str,
    scenario: str,
    changes_json_path: Path | str,
    *,
    parameter: str = DEFAULT_RESTRICTION_PARAMETER,
    source_sheet: str = DEFAULT_RESTRICTION_SOURCE_SHEET,
) -> int:
    """After a rules_script run, rewrite the Restrictions rows for `scenario`.

    Reads the *_CHANGES.json produced by the rules_script and translates its
    'changes' entries into Restrictions rows. ALL pre-existing rows for this
    scenario are removed first (clear-and-write). Rows belonging to other
    scenarios are untouched.

    Returns the number of Restrictions rows written.
    """
    soasia = Path(soasia_path)
    log_path = Path(changes_json_path)
    if not log_path.exists():
        raise FileNotFoundError(f"Rules-script change log not found: {log_path}")

    log = json.loads(log_path.read_text())
    timestamp = log.get("timestamp") or datetime.now().isoformat()

    # Flatten all change entries across sheets.
    new_rows: list[list[Any]] = []
    for sheet_entry in log.get("sheets", []):
        for change in sheet_entry.get("changes", []):
            new_rows.append(
                [
                    scenario,                          # scenario
                    source_sheet,                      # source_sheet
                    change.get("tech"),                # tech
                    parameter,                         # parameter
                    change.get("year"),                # year
                    change.get("new"),                 # value
                    change.get("reason"),              # rule_applied
                    timestamp,                         # source_run_timestamp
                ]
            )

    wb = load_workbook(soasia)
    if RESTRICTIONS_SHEET not in wb.sheetnames:
        wb.close()
        raise ValueError(
            f"Workbook {soasia.name} has no '{RESTRICTIONS_SHEET}' sheet."
        )
    ws = wb[RESTRICTIONS_SHEET]

    headers = _read_header(ws)
    col = {h: headers.index(h) for h in RESTRICTIONS_HEADERS}

    # Preserve rows for other scenarios; drop rows for this scenario.
    keep: list[list[Any]] = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row is None or all(v is None or v == "" for v in row):
            continue
        scen_val = row[col["scenario"]]
        if scen_val is None:
            continue
        if str(scen_val).strip() == scenario:
            continue
        keep.append(list(row))

    # Clear data rows.
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    # Write headers (idempotent) + preserved rows + new rows.
    for c_idx, h in enumerate(RESTRICTIONS_HEADERS, start=1):
        ws.cell(row=1, column=c_idx, value=h)

    out_row = 2
    for row in keep:
        for c_idx, v in enumerate(row, start=1):
            ws.cell(row=out_row, column=c_idx, value=v)
        out_row += 1
    for row in new_rows:
        for c_idx, v in enumerate(row, start=1):
            ws.cell(row=out_row, column=c_idx, value=v)
        out_row += 1

    wb.save(soasia)
    return len(new_rows)


# =============================================================================
# Utilities
# =============================================================================

def available_rules_scripts(a3_process_dir: Path | str | None = None) -> list[str]:
    """List .py rules_script filenames under rules_scripts/. Returns [] if absent."""
    base = Path(a3_process_dir) if a3_process_dir else SCRIPT_DIR
    rs_dir = base / "rules_scripts"
    if not rs_dir.is_dir():
        return []
    return sorted(
        p.name for p in rs_dir.glob("*.py") if not p.name.startswith("_")
    )


def _cli_list_active(args) -> int:
    """Print active scenarios in topological order, one per line.

    Used by run.py to discover the per-scenario loop without importing this
    module inside its own Python (run.py orchestrates conda invocations).
    Pass --soasia to point at a non-default v18 path; otherwise the default
    SOASIA next to this script is used.
    """
    soasia = Path(args.soasia) if args.soasia else DEFAULT_SOASIA
    if not soasia.is_file():
        # Absent v18 -> single-scenario BAU (legacy mode). Print BAU so the
        # caller's loop runs once with --scenario BAU.
        print(BAU_SCENARIO)
        return 0
    configs = read_control_sheet(soasia)
    ordered = topological_order(configs)
    for cfg in ordered:
        print(cfg.scenario)
    return 0


def topological_order(
    configs: list[ScenarioConfig],
) -> list[ScenarioConfig]:
    """Return active scenarios ordered so dependencies (inherit_restrictions_from)
    run before dependents.

    Inactive scenarios are dropped. Cycles raise ValueError. Self-references
    are treated as no-op (a scenario doesn't depend on itself).
    """
    active = [c for c in configs if c.active]
    by_name = {c.scenario: c for c in active}

    # Drop inherit edges that point to scenarios not active (we don't run them
    # this pass, but their Restrictions rows still get read).
    visited: dict[str, int] = {}  # 0=unseen, 1=in-progress, 2=done
    order: list[ScenarioConfig] = []

    def visit(name: str) -> None:
        state = visited.get(name, 0)
        if state == 2:
            return
        if state == 1:
            raise ValueError(
                f"Inheritance cycle detected at scenario '{name}'."
            )
        if name not in by_name:
            return  # inherited-from scenarios that aren't active this run
        visited[name] = 1
        cfg = by_name[name]
        for src in cfg.inherit_restrictions_from:
            if src == name:
                continue
            visit(src)
        visited[name] = 2
        order.append(cfg)

    for cfg in active:
        visit(cfg.scenario)

    return order


# =============================================================================
# CLI entry point — used by run.py to enumerate active scenarios
# =============================================================================

if __name__ == "__main__":
    import argparse as _argparse

    _parser = _argparse.ArgumentParser(
        description="Helpers for SOASIA v18 multi-scenario support."
    )
    _sub = _parser.add_subparsers(dest="cmd", required=True)
    _list = _sub.add_parser(
        "list-active",
        help="Print active scenarios in topological order (one per line).",
    )
    _list.add_argument("--soasia", default=None,
                       help="Path to SOASIA v18 (default: next to this script).")
    _args = _parser.parse_args()

    if _args.cmd == "list-active":
        sys.exit(_cli_list_active(_args))
    else:
        sys.exit(f"Unknown command: {_args.cmd}")
