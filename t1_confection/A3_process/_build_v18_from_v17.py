# -*- coding: utf-8 -*-
"""
_build_v18_from_v17.py

One-shot migration script: takes SOASIA_OSeMOSYS_Template_v17.xlsx and produces
SOASIA_OSeMOSYS_Template_v18.xlsx with multi-scenario structure:

  1. New sheet `Control` (first sheet): one row per scenario with columns
     scenario, active, rules_script, inherit_restrictions_from, notes.
     Seeded with a single BAU row.

  2. New sheet `Restrictions` (second sheet, after README): persistent long-format
     store of restrictions written by rules_scripts, keyed by scenario.
     Starts empty (only headers).

  3. Column `scenario` inserted as the first column in each of the 15 parametric
     sheets. All existing data rows get scenario="BAU".

  4. Data validations:
       - Control.active        -> dropdown TRUE/FALSE
       - Parametric.scenario   -> dropdown referencing Control!$A:$A

Sheets NOT modified:
  - README              (free-form text)
  - Yearsplit_Template  (timeslice definitions, scenario-independent)
  - DaySplit            (timeslice definitions, scenario-independent)

Run from the A3_process folder:

    python _build_v18_from_v17.py
    python _build_v18_from_v17.py --source <v17.xlsx> --output <v18.xlsx>
    python _build_v18_from_v17.py --overwrite        # re-generate v18 in place
"""

from __future__ import annotations

import argparse
import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


SCRIPT_DIR = Path(__file__).resolve().parent

DEFAULT_SOURCE = SCRIPT_DIR / "SOASIA_OSeMOSYS_Template_v17.xlsx"
DEFAULT_OUTPUT = SCRIPT_DIR / "SOASIA_OSeMOSYS_Template_v18.xlsx"

# Sheets that receive a "scenario" column inserted at position 1.
# Order matches the v17 sheet order so the visual layout is preserved.
PARAMETRIC_SHEETS = [
    "Fixed_Horizon_Parameters",
    "Primary_Techs",
    "Secondary_Techs",
    "Capacities_CF",
    "VariableCost",
    "Demand_Projection",
    "Demand_Profiles",
    "Demand_Techs",
    "Emissions",
    "Interconnectors",
    "Interconnector_Params",
    "Existing_Generation",
    "Planned_Generation",
    "Technology_Costs",
    "RE_Targets_Policies",
]

SHEETS_WITHOUT_SCENARIO = ["README", "Yearsplit_Template", "DaySplit"]

CONTROL_SHEET = "Control"
RESTRICTIONS_SHEET = "Restrictions"

CONTROL_HEADERS = [
    "scenario",
    "active",
    "rules_script",
    "inherit_restrictions_from",
    "notes",
]

RESTRICTIONS_HEADERS = [
    "scenario",
    "source_sheet",
    "tech",
    "parameter",
    "year",
    "value",
    "rule_applied",
    "source_run_timestamp",
]

# Tab colors (ARGB) - mild visual hierarchy
TAB_COLOR_CONTROL = "FF4F81BD"       # blue: configuration
TAB_COLOR_RESTRICTIONS = "FFE26B0A"  # orange: machine-written / user-editable persistence

BAU_SCENARIO = "BAU"
DEFAULT_RULES_SCRIPT = "add_max_cap_investment_lid_rule.py"


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--source", type=Path, default=DEFAULT_SOURCE,
                   help=f"Path to v17 template (default: {DEFAULT_SOURCE.name})")
    p.add_argument("--output", type=Path, default=DEFAULT_OUTPUT,
                   help=f"Path to write v18 (default: {DEFAULT_OUTPUT.name})")
    p.add_argument("--overwrite", action="store_true",
                   help="Overwrite output if it already exists")
    return p.parse_args()


def _last_data_row(ws) -> int:
    """Return the index of the last row that contains any non-empty cell.

    openpyxl's ws.max_row counts the last row with formatting or stale presence,
    not the last row with real data. We need the real-data limit so we don't tag
    empty trailing rows with a scenario.
    """
    for r in range(ws.max_row, 1, -1):
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=r, column=c).value not in (None, ""):
                return r
    return 1  # only header


def add_scenario_column(ws, scenario_value: str = BAU_SCENARIO) -> int:
    """Insert a `scenario` column at position 1 and tag only rows with real data.

    Rows that are entirely empty (between blocks of data, or trailing) stay empty.
    Header row is assumed to be row 1. Returns the number of data rows tagged.
    """
    # Snapshot which original rows actually contain data BEFORE inserting the new column.
    last = _last_data_row(ws)
    data_rows = []
    for r in range(2, last + 1):
        has_data = False
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=r, column=c).value not in (None, ""):
                has_data = True
                break
        if has_data:
            data_rows.append(r)

    ws.insert_cols(1)
    ws.cell(row=1, column=1, value="scenario")
    for r in data_rows:
        ws.cell(row=r, column=1, value=scenario_value)
    return len(data_rows)


def build_control_sheet(wb) -> None:
    """Create the Control sheet at position 0 with one BAU row seeded."""
    if CONTROL_SHEET in wb.sheetnames:
        del wb[CONTROL_SHEET]
    ws = wb.create_sheet(CONTROL_SHEET, 0)
    ws.sheet_properties.tabColor = TAB_COLOR_CONTROL

    for col_idx, header in enumerate(CONTROL_HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Seed BAU row
    ws.cell(row=2, column=1, value=BAU_SCENARIO)
    ws.cell(row=2, column=2, value=True)
    ws.cell(row=2, column=3, value=DEFAULT_RULES_SCRIPT)
    ws.cell(row=2, column=4, value=None)
    ws.cell(row=2, column=5, value="Base scenario")

    # Width hints for readability
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 40

    # Data validation: active -> TRUE/FALSE list
    dv_active = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
    dv_active.add(f"B2:B1048576")
    ws.add_data_validation(dv_active)


def build_restrictions_sheet(wb) -> None:
    """Create the Restrictions sheet (headers only) right after Control."""
    if RESTRICTIONS_SHEET in wb.sheetnames:
        del wb[RESTRICTIONS_SHEET]
    # Insert after Control (index 1)
    ws = wb.create_sheet(RESTRICTIONS_SHEET, 1)
    ws.sheet_properties.tabColor = TAB_COLOR_RESTRICTIONS

    for col_idx, header in enumerate(RESTRICTIONS_HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Width hints
    widths = {"A": 14, "B": 22, "C": 16, "D": 36, "E": 8, "F": 12, "G": 18, "H": 22}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def add_scenario_dropdown(ws, scenario_count_hint: int = 100) -> None:
    """Attach a data validation on column A (scenario) referencing Control scenarios."""
    # Reference to Control!$A$2:$A$<scenario_count_hint+1>, fixed range so users see scenarios
    # they have defined. Using $A:$A would include the header "scenario" itself as a valid value.
    formula = f"=Control!$A$2:$A${scenario_count_hint + 1}"
    dv = DataValidation(type="list", formula1=formula, allow_blank=False)
    # Apply to A2:A<huge> so future rows are covered
    dv.add("A2:A1048576")
    ws.add_data_validation(dv)


def reorder_sheets(wb) -> None:
    """Force order: Control, README, Restrictions, then v17 sheet order for the rest."""
    desired_prefix = [CONTROL_SHEET, "README", RESTRICTIONS_SHEET]
    rest_order = [
        "Fixed_Horizon_Parameters",
        "Primary_Techs",
        "Secondary_Techs",
        "Capacities_CF",
        "VariableCost",
        "Demand_Projection",
        "Demand_Profiles",
        "Demand_Techs",
        "Emissions",
        "Yearsplit_Template",
        "DaySplit",
        "Interconnectors",
        "Interconnector_Params",
        "Existing_Generation",
        "Planned_Generation",
        "Technology_Costs",
        "RE_Targets_Policies",
    ]
    desired_order = desired_prefix + rest_order

    present = [s for s in desired_order if s in wb.sheetnames]
    # Any sheet not in desired_order keeps its current position relative to the tail
    others = [s for s in wb.sheetnames if s not in present]
    final_order = present + others

    # openpyxl exposes _sheets directly; reorder by name.
    name_to_sheet = {s.title: s for s in wb._sheets}
    wb._sheets = [name_to_sheet[name] for name in final_order]


def main() -> int:
    args = parse_args()

    src: Path = args.source
    dst: Path = args.output

    if not src.exists():
        print(f"ERROR: source not found: {src}", file=sys.stderr)
        return 1

    if dst.exists() and not args.overwrite:
        print(f"ERROR: output already exists: {dst} (pass --overwrite to replace)", file=sys.stderr)
        return 1

    print(f"Copying {src.name} -> {dst.name}")
    shutil.copy(src, dst)

    print(f"Loading workbook (this may take a moment for large sheets)...")
    wb = load_workbook(dst)

    sheets_present = set(wb.sheetnames)

    # Sanity check: all parametric sheets we intend to modify must exist
    missing = [s for s in PARAMETRIC_SHEETS if s not in sheets_present]
    if missing:
        print(f"ERROR: expected parametric sheets missing from source: {missing}", file=sys.stderr)
        return 2

    # 1. Insert scenario column into each parametric sheet
    for sheet_name in PARAMETRIC_SHEETS:
        ws = wb[sheet_name]
        tagged = add_scenario_column(ws, BAU_SCENARIO)
        print(f"  [{sheet_name}] scenario column inserted; {tagged} data rows tagged BAU")

    # 2. Build Control and Restrictions
    build_control_sheet(wb)
    print(f"  [Control] created with BAU seed row")
    build_restrictions_sheet(wb)
    print(f"  [Restrictions] created (headers only)")

    # 3. Attach scenario-column dropdown to each parametric sheet
    for sheet_name in PARAMETRIC_SHEETS:
        ws = wb[sheet_name]
        add_scenario_dropdown(ws)

    # 4. Order: Control, README, Restrictions, then the rest
    reorder_sheets(wb)
    print(f"  Sheets reordered: {wb.sheetnames}")

    # 5. Save
    print(f"Saving {dst.name}...")
    wb.save(dst)
    print(f"Done. {dst}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
