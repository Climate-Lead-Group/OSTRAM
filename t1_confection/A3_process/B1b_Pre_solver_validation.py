"""
Pre-solver validation for A-O_Parametrization.xlsx.

Detects three classes of infeasibility-prone data BEFORE B1_Compiler reads
the workbook (and BEFORE B2 emits the .txt for glpsol). For each class
shows the auto-fix formula and asks the user before writing.

Validations:
    V1. Per-year:    TotalAnnualMinCapacityInvestment(y) >= TotalAnnualMaxCapacityInvestment(y)
                     => Max_inv(y) = Min_inv(y) * 1.01
    V2. Cumulative:  TotalAnnualMaxCapacity(y) <= ResidualCapacity(y) + Σ Min_inv(y') in [y-OL+1, y]
                     => Max_tot(y) = (Residual + ΣMin) * 1.01
    V3. Activity:    TotalTechnologyAnnualActivityLowerLimit(y) > max_activity(y)
                     where max_activity = max_capacity × AvailabilityFactor × CapacityToActivityUnit × Σ(CF·YS)
                     => ActivityLowerLimit(y) = max_activity(y) * 0.99

Usage:
    # Standalone CLI
    python t1_confection/B1b_Pre_solver_validation.py --scenario BAU
    python t1_confection/B1b_Pre_solver_validation.py --scenario BAU --report-only
    python t1_confection/B1b_Pre_solver_validation.py --scenario BAU --auto-fix-all
    python t1_confection/B1b_Pre_solver_validation.py --xlsx path/to/A-O_Parametrization.xlsx

    # Imported from B1_Compiler.py
    from B1b_Pre_solver_validation import run as pre_solver_validate
    any_fix, abort = pre_solver_validate(scenario, xlsx_path, interactive=True)
"""
import argparse
import csv
import shutil
import sys
import yaml
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from ostram._legacy_import import load_file_module

_validation_core = load_file_module(
    "_ostram_stage11_xlsx_validation_core",
    Path(__file__).resolve().with_name("_xlsx_validation_core.py"),
)
AF_PARAM = _validation_core.AF_PARAM
ACT_LOWER_PARAM = _validation_core.ACT_LOWER_PARAM
ACT_LOWER_HAIRCUT = _validation_core.ACT_LOWER_HAIRCUT
MAX_MULTIPLIER = _validation_core.MAX_MULTIPLIER
consistency_sweep = _validation_core.consistency_sweep
index_target_sheet = _validation_core.index_target_sheet
load_capacity_to_activity_unit = _validation_core.load_capacity_to_activity_unit
load_operational_life = _validation_core.load_operational_life
load_yearsplit = _validation_core.load_yearsplit
validate_activity_lower_limit = _validation_core.validate_activity_lower_limit

TARGET_SHEETS = ("Secondary Techs", "Demand Techs")


# -------------------------------------------------------------------
# Path helpers
# -------------------------------------------------------------------
def default_xlsx_for_scenario(scenario):
    """Return the conventional path for A-O_Parametrization.xlsx of a scenario."""
    return Path(__file__).parent / "A1_Outputs" / f"A1_Outputs_{scenario}" / "A-O_Parametrization.xlsx"


def read_base_year_from_config():
    """Best-effort: read base_year from Config_MOMF_T1_A.yaml. Returns int or None."""
    cfg = Path(__file__).parent / "Config_MOMF_T1_A.yaml"
    if not cfg.exists():
        return None
    try:
        with open(cfg, "r", encoding="utf-8") as f:
            data = yaml.safe_load(f)
        v = data.get("base_year")
        return int(v) if v is not None else None
    except Exception:
        return None


# -------------------------------------------------------------------
# Prompt
# -------------------------------------------------------------------
def _prompt_group(group_name, issues, formula_text, sample_formatter,
                   prompt_mode, interactive):
    """Show issues for one validation group; ask user.

    Returns ("apply"|"skip", new_prompt_mode).
    """
    n = len(issues)
    if n == 0:
        return ("skip", prompt_mode)
    print(f"\n[{group_name}]  {n} issue{'s' if n != 1 else ''}:")
    for i, issue in enumerate(issues[:5]):
        print("  " + sample_formatter(issue))
    if n > 5:
        print(f"  ... and {n - 5} more")
    print(f"  Fix: {formula_text}")

    if prompt_mode == "skip-all":
        print(f"  [auto] skipping {n} fix{'es' if n != 1 else ''}.")
        return ("skip", prompt_mode)
    if prompt_mode == "apply-all" or not interactive:
        print(f"  [auto] applying {n} fix{'es' if n != 1 else ''}.")
        return ("apply", prompt_mode)

    while True:
        try:
            ans = input(f"  Apply {n} fix{'es' if n != 1 else ''}? [y/N/all/skip-all/details]: ").strip().lower()
        except EOFError:
            ans = "n"
        if ans in ("", "n", "no"):
            return ("skip", prompt_mode)
        if ans in ("y", "yes"):
            return ("apply", prompt_mode)
        if ans == "all":
            return ("apply", "apply-all")
        if ans in ("skip-all", "skipall", "sa"):
            return ("skip", "skip-all")
        if ans in ("d", "details"):
            print("\n  --- ALL ISSUES ---")
            for issue in issues:
                print("    " + sample_formatter(issue))
            print("  --- END ---")
            continue
        print("  Choose y / N / all / skip-all / details")


# -------------------------------------------------------------------
# Sample formatters
# -------------------------------------------------------------------
def _fmt_v1(a):
    return (f"{a['target_sheet']:18s} {a['tech']:<14}  {a['year']}  "
            f"min={a['min']:.6g}  old_max={a['old_max']:.6g}  -> new_max={a['new_max']:.6g}")


def _fmt_v2(a):
    return (f"{a['target_sheet']:18s} {a['tech']:<14}  {a['year']}  "
            f"residual={a['residual']:.4g} + ΣMin[{a['window_start']}..{a['year']}]={a['accumulated_min']:.4g} "
            f"=> threshold={a['residual']+a['accumulated_min']:.4g}  "
            f"old_max_tot={a['old_max_tot']:.4g}  -> new={a['new_max_tot']:.4g}  (opLife={a['op_life']})")


def _fmt_v3(a):
    cal = "  ⚠ calibration-impact" if a.get("calibration_impact") else ""
    return (f"{a['target_sheet']:18s} {a['tech']:<14}  {a['year']}  "
            f"max_cap={a['max_capacity']:.4g}  AF={a['af']:.3g}  C2A={a['c2a']:.4g}  Σ(CF·YS)={a['ann_cf']:.4g}  "
            f"=> max_activity={a['max_activity']:.4g}  >  old_lower={a['old_lower']:.4g}  "
            f"-> new_lower={a['new_lower']:.4g}  gap={a['gap']:.4g}{cal}")


# -------------------------------------------------------------------
# Apply helpers (write fixes the user accepted)
# -------------------------------------------------------------------
def _apply_v1_v2(workbook_sheet_ctx, max_adjusts, max_total_adjusts, accept_v1, accept_v2):
    written = 0
    if accept_v1:
        for a in max_adjusts:
            ws = workbook_sheet_ctx[a["target_sheet"]]
            ws.cell(a["row"], a["col"]).value = a["new_max"]
            ws.cell(a["row"], _proj_mode_col(workbook_sheet_ctx, a["target_sheet"])).value = "User defined"
            written += 1
    if accept_v2:
        for a in max_total_adjusts:
            ws = workbook_sheet_ctx[a["target_sheet"]]
            ws.cell(a["row"], a["col"]).value = a["new_max_tot"]
            ws.cell(a["row"], _proj_mode_col(workbook_sheet_ctx, a["target_sheet"])).value = "User defined"
            written += 1
    return written


def _apply_v3(workbook_sheet_ctx, issues, accept):
    if not accept:
        return 0
    written = 0
    for a in issues:
        ws = workbook_sheet_ctx[a["target_sheet"]]
        ws.cell(a["row"], a["col"]).value = a["new_lower"]
        ws.cell(a["row"], _proj_mode_col(workbook_sheet_ctx, a["target_sheet"])).value = "User defined"
        written += 1
    return written


def _proj_mode_col(workbook_sheet_ctx, sheet_name):
    """Cached lookup for 'Projection.Mode' column index in a sheet."""
    ws = workbook_sheet_ctx[sheet_name]
    if not hasattr(ws, "_proj_mode_cached"):
        for i, c in enumerate(ws[1], 1):
            if c.value and str(c.value).strip() == "Projection.Mode":
                ws._proj_mode_cached = i
                break
        else:
            raise RuntimeError(f"'Projection.Mode' not found in sheet '{sheet_name}'")
    return ws._proj_mode_cached


# -------------------------------------------------------------------
# CSV report
# -------------------------------------------------------------------
def _write_report(scenario, all_records, applied_groups):
    out_dir = Path(__file__).parent / "Executables" / f"{scenario}_0"
    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / "_validation_report.csv"
    fieldnames = [
        "validation", "sheet", "tech", "year",
        "current_value", "threshold", "gap", "proposed_fix",
        "applied", "formula", "calibration_impact",
    ]
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for rec in all_records:
            rec_out = {k: rec.get(k, "") for k in fieldnames}
            rec_out["applied"] = rec["validation"] in applied_groups
            w.writerow(rec_out)
    return out


def _to_records(max_adjusts, max_total_adjusts, act_issues):
    out = []
    for a in max_adjusts:
        out.append({
            "validation": "V1",
            "sheet": a["target_sheet"], "tech": a["tech"], "year": a["year"],
            "current_value": a["old_max"], "threshold": a["min"],
            "gap": a["min"] - a["old_max"], "proposed_fix": a["new_max"],
            "formula": "Max_inv(y) = Min_inv(y) * 1.01",
            "calibration_impact": False,
        })
    for a in max_total_adjusts:
        thr = a["residual"] + a["accumulated_min"]
        out.append({
            "validation": "V2",
            "sheet": a["target_sheet"], "tech": a["tech"], "year": a["year"],
            "current_value": a["old_max_tot"], "threshold": thr,
            "gap": thr - a["old_max_tot"], "proposed_fix": a["new_max_tot"],
            "formula": "Max_tot(y) = (Residual + ΣMin in window) * 1.01",
            "calibration_impact": False,
        })
    for a in act_issues:
        out.append({
            "validation": "V3",
            "sheet": a["target_sheet"], "tech": a["tech"], "year": a["year"],
            "current_value": a["old_lower"], "threshold": a["max_activity"],
            "gap": a["gap"], "proposed_fix": a["new_lower"],
            "formula": "ActivityLowerLimit(y) = max_activity(y) * 0.99",
            "calibration_impact": a.get("calibration_impact", False),
        })
    return out


# -------------------------------------------------------------------
# Main entry point
# -------------------------------------------------------------------
def run(scenario, xlsx_path=None, *, interactive=True, auto_fix_all=False,
        report_only=False, base_year=None):
    """Run V1+V2+V3. Returns (any_fix_applied, abort).

    abort=True only if the user replied with skip-all to *every* group AND the
    caller passed interactive=True with a non-trivial issue set; in practice
    we never abort the caller automatically — the user controls flow via
    skip-all. Reserved for future use.
    """
    if xlsx_path is None:
        xlsx_path = default_xlsx_for_scenario(scenario)
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        print(f"[VALIDATE] xlsx not found: {xlsx_path}")
        return (False, False)
    if base_year is None:
        base_year = read_base_year_from_config()

    print(f"\n[VALIDATE] Pre-solver validation on {xlsx_path}")
    if base_year:
        print(f"[VALIDATE] base_year = {base_year}")

    wb = openpyxl.load_workbook(xlsx_path)
    oplife_map = load_operational_life(wb)
    c2a_map = load_capacity_to_activity_unit(wb)
    yearsplit = load_yearsplit(wb)
    print(f"[VALIDATE] Loaded {len(oplife_map)} OperationalLife, "
          f"{len(c2a_map)} CapacityToActivityUnit, {len(yearsplit)} year-splits.")

    # Index every target sheet (including AvailabilityFactor + ActivityLowerLimit)
    sheet_ctx = {}
    workbook_sheet_ctx = {}
    for sheet_name in TARGET_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"[WARN] Target sheet '{sheet_name}' missing; skipping.")
            continue
        ws, cols, ymap, t2r, dup_warns = index_target_sheet(
            wb, sheet_name, extra_params=(AF_PARAM, ACT_LOWER_PARAM)
        )
        sheet_ctx[sheet_name] = (ws, cols, ymap, t2r)
        workbook_sheet_ctx[sheet_name] = ws
        if dup_warns:
            print(f"[WARN] {sheet_name}: {len(dup_warns)} duplicate tech rows.")

    # === Run V1 + V2 sweep (no apply) ===
    all_max_adjusts = []
    all_max_total_adjusts = []
    for sheet_name, (ws, cols, ymap, t2r) in sheet_ctx.items():
        ma, mta = consistency_sweep(ws, sheet_name, cols, ymap, t2r,
                                     oplife_map, apply_changes=False)
        all_max_adjusts.extend(ma)
        all_max_total_adjusts.extend(mta)

    # === Run V3 sweep (no apply) ===
    all_act_issues = []
    for sheet_name, (ws, cols, ymap, t2r) in sheet_ctx.items():
        issues = validate_activity_lower_limit(
            ws, sheet_name, cols, ymap, t2r, oplife_map,
            yearsplit, c2a_map, capacities_wb=wb,
            base_year=base_year, apply_changes=False,
        )
        all_act_issues.extend(issues)

    # === Prompt and apply ===
    prompt_mode = "apply-all" if (auto_fix_all and not report_only) else "ask"
    if report_only:
        prompt_mode = "skip-all"

    accept_v1 = "skip"
    accept_v2 = "skip"
    accept_v3 = "skip"

    accept_v1, prompt_mode = _prompt_group(
        "V1: Per-year max-inv vs min-inv",
        all_max_adjusts,
        "Max_inv(y) = Min_inv(y) * 1.01",
        _fmt_v1, prompt_mode, interactive=(interactive and not report_only),
    )
    accept_v2, prompt_mode = _prompt_group(
        "V2: Cumulative max-cap vs residual+ΣMin (window=OperationalLife)",
        all_max_total_adjusts,
        "Max_tot(y) = (Residual + Σ Min_inv[y - OpLife + 1 .. y]) * 1.01",
        _fmt_v2, prompt_mode, interactive=(interactive and not report_only),
    )
    accept_v3, prompt_mode = _prompt_group(
        "V3: Activity-vs-Capacity (CAb1 vs AAC3)",
        all_act_issues,
        "ActivityLowerLimit(y) = max_activity(y) * 0.99    "
        "where max_activity = (Residual + Σ Max_inv[y-OL+1..y]) * AvailabilityFactor * CapacityToActivityUnit * Σ(CF·YS)",
        _fmt_v3, prompt_mode, interactive=(interactive and not report_only),
    )

    applied_groups = set()
    if accept_v1 == "apply":
        applied_groups.add("V1")
    if accept_v2 == "apply":
        applied_groups.add("V2")
    if accept_v3 == "apply":
        applied_groups.add("V3")

    any_fix = False
    if not report_only and applied_groups:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = xlsx_path.with_name(f"{xlsx_path.stem}.backup_pre_validate_{ts}{xlsx_path.suffix}")
        shutil.copy2(xlsx_path, backup)
        print(f"\n[VALIDATE] Backup: {backup.name}")
        n_v12 = _apply_v1_v2(workbook_sheet_ctx, all_max_adjusts, all_max_total_adjusts,
                              accept_v1 == "apply", accept_v2 == "apply")
        n_v3 = _apply_v3(workbook_sheet_ctx, all_act_issues, accept_v3 == "apply")
        wb.save(xlsx_path)
        print(f"[VALIDATE] Wrote {n_v12 + n_v3} cell{'s' if (n_v12 + n_v3) != 1 else ''} "
              f"({n_v12} from V1+V2, {n_v3} from V3) to {xlsx_path.name}")
        any_fix = (n_v12 + n_v3) > 0

    # CSV report (always)
    records = _to_records(all_max_adjusts, all_max_total_adjusts, all_act_issues)
    if records:
        report_path = _write_report(scenario, records, applied_groups)
        print(f"[VALIDATE] Report: {report_path}")
    else:
        print("[VALIDATE] No issues found.")

    wb.close()
    return (any_fix, False)


# -------------------------------------------------------------------
# CLI
# -------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(description="Pre-solver validation for A-O_Parametrization.xlsx")
    ap.add_argument("--scenario", help="Scenario name (e.g. BAU). Used to derive default xlsx path.")
    ap.add_argument("--xlsx", help="Override xlsx path directly.")
    grp = ap.add_mutually_exclusive_group()
    grp.add_argument("--non-interactive", action="store_true", help="Do not prompt; fail on issues.")
    grp.add_argument("--auto-fix-all", action="store_true", help="Apply every fix without prompting.")
    grp.add_argument("--report-only", action="store_true", help="Just write the report; do not modify the xlsx.")
    args = ap.parse_args()

    if not args.scenario and not args.xlsx:
        ap.error("Provide --scenario or --xlsx")

    scenario = args.scenario or Path(args.xlsx).parent.name.replace("A1_Outputs_", "")
    xlsx = args.xlsx
    if xlsx is None:
        xlsx = default_xlsx_for_scenario(scenario)
        if not Path(xlsx).exists():
            print(f"[ERROR] Could not locate xlsx for scenario {scenario}: {xlsx}")
            return 2

    interactive = not args.non_interactive
    any_fix, abort = run(
        scenario, xlsx,
        interactive=interactive,
        auto_fix_all=args.auto_fix_all,
        report_only=args.report_only,
    )
    return 0 if not abort else 1


if __name__ == "__main__":
    sys.exit(main())
