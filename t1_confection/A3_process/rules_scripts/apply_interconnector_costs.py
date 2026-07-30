"""
apply_interconnector_costs.py
=============================

WS-3 wiring fix. Makes the v18 `Interconnector_Params` sheet the SOURCE OF TRUTH
for interconnector CapitalCost / FixedCost / OperationalLife.

Background
----------
The pipeline historically read interconnector CapitalCost from the OG_csvs base
(distance-computed legacy values, e.g. BD<->IN_E = 292.487 $/kW). The sourced,
cited values in the v18 template's `Interconnector_Params` sheet (e.g. 380 $/kW)
were never consumed. This step closes that gap: for every interconnector tech
present in `Interconnector_Params`, it overwrites the matching model rows with
the sheet's values, across two sheets with different layouts:

    * Secondary Techs         -> CapitalCost, FixedCost   (year-indexed rows)
    * Fixed Horizon Parameters -> OperationalLife          (single 'Value' cell)

Scope (intentionally narrow): CapitalCost, FixedCost, OperationalLife only.
ResidualCapacity and the *Max*Capacity/Investment caps are owned by
fix_trn_residuals / cap_trn_to_residual / relax_interconnectors and are left
untouched; TransmissionLossFactor / CapacityFactor are out of WS-3 scope.

Scenario support: when run inside A3, the template read is the per-scenario
MATERIALIZED template ($OSTRAM_TEMPLATE_PATH), which already has BAU + any
scenario overrides merged (identity key (Tech, Parameter)). So per-scenario
interconnector costs are a one-row add to Interconnector_Params — no change here.

Source of the Interconnector_Params values (priority order):
  1. --template <path>
  2. $OSTRAM_TEMPLATE_PATH   (A3 stage 0 sets this to the materialized template)
  3. <script_dir>/../OSTRAM_Scenario_Inputs.xlsx   (canonical fallback)

OUTPUT
------
1. Timestamped backup of the input directory.
2. In-place edit of A-O_Parametrization.xlsx.
3. A JSON change log next to the backup.

USAGE
-----
    python apply_interconnector_costs.py --input-dir A1_Outputs/A1_Outputs_BAU
    python apply_interconnector_costs.py --self-test
    python apply_interconnector_costs.py --restore
"""
from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook, Workbook

# --------------------------------------------------------------------------- config
# Year-indexed params live in "Secondary Techs"; scalar params in "Fixed Horizon
# Parameters". Both sheets are scanned; each param is written wherever its row is.
DEFAULT_TARGET_SHEETS = ["Secondary Techs", "Fixed Horizon Parameters"]
PARAM_FILE_NAME = "A-O_Parametrization.xlsx"
IC_SHEET = "Interconnector_Params"
TEMPLATE_NAME = "OSTRAM_Scenario_Inputs.xlsx"

APPLIED_PARAMS = ("CapitalCost", "FixedCost", "OperationalLife")

PROJ_MODE_COL = "Projection.Mode"
PROJ_MODE_EMPTY = "EMPTY"
PROJ_MODE_USER = "User defined"
VALUE_COL = "Value"

TRN_RE = re.compile(r"^TRN[A-Z]{5}[A-Z]{5}$")
BACKUP_TAG = "_PRE_TRN_COSTS_"


# --------------------------------------------------------------------------- worksheet helpers
def find_year_columns(ws) -> dict:
    out = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if isinstance(val, int) and 1900 <= val <= 2200:
            out[val] = col_idx
    return out


def find_named_columns(ws, names) -> dict:
    found = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val in names:
            found[val] = col_idx
    return found


def _values_differ(a, b, tol: float = 1e-9) -> bool:
    if a is None and b is None:
        return False
    if a is None or b is None:
        return True
    try:
        return abs(float(a) - float(b)) > tol
    except (TypeError, ValueError):
        return a != b


# --------------------------------------------------------------------------- read source of truth
def resolve_template(explicit: Path | None) -> Path:
    candidates = []
    if explicit:
        candidates.append(Path(explicit))
    env = os.environ.get("OSTRAM_TEMPLATE_PATH")
    if env:
        candidates.append(Path(env))
    candidates.append(Path(__file__).resolve().parent.parent / TEMPLATE_NAME)
    for c in candidates:
        if c and c.is_file():
            return c
    raise FileNotFoundError(
        "Could not locate a template with Interconnector_Params. Tried: "
        + ", ".join(str(c) for c in candidates)
    )


def read_interconnector_params(template: Path) -> dict:
    """Return {tech: {param: value}} for APPLIED_PARAMS from Interconnector_Params.

    Values are the single constant across the year columns. If the template has a
    'scenario' column (raw v18), only BAU rows are read; the materialized template
    has no scenario column (already resolved) so all rows are read.
    """
    wb = load_workbook(template, read_only=True, data_only=True)
    try:
        if IC_SHEET not in wb.sheetnames:
            raise KeyError(f"'{IC_SHEET}' sheet not in {template}")
        rows = list(wb[IC_SHEET].iter_rows(values_only=True))
    finally:
        wb.close()

    hdr = list(rows[0])
    col = {name: i for i, name in enumerate(hdr) if name is not None}
    tech_i, par_i, scen_i = col.get("Tech"), col.get("Parameter"), col.get("scenario")
    if tech_i is None or par_i is None:
        raise ValueError(f"{IC_SHEET} missing Tech/Parameter header columns; got {hdr}")
    year_idx = [i for i, h in enumerate(hdr) if isinstance(h, int) and 1900 <= h <= 2200]

    out, nonconstant = {}, []
    for r in rows[1:]:
        if not r:
            continue
        if scen_i is not None and r[scen_i] not in (None, "", "BAU"):
            continue
        tech = r[tech_i] if tech_i < len(r) else None
        par = r[par_i] if par_i < len(r) else None
        if not tech or par not in APPLIED_PARAMS:
            continue
        yr_vals = [r[i] for i in year_idx if i < len(r) and r[i] not in (None, "")]
        distinct = sorted(set(yr_vals), key=lambda x: str(x))
        if not distinct:
            continue
        if len(distinct) > 1:
            nonconstant.append((str(tech), par, distinct))
        out.setdefault(str(tech), {})[par] = distinct[0]
    if nonconstant:
        raise ValueError(
            "Non-constant year values for cost params in Interconnector_Params: "
            + "; ".join(f"{t}/{p}={d}" for t, p, d in nonconstant)
        )
    return out


# --------------------------------------------------------------------------- core edit
def apply_costs(ws, ic_values: dict, seen: set) -> dict:
    """Overwrite interconnector cost cells in one sheet. Layout auto-detected:
    year-indexed (Secondary Techs) writes every year cell; scalar (Fixed Horizon
    Parameters) writes the single 'Value' cell. `seen` collects (tech, param)
    pairs written anywhere, so run() can report pairs with no home."""
    hdr = find_named_columns(ws, ["Tech", "Parameter", PROJ_MODE_COL, VALUE_COL])
    tech_c, par_c = hdr.get("Tech"), hdr.get("Parameter")
    pm_c, val_c = hdr.get(PROJ_MODE_COL), hdr.get(VALUE_COL)
    if tech_c is None or par_c is None:
        return {"sheet": ws.title, "skipped": "no Tech/Parameter columns"}
    year_cols = find_year_columns(ws)
    if year_cols:
        layout = "year"
    elif val_c is not None:
        layout = "scalar"
    else:
        return {"sheet": ws.title, "skipped": "no year columns and no Value column"}

    log = {"sheet": ws.title, "layout": layout, "changes": [], "rows_touched": [],
           "pm_flips": []}
    years = sorted(year_cols)

    for row_idx in range(2, ws.max_row + 1):
        tech = ws.cell(row=row_idx, column=tech_c).value
        par = ws.cell(row=row_idx, column=par_c).value
        if not isinstance(tech, str) or not TRN_RE.match(tech):
            continue
        if par not in APPLIED_PARAMS or tech not in ic_values or par not in ic_values[tech]:
            continue
        seen.add((tech, par))
        new_val = ic_values[tech][par]
        row_changed = False
        if layout == "year":
            for y in years:
                cell = ws.cell(row=row_idx, column=year_cols[y])
                if _values_differ(cell.value, new_val):
                    log["changes"].append({"tech": tech, "param": par, "year": y,
                                            "old": cell.value, "new": new_val})
                    cell.value = new_val
                    row_changed = True
        else:  # scalar
            cell = ws.cell(row=row_idx, column=val_c)
            if _values_differ(cell.value, new_val):
                log["changes"].append({"tech": tech, "param": par, "year": None,
                                        "old": cell.value, "new": new_val})
                cell.value = new_val
                row_changed = True
        if row_changed:
            log["rows_touched"].append({"tech": tech, "param": par, "new": new_val})
            if pm_c is not None:
                pm = ws.cell(row=row_idx, column=pm_c)
                if pm.value == PROJ_MODE_EMPTY:
                    pm.value = PROJ_MODE_USER
                    log["pm_flips"].append({"tech": tech, "param": par})
    return log


def edit_parametrization(filepath: Path, sheets: list, ic_values: dict) -> dict:
    wb = load_workbook(filepath)
    file_log = {"file": str(filepath), "sheets": []}
    seen: set = set()
    try:
        for sheet in sheets:
            if sheet not in wb.sheetnames:
                file_log["sheets"].append({"sheet": sheet, "skipped": "not present"})
                continue
            file_log["sheets"].append(apply_costs(wb[sheet], ic_values, seen))
        wb.save(filepath)
    finally:
        wb.close()
    # Pairs sourced in Interconnector_Params but with no row in any target sheet.
    missing = [{"tech": t, "param": p}
               for t, params in ic_values.items() for p in params
               if (t, p) not in seen]
    file_log["missing_pairs"] = missing
    return file_log


# --------------------------------------------------------------------------- backup / restore
def make_backup(input_dir: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = input_dir.parent / f"{input_dir.name}{BACKUP_TAG}{stamp}"
    if backup.exists():
        raise FileExistsError(f"Backup already exists: {backup}")
    shutil.copytree(input_dir, backup)
    return backup


def find_latest_backup(input_dir: Path) -> Path | None:
    parent = input_dir.parent
    cands = sorted((p for p in parent.iterdir()
                    if p.is_dir() and p.name.startswith(f"{input_dir.name}{BACKUP_TAG}")),
                   key=lambda p: p.name)
    return cands[-1] if cands else None


def restore_from_backup(input_dir: Path, backup_dir: Path | None = None) -> Path:
    input_dir = Path(input_dir)
    backup_dir = Path(backup_dir) if backup_dir else find_latest_backup(input_dir)
    if not backup_dir or not backup_dir.is_dir():
        raise FileNotFoundError(f"No {BACKUP_TAG}* backup found next to {input_dir}.")
    if input_dir.is_dir():
        shutil.rmtree(input_dir)
    shutil.copytree(backup_dir, input_dir)
    return backup_dir


# --------------------------------------------------------------------------- orchestration
def run(input_dir, sheets=None, skip_backup=False, template=None) -> dict:
    input_dir = Path(input_dir)
    sheets = sheets or DEFAULT_TARGET_SHEETS
    template_path = resolve_template(template)
    ic_values = read_interconnector_params(template_path)

    paramfile = input_dir / PARAM_FILE_NAME
    if not paramfile.exists():
        raise FileNotFoundError(f"{paramfile} not found")

    backup_dir = None if skip_backup else make_backup(input_dir)
    log = edit_parametrization(paramfile, sheets, ic_values)
    log.update({
        "template": str(template_path),
        "backup_dir": str(backup_dir) if backup_dir else None,
        "timestamp": datetime.now().isoformat(),
        "techs_sourced": sorted(ic_values.keys()),
        "applied_params": list(APPLIED_PARAMS),
    })
    if backup_dir is not None:
        log_path = backup_dir.parent / f"{backup_dir.name}_CHANGES.json"
        log_path.write_text(json.dumps(log, indent=2, default=str))
        log["log_path"] = str(log_path)
    return log


def print_summary(log: dict) -> None:
    bar = "=" * 72
    print(bar)
    print("apply_interconnector_costs — v18 Interconnector_Params -> model")
    print(bar)
    print(f"Template (source): {log.get('template')}")
    print(f"Backup folder    : {log.get('backup_dir', '(skipped)')}")
    print(f"Techs sourced    : {len(log.get('techs_sourced', []))}")
    for s in log.get("sheets", []):
        if "skipped" in s:
            print(f"[SKIPPED] '{s['sheet']}': {s['skipped']}")
            continue
        print(f"Sheet '{s['sheet']}' ({s['layout']}): {len(s['changes'])} cells, "
              f"{len(s['rows_touched'])} rows, {len(s['pm_flips'])} proj-mode flips")
        for r in s["rows_touched"][:80]:
            print(f"    {r['tech']:15} {r['param']:16} -> {r['new']}")
    missing = log.get("missing_pairs", [])
    if missing:
        print(f"Sourced pairs with NO row in any target sheet ({len(missing)}): "
              + ", ".join(f"{m['tech']}/{m['param']}" for m in missing))
    if log.get("log_path"):
        print(f"\nChange log: {log['log_path']}")


# --------------------------------------------------------------------------- self-test
def run_self_test() -> int:
    import tempfile
    bar = "=" * 72
    print(bar); print("apply_interconnector_costs.py — SELF-TEST"); print(bar)
    ok = True
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        # synthetic template with Interconnector_Params (CapEx/FOM/Life)
        twb = Workbook(); tws = twb.active; tws.title = IC_SHEET
        for c, h in enumerate(["scenario", "Tech", "Parameter", "Unit", 2023, 2024, 2025], 1):
            tws.cell(row=1, column=c, value=h)
        trows = [
            ("BAU", "TRNBGDXXINDEA", "CapitalCost", "USD/kW", 380, 380, 380),
            ("BAU", "TRNBGDXXINDEA", "FixedCost", "USD/kW/yr", 5.7, 5.7, 5.7),
            ("BAU", "TRNBGDXXINDEA", "OperationalLife", "years", 40, 40, 40),
            ("BAU", "TRNINDSOLKAXX", "CapitalCost", "USD/kW", 1250, 1250, 1250),
            ("BAU", "TRNINDSOLKAXX", "OperationalLife", "years", 40, 40, 40),
        ]
        for ri, row in enumerate(trows, 2):
            for c, v in enumerate(row, 1):
                tws.cell(row=ri, column=c, value=v)
        tpath = td / TEMPLATE_NAME
        twb.save(tpath); twb.close()

        # synthetic A-O_Parametrization: Secondary Techs (year) + Fixed Horizon Parameters (scalar)
        idir = td / "A1_Outputs" / "A1_Outputs_BAU"; idir.mkdir(parents=True)
        pwb = Workbook()
        st = pwb.active; st.title = "Secondary Techs"
        for c, h in enumerate(["Tech", "Parameter", PROJ_MODE_COL, 2023, 2024, 2025], 1):
            st.cell(row=1, column=c, value=h)
        st_rows = [
            ("TRNBGDXXINDEA", "CapitalCost", "EMPTY", 292.487, 292.487, 292.487),
            ("TRNBGDXXINDEA", "FixedCost", "EMPTY", 1.024, 1.024, 1.024),
            ("TRNBGDXXINDEA", "ResidualCapacity", "User defined", 2.5, 2.5, 2.5),   # must NOT change
            ("TRNINDSOLKAXX", "CapitalCost", "EMPTY", 507.844, 507.844, 507.844),
            ("PWRSPVINDEA", "CapitalCost", "User defined", 560, 560, 560),           # non-TRN, must NOT change
        ]
        for ri, row in enumerate(st_rows, 2):
            for c, v in enumerate(row, 1):
                st.cell(row=ri, column=c, value=v)
        fhp = pwb.create_sheet("Fixed Horizon Parameters")
        for c, h in enumerate(["Tech.Type", "Tech.ID", "Tech", "Tech.Name",
                               "Parameter.ID", "Parameter", "Unit", VALUE_COL], 1):
            fhp.cell(row=1, column=c, value=h)
        fhp_rows = [
            ("Secondary", 156, "TRNBGDXXINDEA", "desc", 2, "OperationalLife", "years", 60),
            ("Secondary", 156, "TRNBGDXXINDEA", "desc", 1, "CapacityToActivityUnit", "", 31.536),  # must NOT change
            ("Secondary", 168, "TRNINDSOLKAXX", "desc", 2, "OperationalLife", "years", 60),
        ]
        for ri, row in enumerate(fhp_rows, 2):
            for c, v in enumerate(row, 1):
                fhp.cell(row=ri, column=c, value=v)
        pwb.save(idir / PARAM_FILE_NAME); pwb.close()

        run(idir, skip_backup=True, template=tpath)

        chk = load_workbook(idir / PARAM_FILE_NAME, data_only=True)
        st = chk["Secondary Techs"]; got = {}
        for r in range(2, st.max_row + 1):
            got[(st.cell(row=r, column=1).value, st.cell(row=r, column=2).value)] = \
                [st.cell(row=r, column=c).value for c in (4, 5, 6)]
        fhp = chk["Fixed Horizon Parameters"]; got_fhp = {}
        for r in range(2, fhp.max_row + 1):
            got_fhp[(fhp.cell(row=r, column=3).value, fhp.cell(row=r, column=6).value)] = \
                fhp.cell(row=r, column=8).value
        chk.close()

        def ck(cond, label):
            nonlocal ok
            print(("  PASS " if cond else "  FAIL ") + label)
            ok = ok and cond
        ck(got.get(("TRNBGDXXINDEA", "CapitalCost")) == [380, 380, 380], "CapEx 292->380 (Secondary Techs)")
        ck(got.get(("TRNBGDXXINDEA", "FixedCost")) == [5.7, 5.7, 5.7], "FOM overwritten")
        ck(got.get(("TRNBGDXXINDEA", "ResidualCapacity")) == [2.5, 2.5, 2.5], "ResidualCapacity UNCHANGED")
        ck(got.get(("TRNINDSOLKAXX", "CapitalCost")) == [1250, 1250, 1250], "submarine CapEx overwritten")
        ck(got.get(("PWRSPVINDEA", "CapitalCost")) == [560, 560, 560], "non-TRN UNCHANGED")
        ck(got_fhp.get(("TRNBGDXXINDEA", "OperationalLife")) == 40, "OperationalLife 60->40 (Fixed Horizon Params)")
        ck(got_fhp.get(("TRNINDSOLKAXX", "OperationalLife")) == 40, "submarine life 60->40")
        ck(got_fhp.get(("TRNBGDXXINDEA", "CapacityToActivityUnit")) == 31.536, "CapacityToActivityUnit UNCHANGED")
    print(bar); print("SELF-TEST", "PASSED" if ok else "FAILED"); print(bar)
    return 0 if ok else 1


# --------------------------------------------------------------------------- main
def main() -> int:
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--input-dir", type=Path, default=Path("A1_Outputs/A1_Outputs_BAU"))
    p.add_argument("--template", type=Path, default=None)
    p.add_argument("--sheets", nargs="+", default=DEFAULT_TARGET_SHEETS)
    p.add_argument("--skip-backup", action="store_true")
    p.add_argument("--self-test", action="store_true")
    p.add_argument("--restore", action="store_true")
    p.add_argument("--restore-from", type=Path, default=None)
    args = p.parse_args()

    if args.self_test:
        return run_self_test()
    if args.restore or args.restore_from is not None:
        try:
            used = restore_from_backup(args.input_dir, args.restore_from)
        except Exception as exc:
            print(f"ERROR: {exc}", file=sys.stderr); return 1
        print(f"Restored {args.input_dir} from {used}"); return 0
    try:
        log = run(args.input_dir, args.sheets, args.skip_backup, args.template)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr); return 1
    print_summary(log)
    return 0


if __name__ == "__main__":
    sys.exit(main())
