"""
apply_base_year_pin.py
======================

WS-4 base-year lock. Pins the base window (default 2023-2026) so ALL scenarios
reproduce the calibrated reference run identically — BOTH generation AND capacity
— in those years; divergence only from 2027+.

Why both generation and capacity
---------------------------------
Pinning activity alone (Lower==Upper) forces identical *generation*, but if the
capacity build is left free (e.g. investment ceiling relaxed) a scenario can
build capacity *ahead* in the base window for its own 2027+ future (e.g.
C_Target_VRE building VRE early), so 2023-2026 *capacity* would differ. To make
the base window truly identical we pin the build too.

Mechanism (per pin year, in "Primary Techs" + "Secondary Techs")
----------------------------------------------------------------
Reads the calibrated A_Calibrated_BAU solve outputs:
  * TotalTechnologyAnnualActivity.csv  -> activity reference
  * NewCapacity.csv                    -> annual new-build reference
and sets:
  TotalTechnologyAnnualActivityLowerLimit = UpperLimit = activity_ref   (generation pinned)
  TotalAnnualMaxCapacityInvestment = TotalAnnualMinCapacityInvestment = newbuild_ref  (build pinned)
  TotalAnnualMaxCapacity = 9999   (ceiling relaxed so the pinned build is never blocked;
                                   harmless, since the build is pinned below it)
Projection.Mode is flipped EMPTY->"User defined" on every written row (otherwise
B1/otoole ignores year-indexed rows).

Applied to every scenario (incl. the reference, which reproduces itself) -> the
2023-2026 state is byte-identical across scenarios. 2027+ rows are untouched.

EXCLUSION: the 18 cross-border interconnectors (INTERCONNECTORS, == cap_trn_to_residual
.TRN_TECHS) are skipped ENTIRELY. cap_trn_to_residual freezes their TotalAnnualMaxCapacity
== ResidualCapacity (zero headroom), so forcing a base-year MinCapacityInvestment on them
makes Residual + Sum(MinCapInvest) exceed MaxCapacity in a later year -> GLPK --check fails.
That is safe: the freeze already makes interconnector capacity identical across scenarios
(verified A==B==C), so not pinning them creates no base-year difference. The internal-
transmission families (PWRTRN/RNWTRN/DSPTRN/RNWNLI/RNWRPO/TRNNLI/TRNRPO) are uncapped and
are NOT excluded -> they get the full activity+build pin so their base-year capacity is
identical by construction too.

ORDERING: the reference is the calibrated solve, which changes with other input
edits (e.g. the 3% loss) -> run this AFTER a first CPLEX solve of the reference
scenario, then re-compile (B1) + re-solve.

USAGE
    python apply_base_year_pin.py --input-dir A1_Outputs/A1_Outputs_B_Optimised_VRE \\
        --from-solve-dir Executables/A_Calibrated_BAU_0/Outputs
    python apply_base_year_pin.py --self-test
    python apply_base_year_pin.py --input-dir <dir> --restore
"""
from __future__ import annotations

import argparse
import csv
import json
import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook, Workbook

PARAM_FILE = "A-O_Parametrization.xlsx"
SHEETS = ("Primary Techs", "Secondary Techs")
DEFAULT_YEARS = (2023, 2024, 2025, 2026)
SENTINEL = 9999

P_LOWER = "TotalTechnologyAnnualActivityLowerLimit"
P_UPPER = "TotalTechnologyAnnualActivityUpperLimit"
P_MAXCAP = "TotalAnnualMaxCapacity"
P_MAXCAPINV = "TotalAnnualMaxCapacityInvestment"
P_MINCAPINV = "TotalAnnualMinCapacityInvestment"
ACTIVITY_PARAMS = (P_LOWER, P_UPPER)
BUILD_PARAMS = (P_MAXCAPINV, P_MINCAPINV)
BACKUP_TAG = "_PRE_BASEYEAR_PIN_"

ACTIVITY_CSV = "TotalTechnologyAnnualActivity.csv"
NEWCAP_CSV = "NewCapacity.csv"

# The 18 cross-border interconnectors (mirrors cap_trn_to_residual.TRN_TECHS,
# which is the source of truth). Their TotalAnnualMaxCapacity is FROZEN ==
# ResidualCapacity by cap_trn_to_residual (zero headroom), so forcing a base-year
# MinCapacityInvestment on them makes Residual + Sum(MinCapInvest) exceed
# MaxCapacity in a later year -> GLPK --check fails ("Residual, Total annual maxcap
# and mincap investments"). They are already scenario-independent (the freeze makes
# their build identical across A/B/C), so they are EXCLUDED from the pin entirely.
# NOTE: this is ONLY the 18 interconnectors, NOT the internal-transmission families
# (PWRTRN/RNWTRN/DSPTRN/…), which are uncapped and therefore safely pinned.
INTERCONNECTORS = frozenset({
    "TRNBGDXXINDEA", "TRNBGDXXINDNE", "TRNBTNXXBGDXX", "TRNBTNXXINDEA",
    "TRNBTNXXINDNE", "TRNINDEAINDNE", "TRNINDEAINDNO", "TRNINDEAINDSO",
    "TRNINDEAINDWE", "TRNINDEANPLXX", "TRNINDNEINDNO", "TRNINDNOINDWE",
    "TRNINDNONPLXX", "TRNINDSOINDWE", "TRNINDSOLKAXX", "TRNLKAXXMDVXX",
    "TRNMDVXXINDSO", "TRNNPLXXBGDXX",
})


def read_ref(path: Path, years) -> dict:
    """{tech: {year: value}} for the pin years, from an otoole result CSV."""
    out = {}
    with open(path, newline="", encoding="utf-8") as f:
        r = csv.DictReader(f)
        cols = {c.lower(): c for c in (r.fieldnames or [])}
        tcol, ycol, vcol = cols.get("technology"), cols.get("year"), cols.get("value")
        if not (tcol and ycol and vcol):
            raise ValueError(f"{path} needs TECHNOLOGY/YEAR/VALUE columns; got {r.fieldnames}")
        for row in r:
            try:
                y = int(float(row[ycol]))
            except (TypeError, ValueError):
                continue
            if y in years:
                out.setdefault(row[tcol], {})[y] = float(row[vcol])
    return out


def read_techs(path: Path) -> set:
    """Valid TECHNOLOGY set from a one-column TECHNOLOGY.csv (skip header)."""
    out = set()
    with open(path, newline="", encoding="utf-8") as f:
        for i, row in enumerate(csv.reader(f)):
            if i == 0 or not row or not str(row[0]).strip():
                continue
            out.add(str(row[0]).strip())
    return out


def _hdr(ws):
    return {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)
            if ws.cell(row=1, column=c).value is not None}


def _year_cols(ws, years):
    out = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        yi = v if isinstance(v, int) else (
            int(float(v)) if isinstance(v, str) and v.strip().split(".")[0].isdigit() else None)
        if yi in years:
            out[yi] = c
    return out


def edit_sheet(ws, activity, newbuild, years, valid_techs, log, interconnectors=INTERCONNECTORS):
    h = _hdr(ws)
    tc, pc, pmc = h.get("Tech"), h.get("Parameter"), h.get("Projection.Mode")
    if tc is None or pc is None:
        return
    ycols = _year_cols(ws, set(years))
    if not ycols:
        return

    def set_cells(r, ref_or_val, counter):
        """ref_or_val: a {year:val} dict (per-tech ref) or a scalar (e.g. SENTINEL)."""
        wrote = False
        for y, c in ycols.items():
            val = ref_or_val[y] if isinstance(ref_or_val, dict) else ref_or_val
            val = float(val) if val is not None else 0.0
            cell = ws.cell(row=r, column=c)
            if cell.value != val:
                cell.value = val
                log[counter] += 1
                wrote = True
        # a year-indexed row is only compiled by B1 if Projection.Mode is set
        if pmc is not None:
            pm = ws.cell(row=r, column=pmc)
            if pm.value in (None, "", "EMPTY"):
                pm.value = "User defined"
                log["pm_flips"] += 1
        return wrote

    for r in range(2, ws.max_row + 1):
        tech = ws.cell(row=r, column=tc).value
        par = ws.cell(row=r, column=pc).value
        if not isinstance(tech, str):
            continue
        if valid_techs is not None and tech not in valid_techs:
            continue   # only pin real technologies; never activate fuel/template rows
        if interconnectors and tech in interconnectors:
            log["ic_rows_skipped"] += 1
            continue   # interconnectors: MaxCapacity frozen==residual; pinning collides
                       # with the residual machinery (and they're already A==B==C)
        if par in ACTIVITY_PARAMS:
            set_cells(r, {y: activity.get(tech, {}).get(y, 0.0) for y in years}, "pin_cells")
        elif par in BUILD_PARAMS:
            set_cells(r, {y: newbuild.get(tech, {}).get(y, 0.0) for y in years}, "build_cells")
        elif par == P_MAXCAP:
            set_cells(r, SENTINEL, "relax_cells")


def make_backup(d: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    b = d.parent / f"{d.name}{BACKUP_TAG}{stamp}"
    if b.exists():
        raise FileExistsError(b)
    shutil.copytree(d, b)
    return b


def restore(d: Path, frm=None) -> Path:
    d = Path(d)
    if frm is None:
        c = sorted(p for p in d.parent.iterdir()
                   if p.is_dir() and p.name.startswith(f"{d.name}{BACKUP_TAG}"))
        if not c:
            raise FileNotFoundError(f"no {BACKUP_TAG}* backup by {d}")
        frm = c[-1]
    if d.is_dir():
        shutil.rmtree(d)
    shutil.copytree(frm, d)
    return frm


def run(input_dir, from_solve_dir, tech_csv=None, years=DEFAULT_YEARS, skip_backup=False) -> dict:
    input_dir = Path(input_dir)
    sd = Path(from_solve_dir)
    activity = read_ref(sd / ACTIVITY_CSV, set(years))
    newbuild = read_ref(sd / NEWCAP_CSV, set(years))
    valid_techs = read_techs(Path(tech_csv)) if tech_csv else None
    pf = input_dir / PARAM_FILE
    if not pf.exists():
        raise FileNotFoundError(pf)
    backup = None if skip_backup else make_backup(input_dir)
    log = {"input_dir": str(input_dir), "from_solve_dir": str(sd), "years": list(years),
           "activity_techs": len(activity), "newbuild_techs": len(newbuild),
           "valid_techs": len(valid_techs) if valid_techs is not None else None,
           "pin_cells": 0, "build_cells": 0, "relax_cells": 0, "pm_flips": 0,
           "ic_rows_skipped": 0,
           "timestamp": datetime.now().isoformat(), "backup_dir": str(backup) if backup else None}
    wb = load_workbook(pf)
    try:
        for sh in SHEETS:
            if sh in wb.sheetnames:
                edit_sheet(wb[sh], activity, newbuild, years, valid_techs, log)
        wb.save(pf)
    finally:
        wb.close()
    if backup is not None:
        p = backup.parent / f"{backup.name}_CHANGES.json"
        p.write_text(json.dumps(log, indent=2, default=str))
        log["log_path"] = str(p)
    return log


def print_summary(log):
    print("=" * 70)
    print(f"apply_base_year_pin  years={log['years']}")
    print("=" * 70)
    print(f"  activity ref techs : {log['activity_techs']}   newbuild ref techs: {log['newbuild_techs']}")
    print(f"  activity pin cells : {log['pin_cells']}")
    print(f"  new-build pin cells: {log['build_cells']}")
    print(f"  maxcap relax cells : {log['relax_cells']}")
    print(f"  proj-mode flips    : {log['pm_flips']}")
    print(f"  interconnector rows skipped (excluded): {log['ic_rows_skipped']}")
    print(f"  backup             : {log.get('backup_dir', '(skipped)')}")


def run_self_test() -> int:
    import tempfile
    print("=" * 70); print("apply_base_year_pin.py SELF-TEST"); print("=" * 70)
    ok = True
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        sd = td / "Outputs"; sd.mkdir()
        (sd / ACTIVITY_CSV).write_text(
            "REGION,TECHNOLOGY,YEAR,VALUE\n"
            "R,PWRCOAINDWE,2023,100\nR,PWRCOAINDWE,2024,110\nR,PWRCOAINDWE,2027,999\n"
            "R,PWRSPVINDWE,2024,50\n"
            "R,TRNINDNOINDWE,2024,500\n"   # interconnector ref (must be IGNORED — excluded)
            "R,PWRTRNINDWE,2024,300\n", encoding="utf-8")  # internal-tx ref (must be PINNED)
        (sd / NEWCAP_CSV).write_text(
            "REGION,TECHNOLOGY,YEAR,VALUE\n"
            "R,PWRCOAINDWE,2023,7\nR,PWRCOAINDWE,2024,3\n"
            "R,PWRSPVINDWE,2024,20\n"
            "R,TRNINDNOINDWE,2024,9\n"     # interconnector build ref (must be IGNORED)
            "R,PWRTRNINDWE,2024,4\n", encoding="utf-8")   # internal-tx build ref (must be PINNED)
        tech_csv = td / "TECHNOLOGY.csv"
        # TRNINDNOINDWE + PWRTRNINDWE are valid techs; ELC* fuel deliberately absent
        tech_csv.write_text("VALUE\nPWRCOAINDWE\nPWRSPVINDWE\nTRNINDNOINDWE\nPWRTRNINDWE\n", encoding="utf-8")
        idir = td / "A1_Outputs_X"; idir.mkdir()
        wb = Workbook(); ws = wb.active; ws.title = "Secondary Techs"
        for c, hh in enumerate(["Tech", "Parameter", 2023, 2024, 2025, 2026, 2027, "Projection.Mode"], 1):
            ws.cell(row=1, column=c, value=hh)
        rows = [
            ("PWRCOAINDWE", P_UPPER, None, None, None, None, None, "EMPTY"),
            ("PWRCOAINDWE", P_LOWER, None, None, None, None, None, "EMPTY"),
            ("PWRCOAINDWE", P_MAXCAPINV, 9999, 9999, 9999, 9999, 9999, "EMPTY"),
            ("PWRCOAINDWE", P_MINCAPINV, 0, 0, 0, 0, 0, "EMPTY"),
            ("PWRCOAINDWE", P_MAXCAP, 5, 5, 5, 5, 5, "EMPTY"),
            ("PWRSPVINDWE", P_UPPER, 2, 2, 2, 2, 2, "User defined"),  # C-style tight cap
            ("PWRSPVINDWE", P_MAXCAPINV, 1, 1, 1, 1, 1, "User defined"),
            ("PWRSPVINDWE", P_MINCAPINV, None, None, None, None, None, "EMPTY"),
            ("ELCBGDXX01", P_MAXCAPINV, None, None, None, None, None, "EMPTY"),  # FUEL row: must stay untouched
            ("ELCBGDXX01", P_UPPER, None, None, None, None, None, "EMPTY"),
            # interconnector: EXCLUDED -> activity/build/maxcap must all stay untouched
            ("TRNINDNOINDWE", P_UPPER, None, None, None, None, None, "EMPTY"),
            ("TRNINDNOINDWE", P_MINCAPINV, 0, 0, 0, 0, 0, "EMPTY"),
            ("TRNINDNOINDWE", P_MAXCAP, 36, 36, 36, 36, 36, "EMPTY"),
            # internal-tx: NOT excluded -> gets the full activity + build pin
            ("PWRTRNINDWE", P_UPPER, None, None, None, None, None, "EMPTY"),
            ("PWRTRNINDWE", P_MINCAPINV, None, None, None, None, None, "EMPTY"),
        ]
        for ri, row in enumerate(rows, 2):
            for c, v in enumerate(row, 1):
                ws.cell(row=ri, column=c, value=v)
        wb.save(idir / PARAM_FILE); wb.close()

        run(idir, sd, tech_csv=tech_csv, years=(2023, 2024, 2025, 2026), skip_backup=True)

        chk = load_workbook(idir / PARAM_FILE, data_only=True).active
        g = {}; gpm = {}
        for r in range(2, chk.max_row + 1):
            k = (chk.cell(row=r, column=1).value, chk.cell(row=r, column=2).value)
            g[k] = [chk.cell(row=r, column=c).value for c in (3, 4, 5, 6, 7)]  # 2023..2027
            gpm[k] = chk.cell(row=r, column=8).value

        def ck(cond, label):
            nonlocal ok; print(("  PASS " if cond else "  FAIL ") + label); ok = ok and cond
        ck(g[("PWRCOAINDWE", P_UPPER)][:4] == [100, 110, 0, 0], "COA activity Upper pinned 100/110/0/0")
        ck(g[("PWRCOAINDWE", P_LOWER)][:4] == [100, 110, 0, 0], "COA activity Lower == Upper")
        ck(g[("PWRCOAINDWE", P_MAXCAPINV)][:4] == [7, 3, 0, 0], "COA new-build MaxInv pinned 7/3/0/0")
        ck(g[("PWRCOAINDWE", P_MINCAPINV)][:4] == [7, 3, 0, 0], "COA new-build MinInv == MaxInv (forced)")
        ck(g[("PWRCOAINDWE", P_MAXCAP)][:4] == [9999, 9999, 9999, 9999], "COA MaxCapacity ceiling relaxed")
        ck(g[("PWRCOAINDWE", P_MAXCAP)][4] == 5, "2027 (outside window) untouched")
        ck(g[("PWRSPVINDWE", P_UPPER)][:4] == [0, 50, 0, 0], "SPV activity pinned 0/50/0/0 (overrides tight cap)")
        ck(g[("PWRSPVINDWE", P_MAXCAPINV)][:4] == [0, 20, 0, 0], "SPV new-build pinned 0/20/0/0")
        ck(gpm[("PWRCOAINDWE", P_UPPER)] == "User defined", "PM EMPTY->User defined on pinned row")
        ck(gpm[("ELCBGDXX01", P_MAXCAPINV)] == "EMPTY", "FUEL row Projection.Mode NOT flipped (stays EMPTY)")
        ck(g[("ELCBGDXX01", P_MAXCAPINV)][:4] == [None, None, None, None], "FUEL row values untouched")
        # interconnector excluded entirely (activity ref 500 & build ref 9 both ignored)
        ck(g[("TRNINDNOINDWE", P_UPPER)][:4] == [None, None, None, None], "INTERCONNECTOR activity NOT pinned (excluded)")
        ck(g[("TRNINDNOINDWE", P_MINCAPINV)][:4] == [0, 0, 0, 0], "INTERCONNECTOR build NOT forced (excluded)")
        ck(g[("TRNINDNOINDWE", P_MAXCAP)][:4] == [36, 36, 36, 36], "INTERCONNECTOR MaxCapacity NOT relaxed (excluded)")
        ck(gpm[("TRNINDNOINDWE", P_UPPER)] == "EMPTY", "INTERCONNECTOR Projection.Mode NOT flipped")
        # internal-tx pinned normally (activity 0/300/0/0, build 0/4/0/0)
        ck(g[("PWRTRNINDWE", P_UPPER)][:4] == [0, 300, 0, 0], "INTERNAL-TX activity pinned 0/300/0/0")
        ck(g[("PWRTRNINDWE", P_MINCAPINV)][:4] == [0, 4, 0, 0], "INTERNAL-TX build pinned 0/4/0/0")
        log2 = run(idir, sd, tech_csv=tech_csv, skip_backup=True)
        ck(log2["pin_cells"] == 0 and log2["build_cells"] == 0 and log2["relax_cells"] == 0
           and log2["pm_flips"] == 0, "idempotent (2nd run 0 changes)")
    print("=" * 70); print("SELF-TEST", "PASSED" if ok else "FAILED"); print("=" * 70)
    return 0 if ok else 1


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--input-dir", type=Path, default=None)
    p.add_argument("--from-solve-dir", type=Path, default=None,
                   help="reference solve Outputs dir (has TotalTechnologyAnnualActivity.csv + NewCapacity.csv)")
    p.add_argument("--tech-csv", type=Path, default=None,
                   help="TECHNOLOGY.csv (valid tech set); the pin only touches these, never fuel/template rows")
    p.add_argument("--years", type=int, nargs="+", default=list(DEFAULT_YEARS))
    p.add_argument("--skip-backup", action="store_true")
    p.add_argument("--self-test", action="store_true")
    p.add_argument("--restore", action="store_true")
    p.add_argument("--restore-from", type=Path, default=None)
    a = p.parse_args()
    if a.self_test:
        return run_self_test()
    if a.restore or a.restore_from is not None:
        try:
            used = restore(a.input_dir, a.restore_from)
        except Exception as e:
            print(f"ERROR: {e}", file=sys.stderr); return 1
        print(f"Restored {a.input_dir} from {used}"); return 0
    if not a.input_dir or not a.from_solve_dir:
        print("ERROR: --input-dir and --from-solve-dir are required", file=sys.stderr); return 1
    try:
        log = run(a.input_dir, a.from_solve_dir, a.tech_csv, tuple(a.years), a.skip_backup)
    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr); return 1
    print_summary(log)
    return 0


if __name__ == "__main__":
    sys.exit(main())
