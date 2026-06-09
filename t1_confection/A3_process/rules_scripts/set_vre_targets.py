"""
set_vre_targets.py
==================

Independent patch — sets minimum renewable generation (or capacity)
floors per country-region and technology, expressed as absolute values
derived from a prior BAU solve's total production.

The user supplies percentage targets in a YAML file (e.g. "India-East
must have >= 30% solar by 2030"). The script reads total production
from a completed BAU run's ProductionByTechnology.csv, multiplies by
the target percentage, and writes TotalTechnologyAnnualActivityLowerLimit
and/or TotalAnnualMinCapacityInvestment values into the parametrization
workbook.

HARD DEPENDENCY: A completed BAU solve must exist at the path specified
in the YAML config. The script fails loudly if it's missing.

RULE
----
For each target entry in the YAML:
    1. Read BAU total production for the country-region per year.
    2. Interpolate the target percentage schedule to all horizon years.
    3. Compute floor = total_prod × pct (capped by max_floor_share).
    4. Write the floor into the workbook.

CONFIGURATION
-------------
Edit set_vre_targets.yaml (next to this script).

OUTPUT
------
1. Timestamped backup of the input directory.
2. In-place edit of A-O_Parametrization.xlsx.
3. A JSON change log next to the backup.

USAGE
-----
    python set_vre_targets.py
    python set_vre_targets.py --input-dir A1_Outputs/A1_Outputs_BAU
    python set_vre_targets.py --self-test
    python set_vre_targets.py --restore
"""

from __future__ import annotations

import argparse
import gc
import json
import shutil
import sys
import tempfile
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook, Workbook

try:
    import yaml as _yaml
    def _load_yaml(path: Path) -> dict:
        with open(path, "r", encoding="utf-8") as f:
            return _yaml.safe_load(f)
except ImportError:
    _yaml = None
    def _load_yaml(path: Path) -> dict:
        raise ImportError("PyYAML is required. Install with: pip install pyyaml")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
DEFAULT_TARGET_SHEETS = ["Secondary Techs"]
PARAM_FILE_NAME = "A-O_Parametrization.xlsx"
YAML_FILE_NAME = "set_vre_targets.yaml"

RES_PARAM = "ResidualCapacity"
MIN_INV_PARAM = "TotalAnnualMinCapacityInvestment"
MAX_INV_PARAM = "TotalAnnualMaxCapacityInvestment"
MAX_CAP_PARAM = "TotalAnnualMaxCapacity"
ACTIVITY_LOWER_PARAM = "TotalTechnologyAnnualActivityLowerLimit"
ACTIVITY_UPPER_PARAM = "TotalTechnologyAnnualActivityUpperLimit"

# Capacity column in the combined B2 output (used to derive realized yield
# PJ/GW for the cap_envelope feature).
PROD_CAP_COL = "TotalCapacityAnnual"

PROJ_MODE_COL = "Projection.Mode"
PROJ_MODE_EMPTY = "EMPTY"
PROJ_MODE_USER = "User defined"

TECH_TYPES_FILE = "TECH_TYPES.csv"
TECH_TYPES_CATEGORY_COL = "Technology (PWR)"
TECH_TYPES_TECH_COL = "Technology"
GENERATION_CATEGORY = "GENERATION"

PWR_TECH_LENGTH = 11
COUNTRY_REGION_SLICE = slice(6, 11)

C2AU = 31.536  # CapacityToActivityUnit: GW → PJ/year

UNTIE_MULTIPLIER = 1.01

BACKUP_TAG = "_PRE_VRE_TARGETS_"

# ---------------------------------------------------------------------------
# BAU production source — column names and accepted file layouts
# ---------------------------------------------------------------------------
# Two layouts are supported transparently:
#   1. otoole's separate-parameter file 'ProductionByTechnology.csv'
#      (long format, production in the "VALUE" column).
#   2. the combined B2 solver output 'Pre_processed_<scenario>_..._output.csv'
#      (wide format, production in the "ProductionByTechnologyAnnual" column,
#      or "ProductionByTechnology" at timeslice level).
PROD_CSV_NAME = "ProductionByTechnology.csv"
PROD_COL_TECH = "TECHNOLOGY"
PROD_COL_YEAR = "YEAR"
PROD_COL_VALUE = "VALUE"

# Production-value columns, in priority order. The first one present wins.
PROD_VALUE_CANDIDATES = (
    "VALUE",                          # otoole separate ProductionByTechnology.csv
    "ProductionByTechnologyAnnual",   # combined B2 output (annual)
    "ProductionByTechnology",         # combined B2 output (timeslice-level)
)

# Glob patterns used to locate the combined B2 output inside a directory
# when no otoole ProductionByTechnology.csv is present.
COMBINED_OUTPUT_GLOBS = ("*output.csv", "Pre_processed*.csv")


# ---------------------------------------------------------------------------
# YAML config loader
# ---------------------------------------------------------------------------
def load_config(yaml_path: Path) -> dict:
    """Load and validate the YAML configuration."""
    cfg = _load_yaml(yaml_path)
    if cfg is None:
        cfg = {}

    bau_results_path = cfg.get("bau_results_path", "")
    constraint_type = cfg.get("constraint_type", "activity")
    if constraint_type not in ("activity", "capacity", "both"):
        raise ValueError(
            f"Invalid constraint_type '{constraint_type}'. "
            f"Expected 'activity', 'capacity', or 'both'."
        )

    max_floor_share = float(cfg.get("max_floor_share", 0.80))

    # cap_envelope feature: when a target sets cap_envelope: true, write a
    # bounded MaxCap/MaxCapInv envelope derived from the floor's capacity
    # trajectory (instead of exempting the tech in the lid). headroom is the
    # fractional slack above the floor-implied capacity; min_inv_gw is the
    # minimum per-year build allowance so small/early adjustments stay feasible.
    cap_envelope_headroom = float(cfg.get("cap_envelope_headroom", 0.20))
    cap_envelope_min_inv_gw = float(cfg.get("cap_envelope_min_inv_gw", 2.0))
    default_capacity_factor = float(cfg.get("default_capacity_factor", 0.10))

    # pin_generation_to_target: when true, cap_envelope targets also get a
    # TotalTechnologyAnnualActivityUpperLimit = max(target_gen, CalBAU_gen),
    # which pins generation AT the NDC share (not above it) while leaving the
    # capacity envelope generous — so the solver keeps its room (no degenerate
    # narrow capacity band). The max() guard prevents force-curtailing planned
    # capacity in regions whose baseline already exceeds the early floor.
    pin_generation_to_target = bool(cfg.get("pin_generation_to_target", False))

    # pin_slack: small fractional gap between the activity lower and upper
    # limits so generation is pinned to a NARROW RANGE, not an exact equality.
    # upper = max(target_gen * (1 + pin_slack), CalBAU_gen). An exact equality
    # (slack 0) is a fixed variable / zero-width interval and can be a numeric
    # knife-edge; ~1% gives the solver a real interval at negligible overshoot.
    pin_slack = float(cfg.get("pin_slack", 0.01))

    # Parse targets
    raw_targets = cfg.get("targets", []) or []
    targets = []
    for entry in raw_targets:
        cr = str(entry.get("cr", ""))
        tech = str(entry.get("tech", ""))
        schedule = entry.get("schedule", {})
        if not cr or not tech or not schedule:
            raise ValueError(f"Target entry missing cr/tech/schedule: {entry}")
        schedule = {int(y): float(v) for y, v in schedule.items()}
        cap_envelope = bool(entry.get("cap_envelope", False))
        targets.append({"cr": cr, "tech": tech, "schedule": schedule,
                        "cap_envelope": cap_envelope})

    return {
        "bau_results_path": bau_results_path,
        "constraint_type": constraint_type,
        "max_floor_share": max_floor_share,
        "cap_envelope_headroom": cap_envelope_headroom,
        "cap_envelope_min_inv_gw": cap_envelope_min_inv_gw,
        "default_capacity_factor": default_capacity_factor,
        "pin_generation_to_target": pin_generation_to_target,
        "pin_slack": pin_slack,
        "targets": targets,
    }


# ---------------------------------------------------------------------------
# BAU results loader
# ---------------------------------------------------------------------------
def _resolve_bau_csv(bau_results_path: Path) -> Path:
    """Resolve the BAU production CSV from a file OR directory path.

    Accepts, in order of preference:
      * a direct path to a .csv file              -> used as-is
      * a directory holding 'ProductionByTechnology.csv' (otoole layout)
      * a directory holding the combined B2 output ('*output.csv')

    Raises a clear error (rather than silently guessing) if a directory
    contains several candidate output CSVs.
    """
    p = Path(bau_results_path)

    if p.is_file():
        return p

    if p.is_dir():
        otoole = p / PROD_CSV_NAME
        if otoole.is_file():
            return otoole
        candidates: list = []
        for pattern in COMBINED_OUTPUT_GLOBS:
            for hit in sorted(p.glob(pattern)):
                if hit.is_file() and hit not in candidates:
                    candidates.append(hit)
        if len(candidates) == 1:
            return candidates[0]
        if len(candidates) > 1:
            raise FileNotFoundError(
                f"Multiple candidate output CSVs in {p}: "
                f"{[c.name for c in candidates]}. "
                f"Set bau_results_path to the exact file."
            )

    raise FileNotFoundError(
        f"BAU results not found at {bau_results_path}. Expected a directory "
        f"containing '{PROD_CSV_NAME}' or a combined '*output.csv', or a "
        f"direct path to the CSV. Run the BAU/CalBAU scenario first."
    )


def load_bau_total_production(bau_results_path: Path, gen_techs: set) -> dict:
    """Read BAU production and return {(cr, year): total_pj}.

    Aggregates production across electricity GENERATION technologies (per
    TECH_TYPES.csv) per country-region per year. Country-region is extracted
    from the TECHNOLOGY column (chars 6..11 for 11-char PWR* codes).

    Reads either otoole's 'ProductionByTechnology.csv' or the combined B2
    solver output transparently (see PROD_VALUE_CANDIDATES). Non-generation
    flows (DSPTRN distribution, TRN* transport, RNW* accounting, storage) are
    excluded so the denominator reflects only electricity generated.
    """
    csv_path = _resolve_bau_csv(bau_results_path)

    df = pd.read_csv(csv_path)

    # Locate the production-value column.
    value_col = next((c for c in PROD_VALUE_CANDIDATES if c in df.columns), None)
    if (PROD_COL_TECH not in df.columns or PROD_COL_YEAR not in df.columns
            or value_col is None):
        raise ValueError(
            f"{csv_path.name} missing required columns. Need '{PROD_COL_TECH}', "
            f"'{PROD_COL_YEAR}', and one of {PROD_VALUE_CANDIDATES}. "
            f"Found: {list(df.columns)}"
        )

    df = df.dropna(subset=[PROD_COL_TECH, PROD_COL_YEAR, value_col]).copy()
    df[PROD_COL_TECH] = df[PROD_COL_TECH].astype(str)
    df[PROD_COL_YEAR] = df[PROD_COL_YEAR].astype(float).astype(int)

    # The combined output carries an annual value that may repeat across the
    # wide file's timeslice rows. Collapse to one value per (tech, year) before
    # summing so timeslices are never double-counted. The otoole long format
    # ("VALUE") and the timeslice-level column are summed as-is.
    if value_col == "ProductionByTechnologyAnnual":
        df = df.groupby([PROD_COL_TECH, PROD_COL_YEAR],
                        as_index=False)[value_col].max()

    # Keep only electricity generation techs (length-11 PWR* in the
    # GENERATION category from TECH_TYPES.csv).
    df = df[df[PROD_COL_TECH].str.len() == PWR_TECH_LENGTH]
    df = df[df[PROD_COL_TECH].isin(gen_techs)]
    df["_cr"] = df[PROD_COL_TECH].str[COUNTRY_REGION_SLICE]

    # Aggregate: total generation per (cr, year)
    agg = df.groupby(["_cr", PROD_COL_YEAR])[value_col].sum()
    result: dict = {}
    for (cr, year), val in agg.items():
        result[(str(cr), int(year))] = float(val)
    return result


def load_bau_tech_yield(bau_results_path: Path, techs: set,
                        default_cf: float) -> tuple:
    """Return (yields, cap_traj) from CalBAU output.

    yields    : {tech: realized PJ/GW}  (CF * CapacityToActivityUnit)
    cap_traj  : {tech: {year: TotalCapacityAnnual GW}}  (CalBAU's installed path)

    The cap trajectory is used by the cap_envelope feature so the envelope is
    never sized below what CalBAU actually built — which reflects planned
    capacity (e.g. Bangladesh's planned solar) and the calibrated baseline.
    Yield falls back to default_cf * C2AU for techs CalBAU never built.
    Reads the same file as load_bau_total_production (otoole or combined B2).
    """
    fallback = default_cf * C2AU
    if not techs:
        return {}, {}
    csv_path = _resolve_bau_csv(bau_results_path)
    df = pd.read_csv(csv_path)

    prod_col = next((c for c in PROD_VALUE_CANDIDATES if c in df.columns), None)
    cap_col = PROD_CAP_COL if PROD_CAP_COL in df.columns else None
    if prod_col is None or cap_col is None:
        return {t: fallback for t in techs}, {}

    df = df.dropna(subset=[PROD_COL_TECH, PROD_COL_YEAR]).copy()
    df[PROD_COL_TECH] = df[PROD_COL_TECH].astype(str)
    df[PROD_COL_YEAR] = df[PROD_COL_YEAR].astype(float).astype(int)

    # Production and capacity sit on different rows in the wide format —
    # aggregate each to one value per (tech, year), then merge.
    p = (df.dropna(subset=[prod_col])
           .groupby([PROD_COL_TECH, PROD_COL_YEAR], as_index=False)[prod_col].max())
    c = (df.dropna(subset=[cap_col])
           .groupby([PROD_COL_TECH, PROD_COL_YEAR], as_index=False)[cap_col].max())

    # Capacity trajectory per tech (for the envelope baseline)
    cap_traj: dict = {}
    for t in techs:
        sub = c[c[PROD_COL_TECH] == t]
        cap_traj[t] = {int(r[PROD_COL_YEAR]): float(r[cap_col])
                       for _, r in sub.iterrows() if r[cap_col] > 0}

    # Realized yield = production / capacity (averaged over years with both)
    m = p.merge(c, on=[PROD_COL_TECH, PROD_COL_YEAR], how="inner")
    m = m[(m[cap_col] > 1e-6) & (m[prod_col] > 1e-9)]
    m["_yield"] = m[prod_col] / m[cap_col]
    by_tech = m.groupby(PROD_COL_TECH)["_yield"].mean()

    yields: dict = {}
    for t in techs:
        if t in by_tech.index and by_tech[t] > 0:
            yields[t] = float(by_tech[t])
        else:
            yields[t] = fallback
    return yields, cap_traj


# ---------------------------------------------------------------------------
# Interpolation
# ---------------------------------------------------------------------------
def interpolate_schedule(schedule: dict, years: list,
                         pre_first: float = 0.0) -> dict:
    """Linearly interpolate a sparse {year: value} schedule.

    Years before first key: pre_first (default 0 — no floor before
    the first target milestone).
    Years after last key: hold last value flat.
    """
    if not schedule:
        return {y: pre_first for y in years}
    sorted_keys = sorted(schedule.keys())
    result: dict = {}
    for y in years:
        if y < sorted_keys[0]:
            result[y] = pre_first
        elif y >= sorted_keys[-1]:
            result[y] = schedule[sorted_keys[-1]]
        elif y in schedule:
            result[y] = schedule[y]
        else:
            for i in range(len(sorted_keys) - 1):
                y_lo, y_hi = sorted_keys[i], sorted_keys[i + 1]
                if y_lo <= y <= y_hi:
                    frac = (y - y_lo) / (y_hi - y_lo)
                    result[y] = schedule[y_lo] + frac * (schedule[y_hi] - schedule[y_lo])
                    break
    return result


def interpolate_production(total_prod: dict, cr: str, years: list) -> dict:
    """Interpolate BAU total production for a cr across all horizon years.

    BAU results typically have annual data, but if any years are missing
    we interpolate linearly between available data points.
    """
    cr_data = {y: v for (c, y), v in total_prod.items() if c == cr}
    if not cr_data:
        return {y: 0.0 for y in years}
    return interpolate_schedule(cr_data, years, pre_first=0.0)


# ---------------------------------------------------------------------------
# Tech-type loading and pattern matching
# ---------------------------------------------------------------------------
def load_generation_techs(tech_types_path: Path) -> set:
    """Load TECH_TYPES.csv and return the set of GENERATION techs."""
    if not tech_types_path.is_file():
        raise FileNotFoundError(f"TECH_TYPES.csv not found at {tech_types_path}")
    df = pd.read_csv(tech_types_path)
    return set(df.loc[df[TECH_TYPES_CATEGORY_COL] == GENERATION_CATEGORY,
                      TECH_TYPES_TECH_COL].dropna())


def expand_tech_pattern(pattern: str, cr: str, gen_techs: set) -> list:
    """Expand a tech pattern + cr to matching tech names.

    'PWRSPV*' + 'INDEA' → ['PWRSPVINDEA'] (if it exists in gen_techs)
    'PWRSPVINDEA' (exact) → ['PWRSPVINDEA'] (if it exists)
    """
    if pattern.endswith("*"):
        prefix = pattern[:-1]  # e.g. 'PWRSPV'
        matched = [t for t in gen_techs
                   if t.startswith(prefix) and len(t) == PWR_TECH_LENGTH
                   and t[COUNTRY_REGION_SLICE] == cr]
    else:
        # Exact match
        matched = [pattern] if pattern in gen_techs else []
    return sorted(matched)


# ---------------------------------------------------------------------------
# Backup / restore
# ---------------------------------------------------------------------------
def _rmtree_robust(path: Path, attempts: int = 5) -> None:
    for i in range(attempts):
        try:
            shutil.rmtree(path)
            return
        except PermissionError:
            gc.collect()
            time.sleep(0.1 * (i + 1))
    shutil.rmtree(path)


def find_latest_backup(input_dir: Path) -> Path | None:
    parent = input_dir.parent
    candidates = sorted(
        (p for p in parent.iterdir()
         if p.is_dir() and p.name.startswith(f"{input_dir.name}{BACKUP_TAG}")),
        key=lambda p: p.name,
    )
    return candidates[-1] if candidates else None


def restore_from_backup(input_dir: Path, backup_dir: Path | None = None) -> Path:
    input_dir = Path(input_dir)
    if backup_dir is None:
        backup_dir = find_latest_backup(input_dir)
        if backup_dir is None:
            raise FileNotFoundError(
                f"No {BACKUP_TAG}* backup found next to {input_dir}."
            )
    else:
        backup_dir = Path(backup_dir)
        if not backup_dir.is_dir():
            raise FileNotFoundError(f"Backup folder does not exist: {backup_dir}")
    if input_dir.is_dir():
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        snapshot = input_dir.parent / f"{input_dir.name}_POST_VRE_TARGETS_pre_restore_{stamp}"
        if not snapshot.exists():
            shutil.copytree(input_dir, snapshot)
        _rmtree_robust(input_dir)
    shutil.copytree(backup_dir, input_dir)
    return backup_dir


def make_backup(input_dir: Path) -> Path:
    if not input_dir.is_dir():
        raise FileNotFoundError(f"Input directory does not exist: {input_dir}")
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = input_dir.parent / f"{input_dir.name}{BACKUP_TAG}{stamp}"
    if backup.exists():
        raise FileExistsError(f"Backup folder already exists: {backup}")
    shutil.copytree(input_dir, backup)
    return backup


# ---------------------------------------------------------------------------
# Worksheet helpers
# ---------------------------------------------------------------------------
def find_year_columns(ws) -> dict:
    year_to_col = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if isinstance(val, int) and 1900 <= val <= 2200:
            year_to_col[val] = col_idx
    return year_to_col


def find_named_columns(ws, names) -> dict:
    found = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val in names:
            found[val] = col_idx
    return found


def values_differ(a, b, tol: float = 1e-12) -> bool:
    if a is None and b is None:
        return False
    if a is None or b is None:
        return True
    try:
        return abs(float(a) - float(b)) > tol
    except (TypeError, ValueError):
        return a != b


# ---------------------------------------------------------------------------
# Row finder / creator
# ---------------------------------------------------------------------------
def find_param_row(ws, tech: str, param: str, tech_col: int,
                   param_col: int) -> int | None:
    """Find the row index for (tech, param). Returns None if not found."""
    for row_idx in range(2, ws.max_row + 1):
        t = ws.cell(row=row_idx, column=tech_col).value
        p = ws.cell(row=row_idx, column=param_col).value
        if t == tech and p == param:
            return row_idx
    return None


def find_or_create_param_row(ws, tech: str, param: str, tech_col: int,
                             param_col: int, proj_mode_col: int | None) -> int:
    """Find the row for (tech, param) or create it at the end of the sheet.

    When creating, sets Tech, Parameter, and Projection.Mode = EMPTY.
    """
    existing = find_param_row(ws, tech, param, tech_col, param_col)
    if existing is not None:
        return existing

    new_row = ws.max_row + 1
    ws.cell(row=new_row, column=tech_col, value=tech)
    ws.cell(row=new_row, column=param_col, value=param)
    if proj_mode_col is not None:
        ws.cell(row=new_row, column=proj_mode_col, value=PROJ_MODE_EMPTY)
    return new_row


# ---------------------------------------------------------------------------
# Core logic
# ---------------------------------------------------------------------------
def apply_vre_targets(ws, config: dict, total_prod: dict,
                      gen_techs: set, year_cols: dict,
                      tech_yield: dict = None,
                      tech_bau_cap: dict = None) -> dict:
    """Write VRE target floors into the worksheet. Returns a log dict."""
    constraint_type = config["constraint_type"]
    max_floor_share = config["max_floor_share"]
    targets = config["targets"]
    tech_yield = tech_yield or {}
    tech_bau_cap = tech_bau_cap or {}
    env_headroom = config.get("cap_envelope_headroom", 0.20)
    env_min_inv = config.get("cap_envelope_min_inv_gw", 2.0)
    default_cf = config.get("default_capacity_factor", 0.10)
    pin_gen = config.get("pin_generation_to_target", False)
    pin_slack = config.get("pin_slack", 0.01)

    headers = find_named_columns(ws, ["Tech", "Parameter", PROJ_MODE_COL])
    tech_col = headers.get("Tech")
    param_col = headers.get("Parameter")
    proj_mode_col = headers.get(PROJ_MODE_COL)
    if tech_col is None or param_col is None:
        raise ValueError(f"Sheet '{ws.title}' missing Tech/Parameter columns.")

    sorted_years = sorted(year_cols.keys())

    log = {
        "sheet": ws.title,
        "years": sorted_years,
        "constraint_type": constraint_type,
        "max_floor_share": max_floor_share,
        "targets_processed": [],
        "changes": [],
        "warnings": [],
        "projection_mode_flips": [],
    }

    # Build MaxCapInv lookup (needed for untie rule in capacity mode)
    maxinv_map: dict = {}
    if constraint_type in ("capacity", "both"):
        for row_idx in range(2, ws.max_row + 1):
            t = ws.cell(row=row_idx, column=tech_col).value
            p = ws.cell(row=row_idx, column=param_col).value
            if p != MAX_INV_PARAM:
                continue
            for year, col in year_cols.items():
                val = ws.cell(row=row_idx, column=col).value
                if val is not None:
                    try:
                        maxinv_map[(t, year)] = float(val)
                    except (TypeError, ValueError):
                        pass

    # Build CapacityFactor lookup (needed for PJ→GW conversion in capacity mode)
    cf_map: dict = {}
    if constraint_type in ("capacity", "both"):
        for row_idx in range(2, ws.max_row + 1):
            t = ws.cell(row=row_idx, column=tech_col).value
            p = ws.cell(row=row_idx, column=param_col).value
            if p != "CapacityFactor":
                continue
            for year, col in year_cols.items():
                val = ws.cell(row=row_idx, column=col).value
                if val is not None:
                    try:
                        cf_map[(t, year)] = float(val)
                    except (TypeError, ValueError):
                        pass

    # Build MinCapInv lookup (needed to clamp the cap_envelope MaxCapInv so it
    # never falls below pre-existing planned capacity, which would violate the
    # MaxCapInv >= MinCapInv solver check). Built whenever any target uses the
    # cap_envelope feature.
    has_envelope = any(t.get("cap_envelope") for t in targets)
    mincapinv_map: dict = {}
    if has_envelope:
        for row_idx in range(2, ws.max_row + 1):
            t = ws.cell(row=row_idx, column=tech_col).value
            p = ws.cell(row=row_idx, column=param_col).value
            if p != MIN_INV_PARAM:
                continue
            for year, col in year_cols.items():
                val = ws.cell(row=row_idx, column=col).value
                if val is not None:
                    try:
                        mincapinv_map[(t, year)] = float(val)
                    except (TypeError, ValueError):
                        pass

    for target in targets:
        cr = target["cr"]
        tech_pattern = target["tech"]
        schedule = target["schedule"]

        matched_techs = expand_tech_pattern(tech_pattern, cr, gen_techs)
        if not matched_techs:
            log["warnings"].append(
                f"No techs matched pattern '{tech_pattern}' in cr '{cr}'."
            )
            continue

        # Interpolate target percentages to all horizon years
        pct_by_year = interpolate_schedule(schedule, sorted_years, pre_first=0.0)

        # Interpolate BAU total production for this cr
        prod_by_year = interpolate_production(total_prod, cr, sorted_years)

        target_log = {
            "cr": cr,
            "tech_pattern": tech_pattern,
            "matched_techs": matched_techs,
            "pct_by_year": {str(y): p for y, p in pct_by_year.items()},
            "prod_by_year": {str(y): p for y, p in prod_by_year.items()},
        }
        log["targets_processed"].append(target_log)

        for tech in matched_techs:
            # --- Activity floor ---
            if constraint_type in ("activity", "both"):
                row = find_or_create_param_row(
                    ws, tech, ACTIVITY_LOWER_PARAM,
                    tech_col, param_col, proj_mode_col
                )
                row_modified = False
                for year in sorted_years:
                    col = year_cols[year]
                    cell = ws.cell(row=row, column=col)
                    old = cell.value
                    pct = pct_by_year.get(year, 0.0)
                    prod = prod_by_year.get(year, 0.0)
                    floor_pj = prod * pct

                    # Safety cap
                    cap = prod * max_floor_share
                    capped = False
                    if floor_pj > cap and cap > 0:
                        floor_pj = cap
                        capped = True
                        log["warnings"].append(
                            f"WARNING: {tech} {year}: floor {prod*pct:.2f} PJ "
                            f"exceeds {max_floor_share*100:.0f}% of total prod "
                            f"({prod:.2f} PJ). Capped to {cap:.2f} PJ."
                        )

                    if values_differ(old, floor_pj):
                        cell.value = floor_pj
                        row_modified = True
                        log["changes"].append({
                            "tech": tech, "year": year, "param": ACTIVITY_LOWER_PARAM,
                            "old": old, "new": floor_pj,
                            "pct": pct, "total_prod": prod,
                            "capped": capped,
                            "reason": "activity_floor",
                        })

                if row_modified and proj_mode_col is not None:
                    mode_cell = ws.cell(row=row, column=proj_mode_col)
                    if mode_cell.value == PROJ_MODE_EMPTY:
                        mode_cell.value = PROJ_MODE_USER
                        log["projection_mode_flips"].append(
                            {"tech": tech, "param": ACTIVITY_LOWER_PARAM}
                        )

            # --- Capacity envelope (bounded MaxCap + MaxCapInv) ---
            # When a target opts in (cap_envelope: true), bound the tech from
            # ABOVE so it follows the floor's capacity trajectory rather than
            # being exempted/uncapped. The reference each year is the MAX of:
            #   - cap_needed(y) = floor_pj(y) / yield   (the NDC requirement), and
            #   - bau_cap(y)    = CalBAU's installed capacity (reflects planned
            #     builds + the calibrated baseline)
            # so the envelope never falls below what already exists / is planned
            # (which would make the solve infeasible). MaxCapInv is additionally
            # clamped above any pre-existing MinCapInv (MaxCapInv >= MinCapInv).
            if target.get("cap_envelope"):
                yld = tech_yield.get(tech, default_cf * C2AU)
                if yld <= 0:
                    yld = default_cf * C2AU
                bau_cap = tech_bau_cap.get(tech, {})

                # reference capacity = max(floor-needed, CalBAU installed)
                ref_cap = {}
                for year in sorted_years:
                    pct = pct_by_year.get(year, 0.0)
                    prod = prod_by_year.get(year, 0.0)
                    fpj = prod * pct
                    capm = prod * max_floor_share
                    if fpj > capm and capm > 0:
                        fpj = capm
                    cap_needed = fpj / yld
                    ref_cap[year] = max(cap_needed, bau_cap.get(year, 0.0))

                maxcap_row = find_or_create_param_row(
                    ws, tech, MAX_CAP_PARAM, tech_col, param_col, proj_mode_col
                )
                maxinv_row = find_or_create_param_row(
                    ws, tech, MAX_INV_PARAM, tech_col, param_col, proj_mode_col
                )
                cap_mod = inv_mod = False
                prev_rc = None
                for year in sorted_years:
                    rc = ref_cap[year]
                    maxcap_val = rc * (1.0 + env_headroom)
                    if prev_rc is None:
                        increment = 0.0   # residual/baseline covers the first year
                    else:
                        increment = max(0.0, rc - prev_rc)
                    maxinv_val = max(env_min_inv, increment) * (1.0 + env_headroom)
                    prev_rc = rc

                    # Clamp MaxCapInv above any pre-existing planned MinCapInv
                    # (otherwise MaxCapInv < MinCapInv fails the solver check).
                    existing_min = mincapinv_map.get((tech, year))
                    if existing_min is not None and existing_min > 0:
                        floor_inv = existing_min * UNTIE_MULTIPLIER
                        if floor_inv > maxinv_val:
                            maxinv_val = floor_inv

                    ccell = ws.cell(row=maxcap_row, column=year_cols[year])
                    if values_differ(ccell.value, maxcap_val):
                        log["changes"].append({
                            "tech": tech, "year": year, "param": MAX_CAP_PARAM,
                            "old": ccell.value, "new": maxcap_val,
                            "ref_cap": rc, "headroom": env_headroom,
                            "reason": "cap_envelope_maxcap",
                        })
                        ccell.value = maxcap_val
                        cap_mod = True

                    icell = ws.cell(row=maxinv_row, column=year_cols[year])
                    if values_differ(icell.value, maxinv_val):
                        log["changes"].append({
                            "tech": tech, "year": year, "param": MAX_INV_PARAM,
                            "old": icell.value, "new": maxinv_val,
                            "reason": "cap_envelope_maxinv",
                        })
                        icell.value = maxinv_val
                        inv_mod = True

                if proj_mode_col is not None:
                    for r, mod in ((maxcap_row, cap_mod), (maxinv_row, inv_mod)):
                        if mod:
                            mc = ws.cell(row=r, column=proj_mode_col)
                            if mc.value == PROJ_MODE_EMPTY:
                                mc.value = PROJ_MODE_USER

                # --- Pin generation at the target (activity upper limit) ---
                # With a generous MaxCap (above), capacity has slack, so the
                # solver stays well-conditioned. The upper limit pins generation
                # to a NARROW RANGE at the NDC share:
                #   upper = max(target_gen * (1 + pin_slack), CalBAU_gen).
                # Where the floor binds (target > CalBAU), upper sits pin_slack
                # above the floor -> generation ~ target, but as a real interval
                # (not an exact lower==upper equality, which is a numeric
                # knife-edge). Where CalBAU exceeds the early floor (e.g. planned
                # solar), upper = CalBAU_gen so planned capacity is NOT curtailed.
                if pin_gen:
                    upper_row = find_or_create_param_row(
                        ws, tech, ACTIVITY_UPPER_PARAM,
                        tech_col, param_col, proj_mode_col
                    )
                    up_mod = False
                    for year in sorted_years:
                        pct = pct_by_year.get(year, 0.0)
                        prod = prod_by_year.get(year, 0.0)
                        floor_pj = prod * pct
                        capm = prod * max_floor_share
                        if floor_pj > capm and capm > 0:
                            floor_pj = capm
                        bau_gen = bau_cap.get(year, 0.0) * yld
                        upper_val = max(floor_pj * (1.0 + pin_slack), bau_gen)
                        ucell = ws.cell(row=upper_row, column=year_cols[year])
                        if values_differ(ucell.value, upper_val):
                            log["changes"].append({
                                "tech": tech, "year": year,
                                "param": ACTIVITY_UPPER_PARAM,
                                "old": ucell.value, "new": upper_val,
                                "floor_pj": floor_pj, "bau_gen": bau_gen,
                                "reason": "pin_generation_upper",
                            })
                            ucell.value = upper_val
                            up_mod = True
                    if up_mod and proj_mode_col is not None:
                        mc = ws.cell(row=upper_row, column=proj_mode_col)
                        if mc.value == PROJ_MODE_EMPTY:
                            mc.value = PROJ_MODE_USER

            # --- Capacity floor ---
            if constraint_type in ("capacity", "both"):
                row = find_or_create_param_row(
                    ws, tech, MIN_INV_PARAM,
                    tech_col, param_col, proj_mode_col
                )
                row_modified = False
                for year in sorted_years:
                    col = year_cols[year]
                    cell = ws.cell(row=row, column=col)
                    old = cell.value
                    pct = pct_by_year.get(year, 0.0)
                    prod = prod_by_year.get(year, 0.0)
                    floor_pj = prod * pct

                    # Safety cap
                    cap = prod * max_floor_share
                    if floor_pj > cap and cap > 0:
                        floor_pj = cap

                    # Convert PJ → GW: floor_gw = floor_pj / (C2AU × CF)
                    cf = cf_map.get((tech, year), 1.0)  # default CF=1.0 if missing
                    if cf <= 0:
                        cf = 1.0
                    floor_gw = floor_pj / (C2AU * cf)

                    if values_differ(old, floor_gw):
                        cell.value = floor_gw
                        row_modified = True
                        log["changes"].append({
                            "tech": tech, "year": year, "param": MIN_INV_PARAM,
                            "old": old, "new": floor_gw,
                            "pct": pct, "total_prod": prod,
                            "floor_pj": floor_pj, "cf": cf,
                            "reason": "capacity_floor",
                        })

                    # Untie rule: MaxCapInv > MinCapInv
                    current_maxinv = maxinv_map.get((tech, year))
                    if floor_gw > 0 and current_maxinv is not None and current_maxinv <= floor_gw:
                        new_maxinv = floor_gw * UNTIE_MULTIPLIER
                        # Find and update MaxCapInv row
                        maxinv_row = find_param_row(
                            ws, tech, MAX_INV_PARAM, tech_col, param_col
                        )
                        if maxinv_row is not None:
                            max_cell = ws.cell(row=maxinv_row, column=year_cols[year])
                            max_cell.value = new_maxinv
                            maxinv_map[(tech, year)] = new_maxinv
                            log["changes"].append({
                                "tech": tech, "year": year, "param": MAX_INV_PARAM,
                                "old": current_maxinv, "new": new_maxinv,
                                "reason": "untie_maxinv",
                            })

                if row_modified and proj_mode_col is not None:
                    mode_cell = ws.cell(row=row, column=proj_mode_col)
                    if mode_cell.value == PROJ_MODE_EMPTY:
                        mode_cell.value = PROJ_MODE_USER
                        log["projection_mode_flips"].append(
                            {"tech": tech, "param": MIN_INV_PARAM}
                        )

    return log


def edit_parametrization(filepath: Path, sheets: list, config: dict,
                         total_prod: dict, gen_techs: set,
                         tech_yield: dict = None,
                         tech_bau_cap: dict = None) -> dict:
    """Apply VRE targets to the parametrization workbook."""
    wb = load_workbook(filepath)
    file_log = {"file": str(filepath), "sheets": []}

    try:
        for sheet in sheets:
            if sheet not in wb.sheetnames:
                file_log["sheets"].append(
                    {"sheet": sheet, "skipped": "sheet not present in workbook"}
                )
                continue
            ws = wb[sheet]
            year_cols = find_year_columns(ws)
            if not year_cols:
                file_log["sheets"].append(
                    {"sheet": sheet, "skipped": "no integer year columns found"}
                )
                continue
            sheet_log = apply_vre_targets(ws, config, total_prod, gen_techs,
                                          year_cols, tech_yield, tech_bau_cap)
            file_log["sheets"].append(sheet_log)
        wb.save(filepath)
    finally:
        wb.close()

    return file_log


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------
def run(input_dir, sheets: list = None, skip_backup: bool = False,
        yaml_path: Path | None = None) -> dict:
    """End-to-end: load config, load BAU results, backup, edit, log."""
    input_dir = Path(input_dir)
    sheets = sheets or DEFAULT_TARGET_SHEETS

    # Locate YAML config
    if yaml_path is None:
        yaml_path = Path(__file__).resolve().parent / YAML_FILE_NAME
    if not yaml_path.is_file():
        raise FileNotFoundError(
            f"YAML config not found at {yaml_path}. "
            f"Create {YAML_FILE_NAME} next to this script."
        )
    config = load_config(yaml_path)

    # Load TECH_TYPES.csv — needed both to gate target tech matching and to
    # filter the BAU production aggregate to electricity generation only.
    script_dir = Path(__file__).resolve().parent
    tech_types_path = script_dir.parent / TECH_TYPES_FILE
    gen_techs = load_generation_techs(tech_types_path)

    # Load BAU total production (HARD DEPENDENCY)
    bau_path = Path(config["bau_results_path"])
    if not bau_path.is_absolute():
        # Resolve relative to input_dir's parent (project root)
        bau_path = input_dir.parent.parent / bau_path
    total_prod = load_bau_total_production(bau_path, gen_techs)

    # Realized per-tech yield (PJ/GW) and CalBAU capacity path for cap_envelope.
    envelope_techs: set = set()
    for t in config["targets"]:
        if t.get("cap_envelope"):
            envelope_techs.update(expand_tech_pattern(t["tech"], t["cr"], gen_techs))
    if envelope_techs:
        tech_yield, tech_bau_cap = load_bau_tech_yield(
            bau_path, envelope_techs, config.get("default_capacity_factor", 0.10))
    else:
        tech_yield, tech_bau_cap = {}, {}

    # Backup
    backup_dir = None if skip_backup else make_backup(input_dir)

    # Edit
    paramfile = input_dir / PARAM_FILE_NAME
    if not paramfile.exists():
        raise FileNotFoundError(f"{paramfile} not found")

    log = edit_parametrization(paramfile, sheets, config, total_prod, gen_techs,
                               tech_yield, tech_bau_cap)
    log["backup_dir"] = str(backup_dir) if backup_dir else None
    log["timestamp"] = datetime.now().isoformat()
    log["config"] = {
        "bau_results_path": str(bau_path),
        "constraint_type": config["constraint_type"],
        "max_floor_share": config["max_floor_share"],
        "targets_count": len(config["targets"]),
    }
    log["bau_prod_summary"] = _summarize_prod(total_prod)
    log["gen_techs_count"] = len(gen_techs)

    if backup_dir is not None:
        log_path = backup_dir.parent / f"{backup_dir.name}_CHANGES.json"
        log_path.write_text(json.dumps(log, indent=2, default=str))
        log["log_path"] = str(log_path)

    return log


def _summarize_prod(total_prod: dict) -> dict:
    """Create a readable summary of total production by cr."""
    by_cr: dict = {}
    for (cr, year), val in sorted(total_prod.items()):
        if cr not in by_cr:
            by_cr[cr] = {}
        by_cr[cr][str(year)] = round(val, 2)
    return by_cr


# ---------------------------------------------------------------------------
# Console output
# ---------------------------------------------------------------------------
def print_summary(log: dict) -> None:
    bar = "=" * 72
    cfg = log.get("config", {})
    print(bar)
    print("set_vre_targets — VRE generation/capacity floors applied")
    print(bar)
    print(f"Backup folder      : {log.get('backup_dir', '(skipped)')}")
    print(f"Edited file        : {log['file']}")
    print(f"Constraint type    : {cfg.get('constraint_type', '?')}")
    print(f"Max floor share    : {cfg.get('max_floor_share', '?')}")
    print(f"BAU results        : {cfg.get('bau_results_path', '?')}")
    print(f"GENERATION techs   : {log.get('gen_techs_count', '?')}")
    print(f"Target entries     : {cfg.get('targets_count', '?')}")

    # BAU production summary
    prod_summary = log.get("bau_prod_summary", {})
    if prod_summary:
        print(f"\nBAU total production by country-region (PJ):")
        for cr, years_data in sorted(prod_summary.items()):
            yr_keys = sorted(years_data.keys())
            first_y, last_y = yr_keys[0], yr_keys[-1]
            print(f"  {cr}: {years_data[first_y]} ({first_y}) → "
                  f"{years_data[last_y]} ({last_y})")

    print()
    for s in log["sheets"]:
        if "skipped" in s:
            print(f"[SKIPPED] '{s['sheet']}': {s['skipped']}")
            continue
        print(f"Sheet: '{s['sheet']}'")
        targets_proc = s.get("targets_processed", [])
        for tp in targets_proc:
            print(f"  Target: {tp['cr']} / {tp['tech_pattern']} "
                  f"→ matched {len(tp['matched_techs'])} techs "
                  f"({', '.join(tp['matched_techs'])})")

        changes = s.get("changes", [])
        from collections import Counter
        reason_counts = Counter(c.get("reason", "?") for c in changes)
        print(f"  Cells written: {len(changes)}")
        for reason, count in sorted(reason_counts.items()):
            print(f"    - {reason:30s} : {count}")
        print(f"  Projection.Mode flips: "
              f"{len(s.get('projection_mode_flips', []))}")

        warnings = s.get("warnings", [])
        if warnings:
            print(f"  Warnings:")
            for w in warnings:
                print(f"    {w}")

    if log.get("log_path"):
        print(f"\nDetailed change log: {log['log_path']}")


# ---------------------------------------------------------------------------
# Self-test
# ---------------------------------------------------------------------------
def run_self_test() -> int:
    """Build synthetic data, run all tests, verify assertions."""
    bar = "=" * 72
    print(bar)
    print("set_vre_targets.py — SELF-TEST")
    print(bar)

    passed = 0
    failed = 0
    total_tests = 5

    # --- Synthetic data constants ---
    techs_gen = ["PWRSPVINDEA", "PWRWONINDEA", "PWRSPVBGDXX", "PWRCOAINDEA"]
    years_3 = [2030, 2035]
    years_5 = [2025, 2030, 2032, 2035, 2040]

    def _build_workbook(tmpdir: Path, years: list,
                        maxinv_preset: dict | None = None) -> Path:
        """Create minimal parametrization workbook."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Secondary Techs"
        headers = ["Tech", "Parameter", "Projection.Mode"] + years
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)

        row = 2
        for tech in techs_gen:
            # ResidualCapacity (dummy — not used by this script but present
            # for structural consistency)
            ws.cell(row=row, column=1, value=tech)
            ws.cell(row=row, column=2, value="ResidualCapacity")
            ws.cell(row=row, column=3, value="EMPTY")
            row += 1

            # MinCapInv (empty)
            ws.cell(row=row, column=1, value=tech)
            ws.cell(row=row, column=2, value=MIN_INV_PARAM)
            ws.cell(row=row, column=3, value="EMPTY")
            row += 1

            # MaxCapInv
            ws.cell(row=row, column=1, value=tech)
            ws.cell(row=row, column=2, value=MAX_INV_PARAM)
            ws.cell(row=row, column=3, value="EMPTY")
            if maxinv_preset and tech in maxinv_preset:
                for ci, y in enumerate(years, 4):
                    ws.cell(row=row, column=ci, value=maxinv_preset[tech])
            row += 1

        path = tmpdir / "input_dir" / PARAM_FILE_NAME
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        wb.close()
        return path

    def _build_tech_types(tmpdir: Path) -> Path:
        path = tmpdir / "TECH_TYPES.csv"
        lines = [f"{TECH_TYPES_CATEGORY_COL},{TECH_TYPES_TECH_COL}"]
        for t in techs_gen:
            lines.append(f"{GENERATION_CATEGORY},{t}")
        path.write_text("\n".join(lines))
        return path

    def _build_prod_csv(tmpdir: Path, prod_data: dict) -> Path:
        """Create synthetic ProductionByTechnology.csv.
        prod_data: {(tech, year): value}
        """
        bau_dir = tmpdir / "bau_results"
        bau_dir.mkdir(parents=True, exist_ok=True)
        path = bau_dir / PROD_CSV_NAME
        lines = ["REGION,TECHNOLOGY,TIMESLICE,YEAR,VALUE"]
        for (tech, year), val in prod_data.items():
            lines.append(f"SOASIA,{tech},S1,{year},{val}")
        path.write_text("\n".join(lines))
        return bau_dir

    def _build_yaml(tmpdir: Path, content: dict) -> Path:
        path = tmpdir / YAML_FILE_NAME
        lines = []
        lines.append(f"bau_results_path: '{content.get('bau_results_path', '')}'")
        lines.append(f"constraint_type: \"{content.get('constraint_type', 'activity')}\"")
        lines.append(f"max_floor_share: {content.get('max_floor_share', 0.80)}")
        lines.append("targets:")
        for t in content.get("targets", []):
            lines.append(f"  - cr: \"{t['cr']}\"")
            lines.append(f"    tech: \"{t['tech']}\"")
            sched_str = ", ".join(f"{y}: {v}" for y, v in sorted(t["schedule"].items()))
            lines.append(f"    schedule: {{{sched_str}}}")
        path.write_text("\n".join(lines))
        return path

    def _read_param_values(filepath: Path, param_name: str) -> dict:
        """Read {(tech, year): value} for a given parameter."""
        wb = load_workbook(filepath, data_only=True)
        ws = wb["Secondary Techs"]
        yc = find_year_columns(ws)
        hdr = find_named_columns(ws, ["Tech", "Parameter"])
        result: dict = {}
        for row_idx in range(2, ws.max_row + 1):
            tech = ws.cell(row=row_idx, column=hdr["Tech"]).value
            param = ws.cell(row=row_idx, column=hdr["Parameter"]).value
            if param != param_name:
                continue
            for year, col in yc.items():
                val = ws.cell(row=row_idx, column=col).value
                if val is not None:
                    result[(tech, year)] = float(val)
        wb.close()
        return result

    # ======================================================================
    # TEST 1: Activity floor computation
    # ======================================================================
    print("\nTest 1 — Activity floor computation")
    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)
        wb_path = _build_workbook(tmpdir, years_3)
        tt_path = _build_tech_types(tmpdir)

        # BAU production: distribute total across techs so cr totals are known
        # INDEA total: 100 PJ (2030), 150 PJ (2035)
        # BGDXX total: 50 PJ (2030), 80 PJ (2035)
        prod_data = {
            ("PWRSPVINDEA", 2030): 30.0, ("PWRWONINDEA", 2030): 20.0,
            ("PWRCOAINDEA", 2030): 50.0,
            ("PWRSPVINDEA", 2035): 50.0, ("PWRWONINDEA", 2035): 30.0,
            ("PWRCOAINDEA", 2035): 70.0,
            ("PWRSPVBGDXX", 2030): 50.0,
            ("PWRSPVBGDXX", 2035): 80.0,
        }
        bau_dir = _build_prod_csv(tmpdir, prod_data)

        yaml_path = _build_yaml(tmpdir, {
            "bau_results_path": str(bau_dir),
            "constraint_type": "activity",
            "targets": [
                {"cr": "INDEA", "tech": "PWRSPV*", "schedule": {2030: 0.30, 2035: 0.40}},
                {"cr": "BGDXX", "tech": "PWRSPV*", "schedule": {2030: 0.10, 2035: 0.15}},
            ],
        })

        config = load_config(yaml_path)
        config["bau_results_path"] = str(bau_dir)
        gen_techs = load_generation_techs(tt_path)
        total_prod = load_bau_total_production(bau_dir, gen_techs)
        edit_parametrization(wb_path, DEFAULT_TARGET_SHEETS, config, total_prod, gen_techs)

        result = _read_param_values(wb_path, ACTIVITY_LOWER_PARAM)

        test_ok = True
        # INDEA total=100@2030, 150@2035; target 30%, 40%
        expected = {
            ("PWRSPVINDEA", 2030): 100.0 * 0.30,   # 30.0
            ("PWRSPVINDEA", 2035): 150.0 * 0.40,   # 60.0
            # BGDXX total=50@2030, 80@2035; target 10%, 15%
            ("PWRSPVBGDXX", 2030): 50.0 * 0.10,    # 5.0
            ("PWRSPVBGDXX", 2035): 80.0 * 0.15,    # 12.0
        }
        for key, exp in expected.items():
            actual = result.get(key)
            if actual is None or abs(actual - exp) > 0.01:
                print(f"  FAIL: {key}: expected {exp}, got {actual}")
                test_ok = False

        if test_ok:
            print("  PASS")
            passed += 1
        else:
            failed += 1

    # ======================================================================
    # TEST 2: Interpolation (pct and production)
    # ======================================================================
    print("\nTest 2 — Interpolation between milestones")
    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)
        wb_path = _build_workbook(tmpdir, years_5)
        tt_path = _build_tech_types(tmpdir)

        # BAU production: only at milestone years, interpolation needed
        prod_data = {
            ("PWRSPVINDEA", 2025): 20.0, ("PWRCOAINDEA", 2025): 60.0,
            ("PWRSPVINDEA", 2030): 30.0, ("PWRCOAINDEA", 2030): 70.0,
            ("PWRSPVINDEA", 2035): 50.0, ("PWRCOAINDEA", 2035): 100.0,
            ("PWRSPVINDEA", 2040): 80.0, ("PWRCOAINDEA", 2040): 120.0,
            ("PWRWONINDEA", 2025): 0.0, ("PWRWONINDEA", 2030): 0.0,
            ("PWRWONINDEA", 2035): 0.0, ("PWRWONINDEA", 2040): 0.0,
            ("PWRSPVBGDXX", 2025): 0.0, ("PWRSPVBGDXX", 2030): 0.0,
            ("PWRSPVBGDXX", 2035): 0.0, ("PWRSPVBGDXX", 2040): 0.0,
        }
        bau_dir = _build_prod_csv(tmpdir, prod_data)

        yaml_path = _build_yaml(tmpdir, {
            "bau_results_path": str(bau_dir),
            "constraint_type": "activity",
            "targets": [
                {"cr": "INDEA", "tech": "PWRSPV*", "schedule": {2030: 0.30, 2035: 0.40}},
            ],
        })

        config = load_config(yaml_path)
        config["bau_results_path"] = str(bau_dir)
        gen_techs = load_generation_techs(tt_path)
        total_prod = load_bau_total_production(bau_dir, gen_techs)
        edit_parametrization(wb_path, DEFAULT_TARGET_SHEETS, config, total_prod, gen_techs)

        result = _read_param_values(wb_path, ACTIVITY_LOWER_PARAM)

        test_ok = True
        # Year 2025: before first milestone (2030) → floor = 0
        actual_2025 = result.get(("PWRSPVINDEA", 2025))
        if actual_2025 is not None and abs(actual_2025) > 0.01:
            print(f"  FAIL: 2025 should be 0 (pre-milestone), got {actual_2025}")
            test_ok = False

        # Year 2032: interpolated
        # INDEA total prod: 2030=100, 2035=150 → 2032=100+(150-100)*(2/5)=120
        # pct: 2030=0.30, 2035=0.40 → 2032=0.30+(0.10)*(2/5)=0.34
        # floor = 120 × 0.34 = 40.8
        actual_2032 = result.get(("PWRSPVINDEA", 2032))
        expected_2032 = 120.0 * 0.34
        if actual_2032 is None or abs(actual_2032 - expected_2032) > 0.1:
            print(f"  FAIL: 2032: expected ~{expected_2032:.1f}, got {actual_2032}")
            test_ok = False

        # Year 2040: after last milestone → hold pct=0.40
        # INDEA total prod @2040 = 200
        actual_2040 = result.get(("PWRSPVINDEA", 2040))
        expected_2040 = 200.0 * 0.40  # 80
        if actual_2040 is None or abs(actual_2040 - expected_2040) > 0.1:
            print(f"  FAIL: 2040: expected ~{expected_2040:.1f}, got {actual_2040}")
            test_ok = False

        if test_ok:
            print("  PASS")
            passed += 1
        else:
            failed += 1

    # ======================================================================
    # TEST 3: Max floor share safety cap
    # ======================================================================
    print("\nTest 3 — Max floor share safety cap")
    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)
        wb_path = _build_workbook(tmpdir, [2030])
        tt_path = _build_tech_types(tmpdir)

        prod_data = {
            ("PWRSPVINDEA", 2030): 30.0, ("PWRCOAINDEA", 2030): 70.0,
            ("PWRWONINDEA", 2030): 0.0, ("PWRSPVBGDXX", 2030): 0.0,
        }
        bau_dir = _build_prod_csv(tmpdir, prod_data)

        yaml_path = _build_yaml(tmpdir, {
            "bau_results_path": str(bau_dir),
            "constraint_type": "activity",
            "max_floor_share": 0.80,
            "targets": [
                # 90% target → should be capped to 80%
                {"cr": "INDEA", "tech": "PWRSPV*", "schedule": {2030: 0.90}},
            ],
        })

        config = load_config(yaml_path)
        config["bau_results_path"] = str(bau_dir)
        gen_techs = load_generation_techs(tt_path)
        total_prod = load_bau_total_production(bau_dir, gen_techs)
        log = edit_parametrization(wb_path, DEFAULT_TARGET_SHEETS, config, total_prod, gen_techs)

        result = _read_param_values(wb_path, ACTIVITY_LOWER_PARAM)

        # INDEA total = 100; 90% would be 90, capped to 80% = 80
        actual = result.get(("PWRSPVINDEA", 2030))
        expected = 100.0 * 0.80
        test_ok = True
        if actual is None or abs(actual - expected) > 0.01:
            print(f"  FAIL: expected {expected} (capped), got {actual}")
            test_ok = False

        # Check warning was emitted
        warnings = []
        for s in log.get("sheets", []):
            warnings.extend(s.get("warnings", []))
        has_warning = any("WARNING" in w and "PWRSPVINDEA" in w for w in warnings)
        if not has_warning:
            print(f"  FAIL: expected a WARNING about capping, got none")
            test_ok = False

        if test_ok:
            print("  PASS")
            passed += 1
        else:
            failed += 1

    # ======================================================================
    # TEST 4: Missing BAU results
    # ======================================================================
    print("\nTest 4 — Missing BAU results (hard fail)")
    test_ok = True
    try:
        load_bau_total_production(Path("/nonexistent/path/to/bau"), set())
        print("  FAIL: should have raised FileNotFoundError")
        test_ok = False
    except FileNotFoundError as e:
        if "BAU results not found" in str(e):
            pass  # expected
        else:
            print(f"  FAIL: wrong error message: {e}")
            test_ok = False

    if test_ok:
        print("  PASS")
        passed += 1
    else:
        failed += 1

    # ======================================================================
    # TEST 5: Untie rule on capacity mode
    # ======================================================================
    print("\nTest 5 — Untie rule (capacity mode, MaxCapInv < MinCapInv)")
    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)
        # Pre-set MaxCapInv = 3.0 for PWRSPVINDEA
        wb_path = _build_workbook(tmpdir, [2030],
                                  maxinv_preset={"PWRSPVINDEA": 3.0})
        tt_path = _build_tech_types(tmpdir)

        prod_data = {
            ("PWRSPVINDEA", 2030): 30.0, ("PWRCOAINDEA", 2030): 70.0,
            ("PWRWONINDEA", 2030): 0.0, ("PWRSPVBGDXX", 2030): 0.0,
        }
        bau_dir = _build_prod_csv(tmpdir, prod_data)

        # Target 50% of 100 PJ = 50 PJ → GW = 50 / (31.536 × 1.0) ≈ 1.585
        # But that's less than 3.0, so untie won't fire.
        # Let's use a bigger target to trigger untie: 80% of 100 = 80 PJ
        # GW = 80 / 31.536 ≈ 2.537 — still less than 3.0.
        # Use an even bigger target or lower MaxCapInv.
        # Simpler: set MaxCapInv = 0.5, target = 50%
        wb_path2 = _build_workbook(tmpdir, [2030],
                                   maxinv_preset={"PWRSPVINDEA": 0.5})

        yaml_path = _build_yaml(tmpdir, {
            "bau_results_path": str(bau_dir),
            "constraint_type": "capacity",
            "targets": [
                {"cr": "INDEA", "tech": "PWRSPV*", "schedule": {2030: 0.50}},
            ],
        })

        config = load_config(yaml_path)
        config["bau_results_path"] = str(bau_dir)
        gen_techs = load_generation_techs(tt_path)
        total_prod = load_bau_total_production(bau_dir, gen_techs)
        edit_parametrization(wb_path2, DEFAULT_TARGET_SHEETS, config, total_prod, gen_techs)

        # floor_pj = 100 * 0.50 = 50 PJ
        # floor_gw = 50 / (31.536 * 1.0) ≈ 1.5855
        # MaxCapInv was 0.5 < 1.5855 → bumped to 1.5855 * 1.01 ≈ 1.6013
        mininv = _read_param_values(wb_path2, MIN_INV_PARAM)
        maxinv = _read_param_values(wb_path2, MAX_INV_PARAM)

        expected_min = 50.0 / C2AU
        actual_min = mininv.get(("PWRSPVINDEA", 2030))
        actual_max = maxinv.get(("PWRSPVINDEA", 2030))
        expected_max = expected_min * UNTIE_MULTIPLIER

        test_ok = True
        if actual_min is None or abs(actual_min - expected_min) > 0.01:
            print(f"  FAIL: MinCapInv expected ~{expected_min:.4f}, got {actual_min}")
            test_ok = False
        if actual_max is None or abs(actual_max - expected_max) > 0.01:
            print(f"  FAIL: MaxCapInv expected ~{expected_max:.4f} (untied), got {actual_max}")
            test_ok = False

        if test_ok:
            print("  PASS")
            passed += 1
        else:
            failed += 1

    # ======================================================================
    # Summary
    # ======================================================================
    print()
    print(bar)
    if failed == 0:
        print(f"SELF-TEST PASSED ({passed}/{total_tests})")
    else:
        print(f"SELF-TEST FAILED ({passed}/{total_tests} passed, {failed} failed)")
    print(bar)
    return 0 if failed == 0 else 1


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> int:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument(
        "--input-dir", type=Path,
        default=Path("A1_Outputs/A1_Outputs_BAU"),
    )
    parser.add_argument(
        "--sheets", nargs="+", default=DEFAULT_TARGET_SHEETS,
    )
    parser.add_argument("--skip-backup", action="store_true")
    parser.add_argument("--self-test", action="store_true")
    parser.add_argument("--restore", action="store_true")
    parser.add_argument("--restore-from", type=Path, default=None)
    parser.add_argument("--yaml", type=Path, default=None)
    args = parser.parse_args()

    if args.self_test:
        return run_self_test()

    if args.restore or args.restore_from is not None:
        try:
            used = restore_from_backup(args.input_dir, args.restore_from)
        except Exception as exc:
            print(f"ERROR: {exc}", file=sys.stderr)
            return 1
        print(f"Restored {args.input_dir} from {used}")
        return 0

    try:
        log = run(args.input_dir, args.sheets, args.skip_backup, args.yaml)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    print_summary(log)
    return 0


if __name__ == "__main__":
    sys.exit(main())
