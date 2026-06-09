"""
fix_trn_residuals.py
====================

Splits cumulative ResidualCapacity values for transmission (TRN) technologies
in an OSeMOSYS A-O Parametrization workbook into:

  1. Pre-cutoff stock  -> remains in ResidualCapacity, made constant across
                          all model years (capacity that already existed).
  2. Post-cutoff commissionings -> moved to either
        TotalAnnualMinCapacityInvestment[year]  (mode='min', exogenous timing)
     or TotalAnnualMaxCapacity[year]            (mode='max', endogenous timing,
                                                 cap = original cumulative profile)

Techs whose ResidualCapacity is already flat across all years are left untouched
(they represent stock with no scheduled additions).

Commissioning schedule sources, in order of precedence:
  (1) --override CSV (columns: tech, year, capacity_added)
  (2) auto-derive from year-over-year deltas in the existing ResidualCapacity row

Outputs:
  - Corrected workbook (--output)
  - Diff log CSV (--diff-csv)
  - Diff log Markdown summary (--diff-md)

Usage:
  python fix_trn_residuals.py \\
      --input  A-O_Parametrization.xlsx \\
      --output A-O_Parametrization_FIXED.xlsx \\
      --diff-csv diff_log.csv \\
      --diff-md  diff_log.md
"""
from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# ------------------------------------------------------------------ constants
SHEET_NAME = "Secondary Techs"
TECH_PREFIX = "TRN"
RESIDUAL_PARAM = "ResidualCapacity"
MIN_INV_PARAM = "TotalAnnualMinCapacityInvestment"
MAX_CAP_PARAM = "TotalAnnualMaxCapacity"
DEFAULT_CUTOFF_YEAR = 2023
FLAT_TOLERANCE = 1e-9
DEFAULT_PROJECTION_MODE = "User defined"

# OSeMOSYS "unbounded" convention. ResidualCapacity values at this magnitude
# are treated as sentinel (don't adopt as a reference correction).
SENTINEL_VALUE = 9999.0

# columns in the metadata block (left of the year columns)
TECH_COL = "Tech"
PARAM_COL = "Parameter"
PROJ_MODE_COL = "Projection.Mode"


# ------------------------------------------------------------------ data model
@dataclass(frozen=True)
class Commissioning:
    """A single capacity addition event for one tech in one year."""
    tech: str
    year: int
    capacity_added: float


@dataclass
class TechFix:
    """Plan for fixing one TRN tech's residual capacity rows."""
    tech: str
    base_value: float                       # value at cutoff year (kept flat in RC)
    original_profile: Dict[int, float]      # year -> profile value (used to derive deltas)
    commissionings: List[Commissioning]     # only events with capacity_added > 0
    flatten_only: bool = True               # if False, this is a magnitude-only correction
                                            # (input residual was already flat,
                                            # but we're changing its level)
    base_source: str = "input"              # "input" or "reference"
    profile_source: str = "input"           # which profile drove `commissionings`


@dataclass
class DiffEntry:
    """One cell-level change for the diff log."""
    sheet: str
    tech: str
    parameter: str
    year: int
    old_value: float
    new_value: float
    source: str = "split"   # "overlay" (from benchmark) or "split" (residual->TAMCI)


# ---------------------------------------------------------------- header parse
def _build_column_index(ws: Worksheet) -> Tuple[Dict[str, int], Dict[int, int]]:
    """
    Return (meta_cols, year_cols) where:
      meta_cols: {column_name -> 1-based column index} for textual headers
      year_cols: {year (int) -> 1-based column index}
    Reads the header from row 1.
    """
    meta_cols: Dict[str, int] = {}
    year_cols: Dict[int, int] = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        v = cell.value
        if isinstance(v, int):
            year_cols[v] = col_idx
        elif isinstance(v, str) and v.strip():
            meta_cols[v.strip()] = col_idx
    return meta_cols, year_cols


# ---------------------------------------------------------------- read profile
def _read_residual_profile(
    ws: Worksheet,
    meta_cols: Dict[str, int],
    year_cols: Dict[int, int],
) -> Dict[str, Tuple[int, Dict[int, float]]]:
    """
    Walk every row of the sheet. For each TRN tech with a ResidualCapacity row,
    return its (row_index, {year: value}) profile.
    """
    out: Dict[str, Tuple[int, Dict[int, float]]] = {}
    tech_c = meta_cols[TECH_COL]
    param_c = meta_cols[PARAM_COL]
    for row_idx in range(2, ws.max_row + 1):
        tech = ws.cell(row=row_idx, column=tech_c).value
        param = ws.cell(row=row_idx, column=param_c).value
        if not isinstance(tech, str) or not tech.startswith(TECH_PREFIX):
            continue
        if param != RESIDUAL_PARAM:
            continue
        profile = {}
        for y, c in year_cols.items():
            v = ws.cell(row=row_idx, column=c).value
            profile[y] = float(v) if v is not None else 0.0
        out[tech] = (row_idx, profile)
    return out


# ---------------------------------------------------------------- find param row
def _find_param_row(
    ws: Worksheet, meta_cols: Dict[str, int], tech: str, parameter: str
) -> Optional[int]:
    tech_c = meta_cols[TECH_COL]
    param_c = meta_cols[PARAM_COL]
    for row_idx in range(2, ws.max_row + 1):
        if (
            ws.cell(row=row_idx, column=tech_c).value == tech
            and ws.cell(row=row_idx, column=param_c).value == parameter
        ):
            return row_idx
    return None


# ---------------------------------------------------------------- planning
def derive_commissionings(
    profile: Dict[int, float], cutoff_year: int
) -> List[Tuple[int, float]]:
    """
    From a year->value cumulative profile, return list of (year, delta) for
    every year > cutoff_year where the value strictly increased over the
    previous year.  Negative or zero deltas are dropped (decommissionings are
    flagged separately by the caller).
    """
    years = sorted(profile.keys())
    out: List[Tuple[int, float]] = []
    for prev_y, y in zip(years, years[1:]):
        if y <= cutoff_year:
            continue
        delta = profile[y] - profile[prev_y]
        if delta > FLAT_TOLERANCE:
            out.append((y, delta))
    return out


def is_flat(profile: Dict[int, float]) -> bool:
    vals = list(profile.values())
    return (max(vals) - min(vals)) <= FLAT_TOLERANCE


def build_fix_plan(
    ws: Worksheet,
    cutoff_year: int,
    overrides: Optional[Dict[str, List[Commissioning]]] = None,
    reference_residuals: Optional[Dict[str, Dict[int, float]]] = None,
) -> Tuple[List[TechFix], List[str], List[str]]:
    """
    Build the per-tech fix plan.

    For each TRN tech (existing in the input's Secondary Techs sheet):

      1. If the input residual is GROWING (max - min > tol):
           - Flatten it. The flat level is:
             * NATY's 2023 value — IF reference is provided AND NATY's 2023
               value differs from input's, AND NATY's 2023 value is neither
               9999 (sentinel) nor 0 (likely missing/wrong data in NATY).
             * Otherwise the input's own 2023 value.
           - Commissioning events (deltas) are derived from the SAME profile
             the base came from (input or reference) — they must be consistent.
      2. If the input residual is already FLAT but the reference says the
         magnitude should be different (subject to the 9999/0 guard):
           - Set residual flat at the reference's value. No commissionings.
      3. Otherwise: skip — no change needed.

    Returns (plans, skipped_techs, warnings).
    """
    meta_cols, year_cols = _build_column_index(ws)
    if cutoff_year not in year_cols:
        raise ValueError(
            f"Cutoff year {cutoff_year} not found in sheet header; "
            f"available years: {sorted(year_cols.keys())}"
        )

    profiles = _read_residual_profile(ws, meta_cols, year_cols)
    plans: List[TechFix] = []
    skipped: List[str] = []
    warnings: List[str] = []

    def _ref_base_acceptable(t: str) -> Optional[float]:
        """Return the reference's cutoff-year value for `t` if it should
        override the input base, else None.  Implements the 9999/0 guard."""
        if reference_residuals is None or t not in reference_residuals:
            return None
        ref_prof = reference_residuals[t]
        if cutoff_year not in ref_prof:
            return None
        ref_base = ref_prof[cutoff_year]
        input_base = profiles[t][1][cutoff_year]
        if abs(ref_base - input_base) <= FLAT_TOLERANCE:
            return None  # no magnitude difference
        if ref_base >= SENTINEL_VALUE - FLAT_TOLERANCE:
            return None  # 9999 guard
        if abs(ref_base) <= FLAT_TOLERANCE:
            return None  # zero guard
        return ref_base

    for tech in sorted(profiles):
        _, profile = profiles[tech]
        input_grows = not is_flat(profile)
        ref_override = _ref_base_acceptable(tech)

        if not input_grows and ref_override is None:
            skipped.append(tech)
            continue

        if input_grows:
            # Case 1: flatten. Pick the profile to drive deltas.
            if ref_override is not None and tech in (reference_residuals or {}):
                # adopt NATY's magnitude AND derive deltas from NATY's profile
                use_profile = reference_residuals[tech]
                base_value = ref_override
                base_source = "reference"
                profile_source = "reference"
            else:
                use_profile = profile
                base_value = profile[cutoff_year]
                base_source = "input"
                profile_source = "input"

            if overrides and tech in overrides:
                commissionings = list(overrides[tech])
            else:
                commissionings = [
                    Commissioning(tech=tech, year=y, capacity_added=d)
                    for y, d in derive_commissionings(use_profile, cutoff_year)
                ]

            # warn on negative deltas in the profile we're using
            years_sorted = sorted(use_profile.keys())
            for prev_y, y in zip(years_sorted, years_sorted[1:]):
                if y <= cutoff_year:
                    continue
                delta = use_profile[y] - use_profile[prev_y]
                if delta < -FLAT_TOLERANCE:
                    warnings.append(
                        f"{tech}: negative delta {delta:+.3f} between {prev_y} "
                        f"and {y} in {profile_source} profile (decommissioning); "
                        f"not handled by this script"
                    )

            plans.append(TechFix(
                tech=tech,
                base_value=base_value,
                original_profile=dict(use_profile),
                commissionings=commissionings,
                flatten_only=True,
                base_source=base_source,
                profile_source=profile_source,
            ))
        else:
            # Case 2: input already flat, but reference says magnitude differs.
            # Just change the level. No commissionings.
            plans.append(TechFix(
                tech=tech,
                base_value=ref_override,
                original_profile=dict(profile),
                commissionings=[],
                flatten_only=False,
                base_source="reference",
                profile_source="input",   # not used
            ))

    return plans, skipped, warnings


# ---------------------------------------------------------------- apply
def apply_fix(
    ws: Worksheet,
    plan: TechFix,
    mode: str,
    cutoff_year: int,
) -> Tuple[List[DiffEntry], Dict[int, float]]:
    """
    Apply a single TechFix to the worksheet.

    Returns (diffs, preexisting_nonzero) where preexisting_nonzero is the
    {year -> value} of the target row's pre-existing nonzero entries (the
    values that were preserved/added-onto, not clobbered).
    """
    if mode not in ("min", "max"):
        raise ValueError(f"mode must be 'min' or 'max', got {mode!r}")

    meta_cols, year_cols = _build_column_index(ws)
    diffs: List[DiffEntry] = []

    # --- 1. flatten/level ResidualCapacity row ---------------------------
    rc_row = _find_param_row(ws, meta_cols, plan.tech, RESIDUAL_PARAM)
    if rc_row is None:
        raise RuntimeError(
            f"ResidualCapacity row for {plan.tech} not found"
        )
    rc_source = "magnitude" if not plan.flatten_only else "split"
    for y, c in year_cols.items():
        old = ws.cell(row=rc_row, column=c).value
        old_f = float(old) if old is not None else 0.0
        new_f = float(plan.base_value)
        if abs(old_f - new_f) > FLAT_TOLERANCE:
            ws.cell(row=rc_row, column=c).value = new_f
            diffs.append(
                DiffEntry(SHEET_NAME, plan.tech, RESIDUAL_PARAM, y,
                          old_f, new_f, source=rc_source)
            )

    # If this is a magnitude-only correction (input was already flat,
    # we're just changing the level), there are no commissionings to write.
    if not plan.flatten_only:
        return diffs, {}

    # --- 2. populate the destination parameter -----------------------------
    target_param = MIN_INV_PARAM if mode == "min" else MAX_CAP_PARAM
    target_row = _find_param_row(ws, meta_cols, plan.tech, target_param)
    if target_row is None:
        raise RuntimeError(
            f"{target_param} row for {plan.tech} not found "
            f"(expected to exist in template)"
        )

    # Read existing values from target row first, so we can preserve any
    # hand-curated populations rather than silently clobbering them.
    existing_target: Dict[int, float] = {}
    for y, c in year_cols.items():
        v = ws.cell(row=target_row, column=c).value
        existing_target[y] = float(v) if v is not None else 0.0

    if mode == "min":
        # ADDITIVE merge: existing TAMCI entries (hand-curated commissionings)
        # are preserved; script-derived deltas are added on top of whatever is
        # already there.  Both represent commissioning events: if a user
        # pre-specified one in year y AND the residual delta also implies one
        # in year y, both are real.
        target_values: Dict[int, float] = dict(existing_target)
        for c in plan.commissionings:
            target_values[c.year] = target_values.get(c.year, 0.0) + c.capacity_added
    else:  # mode == "max"
        # Cumulative cap = original residual profile.  Where the user already
        # has a (tighter) cap, preserve it; otherwise fill from the profile.
        target_values = {}
        for y in year_cols:
            ex = existing_target.get(y, 0.0)
            target_values[y] = ex if abs(ex) > FLAT_TOLERANCE \
                else float(plan.original_profile[y])

    # set Projection.Mode on the target row to a non-empty value if missing
    if PROJ_MODE_COL in meta_cols:
        pm_c = meta_cols[PROJ_MODE_COL]
        existing = ws.cell(row=target_row, column=pm_c).value
        if existing in (None, "", "EMPTY"):
            ws.cell(row=target_row, column=pm_c).value = DEFAULT_PROJECTION_MODE

    for y, c in year_cols.items():
        old = ws.cell(row=target_row, column=c).value
        old_f = float(old) if old is not None else 0.0
        new_f = float(target_values.get(y, 0.0))
        # Always write an explicit numeric value to the target row so empty
        # cells become 0.0 (rather than staying None, which downstream tools
        # may read as NaN). Only log a diff when the value actually changed.
        ws.cell(row=target_row, column=c).value = new_f
        if abs(old_f - new_f) > FLAT_TOLERANCE:
            diffs.append(
                DiffEntry(SHEET_NAME, plan.tech, target_param, y, old_f, new_f)
            )

    preexisting_nonzero = {y: v for y, v in existing_target.items()
                           if abs(v) > FLAT_TOLERANCE}
    return diffs, preexisting_nonzero


# ---------------------------------------------------------------- override I/O
def load_overrides(path: Path) -> Dict[str, List[Commissioning]]:
    """Load a commissioning override CSV (columns: tech, year, capacity_added)."""
    out: Dict[str, List[Commissioning]] = {}
    with path.open(newline="") as f:
        reader = csv.DictReader(f)
        required = {"tech", "year", "capacity_added"}
        if not required.issubset(set(c.strip().lower() for c in reader.fieldnames or [])):
            raise ValueError(
                f"override CSV must have columns {sorted(required)}; "
                f"got {reader.fieldnames}"
            )
        for row in reader:
            t = row["tech"].strip()
            y = int(row["year"])
            cap = float(row["capacity_added"])
            if cap <= 0:
                continue
            out.setdefault(t, []).append(
                Commissioning(tech=t, year=y, capacity_added=cap)
            )
    return out


# ---------------------------------------------------------------- diff writer
def write_diff_csv(diffs: List[DiffEntry], path: Path) -> None:
    with path.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["sheet", "tech", "parameter", "year",
                    "old_value", "new_value", "delta", "source"])
        for d in diffs:
            w.writerow(
                [d.sheet, d.tech, d.parameter, d.year,
                 f"{d.old_value:.6f}", f"{d.new_value:.6f}",
                 f"{d.new_value - d.old_value:+.6f}", d.source]
            )


def write_diff_md(
    diffs: List[DiffEntry],
    plans: List[TechFix],
    skipped: List[str],
    warnings: List[str],
    mode: str,
    cutoff_year: int,
    path: Path,
) -> None:
    magnitude_diffs = [d for d in diffs if d.source == "magnitude"]
    split_diffs = [d for d in diffs if d.source == "split"]

    flatten_plans = [p for p in plans if p.flatten_only]
    magn_only_plans = [p for p in plans if not p.flatten_only]

    lines: List[str] = []
    lines.append(f"# TRN ResidualCapacity Fix — Diff Log")
    lines.append("")
    lines.append(f"- Mode: **{mode}** "
                 f"(`{MIN_INV_PARAM}` if min, `{MAX_CAP_PARAM}` if max)")
    lines.append(f"- Cutoff year: **{cutoff_year}**")
    lines.append(f"- Techs split (residual flattened, deltas → "
                 f"`{MIN_INV_PARAM if mode == 'min' else MAX_CAP_PARAM}`): "
                 f"**{len(flatten_plans)}**")
    lines.append(f"- Techs with magnitude-only correction "
                 f"(already flat in input, level adjusted from reference): "
                 f"**{len(magn_only_plans)}**")
    lines.append(f"- Techs skipped (no change needed): **{len(skipped)}**")
    if magnitude_diffs:
        lines.append(f"- Cell changes from magnitude correction: **{len(magnitude_diffs)}**")
    lines.append(f"- Cell changes from residual splitting: **{len(split_diffs)}**")
    if warnings:
        lines.append("")
        lines.append("## Warnings")
        for w in warnings:
            lines.append(f"- {w}")
    lines.append("")
    lines.append("## Skipped (no change needed — already flat & magnitude OK)")
    for t in skipped:
        lines.append(f"- `{t}`")
    if magn_only_plans:
        lines.append("")
        lines.append("## Magnitude-only corrections "
                     "(input was already flat; level changed from reference)")
        for plan in magn_only_plans:
            old_level = plan.original_profile[cutoff_year]
            lines.append(f"- `{plan.tech}`: {old_level:.3f} → "
                         f"**{plan.base_value:.3f}** (flat across all years)")
    lines.append("")
    lines.append("## Per-tech commissioning schedules (residual splits)")
    for plan in flatten_plans:
        lines.append(f"\n### `{plan.tech}`")
        lines.append(f"- Pre-{cutoff_year} stock (kept in `ResidualCapacity`, "
                     f"flat across all years): **{plan.base_value:.3f}** "
                     f"_(base from {plan.base_source})_")
        if plan.profile_source != plan.base_source:
            lines.append(f"- Commissioning deltas derived from "
                         f"`{plan.profile_source}` profile")
        elif plan.profile_source == "reference":
            lines.append(f"- Commissioning deltas derived from reference profile")
        if plan.commissionings:
            lines.append(f"- Post-{cutoff_year} commissionings "
                         f"(moved to `{MIN_INV_PARAM if mode == 'min' else MAX_CAP_PARAM}`):")
            for c in plan.commissionings:
                lines.append(f"  - {c.year}: +{c.capacity_added:.3f}")
        else:
            lines.append("- (no post-cutoff commissionings derived)")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


# ---------------------------------------------------------------- magnitude reference
def _read_reference_residuals(
    reference_path: Path,
) -> Dict[str, Dict[int, float]]:
    """Read TRN ResidualCapacity rows from a reference workbook's Secondary
    Techs sheet. Returns {tech -> {year -> value}}."""
    bm_wb = load_workbook(reference_path, data_only=True)
    if SHEET_NAME not in bm_wb.sheetnames:
        raise ValueError(
            f"reference missing sheet '{SHEET_NAME}'; has: {bm_wb.sheetnames}"
        )
    bm_ws = bm_wb[SHEET_NAME]
    bm_meta, bm_years = _build_column_index(bm_ws)
    tech_c = bm_meta[TECH_COL]
    param_c = bm_meta[PARAM_COL]

    out: Dict[str, Dict[int, float]] = {}
    for r in range(2, bm_ws.max_row + 1):
        t = bm_ws.cell(row=r, column=tech_c).value
        p = bm_ws.cell(row=r, column=param_c).value
        if not isinstance(t, str) or not t.startswith(TECH_PREFIX):
            continue
        if p != RESIDUAL_PARAM:
            continue
        prof = {}
        for y, c in bm_years.items():
            v = bm_ws.cell(row=r, column=c).value
            prof[y] = float(v) if v is not None else 0.0
        out[t] = prof
    return out


# ---------------------------------------------------------------- pipeline
def run_fix(
    input_path: Path,
    output_path: Path,
    mode: str = "min",
    cutoff_year: int = DEFAULT_CUTOFF_YEAR,
    override_path: Optional[Path] = None,
    reference_path: Optional[Path] = None,
) -> Tuple[List[DiffEntry], List[TechFix], List[str], List[str]]:
    """Open workbook, build plan (consulting reference if given), apply fixes,
    save. Pure pipeline (no CLI)."""
    wb = load_workbook(input_path)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"sheet '{SHEET_NAME}' not in workbook; sheets: {wb.sheetnames}"
        )
    ws = wb[SHEET_NAME]

    diffs: List[DiffEntry] = []
    warnings: List[str] = []

    reference_residuals = None
    if reference_path is not None:
        reference_residuals = _read_reference_residuals(reference_path)

    overrides = load_overrides(override_path) if override_path else None
    plans, skipped, plan_warnings = build_fix_plan(
        ws, cutoff_year, overrides, reference_residuals=reference_residuals
    )
    warnings.extend(plan_warnings)

    target_param = MIN_INV_PARAM if mode == "min" else MAX_CAP_PARAM
    for plan in plans:
        d, preexisting = apply_fix(ws, plan, mode=mode, cutoff_year=cutoff_year)
        diffs.extend(d)
        if preexisting:
            entries = ", ".join(f"{y}: {v:+.3f}" for y, v in sorted(preexisting.items()))
            warnings.append(
                f"{plan.tech}: pre-existing nonzero {target_param} values "
                f"preserved/merged: {{{entries}}}"
            )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return diffs, plans, skipped, warnings


# ---------------------------------------------------------------- CLI
def main(argv: Optional[List[str]] = None) -> int:
    p = argparse.ArgumentParser(description=__doc__.split("\n")[1])
    p.add_argument("--input", type=Path, required=True,
                   help="path to input A-O_Parametrization.xlsx")
    p.add_argument("--output", type=Path, required=True,
                   help="path for corrected workbook")
    p.add_argument("--diff-csv", type=Path, default=None,
                   help="path for cell-level diff log CSV")
    p.add_argument("--diff-md", type=Path, default=None,
                   help="path for human-readable diff log Markdown")
    p.add_argument("--mode", choices=("min", "max"), default="min",
                   help="'min' = exogenous timing via "
                        f"{MIN_INV_PARAM} (default); "
                        f"'max' = endogenous timing capped by {MAX_CAP_PARAM}")
    p.add_argument("--cutoff-year", type=int, default=DEFAULT_CUTOFF_YEAR,
                   help=f"split year (default {DEFAULT_CUTOFF_YEAR}); "
                        "values at this year stay in ResidualCapacity, "
                        "additions after move to the destination parameter")
    p.add_argument("--override", type=Path, default=None,
                   help="optional CSV with columns tech,year,capacity_added "
                        "to override the auto-derived commissioning schedule")
    p.add_argument("--reference", type=Path, default=None,
                   help="optional reference workbook whose ResidualCapacity "
                        "column corrects A-O's TRN base magnitudes. A reference "
                        "value is adopted only if it differs from input AND "
                        "is neither 9999 (sentinel) nor 0 (treated as missing).")
    args = p.parse_args(argv)

    diffs, plans, skipped, warnings = run_fix(
        input_path=args.input,
        output_path=args.output,
        mode=args.mode,
        cutoff_year=args.cutoff_year,
        override_path=args.override,
        reference_path=args.reference,
    )

    if args.diff_csv:
        write_diff_csv(diffs, args.diff_csv)
    if args.diff_md:
        write_diff_md(
            diffs, plans, skipped, warnings,
            mode=args.mode, cutoff_year=args.cutoff_year, path=args.diff_md
        )

    print(f"Fixed {len(plans)} techs, skipped {len(skipped)} flat techs, "
          f"wrote {len(diffs)} cell changes -> {args.output}")
    if warnings:
        print(f"  ({len(warnings)} warning(s); see diff log)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
