"""fix_rnwbio_restore.py
==========================
Restores the 2 VariableCost rows that were lost in the "pull from
first_asia_model" (commit f1aad25):
  - RNWBIOBGDXX  (Biomass Bangladesh, mode 1)  -- inserted BEFORE row of RNWBIOINDEA
  - RNWBIONPLXX  (Biomass Nepal, mode 1)        -- inserted AFTER  row of RNWWASINDWE

It copies them from a "source of truth" file, preserving the canonical
position in the tech ordering (it does not append them at the end). Uses
openpyxl insert_rows() for positional insertion.

Idempotent: if the row already exists in the input, it is skipped.

Usage:
    python fix_rnwbio_restore.py --input <patched.xlsx> --source <luis_or_old.xlsx>
"""
from __future__ import annotations
import argparse
from pathlib import Path
import sys
from openpyxl import load_workbook

TARGET_SHEET = "VariableCost"
TECH_COL_NAME = "Tech"

# Each entry: (tech_to_insert, anchor_tech, position) where position is
# 'before' or 'after' relative to the anchor.
INSERTIONS = [
    ("RNWBIOBGDXX", "RNWBIOINDEA", "before"),
    ("RNWBIONPLXX", "RNWWASINDWE", "after"),
]


def find_first_row(ws, tech_col: int, tech: str) -> int | None:
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, tech_col).value == tech:
            return r
    return None


def get_row_values(ws, row: int, ncols: int) -> list:
    return [ws.cell(row, c).value for c in range(1, ncols + 1)]


def restore_rnwbio(input_path: Path, source_path: Path,
                   output_path: Path | None = None) -> int:
    if output_path is None:
        output_path = input_path

    wb_in = load_workbook(input_path)
    wb_src = load_workbook(source_path, data_only=True)

    if TARGET_SHEET not in wb_in.sheetnames:
        sys.exit(f"ERROR: '{TARGET_SHEET}' not in input workbook")
    if TARGET_SHEET not in wb_src.sheetnames:
        sys.exit(f"ERROR: '{TARGET_SHEET}' not in source workbook")

    ws_in = wb_in[TARGET_SHEET]
    ws_src = wb_src[TARGET_SHEET]

    in_headers = [c.value for c in ws_in[1]]
    src_headers = [c.value for c in ws_src[1]]
    if in_headers != src_headers:
        sys.exit(f"ERROR: header mismatch.\n  input : {in_headers}\n  source: {src_headers}")

    tech_col = in_headers.index(TECH_COL_NAME) + 1
    ncols = len(in_headers)

    rows_added = 0
    for tech, anchor, position in INSERTIONS:
        # Skip if already present
        if find_first_row(ws_in, tech_col, tech) is not None:
            print(f"  {tech}: already in input; skipping")
            continue
        # Read values from source
        src_row = find_first_row(ws_src, tech_col, tech)
        if src_row is None:
            print(f"  {tech}: NOT in source; cannot restore")
            continue
        src_values = get_row_values(ws_src, src_row, ncols)
        # Find anchor in input
        anchor_row = find_first_row(ws_in, tech_col, anchor)
        if anchor_row is None:
            # Defensive skip: anchor missing means the input is a non-canonical
            # A1 base (e.g. older snapshot without RNW* techs). Don't abort —
            # the row insertion just isn't applicable here.
            print(f"  {tech}: SKIP — anchor tech '{anchor}' not found in input "
                  f"(cannot determine insertion position)")
            continue
        # Compute insertion position
        insert_at = anchor_row if position == "before" else anchor_row + 1
        ws_in.insert_rows(insert_at)
        # Write source values into the new row
        for c, v in enumerate(src_values, start=1):
            ws_in.cell(insert_at, c).value = v
        print(f"  {tech}: inserted at row {insert_at} ({position} {anchor} at row {anchor_row})")
        rows_added += 1

    wb_in.save(output_path)
    print(f"Added {rows_added} row(s)")
    print(f"Saved: {output_path}")
    return rows_added


if __name__ == "__main__":
    p = argparse.ArgumentParser(description=__doc__.split("\n")[1])
    p.add_argument("--input", type=Path, required=True)
    p.add_argument("--source", type=Path, required=True,
                   help="workbook to copy RNWBIO rows from (e.g., A-O_Parametrization_REFERENCE_with_RNWBIO.xlsx)")
    p.add_argument("--output", type=Path, default=None)
    args = p.parse_args()
    if not args.input.is_file():
        sys.exit(f"ERROR: input not found: {args.input}")
    if not args.source.is_file():
        sys.exit(f"ERROR: source not found: {args.source}")
    restore_rnwbio(args.input, args.source, args.output)
