"""
relax_interconnectors.py
========================

Independent patch — relaxes TotalAnnualMaxCapacityInvestment for
INTERCONNECTOR (TRN*) technologies beyond their ResidualCapacity
baseline, controlled by a YAML configuration file.

This script is the complement of cap_trn_to_residual.py. Where that
script freezes interconnectors at ResidualCapacity, this one opens them
up — by a multiplicative factor, an additive headroom, absolute
per-link overrides, or full unconstrain (9999).

RULE
----
For each TRN tech in TECH_TYPES.csv (category = INTERCONNECTORS):

    IF tech is listed in yaml.overrides:
        MaxCapInv(tech, year) = interpolated override value
    ELSE:
        mode == "multiplicative":
            MaxCapInv = ResidualCapacity × headroom_factor
        mode == "additive":
            MaxCapInv = ResidualCapacity + headroom_gw
        mode == "unconstrained":
            MaxCapInv = 9999

    Safety floor: MaxCapInv = max(MaxCapInv, ResidualCapacity)

    Second pass (MaxCap opening):
        For any overridden tech whose schedule has at least one non-zero
        investment value, TotalAnnualMaxCapacity is set to 9999 across
        all years. This removes the total-capacity ceiling that
        cap_trn_to_residual imposed (MaxCap = ResCap) so that the
        override-controlled MaxCapInv is the sole investment constraint.
        Techs without overrides or with all-zero overrides are untouched.

CONFIGURATION
-------------
Edit relax_interconnectors.yaml (next to this script) to change mode,
headroom, and per-link overrides. See RELAX_INTERCONNECTORS.md for
full YAML schema and pseudologic.

OUTPUT
------
1. Timestamped backup of the input directory.
2. In-place edit of A-O_Parametrization.xlsx.
3. A JSON change log next to the backup.

USAGE
-----
    python relax_interconnectors.py
    python relax_interconnectors.py --input-dir A1_Outputs/A1_Outputs_BAU
    python relax_interconnectors.py --self-test
    python relax_interconnectors.py --restore
"""

from __future__ import annotations

import argparse
import gc
import json
import shutil
import sys
import tempfile
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook, Workbook

# Try to import yaml; fall back to a minimal parser if unavailable.
try:
    import yaml as _yaml
    def _load_yaml(path: Path) -> dict:
        with open(path, "r", encoding="utf-8") as f:
            return _yaml.safe_load(f)
except ImportError:
    _yaml = None
    def _load_yaml(path: Path) -> dict:
        """Minimal YAML-subset loader for flat key-value configs.
        Handles the relax_interconnectors.yaml schema only. For
        production use, install PyYAML."""
        raise ImportError(
            "PyYAML is required. Install with: pip install pyyaml"
        )

# ---------------------------------------------------------------------------
# Configuration constants
# ---------------------------------------------------------------------------
DEFAULT_TARGET_SHEETS = ["Secondary Techs"]
PARAM_FILE_NAME = "A-O_Parametrization.xlsx"
YAML_FILE_NAME = "relax_interconnectors.yaml"

RES_PARAM = "ResidualCapacity"
MAX_CAP_PARAM = "TotalAnnualMaxCapacity"
MAX_INV_PARAM = "TotalAnnualMaxCapacityInvestment"
MIN_INV_PARAM = "TotalAnnualMinCapacityInvestment"

PROJ_MODE_COL = "Projection.Mode"
PROJ_MODE_EMPTY = "EMPTY"
PROJ_MODE_USER = "User defined"

PLACEHOLDER_VALUE = 9999

# Tech-type filter
TECH_TYPES_FILE = "TECH_TYPES.csv"
TECH_TYPES_CATEGORY_COL = "Technology (PWR)"
TECH_TYPES_TECH_COL = "Technology"
INTERCONNECTOR_CATEGORY = "INTERCONNECTORS"

# TRN tech naming: 13 chars, e.g. TRNINDEAINDNE
TRN_TECH_LENGTH = 13


# ---------------------------------------------------------------------------
# YAML config loader
# ---------------------------------------------------------------------------
def load_config(yaml_path: Path) -> dict:
    """Load and validate the YAML configuration.

    Returns a dict with keys:
        mode             : str — "multiplicative", "additive", or "unconstrained"
        headroom_factor  : float (used by multiplicative)
        headroom_gw      : float (used by additive)
        overrides        : dict {tech: {year: value}}
    """
    cfg = _load_yaml(yaml_path)
    if cfg is None:
        cfg = {}

    mode = cfg.get("mode", "multiplicative")
    if mode not in ("multiplicative", "additive", "unconstrained"):
        raise ValueError(
            f"Invalid mode '{mode}' in {yaml_path}. "
            f"Expected 'multiplicative', 'additive', or 'unconstrained'."
        )

    headroom_factor = float(cfg.get("headroom_factor", 2.0))
    headroom_gw = float(cfg.get("headroom_gw", 1.0))

    # Parse overrides: {tech_str: {year_int: float_value}}
    raw_overrides = cfg.get("overrides", {}) or {}
    overrides: dict = {}
    for tech, year_map in raw_overrides.items():
        tech = str(tech)
        if not isinstance(year_map, dict):
            raise ValueError(
                f"Override for '{tech}' must be a dict of year: value, "
                f"got {type(year_map).__name__}."
            )
        overrides[tech] = {int(y): float(v) for y, v in year_map.items()}

    return {
        "mode": mode,
        "headroom_factor": headroom_factor,
        "headroom_gw": headroom_gw,
        "overrides": overrides,
    }


# ---------------------------------------------------------------------------
# Interpolation
# ---------------------------------------------------------------------------
def interpolate_schedule(schedule: dict, years: list) -> dict:
    """Linearly interpolate a sparse {year: value} schedule to cover all years.

    Years before the first key: clamp to first value.
    Years after the last key: clamp to last value.
    Years between keys: linear interpolation.

    Returns {year: value} for every year in `years`.
    """
    if not schedule:
        return {y: 0.0 for y in years}
    sorted_keys = sorted(schedule.keys())
    result: dict = {}
    for y in years:
        if y <= sorted_keys[0]:
            result[y] = schedule[sorted_keys[0]]
        elif y >= sorted_keys[-1]:
            result[y] = schedule[sorted_keys[-1]]
        else:
            # Find bracketing keys
            for i in range(len(sorted_keys) - 1):
                y_lo, y_hi = sorted_keys[i], sorted_keys[i + 1]
                if y_lo <= y <= y_hi:
                    frac = (y - y_lo) / (y_hi - y_lo)
                    v_lo = schedule[y_lo]
                    v_hi = schedule[y_hi]
                    result[y] = v_lo + frac * (v_hi - v_lo)
                    break
    return result


# ---------------------------------------------------------------------------
# Tech-type loading
# ---------------------------------------------------------------------------
def load_interconnector_techs(tech_types_path: Path) -> set:
    """Load TECH_TYPES.csv and return the set of INTERCONNECTOR techs."""
    tech_types_path = Path(tech_types_path)
    if not tech_types_path.is_file():
        raise FileNotFoundError(
            f"TECH_TYPES.csv not found at {tech_types_path}. "
            f"Place it in A3_process/ (one level above rules_scripts/)."
        )
    df = pd.read_csv(tech_types_path)
    cat_col = TECH_TYPES_CATEGORY_COL
    tech_col = TECH_TYPES_TECH_COL
    missing = [c for c in (cat_col, tech_col) if c not in df.columns]
    if missing:
        raise ValueError(
            f"TECH_TYPES.csv missing columns {missing}. "
            f"Found {list(df.columns)}."
        )
    return set(df.loc[df[cat_col] == INTERCONNECTOR_CATEGORY, tech_col].dropna())


# ---------------------------------------------------------------------------
# Backup / restore (mirrors lid rule pattern)
# ---------------------------------------------------------------------------
BACKUP_TAG = "_PRE_TRN_RELAX_"


def _rmtree_robust(path: Path, attempts: int = 5) -> None:
    for i in range(attempts):
        try:
            shutil.rmtree(path)
            return
        except PermissionError:
            gc.collect()
            time.sleep(0.1 * (i + 1))
    shutil.rmtree(path)


def find_latest_backup(input_dir: Path) -> Path | None:
    parent = input_dir.parent
    candidates = sorted(
        (p for p in parent.iterdir()
         if p.is_dir() and p.name.startswith(f"{input_dir.name}{BACKUP_TAG}")),
        key=lambda p: p.name,
    )
    return candidates[-1] if candidates else None


def restore_from_backup(input_dir: Path, backup_dir: Path | None = None) -> Path:
    input_dir = Path(input_dir)
    if backup_dir is None:
        backup_dir = find_latest_backup(input_dir)
        if backup_dir is None:
            raise FileNotFoundError(
                f"No {BACKUP_TAG}* backup found next to {input_dir}."
            )
    else:
        backup_dir = Path(backup_dir)
        if not backup_dir.is_dir():
            raise FileNotFoundError(f"Backup folder does not exist: {backup_dir}")

    if input_dir.is_dir():
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        snapshot = input_dir.parent / f"{input_dir.name}_POST_TRN_RELAX_pre_restore_{stamp}"
        if not snapshot.exists():
            shutil.copytree(input_dir, snapshot)
        _rmtree_robust(input_dir)
    shutil.copytree(backup_dir, input_dir)
    return backup_dir


def make_backup(input_dir: Path) -> Path:
    if not input_dir.is_dir():
        raise FileNotFoundError(f"Input directory does not exist: {input_dir}")
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = input_dir.parent / f"{input_dir.name}{BACKUP_TAG}{stamp}"
    if backup.exists():
        raise FileExistsError(f"Backup folder already exists: {backup}")
    shutil.copytree(input_dir, backup)
    return backup


# ---------------------------------------------------------------------------
# Worksheet helpers
# ---------------------------------------------------------------------------
def find_year_columns(ws) -> dict:
    """Scan row 1 for integer year headers; return {year: col_idx_1based}."""
    year_to_col = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if isinstance(val, int) and 1900 <= val <= 2200:
            year_to_col[val] = col_idx
    return year_to_col


def find_named_columns(ws, names) -> dict:
    found = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val in names:
            found[val] = col_idx
    return found


def values_differ(a, b, tol: float = 1e-12) -> bool:
    if a is None and b is None:
        return False
    if a is None or b is None:
        return True
    try:
        return abs(float(a) - float(b)) > tol
    except (TypeError, ValueError):
        return a != b


# ---------------------------------------------------------------------------
# Build ResidualCapacity lookup from worksheet
# ---------------------------------------------------------------------------
def build_rescap_map(ws, trn_techs: set, year_cols: dict) -> dict:
    """Return {(tech, year): rescap_value} for TRN techs."""
    headers = find_named_columns(ws, ["Tech", "Parameter"])
    tech_col = headers.get("Tech")
    param_col = headers.get("Parameter")
    if tech_col is None or param_col is None:
        return {}

    rescap_map: dict = {}
    for row_idx in range(2, ws.max_row + 1):
        tech = ws.cell(row=row_idx, column=tech_col).value
        param = ws.cell(row=row_idx, column=param_col).value
        if tech not in trn_techs or param != RES_PARAM:
            continue
        for year, col in year_cols.items():
            val = ws.cell(row=row_idx, column=col).value
            rescap_map[(tech, year)] = float(val) if val is not None and not pd.isna(val) else 0.0
    return rescap_map


def build_min_inv_row_map(ws, trn_techs: set) -> dict:
    """Return {tech: row_idx} for TRN MinCapacityInvestment rows."""
    headers = find_named_columns(ws, ["Tech", "Parameter"])
    tech_col = headers.get("Tech")
    param_col = headers.get("Parameter")
    if tech_col is None or param_col is None:
        return {}

    min_row_map: dict = {}
    for row_idx in range(2, ws.max_row + 1):
        tech = ws.cell(row=row_idx, column=tech_col).value
        param = ws.cell(row=row_idx, column=param_col).value
        if tech in trn_techs and param == MIN_INV_PARAM:
            min_row_map[tech] = row_idx
    return min_row_map


# ---------------------------------------------------------------------------
# Core logic
# ---------------------------------------------------------------------------
def apply_trn_relax(ws, trn_techs: set, config: dict,
                    rescap_map: dict, year_cols: dict) -> dict:
    """Edit MaxCapInv rows for TRN techs in the worksheet.

    Also clears legacy MinCapInv cells that would exceed the computed MaxCapInv
    (the source-of-truth ResidualCapacity / MaxCapInv data comes from the
    intermediary; legacy Min entries inherited from the OG dataset that violate
    Min <= Max are zeroed to avoid OSeMOSYS preprocessor infeasibilities).

    Returns a log dict.
    """
    mode = config["mode"]
    headroom_factor = config["headroom_factor"]
    headroom_gw = config["headroom_gw"]
    overrides = config["overrides"]

    headers = find_named_columns(ws, ["Tech", "Parameter", PROJ_MODE_COL])
    tech_col_idx = headers.get("Tech")
    param_col_idx = headers.get("Parameter")
    proj_mode_col_idx = headers.get(PROJ_MODE_COL)

    if tech_col_idx is None or param_col_idx is None:
        raise ValueError(
            f"Sheet '{ws.title}' missing Tech/Parameter columns."
        )

    sorted_years = sorted(year_cols.keys())

    # Pre-interpolate overrides for all years in the horizon
    interp_overrides: dict = {}
    for tech, schedule in overrides.items():
        interp_overrides[tech] = interpolate_schedule(schedule, sorted_years)

    # Locate Min rows so we can clear stale Min > Max cells in lockstep
    min_inv_row_map = build_min_inv_row_map(ws, trn_techs)

    log = {
        "sheet": ws.title,
        "years": sorted_years,
        "mode": mode,
        "headroom_factor": headroom_factor if mode == "multiplicative" else None,
        "headroom_gw": headroom_gw if mode == "additive" else None,
        "trn_techs_in_scope": sorted(trn_techs),
        "overrides_applied": sorted(overrides.keys()),
        "changes": [],
        "preserved": [],
        "projection_mode_flips": [],
        "skipped_techs": [],
        "min_inv_cleared": [],
    }

    for row_idx in range(2, ws.max_row + 1):
        tech = ws.cell(row=row_idx, column=tech_col_idx).value
        param = ws.cell(row=row_idx, column=param_col_idx).value

        if tech is None or param != MAX_INV_PARAM:
            continue
        if tech not in trn_techs:
            continue

        row_was_modified = False

        for year in sorted_years:
            col = year_cols[year]
            cell = ws.cell(row=row_idx, column=col)
            old = cell.value
            rescap = rescap_map.get((tech, year), 0.0)

            # Compute proposed MaxCapInv
            if tech in interp_overrides:
                proposed = interp_overrides[tech].get(year, rescap)
                reason = "override"
            elif mode == "multiplicative":
                proposed = rescap * headroom_factor
                reason = "multiplicative"
            elif mode == "additive":
                proposed = rescap + headroom_gw
                reason = "additive"
            elif mode == "unconstrained":
                proposed = PLACEHOLDER_VALUE
                reason = "unconstrained"
            else:
                proposed = rescap
                reason = "fallback"

            # Safety floor: never go below ResidualCapacity
            if proposed < rescap:
                proposed = rescap
                reason = f"{reason}_clamped_to_rescap"

            if values_differ(old, proposed):
                cell.value = proposed
                row_was_modified = True
                log["changes"].append({
                    "tech": tech,
                    "year": year,
                    "param": MAX_INV_PARAM,
                    "old": old,
                    "new": proposed,
                    "rescap": rescap,
                    "reason": reason,
                })
            else:
                log["preserved"].append({
                    "tech": tech,
                    "year": year,
                    "value": old,
                })

            # Clear legacy MinCapInv cells that would violate Min <= Max
            min_row_idx = min_inv_row_map.get(tech)
            if min_row_idx is not None:
                min_cell = ws.cell(row=min_row_idx, column=col)
                min_old = min_cell.value
                try:
                    min_val = float(min_old) if min_old is not None and not pd.isna(min_old) else 0.0
                except (TypeError, ValueError):
                    min_val = 0.0
                if min_val > float(proposed) + 1e-12:
                    min_cell.value = 0.0
                    log["min_inv_cleared"].append({
                        "tech": tech,
                        "year": year,
                        "param": MIN_INV_PARAM,
                        "old": min_old,
                        "new": 0.0,
                        "max_inv": proposed,
                        "reason": "min_exceeds_computed_max",
                    })

        # Flip Projection.Mode if modified
        if row_was_modified and proj_mode_col_idx is not None:
            mode_cell = ws.cell(row=row_idx, column=proj_mode_col_idx)
            if mode_cell.value == PROJ_MODE_EMPTY:
                mode_cell.value = PROJ_MODE_USER
                log["projection_mode_flips"].append({"tech": tech})

    # ------------------------------------------------------------------
    # Second pass: open TotalAnnualMaxCapacity for overridden techs
    # ------------------------------------------------------------------
    # For techs with explicit overrides that allow non-zero investment,
    # cap_trn_to_residual may have set MaxCap = ResCap = 0, which blocks
    # any new build even when MaxCapInv > 0.  Open MaxCap to 9999 so
    # MaxCapInv is the sole investment control.
    overrides_with_investment = {
        tech for tech, sched in interp_overrides.items()
        if any(v > 0 for v in sched.values())
    }
    if overrides_with_investment:
        log["maxcap_opened"] = []
        for row_idx in range(2, ws.max_row + 1):
            tech = ws.cell(row=row_idx, column=tech_col_idx).value
            param = ws.cell(row=row_idx, column=param_col_idx).value

            if tech not in overrides_with_investment:
                continue
            if param != MAX_CAP_PARAM:
                continue

            for year in sorted_years:
                col = year_cols[year]
                cell = ws.cell(row=row_idx, column=col)
                old = cell.value
                if old is None or float(old) < PLACEHOLDER_VALUE:
                    cell.value = PLACEHOLDER_VALUE
                    log["maxcap_opened"].append({
                        "tech": tech,
                        "year": year,
                        "old": old,
                        "new": PLACEHOLDER_VALUE,
                    })

            # Flip Projection.Mode
            if proj_mode_col_idx is not None:
                mode_cell = ws.cell(row=row_idx, column=proj_mode_col_idx)
                if mode_cell.value == PROJ_MODE_EMPTY:
                    mode_cell.value = PROJ_MODE_USER

    return log


def edit_parametrization(filepath: Path, sheets: list,
                         trn_techs: set, config: dict) -> dict:
    """Apply the TRN relaxation rule to the parametrization workbook."""
    wb = load_workbook(filepath)
    file_log = {"file": str(filepath), "sheets": []}

    try:
        for sheet in sheets:
            if sheet not in wb.sheetnames:
                file_log["sheets"].append(
                    {"sheet": sheet, "skipped": "sheet not present in workbook"}
                )
                continue

            ws = wb[sheet]
            year_cols = find_year_columns(ws)
            if not year_cols:
                file_log["sheets"].append(
                    {"sheet": sheet, "skipped": "no integer year columns found"}
                )
                continue

            rescap_map = build_rescap_map(ws, trn_techs, year_cols)
            sheet_log = apply_trn_relax(ws, trn_techs, config, rescap_map, year_cols)
            file_log["sheets"].append(sheet_log)

        wb.save(filepath)
    finally:
        wb.close()

    return file_log


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------
def run(input_dir, sheets: list = None, skip_backup: bool = False,
        yaml_path: Path | None = None) -> dict:
    """End-to-end: load config, backup, edit, write log."""
    input_dir = Path(input_dir)
    sheets = sheets or DEFAULT_TARGET_SHEETS

    # Locate YAML config
    if yaml_path is None:
        yaml_path = Path(__file__).resolve().parent / YAML_FILE_NAME
    if not yaml_path.is_file():
        raise FileNotFoundError(
            f"YAML config not found at {yaml_path}. "
            f"Create {YAML_FILE_NAME} next to this script."
        )
    config = load_config(yaml_path)

    # Locate TECH_TYPES.csv (one level up from rules_scripts/)
    script_dir = Path(__file__).resolve().parent
    tech_types_path = script_dir.parent / TECH_TYPES_FILE
    trn_techs = load_interconnector_techs(tech_types_path)

    # Backup
    backup_dir = None if skip_backup else make_backup(input_dir)

    # Edit
    paramfile = input_dir / PARAM_FILE_NAME
    if not paramfile.exists():
        raise FileNotFoundError(f"{paramfile} not found")

    log = edit_parametrization(paramfile, sheets, trn_techs, config)
    log["backup_dir"] = str(backup_dir) if backup_dir else None
    log["timestamp"] = datetime.now().isoformat()
    log["config"] = config
    log["config"]["overrides"] = {
        k: {str(y): v for y, v in sched.items()}
        for k, sched in config["overrides"].items()
    }
    log["tech_types_file"] = str(tech_types_path)
    log["trn_techs_count"] = len(trn_techs)

    if backup_dir is not None:
        log_path = backup_dir.parent / f"{backup_dir.name}_CHANGES.json"
        log_path.write_text(json.dumps(log, indent=2, default=str))
        log["log_path"] = str(log_path)

    return log


# ---------------------------------------------------------------------------
# Console output
# ---------------------------------------------------------------------------
def print_summary(log: dict) -> None:
    """Pretty-print the run summary."""
    bar = "=" * 72
    cfg = log.get("config", {})
    print(bar)
    print("relax_interconnectors — TRN MaxCapInv relaxation applied")
    print(bar)
    print(f"Backup folder : {log.get('backup_dir', '(skipped)')}")
    print(f"Edited file   : {log['file']}")
    print(f"Mode          : {cfg.get('mode', '?')}")
    if cfg.get("mode") == "multiplicative":
        print(f"Headroom      : ×{cfg.get('headroom_factor', '?')}")
    elif cfg.get("mode") == "additive":
        print(f"Headroom      : +{cfg.get('headroom_gw', '?')} GW")
    print(f"TRN techs     : {log.get('trn_techs_count', '?')} loaded "
          f"from {log.get('tech_types_file', '?')}")

    overrides = cfg.get("overrides", {})
    if overrides:
        print(f"Per-link overrides : {len(overrides)} "
              f"({', '.join(sorted(overrides.keys()))})")
    else:
        print(f"Per-link overrides : none")

    print()
    for s in log["sheets"]:
        if "skipped" in s:
            print(f"[SKIPPED] '{s['sheet']}': {s['skipped']}")
            continue
        print(f"Sheet: '{s['sheet']}'")
        years = s.get("years", [])
        print(f"  Years        : {years[0]}..{years[-1]} ({len(years)} years)"
              if years else "  Years        : (none)")
        print(f"  TRN techs in scope : {len(s.get('trn_techs_in_scope', []))}")

        changes = s.get("changes", [])
        preserved = s.get("preserved", [])
        from collections import Counter
        reason_counts = Counter(c.get("reason", "?") for c in changes)
        print(f"  Cells written      : {len(changes)}")
        for reason, count in sorted(reason_counts.items()):
            print(f"    - {reason:30s} : {count}")
        print(f"  Cells preserved    : {len(preserved)}")
        print(f"  Projection.Mode flips : {len(s.get('projection_mode_flips', []))}")

        min_cleared = s.get("min_inv_cleared", [])
        print(f"  Legacy Min cleared : {len(min_cleared)} (Min > computed Max)")
        if min_cleared:
            for m in min_cleared[:5]:
                print(f"    {m['tech']}  {m['year']}  "
                      f"Min={m['old']} → 0.0  (Max={m['max_inv']:.3f})")

        # Sample a few changes
        if changes:
            print(f"  Sample changes (first 5):")
            for c in changes[:5]:
                print(f"    {c['tech']}  {c['year']}  "
                      f"{c.get('old', 'None'):>8} → {c['new']:>8.3f}  "
                      f"(rescap={c.get('rescap', 0):.3f}, {c['reason']})")

    if log.get("log_path"):
        print(f"\nDetailed change log: {log['log_path']}")


# ---------------------------------------------------------------------------
# Self-test
# ---------------------------------------------------------------------------
def run_self_test() -> int:
    """Build synthetic data, run all test cases, verify assertions."""
    import os

    bar = "=" * 72
    print(bar)
    print("relax_interconnectors.py — SELF-TEST")
    print(bar)

    passed = 0
    failed = 0
    total_tests = 4

    # --- Synthetic data setup ---
    techs = ["TRNINDEAINDNE", "TRNINDEAINDSO", "TRNBGDXXINDEA"]
    years = [2025, 2030, 2035]
    # ResidualCapacity per tech (same for all years in this test)
    rescap_values = {
        "TRNINDEAINDNE": 2.0,
        "TRNINDEAINDSO": 3.0,
        "TRNBGDXXINDEA": 1.5,
    }

    def _build_workbook(tmpdir: Path) -> Path:
        """Create a minimal A-O_Parametrization.xlsx with TRN techs."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Secondary Techs"

        # Header row
        headers = ["Tech", "Parameter", "Projection.Mode"] + years
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)

        row = 2
        for tech in techs:
            # ResidualCapacity row
            ws.cell(row=row, column=1, value=tech)
            ws.cell(row=row, column=2, value="ResidualCapacity")
            ws.cell(row=row, column=3, value="EMPTY")
            for ci, y in enumerate(years, 4):
                ws.cell(row=row, column=ci, value=rescap_values[tech])
            row += 1

            # MaxCapInv row (initially empty)
            ws.cell(row=row, column=1, value=tech)
            ws.cell(row=row, column=2, value="TotalAnnualMaxCapacityInvestment")
            ws.cell(row=row, column=3, value="EMPTY")
            # Leave year cells as None
            row += 1

        path = tmpdir / "A1_Outputs" / "A1_Outputs_BAU" / PARAM_FILE_NAME
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        wb.close()
        return path

    def _build_tech_types(tmpdir: Path) -> Path:
        """Create a minimal TECH_TYPES.csv."""
        path = tmpdir / "TECH_TYPES.csv"
        lines = [f"{TECH_TYPES_CATEGORY_COL},{TECH_TYPES_TECH_COL}"]
        for t in techs:
            lines.append(f"{INTERCONNECTOR_CATEGORY},{t}")
        path.write_text("\n".join(lines))
        return path

    def _build_yaml(tmpdir: Path, content: dict) -> Path:
        """Write a YAML config file."""
        path = tmpdir / YAML_FILE_NAME
        # Manual YAML serialization (simple enough for test configs)
        lines = []
        lines.append(f"mode: \"{content.get('mode', 'multiplicative')}\"")
        lines.append(f"headroom_factor: {content.get('headroom_factor', 2.0)}")
        lines.append(f"headroom_gw: {content.get('headroom_gw', 1.0)}")
        overrides = content.get("overrides", {})
        if overrides:
            lines.append("overrides:")
            for tech, sched in overrides.items():
                lines.append(f"  {tech}:")
                for y, v in sorted(sched.items()):
                    lines.append(f"    {y}: {v}")
        else:
            lines.append("overrides: {}")
        path.write_text("\n".join(lines))
        return path

    def _read_maxinv(filepath: Path) -> dict:
        """Read back MaxCapInv values as {(tech, year): value}."""
        wb = load_workbook(filepath, data_only=True)
        ws = wb["Secondary Techs"]
        yc = find_year_columns(ws)
        hdr = find_named_columns(ws, ["Tech", "Parameter"])
        result: dict = {}
        for row_idx in range(2, ws.max_row + 1):
            tech = ws.cell(row=row_idx, column=hdr["Tech"]).value
            param = ws.cell(row=row_idx, column=hdr["Parameter"]).value
            if param != MAX_INV_PARAM:
                continue
            for year, col in yc.items():
                val = ws.cell(row=row_idx, column=col).value
                result[(tech, year)] = float(val) if val is not None else None
        wb.close()
        return result

    def _assert(condition: bool, msg: str):
        nonlocal passed, failed
        if not condition:
            print(f"  FAIL: {msg}")
            failed += 1
        # Don't increment passed here — we count per-test, not per-assertion

    # ======================================================================
    # TEST 1: Multiplicative mode (factor = 2.0)
    # ======================================================================
    print("\nTest 1 — Multiplicative mode (factor = 2.0)")
    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)
        wb_path = _build_workbook(tmpdir)
        tt_path = _build_tech_types(tmpdir)
        yaml_path = _build_yaml(tmpdir, {
            "mode": "multiplicative",
            "headroom_factor": 2.0,
        })

        # Monkey-patch paths for the run
        input_dir = wb_path.parent

        # Direct call to edit_parametrization (skip backup for test)
        trn_techs = load_interconnector_techs(tt_path)
        config = load_config(yaml_path)
        edit_parametrization(wb_path, DEFAULT_TARGET_SHEETS, trn_techs, config)

        result = _read_maxinv(wb_path)

        test1_ok = True
        for tech, rescap in rescap_values.items():
            expected = rescap * 2.0
            for y in years:
                actual = result.get((tech, y))
                if actual is None or abs(actual - expected) > 1e-6:
                    print(f"  FAIL: {tech} {y}: expected {expected}, got {actual}")
                    test1_ok = False

        if test1_ok:
            print("  PASS")
            passed += 1
        else:
            failed += 1

    # ======================================================================
    # TEST 2: Additive mode (headroom_gw = 1.0)
    # ======================================================================
    print("\nTest 2 — Additive mode (headroom_gw = 1.0)")
    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)
        wb_path = _build_workbook(tmpdir)
        tt_path = _build_tech_types(tmpdir)
        yaml_path = _build_yaml(tmpdir, {
            "mode": "additive",
            "headroom_gw": 1.0,
        })

        trn_techs = load_interconnector_techs(tt_path)
        config = load_config(yaml_path)
        edit_parametrization(wb_path, DEFAULT_TARGET_SHEETS, trn_techs, config)

        result = _read_maxinv(wb_path)

        test2_ok = True
        for tech, rescap in rescap_values.items():
            expected = rescap + 1.0
            for y in years:
                actual = result.get((tech, y))
                if actual is None or abs(actual - expected) > 1e-6:
                    print(f"  FAIL: {tech} {y}: expected {expected}, got {actual}")
                    test2_ok = False

        if test2_ok:
            print("  PASS")
            passed += 1
        else:
            failed += 1

    # ======================================================================
    # TEST 3: Per-link override with interpolation
    # ======================================================================
    print("\nTest 3 — Per-link override with interpolation")
    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)
        wb_path = _build_workbook(tmpdir)
        tt_path = _build_tech_types(tmpdir)
        yaml_path = _build_yaml(tmpdir, {
            "mode": "multiplicative",
            "headroom_factor": 2.0,
            "overrides": {
                "TRNINDEAINDNE": {2025: 5.0, 2035: 10.0},
            },
        })

        trn_techs = load_interconnector_techs(tt_path)
        config = load_config(yaml_path)
        edit_parametrization(wb_path, DEFAULT_TARGET_SHEETS, trn_techs, config)

        result = _read_maxinv(wb_path)

        test3_ok = True
        # Override tech: interpolated
        expected_override = {2025: 5.0, 2030: 7.5, 2035: 10.0}
        for y, exp in expected_override.items():
            actual = result.get(("TRNINDEAINDNE", y))
            if actual is None or abs(actual - exp) > 1e-6:
                print(f"  FAIL: TRNINDEAINDNE {y}: expected {exp}, got {actual}")
                test3_ok = False

        # Non-override tech: multiplicative default (3.0 × 2.0 = 6.0)
        for y in years:
            actual = result.get(("TRNINDEAINDSO", y))
            expected = 3.0 * 2.0
            if actual is None or abs(actual - expected) > 1e-6:
                print(f"  FAIL: TRNINDEAINDSO {y}: expected {expected}, got {actual}")
                test3_ok = False

        if test3_ok:
            print("  PASS")
            passed += 1
        else:
            failed += 1

    # ======================================================================
    # TEST 4: Safety floor (override below ResCap)
    # ======================================================================
    print("\nTest 4 — Safety floor (override below ResCap, clamped)")
    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)
        wb_path = _build_workbook(tmpdir)
        tt_path = _build_tech_types(tmpdir)
        yaml_path = _build_yaml(tmpdir, {
            "mode": "multiplicative",
            "headroom_factor": 2.0,
            "overrides": {
                # Override to 0.5 for all years (below ResCap of 1.5)
                "TRNBGDXXINDEA": {2025: 0.5, 2035: 0.5},
            },
        })

        trn_techs = load_interconnector_techs(tt_path)
        config = load_config(yaml_path)
        edit_parametrization(wb_path, DEFAULT_TARGET_SHEETS, trn_techs, config)

        result = _read_maxinv(wb_path)

        test4_ok = True
        # Should be clamped to ResCap = 1.5
        for y in years:
            actual = result.get(("TRNBGDXXINDEA", y))
            expected = 1.5  # ResCap
            if actual is None or abs(actual - expected) > 1e-6:
                print(f"  FAIL: TRNBGDXXINDEA {y}: expected {expected} (clamped), got {actual}")
                test4_ok = False

        if test4_ok:
            print("  PASS")
            passed += 1
        else:
            failed += 1

    # ======================================================================
    # Summary
    # ======================================================================
    print()
    print(bar)
    if failed == 0:
        print(f"SELF-TEST PASSED ({passed}/{total_tests})")
    else:
        print(f"SELF-TEST FAILED ({passed}/{total_tests} passed, {failed} failed)")
    print(bar)
    return 0 if failed == 0 else 1


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> int:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=Path("A1_Outputs/A1_Outputs_BAU"),
        help="Directory containing the AO files (default: A1_Outputs/A1_Outputs_BAU)",
    )
    parser.add_argument(
        "--sheets",
        nargs="+",
        default=DEFAULT_TARGET_SHEETS,
        help=f"Sheets to apply the rule to (default: {DEFAULT_TARGET_SHEETS})",
    )
    parser.add_argument(
        "--skip-backup",
        action="store_true",
        help="Skip backup creation (DANGEROUS — for testing only)",
    )
    parser.add_argument(
        "--self-test",
        action="store_true",
        help="Run built-in self-test with synthetic data, then exit.",
    )
    parser.add_argument(
        "--restore",
        action="store_true",
        help=f"Restore input dir from the most recent {BACKUP_TAG}* backup.",
    )
    parser.add_argument(
        "--restore-from",
        type=Path,
        default=None,
        help="Restore input dir from this specific backup folder.",
    )
    parser.add_argument(
        "--yaml",
        type=Path,
        default=None,
        help="Path to YAML config (default: <script_dir>/relax_interconnectors.yaml).",
    )
    args = parser.parse_args()

    if args.self_test:
        return run_self_test()

    if args.restore or args.restore_from is not None:
        try:
            used = restore_from_backup(args.input_dir, args.restore_from)
        except Exception as exc:
            print(f"ERROR: {exc}", file=sys.stderr)
            return 1
        print(f"Restored {args.input_dir} from {used}")
        return 0

    try:
        log = run(args.input_dir, args.sheets, args.skip_backup, args.yaml)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    print_summary(log)
    return 0


if __name__ == "__main__":
    sys.exit(main())
