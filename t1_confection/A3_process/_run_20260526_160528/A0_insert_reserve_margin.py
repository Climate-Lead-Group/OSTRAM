"""A0_insert_reserve_margin.py
================================
Portable version of A0_Insert_ReserveMargin.py: takes --input argument
instead of relying on hardcoded WORK_DIR.

Adds a 'System Parameters' sheet to the input A-O_Parametrization.xlsx
with a single row: ReserveMargin = 1.15 (flat, all years 2023-2050).
Re-runnable: replaces the existing 'System Parameters' sheet if present.

Usage:
    python A0_insert_reserve_margin.py --input <A-O_Parametrization.xlsx>
"""
from __future__ import annotations
import argparse
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

BASE_YEAR = 2023
END_YEAR = 2050
RESERVE_MARGIN = 1.15
SHEET_NAME = "System Parameters"


def insert(input_path: Path) -> None:
    if not input_path.is_file():
        sys.exit(f"ERROR: input not found: {input_path}")
    wb = openpyxl.load_workbook(input_path)
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)
    years = list(range(BASE_YEAR, END_YEAR + 1))
    headers = ["Parameter", "Unit"] + years
    for c, val in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=val)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    ws.cell(row=2, column=1, value="ReserveMargin")
    ws.cell(row=2, column=2, value="ratio")
    for c, yr in enumerate(years, 3):
        ws.cell(row=2, column=c, value=RESERVE_MARGIN)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 10
    for c in range(3, 3 + len(years)):
        ws.column_dimensions[get_column_letter(c)].width = 8
    wb.save(input_path)
    print(f"Added '{SHEET_NAME}' sheet (ReserveMargin={RESERVE_MARGIN} for {BASE_YEAR}-{END_YEAR})")
    print(f"Saved: {input_path}")


if __name__ == "__main__":
    p = argparse.ArgumentParser(description=__doc__.split("\n")[1])
    p.add_argument("--input", type=Path, required=True)
    args = p.parse_args()
    insert(args.input)
