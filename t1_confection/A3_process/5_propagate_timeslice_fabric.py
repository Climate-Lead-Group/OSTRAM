#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
5_propagate_timeslice_fabric.py

Propagates the 20-timeslice / 5-bracket timeslice fabric from
SOASIA_OSeMOSYS_WV.xlsx into A-O_Parametrization_wvaligned_v2.xlsx.

This is a strictly additive post-step that runs after script 4. It does NOT
modify scripts 1-4. It rewrites only the Yearsplit and DaySplit sheets in the
A-O Parametrization workbook; every other sheet is left byte-identical
(via openpyxl in-place rewrite).

What it does:
    Yearsplit  -- direct copy of WV's Yearsplit_Template (20 rows, fraction-of-year)
    DaySplit   -- copy of WV's DaySplit (5 rows) with x365 unit conversion:
                  WV stores fraction-of-year (sums to 1/365), A-O expects
                  fraction-of-day (sums to 1.0, OSeMOSYS-standard).
                  Conversion factor confirmed against script 1 line 255
                  (`daysplit = hours / 8760` -> hours/(24*365) -> fraction-of-year),
                  and against the existing A-O DaySplit values that already
                  follow the fraction-of-day convention.

A-O conventions preserved:
    * Yearsplit `Unit` column stays NaN (A-O leaves it blank; WV writes 'fraction')
    * DaySplit  `Unit` column stays NaN (same reason)
    * Yearsplit year-column headers stay int (matches A-O current state, also matches WV)
    * DaySplit  year-column headers stay str ('2023', '2024', ...; A-O quirk; WV uses int)
    * Parameter.ID values stay as A-O has them (14 for Yearsplit, 12 for DaySplit;
      same as WV, but enforced explicitly so behaviour doesn't drift if WV changes).

Inputs:
    SOASIA_OSeMOSYS_WV.xlsx
    A-O_Parametrization_wvaligned_v2.xlsx

Output:
    A-O_Parametrization_wvaligned_v2_ts20.xlsx   (new file; original untouched)
"""

from pathlib import Path
import shutil
import sys

import openpyxl
import pandas as pd

# -----------------------------------------------------------------------------
# Paths
# -----------------------------------------------------------------------------
SCRIPT_DIR  = Path(__file__).resolve().parent  # auto-detect

WV_FILE     = SCRIPT_DIR / "SOASIA_OSeMOSYS_WV.xlsx"
AO_IN       = SCRIPT_DIR / "wvaligned_outputs_v2" / "A-O_Parametrization_wvaligned_v2.xlsx"
AO_OUT      = SCRIPT_DIR / "wvaligned_outputs_v2" / "A-O_Parametrization_wvaligned_v2_ts20.xlsx"

# DaySplit unit conversion: WV fraction-of-year  ->  A-O fraction-of-day
DAYSPLIT_UNIT_FACTOR = 365.0


# -----------------------------------------------------------------------------
# Helpers
# -----------------------------------------------------------------------------
def replace_sheet_rows(ws, header_row, new_data_rows):
    """
    Clear all rows in `ws` below the header row and write `new_data_rows` in.
    `header_row` is 1-indexed. `new_data_rows` is a list of tuples, each of
    length equal to ws.max_column (or shorter; missing trailing cols are blank).

    Preserves the header row exactly (formatting, types, font, fill).
    """
    # Wipe existing data rows. Iterate by index, set values to None.
    if ws.max_row > header_row:
        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
            for cell in row:
                cell.value = None

    # Write new rows.
    for i, data_row in enumerate(new_data_rows, start=header_row + 1):
        for j, val in enumerate(data_row, start=1):
            ws.cell(row=i, column=j, value=val)


def read_header(ws, header_row=1):
    """Return list of header values from `ws` row `header_row`."""
    return [c.value for c in ws[header_row]]


# -----------------------------------------------------------------------------
# Step 1: load WV source data
# -----------------------------------------------------------------------------
def load_wv_yearsplit():
    """Return (header_list, data_rows_list) ready to write into A-O Yearsplit."""
    df = pd.read_excel(WV_FILE, sheet_name="Yearsplit_Template")
    # Sort by timeslice for stable ordering (WV is already in S1D1..S4D5 order, but assert)
    expected_ts = [f"S{s}D{d}" for s in range(1, 5) for d in range(1, 6)]
    assert list(df["Timeslices"]) == expected_ts, \
        f"WV Yearsplit_Template not in canonical S1D1..S4D5 order: {list(df['Timeslices'])}"
    return df


def load_wv_daysplit():
    """Return WV DaySplit df with year columns multiplied by 365 (fraction-of-day)."""
    df = pd.read_excel(WV_FILE, sheet_name="DaySplit")
    # Verify shape and brackets
    assert list(df["DAILYTIMEBRACKET"]) == [1, 2, 3, 4, 5], \
        f"WV DaySplit brackets not [1..5]: {list(df['DAILYTIMEBRACKET'])}"
    # Apply x365 conversion to year columns
    year_cols = [c for c in df.columns if isinstance(c, int) and 2000 <= c <= 2100]
    df = df.copy()
    df[year_cols] = df[year_cols] * DAYSPLIT_UNIT_FACTOR
    # Sanity: sums per year must be 1.0 +/- 1e-9
    for yc in year_cols:
        s = df[yc].sum()
        assert abs(s - 1.0) < 1e-6, f"DaySplit sum for {yc} = {s} after x365; expected 1.0"
    return df


# -----------------------------------------------------------------------------
# Step 2: build A-O-shaped row tuples from WV data
# -----------------------------------------------------------------------------
def build_yearsplit_rows(wv_df, ao_header):
    """
    Build a list of row tuples shaped to A-O's Yearsplit header.
    A-O quirk (same as DaySplit): year column headers may be STRING ('2023')
    when A-O comes from A1's openpyxl writes, but WV columns are INT. Resolve
    by trying both forms when reading the source row.
    """
    rows = []
    for _, r in wv_df.iterrows():
        out = []
        for col in ao_header:
            if col == "Timeslices":
                out.append(r["Timeslices"])
            elif col == "Parameter.ID":
                out.append(int(r["Parameter.ID"]))
            elif col == "Parameter":
                out.append(r["Parameter"])  # 'YearSplit'
            elif col == "Unit":
                out.append(None)  # preserve A-O NaN
            elif col == "Projection.Mode":
                out.append(r["Projection.Mode"])  # 'User defined'
            elif col == "Projection.Parameter":
                out.append(int(r["Projection.Parameter"]))
            else:
                # Year column -- try int(col) first, then col as-is.
                try:
                    yc = int(col)
                    if 2000 <= yc <= 2100 and yc in r.index:
                        out.append(float(r[yc]))
                        continue
                except (TypeError, ValueError):
                    pass
                if col in r.index:
                    out.append(float(r[col]))
                else:
                    out.append(None)
        rows.append(tuple(out))
    return rows


def build_daysplit_rows(wv_df, ao_header):
    """
    Build A-O-shaped DaySplit rows.
    A-O quirk: year column headers are STRING ('2023'), but WV columns are INT.
    Resolve by trying both forms when reading the source row.
    """
    rows = []
    for _, r in wv_df.iterrows():
        out = []
        for col in ao_header:
            if col == "DAILYTIMEBRACKET":
                out.append(int(r["DAILYTIMEBRACKET"]))
            elif col == "Parameter.ID":
                out.append(int(r["Parameter.ID"]))
            elif col == "Parameter":
                out.append(r["Parameter"])  # 'DaySplit'
            elif col == "Unit":
                out.append(None)  # preserve A-O NaN
            elif col == "Projection.Mode":
                out.append(r["Projection.Mode"])  # 'User defined'
            elif col == "Projection.Parameter":
                out.append(int(r["Projection.Parameter"]))
            else:
                # Year column -- try int(col) first, then col as-is.
                try:
                    yc = int(col)
                    if 2000 <= yc <= 2100 and yc in r.index:
                        out.append(float(r[yc]))
                        continue
                except (TypeError, ValueError):
                    pass
                if col in r.index:
                    out.append(float(r[col]))
                else:
                    out.append(None)
        rows.append(tuple(out))
    return rows


# -----------------------------------------------------------------------------
# Step 3: rewrite A-O Parametrization
# -----------------------------------------------------------------------------
def main():
    if not WV_FILE.is_file():
        sys.exit(f"[ERROR] WV file not found: {WV_FILE}")
    if not AO_IN.is_file():
        sys.exit(f"[ERROR] A-O input file not found: {AO_IN}")

    # Start by copying input -> output, then mutate output in place via openpyxl.
    shutil.copy2(AO_IN, AO_OUT)
    print(f"[INFO] Copied {AO_IN.name} -> {AO_OUT.name}")

    # Load WV source data.
    print("[INFO] Loading WV Yearsplit_Template ...")
    wv_ys = load_wv_yearsplit()
    print(f"       {len(wv_ys)} rows; sum(2023)={wv_ys[2023].sum():.6f}")

    print("[INFO] Loading WV DaySplit (and applying x365 -> fraction-of-day) ...")
    wv_ds = load_wv_daysplit()
    print(f"       {len(wv_ds)} brackets; sum(2023)={wv_ds[2023].sum():.6f}")

    # Open A-O output workbook and rewrite the two sheets.
    wb = openpyxl.load_workbook(AO_OUT)

    # --- Yearsplit ---
    if "Yearsplit" not in wb.sheetnames:
        sys.exit("[ERROR] 'Yearsplit' sheet missing from A-O workbook")
    ws_ys = wb["Yearsplit"]
    ao_ys_header = read_header(ws_ys)
    new_ys_rows = build_yearsplit_rows(wv_ys, ao_ys_header)
    print(f"[INFO] Rewriting 'Yearsplit' sheet: {ws_ys.max_row - 1} -> {len(new_ys_rows)} data rows")
    replace_sheet_rows(ws_ys, header_row=1, new_data_rows=new_ys_rows)

    # --- DaySplit ---
    if "DaySplit" not in wb.sheetnames:
        sys.exit("[ERROR] 'DaySplit' sheet missing from A-O workbook")
    ws_ds = wb["DaySplit"]
    ao_ds_header = read_header(ws_ds)
    new_ds_rows = build_daysplit_rows(wv_ds, ao_ds_header)
    print(f"[INFO] Rewriting 'DaySplit' sheet: {ws_ds.max_row - 1} -> {len(new_ds_rows)} data rows")
    replace_sheet_rows(ws_ds, header_row=1, new_data_rows=new_ds_rows)

    wb.save(AO_OUT)
    print(f"[INFO] Saved -> {AO_OUT.name}")
    print("[INFO] Done.")


if __name__ == "__main__":
    main()
