# -*- coding: utf-8 -*-
"""
3_update_ao_from_extensions.py  (v18 / pipeline stage 3)

Reads OSTRAM_AO_Extensions.xlsx (the human-filled review file from stage 2)
and the four A-O parametrization workbooks, then writes "_wvaligned" copies
of each A-O workbook. Each output combines:
  * new technology rows appended for every Tab 1 Include=Y entry  (Step 3)
  * refreshed Demand_Projection year-cell values drawn from WV    (Step 2b)
  * a wholesale-replaced Profiles sheet drawn from WV             (Step 2c)

Pipeline:
  1_merge_timeslices_into_WV.py    timeslices source     -> WV
  2_extract_ao_extensions.py       WV  + A-O             -> Extensions xlsx
  -- human fills Include / Add_To_* / Override_* / Notes --
  3_update_ao_from_extensions.py   Extensions + A-O      -> *_wvaligned    <- THIS
  4_apply_manual_fixes.py          *_wvaligned           -> *_wvaligned_v2

Design B (non-negotiable)
-------------------------
Templates supply STRUCTURE only (which sheets, which parameter rows, which
columns). Numerical values come from Tab 2 of the Extensions workbook (the
raw WV rows for the new code). This prevents PWRGEOINDNO from inheriting
hydro CapitalCost when cloning from a PWRHYD template, etc.

Three-tier value lookup (Step 3)
--------------------------------
For each new (Tech, Parameter[, Timeslices]) cell:
  Tier 1 STRICT       -- Tab 2 row exact match on Source_Sheet=WV-equivalent
  Tier 2 CROSS_SHEET  -- Tab 2 row match on (Tech, Parameter) any Source_Sheet
  Tier 3 TEMPLATE     -- Tab 2 has nothing; copy template value verbatim

Color tags on changed/added rows
--------------------------------
  STRICT          light green   (#C6EFCE)  -- Tab 2 strict match        [Step 3]
  CROSS_SHEET     light yellow  (#FFEB9C)  -- Tab 2 cross-sheet match   [Step 3]
  TEMPLATE        orange        (#FCD5B4)  -- template fallback         [Step 3]
  DEMAND_REFRESH  light blue    (#B7D7E8)  -- value sync from WV        [Step 2b]
  (Step 2c rows are not colored -- whole sheet has the same provenance.)

Why Step 2b/2c run BEFORE Step 3
--------------------------------
Step 3 may, in principle, append new rows to Demand_Projection or Profiles
if a future Tab 1 row uses an ELC commodity as a template. Running 2b/2c
first means Step 3's appended rows keep their STRICT/CROSS_SHEET/TEMPLATE
colors untouched. For the current run (templates are PWR/SHP/TRN, not ELC
commodities) this is moot, but the ordering is the safer default.

Setup-A: source workbooks are NEVER modified. Outputs go to a subfolder.

Run in Spyder with F5. Paths and mappings are in USER CONFIGURATION below.
"""

import os
import re
from pathlib import Path
import shutil
import hashlib
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# =============================================================================
# USER CONFIGURATION
# =============================================================================

WORK_DIR = str(Path(__file__).resolve().parent)  # auto-detect (was hardcoded for Spyder)

EXTENSIONS_FILE = WORK_DIR + "/OSTRAM_AO_Extensions.xlsx"
WV_FILE         = WORK_DIR + "/SOASIA_OSeMOSYS_WV.xlsx"

AO_PARAM    = WORK_DIR + "/A-O_Parametrization.xlsx"
AO_AR_BASE  = WORK_DIR + "/A-O_AR_Model_Base_Year.xlsx"
AO_AR_PROJ  = WORK_DIR + "/A-O_AR_Projections.xlsx"
AO_DEMAND   = WORK_DIR + "/A-O_Demand.xlsx"

OUTPUT_SUBDIR = "wvaligned_outputs"   # sits inside WORK_DIR
OUTPUT_SUFFIX = "_wvaligned"

LOG_FILE_NAME = "AO_alignment_log.txt"

# Maps an A-O destination sheet name -> the equivalent WV sheet name as
# recorded in Tab 2's Source_Sheet column. Used by the strict (tier-1) lookup.
AO_TO_WV_SHEET = {
    # A-O_Parametrization
    "Fixed Horizon Parameters": "Fixed_Horizon_Parameters",
    "Primary Techs":            "Primary_Techs",
    "Secondary Techs":          "Secondary_Techs",
    "VariableCost":             "VariableCost",
    "Capacities":               "Capacities_CF",
    # A-O_AR_Model_Base_Year and A-O_AR_Projections
    "Primary":                  "Primary_Techs",
    "Secondary":                "Secondary_Techs",
}

# Region descriptors used in Fuel.I.Name / Fuel.O.Name / Fuel.Name free-text
# columns. Substitution is case-insensitive on the value but emits the
# canonical lower-case "region" form.
REGION_NAME_MAP = {
    "BGDXX": "Bangladesh, region XX",
    "BTNXX": "Bhutan, region XX",
    "INDEA": "India, region EA",
    "INDNE": "India, region NE",
    "INDNO": "India, region NO",
    "INDSO": "India, region SO",
    "INDWE": "India, region WE",
    "LKAXX": "Sri Lanka, region XX",
    "MDVXX": "Maldives, region XX",
    "NPLXX": "Nepal, region XX",
}
VALID_REGIONS = set(REGION_NAME_MAP.keys())

# Three-tier color scheme applied as a row fill on every newly-appended row.
TIER_COLOR = {
    "STRICT":      "FFC6EFCE",  # light green
    "CROSS_SHEET": "FFFFEB9C",  # light yellow (matches stage 1 audit color)
    "TEMPLATE":    "FFFCD5B4",  # orange (matches team's "needs review" color)
    "TAB2_ONLY":   "FFCCE5FF",  # light blue-violet -- Tab2 direct synthesis (Step 3B)
}

# Fourth-tier color for Step 2b: rows whose values were synced from WV's
# Demand_Projection. Distinct from the three Step 3 tiers.
DEMAND_REFRESH_COLOR = "FFB7D7E8"   # light blue

# AR_Projections uses a single Fuel + Direction column pair.
# Direction "Input" rows are InputActivityRatio projections; "Output" rows are
# OutputActivityRatio. Tab 2 uses the Parameter names directly. Accept both
# single-letter and full-word forms in case AR conventions vary.
AR_DIRECTION_TO_PARAM = {
    "I":      "InputActivityRatio",
    "INPUT":  "InputActivityRatio",
    "O":      "OutputActivityRatio",
    "OUTPUT": "OutputActivityRatio",
}

# Code/name column candidates -- mirrors stage 2.
CODE_COLS = ["Tech", "Fuel/Tech", "Technology_Code"]
NAME_COLS = ["Tech.Name", "Name", "Technology_Name"]

# A-O workbooks paired with the Tab 1 flag column that controls them.
AO_WORKBOOKS = [
    ("Param",   AO_PARAM,   "Add_To_Param"),
    ("AR_Base", AO_AR_BASE, "Add_To_AR_Base"),
    ("AR_Proj", AO_AR_PROJ, "Add_To_AR_Proj"),
    ("Demand",  AO_DEMAND,  "Add_To_Demand"),
]

# =============================================================================
# HELPERS
# =============================================================================

def parse_code(code):
    """11-char codes:   PWRSHPINDNO  -> ('PWR', 'SHP', 'INDNO')
       13-char TRN:     TRNBTNXXINDEA -> ('TRN', 'BTNXX', 'INDEA')   [origin, dest]
       Returns (None, None, None) if code matches neither layout.
    """
    s = str(code)
    if len(s) == 11:
        return s[:3], s[3:6], s[6:11]
    if len(s) == 13 and s.startswith("TRN"):
        return s[:3], s[3:8], s[8:13]
    return None, None, None


def find_code_col(headers):
    return next((c for c in CODE_COLS if c in headers), None)


def find_name_col(headers):
    return next((c for c in NAME_COLS if c in headers), None)


def md5_of(path):
    if not os.path.exists(path):
        return None
    h = hashlib.md5()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def regions_of(code):
    """Return the set of 5-char region tokens carried by a tech code."""
    pre, r1, r2 = parse_code(code)
    if pre is None:
        return set()
    if len(str(code)) == 11:
        return {r2}                # category, region
    return {r1, r2}                # origin, destination


def substitute_code_regions(text, template_code, new_code):
    """
    Replace the template's region tokens with the new code's region tokens
    inside any string-shaped cell value (e.g. Fuel.I = 'HYDINDEA',
    Fuel.O = 'ELCINDEA00'). Uses placeholder swap so origin/destination
    overlap doesn't double-substitute. No-op for non-string values.
    """
    if not isinstance(text, str):
        return text

    t_pre, t_r1, t_r2 = parse_code(template_code)
    n_pre, n_r1, n_r2 = parse_code(new_code)
    if t_pre is None or n_pre is None:
        return text
    if len(str(template_code)) != len(str(new_code)):
        return text  # shouldn't happen; spec guarantees same shape via templating

    if len(str(template_code)) == 11:
        # Single region (chars 6-10).
        if t_r2 == n_r2:
            return text
        return text.replace(t_r2, n_r2)

    # 13-char TRN: two regions (origin r1, destination r2). Placeholder swap.
    out = text
    if t_r1 != n_r1 or t_r2 != n_r2:
        out = out.replace(t_r1, "@@TMP_O@@").replace(t_r2, "@@TMP_D@@")
        out = out.replace("@@TMP_O@@", n_r1).replace("@@TMP_D@@", n_r2)
    return out


_ZERO_WIDTH_REGION = "@@RNAME@@"  # placeholder unlikely to appear in real data


def substitute_region_names(text, template_code, new_code):
    """
    For free-text Fuel.*.Name columns. Replace REGION_NAME_MAP[template_region]
    with REGION_NAME_MAP[new_region] in a case-insensitive manner. Same
    placeholder-swap pattern for 13-char TRN codes.
    """
    if not isinstance(text, str):
        return text
    t_regs = regions_of(template_code)
    n_regs = regions_of(new_code)
    if not t_regs or not n_regs:
        return text

    if len(str(template_code)) == 11:
        t_r = next(iter(t_regs)); n_r = next(iter(n_regs))
        if t_r == n_r:
            return text
        return re.sub(re.escape(REGION_NAME_MAP[t_r]),
                      REGION_NAME_MAP[n_r], text, flags=re.IGNORECASE)

    # 13-char: two named regions; placeholder swap to avoid clobbering.
    _, t_r1, t_r2 = parse_code(template_code)
    _, n_r1, n_r2 = parse_code(new_code)
    out = text
    if t_r1 != n_r1 or t_r2 != n_r2:
        out = re.sub(re.escape(REGION_NAME_MAP[t_r1]),
                     _ZERO_WIDTH_REGION + "1", out, flags=re.IGNORECASE)
        out = re.sub(re.escape(REGION_NAME_MAP[t_r2]),
                     _ZERO_WIDTH_REGION + "2", out, flags=re.IGNORECASE)
        out = out.replace(_ZERO_WIDTH_REGION + "1", REGION_NAME_MAP[n_r1])
        out = out.replace(_ZERO_WIDTH_REGION + "2", REGION_NAME_MAP[n_r2])
    return out


def is_year_col(c):
    """True for column labels that are year integers or year-like strings."""
    if isinstance(c, int) and 2000 <= c <= 2100:
        return True
    if isinstance(c, str) and c.isdigit() and 2000 <= int(c) <= 2100:
        return True
    return False


def col_as_year_int(c):
    if isinstance(c, int):
        return c
    return int(c)


def lookup_tab2(tab2, tech, parameter, ao_dest_sheet, timeslice=None):
    """
    Three-tier lookup. Returns (row_dict_or_None, tier_label).
      tier_label in {"STRICT", "CROSS_SHEET", "TEMPLATE"}
    "TEMPLATE" means caller must fall back to the template row's values.
    Timeslice match is applied when the candidate row's Timeslices is non-null.
    """
    wv_sheet = AO_TO_WV_SHEET.get(ao_dest_sheet)

    base = tab2[(tab2["Tech"] == tech) & (tab2["Parameter"] == parameter)]
    if len(base) == 0:
        return None, "TEMPLATE"

    def _ts_filter(df):
        if timeslice is None or pd.isna(timeslice):
            # Prefer rows with no timeslice when caller doesn't supply one;
            # but accept whatever's there if all rows have timeslices.
            if "Timeslices" in df.columns:
                no_ts = df[df["Timeslices"].isna()]
                if len(no_ts) > 0:
                    return no_ts
            return df
        if "Timeslices" in df.columns:
            ts_match = df[df["Timeslices"] == timeslice]
            if len(ts_match) > 0:
                return ts_match
            no_ts = df[df["Timeslices"].isna()]
            if len(no_ts) > 0:
                return no_ts
        return df

    # Tier 1: strict (Tech, Parameter, Source_Sheet matches WV-equivalent).
    if wv_sheet is not None:
        strict = base[base["Source_Sheet"] == wv_sheet]
        strict = _ts_filter(strict)
        if len(strict) > 0:
            return strict.iloc[0].to_dict(), "STRICT"

    # Tier 2: cross-sheet (Tech, Parameter) anywhere.
    cross = _ts_filter(base)
    if len(cross) > 0:
        # Prefer the row with the most non-null year values
        year_cols = [c for c in cross.columns if is_year_col(c)]
        if year_cols:
            cross = cross.assign(_nonnull=cross[year_cols].notna().sum(axis=1))
            cross = cross.sort_values("_nonnull", ascending=False)
            cross = cross.drop(columns="_nonnull")
        return cross.iloc[0].to_dict(), "CROSS_SHEET"

    return None, "TEMPLATE"


def color_row(ws, row_num, color_hex):
    fill = PatternFill(start_color=color_hex, end_color=color_hex,
                       fill_type="solid")
    for cell in ws[row_num]:
        cell.fill = fill


# =============================================================================
# STEP 0 -- SET UP OUTPUT FOLDER (Setup-A)
# =============================================================================

print("=" * 72)
print("3_update_ao_from_extensions.py")
print("=" * 72)

out_dir = os.path.join(WORK_DIR, OUTPUT_SUBDIR)
os.makedirs(out_dir, exist_ok=True)
print(f"Output folder: {out_dir}")

# Hashes BEFORE we do anything so we can verify originals untouched.
src_hashes = {
    AO_PARAM:    md5_of(AO_PARAM),
    AO_AR_BASE:  md5_of(AO_AR_BASE),
    AO_AR_PROJ:  md5_of(AO_AR_PROJ),
    AO_DEMAND:   md5_of(AO_DEMAND),
}

# Output filenames.
def out_path_of(src):
    base = os.path.basename(src)
    stem, ext = os.path.splitext(base)
    return os.path.join(out_dir, stem + OUTPUT_SUFFIX + ext)

OUT_FILES = {label: out_path_of(p) for (label, p, _) in AO_WORKBOOKS}

# Setup-A: copy each source to its output path so we edit the COPY only.
print("\nSetup-A: copying source workbooks to output folder")
for (label, src, _) in AO_WORKBOOKS:
    dst = OUT_FILES[label]
    shutil.copyfile(src, dst)
    print(f"  {label:8s}  {os.path.basename(dst)}")

# =============================================================================
# STEP 1 -- LOAD EXTENSIONS WORKBOOK (TABS 1 + 2)
# =============================================================================

print("\n" + "=" * 72)
print("STEP 1 -- Load Extensions workbook")
print("=" * 72)

t1 = pd.read_excel(EXTENSIONS_FILE, sheet_name="1_Extensions_To_Add")
t2 = pd.read_excel(EXTENSIONS_FILE, sheet_name="2_Parameter_Rows_To_Replicate")

# Normalize Y/N columns to upper-case strings for safe comparison.
yn_cols = ["Add_To_Param", "Add_To_AR_Base", "Add_To_AR_Proj",
           "Add_To_Demand", "Include"]
for c in yn_cols:
    if c in t1.columns:
        t1[c] = t1[c].astype(str).str.strip().str.upper()

include_y = t1[t1["Include"] == "Y"].copy()
print(f"  Tab 1: {len(t1)} rows total, {len(include_y)} with Include=Y")
print(f"  Tab 2: {len(t2)} parameter rows captured from WV")

# =============================================================================
# STEP 2 -- INDEX TEMPLATES IN EACH A-O WORKBOOK
# =============================================================================

print("\n" + "=" * 72)
print("STEP 2 -- Index template locations in each A-O workbook")
print("=" * 72)

# For each output workbook label, build {sheet_name: dataframe} ONCE so we can
# scan template presence and read template rows without re-loading.
ao_data = {}      # label -> {sheet: dataframe}
for (label, src, _) in AO_WORKBOOKS:
    ao_data[label] = {}
    xl = pd.ExcelFile(src)
    for s in xl.sheet_names:
        ao_data[label][s] = pd.read_excel(xl, s)
    print(f"  {label:8s}  {len(ao_data[label])} sheets loaded")

def sheets_with_template(label, template_code):
    """List of sheet names where this template code appears in the workbook,
       AND the sheet has a Tech-like code column. Order preserved as in file."""
    out = []
    for s, df in ao_data[label].items():
        cc = find_code_col(df.columns)
        if cc is None:
            continue
        vals = df[cc].dropna().astype(str).values
        if template_code in vals:
            out.append(s)
    return out

# -----------------------------------------------------------------------------
# Open output workbooks for editing and initialize the audit log.
# Done here (not in Step 3) because Steps 2b and 2c need both the open
# workbook handles and the running log_lines list.
# -----------------------------------------------------------------------------
out_wbs = {}
for (label, src, _) in AO_WORKBOOKS:
    out_wbs[label] = load_workbook(OUT_FILES[label])

log_lines = []
log_lines.append("=" * 72)
log_lines.append("3_update_ao_from_extensions.py -- Run log")
log_lines.append(f"Run timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
log_lines.append("=" * 72)
log_lines.append("")
log_lines.append("INPUT")
log_lines.append(f"  Extensions:  {os.path.basename(EXTENSIONS_FILE)}  "
                 f"({len(t1)} rows, {len(include_y)} Include=Y)")
log_lines.append(f"  WV file:     {os.path.basename(WV_FILE)}")
for (label, src, _) in AO_WORKBOOKS:
    log_lines.append(f"  {label:8s}    {os.path.basename(src)}  md5={src_hashes[src]}")
log_lines.append("")

# Counters for summary (used by Step 3)
rows_added = {label: 0 for (label, _, _) in AO_WORKBOOKS}
tier_counts = {"STRICT": 0, "CROSS_SHEET": 0, "TEMPLATE": 0, "TAB2_ONLY": 0}
fallback_items = []   # (code, parameter, sheet, tier, value_used) for review
auto_named = []       # codes whose name had to be region-derived
skipped_codes = []
resolved = {}         # new_code -> (tmpl, tname) -- populated in Pass 1, consumed in 3B

# =============================================================================
# STEP 2B -- Refresh existing Demand_Projection rows from WV
# =============================================================================
# Row-level value sync, matched by Fuel/Tech. Rows where any year cell
# changes get colored light blue (DEMAND_REFRESH_COLOR). Codes present
# in A-O Demand_Projection but absent from WV are left untouched and
# recorded in the log.
print("\n" + "=" * 72)
print("STEP 2B -- Refresh existing Demand_Projection rows from WV")
print("=" * 72)

demand_refresh_count = 0
demand_refresh_log = []        # (code, n_cells_changed)
demand_refresh_skipped = []    # codes in A-O Demand_Projection not found in WV

wv_dp = pd.read_excel(WV_FILE, sheet_name="Demand_Projection")
wv_cc = find_code_col(wv_dp.columns)
if wv_cc is None:
    print("  WARNING: WV Demand_Projection has no recognized code column; "
          "skipping refresh")
    log_lines.append("STEP 2B: SKIPPED (WV Demand_Projection has no code column)")
else:
    wv_year_cols = [c for c in wv_dp.columns if is_year_col(c)]
    wv_lookup = {}
    for _, r in wv_dp.iterrows():
        code = r[wv_cc]
        if pd.isna(code):
            continue
        wv_lookup[str(code).strip()] = {y: r[y] for y in wv_year_cols}

    wb_demand = out_wbs["Demand"]
    if "Demand_Projection" not in wb_demand.sheetnames:
        print("  WARNING: output Demand workbook has no Demand_Projection sheet; "
              "skipping refresh")
        log_lines.append("STEP 2B: SKIPPED (no Demand_Projection sheet)")
    else:
        ws = wb_demand["Demand_Projection"]
        hdrs = [c.value for c in ws[1]]
        cc = find_code_col(hdrs)
        if cc is None:
            print("  WARNING: Demand_Projection has no recognized code column; "
                  "skipping refresh")
            log_lines.append("STEP 2B: SKIPPED (no code column)")
        else:
            ci = hdrs.index(cc) + 1                                    # 1-indexed
            year_col_idx = {h: hdrs.index(h) + 1
                            for h in hdrs if is_year_col(h)}

            # Walk only rows that exist BEFORE Step 3 appends. At this point
            # ws is exactly the Step-0 verbatim copy of source A-O.
            for r_num in range(2, ws.max_row + 1):
                code_val = ws.cell(row=r_num, column=ci).value
                if code_val is None:
                    continue
                code_str = str(code_val).strip()
                if code_str not in wv_lookup:
                    demand_refresh_skipped.append(code_str)
                    continue
                cells_changed = 0
                for h, col_idx in year_col_idx.items():
                    yint = col_as_year_int(h)
                    wv_v = wv_lookup[code_str].get(yint)
                    if pd.notna(wv_v):
                        old_v = ws.cell(row=r_num, column=col_idx).value
                        try:
                            if old_v is None or float(old_v) != float(wv_v):
                                ws.cell(row=r_num, column=col_idx).value = float(wv_v)
                                cells_changed += 1
                        except (TypeError, ValueError):
                            ws.cell(row=r_num, column=col_idx).value = float(wv_v)
                            cells_changed += 1
                if cells_changed > 0:
                    color_row(ws, r_num, DEMAND_REFRESH_COLOR)
                    demand_refresh_count += 1
                    demand_refresh_log.append((code_str, cells_changed))
                    print(f"  refreshed {code_str:14s}  ({cells_changed} year cells)")

            log_lines.append("")
            log_lines.append("STEP 2B -- Demand_Projection refresh from WV")
            log_lines.append(f"  Rows refreshed: {demand_refresh_count}")
            for code, n in demand_refresh_log:
                log_lines.append(f"    {code:14s}  {n} year cells updated")
            if demand_refresh_skipped:
                log_lines.append(f"  Codes in A-O not in WV (unchanged): "
                                 f"{demand_refresh_skipped}")

print(f"  Total rows refreshed: {demand_refresh_count}")

# =============================================================================
# STEP 2C -- Wholesale replace Profiles from WV Demand_Profiles
# =============================================================================
# Why wholesale: the timeslicing merge changed row count from 120 (3 dayparts
# x 4 seasons x 10 fuels) to 200 (5 dayparts x 4 seasons x 10 fuels). 1-to-1
# row matching is not possible. The columns are identical between A-O Profiles
# and WV Demand_Profiles; only the rows change. Sheet name in A-O is
# "Profiles"; in WV it is "Demand_Profiles" -- preserve the A-O sheet name in
# the output so downstream A-O machinery is unaffected.
print("\n" + "=" * 72)
print("STEP 2C -- Replace Profiles sheet wholesale from WV Demand_Profiles")
print("=" * 72)

profiles_replaced = 0

wv_xl_sheets = pd.ExcelFile(WV_FILE).sheet_names
if "Demand_Profiles" not in wv_xl_sheets:
    print("  WARNING: WV has no Demand_Profiles sheet; skipping replace")
    log_lines.append("STEP 2C: SKIPPED (WV has no Demand_Profiles sheet)")
else:
    wv_pf = pd.read_excel(WV_FILE, sheet_name="Demand_Profiles")

    wb_demand = out_wbs["Demand"]
    if "Profiles" not in wb_demand.sheetnames:
        print("  WARNING: output Demand workbook has no Profiles sheet; "
              "skipping replace")
        log_lines.append("STEP 2C: SKIPPED (no Profiles sheet)")
    else:
        ws = wb_demand["Profiles"]
        hdrs = [c.value for c in ws[1]]

        # Sanity-check columns. Headers should match exactly; if not, we
        # still align by column NAME (not position) when writing rows.
        wv_cols = list(wv_pf.columns)
        if hdrs != wv_cols:
            missing = [c for c in wv_cols if c not in hdrs]
            extra   = [c for c in hdrs   if c not in wv_cols]
            print(f"  NOTE: column header difference detected")
            if missing:
                print(f"    in WV not in Profiles: {missing}")
            if extra:
                print(f"    in Profiles not in WV: {extra}")
            log_lines.append(f"STEP 2C: column header difference -- WV-only "
                             f"{missing}, A-O-only {extra}; aligning by name")

        n_existing = ws.max_row - 1   # exclude header row 1
        # Delete all data rows (keep header row 1 with its formatting)
        if ws.max_row >= 2:
            ws.delete_rows(2, ws.max_row - 1)

        # Append WV rows in the worksheet's own header order.
        # NOTE on coloring: Step 2c rows are intentionally NOT colored.
        # The whole Profiles sheet comes from one source (WV), so per-row
        # color would be 100% blue and convey no information. Provenance
        # is captured in the run log instead. Step 2b coloring is kept
        # because there it provides actual contrast (refreshed-from-WV vs
        # unchanged) at the row level.
        for _, r in wv_pf.iterrows():
            row_vals = []
            for h in hdrs:
                # A-O quirk (same as Yearsplit/DaySplit): year column headers
                # may be STRING ('2023') in A-O while WV columns are INT.
                # Try both forms before giving up.
                v = None
                if h in r.index:
                    v = r.get(h)
                else:
                    try:
                        h_int = int(h)
                        if h_int in r.index:
                            v = r.get(h_int)
                    except (TypeError, ValueError):
                        pass
                if pd.isna(v):
                    v = None
                row_vals.append(v)
            ws.append(row_vals)
            profiles_replaced += 1

        log_lines.append("")
        log_lines.append("STEP 2C -- Profiles wholesale replace from WV Demand_Profiles")
        log_lines.append(f"  Old rows deleted: {n_existing}")
        log_lines.append(f"  New rows from WV: {profiles_replaced}")
        log_lines.append(f"  Timeslices in WV: "
                         f"{sorted(wv_pf['Timeslices'].dropna().unique())}")
        log_lines.append(f"  Fuels in WV: "
                         f"{sorted(wv_pf['Fuel/Tech'].dropna().unique())}")

print(f"  Profiles rows replaced: {profiles_replaced}")

# =============================================================================
# STEP 2D -- Refresh existing AO Parametrization rows from WV
# =============================================================================
# WV is the single source of truth. For five AO Parametrization sheets, refresh
# the value cells of every AO row whose match key also appears in the
# corresponding WV sheet. AO rows with no WV counterpart are left alone and
# logged. WV rows with no AO counterpart are NOT touched here -- Step 3's
# additive pass appends them.
#
# Sheet -> WV source -> match key:
#   Primary Techs            <- Primary_Techs            (Tech, Parameter)         year cells
#   Secondary Techs          <- Secondary_Techs          (Tech, Parameter)         year cells
#   Fixed Horizon Parameters <- Fixed_Horizon_Parameters (Tech, Parameter)         Value cell
#   VariableCost             <- VariableCost             (Tech, Mode.Operation)    year cells
#   Capacities               <- Capacities_CF            (Tech)                    block-replace
#                                                        all 12 AO rows for a tech
#                                                        deleted; 20 WV rows
#                                                        appended in their place
#                                                        (preserves Tech.ID)
#
# Capacities is special: it goes to passthrough at 20 timeslices. Existing AO
# techs that have a WV CF entry get their entire 12-row block replaced by the
# 20 WV rows. Tech.ID from the AO block is preserved. AO techs not in WV
# (PWRCSP*, PWRPETBGDXX) keep their 12 rows untouched.
#
# In-memory cache sync: ao_data[label][sheet] (a pandas dataframe) is consumed
# downstream by Step 3 Pass 1 to find template rows for cloning new techs. If
# Step 2D mutates only the openpyxl worksheet, Step 3 will read stale 12-ts
# values when cloning from refreshed templates (e.g., PWRSHPINDEA cloned from
# PWRHYDINDEA). Both representations are mutated in lockstep below.

print("\n" + "=" * 72)
print("STEP 2D -- Refresh existing AO Parametrization rows from WV")
print("=" * 72)

# Reuse Step 2B's color (semantically identical: row's year cells synced from WV).
PARAM_REFRESH_COLOR = DEMAND_REFRESH_COLOR

# -----------------------------------------------------------------------------
# 2D Helpers
# -----------------------------------------------------------------------------

def _norm_key_part(v):
    """Normalize a key cell value for stable hashing across pandas/openpyxl
    type drift (int 1 vs np.int64(1) vs '1.0', empty -> None)."""
    if v is None:
        return None
    if isinstance(v, float):
        if pd.isna(v):
            return None
        if v.is_integer():
            return str(int(v))
        return str(v)
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return str(v)


def build_wv_lookup(wv_df, key_cols, value_cols):
    """{key_tuple: {value_col_name: value}} -- one entry per WV row.
    Year columns and 'Value' alike are pulled into a single dict per row,
    keyed by the WV column name (int year or 'Value')."""
    out = {}
    for _, r in wv_df.iterrows():
        key = tuple(_norm_key_part(r[c]) if c in r.index else None for c in key_cols)
        if any(p is None for p in key):
            continue
        vals = {}
        for c in value_cols:
            if c in r.index:
                vals[c] = r[c]
        out[key] = vals
    return out


def refresh_year_cells_in_place(ws, ao_df, wv_lookup, key_cols,
                                value_col_intersection,
                                color_hex=PARAM_REFRESH_COLOR,
                                value_is_scalar_col=None):
    """
    Walk worksheet rows. For each row whose match-key is in wv_lookup,
    overwrite cells in `value_col_intersection` (year columns or 'Value').
    Color refreshed rows. Update ao_df in lockstep.
    Return: (n_refreshed, n_cells_changed, ao_only_keys_list).
    """
    hdrs = [c.value for c in ws[1]]

    # Map worksheet col-name -> 1-indexed col number for fast cell access.
    col_idx = {h: i + 1 for i, h in enumerate(hdrs)}

    # Resolve key column positions in the worksheet (must exist).
    for kc in key_cols:
        if kc not in col_idx:
            raise RuntimeError(f"Step 2D: key column {kc!r} missing from worksheet")

    n_refreshed = 0
    n_cells_changed = 0
    ao_only_keys = []
    matched_keys_seen = set()

    # Map worksheet row number -> ao_df index (1-to-1: ao_df rows correspond
    # to ws rows starting at row 2). Held by row order, which pandas preserves
    # when reading; this is the same convention Step 2B uses implicitly.
    ao_df_iter = ao_df.reset_index(drop=True)

    for r_num in range(2, ws.max_row + 1):
        # Build key from worksheet cells (authoritative for current state)
        key = tuple(
            _norm_key_part(ws.cell(row=r_num, column=col_idx[kc]).value)
            for kc in key_cols
        )
        if any(p is None for p in key):
            continue

        if key not in wv_lookup:
            ao_only_keys.append(key)
            continue

        matched_keys_seen.add(key)
        wv_vals = wv_lookup[key]
        cells_changed_this_row = 0

        for vc in value_col_intersection:
            if vc not in col_idx:
                continue
            wv_v = wv_vals.get(vc)
            if pd.isna(wv_v) if not isinstance(wv_v, (int, float, str, type(None))) else (wv_v is None or (isinstance(wv_v, float) and pd.isna(wv_v))):
                continue
            if wv_v is None:
                continue
            old_v = ws.cell(row=r_num, column=col_idx[vc]).value
            try:
                changed = (old_v is None) or (float(old_v) != float(wv_v))
            except (TypeError, ValueError):
                changed = (old_v != wv_v)
            if changed:
                # Coerce numeric WV values to float for spreadsheet consistency
                try:
                    new_v = float(wv_v)
                except (TypeError, ValueError):
                    new_v = wv_v
                ws.cell(row=r_num, column=col_idx[vc]).value = new_v
                # Mirror in ao_df cache (row index == r_num - 2)
                df_row_idx = r_num - 2
                if df_row_idx < len(ao_df_iter) and vc in ao_df_iter.columns:
                    ao_df_iter.at[df_row_idx, vc] = new_v
                cells_changed_this_row += 1

        if cells_changed_this_row > 0:
            color_row(ws, r_num, color_hex)
            n_refreshed += 1
            n_cells_changed += cells_changed_this_row

    return n_refreshed, n_cells_changed, ao_only_keys, matched_keys_seen, ao_df_iter


def replace_capacities_blocks_for_techs(ws, ao_df, wv_cap_df, color_hex=PARAM_REFRESH_COLOR):
    """
    For every tech that exists in BOTH AO Capacities and WV Capacities_CF:
      1. Capture the AO Tech.ID for that tech.
      2. Delete all AO rows for that tech from `ws` (preserve other rows).
      3. Append 20 WV rows at the bottom of `ws`, with Tech.ID set to the
         captured AO ID. All other columns aligned by header name.
    Return: (n_techs_refreshed, n_rows_deleted, n_rows_appended, ao_only_techs).
    """
    hdrs = [c.value for c in ws[1]]
    col_idx = {h: i + 1 for i, h in enumerate(hdrs)}

    # --- Identify match set ---
    ao_techs = set(_norm_key_part(t) for t in ao_df["Tech"].dropna().unique())
    ao_techs.discard(None)
    wv_techs = set(_norm_key_part(t) for t in wv_cap_df["Tech"].dropna().unique())
    wv_techs.discard(None)

    refresh_set = ao_techs & wv_techs
    ao_only_techs = sorted(ao_techs - wv_techs)

    # --- Capture AO Tech.IDs and per-tech row positions before deletion ---
    tech_id_lookup = {}     # norm_tech -> Tech.ID value (raw)
    rows_to_delete_by_tech = {}  # norm_tech -> [row_num, ...]
    tech_col = col_idx.get("Tech")
    techid_col = col_idx.get("Tech.ID")

    for r_num in range(2, ws.max_row + 1):
        t_raw = ws.cell(row=r_num, column=tech_col).value
        t = _norm_key_part(t_raw)
        if t is None or t not in refresh_set:
            continue
        rows_to_delete_by_tech.setdefault(t, []).append(r_num)
        if techid_col is not None and t not in tech_id_lookup:
            tid = ws.cell(row=r_num, column=techid_col).value
            if tid is not None:
                tech_id_lookup[t] = tid

    # --- Delete in reverse order to keep row numbers stable ---
    all_rows_to_delete = sorted(
        (r for rows in rows_to_delete_by_tech.values() for r in rows),
        reverse=True
    )
    for r_num in all_rows_to_delete:
        ws.delete_rows(r_num, 1)

    n_rows_deleted = len(all_rows_to_delete)

    # --- Append 20 WV rows for each refreshed tech, aligned by header name ---
    wv_year_cols = [c for c in wv_cap_df.columns if is_year_col(c)]
    n_rows_appended = 0
    appended_rows_for_color = []
    appended_dataframe_rows = []

    # Iterate WV in its own order so timeslices come out in WV's native order
    # (S1D1..S4D5).  Filter to the refresh set.
    norm_tech_series = wv_cap_df["Tech"].apply(_norm_key_part)
    wv_subset = wv_cap_df[norm_tech_series.isin(refresh_set)]

    for _, wv_r in wv_subset.iterrows():
        t = _norm_key_part(wv_r["Tech"])
        row_vals = []
        for h in hdrs:
            if h == "Tech.ID":
                row_vals.append(tech_id_lookup.get(t))   # preserved AO ID
            elif is_year_col(h):
                yint = col_as_year_int(h)
                v = wv_r.get(yint) if yint in wv_r.index else None
                if pd.isna(v):
                    v = None
                else:
                    try:
                        v = float(v)
                    except (TypeError, ValueError):
                        pass
                row_vals.append(v)
            else:
                v = wv_r.get(h) if h in wv_r.index else None
                if pd.isna(v):
                    v = None
                row_vals.append(v)
        ws.append(row_vals)
        appended_rows_for_color.append(ws.max_row)
        # Build dataframe-aligned dict
        df_row = {h: row_vals[i] for i, h in enumerate(hdrs)}
        appended_dataframe_rows.append(df_row)
        n_rows_appended += 1

    # Color appended rows.  (Note: we DO color these even though Step 2C
    # rationale was "whole sheet from one source = no contrast". Capacities
    # is mixed: refreshed techs (blue) vs untouched AO-only techs (no fill).
    # That contrast IS informative.)
    for r_num in appended_rows_for_color:
        color_row(ws, r_num, color_hex)

    # --- Rebuild ao_df cache to mirror new sheet contents exactly ---
    # Keep AO-only tech rows; replace refreshed tech rows with the appended
    # blocks.  Order: original AO-only rows first (preserve their original
    # order), then the appended WV rows.
    norm_ao_tech = ao_df["Tech"].apply(_norm_key_part)
    keep_mask = ~norm_ao_tech.isin(refresh_set)
    kept = ao_df[keep_mask].reset_index(drop=True)

    if appended_dataframe_rows:
        new_df = pd.concat(
            [kept, pd.DataFrame(appended_dataframe_rows, columns=ao_df.columns)],
            ignore_index=True
        )
    else:
        new_df = kept

    return (len(refresh_set), n_rows_deleted, n_rows_appended,
            ao_only_techs, new_df)


# -----------------------------------------------------------------------------
# 2D Driver
# -----------------------------------------------------------------------------

# Sheet plan: (ao_sheet, wv_sheet, key_cols, value_col_kind, scalar_col_or_None)
# value_col_kind in {"years", "value_scalar"}
STEP_2D_PLAN = [
    ("Primary Techs",            "Primary_Techs",            ["Tech", "Parameter"],         "years",        None),
    ("Secondary Techs",          "Secondary_Techs",          ["Tech", "Parameter"],         "years",        None),
    ("Fixed Horizon Parameters", "Fixed_Horizon_Parameters", ["Tech", "Parameter"],         "value_scalar", "Value"),
    ("VariableCost",             "VariableCost",             ["Tech", "Mode.Operation"],    "years",        None),
]

log_lines.append("")
log_lines.append("=" * 72)
log_lines.append("STEP 2D -- WV->AO refresh on Parametrization sheets")
log_lines.append("=" * 72)

step2d_summary = []   # (ao_sheet, refreshed, ao_only_count, ao_only_keys)

wb_param = out_wbs["Param"]

for ao_sheet, wv_sheet, key_cols, kind, scalar_col in STEP_2D_PLAN:
    if ao_sheet not in wb_param.sheetnames:
        msg = f"  SKIPPED  {ao_sheet}: not in Param workbook"
        print(msg); log_lines.append(msg)
        continue

    try:
        wv_df = pd.read_excel(WV_FILE, sheet_name=wv_sheet)
    except Exception as e:
        msg = f"  SKIPPED  {ao_sheet}: cannot load WV.{wv_sheet} ({e})"
        print(msg); log_lines.append(msg)
        continue

    ao_df = ao_data["Param"][ao_sheet]
    ws    = wb_param[ao_sheet]

    # Determine value-column intersection
    if kind == "years":
        wv_years = [c for c in wv_df.columns if is_year_col(c)]
        ao_years = [c for c in ao_df.columns if is_year_col(c)]
        # Match by integer-year identity (handles int/str header drift)
        wv_years_int = {col_as_year_int(c): c for c in wv_years}
        ao_years_int = {col_as_year_int(c): c for c in ao_years}
        common_years = sorted(set(wv_years_int) & set(ao_years_int))
        # value_cols inside wv_lookup are keyed by WV's column form
        value_cols_for_lookup = [wv_years_int[y] for y in common_years]
        # value_col_intersection passed to refresh fn uses AO's worksheet form
        value_col_intersection = [ao_years_int[y] for y in common_years]
        # If WV/AO use different header types (int vs str), need to translate.
        # We store under WV's key, but refresh_year_cells_in_place looks up by
        # ws header name. Easiest: re-key wv_lookup to AO's column names.
    else:  # value_scalar
        value_cols_for_lookup = [scalar_col]
        value_col_intersection = [scalar_col]
        common_years = []

    wv_lookup_raw = build_wv_lookup(wv_df, key_cols, value_cols_for_lookup)

    # Re-key year columns from WV form to AO form so lookup matches ws hdrs
    if kind == "years":
        wv_to_ao = {wv_years_int[y]: ao_years_int[y] for y in common_years}
        wv_lookup = {}
        for k, vd in wv_lookup_raw.items():
            wv_lookup[k] = {wv_to_ao[c]: v for c, v in vd.items() if c in wv_to_ao}
    else:
        wv_lookup = wv_lookup_raw

    n_refreshed, n_cells, ao_only_keys, matched, new_ao_df = \
        refresh_year_cells_in_place(
            ws, ao_df, wv_lookup, key_cols, value_col_intersection,
            color_hex=PARAM_REFRESH_COLOR
        )

    # Sync the in-memory cache so Step 3 Pass 1 reads refreshed values.
    ao_data["Param"][ao_sheet] = new_ao_df

    n_wv_only = len(set(wv_lookup.keys()) - matched)

    print(f"  {ao_sheet:30s}  refreshed {n_refreshed:5d} rows / "
          f"{n_cells:5d} cells   AO-only={len(ao_only_keys)}  "
          f"WV-only(for Step 3)={n_wv_only}")

    log_lines.append("")
    log_lines.append(f"  {ao_sheet} <- WV.{wv_sheet}  key={tuple(key_cols)}  kind={kind}")
    if kind == "years":
        log_lines.append(f"    Year columns synced: {len(common_years)} "
                         f"({common_years[0] if common_years else '-'}.."
                         f"{common_years[-1] if common_years else '-'})")
    log_lines.append(f"    AO rows total:        {ws.max_row - 1}")
    log_lines.append(f"    AO rows refreshed:    {n_refreshed}  "
                     f"({n_cells} cells changed)")
    log_lines.append(f"    AO rows untouched:    {len(ao_only_keys)}  "
                     f"(no WV counterpart)")
    log_lines.append(f"    WV-only keys:         {n_wv_only}  "
                     f"(Step 3 additive pass will append)")

    if ao_only_keys:
        log_lines.append(f"    Untouched AO keys (sample, up to 20):")
        for k in ao_only_keys[:20]:
            log_lines.append(f"      {k}")
        if len(ao_only_keys) > 20:
            log_lines.append(f"      ... and {len(ao_only_keys) - 20} more")

    step2d_summary.append((ao_sheet, n_refreshed, len(ao_only_keys), ao_only_keys))

# --- Capacities (block-replace, special) ---
ao_sheet_cap = "Capacities"
wv_sheet_cap = "Capacities_CF"

if ao_sheet_cap not in wb_param.sheetnames:
    msg = f"  SKIPPED  {ao_sheet_cap}: not in Param workbook"
    print(msg); log_lines.append(msg)
else:
    try:
        wv_cap_df = pd.read_excel(WV_FILE, sheet_name=wv_sheet_cap)
    except Exception as e:
        msg = f"  SKIPPED  {ao_sheet_cap}: cannot load WV.{wv_sheet_cap} ({e})"
        print(msg); log_lines.append(msg)
    else:
        ao_cap_df = ao_data["Param"][ao_sheet_cap]
        ws_cap = wb_param[ao_sheet_cap]

        n_techs, n_del, n_app, ao_only_techs, new_cap_df = \
            replace_capacities_blocks_for_techs(
                ws_cap, ao_cap_df, wv_cap_df, color_hex=PARAM_REFRESH_COLOR
            )

        # Sync in-memory cache so Step 3 sees the refreshed sheet
        ao_data["Param"][ao_sheet_cap] = new_cap_df

        wv_cap_techs = set(_norm_key_part(t) for t in wv_cap_df["Tech"].dropna().unique())
        wv_cap_techs.discard(None)
        ao_cap_techs_set = set(_norm_key_part(t) for t in ao_cap_df["Tech"].dropna().unique())
        ao_cap_techs_set.discard(None)
        wv_only_cap_techs = sorted(wv_cap_techs - ao_cap_techs_set)

        print(f"  {ao_sheet_cap:30s}  refreshed {n_techs:5d} techs   "
              f"({n_del} rows deleted, {n_app} rows appended)   "
              f"AO-only techs={len(ao_only_techs)}  "
              f"WV-only techs(for Step 3)={len(wv_only_cap_techs)}")

        log_lines.append("")
        log_lines.append(f"  {ao_sheet_cap} <- WV.{wv_sheet_cap}  "
                         f"key=(Tech)  kind=block_replace_20ts")
        log_lines.append(f"    Techs in BOTH (refreshed):    {n_techs}")
        log_lines.append(f"    AO rows deleted (was 12 ea):  {n_del}")
        log_lines.append(f"    WV rows appended (20 ea):     {n_app}")
        log_lines.append(f"    Tech.IDs preserved from AO:   {n_techs}")
        log_lines.append(f"    AO-only techs (untouched):    {len(ao_only_techs)}  "
                         f"-> {ao_only_techs}")
        log_lines.append(f"    WV-only techs (Step 3 to add): {len(wv_only_cap_techs)}")

        step2d_summary.append((ao_sheet_cap, n_techs, len(ao_only_techs), ao_only_techs))

print(f"  Step 2D complete.")

# =============================================================================
# STEP 3 -- BUILD AND APPEND NEW ROWS
# =============================================================================

print("\n" + "=" * 72)
print("STEP 3 -- Build and append new rows")
print("=" * 72)

for idx, row in include_y.reset_index(drop=True).iterrows():
    new_code = row["AO_Code_To_Add"]
    log_lines.append("-" * 72)
    log_lines.append(f"PROCESSING ROW {idx+1}/{len(include_y)}: {new_code}")

    # Resolve template
    tmpl = row.get("Override_Template_AO")
    tmpl_src = "Override_Template_AO"
    if pd.isna(tmpl) or str(tmpl).strip() == "":
        tmpl = row.get("Suggested_Template_AO")
        tmpl_src = "Suggested_Template_AO"
    if pd.isna(tmpl) or str(tmpl).strip() == "":
        log_lines.append(f"  SKIP: no template (both Suggested and Override blank)")
        skipped_codes.append(new_code)
        print(f"  [{idx+1:>2}/{len(include_y)}] {new_code:14s}  SKIP (no template)")
        continue
    tmpl = str(tmpl).strip()
    log_lines.append(f"  Template: {tmpl}  ({tmpl_src})")

    # Resolve Tech.Name
    tname = row.get("Override_Tech.Name_AO")
    tname_src = "Override_Tech.Name_AO"
    if pd.isna(tname) or str(tname).strip() == "":
        tname = row.get("Suggested_Tech.Name_AO")
        tname_src = "Suggested_Tech.Name_AO"
    if pd.isna(tname) or str(tname).strip() == "":
        # Auto-derive from template's Tech.Name in any A-O workbook.
        tmpl_name = None
        for label in ["Param", "AR_Base", "AR_Proj"]:
            for s, df in ao_data[label].items():
                cc = find_code_col(df.columns); nc = find_name_col(df.columns)
                if cc is None or nc is None:
                    continue
                hit = df[df[cc].astype(str) == tmpl]
                if len(hit) > 0:
                    cand = str(hit.iloc[0][nc]) if pd.notna(hit.iloc[0][nc]) else ""
                    if cand and (tmpl_name is None or len(cand) > len(tmpl_name)):
                        tmpl_name = cand
        if tmpl_name:
            tname = substitute_region_names(tmpl_name, tmpl, new_code)
            tname_src = "auto-derived (region substitution from template)"
            auto_named.append((new_code, tname))
        else:
            tname = new_code  # last resort
            tname_src = "auto-derived (template name not found, using code)"
            auto_named.append((new_code, tname))
    else:
        tname = str(tname).strip()
    log_lines.append(f"  Tech.Name: {tname}  ({tname_src})")

    # Record resolved (tmpl, tname) for consumption by Step 3B second pass.
    resolved[new_code] = (tmpl, tname)

    # For each Add_To_X workbook
    for (label, _, flag_col) in AO_WORKBOOKS:
        flag = str(row.get(flag_col, "N")).strip().upper()
        if flag != "Y":
            continue

        # Sheets where template lives in THIS workbook
        target_sheets = sheets_with_template(label, tmpl)
        log_lines.append(f"  {flag_col}=Y: template lives in {len(target_sheets)} sheet(s)")
        if not target_sheets:
            log_lines.append(f"    (no sheets host the template -- nothing to clone)")
            continue

        wb_out = out_wbs[label]

        for sheet_name in target_sheets:
            df = ao_data[label][sheet_name]
            cc = find_code_col(df.columns)
            nc = find_name_col(df.columns)
            tmpl_rows = df[df[cc].astype(str) == tmpl]
            if len(tmpl_rows) == 0:
                continue

            ws = wb_out[sheet_name]
            # Header order from the actual worksheet (authoritative for column
            # positions in case openpyxl/pandas disagree on order).
            ws_headers = [c.value for c in ws[1]]

            n_rows_added = 0
            n_strict = n_cross = n_tmpl = 0

            for _, trow in tmpl_rows.iterrows():
                # Build the new row keyed by the worksheet's own column order.
                new_vals = {}
                for h in ws_headers:
                    new_vals[h] = trow.get(h) if h in trow.index else None

                # Replace Tech / Tech.Name
                if cc and cc in new_vals:
                    new_vals[cc] = new_code
                if nc and nc in new_vals:
                    new_vals[nc] = tname

                # Region-substitute commodity-link string fields if present.
                fuel_link_cols = [h for h in ws_headers if isinstance(h, str)
                                  and h.startswith(("Fuel.I", "Fuel.O", "Fuel"))
                                  and h not in ("Value.Fuel.I", "Value.Fuel.O")]
                # Code-style fields (no spaces, ALLCAPS-ish): Fuel, Fuel.I, Fuel.O
                # Name-style fields (free text):              Fuel.Name, Fuel.I.Name, Fuel.O.Name
                for h in fuel_link_cols:
                    if not isinstance(h, str):
                        continue
                    if h.endswith(".Name") or "Name" in h:
                        new_vals[h] = substitute_region_names(new_vals.get(h), tmpl, new_code)
                    else:
                        new_vals[h] = substitute_code_regions(new_vals.get(h), tmpl, new_code)

                # Three-tier lookup -- determine Parameter and Timeslice.
                # Param/AR_Base/AR_Proj diverge in how Parameter is identified:
                if "Parameter" in ws_headers:
                    param = trow.get("Parameter")
                elif label == "AR_Base":
                    # AR_Base has no Parameter column and no year columns.
                    # Values are static fuel-linkage definitions; we fall back
                    # to template values + region substitution. Tag TEMPLATE.
                    param = None
                elif label == "AR_Proj":
                    # AR_Proj uses Direction (I/O) -> implicit Parameter.
                    direction = trow.get("Direction")
                    param = AR_DIRECTION_TO_PARAM.get(str(direction).strip().upper())
                else:
                    param = None

                ts = trow.get("Timeslices") if "Timeslices" in ws_headers else None

                year_cols = [h for h in ws_headers if is_year_col(h)]
                tier = "TEMPLATE"  # default if no Tab 2 match

                if param is not None:
                    t2_row, tier = lookup_tab2(t2, new_code, param, sheet_name, timeslice=ts)
                    if t2_row is not None:
                        # Year columns: overwrite where Tab 2 has a value.
                        # If Tab 2's year cell is empty but Value is populated
                        # (Fixed-Horizon-style row going into a year-column
                        # destination, e.g. AR_Proj's static InputActivityRatio
                        # = 1.0), broadcast Value across all year cells.
                        for yh in year_cols:
                            yint = col_as_year_int(yh)
                            if yint in t2_row and pd.notna(t2_row[yint]):
                                new_vals[yh] = t2_row[yint]
                            elif "Value" in t2_row and pd.notna(t2_row["Value"]):
                                new_vals[yh] = t2_row["Value"]
                        # Scalar Value column (Fixed Horizon Parameters style):
                        # prefer Tab 2's Value, fall back to its 2023 if Value
                        # is empty but year columns are populated (this is what
                        # gives TRN OperationalLife=40 in FH Params instead of
                        # falling back to the template).
                        if "Value" in ws_headers:
                            v = None
                            if "Value" in t2_row and pd.notna(t2_row["Value"]):
                                v = t2_row["Value"]
                            elif 2023 in t2_row and pd.notna(t2_row[2023]):
                                v = t2_row[2023]
                            if v is not None:
                                new_vals["Value"] = v
                        # Metadata: copy where the destination has the column
                        for meta in ("Unit", "Projection.Mode",
                                     "Projection.Parameter", "Tech.Type"):
                            if meta in ws_headers and meta in t2_row \
                                    and pd.notna(t2_row[meta]):
                                new_vals[meta] = t2_row[meta]

                # Append row in worksheet column order
                row_to_append = [new_vals.get(h) for h in ws_headers]
                ws.append(row_to_append)
                appended_row_num = ws.max_row
                color_row(ws, appended_row_num, TIER_COLOR[tier])

                # Tally
                tier_counts[tier] += 1
                if tier == "STRICT":      n_strict += 1
                elif tier == "CROSS_SHEET": n_cross += 1
                else:                       n_tmpl += 1
                if tier != "STRICT":
                    fallback_items.append((new_code, param or "(no-param)",
                                           f"{label}/{sheet_name}", tier))
                n_rows_added += 1

            log_lines.append(f"    {sheet_name:30s}  +{n_rows_added} rows  "
                             f"(strict={n_strict}, cross={n_cross}, tmpl={n_tmpl})")
            rows_added[label] += n_rows_added

    print(f"  [{idx+1:>2}/{len(include_y)}] {new_code:14s}  done")

# =============================================================================
# STEP 3B -- TAB2_ONLY second pass (template structural-gap fill)
# =============================================================================
# When a template tech has NO rows in a given A-O destination sheet, the
# Pass 1 loop silently skips that sheet.  This second pass detects those
# (new_code, label, ao_dest_sheet) gaps where Tab 2 does have rows (sourced
# from the WV equivalent sheet) and synthesises the missing rows directly.
#
# Currently the only affected mapping is:
#   Tab 2 Source_Sheet = "Capacities_CF"  ->  Param workbook, sheet "Capacities"
# (PWRNGSBGDXX, PWRNGSMDVXX: template PWRNGSINDEA has 0 A-O Capacities rows;
#  PWRWOFMDVXX: template PWRWOFBGDXX has 0 A-O Capacities rows.)
#
# Timeslice handling (passthrough at 20):
#   WV Capacities_CF carries 20 timeslices (5 dayparts per season x 4 seasons).
#   The A-O Capacities sheet now also runs at 20 timeslices, refreshed in
#   Step 2D for existing techs. The historical 20->12 YearSplit-weighted merge
#   has been removed (architectural debt called out in handover_step2d.md
#   section 5.1). Step 3B now emits the 20 WV rows verbatim, one per WV
#   timeslice. Year cells are copied directly from the Tab 2 row matching that
#   timeslice; no averaging is performed.

print("\n" + "=" * 72)
print("STEP 3B -- TAB2_ONLY second pass (template-gap fill)")
print("=" * 72)

# ---- Helpers ----------------------------------------------------------------

def ts_sort_key(ts):                                       # "S1D3" -> (1, 3)
    parts = str(ts).replace("S", "").replace("D", " ").split()
    return (int(parts[0]), int(parts[1])) if len(parts) == 2 else (0, 0)


# ---- Build inverse mapping: WV source-sheet -> [(label, ao_dest_sheet), ...] --
# Used to look up which workbook+sheet to target for each Tab 2 Source_Sheet.
WV_TO_AO_DEST = {}
for _ao_sheet, _wv_sheet in AO_TO_WV_SHEET.items():
    for (_lbl, _src_path, _) in AO_WORKBOOKS:
        if _ao_sheet in ao_data[_lbl]:
            WV_TO_AO_DEST.setdefault(_wv_sheet, []).append((_lbl, _ao_sheet))
            break   # first matching label wins per (wv_sheet, ao_sheet) pair

# ---- Second pass -----------------------------------------------------------
tab2_only_rows = 0

log_lines.append("")
log_lines.append("=" * 72)
log_lines.append("STEP 3B -- TAB2_ONLY second pass")
log_lines.append("=" * 72)

for idx, row in include_y.reset_index(drop=True).iterrows():
    new_code = row["AO_Code_To_Add"]
    if new_code not in resolved:
        continue   # skipped in Pass 1 (no template); nothing to do here

    tmpl, tname = resolved[new_code]

    for (label, _, flag_col) in AO_WORKBOOKS:
        flag = str(row.get(flag_col, "N")).strip().upper()
        if flag != "Y":
            continue

        # Sheets that Pass 1 covered (template was present in these)
        p1_sheets = set(sheets_with_template(label, tmpl))

        # All Tab 2 source sheets for this new_code
        tab2_srcs = t2[t2["Tech"] == new_code]["Source_Sheet"].dropna().unique()

        for wv_src in tab2_srcs:
            wv_src_str = str(wv_src)
            for (dest_label, ao_dest_sheet) in WV_TO_AO_DEST.get(wv_src_str, []):
                if dest_label != label:
                    continue
                if ao_dest_sheet in p1_sheets:
                    continue   # Pass 1 already handled this sheet

                # Gap confirmed -- Tab 2 has rows but Pass 1 emitted none.
                t2_gap = t2[(t2["Tech"] == new_code) &
                            (t2["Source_Sheet"] == wv_src_str)].copy()
                if len(t2_gap) == 0:
                    continue

                if ao_dest_sheet not in ao_data[label]:
                    log_lines.append(
                        f"  [TAB2_ONLY] {new_code}  {label}/{ao_dest_sheet}: "
                        f"sheet absent from workbook -- skipped")
                    continue

                wb_out    = out_wbs[label]
                ws        = wb_out[ao_dest_sheet]
                ws_headers = [c.value for c in ws[1]]
                cc         = find_code_col(ws_headers)
                nc         = find_name_col(ws_headers)
                year_cols_ws = [h for h in ws_headers if is_year_col(h)]

                # Determine whether Tab 2 rows carry timeslice data.
                ts_vals_in_tab2 = (
                    t2_gap["Timeslices"].dropna().astype(str).tolist()
                    if "Timeslices" in t2_gap.columns else []
                )

                if ts_vals_in_tab2:
                    # --- Timesliced rows: emit 20 WV rows verbatim ---
                    # (Historical 20->12 YearSplit-weighted merge removed;
                    #  see Step 3B header comment / handover_step2d.md s5.1.)
                    wv_ts_list = sorted(set(ts_vals_in_tab2), key=ts_sort_key)

                    n_appended = 0
                    for _, r2 in t2_gap.iterrows():
                        ts_cell = r2.get("Timeslices")
                        if pd.isna(ts_cell):
                            continue
                        ts_str = str(ts_cell)

                        new_vals = {h: None for h in ws_headers}
                        if cc and cc in new_vals:
                            new_vals[cc] = new_code
                        if nc and nc in new_vals:
                            new_vals[nc] = tname
                        if "Timeslices" in ws_headers:
                            new_vals["Timeslices"] = ts_str
                        for meta in ("Tech.ID", "Parameter.ID", "Parameter",
                                     "Unit", "Projection.Mode",
                                     "Projection.Parameter"):
                            if meta in ws_headers and meta in r2.index \
                                    and pd.notna(r2[meta]):
                                new_vals[meta] = r2[meta]
                        for yh in year_cols_ws:
                            yint = col_as_year_int(yh)
                            v = r2.get(yint) if yint in r2.index else None
                            if pd.isna(v):
                                v = None
                            else:
                                try:
                                    v = float(v)
                                except (TypeError, ValueError):
                                    pass
                            new_vals[yh] = v

                        ws.append([new_vals.get(h) for h in ws_headers])
                        color_row(ws, ws.max_row, TIER_COLOR["TAB2_ONLY"])
                        n_appended += 1

                    tier_counts["TAB2_ONLY"] += n_appended
                    rows_added[label]         += n_appended
                    tab2_only_rows            += n_appended
                    _note = f"passthrough {n_appended} ts rows"
                    log_lines.append(
                        f"  [TAB2_ONLY] {new_code:14s}  {label}/{ao_dest_sheet}"
                        f"  +{n_appended} rows  ({_note})")
                    print(f"  [TAB2_ONLY] {new_code:14s}  "
                          f"{label}/{ao_dest_sheet}  +{n_appended} rows"
                          f"  ({_note})")

                else:
                    # Non-timesliced rows from this Tab 2 source sheet.
                    # (Capacities_CF always has timeslices; this branch guards
                    # against future source sheets that don't.)
                    meta_src = t2_gap.iloc[0]
                    new_vals = {h: None for h in ws_headers}
                    if cc and cc in new_vals:
                        new_vals[cc] = new_code
                    if nc and nc in new_vals:
                        new_vals[nc] = tname
                    for meta in ("Tech.ID", "Parameter.ID", "Parameter",
                                 "Unit", "Projection.Mode",
                                 "Projection.Parameter"):
                        if meta in ws_headers and meta in meta_src.index \
                                and pd.notna(meta_src[meta]):
                            new_vals[meta] = meta_src[meta]
                    for yh in year_cols_ws:
                        yint = col_as_year_int(yh)
                        v = meta_src.get(yint)
                        new_vals[yh] = float(v) if pd.notna(v) else None

                    ws.append([new_vals.get(h) for h in ws_headers])
                    color_row(ws, ws.max_row, TIER_COLOR["TAB2_ONLY"])
                    tier_counts["TAB2_ONLY"] += 1
                    rows_added[label]         += 1
                    tab2_only_rows            += 1
                    log_lines.append(
                        f"  [TAB2_ONLY] {new_code:14s}  {label}/{ao_dest_sheet}"
                        f"  +1 row  (non-TS)")
                    print(f"  [TAB2_ONLY] {new_code:14s}  "
                          f"{label}/{ao_dest_sheet}  +1 row  (non-TS)")

if tab2_only_rows == 0:
    print("  (no template gaps found -- nothing to synthesise)")
    log_lines.append("  (no template gaps found)")
else:
    print(f"\n  STEP 3B: {tab2_only_rows} rows appended across all workbooks")

# =============================================================================
# STEP 4 -- SAVE OUTPUT WORKBOOKS
# =============================================================================

print("\n" + "=" * 72)
print("STEP 4 -- Save output workbooks")
print("=" * 72)
for (label, _, _) in AO_WORKBOOKS:
    out_wbs[label].save(OUT_FILES[label])
    print(f"  Saved: {os.path.basename(OUT_FILES[label])}  (+{rows_added[label]} rows)")

# =============================================================================
# STEP 5 -- WRITE AUDIT LOG
# =============================================================================

print("\n" + "=" * 72)
print("STEP 5 -- Write audit log")
print("=" * 72)

log_lines.append("")
log_lines.append("=" * 72)
log_lines.append("SUMMARY")
log_lines.append("=" * 72)
log_lines.append(f"  Codes processed: {len(include_y) - len(skipped_codes)}"
                 f" (skipped: {len(skipped_codes)})")
log_lines.append(f"  Rows added per workbook:")
for label in ("Param", "AR_Base", "AR_Proj", "Demand"):
    log_lines.append(f"    {label:8s}  {rows_added[label]:5d}")
log_lines.append(f"  Tier counts (one tag per appended row):")
log_lines.append(f"    STRICT       (green):   {tier_counts['STRICT']:5d}")
log_lines.append(f"    CROSS_SHEET  (yellow):  {tier_counts['CROSS_SHEET']:5d}")
log_lines.append(f"    TEMPLATE     (orange):  {tier_counts['TEMPLATE']:5d}")
log_lines.append(f"    TAB2_ONLY    (blue-vi): {tier_counts['TAB2_ONLY']:5d}")
log_lines.append("")
log_lines.append(f"  Auto-derived Tech.Names (review for category accuracy):")
if auto_named:
    for code, name in auto_named:
        log_lines.append(f"    {code:14s}  {name}")
else:
    log_lines.append("    (none)")
log_lines.append("")
log_lines.append(f"  Non-strict rows (CROSS_SHEET + TEMPLATE) -- review:")
if fallback_items:
    for code, param, where, tier in fallback_items:
        log_lines.append(f"    [{tier:11s}] {code:14s} {param:30s}  {where}")
else:
    log_lines.append("    (none)")

log_path = os.path.join(out_dir, LOG_FILE_NAME)
with open(log_path, "w", encoding="utf-8") as f:
    f.write("\n".join(log_lines) + "\n")
print(f"  Wrote: {log_path}")

# =============================================================================
# BUILT-IN TESTS
# =============================================================================

print("\n" + "=" * 72)
print("TESTS")
print("=" * 72)

_passed, _failed = 0, 0
def check(name, cond, detail=""):
    global _passed, _failed
    if cond:
        _passed += 1
        print(f"  PASS  {name}")
    else:
        _failed += 1
        print(f"  FAIL  {name}   {detail}")

# (a) Source workbooks byte-identical to before
for (label, src, _) in AO_WORKBOOKS:
    cur = md5_of(src)
    check(f"{label} source unchanged", cur == src_hashes[src],
          f"(was {src_hashes[src][:8]}, now {cur[:8] if cur else 'missing'})")

# (b) All four output workbooks exist
for label in ("Param", "AR_Base", "AR_Proj", "Demand"):
    p = OUT_FILES[label]
    check(f"output exists: {os.path.basename(p)}", os.path.exists(p))

# (c) For each Include=Y row, presence/absence per Add_To flag
for _, row in include_y.iterrows():
    code = row["AO_Code_To_Add"]
    for (label, _, flag_col) in AO_WORKBOOKS:
        flag = str(row.get(flag_col, "N")).strip().upper()
        # Reload OUT to inspect
        out_wb = load_workbook(OUT_FILES[label], read_only=True, data_only=True)
        present = False
        for s in out_wb.sheetnames:
            ws = out_wb[s]
            hdrs = [c.value for c in ws[1]]
            cc = find_code_col(hdrs)
            if cc is None:
                continue
            ci = hdrs.index(cc) + 1
            for r in ws.iter_rows(min_row=2, min_col=ci, max_col=ci, values_only=True):
                if r[0] == code:
                    present = True
                    break
            if present:
                break
        out_wb.close()
        if flag == "Y":
            check(f"{code} present in {label}", present, "(should be there)")
        else:
            check(f"{code} absent from {label}", not present, "(should NOT be there)")

# (d) Row-count delta non-negative per sheet (no rows deleted)
for (label, src, _) in AO_WORKBOOKS:
    src_xl = pd.ExcelFile(src)
    out_xl = pd.ExcelFile(OUT_FILES[label])
    for s in src_xl.sheet_names:
        if s not in out_xl.sheet_names:
            check(f"{label}: sheet '{s}' preserved", False, "(missing in output)")
            continue
        n_src = pd.read_excel(src_xl, s).shape[0]
        n_out = pd.read_excel(out_xl, s).shape[0]
        check(f"{label}/{s} row count >= source ({n_src})",
              n_out >= n_src, f"(got {n_out})")

# (e) AR rows for new codes have NO leftover template region tokens.
# Build per-code expected-region set and check Fuel-* string fields.
for label in ("AR_Base", "AR_Proj"):
    out_wb = load_workbook(OUT_FILES[label], read_only=True, data_only=True)
    leaks = []
    for _, row in include_y.iterrows():
        code = row["AO_Code_To_Add"]
        flag = str(row.get(f"Add_To_{label}", "N")).strip().upper()
        if flag != "Y":
            continue
        tmpl = row.get("Override_Template_AO")
        if pd.isna(tmpl) or str(tmpl).strip() == "":
            tmpl = row.get("Suggested_Template_AO")
        if pd.isna(tmpl) or str(tmpl).strip() == "":
            continue
        tmpl = str(tmpl).strip()
        t_regs = regions_of(tmpl)
        n_regs = regions_of(code)
        leftover = t_regs - n_regs
        if not leftover:
            continue
        for s in out_wb.sheetnames:
            ws = out_wb[s]
            hdrs = [c.value for c in ws[1]]
            cc = find_code_col(hdrs)
            if cc is None:
                continue
            ci = hdrs.index(cc)
            # Only check the rows for this new code (the cloned ones).
            for r in ws.iter_rows(min_row=2, values_only=True):
                if len(r) <= ci or r[ci] != code:
                    continue
                for h, v in zip(hdrs, r):
                    if not isinstance(v, str):
                        continue
                    if not isinstance(h, str):
                        continue
                    if not h.startswith(("Fuel", "Fuel.I", "Fuel.O")):
                        continue
                    for tr in leftover:
                        if tr in v:
                            leaks.append((label, s, code, h, v))
                        # also region-name leak
                        if REGION_NAME_MAP.get(tr) and \
                                REGION_NAME_MAP[tr].lower() in v.lower():
                            leaks.append((label, s, code, h, v))
    out_wb.close()
    check(f"no template region tokens leaked in {label}", len(leaks) == 0,
          f"(found {len(leaks)} leaks: {leaks[:3]})")

# (f) The 5 PWRSHP codes appear in all three target workbooks
shp_codes = ["PWRSHPINDEA", "PWRSHPINDNE", "PWRSHPINDNO",
             "PWRSHPINDSO", "PWRSHPINDWE"]
for label in ("Param", "AR_Base", "AR_Proj"):
    out_wb = load_workbook(OUT_FILES[label], read_only=True, data_only=True)
    found = set()
    for s in out_wb.sheetnames:
        ws = out_wb[s]
        hdrs = [c.value for c in ws[1]]
        cc = find_code_col(hdrs)
        if cc is None:
            continue
        ci = hdrs.index(cc) + 1
        for r in ws.iter_rows(min_row=2, min_col=ci, max_col=ci, values_only=True):
            if r[0] in shp_codes:
                found.add(r[0])
    out_wb.close()
    check(f"all 5 PWRSHP codes in {label}", found == set(shp_codes),
          f"(missing {set(shp_codes) - found})")

# (g) PWRGEOINDNO present and template was PWRHYDINDNO
geo_row = include_y[include_y["AO_Code_To_Add"] == "PWRGEOINDNO"]
if len(geo_row) > 0:
    geo_tmpl = (geo_row.iloc[0].get("Override_Template_AO") or
                geo_row.iloc[0].get("Suggested_Template_AO"))
    check("PWRGEOINDNO uses PWRHYDINDNO as template",
          str(geo_tmpl).strip() == "PWRHYDINDNO",
          f"(got {geo_tmpl})")

# (h) No PWR/SHP/TRN codes from Tab 1 are missing from where they should be
n_with_template = sum(
    1 for _, r in include_y.iterrows()
    if (str(r.get("Override_Template_AO") or "").strip() != "" or
        str(r.get("Suggested_Template_AO") or "").strip() != "")
)
check(f"all Include=Y rows had a template ({len(include_y)} expected)",
      n_with_template == len(include_y),
      f"({len(include_y) - n_with_template} skipped for missing template)")

# (i) Rows-added totals match what was logged
total_logged = sum(rows_added.values())
total_tier   = sum(tier_counts.values())
check("tier counts sum to rows added",
      total_logged == total_tier,
      f"(rows_added={total_logged} vs tiers={total_tier})")

# (j) After Step 2b refresh, every Fuel/Tech in Demand_Projection that exists
#     in WV's Demand_Projection has matching year-column values.
out_wb = load_workbook(OUT_FILES["Demand"], read_only=True, data_only=True)
ws = out_wb["Demand_Projection"]
hdrs = [c.value for c in ws[1]]
cc = find_code_col(hdrs)
ci = hdrs.index(cc) + 1
year_col_idx = {h: hdrs.index(h) + 1 for h in hdrs if is_year_col(h)}

wv_dp_check = pd.read_excel(WV_FILE, sheet_name="Demand_Projection")
wv_cc_check = find_code_col(wv_dp_check.columns)
wv_check_lookup = {}
for _, r in wv_dp_check.iterrows():
    if pd.isna(r[wv_cc_check]):
        continue
    wv_check_lookup[str(r[wv_cc_check]).strip()] = r

mismatches = []
for r_num in range(2, ws.max_row + 1):
    code = ws.cell(row=r_num, column=ci).value
    if code is None:
        continue
    code_str = str(code).strip()
    if code_str not in wv_check_lookup:
        continue
    wv_row = wv_check_lookup[code_str]
    for h, col_idx in year_col_idx.items():
        yint = col_as_year_int(h)
        if yint not in wv_row.index:
            continue
        wv_v = wv_row[yint]
        out_v = ws.cell(row=r_num, column=col_idx).value
        if pd.notna(wv_v) and out_v is not None:
            try:
                if abs(float(out_v) - float(wv_v)) > 1e-6:
                    mismatches.append((code_str, yint, out_v, wv_v))
            except (TypeError, ValueError):
                mismatches.append((code_str, yint, out_v, wv_v))
out_wb.close()
check("Demand_Projection year cells match WV for matched codes",
      len(mismatches) == 0,
      f"({len(mismatches)} mismatches; first: {mismatches[:3]})")

# (k) After Step 2c, output Profiles has the same row count, fuel set,
#     and timeslice set as WV's Demand_Profiles.
out_wb = load_workbook(OUT_FILES["Demand"], read_only=True, data_only=True)
ws = out_wb["Profiles"]
out_data_rows = ws.max_row - 1
out_wb.close()
wv_pf_check = pd.read_excel(WV_FILE, sheet_name="Demand_Profiles")
out_pf_check = pd.read_excel(OUT_FILES["Demand"], sheet_name="Profiles")
check(f"Profiles row count matches WV Demand_Profiles ({len(wv_pf_check)})",
      out_data_rows == len(wv_pf_check),
      f"(got {out_data_rows})")
check("Profiles fuel set matches WV Demand_Profiles",
      set(out_pf_check["Fuel/Tech"].dropna().unique()) ==
      set(wv_pf_check["Fuel/Tech"].dropna().unique()))
check("Profiles timeslice set matches WV Demand_Profiles",
      set(out_pf_check["Timeslices"].dropna().unique()) ==
      set(wv_pf_check["Timeslices"].dropna().unique()))

print(f"\n  {_passed} passed, {_failed} failed")
if _failed == 0:
    print("  ALL TESTS PASSED")
else:
    print(f"  {_failed} TEST(S) FAILED -- review output above")
