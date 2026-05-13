"""fix_pwrpet_clear.py
========================
Reproduce el clear manual que Luis hizo sobre la hoja Capacities:
para cada fila donde Tech == 'PWRPETBGDXX', borra (None) el valor en
TODAS las columnas de año (2023..2050). Esto es 12 timeslices x 28 años
= 336 celdas.

Justificacion: Luis decidio que los CapacityFactor auto-generados para
PWRPETBGDXX (valores muy bajos como 0.0086, 0.0248) no eran realistas
y los borro a mano. Sin valor, OSeMOSYS usa el AvailabilityFactor=0.8
como cota.

Idempotente: re-correrlo es no-op (las celdas ya estan vacias).

Uso:
    python fix_pwrpet_clear.py --input A-O_Parametrization.xlsx
    python fix_pwrpet_clear.py --input <path> --output <path>  # write a copy
"""
from __future__ import annotations
import argparse
from pathlib import Path
import sys
from openpyxl import load_workbook

TARGET_SHEET = "Capacities"
TARGET_TECH = "PWRPETBGDXX"
TECH_COL_NAME = "Tech"
YEAR_MIN, YEAR_MAX = 2023, 2050


def clear_pwrpet(input_path: Path, output_path: Path | None = None) -> int:
    """Clear PWRPETBGDXX year cells in Capacities sheet.

    Returns number of cells cleared.
    """
    if output_path is None:
        output_path = input_path
    wb = load_workbook(input_path)
    if TARGET_SHEET not in wb.sheetnames:
        sys.exit(f"ERROR: sheet '{TARGET_SHEET}' not in workbook")
    ws = wb[TARGET_SHEET]

    headers = {c.value: c.column for c in ws[1] if c.value is not None}
    if TECH_COL_NAME not in headers:
        sys.exit(f"ERROR: column '{TECH_COL_NAME}' not in {TARGET_SHEET} header")
    tech_col = headers[TECH_COL_NAME]

    year_cols = [headers[y] for y in range(YEAR_MIN, YEAR_MAX + 1) if y in headers]
    if len(year_cols) != (YEAR_MAX - YEAR_MIN + 1):
        # Defensive skip: when running on a non-canonical A1 base (e.g.
        # --skip-a3 against an input where Capacities doesn't yet have the
        # full year header), we can't safely target the right cells.
        # No-op rather than aborting the pipeline.
        print(f"  SKIP — expected {YEAR_MAX-YEAR_MIN+1} year columns in "
              f"'{TARGET_SHEET}' header, found {len(year_cols)}. "
              f"No cells cleared.")
        wb.save(output_path)
        print(f"Saved (unchanged): {output_path}")
        return 0

    cells_cleared = 0
    rows_touched = 0
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, tech_col).value != TARGET_TECH:
            continue
        rows_touched += 1
        for c in year_cols:
            cell = ws.cell(r, c)
            if cell.value is not None:
                cell.value = None
                cells_cleared += 1

    wb.save(output_path)
    print(f"Cleared {cells_cleared} cells across {rows_touched} {TARGET_TECH} rows")
    print(f"Saved: {output_path}")
    return cells_cleared


if __name__ == "__main__":
    p = argparse.ArgumentParser(description=__doc__.split("\n")[1])
    p.add_argument("--input", type=Path, required=True)
    p.add_argument("--output", type=Path, default=None)
    args = p.parse_args()
    if not args.input.is_file():
        sys.exit(f"ERROR: input not found: {args.input}")
    clear_pwrpet(args.input, args.output)
