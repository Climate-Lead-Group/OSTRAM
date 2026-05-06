"""
cap_trn_to_residual.py
======================

Pre-lid data fix for transmission interconnection (TRN*) techs in
the 'Secondary Techs' sheet. Caps each TRN tech's TotalAnnualMaxCapacity
to its ResidualCapacity year-by-year, and resets TotalAnnualMaxCapacityInvestment
to 9999 across the full 2023-2050 horizon.

Background
----------
TRN* techs in the c2a-patched workbook have:
  - ResidualCapacity flat across 2023-2050 (corrected upstream by
    fix_trn_residuals.py). This represents the existing physical
    interconnection capacity, e.g. TRNBGDXXINDEA = 2.5 GW every year.
  - TotalAnnualMaxCapacity empty (None) for 17 of 18 techs --
    no upper bound on total installed transmission capacity.
  - TotalAnnualMaxCapacityInvestment a mix of 9999 and small late-year
    values (0.5 or 2 GW/yr) -- these late-year caps are stale unbinding
    artifacts, not real planning constraints.

Modeling intent for this run: TRN capacity should be FROZEN at residual
across the horizon (no expansion of cross-border interconnects), so:
  - TotalAnnualMaxCapacity := ResidualCapacity (year-by-year copy).
    This caps the total installed capacity at the existing physical
    interconnection. Where ResCap is 0 (e.g. lines that never got built),
    MaxCap = 0 disables the tech entirely.
  - TotalAnnualMaxCapacityInvestment := 9999 across all years.
    With MaxCap pinned to ResCap, no investment can ever push total
    capacity above the cap, so the per-year investment limit is moot.
    Setting it to 9999 just removes a separate (and now redundant)
    constraint that the lid script would otherwise stomp on.

Scope
-----
ONLY patches the 18 TRN* techs in TRN_TECHS (allowlist for safety).
For each, requires all three parameter rows present:
  - ResidualCapacity              (read-only, source for MaxCap copy)
  - TotalAnnualMaxCapacity        (overwritten with ResCap values)
  - TotalAnnualMaxCapacityInvestment  (overwritten with 9999)

Does NOT touch:
  - Any non-TRN tech (PWR*, GEN*, etc.)
  - Any other parameter (CapitalCost, FixedCost, MinCapInv, AvailFactor, ...)
  - Any other sheet
  - The ResidualCapacity row itself (always read-only)

Notable cases
-------------
- TRNMDVXXINDSO has ResCap=0 and was previously flagged as killed by
  MaxCap=0 + MaxCapInv=0. Applying this rule uniformly leaves it killed
  via MaxCap=0, but flips MaxCapInv from 0 to 9999. Functionally
  identical (MaxCap=0 dominates) but worth knowing -- printed
  prominently in the run summary.
- TRNINDSOLKAXX, TRNLKAXXMDVXX have ResCap=0 (never-built lines).
  These too will be capped at MaxCap=0 after the patch.

Output
------
- Patched workbook written as a sibling with `_POST_TRN_CAP_<timestamp>`
  appended to the stem. The input file is left untouched.
- Per-tech summary printed to stdout.
- Re-running on the same input is safe -- a fresh sibling is produced
  each invocation.

Usage
-----
    python cap_trn_to_residual.py
    python cap_trn_to_residual.py --input path/to/A-O_Parametrization_FIXED.xlsx
    python cap_trn_to_residual.py --dry-run

Chain with clear_stale_unbinding_caps.py: order doesn't strictly matter
since they touch disjoint techs (TRN vs PWR), but conventional ordering
is FIXED -> POST_TRN_CAP -> POST_CAP_RESET -> lid script.
"""

from __future__ import annotations
import argparse
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
DEFAULT_INPUT = Path("A-O_Parametrization_c2a_patched_FIXED.xlsx")
SHEET_NAME = "Secondary Techs"
TECH_COL_NAME = "Tech"
PARAM_COL_NAME = "Parameter"

PARAM_RESCAP = "ResidualCapacity"
PARAM_MAXCAP = "TotalAnnualMaxCapacity"
PARAM_MAXCAPINV = "TotalAnnualMaxCapacityInvestment"

PLACEHOLDER = 9999
YEARS = list(range(2023, 2051))  # 2023 through 2050 inclusive (full horizon)

# Allowlist: 18 TRN* transmission interconnection techs.
TRN_TECHS = {
    "TRNBGDXXINDEA", "TRNBGDXXINDNE", "TRNBTNXXBGDXX", "TRNBTNXXINDEA",
    "TRNBTNXXINDNE", "TRNINDEAINDNE", "TRNINDEAINDNO", "TRNINDEAINDSO",
    "TRNINDEAINDWE", "TRNINDEANPLXX", "TRNINDNEINDNO", "TRNINDNOINDWE",
    "TRNINDNONPLXX", "TRNINDSOINDWE", "TRNINDSOLKAXX", "TRNLKAXXMDVXX",
    "TRNMDVXXINDSO", "TRNNPLXXBGDXX",
}


# ---------------------------------------------------------------------------
# Implementation
# ---------------------------------------------------------------------------
def make_output_path(input_path: Path) -> Path:
    """Return a timestamped POST_TRN_CAP sibling path. Does not create it."""
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return input_path.with_name(
        f"{input_path.stem}_POST_TRN_CAP_{stamp}{input_path.suffix}"
    )


def patch_workbook(input_path: Path, dry_run: bool) -> None:
    if not input_path.is_file():
        sys.exit(f"ERROR: input file not found: {input_path}")

    print(f"Input: {input_path}")
    print(f"Sheet: {SHEET_NAME}")
    print(f"Allowlisted TRN techs: {len(TRN_TECHS)}")
    print(f"Year range (inclusive): {YEARS[0]}-{YEARS[-1]}")
    print(f"Operations:")
    print(f"  {PARAM_MAXCAP:35s} := {PARAM_RESCAP} (year-by-year)")
    print(f"  {PARAM_MAXCAPINV:35s} := {PLACEHOLDER}")
    print(f"Dry run: {dry_run}")
    print()

    if not dry_run:
        output_path = make_output_path(input_path)
        print(f"Output will be written to: {output_path.name}")
        print()

    wb = load_workbook(input_path)
    if SHEET_NAME not in wb.sheetnames:
        sys.exit(f"ERROR: sheet '{SHEET_NAME}' not found in workbook")
    ws = wb[SHEET_NAME]

    # Map header values to 1-indexed column numbers
    headers: dict = {}
    for cell in ws[1]:
        if cell.value is not None:
            headers[cell.value] = cell.column

    tech_col = headers.get(TECH_COL_NAME)
    param_col = headers.get(PARAM_COL_NAME)
    if tech_col is None or param_col is None:
        sys.exit(
            f"ERROR: missing required header. "
            f"tech_col='{TECH_COL_NAME}'={tech_col}, "
            f"param_col='{PARAM_COL_NAME}'={param_col}"
        )

    year_cols: dict = {}
    for yr in YEARS:
        col = headers.get(yr)
        if col is None:
            sys.exit(f"ERROR: year column {yr} not found in header row")
        year_cols[yr] = col

    # Index TRN rows: tech -> {param -> row (tuple of cells)}
    # We need real cells (not values_only) so we can mutate.
    trn_rows: dict = {}
    for row in ws.iter_rows(min_row=2):
        tech = row[tech_col - 1].value
        param = row[param_col - 1].value
        if tech not in TRN_TECHS:
            continue
        if param not in (PARAM_RESCAP, PARAM_MAXCAP, PARAM_MAXCAPINV):
            continue
        trn_rows.setdefault(tech, {})[param] = row

    # Validate: every allowlisted tech should have all 3 rows.
    log = []        # (tech, rescap_summary, maxcap_changes, maxcapinv_changes, notable_flag)
    skipped = []    # (tech, reason)

    for tech in sorted(TRN_TECHS):
        rows = trn_rows.get(tech, {})
        missing = [p for p in (PARAM_RESCAP, PARAM_MAXCAP, PARAM_MAXCAPINV)
                   if p not in rows]
        if missing:
            skipped.append((tech, f"missing parameter rows: {missing}"))
            continue

        rescap_row = rows[PARAM_RESCAP]
        maxcap_row = rows[PARAM_MAXCAP]
        maxcapinv_row = rows[PARAM_MAXCAPINV]

        # Read source ResCap values and check no Nones (would be ambiguous)
        rescap_vals = [rescap_row[year_cols[y] - 1].value for y in YEARS]
        if any(v is None for v in rescap_vals):
            none_years = [y for y, v in zip(YEARS, rescap_vals) if v is None]
            skipped.append(
                (tech, f"ResidualCapacity has None at {len(none_years)} years "
                       f"(first: {none_years[0]}) -- ambiguous, refusing to copy")
            )
            continue

        rescap_uniq = sorted(set(rescap_vals))
        rescap_summary = (f"flat={rescap_vals[0]}" if len(rescap_uniq) == 1
                          else f"varies in {rescap_uniq}")

        # Capture old values for notable-case detection
        old_maxcap = [maxcap_row[year_cols[y] - 1].value for y in YEARS]
        old_maxcapinv = [maxcapinv_row[year_cols[y] - 1].value for y in YEARS]

        # Apply: MaxCap := ResCap (year-by-year)
        maxcap_changes = 0
        for y, src in zip(YEARS, rescap_vals):
            cell = maxcap_row[year_cols[y] - 1]
            if cell.value != src:
                if not dry_run:
                    cell.value = src
                maxcap_changes += 1

        # Apply: MaxCapInv := 9999
        maxcapinv_changes = 0
        for y in YEARS:
            cell = maxcapinv_row[year_cols[y] - 1]
            if cell.value != PLACEHOLDER:
                if not dry_run:
                    cell.value = PLACEHOLDER
                maxcapinv_changes += 1

        # Notable: tech that was previously killed via MaxCap=0 + MaxCapInv=0
        was_killed = (
            all(v == 0 for v in old_maxcap if v is not None)
            and any(v == 0 for v in old_maxcap)
            and all(v == 0 for v in old_maxcapinv if v is not None)
            and any(v == 0 for v in old_maxcapinv)
        )
        notable = "  [was killed via MaxCap=0+MaxCapInv=0]" if was_killed else ""

        log.append((tech, rescap_summary, maxcap_changes, maxcapinv_changes, notable))

    # Report
    print(f"TRN techs processed: {len(log)} / {len(TRN_TECHS)} allowlisted")
    print()

    if log:
        print(f"Patch summary ({len(log)} techs, "
              f"{sum(a + b for _, _, a, b, _ in log)} cells changed):")
        print(f"  {'Tech':16s} {'ResCap':28s} "
              f"{'MaxCap cells':>12s} {'MaxCapInv cells':>16s}")
        print(f"  {'-'*78}")
        for tech, summary, mc, mci, notable in log:
            print(f"  {tech:16s} {summary:28s} {mc:>12d} {mci:>16d}{notable}")

    if skipped:
        print()
        print(f"Skipped {len(skipped)} techs (signature mismatch -- safety abort):")
        for tech, reason in skipped:
            print(f"  {tech}: {reason}")

    if dry_run:
        print()
        print("DRY RUN -- no changes written.")
        return

    wb.save(output_path)
    print()
    print(f"Saved: {output_path}")
    print(f"Input file untouched: {input_path}")


def main() -> None:
    p = argparse.ArgumentParser(description=__doc__.split("\n")[2])
    p.add_argument(
        "--input", type=Path, default=DEFAULT_INPUT,
        help=f"Path to the workbook to patch (default: {DEFAULT_INPUT})",
    )
    p.add_argument(
        "--dry-run", action="store_true",
        help="Print what would change but do not write",
    )
    args = p.parse_args()
    patch_workbook(args.input, args.dry_run)


if __name__ == "__main__":
    main()
