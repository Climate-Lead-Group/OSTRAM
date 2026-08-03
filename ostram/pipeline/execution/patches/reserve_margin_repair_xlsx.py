"""Careful reserve-margin repair using an XLSX fallback workbook.

This script is the XLSX sibling of patch_reserve_margin_repair_careful.py.
It keeps stock and flow capacity limits separate and only replaces sentinel
values, by default 0 and 9999.
"""

from __future__ import annotations

import argparse
import hashlib
import os
import sys
from pathlib import Path

from openpyxl import load_workbook

from ostram.paths import (
    WINDOWS_SAFE_ABSOLUTE_PATH_BUDGET,
    WorkspacePathBudgetError,
    windows_path_units,
)

from .reserve_margin_repair import (
    FLOW_PARAM,
    STOCK_PARAM,
    CapFallback,
    build_tag_targets,
    capacity_floors,
    fmt_number,
    make_fallback_lookup,
    parse_assignment_list,
    parse_float,
    patch_capacity_param,
    patch_reserve_tags,
    select_years,
    stock_flow_warnings,
    techs_with_any_min_investment,
)


_WARNING_PATH_DIGEST_LENGTH = 32
_WARNING_PATH_PREFIX = "RMCarefulXLSX"


def bounded_warnings_path(
    desired_path: str | os.PathLike[str],
    *,
    budget: int = WINDOWS_SAFE_ABSOLUTE_PATH_BUDGET,
) -> Path:
    """Return the supplied warning path, compacting only long absolute names.

    Relative paths remain relative so this ephemeral-output correction never
    infers a workspace from the current working directory.  Long absolute
    paths retain their supplied parent and use a digest of the complete desired
    path to keep independently generated warning files distinct.
    """

    desired = Path(desired_path)
    if budget <= 0:
        raise ValueError(f"path budget must be positive, got {budget}")
    if windows_path_units(desired) < budget:
        return desired
    if not desired.is_absolute():
        raise WorkspacePathBudgetError(
            "over-budget reserve-margin warning paths must be absolute; "
            "refusing to infer an output root from the current working directory"
        )

    normalized = os.path.normcase(os.fspath(desired))
    digest = hashlib.sha256(normalized.encode("utf-8")).hexdigest()[
        :_WARNING_PATH_DIGEST_LENGTH
    ]
    compact_name = f"{_WARNING_PATH_PREFIX}_{digest}.warnings.txt"
    compact = desired.with_name(compact_name)
    compact_units = windows_path_units(compact)
    if compact_units >= budget:
        parent_units = windows_path_units(desired.parent)
        available = budget - parent_units - 2
        raise WorkspacePathBudgetError(
            "reserve-margin warnings parent leaves no Windows-safe filename "
            f"budget: parent={desired.parent!s} parent_length={parent_units} "
            f"budget={budget} available_filename_units={max(0, available)} "
            f"required_filename={compact_name!r}. The absolute warning path "
            f"must be shorter than {budget} UTF-16 code units."
        )
    return compact


def normalize_header(value: object) -> str:
    return str(value).strip() if value is not None else ""


def cell_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def pick_value(row: dict[str, object], *candidates: str) -> str | None:
    lower_map = {key.lower(): key for key in row}
    for candidate in candidates:
        key = lower_map.get(candidate.lower())
        if key is None:
            continue
        value = cell_text(row[key])
        if value is not None:
            return value
    return None


def load_fallback_xlsx(path: Path, sheet_name: str | None = None) -> list[CapFallback]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet {sheet_name!r} not found in {path}")
        sheet = workbook[sheet_name]
    else:
        sheet = workbook[workbook.sheetnames[0]]

    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [normalize_header(value) for value in next(rows)]
    except StopIteration as exc:
        raise ValueError(f"Workbook sheet is empty: {path}") from exc

    if not any(headers):
        raise ValueError(f"Workbook sheet has an empty header row: {path}")

    fallbacks: list[CapFallback] = []
    for row_number, values in enumerate(rows, start=2):
        row = {header: value for header, value in zip(headers, values) if header}
        if not any(cell_text(value) is not None for value in row.values()):
            continue

        cr = pick_value(row, "CR", "COUNTRY_REGION", "REGION_CODE", "REGION")
        if not cr:
            raise ValueError(f"{path}:{row_number}: missing CR value")

        prefix = pick_value(row, "TECH_PREFIX", "TECHNOLOGY_PREFIX", "PREFIX") or "*"
        year = pick_value(row, "YEAR") or "*"
        stock = pick_value(row, STOCK_PARAM, "STOCK", "MAX_CAPACITY", "TOTAL_CAPACITY")
        flow = pick_value(
            row,
            FLOW_PARAM,
            "FLOW",
            "MAX_INVESTMENT",
            "MAX_NEW_CAPACITY",
            "ANNUAL_INVESTMENT",
        )

        if stock is not None:
            fallbacks.append(
                CapFallback(
                    parameter=STOCK_PARAM,
                    prefix=prefix,
                    cr=cr,
                    year=year,
                    value=parse_float(stock, f"{path}:{row_number}:{STOCK_PARAM}"),
                )
            )
        if flow is not None:
            fallbacks.append(
                CapFallback(
                    parameter=FLOW_PARAM,
                    prefix=prefix,
                    cr=cr,
                    year=year,
                    value=parse_float(flow, f"{path}:{row_number}:{FLOW_PARAM}"),
                )
            )

    return fallbacks


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Patch reserve tags and carefully replace sentinel firm-capacity caps "
            "using an XLSX fallback workbook."
        )
    )
    parser.add_argument("input_file", help="Input OSeMOSYS datafile.")
    parser.add_argument("-o", "--output", required=True, help="Patched output datafile.")

    parser.add_argument("--fallback-xlsx", required=True, help="XLSX fallback workbook.")
    parser.add_argument("--xlsx-sheet", help="Worksheet name. Defaults to the first sheet.")

    parser.add_argument("--backstop-credit", type=float, default=1.0)
    parser.add_argument("--ccs-credit", type=float, default=0.9)
    parser.add_argument("--skip-backstop-credit", action="store_true")
    parser.add_argument("--skip-ccs-credit", action="store_true")
    parser.add_argument("--backstop-prefixes", nargs="+", default=["PWRBCK"])
    parser.add_argument("--ccs-prefixes", nargs="+", default=["PWRCCS"])

    parser.add_argument("--target-prefixes", nargs="+", default=["PWRPET", "PWROIL", "PWRNGS"])
    parser.add_argument(
        "--sentinel-values",
        nargs="+",
        type=float,
        default=[0.0, 9999.0],
        help="Only these existing cap values are replaced.",
    )
    parser.add_argument(
        "--stock-fallback",
        nargs="*",
        default=[],
        metavar="CR=VALUE",
        help=f"Optional CLI fallback for {STOCK_PARAM}; overrides general workbook values when more specific.",
    )
    parser.add_argument(
        "--flow-fallback",
        nargs="*",
        default=[],
        metavar="CR=VALUE",
        help=f"Optional CLI fallback for {FLOW_PARAM}; overrides general workbook values when more specific.",
    )
    parser.add_argument(
        "--warnings-file",
        help="Optional file to write warnings to. Warnings are always printed to stderr.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_path = Path(args.input_file)
    output_path = Path(args.output)

    lines = input_path.read_text(encoding="utf-8").splitlines(keepends=True)
    years = select_years(lines)

    backstop_credit = None if args.skip_backstop_credit else args.backstop_credit
    ccs_credit = None if args.skip_ccs_credit else args.ccs_credit
    tag_targets, tag_counts = build_tag_targets(
        lines,
        years,
        backstop_credit,
        ccs_credit,
        args.backstop_prefixes,
        args.ccs_prefixes,
    )

    fallbacks = load_fallback_xlsx(Path(args.fallback_xlsx), args.xlsx_sheet)
    fallbacks.extend(parse_assignment_list(args.stock_fallback, STOCK_PARAM))
    fallbacks.extend(parse_assignment_list(args.flow_fallback, FLOW_PARAM))
    fallback_lookup = make_fallback_lookup(fallbacks)

    patched, tag_updates, tag_inserts = patch_reserve_tags(lines, tag_targets)
    stock_floors, flow_floors = capacity_floors(patched, args.target_prefixes)
    min_investment_techs = techs_with_any_min_investment(patched, args.target_prefixes)
    patched, stock_changed, stock_skipped, stock_warnings = patch_capacity_param(
        patched,
        STOCK_PARAM,
        args.target_prefixes,
        [0.0],
        fallback_lookup,
        stock_floors,
    )
    patched, flow_changed, flow_skipped, flow_warnings = patch_capacity_param(
        patched,
        FLOW_PARAM,
        args.target_prefixes,
        args.sentinel_values,
        fallback_lookup,
        flow_floors,
        min_investment_techs,
    )
    consistency_warnings = stock_flow_warnings(patched, args.target_prefixes)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("".join(patched), encoding="utf-8")

    warnings = stock_warnings + flow_warnings + consistency_warnings
    if warnings:
        for warning in warnings:
            print(f"[WARN] {warning}", file=sys.stderr)
    if args.warnings_file:
        warning_text = "\n".join(warnings) + "\n" if warnings else "No warnings.\n"
        bounded_warnings_path(args.warnings_file).write_text(
            warning_text,
            encoding="utf-8",
        )

    print(f"Wrote patched datafile: {output_path}")
    print(f"Fallback workbook: {Path(args.fallback_xlsx)}")
    print(f"Years patched for reserve tags: {len(years)} ({years[0]}-{years[-1]})")
    if backstop_credit is not None:
        print(
            "Backstop reserve tag: "
            f"{tag_counts.get('backstop_techs', 0)} techs at {fmt_number(backstop_credit)}"
        )
    if ccs_credit is not None:
        print(f"CCS reserve tag: {tag_counts.get('ccs_techs', 0)} techs at {fmt_number(ccs_credit)}")
    print(f"ReserveMarginTagTechnology rows updated: {tag_updates}")
    print(f"ReserveMarginTagTechnology rows inserted: {tag_inserts}")
    print(f"{STOCK_PARAM} sentinel rows changed: {stock_changed}")
    print(f"{STOCK_PARAM} sentinel rows skipped: {stock_skipped}")
    print(f"{FLOW_PARAM} sentinel rows changed: {flow_changed}")
    print(f"{FLOW_PARAM} sentinel rows skipped: {flow_skipped}")
    print(f"Warnings: {len(warnings)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
