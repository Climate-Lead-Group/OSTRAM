from pathlib import Path
# -*- coding: utf-8 -*-
"""
2_extract_ao_extensions.py  (v17 / pipeline stage 2)

Reads SOASIA_OSeMOSYS_WV.xlsx (output of stage 1) and the four A-O
parametrization workbooks; writes OSTRAM_AO_Extensions.xlsx -- the
freshly generated review file that receives the maintained decision overlay
before it drives stage 3.

Pipeline:
  1_merge_timeslices_into_WV.py    v17 + timeslices       -> WV
  2_extract_ao_extensions.py       WV  + A-O              -> Extensions xlsx
  apply_ao_extension_decisions.py  maintained decisions   -> overlaid Extensions
  3_update_ao_from_extensions.py   Extensions xlsx + A-O  -> updated A-O

Output: OSTRAM_AO_Extensions.xlsx
  Tab 1: 1_Extensions_To_Add        Single wide table covering presence,
                                    propagation suggestions, template
                                    suggestions, and human-decision columns.
  Tab 2: 2_Parameter_Rows_To_Replicate  Raw parameter rows from WV for each
                                        extension code (reference for stage 3).
  Tab 3: 3_Signal_Disagreements    Sanity check: orange flags vs. absence.

Detection logic
---------------
  ABSENCE is the primary signal: a code is "missing from A-O" iff it's in WV
  but not in any of the four A-O workbooks. That's binary, file-content fact.

  COLOR (orange = FCD5B4) is the team-applied secondary signal carried over
  from the v15/v16/v17 cleanup. Disagreements between absence and color are
  reported in Tab 3.

Template suggestion (option 3, conservative)
--------------------------------------------
  For each absent code, look in A-O Parametrization for a code with the SAME
  prefix + SAME category (e.g., PWRNGS for a missing PWRNGSBGDXX).
   - Match found -> suggest it as Suggested_Template_AO; copy its Tech.Name.
   - No match    -> leave Suggested_Template_AO blank and tag the row
                    [NEEDS_REVIEW] in Notes. No proxy fabrication.
  Same rule drives the Add_To_* propagation suggestion: Y if the template
  is present in that workbook, N otherwise. Blank if no template.

Run in Spyder with F5. Paths and mappings in USER CONFIGURATION below.
"""

import os
import pandas as pd
from openpyxl import load_workbook

# =============================================================================
# USER CONFIGURATION
# =============================================================================

WORK_DIR = str(Path(os.environ["OSTRAM_STAGE_WORKDIR"]).resolve())

WV_FILE     = WORK_DIR + "/SOASIA_OSeMOSYS_WV.xlsx"

AO_PARAM    = WORK_DIR + "/A-O_Parametrization.xlsx"
AO_AR_BASE  = WORK_DIR + "/A-O_AR_Model_Base_Year.xlsx"
AO_AR_PROJ  = WORK_DIR + "/A-O_AR_Projections.xlsx"
AO_DEMAND   = WORK_DIR + "/A-O_Demand.xlsx"

OUTPUT_FILE = WORK_DIR + "/OSTRAM_AO_Extensions.xlsx"

# Categories introduced or restructured by v17 cleanup. Codes whose category
# is in this list get Include="Y" by default (definite v17-driven additions).
# Everything else defaults to Include="" so nothing happens until you decide.
V17_NEW_CATEGORIES = ["SHP"]

# When no same-category template exists in A-O Parametrization, fall back to
# this analogue category. Used only as a SECOND-CHANCE lookup; the primary
# rule (same prefix + same category) is tried first.
# When a fallback IS used, the row's Suggested_Tech.Name_AO is left blank so
# you must consciously override the name (no silent name extension).
CATEGORY_FALLBACKS = {
    "SHP": "HYD",   # Small Hydropower inherits Large Hydropower pattern
    # Add more here as you decide them, e.g.:
    # "GEO": "BIO",   # Geothermal -> Biomass (baseload analogue)
}

# WV sheets scanned for tech codes when building / verifying the roster.
# Mirrors V17_TECH_SHEETS in 1_merge_timeslices_into_WV.py.
WV_TECH_SHEETS = [
    "Primary_Techs", "Secondary_Techs", "Capacities_CF",
    "VariableCost", "Demand_Techs", "Emissions",
    "Interconnector_Params", "Fixed_Horizon_Parameters",
]

# Orange highlight colors used by the team to flag rows for decision.
ORANGE_COLORS = {"FFFCD5B4", "FFFCD5B5"}

# Code/name column candidates -- A-O uses different headers across files.
CODE_COLS = ["Tech", "Fuel/Tech", "Technology_Code"]
NAME_COLS = ["Tech.Name", "Name", "Technology_Name"]

# A-O workbooks paired with short labels for the propagation matrix.
AO_WORKBOOKS = [
    ("Param",   AO_PARAM),
    ("AR_Base", AO_AR_BASE),
    ("AR_Proj", AO_AR_PROJ),
    ("Demand",  AO_DEMAND),
]

# =============================================================================
# HELPERS
# =============================================================================

def parse_code(code):
    """11-char codes:  'PWRSHPINDNO' -> ('PWR','SHP','INDNO');
                       'RNWBIOBGDXX' -> ('RNW','BIO','BGDXX').
       13-char codes:  'TRNBTNXXINDEA' -> ('TRN','BTNXX','INDEA')   [origin, destination]

       Returns (None, None, None) for codes that don't match either layout.
    """
    s = str(code)
    if len(s) == 11:
        return s[:3], s[3:6], s[6:11]
    if len(s) == 13 and s.startswith("TRN"):
        return s[:3], s[3:8], s[8:13]
    return None, None, None


def find_code_col(df):
    return next((c for c in CODE_COLS if c in df.columns), None)


def find_name_col(df):
    return next((c for c in NAME_COLS if c in df.columns), None)


def collect_codes_from_workbook(path):
    """Return set of every tech code anywhere in a workbook."""
    codes = set()
    xl = pd.ExcelFile(path)
    for s in xl.sheet_names:
        df = pd.read_excel(xl, s)
        col = find_code_col(df)
        if col:
            codes.update(df[col].dropna().astype(str))
    return codes


def collect_codes_with_names(path, sheet_filter=None):
    """{code: tech_name} -- longest tech_name observed across in-scope sheets."""
    out = {}
    xl = pd.ExcelFile(path)
    for s in xl.sheet_names:
        if sheet_filter is not None and s not in sheet_filter:
            continue
        df = pd.read_excel(xl, s)
        cc = find_code_col(df)
        nc = find_name_col(df)
        if not cc:
            continue
        for _, r in df.iterrows():
            code = r[cc]
            if pd.isna(code) or str(code).strip() == "":
                continue
            code = str(code)
            name = str(r[nc]).strip() if nc and pd.notna(r[nc]) else ""
            if code not in out or len(name) > len(out[code]):
                out[code] = name
    return out


def collect_orange_codes(path, sheet_filter):
    """Codes that have at least one orange-fill cell on any in-scope sheet."""
    wb = load_workbook(path, data_only=True)
    orange = set()
    for sheet_name in wb.sheetnames:
        if sheet_name not in sheet_filter:
            continue
        ws = wb[sheet_name]
        headers = [c.value for c in ws[1]]
        idx = None
        for i, h in enumerate(headers):
            if h in CODE_COLS:
                idx = i + 1
                break
        if idx is None:
            continue
        for r in range(2, ws.max_row + 1):
            row_orange = False
            for cell in ws[r]:
                if cell.fill and cell.fill.fgColor and \
                        cell.fill.fgColor.rgb in ORANGE_COLORS:
                    row_orange = True
                    break
            if row_orange:
                v = ws.cell(r, idx).value
                if isinstance(v, str) and v.strip():
                    orange.add(v)
    return orange

# =============================================================================
# STEP 1 -- LOAD WV TECH ROSTER
# =============================================================================

print("=" * 72)
print("STEP 1 -- Load WV Tech roster")
print("=" * 72)

if not os.path.exists(WV_FILE):
    raise FileNotFoundError(
        f"{WV_FILE} not found. Run 1_merge_timeslices_into_WV.py first."
    )

wv_xl = pd.ExcelFile(WV_FILE)

# Prefer Tech_Universe (audit sheet from stage 1) -- canonical roster.
# Fallback: rebuild from raw sheets if Tech_Universe is missing.
if "Tech_Universe" in wv_xl.sheet_names:
    tu = pd.read_excel(WV_FILE, sheet_name="Tech_Universe")
    wv_codes = {}
    sheets_in_wv = {}
    for _, r in tu.iterrows():
        code = str(r["Tech"])
        wv_codes[code] = "" if pd.isna(r.get("Tech.Name")) else str(r["Tech.Name"])
        ss = [s for s in WV_TECH_SHEETS if s in tu.columns and r.get(s) == "Y"]
        sheets_in_wv[code] = ss
    print(f"  Loaded Tech_Universe: {len(wv_codes)} unique codes")
else:
    print("  Tech_Universe sheet not found; rebuilding from raw WV sheets")
    wv_codes = collect_codes_with_names(WV_FILE, sheet_filter=set(WV_TECH_SHEETS))
    sheets_in_wv = {}
    for s in WV_TECH_SHEETS:
        if s not in wv_xl.sheet_names:
            continue
        df = pd.read_excel(WV_FILE, sheet_name=s)
        col = find_code_col(df)
        if not col:
            continue
        for c in df[col].dropna().astype(str).unique():
            sheets_in_wv.setdefault(c, []).append(s)
    print(f"  Built roster from raw sheets: {len(wv_codes)} unique codes")

print("\nDetecting orange highlights in WV...")
orange = collect_orange_codes(WV_FILE, set(WV_TECH_SHEETS))
print(f"  {len(orange)} codes have at least one orange cell")

# =============================================================================
# STEP 2 -- PER-WORKBOOK A-O CODE SETS
# =============================================================================

print("\n" + "=" * 72)
print("STEP 2 -- Scan A-O workbooks (per-workbook presence)")
print("=" * 72)

ao_codes_per_wb = {}
for label, path in AO_WORKBOOKS:
    if not os.path.exists(path):
        print(f"  WARNING: {path} not found; treating as empty")
        ao_codes_per_wb[label] = set()
        continue
    s = collect_codes_from_workbook(path)
    ao_codes_per_wb[label] = s
    print(f"  {label:<8} {len(s):>4} codes")

ao_codes_union = set().union(*ao_codes_per_wb.values())
print(f"  Union: {len(ao_codes_union)} unique codes across all 4 A-O workbooks")

# =============================================================================
# STEP 3 -- ABSENCE = EXTENSION LIST
# =============================================================================

print("\n" + "=" * 72)
print("STEP 3 -- Compute extension list (codes in WV but absent from A-O union)")
print("=" * 72)

absent = sorted(c for c in wv_codes if c not in ao_codes_union)
orange_and_absent  = [c for c in absent if c in orange]
absent_not_orange  = [c for c in absent if c not in orange]
orange_in_ao_union = sorted(c for c in orange if c in ao_codes_union)

print(f"  Absent from A-O union:           {len(absent)}")
print(f"    of which orange-flagged:       {len(orange_and_absent)}")
print(f"    of which NOT orange-flagged:   {len(absent_not_orange)}")
print(f"  Orange but already in A-O:       {len(orange_in_ao_union)} (stale flags?)")

# =============================================================================
# STEP 4 -- TEMPLATE SUGGESTIONS (same prefix + same category, in A-O Param)
# =============================================================================

print("\n" + "=" * 72)
print("STEP 4 -- Suggest A-O templates")
print("         (1) same prefix+category in Param;")
print("         (2) CATEGORY_FALLBACKS analogue if (1) fails;")
print("         (3) TRN partial-region match if both above fail")
print("=" * 72)

# Index Param codes by (prefix, category) for fast same-category lookup.
# This works for both 11-char codes (PWR/RNW/MIN) and 13-char TRN codes:
# parse_code returns (prefix, category, region) where for TRN, "category"
# is the 5-char origin region.
param_codes_by_pc = {}
for c in ao_codes_per_wb.get("Param", set()):
    p, cat, _ = parse_code(c)
    if p and cat:
        param_codes_by_pc.setdefault((p, cat), []).append(c)

# For TRN matching: also index by destination region (the 2nd region in
# the 13-char code). This lets us find "any TRN whose destination matches".
param_trn_by_dest = {}  # destination_region -> list of A-O TRN codes ending in it
for c in ao_codes_per_wb.get("Param", set()):
    p, origin, dest = parse_code(c)
    if p == "TRN":
        param_trn_by_dest.setdefault(dest, []).append(c)

# Param Tech.Names are needed to populate Suggested_Tech.Name_AO.
param_names = collect_codes_with_names(AO_PARAM) if os.path.exists(AO_PARAM) else {}

n_with_primary, n_with_fallback, n_with_trn, n_blank = 0, 0, 0, 0
suggestions = {}
for code in absent:
    p, cat, region = parse_code(code)
    template = ""
    fallback_kind = ""        # "" | "category_fallback" | "trn_partial"
    fallback_detail = ""

    # --- Primary rule: same prefix + same category, prefer same region ---
    candidates = param_codes_by_pc.get((p, cat), []) if p else []
    if candidates:
        same_region = [x for x in candidates if region and x.endswith(region)]
        template = sorted(same_region)[0] if same_region else sorted(candidates)[0]
        n_with_primary += 1

    # --- Fallback 1: configured category fallback (e.g. SHP -> HYD) ---
    elif p and cat in CATEGORY_FALLBACKS:
        fb_cat = CATEGORY_FALLBACKS[cat]
        fb_candidates = param_codes_by_pc.get((p, fb_cat), [])
        if fb_candidates:
            same_region = [x for x in fb_candidates if region and x.endswith(region)]
            template = sorted(same_region)[0] if same_region else sorted(fb_candidates)[0]
            fallback_kind = "category_fallback"
            fallback_detail = f"{cat}->{fb_cat}"
            n_with_fallback += 1

    # --- Fallback 2: TRN partial match (any A-O TRN sharing a region with the
    # new code). For new TRN<origin><dest>, look at all four region-overlap
    # cases: A-O code's origin == new origin/dest, or A-O code's dest ==
    # new origin/dest. This catches cross-border peers regardless of which
    # leg of the route matches.
    if not template and p == "TRN":
        new_origin, new_dest = cat, region   # cat is origin for TRN
        cands = set()
        for c2 in ao_codes_per_wb.get("Param", set()):
            p2, o2, d2 = parse_code(c2)
            if p2 != "TRN":
                continue
            if o2 in (new_origin, new_dest) or d2 in (new_origin, new_dest):
                cands.add(c2)
        # Score: prefer cross-border templates (both legs end in XX or are
        # India sub-regions) over national-transmission codes like TRNNLI*.
        # Country-style 5-char regions in this project end in 'XX' or start
        # with 'IND' (the India sub-regions). Anything else is non-canonical.
        def is_country_region(r):
            return r.endswith("XX") or (r and r.startswith("IND"))
        def score(c2):
            _, o2, d2 = parse_code(c2)
            return (is_country_region(o2) and is_country_region(d2),  # 1: cross-border
                    o2 in (new_origin, new_dest),                      # 2: origin overlap
                    d2 in (new_origin, new_dest),                      # 3: dest overlap
                    -ord(c2[0]) if c2 else 0)                          # 4: alpha tiebreak
        if cands:
            template = sorted(cands, key=score, reverse=True)[0]
            fallback_kind = "trn_partial"
            fallback_detail = f"region overlap with {template}"
            n_with_trn += 1

    if not template:
        n_blank += 1

    # Tech.Name suggestion: copy verbatim only when the PRIMARY rule matched.
    # When a fallback was used, leave name blank to force a deliberate override.
    if template and not fallback_kind:
        sug_name = param_names.get(template, "")
    else:
        sug_name = ""

    add_flags = {}
    for label, _ in AO_WORKBOOKS:
        if not template:
            add_flags[label] = ""
        else:
            add_flags[label] = "Y" if template in ao_codes_per_wb.get(label, set()) else "N"

    suggestions[code] = {
        "template":        template,
        "name":            sug_name,
        "add":             add_flags,
        "fallback_kind":   fallback_kind,
        "fallback_detail": fallback_detail,
    }

print(f"  {n_with_primary}  codes matched by primary rule (same prefix+category)")
print(f"  {n_with_fallback}  codes matched by CATEGORY_FALLBACKS")
print(f"  {n_with_trn}  codes matched by TRN partial-region rule")
print(f"  {n_blank}  codes left blank for manual review (NEEDS_REVIEW)")

# =============================================================================
# STEP 5 -- BUILD TAB 1 (Extensions_To_Add)
# =============================================================================

print("\n" + "=" * 72)
print("STEP 5 -- Build Tab 1 (Extensions_To_Add)")
print("=" * 72)

rows_t1 = []
for code in absent:
    p, cat, region = parse_code(code)
    s = suggestions[code]
    include_default = "Y" if cat in V17_NEW_CATEGORIES else ""

    # Notes: explain how the template was found (or why it wasn't).
    if not s["template"]:
        notes = "[NEEDS_REVIEW: no same-category template in A-O Parametrization]"
    elif s["fallback_kind"] == "category_fallback":
        notes = (f"[FALLBACK: category {s['fallback_detail']} via CATEGORY_FALLBACKS; "
                 f"template={s['template']}; confirm Tech.Name override]")
    elif s["fallback_kind"] == "trn_partial":
        notes = (f"[FALLBACK: TRN partial-region match — {s['fallback_detail']}; "
                 f"confirm Tech.Name override]")
    else:
        notes = ""
    rows_t1.append({
        "AO_Code_To_Add":          code,
        "Tech_Name_WV":            wv_codes.get(code, ""),
        "Sheets_in_WV":            ", ".join(sheets_in_wv.get(code, [])),
        "Orange_Flagged":          "Y" if code in orange else "N",
        "In_Param":                "Y" if code in ao_codes_per_wb.get("Param",   set()) else "N",
        "In_AR_Base":              "Y" if code in ao_codes_per_wb.get("AR_Base", set()) else "N",
        "In_AR_Proj":              "Y" if code in ao_codes_per_wb.get("AR_Proj", set()) else "N",
        "In_Demand":               "Y" if code in ao_codes_per_wb.get("Demand",  set()) else "N",
        "Suggested_Template_AO":   s["template"],
        "Suggested_Tech.Name_AO":  s["name"],
        "Add_To_Param":            s["add"]["Param"],
        "Add_To_AR_Base":          s["add"]["AR_Base"],
        "Add_To_AR_Proj":          s["add"]["AR_Proj"],
        "Add_To_Demand":           s["add"]["Demand"],
        "Include":                 include_default,
        "Override_Template_AO":    "",
        "Override_Tech.Name_AO":   "",
        "Notes":                   notes,
    })

cols_t1 = [
    "AO_Code_To_Add", "Tech_Name_WV", "Sheets_in_WV", "Orange_Flagged",
    "In_Param", "In_AR_Base", "In_AR_Proj", "In_Demand",
    "Suggested_Template_AO", "Suggested_Tech.Name_AO",
    "Add_To_Param", "Add_To_AR_Base", "Add_To_AR_Proj", "Add_To_Demand",
    "Include", "Override_Template_AO", "Override_Tech.Name_AO", "Notes",
]
df_t1 = pd.DataFrame(rows_t1, columns=cols_t1)
print(f"  {len(df_t1)} rows / {len(cols_t1)} columns")

# =============================================================================
# STEP 6 -- BUILD TAB 2 (Parameter_Rows_To_Replicate)
# =============================================================================

print("\n" + "=" * 72)
print("STEP 6 -- Build Tab 2 (Parameter_Rows_To_Replicate)")
print("=" * 72)

absent_set = set(absent)
parts = []
for s in WV_TECH_SHEETS:
    if s not in wv_xl.sheet_names:
        continue
    df = pd.read_excel(WV_FILE, sheet_name=s)
    cc = find_code_col(df)
    if not cc:
        continue
    sub = df[df[cc].astype(str).isin(absent_set)].copy()
    if len(sub) == 0:
        continue
    sub.insert(0, "Source_Sheet", s)
    parts.append(sub)

if parts:
    df_t2 = pd.concat(parts, ignore_index=True, sort=False)
else:
    df_t2 = pd.DataFrame({"Note": ["No parameter rows found"]})
print(f"  {len(df_t2)} parameter rows captured")

# =============================================================================
# STEP 7 -- BUILD TAB 3 (Signal_Disagreements)
# =============================================================================

print("\n" + "=" * 72)
print("STEP 7 -- Build Tab 3 (Signal_Disagreements)")
print("=" * 72)

dis = []
for c in orange_in_ao_union:
    dis.append({
        "Code":   c,
        "Issue":  "Orange-flagged in WV but already in A-O union",
        "Action": "Probably stale flag - clear orange highlight in WV",
    })
for c in absent_not_orange:
    dis.append({
        "Code":   c,
        "Issue":  "Absent from A-O union but NOT orange-flagged in WV",
        "Action": "Verify - team may have missed flagging during cleanup",
    })
df_t3 = pd.DataFrame(dis) if dis else pd.DataFrame(columns=["Code", "Issue", "Action"])
print(f"  {len(df_t3)} disagreement rows")

# =============================================================================
# STEP 8 -- WRITE OUTPUT
# =============================================================================

print("\n" + "=" * 72)
print("STEP 8 -- Write OSTRAM_AO_Extensions.xlsx")
print("=" * 72)

with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as w:
    df_t1.to_excel(w, sheet_name="1_Extensions_To_Add",           index=False)
    df_t2.to_excel(w, sheet_name="2_Parameter_Rows_To_Replicate", index=False)
    df_t3.to_excel(w, sheet_name="3_Signal_Disagreements",        index=False)

print(f"  Saved: {OUTPUT_FILE}")

# =============================================================================
# BUILT-IN TESTS
# =============================================================================

print("\n" + "=" * 72)
print("TESTS")
print("=" * 72)

_passed, _failed = 0, 0
def check(name, cond, detail=""):
    global _passed, _failed
    if cond:
        _passed += 1
        print(f"  PASS  {name}")
    else:
        _failed += 1
        print(f"  FAIL  {name}   {detail}")

# (a) Source files untouched (existence check)
check("WV file still exists", os.path.exists(WV_FILE))
for label, path in AO_WORKBOOKS:
    check(f"A-O {label} still exists", os.path.exists(path))

# (b) Output produced and readable
check("Output file written", os.path.exists(OUTPUT_FILE))
out = pd.ExcelFile(OUTPUT_FILE)
check("Tab 1 present", "1_Extensions_To_Add"           in out.sheet_names)
check("Tab 2 present", "2_Parameter_Rows_To_Replicate" in out.sheet_names)
check("Tab 3 present", "3_Signal_Disagreements"        in out.sheet_names)

# (c) Row counts match in-memory dataframes
out_t1 = pd.read_excel(OUTPUT_FILE, sheet_name="1_Extensions_To_Add")
out_t2 = pd.read_excel(OUTPUT_FILE, sheet_name="2_Parameter_Rows_To_Replicate")
out_t3 = pd.read_excel(OUTPUT_FILE, sheet_name="3_Signal_Disagreements")
check(f"Tab 1 rows = {len(df_t1)}", len(out_t1) == len(df_t1))
check(f"Tab 2 rows = {len(df_t2)}", len(out_t2) == len(df_t2))
check(f"Tab 3 rows = {len(df_t3)}", len(out_t3) == len(df_t3))

# (d) No Tab 1 code is already in A-O union (would mean false-positive extension)
t1_codes = set(out_t1["AO_Code_To_Add"].dropna().astype(str))
spurious = t1_codes & ao_codes_union
check("No Tab 1 code already in A-O union", len(spurious) == 0,
      f"(found {len(spurious)}: {sorted(spurious)[:5]}...)")

# (e) Every Tab 1 code is present in WV
gap = t1_codes - set(wv_codes)
check("All Tab 1 codes present in WV", len(gap) == 0,
      f"(stragglers: {sorted(gap)[:5]}...)")

# (f) Codes flagged Include=Y must have category in V17_NEW_CATEGORIES
include_y = out_t1[out_t1["Include"] == "Y"]["AO_Code_To_Add"].astype(str)
bad_include = []
for c in include_y:
    _, cat, _ = parse_code(c)
    if cat not in V17_NEW_CATEGORIES:
        bad_include.append(c)
check("Include=Y codes match V17_NEW_CATEGORIES", len(bad_include) == 0,
      f"(stragglers: {bad_include[:5]}...)")

# (g) Per-workbook In_* flags are consistent with ao_codes_per_wb
inconsistent = []
for _, r in out_t1.iterrows():
    code = str(r["AO_Code_To_Add"])
    for col, lbl in [("In_Param","Param"), ("In_AR_Base","AR_Base"),
                     ("In_AR_Proj","AR_Proj"), ("In_Demand","Demand")]:
        expected = "Y" if code in ao_codes_per_wb.get(lbl, set()) else "N"
        if r[col] != expected:
            inconsistent.append((code, col, expected, r[col]))
check("Per-workbook In_* flags consistent", len(inconsistent) == 0,
      f"(first issue: {inconsistent[:1]})")

# (h) PWRSHP* sanity: every PWRSHP in WV must appear in Tab 1 (since v17-driven)
shp_in_wv = sorted(c for c in wv_codes if str(c).startswith("PWRSHP"))
shp_in_t1 = sorted(c for c in t1_codes  if c.startswith("PWRSHP"))
if shp_in_wv:
    check(f"All {len(shp_in_wv)} PWRSHP codes appear in Tab 1",
          len(shp_in_t1) == len(shp_in_wv),
          f"(WV={shp_in_wv}; T1={shp_in_t1})")

# (i) Add_To_* is blank wherever Suggested_Template_AO is blank
viol = []
for _, r in out_t1.iterrows():
    if str(r["Suggested_Template_AO"]).strip() == "":
        for col in ("Add_To_Param","Add_To_AR_Base","Add_To_AR_Proj","Add_To_Demand"):
            if str(r[col]).strip() != "":
                viol.append((str(r["AO_Code_To_Add"]), col, r[col]))
check("Add_To_* blank when no template suggested", len(viol) == 0,
      f"(stragglers: {viol[:3]})")

# (j) When a FALLBACK was used, Suggested_Tech.Name_AO must be blank.
#     Re-read with keep_default_na=False so empty cells stay as "" (not NaN).
out_t1_raw = pd.read_excel(OUTPUT_FILE, sheet_name="1_Extensions_To_Add", keep_default_na=False)
fb_rows = out_t1_raw[out_t1_raw["Notes"].astype(str).str.startswith("[FALLBACK")]
fb_with_name = fb_rows[fb_rows["Suggested_Tech.Name_AO"].astype(str).str.strip() != ""]
check("Fallback rows leave Suggested_Tech.Name_AO blank", len(fb_with_name) == 0,
      f"(stragglers: {fb_with_name['AO_Code_To_Add'].tolist()[:3]})")

print(f"\n  {_passed} passed, {_failed} failed")
if _failed == 0:
    print("  ALL TESTS PASSED")
else:
    print(f"  {_failed} TEST(S) FAILED - review output above")
