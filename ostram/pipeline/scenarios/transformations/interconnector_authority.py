"""Fail-closed loaders for migrated interconnector minimum authorities.

The production numeric values live only in ``Interconnector_Params`` in the
v18 workbook.  This module contains schema, domain, metadata, and semantic
digest guards, but deliberately contains no numeric schedule fallback.

Both supported workbook shapes are accepted:

* the raw v18 sheet, whose first column is ``scenario``; and
* a scenario-materialized sheet, from which ``scenario`` has been removed.
"""
from __future__ import annotations

import hashlib
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Dict, Mapping

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


AUTHORITY_SHEET_NAME = "Interconnector_Params"
AUTHORITY_YEARS = tuple(range(2023, 2051))
BAU_SCENARIO = "BAU"

MINIMUM_CONTRIBUTION_PARAMETER = "TotalAnnualMinCapacityInvestment"
MINIMUM_CONTRIBUTION_TECHS = frozenset({
    "TRNBTNXXBGDXX",
    "TRNBTNXXINDEA",
    "TRNBTNXXINDNE",
    "TRNINDEAINDWE",
    "TRNINDEANPLXX",
    "TRNINDNOINDWE",
    "TRNINDNONPLXX",
    "TRNINDSOINDWE",
    "TRNINDSOLKAXX",
    "TRNLKAXXMDVXX",
    "TRNNPLXXBGDXX",
})
MINIMUM_CONTRIBUTION_SEMANTIC_SHA256 = (
    "d5e7860480abdda57e207b1639032e9744d5d71096efad5028d372e174ffc958"
)
MINIMUM_CONTRIBUTION_SOURCE = (
    "Effective predecessor Stage 3 additive FUTURE contribution"
)

MINIMUM_BOUNDARY_PARAMETER = "MinimumInvestmentClampBoundary"
MINIMUM_BOUNDARY_TECHS = frozenset({
    "TRNBGDXXINDEA",
    "TRNNPLXXBGDXX",
})
MINIMUM_BOUNDARY_SEMANTIC_SHA256 = (
    "d155a0c4f68b7ef2875b7a14bf2e7d5341e1ff09af16ca4ee7f868d85dcc6ca7"
)
MINIMUM_BOUNDARY_SOURCE = "Minimum-only LinkFreeze compatibility boundary"

_COMMON_HEADERS = (
    "Tech.ID",
    "Tech",
    "Tech.Name",
    "Parameter.ID",
    "Parameter",
    "Unit",
    "Projection.Mode",
    "Projection.Parameter",
    *AUTHORITY_YEARS,
    "Source",
)
_RAW_HEADERS = ("scenario", *_COMMON_HEADERS)
_MATERIALIZED_HEADERS = _COMMON_HEADERS

# These guards identify the intended rows independently of their numeric
# schedules.  The values were mechanically copied from the effective target
# TotalAnnualMinCapacityInvestment metadata.
_CONTRIBUTION_TECH_METADATA = {
    "TRNBTNXXBGDXX": (
        10,
        "Transmission interconnection from Bhutan, region XX to Bangladesh, region XX",
    ),
    "TRNBTNXXINDEA": (
        10,
        "Transmission interconnection from Bhutan, region XX to India, region EA",
    ),
    "TRNBTNXXINDNE": (
        10,
        "Transmission interconnection from Bhutan, region XX to India, region NE",
    ),
    "TRNINDEAINDWE": (
        195,
        "Transmission interconnection from India, region EA to India, region WE",
    ),
    "TRNINDEANPLXX": (
        101,
        "Transmission interconnection from India, region EA to Nepal, region XX",
    ),
    "TRNINDNOINDWE": (
        93,
        "Transmission interconnection from India, region NO to India, region WE",
    ),
    "TRNINDNONPLXX": (
        123,
        "Transmission interconnection from India, region NO to Nepal, region XX",
    ),
    "TRNINDSOINDWE": (
        141,
        "Transmission interconnection from India, region SO to India, region WE",
    ),
    "TRNINDSOLKAXX": (
        22,
        "Transmission interconnection from India, region SO to Sri Lanka, region XX",
    ),
    "TRNLKAXXMDVXX": (
        110,
        "Transmission interconnection from Sri Lanka, region XX to Maldives, region XX",
    ),
    "TRNNPLXXBGDXX": (
        205,
        "Transmission interconnection from Nepal, region XX to Bangladesh, region XX",
    ),
}
_BOUNDARY_TECH_METADATA = {
    "TRNBGDXXINDEA": (
        205,
        "Transmission interconnection from Bangladesh, region XX to India, region EA",
    ),
    "TRNNPLXXBGDXX": (
        205,
        "Transmission interconnection from Nepal, region XX to Bangladesh, region XX",
    ),
}


def _numeric_decimal(value: object, context: str) -> Decimal:
    """Return an exact finite, non-negative Decimal for a numeric cell."""
    if isinstance(value, bool) or not isinstance(value, (int, float, Decimal)):
        if isinstance(value, str) and value.startswith("="):
            detail = "formula"
        elif value is None:
            detail = "blank"
        else:
            detail = f"non-numeric value {value!r}"
        raise ValueError(f"{context}: expected a numeric value, got {detail}")
    try:
        number = Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(
            f"{context}: expected a numeric value, got {value!r}"
        ) from exc
    if not number.is_finite() or number < 0:
        raise ValueError(
            f"{context}: expected a finite non-negative value, got {value!r}"
        )
    return number


def _decimal_text(value: object, context: str) -> str:
    """Return the canonical text used by the authority semantic digests."""
    number = _numeric_decimal(value, context)
    if number == 0:
        return "0"
    return format(number.normalize(), "f")


def authority_semantic_sha256(
    authority: Mapping[str, Mapping[int, object]],
) -> str:
    """Hash a canonical technology/year/value authority domain."""
    payload = "".join(
        f"{tech}|{year}|"
        f"{_decimal_text(authority[tech][year], f'{tech}/{year}')}\n"
        for tech in sorted(authority)
        for year in sorted(authority[tech])
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _header_columns(ws: Worksheet) -> tuple[Dict[object, int], bool]:
    headers = tuple(cell.value for cell in ws[1])
    if headers == _RAW_HEADERS:
        raw_shape = True
    elif headers == _MATERIALIZED_HEADERS:
        raw_shape = False
    else:
        expected_raw = list(_RAW_HEADERS)
        expected_materialized = list(_MATERIALIZED_HEADERS)
        raise ValueError(
            f"{AUTHORITY_SHEET_NAME} has an unexpected header shape; "
            f"expected raw {expected_raw!r} or materialized "
            f"{expected_materialized!r}, got {list(headers)!r}"
        )
    return {header: index for index, header in enumerate(headers)}, raw_shape


def _validate_mapping(
    authority: Mapping[str, Mapping[int, object]],
    *,
    label: str,
    techs: frozenset[str],
    digest: str,
) -> None:
    actual_techs = set(authority)
    if actual_techs != set(techs):
        raise ValueError(
            f"{label} technology domain mismatch: "
            f"missing={sorted(techs - actual_techs)}, "
            f"extra={sorted(actual_techs - techs)}"
        )
    expected_years = set(AUTHORITY_YEARS)
    for tech in sorted(techs):
        actual_years = set(authority[tech])
        if actual_years != expected_years:
            raise ValueError(
                f"{label} year domain mismatch for {tech}: "
                f"missing={sorted(expected_years - actual_years)}, "
                f"extra={sorted(actual_years - expected_years)}"
            )
        for year in AUTHORITY_YEARS:
            _numeric_decimal(authority[tech][year], f"{tech}/{year}")
    actual_digest = authority_semantic_sha256(authority)
    if actual_digest != digest:
        raise ValueError(
            f"{label} semantic digest mismatch: "
            f"expected {digest}, got {actual_digest}"
        )


def validate_minimum_contribution_authority(
    authority: Mapping[str, Mapping[int, object]],
) -> None:
    """Validate an already-loaded minimum/FUTURE contribution mapping."""
    _validate_mapping(
        authority,
        label="minimum contribution authority",
        techs=MINIMUM_CONTRIBUTION_TECHS,
        digest=MINIMUM_CONTRIBUTION_SEMANTIC_SHA256,
    )


def validate_minimum_boundary_authority(
    authority: Mapping[str, Mapping[int, object]],
) -> None:
    """Validate an already-loaded minimum-clamp boundary mapping."""
    _validate_mapping(
        authority,
        label="minimum boundary authority",
        techs=MINIMUM_BOUNDARY_TECHS,
        digest=MINIMUM_BOUNDARY_SEMANTIC_SHA256,
    )


def _load_family(
    authority_path: Path,
    *,
    parameter: str,
    techs: frozenset[str],
    digest: str,
    source: str,
    tech_metadata: Mapping[str, tuple[int, str]],
    label: str,
) -> Dict[str, Dict[int, float]]:
    path = Path(authority_path)
    if not path.is_file():
        raise FileNotFoundError(f"{label} workbook not found: {path}")

    workbook = load_workbook(
        path, data_only=False, read_only=True, keep_links=False
    )
    try:
        if AUTHORITY_SHEET_NAME not in workbook.sheetnames:
            raise ValueError(
                f"authority workbook missing sheet {AUTHORITY_SHEET_NAME!r}"
            )
        worksheet = workbook[AUTHORITY_SHEET_NAME]
        columns, raw_shape = _header_columns(worksheet)

        authority: Dict[str, Dict[int, float]] = {}
        for row_index, values in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            if values[columns["Parameter"]] != parameter:
                continue

            tech = values[columns["Tech"]]
            if not isinstance(tech, str):
                raise ValueError(
                    f"{label} row {row_index}: Tech must be non-blank text"
                )
            if tech in authority:
                raise ValueError(
                    f"{label} duplicate row for {tech} at row {row_index}"
                )
            if raw_shape and values[columns["scenario"]] != BAU_SCENARIO:
                raise ValueError(
                    f"{label} {tech}: scenario must be {BAU_SCENARIO!r}"
                )
            if tech not in tech_metadata:
                raise ValueError(f"{label} unexpected technology {tech!r}")

            expected_id, expected_name = tech_metadata[tech]
            metadata_checks = {
                "Tech.ID": expected_id,
                "Tech.Name": expected_name,
                "Parameter.ID": 7,
                "Unit": None,
                "Projection.Mode": "User defined",
                "Source": source,
            }
            for field, expected in metadata_checks.items():
                actual = values[columns[field]]
                if actual != expected:
                    raise ValueError(
                        f"{label} {tech}: {field} must be {expected!r}, "
                        f"got {actual!r}"
                    )

            projection_parameter = _numeric_decimal(
                values[columns["Projection.Parameter"]],
                f"{label} {tech}/Projection.Parameter",
            )
            if projection_parameter != 0:
                raise ValueError(
                    f"{label} {tech}: Projection.Parameter must be 0"
                )

            profile: Dict[int, float] = {}
            for year in AUTHORITY_YEARS:
                number = _numeric_decimal(
                    values[columns[year]], f"{label} {tech}/{year}"
                )
                profile[year] = float(number)
            authority[tech] = profile
    finally:
        workbook.close()

    _validate_mapping(
        authority, label=label, techs=techs, digest=digest
    )
    return authority


def load_minimum_contribution_authority(
    authority_path: Path,
) -> Dict[str, Dict[int, float]]:
    """Load the exact 11-row additive minimum/FUTURE authority."""
    return _load_family(
        authority_path,
        parameter=MINIMUM_CONTRIBUTION_PARAMETER,
        techs=MINIMUM_CONTRIBUTION_TECHS,
        digest=MINIMUM_CONTRIBUTION_SEMANTIC_SHA256,
        source=MINIMUM_CONTRIBUTION_SOURCE,
        tech_metadata=_CONTRIBUTION_TECH_METADATA,
        label="minimum contribution authority",
    )


def load_minimum_boundary_authority(
    authority_path: Path,
) -> Dict[str, Dict[int, float]]:
    """Load the exact two-row LinkFreeze minimum-clamp boundary authority."""
    return _load_family(
        authority_path,
        parameter=MINIMUM_BOUNDARY_PARAMETER,
        techs=MINIMUM_BOUNDARY_TECHS,
        digest=MINIMUM_BOUNDARY_SEMANTIC_SHA256,
        source=MINIMUM_BOUNDARY_SOURCE,
        tech_metadata=_BOUNDARY_TECH_METADATA,
        label="minimum boundary authority",
    )
