"""
patch_ao_c2a.py
================
Patch the A-O parametrization workbook to set CapacityToActivityUnit = 31.536
for every GENERATION and STORAGE technology that doesn't already have it.

Why:
    OSeMOSYS defaults CapacityToActivityUnit to 1, which causes a 31.5x
    over-build of any tech parametrized with capacity in GW (the convention
    used by every PWR* and TRN* tech in this model). The four families
    PWRCCS, PWRCOG, PWRPET, plus all transmission, were already fixed.
    The remaining 134 GENERATION+STORAGE techs default to 1 and need
    the explicit value 31.536. MIN*/RNW* primary-supply techs have no
    capacity-side parameters in this model, so c2a is moot for them and
    they are intentionally left alone (they aren't in this sheet anyway).

Behaviour:
    - Reads taxonomy from TECH_TYPES.csv to identify GEN+STO techs.
    - Reads the 'Fixed Horizon Parameters' sheet of the A-O workbook.
    - Appends a new CapacityToActivityUnit row for each target tech that
      doesn't already have one. Existing values are NEVER overwritten.
    - Tech.Type, Tech.ID, Tech, Tech.Name are copied from the tech's
      existing rows (typically its OperationalLife row).
    - Parameter.ID is set to (max existing Parameter.ID for that tech) + 1.
    - Cell formatting is copied from a reference existing c2a row so the
      new rows blend in.
    - Writes a sibling output file. The source file is never modified.
    - Idempotent: re-running on an already-patched file is a no-op.

Usage:
    python patch_ao_c2a.py
    python patch_ao_c2a.py --src A-O_Parametrization.xlsx \\
                           --tax TECH_TYPES.csv \\
                           --out A-O_Parametrization_c2a_patched.xlsx
    python patch_ao_c2a.py --dry-run   # show plan, write nothing
"""

from __future__ import annotations

import argparse
import csv
import sys
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell import Cell

# ---------- configuration ----------
DEFAULT_SRC = "A-O_Parametrization.xlsx"
DEFAULT_TAX = "TECH_TYPES.csv"
DEFAULT_OUT = "A-O_Parametrization_c2a_patched.xlsx"

SHEET_NAME = "Fixed Horizon Parameters"
TARGET_TYPES = {"GENERATION", "STORAGE_SHORT", "STORAGE_LONG"}
C2A_PARAM_NAME = "CapacityToActivityUnit"
C2A_VALUE = 31.536

# Column layout (1-indexed) — verified against this workbook
COL_TYPE = 1     # 'Tech.Type'      (e.g. 'Secondary', 'Demand')
COL_TECHID = 2   # 'Tech.ID'        (numeric, unique per tech)
COL_TECH = 3     # 'Tech'           (e.g. 'PWRCOABGDXX')
COL_TECHNAME = 4 # 'Tech.Name'      (description)
COL_PARAMID = 5  # 'Parameter.ID'   (numeric, sequential within tech)
COL_PARAM = 6    # 'Parameter'      (e.g. 'CapacityToActivityUnit')
COL_UNIT = 7     # 'Unit'
COL_VALUE = 8    # 'Value'


# ---------- helpers ----------
def load_taxonomy(path: Path) -> dict[str, str]:
    """Return {tech_name: tech_type} from TECH_TYPES.csv."""
    out = {}
    with path.open(encoding="utf-8-sig", newline="") as f:
        rdr = csv.reader(f)
        next(rdr, None)  # header
        for row in rdr:
            if len(row) < 2:
                continue
            ttype, tname = row[0].strip(), row[1].strip()
            if ttype and tname:
                out[tname] = ttype
    return out


def index_sheet(ws) -> tuple[dict, Cell | None]:
    """
    Walk the sheet once. Return:
        meta: {tech: {'type', 'id', 'name', 'max_paramid', 'has_c2a',
                      'style_row'}}
        ref_c2a_row: row number of any existing c2a row (for style reference)
    """
    meta: dict[str, dict] = {}
    ref_c2a_row: int | None = None

    for r in range(2, ws.max_row + 1):
        tech = ws.cell(r, COL_TECH).value
        if tech is None:
            continue
        param = ws.cell(r, COL_PARAM).value
        pid = ws.cell(r, COL_PARAMID).value

        m = meta.setdefault(
            tech,
            {
                "type": ws.cell(r, COL_TYPE).value,
                "id": ws.cell(r, COL_TECHID).value,
                "name": ws.cell(r, COL_TECHNAME).value,
                "max_paramid": 0,
                "has_c2a": False,
                "ref_row": r,  # fallback row to copy style from
            },
        )
        try:
            pid_int = int(pid)
            if pid_int > m["max_paramid"]:
                m["max_paramid"] = pid_int
        except (TypeError, ValueError):
            pass

        if param == C2A_PARAM_NAME:
            m["has_c2a"] = True
            if ref_c2a_row is None:
                ref_c2a_row = r

    return meta, ref_c2a_row


def copy_row_style(ws, src_row: int, dst_row: int, ncols: int) -> None:
    """Copy font/fill/border/alignment/number_format from src row to dst row."""
    for c in range(1, ncols + 1):
        src = ws.cell(src_row, c)
        dst = ws.cell(dst_row, c)
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)


# ---------- main ----------
def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--src", default=DEFAULT_SRC,
                    help=f"Source workbook (default: {DEFAULT_SRC})")
    ap.add_argument("--tax", default=DEFAULT_TAX,
                    help=f"TECH_TYPES.csv path (default: {DEFAULT_TAX})")
    ap.add_argument("--out", default=DEFAULT_OUT,
                    help=f"Output workbook (default: {DEFAULT_OUT})")
    ap.add_argument("--dry-run", action="store_true",
                    help="Print plan only, don't write the output file")
    args = ap.parse_args()

    src = Path(args.src)
    tax = Path(args.tax)
    out = Path(args.out)

    if not src.exists():
        print(f"ERROR: source workbook not found: {src}", file=sys.stderr)
        return 1
    if not tax.exists():
        print(f"ERROR: taxonomy file not found: {tax}", file=sys.stderr)
        return 1
    if out.resolve() == src.resolve():
        print("ERROR: --out must differ from --src (non-destructive policy)",
              file=sys.stderr)
        return 1

    # 1. Load taxonomy and select target techs
    tech_type = load_taxonomy(tax)
    target_techs = {t for t, ttype in tech_type.items() if ttype in TARGET_TYPES}
    print(f"Loaded taxonomy: {len(tech_type)} techs total, "
          f"{len(target_techs)} in target families {sorted(TARGET_TYPES)}")

    # 2. Open workbook (load with formulas/styles preserved)
    wb = load_workbook(src)
    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: sheet {SHEET_NAME!r} not in workbook. "
              f"Sheets: {wb.sheetnames}", file=sys.stderr)
        return 1
    ws = wb[SHEET_NAME]

    # 3. Index existing rows
    meta, ref_c2a_row = index_sheet(ws)
    if ref_c2a_row is None:
        print("WARNING: no existing CapacityToActivityUnit row found to "
              "copy style from. New rows will use default formatting.")

    # 4. Plan
    in_sheet = [t for t in target_techs if t in meta]
    not_in_sheet = sorted(t for t in target_techs if t not in meta)
    already_ok = sorted(t for t in in_sheet if meta[t]["has_c2a"])
    to_patch = sorted(t for t in in_sheet if not meta[t]["has_c2a"])

    print(f"\nPlan:")
    print(f"  Target techs in sheet:        {len(in_sheet)}")
    print(f"  Already have c2a (skipped):   {len(already_ok)}")
    print(f"  Will add c2a row:             {len(to_patch)}")
    print(f"  Target techs NOT in sheet:    {len(not_in_sheet)}")
    if not_in_sheet:
        # Should be 0 for this model — flag if not.
        print(f"    (these will not be patched — verify they don't need it):")
        for t in not_in_sheet[:20]:
            print(f"      {t}  [{tech_type[t]}]")
        if len(not_in_sheet) > 20:
            print(f"      ... and {len(not_in_sheet) - 20} more")

    # Breakdown by family
    from collections import Counter
    fam = Counter(tech_type[t] for t in to_patch)
    print(f"  Breakdown by tech type:")
    for k, v in sorted(fam.items()):
        print(f"    {k}: {v}")

    if not to_patch:
        print("\nNothing to patch. (Re-run is a no-op — the file already has "
              "c2a set for every target tech.)")
        return 0

    if args.dry_run:
        print("\n--dry-run set; not writing output.")
        return 0

    # 5. Append rows. Sort by Tech.ID for tidiness.
    to_patch_sorted = sorted(to_patch, key=lambda t: (meta[t]["id"] or 0, t))
    next_row = ws.max_row + 1
    first_new_row = next_row
    for tech in to_patch_sorted:
        m = meta[tech]
        new_pid = (m["max_paramid"] or 0) + 1
        ws.cell(next_row, COL_TYPE, m["type"])
        ws.cell(next_row, COL_TECHID, m["id"])
        ws.cell(next_row, COL_TECH, tech)
        ws.cell(next_row, COL_TECHNAME, m["name"])
        ws.cell(next_row, COL_PARAMID, new_pid)
        ws.cell(next_row, COL_PARAM, C2A_PARAM_NAME)
        ws.cell(next_row, COL_UNIT, None)
        ws.cell(next_row, COL_VALUE, C2A_VALUE)

        # Match formatting of existing c2a rows
        style_src = ref_c2a_row if ref_c2a_row else m["ref_row"]
        copy_row_style(ws, style_src, next_row, ws.max_column)

        next_row += 1

    # 6. Save
    wb.save(out)
    print(f"\nWrote: {out}")
    print(f"  Added {len(to_patch_sorted)} rows (rows {first_new_row}–{next_row - 1})")
    print(f"  Source file untouched: {src}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
