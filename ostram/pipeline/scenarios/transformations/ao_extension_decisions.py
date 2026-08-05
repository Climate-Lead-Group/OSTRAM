"""Apply maintained AO extension decisions to a freshly generated review file.

``2_extract_ao_extensions.py`` owns the mechanically derived columns and rows.
This helper overlays only the five human-decision fields maintained in
``OSTRAM_Scenario_Inputs.xlsx::AO_Extension_Decisions`` and excludes generated
copies of ``Interconnector_Params`` rows that are applied by later governed
authority stages. Both the generated workbook and the scenario-specific
authority passed to this helper are runtime working copies; the maintained
authority is never saved here.
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook
from ostram.paths import resolve_paths


DECISIONS_SHEET = "AO_Extension_Decisions"
EXTENSIONS_SHEET = "1_Extensions_To_Add"
PARAMETER_ROWS_SHEET = "2_Parameter_Rows_To_Replicate"
DECISION_HEADERS = (
    "AO_Code_To_Add",
    "Include",
    "Override_Template_AO",
    "Override_Tech.Name_AO",
    "Notes",
)
CODE_HEADERS = ("Tech", "Fuel/Tech", "Technology_Code")
PROPAGATION_TARGETS = {
    "Add_To_Param": "A-O_Parametrization.xlsx",
    "Add_To_AR_Base": "A-O_AR_Model_Base_Year.xlsx",
    "Add_To_AR_Proj": "A-O_AR_Projections.xlsx",
    "Add_To_Demand": "A-O_Demand.xlsx",
}
LATE_AUTHORITY_SOURCE_SHEET = "Interconnector_Params"
LATE_AUTHORITY_PARAMETERS = {
    "MinimumInvestmentClampBoundary",
    "TotalAnnualMinCapacityInvestment",
}


def _headers(worksheet) -> dict[str, int]:
    return {
        str(cell.value): index
        for index, cell in enumerate(worksheet[1], start=1)
        if cell.value is not None
    }


def _csv_decision_rows(authority_path: Path) -> list[dict[str, object]]:
    with authority_path.open("r", encoding="utf-8-sig", newline="") as stream:
        reader = csv.DictReader(stream)
        if reader.fieldnames != list(DECISION_HEADERS):
            raise ValueError(
                f"AO decision CSV headers must be exactly {list(DECISION_HEADERS)}, "
                f"got {reader.fieldnames}"
            )
        decisions: list[dict[str, object]] = []
        seen: set[str] = set()
        for row_index, row in enumerate(reader, start=2):
            code = str(row["AO_Code_To_Add"] or "").strip()
            if not code:
                if any(str(value or "").strip() for value in row.values()):
                    raise ValueError(f"AO decision CSV row {row_index} has no code")
                continue
            if code in seen:
                raise ValueError(f"duplicate AO extension decision for {code!r}")
            include = str(row["Include"] or "").strip().upper()
            if include not in {"Y", "N"}:
                raise ValueError(
                    f"AO decision CSV row {row_index} Include must be Y or N"
                )
            seen.add(code)
            decisions.append(
                {
                    name: (code if name == "AO_Code_To_Add" else row[name])
                    for name in DECISION_HEADERS
                }
            )
        return decisions


def _decision_rows(authority_path: Path) -> list[dict[str, object]]:
    if authority_path.suffix.lower() == ".csv":
        return _csv_decision_rows(authority_path)
    if authority_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError(
            f"AO decision authority must be CSV or XLSX: {authority_path}"
        )
    workbook = load_workbook(authority_path, read_only=True, data_only=True)
    try:
        if DECISIONS_SHEET not in workbook.sheetnames:
            raise ValueError(
                f"{authority_path.name} has no {DECISIONS_SHEET!r} sheet"
            )
        worksheet = workbook[DECISIONS_SHEET]
        columns = _headers(worksheet)
        missing = [name for name in DECISION_HEADERS if name not in columns]
        if missing:
            raise ValueError(
                f"{DECISIONS_SHEET} missing required columns: {missing}"
            )

        decisions: list[dict[str, object]] = []
        seen: set[str] = set()
        for values in worksheet.iter_rows(min_row=2, values_only=True):
            code_value = values[columns["AO_Code_To_Add"] - 1]
            if code_value is None or str(code_value).strip() == "":
                continue
            code = str(code_value).strip()
            if code in seen:
                raise ValueError(
                    f"duplicate AO extension decision for {code!r}"
                )
            seen.add(code)
            decisions.append(
                {
                    name: (
                        code
                        if name == "AO_Code_To_Add"
                        else values[columns[name] - 1]
                    )
                    for name in DECISION_HEADERS
                }
            )
        return decisions
    finally:
        workbook.close()


def _workbook_codes(path: Path) -> set[str]:
    if not path.is_file():
        raise FileNotFoundError(f"required A-O workbook not found: {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        codes: set[str] = set()
        for worksheet in workbook.worksheets:
            columns = _headers(worksheet)
            code_header = next(
                (name for name in CODE_HEADERS if name in columns),
                None,
            )
            if code_header is None:
                continue
            column = columns[code_header]
            for row in worksheet.iter_rows(
                min_row=2,
                min_col=column,
                max_col=column,
                values_only=True,
            ):
                if row[0] is not None and str(row[0]).strip() != "":
                    codes.add(str(row[0]).strip())
        return codes
    finally:
        workbook.close()


def _remove_late_authority_rows(workbook) -> int:
    """Exclude generated duplicates owned by later governed authority stages."""
    if PARAMETER_ROWS_SHEET not in workbook.sheetnames:
        raise ValueError(
            f"generated extension workbook has no {PARAMETER_ROWS_SHEET!r} sheet"
        )
    worksheet = workbook[PARAMETER_ROWS_SHEET]
    columns = _headers(worksheet)
    required = ("Source_Sheet", "Parameter")
    missing = [name for name in required if name not in columns]
    if missing:
        raise ValueError(
            f"{PARAMETER_ROWS_SHEET} missing required columns: {missing}"
        )

    removed = 0
    for row_index in range(worksheet.max_row, 1, -1):
        source = worksheet.cell(
            row=row_index,
            column=columns["Source_Sheet"],
        ).value
        parameter = worksheet.cell(
            row=row_index,
            column=columns["Parameter"],
        ).value
        if (
            str(source).strip() == LATE_AUTHORITY_SOURCE_SHEET
            and str(parameter).strip() in LATE_AUTHORITY_PARAMETERS
        ):
            worksheet.delete_rows(row_index, 1)
            removed += 1
    return removed


def apply_decisions(
    extensions_path: Path | str,
    authority_path: Path | str,
) -> int:
    """Overlay maintained decision cells without changing generated row order."""
    extensions = Path(extensions_path)
    authority = Path(authority_path)
    decisions = _decision_rows(authority)

    workbook = load_workbook(extensions)
    try:
        if EXTENSIONS_SHEET not in workbook.sheetnames:
            raise ValueError(
                f"{extensions.name} has no {EXTENSIONS_SHEET!r} sheet"
            )
        worksheet = workbook[EXTENSIONS_SHEET]
        columns = _headers(worksheet)
        missing = [name for name in DECISION_HEADERS if name not in columns]
        if missing:
            raise ValueError(
                f"{EXTENSIONS_SHEET} missing required columns: {missing}"
            )
        missing_propagation = [
            name for name in PROPAGATION_TARGETS if name not in columns
        ]
        if missing_propagation:
            raise ValueError(
                f"{EXTENSIONS_SHEET} missing generated propagation columns: "
                f"{missing_propagation}"
            )

        generated_rows: dict[str, int] = {}
        for row_index in range(2, worksheet.max_row + 1):
            value = worksheet.cell(
                row=row_index,
                column=columns["AO_Code_To_Add"],
            ).value
            if value is None or str(value).strip() == "":
                continue
            code = str(value).strip()
            if code in generated_rows:
                raise ValueError(
                    f"generated extension workbook has duplicate code {code!r}"
                )
            generated_rows[code] = row_index

        missing_codes = [
            row["AO_Code_To_Add"]
            for row in decisions
            if row["AO_Code_To_Add"] not in generated_rows
        ]
        if missing_codes:
            raise ValueError(
                "maintained AO decisions do not match freshly generated "
                f"extension codes: {missing_codes}"
            )

        for decision in decisions:
            row_index = generated_rows[str(decision["AO_Code_To_Add"])]
            for name in DECISION_HEADERS[1:]:
                value = decision[name]
                if name != "Include" and (
                    value is None or str(value).strip() == ""
                ):
                    continue
                worksheet.cell(
                    row=row_index,
                    column=columns[name],
                ).value = value

        # An override template is a human decision. Its Add_To_* flags are not:
        # derive those flags from live workbook presence on every run.
        override_decisions = [
            decision
            for decision in decisions
            if decision["Override_Template_AO"] is not None
            and str(decision["Override_Template_AO"]).strip() != ""
        ]
        if override_decisions:
            codes_by_flag = {
                flag: _workbook_codes(extensions.parent / filename)
                for flag, filename in PROPAGATION_TARGETS.items()
            }
            for decision in override_decisions:
                row_index = generated_rows[
                    str(decision["AO_Code_To_Add"])
                ]
                template = str(
                    decision["Override_Template_AO"]
                ).strip()
                for flag, codes in codes_by_flag.items():
                    worksheet.cell(
                        row=row_index,
                        column=columns[flag],
                    ).value = "Y" if template in codes else "N"
        removed = _remove_late_authority_rows(workbook)
        workbook.save(extensions)
        if removed:
            print(
                f"Excluded {removed} generated row(s) owned by later "
                "interconnector authority stages"
            )
        return len(decisions)
    finally:
        workbook.close()


def _default_authority() -> Path:
    # The activated bundle makes this either the full workbook authority or a
    # profile-specific CSV sidecar.  No filename probing or fallback occurs.
    return resolve_paths().ao_decisions


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    script_dir = Path(__file__).resolve().parent
    parser.add_argument(
        "--extensions",
        type=Path,
        default=script_dir / "OSTRAM_AO_Extensions.xlsx",
    )
    parser.add_argument("--authority", type=Path, default=None)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    authority = args.authority if args.authority is not None else _default_authority()
    count = apply_decisions(args.extensions, authority)
    print(
        f"Applied {count} maintained AO decision row(s) to "
        f"{args.extensions.name}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
