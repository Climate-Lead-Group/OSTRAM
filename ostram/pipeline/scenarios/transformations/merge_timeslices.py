# -*- coding: utf-8 -*-
"""
merge_timeslices_into_WV.py

Reads OSTRAM_Timeslice_Inputs.xlsx and produces SOASIA_OSeMOSYS_WV.xlsx, a
working copy of the scenario template supplied through OSTRAM_TEMPLATE_PATH,
with the four timeslice-related sheets rebuilt (tabs flipped from RED -> PINK)
and two audit sheets appended (tabs in LIGHT YELLOW):

  Pink (rebuilt from Timeslice Inputs):
    Yearsplit_Template   <- TS YearSplit                      (broadcast across 2023-2050)
    DaySplit             <- derived from TS Config dayparts   (hours / 8760)
    Demand_Profiles      <- TS {region}_Dem, one block per region
                           (entities sourced from the materialized template,
                           codes ELC<REGION>03)
    Capacities_CF        <- TS {region}_CF, driven by Secondary_Techs PWR roster

  Light yellow (auto-generated audit sheets):
    Tech_Universe        Canonical roster of every PWR/MIN/RNW/TRN tech in the
                         materialized template, with tick-marks per source sheet,
                         NeedsCF flag and CF_Status.
    CF_Provenance        Per-tech CF resolution: parsed category/region, TS sheet
                         used, TS tech_type matched, TS column used, Approximation_Flag,
                         and Resolution bucket (TS_MATCH / TS_DEFAULT / NO_TS_SOURCE / UNRESOLVED).

The source materialized template is NEVER modified (copy first, write to copy only).

The A3 orchestrator supplies the required template path.
"""

import os
import re
from pathlib import Path
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from ostram.pipeline.scenarios.transformations.fix_elc_pmode_revert import (
    configured_country_region_map,
)

# =============================================================================
# USER CONFIGURATION
# =============================================================================

WORK_DIR = str(Path(os.environ["OSTRAM_STAGE_WORKDIR"]).resolve())

# The A3 orchestrator materializes the selected scenario from the maintained
# v18 workbook and supplies that disposable template explicitly.
TEMPLATE_PATH = os.environ.get("OSTRAM_TEMPLATE_PATH")
if not TEMPLATE_PATH:
    raise RuntimeError(
        "OSTRAM_TEMPLATE_PATH is required; run through python -m ostram transform"
    )
TIMESLICE_FILE = WORK_DIR + "/OSTRAM_Timeslice_Inputs.xlsx"
WV_FILE        = WORK_DIR + "/SOASIA_OSeMOSYS_WV.xlsx"

YEAR_START = 2023
YEAR_END   = 2050

# Tab colors (ARGB)
TAB_COLOR_PINK = "FFFFB6C1"  # rebuilt data sheets (were RED in v17)
TAB_COLOR_META = "FFFFEB9C"  # auto-generated audit sheets (light yellow)

# v17 region (5-char Tech infix) -> TS sheet prefix, derived from the active
# profile country authority.  The full country list reproduces the historical
# ten-region mapping exactly; reduced profiles load only their own sheets.
REGION_MAP = configured_country_region_map()
REGION_MAP_INV = {v: k for k, v in REGION_MAP.items()}

# v17 tech category (3-char infix after PWR) -> TS tech_type candidates,
# tried in priority order. Empty list -> use CF_DEFAULT for that category.
# (CSP and WAV are present as dormant placeholders; they don't appear in v17
# Secondary_Techs but stay here so the mapping is robust to future additions.)
CF_TECH_MAP = {
    "SPV": ["Solar"],
    "CSP": ["Solar"],                                                      # CSP uses PV solar profile (approximation)
    "WON": ["Wind"],
    "WOF": ["Wind"],                                                       # offshore -> onshore profile (approximation)
    "HYD": ["Hydro_HYD", "Hydro_HDR", "Hydro_HRO",
            "Major_Hydro", "Mini_Hydro", "Hydro"],                         # all hydro variants identical in TS data
    "SHP": ["Hydro_HRO", "Mini_Hydro", "Hydro_HYD", "Hydro"],              # Small Hydropower -> RoR-style profile preferred
    "BIO": ["Biomass"],
    "COA": ["COA", "Coal"],
    "NGS": ["NGA", "GAS", "Gas"],
    "OIL": ["OIL", "Oil_HFO", "Oil_HSD", "Oil_CEB", "Oil_IPP", "Diesel"],
    "GEO": [],                                                             # no TS source -> CF_DEFAULT
    "URN": [],                                                             # no TS source -> CF_DEFAULT
    "WAS": [],                                                             # no TS source -> CF_DEFAULT
    "SDS": [],                                                             # storage placeholder -> CF_DEFAULT
    "LDS": [],                                                             # storage placeholder -> CF_DEFAULT
    "WAV": [],                                                             # tidal/marine placeholder -> CF_DEFAULT
}
CF_DEFAULT = 1.0  # dispatchable / storage / no-source placeholder

# CF column priority within a tt-matched row. cf_ninja is FIRST because
# Renewables.ninja is modeled from reanalysis and is bounded to [0,1] by
# construction -- this avoids the numerical noise that contaminates
# cf_dispatch for renewables (e.g., LKA solar with slightly negative values
# from CF inversion, LKA wind exceeding 1.0 from capacity-denominator
# mismatches in utility dispatch records).
#
# In practice cf_ninja is only populated for Solar/Wind tech_types, so this
# priority change only affects renewables: hydro still falls through to
# cf_dispatch (where dispatch realism matters and ninja has no entry), and
# thermal/import still falls through to cf_default.
CF_COLUMN_PRIORITY = ["cf_ninja", "cf_dispatch", "cf_da_workbook", "cf_default"]

# Surfaced in CF_Provenance for techs using approximate or placeholder sources.
APPROXIMATION_FLAGS = {
    "CSP": "CSP uses PV Solar profile (thermal inertia smoothing not modeled)",
    "WOF": "Offshore wind uses onshore Wind profile (sea-surface uplift ignored)",
    "SHP": "Small Hydropower uses RoR-equivalent hydro profile",
    "GEO": "No TS source - CF=1.0 placeholder",
    "URN": "No TS source - CF=1.0 placeholder",
    "WAS": "No TS source - CF=1.0 placeholder",
    "SDS": "Short-duration storage - no TS source, CF=1.0 placeholder",
    "LDS": "Long-duration storage - no TS source, CF=1.0 placeholder",
    "WAV": "Tidal/wave - no TS source, CF=1.0 placeholder",
}

# v17 sheets scanned to build the Tech_Universe roster.
V17_TECH_SHEETS = [
    "Primary_Techs", "Secondary_Techs", "Capacities_CF",
    "VariableCost", "Demand_Techs", "Emissions",
    "Interconnector_Params", "Fixed_Horizon_Parameters",
]

# Recognised tech prefixes:
#   PWR = power generation / storage  (need CFs)
#   MIN = mining / primary supply (coal, gas, oil, uranium)
#   TRN = transmission interconnectors
#   RNW = renewables resource accounting
TECH_PREFIXES = ("PWR", "MIN", "TRN", "RNW")

YEARS = list(range(YEAR_START, YEAR_END + 1))

# =============================================================================
# HELPERS
# =============================================================================

def ts_sort_key(s):
    """Natural sort for timeslice strings like 'S1D1', 'S1D2', ..., 'S4D5'."""
    m = re.match(r"S(\d+)D(\d+)", str(s))
    return (int(m.group(1)), int(m.group(2))) if m else (99, 99)

def broadcast_years(base_row, value):
    """Attach {year: value} pairs for all YEARS to a row dict."""
    for y in YEARS:
        base_row[y] = value
    return base_row

def replace_sheet(wb, sheet_name, df, tab_color):
    """Clear an existing sheet in-place and write df; set tab color."""
    ws = wb[sheet_name]
    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    ws.sheet_properties.tabColor = tab_color

def create_or_replace_sheet(wb, sheet_name, df, tab_color):
    """Create sheet (or clear existing) and write df."""
    if sheet_name in wb.sheetnames:
        replace_sheet(wb, sheet_name, df, tab_color)
    else:
        ws = wb.create_sheet(sheet_name)
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)
        ws.sheet_properties.tabColor = tab_color

def parse_v17_tech(tech):
    """'PWRSPVBGDXX' -> ('SPV', 'BGDXX'); 'PWRHYDINDNE' -> ('HYD', 'INDNE').

    v17 format: PWR + 3-char category + 5-char region. 11 chars total.
    """
    t = str(tech)
    if len(t) != 11 or not t.startswith("PWR"):
        return None, None
    return t[3:6], t[6:11]

def v17_fuel_to_region(fuel_tech):
    """'ELCBGDXX03' -> 'BGDXX';  'ELCINDNO03' -> 'INDNO'."""
    s = str(fuel_tech)
    if len(s) != 10 or not s.startswith("ELC"):
        return None
    return s[3:8]  # 5-char region between ELC...XX/sub-region and 03 suffix

# =============================================================================
# STEP 1 -- COPY V17 -> WV (Setup-A: source untouched)
# =============================================================================

print("=" * 72)
print("STEP 1 -- Copy v17 Template -> WV")
print("=" * 72)

if not os.path.exists(TEMPLATE_PATH):
    raise RuntimeError(f"materialized template not found: {TEMPLATE_PATH}")
if not os.path.exists(TIMESLICE_FILE):
    raise RuntimeError(f"Timeslice file not found: {TIMESLICE_FILE}")

shutil.copy(TEMPLATE_PATH, WV_FILE)
print(f"  source: {TEMPLATE_PATH}")
print(f"  target: {WV_FILE}")

wb = load_workbook(WV_FILE)

# =============================================================================
# STEP 2 -- Yearsplit_Template
# =============================================================================

print("\n" + "=" * 72)
print("STEP 2 -- Yearsplit_Template")
print("=" * 72)

ys_lean = pd.read_excel(TIMESLICE_FILE, sheet_name="YearSplit")
for c in ["timeslice", "yearsplit"]:
    if c not in ys_lean.columns:
        raise RuntimeError(f"YearSplit sheet missing column '{c}'")
ys_lean = ys_lean.sort_values(
    "timeslice", key=lambda c: c.map(ts_sort_key)
).reset_index(drop=True)

rows = []
for _, r in ys_lean.iterrows():
    row = {
        "Timeslices":           r["timeslice"],
        "Parameter.ID":         14,
        "Parameter":            "YearSplit",
        "Unit":                 "fraction",
        "Projection.Mode":      "User defined",
        "Projection.Parameter": 0,
    }
    rows.append(broadcast_years(row, float(r["yearsplit"])))

cols_ys = ["Timeslices", "Parameter.ID", "Parameter", "Unit",
           "Projection.Mode", "Projection.Parameter"] + YEARS
df_ys = pd.DataFrame(rows, columns=cols_ys)
replace_sheet(wb, "Yearsplit_Template", df_ys, TAB_COLOR_PINK)
print(f"  Wrote {len(df_ys)} timeslice rows. "
      f"Sum(YearSplit) = {df_ys[YEAR_START].sum():.6f}")

# =============================================================================
# STEP 3 -- DaySplit (derived from Config sheet)
# =============================================================================

print("\n" + "=" * 72)
print("STEP 3 -- DaySplit (derived from Config hour-ranges)")
print("=" * 72)

cfg_raw = pd.read_excel(TIMESLICE_FILE, sheet_name="Config", header=None)
cfg_raw.columns = ["key", "value"]
dp_cfg = cfg_raw[cfg_raw["key"].astype(str).str.startswith("daypart_")].copy()
dp_cfg["bracket"] = dp_cfg["key"].str.replace("daypart_D", "", regex=False).astype(int)

def parse_hours(label):
    m = re.search(r"\((\d{1,2})-(\d{1,2})\)", str(label))
    if not m:
        raise RuntimeError(f"Cannot parse hours from daypart label: {label!r}")
    return int(m.group(2)) - int(m.group(1))

dp_cfg["hours"]    = dp_cfg["value"].apply(parse_hours)
dp_cfg["daysplit"] = dp_cfg["hours"] / 8760.0
dp_cfg = dp_cfg.sort_values("bracket").reset_index(drop=True)

if dp_cfg["hours"].sum() != 24:
    raise RuntimeError(f"DaySplit hours must sum to 24; got {dp_cfg['hours'].sum()}\n{dp_cfg}")

rows = []
for _, r in dp_cfg.iterrows():
    row = {
        "DAILYTIMEBRACKET":     int(r["bracket"]),
        "Parameter.ID":         12,
        "Parameter":            "DaySplit",
        "Unit":                 "fraction",
        "Projection.Mode":      "User defined",
        "Projection.Parameter": 0,
    }
    rows.append(broadcast_years(row, float(r["daysplit"])))

cols_ds = ["DAILYTIMEBRACKET", "Parameter.ID", "Parameter", "Unit",
           "Projection.Mode", "Projection.Parameter"] + YEARS
df_ds = pd.DataFrame(rows, columns=cols_ds)
replace_sheet(wb, "DaySplit", df_ds, TAB_COLOR_PINK)
print(f"  Wrote {len(df_ds)} brackets ({dp_cfg['hours'].sum()}h total). "
      f"Sum(DaySplit) = {df_ds[YEAR_START].sum():.6f} (expect {24/8760:.6f})")

# =============================================================================
# STEP 4 -- Demand_Profiles
# =============================================================================
# Use the existing Fuel/Tech codes from v17 Demand_Profiles (ELC<REGION>03);
# each entity's profile comes from {ts_region}_Dem.  A region configured in
# the country authority but absent from the v17 sheet (a country added with
# `country template` + `country merge` whose demand entity was never inserted
# into the scenario workbook) gets its canonical ELC<REGION>03 entity added
# here, so the new country still receives SpecifiedDemandProfile rows.

print("\n" + "=" * 72)
print("STEP 4 -- Demand_Profiles")
print("=" * 72)

dp_v17 = pd.read_excel(TEMPLATE_PATH, sheet_name="Demand_Profiles")
dp_v17 = dp_v17.dropna(subset=["Timeslices", "Fuel/Tech"]).reset_index(drop=True)

# Carry over the entity-defining columns from v17 (Fuel/Tech, Name, Demand/Share, Refs)
ent_cols = ["Fuel/Tech", "Name", "Demand/Share", "Ref.Cap.BY", "Ref.OAR.BY", "Ref.km.BY"]
entities = dp_v17[ent_cols].drop_duplicates(subset=["Fuel/Tech"]).reset_index(drop=True)
entities["region"]    = entities["Fuel/Tech"].apply(v17_fuel_to_region)
entities["ts_prefix"] = entities["region"].map(REGION_MAP)

unmapped = entities[entities["ts_prefix"].isna()]
if not unmapped.empty:
    raise RuntimeError(f"Cannot map these v17 Fuel/Tech codes:\n{unmapped[['Fuel/Tech','region']]}")

print(f"  {len(entities)} demand entities found in v17 Demand_Profiles")

# Configured regions with no demand entity in the v17 sheet: add the
# canonical ELC<REGION>03 entity so the region is not silently skipped.
# A configured region whose _Dem sheet is missing from the timeslice
# workbook is a hard error -- `country template` should have cloned it.
_ts_sheet_names = set(pd.ExcelFile(TIMESLICE_FILE).sheet_names)
_covered_regions = set(entities["region"].dropna())
_added_entities = []
for _region, _prefix in sorted(REGION_MAP.items()):
    if _region in _covered_regions:
        continue
    _dem_sheet = f"{_prefix}_Dem"
    if _dem_sheet not in _ts_sheet_names:
        raise RuntimeError(
            f"Region {_region} is configured in the country authority but "
            f"sheet '{_dem_sheet}' is missing from the timeslice workbook "
            f"({TIMESLICE_FILE}). Run `country template` so the timeslice "
            f"sheets are cloned for the new country, then re-run."
        )
    _added_entities.append({
        "Fuel/Tech":    f"ELC{_region}03",
        "Name":         f"Output demand of transmission lines in {_region}",
        "Demand/Share": "Demand",
        "Ref.Cap.BY":   "not needed",
        "Ref.OAR.BY":   "not needed",
        "Ref.km.BY":    "not needed",
        "region":       _region,
        "ts_prefix":    _prefix,
    })
    print(f"  + Added demand entity ELC{_region}03 for configured region "
          f"{_region} (profile from sheet '{_dem_sheet}'; entity not "
          f"present in v17 Demand_Profiles)")
if _added_entities:
    entities = pd.concat(
        [entities, pd.DataFrame(_added_entities)], ignore_index=True
    )
    print(f"  {len(entities)} demand entities total after adding "
          f"{len(_added_entities)} configured region(s)")

new_rows = []
for _, ent in entities.iterrows():
    dem = pd.read_excel(TIMESLICE_FILE, sheet_name=f"{ent['ts_prefix']}_Dem")
    dem = dem.sort_values(
        "timeslice", key=lambda c: c.map(ts_sort_key)
    ).reset_index(drop=True)
    for _, d in dem.iterrows():
        row = {
            "Timeslices":           d["timeslice"],
            "Demand/Share":         ent["Demand/Share"] if pd.notna(ent["Demand/Share"]) else "Demand",
            "Fuel/Tech":            ent["Fuel/Tech"],
            "Name":                 ent["Name"],
            "Ref.Cap.BY":           ent["Ref.Cap.BY"] if pd.notna(ent["Ref.Cap.BY"]) else "not needed",
            "Ref.OAR.BY":           ent["Ref.OAR.BY"] if pd.notna(ent["Ref.OAR.BY"]) else "not needed",
            "Ref.km.BY":            ent["Ref.km.BY"]  if pd.notna(ent["Ref.km.BY"])  else "not needed",
            "Projection.Mode":      "User defined",
            "Projection.Parameter": 0,
        }
        new_rows.append(broadcast_years(row, float(d["demand_fraction"])))

cols_dp = ["Timeslices", "Demand/Share", "Fuel/Tech", "Name",
           "Ref.Cap.BY", "Ref.OAR.BY", "Ref.km.BY",
           "Projection.Mode", "Projection.Parameter"] + YEARS
df_dp = pd.DataFrame(new_rows, columns=cols_dp)
replace_sheet(wb, "Demand_Profiles", df_dp, TAB_COLOR_PINK)
print(f"  Wrote {len(df_dp)} rows ({len(entities)} entities x {len(ys_lean)} timeslices)")

# =============================================================================
# STEP 5 -- Capacities_CF (drives provenance; driven by Secondary_Techs roster)
# =============================================================================

print("\n" + "=" * 72)
print("STEP 5 -- Capacities_CF (driven by Secondary_Techs PWR roster)")
print("=" * 72)

st_v17 = pd.read_excel(TEMPLATE_PATH, sheet_name="Secondary_Techs")
st_v17 = st_v17.dropna(subset=["Tech"]).copy()

# Unique PWR codes -> first non-null (Tech.ID, Tech.Name)
pwr_mask = st_v17["Tech"].astype(str).str.startswith("PWR")
roster = (st_v17.loc[pwr_mask, ["Tech.ID", "Tech", "Tech.Name"]]
                .drop_duplicates(subset=["Tech"])
                .reset_index(drop=True))
print(f"  {len(roster)} unique PWR Tech codes in Secondary_Techs")

# Pre-load all CF sheets
cf_sheets = {}
for prefix in REGION_MAP.values():
    cf_sheets[prefix] = pd.read_excel(TIMESLICE_FILE, sheet_name=f"{prefix}_CF")

# Build a default timeslice list from YearSplit (authoritative)
default_timeslices = sorted(ys_lean["timeslice"].astype(str).tolist(), key=ts_sort_key)

def lookup_cf(category, ts_prefix):
    """Resolve CF for (category, region) -> ({timeslice: cf}, resolution_bucket, tt_used, col_used).

    Buckets:
      TS_MATCH       - matched a tech_type in the regional CF sheet
      TS_DEFAULT     - category has TS candidates but none present in this region's sheet
      NO_TS_SOURCE   - category has NO TS candidates (placeholder/unsupported)
      UNRESOLVED     - parse failure / unknown category (loud)
    """
    if category not in CF_TECH_MAP:
        return ({ts: CF_DEFAULT for ts in default_timeslices},
                "UNRESOLVED", "(unknown category)", f"CF_DEFAULT={CF_DEFAULT}")

    candidates = CF_TECH_MAP[category]
    cf_df = cf_sheets[ts_prefix]

    if not candidates:
        return ({ts: CF_DEFAULT for ts in default_timeslices},
                "NO_TS_SOURCE", "(none)", f"CF_DEFAULT={CF_DEFAULT}")

    for tt in candidates:
        sub = cf_df[cf_df["tech_type"] == tt]
        if sub.empty:
            continue
        out = {}
        col_used_per_row = []
        for _, r in sub.iterrows():
            chosen = None
            for col in CF_COLUMN_PRIORITY:
                v = r.get(col)
                if pd.notna(v):
                    chosen = (col, float(v))
                    break
            if chosen is None:
                # all priority columns NaN for this row -> fall back to default
                out[r["timeslice"]] = CF_DEFAULT
                col_used_per_row.append("CF_DEFAULT")
            else:
                out[r["timeslice"]] = chosen[1]
                col_used_per_row.append(chosen[0])
        if out:
            # Report the dominant column for provenance (first non-default in priority order)
            for col in CF_COLUMN_PRIORITY:
                if col in col_used_per_row:
                    col_used = col
                    break
            else:
                col_used = "CF_DEFAULT"
            return out, "TS_MATCH", tt, col_used

    # candidates listed but none present in this region's sheet
    return ({ts: CF_DEFAULT for ts in default_timeslices},
            "TS_DEFAULT", "(no candidate present in region)", f"CF_DEFAULT={CF_DEFAULT}")

new_rows = []
provenance = []
tally = {"TS_MATCH": 0, "TS_DEFAULT": 0, "NO_TS_SOURCE": 0, "UNRESOLVED": 0}
unresolved_log = []

for _, t in roster.iterrows():
    tech = t["Tech"]
    tech_id = t["Tech.ID"]
    tech_name = t["Tech.Name"] if pd.notna(t["Tech.Name"]) else ""
    category, region = parse_v17_tech(tech)
    ts_prefix = REGION_MAP.get(region) if region else None

    prov = {
        "Tech":               tech,
        "Tech.Name":          tech_name,
        "Parsed.Category":    category if category else "(parse-fail)",
        "Parsed.Region":      region if region else "(parse-fail)",
        "TS.Prefix":          ts_prefix or "-",
        "Resolution":         "",
        "TS.Sheet":           "-",
        "TS.Tech_Type":       "-",
        "TS.Column":          "-",
        "Approximation_Flag": APPROXIMATION_FLAGS.get(category, "") if category else "",
    }

    if category is None or region is None:
        unresolved_log.append(f"    {tech}: parse failure (category={category}, region={region})")
        tally["UNRESOLVED"] += 1
        prov["Resolution"] = "UNRESOLVED - tech code parse failure"
        cf_values = {ts: CF_DEFAULT for ts in default_timeslices}
        provenance.append(prov)
    elif ts_prefix is None:
        unresolved_log.append(f"    {tech}: region '{region}' not in REGION_MAP")
        tally["UNRESOLVED"] += 1
        prov["Resolution"] = "UNRESOLVED - region not mapped"
        cf_values = {ts: CF_DEFAULT for ts in default_timeslices}
        provenance.append(prov)
    else:
        cf_values, bucket, tt_used, col_used = lookup_cf(category, ts_prefix)
        tally[bucket] += 1
        prov["Resolution"]   = bucket
        prov["TS.Sheet"]     = f"{ts_prefix}_CF"
        prov["TS.Tech_Type"] = tt_used
        prov["TS.Column"]    = col_used
        if bucket == "UNRESOLVED":
            unresolved_log.append(f"    {tech}: unknown category '{category}'")
        provenance.append(prov)

    # Always write rows (CF=1.0 placeholder for non-TS-matched cases)
    for ts in sorted(cf_values.keys(), key=ts_sort_key):
        row = {
            "Timeslices":           ts,
            "Tech.ID":              int(tech_id) if pd.notna(tech_id) else 0,
            "Tech":                 tech,
            "Tech.Name":            tech_name,
            "Parameter.ID":         13,
            "Parameter":            "CapacityFactor",
            "Unit":                 "fraction",
            "Projection.Mode":      "User defined",
            "Projection.Parameter": 0,
        }
        new_rows.append(broadcast_years(row, cf_values[ts]))

cols_cf = ["Timeslices", "Tech.ID", "Tech", "Tech.Name",
           "Parameter.ID", "Parameter", "Unit",
           "Projection.Mode", "Projection.Parameter"] + YEARS
df_cf = pd.DataFrame(new_rows, columns=cols_cf)
replace_sheet(wb, "Capacities_CF", df_cf, TAB_COLOR_PINK)
print(f"  Wrote {len(df_cf)} rows ({len(roster)} techs x {len(default_timeslices)} timeslices)")
print(f"  Resolution tally: TS_MATCH={tally['TS_MATCH']}, "
      f"TS_DEFAULT={tally['TS_DEFAULT']}, "
      f"NO_TS_SOURCE={tally['NO_TS_SOURCE']}, "
      f"UNRESOLVED={tally['UNRESOLVED']}")
if unresolved_log:
    print("  WARN UNRESOLVED techs:")
    for u in unresolved_log:
        print(u)

# =============================================================================
# STEP 6 -- Tech_Universe (audit sheet)
# =============================================================================

print("\n" + "=" * 72)
print("STEP 6 -- Tech_Universe (audit)")
print("=" * 72)

appears_in = {}   # tech -> set of v17 sheet names
canon_name = {}   # tech -> canonical Tech.Name (first non-null encountered)

for sheet in V17_TECH_SHEETS:
    # BUG-1 FIX: Capacities_CF was rebuilt in-memory by Step 5 (df_cf).
    # Re-reading it from TEMPLATE_PATH would return the OLD rows (WON/WOF only),
    # causing every other PWR tech that gained CF rows in Step 5 to stay
    # unmarked in Tech_Universe.  Use df_cf directly for this sheet; all other
    # sheets are untouched by Script 1 and the template read remains correct.
    if sheet == "Capacities_CF":
        df = df_cf.copy()
    else:
        try:
            df = pd.read_excel(TEMPLATE_PATH, sheet_name=sheet)
        except Exception as e:
            print(f"  (skipping '{sheet}': {e})")
            continue
    if "Tech" not in df.columns:
        continue
    name_col = "Tech.Name" if "Tech.Name" in df.columns else None
    for _, r in df.dropna(subset=["Tech"]).iterrows():
        t = str(r["Tech"])
        if not t.startswith(TECH_PREFIXES):
            continue
        appears_in.setdefault(t, set()).add(sheet)
        if name_col and pd.notna(r[name_col]) and t not in canon_name:
            canon_name[t] = r[name_col]

# Map each tech in the rebuilt CF -> CF_Status from provenance.
# NeedsCF is anchored to the canonical generation roster (Secondary_Techs PWRs)
# so transmission techs like PWRTRN* (which live only in Demand_Techs /
# Fixed_Horizon_Parameters) are correctly NOT flagged as needing a CF.
status_by_tech = {p["Tech"]: p["Resolution"] for p in provenance}
generation_roster = set(roster["Tech"])

rows_tu = []
for t in sorted(appears_in.keys()):
    row = {
        "Tech":      t,
        "Prefix":    t[:3],
        "Tech.Name": canon_name.get(t, ""),
    }
    for s in V17_TECH_SHEETS:
        row[s] = "Y" if s in appears_in[t] else ""
    needs_cf = "Y" if t in generation_roster else ""
    row["NeedsCF"]   = needs_cf
    if needs_cf:
        row["CF_Status"] = status_by_tech.get(t, "(not processed)")
    else:
        row["CF_Status"] = "-"
    rows_tu.append(row)

cols_tu = ["Tech", "Prefix", "Tech.Name"] + V17_TECH_SHEETS + ["NeedsCF", "CF_Status"]
df_tu = pd.DataFrame(rows_tu, columns=cols_tu)
create_or_replace_sheet(wb, "Tech_Universe", df_tu, TAB_COLOR_META)

prefix_counts = df_tu["Prefix"].value_counts().to_dict()
breakdown = ", ".join(f"{p}={n}" for p, n in sorted(prefix_counts.items()))
print(f"  Wrote {len(df_tu)} distinct techs ({breakdown}) from {len(V17_TECH_SHEETS)} v17 sheets")

# =============================================================================
# STEP 7 -- CF_Provenance (audit sheet)
# =============================================================================

print("\n" + "=" * 72)
print("STEP 7 -- CF_Provenance (audit)")
print("=" * 72)

cols_cp = ["Tech", "Tech.Name", "Parsed.Category", "Parsed.Region",
           "TS.Prefix", "Resolution", "TS.Sheet", "TS.Tech_Type",
           "TS.Column", "Approximation_Flag"]
df_cp = pd.DataFrame(provenance, columns=cols_cp)
create_or_replace_sheet(wb, "CF_Provenance", df_cp, TAB_COLOR_META)

print(f"  Wrote {len(df_cp)} rows. Resolution breakdown:")
for res, n in df_cp["Resolution"].value_counts().items():
    print(f"    {res:45s} {n:4d}")
n_approx = df_cp["Approximation_Flag"].ne("").sum()
print(f"  Techs carrying an Approximation_Flag: {n_approx}")

# =============================================================================
# STEP 8 -- SAVE
# =============================================================================

print("\n" + "=" * 72)
print("STEP 8 -- Save WV")
print("=" * 72)
wb.save(WV_FILE)
print(f"  Saved: {WV_FILE}")

# =============================================================================
# BUILT-IN TESTS
# =============================================================================

print("\n" + "=" * 72)
print("TESTS")
print("=" * 72)

_passed, _failed = 0, 0
def check(name, cond, detail=""):
    global _passed, _failed
    if cond:
        _passed += 1
        print(f"  PASS  {name}")
    else:
        _failed += 1
        print(f"  FAIL  {name}   {detail}")

wb2 = load_workbook(WV_FILE)

# (a) v17 source untouched
check("materialized source file still exists", os.path.exists(TEMPLATE_PATH))

# (b) 4 pink sheets present + correct color
for s in ["Yearsplit_Template", "Capacities_CF", "Demand_Profiles", "DaySplit"]:
    check(f"sheet '{s}' exists in WV", s in wb2.sheetnames)
    tc = wb2[s].sheet_properties.tabColor
    val = tc.value if tc is not None else None
    check(f"'{s}' tab is PINK", val == TAB_COLOR_PINK, f"(got {val})")

# (c) 2 audit sheets present + correct color
for s in ["Tech_Universe", "CF_Provenance"]:
    check(f"audit sheet '{s}' exists", s in wb2.sheetnames)
    tc = wb2[s].sheet_properties.tabColor
    val = tc.value if tc is not None else None
    check(f"'{s}' tab is META (yellow)", val == TAB_COLOR_META, f"(got {val})")

# (d) Row counts match in-memory dfs
ys_d = pd.read_excel(WV_FILE, sheet_name="Yearsplit_Template")
ds_d = pd.read_excel(WV_FILE, sheet_name="DaySplit")
dp_d = pd.read_excel(WV_FILE, sheet_name="Demand_Profiles")
cf_d = pd.read_excel(WV_FILE, sheet_name="Capacities_CF")
tu_d = pd.read_excel(WV_FILE, sheet_name="Tech_Universe")
cp_d = pd.read_excel(WV_FILE, sheet_name="CF_Provenance")

check(f"Yearsplit rows = {len(df_ys)}",       len(ys_d) == len(df_ys), f"(got {len(ys_d)})")
check(f"DaySplit rows = {len(df_ds)}",        len(ds_d) == len(df_ds), f"(got {len(ds_d)})")
check(f"Demand_Profiles rows = {len(df_dp)}", len(dp_d) == len(df_dp), f"(got {len(dp_d)})")
check(f"Capacities_CF rows = {len(df_cf)}",   len(cf_d) == len(df_cf), f"(got {len(cf_d)})")
check(f"Tech_Universe rows = {len(df_tu)}",   len(tu_d) == len(df_tu), f"(got {len(tu_d)})")
check(f"CF_Provenance rows = {len(df_cp)}",   len(cp_d) == len(df_cp), f"(got {len(cp_d)})")

# (e) Sum invariants
check("YearSplit sums to 1.0 (+/- 0.001)",
      abs(ys_d[YEAR_START].sum() - 1.0) < 0.001,
      f"(got {ys_d[YEAR_START].sum():.6f})")
check("DaySplit sums to 24/8760",
      abs(ds_d[YEAR_START].sum() - 24/8760) < 1e-9,
      f"(got {ds_d[YEAR_START].sum():.6f})")
for ft in dp_d["Fuel/Tech"].dropna().unique():
    s = dp_d[dp_d["Fuel/Tech"] == ft][YEAR_START].sum()
    check(f"Demand_Profiles[{ft}] sums to ~1.0",
          abs(s - 1.0) < 0.01, f"(got {s:.6f})")

# (f) CF values in [0, 1] -- WARNING (not a failure): handover says "warn if not".
#     Out-of-range values originate in the timeslice source data, not the script.
year_cols = [c for c in cf_d.columns if isinstance(c, int)]
oor_mask = (cf_d[year_cols] < 0).any(axis=1) | (cf_d[year_cols] > 1).any(axis=1)
oor = cf_d[oor_mask]
if len(oor) == 0:
    _passed += 1
    print("  PASS  Capacities_CF values in [0, 1]")
else:
    print(f"  WARN  Capacities_CF has {len(oor)} out-of-range rows (source data issue, not script bug)")
    print(f"        First {min(len(oor),20)} of {len(oor)} offenders:")
    for _, r in oor[["Timeslices", "Tech", YEAR_START]].head(20).iterrows():
        print(f"          {r['Timeslices']:6s}  {r['Tech']:18s}  {r[YEAR_START]:+.6f}")

# (g) No NaN in year columns of any rebuilt sheet
for name, d in [("Yearsplit_Template", ys_d), ("DaySplit", ds_d),
                ("Demand_Profiles", dp_d), ("Capacities_CF", cf_d)]:
    ycs = [c for c in d.columns if isinstance(c, int)]
    n_nan = d[ycs].isna().sum().sum()
    check(f"{name} has no NaN in year columns", n_nan == 0,
          f"(found {n_nan} NaN cells)")

# (h) Every Secondary_Techs PWR/SHP code present in rebuilt Capacities_CF
roster_techs = set(roster["Tech"])
cf_techs_set = set(cf_d["Tech"].dropna())
missing = roster_techs - cf_techs_set
check("All Secondary_Techs PWR codes present in new Capacities_CF",
      len(missing) == 0, f"(missing {len(missing)}: {sorted(missing)[:5]}...)")

# (i) Every NeedsCF tech in Tech_Universe has a CF_Provenance row
tu_needs_cf = set(tu_d[tu_d["NeedsCF"] == "Y"]["Tech"])
cp_techs    = set(cp_d["Tech"].dropna())
gap = tu_needs_cf - cp_techs
check("All NeedsCF techs from Tech_Universe appear in CF_Provenance",
      len(gap) == 0, f"(gap of {len(gap)}: {sorted(gap)[:5]}...)")

# (j) No UNRESOLVED rows in CF_Provenance
n_unres = cp_d["Resolution"].astype(str).str.startswith("UNRESOLVED").sum()
check("No UNRESOLVED techs in CF_Provenance",
      n_unres == 0, f"({n_unres} unresolved - review CF_TECH_MAP / REGION_MAP)")

# (k) Non-target sheets preserved
for s in ["Fixed_Horizon_Parameters", "Primary_Techs", "Secondary_Techs",
          "VariableCost", "Demand_Techs", "Emissions",
          "Existing_Generation", "Planned_Generation",
          "Technology_Costs", "RE_Targets_Policies",
          "Interconnectors", "Interconnector_Params", "README"]:
    check(f"preserved sheet '{s}' still exists", s in wb2.sheetnames)

print(f"\n  {_passed} passed, {_failed} failed")
if _failed == 0:
    print("  ALL TESTS PASSED")
else:
    print(f"  {_failed} TEST(S) FAILED - review output above")
