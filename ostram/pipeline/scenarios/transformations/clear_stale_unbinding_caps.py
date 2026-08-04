"""
clear_stale_unbinding_caps.py
==============================

Pre-lid data fix. Resets stale 'unbinding cap' values in
TotalAnnualMaxCapacityInvestment cells of PWRSPV* and PWRWON* techs
that are flat across the full 2024-2050 horizon, so the lid script
(add_max_cap_investment_lid_rule.py) can write the year-schedule
into them.

Background
----------
The lid script preserves any existing non-9999 cell value as a
'manual calibration'. A subset of PWR techs in the c2a-patched
workbook have flat (non-9999) numbers in MaxCapInv for some
contiguous block starting at 2024. Two variants of the same
disease:

  Pattern A -- full-horizon cliff (PWRSPV*, PWRWON*, 14 techs):
    2023=9999, 2024-2050 = a single flat number
    e.g. PWRWONINDNO = 186.004 every year
    Causes a sudden 30x jump in lid headroom from 2023 to 2024
    and lets the deferred 2023 surge land as a ~440 GW dump.

  Pattern B -- planning-window cap (PWRHYD*, PWRSHP*, 13 techs):
    2023=9999, 2024-2030 = a single flat number,
    2031-2050 = 9999
    e.g. PWRHYDINDNO = 36.776 for 2024-2030 then 9999 thereafter
    Causes the same lid cliff in 2024-2030 then a transition.

Both patterns are stale unbinding caps unrelated to actual
project schedules. The real project commit data lives in
TotalAnnualMinCapacityInvestment (year-specific non-flat values
matching plant commit dates). The lid script's untie rule
(MinCap >= proposed lid -> bump lid to MinCap * 1.01) preserves
those floors automatically, so resetting the MaxCapInv values
does not lose any planning information.

This patch resets MaxCapInv 2024-2050 to 9999 for all 27 affected
techs. On the next lid-script run, the year-schedule writes into
every cell, with the untie rule raising the lid above the MinCap
floor wherever a real project commitment is scheduled.

Scope
-----
ONLY patches techs in TECHS_TO_PATCH (27 techs) that satisfy ALL of:
  - Parameter == 'TotalAnnualMaxCapacityInvestment'
  - 2023 cell == 9999
  - At least one cell in 2024-2050 has a non-9999 value
  - All non-9999 values in 2024-2050 are equal (single flat number)

Does NOT touch:
  - TotalAnnualMinCapacityInvestment (where real project floors live)
  - Any other parameter or sheet
  - TRN* transmission interconnects (lid script skips these by design)
  - Any tech outside TECHS_TO_PATCH, even if it matches the
    pattern -- explicit allowlist for safety

Output
------
- Patched workbook written as a sibling with `_POST_CAP_RESET` appended to the
  stem. Long mutable-workspace paths are deterministically compacted while
  retaining that stage identity. The input file is left untouched.
- Per-tech summary printed to stdout.
- The lid script should be pointed at the new POST_CAP_RESET file.

Usage
-----
    python clear_stale_unbinding_caps.py
    python clear_stale_unbinding_caps.py --input path/to/A-O_Parametrization.xlsx
    python clear_stale_unbinding_caps.py --dry-run

The input file is never modified. A deterministic `_POST_CAP_RESET` sibling is
created next to it. The containing A3 run directory already provides run-level
isolation.
"""

from __future__ import annotations
import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook

from ostram.paths import bounded_workspace_workbook_path

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
DEFAULT_INPUT = Path("A-O_Parametrization.xlsx")
SHEET_NAME = "Secondary Techs"
TECH_COL_NAME = "Tech"
PARAM_COL_NAME = "Parameter"
TARGET_PARAMETER = "TotalAnnualMaxCapacityInvestment"
PLACEHOLDER = 9999
RESET_YEARS = list(range(2024, 2051))  # 2024 through 2050 inclusive

# Allowlist: techs whose stale flat caps in MaxCapInv should be reset
# to 9999 so the lid script can write the year-schedule.
#
# Pattern A -- full-horizon cliff (PWRSPV*, PWRWON*): 14 techs,
#   2023=9999, 2024-2050 all equal a single flat number.
# Pattern B -- planning-window cap (PWRHYD*, PWRSHP*): 13 techs,
#   2023=9999, 2024-2030 a single flat number, 2031-2050 = 9999.
#   Real project floors are in MinCapInv; resetting MaxCapInv loses
#   no planning data because the lid script's untie rule re-derives
#   a lid above MinCap automatically.
TECHS_TO_PATCH = {
    # Pattern A
    "PWRSPVBGDXX", "PWRSPVBTNXX", "PWRSPVINDEA", "PWRSPVINDNE",
    "PWRSPVINDNO", "PWRSPVINDSO", "PWRSPVINDWE", "PWRSPVNPLXX",
    "PWRWONBGDXX", "PWRWONBTNXX", "PWRWONINDNO", "PWRWONINDSO",
    "PWRWONINDWE", "PWRWONMDVXX",
    # Pattern B
    "PWRHYDBGDXX", "PWRHYDBTNXX", "PWRHYDINDEA", "PWRHYDINDNE",
    "PWRHYDINDNO", "PWRHYDINDSO", "PWRHYDINDWE", "PWRHYDNPLXX",
    "PWRSHPINDEA", "PWRSHPINDNE", "PWRSHPINDNO", "PWRSHPINDSO",
    "PWRSHPINDWE",
}


# ---------------------------------------------------------------------------
# Implementation
# ---------------------------------------------------------------------------
def make_output_path(input_path: Path) -> Path:
    """Return the bounded POST_CAP_RESET sibling path. Does not create it."""
    return bounded_workspace_workbook_path(
        input_path.with_name(
            f"{input_path.stem}_POST_CAP_RESET{input_path.suffix}"
        ),
        stage_identity="POST_CAP_RESET",
    )


def patch_workbook(
    input_path: Path,
    dry_run: bool,
    output_path: Path | None = None,
) -> Path | None:
    if not input_path.is_file():
        sys.exit(f"ERROR: input file not found: {input_path}")

    print(f"Input: {input_path}")
    print(f"Sheet: {SHEET_NAME}")
    print(f"Parameter: {TARGET_PARAMETER}")
    print(f"Allowlisted techs: {len(TECHS_TO_PATCH)}")
    print(f"Year range to reset: {RESET_YEARS[0]}-{RESET_YEARS[-1]}")
    print(f"Dry run: {dry_run}")
    print()

    if not dry_run:
        output_path = (
            make_output_path(input_path)
            if output_path is None
            else bounded_workspace_workbook_path(
                output_path,
                stage_identity="POST_CAP_RESET",
            )
        )
        print(f"Output will be written to: {output_path.name}")
        print()

    wb = load_workbook(input_path)
    if SHEET_NAME not in wb.sheetnames:
        sys.exit(f"ERROR: sheet '{SHEET_NAME}' not found in workbook")
    ws = wb[SHEET_NAME]

    # Map header values to 1-indexed column numbers
    headers: dict = {}
    for cell in ws[1]:
        if cell.value is not None:
            headers[cell.value] = cell.column

    tech_col = headers.get(TECH_COL_NAME)
    param_col = headers.get(PARAM_COL_NAME)
    if tech_col is None or param_col is None:
        sys.exit(
            f"ERROR: missing required header. "
            f"tech_col='{TECH_COL_NAME}'={tech_col}, "
            f"param_col='{PARAM_COL_NAME}'={param_col}"
        )

    year_cols: dict = {}
    for yr in [2023, *RESET_YEARS]:
        col = headers.get(yr)
        if col is None:
            sys.exit(f"ERROR: year column {yr} not found in header row")
        year_cols[yr] = col

    # Walk the rows; collect changes; apply
    log = []  # (tech, prev_flat_value, n_cells_changed)
    skipped = []  # (tech, reason)
    seen_techs = set()

    for row in ws.iter_rows(min_row=2, values_only=False):
        tech = row[tech_col - 1].value
        param = row[param_col - 1].value
        if param != TARGET_PARAMETER:
            continue
        if tech not in TECHS_TO_PATCH:
            continue
        seen_techs.add(tech)

        val_2023 = row[year_cols[2023] - 1].value
        vals_post = [row[year_cols[y] - 1].value for y in RESET_YEARS]
        non_placeholder_post = {v for v in vals_post if v is not None and v != PLACEHOLDER}

        # Safety checks: signature is "2023=9999, 2024-2050 contains
        # exactly one distinct non-9999 value (flat block, possibly
        # interleaved with 9999 placeholders for Pattern B)".
        if val_2023 != PLACEHOLDER:
            skipped.append((tech, f"2023 != {PLACEHOLDER} (got {val_2023})"))
            continue
        if len(non_placeholder_post) == 0:
            skipped.append((tech, "2024-2050 already all 9999 -- nothing to reset"))
            continue
        if len(non_placeholder_post) != 1:
            skipped.append(
                (tech, f"2024-2050 has multiple distinct non-9999 values: "
                       f"{sorted(non_placeholder_post)}")
            )
            continue

        flat_val = next(iter(non_placeholder_post))
        n_changed = 0
        for yr in RESET_YEARS:
            cell = row[year_cols[yr] - 1]
            if cell.value != PLACEHOLDER:
                if not dry_run:
                    cell.value = PLACEHOLDER
                n_changed += 1
        log.append((tech, flat_val, n_changed))

    # Report
    print(f"Techs matched in sheet: {len(seen_techs)} / {len(TECHS_TO_PATCH)} allowlisted")
    not_seen = TECHS_TO_PATCH - seen_techs
    if not_seen:
        print(f"  Not found in sheet: {sorted(not_seen)}")
    print()

    if log:
        print(f"Reset summary ({len(log)} techs, {sum(n for _, _, n in log)} cells):")
        print(f"  {'Tech':14s} {'Was':>10s}  ->  {'Now':>10s}   (cells changed)")
        print(f"  {'-'*50}")
        for tech, prev, n in sorted(log):
            print(f"  {tech:14s} {prev:>10}  ->  {PLACEHOLDER:>10}   ({n})")
    else:
        print("No rows matched the cliff signature -- nothing to do.")

    if skipped:
        print()
        print(f"Skipped {len(skipped)} rows (signature mismatch -- safety abort for these):")
        for tech, reason in skipped:
            print(f"  {tech}: {reason}")

    if dry_run:
        print()
        print("DRY RUN -- no changes written.")
        return None

    assert output_path is not None
    wb.save(output_path)
    print()
    print(f"Saved: {output_path}")
    print(f"Input file untouched: {input_path}")
    return output_path


def main() -> None:
    p = argparse.ArgumentParser(description=__doc__.split("\n")[2])
    p.add_argument(
        "--input", type=Path, default=DEFAULT_INPUT,
        help=f"Path to the workbook to patch (default: {DEFAULT_INPUT})",
    )
    p.add_argument(
        "--dry-run", action="store_true",
        help="Print what would change but do not write",
    )
    p.add_argument(
        "--output", type=Path, default=None,
        help="Explicit mutable-workspace workbook output path",
    )
    args = p.parse_args()
    patch_workbook(args.input, args.dry_run, args.output)


if __name__ == "__main__":
    main()
