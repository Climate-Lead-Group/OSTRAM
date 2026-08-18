"""
fix_trn_residuals.py
====================

Maintains transmission (TRN) commissioning constraints in an OSeMOSYS
A-O Parametrization workbook.

In the production ``mode='min'`` route, the exact v18
``TotalAnnualMinCapacityInvestment`` contribution authority is additively
merged into the existing minimum rows. The complete v18 RC Authority V1 table
is then applied to final cross-border ``ResidualCapacity``.

An explicit ``--override`` retains the generic commissioning override path.
``mode='max'`` retains the generic input-profile cap behavior. Production
numeric authority is supplied only through the raw or materialized v18 workbook
passed with ``--authority``.

Outputs:
  - Corrected workbook (--output)
  - Diff log CSV (--diff-csv)
  - Diff log Markdown summary (--diff-md)

Usage:
  python fix_trn_residuals.py \\
      --input  A-O_Parametrization.xlsx \\
      --output A-O_Parametrization_FIXED.xlsx \\
      --diff-csv diff_log.csv \\
      --diff-md  diff_log.md
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import math
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Dict, List, Mapping, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from ostram.profiles import DEFAULT_PROFILE, active_profile_id
from .interconnector_authority import (
    MINIMUM_CONTRIBUTION_TECHS,
    load_minimum_boundary_authority,
    load_minimum_contribution_authority,
    validate_minimum_contribution_authority,
)

# ------------------------------------------------------------------ constants
SHEET_NAME = "Secondary Techs"
AUTHORITY_SHEET_NAME = "Interconnector_Params"
TECH_PREFIX = "TRN"
RESIDUAL_PARAM = "ResidualCapacity"
MIN_INV_PARAM = "TotalAnnualMinCapacityInvestment"
MAX_CAP_PARAM = "TotalAnnualMaxCapacity"
DEFAULT_CUTOFF_YEAR = 2023
FLAT_TOLERANCE = 1e-9
DEFAULT_PROJECTION_MODE = "User defined"

# columns in the metadata block (left of the year columns)
TECH_COL = "Tech"
PARAM_COL = "Parameter"
PROJ_MODE_COL = "Projection.Mode"

# RC Authority V1 is a dense 18-technology x 28-year table in the materialized
# v18 workbook. The digest commits the complete numeric table without making
# this script a second source of the values.
AUTHORITY_YEARS = tuple(range(2023, 2051))
AUTHORITY_TECHS = frozenset({
    "TRNBGDXXINDEA", "TRNBGDXXINDNE", "TRNBTNXXBGDXX",
    "TRNBTNXXINDEA", "TRNBTNXXINDNE", "TRNINDEAINDNE",
    "TRNINDEAINDNO", "TRNINDEAINDSO", "TRNINDEAINDWE",
    "TRNINDEANPLXX", "TRNINDNEINDNO", "TRNINDNOINDWE",
    "TRNINDNONPLXX", "TRNINDSOINDWE", "TRNINDSOLKAXX",
    "TRNLKAXXMDVXX", "TRNMDVXXINDSO", "TRNNPLXXBGDXX",
})
AUTHORITY_SEMANTIC_SHA256 = (
    "6c4420017b4a2df0b0e4ab6cc11f0bb45c79aa139bed858bdea5fec1aa54584b"
)


# ------------------------------------------------------------------ data model
@dataclass(frozen=True)
class Commissioning:
    """A single capacity addition event for one tech in one year."""
    tech: str
    year: int
    capacity_added: float


@dataclass
class TechFix:
    """Plan for fixing one TRN tech's residual capacity rows."""
    tech: str
    base_value: float                       # value at cutoff year (kept flat in RC)
    original_profile: Dict[int, float]      # year -> profile value (used to derive deltas)
    commissionings: List[Commissioning]     # only events with capacity_added > 0
    flatten_only: bool = True
    base_source: str = "input"
    profile_source: str = "input"           # which profile drove `commissionings`


@dataclass
class DiffEntry:
    """One cell-level change for the diff log."""
    sheet: str
    tech: str
    parameter: str
    year: int
    old_value: float
    new_value: float
    source: str = "split"   # "split" or final "authority_v1"


# ---------------------------------------------------------------- header parse
def _build_column_index(ws: Worksheet) -> Tuple[Dict[str, int], Dict[int, int]]:
    """
    Return (meta_cols, year_cols) where:
      meta_cols: {column_name -> 1-based column index} for textual headers
      year_cols: {year (int) -> 1-based column index}
    Reads the header from row 1.
    """
    meta_cols: Dict[str, int] = {}
    year_cols: Dict[int, int] = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        v = cell.value
        if isinstance(v, int):
            year_cols[v] = col_idx
        elif isinstance(v, str) and v.strip():
            meta_cols[v.strip()] = col_idx
    return meta_cols, year_cols


# ---------------------------------------------------------------- read profile
def _read_residual_profile(
    ws: Worksheet,
    meta_cols: Dict[str, int],
    year_cols: Dict[int, int],
) -> Dict[str, Tuple[int, Dict[int, float]]]:
    """
    Walk every row of the sheet. For each TRN tech with a ResidualCapacity row,
    return its (row_index, {year: value}) profile.
    """
    out: Dict[str, Tuple[int, Dict[int, float]]] = {}
    tech_c = meta_cols[TECH_COL]
    param_c = meta_cols[PARAM_COL]
    for row_idx in range(2, ws.max_row + 1):
        tech = ws.cell(row=row_idx, column=tech_c).value
        param = ws.cell(row=row_idx, column=param_c).value
        if not isinstance(tech, str) or not tech.startswith(TECH_PREFIX):
            continue
        if param != RESIDUAL_PARAM:
            continue
        profile = {}
        for y, c in year_cols.items():
            v = ws.cell(row=row_idx, column=c).value
            profile[y] = float(v) if v is not None else 0.0
        out[tech] = (row_idx, profile)
    return out


# ---------------------------------------------------------------- find param row
def _find_param_row(
    ws: Worksheet, meta_cols: Dict[str, int], tech: str, parameter: str
) -> Optional[int]:
    tech_c = meta_cols[TECH_COL]
    param_c = meta_cols[PARAM_COL]
    for row_idx in range(2, ws.max_row + 1):
        if (
            ws.cell(row=row_idx, column=tech_c).value == tech
            and ws.cell(row=row_idx, column=param_c).value == parameter
        ):
            return row_idx
    return None


def _decimal_text(value: object, context: str) -> str:
    """Return a canonical finite, non-negative decimal string."""
    if isinstance(value, bool) or value is None:
        raise ValueError(f"{context}: expected a numeric value, got {value!r}")
    try:
        decimal_value = Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(
            f"{context}: expected a numeric value, got {value!r}"
        ) from exc
    if not decimal_value.is_finite() or decimal_value < 0:
        raise ValueError(
            f"{context}: expected a finite non-negative value, got {value!r}"
        )
    if decimal_value == 0:
        return "0"
    return format(decimal_value.normalize(), "f")


def authority_semantic_sha256(
    authority: Mapping[str, Mapping[int, float]],
) -> str:
    """Hash the canonical technology/year/value authority domain."""
    payload = "".join(
        f"{tech}|{year}|"
        f"{_decimal_text(authority[tech][year], f'{tech}/{year}')}\n"
        for tech in sorted(authority)
        for year in sorted(authority[tech])
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _validate_authority_mapping(
    authority: Mapping[str, Mapping[int, float]],
) -> None:
    techs = set(authority)
    strict_full = active_profile_id() == DEFAULT_PROFILE
    if strict_full and techs != set(AUTHORITY_TECHS):
        missing = sorted(AUTHORITY_TECHS - techs)
        extra = sorted(techs - AUTHORITY_TECHS)
        raise ValueError(
            "RC Authority V1 technology domain mismatch: "
            f"missing={missing}, extra={extra}"
        )
    if not techs or any(not tech.startswith(TECH_PREFIX) for tech in techs):
        raise ValueError("RC authority requires a non-empty TRN technology domain")
    expected_years = set(AUTHORITY_YEARS)
    for tech in sorted(techs):
        years = set(authority[tech])
        if strict_full and years != expected_years:
            raise ValueError(
                f"RC Authority V1 year domain mismatch for {tech}: "
                f"missing={sorted(expected_years - years)}, "
                f"extra={sorted(years - expected_years)}"
            )
        if not years:
            raise ValueError(
                f"RC Authority V1 has no year data for {tech}"
            )
        for year in sorted(years):
            _decimal_text(authority[tech][year], f"{tech}/{year}")
    digest = authority_semantic_sha256(authority)
    if strict_full and digest != AUTHORITY_SEMANTIC_SHA256:
        raise ValueError(
            "RC Authority V1 semantic digest mismatch: "
            f"expected {AUTHORITY_SEMANTIC_SHA256}, got {digest}"
        )


def load_rc_authority(authority_path: Path) -> Dict[str, Dict[int, float]]:
    """Load and fully validate RC Authority V1 from materialized v18."""
    if not authority_path.is_file():
        raise FileNotFoundError(
            f"RC Authority V1 workbook not found: {authority_path}"
        )
    wb = load_workbook(
        authority_path, data_only=False, read_only=True, keep_links=False
    )
    try:
        if AUTHORITY_SHEET_NAME not in wb.sheetnames:
            raise ValueError(
                f"authority workbook missing sheet '{AUTHORITY_SHEET_NAME}'"
            )
        ws = wb[AUTHORITY_SHEET_NAME]
        meta_cols, year_cols = _build_column_index(ws)
        strict_full = active_profile_id() == DEFAULT_PROFILE
        # Core columns are always required; extended metadata columns are
        # only present (and validated) in the full/default profile workbook.
        core_meta = {TECH_COL, PARAM_COL, "Unit"}
        extended_meta = {"Parameter.ID", PROJ_MODE_COL, "Projection.Parameter"}
        required_meta = core_meta | extended_meta if strict_full else core_meta
        missing_meta = sorted(required_meta - set(meta_cols))
        if missing_meta:
            raise ValueError(
                f"RC Authority V1 missing metadata columns: {missing_meta}"
            )
        available_years = set(year_cols) & set(AUTHORITY_YEARS)
        if strict_full and set(year_cols) != set(AUTHORITY_YEARS):
            raise ValueError(
                "RC Authority V1 workbook year columns must be exactly "
                f"2023-2050; got {sorted(year_cols)}"
            )
        if not available_years:
            raise ValueError(
                "RC Authority V1 workbook has no recognised year columns"
            )

        authority: Dict[str, Dict[int, float]] = {}
        for row_idx, values in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            tech = values[meta_cols[TECH_COL] - 1]
            param = values[meta_cols[PARAM_COL] - 1]
            if not isinstance(tech, str) or not tech.startswith(TECH_PREFIX):
                continue
            if param != RESIDUAL_PARAM:
                continue
            if tech in authority:
                raise ValueError(
                    f"RC Authority V1 duplicate row for {tech} at row {row_idx}"
                )
            if "Parameter.ID" in meta_cols:
                if values[meta_cols["Parameter.ID"] - 1] != 3:
                    raise ValueError(
                        f"RC Authority V1 {tech}: Parameter.ID must be 3"
                    )
            if values[meta_cols["Unit"] - 1] != "GW":
                raise ValueError(f"RC Authority V1 {tech}: Unit must be GW")
            if PROJ_MODE_COL in meta_cols:
                if (
                    values[meta_cols[PROJ_MODE_COL] - 1]
                    != DEFAULT_PROJECTION_MODE
                ):
                    raise ValueError(
                        f"RC Authority V1 {tech}: Projection.Mode must be "
                        f"{DEFAULT_PROJECTION_MODE!r}"
                    )
            if "Projection.Parameter" in meta_cols:
                projection_parameter = values[
                    meta_cols["Projection.Parameter"] - 1
                ]
                if _decimal_text(
                    projection_parameter, f"{tech}/Projection.Parameter"
                ) != "0":
                    raise ValueError(
                        f"RC Authority V1 {tech}: Projection.Parameter must be 0"
                    )
            profile: Dict[int, float] = {}
            for year in available_years:
                raw = values[year_cols[year] - 1]
                canonical = _decimal_text(raw, f"{tech}/{year}")
                profile[year] = float(Decimal(canonical))
            authority[tech] = profile
    finally:
        wb.close()

    _validate_authority_mapping(authority)
    return authority


def apply_rc_authority(
    ws: Worksheet,
    authority: Mapping[str, Mapping[int, float]],
) -> List[DiffEntry]:
    """Write the validated v18 RC table into model Secondary Techs."""
    _validate_authority_mapping(authority)
    authority_techs = frozenset(authority)
    # Determine which years the authority actually covers (may be a subset
    # of AUTHORITY_YEARS for reduced-profile workbooks).
    authority_years = set()
    for tech_years in authority.values():
        authority_years |= set(tech_years)
    meta_cols, year_cols = _build_column_index(ws)
    if not authority_years.issubset(year_cols):
        missing = sorted(authority_years - set(year_cols))
        raise ValueError(f"model workbook missing authority years: {missing}")

    rows: Dict[str, int] = {}
    for row_idx in range(2, ws.max_row + 1):
        tech = ws.cell(row_idx, meta_cols[TECH_COL]).value
        param = ws.cell(row_idx, meta_cols[PARAM_COL]).value
        if tech not in authority_techs or param != RESIDUAL_PARAM:
            continue
        if tech in rows:
            raise ValueError(
                f"model workbook has duplicate {tech}/{RESIDUAL_PARAM} rows"
            )
        rows[tech] = row_idx
    if set(rows) != set(authority_techs):
        raise ValueError(
            "model workbook RC row domain mismatch: "
            f"missing={sorted(authority_techs - set(rows))}"
        )

    diffs: List[DiffEntry] = []
    for tech in sorted(authority_techs):
        row_idx = rows[tech]
        for year in sorted(authority[tech]):
            if year not in year_cols:
                continue
            cell = ws.cell(row_idx, year_cols[year])
            old = cell.value
            old_f = float(old) if old is not None else 0.0
            new_f = float(authority[tech][year])
            cell.value = new_f
            if not math.isclose(
                old_f, new_f, rel_tol=0.0, abs_tol=FLAT_TOLERANCE
            ):
                diffs.append(
                    DiffEntry(
                        SHEET_NAME, tech, RESIDUAL_PARAM, year,
                        old_f, new_f, source="authority_v1",
                    )
                )
    return diffs


# ---------------------------------------------------------------- planning
def derive_commissionings(
    profile: Dict[int, float], cutoff_year: int
) -> List[Tuple[int, float]]:
    """
    From a year->value cumulative profile, return list of (year, delta) for
    every year > cutoff_year where the value strictly increased over the
    previous year.  Negative or zero deltas are dropped (decommissionings are
    flagged separately by the caller).
    """
    years = sorted(profile.keys())
    out: List[Tuple[int, float]] = []
    for prev_y, y in zip(years, years[1:]):
        if y <= cutoff_year:
            continue
        delta = profile[y] - profile[prev_y]
        if delta > FLAT_TOLERANCE:
            out.append((y, delta))
    return out


def is_flat(profile: Dict[int, float]) -> bool:
    vals = list(profile.values())
    return (max(vals) - min(vals)) <= FLAT_TOLERANCE


def build_fix_plan(
    ws: Worksheet,
    cutoff_year: int,
    overrides: Optional[Dict[str, List[Commissioning]]] = None,
) -> Tuple[List[TechFix], List[str], List[str]]:
    """Build the generic input-profile plan used by override/max routes."""
    meta_cols, year_cols = _build_column_index(ws)
    if cutoff_year not in year_cols:
        raise ValueError(
            f"Cutoff year {cutoff_year} not found in sheet header; "
            f"available years: {sorted(year_cols.keys())}"
        )

    profiles = _read_residual_profile(ws, meta_cols, year_cols)
    plans: List[TechFix] = []
    skipped: List[str] = []
    warnings: List[str] = []

    for tech in sorted(profiles):
        _, profile = profiles[tech]
        if is_flat(profile):
            skipped.append(tech)
            continue

        if overrides and tech in overrides:
            commissionings = list(overrides[tech])
        else:
            commissionings = [
                Commissioning(tech=tech, year=year, capacity_added=addition)
                for year, addition in derive_commissionings(
                    profile, cutoff_year
                )
            ]

        years_sorted = sorted(profile)
        for previous_year, year in zip(years_sorted, years_sorted[1:]):
            if year <= cutoff_year:
                continue
            delta = profile[year] - profile[previous_year]
            if delta < -FLAT_TOLERANCE:
                warnings.append(
                    f"{tech}: negative delta {delta:+.3f} between "
                    f"{previous_year} and {year} in input profile "
                    "(decommissioning); not handled by this script"
                )

        plans.append(TechFix(
            tech=tech,
            base_value=profile[cutoff_year],
            original_profile=dict(profile),
            commissionings=commissionings,
            flatten_only=True,
            base_source="input",
            profile_source="input",
        ))

    return plans, skipped, warnings


def build_minimum_contribution_plan(
    ws: Worksheet,
    contribution_authority: Mapping[str, Mapping[int, float]],
    cutoff_year: int,
    overrides: Optional[Dict[str, List[Commissioning]]] = None,
) -> Tuple[List[TechFix], List[str], List[str]]:
    """Build the exact v18 minimum-contribution participation/write plan.

    The 11 authority rows participate even when their complete contribution is
    zero. An explicit override replaces the authority schedule for that
    technology. As in the predecessor route, overrides for skipped flat input
    technologies do not expand the minimum-investment write domain.
    """
    validate_minimum_contribution_authority(contribution_authority)
    meta_cols, year_cols = _build_column_index(ws)
    if cutoff_year not in year_cols:
        raise ValueError(
            f"Cutoff year {cutoff_year} not found in sheet header; "
            f"available years: {sorted(year_cols)}"
        )
    profiles = _read_residual_profile(ws, meta_cols, year_cols)
    missing_profiles = sorted(MINIMUM_CONTRIBUTION_TECHS - set(profiles))
    if missing_profiles:
        raise ValueError(
            "model workbook missing minimum-contribution technologies: "
            f"{missing_profiles}"
        )

    plan_techs = set(MINIMUM_CONTRIBUTION_TECHS)
    if overrides:
        plan_techs.update(
            tech
            for tech in overrides
            if tech in profiles and not is_flat(profiles[tech][1])
        )

    plans: List[TechFix] = []
    for tech in sorted(plan_techs):
        profile = profiles[tech][1]
        if overrides and tech in overrides:
            commissionings = list(overrides[tech])
            profile_source = "override"
        else:
            commissionings = [
                Commissioning(
                    tech=tech, year=year, capacity_added=float(capacity)
                )
                for year, capacity in sorted(
                    contribution_authority[tech].items()
                )
                if float(capacity) > FLAT_TOLERANCE
            ]
            profile_source = "v18 minimum contribution"
        plans.append(TechFix(
            tech=tech,
            base_value=profile[cutoff_year],
            original_profile=dict(profile),
            commissionings=commissionings,
            flatten_only=True,
            base_source="input",
            profile_source=profile_source,
        ))

    skipped = sorted(set(profiles) - plan_techs)
    return plans, skipped, []


# ---------------------------------------------------------------- apply
def apply_fix(
    ws: Worksheet,
    plan: TechFix,
    mode: str,
    cutoff_year: int,
    write_residual: bool = True,
) -> Tuple[List[DiffEntry], Dict[int, float]]:
    """
    Apply a single TechFix to the worksheet.

    Returns (diffs, preexisting_nonzero) where preexisting_nonzero is the
    {year -> value} of the target row's pre-existing nonzero entries (the
    values that were preserved/added-onto, not clobbered).
    """
    if mode not in ("min", "max"):
        raise ValueError(f"mode must be 'min' or 'max', got {mode!r}")

    meta_cols, year_cols = _build_column_index(ws)
    diffs: List[DiffEntry] = []

    # --- 1. flatten/level ResidualCapacity row ---------------------------
    rc_row = _find_param_row(ws, meta_cols, plan.tech, RESIDUAL_PARAM)
    if rc_row is None:
        raise RuntimeError(
            f"ResidualCapacity row for {plan.tech} not found"
        )
    if write_residual:
        rc_source = "magnitude" if not plan.flatten_only else "split"
        for y, c in year_cols.items():
            old = ws.cell(row=rc_row, column=c).value
            old_f = float(old) if old is not None else 0.0
            new_f = float(plan.base_value)
            if abs(old_f - new_f) > FLAT_TOLERANCE:
                ws.cell(row=rc_row, column=c).value = new_f
                diffs.append(
                    DiffEntry(SHEET_NAME, plan.tech, RESIDUAL_PARAM, y,
                              old_f, new_f, source=rc_source)
                )

    # If this is a magnitude-only correction (input was already flat,
    # we're just changing the level), there are no commissionings to write.
    if not plan.flatten_only:
        return diffs, {}

    # --- 2. populate the destination parameter -----------------------------
    target_param = MIN_INV_PARAM if mode == "min" else MAX_CAP_PARAM
    target_row = _find_param_row(ws, meta_cols, plan.tech, target_param)
    if target_row is None:
        raise RuntimeError(
            f"{target_param} row for {plan.tech} not found "
            f"(expected to exist in template)"
        )

    # Read existing values from target row first, so we can preserve any
    # hand-curated populations rather than silently clobbering them.
    existing_target: Dict[int, float] = {}
    for y, c in year_cols.items():
        v = ws.cell(row=target_row, column=c).value
        existing_target[y] = float(v) if v is not None else 0.0

    if mode == "min":
        # ADDITIVE merge: existing TAMCI entries (hand-curated commissionings)
        # are preserved; script-derived deltas are added on top of whatever is
        # already there.  Both represent commissioning events: if a user
        # pre-specified one in year y AND the residual delta also implies one
        # in year y, both are real.
        target_values: Dict[int, float] = dict(existing_target)
        for c in plan.commissionings:
            target_values[c.year] = target_values.get(c.year, 0.0) + c.capacity_added
    else:  # mode == "max"
        # Cumulative cap = original residual profile.  Where the user already
        # has a (tighter) cap, preserve it; otherwise fill from the profile.
        target_values = {}
        for y in year_cols:
            ex = existing_target.get(y, 0.0)
            target_values[y] = ex if abs(ex) > FLAT_TOLERANCE \
                else float(plan.original_profile[y])

    # set Projection.Mode on the target row to a non-empty value if missing
    if PROJ_MODE_COL in meta_cols:
        pm_c = meta_cols[PROJ_MODE_COL]
        existing = ws.cell(row=target_row, column=pm_c).value
        if existing in (None, "", "EMPTY"):
            ws.cell(row=target_row, column=pm_c).value = DEFAULT_PROJECTION_MODE

    for y, c in year_cols.items():
        old = ws.cell(row=target_row, column=c).value
        old_f = float(old) if old is not None else 0.0
        new_f = float(target_values.get(y, 0.0))
        # Always write an explicit numeric value to the target row so empty
        # cells become 0.0 (rather than staying None, which downstream tools
        # may read as NaN). Only log a diff when the value actually changed.
        ws.cell(row=target_row, column=c).value = new_f
        if abs(old_f - new_f) > FLAT_TOLERANCE:
            diffs.append(
                DiffEntry(SHEET_NAME, plan.tech, target_param, y, old_f, new_f)
            )

    preexisting_nonzero = {y: v for y, v in existing_target.items()
                           if abs(v) > FLAT_TOLERANCE}
    return diffs, preexisting_nonzero


# ---------------------------------------------------------------- override I/O
def load_overrides(path: Path) -> Dict[str, List[Commissioning]]:
    """Load a commissioning override CSV (columns: tech, year, capacity_added)."""
    out: Dict[str, List[Commissioning]] = {}
    with path.open(newline="") as f:
        reader = csv.DictReader(f)
        required = {"tech", "year", "capacity_added"}
        if not required.issubset(set(c.strip().lower() for c in reader.fieldnames or [])):
            raise ValueError(
                f"override CSV must have columns {sorted(required)}; "
                f"got {reader.fieldnames}"
            )
        for row in reader:
            t = row["tech"].strip()
            y = int(row["year"])
            cap = float(row["capacity_added"])
            if cap <= 0:
                continue
            out.setdefault(t, []).append(
                Commissioning(tech=t, year=y, capacity_added=cap)
            )
    return out


# ---------------------------------------------------------------- diff writer
def write_diff_csv(diffs: List[DiffEntry], path: Path) -> None:
    with path.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["sheet", "tech", "parameter", "year",
                    "old_value", "new_value", "delta", "source"])
        for d in diffs:
            w.writerow(
                [d.sheet, d.tech, d.parameter, d.year,
                 f"{d.old_value:.6f}", f"{d.new_value:.6f}",
                 f"{d.new_value - d.old_value:+.6f}", d.source]
            )


def write_diff_md(
    diffs: List[DiffEntry],
    plans: List[TechFix],
    skipped: List[str],
    warnings: List[str],
    mode: str,
    cutoff_year: int,
    path: Path,
) -> None:
    split_diffs = [d for d in diffs if d.source == "split"]
    authority_diffs = [d for d in diffs if d.source == "authority_v1"]

    flatten_plans = [p for p in plans if p.flatten_only]

    lines: List[str] = []
    lines.append("# TRN Interconnector Authority — Diff Log")
    lines.append("")
    lines.append(f"- Mode: **{mode}** "
                 f"(`{MIN_INV_PARAM}` if min, `{MAX_CAP_PARAM}` if max)")
    lines.append(f"- Cutoff year: **{cutoff_year}**")
    lines.append(f"- Techs written to "
                 f"`{MIN_INV_PARAM if mode == 'min' else MAX_CAP_PARAM}`: "
                 f"**{len(flatten_plans)}**")
    lines.append(f"- Techs skipped (no change needed): **{len(skipped)}**")
    lines.append(
        f"- Cell changes from commissioning schedules: **{len(split_diffs)}**"
    )
    lines.append(
        f"- Cell changes from RC Authority V1: **{len(authority_diffs)}**"
    )
    if warnings:
        lines.append("")
        lines.append("## Warnings")
        for w in warnings:
            lines.append(f"- {w}")
    lines.append("")
    lines.append("## Skipped (outside the selected write plan)")
    for t in skipped:
        lines.append(f"- `{t}`")
    lines.append("")
    lines.append("## Per-tech commissioning schedules")
    for plan in flatten_plans:
        lines.append(f"\n### `{plan.tech}`")
        lines.append(f"- Schedule source: `{plan.profile_source}`")
        if plan.commissionings:
            lines.append(f"- Post-{cutoff_year} commissionings "
                         f"(moved to `{MIN_INV_PARAM if mode == 'min' else MAX_CAP_PARAM}`):")
            for c in plan.commissionings:
                lines.append(f"  - {c.year}: +{c.capacity_added:.3f}")
        else:
            lines.append("- (no post-cutoff commissionings derived)")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


# ---------------------------------------------------------------- pipeline
def run_fix(
    input_path: Path,
    output_path: Path,
    mode: str = "min",
    cutoff_year: int = DEFAULT_CUTOFF_YEAR,
    override_path: Optional[Path] = None,
    authority_path: Optional[Path] = None,
) -> Tuple[List[DiffEntry], List[TechFix], List[str], List[str]]:
    """Validate v18 authority, apply the selected plan, and save."""
    if mode not in ("min", "max"):
        raise ValueError(f"mode must be 'min' or 'max', got {mode!r}")
    if authority_path is None:
        raise ValueError("v18 interconnector authority workbook is required")

    # All production authority families fail closed before the target workbook
    # is opened or can be mutated.
    authority = load_rc_authority(authority_path)
    strict_full = active_profile_id() == DEFAULT_PROFILE
    if strict_full:
        contribution_authority = load_minimum_contribution_authority(
            authority_path
        )
        load_minimum_boundary_authority(authority_path)
    else:
        # Reduced profile workbooks retain their own RC rows but intentionally
        # omit the full model's 11-row additive contribution family and
        # two-row LinkFreeze boundary family.
        contribution_authority = {}

    wb = load_workbook(input_path)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"sheet '{SHEET_NAME}' not in workbook; sheets: {wb.sheetnames}"
        )
    ws = wb[SHEET_NAME]

    diffs: List[DiffEntry] = []
    warnings: List[str] = []

    overrides = load_overrides(override_path) if override_path else None
    if mode == "min" and contribution_authority:
        plans, skipped, plan_warnings = build_minimum_contribution_plan(
            ws, contribution_authority, cutoff_year, overrides
        )
    elif mode == "min":
        plans, skipped, plan_warnings = [], sorted(authority), []
    else:
        plans, skipped, plan_warnings = build_fix_plan(
            ws, cutoff_year, overrides
        )
    warnings.extend(plan_warnings)

    target_param = MIN_INV_PARAM if mode == "min" else MAX_CAP_PARAM
    for plan in plans:
        d, preexisting = apply_fix(
            ws, plan, mode=mode, cutoff_year=cutoff_year,
            write_residual=False,
        )
        diffs.extend(d)
        if preexisting:
            entries = ", ".join(f"{y}: {v:+.3f}" for y, v in sorted(preexisting.items()))
            warnings.append(
                f"{plan.tech}: pre-existing nonzero {target_param} values "
                f"preserved/merged: {{{entries}}}"
            )

    diffs.extend(apply_rc_authority(ws, authority))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return diffs, plans, skipped, warnings


# ---------------------------------------------------------------- CLI
def main(argv: Optional[List[str]] = None) -> int:
    p = argparse.ArgumentParser(description=__doc__.split("\n")[1])
    p.add_argument("--input", type=Path, required=True,
                   help="path to input A-O_Parametrization.xlsx")
    p.add_argument("--output", type=Path, required=True,
                   help="path for corrected workbook")
    p.add_argument("--diff-csv", type=Path, default=None,
                   help="path for cell-level diff log CSV")
    p.add_argument("--diff-md", type=Path, default=None,
                   help="path for human-readable diff log Markdown")
    p.add_argument("--mode", choices=("min", "max"), default="min",
                   help="'min' = exogenous timing via "
                        f"{MIN_INV_PARAM} (default); "
                        f"'max' = endogenous timing capped by {MAX_CAP_PARAM}")
    p.add_argument("--cutoff-year", type=int, default=DEFAULT_CUTOFF_YEAR,
                   help=f"split year (default {DEFAULT_CUTOFF_YEAR}); "
                        "values at this year stay in ResidualCapacity, "
                        "additions after move to the destination parameter")
    p.add_argument("--override", type=Path, default=None,
                   help="optional CSV with columns tech,year,capacity_added "
                        "to override the selected commissioning schedule")
    p.add_argument("--authority", type=Path, required=True,
                   help="raw or materialized v18 workbook containing the "
                        "digest-validated interconnector authorities")
    args = p.parse_args(argv)

    diffs, plans, skipped, warnings = run_fix(
        input_path=args.input,
        output_path=args.output,
        mode=args.mode,
        cutoff_year=args.cutoff_year,
        override_path=args.override,
        authority_path=args.authority,
    )

    if args.diff_csv:
        write_diff_csv(diffs, args.diff_csv)
    if args.diff_md:
        write_diff_md(
            diffs, plans, skipped, warnings,
            mode=args.mode, cutoff_year=args.cutoff_year, path=args.diff_md
        )

    print(f"Fixed {len(plans)} techs, skipped {len(skipped)} flat techs, "
          f"wrote {len(diffs)} cell changes -> {args.output}")
    if warnings:
        print(f"  ({len(warnings)} warning(s); see diff log)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
