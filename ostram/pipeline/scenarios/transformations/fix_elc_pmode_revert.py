"""fix_elc_pmode_revert.py
=============================
Reproduces the manual edit that Luis made between run 2 of
add_max_capacity_investment_rule.py (10:55) and B1b_Pre_solver_validation.py
(12:01) on 2026-04-30.

WHAT IT DOES
------------
In the 'Secondary Techs' sheet, reverts the 'Projection.Mode' column from
"User defined" to "EMPTY" for each ELC*01 dispatch node derived from the
active profile's ``countries`` list, across two parameters:
  Parameters (2):
    TotalAnnualMaxCapacity, TotalAnnualMaxCapacityInvestment

WHY
---
The second run of add_max_capacity_investment_rule.py (commit 2be1616)
flips Projection.Mode EMPTY -> User defined for all techs ZEROED
in MaxCap/MaxCapInv. The ELC*01 techs are ZEROED (lockout: the model must not
invest in a "placeholder electricity tech", only use the PWR* ones).
Luis decided that 'User defined' was misleading for those lockout rows
(the year cells are 0, with no real configurable value) and reverted them to
"EMPTY" by hand. The rows are functionally disabled either way.

Idempotent: re-running is a no-op (they are already EMPTY).

Usage:
    python fix_elc_pmode_revert.py --input <A-O_Parametrization.xlsx>
    python fix_elc_pmode_revert.py --input <in> --output <out>  # write a copy
"""
from __future__ import annotations
import argparse
from pathlib import Path
import re
import sys
from openpyxl import load_workbook
import yaml

from ostram.paths import resolve_paths

TARGET_SHEET = "Secondary Techs"
TECH_COL_NAME = "Tech"
PARAM_COL_NAME = "Parameter"
PMODE_COL_NAME = "Projection.Mode"

OLD_VALUE = "User defined"
NEW_VALUE = "EMPTY"

TARGET_PARAMS = {
    "TotalAnnualMaxCapacity", "TotalAnnualMaxCapacityInvestment",
}


_COUNTRY_REGION = re.compile(r"^[A-Z]{3}(?:[A-Z]{2})?$")


def country_region_map(countries: object) -> dict[str, str]:
    """Map model regions to their timeslice workbook sheet prefixes."""

    if not isinstance(countries, list) or not countries:
        raise ValueError("country configuration requires a non-empty countries list")
    result: dict[str, str] = {}
    for value in countries:
        if not isinstance(value, str) or not _COUNTRY_REGION.fullmatch(value):
            raise ValueError(f"invalid country region in countries list: {value!r}")
        region = value + "XX" if len(value) == 3 else value
        result[region] = value
    return result


def elc_dispatch_techs(countries: object) -> frozenset[str]:
    """Return ELC*01 nodes from three- or five-character country regions."""

    return frozenset(f"ELC{region}01" for region in country_region_map(countries))


def configured_country_region_map(config_path: Path | None = None) -> dict[str, str]:
    path = resolve_paths().country_config if config_path is None else Path(config_path)
    raw = yaml.safe_load(path.read_text(encoding="utf-8"))
    if not isinstance(raw, dict):
        raise ValueError(f"country configuration must be a mapping: {path}")
    return country_region_map(raw.get("countries"))


def configured_elc_dispatch_techs(config_path: Path | None = None) -> frozenset[str]:
    return frozenset(
        f"ELC{region}01"
        for region in configured_country_region_map(config_path)
    )


def revert_pmode(
    input_path: Path,
    output_path: Path | None = None,
    *,
    target_techs: frozenset[str] | None = None,
) -> int:
    """Revert Projection.Mode for ELC*01 lockout rows. Returns cells changed."""
    if output_path is None:
        output_path = input_path
    wb = load_workbook(input_path)
    if TARGET_SHEET not in wb.sheetnames:
        sys.exit(f"ERROR: sheet '{TARGET_SHEET}' not in workbook")
    ws = wb[TARGET_SHEET]

    headers = {c.value: c.column for c in ws[1] if c.value is not None}
    for col in (TECH_COL_NAME, PARAM_COL_NAME, PMODE_COL_NAME):
        if col not in headers:
            sys.exit(f"ERROR: required column '{col}' missing in '{TARGET_SHEET}'")
    tech_col = headers[TECH_COL_NAME]
    param_col = headers[PARAM_COL_NAME]
    pmode_col = headers[PMODE_COL_NAME]
    active_target_techs = (
        configured_elc_dispatch_techs()
        if target_techs is None
        else target_techs
    )

    cells_changed = 0
    rows_touched: list[tuple[int, str, str, object, object]] = []
    rows_skipped_already_empty: list[tuple[int, str, str]] = []
    for r in range(2, ws.max_row + 1):
        tech = ws.cell(r, tech_col).value
        if tech not in active_target_techs:
            continue
        param = ws.cell(r, param_col).value
        if param not in TARGET_PARAMS:
            continue
        cell = ws.cell(r, pmode_col)
        old = cell.value
        if old == NEW_VALUE:
            rows_skipped_already_empty.append((r, str(tech), str(param)))
            continue
        cell.value = NEW_VALUE
        rows_touched.append((r, str(tech), str(param), old, NEW_VALUE))
        cells_changed += 1

    wb.save(output_path)

    print(f"Reverted {cells_changed} cells (Projection.Mode -> '{NEW_VALUE}')")
    if rows_touched:
        for r, t, p, old, new in rows_touched:
            print(f"  row {r:>5}  {t:<14}  {p:<35}  {old!r} -> {new!r}")
    if rows_skipped_already_empty:
        print(f"Skipped {len(rows_skipped_already_empty)} cells already at '{NEW_VALUE}' "
              f"(idempotent re-run)")
    print(f"Saved: {output_path}")
    return cells_changed


if __name__ == "__main__":
    p = argparse.ArgumentParser(description=__doc__.split("\n")[1])
    p.add_argument("--input", type=Path, required=True)
    p.add_argument("--output", type=Path, default=None)
    args = p.parse_args()
    if not args.input.is_file():
        sys.exit(f"ERROR: input not found: {args.input}")
    revert_pmode(args.input, args.output)
