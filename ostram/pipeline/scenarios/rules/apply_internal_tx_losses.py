"""
apply_internal_tx_losses.py
===========================

WS-4. Give the six INTERNAL (intra-node) transmission families a non-zero
transmission loss, matching how the interconnectors carry losses (via the
output activity ratio in the A-O_AR files).

Background
----------
A2_AddTx injects the internal families (RNWTRN/RNWNLI/RNWRPO, PWRTRN/TRNNLI/
TRNRPO) with output activity = 1.0 -> 0% loss, which is not physical. The
interconnectors already carry per-corridor losses (OAR 0.93-0.98) in
A-O_AR_Projections "Demand Techs" (year-indexed Direction=Output rows) and
A-O_AR_Model_Base_Year "Demand Techs" (Value.Fuel.O). This stage sets the
internal families' output activity to (1 - loss), so B1 compiles
OutputActivityRatio = 1 - loss for them.

Loss value: `internal_transmission.transmission_loss` in Config_country_codes.yaml
(default 0.03 = 3%, CEA all-India *transmission* loss ~3-4%; distribution is out
of scope). Uniform across nodes (per the WS-3 uniform-cost decision). OAR = 0.97.

Scope: the 6 internal families only, on the two AR files' "Demand Techs" sheet.
Interconnectors, DSPTRN, generators, storage untouched.

OUTPUT: timestamped backup, in-place edit of the two AR files, JSON change log.

USAGE
    python apply_internal_tx_losses.py --input-dir <stage5 or A1_Outputs/...>
    python apply_internal_tx_losses.py --self-test
    python apply_internal_tx_losses.py --input-dir <dir> --restore
"""
from __future__ import annotations

import argparse
import json
import os
import shutil
import sys
from datetime import datetime
from pathlib import Path

import yaml
from openpyxl import load_workbook, Workbook

PROJ_FILE = "A-O_AR_Projections.xlsx"
BASE_FILE = "A-O_AR_Model_Base_Year.xlsx"
SHEET = "Demand Techs"
CONFIG_NAME = "Config_country_codes.yaml"

INTERNAL_FAMILIES = ("RNWTRN", "RNWNLI", "RNWRPO", "PWRTRN", "TRNNLI", "TRNRPO")
DEFAULT_LOSS = 0.03
BACKUP_TAG = "_PRE_INTERNAL_LOSS_"

COL_TECH = "Tech"
COL_DIRECTION = "Direction"
COL_VALUE_O = "Value.Fuel.O"


def _is_internal(t) -> bool:
    return isinstance(t, str) and len(t) == 11 and t[:6] in INTERNAL_FAMILIES


def _values_differ(a, b, tol=1e-9) -> bool:
    if a is None or b is None:
        return a is not b
    try:
        return abs(float(a) - float(b)) > tol
    except (TypeError, ValueError):
        return a != b


# --------------------------------------------------------------------------- config
def resolve_config(explicit):
    cands = []
    if explicit:
        cands.append(Path(explicit))
    env = os.environ.get("OSTRAM_CONFIG_PATH")
    if env:
        cands.append(Path(env))
    cands.append(Path(__file__).resolve().parent.parent.parent / CONFIG_NAME)
    for c in cands:
        if c and c.is_file():
            return c
    raise FileNotFoundError("Config_country_codes.yaml not found; tried " + ", ".join(map(str, cands)))


def read_loss(config_path) -> float:
    with open(config_path, "r", encoding="utf-8") as fh:
        data = yaml.safe_load(fh) or {}
    blk = data.get("internal_transmission", {}) or {}
    loss = blk.get("transmission_loss", DEFAULT_LOSS)
    loss = float(loss)
    if not (0.0 <= loss < 1.0):
        raise ValueError(f"transmission_loss must be in [0,1); got {loss}")
    return loss


# --------------------------------------------------------------------------- header helper
def _hdr_map(ws):
    return {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)
            if ws.cell(row=1, column=c).value is not None}


def _year_cols(ws):
    out = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if isinstance(v, int) and 1900 <= v <= 2200:
            out[v] = c
        elif isinstance(v, str) and v.strip().split(".")[0].isdigit() and 1900 <= int(float(v)) <= 2200:
            out[int(float(v))] = c
    return out


# --------------------------------------------------------------------------- editors
def edit_projections(path: Path, oar: float) -> dict:
    """A-O_AR_Projections 'Demand Techs': set year cells of internal Output rows to oar."""
    wb = load_workbook(path)
    log = {"file": str(path), "sheet": SHEET, "cells": 0, "rows": 0}
    try:
        if SHEET not in wb.sheetnames:
            log["skipped"] = "sheet absent"; return log
        ws = wb[SHEET]
        h = _hdr_map(ws)
        tc, dc = h.get(COL_TECH), h.get(COL_DIRECTION)
        ycols = _year_cols(ws)
        if tc is None or dc is None or not ycols:
            log["skipped"] = f"missing Tech/Direction/year cols (tc={tc} dc={dc} yrs={len(ycols)})"; return log
        for r in range(2, ws.max_row + 1):
            if not _is_internal(ws.cell(row=r, column=tc).value):
                continue
            if str(ws.cell(row=r, column=dc).value) != "Output":
                continue
            touched = False
            for y, c in ycols.items():
                cell = ws.cell(row=r, column=c)
                if _values_differ(cell.value, oar):
                    cell.value = oar
                    log["cells"] += 1
                    touched = True
            if touched:
                log["rows"] += 1
        wb.save(path)
    finally:
        wb.close()
    return log


def edit_base_year(path: Path, oar: float) -> dict:
    """A-O_AR_Model_Base_Year 'Demand Techs': set Value.Fuel.O of internal rows to oar."""
    wb = load_workbook(path)
    log = {"file": str(path), "sheet": SHEET, "cells": 0, "rows": 0}
    try:
        if SHEET not in wb.sheetnames:
            log["skipped"] = "sheet absent"; return log
        ws = wb[SHEET]
        h = _hdr_map(ws)
        tc, vo = h.get(COL_TECH), h.get(COL_VALUE_O)
        if tc is None or vo is None:
            log["skipped"] = f"missing Tech/Value.Fuel.O cols (tc={tc} vo={vo})"; return log
        for r in range(2, ws.max_row + 1):
            if not _is_internal(ws.cell(row=r, column=tc).value):
                continue
            cell = ws.cell(row=r, column=vo)
            if cell.value is not None and _values_differ(cell.value, oar):
                cell.value = oar
                log["cells"] += 1
                log["rows"] += 1
        wb.save(path)
    finally:
        wb.close()
    return log


# --------------------------------------------------------------------------- backup / restore
def make_backup(d: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    b = d.parent / f"{d.name}{BACKUP_TAG}{stamp}"
    if b.exists():
        raise FileExistsError(b)
    shutil.copytree(d, b)
    return b


def restore(d: Path, frm=None) -> Path:
    d = Path(d)
    if frm is None:
        cands = sorted(p for p in d.parent.iterdir()
                       if p.is_dir() and p.name.startswith(f"{d.name}{BACKUP_TAG}"))
        if not cands:
            raise FileNotFoundError(f"no {BACKUP_TAG}* backup by {d}")
        frm = cands[-1]
    if d.is_dir():
        shutil.rmtree(d)
    shutil.copytree(frm, d)
    return frm


# --------------------------------------------------------------------------- orchestration
def run(input_dir, skip_backup=False, config=None) -> dict:
    input_dir = Path(input_dir)
    loss = read_loss(resolve_config(config))
    oar = round(1.0 - loss, 6)
    proj, base = input_dir / PROJ_FILE, input_dir / BASE_FILE
    if not proj.exists():
        raise FileNotFoundError(proj)
    backup = None if skip_backup else make_backup(input_dir)
    log = {"loss": loss, "oar": oar, "families": list(INTERNAL_FAMILIES),
           "timestamp": datetime.now().isoformat(),
           "backup_dir": str(backup) if backup else None,
           "projections": edit_projections(proj, oar),
           "base_year": edit_base_year(base, oar) if base.exists() else {"skipped": "absent"}}
    if backup is not None:
        p = backup.parent / f"{backup.name}_CHANGES.json"
        p.write_text(json.dumps(log, indent=2, default=str))
        log["log_path"] = str(p)
    return log


def print_summary(log):
    print("=" * 68)
    print(f"apply_internal_tx_losses  loss={log['loss']:.3f}  OAR={log['oar']}")
    print("=" * 68)
    for k in ("projections", "base_year"):
        s = log[k]
        if "skipped" in s:
            print(f"  {k}: SKIPPED ({s['skipped']})")
        else:
            print(f"  {k}: {s['cells']} cells / {s['rows']} rows -> {log['oar']}")
    print(f"  backup: {log.get('backup_dir', '(skipped)')}")


# --------------------------------------------------------------------------- self-test
def run_self_test() -> int:
    import tempfile
    print("=" * 68); print("apply_internal_tx_losses.py SELF-TEST"); print("=" * 68)
    ok = True
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        (td / CONFIG_NAME).write_text(
            yaml.safe_dump({"internal_transmission": {"transmission_loss": 0.03}}), encoding="utf-8")
        idir = td / "A1_Outputs" / "A1_Outputs_BAU"; idir.mkdir(parents=True)
        # Projections: Demand Techs with internal Output rows (=1) + an interconnector (must NOT change)
        pw = Workbook(); ws = pw.active; ws.title = SHEET
        for c, hh in enumerate(["Mode.Operation", "Tech", "Fuel", "Direction", 2023, 2024, 2025], 1):
            ws.cell(row=1, column=c, value=hh)
        rows = [
            (1, "RNWTRNINDWE", "ELCINDWE02", "Output", 1, 1, 1),   # -> 0.97
            (1, "RNWTRNINDWE", "ELCINDWE00", "Input", 1, 1, 1),    # input stays 1
            (1, "PWRTRNMDVXX", "ELCMDVXX02", "Output", 1, 1, 1),   # -> 0.97
            (1, "TRNBGDXXINDEA", "ELCINDEA04", "Output", 0.983, 0.983, 0.983),  # interconnector, must NOT change
            (1, "PWRSPVINDWE", "ELCINDWE00", "Output", 1, 1, 1),   # non-tx, must NOT change
        ]
        for ri, row in enumerate(rows, 2):
            for c, v in enumerate(row, 1):
                ws.cell(row=ri, column=c, value=v)
        pw.save(idir / PROJ_FILE); pw.close()
        # Base year: Demand Techs with Value.Fuel.O
        bw = Workbook(); bs = bw.active; bs.title = SHEET
        for c, hh in enumerate(["Tech", "Fuel.O", COL_VALUE_O], 1):
            bs.cell(row=1, column=c, value=hh)
        brows = [("RNWTRNINDWE", "ELCINDWE02", 1), ("PWRTRNMDVXX", "ELCMDVXX02", 1),
                 ("PWRSPVINDWE", "ELCINDWE00", 1)]  # last is non-tx, must NOT change
        for ri, row in enumerate(brows, 2):
            for c, v in enumerate(row, 1):
                bs.cell(row=ri, column=c, value=v)
        bw.save(idir / BASE_FILE); bw.close()

        run(idir, skip_backup=True, config=td / CONFIG_NAME)

        pv = load_workbook(idir / PROJ_FILE, data_only=True).active
        got = {}
        for r in range(2, pv.max_row + 1):
            got[(pv.cell(row=r, column=2).value, pv.cell(row=r, column=4).value)] = pv.cell(row=r, column=5).value
        bv = load_workbook(idir / BASE_FILE, data_only=True).active
        gotb = {bv.cell(row=r, column=1).value: bv.cell(row=r, column=3).value for r in range(2, bv.max_row + 1)}

        def ck(cond, label):
            nonlocal ok; print(("  PASS " if cond else "  FAIL ") + label); ok = ok and cond
        ck(got.get(("RNWTRNINDWE", "Output")) == 0.97, "internal RNWTRN Output 1 -> 0.97")
        ck(got.get(("PWRTRNMDVXX", "Output")) == 0.97, "internal PWRTRN Output 1 -> 0.97")
        ck(got.get(("RNWTRNINDWE", "Input")) == 1, "internal Input stays 1")
        ck(got.get(("TRNBGDXXINDEA", "Output")) == 0.983, "interconnector UNCHANGED")
        ck(got.get(("PWRSPVINDWE", "Output")) == 1, "non-tx UNCHANGED")
        ck(gotb.get("RNWTRNINDWE") == 0.97, "base-year internal Value.Fuel.O -> 0.97")
        ck(gotb.get("PWRSPVINDWE") == 1, "base-year non-tx UNCHANGED")
        log2 = run(idir, skip_backup=True, config=td / CONFIG_NAME)
        ck(log2["projections"]["cells"] == 0 and log2["base_year"]["cells"] == 0, "idempotent (2nd run 0 cells)")
    print("=" * 68); print("SELF-TEST", "PASSED" if ok else "FAILED"); print("=" * 68)
    return 0 if ok else 1


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--input-dir", type=Path, default=Path("A1_Outputs/A1_Outputs_BAU"))
    p.add_argument("--config", type=Path, default=None)
    p.add_argument("--skip-backup", action="store_true")
    p.add_argument("--self-test", action="store_true")
    p.add_argument("--restore", action="store_true")
    p.add_argument("--restore-from", type=Path, default=None)
    a = p.parse_args()
    if a.self_test:
        return run_self_test()
    if a.restore or a.restore_from is not None:
        try:
            used = restore(a.input_dir, a.restore_from)
        except Exception as e:
            print(f"ERROR: {e}", file=sys.stderr); return 1
        print(f"Restored {a.input_dir} from {used}"); return 0
    try:
        log = run(a.input_dir, a.skip_backup, a.config)
    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr); return 1
    print_summary(log)
    return 0


if __name__ == "__main__":
    sys.exit(main())
