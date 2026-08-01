"""
set_interconnector_direction.py
===============================

Controls interconnector directionality by zeroing one flow mode's
InputActivityRatio and OutputActivityRatio in the AR source files.

Each TRN interconnector has two modes of operation:
  - Forward (mode whose input fuel belongs to the source region)
  - Reverse (mode whose input fuel belongs to the destination region)

The direction is relative to the tech name TRN<SRC><DST>:
  forward       = keep SRC -> DST, disable reverse mode
  reverse       = keep DST -> SRC, disable forward mode
  bidirectional = no-op (both modes active, the default)

This script is the complement of relax_interconnectors.py.  Where relax
controls *how much* capacity a link can expand to, this script controls
*which way* power can flow.  They compose cleanly: you can cap a link at
2.5 GW AND force it one-way in the same scenario.

Runs AFTER relax_interconnectors.py in the rules chain.

CONFIGURATION
-------------
Edit set_interconnector_direction.yaml (per-scenario configs in
rules_scripts/configs/<scenario>/ take precedence).

OUTPUT
------
1. Timestamped backup of the input directory.
2. In-place edit of A-O_AR_Projections.xlsx (zeros disabled mode years).
3. In-place edit of A-O_AR_Model_Base_Year.xlsx (zeros disabled mode values).
4. A JSON change log next to the backup.

USAGE
-----
    python set_interconnector_direction.py --input-dir A1_Outputs/A1_Outputs_BAU
    python set_interconnector_direction.py --restore
"""

from __future__ import annotations

import argparse
import gc
import json
import shutil
import sys
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

try:
    import yaml as _yaml
    def _load_yaml(path: Path) -> dict:
        with open(path, "r", encoding="utf-8") as f:
            return _yaml.safe_load(f)
except ImportError:
    _yaml = None
    def _load_yaml(path: Path) -> dict:
        raise ImportError("PyYAML is required. Install with: pip install pyyaml")


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
YAML_FILE_NAME = "set_interconnector_direction.yaml"

AR_PROJ_FILE = "A-O_AR_Projections.xlsx"
AR_BASE_FILE = "A-O_AR_Model_Base_Year.xlsx"
AR_SHEET = "Secondary"

COL_MODE_OP = "Mode.Operation"
COL_TECH = "Tech"
COL_FUEL = "Fuel"
COL_DIRECTION = "Direction"
COL_PMODE = "Projection.Mode"
PMODE_USER = "User defined"

COL_BASE_FUEL_I = "Fuel.I"
COL_BASE_VALUE_I = "Value.Fuel.I"
COL_BASE_FUEL_O = "Fuel.O"
COL_BASE_VALUE_O = "Value.Fuel.O"

TECH_TYPES_FILE = "TECH_TYPES.csv"
TECH_TYPES_CATEGORY_COL = "Technology (PWR)"
TECH_TYPES_TECH_COL = "Technology"
INTERCONNECTOR_CATEGORY = "INTERCONNECTORS"

TRN_PREFIX = "TRN"
REGION_LEN = 5

VALID_DIRECTIONS = {"forward", "reverse", "bidirectional"}

BACKUP_TAG = "_PRE_IC_DIR_"


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
def load_config(yaml_path: Path) -> dict:
    raw = _load_yaml(yaml_path)
    if raw is None:
        raw = {}
    directions = raw.get("directions", {})
    if directions is None:
        directions = {}

    config = {"directions": {}}
    for tech, direction in directions.items():
        direction = str(direction).strip().lower()
        if direction not in VALID_DIRECTIONS:
            raise ValueError(
                f"Invalid direction '{direction}' for {tech}. "
                f"Valid: {sorted(VALID_DIRECTIONS)}"
            )
        if direction != "bidirectional":
            config["directions"][tech] = direction

    return config


def load_interconnector_techs(tech_types_path: Path) -> set:
    tech_types_path = Path(tech_types_path)
    if not tech_types_path.is_file():
        raise FileNotFoundError(
            f"TECH_TYPES.csv not found at {tech_types_path}."
        )
    df = pd.read_csv(tech_types_path)
    return set(
        df.loc[df[TECH_TYPES_CATEGORY_COL] == INTERCONNECTOR_CATEGORY,
               TECH_TYPES_TECH_COL].dropna()
    )


# ---------------------------------------------------------------------------
# Backup / restore
# ---------------------------------------------------------------------------
def _rmtree_robust(path: Path, attempts: int = 5) -> None:
    for i in range(attempts):
        try:
            shutil.rmtree(path)
            return
        except PermissionError:
            gc.collect()
            time.sleep(0.1 * (i + 1))
    shutil.rmtree(path)


def make_backup(input_dir: Path) -> Path:
    if not input_dir.is_dir():
        raise FileNotFoundError(f"Input directory does not exist: {input_dir}")
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = input_dir.parent / f"{input_dir.name}{BACKUP_TAG}{stamp}"
    if backup.exists():
        raise FileExistsError(f"Backup already exists: {backup}")
    shutil.copytree(input_dir, backup)
    return backup


def find_latest_backup(input_dir: Path) -> Path | None:
    parent = input_dir.parent
    candidates = sorted(
        (p for p in parent.iterdir()
         if p.is_dir() and p.name.startswith(f"{input_dir.name}{BACKUP_TAG}")),
        key=lambda p: p.name,
    )
    return candidates[-1] if candidates else None


def restore_from_backup(input_dir: Path, backup_dir: Path | None = None) -> Path:
    input_dir = Path(input_dir)
    if backup_dir is None:
        backup_dir = find_latest_backup(input_dir)
        if backup_dir is None:
            raise FileNotFoundError(
                f"No {BACKUP_TAG}* backup found next to {input_dir}."
            )
    else:
        backup_dir = Path(backup_dir)
        if not backup_dir.is_dir():
            raise FileNotFoundError(f"Backup folder does not exist: {backup_dir}")
    if input_dir.is_dir():
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        snapshot = input_dir.parent / (
            f"{input_dir.name}_POST_IC_DIR_pre_restore_{stamp}"
        )
        if not snapshot.exists():
            shutil.copytree(input_dir, snapshot)
        _rmtree_robust(input_dir)
    shutil.copytree(backup_dir, input_dir)
    return backup_dir


# ---------------------------------------------------------------------------
# Worksheet helpers
# ---------------------------------------------------------------------------
def find_year_columns(ws) -> dict:
    year_to_col = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if isinstance(val, (int, float)) and 1900 <= val <= 2200:
            year_to_col[int(val)] = col_idx
        elif isinstance(val, str):
            try:
                y = int(val)
                if 1900 <= y <= 2200:
                    year_to_col[y] = col_idx
            except ValueError:
                pass
    return year_to_col


def find_named_columns(ws, names: list) -> dict:
    found = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val in names:
            found[val] = col_idx
    return found


# ---------------------------------------------------------------------------
# Direction detection
# ---------------------------------------------------------------------------
def parse_tech_regions(tech: str) -> tuple:
    """Extract (source_region, dest_region) from TRN<SRC><DST>."""
    body = tech[len(TRN_PREFIX):]
    return body[:REGION_LEN], body[REGION_LEN:]


def determine_forward_mode(tech: str, mode_input_fuels: dict) -> int:
    """Given {mode_int: input_fuel_str}, return which mode is forward.

    Forward = the mode whose input fuel contains the source region code.
    """
    src, _ = parse_tech_regions(tech)
    for mode, fuel in mode_input_fuels.items():
        if src in fuel:
            return mode
    raise ValueError(
        f"Cannot determine forward mode for {tech}: "
        f"no input fuel contains source region '{src}'. "
        f"Fuels: {mode_input_fuels}"
    )


# ---------------------------------------------------------------------------
# AR Projections editor
# ---------------------------------------------------------------------------
def edit_ar_projections(filepath: Path, trn_techs: set,
                        config: dict, study_start_year: int | None = None) -> dict:
    wb = load_workbook(filepath)
    log = {"file": str(filepath), "changes": [], "warnings": [],
           "study_start_year": study_start_year}

    try:
        if AR_SHEET not in wb.sheetnames:
            log["warnings"].append(f"Sheet '{AR_SHEET}' not found")
            return log
        ws = wb[AR_SHEET]

        cols = find_named_columns(ws, [COL_MODE_OP, COL_TECH, COL_FUEL,
                                       COL_DIRECTION, COL_PMODE])
        missing = [c for c in [COL_MODE_OP, COL_TECH, COL_FUEL, COL_DIRECTION]
                   if c not in cols]
        if missing:
            log["warnings"].append(f"Missing columns: {missing}")
            return log

        mode_col = cols[COL_MODE_OP]
        tech_col = cols[COL_TECH]
        fuel_col = cols[COL_FUEL]
        dir_col = cols[COL_DIRECTION]
        pmode_col = cols.get(COL_PMODE)
        year_cols = find_year_columns(ws)

        if not year_cols:
            log["warnings"].append("No year columns found")
            return log

        directions = config["directions"]

        # Discovery pass: collect rows per configured tech
        tech_rows = defaultdict(list)
        for row_idx in range(2, ws.max_row + 1):
            tech = ws.cell(row=row_idx, column=tech_col).value
            if tech not in directions or tech not in trn_techs:
                continue
            mode_op = ws.cell(row=row_idx, column=mode_col).value
            fuel = ws.cell(row=row_idx, column=fuel_col).value
            direction = ws.cell(row=row_idx, column=dir_col).value
            tech_rows[tech].append({
                "row": row_idx,
                "mode": int(mode_op) if mode_op is not None else None,
                "fuel": str(fuel) if fuel else "",
                "direction": direction,
            })

        # Edit pass
        for tech, desired in directions.items():
            if tech not in tech_rows:
                log["warnings"].append(
                    f"{tech}: not found in '{AR_SHEET}' sheet "
                    f"(or not in TECH_TYPES)"
                )
                continue

            rows = tech_rows[tech]
            input_fuels = {
                r["mode"]: r["fuel"]
                for r in rows if r["direction"] == "Input"
            }
            if len(input_fuels) < 2:
                log["warnings"].append(
                    f"{tech}: expected 2 input-fuel modes, "
                    f"found {len(input_fuels)}"
                )
                continue

            forward_mode = determine_forward_mode(tech, input_fuels)
            all_modes = set(r["mode"] for r in rows)
            reverse_modes = all_modes - {forward_mode}

            if desired == "forward":
                disabled = reverse_modes
            elif desired == "reverse":
                disabled = {forward_mode}
            else:
                continue

            cells_zeroed = 0
            rows_touched = 0
            for r in rows:
                if r["mode"] not in disabled:
                    continue
                rows_touched += 1
                row_idx = r["row"]
                for year, col in sorted(year_cols.items()):
                    # STUDY PERIOD restriction: leave base-window years (< study_start_year)
                    # bidirectional so the pinned calibrated base-year flows stay feasible.
                    if study_start_year is not None and year < study_start_year:
                        continue
                    cell = ws.cell(row=row_idx, column=col)
                    if cell.value not in (0, 0.0, None):
                        cell.value = 0.0
                        cells_zeroed += 1
                    elif cell.value is None:
                        cell.value = 0.0
                if pmode_col is not None:
                    ws.cell(row=row_idx, column=pmode_col).value = PMODE_USER

            src, dst = parse_tech_regions(tech)
            if desired == "forward":
                kept_label = f"{src} -> {dst}"
            else:
                kept_label = f"{dst} -> {src}"

            log["changes"].append({
                "tech": tech,
                "direction": desired,
                "kept": kept_label,
                "forward_mode": forward_mode,
                "disabled_modes": sorted(disabled),
                "rows_touched": rows_touched,
                "cells_zeroed": cells_zeroed,
            })

        wb.save(filepath)
    finally:
        wb.close()

    return log


# ---------------------------------------------------------------------------
# AR Base Year editor
# ---------------------------------------------------------------------------
def edit_ar_base_year(filepath: Path, trn_techs: set,
                      config: dict, mode_map: dict) -> dict:
    """Zero disabled mode values in the AR Base Year workbook.

    mode_map: {tech: forward_mode_int} built from the projections pass.
    """
    wb = load_workbook(filepath)
    log = {"file": str(filepath), "changes": [], "warnings": []}

    try:
        if AR_SHEET not in wb.sheetnames:
            log["warnings"].append(f"Sheet '{AR_SHEET}' not found")
            return log
        ws = wb[AR_SHEET]

        cols = find_named_columns(
            ws,
            [COL_MODE_OP, COL_TECH, COL_BASE_VALUE_I, COL_BASE_VALUE_O],
        )
        required = [COL_MODE_OP, COL_TECH, COL_BASE_VALUE_I, COL_BASE_VALUE_O]
        missing = [c for c in required if c not in cols]
        if missing:
            log["warnings"].append(f"Missing columns: {missing}")
            return log

        mode_col = cols[COL_MODE_OP]
        tech_col = cols[COL_TECH]
        val_i_col = cols[COL_BASE_VALUE_I]
        val_o_col = cols[COL_BASE_VALUE_O]

        directions = config["directions"]

        for row_idx in range(2, ws.max_row + 1):
            tech = ws.cell(row=row_idx, column=tech_col).value
            if tech not in directions or tech not in trn_techs:
                continue

            mode_op = ws.cell(row=row_idx, column=mode_col).value
            mode_int = int(mode_op) if mode_op is not None else None

            desired = directions[tech]
            forward_mode = mode_map.get(tech)
            if forward_mode is None:
                continue

            if desired == "forward":
                is_disabled = mode_int != forward_mode
            elif desired == "reverse":
                is_disabled = mode_int == forward_mode
            else:
                continue

            if not is_disabled:
                continue

            old_i = ws.cell(row=row_idx, column=val_i_col).value
            old_o = ws.cell(row=row_idx, column=val_o_col).value
            ws.cell(row=row_idx, column=val_i_col).value = 0.0
            ws.cell(row=row_idx, column=val_o_col).value = 0.0

            log["changes"].append({
                "tech": tech,
                "mode": mode_int,
                "old_value_i": old_i,
                "old_value_o": old_o,
            })

        wb.save(filepath)
    finally:
        wb.close()

    return log


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------
def run(input_dir, skip_backup: bool = False,
        yaml_path: Path | None = None,
        study_start_year: int | None = None) -> dict:
    input_dir = Path(input_dir)

    if yaml_path is None:
        yaml_path = Path(__file__).resolve().parent / YAML_FILE_NAME
    if not yaml_path.is_file():
        raise FileNotFoundError(f"YAML config not found at {yaml_path}")
    config = load_config(yaml_path)

    if not config["directions"]:
        print("No direction overrides configured -- nothing to do.")
        return {"skipped": True}

    script_dir = Path(__file__).resolve().parent
    tech_types_path = script_dir.parent / TECH_TYPES_FILE
    trn_techs = load_interconnector_techs(tech_types_path)

    backup_dir = None if skip_backup else make_backup(input_dir)

    # Edit AR Projections
    proj_file = input_dir / AR_PROJ_FILE
    if not proj_file.exists():
        raise FileNotFoundError(f"{proj_file} not found")
    proj_log = edit_ar_projections(proj_file, trn_techs, config, study_start_year)

    # Build mode map for the base year pass
    mode_map = {
        ch["tech"]: ch["forward_mode"]
        for ch in proj_log.get("changes", [])
    }

    # Edit AR Base Year -- SKIP entirely when a study_start_year is set: the base-year AR
    # file governs the pinned base window (< study_start_year), which must stay at the
    # calibrated (bidirectional) mix so its demand balance remains feasible.
    base_file = input_dir / AR_BASE_FILE
    base_log = {"skipped": True}
    if study_start_year is not None:
        print(f"[study_start_year={study_start_year}] skipping AR Base Year edit "
              f"(base window stays at calibrated/bidirectional mix)")
        base_log = {"skipped": True, "reason": f"study_start_year={study_start_year}"}
    elif base_file.exists():
        base_log = edit_ar_base_year(base_file, trn_techs, config, mode_map)
    else:
        print(f"WARNING: {AR_BASE_FILE} not found -- skipping base year edit")

    log = {
        "timestamp": datetime.now().isoformat(),
        "backup_dir": str(backup_dir) if backup_dir else None,
        "config": config,
        "projections": proj_log,
        "base_year": base_log,
    }

    if backup_dir is not None:
        log_path = backup_dir.parent / f"{backup_dir.name}_CHANGES.json"
        log_path.write_text(json.dumps(log, indent=2, default=str))
        log["log_path"] = str(log_path)

    return log


# ---------------------------------------------------------------------------
# Console output
# ---------------------------------------------------------------------------
def print_summary(log: dict) -> None:
    bar = "=" * 72
    print(bar)
    print("set_interconnector_direction -- directionality control applied")
    print(bar)
    print(f"Backup folder : {log.get('backup_dir', '(skipped)')}")

    proj = log.get("projections", {})
    print(f"\nAR Projections : {proj.get('file', '?')}")
    for ch in proj.get("changes", []):
        print(f"  {ch['tech']}  direction={ch['direction']}  "
              f"kept={ch['kept']}  "
              f"disabled mode(s)={ch['disabled_modes']}  "
              f"cells zeroed={ch['cells_zeroed']}")
    for w in proj.get("warnings", []):
        print(f"  WARNING: {w}")

    base = log.get("base_year", {})
    if "skipped" not in base:
        print(f"\nAR Base Year   : {base.get('file', '?')}")
        for ch in base.get("changes", []):
            print(f"  {ch['tech']}  mode={ch['mode']}  "
                  f"Value.I: {ch['old_value_i']} -> 0  "
                  f"Value.O: {ch['old_value_o']} -> 0")
        for w in base.get("warnings", []):
            print(f"  WARNING: {w}")

    print(bar)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main() -> int:
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--input-dir", type=Path,
        default=Path("A1_Outputs/A1_Outputs_BAU"),
        help="Directory containing the A-O files",
    )
    parser.add_argument(
        "--skip-backup", action="store_true",
        help="Skip backup (DANGEROUS -- testing only)",
    )
    parser.add_argument(
        "--restore", action="store_true",
        help=f"Restore from most recent {BACKUP_TAG}* backup",
    )
    parser.add_argument(
        "--restore-from", type=Path, default=None,
        help="Restore from this specific backup folder",
    )
    parser.add_argument(
        "--yaml", type=Path, default=None,
        help="Path to YAML config (default: next to this script)",
    )
    parser.add_argument(
        "--study-start-year", type=int, default=None,
        help="Apply the direction lever only to years >= this (study period); leave the "
             "pinned base window (< this year) bidirectional. Omit for all-years (default).",
    )
    args = parser.parse_args()

    if args.restore or args.restore_from is not None:
        try:
            used = restore_from_backup(args.input_dir, args.restore_from)
        except Exception as exc:
            print(f"ERROR: {exc}", file=sys.stderr)
            return 1
        print(f"Restored {args.input_dir} from {used}")
        return 0

    try:
        log = run(args.input_dir, args.skip_backup, args.yaml, args.study_start_year)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    if not log.get("skipped"):
        print_summary(log)
    return 0


if __name__ == "__main__":
    sys.exit(main())
