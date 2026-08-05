"""Apply the audited 2023-2026 PWR/MIN calibration allowlist.

The accepted WS-4 recipe formerly read a solver result and broadly rewrote
whole technology bands.  Production now consumes a frozen, key-complete table
derived from the independently audited restoration evidence.  Only an exact
``(REGION, TECHNOLOGY, PARAMETER, YEAR)`` rule may change a workbook cell.

Frozen lineage:

* corrected evidence ``canonical_source_rules.csv`` SHA-256:
  ``9c28f9d43c3037daa668554a94061e829d0974662a746efa48d4a2dc341b9ca6``
* production projection ``pwr_min_2023_2026_pin.csv`` SHA-256:
  ``cdcb0aeb570486b40ab96be68f6db031af54afa3ac02e4832a456522ca73a17c``

The transformation is deliberately unable to create rows or columns, touch
Maldives, write outside 2023-2026, or synthesize an absent value as zero or
9999.  ``Projection.Mode`` is row-wide, so an EMPTY row can be activated only
when every non-target year cell is blank; otherwise the run fails before any
workbook mutation.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import os
import shutil
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from ostram.paths import resolve_paths
from ostram.profiles import DEFAULT_PROFILE, active_profile_id, profile_policy

from openpyxl import load_workbook


PARAM_FILE = "A-O_Parametrization.xlsx"
SHEETS = ("Primary Techs", "Secondary Techs")
PIN_YEARS = frozenset({2023, 2024, 2025, 2026})
PIN_ROOT_SCENARIOS = frozenset(
    {"A_Calibrated_BAU", "B_Optimised_VRE", "C_Target_VRE"}
)
RULES_CSV = (
    resolve_paths().scenario_config_root / "rules" / "pwr_min_2023_2026_pin.csv"
)
RULES_SHA256 = (
    "cdcb0aeb570486b40ab96be68f6db031af54afa3ac02e4832a456522ca73a17c"
)
CANONICAL_SOURCE_RULES_SHA256 = (
    "9c28f9d43c3037daa668554a94061e829d0974662a746efa48d4a2dc341b9ca6"
)
BACKUP_TAG = "_PRE_PWR_MIN_PIN_"

P_MAX_CAP = "TotalAnnualMaxCapacity"
P_MAX_INV = "TotalAnnualMaxCapacityInvestment"
P_MIN_INV = "TotalAnnualMinCapacityInvestment"
P_ACTIVITY_LOWER = "TotalTechnologyAnnualActivityLowerLimit"
P_ACTIVITY_UPPER = "TotalTechnologyAnnualActivityUpperLimit"
ALLOWED_PARAMETERS = frozenset(
    {
        P_MAX_CAP,
        P_MAX_INV,
        P_MIN_INV,
        P_ACTIVITY_LOWER,
        P_ACTIVITY_UPPER,
    }
)
ACTIVITY_PARAMETERS = frozenset({P_ACTIVITY_LOWER, P_ACTIVITY_UPPER})
ALLOWED_COUNTRIES = frozenset({"BGD", "BTN", "IND", "LKA", "NPL"})
ORDERED_INDICES = ("REGION", "TECHNOLOGY", "YEAR")
EXPECTED_AUTHORITY = "BENCHMARK_SUPPORTED"
EXPECTED_LINEAGE = "ACCEPTED_WS4_BASE_YEAR_PIN_2023_2026"
EXPECTED_FIELDS = (
    "source_rule_id",
    "parameter",
    "ordered_parameter_indices",
    "region",
    "technology",
    "semantic_technology_group",
    "canonical_country",
    "year",
    "required_root_present",
    "required_root_value",
    "required_root_state",
    "verified_physical_unit",
    "verified_compiled_unit",
    "root_scenarios_with_actual_change",
    "authority_classification",
    "authority_lineage_class",
)
EXPECTED_PARAMETER_COUNTS = {
    P_MAX_CAP: 620,
    P_MAX_INV: 388,
    P_MIN_INV: 144,
    P_ACTIVITY_LOWER: 216,
    P_ACTIVITY_UPPER: 588,
}
EXPECTED_STATE_COUNTS = {"POSITIVE": 1356, "ZERO": 600}
EXPECTED_SCENARIO_COUNTS = {
    "A_Calibrated_BAU": 1915,
    "B_Optimised_VRE": 1956,
    "C_Target_VRE": 1956,
}


@dataclass(frozen=True)
class PinRule:
    source_rule_id: str
    parameter: str
    region: str
    technology: str
    semantic_technology_group: str
    canonical_country: str
    year: int
    present: bool
    value: Decimal
    state: str
    physical_unit: str
    compiled_unit: str
    root_scenarios: tuple[str, ...]
    authority_classification: str
    authority_lineage_class: str

    @property
    def complete_key(self) -> tuple[str, str, str, int]:
        return self.region, self.technology, self.parameter, self.year


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _parse_decimal(raw: str, *, label: str) -> Decimal:
    try:
        value = Decimal(raw)
    except (InvalidOperation, ValueError) as error:
        raise ValueError(f"{label} is not a decimal: {raw!r}") from error
    if not value.is_finite():
        raise ValueError(f"{label} must be finite: {raw!r}")
    return value


def _validate_units(
    parameter: str,
    technology: str,
    physical_unit: str,
    compiled_unit: str,
) -> None:
    if parameter in ACTIVITY_PARAMETERS:
        expected = ("PJ_per_year_activity", "PJ_per_year")
    elif technology.startswith("MIN"):
        expected = (
            "native_supply_capacity_unit",
            "native_supply_capacity_unit",
        )
    else:
        expected = ("GW", "GW")
    actual = (physical_unit, compiled_unit)
    if actual != expected:
        raise ValueError(
            f"invalid unit pair for {technology}/{parameter}: "
            f"{actual!r}, expected {expected!r}"
        )


def _parse_rule(row: dict[str, str], row_number: int) -> PinRule:
    label = f"rules row {row_number}"
    try:
        indices = tuple(json.loads(row["ordered_parameter_indices"]))
    except (json.JSONDecodeError, TypeError) as error:
        raise ValueError(f"{label} has invalid ordered indices") from error
    if indices != ORDERED_INDICES:
        raise ValueError(f"{label} ordered indices are not {ORDERED_INDICES!r}")
    if row["region"] != "GLOBAL":
        raise ValueError(f"{label} region is not GLOBAL")
    parameter = row["parameter"]
    if parameter not in ALLOWED_PARAMETERS:
        raise ValueError(f"{label} has unsupported parameter {parameter!r}")
    try:
        year = int(row["year"])
    except ValueError as error:
        raise ValueError(f"{label} has invalid year {row['year']!r}") from error
    if year not in PIN_YEARS:
        raise ValueError(f"{label} year {year} is outside 2023-2026")
    technology = row["technology"]
    if not technology.startswith(("PWR", "MIN")):
        raise ValueError(f"{label} technology is not PWR/MIN: {technology!r}")
    if "MDV" in technology:
        raise ValueError(f"{label} contains a Maldives technology")
    country = row["canonical_country"]
    if country not in ALLOWED_COUNTRIES:
        raise ValueError(f"{label} has unsupported country {country!r}")
    if row["required_root_present"] != "true":
        raise ValueError(
            f"{label} must be explicitly present; absence cannot be encoded"
        )
    value = _parse_decimal(
        row["required_root_value"],
        label=f"{label} required_root_value",
    )
    if value < 0:
        raise ValueError(f"{label} value must be non-negative")
    state = row["required_root_state"]
    expected_state = "ZERO" if value == 0 else "POSITIVE"
    if state != expected_state:
        raise ValueError(
            f"{label} state/value mismatch: {state!r} vs {value}"
        )
    if value == Decimal("9999") and parameter != P_MAX_CAP:
        raise ValueError(f"{label} uses 9999 outside {P_MAX_CAP}")
    scenarios = tuple(
        item
        for item in row["root_scenarios_with_actual_change"].split(";")
        if item
    )
    if not scenarios or len(scenarios) != len(set(scenarios)):
        raise ValueError(f"{label} has invalid root-scenario membership")
    if not set(scenarios).issubset(PIN_ROOT_SCENARIOS):
        raise ValueError(f"{label} has an unsupported root scenario")
    if row["authority_classification"] != EXPECTED_AUTHORITY:
        raise ValueError(f"{label} is not benchmark-supported")
    if row["authority_lineage_class"] != EXPECTED_LINEAGE:
        raise ValueError(f"{label} has an unsupported authority lineage")
    expected_id = (
        f"PWR_MIN_PIN::{parameter}::GLOBAL::{technology}::{year}"
    )
    if row["source_rule_id"] != expected_id:
        raise ValueError(
            f"{label} source_rule_id does not match its complete key"
        )
    _validate_units(
        parameter,
        technology,
        row["verified_physical_unit"],
        row["verified_compiled_unit"],
    )
    return PinRule(
        source_rule_id=row["source_rule_id"],
        parameter=parameter,
        region="GLOBAL",
        technology=technology,
        semantic_technology_group=row["semantic_technology_group"],
        canonical_country=country,
        year=year,
        present=True,
        value=value,
        state=state,
        physical_unit=row["verified_physical_unit"],
        compiled_unit=row["verified_compiled_unit"],
        root_scenarios=scenarios,
        authority_classification=row["authority_classification"],
        authority_lineage_class=row["authority_lineage_class"],
    )


def _validate_production_contract(rules: tuple[PinRule, ...]) -> None:
    if len(rules) != 1956:
        raise ValueError(f"production rule count is {len(rules)}, expected 1956")
    if len({(rule.technology, rule.parameter) for rule in rules}) != 517:
        raise ValueError("production rules do not resolve to exactly 517 rows")
    if len({rule.technology for rule in rules}) != 182:
        raise ValueError("production rules do not contain exactly 182 technologies")
    if Counter(rule.parameter for rule in rules) != EXPECTED_PARAMETER_COUNTS:
        raise ValueError("production parameter distribution mismatch")
    if Counter(rule.state for rule in rules) != EXPECTED_STATE_COUNTS:
        raise ValueError("production state distribution mismatch")
    scenario_counts = {
        scenario: sum(scenario in rule.root_scenarios for rule in rules)
        for scenario in PIN_ROOT_SCENARIOS
    }
    if scenario_counts != EXPECTED_SCENARIO_COUNTS:
        raise ValueError(
            f"production scenario distribution mismatch: {scenario_counts}"
        )


def load_pin_rules(
    rules_csv: Path | str = RULES_CSV,
    *,
    enforce_production_contract: bool | None = None,
) -> tuple[PinRule, ...]:
    """Load and fail-close validate the complete source-rule allowlist."""
    path = Path(rules_csv)
    if not path.is_file():
        raise FileNotFoundError(path)
    is_default = path.resolve() == RULES_CSV.resolve()
    profile_id = active_profile_id()
    if enforce_production_contract is None:
        enforce_production_contract = is_default and profile_id == DEFAULT_PROFILE
    if is_default:
        expected_hash = profile_policy("pwr_min_pin_rules_sha256")
        if expected_hash is None and profile_id == DEFAULT_PROFILE:
            expected_hash = RULES_SHA256
        if not (
            isinstance(expected_hash, str)
            and len(expected_hash) == 64
            and all(character in "0123456789abcdef" for character in expected_hash)
        ):
            raise ValueError(
                f"profile {profile_id!r} does not declare a valid "
                "pwr_min_pin_rules_sha256 policy"
            )
        actual_hash = _sha256(path)
        if actual_hash != expected_hash:
            raise ValueError(
                f"profile rule hash mismatch: {actual_hash} != {expected_hash}"
            )
    with path.open("r", encoding="utf-8-sig", newline="") as stream:
        reader = csv.DictReader(stream)
        fields = tuple(reader.fieldnames or ())
        if fields != EXPECTED_FIELDS:
            raise ValueError(
                f"rule header mismatch: {fields!r} != {EXPECTED_FIELDS!r}"
            )
        rules = tuple(
            _parse_rule(row, row_number)
            for row_number, row in enumerate(reader, start=2)
        )
    ids = [rule.source_rule_id for rule in rules]
    keys = [rule.complete_key for rule in rules]
    if len(ids) != len(set(ids)):
        raise ValueError("duplicate source_rule_id in pin rules")
    if len(keys) != len(set(keys)):
        raise ValueError("duplicate complete source key in pin rules")
    if enforce_production_contract:
        _validate_production_contract(rules)
    return rules


def _headers(worksheet) -> dict[object, int]:
    headers: dict[object, int] = {}
    for column in range(1, worksheet.max_column + 1):
        value = worksheet.cell(row=1, column=column).value
        if value is None:
            continue
        if value in headers:
            raise ValueError(
                f"{worksheet.title!r} has duplicate header {value!r}"
            )
        headers[value] = column
    return headers


def _year_columns(headers: dict[object, int]) -> dict[int, int]:
    result: dict[int, int] = {}
    for raw, column in headers.items():
        year: int | None = None
        if isinstance(raw, int) and not isinstance(raw, bool):
            year = raw
        elif isinstance(raw, str) and raw.strip().isdigit():
            year = int(raw.strip())
        if year is None:
            continue
        if year in result:
            raise ValueError(f"duplicate year header {year}")
        result[year] = column
    return result


def _cell_decimal(value: object, *, label: str) -> Decimal:
    if value is None or value == "" or isinstance(value, bool):
        raise ValueError(f"{label} is not an explicit numeric value")
    if isinstance(value, str) and value.startswith("="):
        raise ValueError(f"{label} is a formula")
    try:
        result = Decimal(str(value))
    except (InvalidOperation, ValueError) as error:
        raise ValueError(f"{label} is not numeric: {value!r}") from error
    if not result.is_finite():
        raise ValueError(f"{label} is not finite")
    return result


def _excel_number(value: Decimal) -> int | float:
    if value == value.to_integral_value():
        return int(value)
    result = float(value)
    if not math.isfinite(result):
        raise ValueError(f"cannot serialize non-finite value {value}")
    return result


def make_backup(input_dir: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = input_dir.parent / f"{input_dir.name}{BACKUP_TAG}{stamp}"
    if backup.exists():
        raise FileExistsError(backup)
    shutil.copytree(input_dir, backup)
    return backup


def restore(input_dir: Path | str, restore_from: Path | str | None = None) -> Path:
    destination = Path(input_dir)
    if restore_from is None:
        candidates = sorted(
            path
            for path in destination.parent.iterdir()
            if path.is_dir()
            and path.name.startswith(f"{destination.name}{BACKUP_TAG}")
        )
        if not candidates:
            raise FileNotFoundError(
                f"no {BACKUP_TAG}* backup found beside {destination}"
            )
        source = candidates[-1]
    else:
        source = Path(restore_from)
    if not source.is_dir():
        raise FileNotFoundError(source)
    if destination.exists():
        shutil.rmtree(destination)
    shutil.copytree(source, destination)
    return source


def apply_pin_rules(
    input_dir: Path | str,
    scenario: str,
    rules_csv: Path | str = RULES_CSV,
    *,
    skip_backup: bool = False,
    enforce_production_contract: bool | None = None,
) -> dict[str, object]:
    """Apply exact rules for one canonical root scenario.

    All structural, key, and row-wide projection checks complete before the
    first workbook assignment.  A validation failure therefore leaves the
    input workbook byte-for-byte unchanged.
    """
    if scenario not in PIN_ROOT_SCENARIOS:
        raise ValueError(f"unsupported pin scenario: {scenario!r}")
    rules_path = Path(rules_csv)
    all_rules = load_pin_rules(
        rules_path,
        enforce_production_contract=enforce_production_contract,
    )
    rules = tuple(
        rule for rule in all_rules if scenario in rule.root_scenarios
    )
    if not rules:
        raise ValueError(f"no pin rules apply to scenario {scenario!r}")
    if (
        rules_path.resolve() == RULES_CSV.resolve()
        and active_profile_id() == DEFAULT_PROFILE
        and len(rules) != EXPECTED_SCENARIO_COUNTS[scenario]
    ):
        raise ValueError(
            f"{scenario} rule count is {len(rules)}, expected "
            f"{EXPECTED_SCENARIO_COUNTS[scenario]}"
        )

    directory = Path(input_dir)
    workbook_path = directory / PARAM_FILE
    if not workbook_path.is_file():
        raise FileNotFoundError(workbook_path)
    workbook = load_workbook(workbook_path)
    temp_path = workbook_path.with_name(
        f".{workbook_path.stem}.pwr-min-pin.tmp.xlsx"
    )
    if temp_path.exists():
        workbook.close()
        raise FileExistsError(temp_path)

    rules_by_row: dict[tuple[str, str], list[PinRule]] = defaultdict(list)
    for rule in rules:
        rules_by_row[(rule.technology, rule.parameter)].append(rule)

    try:
        missing_sheets = [sheet for sheet in SHEETS if sheet not in workbook]
        if missing_sheets:
            raise ValueError(f"missing required sheets: {missing_sheets}")
        locations: dict[tuple[str, str], list[tuple[object, int]]] = defaultdict(
            list
        )
        sheet_metadata: dict[str, tuple[dict[object, int], dict[int, int]]] = {}
        for sheet_name in SHEETS:
            worksheet = workbook[sheet_name]
            headers = _headers(worksheet)
            required_headers = {
                "Tech",
                "Parameter",
                "Projection.Mode",
                "Projection.Parameter",
            }
            missing_headers = sorted(required_headers - set(headers))
            if missing_headers:
                raise ValueError(
                    f"{sheet_name!r} is missing headers {missing_headers}"
                )
            years = _year_columns(headers)
            sheet_metadata[sheet_name] = headers, years
            tech_column = headers["Tech"]
            parameter_column = headers["Parameter"]
            for row_number in range(2, worksheet.max_row + 1):
                key = (
                    worksheet.cell(row=row_number, column=tech_column).value,
                    worksheet.cell(
                        row=row_number, column=parameter_column
                    ).value,
                )
                if key in rules_by_row:
                    locations[key].append((worksheet, row_number))

        missing_rows = sorted(
            key for key in rules_by_row if len(locations.get(key, ())) == 0
        )
        duplicate_rows = {
            key: [(worksheet.title, row) for worksheet, row in found]
            for key, found in locations.items()
            if len(found) > 1
        }
        if missing_rows:
            raise ValueError(f"missing target workbook rows: {missing_rows[:10]}")
        if duplicate_rows:
            raise ValueError(f"duplicate target workbook rows: {duplicate_rows}")

        assignments: list[tuple[object, int, Decimal, PinRule]] = []
        projection_flips: list[tuple[object, int]] = []
        for key, row_rules in rules_by_row.items():
            worksheet, row_number = locations[key][0]
            headers, year_columns = sheet_metadata[worksheet.title]
            target_years = {rule.year for rule in row_rules}
            missing_years = sorted(target_years - set(year_columns))
            if missing_years:
                raise ValueError(
                    f"{worksheet.title}/{key} is missing years {missing_years}"
                )
            mode_cell = worksheet.cell(
                row=row_number, column=headers["Projection.Mode"]
            )
            mode = mode_cell.value
            if mode == "User defined":
                pass
            elif mode in (None, "", "EMPTY"):
                non_target_values = [
                    (year, worksheet.cell(row=row_number, column=column).value)
                    for year, column in sorted(year_columns.items())
                    if year not in target_years
                    and worksheet.cell(row=row_number, column=column).value
                    not in (None, "")
                ]
                if non_target_values:
                    raise ValueError(
                        f"{worksheet.title}/{key} cannot activate row-wide "
                        f"Projection.Mode; populated non-target years: "
                        f"{non_target_values[:8]}"
                    )
                projection_parameter = worksheet.cell(
                    row=row_number, column=headers["Projection.Parameter"]
                ).value
                if projection_parameter not in (None, "", 0, 0.0):
                    raise ValueError(
                        f"{worksheet.title}/{key} has unsafe "
                        f"Projection.Parameter {projection_parameter!r}"
                    )
                projection_flips.append((mode_cell, row_number))
            else:
                raise ValueError(
                    f"{worksheet.title}/{key} has unsupported "
                    f"Projection.Mode {mode!r}"
                )
            for rule in row_rules:
                cell = worksheet.cell(
                    row=row_number, column=year_columns[rule.year]
                )
                if cell.value not in (None, ""):
                    _cell_decimal(
                        cell.value,
                        label=(
                            f"{worksheet.title}/{rule.technology}/"
                            f"{rule.parameter}/{rule.year}"
                        ),
                    )
                assignments.append((cell, row_number, rule.value, rule))

        backup = None if skip_backup else make_backup(directory)
        changed_value_cells = 0
        zero_cells = 0
        positive_cells = 0
        for cell, _row_number, value, rule in assignments:
            current = (
                None
                if cell.value in (None, "")
                else _cell_decimal(
                    cell.value,
                    label=f"current cell for {rule.source_rule_id}",
                )
            )
            if current != value:
                cell.value = _excel_number(value)
                changed_value_cells += 1
            if rule.state == "ZERO":
                zero_cells += 1
            else:
                positive_cells += 1
        changed_projection_modes = 0
        for mode_cell, _row_number in projection_flips:
            if mode_cell.value != "User defined":
                mode_cell.value = "User defined"
                changed_projection_modes += 1

        for cell, _row_number, value, rule in assignments:
            actual = _cell_decimal(
                cell.value,
                label=f"post-apply cell for {rule.source_rule_id}",
            )
            if actual != value:
                raise RuntimeError(
                    f"post-apply mismatch for {rule.source_rule_id}: "
                    f"{actual} != {value}"
                )

        changed = changed_value_cells + changed_projection_modes
        if changed:
            workbook.save(temp_path)
            workbook.close()
            os.replace(temp_path, workbook_path)
        else:
            workbook.close()
        return {
            "status": "PASS",
            "scenario": scenario,
            "input_dir": str(directory),
            "workbook": str(workbook_path),
            "rules_csv": str(rules_path),
            "rules_sha256": _sha256(rules_path),
            "canonical_source_rules_sha256": CANONICAL_SOURCE_RULES_SHA256,
            "rules_loaded": len(all_rules),
            "rules_applied": len(rules),
            "workbook_rows_matched": len(rules_by_row),
            "zero_rules": zero_cells,
            "positive_rules": positive_cells,
            "changed_value_cells": changed_value_cells,
            "changed_projection_modes": changed_projection_modes,
            "saved": bool(changed),
            "backup_dir": str(backup) if backup is not None else None,
        }
    finally:
        try:
            workbook.close()
        finally:
            if temp_path.exists():
                temp_path.unlink()


def run(
    input_dir: Path | str,
    scenario: str,
    rules_csv: Path | str = RULES_CSV,
    *,
    skip_backup: bool = False,
) -> dict[str, object]:
    """Compatibility wrapper around :func:`apply_pin_rules`."""
    return apply_pin_rules(
        input_dir,
        scenario,
        rules_csv,
        skip_backup=skip_backup,
    )


def print_summary(log: dict[str, object]) -> None:
    print(
        "apply_base_year_pin "
        f"scenario={log['scenario']} status={log['status']} "
        f"rules={log['rules_applied']} rows={log['workbook_rows_matched']} "
        f"value_changes={log['changed_value_cells']} "
        f"projection_mode_changes={log['changed_projection_modes']}"
    )


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input-dir", type=Path)
    parser.add_argument("--scenario")
    parser.add_argument("--rules-csv", type=Path, default=RULES_CSV)
    parser.add_argument("--skip-backup", action="store_true")
    parser.add_argument("--restore", action="store_true")
    parser.add_argument("--restore-from", type=Path)
    args = parser.parse_args()
    if args.restore or args.restore_from is not None:
        if args.input_dir is None:
            parser.error("--input-dir is required for restore")
        try:
            source = restore(args.input_dir, args.restore_from)
        except Exception as error:
            print(f"ERROR: {error}", file=sys.stderr)
            return 1
        print(f"Restored {args.input_dir} from {source}")
        return 0
    if args.input_dir is None or args.scenario is None:
        parser.error("--input-dir and --scenario are required")
    try:
        log = apply_pin_rules(
            args.input_dir,
            args.scenario,
            args.rules_csv,
            skip_backup=args.skip_backup,
        )
    except Exception as error:
        print(f"ERROR: {error}", file=sys.stderr)
        return 1
    print_summary(log)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
