"""
set_retirement_schedule.py
==========================

Independent patch — draws down ResidualCapacity for thermal generation
technologies over time, modelling planned retirements and end-of-life
decommissioning.

Two retirement mechanisms, applied in order:
  A) Age-based: linear or back-loaded (logistic) drawdown per fuel type.
  B) Scheduled: explicit per-tech ResidualCapacity trajectories from
     policy documents (overrides age-based).

Exempt tech patterns (renewables, hydro, backstop) are never touched.

Safety constraints enforce non-negative values and year-over-year
monotonicity (ResidualCapacity can only decrease or stay flat).

This script modifies ResidualCapacity ONLY. It never touches MaxCapInv,
MinCapInv, or ActivityLimits.

CONFIGURATION
-------------
Edit retirement_schedule.yaml (next to this script).

USAGE
-----
    python set_retirement_schedule.py
    python set_retirement_schedule.py --self-test
    python set_retirement_schedule.py --restore
"""

from __future__ import annotations

import argparse
import gc
import json
import shutil
import sys
import tempfile
import time
from datetime import datetime
from math import exp
from pathlib import Path

import pandas as pd
from ostram.paths import resolve_paths
from openpyxl import load_workbook, Workbook

try:
    import yaml as _yaml
    def _load_yaml(path: Path) -> dict:
        with open(path, "r", encoding="utf-8") as f:
            return _yaml.safe_load(f)
except ImportError:
    _yaml = None
    def _load_yaml(path: Path) -> dict:
        raise ImportError("PyYAML is required. Install with: pip install pyyaml")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
DEFAULT_TARGET_SHEETS = ["Secondary Techs"]
PARAM_FILE_NAME = "A-O_Parametrization.xlsx"
YAML_FILE_NAME = "retirement_schedule.yaml"

RES_PARAM = "ResidualCapacity"

PROJ_MODE_COL = "Projection.Mode"
PROJ_MODE_EMPTY = "EMPTY"
PROJ_MODE_USER = "User defined"

TECH_TYPES_FILE = "TECH_TYPES.csv"
TECH_TYPES_CATEGORY_COL = "Technology (PWR)"
TECH_TYPES_TECH_COL = "Technology"
GENERATION_CATEGORY = "GENERATION"

PWR_TECH_LENGTH = 11
COUNTRY_REGION_SLICE = slice(6, 11)

BACKUP_TAG = "_PRE_RETIRE_"


# ---------------------------------------------------------------------------
# Retirement profile functions
# ---------------------------------------------------------------------------
def retired_fraction_linear(years_elapsed: float, lifetime: float) -> float:
    """Linear retirement: uniform commissioning over [base-lifetime, base].
    Each year, 1/lifetime of the base fleet retires."""
    if lifetime <= 0:
        return 1.0
    return min(1.0, max(0.0, years_elapsed / lifetime))


def retired_fraction_backloaded(years_elapsed: float, lifetime: float) -> float:
    """Back-loaded (logistic) retirement. Most retirements in the last
    third of life. Midpoint at 70% of lifetime, steepness = 6/lifetime."""
    if lifetime <= 0:
        return 1.0
    midpoint = lifetime * 0.7
    steepness = 6.0 / lifetime
    return 1.0 / (1.0 + exp(-steepness * (years_elapsed - midpoint)))


PROFILE_FUNCS = {
    "linear": retired_fraction_linear,
    "back-loaded": retired_fraction_backloaded,
}


# ---------------------------------------------------------------------------
# YAML config loader
# ---------------------------------------------------------------------------
def load_config(yaml_path: Path) -> dict:
    cfg = _load_yaml(yaml_path)
    if cfg is None:
        cfg = {}

    base_year = int(cfg.get("base_year", 2023))

    # Fuel slice for extracting fuel code from tech name
    fuel_slice_raw = cfg.get("fuel_slice", [3, 6])
    fuel_slice = slice(int(fuel_slice_raw[0]), int(fuel_slice_raw[1]))

    # Age-based retirement rules
    raw_age = cfg.get("age_based", {}) or {}
    age_based: dict = {}
    for fuel, params in raw_age.items():
        fuel = str(fuel).upper()
        lifetime = int(params.get("lifetime_years", 30))
        profile = str(params.get("retirement_profile", "linear"))
        if profile not in PROFILE_FUNCS:
            raise ValueError(
                f"age_based.{fuel}: unknown profile '{profile}'. "
                f"Expected one of: {list(PROFILE_FUNCS.keys())}"
            )
        age_based[fuel] = {"lifetime_years": lifetime, "retirement_profile": profile}

    # Scheduled overrides
    raw_sched = cfg.get("scheduled", []) or []
    scheduled = []
    for entry in raw_sched:
        cr = str(entry.get("cr", ""))
        tech = str(entry.get("tech", ""))
        schedule = entry.get("schedule", {})
        note = str(entry.get("note", ""))
        if not cr or not tech or not schedule:
            raise ValueError(f"scheduled entry missing cr/tech/schedule: {entry}")
        schedule = {int(y): float(v) for y, v in schedule.items()}
        scheduled.append({"cr": cr, "tech": tech, "schedule": schedule, "note": note})

    # Exempt patterns
    exempt = [str(p) for p in (cfg.get("exempt", []) or [])]

    return {
        "base_year": base_year,
        "fuel_slice": fuel_slice,
        "age_based": age_based,
        "scheduled": scheduled,
        "exempt": exempt,
    }


# ---------------------------------------------------------------------------
# Interpolation
# ---------------------------------------------------------------------------
def interpolate_schedule(schedule: dict, years: list) -> dict:
    """Linearly interpolate. Before first key: first value. After last: last."""
    if not schedule:
        return {y: 0.0 for y in years}
    sorted_keys = sorted(schedule.keys())
    result: dict = {}
    for y in years:
        if y <= sorted_keys[0]:
            result[y] = schedule[sorted_keys[0]]
        elif y >= sorted_keys[-1]:
            result[y] = schedule[sorted_keys[-1]]
        elif y in schedule:
            result[y] = schedule[y]
        else:
            for i in range(len(sorted_keys) - 1):
                y_lo, y_hi = sorted_keys[i], sorted_keys[i + 1]
                if y_lo <= y <= y_hi:
                    frac = (y - y_lo) / (y_hi - y_lo)
                    result[y] = schedule[y_lo] + frac * (schedule[y_hi] - schedule[y_lo])
                    break
    return result


# ---------------------------------------------------------------------------
# Tech helpers
# ---------------------------------------------------------------------------
def load_generation_techs(tech_types_path: Path) -> set:
    if not tech_types_path.is_file():
        raise FileNotFoundError(f"TECH_TYPES.csv not found at {tech_types_path}")
    df = pd.read_csv(tech_types_path)
    return set(df.loc[df[TECH_TYPES_CATEGORY_COL] == GENERATION_CATEGORY,
                      TECH_TYPES_TECH_COL].dropna())


def expand_tech_pattern(pattern: str, cr: str, gen_techs: set) -> list:
    if pattern.endswith("*"):
        prefix = pattern[:-1]
        matched = [t for t in gen_techs
                   if t.startswith(prefix) and len(t) == PWR_TECH_LENGTH
                   and t[COUNTRY_REGION_SLICE] == cr]
    else:
        matched = [pattern] if pattern in gen_techs else []
    return sorted(matched)


def is_exempt(tech: str, exempt_patterns: list) -> bool:
    """Check if a tech matches any exempt pattern."""
    for pat in exempt_patterns:
        if pat.endswith("*"):
            if tech.startswith(pat[:-1]):
                return True
        elif tech == pat:
            return True
    return False


def fuel_code(tech: str, fuel_slice: slice) -> str:
    """Extract fuel code from tech name. PWRCOAINDEA → COA"""
    return tech[fuel_slice].upper()


# ---------------------------------------------------------------------------
# Backup / restore
# ---------------------------------------------------------------------------
def _rmtree_robust(path: Path, attempts: int = 5) -> None:
    for i in range(attempts):
        try:
            shutil.rmtree(path)
            return
        except PermissionError:
            gc.collect()
            time.sleep(0.1 * (i + 1))
    shutil.rmtree(path)


def find_latest_backup(input_dir: Path) -> Path | None:
    parent = input_dir.parent
    candidates = sorted(
        (p for p in parent.iterdir()
         if p.is_dir() and p.name.startswith(f"{input_dir.name}{BACKUP_TAG}")),
        key=lambda p: p.name,
    )
    return candidates[-1] if candidates else None


def restore_from_backup(input_dir: Path, backup_dir: Path | None = None) -> Path:
    input_dir = Path(input_dir)
    if backup_dir is None:
        backup_dir = find_latest_backup(input_dir)
        if backup_dir is None:
            raise FileNotFoundError(f"No {BACKUP_TAG}* backup found next to {input_dir}.")
    else:
        backup_dir = Path(backup_dir)
        if not backup_dir.is_dir():
            raise FileNotFoundError(f"Backup folder does not exist: {backup_dir}")
    if input_dir.is_dir():
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        snapshot = input_dir.parent / f"{input_dir.name}_POST_RETIRE_pre_restore_{stamp}"
        if not snapshot.exists():
            shutil.copytree(input_dir, snapshot)
        _rmtree_robust(input_dir)
    shutil.copytree(backup_dir, input_dir)
    return backup_dir


def make_backup(input_dir: Path) -> Path:
    if not input_dir.is_dir():
        raise FileNotFoundError(f"Input directory does not exist: {input_dir}")
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = input_dir.parent / f"{input_dir.name}{BACKUP_TAG}{stamp}"
    if backup.exists():
        raise FileExistsError(f"Backup folder already exists: {backup}")
    shutil.copytree(input_dir, backup)
    return backup


# ---------------------------------------------------------------------------
# Worksheet helpers
# ---------------------------------------------------------------------------
def find_year_columns(ws) -> dict:
    year_to_col = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if isinstance(val, int) and 1900 <= val <= 2200:
            year_to_col[val] = col_idx
    return year_to_col


def find_named_columns(ws, names) -> dict:
    found = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val in names:
            found[val] = col_idx
    return found


def values_differ(a, b, tol: float = 1e-12) -> bool:
    if a is None and b is None:
        return False
    if a is None or b is None:
        return True
    try:
        return abs(float(a) - float(b)) > tol
    except (TypeError, ValueError):
        return a != b


# ---------------------------------------------------------------------------
# Core logic
# ---------------------------------------------------------------------------
def apply_retirement(ws, config: dict, gen_techs: set,
                     year_cols: dict) -> dict:
    """Apply age-based and scheduled retirement to ResidualCapacity rows."""
    base_year = config["base_year"]
    fuel_sl = config["fuel_slice"]
    age_based = config["age_based"]
    scheduled = config["scheduled"]
    exempt = config["exempt"]

    headers = find_named_columns(ws, ["Tech", "Parameter", PROJ_MODE_COL])
    tech_col = headers.get("Tech")
    param_col = headers.get("Parameter")
    proj_mode_col = headers.get(PROJ_MODE_COL)
    if tech_col is None or param_col is None:
        raise ValueError(f"Sheet '{ws.title}' missing Tech/Parameter columns.")

    sorted_years = sorted(year_cols.keys())

    log = {
        "sheet": ws.title,
        "years": sorted_years,
        "base_year": base_year,
        "age_based_rules": {k: v for k, v in age_based.items()},
        "scheduled_count": len(scheduled),
        "exempt_patterns": exempt,
        "changes": [],
        "exempt_skipped": [],
        "non_gen_skipped": [],
        "no_rule_skipped": [],
        "projection_mode_flips": [],
    }

    # --- STEP A: Compute age-based retirement for all eligible techs ---
    # Build a map: {tech: {year: proposed_rescap}} from age-based rules.
    # Then scheduled overrides will replace entries in this map.
    proposed: dict = {}  # {tech: {year: value}}

    for row_idx in range(2, ws.max_row + 1):
        tech = ws.cell(row=row_idx, column=tech_col).value
        param = ws.cell(row=row_idx, column=param_col).value
        if tech is None or param != RES_PARAM:
            continue
        if tech not in gen_techs:
            if tech not in log["non_gen_skipped"]:
                log["non_gen_skipped"].append(tech)
            continue
        if len(tech) != PWR_TECH_LENGTH:
            continue
        if is_exempt(tech, exempt):
            if tech not in log["exempt_skipped"]:
                log["exempt_skipped"].append(tech)
            continue

        # Read base-year ResidualCapacity
        base_col = year_cols.get(base_year)
        if base_col is None:
            # No base year column → skip age-based, may still get scheduled
            continue
        base_val = ws.cell(row=row_idx, column=base_col).value
        base_cap = float(base_val) if base_val is not None else 0.0

        if base_cap <= 0:
            continue

        # Determine fuel category
        fc = fuel_code(tech, fuel_sl)
        if fc not in age_based:
            if tech not in log["no_rule_skipped"]:
                log["no_rule_skipped"].append(tech)
            continue

        rule = age_based[fc]
        lifetime = rule["lifetime_years"]
        profile_fn = PROFILE_FUNCS[rule["retirement_profile"]]

        tech_sched: dict = {}
        for y in sorted_years:
            elapsed = y - base_year
            if elapsed <= 0:
                tech_sched[y] = base_cap
            else:
                frac = profile_fn(elapsed, lifetime)
                tech_sched[y] = base_cap * (1.0 - frac)
        proposed[tech] = tech_sched

    # --- STEP B: Scheduled overrides ---
    scheduled_techs: set = set()
    for entry in scheduled:
        cr = entry["cr"]
        tech_pattern = entry["tech"]
        schedule = entry["schedule"]

        matched = expand_tech_pattern(tech_pattern, cr, gen_techs)
        matched = [t for t in matched if not is_exempt(t, exempt)]

        full_sched = interpolate_schedule(schedule, sorted_years)

        for tech in matched:
            proposed[tech] = {y: full_sched.get(y, 0.0) for y in sorted_years}
            scheduled_techs.add(tech)

    # --- STEP C: Safety constraints and write-back ---
    for row_idx in range(2, ws.max_row + 1):
        tech = ws.cell(row=row_idx, column=tech_col).value
        param = ws.cell(row=row_idx, column=param_col).value
        if tech is None or param != RES_PARAM or tech not in proposed:
            continue

        tech_sched = proposed[tech]
        row_modified = False
        prev_value = None

        for y in sorted_years:
            col = year_cols[y]
            cell = ws.cell(row=row_idx, column=col)
            old = cell.value
            new_val = tech_sched.get(y, 0.0)

            # Non-negative floor
            new_val = max(0.0, new_val)

            # Monotonicity: cannot increase year-over-year
            if prev_value is not None and new_val > prev_value:
                new_val = prev_value

            prev_value = new_val

            if values_differ(old, new_val):
                cell.value = new_val
                row_modified = True
                mechanism = "scheduled" if tech in scheduled_techs else "age_based"
                old_f = float(old) if old is not None else None
                base_col = year_cols.get(base_year)
                base_v = ws.cell(row=row_idx, column=base_col).value if base_col else None
                base_cap = float(base_v) if base_v is not None else 0.0
                retired_frac = (1.0 - new_val / base_cap) if base_cap > 0 else 0.0
                log["changes"].append({
                    "tech": tech, "year": y, "param": RES_PARAM,
                    "old": old_f, "new": round(new_val, 6),
                    "mechanism": mechanism,
                    "retired_fraction": round(retired_frac, 4),
                })

        if row_modified and proj_mode_col is not None:
            mode_cell = ws.cell(row=row_idx, column=proj_mode_col)
            if mode_cell.value == PROJ_MODE_EMPTY:
                mode_cell.value = PROJ_MODE_USER
                log["projection_mode_flips"].append({"tech": tech})

    return log


def edit_parametrization(filepath: Path, sheets: list, config: dict,
                         gen_techs: set) -> dict:
    wb = load_workbook(filepath)
    file_log = {"file": str(filepath), "sheets": []}

    try:
        for sheet in sheets:
            if sheet not in wb.sheetnames:
                file_log["sheets"].append(
                    {"sheet": sheet, "skipped": "sheet not present in workbook"}
                )
                continue
            ws = wb[sheet]
            year_cols = find_year_columns(ws)
            if not year_cols:
                file_log["sheets"].append(
                    {"sheet": sheet, "skipped": "no integer year columns found"}
                )
                continue
            sheet_log = apply_retirement(ws, config, gen_techs, year_cols)
            file_log["sheets"].append(sheet_log)
        wb.save(filepath)
    finally:
        wb.close()

    return file_log


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------
def run(input_dir, sheets: list = None, skip_backup: bool = False,
        yaml_path: Path | None = None) -> dict:
    input_dir = Path(input_dir)
    sheets = sheets or DEFAULT_TARGET_SHEETS

    if yaml_path is None:
        yaml_path = Path(__file__).resolve().parent / YAML_FILE_NAME
    if not yaml_path.is_file():
        raise FileNotFoundError(
            f"YAML config not found at {yaml_path}. "
            f"Create {YAML_FILE_NAME} next to this script."
        )
    config = load_config(yaml_path)

    tech_types_path = resolve_paths().interconnector_taxonomy
    gen_techs = load_generation_techs(tech_types_path)

    backup_dir = None if skip_backup else make_backup(input_dir)

    paramfile = input_dir / PARAM_FILE_NAME
    if not paramfile.exists():
        raise FileNotFoundError(f"{paramfile} not found")

    log = edit_parametrization(paramfile, sheets, config, gen_techs)
    log["backup_dir"] = str(backup_dir) if backup_dir else None
    log["timestamp"] = datetime.now().isoformat()
    log["config_summary"] = {
        "base_year": config["base_year"],
        "age_based_fuels": list(config["age_based"].keys()),
        "scheduled_count": len(config["scheduled"]),
        "exempt_patterns": config["exempt"],
    }
    log["gen_techs_count"] = len(gen_techs)

    if backup_dir is not None:
        log_path = backup_dir.parent / f"{backup_dir.name}_CHANGES.json"
        log_path.write_text(json.dumps(log, indent=2, default=str))
        log["log_path"] = str(log_path)

    return log


# ---------------------------------------------------------------------------
# Console output
# ---------------------------------------------------------------------------
def print_summary(log: dict) -> None:
    bar = "=" * 72
    cfg = log.get("config_summary", {})
    print(bar)
    print("set_retirement_schedule — ResidualCapacity drawdown applied")
    print(bar)
    print(f"Backup folder      : {log.get('backup_dir', '(skipped)')}")
    print(f"Edited file        : {log['file']}")
    print(f"Base year          : {cfg.get('base_year', '?')}")
    print(f"Age-based fuels    : {', '.join(cfg.get('age_based_fuels', []))}")
    print(f"Scheduled entries  : {cfg.get('scheduled_count', 0)}")
    print(f"Exempt patterns    : {', '.join(cfg.get('exempt_patterns', []))}")
    print(f"GENERATION techs   : {log.get('gen_techs_count', '?')}")

    print()
    for s in log["sheets"]:
        if "skipped" in s:
            print(f"[SKIPPED] '{s['sheet']}': {s['skipped']}")
            continue
        print(f"Sheet: '{s['sheet']}'")
        years = s.get("years", [])
        if years:
            print(f"  Years: {years[0]}..{years[-1]} ({len(years)} years)")

        changes = s.get("changes", [])
        from collections import Counter
        mech_counts = Counter(c.get("mechanism", "?") for c in changes)
        print(f"  Cells modified: {len(changes)}")
        for mech, count in sorted(mech_counts.items()):
            print(f"    - {mech:20s} : {count}")

        exempt_s = s.get("exempt_skipped", [])
        no_rule = s.get("no_rule_skipped", [])
        print(f"  Exempt techs skipped : {len(exempt_s)}"
              + (f" ({', '.join(exempt_s[:5])}{'...' if len(exempt_s) > 5 else ''})"
                 if exempt_s else ""))
        print(f"  No-rule techs skipped: {len(no_rule)}"
              + (f" ({', '.join(no_rule[:5])}{'...' if len(no_rule) > 5 else ''})"
                 if no_rule else ""))

        # Show per-tech summary: base_cap → final_cap (retired %)
        if changes:
            tech_summary: dict = {}
            for c in changes:
                tech = c["tech"]
                if tech not in tech_summary:
                    tech_summary[tech] = {"first_year": c["year"], "last_year": c["year"],
                                          "first_new": c["new"], "last_new": c["new"],
                                          "mechanism": c["mechanism"]}
                tech_summary[tech]["last_year"] = c["year"]
                tech_summary[tech]["last_new"] = c["new"]

            print(f"\n  Per-tech retirement summary:")
            for tech, info in sorted(tech_summary.items()):
                print(f"    {tech} ({info['mechanism']}):"
                      f"  {info['first_year']}={info['first_new']:.2f}"
                      f" → {info['last_year']}={info['last_new']:.2f}")

        flips = s.get("projection_mode_flips", [])
        if flips:
            print(f"  Projection.Mode flips: {len(flips)}")

    if log.get("log_path"):
        print(f"\nDetailed change log: {log['log_path']}")


# ---------------------------------------------------------------------------
# Self-test
# ---------------------------------------------------------------------------
def run_self_test() -> int:
    bar = "=" * 72
    print(bar)
    print("set_retirement_schedule.py — SELF-TEST")
    print(bar)

    passed = 0
    failed = 0
    total_tests = 5

    techs_all = ["PWRCOAINDEA", "PWRGASBGDXX", "PWRPETINDSO",
                 "PWRHYDNPLXX", "PWRSPVINDEA"]
    rescap_base = {
        "PWRCOAINDEA": 20.0,
        "PWRGASBGDXX": 5.0,
        "PWRPETINDSO": 2.0,
        "PWRHYDNPLXX": 8.0,
        "PWRSPVINDEA": 3.0,
    }
    years = [2023, 2025, 2030, 2035, 2040, 2045, 2050]

    def _build_workbook(tmpdir: Path) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = "Secondary Techs"
        headers = ["Tech", "Parameter", "Projection.Mode"] + years
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        row = 2
        for tech in techs_all:
            ws.cell(row=row, column=1, value=tech)
            ws.cell(row=row, column=2, value="ResidualCapacity")
            ws.cell(row=row, column=3, value="EMPTY")
            for ci, y in enumerate(years, 4):
                ws.cell(row=row, column=ci, value=rescap_base[tech])
            row += 1
        path = tmpdir / "input_dir" / PARAM_FILE_NAME
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        wb.close()
        return path

    def _build_tech_types(tmpdir: Path) -> Path:
        path = tmpdir / "TECH_TYPES.csv"
        lines = [f"{TECH_TYPES_CATEGORY_COL},{TECH_TYPES_TECH_COL}"]
        for t in techs_all:
            lines.append(f"{GENERATION_CATEGORY},{t}")
        path.write_text("\n".join(lines))
        return path

    def _build_yaml(tmpdir: Path, content: dict) -> Path:
        path = tmpdir / YAML_FILE_NAME
        lines = []
        lines.append(f"base_year: {content.get('base_year', 2023)}")
        fs = content.get("fuel_slice", [3, 6])
        lines.append(f"fuel_slice: [{fs[0]}, {fs[1]}]")

        ab = content.get("age_based", {})
        if ab:
            lines.append("age_based:")
            for fuel, params in ab.items():
                lines.append(f"  {fuel}:")
                lines.append(f"    lifetime_years: {params['lifetime_years']}")
                lines.append(f"    retirement_profile: \"{params['retirement_profile']}\"")
        else:
            lines.append("age_based: {}")

        sched = content.get("scheduled", [])
        if sched:
            lines.append("scheduled:")
            for e in sched:
                lines.append(f"  - cr: \"{e['cr']}\"")
                lines.append(f"    tech: \"{e['tech']}\"")
                s_str = ", ".join(f"{y}: {v}" for y, v in sorted(e["schedule"].items()))
                lines.append(f"    schedule: {{{s_str}}}")
                if e.get("note"):
                    lines.append(f"    note: \"{e['note']}\"")
        else:
            lines.append("scheduled: []")

        exempt = content.get("exempt", [])
        if exempt:
            lines.append("exempt:")
            for p in exempt:
                lines.append(f"  - \"{p}\"")
        else:
            lines.append("exempt: []")

        path.write_text("\n".join(lines))
        return path

    def _read_rescap(filepath: Path) -> dict:
        wb = load_workbook(filepath, data_only=True)
        ws = wb["Secondary Techs"]
        yc = find_year_columns(ws)
        hdr = find_named_columns(ws, ["Tech", "Parameter"])
        result: dict = {}
        for row_idx in range(2, ws.max_row + 1):
            t = ws.cell(row=row_idx, column=hdr["Tech"]).value
            p = ws.cell(row=row_idx, column=hdr["Parameter"]).value
            if p != RES_PARAM:
                continue
            for year, col in yc.items():
                val = ws.cell(row=row_idx, column=col).value
                if val is not None:
                    result[(t, year)] = float(val)
        wb.close()
        return result

    # ======================================================================
    # TEST 1: Linear age-based retirement (coal, lifetime=35)
    # ======================================================================
    print("\nTest 1 — Linear age-based retirement (coal, lifetime=35)")
    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)
        wb_path = _build_workbook(tmpdir)
        tt_path = _build_tech_types(tmpdir)
        yaml_path = _build_yaml(tmpdir, {
            "base_year": 2023,
            "age_based": {
                "COA": {"lifetime_years": 35, "retirement_profile": "linear"},
                "GAS": {"lifetime_years": 30, "retirement_profile": "linear"},
                "PET": {"lifetime_years": 25, "retirement_profile": "linear"},
            },
            "exempt": ["PWRHYD*", "PWRSPV*"],
        })

        config = load_config(yaml_path)
        gen_techs = load_generation_techs(tt_path)
        edit_parametrization(wb_path, DEFAULT_TARGET_SHEETS, config, gen_techs)

        result = _read_rescap(wb_path)
        test_ok = True
        base = 20.0

        checks = {
            2023: base * (1.0 - 0 / 35),       # 20.0
            2030: base * (1.0 - 7 / 35),       # 16.0
            2040: base * (1.0 - 17 / 35),      # 10.2857...
            2050: base * (1.0 - 27 / 35),      # 4.5714...
        }
        for y, exp in checks.items():
            actual = result.get(("PWRCOAINDEA", y))
            if actual is None or abs(actual - exp) > 0.05:
                print(f"  FAIL: PWRCOAINDEA {y}: expected ~{exp:.2f}, got {actual}")
                test_ok = False

        # Gas: PWRGASBGDXX, lifetime=30, base=5.0
        gas_2050 = result.get(("PWRGASBGDXX", 2050))
        gas_exp = 5.0 * (1.0 - 27 / 30)
        if gas_2050 is None or abs(gas_2050 - gas_exp) > 0.05:
            print(f"  FAIL: PWRGASBGDXX 2050: expected ~{gas_exp:.2f}, got {gas_2050}")
            test_ok = False

        if test_ok:
            print("  PASS")
            passed += 1
        else:
            failed += 1

    # ======================================================================
    # TEST 2: Scheduled override replaces age-based
    # ======================================================================
    print("\nTest 2 — Scheduled override replaces age-based")
    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)
        wb_path = _build_workbook(tmpdir)
        tt_path = _build_tech_types(tmpdir)
        yaml_path = _build_yaml(tmpdir, {
            "base_year": 2023,
            "age_based": {
                "COA": {"lifetime_years": 35, "retirement_profile": "linear"},
            },
            "scheduled": [
                {"cr": "INDEA", "tech": "PWRCOA*",
                 "schedule": {2023: 20.0, 2030: 15.0, 2050: 0.0}},
            ],
            "exempt": ["PWRHYD*", "PWRSPV*"],
        })

        config = load_config(yaml_path)
        gen_techs = load_generation_techs(tt_path)
        edit_parametrization(wb_path, DEFAULT_TARGET_SHEETS, config, gen_techs)

        result = _read_rescap(wb_path)
        test_ok = True

        # 2030: scheduled = 15.0 (not age-based 16.0)
        v = result.get(("PWRCOAINDEA", 2030))
        if v is None or abs(v - 15.0) > 0.01:
            print(f"  FAIL: 2030 expected 15.0 (scheduled), got {v}")
            test_ok = False

        # 2040: interpolated between 2030=15 and 2050=0 → 7.5
        v = result.get(("PWRCOAINDEA", 2040))
        if v is None or abs(v - 7.5) > 0.01:
            print(f"  FAIL: 2040 expected 7.5 (interpolated), got {v}")
            test_ok = False

        # 2050: 0.0
        v = result.get(("PWRCOAINDEA", 2050))
        if v is None or abs(v - 0.0) > 0.01:
            print(f"  FAIL: 2050 expected 0.0, got {v}")
            test_ok = False

        if test_ok:
            print("  PASS")
            passed += 1
        else:
            failed += 1

    # ======================================================================
    # TEST 3: Exempt techs untouched
    # ======================================================================
    print("\nTest 3 — Exempt techs untouched")
    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)
        wb_path = _build_workbook(tmpdir)
        tt_path = _build_tech_types(tmpdir)
        yaml_path = _build_yaml(tmpdir, {
            "base_year": 2023,
            "age_based": {
                "COA": {"lifetime_years": 35, "retirement_profile": "linear"},
            },
            "exempt": ["PWRHYD*", "PWRSPV*"],
        })

        config = load_config(yaml_path)
        gen_techs = load_generation_techs(tt_path)
        edit_parametrization(wb_path, DEFAULT_TARGET_SHEETS, config, gen_techs)

        result = _read_rescap(wb_path)
        test_ok = True

        for y in years:
            v_hyd = result.get(("PWRHYDNPLXX", y))
            if v_hyd is None or abs(v_hyd - 8.0) > 0.001:
                print(f"  FAIL: PWRHYDNPLXX {y} expected 8.0 (exempt), got {v_hyd}")
                test_ok = False
            v_spv = result.get(("PWRSPVINDEA", y))
            if v_spv is None or abs(v_spv - 3.0) > 0.001:
                print(f"  FAIL: PWRSPVINDEA {y} expected 3.0 (exempt), got {v_spv}")
                test_ok = False

        if test_ok:
            print("  PASS")
            passed += 1
        else:
            failed += 1

    # ======================================================================
    # TEST 4: Monotonicity enforcement
    # ======================================================================
    print("\nTest 4 — Monotonicity enforcement")
    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)
        wb_path = _build_workbook(tmpdir)
        tt_path = _build_tech_types(tmpdir)
        # Scheduled override with a non-monotonic dip then rise
        yaml_path = _build_yaml(tmpdir, {
            "base_year": 2023,
            "scheduled": [
                {"cr": "INDEA", "tech": "PWRCOA*",
                 "schedule": {2023: 20.0, 2030: 10.0, 2035: 15.0, 2050: 5.0}},
            ],
            "exempt": ["PWRHYD*", "PWRSPV*"],
        })

        config = load_config(yaml_path)
        gen_techs = load_generation_techs(tt_path)
        edit_parametrization(wb_path, DEFAULT_TARGET_SHEETS, config, gen_techs)

        result = _read_rescap(wb_path)
        test_ok = True

        # 2030=10.0, 2035 would be 15.0 but monotonicity clamps to 10.0
        v_2035 = result.get(("PWRCOAINDEA", 2035))
        if v_2035 is None or abs(v_2035 - 10.0) > 0.01:
            print(f"  FAIL: 2035 expected 10.0 (clamped by monotonicity), got {v_2035}")
            test_ok = False

        # Subsequent years also clamped: 2040 interpolated from schedule
        # would be between 15 and 5 = 10.0, but monotonicity from 2035=10.0 caps it
        v_2040 = result.get(("PWRCOAINDEA", 2040))
        if v_2040 is not None and v_2040 > 10.0 + 0.01:
            print(f"  FAIL: 2040 should be <= 10.0 (monotonicity), got {v_2040}")
            test_ok = False

        if test_ok:
            print("  PASS")
            passed += 1
        else:
            failed += 1

    # ======================================================================
    # TEST 5: Non-negative floor
    # ======================================================================
    print("\nTest 5 — Non-negative floor")
    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)
        wb_path = _build_workbook(tmpdir)
        tt_path = _build_tech_types(tmpdir)
        # PET lifetime=25, base=2.0, by 2050 (27 years elapsed) frac>1.0
        # Should clamp to 0.0
        yaml_path = _build_yaml(tmpdir, {
            "base_year": 2023,
            "age_based": {
                "PET": {"lifetime_years": 25, "retirement_profile": "linear"},
            },
            "exempt": ["PWRHYD*", "PWRSPV*"],
        })

        config = load_config(yaml_path)
        gen_techs = load_generation_techs(tt_path)
        edit_parametrization(wb_path, DEFAULT_TARGET_SHEETS, config, gen_techs)

        result = _read_rescap(wb_path)
        test_ok = True

        # 2050: 27 years elapsed, lifetime=25 → frac = 27/25 = 1.08 → clamped to 1.0
        # ResCap = 2.0 * (1 - 1.0) = 0.0
        v = result.get(("PWRPETINDSO", 2050))
        if v is None or abs(v - 0.0) > 0.001:
            print(f"  FAIL: PWRPETINDSO 2050 expected 0.0 (non-neg floor), got {v}")
            test_ok = False

        # Also check 2045: 22 years, frac=22/25=0.88, rescap=2.0*0.12=0.24
        v_2045 = result.get(("PWRPETINDSO", 2045))
        exp_2045 = 2.0 * (1.0 - 22 / 25)
        if v_2045 is None or abs(v_2045 - exp_2045) > 0.01:
            print(f"  FAIL: PWRPETINDSO 2045 expected ~{exp_2045:.3f}, got {v_2045}")
            test_ok = False

        if test_ok:
            print("  PASS")
            passed += 1
        else:
            failed += 1

    # ======================================================================
    print()
    print(bar)
    if failed == 0:
        print(f"SELF-TEST PASSED ({passed}/{total_tests})")
    else:
        print(f"SELF-TEST FAILED ({passed}/{total_tests} passed, {failed} failed)")
    print(bar)
    return 0 if failed == 0 else 1


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> int:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("--input-dir", type=Path,
                        default=Path("A1_Outputs/A1_Outputs_BAU"))
    parser.add_argument("--sheets", nargs="+", default=DEFAULT_TARGET_SHEETS)
    parser.add_argument("--skip-backup", action="store_true")
    parser.add_argument("--self-test", action="store_true")
    parser.add_argument("--restore", action="store_true")
    parser.add_argument("--restore-from", type=Path, default=None)
    parser.add_argument("--yaml", type=Path, default=None)
    args = parser.parse_args()

    if args.self_test:
        return run_self_test()

    if args.restore or args.restore_from is not None:
        try:
            used = restore_from_backup(args.input_dir, args.restore_from)
        except Exception as exc:
            print(f"ERROR: {exc}", file=sys.stderr)
            return 1
        print(f"Restored {args.input_dir} from {used}")
        return 0

    try:
        log = run(args.input_dir, args.sheets, args.skip_backup, args.yaml)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    print_summary(log)
    return 0


if __name__ == "__main__":
    sys.exit(main())
