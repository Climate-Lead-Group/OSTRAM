"""
apply_internal_transmission.py
==============================

WS-3 D5. Calibrates the INTERNAL (intra-node) transmission families in the
model's `A-O_Parametrization.xlsx`, making `Config_country_codes.yaml` +
a desk-checked per-node residuals file the source of truth for:

    * ResidualCapacity   -> Demand Techs            (year-indexed rows)
    * CapitalCost         -> Demand Techs            (year-indexed rows)
    * FixedCost           -> Demand Techs            (year-indexed rows)
    * OperationalLife     -> Fixed Horizon Parameters (single 'Value' cell)

for the six internal-transmission families (10 nodes each = 60 techs):

    RE     : RNWTRN, RNWNLI, RNWRPO   (renewable-carrying lines)
    non-RE : PWRTRN, TRNNLI, TRNRPO

Background / why a late A3 stage
--------------------------------
A2_AddTx injects these families flat from the YAML (CapEx 100, FOM 4, ResCap 5,
life 20) into `Demand Techs` / `Fixed Horizon Parameters`. But:
  * ResidualCapacity is a flat 5 GW placeholder for every node (Maldives ~=
    India-West), which is not physical.
  * OperationalLife is rewritten to 50/20 by the Stage-1 template merge
    (3_update_ao_from_extensions.py), so a YAML/snapshot life edit does NOT
    stick.
  * There is no RE-vs-non-RE cost differentiation.

This stage runs LAST (after stage 5 and the interconnector-cost stage), so its
writes are authoritative and cannot be clobbered downstream — the same pattern
proven by apply_interconnector_costs.py. It leaves the flat post-A2 snapshot
untouched and does the calibration transparently in the delivered workbook.

Values applied
--------------
  * ResidualCapacity (existing stock, base year, held flat across the horizon):
        RNWTRN<node> = residuals[node].RNWTRN   (RE available at peak x margin)
        PWRTRN<node> = residuals[node].PWRTRN   (peak x margin - RE x margin)
        RNWNLI/RNWRPO/TRNNLI/TRNRPO = 0          (new-build / repower mechanisms)
    Per-node values come from the frozen, desk-checked
    `internal_tx_residuals.csv` (peak x 1.2).
  * CapitalCost / FixedCost (uniform across nodes — intra-node tx is accounting;
    the study is about interties):
        non-RE (PWRTRN/TRNNLI/TRNRPO) = base_capital_cost / base_fixed_cost
        RE     (RNWTRN/RNWNLI/RNWRPO) = base x re_capex_multiplier
    Basis: per-kW RE transmission premium ~1.5-2.3x (LBNL 2019); default 2.0.
    The multiplier is a single exposed YAML knob so WS-1 can slide it.
  * OperationalLife = operational_life (default 40) for all 6 families.

Scope (intentionally narrow): the four params above, on the six internal
families only. Interconnectors (13-char TRN*****), DSPTRN, generators (PWR***),
storage, losses, and the Max*/Investment caps are all left untouched.

OUTPUT
------
1. Timestamped backup of the input directory (unless --skip-backup).
2. In-place edit of A-O_Parametrization.xlsx.
3. A JSON change log next to the backup.
4. A desk-check CSV (old -> new for every cell) when --desk-check-csv is given
   or on --dry-run (which writes the CSV and makes NO workbook edits).

USAGE
-----
    python apply_internal_transmission.py --input-dir A1_Outputs/A1_Outputs_BAU
    python apply_internal_transmission.py --input-dir <dir> --dry-run \\
        --desk-check-csv desk_check.csv          # preview only, no mutation
    python apply_internal_transmission.py --self-test
    python apply_internal_transmission.py --input-dir <dir> --restore
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import shutil
import sys
from datetime import datetime
from pathlib import Path

import yaml
from openpyxl import load_workbook, Workbook

# --------------------------------------------------------------------------- config
PARAM_FILE_NAME = "A-O_Parametrization.xlsx"
DEFAULT_TARGET_SHEETS = ["Demand Techs", "Fixed Horizon Parameters"]

# Where ResidualCapacity / CapitalCost / FixedCost live (year-indexed) vs.
# OperationalLife (scalar). Both sheets are scanned; each param is written
# wherever its row is found, with the layout auto-detected per sheet.
YEAR_PARAMS = ("ResidualCapacity", "CapitalCost", "FixedCost")
SCALAR_PARAMS = ("OperationalLife",)
APPLIED_PARAMS = YEAR_PARAMS + SCALAR_PARAMS

RE_FAMILIES = ("RNWTRN", "RNWNLI", "RNWRPO")
NONRE_FAMILIES = ("PWRTRN", "TRNNLI", "TRNRPO")
ALL_FAMILIES = RE_FAMILIES + NONRE_FAMILIES
# Only the "Existing" families carry base-year residual stock; the rest are
# repower/new-line build mechanisms and start at 0.
RESIDUAL_FAMILY = {"RNWTRN": "RNWTRN", "PWRTRN": "PWRTRN"}

PROJ_MODE_COL = "Projection.Mode"
PROJ_MODE_EMPTY = "EMPTY"
PROJ_MODE_USER = "User defined"
VALUE_COL = "Value"

# Defaults mirror the YAML block; used only if the YAML omits a key.
DEFAULTS = {
    "base_capital_cost": 100.0,
    "base_fixed_cost": 4.0,
    "re_capex_multiplier": 2.0,
    "operational_life": 40,
}

BACKUP_TAG = "_PRE_INTERNAL_TX_"
CONFIG_NAME = "Config_country_codes.yaml"
RESIDUALS_NAME = "internal_tx_residuals.csv"


def _is_internal_tx(tech) -> bool:
    """True for the 60 internal-transmission techs (family(6) + node(5) = 11).
    Excludes interconnectors (13-char), DSPTRN, generators, storage."""
    return isinstance(tech, str) and len(tech) == 11 and tech[:6] in ALL_FAMILIES


def _family_of(tech: str) -> str:
    return tech[:6]


def _node_of(tech: str) -> str:
    return tech[6:]


# --------------------------------------------------------------------------- sources
def resolve_config(explicit: Path | None) -> Path:
    """Locate Config_country_codes.yaml (knobs). Priority: --config, env, canonical."""
    candidates = []
    if explicit:
        candidates.append(Path(explicit))
    env = os.environ.get("OSTRAM_CONFIG_PATH")
    if env:
        candidates.append(Path(env))
    # rules_scripts/ -> A3_process/ -> t1_confection/Config_country_codes.yaml
    candidates.append(Path(__file__).resolve().parent.parent.parent / CONFIG_NAME)
    for c in candidates:
        if c and c.is_file():
            return c
    raise FileNotFoundError(
        "Could not locate Config_country_codes.yaml. Tried: "
        + ", ".join(str(c) for c in candidates)
    )


def resolve_residuals(explicit: Path | None) -> Path:
    if explicit:
        p = Path(explicit)
        if p.is_file():
            return p
    p = Path(__file__).resolve().parent / RESIDUALS_NAME
    if p.is_file():
        return p
    raise FileNotFoundError(
        f"Could not locate {RESIDUALS_NAME}. Tried --residuals and {p}"
    )


def read_config(config_path: Path) -> dict:
    """Read the internal_transmission knob block; fall back to DEFAULTS per key."""
    with open(config_path, "r", encoding="utf-8") as fh:
        data = yaml.safe_load(fh) or {}
    blk = data.get("internal_transmission", {}) or {}
    knobs = dict(DEFAULTS)
    for k in DEFAULTS:
        if k in blk and blk[k] is not None:
            knobs[k] = blk[k]
    return knobs


def read_residuals(residuals_path: Path) -> dict:
    """Return {node: {'RNWTRN': gw, 'PWRTRN': gw}} from the frozen CSV.
    Comment lines starting with '#' are skipped."""
    out: dict = {}
    with open(residuals_path, newline="", encoding="utf-8") as f:
        rows = [r for r in csv.reader(f) if r and not str(r[0]).startswith("#")]
    if not rows:
        raise ValueError(f"{residuals_path} has no data rows")
    hdr = [h.strip() for h in rows[0]]
    ni, ri, pi = hdr.index("node"), hdr.index("RNWTRN"), hdr.index("PWRTRN")
    for r in rows[1:]:
        if len(r) <= max(ni, ri, pi) or not r[ni].strip():
            continue
        out[r[ni].strip()] = {"RNWTRN": float(r[ri]), "PWRTRN": float(r[pi])}
    return out


def build_target_values(knobs: dict, residuals: dict, nodes: list) -> dict:
    """Return {tech: {param: value}} for all 60 internal-tx techs."""
    base_cap = float(knobs["base_capital_cost"])
    base_fom = float(knobs["base_fixed_cost"])
    mult = float(knobs["re_capex_multiplier"])
    life = knobs["operational_life"]

    out: dict = {}
    for fam in ALL_FAMILIES:
        is_re = fam in RE_FAMILIES
        cap = base_cap * (mult if is_re else 1.0)
        fom = base_fom * (mult if is_re else 1.0)
        for node in nodes:
            tech = f"{fam}{node}"
            vals = {
                "CapitalCost": cap,
                "FixedCost": fom,
                "OperationalLife": life,
            }
            if fam in RESIDUAL_FAMILY:
                key = RESIDUAL_FAMILY[fam]
                vals["ResidualCapacity"] = float(residuals.get(node, {}).get(key, 0.0))
            else:
                vals["ResidualCapacity"] = 0.0
            out[tech] = vals
    return out


# --------------------------------------------------------------------------- worksheet helpers
def find_year_columns(ws) -> dict:
    out = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if isinstance(val, int) and 1900 <= val <= 2200:
            out[val] = col_idx
    return out


def find_named_columns(ws, names) -> dict:
    found = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val in names:
            found[val] = col_idx
    return found


def _values_differ(a, b, tol: float = 1e-9) -> bool:
    if a is None and b is None:
        return False
    if a is None or b is None:
        return True
    try:
        return abs(float(a) - float(b)) > tol
    except (TypeError, ValueError):
        return a != b


# --------------------------------------------------------------------------- core edit
def apply_values(ws, target_values: dict, seen: set, dry_run: bool) -> dict:
    """Overwrite internal-tx cells in one sheet. Layout auto-detected: year-indexed
    (Demand Techs) writes every year cell; scalar (Fixed Horizon Parameters) writes
    the single 'Value' cell. `seen` collects (tech, param) written anywhere."""
    hdr = find_named_columns(ws, ["Tech", "Parameter", PROJ_MODE_COL, VALUE_COL])
    tech_c, par_c = hdr.get("Tech"), hdr.get("Parameter")
    pm_c, val_c = hdr.get(PROJ_MODE_COL), hdr.get(VALUE_COL)
    if tech_c is None or par_c is None:
        return {"sheet": ws.title, "skipped": "no Tech/Parameter columns"}
    year_cols = find_year_columns(ws)
    if year_cols:
        layout = "year"
    elif val_c is not None:
        layout = "scalar"
    else:
        return {"sheet": ws.title, "skipped": "no year columns and no Value column"}

    log = {"sheet": ws.title, "layout": layout, "changes": [], "rows_touched": [],
           "pm_flips": []}
    years = sorted(year_cols)

    for row_idx in range(2, ws.max_row + 1):
        tech = ws.cell(row=row_idx, column=tech_c).value
        par = ws.cell(row=row_idx, column=par_c).value
        if not _is_internal_tx(tech):
            continue
        if par not in APPLIED_PARAMS or tech not in target_values or par not in target_values[tech]:
            continue
        seen.add((tech, par))
        new_val = target_values[tech][par]
        row_changed = False
        if layout == "year":
            for y in years:
                cell = ws.cell(row=row_idx, column=year_cols[y])
                if _values_differ(cell.value, new_val):
                    log["changes"].append({"tech": tech, "param": par, "year": y,
                                            "old": cell.value, "new": new_val})
                    if not dry_run:
                        cell.value = new_val
                    row_changed = True
        else:  # scalar
            cell = ws.cell(row=row_idx, column=val_c)
            if _values_differ(cell.value, new_val):
                log["changes"].append({"tech": tech, "param": par, "year": None,
                                        "old": cell.value, "new": new_val})
                if not dry_run:
                    cell.value = new_val
                row_changed = True
        if row_changed:
            log["rows_touched"].append({"tech": tech, "param": par, "new": new_val})
            if pm_c is not None and not dry_run:
                pm = ws.cell(row=row_idx, column=pm_c)
                if pm.value in (None, "", PROJ_MODE_EMPTY):
                    pm.value = PROJ_MODE_USER
                    log["pm_flips"].append({"tech": tech, "param": par})
    return log


def edit_parametrization(filepath: Path, sheets: list, target_values: dict,
                         dry_run: bool) -> dict:
    wb = load_workbook(filepath)
    file_log = {"file": str(filepath), "sheets": []}
    seen: set = set()
    try:
        for sheet in sheets:
            if sheet not in wb.sheetnames:
                file_log["sheets"].append({"sheet": sheet, "skipped": "not present"})
                continue
            file_log["sheets"].append(apply_values(wb[sheet], target_values, seen, dry_run))
        if not dry_run:
            wb.save(filepath)
    finally:
        wb.close()
    missing = [{"tech": t, "param": p}
               for t, params in target_values.items() for p in params
               if (t, p) not in seen]
    file_log["missing_pairs"] = missing
    return file_log


def write_desk_check_csv(file_log: dict, path: Path) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["sheet", "tech", "family", "node", "parameter", "year",
                    "old_value", "new_value"])
        for s in file_log.get("sheets", []):
            for c in s.get("changes", []):
                t = c["tech"]
                w.writerow([s["sheet"], t, _family_of(t), _node_of(t), c["param"],
                            c.get("year"), c.get("old"), c.get("new")])


# --------------------------------------------------------------------------- backup / restore
def make_backup(input_dir: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = input_dir.parent / f"{input_dir.name}{BACKUP_TAG}{stamp}"
    if backup.exists():
        raise FileExistsError(f"Backup already exists: {backup}")
    shutil.copytree(input_dir, backup)
    return backup


def find_latest_backup(input_dir: Path) -> Path | None:
    parent = input_dir.parent
    cands = sorted((p for p in parent.iterdir()
                    if p.is_dir() and p.name.startswith(f"{input_dir.name}{BACKUP_TAG}")),
                   key=lambda p: p.name)
    return cands[-1] if cands else None


def restore_from_backup(input_dir: Path, backup_dir: Path | None = None) -> Path:
    input_dir = Path(input_dir)
    backup_dir = Path(backup_dir) if backup_dir else find_latest_backup(input_dir)
    if not backup_dir or not backup_dir.is_dir():
        raise FileNotFoundError(f"No {BACKUP_TAG}* backup found next to {input_dir}.")
    if input_dir.is_dir():
        shutil.rmtree(input_dir)
    shutil.copytree(backup_dir, input_dir)
    return backup_dir


# --------------------------------------------------------------------------- orchestration
def run(input_dir, sheets=None, skip_backup=False, config=None, residuals=None,
        dry_run=False, desk_check_csv=None, nodes=None) -> dict:
    input_dir = Path(input_dir)
    sheets = sheets or DEFAULT_TARGET_SHEETS
    config_path = resolve_config(config)
    residuals_path = resolve_residuals(residuals)
    knobs = read_config(config_path)
    resid = read_residuals(residuals_path)
    node_list = nodes or sorted(resid.keys())
    target_values = build_target_values(knobs, resid, node_list)

    paramfile = input_dir / PARAM_FILE_NAME
    if not paramfile.exists():
        raise FileNotFoundError(f"{paramfile} not found")

    backup_dir = None
    if not dry_run and not skip_backup:
        backup_dir = make_backup(input_dir)

    file_log = edit_parametrization(paramfile, sheets, target_values, dry_run)
    file_log.update({
        "config": str(config_path),
        "residuals": str(residuals_path),
        "knobs": knobs,
        "nodes": node_list,
        "dry_run": dry_run,
        "backup_dir": str(backup_dir) if backup_dir else None,
        "timestamp": datetime.now().isoformat(),
        "families": {"RE": list(RE_FAMILIES), "non_RE": list(NONRE_FAMILIES)},
        "applied_params": list(APPLIED_PARAMS),
    })
    if desk_check_csv:
        write_desk_check_csv(file_log, Path(desk_check_csv))
        file_log["desk_check_csv"] = str(desk_check_csv)
    if backup_dir is not None:
        log_path = backup_dir.parent / f"{backup_dir.name}_CHANGES.json"
        log_path.write_text(json.dumps(file_log, indent=2, default=str))
        file_log["log_path"] = str(log_path)
    return file_log


def print_summary(log: dict) -> None:
    bar = "=" * 72
    print(bar)
    print("apply_internal_transmission — YAML + residuals -> model"
          + ("  [DRY RUN]" if log.get("dry_run") else ""))
    print(bar)
    print(f"Config (knobs)   : {log.get('config')}")
    print(f"Residuals        : {log.get('residuals')}")
    print(f"Knobs            : {log.get('knobs')}")
    print(f"Backup folder    : {log.get('backup_dir', '(skipped)')}")
    for s in log.get("sheets", []):
        if "skipped" in s:
            print(f"[SKIPPED] '{s['sheet']}': {s['skipped']}")
            continue
        print(f"Sheet '{s['sheet']}' ({s['layout']}): {len(s['changes'])} cells, "
              f"{len(s['rows_touched'])} rows, {len(s.get('pm_flips', []))} proj-mode flips")
    missing = log.get("missing_pairs", [])
    if missing:
        print(f"Sourced pairs with NO row in any target sheet ({len(missing)}): "
              + ", ".join(f"{m['tech']}/{m['param']}" for m in missing[:20])
              + (" ..." if len(missing) > 20 else ""))
    if log.get("desk_check_csv"):
        print(f"Desk-check CSV   : {log['desk_check_csv']}")
    if log.get("log_path"):
        print(f"Change log       : {log['log_path']}")


# --------------------------------------------------------------------------- self-test
def run_self_test() -> int:
    import tempfile
    bar = "=" * 72
    print(bar); print("apply_internal_transmission.py — SELF-TEST"); print(bar)
    ok = True
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        # synthetic config
        cfg = {
            "internal_transmission": {
                "base_capital_cost": 100, "base_fixed_cost": 4,
                "re_capex_multiplier": 2.0, "operational_life": 40,
            }
        }
        cfg_path = td / CONFIG_NAME
        cfg_path.write_text(yaml.safe_dump(cfg), encoding="utf-8")
        # synthetic residuals (2 nodes)
        resid_path = td / RESIDUALS_NAME
        resid_path.write_text(
            "# comment line\nnode,RNWTRN,PWRTRN,status_flags\n"
            "INDWE,19.552,68.075,ok\nMDVXX,0.001,0.295,SMALL_GRID\n",
            encoding="utf-8")

        # synthetic A-O_Parametrization: Demand Techs (year) + FHP (scalar)
        idir = td / "A1_Outputs" / "A1_Outputs_BAU"; idir.mkdir(parents=True)
        pwb = Workbook()
        dt = pwb.active; dt.title = "Demand Techs"
        for c, h in enumerate(["Tech", "Parameter", PROJ_MODE_COL, 2023, 2024, 2025], 1):
            dt.cell(row=1, column=c, value=h)
        dt_rows = [
            ("RNWTRNINDWE", "ResidualCapacity", "User defined", 5, 5, 5),   # -> 19.552
            ("RNWTRNINDWE", "CapitalCost", "User defined", 100, 100, 100),  # RE -> 200
            ("RNWTRNINDWE", "FixedCost", "User defined", 4, 4, 4),          # RE -> 8
            ("PWRTRNMDVXX", "ResidualCapacity", "User defined", 5, 5, 5),   # -> 0.295
            ("PWRTRNMDVXX", "CapitalCost", "User defined", 100, 100, 100),  # non-RE -> 100 (unchanged)
            ("RNWNLIINDWE", "ResidualCapacity", "EMPTY", "", "", ""),       # NLI -> 0
            ("DSPTRNINDWE", "ResidualCapacity", "User defined", 9999, 9999, 9999),  # must NOT change
            ("PWRSPVINDWE", "CapitalCost", "User defined", 560, 560, 560),  # non-tx, must NOT change
        ]
        for ri, row in enumerate(dt_rows, 2):
            for c, v in enumerate(row, 1):
                dt.cell(row=ri, column=c, value=v)
        fhp = pwb.create_sheet("Fixed Horizon Parameters")
        for c, h in enumerate(["Tech.Type", "Tech.ID", "Tech", "Parameter", "Unit", VALUE_COL], 1):
            fhp.cell(row=1, column=c, value=h)
        fhp_rows = [
            ("Demand", 1, "RNWTRNINDWE", "OperationalLife", "years", 50),   # -> 40
            ("Demand", 2, "PWRTRNMDVXX", "OperationalLife", "years", 20),   # -> 40
            ("Demand", 3, "RNWTRNINDWE", "CapacityToActivityUnit", "", 31.536),  # must NOT change
            ("Demand", 4, "DSPTRNINDWE", "OperationalLife", "years", 20),   # must NOT change
        ]
        for ri, row in enumerate(fhp_rows, 2):
            for c, v in enumerate(row, 1):
                fhp.cell(row=ri, column=c, value=v)
        pwb.save(idir / PARAM_FILE_NAME); pwb.close()

        run(idir, skip_backup=True, config=cfg_path, residuals=resid_path)

        chk = load_workbook(idir / PARAM_FILE_NAME, data_only=True)
        dt = chk["Demand Techs"]; got = {}
        for r in range(2, dt.max_row + 1):
            got[(dt.cell(row=r, column=1).value, dt.cell(row=r, column=2).value)] = \
                [dt.cell(row=r, column=c).value for c in (4, 5, 6)]
        fhp = chk["Fixed Horizon Parameters"]; got_f = {}
        for r in range(2, fhp.max_row + 1):
            got_f[(fhp.cell(row=r, column=3).value, fhp.cell(row=r, column=4).value)] = \
                fhp.cell(row=r, column=6).value
        chk.close()

        def ck(cond, label):
            nonlocal ok
            print(("  PASS " if cond else "  FAIL ") + label)
            ok = ok and cond
        ck(got.get(("RNWTRNINDWE", "ResidualCapacity")) == [19.552, 19.552, 19.552], "RNWTRN residual 5 -> per-node 19.552")
        ck(got.get(("PWRTRNMDVXX", "ResidualCapacity")) == [0.295, 0.295, 0.295], "PWRTRN residual 5 -> per-node 0.295")
        ck(got.get(("RNWNLIINDWE", "ResidualCapacity")) == [0.0, 0.0, 0.0], "RNWNLI residual empty -> 0")
        ck(got.get(("RNWTRNINDWE", "CapitalCost")) == [200.0, 200.0, 200.0], "RE CapitalCost 100 -> 200 (2x)")
        ck(got.get(("RNWTRNINDWE", "FixedCost")) == [8.0, 8.0, 8.0], "RE FixedCost 4 -> 8 (2x)")
        ck(got.get(("PWRTRNMDVXX", "CapitalCost")) == [100.0, 100.0, 100.0], "non-RE CapitalCost stays 100")
        ck(got.get(("DSPTRNINDWE", "ResidualCapacity")) == [9999, 9999, 9999], "DSPTRN UNCHANGED")
        ck(got.get(("PWRSPVINDWE", "CapitalCost")) == [560, 560, 560], "non-tx PWRSPV UNCHANGED")
        ck(got_f.get(("RNWTRNINDWE", "OperationalLife")) == 40, "RE life 50 -> 40")
        ck(got_f.get(("PWRTRNMDVXX", "OperationalLife")) == 40, "non-RE life 20 -> 40")
        ck(got_f.get(("RNWTRNINDWE", "CapacityToActivityUnit")) == 31.536, "CapacityToActivityUnit UNCHANGED")
        ck(got_f.get(("DSPTRNINDWE", "OperationalLife")) == 20, "DSPTRN life UNCHANGED")

        # idempotency: a second run should change nothing
        log2 = run(idir, skip_backup=True, config=cfg_path, residuals=resid_path)
        n2 = sum(len(s.get("changes", [])) for s in log2["sheets"])
        ck(n2 == 0, f"idempotent (2nd run: {n2} changes)")
    print(bar); print("SELF-TEST", "PASSED" if ok else "FAILED"); print(bar)
    return 0 if ok else 1


# --------------------------------------------------------------------------- main
def main() -> int:
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--input-dir", type=Path, default=Path("A1_Outputs/A1_Outputs_BAU"))
    p.add_argument("--config", type=Path, default=None, help="Config_country_codes.yaml")
    p.add_argument("--residuals", type=Path, default=None, help="internal_tx_residuals.csv")
    p.add_argument("--sheets", nargs="+", default=DEFAULT_TARGET_SHEETS)
    p.add_argument("--skip-backup", action="store_true")
    p.add_argument("--dry-run", action="store_true",
                   help="compute + write the desk-check CSV but do NOT edit the workbook")
    p.add_argument("--desk-check-csv", type=Path, default=None,
                   help="write an old->new CSV of every intended cell change")
    p.add_argument("--self-test", action="store_true")
    p.add_argument("--restore", action="store_true")
    p.add_argument("--restore-from", type=Path, default=None)
    args = p.parse_args()

    if args.self_test:
        return run_self_test()
    if args.restore or args.restore_from is not None:
        try:
            used = restore_from_backup(args.input_dir, args.restore_from)
        except Exception as exc:
            print(f"ERROR: {exc}", file=sys.stderr); return 1
        print(f"Restored {args.input_dir} from {used}"); return 0
    try:
        log = run(args.input_dir, args.sheets, args.skip_backup, args.config,
                  args.residuals, args.dry_run, args.desk_check_csv)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr); return 1
    print_summary(log)
    return 0


if __name__ == "__main__":
    sys.exit(main())
