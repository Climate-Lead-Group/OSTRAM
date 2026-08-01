"""
set_min_capacity_floors.py
==========================

Independent patch — writes minimum capacity/generation floors and
optional maximum ceilings from exogenous policy documents (CCDRs,
national energy plans, IRPs) into the parametrization workbook.

All target values come from a user-curated YAML file. Unlike
set_vre_targets.py this script has NO dependency on prior model runs —
it is purely a "policy injection" layer.

RULE
----
For each entry in the YAML (floors and ceilings):
    Expand the tech pattern + cr to matching techs in TECH_TYPES.csv.
    Interpolate the schedule to all horizon years.
    Write the value into the appropriate parameter row:
        min_capacity  → TotalAnnualMinCapacityInvestment (GW)
        min_activity  → TotalTechnologyAnnualActivityLowerLimit (PJ)
        max_capacity  → TotalAnnualMaxCapacityInvestment (GW) [tighten only]
        max_activity  → TotalTechnologyAnnualActivityUpperLimit (PJ)

    For min_capacity: untie rule ensures MaxCapInv > MinCapInv.
    For max_capacity: only writes if value < current MaxCapInv (never loosens).

CONFIGURATION
-------------
Edit bau_calibration.yaml (next to this script).

USAGE
-----
    python set_min_capacity_floors.py
    python set_min_capacity_floors.py --self-test
    python set_min_capacity_floors.py --restore
"""

from __future__ import annotations

import argparse
import gc
import json
import shutil
import sys
import tempfile
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
from ostram.paths import resolve_paths
from openpyxl import load_workbook, Workbook

try:
    import yaml as _yaml
    def _load_yaml(path: Path) -> dict:
        with open(path, "r", encoding="utf-8") as f:
            return _yaml.safe_load(f)
except ImportError:
    _yaml = None
    def _load_yaml(path: Path) -> dict:
        raise ImportError("PyYAML is required. Install with: pip install pyyaml")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
DEFAULT_TARGET_SHEETS = ["Secondary Techs"]
PARAM_FILE_NAME = "A-O_Parametrization.xlsx"
YAML_FILE_NAME = "bau_calibration.yaml"

PARAM_MAP = {
    "min_capacity": "TotalAnnualMinCapacityInvestment",
    "min_activity": "TotalTechnologyAnnualActivityLowerLimit",
    "max_capacity": "TotalAnnualMaxCapacityInvestment",
    "max_activity": "TotalTechnologyAnnualActivityUpperLimit",
}

MAX_INV_PARAM = "TotalAnnualMaxCapacityInvestment"
MIN_INV_PARAM = "TotalAnnualMinCapacityInvestment"

PROJ_MODE_COL = "Projection.Mode"
PROJ_MODE_EMPTY = "EMPTY"
PROJ_MODE_USER = "User defined"

TECH_TYPES_FILE = "TECH_TYPES.csv"
TECH_TYPES_CATEGORY_COL = "Technology (PWR)"
TECH_TYPES_TECH_COL = "Technology"
GENERATION_CATEGORY = "GENERATION"

PWR_TECH_LENGTH = 11
COUNTRY_REGION_SLICE = slice(6, 11)

UNTIE_MULTIPLIER = 1.01

BACKUP_TAG = "_PRE_BAU_CAL_"


# ---------------------------------------------------------------------------
# YAML config loader
# ---------------------------------------------------------------------------
def load_config(yaml_path: Path) -> dict:
    """Load and validate bau_calibration.yaml."""
    cfg = _load_yaml(yaml_path)
    if cfg is None:
        cfg = {}

    warn_on_untie = bool(cfg.get("warn_on_untie", True))

    def _parse_entries(raw: list | None, section: str) -> list:
        if not raw:
            return []
        entries = []
        for i, entry in enumerate(raw):
            cr = str(entry.get("cr", ""))
            tech = str(entry.get("tech", ""))
            param = str(entry.get("param", ""))
            schedule = entry.get("schedule", {})
            note = str(entry.get("note", ""))
            if not cr or not tech or not param or not schedule:
                raise ValueError(
                    f"{section}[{i}] missing cr/tech/param/schedule: {entry}"
                )
            if param not in PARAM_MAP:
                raise ValueError(
                    f"{section}[{i}] invalid param '{param}'. "
                    f"Expected one of: {list(PARAM_MAP.keys())}"
                )
            schedule = {int(y): float(v) for y, v in schedule.items()}
            entries.append({
                "cr": cr, "tech": tech, "param": param,
                "schedule": schedule, "note": note,
            })
        return entries

    floors = _parse_entries(cfg.get("floors"), "floors")
    ceilings = _parse_entries(cfg.get("ceilings"), "ceilings")

    return {
        "floors": floors,
        "ceilings": ceilings,
        "warn_on_untie": warn_on_untie,
    }


# ---------------------------------------------------------------------------
# Interpolation
# ---------------------------------------------------------------------------
def interpolate_schedule(schedule: dict, years: list,
                         pre_first: float = 0.0) -> dict:
    """Linearly interpolate a sparse {year: value} schedule.

    Before first key: pre_first. After last key: hold last value.
    """
    if not schedule:
        return {y: pre_first for y in years}
    sorted_keys = sorted(schedule.keys())
    result: dict = {}
    for y in years:
        if y < sorted_keys[0]:
            result[y] = pre_first
        elif y >= sorted_keys[-1]:
            result[y] = schedule[sorted_keys[-1]]
        elif y in schedule:
            result[y] = schedule[y]
        else:
            for i in range(len(sorted_keys) - 1):
                y_lo, y_hi = sorted_keys[i], sorted_keys[i + 1]
                if y_lo <= y <= y_hi:
                    frac = (y - y_lo) / (y_hi - y_lo)
                    result[y] = schedule[y_lo] + frac * (schedule[y_hi] - schedule[y_lo])
                    break
    return result


# ---------------------------------------------------------------------------
# Tech-type loading and pattern matching
# ---------------------------------------------------------------------------
def load_generation_techs(tech_types_path: Path) -> set:
    if not tech_types_path.is_file():
        raise FileNotFoundError(f"TECH_TYPES.csv not found at {tech_types_path}")
    df = pd.read_csv(tech_types_path)
    return set(df.loc[df[TECH_TYPES_CATEGORY_COL] == GENERATION_CATEGORY,
                      TECH_TYPES_TECH_COL].dropna())


def expand_tech_pattern(pattern: str, cr: str, gen_techs: set) -> list:
    """Expand 'PWRSPV*' + 'INDEA' → ['PWRSPVINDEA'] if in gen_techs."""
    if pattern.endswith("*"):
        prefix = pattern[:-1]
        matched = [t for t in gen_techs
                   if t.startswith(prefix) and len(t) == PWR_TECH_LENGTH
                   and t[COUNTRY_REGION_SLICE] == cr]
    else:
        matched = [pattern] if pattern in gen_techs else []
    return sorted(matched)


# ---------------------------------------------------------------------------
# Backup / restore
# ---------------------------------------------------------------------------
def _rmtree_robust(path: Path, attempts: int = 5) -> None:
    for i in range(attempts):
        try:
            shutil.rmtree(path)
            return
        except PermissionError:
            gc.collect()
            time.sleep(0.1 * (i + 1))
    shutil.rmtree(path)


def find_latest_backup(input_dir: Path) -> Path | None:
    parent = input_dir.parent
    candidates = sorted(
        (p for p in parent.iterdir()
         if p.is_dir() and p.name.startswith(f"{input_dir.name}{BACKUP_TAG}")),
        key=lambda p: p.name,
    )
    return candidates[-1] if candidates else None


def restore_from_backup(input_dir: Path, backup_dir: Path | None = None) -> Path:
    input_dir = Path(input_dir)
    if backup_dir is None:
        backup_dir = find_latest_backup(input_dir)
        if backup_dir is None:
            raise FileNotFoundError(f"No {BACKUP_TAG}* backup found next to {input_dir}.")
    else:
        backup_dir = Path(backup_dir)
        if not backup_dir.is_dir():
            raise FileNotFoundError(f"Backup folder does not exist: {backup_dir}")
    if input_dir.is_dir():
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        snapshot = input_dir.parent / f"{input_dir.name}_POST_BAU_CAL_pre_restore_{stamp}"
        if not snapshot.exists():
            shutil.copytree(input_dir, snapshot)
        _rmtree_robust(input_dir)
    shutil.copytree(backup_dir, input_dir)
    return backup_dir


def make_backup(input_dir: Path) -> Path:
    if not input_dir.is_dir():
        raise FileNotFoundError(f"Input directory does not exist: {input_dir}")
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = input_dir.parent / f"{input_dir.name}{BACKUP_TAG}{stamp}"
    if backup.exists():
        raise FileExistsError(f"Backup folder already exists: {backup}")
    shutil.copytree(input_dir, backup)
    return backup


# ---------------------------------------------------------------------------
# Worksheet helpers
# ---------------------------------------------------------------------------
def find_year_columns(ws) -> dict:
    year_to_col = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if isinstance(val, int) and 1900 <= val <= 2200:
            year_to_col[val] = col_idx
    return year_to_col


def find_named_columns(ws, names) -> dict:
    found = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val in names:
            found[val] = col_idx
    return found


def values_differ(a, b, tol: float = 1e-12) -> bool:
    if a is None and b is None:
        return False
    if a is None or b is None:
        return True
    try:
        return abs(float(a) - float(b)) > tol
    except (TypeError, ValueError):
        return a != b


def find_param_row(ws, tech: str, param: str, tech_col: int,
                   param_col: int) -> int | None:
    for row_idx in range(2, ws.max_row + 1):
        t = ws.cell(row=row_idx, column=tech_col).value
        p = ws.cell(row=row_idx, column=param_col).value
        if t == tech and p == param:
            return row_idx
    return None


def find_or_create_param_row(ws, tech: str, param: str, tech_col: int,
                             param_col: int, proj_mode_col: int | None) -> int:
    existing = find_param_row(ws, tech, param, tech_col, param_col)
    if existing is not None:
        return existing
    new_row = ws.max_row + 1
    ws.cell(row=new_row, column=tech_col, value=tech)
    ws.cell(row=new_row, column=param_col, value=param)
    if proj_mode_col is not None:
        ws.cell(row=new_row, column=proj_mode_col, value=PROJ_MODE_EMPTY)
    return new_row


# ---------------------------------------------------------------------------
# Core logic
# ---------------------------------------------------------------------------
def apply_entries(ws, entries: list, section_name: str,
                  gen_techs: set, year_cols: dict,
                  warn_on_untie: bool) -> dict:
    """Apply a list of floor or ceiling entries to the worksheet.

    Returns a log dict fragment.
    """
    headers = find_named_columns(ws, ["Tech", "Parameter", PROJ_MODE_COL])
    tech_col = headers.get("Tech")
    param_col = headers.get("Parameter")
    proj_mode_col = headers.get(PROJ_MODE_COL)
    if tech_col is None or param_col is None:
        raise ValueError(f"Sheet '{ws.title}' missing Tech/Parameter columns.")

    sorted_years = sorted(year_cols.keys())

    log = {
        "section": section_name,
        "entries_processed": [],
        "changes": [],
        "warnings": [],
        "projection_mode_flips": [],
    }

    # Build a live MaxCapInv cache for untie and ceiling-tighten logic
    maxinv_cache: dict = {}
    for row_idx in range(2, ws.max_row + 1):
        t = ws.cell(row=row_idx, column=tech_col).value
        p = ws.cell(row=row_idx, column=param_col).value
        if p != MAX_INV_PARAM:
            continue
        for year, col in year_cols.items():
            val = ws.cell(row=row_idx, column=col).value
            if val is not None:
                try:
                    maxinv_cache[(t, year)] = float(val)
                except (TypeError, ValueError):
                    pass

    for entry in entries:
        cr = entry["cr"]
        tech_pattern = entry["tech"]
        param_key = entry["param"]
        schedule = entry["schedule"]
        note = entry.get("note", "")
        wb_param = PARAM_MAP[param_key]

        matched_techs = expand_tech_pattern(tech_pattern, cr, gen_techs)
        if not matched_techs:
            log["warnings"].append(
                f"No techs matched pattern '{tech_pattern}' in cr '{cr}'."
            )
            continue

        # For floors: 0 before first milestone. For ceilings: hold first value.
        is_ceiling = param_key.startswith("max_")
        pre_first = list(sorted(schedule.values()))[0] if (is_ceiling and schedule) else 0.0
        full_sched = interpolate_schedule(schedule, sorted_years, pre_first=pre_first)

        entry_log = {
            "cr": cr, "tech_pattern": tech_pattern, "param": param_key,
            "wb_param": wb_param, "matched_techs": matched_techs,
            "note": note,
            "schedule_interpolated": {str(y): v for y, v in full_sched.items()},
        }
        log["entries_processed"].append(entry_log)

        for tech in matched_techs:
            row = find_or_create_param_row(
                ws, tech, wb_param, tech_col, param_col, proj_mode_col
            )
            row_modified = False

            for year in sorted_years:
                col = year_cols[year]
                cell = ws.cell(row=row, column=col)
                old = cell.value
                proposed = full_sched.get(year, 0.0)

                # Ceiling logic: only tighten, never loosen
                if param_key == "max_capacity":
                    current_max = maxinv_cache.get((tech, year))
                    if current_max is not None and proposed >= current_max:
                        # Ceiling is above current → no change
                        log["changes"].append({
                            "tech": tech, "year": year, "param": wb_param,
                            "old": old, "new": old,
                            "proposed": proposed,
                            "reason": "ceiling_not_tighter",
                        })
                        continue

                if values_differ(old, proposed):
                    cell.value = proposed
                    row_modified = True
                    reason = "floor_write" if not is_ceiling else "ceiling_write"
                    log["changes"].append({
                        "tech": tech, "year": year, "param": wb_param,
                        "old": old, "new": proposed, "reason": reason,
                        "note": note,
                    })

                    # Update cache if we wrote MaxCapInv
                    if wb_param == MAX_INV_PARAM:
                        maxinv_cache[(tech, year)] = proposed

                # Untie rule for min_capacity: MaxCapInv > MinCapInv
                if param_key == "min_capacity" and proposed > 0:
                    current_max = maxinv_cache.get((tech, year))
                    if current_max is not None and current_max <= proposed:
                        new_max = proposed * UNTIE_MULTIPLIER
                        max_row = find_param_row(ws, tech, MAX_INV_PARAM,
                                                 tech_col, param_col)
                        if max_row is not None:
                            max_cell = ws.cell(row=max_row, column=year_cols[year])
                            max_cell.value = new_max
                            maxinv_cache[(tech, year)] = new_max
                            log["changes"].append({
                                "tech": tech, "year": year, "param": MAX_INV_PARAM,
                                "old": current_max, "new": new_max,
                                "reason": "untie_maxinv",
                            })
                            if warn_on_untie:
                                msg = (f"WARNING: {tech} {year}: MinCapInv={proposed:.3f} "
                                       f">= MaxCapInv={current_max:.3f} → bumped to "
                                       f"{new_max:.3f}")
                                log["warnings"].append(msg)

            # Flip Projection.Mode
            if row_modified and proj_mode_col is not None:
                mode_cell = ws.cell(row=row, column=proj_mode_col)
                if mode_cell.value == PROJ_MODE_EMPTY:
                    mode_cell.value = PROJ_MODE_USER
                    log["projection_mode_flips"].append(
                        {"tech": tech, "param": wb_param}
                    )

    return log


def edit_parametrization(filepath: Path, sheets: list, config: dict,
                         gen_techs: set) -> dict:
    """Apply BAU calibration floors + ceilings to the workbook."""
    wb = load_workbook(filepath)
    file_log = {"file": str(filepath), "sheets": []}

    try:
        for sheet in sheets:
            if sheet not in wb.sheetnames:
                file_log["sheets"].append(
                    {"sheet": sheet, "skipped": "sheet not present in workbook"}
                )
                continue
            ws = wb[sheet]
            year_cols = find_year_columns(ws)
            if not year_cols:
                file_log["sheets"].append(
                    {"sheet": sheet, "skipped": "no integer year columns found"}
                )
                continue

            sheet_log = {"sheet": sheet, "years": sorted(year_cols.keys()), "sections": []}

            # Process floors
            if config["floors"]:
                floor_log = apply_entries(
                    ws, config["floors"], "floors", gen_techs, year_cols,
                    config["warn_on_untie"]
                )
                sheet_log["sections"].append(floor_log)

            # Process ceilings
            if config["ceilings"]:
                ceil_log = apply_entries(
                    ws, config["ceilings"], "ceilings", gen_techs, year_cols,
                    config["warn_on_untie"]
                )
                sheet_log["sections"].append(ceil_log)

            file_log["sheets"].append(sheet_log)
        wb.save(filepath)
    finally:
        wb.close()

    return file_log


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------
def run(input_dir, sheets: list = None, skip_backup: bool = False,
        yaml_path: Path | None = None) -> dict:
    input_dir = Path(input_dir)
    sheets = sheets or DEFAULT_TARGET_SHEETS

    if yaml_path is None:
        yaml_path = Path(__file__).resolve().parent / YAML_FILE_NAME
    if not yaml_path.is_file():
        raise FileNotFoundError(
            f"YAML config not found at {yaml_path}. "
            f"Create {YAML_FILE_NAME} next to this script."
        )
    config = load_config(yaml_path)

    tech_types_path = resolve_paths().scenario_config_root / "technology_types.csv"
    gen_techs = load_generation_techs(tech_types_path)

    backup_dir = None if skip_backup else make_backup(input_dir)

    paramfile = input_dir / PARAM_FILE_NAME
    if not paramfile.exists():
        raise FileNotFoundError(f"{paramfile} not found")

    log = edit_parametrization(paramfile, sheets, config, gen_techs)
    log["backup_dir"] = str(backup_dir) if backup_dir else None
    log["timestamp"] = datetime.now().isoformat()
    log["config_summary"] = {
        "floors_count": len(config["floors"]),
        "ceilings_count": len(config["ceilings"]),
        "warn_on_untie": config["warn_on_untie"],
    }
    log["gen_techs_count"] = len(gen_techs)

    if backup_dir is not None:
        log_path = backup_dir.parent / f"{backup_dir.name}_CHANGES.json"
        log_path.write_text(json.dumps(log, indent=2, default=str))
        log["log_path"] = str(log_path)

    return log


# ---------------------------------------------------------------------------
# Console output
# ---------------------------------------------------------------------------
def print_summary(log: dict) -> None:
    bar = "=" * 72
    cfg = log.get("config_summary", {})
    print(bar)
    print("set_min_capacity_floors — BAU calibration floors/ceilings applied")
    print(bar)
    print(f"Backup folder    : {log.get('backup_dir', '(skipped)')}")
    print(f"Edited file      : {log['file']}")
    print(f"Floor entries    : {cfg.get('floors_count', 0)}")
    print(f"Ceiling entries  : {cfg.get('ceilings_count', 0)}")
    print(f"Warn on untie    : {cfg.get('warn_on_untie', True)}")
    print(f"GENERATION techs : {log.get('gen_techs_count', '?')}")

    print()
    for s in log["sheets"]:
        if "skipped" in s:
            print(f"[SKIPPED] '{s['sheet']}': {s['skipped']}")
            continue
        print(f"Sheet: '{s['sheet']}'")
        years = s.get("years", [])
        if years:
            print(f"  Years: {years[0]}..{years[-1]} ({len(years)} years)")

        for sec in s.get("sections", []):
            section_name = sec.get("section", "?")
            entries = sec.get("entries_processed", [])
            changes = sec.get("changes", [])
            warnings = sec.get("warnings", [])

            print(f"\n  [{section_name.upper()}]")
            for ep in entries:
                print(f"    {ep['cr']} / {ep['tech_pattern']} → "
                      f"{ep['wb_param']} — matched {len(ep['matched_techs'])} techs "
                      f"({', '.join(ep['matched_techs'])})")
                if ep.get("note"):
                    print(f"      note: {ep['note']}")

            from collections import Counter
            reason_counts = Counter(c.get("reason", "?") for c in changes)
            total_written = sum(1 for c in changes
                                if c.get("reason") not in ("ceiling_not_tighter",))
            print(f"    Cells written: {total_written}")
            for reason, count in sorted(reason_counts.items()):
                print(f"      - {reason:30s} : {count}")

            if warnings:
                for w in warnings:
                    print(f"    {w}")

            flips = sec.get("projection_mode_flips", [])
            if flips:
                print(f"    Projection.Mode flips: {len(flips)}")

    if log.get("log_path"):
        print(f"\nDetailed change log: {log['log_path']}")


# ---------------------------------------------------------------------------
# Self-test
# ---------------------------------------------------------------------------
def run_self_test() -> int:
    bar = "=" * 72
    print(bar)
    print("set_min_capacity_floors.py — SELF-TEST")
    print(bar)

    passed = 0
    failed = 0
    total_tests = 6

    techs = ["PWRSPVINDEA", "PWRCOAINDEA", "PWRGASBGDXX", "PWRHYDNPLXX"]
    years = [2025, 2030, 2035]

    def _build_workbook(tmpdir: Path,
                        maxinv_presets: dict | None = None) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = "Secondary Techs"
        headers = ["Tech", "Parameter", "Projection.Mode"] + years
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)

        row = 2
        for tech in techs:
            # ResidualCapacity
            ws.cell(row=row, column=1, value=tech)
            ws.cell(row=row, column=2, value="ResidualCapacity")
            ws.cell(row=row, column=3, value="EMPTY")
            row += 1
            # MinCapInv
            ws.cell(row=row, column=1, value=tech)
            ws.cell(row=row, column=2, value=MIN_INV_PARAM)
            ws.cell(row=row, column=3, value="EMPTY")
            row += 1
            # MaxCapInv
            ws.cell(row=row, column=1, value=tech)
            ws.cell(row=row, column=2, value=MAX_INV_PARAM)
            ws.cell(row=row, column=3, value="EMPTY")
            if maxinv_presets and tech in maxinv_presets:
                for ci, y in enumerate(years, 4):
                    preset = maxinv_presets[tech]
                    if isinstance(preset, dict):
                        ws.cell(row=row, column=ci, value=preset.get(y, None))
                    else:
                        ws.cell(row=row, column=ci, value=preset)
            row += 1

        path = tmpdir / "input_dir" / PARAM_FILE_NAME
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        wb.close()
        return path

    def _build_tech_types(tmpdir: Path) -> Path:
        path = tmpdir / "TECH_TYPES.csv"
        lines = [f"{TECH_TYPES_CATEGORY_COL},{TECH_TYPES_TECH_COL}"]
        for t in techs:
            lines.append(f"{GENERATION_CATEGORY},{t}")
        path.write_text("\n".join(lines))
        return path

    def _build_yaml(tmpdir: Path, content: dict) -> Path:
        path = tmpdir / YAML_FILE_NAME
        lines = []
        lines.append(f"warn_on_untie: {str(content.get('warn_on_untie', True)).lower()}")

        for section in ("floors", "ceilings"):
            entries = content.get(section, [])
            if not entries:
                lines.append(f"{section}: []")
                continue
            lines.append(f"{section}:")
            for e in entries:
                lines.append(f"  - cr: \"{e['cr']}\"")
                lines.append(f"    tech: \"{e['tech']}\"")
                lines.append(f"    param: \"{e['param']}\"")
                sched = ", ".join(f"{y}: {v}" for y, v in sorted(e["schedule"].items()))
                lines.append(f"    schedule: {{{sched}}}")
                if e.get("note"):
                    lines.append(f"    note: \"{e['note']}\"")

        path.write_text("\n".join(lines))
        return path

    def _read_param_values(filepath: Path, param_name: str) -> dict:
        wb = load_workbook(filepath, data_only=True)
        ws = wb["Secondary Techs"]
        yc = find_year_columns(ws)
        hdr = find_named_columns(ws, ["Tech", "Parameter"])
        result: dict = {}
        for row_idx in range(2, ws.max_row + 1):
            t = ws.cell(row=row_idx, column=hdr["Tech"]).value
            p = ws.cell(row=row_idx, column=hdr["Parameter"]).value
            if p != param_name:
                continue
            for year, col in yc.items():
                val = ws.cell(row=row_idx, column=col).value
                if val is not None:
                    result[(t, year)] = float(val)
        wb.close()
        return result

    # ======================================================================
    # TEST 1: Floor writes MinCapInv
    # ======================================================================
    print("\nTest 1 — Floor writes MinCapInv")
    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)
        wb_path = _build_workbook(tmpdir)
        tt_path = _build_tech_types(tmpdir)
        yaml_path = _build_yaml(tmpdir, {
            "floors": [
                {"cr": "BGDXX", "tech": "PWRGAS*", "param": "min_capacity",
                 "schedule": {2030: 3.5}},
            ],
        })

        config = load_config(yaml_path)
        gen_techs = load_generation_techs(tt_path)
        edit_parametrization(wb_path, DEFAULT_TARGET_SHEETS, config, gen_techs)

        result = _read_param_values(wb_path, MIN_INV_PARAM)
        test_ok = True

        # 2025: before milestone → 0
        v_2025 = result.get(("PWRGASBGDXX", 2025))
        if v_2025 is not None and abs(v_2025) > 0.001:
            print(f"  FAIL: 2025 expected 0 (pre-milestone), got {v_2025}")
            test_ok = False

        # 2030: 3.5
        v_2030 = result.get(("PWRGASBGDXX", 2030))
        if v_2030 is None or abs(v_2030 - 3.5) > 0.001:
            print(f"  FAIL: 2030 expected 3.5, got {v_2030}")
            test_ok = False

        # 2035: held flat at 3.5
        v_2035 = result.get(("PWRGASBGDXX", 2035))
        if v_2035 is None or abs(v_2035 - 3.5) > 0.001:
            print(f"  FAIL: 2035 expected 3.5 (held flat), got {v_2035}")
            test_ok = False

        if test_ok:
            print("  PASS")
            passed += 1
        else:
            failed += 1

    # ======================================================================
    # TEST 2: Floor triggers untie rule
    # ======================================================================
    print("\nTest 2 — Floor triggers untie rule")
    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)
        wb_path = _build_workbook(tmpdir, maxinv_presets={"PWRSPVINDEA": 2.0})
        tt_path = _build_tech_types(tmpdir)
        yaml_path = _build_yaml(tmpdir, {
            "warn_on_untie": True,
            "floors": [
                {"cr": "INDEA", "tech": "PWRSPV*", "param": "min_capacity",
                 "schedule": {2030: 5.0}},
            ],
        })

        config = load_config(yaml_path)
        gen_techs = load_generation_techs(tt_path)
        log = edit_parametrization(wb_path, DEFAULT_TARGET_SHEETS, config, gen_techs)

        mininv = _read_param_values(wb_path, MIN_INV_PARAM)
        maxinv = _read_param_values(wb_path, MAX_INV_PARAM)

        test_ok = True
        # 2030: MinCapInv = 5.0, MaxCapInv was 2.0 → bumped to 5.05
        v_min = mininv.get(("PWRSPVINDEA", 2030))
        if v_min is None or abs(v_min - 5.0) > 0.001:
            print(f"  FAIL: MinCapInv expected 5.0, got {v_min}")
            test_ok = False

        v_max = maxinv.get(("PWRSPVINDEA", 2030))
        expected_max = 5.0 * UNTIE_MULTIPLIER
        if v_max is None or abs(v_max - expected_max) > 0.001:
            print(f"  FAIL: MaxCapInv expected {expected_max}, got {v_max}")
            test_ok = False

        # Check warning
        warnings = []
        for s in log.get("sheets", []):
            for sec in s.get("sections", []):
                warnings.extend(sec.get("warnings", []))
        has_warning = any("WARNING" in w and "PWRSPVINDEA" in w for w in warnings)
        if not has_warning:
            print(f"  FAIL: expected untie WARNING")
            test_ok = False

        if test_ok:
            print("  PASS")
            passed += 1
        else:
            failed += 1

    # ======================================================================
    # TEST 3: Ceiling tightens MaxCapInv
    # ======================================================================
    print("\nTest 3 — Ceiling tightens MaxCapInv")
    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)
        # MaxCapInv for PWRCOAINDEA = 9999 (or empty → we use 9999)
        wb_path = _build_workbook(tmpdir, maxinv_presets={"PWRCOAINDEA": 9999})
        tt_path = _build_tech_types(tmpdir)
        yaml_path = _build_yaml(tmpdir, {
            "ceilings": [
                {"cr": "INDEA", "tech": "PWRCOA*", "param": "max_capacity",
                 "schedule": {2030: 1.0}},
            ],
        })

        config = load_config(yaml_path)
        gen_techs = load_generation_techs(tt_path)
        edit_parametrization(wb_path, DEFAULT_TARGET_SHEETS, config, gen_techs)

        maxinv = _read_param_values(wb_path, MAX_INV_PARAM)
        test_ok = True

        v = maxinv.get(("PWRCOAINDEA", 2030))
        if v is None or abs(v - 1.0) > 0.001:
            print(f"  FAIL: MaxCapInv expected 1.0 (tightened), got {v}")
            test_ok = False

        if test_ok:
            print("  PASS")
            passed += 1
        else:
            failed += 1

    # ======================================================================
    # TEST 4: Ceiling does not loosen
    # ======================================================================
    print("\nTest 4 — Ceiling does not loosen existing MaxCapInv")
    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)
        # MaxCapInv for PWRCOAINDEA: 0.5 at all years
        wb_path = _build_workbook(tmpdir, maxinv_presets={"PWRCOAINDEA": 0.5})
        tt_path = _build_tech_types(tmpdir)
        yaml_path = _build_yaml(tmpdir, {
            "ceilings": [
                {"cr": "INDEA", "tech": "PWRCOA*", "param": "max_capacity",
                 "schedule": {2035: 2.0}},
            ],
        })

        config = load_config(yaml_path)
        gen_techs = load_generation_techs(tt_path)
        edit_parametrization(wb_path, DEFAULT_TARGET_SHEETS, config, gen_techs)

        maxinv = _read_param_values(wb_path, MAX_INV_PARAM)
        test_ok = True

        # 2035: ceiling=2.0 but current=0.5, so stays 0.5
        v = maxinv.get(("PWRCOAINDEA", 2035))
        if v is None or abs(v - 0.5) > 0.001:
            print(f"  FAIL: MaxCapInv expected 0.5 (not loosened), got {v}")
            test_ok = False

        if test_ok:
            print("  PASS")
            passed += 1
        else:
            failed += 1

    # ======================================================================
    # TEST 5: Interpolation
    # ======================================================================
    print("\nTest 5 — Interpolation between milestones")
    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)
        wb_path = _build_workbook(tmpdir)
        tt_path = _build_tech_types(tmpdir)
        yaml_path = _build_yaml(tmpdir, {
            "floors": [
                {"cr": "NPLXX", "tech": "PWRHYD*", "param": "min_activity",
                 "schedule": {2025: 10.0, 2035: 30.0}},
            ],
        })

        config = load_config(yaml_path)
        gen_techs = load_generation_techs(tt_path)
        edit_parametrization(wb_path, DEFAULT_TARGET_SHEETS, config, gen_techs)

        result = _read_param_values(wb_path, PARAM_MAP["min_activity"])
        test_ok = True

        # 2025: 10.0
        v = result.get(("PWRHYDNPLXX", 2025))
        if v is None or abs(v - 10.0) > 0.001:
            print(f"  FAIL: 2025 expected 10.0, got {v}")
            test_ok = False

        # 2030: interpolated = 10 + (30-10) × 5/10 = 20.0
        v = result.get(("PWRHYDNPLXX", 2030))
        if v is None or abs(v - 20.0) > 0.01:
            print(f"  FAIL: 2030 expected 20.0 (interpolated), got {v}")
            test_ok = False

        # 2035: 30.0
        v = result.get(("PWRHYDNPLXX", 2035))
        if v is None or abs(v - 30.0) > 0.001:
            print(f"  FAIL: 2035 expected 30.0, got {v}")
            test_ok = False

        if test_ok:
            print("  PASS")
            passed += 1
        else:
            failed += 1

    # ======================================================================
    # TEST 6: Tech pattern expansion (cr filter)
    # ======================================================================
    print("\nTest 6 — Tech pattern expansion respects cr filter")
    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)
        # Add an extra tech to prove cr filtering works
        extra_techs = techs + ["PWRSPVBGDXX"]
        wb = Workbook()
        ws = wb.active
        ws.title = "Secondary Techs"
        headers = ["Tech", "Parameter", "Projection.Mode"] + years
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        row = 2
        for tech in extra_techs:
            ws.cell(row=row, column=1, value=tech)
            ws.cell(row=row, column=2, value=MIN_INV_PARAM)
            ws.cell(row=row, column=3, value="EMPTY")
            row += 1
        wb_path = tmpdir / "input_dir" / PARAM_FILE_NAME
        wb_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(wb_path)
        wb.close()

        # Tech types includes both PWRSPVINDEA and PWRSPVBGDXX
        tt_path = tmpdir / "TECH_TYPES.csv"
        lines = [f"{TECH_TYPES_CATEGORY_COL},{TECH_TYPES_TECH_COL}"]
        for t in extra_techs:
            lines.append(f"{GENERATION_CATEGORY},{t}")
        tt_path.write_text("\n".join(lines))

        yaml_path = _build_yaml(tmpdir, {
            "floors": [
                # cr=INDEA, so only PWRSPVINDEA should match, not PWRSPVBGDXX
                {"cr": "INDEA", "tech": "PWRSPV*", "param": "min_capacity",
                 "schedule": {2030: 7.0}},
            ],
        })

        config = load_config(yaml_path)
        gen_techs = load_generation_techs(tt_path)
        edit_parametrization(wb_path, DEFAULT_TARGET_SHEETS, config, gen_techs)

        result = _read_param_values(wb_path, MIN_INV_PARAM)
        test_ok = True

        # PWRSPVINDEA should have 7.0 at 2030
        v_indea = result.get(("PWRSPVINDEA", 2030))
        if v_indea is None or abs(v_indea - 7.0) > 0.001:
            print(f"  FAIL: PWRSPVINDEA 2030 expected 7.0, got {v_indea}")
            test_ok = False

        # PWRSPVBGDXX should NOT have been written
        v_bgd = result.get(("PWRSPVBGDXX", 2030))
        if v_bgd is not None:
            print(f"  FAIL: PWRSPVBGDXX should not be written, got {v_bgd}")
            test_ok = False

        if test_ok:
            print("  PASS")
            passed += 1
        else:
            failed += 1

    # ======================================================================
    # Summary
    # ======================================================================
    print()
    print(bar)
    if failed == 0:
        print(f"SELF-TEST PASSED ({passed}/{total_tests})")
    else:
        print(f"SELF-TEST FAILED ({passed}/{total_tests} passed, {failed} failed)")
    print(bar)
    return 0 if failed == 0 else 1


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> int:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("--input-dir", type=Path,
                        default=Path("A1_Outputs/A1_Outputs_BAU"))
    parser.add_argument("--sheets", nargs="+", default=DEFAULT_TARGET_SHEETS)
    parser.add_argument("--skip-backup", action="store_true")
    parser.add_argument("--self-test", action="store_true")
    parser.add_argument("--restore", action="store_true")
    parser.add_argument("--restore-from", type=Path, default=None)
    parser.add_argument("--yaml", type=Path, default=None)
    args = parser.parse_args()

    if args.self_test:
        return run_self_test()

    if args.restore or args.restore_from is not None:
        try:
            used = restore_from_backup(args.input_dir, args.restore_from)
        except Exception as exc:
            print(f"ERROR: {exc}", file=sys.stderr)
            return 1
        print(f"Restored {args.input_dir} from {used}")
        return 0

    try:
        log = run(args.input_dir, args.sheets, args.skip_backup, args.yaml)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    print_summary(log)
    return 0


if __name__ == "__main__":
    sys.exit(main())
