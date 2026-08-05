"""Synchronize country technology rows into a prepared scenario workbook."""

from __future__ import annotations

import argparse
from dataclasses import dataclass
import os
from pathlib import Path
import re
from typing import Iterable, Sequence

from openpyxl import load_workbook

from ostram.paths import resolve_paths
from ostram.profiles import PROFILE_WORKSPACE_ENV
from ostram.pipeline.scenarios.rules.set_interconnector_direction import (
    parse_tech_regions,
)


COUNTRY = re.compile(r"^[A-Z]{3}$")
SHEET_MAP = (
    ("Secondary Techs", "Secondary_Techs"),
    ("Primary Techs", "Primary_Techs"),
    ("Demand Techs", "Demand_Techs"),
    ("VariableCost", "VariableCost"),
    ("Capacities", "Capacities_CF"),
    ("Fixed Horizon Parameters", "Fixed_Horizon_Parameters"),
)


@dataclass(frozen=True)
class SyncChange:
    source_sheet: str
    target_sheet: str
    scenarios: tuple[str, ...]
    technologies: tuple[str, ...]
    rows_added: int


def _headers(worksheet) -> list[object]:
    return [cell.value for cell in worksheet[1]]


def _normalized_header(value: object) -> str:
    return str(value).strip().replace(",", ".") if value is not None else ""


def _unique_columns(headers: Iterable[object], *, sheet: str) -> dict[str, int]:
    result: dict[str, int] = {}
    for index, raw in enumerate(headers):
        name = _normalized_header(raw)
        if not name:
            continue
        if name in result:
            raise ValueError(f"{sheet} has duplicate normalized header {name!r}")
        result[name] = index
    return result


def technology_regions(technology: object) -> tuple[str, ...]:
    """Return structurally encoded five-character regions for a technology."""

    tech = str(technology or "").strip().upper()
    if tech.startswith("TRN") and len(tech) == 13:
        return tuple(parse_tech_regions(tech))
    if tech.startswith("PWR") and len(tech) >= 11:
        region = tech[6:11]
        if region[:3].isalpha() and region[3:].isalnum():
            return (region,)
    if len(tech) >= 5:
        region = tech[-5:]
        if region[:3].isalpha() and region[3:].isalnum():
            return (region,)
    return ()


def technology_belongs_to_country(technology: object, country: str) -> bool:
    return any(region[:3] == country for region in technology_regions(technology))


def _control_scenarios(workbook) -> tuple[str, ...]:
    if "Control" not in workbook.sheetnames:
        raise ValueError("scenario workbook is missing Control sheet")
    worksheet = workbook["Control"]
    columns = _unique_columns(_headers(worksheet), sheet="Control")
    scenario_header = next(
        (name for name in ("scenario", "Scenario") if name in columns), None
    )
    if scenario_header is None:
        raise ValueError("Control sheet is missing scenario header")
    scenarios: list[str] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        value = row[columns[scenario_header]]
        if value is not None and str(value).strip():
            scenarios.append(str(value).strip())
    duplicates = sorted({name for name in scenarios if scenarios.count(name) > 1})
    if not scenarios or duplicates:
        raise ValueError(
            "Control scenarios must be non-empty and unique; "
            f"duplicates={duplicates}"
        )
    return tuple(scenarios)


def _mapped_rows(source, target, country: str) -> list[tuple[object, ...]]:
    source_headers = _headers(source)
    target_headers = _headers(target)
    source_columns = _unique_columns(source_headers, sheet=source.title)
    target_columns = _unique_columns(target_headers, sheet=target.title)
    if "Tech" not in source_columns or "Tech" not in target_columns:
        raise ValueError(
            f"{source.title} -> {target.title} requires an exact Tech header"
        )
    if "scenario" not in target_columns:
        raise ValueError(f"{target.title} requires an exact scenario header")
    target_data_headers = set(target_columns) - {"scenario"}
    if set(source_columns) != target_data_headers:
        raise ValueError(
            f"header mismatch for {source.title} -> {target.title}: "
            f"missing={sorted(target_data_headers - set(source_columns))}, "
            f"extra={sorted(set(source_columns) - target_data_headers)}"
        )
    rows: list[tuple[object, ...]] = []
    for values in source.iter_rows(min_row=2, values_only=True):
        tech = values[source_columns["Tech"]]
        if not technology_belongs_to_country(tech, country):
            continue
        rows.append(
            tuple(values[source_columns[_normalized_header(header)]]
                  for header in target_headers
                  if _normalized_header(header) != "scenario")
        )
    return rows


def _sync_sheet(source, target, country: str, scenarios: Sequence[str]) -> SyncChange:
    target_headers = _headers(target)
    target_columns = _unique_columns(target_headers, sheet=target.title)
    scenario_index = target_columns["scenario"]
    tech_index = target_columns["Tech"]
    data_indices = [
        index for index, header in enumerate(target_headers)
        if _normalized_header(header) != "scenario"
    ]
    source_rows = _mapped_rows(source, target, country)
    existing: set[tuple[str, tuple[object, ...]]] = set()
    for values in target.iter_rows(min_row=2, values_only=True):
        scenario = values[scenario_index]
        if scenario is None:
            continue
        existing.add(
            (str(scenario).strip(), tuple(values[index] for index in data_indices))
        )

    added = 0
    affected_techs: set[str] = set()
    for scenario in scenarios:
        for data in source_rows:
            signature = (scenario, data)
            if signature in existing:
                continue
            output: list[object] = []
            data_iterator = iter(data)
            for header in target_headers:
                output.append(
                    scenario
                    if _normalized_header(header) == "scenario"
                    else next(data_iterator)
                )
            target.append(output)
            existing.add(signature)
            added += 1
            affected_techs.add(str(output[tech_index]))

    expected = {(scenario, data) for scenario in scenarios for data in source_rows}
    missing = expected - existing
    if missing:
        raise AssertionError(
            f"post-sync validation failed for {target.title}: {len(missing)} rows missing"
        )
    return SyncChange(
        source.title,
        target.title,
        tuple(scenarios),
        tuple(sorted(affected_techs)),
        added,
    )


def synchronize_country(
    *,
    country: str,
    ao_path: Path,
    scenario_path: Path,
    dry_run: bool = False,
) -> tuple[SyncChange, ...]:
    country = country.strip().upper()
    if not COUNTRY.fullmatch(country):
        raise ValueError(f"country must be an ISO-3 code: {country!r}")
    ao_path = Path(ao_path).resolve()
    scenario_path = Path(scenario_path).resolve()
    if not ao_path.is_file():
        raise FileNotFoundError(f"A-O workbook not found: {ao_path}")
    if not scenario_path.is_file():
        raise FileNotFoundError(f"scenario workbook not found: {scenario_path}")

    source_workbook = load_workbook(ao_path, read_only=True, data_only=False)
    target_workbook = load_workbook(scenario_path, data_only=False)
    try:
        scenarios = _control_scenarios(target_workbook)
        changes: list[SyncChange] = []
        for source_name, target_name in SHEET_MAP:
            if source_name not in source_workbook.sheetnames:
                raise ValueError(f"A-O workbook is missing {source_name!r}")
            if target_name not in target_workbook.sheetnames:
                raise ValueError(f"scenario workbook is missing {target_name!r}")
            changes.append(
                _sync_sheet(
                    source_workbook[source_name],
                    target_workbook[target_name],
                    country,
                    scenarios,
                )
            )
        if not dry_run:
            temporary = scenario_path.with_name(
                f".{scenario_path.stem}.syncing{scenario_path.suffix}"
            )
            if temporary.exists():
                raise FileExistsError(f"interrupted sync output exists: {temporary}")
            target_workbook.save(temporary)
            os.replace(temporary, scenario_path)
        return tuple(changes)
    finally:
        source_workbook.close()
        target_workbook.close()


def _default_ao(paths) -> Path:
    snapshots = sorted(paths.a1_outputs.glob("_post_a2_snapshot_*/A-O_Parametrization.xlsx"))
    if len(snapshots) != 1:
        raise FileNotFoundError(
            "--ao is required unless exactly one post-A2 parametrization exists"
        )
    return snapshots[0]


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(prog="python -m ostram scenario")
    subparsers = parser.add_subparsers(dest="action", required=True)
    sync = subparsers.add_parser("sync-country")
    sync.add_argument("--country", required=True)
    sync.add_argument("--ao", type=Path)
    sync.add_argument("--target", type=Path)
    sync.add_argument("--dry-run", action="store_true")
    args = parser.parse_args(argv)
    paths = resolve_paths()
    target = (args.target or paths.scenario_workbook).resolve()
    prepared_root_raw = os.environ.get(PROFILE_WORKSPACE_ENV)
    if not prepared_root_raw:
        raise RuntimeError("scenario sync-country requires an activated profile workspace")
    prepared_root = Path(prepared_root_raw).resolve()
    try:
        target.relative_to(prepared_root)
    except ValueError as error:
        raise ValueError(
            "scenario sync-country refuses to mutate a committed profile seed; "
            f"target must be inside {prepared_root}"
        ) from error
    changes = synchronize_country(
        country=args.country,
        ao_path=(args.ao.resolve() if args.ao else _default_ao(paths)),
        scenario_path=target,
        dry_run=args.dry_run,
    )
    total = sum(change.rows_added for change in changes)
    print(f"{'Would add' if args.dry_run else 'Added'} {total} scenario-country rows")
    for change in changes:
        print(
            f"  {change.source_sheet} -> {change.target_sheet}: "
            f"{change.rows_added} rows; technologies={list(change.technologies)}"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
