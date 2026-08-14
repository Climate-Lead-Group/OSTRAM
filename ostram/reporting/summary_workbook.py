"""Readable multi-sheet Excel summary of OSTRAM combined result CSVs.

A thin xlsx writer over the training dashboard's aggregation layer
(:mod:`ostram.reporting.training_dashboard`). Sheets that the dashboard
already computes (capacity, generation, interconnector series, cost KPIs)
are written from :func:`aggregate_metrics` output, so workbook numbers equal
dashboard numbers by construction. Three extensions follow the same coding
pattern as the dashboard's ``cost_kpis``:

1. operating costs with the fixed/variable split
   (``AnnualFixedOperatingCost`` + ``AnnualVariableOperatingCost``);
2. per-year ``TotalDiscountedCost`` with a cumulative column;
3. seasonal peaks from timeslice-level ``ProductionByTechnology``
   (peak PV output, peak interconnector flow by season, yearly corridor
   flow as a cross-check of the Power Flows sheet).

Input contract: the per-scenario combined input/output CSVs the report
route consumes (wide by parameter, merged on the otoole set columns),
including labelled captures under ``reports/snapshots/``. Accepts folders,
CSV files, or zip archives containing CSVs. Scenario identity comes from
the ``Scenario`` column when present, else from the file name.

Non-destructive: reads inputs, writes ``OSTRAM_Summary_<timestamp>.xlsx``,
never mutates anything.

Usage::

    python -m ostram.reporting.summary_workbook <folder-or-zip-or-csv ...> \
        [--out PATH] [--manifest examples/unescap/profile.yaml]
"""

from __future__ import annotations

import argparse
from contextlib import ExitStack
from datetime import datetime, timezone
from pathlib import Path
import re
import tempfile
from typing import Iterable, Mapping, Sequence
import zipfile

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ostram.reporting.training_dashboard import (
    BACKSTOP_PREFIX,
    INTERNAL_TRANSMISSION_PREFIX,
    ORDERED_FAMILIES,
    TECH_FAMILIES,
    _dedup,
    _scenario_frame,
    aggregate_metrics,
    is_storage_tech,
    storage_prefixes,
    tech_family,
)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Columns the workbook reads from a combined CSV. A superset of the
# dashboard's NEEDED_COLUMNS: the extensions add the operating-cost results,
# storage capex, the timeslice-level production result and YearSplit.
NEEDED_COLUMNS = (
    "Scenario", "REGION", "YEAR", "TECHNOLOGY", "EMISSION", "FUEL",
    "TIMESLICE", "SEASON", "STORAGE",
    "ProductionByTechnologyAnnual", "ProductionByTechnology",
    "TotalCapacityAnnual", "NewCapacity", "TotalAnnualMaxCapacityInvestment",
    "AnnualEmissions", "TotalDiscountedCost", "CapitalInvestment",
    "CapitalInvestmentStorage",
    "AnnualFixedOperatingCost", "AnnualVariableOperatingCost",
    "YearSplit", "Trade",
)

# Unit conversions. Production results are PJ; capacities are GW.
PJ_TO_TWH = 0.2777778
PJ_TO_GWH = 277.7778
# One GW running for a full year produces 31.536 PJ, so the average power in
# a timeslice is  GW = value_PJ / (YearSplit x 31.536).
PJ_PER_GW_YEAR = 31.536

OTHER_CATEGORY = "Other (uncategorised)"
EXTRA_CATEGORIES = (
    "Storage", "Interconnector", "Backstop", "Transmission (internal)",
    OTHER_CATEGORY,
)

_INTERCONNECTOR = re.compile(r"^TRN[A-Z0-9]{10}$")
_SEASON = re.compile(r"^(S\d+)")

# Number formats per unit (comma separators, 0-2 decimals by unit).
FORMATS = {
    "MUSD": "#,##0",
    "GW": "#,##0.00",
    "TWh": "#,##0.00",
    "Mt": "#,##0.00",
    "GWh": "#,##0",
}

_TITLE_FONT = Font(bold=True, size=13)
_UNIT_FONT = Font(italic=True, color="666666")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="00414C")
_BLOCK_FONT = Font(bold=True, size=11)


# ---------------------------------------------------------------------------
# Input collection and loading
# ---------------------------------------------------------------------------

def _load_frame(path: Path) -> pd.DataFrame:
    """Load only the columns the workbook needs; coerce YEAR to Int64."""
    header = pd.read_csv(path, nrows=0)
    usable = [column for column in NEEDED_COLUMNS if column in header.columns]
    frame = pd.read_csv(path, usecols=usable or None, low_memory=False)
    if "YEAR" in frame.columns:
        frame["YEAR"] = pd.to_numeric(frame["YEAR"], errors="coerce").astype("Int64")
    return frame


def _collect_csv_paths(
    inputs: Sequence[Path], stack: ExitStack
) -> list[tuple[str, Path]]:
    """Resolve folders/zips/files to ``(label, csv_path)`` pairs.

    For a folder, canonical ``*Combined_Inputs_Outputs.csv`` files win when
    present (this skips the orchestrator's dated duplicate copies); otherwise
    every ``*.csv`` in the folder is taken (the snapshot-folder case).
    """
    collected: list[tuple[str, Path]] = []

    def _add(path: Path) -> None:
        collected.append((path.stem, path))

    for raw in inputs:
        path = Path(raw)
        if path.is_dir():
            canonical = sorted(path.rglob("*Combined_Inputs_Outputs.csv"))
            for item in canonical or sorted(path.rglob("*.csv")):
                _add(item)
        elif path.suffix.lower() == ".zip":
            extracted = Path(stack.enter_context(tempfile.TemporaryDirectory()))
            with zipfile.ZipFile(path) as archive:
                members = [
                    name for name in archive.namelist()
                    if name.lower().endswith(".csv")
                ]
                archive.extractall(extracted, members=members)
            for item in sorted(extracted.rglob("*.csv")):
                _add(item)
        elif path.is_file():
            _add(path)
        else:
            raise FileNotFoundError(f"input not found: {path}")

    if not collected:
        raise FileNotFoundError(f"no CSV inputs found under: {list(inputs)}")
    return collected


def _scenario_bundles(
    sources: list[tuple[str, Path, pd.DataFrame]],
) -> list[dict[str, object]]:
    """One bundle per (source file, scenario): label, scenario id, sub-frame.

    Scenario ids come from the ``Scenario`` column when present, else the
    file name. When the same scenario id appears in more than one input file
    (before/after snapshots of one scenario), labels are disambiguated with
    the file stem so both stay visible side by side.
    """
    raw: list[tuple[str, str, pd.DataFrame]] = []
    for stem, _path, frame in sources:
        if "Scenario" in frame.columns:
            scenarios = sorted(frame["Scenario"].dropna().astype(str).unique())
        else:
            scenarios = [stem]
        for scenario in scenarios:
            raw.append((stem, scenario, frame))

    counts: dict[str, int] = {}
    for _stem, scenario, _frame in raw:
        counts[scenario] = counts.get(scenario, 0) + 1

    bundles: list[dict[str, object]] = []
    for stem, scenario, frame in raw:
        label = scenario if counts[scenario] == 1 else f"{stem} · {scenario}"
        bundles.append({
            "label": label,
            "scenario": scenario,
            "frame": _scenario_frame(frame, scenario),
        })
    return bundles


def _detect_interconnectors(frames: Iterable[pd.DataFrame]) -> list[str]:
    """13-character TRN technology codes observed in the data."""
    found: set[str] = set()
    for frame in frames:
        if "TECHNOLOGY" not in frame.columns:
            continue
        for tech in frame["TECHNOLOGY"].dropna().astype(str).unique():
            if _INTERCONNECTOR.fullmatch(tech):
                found.add(tech)
    return sorted(found)


# ---------------------------------------------------------------------------
# Categorisation (printed on the Readme sheet)
# ---------------------------------------------------------------------------

def categorize_tech(technology: object, storage: tuple[str, ...]) -> str:
    """Bucket any technology code; never drops a tech silently."""
    value = str(technology)
    if is_storage_tech(value, storage):
        return "Storage"
    if value.startswith(INTERNAL_TRANSMISSION_PREFIX):
        return "Transmission (internal)"
    if value.startswith(BACKSTOP_PREFIX):
        return "Backstop"
    family = tech_family(value)
    if family:
        return family
    if _INTERCONNECTOR.fullmatch(value):
        return "Interconnector"
    return OTHER_CATEGORY


def _category_order(present: Iterable[str]) -> list[str]:
    ordered = [f for f in ORDERED_FAMILIES if f in present]
    ordered.extend(c for c in EXTRA_CATEGORIES if c in present)
    ordered.extend(sorted(set(present) - set(ordered)))
    return ordered


# ---------------------------------------------------------------------------
# Extensions (same coding pattern as training_dashboard.cost_kpis)
# ---------------------------------------------------------------------------

def stacked_by_category(
    scenario_frame: pd.DataFrame,
    value_column: str,
    storage: tuple[str, ...],
    uncategorised: set[str],
) -> dict[int, dict[str, float]]:
    """{year: {category: value}} over ALL technologies (nothing dropped)."""
    required = {value_column, "YEAR", "TECHNOLOGY"}
    if not required.issubset(scenario_frame.columns):
        return {}
    sub = scenario_frame[
        scenario_frame[value_column].notna()
        & scenario_frame["YEAR"].notna()
        & scenario_frame["TECHNOLOGY"].notna()
    ]
    sub = _dedup(sub, ("YEAR", "TECHNOLOGY", value_column)).copy()
    if sub.empty:
        return {}
    sub["category"] = sub["TECHNOLOGY"].map(
        lambda tech: categorize_tech(tech, storage)
    )
    for tech in sub.loc[sub["category"] == OTHER_CATEGORY, "TECHNOLOGY"]:
        uncategorised.add(str(tech))
    out: dict[int, dict[str, float]] = {}
    grouped = sub.groupby(["YEAR", "category"])[value_column].apply(
        lambda values: float(pd.to_numeric(values, errors="coerce").sum())
    )
    for (year, category), value in grouped.items():
        out.setdefault(int(year), {})[category] = round(float(value), 4)
    return out


def storage_capex_block(scenario_frame: pd.DataFrame) -> dict[int, dict[str, float]]:
    """{year: {storage_code: MUSD}} from CapitalInvestmentStorage."""
    required = {"CapitalInvestmentStorage", "YEAR", "STORAGE"}
    if not required.issubset(scenario_frame.columns):
        return {}
    sub = scenario_frame[
        scenario_frame["CapitalInvestmentStorage"].notna()
        & scenario_frame["YEAR"].notna()
        & scenario_frame["STORAGE"].notna()
    ]
    sub = sub[sub["STORAGE"].astype(str) != "nan"]
    sub = _dedup(sub, ("YEAR", "STORAGE", "CapitalInvestmentStorage"))
    out: dict[int, dict[str, float]] = {}
    grouped = sub.groupby(["YEAR", "STORAGE"])["CapitalInvestmentStorage"].sum()
    for (year, code), value in grouped.items():
        out.setdefault(int(year), {})[str(code)] = round(float(value), 4)
    return out


def system_cost_series(scenario_frame: pd.DataFrame) -> dict[int, float]:
    """Per-year TotalDiscountedCost (the dashboard only sums the horizon)."""
    required = {"TotalDiscountedCost", "YEAR"}
    if not required.issubset(scenario_frame.columns):
        return {}
    sub = scenario_frame[
        scenario_frame["TotalDiscountedCost"].notna()
        & scenario_frame["YEAR"].notna()
    ]
    sub = _dedup(sub, ("REGION", "YEAR", "TotalDiscountedCost"))
    grouped = sub.groupby("YEAR")["TotalDiscountedCost"].sum()
    return {int(year): float(value) for year, value in grouped.items()}


def emissions_by_code(scenario_frame: pd.DataFrame) -> dict[str, dict[int, float]]:
    """{emission_code: {year: value}} (the dashboard sums across codes)."""
    required = {"AnnualEmissions", "EMISSION", "YEAR"}
    if not required.issubset(scenario_frame.columns):
        return {}
    sub = scenario_frame[
        scenario_frame["AnnualEmissions"].notna()
        & scenario_frame["EMISSION"].notna()
        & scenario_frame["YEAR"].notna()
    ]
    sub = _dedup(sub, ("YEAR", "EMISSION", "AnnualEmissions"))
    out: dict[str, dict[int, float]] = {}
    grouped = sub.groupby(["EMISSION", "YEAR"])["AnnualEmissions"].sum()
    for (code, year), value in grouped.items():
        out.setdefault(str(code), {})[int(year)] = float(value)
    return out


def corridor_flows_annual(
    scenario_frame: pd.DataFrame, technologies: Sequence[str]
) -> dict[str, dict[int, float]]:
    """{corridor_tech: {year: PJ}} from annual TRN production."""
    required = {"ProductionByTechnologyAnnual", "YEAR", "TECHNOLOGY"}
    if not technologies or not required.issubset(scenario_frame.columns):
        return {}
    sub = scenario_frame[
        scenario_frame["ProductionByTechnologyAnnual"].notna()
        & scenario_frame["YEAR"].notna()
        & scenario_frame["TECHNOLOGY"].astype(str).isin(set(technologies))
    ]
    sub = _dedup(sub, ("YEAR", "TECHNOLOGY", "ProductionByTechnologyAnnual"))
    out: dict[str, dict[int, float]] = {}
    grouped = sub.groupby(["TECHNOLOGY", "YEAR"])["ProductionByTechnologyAnnual"].sum()
    for (tech, year), value in grouped.items():
        out.setdefault(str(tech), {})[int(year)] = float(value)
    return out


def timeslice_season(timeslice: object) -> str:
    """S1D3 -> S1; anything unrecognised is returned verbatim."""
    match = _SEASON.match(str(timeslice))
    return match.group(1) if match else str(timeslice)


def year_split_map(frame: pd.DataFrame) -> dict[tuple[str, int], float]:
    """{(timeslice, year): fraction} from the YearSplit input parameter."""
    required = {"YearSplit", "TIMESLICE", "YEAR"}
    if not required.issubset(frame.columns):
        return {}
    sub = frame[
        frame["YearSplit"].notna()
        & frame["TIMESLICE"].notna()
        & frame["YEAR"].notna()
    ]
    sub = _dedup(sub, ("TIMESLICE", "YEAR", "YearSplit"))
    return {
        (str(row["TIMESLICE"]), int(row["YEAR"])): float(row["YearSplit"])
        for _, row in sub.iterrows()
    }


def seasonal_peaks(
    scenario_frame: pd.DataFrame,
    interconnector_techs: Sequence[str],
    year_split: Mapping[tuple[str, int], float],
) -> dict[str, object]:
    """Timeslice-level extraction for the Seasonal Peaks sheet.

    Returns ``pv_peaks`` rows ``(year, gw, timeslice, season)``,
    ``corridor_by_season`` as ``{year: {season: gw}}``, plus a ``gaps``
    list naming whatever was missing (the defensive GAP fallback).
    """
    gaps: list[str] = []
    required = {"ProductionByTechnology", "TIMESLICE", "YEAR", "TECHNOLOGY"}
    missing = sorted(required - set(scenario_frame.columns))
    if missing:
        gaps.append(
            "timeslice-level production is unavailable: missing column(s) "
            + ", ".join(missing)
        )
        return {"pv_peaks": [], "corridor_by_season": {}, "gaps": gaps}
    if not year_split:
        gaps.append(
            "YearSplit rows are unavailable, so timeslice energy (PJ) cannot "
            "be converted to power (GW)"
        )
        return {"pv_peaks": [], "corridor_by_season": {}, "gaps": gaps}

    sub = scenario_frame[
        scenario_frame["ProductionByTechnology"].notna()
        & scenario_frame["TIMESLICE"].notna()
        & scenario_frame["YEAR"].notna()
        & scenario_frame["TECHNOLOGY"].notna()
    ]
    sub = sub[sub["TIMESLICE"].astype(str) != "nan"]
    identity = ("YEAR", "TIMESLICE", "TECHNOLOGY", "FUEL", "ProductionByTechnology")
    sub = _dedup(sub, identity)
    if sub.empty:
        gaps.append("no timeslice-level ProductionByTechnology rows were found")
        return {"pv_peaks": [], "corridor_by_season": {}, "gaps": gaps}

    def _to_gw(group: pd.DataFrame) -> dict[tuple[int, str], float]:
        """{(year, timeslice): GW} for the selected rows."""
        out: dict[tuple[int, str], float] = {}
        grouped = group.groupby(["YEAR", "TIMESLICE"])["ProductionByTechnology"].sum()
        for (year, timeslice), value in grouped.items():
            fraction = year_split.get((str(timeslice), int(year)))
            if not fraction:
                continue
            out[(int(year), str(timeslice))] = (
                float(value) / (fraction * PJ_PER_GW_YEAR)
            )
        return out

    pv_rows = sub[sub["TECHNOLOGY"].astype(str).str.startswith("PWRSPV")]
    pv_peaks: list[tuple[int, float, str, str]] = []
    pv_gw = _to_gw(pv_rows)
    for year in sorted({year for year, _ in pv_gw}):
        timeslice, gw = max(
            ((ts, value) for (y, ts), value in pv_gw.items() if y == year),
            key=lambda item: item[1],
        )
        pv_peaks.append((year, gw, timeslice, timeslice_season(timeslice)))
    if not pv_peaks:
        gaps.append("no PWRSPV timeslice production was found for the PV peak block")

    corridor_by_season: dict[int, dict[str, float]] = {}
    if interconnector_techs:
        corridor_rows = sub[
            sub["TECHNOLOGY"].astype(str).isin(set(interconnector_techs))
        ]
        for (year, timeslice), gw in _to_gw(corridor_rows).items():
            season = timeslice_season(timeslice)
            seasons = corridor_by_season.setdefault(year, {})
            seasons[season] = max(seasons.get(season, 0.0), gw)
    if not corridor_by_season:
        gaps.append(
            "no interconnector timeslice production was found for the "
            "seasonal corridor-peak block"
        )
    return {"pv_peaks": pv_peaks, "corridor_by_season": corridor_by_season, "gaps": gaps}


# ---------------------------------------------------------------------------
# Workbook writer
# ---------------------------------------------------------------------------

def _init_sheet(workbook: Workbook, name: str, title: str, unit_note: str):
    sheet = workbook.create_sheet(name)
    sheet.cell(row=1, column=1, value=title).font = _TITLE_FONT
    sheet.cell(row=2, column=1, value=unit_note).font = _UNIT_FONT
    return sheet


def _write_table(
    sheet,
    start_row: int,
    columns: Sequence[str],
    rows: Sequence[Sequence[object]],
    value_format: str,
    first_value_column: int = 3,
    autofilter: bool = False,
) -> int:
    """Write one header+data table; return the next free row."""
    for index, column in enumerate(columns, start=1):
        cell = sheet.cell(row=start_row, column=index, value=column)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for offset, row in enumerate(rows, start=1):
        for index, value in enumerate(row, start=1):
            cell = sheet.cell(row=start_row + offset, column=index, value=value)
            if index >= first_value_column and isinstance(value, (int, float)):
                cell.number_format = value_format
    last_row = start_row + len(rows)
    if autofilter and rows:
        sheet.auto_filter.ref = (
            f"A{start_row}:{get_column_letter(len(columns))}{last_row}"
        )
    for index, column in enumerate(columns, start=1):
        letter = get_column_letter(index)
        width = 30 if index == 1 else (10 if index == 2 else 14)
        sheet.column_dimensions[letter].width = max(
            sheet.column_dimensions[letter].width or 0, width, len(str(column)) + 2
        )
    return last_row + 2


def _block_title(sheet, row: int, text: str) -> int:
    sheet.cell(row=row, column=1, value=text).font = _BLOCK_FONT
    return row + 1


def _category_table_rows(
    per_bundle: Mapping[str, dict[int, dict[str, float]]],
    categories: Sequence[str],
    scale: float = 1.0,
) -> list[list[object]]:
    rows: list[list[object]] = []
    for label, by_year in per_bundle.items():
        for year in sorted(by_year):
            values = by_year[year]
            row: list[object] = [label, year]
            total = 0.0
            for category in categories:
                value = values.get(category)
                if value is None:
                    row.append(None)
                else:
                    scaled = value * scale
                    row.append(scaled)
                    total += scaled
            row.append(total)
            rows.append(row)
    return rows


def _categories_present(
    per_bundle: Mapping[str, dict[int, dict[str, float]]],
) -> list[str]:
    present: set[str] = set()
    for by_year in per_bundle.values():
        for values in by_year.values():
            present.update(values)
    return _category_order(present)


def build_workbook(
    sources: list[tuple[str, Path, pd.DataFrame]],
    output: Path,
    metadata: Mapping[str, object] | None = None,
) -> Path:
    """Aggregate every scenario in ``sources`` and write the workbook."""
    metadata = dict(metadata or {})
    storage = storage_prefixes(metadata)
    bundles = _scenario_bundles(sources)

    declared = metadata.get("interconnectors")
    if isinstance(declared, list) and declared:
        interconnector_techs = sorted(
            str(entry["technology"]) if isinstance(entry, Mapping) else str(entry)
            for entry in declared
        )
    else:
        interconnector_techs = _detect_interconnectors(
            frame for _stem, _path, frame in sources
        )

    uncategorised: set[str] = set()
    skipped: list[str] = []
    gap_notes: list[str] = []
    computed: list[dict[str, object]] = []
    for bundle in bundles:
        frame: pd.DataFrame = bundle["frame"]  # type: ignore[assignment]
        metrics = aggregate_metrics(
            frame,
            scenario=str(bundle["scenario"]),
            region="System",
            interconnectors=[{"technology": t} for t in interconnector_techs],
            metadata=metadata,
        )
        if not metrics["available"]:
            skipped.append(str(bundle["label"]))
            continue
        peaks = seasonal_peaks(frame, interconnector_techs, year_split_map(frame))
        for gap in peaks["gaps"]:
            gap_notes.append(f"{bundle['label']}: {gap}")
        computed.append({
            **bundle,
            "metrics": metrics,
            "capex": stacked_by_category(frame, "CapitalInvestment", storage, uncategorised),
            "capex_storage": storage_capex_block(frame),
            "opex_fixed": stacked_by_category(frame, "AnnualFixedOperatingCost", storage, uncategorised),
            "opex_variable": stacked_by_category(frame, "AnnualVariableOperatingCost", storage, uncategorised),
            "system_cost": system_cost_series(frame),
            "emissions": emissions_by_code(frame),
            "corridor_annual": corridor_flows_annual(frame, interconnector_techs),
            "peaks": peaks,
        })

    if not computed:
        raise ValueError(
            "no scenario in the inputs has solver results; nothing to summarise"
        )

    workbook = Workbook()
    workbook.remove(workbook.active)

    # --- Sheet 1: Capital expenditure -----------------------------------
    capex = {str(b["label"]): b["capex"] for b in computed}
    categories = _categories_present(capex)
    sheet = _init_sheet(
        workbook, "1 Capital Expenditure",
        "Capital expenditure by technology category (CapitalInvestment)",
        "Unit: MUSD — million US dollars",
    )
    next_row = _write_table(
        sheet, 3, ["Scenario", "Year", *categories, "Total"],
        _category_table_rows(capex, categories),
        FORMATS["MUSD"], autofilter=True,
    )
    storage_capex = {str(b["label"]): b["capex_storage"] for b in computed}
    storage_codes = sorted({
        code for by_year in storage_capex.values()
        for values in by_year.values() for code in values
    })
    if storage_codes:
        next_row = _block_title(
            sheet, next_row,
            "Storage capital investment (CapitalInvestmentStorage, MUSD)",
        )
        _write_table(
            sheet, next_row, ["Scenario", "Year", *storage_codes, "Total"],
            _category_table_rows(storage_capex, storage_codes),
            FORMATS["MUSD"],
        )
    sheet.freeze_panes = "C4"

    # --- Sheet 2: Operating costs ----------------------------------------
    sheet = _init_sheet(
        workbook, "2 Operating Costs",
        "Operating costs by technology category "
        "(AnnualFixedOperatingCost + AnnualVariableOperatingCost)",
        "Unit: MUSD — million US dollars; fixed and variable shown "
        "separately, then summed",
    )
    opex_kinds: dict[str, dict[int, dict[str, float]]] = {}
    for bundle in computed:
        label = str(bundle["label"])
        fixed: dict[int, dict[str, float]] = bundle["opex_fixed"]  # type: ignore[assignment]
        variable: dict[int, dict[str, float]] = bundle["opex_variable"]  # type: ignore[assignment]
        total: dict[int, dict[str, float]] = {}
        for source in (fixed, variable):
            for year, values in source.items():
                bucket = total.setdefault(year, {})
                for category, value in values.items():
                    bucket[category] = bucket.get(category, 0.0) + value
        opex_kinds[f"{label} — Fixed"] = fixed
        opex_kinds[f"{label} — Variable"] = variable
        opex_kinds[f"{label} — Fixed + Variable"] = total
    categories = _categories_present(opex_kinds)
    rows = []
    for key, by_year in opex_kinds.items():
        label, _, kind = key.rpartition(" — ")
        for year in sorted(by_year):
            values = by_year[year]
            row: list[object] = [label, kind, year]
            total_value = 0.0
            for category in categories:
                value = values.get(category)
                row.append(value)
                total_value += value or 0.0
            row.append(total_value)
            rows.append(row)
    _write_table(
        sheet, 3, ["Scenario", "Cost type", "Year", *categories, "Total"],
        rows, FORMATS["MUSD"], first_value_column=4, autofilter=True,
    )
    sheet.freeze_panes = "D4"

    # --- Sheet 3: Total system cost --------------------------------------
    sheet = _init_sheet(
        workbook, "3 Total System Cost",
        "Total discounted system cost per year (TotalDiscountedCost)",
        "Unit: MUSD — million US dollars; cumulative column runs within "
        "each scenario",
    )
    rows = []
    for bundle in computed:
        series: dict[int, float] = bundle["system_cost"]  # type: ignore[assignment]
        running = 0.0
        for year in sorted(series):
            running += series[year]
            rows.append([str(bundle["label"]), year, series[year], running])
        horizon = bundle["metrics"]["cost"].get("total_discounted", 0.0)  # type: ignore[index]
        if series and abs(running - horizon) > max(1e-6 * abs(horizon), 1e-6):
            gap_notes.append(
                f"{bundle['label']}: per-year TotalDiscountedCost sums to "
                f"{running:.4f} but the dashboard horizon total is {horizon:.4f}"
            )
    _write_table(
        sheet, 3,
        ["Scenario", "Year", "Total discounted cost", "Cumulative"],
        rows, FORMATS["MUSD"], autofilter=True,
    )
    sheet.freeze_panes = "C4"

    # --- Sheet 4: Capacity mix -------------------------------------------
    capacity: dict[str, dict[int, dict[str, float]]] = {}
    for bundle in computed:
        metrics = bundle["metrics"]
        by_year = {
            year: dict(values)
            for year, values in metrics["capacity"].items()  # type: ignore[index]
        }
        for year, value in metrics["storage"].items():  # type: ignore[index]
            by_year.setdefault(year, {})["Storage"] = value
        inter = metrics["interconnectors"].get("TotalCapacityAnnual", {})  # type: ignore[index]
        for year, value in inter.items():
            by_year.setdefault(year, {})["Interconnector"] = value
        capacity[str(bundle["label"])] = by_year
    categories = _categories_present(capacity)
    sheet = _init_sheet(
        workbook, "4 Capacity Mix",
        "Installed capacity by technology category (TotalCapacityAnnual)",
        "Unit: GW — gigawatts",
    )
    _write_table(
        sheet, 3, ["Scenario", "Year", *categories, "Total"],
        _category_table_rows(capacity, categories),
        FORMATS["GW"], autofilter=True,
    )
    sheet.freeze_panes = "C4"

    # --- Sheet 5: Generation mix ------------------------------------------
    generation = {
        str(b["label"]): b["metrics"]["generation"]  # type: ignore[index]
        for b in computed
    }
    categories = _categories_present(generation)
    sheet = _init_sheet(
        workbook, "5 Generation Mix",
        "Electricity generation by technology category "
        "(ProductionByTechnologyAnnual)",
        f"Unit: TWh — terawatt hours (PJ × {PJ_TO_TWH})",
    )
    _write_table(
        sheet, 3, ["Scenario", "Year", *categories, "Total"],
        _category_table_rows(generation, categories, scale=PJ_TO_TWH),
        FORMATS["TWh"], autofilter=True,
    )
    sheet.freeze_panes = "C4"

    # --- Sheet 6: Emissions -------------------------------------------------
    sheet = _init_sheet(
        workbook, "6 Emissions",
        "Annual emissions by emission code (AnnualEmissions)",
        "Unit: Mt CO2 — million tonnes",
    )
    emission_codes = sorted({
        code for b in computed for code in b["emissions"]  # type: ignore[union-attr]
    })
    rows = []
    for bundle in computed:
        by_code: dict[str, dict[int, float]] = bundle["emissions"]  # type: ignore[assignment]
        years = sorted({year for series in by_code.values() for year in series})
        for year in years:
            row: list[object] = [str(bundle["label"]), year]
            total_value = 0.0
            for code in emission_codes:
                value = by_code.get(code, {}).get(year)
                row.append(value)
                total_value += value or 0.0
            row.append(total_value)
            rows.append(row)
    _write_table(
        sheet, 3, ["Scenario", "Year", *emission_codes, "Total"],
        rows, FORMATS["Mt"], autofilter=True,
    )
    sheet.freeze_panes = "C4"

    # --- Sheet 7: Power flows ------------------------------------------------
    sheet = _init_sheet(
        workbook, "7 Power Flows",
        "Interconnector corridor flows "
        "(ProductionByTechnologyAnnual over TRN technologies)",
        f"Unit: GWh — gigawatt hours (PJ × {PJ_TO_GWH})",
    )
    rows = []
    for bundle in computed:
        corridors: dict[str, dict[int, float]] = bundle["corridor_annual"]  # type: ignore[assignment]
        years = sorted({year for series in corridors.values() for year in series})
        for year in years:
            row: list[object] = [str(bundle["label"]), year]
            total_value = 0.0
            for tech in interconnector_techs:
                value = corridors.get(tech, {}).get(year)
                scaled = None if value is None else value * PJ_TO_GWH
                row.append(scaled)
                total_value += scaled or 0.0
            row.append(total_value)
            rows.append(row)
    _write_table(
        sheet, 3, ["Scenario", "Year", *interconnector_techs, "Total"],
        rows, FORMATS["GWh"], autofilter=True,
    )
    sheet.freeze_panes = "C4"

    # --- Sheet 8: Seasonal peaks ----------------------------------------------
    sheet = _init_sheet(
        workbook, "8 Seasonal Peaks",
        "Timeslice-level peaks (ProductionByTechnology)",
        f"Unit: GW — average power in a timeslice: "
        f"PJ / (YearSplit × {PJ_PER_GW_YEAR}); seasons are the model's "
        "S1-S4, not a binary rainy/dry split",
    )
    next_row = _block_title(sheet, 3, "Peak PV output per year (GW)")
    rows = []
    for bundle in computed:
        peaks: dict[str, object] = bundle["peaks"]  # type: ignore[assignment]
        for year, gw, timeslice, season in peaks["pv_peaks"]:  # type: ignore[index]
            rows.append([str(bundle["label"]), year, gw, timeslice, season])
    next_row = _write_table(
        sheet, next_row,
        ["Scenario", "Year", "Peak PV (GW)", "Timeslice", "Season"],
        rows, FORMATS["GW"], autofilter=True,
    )
    seasons = sorted({
        season
        for b in computed
        for by_season in b["peaks"]["corridor_by_season"].values()  # type: ignore[index]
        for season in by_season
    })
    next_row = _block_title(
        sheet, next_row, "Peak interconnector flow per year, by season (GW)"
    )
    rows = []
    for bundle in computed:
        by_year: dict[int, dict[str, float]] = (
            bundle["peaks"]["corridor_by_season"]  # type: ignore[index]
        )
        for year in sorted(by_year):
            row: list[object] = [str(bundle["label"]), year]
            row.extend(by_year[year].get(season) for season in seasons)
            rows.append(row)
    next_row = _write_table(
        sheet, next_row, ["Scenario", "Year", *seasons], rows, FORMATS["GW"],
    )
    next_row = _block_title(
        sheet, next_row,
        "Yearly corridor flow (GWh) — cross-check of sheet 7",
    )
    rows = []
    for bundle in computed:
        corridors = bundle["corridor_annual"]  # type: ignore[assignment]
        years = sorted({year for series in corridors.values() for year in series})
        for year in years:
            row = [str(bundle["label"]), year]
            row.extend(
                (corridors.get(tech, {}).get(year) or 0.0) * PJ_TO_GWH
                for tech in interconnector_techs
            )
            rows.append(row)
    next_row = _write_table(
        sheet, next_row, ["Scenario", "Year", *interconnector_techs],
        rows, FORMATS["GWh"],
    )
    if gap_notes:
        next_row = _block_title(sheet, next_row, "DATA GAPS")
        for note in gap_notes:
            sheet.cell(row=next_row, column=1, value=note).font = _UNIT_FONT
            next_row += 1
    sheet.freeze_panes = "A4"

    # --- Sheet 0: Readme (built last, placed first) ---------------------------
    sheet = _init_sheet(
        workbook, "Readme", "OSTRAM Summary Workbook",
        "Generated by ostram.reporting.summary_workbook — read-only summary "
        "of combined result CSVs",
    )
    workbook.move_sheet("Readme", -(len(workbook.sheetnames) - 1))
    row = 4

    def _line(text: str, bold: bool = False) -> None:
        nonlocal row
        cell = sheet.cell(row=row, column=1, value=text)
        if bold:
            cell.font = _BLOCK_FONT
        row += 1

    _line(f"Generated (UTC): {datetime.now(timezone.utc).isoformat()}")
    _line("")
    _line("Input files", bold=True)
    for _stem, path, _frame in sources:
        stat = path.stat()
        _line(
            f"  {path} — {stat.st_size:,} bytes, modified "
            f"{datetime.fromtimestamp(stat.st_mtime, timezone.utc).isoformat()}"
        )
    _line("")
    _line("Scenarios", bold=True)
    for bundle in computed:
        _line(f"  {bundle['label']} (results present)")
    for label in skipped:
        _line(f"  {label} — SKIPPED: no solver output values in the input rows")
    _line("")
    _line("Unit conventions", bold=True)
    _line("  Costs: MUSD (million US dollars), as produced by the model")
    _line("  Capacity: GW")
    _line(f"  Generation: TWh = PJ × {PJ_TO_TWH}")
    _line(f"  Corridor flows: GWh = PJ × {PJ_TO_GWH}")
    _line(
        f"  Timeslice power: GW = PJ / (YearSplit × {PJ_PER_GW_YEAR}), "
        "with YearSplit read from the same combined CSV"
    )
    _line("  Emissions: Mt CO2, as produced by the model")
    _line("")
    _line("Technology categorisation", bold=True)
    _line(
        "  Category is decided by code prefix, in this order: storage "
        "prefixes, internal transmission, backstop, generation families, "
        "13-character TRN interconnectors, then 'Other (uncategorised)'."
    )
    for prefix, (label, _color) in TECH_FAMILIES.items():
        _line(f"  {prefix}* -> {label}")
    _line(f"  {' / '.join(f'{p}*' for p in storage)} -> Storage")
    _line(f"  {INTERNAL_TRANSMISSION_PREFIX}* -> Transmission (internal)")
    _line(f"  {BACKSTOP_PREFIX}* -> Backstop")
    _line("  TRN + 10 characters -> Interconnector")
    _line("")
    if uncategorised:
        _line("Uncategorised technologies (bucketed under "
              f"'{OTHER_CATEGORY}')", bold=True)
        for tech in sorted(uncategorised):
            _line(f"  {tech}")
    else:
        _line("Uncategorised technologies: none", bold=True)
    _line("")
    _line("Notes", bold=True)
    _line(
        "  Sheet 8 seasons are the model's four seasons (S1-S4), not a "
        "binary rainy/dry split."
    )
    _line(
        "  Corridor flows use TRN-technology production. The combined CSV "
        "keeps a single REGION column, so the two-region Trade result "
        "cannot express corridor direction there and is not used."
    )
    _line(
        "  Sheets 4/5/7 and the horizon cost totals are computed by the "
        "training dashboard's own aggregation (aggregate_metrics), so they "
        "match the HTML dashboard by construction."
    )
    if interconnector_techs:
        _line(f"  Interconnectors: {', '.join(interconnector_techs)}")
    else:
        _line("  Interconnectors: none declared or detected")
    if gap_notes:
        _line("")
        _line("Data gaps", bold=True)
        for note in gap_notes:
            _line(f"  {note}")
    sheet.column_dimensions["A"].width = 100
    sheet.freeze_panes = "A3"

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return output


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _default_output() -> Path:
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    return Path.cwd() / f"OSTRAM_Summary_{stamp}.xlsx"


def _load_manifest_metadata(path: Path) -> Mapping[str, object]:
    import yaml

    document = yaml.safe_load(path.read_text(encoding="utf-8"))
    if not isinstance(document, Mapping):
        raise ValueError(f"not a mapping document: {path}")
    metadata = document.get("metadata", document)
    if not isinstance(metadata, Mapping):
        raise ValueError(f"manifest metadata is not a mapping: {path}")
    return metadata


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        prog="python -m ostram.reporting.summary_workbook",
        description=__doc__.splitlines()[0],
    )
    parser.add_argument(
        "inputs", nargs="+", type=Path,
        help="combined result CSVs, folders containing them, or zip archives",
    )
    parser.add_argument(
        "--out", type=Path, default=None,
        help="output .xlsx path or directory "
             "(default: ./OSTRAM_Summary_<timestamp>.xlsx)",
    )
    parser.add_argument(
        "--manifest", type=Path, default=None,
        help="optional profile manifest (profile.yaml); its metadata supplies "
             "storage prefixes and declared interconnectors",
    )
    args = parser.parse_args(argv)

    output = args.out
    if output is None:
        output = _default_output()
    elif output.is_dir() or output.suffix.lower() != ".xlsx":
        output = output / _default_output().name
    if output.exists():
        parser.error(f"refusing to overwrite existing file: {output}")

    metadata = _load_manifest_metadata(args.manifest) if args.manifest else None

    with ExitStack() as stack:
        pairs = _collect_csv_paths(args.inputs, stack)
        sources = [(stem, path, _load_frame(path)) for stem, path in pairs]
        written = build_workbook(sources, output, metadata=metadata)
    print(f"Summary workbook: {written}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
